"""
Comparación de clientes - CERVEZAS.

Script standalone que genera un Excel comparando la venta en bultos
de cada cliente para Enero, Febrero y Marzo 2025 vs 2026.

Hojas:
  - Comparacion Clientes: bultos vendidos + bonificaciones (pareado 25|26 por mes)
  - Por Marca: apertura por marca con groups colapsables por cada marca
  - Por Lista Precio: resumen agrupado por lista de precio
  - Por Ruta: resumen agrupado por ruta

Uso:
    python comparacion_clientes.py
    python comparacion_clientes.py --sucursales "CASA CENTRAL,SUCURSAL CAFAYATE"
"""
import argparse

import pandas as pd

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.core.excel_writer import ExcelWriter, SheetStyle, ColumnFormat, ColumnGroup

FILL_TOTAL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

GENERICO = "CERVEZAS"
CLIENTES_XLSX = "/home/nahuel/VM shared/clientes.xlsx"

PERIODOS = {
    "Ene 2025": ("2025-01-01", "2025-01-31"),
    "Ene 2026": ("2026-01-01", "2026-01-31"),
    "Feb 2025": ("2025-02-01", "2025-02-28"),
    "Feb 2026": ("2026-02-01", "2026-02-28"),
    "Mar 2025": ("2025-03-01", "2025-03-31"),
    "Mar 2026": ("2026-03-01", "2026-03-31"),
}



def get_clientes(loader: DataLoader) -> pd.DataFrame:
    """Obtiene clientes que tuvieron ventas de CERVEZAS en alguno de los 4 períodos."""
    params = {"generico": GENERICO}
    condiciones = []
    for i, (desde, hasta) in enumerate(PERIODOS.values(), 1):
        params[f"desde_{i}"] = desde
        params[f"hasta_{i}"] = hasta
        condiciones.append(f"fv.fecha_comprobante BETWEEN :desde_{i} AND :hasta_{i}")

    query = f"""
    SELECT DISTINCT
        fv.id_cliente,
        fv.id_sucursal
    FROM gold.fact_ventas fv
    LEFT JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
    WHERE fv.id_sucursal = 1
      AND da.generico = :generico
      AND ({' OR '.join(condiciones)})
    ORDER BY fv.id_cliente
    """
    return loader.execute_query(query, params)


def cargar_datos_clientes() -> pd.DataFrame:
    """Carga datos de clientes desde el Excel externo (ruta, preventista, subcanal, etc.)."""
    df = pd.read_excel(CLIENTES_XLSX)
    df = df.rename(columns={
        "Sucursal": "id_sucursal",
        "Cliente": "id_cliente",
        "Razon social": "razon_social",
        "Subcanal": "id_subcanal",
        "Descripcion subcanal": "subcanal",
        "Lista de precios": "id_lista_precio",
        "Descripcion lista de precios": "desc_lista_precio",
        "Fuerza de venta 1 Ruta de venta": "id_ruta",
        "Fuerza de venta 1 Descripcion ruta de venta": "desc_ruta",
        "Fuerza de venta 1 Personal comercial": "id_preventista",
        "Fuerza de venta 1 Descripcion personal comercial": "preventista",
    })
    # Convertir claves a float para que matcheen con la BD
    df["id_cliente"] = pd.to_numeric(df["id_cliente"], errors="coerce")
    df["id_sucursal"] = pd.to_numeric(df["id_sucursal"], errors="coerce")
    # Quedarme con las columnas útiles
    cols = ["id_sucursal", "id_cliente", "razon_social", "subcanal",
            "id_lista_precio", "desc_lista_precio", "id_ruta", "desc_ruta",
            "id_preventista", "preventista"]
    return df[cols]


def get_ventas_periodo(loader: DataLoader, fecha_desde: str, fecha_hasta: str) -> pd.DataFrame:
    """Obtiene bultos vendidos por cliente en un período para CERVEZAS."""
    params = {"desde": fecha_desde, "hasta": fecha_hasta, "generico": GENERICO}

    query = """
    SELECT
        fv.id_cliente,
        fv.id_sucursal,
        SUM(fv.cantidades_total) AS bultos
    FROM gold.fact_ventas fv
    LEFT JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
    WHERE fv.id_sucursal = 1
      AND fv.fecha_comprobante BETWEEN :desde AND :hasta
      AND da.generico = :generico
    GROUP BY fv.id_cliente, fv.id_sucursal
    """
    return loader.execute_query(query, params)


def get_bonificaciones_periodo(loader: DataLoader, fecha_desde: str, fecha_hasta: str) -> pd.DataFrame:
    """Obtiene bultos bonificados por cliente en un período para CERVEZAS.

    Calcula: SUM(cantidades_total * bonificacion / 100) por cliente.
    """
    params = {"desde": fecha_desde, "hasta": fecha_hasta, "generico": GENERICO}

    query = """
    SELECT
        fv.id_cliente,
        fv.id_sucursal,
        SUM(fv.cantidades_total * fv.bonificacion / 100.0) AS bultos_bonificados
    FROM gold.fact_ventas fv
    LEFT JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
    WHERE fv.id_sucursal = 1
      AND fv.fecha_comprobante BETWEEN :desde AND :hasta
      AND da.generico = :generico
      AND fv.bonificacion > 0
    GROUP BY fv.id_cliente, fv.id_sucursal
    """
    return loader.execute_query(query, params)



def get_ventas_marca_periodo(loader: DataLoader, fecha_desde: str, fecha_hasta: str) -> pd.DataFrame:
    """Obtiene bultos vendidos por cliente y marca en un período para CERVEZAS."""
    params = {"desde": fecha_desde, "hasta": fecha_hasta, "generico": GENERICO}

    query = """
    SELECT
        fv.id_cliente,
        fv.id_sucursal,
        da.marca,
        SUM(fv.cantidades_total) AS bultos
    FROM gold.fact_ventas fv
    LEFT JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
    WHERE fv.id_sucursal = 1
      AND fv.fecha_comprobante BETWEEN :desde AND :hasta
      AND da.generico = :generico
    GROUP BY fv.id_cliente, fv.id_sucursal, da.marca
    """
    return loader.execute_query(query, params)


def get_ventas_articulo_periodo(loader: DataLoader, fecha_desde: str, fecha_hasta: str) -> pd.DataFrame:
    """Obtiene bultos vendidos por cliente, artículo y marca en un período."""
    params = {"desde": fecha_desde, "hasta": fecha_hasta, "generico": GENERICO}

    query = """
    SELECT
        fv.id_cliente,
        fv.id_sucursal,
        fv.id_articulo,
        da.des_articulo,
        da.marca,
        SUM(fv.cantidades_total) AS bultos
    FROM gold.fact_ventas fv
    LEFT JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
    WHERE fv.id_sucursal = 1
      AND fv.fecha_comprobante BETWEEN :desde AND :hasta
      AND da.generico = :generico
    GROUP BY fv.id_cliente, fv.id_sucursal, fv.id_articulo, da.des_articulo, da.marca
    """
    return loader.execute_query(query, params)


def agregar_fila_total(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega una fila TOTAL al final del DataFrame sumando columnas numéricas."""
    total = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            total[col] = df[col].sum()
        else:
            total[col] = ""
    # Poner "TOTAL" en la primera columna de texto
    for col in df.columns:
        if not pd.api.types.is_numeric_dtype(df[col]):
            total[col] = "TOTAL"
            break
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


def pintar_fila_total(ws, num_rows, num_cols, row_offset=1):
    """Pinta la última fila de datos con fondo amarillo."""
    fila = num_rows + row_offset  # +1 por header
    for c in range(1, num_cols + 1):
        ws.cell(row=fila, column=c).fill = FILL_TOTAL
        ws.cell(row=fila, column=c).font = Font(bold=True)


def generar_comparacion(sucursales: list[str] | None = None) -> str:
    """Genera el reporte de comparación de clientes."""
    loader = DataLoader()

    # 1. Obtener clientes con ventas desde la BD
    df_clientes = get_clientes(loader)
    print(f"Clientes con ventas: {len(df_clientes)}")

    # 2. Traer campo anulado de dim_cliente
    df_anulado = loader.execute_query("""
        SELECT id_cliente, id_sucursal, anulado, fantasia
        FROM gold.dim_cliente
        WHERE id_sucursal = 1
    """)
    df_clientes = df_clientes.merge(df_anulado, on=["id_cliente", "id_sucursal"], how="left")

    # 3. Cruzar con datos del Excel (ruta, preventista, subcanal, etc.)
    df_xlsx = cargar_datos_clientes()
    df_xlsx = df_xlsx.dropna(subset=["id_cliente", "id_sucursal"])
    df_xlsx["id_cliente"] = df_xlsx["id_cliente"].astype(int)
    df_xlsx["id_sucursal"] = df_xlsx["id_sucursal"].astype(int)
    df_clientes = df_clientes.merge(df_xlsx, on=["id_cliente", "id_sucursal"], how="left")
    print(f"Clientes tras cruce con Excel: {len(df_clientes)}")

    # 3. Obtener ventas por período
    df_base = df_clientes.copy()

    for nombre_periodo, (desde, hasta) in PERIODOS.items():
        df_ventas = get_ventas_periodo(loader, desde, hasta)
        df_base = df_base.merge(
            df_ventas.rename(columns={"bultos": nombre_periodo}),
            on=["id_cliente", "id_sucursal"],
            how="left",
        )
        df_base[nombre_periodo] = df_base[nombre_periodo].fillna(0)

    # 4. Obtener bonificaciones por período
    for nombre_periodo, (desde, hasta) in PERIODOS.items():
        df_b = get_bonificaciones_periodo(loader, desde, hasta)
        col_bonif = f"Bonif {nombre_periodo}"
        df_base = df_base.merge(
            df_b.rename(columns={"bultos_bonificados": col_bonif}),
            on=["id_cliente", "id_sucursal"],
            how="left",
        )
        df_base[col_bonif] = df_base[col_bonif].fillna(0)

    # 5. Detalle por artículo (solo lo que se compró, sin producto cartesiano)
    all_art = []
    for nombre_periodo, (desde, hasta) in PERIODOS.items():
        df_a = get_ventas_articulo_periodo(loader, desde, hasta)
        df_a = df_a.rename(columns={"bultos": nombre_periodo})
        all_art.append(df_a)

    # Merge progresivo por cliente+artículo (outer para no perder ningún período)
    df_detalle = all_art[0]
    for df_a in all_art[1:]:
        df_detalle = df_detalle.merge(
            df_a, on=["id_cliente", "id_sucursal", "id_articulo", "des_articulo", "marca"],
            how="outer",
        )

    # Rellenar NaN con 0 en columnas de período
    for periodo in PERIODOS:
        if periodo in df_detalle.columns:
            df_detalle[periodo] = df_detalle[periodo].fillna(0)

    # Agregar datos del cliente
    cols_cliente_detalle = ["id_cliente", "id_sucursal", "anulado", "fantasia", "razon_social",
                            "subcanal", "id_lista_precio", "desc_lista_precio",
                            "id_ruta", "desc_ruta", "preventista"]
    cols_disp_det = [c for c in cols_cliente_detalle if c in df_base.columns]
    df_detalle = df_detalle.merge(
        df_base[cols_disp_det].drop_duplicates(subset=["id_cliente", "id_sucursal"]),
        on=["id_cliente", "id_sucursal"],
        how="left",
    )

    # Ordenar columnas: info cliente + marca + artículo + períodos
    cols_drop_det = ["id_sucursal", "id_subcanal", "id_preventista"]
    cols_info_det = [c for c in cols_disp_det if c not in cols_drop_det]
    df_detalle = df_detalle[cols_info_det + ["marca", "des_articulo"] + list(PERIODOS.keys())]
    df_detalle = df_detalle.sort_values(["id_cliente", "marca", "des_articulo"]).reset_index(drop=True)
    print(f"Detalle artículos: {len(df_detalle)} filas")

    # 6. Apertura por marca: una columna por marca×período
    all_marcas = []
    for nombre_periodo, (desde, hasta) in PERIODOS.items():
        df_m = get_ventas_marca_periodo(loader, desde, hasta)
        df_m["periodo"] = nombre_periodo
        df_m["col"] = df_m["marca"] + " " + nombre_periodo
        all_marcas.append(df_m)
    df_marcas_all = pd.concat(all_marcas, ignore_index=True)

    # Pivot: fila por cliente, columna por "marca periodo"
    df_marcas_pivot = df_marcas_all.pivot_table(
        index=["id_cliente", "id_sucursal"],
        columns="col",
        values="bultos",
        fill_value=0,
    ).reset_index()
    df_marcas_pivot.columns.name = None

    # Obtener marcas únicas ordenadas
    marcas_unicas = sorted(df_marcas_all["marca"].dropna().unique())

    # Merge con datos del cliente
    cols_cliente_info = ["id_cliente", "id_sucursal", "anulado", "fantasia", "razon_social",
                         "subcanal", "id_lista_precio", "desc_lista_precio",
                         "id_ruta", "desc_ruta", "preventista"]
    cols_disponibles = [c for c in cols_cliente_info if c in df_base.columns]
    df_marcas_sheet = df_marcas_pivot.merge(
        df_base[cols_disponibles].drop_duplicates(subset=["id_cliente", "id_sucursal"]),
        on=["id_cliente", "id_sucursal"],
        how="left",
    )

    # Agregar columna total por período (visible) + marcas (colapsadas)
    cols_drop_marca = ["id_sucursal", "id_subcanal", "id_preventista"]
    cols_info_final = [c for c in cols_disponibles if c not in cols_drop_marca]

    # Traer totales por período desde df_base
    for periodo in PERIODOS:
        if periodo in df_base.columns:
            total_map = df_base.set_index("id_cliente")[periodo]
            df_marcas_sheet[periodo] = df_marcas_sheet["id_cliente"].map(total_map).fillna(0)

    # Reordenar: cliente info + [total_periodo, marca1, marca2, ...] por cada período
    cols_marca_ordenadas = []
    for periodo in PERIODOS:
        cols_marca_ordenadas.append(periodo)  # total visible
        for marca in marcas_unicas:
            col = f"{marca} {periodo}"
            if col not in df_marcas_sheet.columns:
                df_marcas_sheet[col] = 0
            cols_marca_ordenadas.append(col)

    df_marcas_sheet = df_marcas_sheet[cols_info_final + cols_marca_ordenadas]

    # Renombrar columnas de marca: "Marca Ene 2025" → "Marca"
    rename_map = {}
    for periodo in PERIODOS:
        for marca in marcas_unicas:
            rename_map[f"{marca} {periodo}"] = marca
    df_marcas_sheet = df_marcas_sheet.rename(columns=rename_map)

    # Guardar posiciones de groups por índice (nombres duplicados no sirven con ColumnGroup)
    headers_marca = list(df_marcas_sheet.columns)
    grupos_marca_idx = []  # lista de (start_col_idx, end_col_idx) base 1
    for periodo in PERIODOS:
        idx_total = headers_marca.index(periodo)
        start_idx = idx_total + 2  # +1 por base 1, +1 para saltar el total
        end_idx = start_idx + len(marcas_unicas) - 1
        grupos_marca_idx.append((start_idx, end_idx))

    print(f"Apertura por marca: {len(df_marcas_sheet)} clientes × {len(marcas_unicas)} marcas")

    # 6. Resúmenes agrupados
    cols_periodo = list(PERIODOS.keys())

    def resumen_por(df, col_grupo):
        agg = df.groupby(col_grupo, dropna=False)[cols_periodo].sum().reset_index()
        agg["Clientes"] = df.groupby(col_grupo, dropna=False)["id_cliente"].nunique().values
        return agg

    df_por_lista = resumen_por(df_base, "desc_lista_precio")
    df_por_ruta = resumen_por(df_base, "desc_ruta")

    # Segregados: resumen por Marca, Subcanal, Lista de precio
    # Por Marca: pivotar df_marcas_all (ya tiene marca, bultos, col="marca periodo")
    df_marca_total = df_marcas_all.groupby(["marca", "periodo"], dropna=False)["bultos"].sum().reset_index()
    df_marca_total = df_marca_total.pivot_table(
        index="marca", columns="periodo", values="bultos", fill_value=0
    ).reset_index()
    df_marca_total.columns.name = None
    # Reordenar columnas según PERIODOS
    df_marca_total = df_marca_total[["marca"] + [p for p in PERIODOS if p in df_marca_total.columns]]

    df_por_subcanal = resumen_por(df_base, "subcanal")
    df_por_lista_seg = resumen_por(df_base, "desc_lista_precio")


    # 8. Sacar columnas internas de la hoja principal
    cols_drop = ["id_sucursal", "id_subcanal", "id_preventista"]
    df_base = df_base.drop(columns=[c for c in cols_drop if c in df_base.columns])

    # 7. Generar Excel
    col_width_8 = ColumnFormat(width=8)
    cols_cliente_fmt = {c: col_width_8 for c in
                        ["id_cliente", "anulado", "fantasia", "razon_social", "subcanal",
                         "id_lista_precio", "desc_lista_precio", "id_ruta", "desc_ruta", "preventista"]}

    bonif_cols = [f"Bonif {p}" for p in PERIODOS]
    grupos_ventas = [ColumnGroup(
        start_col=bonif_cols[0], end_col=bonif_cols[-1], collapsed=True
    )]

    style_ventas = SheetStyle(
        numeric_format="#,##0",
        as_table=True,
        table_style="TableStyleMedium9",
        column_formats=cols_cliente_fmt,
        column_groups=grupos_ventas,
    )

    style_resumen = SheetStyle(
        numeric_format="#,##0",
        as_table=True,
        table_style="TableStyleMedium9",
    )

    # Agregar fila de totales a todos los DataFrames
    df_base = agregar_fila_total(df_base)
    df_detalle = agregar_fila_total(df_detalle)
    df_marcas_sheet = agregar_fila_total(df_marcas_sheet)
    df_por_lista = agregar_fila_total(df_por_lista)
    df_por_ruta = agregar_fila_total(df_por_ruta)

    writer = ExcelWriter("Comparacion Clientes CERVEZAS")
    writer.add_sheet(df_base, sheet_name="Comparacion Clientes", style=style_ventas)
    style_marcas = SheetStyle(
        numeric_format="#,##0",
        as_table=False,
        column_formats=cols_cliente_fmt,
    )

    style_detalle = SheetStyle(
        numeric_format="#,##0",
        as_table=True,
        table_style="TableStyleMedium9",
        column_formats=cols_cliente_fmt,
    )

    writer.add_sheet(df_detalle, sheet_name="Detalle Articulo", style=style_detalle)
    writer.add_sheet(df_marcas_sheet, sheet_name="Por Marca", style=style_marcas)

    # Aplicar groups manuales por índice (nombres de columna duplicados)
    ws_marca = writer.workbook["Por Marca"]
    for start_idx, end_idx in grupos_marca_idx:
        ws_marca.column_dimensions.group(
            get_column_letter(start_idx),
            get_column_letter(end_idx),
            hidden=True,
        )
    ws_marca.sheet_properties.outlinePr.summaryRight = False

    writer.add_sheet(df_por_lista, sheet_name="Por Lista Precio", style=style_resumen)
    writer.add_sheet(df_por_ruta, sheet_name="Por Ruta", style=style_resumen)

    # Hoja Segregados: 3 tablas apiladas (Marca, Subcanal, Lista de Precio)
    ws_seg = writer.workbook.create_sheet(title="Segregados")
    header_fill = PatternFill(start_color="A92C1F", end_color="A92C1F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)
    seg_tables = [
        ("Por Marca", df_marca_total),
        ("Por Subcanal", df_por_subcanal),
        ("Por Lista de Precio", df_por_lista_seg),
    ]
    current_row = 1
    seg_ranges = []  # (start_row_data, num_rows, headers) para pintar después
    for titulo, df_seg in seg_tables:
        # Título
        cell = ws_seg.cell(row=current_row, column=1, value=titulo)
        cell.font = title_font
        current_row += 1
        # Headers
        headers = list(df_seg.columns)
        for c_idx, h in enumerate(headers, 1):
            cell = ws_seg.cell(row=current_row, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        current_row += 1
        data_start = current_row
        # Datos
        for _, row_data in df_seg.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_seg.cell(row=current_row, column=c_idx, value=val)
                if isinstance(val, (int, float)) and val is not None:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)
            current_row += 1
        # Fila de totales con fondo amarillo
        for c_idx, h in enumerate(headers, 1):
            if pd.api.types.is_numeric_dtype(df_seg[h]):
                val = df_seg[h].sum()
            else:
                val = "TOTAL" if c_idx == 1 else ""
            cell = ws_seg.cell(row=current_row, column=c_idx, value=val)
            cell.fill = FILL_TOTAL
            cell.font = Font(bold=True)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")
        seg_ranges.append((data_start, len(df_seg) + 1, headers))  # +1 por fila total
        current_row += 2  # fila total + fila vacía separadora

    # Auto-fit columnas de Segregados
    for col_cells in ws_seg.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_seg.column_dimensions[col_cells[0].column_letter].width = max_len + 2

    # Pintar columnas por par de meses (solo celdas de datos, no headers)
    fill_ene = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # azul claro
    fill_feb = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # verde claro
    fill_mar = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")  # durazno claro
    fill_por_mes = {"Ene": fill_ene, "Feb": fill_feb, "Mar": fill_mar}

    def pintar_columnas_mes(ws, headers, num_rows):
        for col_idx, header in enumerate(headers, 1):
            h = str(header)
            fill = None
            for prefijo, f in fill_por_mes.items():
                if prefijo in h:
                    fill = f
                    break
            if not fill:
                continue
            for row in range(2, num_rows + 2):
                ws.cell(row=row, column=col_idx).fill = fill

    for sheet_name, df in [
        ("Comparacion Clientes", df_base),
        ("Detalle Articulo", df_detalle),
        ("Por Marca", df_marcas_sheet),
        ("Por Lista Precio", df_por_lista),
        ("Por Ruta", df_por_ruta),
    ]:
        ws = writer.workbook[sheet_name]
        pintar_columnas_mes(ws, list(df.columns), len(df))

    # Pintar fila total con amarillo en hojas normales
    for sheet_name, df in [
        ("Comparacion Clientes", df_base),
        ("Detalle Articulo", df_detalle),
        ("Por Marca", df_marcas_sheet),
        ("Por Lista Precio", df_por_lista),
        ("Por Ruta", df_por_ruta),
    ]:
        ws = writer.workbook[sheet_name]
        pintar_fila_total(ws, len(df), len(df.columns))

    # Pintar columnas de Segregados (cada tabla tiene su propio rango)
    for data_start, num_rows, headers in seg_ranges:
        for col_idx, header in enumerate(headers, 1):
            h = str(header)
            fill = None
            for prefijo, f in fill_por_mes.items():
                if prefijo in h:
                    fill = f
                    break
            if not fill:
                continue
            for row in range(data_start, data_start + num_rows):
                ws_seg.cell(row=row, column=col_idx).fill = fill

    ruta = writer.save()

    print(f"Reporte generado: {ruta}")
    return str(ruta)


def main():
    parser = argparse.ArgumentParser(description="Comparación de clientes CERVEZAS")
    parser.add_argument(
        "--sucursales",
        type=str,
        default=None,
        help="Sucursales separadas por coma (ej: 'CASA CENTRAL,SUCURSAL CAFAYATE')",
    )
    args = parser.parse_args()

    sucursales = None
    if args.sucursales:
        sucursales = [s.strip() for s in args.sucursales.split(",")]

    generar_comparacion(sucursales)


if __name__ == "__main__":
    main()
