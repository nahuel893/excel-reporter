#!/usr/bin/env python3
"""
build_dashboard.py — Genera un dashboard HTML standalone (single-file) desde
el BASE control xlsx de acciones-comerciales.

Output: dashboards/acciones-comerciales/dashboard.html
- HTML único, autocontenido.
- SVG charts hand-crafted (cero CDN, cero Chart.js).
- JSON embebido inline en <script type="application/json">.
- Tipografía via Google Fonts (Fraunces / Manrope / JetBrains Mono) con
  fallback a system fonts si no hay red.

Uso:
    python build_dashboard.py                                    # usa defaults
    python build_dashboard.py path/al.xlsx                       # otro xlsx
    python build_dashboard.py path/al.xlsx path/al/output.html   # custom out
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# --- CCU domain ---
CCU_GENERICOS = {
    "CERVEZAS",
    "AGUAS DANONE",
    "VINOS CCU",
    "SIDRAS Y LICORES",
    "PERNOD RICARD",
}

# --- paths ---
REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = (
    REPO_ROOT
    / "data/output/acciones-comerciales/2026-07/BASE control Acciones Comerciales - JULIO 2026.xlsx"
)
DEFAULT_OUT = Path(__file__).resolve().parent / "dashboard.html"
JSON_OUT = Path(__file__).resolve().parent / "data" / "dashboard.json"
# Informe final — segunda fuente, opcional (ver load_informe)
DEFAULT_INFORME = (
    REPO_ROOT
    / "data/backups/acciones-comerciales-2026-07-20/engine-informe"
    / "INFO - ACCIONES BADIE JULIO 2026.xlsm"
)


# ============== helpers ==============
def _f(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s[:4].isdigit() or len(s) < 10:
        return None
    return s[:10]


def _round_money(v: float) -> float:
    return round(v, 2)


# ============== streamed aggregators ==============
def agg_fact_net(ws):
    """Returns (mix_gen, mix_suc, gen_by_suc).
    gen_by_suc[sucursal][generico] = {fact, desc, articulos} — powers the filter."""
    mix_gen: dict[str, dict] = defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "articulos": 0})
    mix_suc: dict[str, dict] = defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "articulos": 0})
    gen_by_suc: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "articulos": 0})
    )
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or r[0] == "TOTAL GENERAL":
            continue
        sucursal, generico = r[0], r[4]
        fact_v, desc_v = _f(r[5]), _f(r[6])
        for bucket in (mix_gen[generico], mix_suc[sucursal], gen_by_suc[sucursal][generico]):
            bucket["fact"] += fact_v; bucket["desc"] += desc_v; bucket["articulos"] += 1
    return mix_gen, mix_suc, gen_by_suc


def agg_art_accion(ws):
    """Returns (acc_desc, acc_by_suc). acc_by_suc[sucursal][accion] = {desc, articulos, ...}"""
    acc_desc: dict[str, dict] = defaultdict(lambda: {"desc": 0.0, "articulos": 0, "mvb": "OTRAS"})
    acc_by_suc: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"desc": 0.0, "articulos": 0, "mvb": "OTRAS"})
    )
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or r[0] == "TOTAL GENERAL" or not r[3]:
            continue
        sucursal = r[0]
        accion = r[3]
        for bucket in (acc_desc[accion], acc_by_suc[sucursal][accion]):
            bucket["desc"] += _f(r[6]); bucket["articulos"] += 1
            if r[5]: bucket["mvb"] = r[5]
            bucket["descripcion"] = r[4]
    return acc_desc, acc_by_suc


def agg_cliente_fecha(ws, cli_desc, cli_by_suc):
    """cli_desc: global top. cli_by_suc[sucursal][(cod, rs)] = {desc, ops}."""
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        cod, rs = r[2], r[3] or ""
        if cod is None or cod == "":
            continue
        if rs == "TOTAL GENERAL" or str(r[1]) == "TOTAL GENERAL":
            continue
        sucursal = r[1]
        k = (cod, rs)
        # global
        b = cli_desc[k]
        b["desc"] += _f(r[9]); b["ops"] += 1; b["sucursal"] = sucursal
        # per-sucursal
        s_b = cli_by_suc[sucursal][k]
        s_b["desc"] += _f(r[9]); s_b["ops"] += 1


def agg_wapi_daily(ws, daily, dow=None, tipo_desc=None, dq=None, daily_by_suc=None):
    """Streamed wapi aggregator. Populates:
      - daily: per-day fact/desc/ops
      - dow: per-weekday desc/ops (Lun..Dom)
      - tipo_desc: per-tipo_descuento fact/desc/ops
      - dq: data-quality counters (rows total / sin_zona / sin_marca / sin_calibre)
    """
    if dow is None: dow = defaultdict(lambda: {"desc": 0.0, "ops": 0})
    if tipo_desc is None: tipo_desc = defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "ops": 0})
    if dq is None: dq = {"rows": 0, "sin_zona": 0, "sin_marca": 0, "sin_calibre": 0}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        fecha = _iso(r[0])
        if not fecha:
            continue
        fact_v = _f(r[26]); desc_v = _f(r[27])
        tipo = (r[28] or "SIN CLASIFICAR") if len(r) > 28 else "SIN CLASIFICAR"
        zona = r[25] if len(r) > 25 else None
        marca = r[8] if len(r) > 8 else None
        calibre = r[9] if len(r) > 9 else None

        d = daily[fecha]
        d["fact"] += fact_v; d["desc"] += desc_v; d["ops"] += 1

        # per-sucursal daily (powers the filter)
        sucursal = r[21] if len(r) > 21 else None
        if daily_by_suc is not None and sucursal:
            sd = daily_by_suc[sucursal][fecha]
            sd["fact"] += fact_v; sd["desc"] += desc_v; sd["ops"] += 1

        # weekday
        try:
            wd = datetime.strptime(fecha, "%Y-%m-%d").weekday()  # 0=Lun
            wd_b = dow[wd]
            wd_b["desc"] += desc_v; wd_b["ops"] += 1
        except ValueError:
            pass

        # tipo descuento
        td = tipo_desc[tipo]
        td["fact"] += fact_v; td["desc"] += desc_v; td["ops"] += 1

        # data quality
        dq["rows"] += 1
        if zona is None or zona == "": dq["sin_zona"] += 1
        if marca is None or marca == "": dq["sin_marca"] += 1
        if calibre is None or calibre == "": dq["sin_calibre"] += 1


# ---- Feriados AR 2026 (subset for Julio) ----
# Source: calendario oficial. Extend as needed.
FERIADOS_AR_2026 = {
    "2026-07-09",  # Día de la Independencia
}


# ============== INFORME loader (segunda fuente, alcance distinto) ==============
# El informe final (INFO - ACCIONES BADIE {MES} {AÑO}.xlsm) contiene dimensiones
# de negocio que la BASE control NO produce: el modelo de TASA (compras vs
# reintegro CCU), el split REVERSA, la atribucion "quien reconoce" y el CANAL.
#
# OJO — alcance distinto al de la BASE:
#   BASE    : corte 21-jul, TODOS los genericos.
#   INFORME : corte 17-jul (cervezas) / 16-jul (aguas), SOLO genericos CCU.
# Por eso los totales no coinciden y se renderizan en un bloque aparte rotulado.

# Coordenadas fijas dentro de la hoja "COMPRAS & DESC" (1-indexed como Excel).
_TASA_BLOQUES = [
    {"generico": "CERVEZAS", "fila_header": 19, "filas": (20, 23), "fila_total": 24},
    {"generico": "AGUAS DANONE", "fila_header": 26, "filas": (27, 30), "fila_total": 31},
]


def load_informe(path: Path) -> dict:
    """Extrae las dimensiones del informe final. Devuelve {} si no se puede leer."""
    try:
        import pandas as pd
    except ImportError:
        print("  WARN: pandas no disponible — se omiten las dimensiones del informe")
        return {}
    if not path.exists():
        print(f"  WARN: informe no encontrado en {path} — se omiten sus dimensiones")
        return {}

    out: dict = {}
    print(f"  Leyendo informe: {path.name}")

    # ---------- 1) Modelo de TASA (hoja COMPRAS & DESC) ----------
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        ws = wb["COMPRAS & DESC"]
        filas = list(ws.iter_rows(min_row=1, max_row=32, values_only=True))

        corte_cervezas = _iso(filas[0][2]) if len(filas[0]) > 2 else None
        corte_aguas = _iso(filas[0][3]) if len(filas[0]) > 3 else None

        tasa: list[dict] = []
        for bloque in _TASA_BLOQUES:
            desde, hasta = bloque["filas"]
            for idx in range(desde - 1, hasta):
                r = filas[idx]
                if not r or not r[0]:
                    continue
                compras = _f(r[2])
                descuentos = _f(r[1])
                tasa_pct = _f(r[3])
                generada = _f(r[5])
                tasa.append({
                    "generico": bloque["generico"],
                    "zona": r[0],
                    "descuentos": _round_money(descuentos),
                    "compras": _round_money(compras),
                    "tasa_pct": round(tasa_pct, 4),
                    "tasa_generada": _round_money(generada),
                    # Diferencia = Tasa Generada − Descuentos  (verificado contra el informe)
                    "diferencia": _round_money(generada - descuentos),
                    # Compra Necesaria = Compras − (Descuentos / Tasa%); negativo = faltante
                    "compra_necesaria": _round_money(compras - (descuentos / tasa_pct)) if tasa_pct else 0,
                    "cubre": generada >= descuentos,
                })
            rt = filas[bloque["fila_total"] - 1]
            gen_t = _f(rt[5])
            desc_t = _f(rt[1])
            tasa.append({
                "generico": bloque["generico"],
                "zona": "TOTAL",
                "descuentos": _round_money(desc_t),
                "compras": _round_money(_f(rt[2])),
                "tasa_pct": None,
                "tasa_generada": _round_money(gen_t),
                "diferencia": _round_money(gen_t - desc_t),
                "compra_necesaria": _round_money(_f(rt[4])),
                "cubre": gen_t >= desc_t,
                "es_total": True,
            })
        out["tasa_model"] = tasa
        out["tasa_cortes"] = {"CERVEZAS": corte_cervezas, "AGUAS DANONE": corte_aguas}
        wb.close()
    except Exception as e:  # noqa: BLE001 — el informe es opcional, no debe romper el build
        print(f"  WARN: no se pudo leer COMPRAS & DESC ({e})")

    # ---------- 2) REVERSA (ART-ACCION, header en fila 1) ----------
    try:
        aa = pd.read_excel(path, sheet_name="ART-ACCION", header=1)
        aa["Suma de Descuento"] = pd.to_numeric(aa["Suma de Descuento"], errors="coerce").fillna(0)
        aa = aa[aa["Acción"].notna()]
        es_rev = aa["Es Reversa?"].astype(str).str.upper().str.strip().eq("REVERSA")
        out["reversa_split"] = [
            {
                "tipo": "SI reversa",
                "detalle": "CCU lo reintegra",
                "descuento": _round_money(float(aa.loc[es_rev, "Suma de Descuento"].sum())),
                "acciones": int(aa.loc[es_rev, "Acción"].nunique()),
            },
            {
                "tipo": "NO reversa",
                "detalle": "lo absorbe BADIE",
                "descuento": _round_money(float(aa.loc[~es_rev, "Suma de Descuento"].sum())),
                "acciones": int(aa.loc[~es_rev, "Acción"].nunique()),
            },
        ]
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: no se pudo leer ART-ACCION/reversa ({e})")

    # ---------- 3) CANAL (CLIENTE-FECHA, header en fila 2) ----------
    try:
        cf = pd.read_excel(path, sheet_name="CLIENTE-FECHA", header=2)
        cf["Suma de Descuento"] = pd.to_numeric(cf["Suma de Descuento"], errors="coerce").fillna(0)
        cf = cf[cf["Razón Social"].notna() & cf["Razón Social"].astype(str).ne("(blank)")]
        cf["LISTA PRECIO"] = cf["LISTA PRECIO"].fillna("SIN CANAL (#N/D)")
        cg = (
            cf.groupby("LISTA PRECIO")["Suma de Descuento"]
            .agg(descuento="sum", lineas="size")
            .sort_values("descuento", ascending=False)
        )
        total_canal = float(cg["descuento"].sum())
        out["canal"] = [
            {
                "canal": str(k),
                "descuento": _round_money(float(v["descuento"])),
                "lineas": int(v["lineas"]),
                "share": round(float(v["descuento"]) / total_canal, 4) if total_canal else 0,
                "sin_canal": "#N/D" in str(k),
            }
            for k, v in cg.iterrows()
        ]
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: no se pudo leer CLIENTE-FECHA/canal ({e})")

    # ---------- 4) QUIEN RECONOCE (atribucion de costo) ----------
    try:
        det = pd.read_excel(path, sheet_name="Detalle_Clientes_Quien_Reco (2)")
        det.columns = [str(c).strip() for c in det.columns]
        det["TOTAL PLATA"] = pd.to_numeric(det["TOTAL PLATA"], errors="coerce").fillna(0)
        qr = (
            det.groupby("Quien reconoce", dropna=False)["TOTAL PLATA"]
            .agg(total="sum", lineas="size")
            .sort_values("total", ascending=False)
        )
        qr = qr[qr["total"] > 0]
        tot_qr = float(qr["total"].sum())
        out["quien_reconoce"] = [
            {
                "quien": str(k),
                "total": _round_money(float(v["total"])),
                "lineas": int(v["lineas"]),
                "share": round(float(v["total"]) / tot_qr, 4) if tot_qr else 0,
                "es_area": "REC" in str(k).upper(),
            }
            for k, v in qr.iterrows()
        ]
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: no se pudo leer Detalle_Clientes_Quien_Reco ({e})")

    # ---------- 5) Maestros: zona y supervisor por sucursal ----------
    try:
        su = pd.read_excel(path, sheet_name="sucu", header=None)
        out["zona_map"] = {
            str(r[1]).strip(): str(r[3]).strip()
            for r in su.itertuples(index=False)
            if len(r) > 3 and r[1] is not None and str(r[1]) != "nan"
        }
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: no se pudo leer sucu ({e})")
    try:
        h1 = pd.read_excel(path, sheet_name="Hoja1", header=None)
        out["supervisor_map"] = {
            str(r[0]).strip(): str(r[1]).strip()
            for r in h1.itertuples(index=False)
            if len(r) > 1 and r[0] is not None and str(r[0]) != "nan"
        }
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: no se pudo leer Hoja1 ({e})")

    return out


def _mix_gen_block(gen_map: dict) -> dict:
    """Shared shaper: {generico: {...}} sorted by facturación desc."""
    return {
        g: {
            "facturacion_neta": _round_money(v["fact"]),
            "descuentos": _round_money(v["desc"]),
            "ratio_d_f": round(v["desc"] / v["fact"], 4) if v["fact"] else 0,
            "articulos": v["articulos"],
            "es_ccu": g in CCU_GENERICOS,
            "es_outlier": v["fact"] > 0 and (v["desc"] / v["fact"]) > 0.20,
        }
        for g, v in sorted(gen_map.items(), key=lambda kv: -kv[1]["fact"])
    }


def _ccu_split_block(mix_generico: dict) -> dict:
    """Shared shaper: CCU vs No-CCU rollup from a mix_generico block."""
    ccu_f = sum(v["facturacion_neta"] for v in mix_generico.values() if v["es_ccu"])
    no_f = sum(v["facturacion_neta"] for v in mix_generico.values() if not v["es_ccu"])
    ccu_d = sum(v["descuentos"] for v in mix_generico.values() if v["es_ccu"])
    no_d = sum(v["descuentos"] for v in mix_generico.values() if not v["es_ccu"])
    return {
        "ccu": {
            "facturacion_neta": _round_money(ccu_f),
            "descuentos": _round_money(ccu_d),
            "ratio_d_f": round(ccu_d / ccu_f, 4) if ccu_f else 0,
        },
        "no_ccu": {
            "facturacion_neta": _round_money(no_f),
            "descuentos": _round_money(no_d),
            "ratio_d_f": round(no_d / no_f, 4) if no_f else 0,
        },
    }


def build_payload(wb, informe: dict | None = None) -> dict:
    mix_gen, mix_suc, gen_by_suc = agg_fact_net(wb["FACT_NET"])
    cli_desc: dict[tuple, dict] = defaultdict(lambda: {"desc": 0.0, "ops": 0, "sucursal": ""})
    cli_by_suc: dict[str, dict[tuple, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"desc": 0.0, "ops": 0})
    )
    agg_cliente_fecha(wb["CLIENTE-FECHA"], cli_desc, cli_by_suc)
    acc_desc, acc_by_suc = agg_art_accion(wb["ART-ACCION"])
    daily: dict[str, dict] = defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "ops": 0})
    daily_by_suc: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "ops": 0})
    )
    dow: dict[int, dict] = defaultdict(lambda: {"desc": 0.0, "ops": 0})
    tipo_desc: dict[str, dict] = defaultdict(lambda: {"fact": 0.0, "desc": 0.0, "ops": 0})
    dq: dict = {"rows": 0, "sin_zona": 0, "sin_marca": 0, "sin_calibre": 0}
    agg_wapi_daily(wb["wapi"], daily, dow=dow, tipo_desc=tipo_desc, dq=dq,
                   daily_by_suc=daily_by_suc)

    fact_total = 0.0
    desc_total = 0.0
    for r in wb["FACT_NET"].iter_rows(min_row=2, values_only=True):
        if r and r[0] == "TOTAL GENERAL":
            fact_total += _f(r[5]); desc_total += _f(r[6])

    global_ratio = (desc_total / fact_total) if fact_total else 0

    out: dict = {}
    out["hero"] = {
        "facturacion_neta": _round_money(fact_total),
        "descuentos": _round_money(desc_total),
        "ratio_descuento_facturacion": round(global_ratio, 4),
        "acciones_distintas": len(acc_desc),
        "clientes_distintos": len(cli_desc),
        "articulos_distintos": sum(v["articulos"] for v in mix_gen.values()),
    }
    out["mix_generico"] = {
        g: {
            "facturacion_neta": _round_money(v["fact"]),
            "descuentos": _round_money(v["desc"]),
            "ratio_d_f": round(v["desc"] / v["fact"], 4) if v["fact"] else 0,
            "articulos": v["articulos"],
            "es_ccu": g in CCU_GENERICOS,
            "es_outlier": v["fact"] > 0 and (v["desc"] / v["fact"]) > 0.20,
        }
        for g, v in sorted(mix_gen.items(), key=lambda kv: -kv[1]["fact"])
    }
    out["mix_sucursal"] = {
        s: {
            "facturacion_neta": _round_money(v["fact"]),
            "descuentos": _round_money(v["desc"]),
            "ratio_d_f": round(v["desc"] / v["fact"], 4) if v["fact"] else 0,
            "articulos": v["articulos"],
        }
        for s, v in sorted(mix_suc.items(), key=lambda kv: -kv[1]["fact"])
    }
    # ---- v0.2: sucursal efficiency (vs global ratio) ----
    out["sucursal_efficiency"] = [
        {
            "sucursal": s,
            "facturacion_neta": _round_money(v["fact"]),
            "descuentos": _round_money(v["desc"]),
            "ratio_d_f": round(v["desc"] / v["fact"], 4) if v["fact"] else 0,
            "vs_global_pp": round(((v["desc"] / v["fact"]) - global_ratio) * 100, 2) if v["fact"] else 0,
        }
        for s, v in sorted(mix_suc.items(), key=lambda kv: -kv[1]["fact"])
    ]

    # ---- v0.2: mega-acciones (single-art-concentradas) ----
    mega = []
    for a, v in acc_desc.items():
        if v["articulos"] <= 2 and v["desc"] > 5_000_000:
            mega.append({
                "accion": a,
                "descripcion": v.get("descripcion", ""),
                "mvb": v["mvb"],
                "descuento": _round_money(v["desc"]),
                "articulos": v["articulos"],
                "desc_per_art": _round_money(v["desc"] / v["articulos"]),
            })
    out["mega_acciones"] = sorted(mega, key=lambda x: -x["desc_per_art"])[:8]

    out["top_acciones"] = [
        {
            "accion": a,
            "descripcion": v.get("descripcion", ""),
            "mvb": v["mvb"],
            "descuento": _round_money(v["desc"]),
            "articulos": v["articulos"],
            "is_mega": (v["articulos"] <= 2 and v["desc"] > 5_000_000),
        }
        for a, v in sorted(acc_desc.items(), key=lambda kv: -kv[1]["desc"])[:20]
    ]

    out["top_clientes"] = [
        {
            "cod_cliente": k[0],
            "razon_social": k[1],
            "sucursal": v["sucursal"],
            "descuento": _round_money(v["desc"]),
            "operaciones": v["ops"],
        }
        for k, v in sorted(cli_desc.items(), key=lambda kv: -kv[1]["desc"])[:25]
    ]

    # ---- v0.2: per-sucursal aggregates (full set — the filter swaps EVERY section) ----
    out["by_sucursal"] = {}
    for suc in sorted(mix_suc.keys()):
        suc_mix_gen = _mix_gen_block(gen_by_suc.get(suc, {}))
        suc_fact = mix_suc[suc]["fact"]
        suc_desc = mix_suc[suc]["desc"]
        suc_daily_map = daily_by_suc.get(suc, {})
        out["by_sucursal"][suc] = {
            "hero": {
                "facturacion_neta": _round_money(suc_fact),
                "descuentos": _round_money(suc_desc),
                "ratio_descuento_facturacion": round(suc_desc / suc_fact, 4) if suc_fact else 0,
                "acciones_distintas": len(acc_by_suc.get(suc, {})),
                "clientes_distintos": len(cli_by_suc.get(suc, {})),
                "articulos_distintos": mix_suc[suc]["articulos"],
            },
            "mix_generico": suc_mix_gen,
            "ccu_split": _ccu_split_block(suc_mix_gen),
            "daily": [
                {
                    "fecha": d,
                    "facturacion_neta": _round_money(v["fact"]),
                    "descuentos": _round_money(v["desc"]),
                    "operaciones": v["ops"],
                    "es_feriado": d in FERIADOS_AR_2026,
                }
                for d, v in sorted(suc_daily_map.items())
            ],
            "top_acciones": [
                {
                    "accion": a,
                    "descripcion": v.get("descripcion", ""),
                    "mvb": v["mvb"],
                    "descuento": _round_money(v["desc"]),
                    "articulos": v["articulos"],
                    "is_mega": (v["articulos"] <= 2 and v["desc"] > 5_000_000),
                }
                for a, v in sorted(acc_by_suc.get(suc, {}).items(), key=lambda kv: -kv[1]["desc"])[:15]
            ],
            "top_clientes": [
                {
                    "cod_cliente": k[0],
                    "razon_social": k[1],
                    "sucursal": suc,
                    "descuento": _round_money(v["desc"]),
                    "operaciones": v["ops"],
                }
                for k, v in sorted(cli_by_suc.get(suc, {}).items(), key=lambda kv: -kv[1]["desc"])[:20]
            ],
        }
    out["sucursal_options"] = sorted(mix_suc.keys())
    out["daily"] = [
        {
            "fecha": d,
            "facturacion_neta": _round_money(v["fact"]),
            "descuentos": _round_money(v["desc"]),
            "operaciones": v["ops"],
            "es_feriado": d in FERIADOS_AR_2026,
        }
        for d, v in sorted(daily.items())
    ]
    fechas = sorted(daily.keys())
    out["meta"] = {
        "periodo_desde": fechas[0] if fechas else None,
        "periodo_hasta": fechas[-1] if fechas else None,
        "generado_en": datetime.now().astimezone().isoformat(),
        "fuente": "data/output/acciones-comerciales/2026-07/BASE control Acciones Comerciales - JULIO 2026.xlsx",
        "dias_en_periodo": len(fechas),
        "feriados_en_periodo": [d for d in fechas if d in FERIADOS_AR_2026],
    }

    # ---- v0.2: day of week (descuento agregado por día de la semana) ----
    DOW_NAMES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    total_desc = sum(v["desc"] for v in dow.values())
    out["day_of_week"] = []
    for i in range(7):
        v = dow.get(i) or {"desc": 0.0, "ops": 0}
        out["day_of_week"].append({
            "wd": i,
            "label": DOW_NAMES[i],
            "es_fin_de_semana": i >= 5,
            "descuento": _round_money(v["desc"]),
            "ops": v["ops"],
            "share": round(v["desc"] / total_desc, 4) if total_desc else 0,
        })

    # ---- v0.2: tipo descuento split ----
    out["tipo_descuento_split"] = [
        {
            "tipo": k,
            "fact": _round_money(v["fact"]),
            "descuento": _round_money(v["desc"]),
            "ops": v["ops"],
            "ratio_d_f": round(v["desc"] / v["fact"], 4) if v["fact"] else 0,
        }
        for k, v in sorted(tipo_desc.items(), key=lambda kv: -kv[1]["desc"])
    ]

    # ---- v0.2: data quality ----
    rows = dq["rows"]
    out["data_quality"] = {
        "rows_total": rows,
        "sin_zona_pct": round(dq["sin_zona"] / rows * 100, 1) if rows else 0,
        "sin_marca_pct": round(dq["sin_marca"] / rows * 100, 1) if rows else 0,
        "sin_calibre_pct": round(dq["sin_calibre"] / rows * 100, 1) if rows else 0,
    }

    ccu_fact = sum(v["facturacion_neta"] for v in out["mix_generico"].values() if v["es_ccu"])
    no_ccu_fact = sum(v["facturacion_neta"] for v in out["mix_generico"].values() if not v["es_ccu"])
    ccu_desc = sum(v["descuentos"] for v in out["mix_generico"].values() if v["es_ccu"])
    no_ccu_desc = sum(v["descuentos"] for v in out["mix_generico"].values() if not v["es_ccu"])
    out["ccu_split"] = {
        "ccu": {
            "facturacion_neta": _round_money(ccu_fact),
            "descuentos": _round_money(ccu_desc),
            "ratio_d_f": round(ccu_desc / ccu_fact, 4) if ccu_fact else 0,
        },
        "no_ccu": {
            "facturacion_neta": _round_money(no_ccu_fact),
            "descuentos": _round_money(no_ccu_desc),
            "ratio_d_f": round(no_ccu_desc / no_ccu_fact, 4) if no_ccu_fact else 0,
        },
    }

    # ---- dimensiones del INFORME (segunda fuente, alcance distinto) ----
    if informe:
        out["informe"] = informe
        # rollup de tasa para el hero del bloque
        tot = [t for t in informe.get("tasa_model", []) if t.get("es_total")]
        if tot:
            out["informe"]["tasa_rollup"] = {
                "descuentos": _round_money(sum(t["descuentos"] for t in tot)),
                "compras": _round_money(sum(t["compras"] for t in tot)),
                "tasa_generada": _round_money(sum(t["tasa_generada"] for t in tot)),
                "diferencia": _round_money(sum(t["diferencia"] for t in tot)),
            }
    return out


# ============== HTML template ==============
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Acciones Comerciales — Dashboard · Julio 2026</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght,SOFT,WONK@9..144,300..900,0..100,0..1&family=Manrope:wght@300;400;500;600;700&family=JetBrains+Mono:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  :root {
    --paper: #f4efe4;
    --paper-deep: #ece4d2;
    --ink: #1a1612;
    --ink-soft: #4a4035;
    --ink-mute: #8a7d6a;
    --rule: #1a1612;
    --rule-soft: rgba(26, 22, 18, 0.18);
    --badie: #b8351c;
    --badie-deep: #8a2613;
    --teal: #2c4a52;
    --teal-soft: #4a6c75;
    --warn: #c98717;
    --ok: #4a6c3a;
    --display: 'Fraunces', 'Hoefler Text', 'Iowan Old Style', Baskerville, Georgia, serif;
    --body: 'Manrope', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    --mono: 'JetBrains Mono', ui-monospace, 'SF Mono', 'Fira Code', Consolas, monospace;
  }
  *, *::before, *::after { box-sizing: border-box; }
  html { font-size: 15px; }
  body {
    margin: 0;
    padding: 0;
    background: var(--paper);
    color: var(--ink);
    font-family: var(--body);
    font-weight: 400;
    line-height: 1.45;
    -webkit-font-smoothing: antialiased;
    background-image:
      radial-gradient(ellipse at 12% 8%, rgba(184, 53, 28, 0.04), transparent 40%),
      radial-gradient(ellipse at 88% 92%, rgba(44, 74, 82, 0.05), transparent 45%),
      url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='220' height='220'><filter id='n'><feTurbulence baseFrequency='0.85' numOctaves='2' stitchTiles='stitch'/><feColorMatrix values='0 0 0 0 0.10 0 0 0 0 0.08 0 0 0 0 0.05 0 0 0 0.06 0'/></filter><rect width='100%' height='100%' filter='url(%23n)'/></svg>");
  }
  main {
    max-width: 1480px;
    margin: 0 auto;
    padding: 56px 48px 96px;
  }
  /* masthead */
  .masthead {
    border-top: 3px double var(--ink);
    border-bottom: 1px solid var(--ink);
    padding: 18px 0 16px;
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    gap: 24px;
    flex-wrap: wrap;
  }
  .masthead-title {
    font-family: var(--display);
    font-variation-settings: "opsz" 144, "SOFT" 30, "WONK" 0;
    font-weight: 400;
    font-size: clamp(34px, 5vw, 56px);
    line-height: 0.95;
    letter-spacing: -0.02em;
    color: var(--ink);
  }
  .masthead-title em {
    font-style: italic;
    font-variation-settings: "opsz" 144, "SOFT" 100, "WONK" 1;
    color: var(--badie);
  }
  .masthead-meta {
    font-family: var(--mono);
    font-size: 11px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: var(--ink-soft);
    text-align: right;
  }
  .masthead-meta b {
    color: var(--ink);
    font-weight: 500;
  }
  .masthead-sub {
    font-family: var(--display);
    font-style: italic;
    font-size: 14px;
    color: var(--ink-soft);
    margin-top: 4px;
  }
  .masthead-sub #scope-tag {
    font-family: var(--mono);
    font-style: normal;
    font-size: 11px;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--paper);
    background: var(--badie);
    padding: 2px 8px;
  }
  /* section heading */
  .sec {
    margin-top: 56px;
    border-top: 1px solid var(--ink);
    padding-top: 18px;
  }
  .sec-head {
    display: flex;
    align-items: baseline;
    justify-content: space-between;
    gap: 18px;
    margin-bottom: 22px;
    border-bottom: 1px solid var(--rule-soft);
    padding-bottom: 10px;
  }
  .sec-title {
    font-family: var(--display);
    font-variation-settings: "opsz" 144, "SOFT" 50;
    font-weight: 500;
    font-size: 22px;
    letter-spacing: -0.01em;
  }
  .sec-title em {
    font-style: italic;
    color: var(--badie);
    font-weight: 400;
  }
  .sec-kicker {
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--ink-mute);
  }
  /* --- bloque informe (segunda fuente) --- */
  .scope-warn {
    border-left: 3px solid var(--warn);
    background: rgba(201, 135, 23, 0.07);
    padding: 12px 16px;
    font-size: 12.5px;
    line-height: 1.5;
    color: var(--ink-soft);
    margin-bottom: 24px;
  }
  .scope-warn b { color: var(--ink); }
  .scope-warn i { font-style: italic; }
  .pnl-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    border: 1px solid var(--ink);
  }
  .pnl-cell {
    padding: 20px 22px;
    border-right: 1px solid var(--rule-soft);
  }
  .pnl-cell:last-child { border-right: none; }
  .pnl-cell.danger { background: rgba(184, 53, 28, 0.06); }
  .pnl-num {
    font-family: var(--display);
    font-variation-settings: "opsz" 144, "SOFT" 0;
    font-weight: 400;
    font-size: clamp(26px, 2.9vw, 38px);
    line-height: 1.05;
    letter-spacing: -0.02em;
    font-variant-numeric: tabular-nums;
    margin-top: 6px;
  }
  .pnl-cell.danger .pnl-num { color: var(--badie); }
  /* tasa rows */
  .tasa-row {
    display: grid;
    grid-template-columns: 132px 1fr 96px;
    align-items: center;
    gap: 12px;
    margin: 9px 0;
    font-size: 12.5px;
  }
  .tasa-row.total {
    border-top: 1px solid var(--ink);
    margin-top: 14px;
    padding-top: 12px;
    font-weight: 600;
  }
  .tasa-row .zona { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .tasa-row .zona small {
    font-family: var(--mono); font-size: 10px; color: var(--ink-mute); display: block;
  }
  .tasa-bar {
    position: relative; height: 22px; background: rgba(26,22,18,0.05);
  }
  .tasa-bar .need {
    position: absolute; left: 0; top: 0; height: 100%;
    background: repeating-linear-gradient(45deg, rgba(184,53,28,0.16) 0 5px, rgba(184,53,28,0.05) 5px 10px);
  }
  .tasa-bar .got { position: absolute; left: 0; top: 0; height: 100%; background: var(--ok); }
  .tasa-bar .got.short { background: var(--badie); }
  .tasa-bar .marker {
    position: absolute; top: -3px; width: 2px; height: 28px; background: var(--ink);
  }
  .tasa-row .delta {
    font-family: var(--mono); font-size: 11.5px; text-align: right;
    font-variant-numeric: tabular-nums;
  }
  .tasa-row .delta.neg { color: var(--badie); }
  .tasa-row .delta.pos { color: var(--ok); }
  .gen-head {
    font-family: var(--mono); font-size: 10.5px; letter-spacing: 0.16em;
    text-transform: uppercase; color: var(--ink-mute);
    margin: 18px 0 6px; padding-bottom: 5px; border-bottom: 1px solid var(--rule-soft);
  }
  .gen-head:first-child { margin-top: 0; }
  /* split bar (reversa) */
  .split-bar { display: flex; height: 40px; border: 1px solid var(--ink); margin-bottom: 14px; }
  .split-seg { display: flex; align-items: center; justify-content: center;
    font-family: var(--mono); font-size: 11px; color: var(--paper); overflow: hidden; }
  .split-seg.badie { background: var(--badie); }
  .split-seg.ok { background: var(--teal); }
  /* filter bar */
  .filter-bar {
    margin-top: 22px;
    padding: 14px 18px;
    background: rgba(255, 252, 244, 0.7);
    border: 1px solid var(--ink);
    display: flex;
    align-items: center;
    gap: 18px;
    flex-wrap: wrap;
  }
  .filter-bar-label {
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--ink-soft);
    border-right: 1px solid var(--rule-soft);
    padding-right: 18px;
  }
  .filter-chips {
    display: flex;
    gap: 6px;
    flex-wrap: wrap;
    flex: 1;
  }
  .chip {
    font-family: var(--body);
    font-size: 12px;
    font-weight: 500;
    padding: 6px 12px;
    background: transparent;
    color: var(--ink-soft);
    border: 1px solid var(--rule-soft);
    cursor: pointer;
    letter-spacing: 0.02em;
    transition: all 0.12s ease;
  }
  .chip:hover {
    border-color: var(--ink);
    color: var(--ink);
  }
  .chip.active {
    background: var(--ink);
    color: var(--paper);
    border-color: var(--ink);
  }
  .filter-hint {
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.06em;
    color: var(--ink-mute);
  }
  /* hero KPIs */
  .hero-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0;
    margin-top: 32px;
    border-top: 1px solid var(--ink);
    border-bottom: 1px solid var(--ink);
  }
  .hero-cell {
    padding: 22px 22px 24px;
    border-right: 1px solid var(--rule-soft);
    position: relative;
  }
  .hero-cell:last-child { border-right: none; }
  .hero-cell::before {
    content: "";
    position: absolute;
    left: 0; top: -1px;
    width: 28px; height: 3px;
    background: var(--badie);
  }
  .hero-cell.cool::before { background: var(--teal); }
  .hero-cell.neutral::before { background: var(--ink); }
  .hero-kicker {
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--ink-mute);
    margin-bottom: 8px;
  }
  .hero-num {
    font-family: var(--display);
    font-variation-settings: "opsz" 144, "SOFT" 0, "WONK" 0;
    font-weight: 400;
    font-size: clamp(34px, 3.6vw, 48px);
    line-height: 1;
    letter-spacing: -0.025em;
    color: var(--ink);
    font-variant-numeric: tabular-nums;
    display: flex;
    align-items: baseline;
    gap: 6px;
  }
  .hero-num .unit {
    font-family: var(--body);
    font-size: 14px;
    font-weight: 500;
    color: var(--ink-soft);
    letter-spacing: 0.04em;
  }
  .hero-sub {
    font-family: var(--mono);
    font-size: 11px;
    color: var(--ink-soft);
    margin-top: 8px;
    letter-spacing: 0.02em;
  }
  /* CCU split */
  .ccu-split {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 0;
    border: 1px solid var(--ink);
  }
  .ccu-cell {
    padding: 28px 32px;
    position: relative;
  }
  .ccu-cell:first-child { border-right: 1px solid var(--ink); }
  .ccu-cell .label {
    font-family: var(--mono);
    font-size: 11px;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--ink-soft);
  }
  .ccu-cell .num {
    font-family: var(--display);
    font-variation-settings: "opsz" 144, "SOFT" 0;
    font-weight: 400;
    font-size: clamp(38px, 4.5vw, 60px);
    line-height: 1;
    letter-spacing: -0.025em;
    margin-top: 12px;
    color: var(--ink);
    font-variant-numeric: tabular-nums;
  }
  .ccu-cell .row2 {
    margin-top: 14px;
    display: flex;
    gap: 24px;
    align-items: baseline;
    flex-wrap: wrap;
  }
  .ccu-cell .row2 .item {
    font-family: var(--mono);
    font-size: 12px;
    color: var(--ink-soft);
  }
  .ccu-cell .row2 .item b {
    color: var(--ink);
    font-weight: 500;
    font-size: 14px;
    font-variant-numeric: tabular-nums;
  }
  .ccu-cell .badge {
    display: inline-block;
    padding: 3px 8px;
    background: var(--badie);
    color: var(--paper);
    font-family: var(--mono);
    font-size: 9.5px;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    margin-bottom: 8px;
  }
  .ccu-cell.cool .badge { background: var(--teal); }
  /* two-col grid for charts */
  .row-2 {
    display: grid;
    grid-template-columns: 2fr 1fr;
    gap: 32px;
    align-items: start;
  }
  .row-2-eq {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 32px;
    align-items: start;
  }
  /* chart card */
  .chart-card {
    background: rgba(255, 252, 244, 0.55);
    border: 1px solid var(--rule-soft);
    padding: 18px 20px 14px;
  }
  .chart-card .chart-head {
    display: flex;
    align-items: baseline;
    justify-content: space-between;
    margin-bottom: 14px;
  }
  .chart-card .chart-title {
    font-family: var(--display);
    font-style: italic;
    font-weight: 500;
    font-size: 16px;
    color: var(--ink);
  }
  .chart-card .chart-legend {
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--ink-mute);
    display: flex;
    gap: 14px;
  }
  .chart-card .swatch {
    display: inline-block;
    width: 9px;
    height: 9px;
    margin-right: 5px;
    vertical-align: middle;
    border-radius: 1px;
  }
  svg { display: block; max-width: 100%; }
  /* tables */
  table.tab {
    width: 100%;
    border-collapse: collapse;
    font-family: var(--body);
    font-size: 13px;
  }
  table.tab thead th {
    font-family: var(--mono);
    font-size: 10.5px;
    font-weight: 500;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    text-align: left;
    color: var(--ink-mute);
    padding: 6px 10px;
    border-bottom: 1.5px solid var(--ink);
  }
  table.tab thead th.num { text-align: right; }
  table.tab tbody td {
    padding: 9px 10px;
    border-bottom: 1px solid var(--rule-soft);
    vertical-align: top;
  }
  table.tab tbody tr:last-child td { border-bottom: none; }
  table.tab tbody td.num {
    text-align: right;
    font-family: var(--mono);
    font-variant-numeric: tabular-nums;
  }
  table.tab tbody td .pct {
    font-family: var(--mono);
    font-size: 11px;
    color: var(--ink-mute);
    margin-left: 4px;
  }
  table.tab tbody td.code {
    font-family: var(--mono);
    font-size: 11px;
    color: var(--ink-mute);
  }
  table.tab tbody td.badie { color: var(--badie); font-weight: 500; }
  /* bar */
  .bar-row {
    display: grid;
    grid-template-columns: 110px 1fr 88px;
    align-items: center;
    gap: 12px;
    margin: 6px 0;
    font-size: 12.5px;
  }
  .bar-row .label {
    font-family: var(--body);
    color: var(--ink);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .bar-row .label small {
    font-family: var(--mono);
    font-size: 10px;
    color: var(--ink-mute);
    margin-left: 4px;
  }
  .bar-row .bar-track {
    height: 18px;
    background: rgba(26, 22, 18, 0.05);
    position: relative;
  }
  .bar-row .bar-fill {
    height: 100%;
    background: var(--ink);
  }
  .bar-row .bar-fill.ccu { background: var(--teal); }
  .bar-row .bar-fill.badie { background: var(--badie); }
  .bar-row .val {
    font-family: var(--mono);
    font-size: 11.5px;
    color: var(--ink-soft);
    text-align: right;
    font-variant-numeric: tabular-nums;
  }
  /* dot indicator */
  .dot {
    display: inline-block;
    width: 6px; height: 6px;
    border-radius: 50%;
    margin-right: 6px;
    vertical-align: middle;
  }
  .dot.ccu { background: var(--teal); }
  .bar-row .label small {
    font-family: var(--mono); font-size: 9.5px; color: var(--badie);
    letter-spacing: 0.06em; margin-left: 4px;
  }
  .dot.badie { background: var(--badie); }
  .dot.ok { background: var(--ok); }
  .dot.warn { background: var(--warn); }
  /* footnote */
  .footnote {
    margin-top: 64px;
    border-top: 3px double var(--ink);
    padding-top: 14px;
    font-family: var(--mono);
    font-size: 10.5px;
    letter-spacing: 0.04em;
    color: var(--ink-mute);
    display: flex;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 16px;
  }
  .footnote b { color: var(--ink-soft); font-weight: 500; }
  /* reveal animation */
  @keyframes reveal {
    from { opacity: 0; transform: translateY(6px); }
    to   { opacity: 1; transform: translateY(0); }
  }
  .reveal { animation: reveal 0.7s cubic-bezier(0.2, 0.7, 0.2, 1) both; }
  .reveal.d1 { animation-delay: 0.06s; }
  .reveal.d2 { animation-delay: 0.12s; }
  .reveal.d3 { animation-delay: 0.18s; }
  .reveal.d4 { animation-delay: 0.24s; }
  .reveal.d5 { animation-delay: 0.30s; }
  /* responsive */
  @media (max-width: 1100px) {
    .hero-grid { grid-template-columns: repeat(2, 1fr); }
    .hero-cell:nth-child(2) { border-right: none; }
    .hero-cell:nth-child(1), .hero-cell:nth-child(2) { border-bottom: 1px solid var(--rule-soft); }
    .row-2, .row-2-eq { grid-template-columns: 1fr; }
    .ccu-split { grid-template-columns: 1fr; }
    .ccu-cell:first-child { border-right: none; border-bottom: 1px solid var(--ink); }
  }
  @media (max-width: 640px) {
    main { padding: 32px 18px 56px; }
    .hero-grid { grid-template-columns: 1fr; }
    .hero-cell { border-right: none; border-bottom: 1px solid var(--rule-soft); }
    .hero-cell:last-child { border-bottom: none; }
  }
</style>
</head>
<body>
<main>

  <header class="masthead reveal">
    <div>
      <div class="masthead-title">Acciones <em>Comerciales</em></div>
      <div class="masthead-sub">Tablero ejecutivo · Julio 2026 · BADIE S.A. · <span id="scope-tag">Toda la red</span></div>
    </div>
    <div class="masthead-meta">
      <div><b>Periodo</b> __PERIODO_DESDE__ → __PERIODO_HASTA__</div>
      <div><b>__DIAS__ días</b> en informe</div>
      <div><b>Generado</b> __GENERADO_EN__</div>
    </div>
  </header>

  <section class="filter-bar reveal">
    <div class="filter-bar-label">Filtro</div>
    <div class="filter-chips" id="filter-sucursal" role="tablist">
      <button class="chip active" data-sucursal="__ALL__" type="button">Toda la red</button>
      __FILTER_CHIPS__
    </div>
    <div class="filter-hint" id="filter-hint">Vista global — 14 sucursales</div>
  </section>

  <section class="hero-grid">
    <div class="hero-cell reveal d1">
      <div class="hero-kicker">Facturación Neta</div>
      <div class="hero-num"><span id="kpi-fact">—</span><span class="unit">ARS</span></div>
      <div class="hero-sub" id="kpi-fact-sub">—</div>
    </div>
    <div class="hero-cell badie reveal d2">
      <div class="hero-kicker">Descuentos otorgados</div>
      <div class="hero-num"><span id="kpi-desc">—</span><span class="unit">ARS</span></div>
      <div class="hero-sub" id="kpi-desc-sub">—</div>
    </div>
    <div class="hero-cell cool reveal d3">
      <div class="hero-kicker">Porcentaje de Descuento</div>
      <div class="hero-num"><span id="kpi-ratio">—</span><span class="unit">%</span></div>
      <div class="hero-sub" id="kpi-ratio-sub">—</div>
    </div>
    <div class="hero-cell neutral reveal d4">
      <div class="hero-kicker">Universo</div>
      <div class="hero-num"><span id="kpi-uni">—</span></div>
      <div class="hero-sub" id="kpi-uni-sub">—</div>
    </div>
  </section>

  <section class="sec">
    <div class="sec-head reveal">
      <div class="sec-title">Estrategia <em>CCU</em></div>
      <div class="sec-kicker">5 genéricos CCU · mix estratégico</div>
    </div>
    <div class="ccu-split reveal">
      <div class="ccu-cell">
        <span class="badge">CCU · 5 genéricos</span>
        <div class="num" id="kpi-ccu-fact">—</div>
        <div class="row2">
          <div class="item"><b id="kpi-ccu-desc">—</b> descuentos</div>
          <div class="item">Desc. / Fact. <b id="kpi-ccu-ratio">—</b></div>
          <div class="item">Participa <b id="kpi-ccu-share">—</b> de la facturación</div>
        </div>
      </div>
      <div class="ccu-cell cool">
        <span class="badge">NO CCU · 9 genéricos</span>
        <div class="num" id="kpi-no-fact">—</div>
        <div class="row2">
          <div class="item"><b id="kpi-no-desc">—</b> descuentos</div>
          <div class="item">Desc. / Fact. <b id="kpi-no-ratio">—</b></div>
          <div class="item">Participa <b id="kpi-no-share">—</b> de la facturación</div>
        </div>
      </div>
    </div>
  </section>

  <section class="sec">
    <div class="sec-head reveal">
      <div class="sec-title">Pulse <em>diario</em></div>
      <div class="sec-kicker">Facturación neta &amp; descuentos por día · __DIAS__ días</div>
    </div>
    <div class="chart-card reveal">
      <div class="chart-head">
        <div class="chart-title">Día a día · Julio</div>
        <div class="chart-legend">
          <span><span class="swatch" style="background:var(--ink)"></span>Facturación</span>
          <span><span class="swatch" style="background:var(--badie)"></span>Descuentos</span>
        </div>
      </div>
      <div id="chart-daily"></div>
    </div>
  </section>

  <div class="row-2" style="margin-top:32px">
    <section class="sec" style="margin-top:0">
      <div class="sec-head reveal">
        <div class="sec-title">Mix por <em>genérico</em></div>
        <div class="sec-kicker">Share de facturación · top 14</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-mix-gen"></div>
      </div>
    </section>
    <section class="sec" style="margin-top:0">
      <div class="sec-head reveal">
        <div class="sec-title">Top <em>acciones</em></div>
        <div class="sec-kicker">por descuento · top 12</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-top-acc"></div>
      </div>
    </section>
  </div>

  <div class="row-2-eq" style="margin-top:8px">
    <section class="sec">
      <div class="sec-head reveal">
        <div class="sec-title">Mix por <em>sucursal</em></div>
        <div class="sec-kicker">facturación · 13 puntos de venta</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-mix-suc"></div>
      </div>
    </section>
    <section class="sec">
      <div class="sec-head reveal">
        <div class="sec-title">Porcentaje de Descuento <em>(Desc. / Fact.)</em></div>
        <div class="sec-kicker">presión promocional por genérico</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-ratios"></div>
      </div>
    </section>
  </div>

  <!-- ============ BLOQUE INFORME (segunda fuente · alcance distinto) ============ -->
  <section class="sec" id="sec-informe">
    <div class="sec-head reveal">
      <div class="sec-title">Retorno de la <em>inversión promocional</em></div>
      <div class="sec-kicker">modelo de tasa · compras vs reintegro CCU</div>
    </div>

    <div class="scope-warn reveal">
      <b>Otra fuente, otro alcance.</b> Esta sección sale del <i>informe final</i>, no de la BASE:
      corte al <span id="inf-corte">—</span>, sólo genéricos CCU. Los totales no coinciden
      con los de arriba y <b>no deben sumarse entre sí</b>. El filtro por sucursal no aplica acá
      (el modelo de tasa se liquida por zona).
    </div>

    <div class="pnl-grid reveal">
      <div class="pnl-cell">
        <div class="hero-kicker">Compras a CCU</div>
        <div class="pnl-num" id="pnl-compras">—</div>
      </div>
      <div class="pnl-cell">
        <div class="hero-kicker">Tasa generada</div>
        <div class="pnl-num" id="pnl-generada">—</div>
        <div class="hero-sub">lo que reintegra CCU</div>
      </div>
      <div class="pnl-cell">
        <div class="hero-kicker">Descuentos otorgados</div>
        <div class="pnl-num" id="pnl-desc">—</div>
      </div>
      <div class="pnl-cell danger">
        <div class="hero-kicker">Diferencia</div>
        <div class="pnl-num" id="pnl-dif">—</div>
        <div class="hero-sub" id="pnl-dif-sub">tasa generada − descuentos</div>
      </div>
    </div>

    <div class="chart-card reveal" style="margin-top:22px">
      <div class="chart-head">
        <div class="chart-title">Por zona · ¿la tasa cubre el descuento?</div>
        <div class="chart-legend">
          <span><span class="swatch" style="background:var(--ok)"></span>cubre</span>
          <span><span class="swatch" style="background:var(--badie)"></span>no cubre</span>
        </div>
      </div>
      <div id="chart-tasa"></div>
    </div>
  </section>

  <div class="row-2-eq" style="margin-top:8px">
    <section class="sec">
      <div class="sec-head reveal">
        <div class="sec-title">¿Quién <em>absorbe</em> el descuento?</div>
        <div class="sec-kicker">reversa · reintegro CCU vs costo BADIE</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-reversa"></div>
      </div>
    </section>
    <section class="sec">
      <div class="sec-head reveal">
        <div class="sec-title">Atribución · <em>quién reconoce</em></div>
        <div class="sec-kicker">área que se hace cargo del costo</div>
      </div>
      <div class="chart-card reveal">
        <div id="chart-quien"></div>
      </div>
    </section>
  </div>

  <section class="sec">
    <div class="sec-head reveal">
      <div class="sec-title">Descuento por <em>canal</em></div>
      <div class="sec-kicker">lista de precios · 11 canales comerciales</div>
    </div>
    <div class="chart-card reveal">
      <div id="chart-canal"></div>
    </div>
  </section>
  <!-- ============ FIN BLOQUE INFORME ============ -->

  <section class="sec">
    <div class="sec-head reveal">
      <div class="sec-title">Top <em>clientes</em></div>
      <div class="sec-kicker">descuento recibido · top 25</div>
    </div>
    <div class="chart-card reveal">
      <table class="tab" id="tab-clientes">
        <thead>
          <tr>
            <th style="width:42px">#</th>
            <th>Razón Social</th>
            <th>Código</th>
            <th>Sucursal</th>
            <th class="num">Operaciones</th>
            <th class="num">Descuento</th>
            <th class="num">% del total</th>
          </tr>
        </thead>
        <tbody></tbody>
      </table>
    </div>
  </section>

  <footer class="footnote">
    <div>Fuente: <b>__FUENTE__</b> · sólo lectura</div>
    <div>Generado por <b>build_dashboard.py</b> · <span id="ft-stamp">—</span></div>
  </footer>

</main>

<script type="application/json" id="dashboard-data">__JSON_DATA__</script>
<script>
(function () {
  'use strict';

  var DATA = JSON.parse(document.getElementById('dashboard-data').textContent);
  var CURRENT_FILTER = '__ALL__';  // '__ALL__' or a sucursal name

  function getActiveDataSet(name) {
    if (CURRENT_FILTER === '__ALL__') return DATA[name];
    var bySuc = DATA.by_sucursal && DATA.by_sucursal[CURRENT_FILTER];
    return bySuc ? bySuc[name] : DATA[name];
  }

  // ---- formatters ----
  // Compact currency in MILLONES (no billones). Whole millones w/ thousands separator.
  var fmtARS = function (v) {
    if (v === null || v === undefined || isNaN(v)) return '—';
    var abs = Math.abs(v);
    var sign = v < 0 ? '-' : '';
    if (abs >= 1e6) return '$ ' + sign + Math.round(abs / 1e6).toLocaleString('en-US') + ' M';
    if (abs >= 1e3) return '$ ' + sign + Math.round(abs / 1e3).toLocaleString('en-US') + ' K';
    return '$ ' + sign + Math.round(abs).toLocaleString('en-US');
  };
  var fmtARSFull = function (v) {
    if (v === null || v === undefined || isNaN(v)) return '—';
    return v.toLocaleString('es-AR', { maximumFractionDigits: 2 });
  };
  var fmtPct = function (v, d) {
    d = d === undefined ? 2 : d;
    return (v * 100).toFixed(d) + '%';
  };

  // ---- hero KPIs ----
  function renderHero() {
    var h = getActiveDataSet('hero');
    document.getElementById('kpi-fact').textContent = fmtARS(h.facturacion_neta);
    document.getElementById('kpi-fact-sub').textContent =
      'ARS ' + fmtARSFull(h.facturacion_neta);
    document.getElementById('kpi-desc').textContent = fmtARS(h.descuentos);
    document.getElementById('kpi-desc-sub').textContent =
      'ARS ' + fmtARSFull(h.descuentos) + ' entregados en el período';
    document.getElementById('kpi-ratio').textContent = (h.ratio_descuento_facturacion * 100).toFixed(2);
    document.getElementById('kpi-ratio-sub').textContent = 'Desc. / Fact.';
    document.getElementById('kpi-uni').innerHTML =
      '<span style="font-size:0.55em;color:var(--ink-mute);font-family:var(--mono);letter-spacing:0.1em">ACC</span> ' +
      h.acciones_distintas.toLocaleString('es-AR') +
      ' <span style="color:var(--ink-mute);font-family:var(--mono);font-size:0.55em">·</span> ' +
      '<span style="font-size:0.55em;color:var(--ink-mute);font-family:var(--mono);letter-spacing:0.1em">CLI</span> ' +
      h.clientes_distintos.toLocaleString('es-AR') +
      ' <span style="color:var(--ink-mute);font-family:var(--mono);font-size:0.55em">·</span> ' +
      '<span style="font-size:0.55em;color:var(--ink-mute);font-family:var(--mono);letter-spacing:0.1em">ART</span> ' +
      h.articulos_distintos.toLocaleString('es-AR');
    document.getElementById('kpi-uni-sub').textContent =
      'acciones · clientes · artículos distintos';
  }

  // ---- CCU split ----
  function renderCcuSplit() {
    var c = getActiveDataSet('ccu_split');
    var totalFact = c.ccu.facturacion_neta + c.no_ccu.facturacion_neta;
    document.getElementById('kpi-ccu-fact').textContent = fmtARS(c.ccu.facturacion_neta);
    document.getElementById('kpi-ccu-desc').textContent = fmtARS(c.ccu.descuentos);
    document.getElementById('kpi-ccu-ratio').textContent = fmtPct(c.ccu.ratio_d_f, 2);
    document.getElementById('kpi-ccu-share').textContent = totalFact ? fmtPct(c.ccu.facturacion_neta / totalFact, 1) : '—';

    document.getElementById('kpi-no-fact').textContent = fmtARS(c.no_ccu.facturacion_neta);
    document.getElementById('kpi-no-desc').textContent = fmtARS(c.no_ccu.descuentos);
    document.getElementById('kpi-no-ratio').textContent = fmtPct(c.no_ccu.ratio_d_f, 2);
    document.getElementById('kpi-no-share').textContent = totalFact ? fmtPct(c.no_ccu.facturacion_neta / totalFact, 1) : '—';
  }

  // ---- daily chart ----
  function renderDaily() {
    var daily = getActiveDataSet('daily');
    if (!daily || daily.length === 0) {
      document.getElementById('chart-daily').innerHTML =
        '<div style="padding:32px;color:var(--ink-mute);font-style:italic">Sin movimiento diario para esta sucursal.</div>';
      return;
    }
    var W = Math.max(700, document.getElementById('chart-daily').clientWidth - 8);
    var H = 280;
    var padL = 56, padR = 56, padT = 22, padB = 38;
    var innerW = W - padL - padR;
    var innerH = H - padT - padB;

    var maxFact = Math.max.apply(null, daily.map(function (d) { return d.facturacion_neta; }));
    var maxDesc = Math.max.apply(null, daily.map(function (d) { return d.descuentos; }));
    var n = daily.length;
    var xStep = innerW / Math.max(1, n - 1);

    var xAt = function (i) { return padL + i * xStep; };
    var yFactAt = function (v) { return padT + innerH - (v / maxFact) * innerH; };
    var yDescAt = function (v) { return padT + innerH - (v / maxDesc) * innerH; };

    var svg = '<svg viewBox="0 0 ' + W + ' ' + H + '" preserveAspectRatio="none" style="height:auto">';

    // gridlines (left axis = fact)
    var ticks = 4;
    for (var t = 0; t <= ticks; t++) {
      var y = padT + (innerH * t) / ticks;
      var factV = maxFact * (1 - t / ticks);
      var descV = maxDesc * (1 - t / ticks);
      svg += '<line x1="' + padL + '" y1="' + y + '" x2="' + (W - padR) + '" y2="' + y + '" stroke="rgba(26,22,18,0.08)" stroke-width="1"/>';
      svg += '<text x="' + (padL - 8) + '" y="' + (y + 3) + '" text-anchor="end" font-family="JetBrains Mono, monospace" font-size="9.5" fill="#8a7d6a">' + fmtARS(factV) + '</text>';
      svg += '<text x="' + (W - padR + 8) + '" y="' + (y + 3) + '" text-anchor="start" font-family="JetBrains Mono, monospace" font-size="9.5" fill="#b8351c" opacity="0.85">' + fmtARS(descV) + '</text>';
    }

    // x labels (every other day)
    for (var i = 0; i < n; i++) {
      if (i % 2 !== 0 && i !== n - 1) continue;
      var x = xAt(i);
      var day = daily[i].fecha.slice(8);
      svg += '<text x="' + x + '" y="' + (H - padB + 14) + '" text-anchor="middle" font-family="JetBrains Mono, monospace" font-size="9.5" fill="#8a7d6a">' + day + '</text>';
      svg += '<line x1="' + x + '" y1="' + (padT + innerH) + '" x2="' + x + '" y2="' + (padT + innerH + 4) + '" stroke="#8a7d6a" stroke-width="1"/>';
    }

    // fact area (subtle fill + line)
    var pathFactArea = 'M ' + xAt(0) + ' ' + (padT + innerH);
    for (var j = 0; j < n; j++) pathFactArea += ' L ' + xAt(j) + ' ' + yFactAt(daily[j].facturacion_neta);
    pathFactArea += ' L ' + xAt(n - 1) + ' ' + (padT + innerH) + ' Z';
    svg += '<path d="' + pathFactArea + '" fill="rgba(26,22,18,0.06)" stroke="none"/>';

    var pathFact = '';
    for (var k = 0; k < n; k++) {
      pathFact += (k === 0 ? 'M ' : ' L ') + xAt(k) + ' ' + yFactAt(daily[k].facturacion_neta);
    }
    svg += '<path d="' + pathFact + '" fill="none" stroke="#1a1612" stroke-width="2.2" stroke-linejoin="round" stroke-linecap="round"/>';

    // descuentos line
    var pathDesc = '';
    for (var m = 0; m < n; m++) {
      pathDesc += (m === 0 ? 'M ' : ' L ') + xAt(m) + ' ' + yDescAt(daily[m].descuentos);
    }
    svg += '<path d="' + pathDesc + '" fill="none" stroke="#b8351c" stroke-width="2" stroke-linejoin="round" stroke-linecap="round" stroke-dasharray="0"/>';

    // dots on fact line
    for (var p = 0; p < n; p++) {
      svg += '<circle cx="' + xAt(p) + '" cy="' + yFactAt(daily[p].facturacion_neta) + '" r="2.6" fill="#f4efe4" stroke="#1a1612" stroke-width="1.5"/>';
      svg += '<circle cx="' + xAt(p) + '" cy="' + yDescAt(daily[p].descuentos) + '" r="2.2" fill="#b8351c"/>';
    }

    // weekend marker (saturday/sunday)
    for (var q = 0; q < n; q++) {
      var dt = new Date(daily[q].fecha + 'T12:00:00');
      var dow = dt.getUTCDay();
      if (dow === 0 || dow === 6) {
        svg += '<rect x="' + (xAt(q) - xStep / 2) + '" y="' + padT + '" width="' + xStep + '" height="' + innerH + '" fill="rgba(184,53,28,0.04)"/>';
      }
    }

    svg += '</svg>';
    document.getElementById('chart-daily').innerHTML = svg;
  }

  // ---- mix generico: horizontal stacked bars ----
  function renderMixGen() {
    var data = getActiveDataSet('mix_generico');
    var entries = [];
    var keys = Object.keys(data);
    for (var i = 0; i < keys.length; i++) entries.push(data[keys[i]]);
    if (entries.length === 0) {
      document.getElementById('chart-mix-gen').innerHTML =
        '<div style="padding:24px;color:var(--ink-mute);font-style:italic">Sin datos para esta sucursal.</div>';
      return;
    }
    var maxFact = Math.max.apply(null, entries.map(function (e) { return e.facturacion_neta; }));
    var totalFact = entries.reduce(function (s, e) { return s + e.facturacion_neta; }, 0);

    var html = '';
    for (var j = 0; j < entries.length; j++) {
      var e = entries[j];
      var pct = e.facturacion_neta / maxFact;
      var share = (e.facturacion_neta / totalFact) * 100;
      var cls = e.es_ccu ? 'bar-fill ccu' : 'bar-fill';
      html += '<div class="bar-row">' +
        '<div class="label">' + keys[j] + (e.es_ccu ? '<span class="dot ccu"></span>' : '') + '</div>' +
        '<div class="bar-track"><div class="' + cls + '" style="width:' + (pct * 100).toFixed(2) + '%"></div></div>' +
        '<div class="val">' + fmtARS(e.facturacion_neta) + ' <span style="color:var(--ink-mute);font-size:10px">' + share.toFixed(1) + '%</span></div>' +
      '</div>';
    }
    document.getElementById('chart-mix-gen').innerHTML = html;
  }

  // ---- top acciones: horizontal bars (descuento) ----
  function renderTopAcc() {
    var data = getActiveDataSet('top_acciones').slice(0, 12);
    var max = Math.max.apply(null, data.map(function (d) { return d.descuento; }));
    var html = '';
    for (var i = 0; i < data.length; i++) {
      var d = data[i];
      var pct = d.descuento / max;
      var lbl = d.descripcion.length > 28 ? d.descripcion.slice(0, 28) + '…' : d.descripcion;
      html += '<div class="bar-row" title="' + d.descripcion.replace(/"/g, '&quot;') + '">' +
        '<div class="label"><small>' + d.accion + '</small> ' + lbl + '</div>' +
        '<div class="bar-track"><div class="bar-fill badie" style="width:' + (pct * 100).toFixed(2) + '%"></div></div>' +
        '<div class="val">' + fmtARS(d.descuento) + '</div>' +
      '</div>';
    }
    if (data.length === 0) html = '<div style="padding:24px;color:var(--ink-mute);font-style:italic">Sin acciones para esta sucursal.</div>';
    document.getElementById('chart-top-acc').innerHTML = html;
  }

  // ---- mix sucursal: vertical bar chart ----
  function renderMixSuc() {
    var data = DATA.mix_sucursal;
    var entries = [];
    var keys = Object.keys(data);
    for (var i = 0; i < keys.length; i++) entries.push({ key: keys[i], val: data[keys[i]] });
    var W = Math.max(360, document.getElementById('chart-mix-suc').clientWidth - 8);
    var H = 360;
    var padL = 50, padR = 16, padT = 14, padB = 110;
    var innerW = W - padL - padR;
    var innerH = H - padT - padB;
    var max = Math.max.apply(null, entries.map(function (e) { return e.val.facturacion_neta; }));
    var ticks = 4;
    var n = entries.length;
    var band = innerW / n;
    var barW = Math.max(8, Math.min(34, band * 0.62));

    var svg = '<svg viewBox="0 0 ' + W + ' ' + H + '" style="height:auto">';
    // gridlines
    for (var t = 0; t <= ticks; t++) {
      var y = padT + (innerH * t) / ticks;
      var v = max * (1 - t / ticks);
      svg += '<line x1="' + padL + '" y1="' + y + '" x2="' + (W - padR) + '" y2="' + y + '" stroke="rgba(26,22,18,0.08)" stroke-width="1"/>';
      svg += '<text x="' + (padL - 6) + '" y="' + (y + 3) + '" text-anchor="end" font-family="JetBrains Mono, monospace" font-size="9.5" fill="#8a7d6a">' + fmtARS(v) + '</text>';
    }
    // bars
    for (var i2 = 0; i2 < n; i2++) {
      var e2 = entries[i2];
      var x = padL + i2 * band + (band - barW) / 2;
      var h2 = (e2.val.facturacion_neta / max) * innerH;
      var y2 = padT + innerH - h2;
      // Cross-sucursal comparison stays global; the active filter is highlighted instead.
      var isSel = (CURRENT_FILTER !== '__ALL__' && e2.key === CURRENT_FILTER);
      var barFill = (CURRENT_FILTER === '__ALL__') ? '#1a1612' : (isSel ? '#b8351c' : 'rgba(26,22,18,0.22)');
      svg += '<rect x="' + x + '" y="' + y2 + '" width="' + barW + '" height="' + h2 + '" fill="' + barFill + '"/>';
      // ratio indicator (small line)
      var ratioY = padT + innerH - (e2.val.ratio_d_f * 0.4) * innerH; // visual scale
      svg += '<rect x="' + x + '" y="' + ratioY + '" width="' + barW + '" height="2" fill="#b8351c"/>';
      // x label (rotated)
      var cx = x + barW / 2;
      var lbl = e2.key.replace(/^\d+\s*-\s*/, '');
      if (lbl.length > 14) lbl = lbl.slice(0, 12) + '…';
      svg += '<text x="' + cx + '" y="' + (padT + innerH + 12) + '" text-anchor="end" transform="rotate(-50 ' + cx + ' ' + (padT + innerH + 12) + ')" font-family="Manrope, sans-serif" font-size="10" fill="#1a1612">' + lbl + '</text>';
      // value above bar
      svg += '<text x="' + cx + '" y="' + (y2 - 6) + '" text-anchor="middle" font-family="JetBrains Mono, monospace" font-size="9.5" fill="#4a4035">' + fmtARS(e2.val.facturacion_neta) + '</text>';
    }
    // baseline
    svg += '<line x1="' + padL + '" y1="' + (padT + innerH) + '" x2="' + (W - padR) + '" y2="' + (padT + innerH) + '" stroke="#1a1612" stroke-width="1.5"/>';
    svg += '</svg>';
    document.getElementById('chart-mix-suc').innerHTML = svg;
  }

  // ---- ratios: vertical lollipop chart ----
  function renderRatios() {
    var mg = getActiveDataSet('mix_generico');
    var entries = [];
    var keys = Object.keys(mg);
    for (var i = 0; i < keys.length; i++) {
      var v = mg[keys[i]];
      entries.push({ name: keys[i], ratio: v.ratio_d_f, fact: v.facturacion_neta, ccu: v.es_ccu });
    }
    if (entries.length === 0) {
      document.getElementById('chart-ratios').innerHTML =
        '<div style="padding:24px;color:var(--ink-mute);font-style:italic">Sin datos para esta sucursal.</div>';
      return;
    }
    var max = Math.max.apply(null, entries.map(function (e) { return e.ratio; })) || 1;
    var W = Math.max(360, document.getElementById('chart-ratios').clientWidth - 8);
    var H = 360;
    var padL = 130, padR = 28, padT = 14, padB = 14;
    var innerW = W - padL - padR;
    var innerH = H - padT - padB;
    var n = entries.length;
    var rowH = innerH / n;

    var svg = '<svg viewBox="0 0 ' + W + ' ' + H + '" style="height:auto">';
    // reference line at overall ratio (of the ACTIVE dataset)
    var overallRatio = getActiveDataSet('hero').ratio_descuento_facturacion;
    var xRef = padL + (overallRatio / max) * innerW;
    svg += '<line x1="' + xRef + '" y1="' + padT + '" x2="' + xRef + '" y2="' + (padT + innerH) + '" stroke="#b8351c" stroke-width="1" stroke-dasharray="3 3" opacity="0.7"/>';
    svg += '<text x="' + xRef + '" y="' + (padT + 10) + '" text-anchor="middle" font-family="JetBrains Mono, monospace" font-size="9" fill="#b8351c">GLOBAL ' + (overallRatio * 100).toFixed(1) + '%</text>';
    for (var j = 0; j < n; j++) {
      var e = entries[j];
      var yMid = padT + j * rowH + rowH / 2;
      var xEnd = padL + (e.ratio / max) * innerW;
      // label
      svg += '<text x="' + (padL - 10) + '" y="' + (yMid + 3) + '" text-anchor="end" font-family="Manrope, sans-serif" font-size="11" fill="#1a1612">' + e.name + (e.ccu ? ' ●' : '') + '</text>';
      // track
      svg += '<line x1="' + padL + '" y1="' + yMid + '" x2="' + (padL + innerW) + '" y2="' + yMid + '" stroke="rgba(26,22,18,0.06)" stroke-width="1"/>';
      // dot
      svg += '<line x1="' + padL + '" y1="' + yMid + '" x2="' + xEnd + '" y2="' + yMid + '" stroke="' + (e.ccu ? '#2c4a52' : '#1a1612') + '" stroke-width="2.5"/>';
      svg += '<circle cx="' + xEnd + '" cy="' + yMid + '" r="5" fill="' + (e.ccu ? '#2c4a52' : '#1a1612') + '"/>';
      svg += '<circle cx="' + xEnd + '" cy="' + yMid + '" r="2" fill="#f4efe4"/>';
      // pct
      svg += '<text x="' + (xEnd + 10) + '" y="' + (yMid + 3) + '" text-anchor="start" font-family="JetBrains Mono, monospace" font-size="10" fill="#1a1612">' + (e.ratio * 100).toFixed(2) + '%</text>';
    }
    svg += '</svg>';
    document.getElementById('chart-ratios').innerHTML = svg;
  }

  // ---- top clientes table ----
  function renderClientes() {
    var top = getActiveDataSet('top_clientes');
    // % is always against the active scope's total discount
    var totalDesc = getActiveDataSet('hero').descuentos;
    var tbody = document.querySelector('#tab-clientes tbody');
    var html = '';
    for (var i = 0; i < top.length; i++) {
      var c2 = top[i];
      var share = totalDesc ? (c2.descuento / totalDesc) * 100 : 0;
      var suc = c2.sucursal || CURRENT_FILTER;
      var badge = suc.indexOf('CASA CENTRAL') !== -1 ? 'badie' : '';
      html += '<tr>' +
        '<td class="code">' + (i + 1).toString().padStart(2, '0') + '</td>' +
        '<td>' + c2.razon_social + '</td>' +
        '<td class="code">' + c2.cod_cliente + '</td>' +
        '<td class="' + badge + '">' + suc + '</td>' +
        '<td class="num">' + c2.operaciones.toLocaleString('es-AR') + '</td>' +
        '<td class="num">' + fmtARS(c2.descuento) + '</td>' +
        '<td class="num">' + share.toFixed(2) + '%</td>' +
      '</tr>';
    }
    if (top.length === 0) html = '<tr><td colspan="7" style="padding:24px;color:var(--ink-mute);font-style:italic;text-align:center">Sin clientes para esta sucursal.</td></tr>';
    tbody.innerHTML = html;
  }

  // ---- footer stamp ----
  document.getElementById('ft-stamp').textContent = new Date().toLocaleString('es-AR');

  // ================= BLOQUE INFORME (no responde al filtro por sucursal) =======
  // El modelo de tasa se liquida por ZONA contra CCU, no por sucursal — por eso
  // estas secciones son siempre globales y quedan rotuladas como tal.
  function renderInforme() {
    var inf = DATA.informe;
    var sec = document.getElementById('sec-informe');
    if (!inf || !inf.tasa_model) { if (sec) sec.style.display = 'none'; return; }

    var cortes = inf.tasa_cortes || {};
    var el = document.getElementById('inf-corte');
    if (el) {
      var cs = [];
      for (var k in cortes) { if (cortes[k]) cs.push(k + ' ' + cortes[k]); }
      el.textContent = cs.join(' · ') || '—';
    }

    var r = inf.tasa_rollup || {};
    document.getElementById('pnl-compras').textContent  = fmtARS(r.compras);
    document.getElementById('pnl-generada').textContent = fmtARS(r.tasa_generada);
    document.getElementById('pnl-desc').textContent     = fmtARS(r.descuentos);
    var dif = r.diferencia || 0;
    document.getElementById('pnl-dif').textContent = fmtARS(dif);
    document.getElementById('pnl-dif-sub').textContent = dif < 0
      ? 'BADIE pone de más'
      : 'la tasa cubre el descuento';

    // ---- tasa por zona: barra = descuento; verde = cubierto por tasa ----
    var rows = inf.tasa_model;
    var maxV = 0;
    for (var i = 0; i < rows.length; i++) {
      maxV = Math.max(maxV, rows[i].descuentos, rows[i].tasa_generada);
    }
    var html = '', lastGen = null;
    for (var j = 0; j < rows.length; j++) {
      var t = rows[j];
      if (t.generico !== lastGen) {
        html += '<div class="gen-head">' + t.generico + '</div>';
        lastGen = t.generico;
      }
      var wDesc = maxV ? (t.descuentos / maxV) * 100 : 0;
      var wGot  = maxV ? (t.tasa_generada / maxV) * 100 : 0;
      var cubre = t.tasa_generada >= t.descuentos;
      var pct = t.tasa_pct !== null && t.tasa_pct !== undefined
        ? '<small>tasa ' + (t.tasa_pct * 100).toFixed(1) + '%</small>' : '<small>&nbsp;</small>';
      html += '<div class="tasa-row' + (t.es_total ? ' total' : '') + '">' +
        '<div class="zona">' + t.zona + pct + '</div>' +
        '<div class="tasa-bar" title="Descuento ' + fmtARS(t.descuentos) +
          ' · Tasa generada ' + fmtARS(t.tasa_generada) + '">' +
          '<div class="need" style="width:' + wDesc.toFixed(2) + '%"></div>' +
          '<div class="got' + (cubre ? '' : ' short') + '" style="width:' + wGot.toFixed(2) + '%"></div>' +
          '<div class="marker" style="left:' + wDesc.toFixed(2) + '%"></div>' +
        '</div>' +
        '<div class="delta ' + (t.diferencia < 0 ? 'neg' : 'pos') + '">' + fmtARS(t.diferencia) + '</div>' +
      '</div>';
    }
    document.getElementById('chart-tasa').innerHTML = html;

    // ---- reversa ----
    var rev = inf.reversa_split || [];
    if (rev.length) {
      var totRev = rev.reduce(function (s, x) { return s + x.descuento; }, 0);
      var rh = '<div class="split-bar">';
      for (var k2 = 0; k2 < rev.length; k2++) {
        var w = totRev ? (rev[k2].descuento / totRev) * 100 : 0;
        var cls = rev[k2].tipo.indexOf('NO') === 0 ? 'badie' : 'ok';
        rh += '<div class="split-seg ' + cls + '" style="width:' + w.toFixed(2) + '%">' +
              (w > 12 ? w.toFixed(0) + '%' : '') + '</div>';
      }
      rh += '</div>';
      for (var k3 = 0; k3 < rev.length; k3++) {
        var x = rev[k3];
        var dot = x.tipo.indexOf('NO') === 0 ? 'badie' : 'ccu';
        rh += '<div class="bar-row" style="grid-template-columns:150px 1fr 96px">' +
          '<div class="label"><span class="dot ' + dot + '"></span>' + x.tipo + '</div>' +
          '<div class="val" style="text-align:left;color:var(--ink-mute)">' + x.detalle +
            ' · ' + x.acciones + ' acc.</div>' +
          '<div class="val">' + fmtARS(x.descuento) + '</div>' +
        '</div>';
      }
      document.getElementById('chart-reversa').innerHTML = rh;
    }

    // ---- quien reconoce ----
    var qr = inf.quien_reconoce || [];
    if (qr.length) {
      var maxQ = Math.max.apply(null, qr.map(function (q) { return q.total; }));
      var qh = '';
      for (var m = 0; m < qr.length; m++) {
        var q = qr[m];
        qh += '<div class="bar-row">' +
          '<div class="label">' + q.quien + '</div>' +
          '<div class="bar-track"><div class="bar-fill' + (q.es_area ? ' ccu' : '') +
            '" style="width:' + ((q.total / maxQ) * 100).toFixed(2) + '%"></div></div>' +
          '<div class="val">' + fmtARS(q.total) +
            ' <span style="color:var(--ink-mute);font-size:10px">' + (q.share * 100).toFixed(0) + '%</span></div>' +
        '</div>';
      }
      document.getElementById('chart-quien').innerHTML = qh;
    }

    // ---- canal ----
    var cn = inf.canal || [];
    if (cn.length) {
      var maxC = Math.max.apply(null, cn.map(function (c) { return c.descuento; }));
      var ch = '';
      for (var n = 0; n < cn.length; n++) {
        var c2 = cn[n];
        ch += '<div class="bar-row">' +
          '<div class="label">' + c2.canal + (c2.sin_canal ? ' <small>sin mapear</small>' : '') + '</div>' +
          '<div class="bar-track"><div class="bar-fill' + (c2.sin_canal ? ' badie' : '') +
            '" style="width:' + ((c2.descuento / maxC) * 100).toFixed(2) + '%"></div></div>' +
          '<div class="val">' + fmtARS(c2.descuento) +
            ' <span style="color:var(--ink-mute);font-size:10px">' + (c2.share * 100).toFixed(1) + '%</span></div>' +
        '</div>';
      }
      document.getElementById('chart-canal').innerHTML = ch;
    }
  }

  // ---- render everything from the ACTIVE dataset ----
  function renderAll() {
    renderHero();
    renderCcuSplit();
    renderDaily();
    renderMixGen();
    renderTopAcc();
    renderMixSuc();
    renderRatios();
    renderClientes();
  }

  // ---- filter: chip click handler ----
  function applyFilter(sucursal) {
    CURRENT_FILTER = sucursal;
    var chips = document.querySelectorAll('#filter-sucursal .chip');
    for (var i = 0; i < chips.length; i++) {
      var isActive = chips[i].getAttribute('data-sucursal') === sucursal;
      if (isActive) { chips[i].classList.add('active'); }
      else { chips[i].classList.remove('active'); }
    }
    var hint = document.getElementById('filter-hint');
    var scopeTag = document.getElementById('scope-tag');
    if (sucursal === '__ALL__') {
      hint.textContent = 'Vista global — ' + (DATA.sucursal_options || []).length + ' sucursales';
      if (scopeTag) scopeTag.textContent = 'Toda la red';
    } else {
      var h = (DATA.by_sucursal[sucursal] || {}).hero || {};
      hint.textContent = (h.clientes_distintos || 0).toLocaleString('es-AR') +
        ' clientes · ' + (h.acciones_distintas || 0) + ' acciones · ' +
        (h.articulos_distintos || 0) + ' artículos';
      if (scopeTag) scopeTag.textContent = sucursal;
    }
    renderAll();
  }

  var allChips = document.querySelectorAll('#filter-sucursal .chip');
  for (var ci = 0; ci < allChips.length; ci++) {
    (function (chip) {
      chip.addEventListener('click', function () {
        applyFilter(chip.getAttribute('data-sucursal'));
      });
    })(allChips[ci]);
  }

  // ---- boot ----
  renderAll();
  renderInforme();   // fuera de renderAll: no responde al filtro por sucursal

  // re-render on resize (debounced)
  var rt;
  window.addEventListener('resize', function () {
    clearTimeout(rt);
    rt = setTimeout(function () {
      renderDaily();
      renderMixSuc();
      renderRatios();
    }, 120);
  });
})();
</script>
</body>
</html>
"""


def render_html(payload: dict) -> str:
    json_data = json.dumps(payload, ensure_ascii=False, indent=2)
    meta = payload["meta"]
    html = HTML_TEMPLATE
    html = html.replace("__JSON_DATA__", json_data)
    html = html.replace("__PERIODO_DESDE__", meta["periodo_desde"] or "—")
    html = html.replace("__PERIODO_HASTA__", meta["periodo_hasta"] or "—")
    html = html.replace("__DIAS__", str(meta.get("dias_en_periodo", "—")))
    html = html.replace("__GENERADO_EN__", meta["generado_en"])
    html = html.replace("__FUENTE__", meta["fuente"])
    # Filter chips: one per sucursal
    import re
    chips = []
    for s in payload.get("sucursal_options", []):
        label = re.sub(r"^\d+\s*-\s*", "", s)
        # escape for HTML attr
        label_safe = label.replace("&", "&amp;").replace('"', "&quot;")
        s_safe = s.replace("&", "&amp;").replace('"', "&quot;")
        chips.append(
            f'<button class="chip" data-sucursal="{s_safe}" type="button">{label_safe}</button>'
        )
    html = html.replace("__FILTER_CHIPS__", "\n      ".join(chips))
    return html


# ============== main ==============
def main(argv: list[str]) -> int:
    xlsx = Path(argv[1]) if len(argv) > 1 else DEFAULT_XLSX
    out_html = Path(argv[2]) if len(argv) > 2 else DEFAULT_OUT
    if not xlsx.exists():
        print(f"ERROR: BASE xlsx not found at {xlsx}", file=sys.stderr)
        return 1

    print(f"Loading {xlsx} (streaming mode) ...")
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    informe = load_informe(Path(argv[3]) if len(argv) > 3 else DEFAULT_INFORME)
    payload = build_payload(wb, informe=informe)

    # Always also dump the JSON for downstream consumers
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"Wrote {JSON_OUT} ({JSON_OUT.stat().st_size:,} bytes)")

    html = render_html(payload)
    out_html.write_text(html, encoding="utf-8")
    print(f"Wrote {out_html} ({out_html.stat().st_size:,} bytes)")

    h = payload["hero"]
    print(
        f"  facturacion_neta={h['facturacion_neta']:,.2f}  "
        f"descuentos={h['descuentos']:,.2f}  ratio_d_f={h['ratio_descuento_facturacion']:.2%}  "
        f"acciones={h['acciones_distintas']}  clientes={h['clientes_distintos']}  "
        f"articulos={h['articulos_distintos']}"
    )
    print(f"  periodo: {payload['meta']['periodo_desde']} → {payload['meta']['periodo_hasta']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
