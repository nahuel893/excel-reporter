#!/usr/bin/env python3
"""
build_informe_excel.py — Informe analitico Acciones Comerciales (workbook Excel).

Tres fuentes, declaradas y NO mezcladas en un mismo numero:
  A) BASE control xlsx  — todos los genericos, corte = PERIODO_HASTA del run
  B) INFORME final xlsm — solo CCU; el corte se LEE de COMPRAS & DESC C1/D1
  C) gold (PostgreSQL)  — ventas y cobertura del periodo

Reglas del proyecto que este script respeta:
  - NUNCA redondea/trunca datos. Solo aplica number_format de Excel.
  - Toda tabla lleva fila TOTAL GENERAL con estilo propio.
  - Join con dim_* SIEMPRE por clave compuesta (id + id_sucursal).

Uso:
    python build_informe_excel.py [salida.xlsx]
"""
from __future__ import annotations

import sys
import warnings
from datetime import date, datetime
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO = Path(__file__).resolve().parents[2]
BASE_XLSX = REPO / "data/output/acciones-comerciales/2026-07/BASE control Acciones Comerciales - JULIO 2026.xlsx"
INFORME_XLSM = REPO / "data/backups/acciones-comerciales-2026-07-20/engine-informe/INFO - ACCIONES BADIE JULIO 2026.xlsm"
OUT_DEFAULT = REPO / "data/output/acciones-comerciales/2026-07/Informe Analitico Acciones Comerciales - JULIO 2026.xlsx"
IMG_DIR = Path(__file__).resolve().parent / "_img"
# compras.xls: insumo que baja el watcher solo. De aca sale la columna
# Compras del modelo de tasa, en vez de leerla del .xlsm manual.
COMPRAS_DIR = Path("/home/nahuel/VM shared/archivos_diarios/acciones")
COMPRAS_XLS = COMPRAS_DIR / "compras.xls"   # nombre canonico; ver _resolver_compras

PERIODO_DESDE, PERIODO_HASTA = "2026-07-01", "2026-07-21"
CCU = {"CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"}

# ─────────────────────────────────────────────────────────────────────────
# id de lista de precios -> descripcion. Va fijo porque NO hay de donde
# leerlo: gold.dim_lista_precio esta VACIA y dim_cliente.des_lista_precio
# viene en blanco en las 11 listas. El .xlsm trae el nombre pero no el id.
#
# Derivado (no tipeado): se cruzo el padron LISTA del .xlsm contra
# gold.dim_cliente.id_lista_precio por cliente. Cada descripcion mapea a un
# id dominante con 96,9%-100% de los clientes; el resto es el fan-out
# conocido de joinear id_cliente sin id_sucursal (REGLA DE ORO). Coincide
# con el universo ON PREMISE ya documentado: listas 4 y 8.
#
# Si aparece una lista nueva, la hoja Canal la muestra sin id en vez de
# romper — un id inventado seria peor que no tenerlo.
# ─────────────────────────────────────────────────────────────────────────
LISTA_PRECIO_ID = {
    "LISTA SALTA MAYORISTA": 1,
    "LISTA SALTA MINORISTA": 3,
    "LISTA SALTA ON PREMISE": 4,
    "LISTA SALTA AUTOSERVICIOS": 5,
    "INTERIOR MAYORISTA": 6,
    "INTERIOR MINORISTA": 7,
    "INTERIOR ON PREMISE": 8,
    "INTERIOR AUTOSERVICIOS": 9,
    "SUB DISTRIBUIDORES SALTA CAPIT": 11,
    "SUB DISTRIBUIDORES INTERIOR": 12,
    "NORTE EMPRENDIMIENTOS SAS": 14,
}


# ─────────────────────────────────────────────────────────────────────────
# Tasa% NEGOCIADA con CCU que pisa lo que trae el .xlsm.
#
# La Tasa% es el unico numero del modelo que no produce ningun sistema: se
# acuerda con CCU y llega por mensaje. Cuando cambia, el .xlsm manual tarda
# en reflejarlo, asi que se declara aca — con fecha y origen, para que
# dentro de dos meses se sepa de donde salio cada valor.
#
# Informado por CCU el 2026-08-18 ("por todo concepto"):
#   CCU (CERVEZAS)     Jujuy Interior 5,3%   Salta Int. Norte 7,5%
#   ADO (AGUAS DANONE) Jujuy Interior 9,5%   Salta Int. Norte 9,1%
# Corregido el 2026-08-19:
#   CCU (CERVEZAS)     Salta Capital 13,4%   (el .xlsm traia 12,8%)
#
# SALTA INT. SUR no se informo: sigue saliendo del .xlsm. Una zona ausente
# aca NO se pisa.
# ─────────────────────────────────────────────────────────────────────────
TASA_OVERRIDE = {
    ("CERVEZAS", "SALTA CAPITAL"): 0.134,
    ("CERVEZAS", "JUJUY INTERIOR"): 0.053,
    ("CERVEZAS", "SALTA INT. NORTE"): 0.075,
    ("AGUAS DANONE", "JUJUY INTERIOR"): 0.095,
    ("AGUAS DANONE", "SALTA INT. NORTE"): 0.091,
}
TASA_OVERRIDE_FUENTE = "informada por CCU (18/19-ago-2026)"


def etiquetar_lista(nombre) -> str:
    """'LISTA SALTA ON PREMISE' -> '4 - LISTA SALTA ON PREMISE'.

    Misma convencion que SUCURSAL ('1 - CASA CENTRAL'). Sin id conocido
    devuelve el nombre pelado.
    """
    s = str(nombre).strip()
    i = LISTA_PRECIO_ID.get(s.upper())
    return f"{i} - {s}" if i is not None else s

# ---------------- paleta ----------------
# Nucleo: la misma del dashboard. Pasteles: agregados para las series.
PAPER, INK, BADIE, TEAL = "F4EFE4", "1A1612", "B8351C", "2C4A52"
WARN, OK, INK_SOFT, INK_MUTE = "C98717", "4A6C3A", "4A4035", "8A7D6A"
RULE = "D8CFBE"
PASTEL = ["E8B4A8", "A2B9C4", "A8C4A2", "E8D0A8", "C4A8C4",
          "B4D4C4", "D9CDB8", "BDD0DB", "E0A89A", "C2CBB4"]

F_MONEY, F_MONEY0, F_PCT, F_INT, F_DEC = '$ #,##0', '$ #,##0;[Red]-$ #,##0', '0.00%', '#,##0', '#,##0.00'

thin = Side(style="thin", color=RULE)
med = Side(style="medium", color=INK)


# ---------------- helpers de estilo ----------------
def hdr(ws, row, headers, start=1, fill=INK, color=PAPER):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=med)
    ws.row_dimensions[row].height = 28


def title(ws, row, text, sub=None, col=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=16, bold=True, color=INK)
    if sub:
        s = ws.cell(row=row + 1, column=col, value=sub)
        s.font = Font(name="Calibri", size=9, italic=True, color=INK_MUTE)
    return row + (3 if sub else 2)


def band(ws, row, text, col=1, span=8, fill=TEAL):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color=PAPER)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i in range(1, span):
        d = ws.cell(row=row, column=col + i)
        d.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[row].height = 20
    return row + 1


def widths(ws, spec: dict):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def write_df(ws, df: pd.DataFrame, start_row: int, formats: dict,
             total_row: bool = True, start_col: int = 1,
             table_name: str | None = None, zebra: bool = True,
             no_total: tuple = ()):
    """Escribe un DataFrame con header, zebra, TOTAL GENERAL y tabla nativa.

    `no_total`: columnas numericas que NO deben sumarse en la fila TOTAL porque
    la suma no significa nada — codigos/IDs y ratios. Los porcentajes (F_PCT) ya
    se excluyen solos."""
    hdr(ws, start_row, list(df.columns), start=start_col)
    r = start_row + 1
    for _, rec in df.iterrows():
        for i, col in enumerate(df.columns):
            c = ws.cell(row=r, column=start_col + i, value=rec[col])
            c.font = Font(name="Calibri", size=9, color=INK)
            c.border = Border(bottom=thin)
            if zebra and (r - start_row) % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FBF8F1")
            if col in formats:
                c.number_format = formats[col]
                c.alignment = Alignment(horizontal="right")
        r += 1

    if total_row:
        num = df.select_dtypes("number").columns
        for i, col in enumerate(df.columns):
            c = ws.cell(row=r, column=start_col + i)
            c.font = Font(name="Calibri", size=9, bold=True, color=PAPER)
            c.fill = PatternFill("solid", fgColor=INK)
            c.border = Border(top=med)
            if i == 0:
                c.value = "TOTAL GENERAL"
            elif col in num:
                # ratios/porcentajes/IDs no se suman: se dejan vacios
                if formats.get(col) == F_PCT or col in no_total:
                    c.value = None
                else:
                    c.value = float(df[col].sum())
                    c.number_format = formats.get(col, F_INT)
                    c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 18
        r += 1

    if table_name:
        ref = (f"{get_column_letter(start_col)}{start_row}:"
               f"{get_column_letter(start_col + len(df.columns) - 1)}{r - 2}")
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.add_table(t)
    return r


def style_chart(ch, title_txt, x_title=None, y_title=None, w=24, h=11):
    ch.title = title_txt
    ch.style = None
    ch.width, ch.height = w, h
    if x_title:
        ch.x_axis.title = x_title
    if y_title:
        ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    return ch


def _f(v) -> float:
    """Numero o 0.0. Nunca revienta por un string suelto en una celda."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ---------------- carga de datos ----------------
def load_base():
    from openpyxl import load_workbook
    wb = load_workbook(BASE_XLSX, read_only=True, data_only=True)

    fact = []
    for r in wb["FACT_NET"].iter_rows(min_row=2, values_only=True):
        if r and r[0] and r[0] != "TOTAL GENERAL":
            fact.append({"sucursal": r[0], "codigo": r[1], "articulo": r[2],
                         "marca": r[3], "generico": r[4],
                         "facturacion": float(r[5] or 0), "descuentos": float(r[6] or 0)})
    fact = pd.DataFrame(fact)

    acc = []
    for r in wb["ART-ACCION"].iter_rows(min_row=2, values_only=True):
        if r and r[0] and r[0] != "TOTAL GENERAL" and r[3]:
            acc.append({"sucursal": r[0], "codigo": r[1], "articulo": r[2],
                        "accion": r[3], "desc_accion": r[4], "mvb": r[5],
                        "descuento": float(r[6] or 0)})
    acc = pd.DataFrame(acc)

    # Las columnas del wapi se resuelven por NOMBRE DE HEADER, no por posicion.
    # Con indices fijos, agregar una columna al enriquecido (paso 19-bis sumo
    # "PRECIO FINAL (terna)" y "Origen Precio") corre todo y el lector empieza
    # a sumar strings como si fueran plata — falla ruidosa si hay suerte,
    # silenciosa si la columna corrida tambien es numerica.
    ws_wapi = wb["wapi"]
    hdr_wapi = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws_wapi.iter_rows(min_row=1, max_row=1))
    ]

    def _idx(nombre: str) -> int:
        if nombre in hdr_wapi:
            return hdr_wapi.index(nombre)
        raise KeyError(
            f"La hoja wapi no trae la columna {nombre!r}. "
            f"Columnas presentes: {hdr_wapi}"
        )

    i_fecha, i_total2, i_desc = _idx("Fecha"), _idx("Total2"), _idx("Descuento")

    # Columnas del wapi que van al informe como hoja cruda. Son las que
    # alimentan algun numero (Fecha/Total2/Descuento) mas las que hacen falta
    # para ubicar la fila (comprobante, cliente, articulo, accion, sucursal).
    # Las 32 completas serian 1,8 millones de celdas: el archivo se vuelve
    # inmanejable y las 20 que sobran no explican ningun numero del informe.
    _WAPI_COLS = ["Fecha", "Comprobante", "Cod. Cliente", "Razón Social",
                  "Artículo CMQ", "Descripción", "Marca", "Acción",
                  "Descripción Acción", "SUCURSAL", "Total2", "Descuento",
                  "Tipo Descuento"]
    wapi_idx = [(n, hdr_wapi.index(n)) for n in _WAPI_COLS if n in hdr_wapi]

    daily = {}
    wapi_crudo = []
    for r in ws_wapi.iter_rows(min_row=2, values_only=True):
        if not r or not r[i_fecha]:
            continue
        f = r[i_fecha]
        f = f.date().isoformat() if hasattr(f, "date") else str(f)[:10]
        if not f[:4].isdigit():
            continue
        d = daily.setdefault(f, {"fecha": f, "facturacion": 0.0, "descuentos": 0.0, "operaciones": 0})
        d["facturacion"] += _f(r[i_total2])
        d["descuentos"] += _f(r[i_desc])
        d["operaciones"] += 1
        rec = {}
        for nombre, i in wapi_idx:
            v = r[i]
            rec[nombre] = f if nombre == "Fecha" else v
        wapi_crudo.append(rec)
    daily = pd.DataFrame(sorted(daily.values(), key=lambda x: x["fecha"]))
    wapi_crudo = pd.DataFrame(wapi_crudo)

    # CLIENTE-FECHA: grano cliente x accion x articulo (lo que pide la hoja Clientes)
    cli = []
    for r in wb["CLIENTE-FECHA"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or r[2] is None or r[2] == "":
            continue
        if str(r[3]) == "TOTAL GENERAL" or str(r[1]) == "TOTAL GENERAL":
            continue
        cli.append({"sucursal": r[1], "cod_cliente": r[2], "razon_social": r[3],
                    "cod_articulo": r[4], "articulo": r[5], "generico": r[6],
                    "accion": r[7], "desc_accion": r[8], "descuento": float(r[9] or 0)})
    cli = pd.DataFrame(cli)

    # ACC-GEN: pivot sucursal x accion x generico (la BASE ya lo construye)
    accgen = []
    if "ACC-GEN" in wb.sheetnames:
        ws = wb["ACC-GEN"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        gens = [h for h in hdr[5:] if h]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or r[0] == "TOTAL GENERAL":
                continue
            rec = {"sucursal": r[0], "accion": r[1], "desc_accion": r[2], "mvb": r[3]}
            for i, g in enumerate(gens):
                v = r[5 + i]
                rec[g] = float(v) if isinstance(v, (int, float)) else 0.0
            accgen.append(rec)
    accgen = pd.DataFrame(accgen)

    wb.close()
    return fact, acc, daily, cli, accgen, wapi_crudo


def load_informe():
    out = {}
    from openpyxl import load_workbook
    wb = load_workbook(INFORME_XLSM, read_only=True, data_only=True, keep_vba=False)
    rows = list(wb["COMPRAS & DESC"].iter_rows(min_row=1, max_row=32, values_only=True))

    # Corte REAL del informe: celdas C1 (cervezas) y D1 (aguas). Se lee del archivo
    # en vez de hardcodearlo — un rotulo de periodo que miente es peor que no tenerlo.
    def _fecha(v):
        if v is None:
            return None
        if hasattr(v, "date"):
            return v.date().isoformat()
        s = str(v)[:10]
        return s if s[:4].isdigit() else None

    r0 = rows[0] if rows else ()
    out["cortes"] = {
        "CERVEZAS": _fecha(r0[2]) if len(r0) > 2 else None,
        "AGUAS DANONE": _fecha(r0[3]) if len(r0) > 3 else None,
    }

    # --- maestros para RECALCULAR la tasa desde los insumos frescos ---
    # Tasa% por (generico, zona): valor NEGOCIADO con CCU, tipeado a mano en
    # la columna D. No lo produce ningun sistema y cambia una vez por mes, asi
    # que leerlo de un .xlsm de hace unos dias no lo desactualiza.
    tasa_pct = {}
    for gen, ini, fin in [("CERVEZAS", 20, 23), ("AGUAS DANONE", 27, 30)]:
        for idx in range(ini - 1, fin):
            r = rows[idx]
            if r and r[0] and isinstance(r[3], (int, float)) and r[3]:
                tasa_pct[(gen, str(r[0]).strip())] = float(r[3])
    # Pisar con lo negociado. Se loguea cada cambio: una tasa que se mueve
    # sola y sin dejar rastro es la peor forma de equivocarse en este informe.
    for clave, pct in TASA_OVERRIDE.items():
        previo = tasa_pct.get(clave)
        if previo is not None and abs(previo - pct) < 1e-12:
            continue
        tasa_pct[clave] = pct
        anterior = f"{previo:.3%}" if previo is not None else "sin valor en el .xlsm"
        print(f"    tasa {clave[0]} / {clave[1]}: {anterior} -> {pct:.3%} "
              f"({TASA_OVERRIDE_FUENTE})")
    out["tasa_pct"] = tasa_pct

    # sucursal -> ZONA del modelo de tasa, del bloque superior de la hoja
    # (col A sucursal, col B zona) — es lo que agrupa el SUMIF del original.
    zona_por_suc = {}
    for idx in range(2, 16):
        r = rows[idx] if idx < len(rows) else None
        if r and r[0] and r[1]:
            zona_por_suc[str(r[0]).strip()] = str(r[1]).strip()
    out["zona_por_sucursal"] = zona_por_suc

    tasa = []
    for gen, ini, fin, tot in [("CERVEZAS", 20, 23, 24), ("AGUAS DANONE", 27, 30, 31)]:
        for i in range(ini - 1, fin):
            r = rows[i]
            if r and r[0]:
                desc, comp, pct, genr = float(r[1] or 0), float(r[2] or 0), float(r[3] or 0), float(r[5] or 0)
                tasa.append({"generico": gen, "zona": r[0], "descuentos": desc, "compras": comp,
                             "tasa_pct": pct, "tasa_generada": genr, "diferencia": genr - desc,
                             "compra_necesaria": (comp - desc / pct) if pct else 0})
        # 'Compra' (G) y 'Extra Tasa' (H) estan vacias todos los meses, y por eso
        # computar_tasa arma la Diferencia como Generada - Descuentos. Si alguien
        # las carga, esa Diferencia deja de ser la del original SIN avisar. Se
        # chequea aca, que es donde se ve el dato.
        for i in range(ini - 1, fin):
            rr = rows[i]
            extra = [v for v in (rr[6] if len(rr) > 6 else None,
                                 rr[7] if len(rr) > 7 else None)
                     if isinstance(v, (int, float)) and v]
            if extra:
                print(f"    ALERTA: '{gen} / {rr[0]}' tiene Compra/Extra Tasa cargada "
                      f"({extra}) en 'COMPRAS & DESC'. La Diferencia recalculada NO las "
                      f"suma — revisar antes de usar el cuadro.")
        rt = rows[tot - 1]
        tasa.append({"generico": gen, "zona": "— TOTAL " + gen, "descuentos": float(rt[1] or 0),
                     "compras": float(rt[2] or 0), "tasa_pct": None,
                     "tasa_generada": float(rt[5] or 0),
                     "diferencia": float(rt[5] or 0) - float(rt[1] or 0),
                     "compra_necesaria": float(rt[4] or 0)})
    out["tasa"] = pd.DataFrame(tasa)
    wb.close()

    aa = pd.read_excel(INFORME_XLSM, sheet_name="ART-ACCION", header=1)
    aa["Suma de Descuento"] = pd.to_numeric(aa["Suma de Descuento"], errors="coerce").fillna(0)
    aa = aa[aa["Acción"].notna()]
    # La columna derivada 'Es Reversa?' no siempre existe (el informe recalculado
    # del 30-jul la perdio). Fuente autoritativa: la hoja AccRever, col C.
    if "Es Reversa?" in aa.columns:
        aa["es_reversa"] = aa["Es Reversa?"].astype(str).str.upper().str.strip().eq("REVERSA")
    else:
        ar = pd.read_excel(INFORME_XLSM, sheet_name="AccRever")
        ar.columns = [str(c).strip() for c in ar.columns]
        flag_col = ar.columns[2]
        rev_set = set(
            ar.loc[ar[flag_col].astype(str).str.upper().str.strip().eq("REVERSA"), "Acción"].dropna()
        )
        aa["es_reversa"] = aa["Acción"].isin(rev_set)
        print(f"    reversa derivada de AccRever ({len(rev_set)} acciones marcadas)")
    out["reversa_map"] = dict(zip(aa["Acción"], aa["es_reversa"]))

    # El maestro AccRever NO cubre todas las acciones con movimiento. Las ausentes
    # caen en "NO reversa" por DEFECTO — eso es un supuesto, no un dato, y se
    # muestra en su propia fila para no inflar el "lo absorbe BADIE" con algo
    # que en realidad no esta clasificado.
    try:
        _ar = pd.read_excel(INFORME_XLSM, sheet_name="AccRever")
        _ar.columns = [str(c).strip() for c in _ar.columns]
        maestro = set(_ar["Acción"].dropna())
        rev_set = set(
            _ar.loc[_ar[_ar.columns[2]].astype(str).str.upper().str.strip().eq("REVERSA"), "Acción"].dropna()
        )
    except Exception:
        maestro, rev_set = set(aa.loc[aa.es_reversa, "Acción"]), set(aa.loc[aa.es_reversa, "Acción"])

    def _clase(acc):
        if acc in rev_set:
            return "SI reversa · lo reintegra CCU"
        if acc in maestro:
            return "NO reversa · confirmado en maestro"
        return "NO reversa · SUPUESTO (acción ausente del maestro)"

    aa["_clase"] = aa["Acción"].map(_clase)
    orden = ["SI reversa · lo reintegra CCU",
             "NO reversa · confirmado en maestro",
             "NO reversa · SUPUESTO (acción ausente del maestro)"]
    g = (aa.groupby("_clase")
           .agg(descuento=("Suma de Descuento", "sum"), codigos=("Acción", "nunique"))
           .reindex(orden).fillna(0).reset_index())
    out["reversa"] = pd.DataFrame({
        "tipo": g["_clase"],
        "descuento": g["descuento"].astype(float),
        "codigos": g["codigos"].astype(int),
    })
    out["reversa_cobertura"] = {
        "acciones_con_movimiento": int(aa["Acción"].nunique()),
        "en_maestro": int(aa["Acción"].isin(maestro).groupby(aa["Acción"]).first().sum()),
        "maestro_total": len(maestro),
    }

    # Padron cliente -> lista de precios. Es un MAESTRO (a que canal pertenece
    # cada cliente), no una cifra del periodo: leerlo del .xlsm no lo
    # desactualiza. Los descuentos que se cruzan contra el salen de la BASE
    # fresca, no de la hoja CLIENTE-FECHA del .xlsm — ver computar_canal.
    li = pd.read_excel(INFORME_XLSM, sheet_name="LISTA")
    li.columns = [str(c).strip() for c in li.columns]
    li = li[[li.columns[0], li.columns[1]]].dropna()
    li.columns = ["_cli", "LISTA PRECIO"]
    li["_cli"] = pd.to_numeric(li["_cli"], errors="coerce")
    li = li.dropna(subset=["_cli"]).drop_duplicates("_cli")
    out["lista_precio"] = li
    print(f"    padron LISTA: {len(li):,} clientes -> canal")

    # Fallback: si algo sale mal con la BASE, computar_canal degrada a esto.
    cf = pd.read_excel(INFORME_XLSM, sheet_name="CLIENTE-FECHA", header=2)
    cf["Suma de Descuento"] = pd.to_numeric(cf["Suma de Descuento"], errors="coerce").fillna(0)
    cf = cf[cf["Razón Social"].notna() & cf["Razón Social"].astype(str).ne("(blank)")]
    if "LISTA PRECIO" not in cf.columns:
        cf["_cli"] = pd.to_numeric(cf["Cod. Cliente"], errors="coerce")
        cf = cf.merge(li, on="_cli", how="left")
    cf["LISTA PRECIO"] = (cf["LISTA PRECIO"].fillna("SIN CANAL (#N/D)")
                          .map(etiquetar_lista))
    out["canal"] = (cf.groupby("LISTA PRECIO")["Suma de Descuento"]
                    .agg(descuento="sum", lineas="size").reset_index()
                    .rename(columns={"LISTA PRECIO": "canal"})
                    .sort_values("descuento", ascending=False))
    out["canal_gen"] = (cf.groupby(["LISTA PRECIO", "Calibre"])["Suma de Descuento"].sum()
                        .reset_index().rename(columns={"LISTA PRECIO": "canal", "Calibre": "generico",
                                                       "Suma de Descuento": "descuento"}))

    su = pd.read_excel(INFORME_XLSM, sheet_name="sucu", header=None)
    out["zona_map"] = {str(r[1]).strip(): str(r[3]).strip() for r in su.itertuples(index=False)
                       if len(r) > 3 and str(r[1]) != "nan"}

    # cliente interno (346110, 2348166, ...) -> sucursal. Es la clave con la
    # que compras.xls identifica la sucursal: VLOOKUP(Cliente, sucu!C:E, 3, 0).
    cli_suc = {}
    for r in su.itertuples(index=False):
        if len(r) < 5:
            continue
        try:
            cli_suc[int(float(r[2]))] = str(r[4]).strip()
        except (TypeError, ValueError):
            continue
    out["sucursal_por_cliente_interno"] = cli_suc

    # Tabla de signos por codigo de movimiento ('extra y tasa' cols I/J/K).
    # Los codigos ausentes (comodatos 615/616) daban #N/A en el VLOOKUP del
    # original y por eso el SUMIFS nunca los sumaba: aca se descartan igual.
    try:
        et = pd.read_excel(INFORME_XLSM, sheet_name="extra y tasa", header=None)
        signos = {}
        for r in et.itertuples(index=False):
            if len(r) < 11:
                continue
            cod, flag, val = r[8], r[9], r[10]
            if cod is None or str(cod) in ("nan", "tipo", "Comprobante"):
                continue
            try:
                signos[str(cod).strip()] = (float(val), str(flag).strip())
            except (TypeError, ValueError):
                continue
        out["signos_movimiento"] = signos
    except Exception as e:  # noqa: BLE001
        print(f"  WARN: 'extra y tasa' ilegible ({e})")

    h1 = pd.read_excel(INFORME_XLSM, sheet_name="Hoja1", header=None)
    out["sup_map"] = {str(r[0]).strip(): str(r[1]).strip() for r in h1.itertuples(index=False)
                      if len(r) > 1 and str(r[0]) != "nan"}
    return out


def load_gold():
    sys.path.insert(0, str(REPO))
    from src.core.data_loader import DataLoader
    dl = DataLoader()
    ventas = pd.read_sql(f"""
        SELECT ds.descripcion AS sucursal, da.generico, da.marca,
               SUM(fv.facturacion_neta) AS facturacion,
               SUM(fv.descuentos)       AS descuentos,
               SUM(fv.cantidades_total) AS bultos,
               SUM(fv.cantidad_total_htls) AS htls,
               COUNT(DISTINCT fv.id_cliente)  AS clientes,
               COUNT(DISTINCT fv.id_articulo) AS articulos
        FROM gold.fact_ventas fv
        JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
        JOIN gold.dim_sucursal ds ON fv.id_sucursal = ds.id_sucursal
        WHERE fv.fecha_comprobante BETWEEN '{PERIODO_DESDE}' AND '{PERIODO_HASTA}'
          AND fv.anulado = false
        GROUP BY 1,2,3""", dl.engine)
    cob = pd.read_sql("""
        SELECT ds_sucursal AS sucursal, generico, des_lista_precio AS lista,
               SUM(clientes_compradores) AS clientes_cob, SUM(volumen_total) AS volumen
        FROM gold.cob_sucursal_lista_generico
        WHERE periodo = '2026-07-01'
        GROUP BY 1,2,3""", dl.engine)
    # Evolucion diaria: facturacion y descuentos POR DIA, abiertos en CCU y
    # resto. Antes esta hoja salia del wapi, que trae SOLO las lineas con
    # accion promocional: la facturacion quedaba corta ($3.424 M contra los
    # $4.572 M reales de CCU en agosto-2026) y el % Desc. salia inflado
    # porque dividia el descuento por un subconjunto de la venta, no por la
    # venta. Los descuentos coincidian por las dos vias (0,04%), asi que lo
    # unico que faltaba era el denominador.
    # Clientes por sucursal al GRANO SUCURSAL. El `clientes` de `ventas` esta
    # contado por (sucursal, generico, marca): sumarlo cuenta al mismo cliente
    # una vez por cada marca que compro. Daba 59.595 contra 11.415 reales
    # (agosto-2026, 5,2x). La cobertura NO es aditiva entre marcas ni
    # genericos — hay que contarla en el corte donde se la quiere mostrar.
    cli_suc = pd.read_sql(f"""
        SELECT ds.descripcion AS sucursal,
               COUNT(DISTINCT fv.id_cliente) AS clientes
        FROM gold.fact_ventas fv
        JOIN gold.dim_sucursal ds ON fv.id_sucursal = ds.id_sucursal
        WHERE fv.fecha_comprobante BETWEEN '{PERIODO_DESDE}' AND '{PERIODO_HASTA}'
          AND fv.anulado = false
        GROUP BY 1""", dl.engine)

    ccu_sql = ", ".join(f"'{g}'" for g in sorted(CCU))
    diario = pd.read_sql(f"""
        SELECT fv.fecha_comprobante AS fecha,
               (da.generico IN ({ccu_sql})) AS es_ccu,
               SUM(fv.facturacion_neta) AS facturacion,
               SUM(fv.descuentos)       AS descuentos,
               COUNT(*)                 AS lineas
        FROM gold.fact_ventas fv
        JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
        WHERE fv.fecha_comprobante BETWEEN '{PERIODO_DESDE}' AND '{PERIODO_HASTA}'
          AND fv.anulado = false
        GROUP BY 1,2 ORDER BY 1,2""", dl.engine)
    # El mapa articulo -> generico sale del mismo cursor: lo necesita el
    # recalculo de la tasa y no justifica una segunda conexion.
    art = pd.read_sql("SELECT id_articulo, generico FROM gold.dim_articulo", dl.engine)
    return ventas, cob, dict(zip(art.id_articulo, art.generico)), diario, cli_suc


# ---------------- graficos matplotlib (lo que Excel no hace bien) ----------------
def make_corr_heatmap(mat: pd.DataFrame, path: Path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib as mpl
    import matplotlib.pyplot as plt
    import numpy as np

    corr = mat.corr(method="pearson")
    n = len(corr)
    fig, ax = plt.subplots(figsize=(1.05 * n + 2.2, 0.95 * n + 1.8), layout="constrained")
    # Diverging centrado en 0: el signo de la correlacion es la lectura principal.
    norm = mpl.colors.TwoSlopeNorm(vmin=-1, vcenter=0, vmax=1)
    cmap = mpl.colormaps["RdBu_r"].with_extremes(bad="#D9D2C4")
    im = ax.imshow(corr.values, cmap=cmap, norm=norm, interpolation="nearest")

    ax.set_xticks(range(n), corr.columns, rotation=42, ha="right", fontsize=8.5)
    ax.set_yticks(range(n), corr.index, fontsize=8.5)
    for i in range(n):
        for j in range(n):
            v = corr.values[i, j]
            if np.isnan(v):
                continue
            # Color redundante: ademas del fondo, el texto lleva el valor con signo.
            ax.text(j, i, f"{v:+.2f}", ha="center", va="center", fontsize=7.6,
                    color="white" if abs(v) > 0.62 else "#1A1612")
    ax.set_title(f"Correlación de Pearson entre métricas por sucursal  (n={len(mat)})",
                 fontsize=10.5, pad=11, color="#1A1612")
    cb = fig.colorbar(im, ax=ax, shrink=0.78)
    cb.set_label("r de Pearson", fontsize=8.5)
    cb.ax.tick_params(labelsize=8)
    for s in ax.spines.values():
        s.set_visible(False)
    fig.savefig(path, dpi=170, facecolor="white")
    plt.close(fig)
    return corr


def make_scatter_cob(df: pd.DataFrame, path: Path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(8.6, 5.4), layout="constrained")
    ccu = df[df.es_ccu]
    noc = df[~df.es_ccu]
    # Color + forma: no depender solo del color (accesibilidad).
    ax.scatter(noc.clientes_cob, noc.pct_desc, s=44, marker="s",
               c="#A2B9C4", edgecolor="#2C4A52", linewidth=0.8, label="No CCU", zorder=3)
    ax.scatter(ccu.clientes_cob, ccu.pct_desc, s=54, marker="o",
               c="#E8B4A8", edgecolor="#B8351C", linewidth=0.9, label="CCU", zorder=4)
    for _, r in df.nlargest(7, "pct_desc").iterrows():
        ax.annotate(r["etiqueta"], (r.clientes_cob, r.pct_desc), fontsize=7.2,
                    xytext=(4, 4), textcoords="offset points", color="#4A4035")
    ax.set_xlabel("Clientes con cobertura (gold.cob_sucursal_lista_generico)", fontsize=9)
    ax.set_ylabel("% Descuento  (Desc. / Fact.)", fontsize=9)
    ax.set_title("¿Más cobertura implica más presión de descuento?", fontsize=11, color="#1A1612")
    ax.yaxis.set_major_formatter(lambda v, p: f"{v*100:.0f}%")
    ax.grid(alpha=0.22, linewidth=0.6)
    ax.legend(frameon=False, fontsize=8.5)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    fig.savefig(path, dpi=170, facecolor="white")
    plt.close(fig)


# ---------------- construccion del workbook ----------------
# ─────────────────────────────────────────────────────────────────────────
# Modelo de tasa calculado desde los INSUMOS, no leido del .xlsm manual.
#
# Replica exacta de las formulas del informe original:
#
#   compras!AG articulo = LEFT(Productos, SEARCH("-")-1) * 1
#   compras!AH generico = VLOOKUP(articulo, art!A:F, 4, 0)
#   compras!AI Sucursal = VLOOKUP(Cliente, sucu!C:E, 3, 0)
#   compras!AJ Suma?    = VLOOKUP(Codigo Movimiento, 'extra y tasa'!I:J, 2, 0)
#   compras!AK Total    = VLOOKUP(Codigo Movimiento, 'extra y tasa'!I:K, 3, 0)
#                         * [Total Facturado Producto]
#
#   Compras = SUMIFS(AK, AH=generico, AI=sucursal, AJ="suma")
#   Tasa Generada    = Compras * Tasa%
#   Diferencia       = (Tasa + Reversa + Extra Tasa) - Descuentos
#   Compra Necesaria = Diferencia / Tasa%
#
# Los codigos de movimiento que NO estan en la tabla de signos (comodatos
# 615/616) daban #N/A en el VLOOKUP y por eso el SUMIFS nunca los sumaba.
# Aca se replica descartandolos, no asignandoles 0 — el efecto es el mismo
# pero la intencion queda explicita.
#
# La Tasa% se sigue leyendo del .xlsm porque es un valor NEGOCIADO con CCU
# que cambia una vez por mes: no lo produce ningun sistema. Leerlo de un
# archivo de hace unos dias no lo desactualiza.
# ─────────────────────────────────────────────────────────────────────────

_COMPRAS_HEADER_ROW = 3          # 0-indexed: fila 4 de Excel
_COL_COD_MOVIMIENTO = 3
_COL_FECHA_MOVIMIENTO = 9
_COL_PRODUCTOS = 11
_COL_TOTAL_FACTURADO = 14
_COL_CLIENTE = 23


# Cabecera de la fila 4 que identifica un export de compras. Se compara por
# contenido y no por nombre de archivo porque el ERP exporta con nombres
# distintos cada vez ("19-08-1.xls", "FC-NC_0 ...") y quedarse pegado a
# "compras.xls" hacia que el informe usara un archivo viejo en silencio.
_COMPRAS_FIRMA = ("Codigo Movimiento", "Fecha Movimiento", "Productos",
                  "Total Facturado Producto", "Cliente")


def _es_export_compras(path: Path) -> bool:
    import xlrd
    try:
        sh = xlrd.open_workbook(str(path)).sheet_by_index(0)
        if sh.nrows <= _COMPRAS_HEADER_ROW:
            return False
        fila = [str(sh.cell_value(_COMPRAS_HEADER_ROW, c)).strip()
                for c in (_COL_COD_MOVIMIENTO, _COL_FECHA_MOVIMIENTO, _COL_PRODUCTOS,
                          _COL_TOTAL_FACTURADO, _COL_CLIENTE)]
        return tuple(fila) == _COMPRAS_FIRMA
    except Exception:  # noqa: BLE001
        return False


def _resolver_compras(directorio: Path, canonico: Path) -> Path | None:
    """El export de compras mas RECIENTE del directorio, elegido por contenido.

    Devuelve None si no hay ninguno. Si el elegido no es el canonico se avisa,
    para que quede claro de que archivo salieron las Compras del informe.
    """
    candidatos = []
    for f in sorted(directorio.glob("*.xls")):
        if f.name.startswith("~$") or f.name.startswith("."):
            continue
        if _es_export_compras(f):
            candidatos.append(f)
    if not candidatos:
        return canonico if canonico.exists() else None
    elegido = max(candidatos, key=lambda f: f.stat().st_mtime)
    if elegido != canonico:
        otros = [f.name for f in candidatos if f != elegido]
        print(f"    compras: se usa {elegido.name} (el mas reciente)"
              + (f"; descartados por viejos: {', '.join(otros)}" if otros else ""))
    return elegido



def _leer_compras(path: Path, signos: dict, sucursal_por_cliente: dict,
                  generico_por_articulo: dict,
                  desde: str, hasta: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """compras.xls -> (modelo, crudo).

    `modelo` trae generico, sucursal y total con signo: es lo que consume el
    cuadro de tasa. `crudo` trae TODAS las filas del periodo, entren o no, con
    el motivo por el que cada una quedo afuera. Sin ese segundo DataFrame la
    unica forma de contestar "por que Compras da esto" es leer este codigo;
    con el, la respuesta esta en una hoja del propio informe.

    Se acota a [desde, hasta] por Fecha Movimiento. El SUMIFS del original NO
    filtra por fecha: confia en que el operador exporto justo el periodo. Si el
    export trae otro mes, el Excel manual cruza Compras de un periodo contra
    Descuentos de otro y la Diferencia no significa nada. Aca se filtra, asi
    que las dos columnas miden siempre la misma ventana.
    """
    import xlrd

    bk = xlrd.open_workbook(str(path))
    sh = bk.sheet_by_index(0)
    d0, d1 = date.fromisoformat(desde), date.fromisoformat(hasta)
    filas = []
    crudo = []
    fuera_de_rango = 0
    cubre_desde = cubre_hasta = None
    desconocidos: dict[str, float] = {}
    for i in range(_COMPRAS_HEADER_ROW + 1, sh.nrows):
        f = sh.cell_value(i, _COL_FECHA_MOVIMIENTO)
        if not isinstance(f, float) or f <= 0:
            continue                      # sin fecha no se puede ubicar en el periodo
        fecha = xlrd.xldate.xldate_as_datetime(f, bk.datemode).date()
        if not (d0 <= fecha <= d1):
            fuera_de_rango += 1
            continue
        cubre_desde = fecha if cubre_desde is None else min(cubre_desde, fecha)
        cubre_hasta = fecha if cubre_hasta is None else max(cubre_hasta, fecha)

        cod = str(sh.cell_value(i, _COL_COD_MOVIMIENTO)).strip()
        try:
            bruto = float(sh.cell_value(i, _COL_TOTAL_FACTURADO) or 0)
        except (TypeError, ValueError):
            bruto = 0.0
        productos = str(sh.cell_value(i, _COL_PRODUCTOS))
        # Fila del crudo: se completa a medida que se resuelve y se cierra con
        # el motivo. `fila_excel` es 1-based para poder ir al .xls original.
        reg = {"fila_excel": i + 1, "fecha": fecha.isoformat(), "cod_movimiento": cod,
               "cliente": None, "sucursal": None, "producto": productos,
               "articulo": None, "generico": None, "bruto": bruto,
               "signo": None, "importe": 0.0, "entra": "NO", "motivo": ""}

        def _descartar(motivo: str):
            reg["motivo"] = motivo
            crudo.append(reg)

        if cod not in signos:
            # #N/A en el VLOOKUP del original, asi que el SUMIFS no lo suma.
            # Se replica, pero se MIDE: un codigo nuevo con plata importante
            # tiene que aparecer en pantalla, no desaparecer sin ruido.
            desconocidos[cod] = desconocidos.get(cod, 0.0) + abs(bruto)
            _descartar("codigo sin signo en 'extra y tasa' (#N/A en el original)")
            continue
        signo, flag = signos[cod]
        reg["signo"] = signo
        if flag.strip().lower() != "suma":
            _descartar(f"'extra y tasa' marca el codigo como '{flag.strip()}', no 'suma'")
            continue                      # el SUMIFS filtra por AJ="suma"

        if "-" not in productos:
            _descartar("Productos no tiene el formato '<articulo>-<descripcion>'")
            continue
        try:
            articulo = int(float(productos.split("-", 1)[0].strip()))
        except ValueError:
            _descartar("el codigo de articulo de Productos no es numerico")
            continue
        reg["articulo"] = articulo

        cliente = sh.cell_value(i, _COL_CLIENTE)
        try:
            cliente = int(float(cliente))
        except (TypeError, ValueError):
            _descartar("Cliente vacio o no numerico")
            continue
        reg["cliente"] = cliente

        gen = generico_por_articulo.get(articulo)
        suc = sucursal_por_cliente.get(cliente)
        reg["generico"], reg["sucursal"] = gen, suc
        if gen is None or suc is None:
            falta = "articulo sin generico en gold" if gen is None else "cliente sin sucursal en 'sucu'"
            _descartar(f"{falta} (#N/A en el original)")
            continue                      # sin match -> el VLOOKUP daba #N/A

        importe = signo * bruto
        reg["importe"], reg["entra"], reg["motivo"] = importe, "SI", "entra al modelo"
        crudo.append(reg)
        filas.append({"generico": gen, "sucursal": suc, "total": importe})

    if fuera_de_rango:
        print(f"    {path.name}: {fuera_de_rango} filas fuera de {desde}..{hasta}, descartadas")
    if desconocidos:
        sumado = sum(abs(f["total"]) for f in filas)
        for cod, monto in sorted(desconocidos.items(), key=lambda kv: -kv[1]):
            peso = (monto / sumado) if sumado else 0.0
            nivel = "ALERTA" if peso >= 0.05 else "aviso"
            print(f"    {nivel}: codigo de movimiento '{cod}' sin signo en "
                  f"'extra y tasa' — ${monto:,.0f} ({peso:.1%} de las Compras) "
                  f"NO entran en el modelo, igual que en el informe original")
    if cubre_desde is None:
        print(f"    WARN: {path.name} no tiene NINGUNA fila en {desde}..{hasta}; "
              f"la tasa sale del .xlsm")
    elif cubre_desde > d0 or cubre_hasta < d1:
        print(f"    WARN: {path.name} cubre {cubre_desde}..{cubre_hasta} pero el "
              f"periodo es {desde}..{hasta}. Las Compras quedan CORTAS contra los "
              f"Descuentos: la Diferencia subestima. Reexportar el archivo.")
    return pd.DataFrame(filas), pd.DataFrame(crudo)


def computar_tasa(compras_path: Path, inf: dict, fact: pd.DataFrame,
                  generico_por_articulo: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame] | None:
    """Recalcula el modelo de tasa desde los insumos frescos.

    Descuentos salen de la BASE (frescos). Compras de compras.xls (fresco).
    Solo la Tasa% viene del .xlsm, porque es un valor negociado, no calculado.
    Devuelve (modelo_por_zona, apertura_por_sucursal, compras_crudo), o None si
    falta algun insumo — el informe cae al bloque leido del .xlsm, que es
    preferible a inventar."""
    signos = inf.get("signos_movimiento")
    suc_por_cliente = inf.get("sucursal_por_cliente_interno")
    tasas = inf.get("tasa_pct")           # {(generico, zona): pct}
    zona_por_suc = inf.get("zona_por_sucursal")
    if not (signos and suc_por_cliente and tasas and zona_por_suc):
        return None
    if not compras_path.exists():
        print(f"  WARN: no esta {compras_path.name}; la tasa sale del .xlsm")
        return None

    compras, crudo = _leer_compras(compras_path, signos, suc_por_cliente,
                                   generico_por_articulo, PERIODO_DESDE, PERIODO_HASTA)
    if compras.empty:
        return None
    if not crudo.empty:
        crudo["zona"] = crudo["sucursal"].map(zona_por_suc)

    # Compras por (generico, zona) — el original agrega por sucursal y despues
    # suma por zona con SUMIF; aca se hace en un paso, mismo resultado.
    compras["zona"] = compras["sucursal"].map(zona_por_suc)
    compras_zona = (compras.dropna(subset=["zona"])
                    .groupby(["generico", "zona"])["total"].sum())

    # Descuentos por (generico, zona) desde la BASE.
    fz = fact.copy()
    fz["zona"] = fz["sucursal"].map(zona_por_suc)
    desc_zona = (fz.dropna(subset=["zona"])
                 .groupby(["generico", "zona"])["descuentos"].sum())

    # Se recorre `tasas` en orden de insercion (CERVEZAS y despues AGUAS, cada
    # una con sus zonas en el orden de la hoja) y el TOTAL de cada generico se
    # intercala al cierre de su bloque. Es el mismo layout que produce el
    # camino que lee el .xlsm: la hoja no puede reordenarse segun cual de los
    # dos corrio.
    filas = []
    bloque = []
    gen_actual = None

    def _cerrar_bloque():
        if not bloque:
            return
        filas.extend(bloque)
        filas.append({
            "generico": gen_actual, "zona": "— TOTAL " + gen_actual,
            "descuentos": sum(f["descuentos"] for f in bloque),
            "compras": sum(f["compras"] for f in bloque),
            "tasa_pct": None,
            "tasa_generada": sum(f["tasa_generada"] for f in bloque),
            "diferencia": sum(f["diferencia"] for f in bloque),
            "compra_necesaria": sum(f["compra_necesaria"] for f in bloque),
        })
        bloque.clear()

    for (gen, zona), pct in tasas.items():
        if gen != gen_actual:
            _cerrar_bloque()
            gen_actual = gen
        comp = float(compras_zona.get((gen, zona), 0.0))
        desc = float(desc_zona.get((gen, zona), 0.0))
        generada = comp * pct
        diferencia = generada - desc          # Reversa y Extra Tasa vienen vacias
        bloque.append({
            "generico": gen, "zona": zona,
            "descuentos": desc, "compras": comp,
            "tasa_pct": pct, "tasa_generada": generada,
            "diferencia": diferencia,
            "compra_necesaria": (diferencia / pct) if pct else 0.0,
        })
    _cerrar_bloque()

    if not filas:
        return None

    # Apertura por SUCURSAL — el bloque superior de 'COMPRAS & DESC', que es
    # de donde el original saca las cifras de zona con un SUMIF. Sin el, la
    # unica forma de ver que sucursal mueve una zona es abrir la BASE.
    # El universo de filas son las 14 sucursales del mapa sucursal->zona, no
    # las que tengan movimiento: una sucursal en cero es informacion (el
    # original la lista con 0), no una fila que sobra.
    compras_suc = compras.groupby(["generico", "sucursal"])["total"].sum()
    desc_suc = fact.groupby(["generico", "sucursal"])["descuentos"].sum()
    orden_suc = sorted(zona_por_suc, key=lambda s: (zona_por_suc[s], s))
    apertura = []
    for gen in dict.fromkeys(g for g, _ in tasas):
        bloque_suc = []
        for suc in orden_suc:
            bloque_suc.append({
                "generico": gen, "sucursal": suc, "zona": zona_por_suc[suc],
                "descuentos": float(desc_suc.get((gen, suc), 0.0)),
                "compras": float(compras_suc.get((gen, suc), 0.0)),
            })
        apertura.extend(bloque_suc)
        apertura.append({
            "generico": gen, "sucursal": "— TOTAL " + gen, "zona": "",
            "descuentos": sum(f["descuentos"] for f in bloque_suc),
            "compras": sum(f["compras"] for f in bloque_suc),
        })

    return pd.DataFrame(filas), pd.DataFrame(apertura), crudo


def _hoja_compras_crudo(wb, inf: dict, layout: dict) -> None:
    """Hoja con TODAS las filas de compras.xls del periodo, entren o no.

    Es la contraparte de 'Datos' para el otro lado del modelo: el Modelo Tasa
    saca de aca la columna Compras con un SUMIFS, y la columna Motivo explica
    fila por fila por que una compra no sumo. Sin esto, los codigos que quedan
    afuera (255, 604, 615, 616...) son un renglon en la consola que nadie ve.
    """
    crudo = inf.get("compras_crudo")
    ws = wb.create_sheet("Compras (crudo)")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 9, "C": 11, "D": 30, "E": 11, "F": 30, "G": 18,
                "H": 15, "I": 10, "J": 40, "K": 16, "L": 7, "M": 16, "N": 8, "O": 46})
    archivo = inf.get("compras_archivo", "compras.xls")
    r = title(ws, 2, "Compras — filas crudas",
              f"Fuente C · {archivo} · período {_label_periodo()} · "
              "de acá sale la columna Compras del Modelo Tasa")
    if crudo is None or not len(crudo):
        ws.cell(row=r, column=2, value="No se pudo leer el export de compras.").font = Font(
            size=10, italic=True, color=BADIE)
        return

    ws.cell(row=r, column=2, value=(
        "Una fila suma al modelo sólo si ¿Entra? = SI. El resto queda con el motivo escrito "
        "al lado: es exactamente lo que el informe manual descarta como #N/A, pero acá se ve."
    )).font = Font(size=9, italic=True, color=BADIE)
    r += 2

    df = crudo.copy()
    if "zona" not in df.columns:
        df["zona"] = None
    df = df.rename(columns={
        "fila_excel": "Fila .xls", "fecha": "Fecha", "cod_movimiento": "Cód. Movimiento",
        "cliente": "Cliente", "sucursal": "Sucursal", "zona": "Zona",
        "generico": "Genérico", "articulo": "Artículo", "producto": "Producto",
        "bruto": "Total facturado", "signo": "Signo", "importe": "Importe (con signo)",
        "entra": "¿Entra?", "motivo": "Motivo"})
    cols = ["Fila .xls", "Fecha", "Cód. Movimiento", "Cliente", "Sucursal", "Zona",
            "Genérico", "Artículo", "Producto", "Total facturado", "Signo",
            "Importe (con signo)", "¿Entra?", "Motivo"]
    df = df[[c for c in cols if c in df.columns]]
    start = r
    r = write_df(ws, df, r, {"Total facturado": F_MONEY, "Importe (con signo)": F_MONEY,
                             "Cliente": F_INT, "Artículo": F_INT, "Fila .xls": F_INT,
                             "Signo": F_INT},
                 start_col=2, table_name="tblComprasCrudo", zebra=False,
                 no_total=("Cliente", "Artículo", "Fila .xls", "Signo"))
    ws.freeze_panes = ws.cell(row=start + 1, column=2)
    ws.auto_filter.ref = f"B{start}:O{r-2}"
    # B=Fila C=Fecha D=Cod E=Cliente F=Sucursal G=Zona H=Generico I=Articulo
    # J=Producto K=Total facturado L=Signo M=Importe N=¿Entra? O=Motivo
    layout["compras"] = (start + 1, r - 2)


def _hoja_wapi_crudo(wb, wapi: pd.DataFrame) -> None:
    """Hoja con el wapi tal como llega, recortado a las columnas que se usan.

    Es la fuente de la hoja «Evolución Diaria» y de los Descuentos de la BASE.
    Van las 13 columnas que explican algun numero del informe; las otras 19 del
    export no las lee nadie y multiplicarian el peso del archivo.
    """
    ws = wb.create_sheet("wapi (crudo)")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 11, "C": 14, "D": 11, "E": 34, "F": 11, "G": 34,
                "H": 18, "I": 10, "J": 34, "K": 24, "L": 16, "M": 14, "N": 16})
    r = title(ws, 2, "wapi — filas crudas",
              "Fuente A · hoja wapi de la BASE control · SÓLO líneas con acción CCU · "
              "de acá sale «Evolución Diaria»")
    if wapi is None or not len(wapi):
        ws.cell(row=r, column=2, value="La BASE no trajo filas de wapi.").font = Font(
            size=10, italic=True, color=BADIE)
        return
    ws.cell(row=r, column=2, value=(
        f"{len(wapi):,} filas. Es el archivo que exporta CCU, no la venta completa de BADIE: "
        "por eso la Facturación de esta hoja no coincide con «Genéricos»."
    )).font = Font(size=9, italic=True, color=BADIE)
    r += 2
    start = r
    r = write_df(ws, wapi, r, {"Total2": F_MONEY, "Descuento": F_MONEY,
                               "Cod. Cliente": F_INT, "Artículo CMQ": F_INT},
                 start_col=2, table_name="tblWapiCrudo", zebra=False,
                 no_total=("Cod. Cliente", "Artículo CMQ"))
    ws.freeze_panes = ws.cell(row=start + 1, column=2)
    ultima = get_column_letter(1 + len(wapi.columns))
    ws.auto_filter.ref = f"B{start}:{ultima}{r-2}"


def _formulas_modelo_tasa(wb, inf: dict, layout: dict) -> None:
    """Reescribe el cuadro de tasa con formulas que apuntan a los datos crudos.

    Cadena completa: la fila de sucursal es un SUMIFS contra 'Datos' y
    'Compras (crudo)'; la fila de zona es la SUMA de sus sucursales; el TOTAL
    del generico es la suma de sus zonas. Es la misma estructura que el .xlsm
    original (que agrega por sucursal y sube a zona con SUMIF), con la
    diferencia de que aca las hojas de origen viajan dentro del archivo.

    No se toca ningun valor: cada formula tiene que dar exactamente el numero
    que ya estaba. Si no da, es un bug de la formula, no un criterio nuevo.
    """
    if not layout.get("tasa_rows") or not layout.get("apertura"):
        return
    if not layout.get("datos") or not layout.get("compras"):
        return                              # sin hojas crudas no hay a que apuntar

    ws = wb["Modelo Tasa"]
    d0, d1 = layout["datos"]
    c0, c1 = layout["compras"]
    DAT = f"'Datos'!$B${d0}:$B${d1}"        # Sucursal
    DAT_GEN = f"'Datos'!$I${d0}:$I${d1}"    # Genérico
    DAT_DESC = f"'Datos'!$K${d0}:$K${d1}"   # Descuentos
    CMP = f"'Compras (crudo)'!$F${c0}:$F${c1}"        # Sucursal
    CMP_GEN = f"'Compras (crudo)'!$H${c0}:$H${c1}"    # Genérico
    CMP_IMP = f"'Compras (crudo)'!$M${c0}:$M${c1}"    # Importe con signo
    CMP_ENT = f"'Compras (crudo)'!$N${c0}:$N${c1}"    # ¿Entra?

    # --- apertura por sucursal: el piso de la cadena ---
    for gen, (ini, fin) in layout["apertura"].items():
        for rr in range(ini, fin + 1):
            suc = ws.cell(row=rr, column=2).value
            if isinstance(suc, str) and suc.startswith("—"):
                ws.cell(row=rr, column=4, value=f"=SUM(D{ini}:D{rr - 1})")
                ws.cell(row=rr, column=5, value=f"=SUM(E{ini}:E{rr - 1})")
                continue
            ws.cell(row=rr, column=4,
                    value=f'=SUMIFS({DAT_DESC},{DAT},$B{rr},{DAT_GEN},"{gen}")')
            ws.cell(row=rr, column=5,
                    value=f'=SUMIFS({CMP_IMP},{CMP},$B{rr},{CMP_GEN},"{gen}",{CMP_ENT},"SI")')

    # --- cuadro por zona: suma de las sucursales de esa zona ---
    zona_col = {}                            # (gen, zona) -> rango de filas de apertura
    for gen, (ini, fin) in layout["apertura"].items():
        for rr in range(ini, fin + 1):
            z = ws.cell(row=rr, column=3).value
            if isinstance(z, str) and z:
                zona_col.setdefault((gen, z), []).append(rr)

    pendiente_total = []
    for row, gen, zona, es_total in layout["tasa_rows"]:
        if es_total:
            if pendiente_total:
                p0, p1 = pendiente_total[0], pendiente_total[-1]
                for col in (4, 5, 7, 8, 9):
                    letra = get_column_letter(col)
                    ws.cell(row=row, column=col, value=f"=SUM({letra}{p0}:{letra}{p1})")
                pendiente_total = []
            continue
        pendiente_total.append(row)
        filas = zona_col.get((gen, zona), [])
        if filas:
            refs_d = ",".join(f"D{x}" for x in filas)
            refs_e = ",".join(f"E{x}" for x in filas)
            ws.cell(row=row, column=4, value=f"=SUM({refs_d})")
            ws.cell(row=row, column=5, value=f"=SUM({refs_e})")
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}")          # Tasa generada
        ws.cell(row=row, column=8, value=f"=G{row}-D{row}")          # Diferencia
        ws.cell(row=row, column=9,                                    # Compra necesaria
                value=f'=IF(F{row}=0,0,H{row}/F{row})')
        ws.cell(row=row, column=10, value=f'=IF(H{row}>=0,"SI","NO")')

    # El pie del cuadro deja escrito a que hoja apunta cada columna: si alguien
    # abre el archivo en seis meses, no tiene que reconstruirlo leyendo formulas.
    fin = max(x[0] for x in layout["tasa_rows"]) + 1
    ws.cell(row=fin, column=2, value=(
        "Descuentos → hoja «Datos» (FACT_NET)   ·   Compras → hoja «Compras (crudo)», "
        "sólo ¿Entra?=SI   ·   Tasa % → informe manual / informada por CCU"
    )).font = Font(name="Consolas", size=8, color=TEAL)


def computar_canal(cli: pd.DataFrame, inf: dict) -> tuple | None:
    """Descuento por canal desde la BASE (fresca), no desde el .xlsm.

    La hoja CLIENTE-FECHA existe en los dos archivos con el MISMO layout, y
    leerla del .xlsm era el bug: ese archivo se refresca a mano, asi que la
    hoja Canal se quedaba en el corte del ultimo refresh mientras el resto del
    informe avanzaba. Del .xlsm se sigue tomando solo el padron LISTA
    (cliente -> canal), que es un maestro y no una cifra del periodo.

    Devuelve (canal, canal_gen) o None si falta algun insumo — degradar al
    bloque del .xlsm es preferible a inventar una apertura.
    """
    li = inf.get("lista_precio")
    if li is None or cli is None or cli.empty:
        return None

    cf = cli.copy()
    cf = cf[cf["razon_social"].notna() & cf["razon_social"].astype(str).ne("(blank)")]
    if cf.empty:
        return None
    cf["_cli"] = pd.to_numeric(cf["cod_cliente"], errors="coerce")
    cf = cf.merge(li, on="_cli", how="left")
    cf["LISTA PRECIO"] = (cf["LISTA PRECIO"].fillna("SIN CANAL (#N/D)")
                          .map(etiquetar_lista))

    canal = (cf.groupby("LISTA PRECIO")["descuento"]
             .agg(descuento="sum", lineas="size").reset_index()
             .rename(columns={"LISTA PRECIO": "canal"})
             .sort_values("descuento", ascending=False))
    canal_gen = (cf.groupby(["LISTA PRECIO", "generico"])["descuento"].sum()
                 .reset_index().rename(columns={"LISTA PRECIO": "canal"}))
    return canal, canal_gen


def _origen_tasa(inf: dict) -> str:
    """Rotula de donde salio cada Tasa%: del .xlsm o declarada aca.

    Con dos origenes conviviendo, decir 'del informe manual' a secas seria
    falso para las zonas pisadas por TASA_OVERRIDE.
    """
    tasas = inf.get("tasa_pct") or {}
    pisadas = [z for (g, z) in TASA_OVERRIDE if (g, z) in tasas]
    if not pisadas:
        return "del informe manual"
    if len(pisadas) == len(tasas):
        return TASA_OVERRIDE_FUENTE
    return f"{len(pisadas)} de {len(tasas)} zonas {TASA_OVERRIDE_FUENTE}; el resto del informe manual"


def _label_periodo() -> str:
    """'01→30 jul' derivado de las constantes del run, nunca hardcodeado."""
    MESES = ["ene", "feb", "mar", "abr", "may", "jun",
             "jul", "ago", "sep", "oct", "nov", "dic"]
    d0, d1 = PERIODO_DESDE, PERIODO_HASTA
    m0, m1 = int(d0[5:7]), int(d1[5:7])
    if m0 == m1:
        return f"{d0[8:10]}→{d1[8:10]} {MESES[m1 - 1]}"
    return f"{d0[8:10]} {MESES[m0 - 1]}→{d1[8:10]} {MESES[m1 - 1]}"


def _label_mes_anio() -> str:
    MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return f"{MESES[int(PERIODO_HASTA[5:7]) - 1]} {PERIODO_HASTA[:4]}"


def _label_corte(cortes: dict) -> str:
    """Rotulo del corte de la fuente B, derivado del archivo (nunca hardcodeado)."""
    cs = {k: v for k, v in (cortes or {}).items() if v}
    if not cs:
        return "corte no declarado en el archivo"
    vals = set(cs.values())
    if len(vals) == 1:
        return f"corte {vals.pop()}"
    return " / ".join(f"{v} ({k.split()[0].lower()})" for k, v in cs.items())


def build(out_path: Path):
    IMG_DIR.mkdir(exist_ok=True)
    print("Cargando BASE ...")
    fact, acc, daily, cli, accgen, wapi_crudo = load_base()

    # El periodo se DERIVA del dato real de la BASE, no de constantes ni de
    # argumentos: si el rotulo y el contenido pueden desincronizarse, tarde o
    # temprano lo hacen. Con esto el encabezado siempre describe lo que hay.
    global PERIODO_DESDE, PERIODO_HASTA
    if len(daily):
        PERIODO_DESDE = str(daily["fecha"].min())
        PERIODO_HASTA = str(daily["fecha"].max())
        print(f"Periodo derivado de la BASE: {PERIODO_DESDE} -> {PERIODO_HASTA}")
    print("Cargando INFORME ...")
    inf = load_informe()

    print("Cargando gold ...")
    ventas, cob, generico_por_articulo, diario_gold, clientes_suc = load_gold()

    # Modelo de tasa recalculado desde los insumos. Si algo falta, se cae al
    # bloque leido del .xlsm — mejor un dato viejo declarado que uno inventado.
    try:
        _compras_path = _resolver_compras(COMPRAS_DIR, COMPRAS_XLS)
        if _compras_path is None:
            raise FileNotFoundError(f"no hay ningun export de compras en {COMPRAS_DIR}")
        _res = computar_tasa(_compras_path, inf, fact, generico_por_articulo)
        _tasa, _apertura, _crudo = _res if _res is not None else (None, None, None)
        if _tasa is not None and len(_tasa):
            inf["tasa"] = _tasa
            inf["tasa_apertura"] = _apertura
            inf["compras_crudo"] = _crudo
            inf["compras_archivo"] = _compras_path.name
            inf["tasa_calculada"] = True
            print(f"    tasa RECALCULADA desde {_compras_path.name} + BASE")
        else:
            print("    tasa leida del .xlsm (no se pudo recalcular)")
    except Exception as e:  # noqa: BLE001
        print(f"    WARN: no se pudo recalcular la tasa ({e}); se usa la del .xlsm")

    # Canal desde la BASE fresca. El .xlsm solo aporta el padron LISTA.
    try:
        _c = computar_canal(cli, inf)
        if _c is not None:
            _viejo = float(inf["canal"].descuento.sum())
            inf["canal"], inf["canal_gen"] = _c
            _nuevo = float(inf["canal"].descuento.sum())
            inf["canal_calculado"] = True
            print(f"    canal RECALCULADO desde la BASE: ${_nuevo:,.0f} "
                  f"(el .xlsm traia ${_viejo:,.0f})")
        else:
            print("    canal leido del .xlsm (no se pudo recalcular)")
    except Exception as e:  # noqa: BLE001
        print(f"    WARN: no se pudo recalcular el canal ({e}); se usa el del .xlsm")

    wb = Workbook()
    wb.remove(wb.active)

    # ============ 1. PORTADA ============
    ws = wb.create_sheet("Portada")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 34, "C": 22, "D": 22, "E": 22, "F": 22, "G": 26})
    ws["B2"] = "ACCIONES COMERCIALES"
    ws["B2"].font = Font(name="Calibri", size=28, bold=True, color=INK)
    ws["B3"] = f"Informe analítico · {_label_mes_anio()} · BADIE S.A."
    ws["B3"].font = Font(name="Calibri", size=12, italic=True, color=BADIE)
    ws["B4"] = f"Generado {datetime.now():%d/%m/%Y %H:%M}"
    ws["B4"].font = Font(name="Calibri", size=9, color=INK_MUTE)

    r = band(ws, 6, "FUENTES  ·  alcances distintos, no sumar entre sí", col=2, span=6, fill=WARN)
    fuentes = pd.DataFrame([
        {"Fuente": "A · BASE control", "Corte": _label_periodo(), "Alcance": "Todos los genéricos",
         "Aporta": "Facturación, descuentos, acciones, clientes, serie diaria"},
        {"Fuente": "B · Informe final", "Corte": _label_corte(inf.get("cortes")).replace("corte ", ""), "Alcance": "Sólo genéricos CCU",
         "Aporta": "Modelo de tasa, reversa, canal, maestros zona/supervisor"},
        {"Fuente": "C · gold (Postgres)", "Corte": _label_periodo(), "Alcance": "Todos los genéricos",
         "Aporta": "Bultos, HTLs, clientes/artículos únicos, cobertura"},
    ])
    r = write_df(ws, fuentes, r, {}, total_row=False, start_col=2)
    ws.cell(row=r, column=2, value="Cada hoja declara de qué fuente sale. Los totales entre fuentes NO coinciden y no deben sumarse.").font = Font(size=9, italic=True, color=BADIE)

    r += 2
    r = band(ws, r, "ÍNDICE", col=2, span=6, fill=TEAL)
    idx = [
        ("Resumen", "KPIs del período y hallazgos principales", "A+B"),
        ("Modelo Tasa", "Compras vs reintegro CCU · el P&L real", "B"),
        ("Evolución Diaria", "Serie facturación / descuentos / % desc", "A"),
        ("Genéricos", "Mix y presión promocional por genérico", "A"),
        ("Sucursales", "Ranking con zona y supervisor", "A+B+C"),
        ("Canal", "Descuento por lista de precios", "B"),
        ("Códigos de Acción", "Top acciones + flag reversa", "A+B"),
        ("Matriz Suc x Gen", "Heatmap de descuento cruzado", "A"),
        ("Correlaciones", "Matriz de Pearson entre métricas", "A+C"),
        ("Cobertura", "Cobertura gold vs presión de descuento", "C"),
        ("Tablero CCU", "Réplica del TABLERO original · sólo CCU", "A"),
        ("Acción x Genérico", "Réplica de ACC-GEN", "A"),
        ("Reversa Clientes", "Réplica de Reversa · categorías manuales", "B"),
        ("Clientes", "Clientes con acción y artículo", "A+B"),
        ("Datos", "Tablas base para tabla dinámica", "A"),
    ]
    hdr(ws, r, ["Hoja", "Contenido", "Fuente"], start=2)
    r += 1
    for name, desc, src in idx:
        c = ws.cell(row=r, column=2, value=name)
        c.font = Font(name="Calibri", size=10, bold=True, color=TEAL, underline="single")
        c.hyperlink = f"#'{name}'!A1"
        ws.cell(row=r, column=3, value=desc).font = Font(size=9, color=INK_SOFT)
        ws.cell(row=r, column=6, value=src).font = Font(name="Consolas", size=8, color=INK_MUTE)
        r += 1

    # ============ 2. RESUMEN ============
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20})
    r = title(ws, 2, "Resumen ejecutivo", f"Período {PERIODO_DESDE} → {PERIODO_HASTA}")

    tot_f, tot_d = fact.facturacion.sum(), fact.descuentos.sum()
    tasa_tot = inf["tasa"][inf["tasa"].zona.str.startswith("—")]
    kpis = [
        ("Facturación neta", tot_f, F_MONEY, "fuente A"),
        ("Descuentos otorgados", tot_d, F_MONEY, "fuente A"),
        ("% Descuento (D/F)", tot_d / tot_f, F_PCT, "fuente A"),
        ("Compras a CCU", tasa_tot.compras.sum(), F_MONEY, "fuente B"),
        ("Tasa generada", tasa_tot.tasa_generada.sum(), F_MONEY, "fuente B"),
        ("Diferencia (tasa − desc.)", tasa_tot.diferencia.sum(), F_MONEY0, "fuente B"),
    ]
    r = band(ws, r, "INDICADORES CLAVE", col=2, span=6)
    for i, (lbl, val, fmt, src) in enumerate(kpis):
        col = 2 + (i % 3) * 2
        row = r + (i // 3) * 4
        c = ws.cell(row=row, column=col, value=lbl)
        c.font = Font(size=9, bold=True, color=INK_MUTE)
        v = ws.cell(row=row + 1, column=col, value=val)
        neg = isinstance(val, (int, float)) and val < 0
        v.font = Font(size=19, bold=True, color=BADIE if neg else INK)
        v.number_format = fmt
        s = ws.cell(row=row + 2, column=col, value=src)
        s.font = Font(name="Consolas", size=7.5, color=INK_MUTE)
    r += 9

    r = band(ws, r, "HALLAZGOS", col=2, span=6, fill=BADIE)
    dif = tasa_tot.diferencia.sum()
    rev = inf["reversa"]
    no_rev = rev.loc[rev.tipo.str.startswith("NO"), "descuento"].iloc[0]
    pct_no_rev = no_rev / rev.descuento.sum()
    peor = inf["tasa"][~inf["tasa"].zona.str.startswith("—")].nsmallest(1, "diferencia").iloc[0]
    canal_top = inf["canal"].iloc[0]
    sin_canal = inf["canal"][inf["canal"].canal.str.contains("#N/D")]
    gen_top = fact.groupby("generico").descuentos.sum().nlargest(1)
    hall = [
        f"La tasa que reintegra CCU NO cubre los descuentos: diferencia de {dif:,.0f} en el período (fuente B).",
        f"El {pct_no_rev:.0%} del descuento ({no_rev:,.0f}) NO es reversible: lo absorbe BADIE (fuente B).",
        f"Zona más deficitaria: {peor.zona} / {peor.generico} con {peor.diferencia:,.0f} (fuente B).",
        f"Canal con mayor descuento: {canal_top.canal} con {canal_top.descuento:,.0f} (fuente B).",
        (f"Hueco de datos: {sin_canal.descuento.iloc[0]:,.0f} de descuento sin canal asignado (#N/D)."
         if len(sin_canal) else "Todos los descuentos tienen canal asignado."),
        f"Genérico que más descuento concentra: {gen_top.index[0]} con {gen_top.iloc[0]:,.0f} (fuente A).",
        "Cupos del período NO disponibles en gold.fact_cupos → no se pudo cruzar objetivo vs real.",
    ]
    for h in hall:
        c = ws.cell(row=r, column=2, value="•  " + h)
        c.font = Font(size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 17
        r += 1

    # ============ 3. MODELO TASA ============
    ws = wb.create_sheet("Modelo Tasa")
    layout: dict = {}          # coordenadas para el pase de formulas del final
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 20, "C": 22, "D": 16, "E": 16, "F": 11, "G": 16, "H": 17, "I": 19})
    r = title(ws, 2, "Modelo de tasa — ¿la inversión promocional se paga sola?",
              ("Descuentos y Compras RECALCULADOS de los insumos · "
               f"Tasa % {_origen_tasa(inf)} · sólo CCU · período {_label_periodo()}"
               if inf.get("tasa_calculada")
               else f"Fuente B · informe final · sólo CCU · {_label_corte(inf.get('cortes'))}"))
    ws.cell(row=r, column=2, value="Tasa Generada = Compras × Tasa%     ·     Diferencia = Tasa Generada − Descuentos     ·     Compra Necesaria = Compras − (Descuentos / Tasa%)").font = Font(name="Consolas", size=8.5, color=TEAL)
    r += 2

    t = inf["tasa"].copy()
    t["cubre"] = t.diferencia.apply(lambda v: "SI" if v >= 0 else "NO")
    tdf = t.rename(columns={"generico": "Genérico", "zona": "Zona", "descuentos": "Descuentos",
                            "compras": "Compras", "tasa_pct": "Tasa %", "tasa_generada": "Tasa generada",
                            "diferencia": "Diferencia", "compra_necesaria": "Compra necesaria",
                            "cubre": "¿Cubre?"})[
        ["Genérico", "Zona", "Descuentos", "Compras", "Tasa %", "Tasa generada", "Diferencia", "Compra necesaria", "¿Cubre?"]]
    start = r
    r = write_df(ws, tdf, r, {"Descuentos": F_MONEY, "Compras": F_MONEY, "Tasa %": F_PCT,
                              "Tasa generada": F_MONEY, "Diferencia": F_MONEY0,
                              "Compra necesaria": F_MONEY0}, total_row=False, start_col=2)
    # Coordenadas del cuadro, para reescribirlo con formulas una vez que existan
    # las hojas a las que apuntan. Se guarda la fila de cada (generico, zona) y
    # si es un TOTAL, que se resuelve con SUM del bloque y no con SUMIFS.
    layout["tasa_rows"] = [
        (start + 1 + i, x.generico, x.zona, str(x.zona).startswith("—"))
        for i, (_, x) in enumerate(t.iterrows())
    ]
    # resaltar filas TOTAL y la columna Diferencia
    for rr in range(start + 1, r):
        z = ws.cell(row=rr, column=3).value
        if isinstance(z, str) and z.startswith("—"):
            for cc in range(2, 11):
                ws.cell(row=rr, column=cc).font = Font(size=9, bold=True, color=INK)
                ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor="EFE7D6")
        d = ws.cell(row=rr, column=8).value
        if isinstance(d, (int, float)):
            ws.cell(row=rr, column=8).font = Font(size=9, bold=True, color=BADIE if d < 0 else OK)
            ws.cell(row=rr, column=10).font = Font(size=9, bold=True, color=BADIE if d < 0 else OK)

    ch = BarChart()
    ch.type, ch.grouping = "col", "clustered"
    zonas = t[~t.zona.str.startswith("—")]
    off = start + 1
    idxs = [i for i, (_, x) in enumerate(t.iterrows()) if not x.zona.startswith("—")]
    data = Reference(ws, min_col=4, max_col=4, min_row=start, max_row=r - 1)
    data2 = Reference(ws, min_col=7, max_col=7, min_row=start, max_row=r - 1)
    cats = Reference(ws, min_col=3, max_col=3, min_row=start + 1, max_row=r - 1)
    ch.add_data(data, titles_from_data=True)
    ch.add_data(data2, titles_from_data=True)
    ch.set_categories(cats)
    ch.series[0].graphicalProperties.solidFill = BADIE
    ch.series[1].graphicalProperties.solidFill = TEAL
    style_chart(ch, "Descuentos otorgados vs Tasa generada, por zona", y_title="ARS", w=25, h=11)
    ws.add_chart(ch, f"B{r + 2}")

    ch2 = BarChart()
    ch2.type = "bar"
    d3 = Reference(ws, min_col=8, max_col=8, min_row=start, max_row=r - 1)
    ch2.add_data(d3, titles_from_data=True)
    ch2.set_categories(cats)
    ch2.series[0].graphicalProperties.solidFill = WARN
    style_chart(ch2, "Diferencia por zona (negativo = BADIE pone de más)", w=25, h=10)
    ws.add_chart(ch2, f"B{r + 25}")

    # --- apertura por sucursal (el bloque superior de 'COMPRAS & DESC') ---
    # Va DEBAJO de los graficos para no separar el modelo de su lectura. Es la
    # apertura de donde salen las cifras de zona: sin ella no hay forma de ver
    # que sucursal mueve una zona sin abrir la BASE.
    ap = inf.get("tasa_apertura")
    if ap is not None and len(ap):
        r = max(r + 48, r + 48)
        r = band(ws, r, "APERTURA POR SUCURSAL — de acá salen las cifras de zona", col=2, span=6)
        r += 1
        for gen in dict.fromkeys(ap["generico"]):
            g = ap[ap["generico"] == gen].copy()
            gdf = g.rename(columns={"sucursal": "Sucursal", "zona": "Zona",
                                    "descuentos": "Descuentos", "compras": "Compras"})[
                ["Sucursal", "Zona", "Descuentos", "Compras"]]
            ws.cell(row=r, column=2, value=gen).font = Font(size=11, bold=True, color=TEAL)
            r += 1
            ini = r
            r = write_df(ws, gdf, r, {"Descuentos": F_MONEY, "Compras": F_MONEY0},
                         total_row=False, start_col=2)
            # Primera y ultima fila de datos del bloque (la ultima es el TOTAL
            # del generico, que se resuelve con SUM y no con SUMIFS).
            layout.setdefault("apertura", {})[gen] = (ini + 1, r - 1)
            for rr in range(ini + 1, r):
                sv = ws.cell(row=rr, column=2).value
                if isinstance(sv, str) and sv.startswith("—"):
                    for cc in range(2, 6):
                        ws.cell(row=rr, column=cc).font = Font(size=9, bold=True, color=INK)
                        ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor="EFE7D6")
            r += 2

    # ============ 4. EVOLUCION DIARIA ============
    ws = wb.create_sheet("Evolución Diaria")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 13, "C": 19, "D": 17, "E": 11, "F": 19, "G": 17,
                "H": 11, "I": 19})
    # La facturacion sale de gold (fact_ventas), NO del wapi. El wapi trae solo
    # las lineas CON accion promocional: como denominador del % Desc. subestima
    # la venta y el porcentaje sale inflado. Los descuentos son los mismos por
    # las dos vias (0,04% de diferencia en agosto-2026) porque el descuento solo
    # existe donde hay accion; lo que faltaba era la venta contra la que medirlo.
    r = title(ws, 2, "Evolución diaria",
              "Fuente C · gold.fact_ventas — venta COMPLETA, abierta en CCU y resto · "
              "el % se mide contra la venta del genérico, no contra las líneas promocionadas")
    dg = diario_gold.copy()
    dg["fecha"] = dg["fecha"].astype(str)
    dg["es_ccu"] = dg["es_ccu"].astype(bool)
    piv = dg.pivot_table(index="fecha", columns="es_ccu",
                         values=["facturacion", "descuentos"], aggfunc="sum", fill_value=0.0)
    fechas = sorted(piv.index)

    def _v(campo, ccu, f):
        try:
            return float(piv.loc[f, (campo, ccu)])
        except KeyError:
            return 0.0

    dd = pd.DataFrame({
        "Fecha": fechas,
        "Fact. CCU": [_v("facturacion", True, f) for f in fechas],
        "Desc. CCU": [_v("descuentos", True, f) for f in fechas],
        "Fact. resto": [_v("facturacion", False, f) for f in fechas],
        "Desc. resto": [_v("descuentos", False, f) for f in fechas],
    })
    dd["% Desc. CCU"] = dd["Desc. CCU"] / dd["Fact. CCU"].replace(0, pd.NA)
    dd["% Desc. resto"] = dd["Desc. resto"] / dd["Fact. resto"].replace(0, pd.NA)
    dd["Fact. TOTAL"] = dd["Fact. CCU"] + dd["Fact. resto"]
    dd = dd[["Fecha", "Fact. CCU", "Desc. CCU", "% Desc. CCU",
             "Fact. resto", "Desc. resto", "% Desc. resto", "Fact. TOTAL"]]
    start = r
    r = write_df(ws, dd, r,
                 {"Fact. CCU": F_MONEY, "Desc. CCU": F_MONEY, "% Desc. CCU": F_PCT,
                  "Fact. resto": F_MONEY, "Desc. resto": F_MONEY, "% Desc. resto": F_PCT,
                  "Fact. TOTAL": F_MONEY},
                 start_col=2, table_name="tblDiario")
    # Los % de la fila de total no son sumables: se recalculan como cociente de
    # los totales, que es lo que significan.
    _tot = r - 1
    ws.cell(row=_tot, column=5, value=f"=IF(D{_tot}=0,\"\",D{_tot}/C{_tot})").number_format = F_PCT
    ws.cell(row=_tot, column=8, value=f"=IF(G{_tot}=0,\"\",G{_tot}/F{_tot})").number_format = F_PCT

    lc = LineChart()
    d = Reference(ws, min_col=3, max_col=4, min_row=start, max_row=r - 2)
    cats = Reference(ws, min_col=2, max_col=2, min_row=start + 1, max_row=r - 2)
    lc.add_data(d, titles_from_data=True)
    lc.set_categories(cats)
    lc.series[0].graphicalProperties.line.solidFill = INK
    lc.series[0].graphicalProperties.line.width = 26000
    lc.series[0].marker = Marker(symbol="circle", size=5)
    lc.series[1].graphicalProperties.line.solidFill = BADIE
    lc.series[1].graphicalProperties.line.width = 22000
    lc.series[1].marker = Marker(symbol="diamond", size=5)
    style_chart(lc, "CCU — facturación y descuentos por día", x_title="Fecha", y_title="ARS", w=26, h=11)
    ws.add_chart(lc, f"K{start}")

    lc2 = LineChart()
    d2 = Reference(ws, min_col=5, max_col=5, min_row=start, max_row=r - 2)
    lc2.add_data(d2, titles_from_data=True)
    lc2.set_categories(cats)
    lc2.series[0].graphicalProperties.line.solidFill = WARN
    lc2.series[0].graphicalProperties.line.width = 24000
    lc2.series[0].marker = Marker(symbol="triangle", size=5)
    style_chart(lc2, "CCU — % descuento diario (Desc. / Fact. del genérico)", x_title="Fecha", w=26, h=10)
    ws.add_chart(lc2, f"K{start + 23}")

    # Conciliacion calculada contra «Genéricos» y contra el wapi. Los numeros se
    # calculan, no se declaran: si un dia dejan de cerrar, la hoja lo muestra.
    rn = r + 2
    rn = band(ws, rn, "CONCILIACIÓN — de dónde sale cada total", col=2, span=7)
    f_ccu_d, d_ccu_d = float(dd["Fact. CCU"].sum()), float(dd["Desc. CCU"].sum())
    f_tot_d = float(dd["Fact. TOTAL"].sum())
    ccu_mask = fact["generico"].isin(CCU)
    f_ccu_g = float(fact.loc[ccu_mask, "facturacion"].sum())
    d_ccu_g = float(fact.loc[ccu_mask, "descuentos"].sum())
    f_all_g, d_all_g = float(fact["facturacion"].sum()), float(fact["descuentos"].sum())
    f_wapi = float(daily["facturacion"].sum()) if len(daily) else 0.0
    d_wapi = float(daily["descuentos"].sum()) if len(daily) else 0.0
    ws.cell(row=rn - 1, column=4, value="Facturación").font = Font(size=8, bold=True, color=INK_MUTE)
    ws.cell(row=rn - 1, column=5, value="Descuentos").font = Font(size=8, bold=True, color=INK_MUTE)
    for etiqueta, fv, dv in [
        ("Esta hoja — columna CCU", f_ccu_d, d_ccu_d),
        ("«Genéricos» acotado a los 5 genéricos CCU", f_ccu_g, d_ccu_g),
        ("Esta hoja — CCU + resto", f_tot_d, float(dd["Desc. CCU"].sum() + dd["Desc. resto"].sum())),
        ("«Genéricos» completo (toda la venta BADIE)", f_all_g, d_all_g),
        ("wapi (sólo líneas CON acción) — ya no se usa acá", f_wapi, d_wapi),
    ]:
        ws.cell(row=rn, column=2, value=etiqueta).font = Font(size=9, color=INK)
        c = ws.cell(row=rn, column=4, value=fv); c.number_format = F_MONEY; c.font = Font(size=9)
        c = ws.cell(row=rn, column=5, value=dv); c.number_format = F_MONEY; c.font = Font(size=9)
        rn += 1
    dif_f = (f_ccu_d / f_ccu_g - 1) if f_ccu_g else 0
    dif_d = (d_wapi / d_ccu_d - 1) if d_ccu_d else 0
    falta = f_ccu_d - f_wapi
    rn += 1
    ws.cell(row=rn, column=2, value=(
        f"La columna CCU cierra contra «Genéricos» ({dif_f:+.3%}): las dos miden la venta de los 5 "
        f"genéricos CCU en gold. El wapi mide otra cosa — sólo las líneas CON acción promocional — y "
        f"por eso queda ${falta:,.0f} corto: es venta de producto CCU que no tuvo acción. Los "
        f"DESCUENTOS sí coinciden por las dos vías ({dif_d:+.2%}), porque el descuento sólo existe "
        f"donde hay acción. Usar el wapi como denominador del % inflaba el porcentaje."
    )).font = Font(size=9, italic=True, color=INK_MUTE)
    ws.cell(row=rn, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=rn, start_column=2, end_row=rn + 2, end_column=9)

    # ============ 5. GENERICOS ============
    ws = wb.create_sheet("Genéricos")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 22, "C": 18, "D": 17, "E": 12, "F": 11, "G": 10})
    r = title(ws, 2, "Mix y presión promocional por genérico",
              "Fuente A · BASE control · hoja FACT_NET — venta COMPLETA de BADIE, "
              "todos los genéricos (no sólo CCU)")
    g = (fact.groupby("generico").agg(Facturación=("facturacion", "sum"),
                                      Descuentos=("descuentos", "sum"),
                                      Artículos=("codigo", "nunique")).reset_index()
         .rename(columns={"generico": "Genérico"}).sort_values("Facturación", ascending=False))
    g["% Desc."] = g.Descuentos / g.Facturación.replace(0, pd.NA)
    g["Share"] = g.Facturación / g.Facturación.sum()
    g["CCU"] = g["Genérico"].isin(CCU).map({True: "CCU", False: "—"})
    start = r
    r = write_df(ws, g, r, {"Facturación": F_MONEY, "Descuentos": F_MONEY, "Artículos": F_INT,
                            "% Desc.": F_PCT, "Share": F_PCT}, start_col=2, table_name="tblGen")
    ws.conditional_formatting.add(f"F{start+1}:F{r-2}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=50, mid_color=PASTEL[3],
                       end_type="max", end_color=BADIE))

    bc = BarChart()
    bc.type = "bar"
    d = Reference(ws, min_col=3, max_col=4, min_row=start, max_row=r - 2)
    cats = Reference(ws, min_col=2, max_col=2, min_row=start + 1, max_row=r - 2)
    bc.add_data(d, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = TEAL
    bc.series[1].graphicalProperties.solidFill = BADIE
    style_chart(bc, "Facturación y descuentos por genérico", w=24, h=12)
    ws.add_chart(bc, f"I{start}")

    bc2 = BarChart()
    bc2.type = "bar"
    d2 = Reference(ws, min_col=6, max_col=6, min_row=start, max_row=r - 2)
    bc2.add_data(d2, titles_from_data=True)
    bc2.set_categories(cats)
    bc2.series[0].graphicalProperties.solidFill = PASTEL[0]
    style_chart(bc2, "% Descuento por genérico", w=24, h=12)
    ws.add_chart(bc2, f"I{start + 26}")

    # ============ 6. SUCURSALES ============
    ws = wb.create_sheet("Sucursales")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30, "C": 17, "D": 12, "E": 18, "F": 17, "G": 11, "H": 12, "I": 12, "J": 12})
    r = title(ws, 2, "Ranking de sucursales", "Fuentes A (fact/desc) + B (zona, supervisor) + C (bultos, clientes de gold)")
    s = (fact.groupby("sucursal").agg(Facturación=("facturacion", "sum"),
                                      Descuentos=("descuentos", "sum"),
                                      Artículos=("codigo", "nunique")).reset_index())
    s["Zona"] = s.sucursal.map(inf["zona_map"]).fillna("—")
    s["Supervisor"] = s.sucursal.map(inf["sup_map"]).fillna("—")
    # Bultos SI se suma (es volumen). Clientes NO: viene contado por
    # (sucursal, generico, marca) y sumarlo cuenta al mismo cliente una vez por
    # marca. Se toma del conteo hecho al grano sucursal.
    gv = ventas.groupby("sucursal").agg(Bultos=("bultos", "sum")).reset_index()
    gv = gv.merge(clientes_suc.rename(columns={"clientes": "Clientes"}), on="sucursal", how="left")
    gv["sucursal_norm"] = gv.sucursal.str.upper().str.strip()
    s["sucursal_norm"] = s.sucursal.str.replace(r"^\d+\s*-\s*", "", regex=True).str.upper().str.strip()
    s = s.merge(gv[["sucursal_norm", "Bultos", "Clientes"]], on="sucursal_norm", how="left")
    s["% Desc."] = s.Descuentos / s.Facturación.replace(0, pd.NA)
    glob = s.Descuentos.sum() / s.Facturación.sum()
    s["vs Global"] = s["% Desc."] - glob
    s = s.rename(columns={"sucursal": "Sucursal"}).sort_values("Facturación", ascending=False)
    s = s[["Sucursal", "Zona", "Supervisor", "Facturación", "Descuentos", "% Desc.", "vs Global", "Bultos", "Clientes"]]
    start = r
    r = write_df(ws, s, r, {"Facturación": F_MONEY, "Descuentos": F_MONEY, "% Desc.": F_PCT,
                            "vs Global": F_PCT, "Bultos": F_DEC, "Clientes": F_INT},
                 start_col=2, table_name="tblSuc")
    ws.conditional_formatting.add(f"H{start+1}:H{r-2}",
        ColorScaleRule(start_type="min", start_color=OK, mid_type="num", mid_value=0,
                       mid_color="FFFFFF", end_type="max", end_color=BADIE))
    ws.conditional_formatting.add(f"E{start+1}:E{r-2}",
        DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True))

    bc = BarChart()
    bc.type = "col"
    d = Reference(ws, min_col=5, max_col=6, min_row=start, max_row=r - 2)
    cats = Reference(ws, min_col=2, max_col=2, min_row=start + 1, max_row=r - 2)
    bc.add_data(d, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = TEAL
    bc.series[1].graphicalProperties.solidFill = BADIE
    style_chart(bc, "Facturación y descuentos por sucursal", y_title="ARS", w=27, h=11)
    ws.add_chart(bc, f"B{r + 2}")

    # ============ 7. CANAL ============
    ws = wb.create_sheet("Canal")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 34, "C": 18, "D": 12, "E": 11})
    r = title(ws, 2, "Descuento por canal (lista de precios)",
              ("Descuentos de la BASE · padrón LISTA del informe manual · " + _label_periodo())
              if inf.get("canal_calculado")
              else "Fuente B · informe final · columna LISTA PRECIO")
    cn = inf["canal"].rename(columns={"canal": "Canal", "descuento": "Descuento", "lineas": "Líneas"})
    cn["Share"] = cn.Descuento / cn.Descuento.sum()
    start = r
    r = write_df(ws, cn, r, {"Descuento": F_MONEY, "Líneas": F_INT, "Share": F_PCT},
                 start_col=2, table_name="tblCanal")
    for rr in range(start + 1, r - 1):
        if "#N/D" in str(ws.cell(row=rr, column=2).value):
            for cc in range(2, 6):
                ws.cell(row=rr, column=cc).font = Font(size=9, bold=True, color=BADIE)
    bc = BarChart()
    bc.type = "bar"
    d = Reference(ws, min_col=3, max_col=3, min_row=start, max_row=r - 2)
    cats = Reference(ws, min_col=2, max_col=2, min_row=start + 1, max_row=r - 2)
    bc.add_data(d, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = PASTEL[1]
    style_chart(bc, "Descuento por canal comercial", w=25, h=12)
    ws.add_chart(bc, f"G{start}")

    # pivot canal x generico
    r += 2
    r = band(ws, r, "CANAL × GENÉRICO", col=2, span=8)
    cg = inf["canal_gen"].pivot_table(index="canal", columns="generico",
                                      values="descuento", aggfunc="sum", fill_value=0).reset_index()
    cg = cg.rename(columns={"canal": "Canal"})
    p_start = r
    r = write_df(ws, cg, r, {c: F_MONEY for c in cg.columns if c != "Canal"},
                 start_col=2, table_name="tblCanalGen")
    last = get_column_letter(2 + len(cg.columns) - 1)
    ws.conditional_formatting.add(f"C{p_start+1}:{last}{r-2}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=60, mid_color=PASTEL[5],
                       end_type="max", end_color=TEAL))

    # ============ 8. CODIGOS DE ACCION ============
    ws = wb.create_sheet("Códigos de Acción")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 46, "C": 48, "D": 14, "E": 18, "F": 11, "G": 11, "H": 13})
    r = title(ws, 2, "Códigos de acción", "Fuente A (montos) + B (flag reversa)")
    a = (acc.groupby(["accion", "desc_accion"]).agg(Descuento=("descuento", "sum"),
                                                    Artículos=("codigo", "nunique"),
                                                    Sucursales=("sucursal", "nunique")).reset_index())
    _maestro = set(inf.get("reversa_map", {}))
    a["Reversa"] = a.accion.map(
        lambda x: "SI" if inf.get("reversa_map", {}).get(x) is True
        else ("NO" if x in _maestro else "?"))
    a["$ / Artículo"] = a.Descuento / a.Artículos.replace(0, pd.NA)
    a = a.rename(columns={"accion": "Código", "desc_accion": "Descripción"}).sort_values("Descuento", ascending=False)
    a = a[["Código", "Descripción", "Reversa", "Descuento", "Artículos", "Sucursales", "$ / Artículo"]]
    ws.cell(row=r, column=2, value=(
        f"{len(a):,} códigos de acción con movimiento — se listan TODOS.  "
        "«Artículos» = artículos distintos alcanzados por la acción · "
        "«Sucursales» = sucursales distintas donde se aplicó (máx. 13; incluye alguna con monto 0 o negativo).  "
        "«Reversa»: SI = CCU lo reintegra · NO = confirmado que lo absorbe BADIE · "
        "? = la acción no figura en el maestro AccRever, no está clasificada."
    )).font = Font(size=9, italic=True, color=INK_MUTE)
    r += 2
    start = r
    r = write_df(ws, a, r, {"Descuento": F_MONEY, "Artículos": F_INT, "Sucursales": F_INT,
                            "$ / Artículo": F_MONEY}, start_col=2, table_name="tblAcc",
                 no_total=("Artículos", "Sucursales", "$ / Artículo"))
    for rr in range(start + 1, r - 1):
        v = ws.cell(row=rr, column=4).value
        if v == "SI":
            ws.cell(row=rr, column=4).fill = PatternFill("solid", fgColor=PASTEL[5])
        elif v == "NO":
            ws.cell(row=rr, column=4).fill = PatternFill("solid", fgColor=PASTEL[8])
        elif v == "?":
            ws.cell(row=rr, column=4).fill = PatternFill("solid", fgColor=PASTEL[6])
        ws.cell(row=rr, column=4).alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(f"E{start+1}:E{r-2}",
        DataBarRule(start_type="min", end_type="max", color=BADIE, showValue=True))
    ws.auto_filter.ref = f"B{start}:H{r-2}"
    ws.freeze_panes = ws.cell(row=start + 1, column=2)

    r += 1
    r = band(ws, r, "REVERSA — ¿quién absorbe el descuento?", col=2, span=7, fill=BADIE)
    cov = inf.get("reversa_cobertura", {})
    ws.cell(row=r, column=2, value=(
        f"«Códigos» = cantidad de códigos de acción DISTINTOS en cada clase. "
        f"El maestro AccRever sólo clasifica {cov.get('maestro_total', 0)} acciones, "
        f"pero este período tuvo {cov.get('acciones_con_movimiento', 0)} con movimiento: "
        "las que faltan caen en la tercera fila como SUPUESTO, no como dato."
    )).font = Font(size=9, italic=True, color=BADIE)
    r += 2
    rv = inf["reversa"].rename(columns={"tipo": "Tipo", "descuento": "Descuento", "codigos": "Códigos"})
    rv["Share"] = rv.Descuento / rv.Descuento.sum()
    r = write_df(ws, rv, r, {"Descuento": F_MONEY, "Códigos": F_INT, "Share": F_PCT},
                 total_row=False, start_col=2)

    # ============ 9. MATRIZ SUC x GEN ============
    ws = wb.create_sheet("Matriz Suc x Gen")
    ws.sheet_view.showGridLines = False
    r = title(ws, 2, "Matriz sucursal × genérico", "Fuente A · descuento cruzado · escala de color por intensidad")
    m = fact.pivot_table(index="sucursal", columns="generico", values="descuentos",
                         aggfunc="sum", fill_value=0)
    m = m.loc[m.sum(axis=1).sort_values(ascending=False).index,
              m.sum(axis=0).sort_values(ascending=False).index].reset_index()
    m = m.rename(columns={"sucursal": "Sucursal"})
    widths(ws, {"A": 3, "B": 32})
    for i in range(len(m.columns) - 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15
    start = r
    r = write_df(ws, m, r, {c: F_MONEY for c in m.columns if c != "Sucursal"},
                 start_col=2, table_name="tblMatriz", zebra=False)
    last = get_column_letter(2 + len(m.columns) - 1)
    ws.conditional_formatting.add(f"C{start+1}:{last}{r-2}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       mid_type="percentile", mid_value=70, mid_color=PASTEL[3],
                       end_type="max", end_color=BADIE))
    ws.freeze_panes = ws.cell(row=start + 1, column=3)

    # ============ 10. CORRELACIONES ============
    ws = wb.create_sheet("Correlaciones")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30})
    r = title(ws, 2, "Correlaciones entre métricas", "Fuentes A + C · Pearson sobre las 13 sucursales")
    ws.cell(row=r, column=2, value="Con n=13 sucursales, |r| por debajo de ~0.55 no es concluyente. La correlación no implica causalidad.").font = Font(size=9, italic=True, color=BADIE)
    r += 2

    base = s.copy()
    base["Bultos"] = pd.to_numeric(base.Bultos, errors="coerce")
    base["Clientes"] = pd.to_numeric(base.Clientes, errors="coerce")
    accs = acc.groupby("sucursal").accion.nunique().rename("Acciones")
    base = base.merge(accs, left_on="Sucursal", right_index=True, how="left")
    cobs = cob.groupby("sucursal").clientes_cob.sum().rename("Cobertura")
    base["sn"] = base.Sucursal.str.replace(r"^\d+\s*-\s*", "", regex=True).str.upper().str.strip()
    base = base.merge(cobs.rename_axis("sn").reset_index().assign(sn=lambda d: d.sn.str.upper().str.strip()),
                      on="sn", how="left")
    mat = base[["Facturación", "Descuentos", "% Desc.", "Bultos", "Clientes", "Acciones", "Cobertura"]].astype(float)
    mat = mat.dropna(axis=1, how="all")

    img_corr = IMG_DIR / "corr.png"
    corr = make_corr_heatmap(mat, img_corr)
    ws.add_image(XLImage(str(img_corr)), f"B{r}")
    r_tbl = r + 34

    r_tbl = band(ws, r_tbl, "MATRIZ DE PEARSON (valores)", col=2, span=len(corr) + 1)
    cdf = corr.reset_index().rename(columns={"index": "Métrica"})
    for i in range(len(cdf.columns) - 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    cs = r_tbl
    r_tbl = write_df(ws, cdf, r_tbl, {c: "0.000" for c in cdf.columns if c != "Métrica"},
                     total_row=False, start_col=2, zebra=False)
    lastc = get_column_letter(2 + len(cdf.columns) - 1)
    ws.conditional_formatting.add(f"C{cs+1}:{lastc}{r_tbl-1}",
        ColorScaleRule(start_type="num", start_value=-1, start_color="4A6C9B",
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="num", end_value=1, end_color=BADIE))

    # ============ 11. COBERTURA ============
    ws = wb.create_sheet("Cobertura")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30, "C": 22, "D": 16, "E": 16, "F": 12, "G": 14})
    r = title(ws, 2, "Cobertura vs presión de descuento", "Fuente C · gold.cob_sucursal_lista_generico + fact_ventas")
    cg2 = (cob.groupby(["sucursal", "generico"]).agg(clientes_cob=("clientes_cob", "sum"),
                                                     volumen=("volumen", "sum")).reset_index())
    vg = ventas.groupby(["sucursal", "generico"]).agg(facturacion=("facturacion", "sum"),
                                                      descuentos=("descuentos", "sum")).reset_index()
    cx = vg.merge(cg2, on=["sucursal", "generico"], how="left")
    cx = cx[cx.facturacion > 0].copy()
    # to_numeric: replace(0, NA) devuelve object y rompe nlargest aguas abajo
    cx["pct_desc"] = pd.to_numeric(cx.descuentos / cx.facturacion, errors="coerce")
    cx["clientes_cob"] = pd.to_numeric(cx.clientes_cob, errors="coerce")
    cx["es_ccu"] = cx.generico.isin(CCU)
    cx["etiqueta"] = cx.sucursal.str[:11] + "/" + cx.generico.str[:9]
    sc = cx.dropna(subset=["clientes_cob", "pct_desc"])
    if len(sc) > 3:
        img_sc = IMG_DIR / "cob.png"
        make_scatter_cob(sc, img_sc)
        ws.add_image(XLImage(str(img_sc)), f"I{r}")

    cxo = cx.rename(columns={"sucursal": "Sucursal", "generico": "Genérico",
                             "facturacion": "Facturación", "descuentos": "Descuentos",
                             "clientes_cob": "Clientes cob.", "volumen": "Volumen",
                             "pct_desc": "% Desc."}).sort_values("Descuentos", ascending=False)
    cxo = cxo[["Sucursal", "Genérico", "Facturación", "Descuentos", "% Desc.", "Clientes cob.", "Volumen"]].head(60)
    r = write_df(ws, cxo, r, {"Facturación": F_MONEY, "Descuentos": F_MONEY, "% Desc.": F_PCT,
                              "Clientes cob.": F_INT, "Volumen": F_DEC},
                 start_col=2, table_name="tblCob")

    # ============ 12. CLIENTES ============
    ws = wb.create_sheet("Clientes")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 11, "C": 34, "D": 26, "E": 17, "F": 10, "G": 10, "H": 10, "I": 11})
    r = title(ws, 2, "Clientes que reciben descuento",
              "Fuente A · BASE control · hoja CLIENTE-FECHA · grano cliente × acción × artículo")

    cres = (cli.groupby(["cod_cliente", "razon_social", "sucursal"])
            .agg(Descuento=("descuento", "sum"),
                 Acciones=("accion", "nunique"),
                 Artículos=("cod_articulo", "nunique"),
                 Movimientos=("descuento", "size")).reset_index()
            .rename(columns={"cod_cliente": "Código", "razon_social": "Razón Social",
                             "sucursal": "Sucursal"})
            .sort_values("Descuento", ascending=False))
    cres["Share"] = cres.Descuento / cres.Descuento.sum()
    cres["$ / Movim."] = cres.Descuento / cres["Movimientos"].replace(0, pd.NA)
    tot_cli = len(cres)
    # Sin tope: la tabla lista TODOS los clientes, asi la fila TOTAL GENERAL
    # coincide con el universo real y no con un recorte.
    top = cres[["Código", "Razón Social", "Sucursal", "Descuento",
                "Acciones", "Artículos", "Movimientos", "Share", "$ / Movim."]]
    ws.cell(row=r, column=2, value=(
        f"{tot_cli:,} clientes con descuento — se listan TODOS. "
        "«Movimientos» = cantidad de registros de CLIENTE-FECHA del cliente "
        "(un registro = un descuento aplicado en una fecha, sobre un artículo, por una acción)."
    )).font = Font(size=9, italic=True, color=INK_MUTE)
    r += 2
    start = r
    r = write_df(ws, top, r, {"Código": F_INT, "Descuento": F_MONEY, "Acciones": F_INT,
                              "Artículos": F_INT, "Movimientos": F_INT, "Share": F_PCT,
                              "$ / Movim.": F_MONEY}, start_col=2, table_name="tblClientes",
                 no_total=("Código", "Acciones", "Artículos", "$ / Movim."))
    ws.auto_filter.ref = f"B{start}:J{r-2}"
    ws.conditional_formatting.add(f"E{start+1}:E{r-2}",
        DataBarRule(start_type="min", end_type="max", color=BADIE, showValue=True))
    ws.freeze_panes = ws.cell(row=start + 1, column=2)

    bc = BarChart()
    bc.type = "bar"
    d = Reference(ws, min_col=5, max_col=5, min_row=start, max_row=min(start + 20, r - 2))
    cats = Reference(ws, min_col=3, max_col=3, min_row=start + 1, max_row=min(start + 20, r - 2))
    bc.add_data(d, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = PASTEL[4]
    style_chart(bc, "Top 20 clientes por descuento recibido", w=25, h=13)
    ws.add_chart(bc, f"K{start}")

    # ---- detalle cliente x accion x articulo ----
    r += 2
    det = (cli.groupby(["cod_cliente", "razon_social", "sucursal", "accion",
                        "desc_accion", "cod_articulo", "articulo", "generico"])
           .agg(Descuento=("descuento", "sum"), Días=("descuento", "size")).reset_index()
           .rename(columns={"cod_cliente": "Código", "razon_social": "Razón Social",
                            "sucursal": "Sucursal", "accion": "Acción",
                            "desc_accion": "Descripción Acción", "cod_articulo": "Cód. Art.",
                            "articulo": "Artículo", "generico": "Genérico"})
           .sort_values("Descuento", ascending=False))
    det["Reversa"] = det["Acción"].map(inf["reversa_map"]).map({True: "SI", False: "NO"}).fillna("—")
    # Sin tope tambien aca: el grano es cliente x accion x articulo, y lo unico
    # que queda variando dentro de cada grupo es la fecha -> «Días».
    det_top = det[["Código", "Razón Social", "Sucursal", "Acción", "Descripción Acción",
                   "Reversa", "Cód. Art.", "Artículo", "Genérico", "Descuento", "Días"]]
    r = band(ws, r, f"DETALLE  ·  cliente × acción × artículo  —  {len(det_top):,} filas (todas)",
             col=2, span=8)
    ws.cell(row=r, column=2, value=(
        "«Días» = en cuántas fechas distintas ese cliente recibió esa acción sobre ese artículo."
    )).font = Font(size=9, italic=True, color=INK_MUTE)
    r += 2
    widths(ws, {"J": 12, "K": 34, "L": 20, "M": 16, "N": 10})
    dstart = r
    r = write_df(ws, det_top, r, {"Código": F_INT, "Cód. Art.": F_INT,
                                  "Descuento": F_MONEY, "Días": F_INT},
                 start_col=2, table_name="tblCliDet", zebra=False,
                 no_total=("Código", "Cód. Art."))
    ws.auto_filter.ref = f"B{dstart}:L{r-2}"
    for rr in range(dstart + 1, r - 1):
        v = ws.cell(row=rr, column=7).value
        if v == "SI":
            ws.cell(row=rr, column=7).fill = PatternFill("solid", fgColor=PASTEL[5])
        elif v == "NO":
            ws.cell(row=rr, column=7).fill = PatternFill("solid", fgColor=PASTEL[8])
        ws.cell(row=rr, column=7).alignment = Alignment(horizontal="center")

    # ============ TABLERO (paridad con la hoja TABLERO del original) ============
    ws = wb.create_sheet("Tablero CCU")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 26, "C": 20, "D": 14, "E": 4, "F": 20, "G": 14})
    r = title(ws, 2, "Tablero CCU",
              "Réplica de la hoja TABLERO del informe original · sólo los 5 genéricos CCU")
    ws.cell(row=r, column=2, value=(
        "El TABLERO original corta SÓLO por genéricos CCU. La hoja «Genéricos» de este "
        "informe abre el universo completo, por eso sus totales son mayores. No son "
        "cifras contradictorias: son alcances distintos."
    )).font = Font(size=9, italic=True, color=BADIE)
    r += 2

    tb = (fact[fact.generico.isin(CCU)]
          .groupby("generico")
          .agg(Facturación=("facturacion", "sum"), Descuentos=("descuentos", "sum"))
          .reset_index().rename(columns={"generico": "Genérico"})
          .sort_values("Facturación", ascending=False))
    tot_f, tot_d = tb.Facturación.sum(), tb.Descuentos.sum()
    tb["$ del Total"] = tb.Facturación / tot_f if tot_f else 0
    tb["% del Total"] = tb.Descuentos / tot_d if tot_d else 0
    tb = tb[["Genérico", "Facturación", "$ del Total", "Descuentos", "% del Total"]]
    start = r
    r = write_df(ws, tb, r, {"Facturación": F_MONEY, "$ del Total": F_PCT,
                             "Descuentos": F_MONEY, "% del Total": F_PCT},
                 start_col=2, table_name="tblTablero")
    ws.cell(row=r + 1, column=2, value="Porcentaje global (Desc. / Fact.)").font = Font(size=10, bold=True, color=INK)
    c = ws.cell(row=r + 1, column=3, value=(tot_d / tot_f) if tot_f else 0)
    c.number_format = F_PCT
    c.font = Font(size=12, bold=True, color=BADIE)

    bc = BarChart(); bc.type = "bar"
    d = Reference(ws, min_col=3, max_col=3, min_row=start, max_row=r - 2)
    cats = Reference(ws, min_col=2, max_col=2, min_row=start + 1, max_row=r - 2)
    bc.add_data(d, titles_from_data=True); bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = TEAL
    style_chart(bc, "Facturación por genérico CCU", w=22, h=10)
    ws.add_chart(bc, f"I{start}")

    # ============ ACC-GEN (paridad con la hoja ACC-GEN del original) ============
    ws = wb.create_sheet("Acción x Genérico")
    ws.sheet_view.showGridLines = False
    r = title(ws, 2, "Acción × Genérico por sucursal",
              "Réplica de la hoja ACC-GEN del original · fuente A · BASE control")
    if len(accgen):
        gcols = [c for c in accgen.columns if c not in ("sucursal", "accion", "desc_accion", "mvb")]
        ag = accgen.rename(columns={"sucursal": "Sucursal", "accion": "Acción",
                                    "desc_accion": "Descripción Acción", "mvb": "mvb"})
        ag = ag[["Sucursal", "Acción", "Descripción Acción", "mvb"] + gcols]
        widths(ws, {"A": 3, "B": 30, "C": 11, "D": 44, "E": 12})
        for i in range(len(gcols)):
            ws.column_dimensions[get_column_letter(6 + i)].width = 17
        ws.cell(row=r, column=2, value=f"{len(ag):,} combinaciones sucursal × acción — se listan TODAS.").font = Font(size=9, italic=True, color=INK_MUTE)
        r += 2
        astart = r
        r = write_df(ws, ag, r, {g: F_MONEY for g in gcols},
                     start_col=2, table_name="tblAccGen", zebra=False)
        last = get_column_letter(2 + len(ag.columns) - 1)
        ws.conditional_formatting.add(f"F{astart+1}:{last}{r-2}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                           mid_type="percentile", mid_value=70, mid_color=PASTEL[5],
                           end_type="max", end_color=TEAL))
        ws.freeze_panes = ws.cell(row=astart + 1, column=4)
        ws.auto_filter.ref = f"B{astart}:{last}{r-2}"
    else:
        ws.cell(row=r, column=2, value="La BASE no trajo la hoja ACC-GEN.").font = Font(size=10, italic=True, color=BADIE)

    # ============ REVERSA por categoría de cliente ============
    ws = wb.create_sheet("Reversa Clientes")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 11, "C": 34, "D": 30, "E": 24, "F": 18})
    r = title(ws, 2, "Reversa por categoría de cliente",
              "Réplica de la hoja Reversa del original · fuente B")
    rvc = inf.get("reversa_clientes")
    if rvc is not None and len(rvc):
        ws.cell(row=r, column=2, value=(
            "La categoría (CLIENTES ON A · CLUBES DE RUGBY · CLUBES DE FÚTBOL) es una "
            "clasificación MANUAL que vive sólo en el informe: no se puede derivar de gold "
            "ni de la BASE. Si se agrega un cliente nuevo, hay que cargarlo ahí."
        )).font = Font(size=9, italic=True, color=BADIE)
        r += 2
        res = (rvc.groupby("categoria")
               .agg(Clientes=("cod_cliente", "nunique"), Total=("total", "sum"))
               .reset_index().rename(columns={"categoria": "Categoría"})
               .sort_values("Total", ascending=False))
        r = band(ws, r, "RESUMEN POR CATEGORÍA", col=2, span=5, fill=BADIE)
        r = write_df(ws, res, r, {"Clientes": F_INT, "Total": F_MONEY}, start_col=2)
        r += 1
        r = band(ws, r, "DETALLE POR CLIENTE", col=2, span=5)
        det = (rvc.rename(columns={"cod_cliente": "Código", "razon_social": "Razón Social",
                                   "fantasia": "Fantasía", "categoria": "Categoría",
                                   "total": "Total"})
               [["Código", "Razón Social", "Fantasía", "Categoría", "Total"]]
               .sort_values("Total", ascending=False))
        det["Código"] = det["Código"].astype(int)
        dstart = r
        r = write_df(ws, det, r, {"Código": F_INT, "Total": F_MONEY},
                     start_col=2, table_name="tblRevCli", no_total=("Código",))
        ws.conditional_formatting.add(f"F{dstart+1}:F{r-2}",
            DataBarRule(start_type="min", end_type="max", color=BADIE, showValue=True))
        ws.auto_filter.ref = f"B{dstart}:F{r-2}"
    else:
        ws.cell(row=r, column=2, value="El informe no trajo la hoja Reversa.").font = Font(size=10, italic=True, color=BADIE)

    # ============ 13. DATOS ============
    ws = wb.create_sheet("Datos")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30, "C": 18, "D": 14, "E": 20, "F": 10, "G": 34,
                "H": 18, "I": 18, "J": 16, "K": 16})
    r = title(ws, 2, "Datos base",
              "Fuente A · FACT_NET completo · de acá salen los Descuentos del Modelo Tasa")
    fd = fact.rename(columns={"sucursal": "Sucursal", "codigo": "Código", "articulo": "Artículo",
                              "marca": "Marca", "generico": "Genérico",
                              "facturacion": "Facturación", "descuentos": "Descuentos"})
    # OJO: `zona_map` sale de sucu!D y NO es la zona — es el alias corto de la
    # sucursal (ORAN, LEDESMA, MAIMARA...). La zona del modelo de tasa es la de
    # sucu leida en el bloque superior de 'COMPRAS & DESC' (4 zonas). Se muestran
    # las dos: la zona para poder agrupar, el alias porque es como la nombra el
    # informe manual.
    fd["Zona"] = fd.Sucursal.map(inf["zona_por_sucursal"]).fillna("—")
    fd["Alias"] = fd.Sucursal.map(inf["zona_map"]).fillna("—")
    fd["Supervisor"] = fd.Sucursal.map(inf["sup_map"]).fillna("—")
    fd = fd[["Sucursal", "Zona", "Alias", "Supervisor", "Código", "Artículo", "Marca", "Genérico", "Facturación", "Descuentos"]]
    start = r
    r = write_df(ws, fd, r, {"Facturación": F_MONEY, "Descuentos": F_MONEY, "Código": F_INT},
                 start_col=2, table_name="tblDatos", zebra=False, no_total=("Código",))
    ws.freeze_panes = ws.cell(row=start + 1, column=2)
    ws.auto_filter.ref = f"B{start}:K{r-2}"
    # B=Sucursal C=Zona D=Alias E=Supervisor F=Codigo G=Articulo H=Marca
    # I=Generico J=Facturacion K=Descuentos. write_df cierra con fila de total,
    # asi que los datos van hasta r-2.
    layout["datos"] = (start + 1, r - 2)

    # ============ 14. INSUMOS CRUDOS ============
    _hoja_compras_crudo(wb, inf, layout)
    _hoja_wapi_crudo(wb, wapi_crudo)

    # Con las hojas crudas ya escritas, el cuadro de tasa se reescribe como
    # formulas: cada numero pasa a mostrar de donde sale.
    _formulas_modelo_tasa(wb, inf, layout)

    # tab colors
    for name, col in [("Portada", INK), ("Resumen", BADIE), ("Modelo Tasa", BADIE),
                      ("Evolución Diaria", TEAL), ("Genéricos", TEAL), ("Sucursales", TEAL),
                      ("Canal", TEAL), ("Códigos de Acción", TEAL), ("Matriz Suc x Gen", WARN),
                      ("Correlaciones", WARN), ("Cobertura", WARN), ("Clientes", TEAL), ("Tablero CCU", BADIE), ("Acción x Genérico", TEAL), ("Reversa Clientes", BADIE), ("Datos", INK_MUTE)]:
        wb[name].sheet_properties.tabColor = col

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, corr


def main(argv):
    global BASE_XLSX, PERIODO_HASTA, INFORME_XLSM
    out = Path(argv[1]) if len(argv) > 1 else OUT_DEFAULT
    if len(argv) > 2:                      # ruta alternativa a la BASE control
        BASE_XLSX = Path(argv[2])
    if len(argv) > 3:                      # corte del periodo (para gold)
        PERIODO_HASTA = argv[3]
    if len(argv) > 4:                      # informe final (fuente B)
        INFORME_XLSM = Path(argv[4])
    print(f"BASE    : {BASE_XLSX}")
    print(f"INFORME : {INFORME_XLSM}")
    print(f"Periodo : {PERIODO_DESDE} -> {PERIODO_HASTA}")
    if out.exists():
        print(f"AVISO: ya existe {out.name} — se sobrescribe la salida generada por este script.")
    p, corr = build(out)
    print(f"\nOK  {p}")
    print(f"    {p.stat().st_size:,} bytes")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
