"""Add (or replace) the 'Cobertura CCU' sheet on a GUEMES workbook, IN-PLACE.

Layout: preventistas as ROWS, marcas as COLUMNS. One section per CCU generico,
with CERVEZAS split across three sections (too many marcas for one row of
headers). Each section ends with a per-generico TOTAL column group (except
the first two CERVEZAS thirds, which share their total with the third) and a
sucursal-wide TOTAL GUEMES row.

Formulas are written from scratch against COBER / CuposCober, scoped to
GUEMES ('16 - SUCURSAL GUEMES'). Nothing is read from or copied out of the
legacy 'Cober Nueva' sheet — that sheet hardcodes '1 - CASA CENTRAL' filters
and has known-broken cross-block references, which is exactly why this sheet
exists as an independent replacement.

Data model:
    COBER        A=Sucursal B=Descripcion Vendedor C=Ruta D=Descripcion_Marca
                 E=Numero_Clientes                              (marca grain)
                 H=Sucursal I=Descripcion Vendedor J=Ruta K=GENERICO
                 L=Numero_Clientes                            (generico grain)
    CuposCober   B=Ruta C=Descripción Vendedor D=MARCA E=ZONA (sucursal) H=CUPO

The generico PDV is read from COBER's generico grain (H/I/K/L), NOT summed
from the marca columns — a client covered in two marcas of the same generico
must count ONCE. The generico OBJ, in contrast, IS rolled up from the marca
cupos via a SUMIFS array literal, because the generico-level cupo is not
published (CuposCoberGen is empty for GUEMES). That rollup must cover ALL
marcas of the generico, including ones with a published cupo that are not
rendered as columns (see OBJ_ONLY_MARCAS) — otherwise the coverage ratio is
overstated because PDV counts the full generico universe but OBJ would not.

Idempotent: safe to re-run. Backs up the target file first (data/ is
gitignored).

Usage:
    python scripts/add_cobertura_ccu_sheet.py "data/input/avances/AVANCE GUEMES.xlsx"
"""
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

SHEET = "Cobertura CCU"
SUC = "16 - SUCURSAL GUEMES"
PREVENTISTAS = ["JORGE RAMOS", "TALLO GABRIELA", "DIRECTA"]

# (section title, generico, marcas, emit_generico_total)
# CERVEZAS spans three sections; only the third carries the CERVEZAS total.
SECTIONS = [
    ("CERVEZAS (1/3)", "CERVEZAS",
     ["SALTA", "HEINEKEN", "IMPERIAL", "MILLER"], False),
    ("CERVEZAS (2/3)", "CERVEZAS",
     ["BIECKERT", "SCHNEIDER", "SOL", "AMSTEL", "KUNSTMAN", "BLUE MOON", "SALTA CAUTIVA1"], False),
    ("CERVEZAS (3/3)", "CERVEZAS",
     ["GROLSCH", "IGUANA", "ISENBECK", "WARSTEINER", "NORTE", "PALERMO"], True),
    ("AGUAS DANONE", "AGUAS DANONE",
     ["LEVITE", "VILLAVICENCIO", "VILLA DEL SUR", "BRIO", "SER", "FULL SPORT"], True),
    ("VINOS CCU", "VINOS CCU",
     ["COLON", "LA CELIA", "GRAFFIGNA", "EUGENIO BUSTOS", "O-61", "SANTA SILVIA"], True),
    ("SIDRAS Y LICORES", "SIDRAS Y LICORES",
     ["REAL", "LA VICTORIA", "SAENZ BRIONES", "EL ABUELO", "PEHUENIA", "MISTRAL", "CONTROL C"], True),
]

# Marcas with published cupos in CuposCober that are NOT rendered as columns.
# Empty: every marca verified against gold.dim_articulo with a published cupo
# is now a displayed column (see SECTIONS above), so nothing is OBJ-only
# anymore. Kept as a dict (rather than removed) because GENERICO_MARCAS and
# existing call sites still do `.get(generico, [])` on it.
OBJ_ONLY_MARCAS: dict[str, list[str]] = {}

# All marcas that roll up into each generico total (CERVEZAS = both halves +
# OBJ_ONLY_MARCAS). Extra marcas from OBJ_ONLY_MARCAS never become columns —
# they only enter the array literal of the "TOTAL {generico}" OBJ formula.
GENERICO_MARCAS: dict[str, list[str]] = {}
for _title, _generico, _marcas, _ in SECTIONS:
    GENERICO_MARCAS.setdefault(_generico, []).extend(_marcas)
for _generico, _extra_marcas in OBJ_ONLY_MARCAS.items():
    GENERICO_MARCAS.setdefault(_generico, []).extend(_extra_marcas)

SUBHEADERS = ["PDV", "OBJ", "Faltan", "%"]

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FILL = PatternFill("solid", fgColor="4472C4")
_MARCA_FILL = PatternFill("solid", fgColor="DDEBF7")
_GENERICO_FILL = PatternFill("solid", fgColor="C6E0B4")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_CENTER = Alignment(horizontal="center", vertical="center")


def _marca_array(marcas: list[str]) -> str:
    """Render an Excel array literal, e.g. {"SALTA";"HEINEKEN";...}."""
    return "{" + ";".join(f'"{m}"' for m in marcas) + "}"


def _pdv_marca_formula(row: int, col_letter: str, header_row: int, *, with_preventista: bool = True) -> str:
    criteria = f',COBER!$B:$B,$A{row}' if with_preventista else ""
    return (
        f'=SUMIFS(COBER!$E:$E,COBER!$A:$A,"{SUC}"{criteria},'
        f'COBER!$D:$D,{col_letter}${header_row})'
    )


def _obj_marca_formula(row: int, col_letter: str, header_row: int, *, with_preventista: bool = True) -> str:
    criteria = f',CuposCober!$C:$C,$A{row}' if with_preventista else ""
    return (
        f'=SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"{SUC}"{criteria},'
        f'CuposCober!$D:$D,{col_letter}${header_row})'
    )


def _pdv_generico_formula(row: int, generico: str, *, with_preventista: bool = True) -> str:
    """Read the generico grain (H/I/K/L) so a client covered under two marcas
    of the same generico is counted once — this must NOT be a sum of the
    marca PDV cells."""
    criteria = f',COBER!$I:$I,$A{row}' if with_preventista else ""
    return (
        f'=SUMIFS(COBER!$L:$L,COBER!$H:$H,"{SUC}"{criteria},'
        f'COBER!$K:$K,"{generico}")'
    )


def _obj_generico_formula(row: int, generico: str, *, with_preventista: bool = True) -> str:
    """Generico-level cupo is not published; roll it up from the marca cupos."""
    criteria = f',CuposCober!$C:$C,$A{row}' if with_preventista else ""
    return (
        f'=SUM(SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"{SUC}"{criteria},'
        f'CuposCober!$D:$D,{_marca_array(GENERICO_MARCAS[generico])}))'
    )


def _write_group(ws, row: int, col: int, pdv_formula: str, obj_formula: str, *, fill=None) -> None:
    """Write a PDV/OBJ/Faltan/% quartet starting at `col`."""
    pdv_letter = get_column_letter(col)
    obj_letter = get_column_letter(col + 1)
    ws.cell(row=row, column=col, value=pdv_formula)
    ws.cell(row=row, column=col + 1, value=obj_formula)
    ws.cell(row=row, column=col + 2, value=f"={obj_letter}{row}-{pdv_letter}{row}")
    # Blank (not a fake 0%) when OBJ is 0 — e.g. DIRECTA has PDV but no
    # published cupo; a literal 0% would read as "sold nothing", the
    # opposite of the truth. OBJ is always a SUMIFS result (never an
    # error), so IFERROR is unnecessary here.
    ws.cell(row=row, column=col + 3, value=f'=IF({obj_letter}{row}=0,"",{pdv_letter}{row}/{obj_letter}{row})')
    for i in range(4):
        cell = ws.cell(row=row, column=col + i)
        # Percentage column shows exactly 2 decimals; the value is never
        # rounded (formatting only). Everything else is an integer count.
        cell.number_format = "0.00%" if i == 3 else "#,##0"
        cell.border = _BORDER
        if fill:
            cell.fill = fill


def _build_section(
    ws, top_row: int, title: str, generico: str, marcas: list[str], emit_generico_total: bool
) -> tuple[int, int]:
    """Write one section starting at `top_row`. Returns (next_top_row, section_width)."""
    groups = [(m, "marca") for m in marcas]
    if emit_generico_total:
        groups.append((f"TOTAL {generico}", "generico"))
    width = 1 + 4 * len(groups)

    ws.merge_cells(start_row=top_row, start_column=1, end_row=top_row, end_column=width)
    title_cell = ws.cell(row=top_row, column=1, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = _CENTER

    header_row = top_row + 1
    subheader_row = top_row + 2
    for group_index, (label, kind) in enumerate(groups):
        col = 2 + 4 * group_index
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 3)
        header_cell = ws.cell(row=header_row, column=col, value=label)
        header_cell.font = Font(bold=True, size=9)
        header_cell.fill = _GENERICO_FILL if kind == "generico" else _MARCA_FILL
        header_cell.alignment = _CENTER
        for i in range(4):
            ws.cell(row=header_row, column=col + i).border = _BORDER
            sub_cell = ws.cell(row=subheader_row, column=col + i, value=SUBHEADERS[i])
            sub_cell.font = Font(bold=True, size=9)
            sub_cell.fill = _SUBHEADER_FILL
            sub_cell.border = _BORDER
            sub_cell.alignment = _CENTER

    preventista_header = ws.cell(row=subheader_row, column=1, value="Preventista")
    preventista_header.font = Font(bold=True, size=9)
    preventista_header.fill = _SUBHEADER_FILL
    preventista_header.border = _BORDER

    first_data_row = subheader_row + 1
    for preventista_index, preventista in enumerate(PREVENTISTAS):
        row = first_data_row + preventista_index
        name_cell = ws.cell(row=row, column=1, value=preventista)
        name_cell.border = _BORDER
        for group_index, (label, kind) in enumerate(groups):
            col = 2 + 4 * group_index
            col_letter = get_column_letter(col)
            if kind == "generico":
                _write_group(ws, row, col, _pdv_generico_formula(row, generico), _obj_generico_formula(row, generico))
            else:
                _write_group(
                    ws, row, col,
                    _pdv_marca_formula(row, col_letter, header_row),
                    _obj_marca_formula(row, col_letter, header_row),
                )

    total_row = first_data_row + len(PREVENTISTAS)
    total_name_cell = ws.cell(row=total_row, column=1, value="TOTAL GUEMES")
    total_name_cell.font = Font(bold=True)
    total_name_cell.fill = _TOTAL_FILL
    total_name_cell.border = _BORDER
    for group_index, (label, kind) in enumerate(groups):
        col = 2 + 4 * group_index
        col_letter = get_column_letter(col)
        if kind == "generico":
            _write_group(
                ws, total_row, col,
                _pdv_generico_formula(total_row, generico, with_preventista=False),
                _obj_generico_formula(total_row, generico, with_preventista=False),
                fill=_TOTAL_FILL,
            )
        else:
            _write_group(
                ws, total_row, col,
                _pdv_marca_formula(total_row, col_letter, header_row, with_preventista=False),
                _obj_marca_formula(total_row, col_letter, header_row, with_preventista=False),
                fill=_TOTAL_FILL,
            )
        for i in range(4):
            ws.cell(row=total_row, column=col + i).font = Font(bold=True)

    return total_row + 2, width


def build_cobertura_ccu(wb: Workbook) -> None:
    """Create (or replace) the 'Cobertura CCU' sheet on `wb`, in place.

    Idempotent: if the sheet already exists it is deleted and rebuilt from
    scratch, so re-running never duplicates rows, sections, or shifts data.
    """
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 0)

    row = 1
    widest_col = 0
    for title, generico, marcas, emit_generico_total in SECTIONS:
        row, width = _build_section(ws, row, title, generico, marcas, emit_generico_total)
        widest_col = max(widest_col, width)

    ws.column_dimensions["A"].width = 20
    for col in range(2, widest_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8.5
    ws.sheet_view.showGridLines = False


def main(path_str: str) -> int:
    path = Path(path_str)
    if not path.exists():
        print(f"ERROR: file not found: {path}")
        return 1

    wb = load_workbook(str(path))

    # Validate BEFORE backing up: an aborted run must not litter a ~5 MB
    # backup file. Both checks below run first, save/backup happens after.
    if "COBER" not in wb.sheetnames or "CuposCober" not in wb.sheetnames:
        print("ABORT: workbook is missing COBER and/or CuposCober — not a valid GUEMES template.")
        return 1

    cupos_ws = wb["CuposCober"]
    if not any(isinstance(cell.value, str) and SUC in cell.value for cell in cupos_ws["E"]):
        print(
            f"ABORT: {SUC!r} not found in CuposCober column E — refusing to "
            "build a sheet of zeros onto the wrong workbook."
        )
        return 1

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    # Stem must contain "backup" so `_is_backup_name` (src/services/avances/
    # service.py) recognizes and excludes it from next-month seeding.
    backup = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    print(f"Backup: {backup.name}")

    build_cobertura_ccu(wb)

    wb.save(str(path))
    print(f"'{SHEET}' sheet written to {path}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: python scripts/add_cobertura_ccu_sheet.py "<AVANCE GUEMES.xlsx>"')
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
