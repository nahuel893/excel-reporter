"""Apertura por ruta del cupo de COBERTURA del interior (11 sucursales).

Hermano de `apertura_cobertura_por_ruta.py`, que cubre Casa Central, Valle y
Guemes. Este toma el archivo que manda CCU para el interior —una hoja por
sucursal— y baja el cupo de clientes a ruta, ponderando por la cobertura real
del mes anterior.

Dos diferencias con el de Casa Central que NO son un descuido
-------------------------------------------------------------
1. El Cupo esta en la columna D, no en la E: la fuente del interior tiene
   cuatro columnas (Marca | Cob JULIO 2026 | Cob AGOSTO 2025 | Cupo) contra
   las cinco de la otra.
2. La cobertura de referencia se lee filtrando `id_fuerza_ventas = 1`.
   `cob_preventista_*` trae fv1 y fv4 y el MISMO cliente aparece en las dos,
   atribuido a su ruta de cerveza y a su ruta de Branca. Sin el filtro los
   pesos salen casi al doble. (Verificado: CERVEZAS suc.1 julio da 4.325 con
   filtro y 8.141 sin el, sobre 4.996 clientes que existen.)

Uso
---
    python scripts/apertura_cobertura_interior.py [--force]
"""
from __future__ import annotations

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

PERIODO = "2026-08"
PERIODO_HISTORIA = "2026-07-01"
FUENTE = ROOT / "data/output/cupos" / PERIODO / "Cobertura y Cupos CCU - Interior.xlsx"
SALIDA = ROOT / "data/output/cupos" / PERIODO / "CUPO COBERTURA POR RUTA INTERIOR - AGOSTO 2026.xlsx"

COL_CUPO = 4                 # columna D
DECIMALES = 4                # cobertura reparte clientes: 2 pierden resolucion
GENERICOS = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES",
             "PERNOD RICARD"]
# Un cupo de 5 clientes entre 20 rutas deja a casi todas en cero. El piso
# garantiza que cada ruta tenga algo que perseguir.
GENERICOS_CON_PISO = {"VINOS CCU", "SIDRAS Y LICORES"}
PISO_CLIENTES = 1.0
SIN_CUPO = {"DIRECTA", "RUIZ MARCELO", "VENDEDOR CHOPERAS"}

# MULTI CCU no existe como generico ni tiene fila propia en la fuente: es la
# UNION de estos tres y toma el cupo de VINOS CCU en cada sucursal. Mismo
# criterio que en Casa Central/Valle/Guemes (apertura_cobertura_por_ruta.py).
MULTICCU = "MULTI CCU"
MULTICCU_GENERICOS = ("VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD")
MULTICCU_CUPO_DE = "VINOS CCU"
AMARILLO = "FFFF00"


def _txt(v) -> str:
    if v is None or (isinstance(v, float) and v != v):
        return ""
    s = unicodedata.normalize("NFKD", str(v).strip())
    return "".join(c for c in s if not unicodedata.combining(c))


def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    pos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(pos)
    if suma <= 0:
        partes = [round(total / n, DECIMALES)] * (n - 1)
        partes.append(round(total - sum(partes), DECIMALES))
        return partes
    partes = [round(total * p / suma, DECIMALES) for p in pos]
    residuo = round(total - sum(partes), DECIMALES)
    if residuo:
        i = max(range(n), key=lambda k: pos[k])
        partes[i] = round(partes[i] + residuo, DECIMALES)
    return partes


def split_con_piso(total: float, pesos: list[float], piso: float) -> list[float]:
    """Cada ruta arranca en `piso`; el excedente se reparte por historia.

    Si el cupo no alcanza a cubrir el piso de todas, el cupo SUBE — se prefiere
    pedir de mas antes que dejar rutas sin objetivo.
    """
    n = len(pesos)
    if n == 0:
        return []
    excedente = total - piso * n
    if excedente <= 0:
        return [piso] * n
    return [round(piso + p, DECIMALES) for p in split_proporcional(excedente, pesos)]


def leer_cupos(path: Path):
    """(sucursal, generico, marca) -> cupo, y (sucursal, generico) -> cupo."""
    wb = openpyxl.load_workbook(path, data_only=True)
    por_marca, por_gen, anotaciones = {}, {}, []
    try:
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            actual = None
            for r in range(1, ws.max_row + 1):
                a = _txt(ws.cell(row=r, column=1).value)
                cupo = ws.cell(row=r, column=COL_CUPO).value
                if a.upper() in GENERICOS and not _txt(ws.cell(row=r, column=2).value):
                    actual = a.upper()
                    continue
                if not a or a == "Marca" or actual is None:
                    continue
                if a.upper().startswith("TOTAL "):
                    if isinstance(cupo, (int, float)):
                        por_gen[(hoja, actual)] = float(cupo)
                    actual = None
                    continue
                if isinstance(cupo, (int, float)):
                    por_marca[(hoja, actual, a.upper())] = float(cupo)
                elif cupo is not None:
                    anotaciones.append(f"{hoja} / {actual} / {a}: {cupo!r}")
        return por_marca, por_gen, anotaciones
    finally:
        wb.close()


def cargar_rutas(dl: DataLoader, sucursales: list[str]):
    df = dl.execute_query("""
        SELECT ds.descripcion AS sucursal, dc.id_sucursal,
               dc.id_ruta_fv1 AS ruta,
               MIN(dc.des_ruta_fv1) AS des_ruta,
               UPPER(TRIM(dc.des_personal_fv1)) AS preventista
        FROM gold.dim_cliente dc
        JOIN gold.dim_sucursal ds ON ds.id_sucursal = dc.id_sucursal
        WHERE COALESCE(dc.anulado, false) = false
          AND ds.descripcion IN :sucs
          AND dc.id_ruta_fv1 IS NOT NULL AND dc.des_personal_fv1 IS NOT NULL
        GROUP BY 1, 2, 3, 5
    """, {"sucs": tuple(sucursales)})
    rutas: dict[str, list] = {}
    for f in df.itertuples(index=False):
        if _txt(f.preventista).upper() in SIN_CUPO:
            continue
        rutas.setdefault(f.sucursal, []).append(
            (int(f.id_sucursal), int(f.ruta),
             _txt(f.des_ruta) or f"RUTA {int(f.ruta)}", _txt(f.preventista)))
    for v in rutas.values():
        v.sort(key=lambda x: x[1])
    return rutas


def cargar_cobertura(dl: DataLoader, tabla: str, campo: str, sucursales: list[str]):
    """(id_sucursal, ruta, clave) -> clientes. SOLO fuerza de ventas 1."""
    df = dl.execute_query(f"""
        SELECT c.id_sucursal, c.id_ruta, UPPER(TRIM(c.{campo})) AS clave,
               SUM(c.clientes_compradores) AS clientes
        FROM gold.{tabla} c
        JOIN gold.dim_sucursal ds ON ds.id_sucursal = c.id_sucursal
        WHERE c.periodo = :p AND ds.descripcion IN :sucs
          AND c.id_fuerza_ventas = 1
        GROUP BY 1, 2, 3
    """, {"p": PERIODO_HISTORIA, "sucs": tuple(sucursales)})
    return {(int(f.id_sucursal), int(f.id_ruta), f.clave): float(f.clientes or 0)
            for f in df.itertuples(index=False)}


def cargar_cobertura_multiccu(dl: DataLoader, sucursales: list[str]):
    """(id_sucursal, ruta, 'MULTI CCU') -> clientes de la UNION.

    No se puede sumar la cobertura de los tres genericos: quien compra vinos y
    pernod se contaria dos veces. Hay que ir al grano de cliente — totalizar su
    volumen DENTRO del corte de los tres y recien ahi contar los que dan > 0.

    La ruta sale del `dim_cliente` de HOY, igual que el resto del reparto, y se
    filtra `id_fuerza_ventas = 1` para reproducir `cob_preventista_*`.
    """
    df = dl.execute_query("""
        SELECT id_sucursal, id_ruta, COUNT(*) AS clientes FROM (
          SELECT dc.id_sucursal, dc.id_ruta_fv1 AS id_ruta, fv.id_cliente,
                 SUM(fv.cantidades_total) AS vol
          FROM gold.fact_ventas fv
          JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                  AND dc.id_sucursal = fv.id_sucursal
          JOIN gold.dim_sucursal ds ON ds.id_sucursal = fv.id_sucursal
          JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
          JOIN gold.dim_vendedor dv ON dv.id_vendedor = dc.id_personal_fv1
                                   AND dv.id_sucursal = dc.id_sucursal
          WHERE ds.descripcion IN :sucs AND fv.anulado = false
            AND dv.id_fuerza_ventas = 1
            AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
            AND da.generico IN :gen
          GROUP BY 1, 2, 3
        ) t WHERE vol > 0 GROUP BY 1, 2
    """, {"d": PERIODO_HISTORIA, "h": f"{PERIODO}-01",
          "gen": MULTICCU_GENERICOS, "sucs": tuple(sucursales)})
    return {(int(f.id_sucursal), int(f.id_ruta), MULTICCU): float(f.clientes)
            for f in df.itertuples(index=False)}


def repartir(cupos: dict, rutas: dict, cobertura: dict, es_marca: bool):
    filas, sin_ref, subidos = [], [], []
    for clave, cupo in sorted(cupos.items()):
        hoja, generico = clave[0], clave[1]
        etiqueta = clave[2] if es_marca else clave[1]
        de_hoja = rutas.get(hoja, [])
        if not de_hoja:
            continue
        pesos = [cobertura.get((s, r, etiqueta), 0.0) for s, r, _, _ in de_hoja]
        if sum(pesos) <= 0 and cupo > 0:
            sin_ref.append(f"{hoja} / {etiqueta}")
        if generico in GENERICOS_CON_PISO:
            partes = split_con_piso(float(cupo), pesos, PISO_CLIENTES)
            asignado = round(sum(partes), DECIMALES)
            if asignado > float(cupo) + 0.01:
                subidos.append(f"{hoja} / {etiqueta}: {cupo:g} -> {asignado:g} "
                               f"({len(de_hoja)} rutas)")
        else:
            partes = split_proporcional(float(cupo), pesos)
        for (id_suc, ruta, des_ruta, prev), parte in zip(de_hoja, partes):
            # Posiciones: 0 ruta | 1 desc | 2 preventista | 3 etiqueta | 4 zona | 5 cupo
            filas.append([ruta, des_ruta if es_marca else None, prev, etiqueta,
                          f"{id_suc} - {hoja}", parte])
    return filas, sin_ref, subidos


def escribir(wb, titulo, filas, es_marca):
    ws = wb.create_sheet(titulo)
    ws.append(["Ruta", "Descripción Ruta" if es_marca else None,
               "Descripción Vendedor" if es_marca else "Preventista",
               "MARCA" if es_marca else "Generico", "ZONA", "CUPO"])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    for f in filas:
        ws.append(f)
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(fila):
            c.border = borde
            if i == 0:
                c.number_format = "0"
            if i == 5:
                c.number_format = "#,##0.0000"
    for col, w in zip("ABCDEF", (9, 26, 26, 20, 28, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    # Amarillo: estas dos hojas son las que se cargan en la base.
    ws.sheet_properties.tabColor = AMARILLO
    return ws


def main() -> int:
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"Ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        return 1
    if not FUENTE.exists():
        print(f"No existe la fuente: {FUENTE}")
        return 2

    cupos_marca, cupos_gen, anotaciones = leer_cupos(FUENTE)
    sucursales = sorted({k[0] for k in cupos_gen})
    dl = DataLoader()
    rutas = cargar_rutas(dl, sucursales)

    print(f"Fuente: {FUENTE.name}")
    print(f"  {len(sucursales)} sucursales | {len(cupos_marca)} cupos por marca "
          f"| {len(cupos_gen)} por generico")
    if anotaciones:
        print(f"  Celdas de Cupo con texto, salteadas ({len(anotaciones)}):")
        for x in anotaciones[:6]:
            print(f"    {x}")
    faltan = [s for s in sucursales if not rutas.get(s)]
    if faltan:
        print(f"  SIN RUTAS en dim_cliente: {faltan}")

    cob_m = cargar_cobertura(dl, "cob_preventista_marca", "marca", sucursales)
    cob_g = cargar_cobertura(dl, "cob_preventista_generico", "generico", sucursales)

    # MULTI CCU: mismo cupo que VINOS CCU en cada sucursal. Las sucursales sin
    # cupo de VINOS CCU (p. ej. JOAQUIN V GONZALEZ) tampoco reciben MULTI CCU.
    for (hoja, generico), cupo in list(cupos_gen.items()):
        if generico == MULTICCU_CUPO_DE:
            cupos_gen[(hoja, MULTICCU)] = cupo
    cob_g.update(cargar_cobertura_multiccu(dl, sucursales))

    print("Reparto por marca:")
    fm, sin_m, sub_m = repartir(cupos_marca, rutas, cob_m, True)
    print("Reparto por generico:")
    fg, sin_g, sub_g = repartir(cupos_gen, rutas, cob_g, False)
    for etiqueta, sin_ref, subidos in (("marca", sin_m, sub_m), ("generico", sin_g, sub_g)):
        if sin_ref:
            print(f"  [{etiqueta}] sin cobertura de referencia (reparto parejo): "
                  f"{len(sin_ref)} — ej. {sorted(set(sin_ref))[:3]}")
        if subidos:
            print(f"  [{etiqueta}] cupo SUBIDO por el piso de {PISO_CLIENTES:g} "
                  f"cliente ({len(subidos)}):")
            for x in subidos[:5]:
                print(f"    {x}")

    wb = Workbook()
    wb.remove(wb.active)
    escribir(wb, "apertura_marca", fm, True)
    escribir(wb, "apertura_generico", fg, False)
    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA)
    print(f"\nGuardado: {SALIDA.name}")
    print(f"  apertura_marca   : {len(fm)} filas")
    print(f"  apertura_generico: {len(fg)} filas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
