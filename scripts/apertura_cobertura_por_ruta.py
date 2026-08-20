"""Abre los cupos de COBERTURA por ruta, a partir del informe de zonas.

Entrada: el xlsx de "Cobertura y Cupos CCU" con una hoja por zona. Cada hoja
trae bloques por generico: filas de marca con su `Cupo`, y una fila
`TOTAL <generico>` con el cupo del generico.

Salida: un xlsx con las dos hojas que espera `medallion-etl`
(`bronze/loaders/cupos_cobertura_loader.py`), que lee **por posicion**:

    apertura_marca     -> [ruta, desc_ruta, preventista, MARCA,    zona, cupo]
    apertura_generico  -> [ruta, desc_ruta, preventista, GENERICO, zona, cupo]

La columna 3 cae en `generico` para la hoja de marca y en `marca` para la de
generico — estan invertidas en la tabla, pero el loader lo resuelve solo. Acá
se emite el formato ORIGEN, sin pre-transformar.

Por que el cupo del generico NO es la suma de sus marcas
--------------------------------------------------------
La cobertura cuenta CLIENTES. Un cliente que compra SALTA e IMPERIAL es uno
solo en CERVEZAS pero dos si se suman las marcas. Por eso el cupo del generico
se toma de la fila TOTAL del informe y nunca se deriva. (En CASA CENTRAL las
marcas de cerveza suman 13.260 contra un TOTAL CERVEZAS de 4.000.)

Si es aditiva entre RUTAS: cada cliente pertenece a una sola. Por eso el
reparto por ruta es legitimo, y se pondera con la cobertura real del mes de
historia (`gold.cob_preventista_*`), no con volumen.

Uso
---
    python scripts/apertura_cobertura_por_ruta.py [--force]
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

# --- Periodos ---------------------------------------------------------------
PERIODO = "2026-08"                 # el mes del cupo
PERIODO_HISTORIA = "2026-07-01"     # cobertura que pondera el reparto
FUENTE = Path("/home/nahuel/Documents/Cobertura y Cupos CCU.xlsx")
SALIDA = ROOT / "data/output/cupos" / PERIODO / "cupos_cobertura agosto 2026.xlsx"

# --- Zonas ------------------------------------------------------------------
# La hoja del informe -> (id_sucursal, etiqueta que va en la columna ZONA).
# VALLE SALTA es sucursal 1 igual que CASA CENTRAL: en la base las dos se
# guardan como "1 - CASA CENTRAL", la zona solo separa que rutas van en cada
# hoja del informe de origen.
ZONAS: dict[str, tuple[int, str]] = {
    "CASA CENTRAL": (1, "1 - CASA CENTRAL"),
    "VALLE SALTA": (1, "1 - CASA CENTRAL"),
    "SUCURSAL GUEMES": (16, "16 - SUCURSAL GUEMES"),
}
RUTAS_VALLE = {81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 118, 119, 120, 122}
RUTAS_SUBD = {93}
# SUB DISTRIBUIDORES no entra en el reparto proporcional: tiene 6 clientes y el
# objetivo es cubrirlos a TODOS en cada categoria y marca, asi que su cupo es
# fijo en 6 y no sale del cupo de ninguna zona. Sus filas se agregan aparte
# (ver filas_subdistribuidor), no via `repartir`.
SUBD_RUTA = 93
SUBD_DES_RUTA = "SUB DISTRIBUIDORES"
SUBD_PREVENTISTA = "SUB DISTRIBUIDOR"
SUBD_ZONA_LABEL = "1 - CASA CENTRAL"
SUBD_CUPO = 6.0
# DIRECTA (100) y CERVECERA (200) no son rutas de preventa.
RUTAS_SIN_CUPO = {77, 100, 200, 999}

# Preventistas de baja. RUIZ MARCELO tiene 217 clientes asignados y 4 activos:
# sus rutas facturaron en julio y recien despues se reasignaron los clientes.
# OJO: `cob_preventista_*` guarda la ruta como estaba en el mes medido, asi que
# sus rutas SI aparecen con cobertura (302/306/324) — a diferencia de
# `fact_ventas`, que se cruza contra el dim_cliente de hoy y lo deja en cero.
# Sin esta exclusion se le asigna objetivo a clientes que ya no atiende.
VENDEDORES_SIN_CUPO = {"RUIZ MARCELO"}

GENERICOS = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES",
             "PERNOD RICARD"]

# Genericos donde TODA ruta lleva al menos un cliente de objetivo, aunque no
# tenga historia: son los de portafolio, y un cero le saca al preventista la
# obligacion de intentarlo. El piso se aplica a sus MARCAS.
GENERICOS_CON_PISO = {"VINOS CCU", "SIDRAS Y LICORES"}
PISO_CLIENTES = 1.0


def _texto(v) -> str:
    if v is None or (isinstance(v, float) and v != v):
        return ""
    return str(v).strip()


# --- Lectura del informe de zonas -------------------------------------------
def leer_cupos(path: Path) -> tuple[dict, dict]:
    """Devuelve (cupos_marca, cupos_generico).

    cupos_marca[(zona, generico, marca)] = cupo
    cupos_generico[(zona, generico)]     = cupo de la fila TOTAL
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        por_marca: dict[tuple[str, str, str], float] = {}
        por_generico: dict[tuple[str, str], float] = {}
        anotaciones: list[str] = []
        for zona in ZONAS:
            if zona not in wb.sheetnames:
                raise ValueError(f"Falta la hoja {zona!r}. Hay: {wb.sheetnames}")
            ws = wb[zona]
            generico_actual = None
            for r in range(1, ws.max_row + 1):
                a = _texto(ws.cell(row=r, column=1).value)
                cupo = ws.cell(row=r, column=5).value
                if a in GENERICOS and not _texto(ws.cell(row=r, column=2).value):
                    generico_actual = a          # cabecera del bloque
                    continue
                if not a or a == "Marca" or generico_actual is None:
                    continue
                if a.upper().startswith("TOTAL "):
                    if isinstance(cupo, (int, float)):
                        por_generico[(zona, generico_actual)] = float(cupo)
                    generico_actual = None       # se cierra el bloque
                    continue
                if isinstance(cupo, (int, float)):
                    por_marca[(zona, generico_actual, a)] = float(cupo)
                elif cupo is not None:
                    # Celdas con texto ("no vendemos", etc.): son anotaciones,
                    # no cupos. Se saltean pero se informan — un cupo perdido
                    # en silencio es una marca que se queda sin objetivo.
                    anotaciones.append(
                        f"{zona} / {generico_actual} / {a}: {cupo!r}")
        if anotaciones:
            print(f"  Celdas de Cupo con texto, salteadas ({len(anotaciones)}):")
            for x in anotaciones:
                print(f"    {x}")
        return por_marca, por_generico
    finally:
        wb.close()


# --- Rutas y cobertura real -------------------------------------------------
def zona_de(id_sucursal: int, ruta: int) -> str | None:
    if id_sucursal == 16:
        return "SUCURSAL GUEMES"
    if ruta in RUTAS_SUBD:
        return None                  # SUB DISTRIBUIDORES no entra en el informe
    return "VALLE SALTA" if ruta in RUTAS_VALLE else "CASA CENTRAL"


def cargar_rutas(dl: DataLoader) -> dict[str, list[tuple[int, str, str]]]:
    """zona -> [(id_ruta, des_ruta, preventista)]."""
    df = dl.execute_query("""
        SELECT dc.id_sucursal, dc.id_ruta_fv1 AS ruta,
               MIN(dc.des_ruta_fv1) AS des_ruta,
               MIN(dv.des_vendedor) AS preventista
        FROM gold.dim_cliente dc
        LEFT JOIN gold.dim_vendedor dv ON dv.id_vendedor = dc.id_personal_fv1
                                      AND dv.id_sucursal = dc.id_sucursal
        WHERE dc.id_sucursal IN (1, 16)
          AND COALESCE(dc.anulado, false) = false
          AND dc.id_ruta_fv1 IS NOT NULL AND dc.id_personal_fv1 IS NOT NULL
        GROUP BY dc.id_sucursal, dc.id_ruta_fv1
        ORDER BY dc.id_sucursal, dc.id_ruta_fv1
    """)
    rutas: dict[str, list[tuple[int, str, str]]] = {z: [] for z in ZONAS}
    excluidas: list[str] = []
    for f in df.itertuples(index=False):
        ruta = int(f.ruta)
        preventista = _texto(f.preventista) or "SIN PREVENTISTA"
        if ruta in RUTAS_SIN_CUPO or preventista in VENDEDORES_SIN_CUPO:
            excluidas.append(f"{ruta} ({preventista})")
            continue
        zona = zona_de(int(f.id_sucursal), ruta)
        if zona is None:
            continue
        rutas[zona].append((ruta, _texto(f.des_ruta) or f"RUTA {ruta}",
                            preventista))
    if excluidas:
        print(f"  Rutas sin cupo ({len(excluidas)}): {', '.join(excluidas)}")
    return rutas


# MULTI CCU no existe como generico: es la union de estos tres. No tiene fila
# propia en el informe, asi que toma el cupo de VINOS CCU en cada zona.
MULTICCU = "MULTI CCU"
MULTICCU_GENERICOS = ("VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD")
MULTICCU_CUPO_DE = "VINOS CCU"


def cargar_cobertura_multiccu(dl: DataLoader) -> dict:
    """(id_sucursal, id_ruta, 'MULTI CCU') -> clientes de la UNION.

    No se puede sumar la cobertura de los tres genericos: quien compra vinos y
    pernod se contaria dos veces. Hay que ir al grano de cliente — totalizar su
    volumen DENTRO del corte de los tres y recien ahi contar los que dan > 0.
    En julio 2026 eso da 1.841 contra 4.127 sumando, o sea menos de la mitad.

    La ruta sale del `dim_cliente` de HOY, igual que el resto del reparto: el
    objetivo va a quien atiende al cliente ahora, no a quien lo atendia.
    """
    df = dl.execute_query("""
        SELECT id_sucursal, id_ruta, COUNT(*) AS clientes FROM (
          SELECT dc.id_sucursal, dc.id_ruta_fv1 AS id_ruta, fv.id_cliente,
                 SUM(fv.cantidades_total) AS vol
          FROM gold.fact_ventas fv
          JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                  AND dc.id_sucursal = fv.id_sucursal
          JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
          JOIN gold.dim_vendedor dv ON dv.id_vendedor = dc.id_personal_fv1
                                   AND dv.id_sucursal = dc.id_sucursal
          WHERE fv.id_sucursal IN (1, 16) AND fv.anulado = false
            AND dv.id_fuerza_ventas = 1
            AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
            AND da.generico IN :gen
          GROUP BY 1, 2, 3
        ) t WHERE vol > 0 GROUP BY 1, 2
    """, {"d": PERIODO_HISTORIA, "h": "2026-08-01", "gen": MULTICCU_GENERICOS})
    return {(int(f.id_sucursal), int(f.id_ruta), MULTICCU): float(f.clientes)
            for f in df.itertuples(index=False)}


def cargar_cobertura(dl: DataLoader, tabla: str, campo: str) -> dict:
    """(id_sucursal, id_ruta, marca|generico) -> clientes compradores."""
    df = dl.execute_query(f"""
        SELECT id_sucursal, id_ruta, {campo} AS clave,
               SUM(clientes_compradores) AS clientes
        FROM gold.{tabla}
        WHERE periodo = :p AND id_sucursal IN (1, 16)
        GROUP BY 1, 2, 3
    """, {"p": PERIODO_HISTORIA})
    return {(int(f.id_sucursal), int(f.id_ruta), _texto(f.clave)): float(f.clientes or 0)
            for f in df.itertuples(index=False)}


# --- Reparto ----------------------------------------------------------------
def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`; el residuo va al de mayor cobertura."""
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, 4)] * (n - 1)
        partes.append(round(total - sum(partes), 4))
        return partes
    partes = [round(total * p / suma, 4) for p in positivos]
    residuo = round(total - sum(partes), 4)
    if residuo:
        i = max(range(n), key=lambda k: positivos[k])
        partes[i] = round(partes[i] + residuo, 4)
    return partes


def split_con_piso(total: float, pesos: list[float], piso: float) -> list[float]:
    """Reparte `total` garantizando `piso` en TODAS las rutas.

    Cada ruta arranca en `piso` y el excedente se reparte por historia. Si el
    cupo no alcanza para cubrir el piso, el cupo SUBE a `piso * rutas`: la
    alternativa seria dejar rutas en cero, que es justo lo que el piso evita.
    """
    n = len(pesos)
    if n == 0:
        return []
    excedente = total - piso * n
    if excedente <= 0:
        return [piso] * n
    partes = split_proporcional(excedente, pesos)
    return [round(piso + p, 4) for p in partes]


def repartir(cupos: dict, rutas: dict, cobertura: dict, es_marca: bool) -> list[list]:
    """Genera las filas en el orden posicional que espera el loader."""
    filas: list[list] = []
    sin_cobertura: list[str] = []
    subidos: list[str] = []
    for clave, cupo in sorted(cupos.items()):
        zona, generico = clave[0], clave[1]
        etiqueta = clave[2] if es_marca else clave[1]
        id_sucursal, zona_label = ZONAS[zona]
        de_zona = rutas[zona]
        if not de_zona:
            continue
        pesos = [cobertura.get((id_sucursal, r, etiqueta), 0.0)
                 for r, _, _ in de_zona]
        if sum(pesos) <= 0 and cupo > 0:
            sin_cobertura.append(f"{zona} / {etiqueta}")

        # El piso corre en los dos niveles: en las marcas y en la fila del
        # generico. Que el generico quede cubierto por el maximo de sus marcas
        # es cierto hoy, pero es un efecto secundario — si manana se saca el
        # piso de alguna marca, el generico volveria a tener ceros.
        if generico in GENERICOS_CON_PISO:
            partes = split_con_piso(float(cupo), pesos, PISO_CLIENTES)
            asignado = round(sum(partes), 2)
            if asignado > float(cupo) + 0.01:
                subidos.append(f"{zona} / {etiqueta}: {cupo:g} -> {asignado:g} "
                               f"({len(de_zona)} rutas)")
        else:
            partes = split_proporcional(float(cupo), pesos)

        for (ruta, des_ruta, preventista), parte in zip(de_zona, partes):
            # Posiciones: 0 ruta | 1 desc | 2 preventista | 3 etiqueta | 4 zona | 5 cupo
            # La hoja de generico deja la descripcion vacia, como esta cargado
            # hoy en fact_cupos_cobertura (descripcion_ruta = NULL).
            filas.append([ruta, des_ruta if es_marca else None, preventista,
                          etiqueta, zona_label, parte])
    if sin_cobertura:
        print(f"  Sin cobertura de referencia (reparto parejo): "
              f"{len(sin_cobertura)} — ej. {sorted(set(sin_cobertura))[:3]}")
    if subidos:
        print(f"  Cupo SUBIDO por el piso de {PISO_CLIENTES:g} cliente "
              f"({len(subidos)}): el cupo no alcanzaba a cubrir todas las rutas")
        for x in subidos:
            print(f"    {x}")
    return filas


def filas_subdistribuidor(cupos_marca: dict, cupos_generico: dict):
    """Filas de cupo fijo para la ruta de SUB DISTRIBUIDORES.

    Una fila por cada marca y por cada generico que exista en el objetivo de
    sucursal 1 (CASA CENTRAL + VALLE SALTA), todas con el mismo cupo. No se
    reparte nada: son 6 clientes y el objetivo es cubrir los 6 en todo.
    """
    zonas_suc1 = {z for z, (id_suc, _) in ZONAS.items() if id_suc == 1}
    marcas = sorted({k[2] for k in cupos_marca if k[0] in zonas_suc1})
    genericos = sorted({k[1] for k in cupos_generico if k[0] in zonas_suc1})
    base = [SUBD_RUTA, SUBD_DES_RUTA, SUBD_PREVENTISTA]
    fm = [[*base, m, SUBD_ZONA_LABEL, SUBD_CUPO] for m in marcas]
    fg = [[SUBD_RUTA, None, SUBD_PREVENTISTA, g, SUBD_ZONA_LABEL, SUBD_CUPO]
          for g in genericos]
    print(f"  SUB DISTRIBUIDORES (ruta {SUBD_RUTA}): cupo fijo {SUBD_CUPO:g} — "
          f"{len(fm)} marcas, {len(fg)} genericos")
    return fm, fg


# --- Excel ------------------------------------------------------------------
HEADERS_MARCA = ["Ruta", "Descripción Ruta", "Descripción Vendedor", "MARCA",
                 "ZONA", "CUPO"]
HEADERS_GENERICO = ["Ruta", "Descripción Ruta", "Preventista", "Generico",
                    "ZONA", "CUPO"]


def escribir(filas_marca: list[list], filas_generico: list[list]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for nombre, headers, filas in (("apertura_marca", HEADERS_MARCA, filas_marca),
                                   ("apertura_generico", HEADERS_GENERICO, filas_generico)):
        ws = wb.create_sheet(nombre)
        ws.append(headers)
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        for f in filas:
            ws.append(f)
        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for i, c in enumerate(fila):
                c.border = borde
                if i == 0:
                    c.number_format = "0"
                if i == 5:
                    c.number_format = "#,##0.0000"
        for col, w in zip("ABCDEF", (8, 26, 24, 20, 22, 12)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
    return wb


def ajustar_generico_al_maximo(filas_generico: list[list], filas_marca: list[list],
                               marca_de_generico: dict[str, str]) -> None:
    """Sube el cupo del generico a la marca mas alta de esa misma ruta.

    Si a una ruta se le piden 19 clientes de LA VICTORIA, esos 19 compraron
    sidras: el generico no puede pedir menos. Es un piso, no una suma — las
    coberturas de marca NO se suman entre si, cada una se mide sola.

    Solo se compara contra el MAXIMO. El cupo del generico sube lo necesario.
    """
    # (ruta, zona, generico) -> mayor cupo de sus marcas
    mayor: dict[tuple, float] = {}
    for ruta, _desc, _prev, marca, zona, cupo in filas_marca:
        generico = marca_de_generico.get(marca)
        if generico not in GENERICOS_CON_PISO:
            continue
        clave = (ruta, zona, generico)
        mayor[clave] = max(mayor.get(clave, 0.0), float(cupo))

    subidas: list[str] = []
    for fila in filas_generico:
        ruta, _desc, _prev, generico, zona, cupo = fila
        if generico not in GENERICOS_CON_PISO:
            continue
        piso = mayor.get((ruta, zona, generico))
        if piso is not None and piso > float(cupo) + 1e-9:
            subidas.append(f"{generico} ruta {ruta}: {float(cupo):g} -> {piso:g}")
            fila[5] = round(piso, 4)
    if subidas:
        print(f"  Generico subido al maximo de sus marcas ({len(subidas)} rutas):")
        for x in subidas[:4]:
            print(f"    {x}")
        if len(subidas) > 4:
            print(f"    ... y {len(subidas) - 4} mas")


def main() -> int:
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"El archivo ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nPara regenerarlo igual, guarda una copia y corre con --force.")
        return 1
    if not FUENTE.exists():
        print(f"Error: no existe la fuente {FUENTE}")
        return 1

    cupos_marca, cupos_generico = leer_cupos(FUENTE)
    print(f"Fuente: {FUENTE.name}")
    print(f"  cupos por marca   : {len(cupos_marca)}")
    print(f"  cupos por generico: {len(cupos_generico)}")

    dl = DataLoader()
    rutas = cargar_rutas(dl)
    for zona, rs in rutas.items():
        print(f"  {zona:18} {len(rs):3} rutas")

    print("Reparto por marca:")
    filas_marca = repartir(cupos_marca, rutas,
                           cargar_cobertura(dl, "cob_preventista_marca", "marca"),
                           es_marca=True)
    # MULTI CCU: mismo cupo que VINOS CCU en cada zona (no tiene fila propia).
    for (zona, generico), cupo in list(cupos_generico.items()):
        if generico == MULTICCU_CUPO_DE:
            cupos_generico[(zona, MULTICCU)] = cupo
    cobertura_gen = cargar_cobertura(dl, "cob_preventista_generico", "generico")
    cobertura_gen.update(cargar_cobertura_multiccu(dl))

    print("Reparto por generico:")
    filas_generico = repartir(cupos_generico, rutas, cobertura_gen, es_marca=False)

    fm_subd, fg_subd = filas_subdistribuidor(cupos_marca, cupos_generico)
    filas_marca += fm_subd
    filas_generico += fg_subd
    # El piso del generico es 1 por ruta y nada mas. Subirlo al maximo de sus
    # marcas era coherente (quien compra LA VICTORIA esta cubierto en SIDRAS)
    # pero desfasaba el cupo: SIDRAS Y LICORES saltaba de 250 a 419 porque
    # LA VICTORIA reparte 1,95 por ruta. Decision de Nahuel: manda el cupo.
    # Consecuencia asumida: en algunas rutas la marca queda por encima del
    # generico. Para reactivarlo, llamar a ajustar_generico_al_maximo().
    _ = ajustar_generico_al_maximo

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    escribir(filas_marca, filas_generico).save(SALIDA)
    print(f"\nGuardado: {SALIDA}")
    print(f"  apertura_marca   : {len(filas_marca)} filas")
    print(f"  apertura_generico: {len(filas_generico)} filas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
