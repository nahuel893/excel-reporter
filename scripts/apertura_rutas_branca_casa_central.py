"""Apertura por ruta de los cupos de BRANCA en CASA CENTRAL (fuerza de ventas 4).

Hermano de `apertura_rutas_casa_central.py`, pero para la OTRA fuerza de
ventas. Los cupos de Branca / vinos los trabaja fv4, que tiene su propia
asignacion de rutas y preventistas (`dim_cliente.id_ruta_fv4`), distinta de la
de cervezas (fv1). Por eso va en un archivo aparte.

Zonas
-----
El cupo viene por zona (CASA CENTRAL / VALLE SALTA / SUB DISTRIBUIDORES), igual
que en cervezas. Pero la zona es una propiedad geografica del CLIENTE, no de la
ruta, y las rutas de fv4 no calzan con las de fv1: 13 de 93 mezclan zonas.

Asi que el mapa de zonas NO se hardcodea: se DEDUCE cruzando los clientes con
su ruta de cervezas (fv1), cuyo mapa de zonas ya esta definido en
config/settings.py. Cada ruta de fv4 va a la zona donde esta la mayoria de sus
clientes. Con los datos de agosto 2026 eso da VALLE = {81,82,83,84,89,90,91,92,
117,120,122} y SUBD = {93}, todas con 87-100% de coincidencia salvo la 117
(55%). Se deduce en cada corrida, asi que sigue los movimientos de cartera.

Categorias
----------
- BRANCA      = generico FRATELLI B completo (fernet + Carpano, Sernova, etc.)
- ARIZU       = generico VINOS, marcas ARIZU / CANCILLER / ANIMANA1 / TORO
- VINOS FINOS = generico VINOS FINOS, EXCLUYENDO la marca QUARA
- QUARA       = marca QUARA dentro de VINOS FINOS
"Jugos rinde 2" viene sin cupo: no genera fila.

Uso
---
    python scripts/apertura_rutas_branca_casa_central.py
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

# --- Periodo ---------------------------------------------------------------
PERIODO = "2026-08"
HIST_DESDE, HIST_HASTA = "2026-07-01", "2026-08-01"
# Todos los informes de cupos van a data/output/cupos/{mes}/.
SALIDA = ROOT / "data/output/cupos" / PERIODO / (
    "CUPO DESAGREGADO POR RUTA BRANCA - AGOSTO 2026.xlsx")

# --- Zonas ------------------------------------------------------------------
# Rutas de CERVEZAS (fv1) que definen cada zona — ver config/settings.py.
# Son la referencia para deducir la zona de las rutas de fv4.
RUTAS_VALLE_FV1 = {81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92,
                   118, 119, 120, 122}
RUTAS_SUBD_FV1 = {93}
ZONAS = ["CASA CENTRAL", "VALLE SALTA", "SUB DISTRIBUIDORES"]

# DIRECTA VINOS no es una ruta de preventa: absorberia cupo que tienen que
# perseguir los preventistas. Mismo criterio que DIRECTA en el informe de fv1.
VENDEDORES_SIN_CUPO = {"DIRECTA VINOS"}

# --- Categorias -------------------------------------------------------------
CATEGORIAS = ["BRANCA", "ARIZU", "VINOS FINOS", "QUARA"]
MARCAS_ARIZU = {"ARIZU", "CANCILLER", "ANIMANA1", "TORO"}

CUPOS: dict[str, dict[str, float]] = {
    "CASA CENTRAL": {"BRANCA": 5000, "ARIZU": 2500, "VINOS FINOS": 370, "QUARA": 450},
    "VALLE SALTA": {"BRANCA": 800, "ARIZU": 500, "VINOS FINOS": 30, "QUARA": 50},
    "SUB DISTRIBUIDORES": {"BRANCA": 200, "ARIZU": 0, "VINOS FINOS": 0, "QUARA": 0},
}


def _texto(valor) -> str:
    """Normaliza a mayusculas tolerando NULL/NaN de gold."""
    if valor is None or (isinstance(valor, float) and valor != valor):
        return ""
    return str(valor).strip().upper()


def clasificar(generico, marca) -> str | None:
    """Mapea una venta a una categoria de cupo de Branca."""
    g, m = _texto(generico), _texto(marca)
    if g == "FRATELLI B":
        return "BRANCA"
    if g == "VINOS":
        return "ARIZU" if m in MARCAS_ARIZU else None
    if g == "VINOS FINOS":
        # QUARA lleva cupo propio; el resto del generico es VINOS FINOS.
        return "QUARA" if m == "QUARA" else "VINOS FINOS"
    return None


def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`, con el residuo al de MAYOR historia.

    Los pesos negativos (notas de credito) se tratan como cero: una devolucion
    no puede restar cupo.
    """
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, 2)] * (n - 1)
        partes.append(round(total - sum(partes), 2))
        return partes
    partes = [round(total * p / suma, 2) for p in positivos]
    residuo = round(total - sum(partes), 2)
    if residuo:
        mayor = max(range(n), key=lambda i: positivos[i])
        partes[mayor] = round(partes[mayor] + residuo, 2)
    return partes


# --- Datos ------------------------------------------------------------------
def deducir_zonas(dl: DataLoader) -> tuple[dict[int, str], list[tuple[int, str, str]]]:
    """Deduce la zona de cada ruta de fv4 desde las rutas de cervezas (fv1).

    Devuelve ({id_ruta_fv4: zona}, [(ruta, des_ruta, preventista)]).
    """
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv4 AS ruta, dc.id_ruta_fv1 AS ruta_cervezas,
               MIN(dc.des_ruta_fv4) AS des_ruta,
               MIN(dc.des_personal_fv4) AS preventista,
               COUNT(*) AS clientes
        FROM gold.dim_cliente dc
        WHERE dc.id_sucursal = 1
          AND COALESCE(dc.anulado, false) = false
          AND dc.id_ruta_fv4 IS NOT NULL
        GROUP BY dc.id_ruta_fv4, dc.id_ruta_fv1
    """)

    conteo: dict[int, dict[str, int]] = {}
    meta: dict[int, tuple[str, str]] = {}
    for f in df.itertuples(index=False):
        ruta = int(f.ruta)
        r1 = f.ruta_cervezas
        if r1 is None or (isinstance(r1, float) and r1 != r1):
            zona = "CASA CENTRAL"
        elif int(r1) in RUTAS_SUBD_FV1:
            zona = "SUB DISTRIBUIDORES"
        elif int(r1) in RUTAS_VALLE_FV1:
            zona = "VALLE SALTA"
        else:
            zona = "CASA CENTRAL"
        conteo.setdefault(ruta, {z: 0 for z in ZONAS})[zona] += int(f.clientes)
        meta.setdefault(ruta, (_texto(f.des_ruta) or f"RUTA {ruta}",
                               _texto(f.preventista) or "SIN PREVENTISTA"))

    zonas = {ruta: max(c, key=c.get) for ruta, c in conteo.items()}
    rutas = [(ruta, *meta[ruta]) for ruta in sorted(conteo)
             if meta[ruta][1] not in VENDEDORES_SIN_CUPO]
    return zonas, rutas


def cargar_historia(dl: DataLoader) -> dict[tuple[int, str], float]:
    """(id_ruta_fv4, categoria) -> bultos vendidos en el periodo."""
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv4 AS ruta, da.generico, da.marca,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.id_sucursal = 1 AND fv.anulado = false
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv4 IS NOT NULL
        GROUP BY dc.id_ruta_fv4, da.generico, da.marca
    """, {"d": HIST_DESDE, "h": HIST_HASTA})
    historia: dict[tuple[int, str], float] = {}
    for f in df.itertuples(index=False):
        cat = clasificar(f.generico, f.marca)
        if cat is None:
            continue
        clave = (int(f.ruta), cat)
        historia[clave] = historia.get(clave, 0.0) + float(f.qty or 0)
    return historia


def repartir(rutas, zonas, historia) -> dict[int, dict[str, float]]:
    """Abre el cupo de cada zona entre sus rutas."""
    valores = {ruta: {c: 0.0 for c in CATEGORIAS} for ruta, _, _ in rutas}
    for zona in ZONAS:
        de_zona = [r for r in rutas if zonas[r[0]] == zona]
        if not de_zona:
            continue
        for cat in CATEGORIAS:
            pesos = [historia.get((ruta, cat), 0.0) for ruta, _, _ in de_zona]
            partes = split_proporcional(float(CUPOS[zona].get(cat, 0)), pesos)
            for (ruta, _, _), parte in zip(de_zona, partes):
                valores[ruta][cat] = parte
    return valores


def validar(rutas, zonas, valores) -> int:
    """Imprime y cuenta las zonas/categorias donde el reparto no cierra."""
    print("Validacion suma rutas == cupo de la zona")
    errores = 0
    for zona in ZONAS:
        de_zona = [r for r, _, _ in rutas if zonas[r] == zona]
        for cat in CATEGORIAS:
            total = round(sum(valores[r][cat] for r in de_zona), 2)
            esperado = float(CUPOS[zona].get(cat, 0))
            if abs(total - esperado) > 0.01:
                print(f"  !! {zona:20} {cat:14} {total} != {esperado}")
                errores += 1
    print(f"  {'OK — todo cierra' if not errores else f'{errores} diferencias'}")
    return errores


# --- Excel ------------------------------------------------------------------
HDR = PatternFill("solid", fgColor="1F4E78")
SUB = PatternFill("solid", fgColor="D9E1F2")
TOT = PatternFill("solid", fgColor="FCE4D6")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
BOLD = Font(bold=True)


def escribir_hoja(ws, headers, filas, etiqueta_idx, primera_num, formatos=None):
    formatos = formatos or {}
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = HDR, Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for fila in filas:
        ws.append(fila)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        etiqueta = str(fila[etiqueta_idx].value or "")
        es_total = etiqueta == "TOTAL GENERAL"
        es_sub = etiqueta.startswith("TOTAL ") and not es_total
        for i, c in enumerate(fila):
            c.border = BORDE
            if i >= primera_num and isinstance(c.value, (int, float)):
                c.number_format = formatos.get(i, "#,##0.00")
            if es_sub:
                c.font, c.fill = BOLD, SUB
            elif es_total:
                c.font, c.fill = BOLD, TOT
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(22, len(str(h)) + 3))
    ws.freeze_panes = ws.cell(row=2, column=primera_num + 1).coordinate


def construir_workbook(rutas, zonas, valores) -> Workbook:
    wb = Workbook()
    total_general = [round(sum(valores[r][c] for r, _, _ in rutas), 2)
                     for c in CATEGORIAS]

    # Hoja 1 — detalle por ruta
    ws = wb.active
    ws.title = "Cupo Ruta"
    filas = []
    for zona in ZONAS:
        de_zona = [r for r in rutas if zonas[r[0]] == zona]
        for ruta, des_ruta, preventista in de_zona:
            filas.append([zona, preventista, ruta, des_ruta]
                         + [valores[ruta][c] for c in CATEGORIAS])
        subtotal = [round(sum(valores[r][c] for r, _, _ in de_zona), 2)
                    for c in CATEGORIAS]
        filas.append([zona, f"TOTAL {zona}", "", "", *subtotal])
    filas.append(["", "TOTAL GENERAL", "", "", *total_general])
    escribir_hoja(ws, ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA"] + CATEGORIAS,
                  filas, etiqueta_idx=1, primera_num=2, formatos={2: "0"})

    # Hoja 2 — categorias en filas contra las zonas
    ws2 = wb.create_sheet("Cupo por Categoria")
    filas2 = []
    for cat in CATEGORIAS:
        por_zona = [round(sum(valores[r][cat] for r, _, _ in rutas
                              if zonas[r] == zona), 2) for zona in ZONAS]
        filas2.append([cat, *por_zona, round(sum(por_zona), 2)])
    filas2.append(["TOTAL GENERAL",
                   *[round(sum(f[i] for f in filas2), 2) for i in range(1, 5)]])
    escribir_hoja(ws2, ["CATEGORIA", *ZONAS, "TOTAL"], filas2,
                  etiqueta_idx=0, primera_num=1)

    # Hoja 3 — formato largo, una fila por preventista x ruta x categoria
    ws3 = wb.create_sheet("Base Pivot")
    filas3 = []
    for zona in ZONAS:
        for ruta, des_ruta, preventista in [r for r in rutas if zonas[r[0]] == zona]:
            for cat in CATEGORIAS:
                filas3.append([zona, preventista, ruta, des_ruta, cat,
                               valores[ruta][cat]])
    escribir_hoja(ws3, ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA", "CATEGORIA",
                        "CUPO"], filas3, etiqueta_idx=1, primera_num=2,
                  formatos={2: "0"})

    # Hoja 4 — agregado por preventista
    ws4 = wb.create_sheet("Resumen Preventista")
    por_prev: dict[tuple[str, str], dict[str, float]] = {}
    for ruta, _, preventista in rutas:
        acc = por_prev.setdefault((zonas[ruta], preventista),
                                  {c: 0.0 for c in CATEGORIAS})
        for c in CATEGORIAS:
            acc[c] += valores[ruta][c]
    filas4 = [[zona, prev] + [round(vals[c], 2) for c in CATEGORIAS]
              for (zona, prev), vals in sorted(por_prev.items())]
    filas4.append(["", "TOTAL GENERAL", *total_general])
    escribir_hoja(ws4, ["ZONA", "PREVENTISTA"] + CATEGORIAS, filas4,
                  etiqueta_idx=1, primera_num=2)

    # Hoja 5 — de donde salio la zona de cada ruta, para poder auditarlo
    ws5 = wb.create_sheet("Mapa Zonas")
    filas5 = [[ruta, des_ruta, preventista, zonas[ruta]]
              for ruta, des_ruta, preventista in rutas]
    escribir_hoja(ws5, ["CÓDIGO", "RUTA", "PREVENTISTA", "ZONA DEDUCIDA"],
                  filas5, etiqueta_idx=1, primera_num=0, formatos={0: "0"})
    return wb


def main() -> int:
    # El xlsx se edita a mano despues de generarlo. Regenerar encima destruye
    # ese trabajo sin aviso, asi que hay que pedirlo explicito con --force.
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"El archivo ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nPuede tener ajustes hechos a mano. Para regenerarlo igual, "
              "guarda una copia y corre con --force.")
        return 1

    dl = DataLoader()
    zonas, rutas = deducir_zonas(dl)
    historia = cargar_historia(dl)
    valores = repartir(rutas, zonas, historia)
    errores = validar(rutas, zonas, valores)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(rutas, zonas, valores).save(SALIDA)

    preventistas = {p for _, _, p in rutas} - {"SUB DISTRIBUIDOR"}
    print(f"\nGuardado: {SALIDA}")
    print(f"Rutas con cupo: {len(rutas)}  |  preventistas: {len(preventistas)}")
    for zona in ZONAS:
        n = len([r for r, _, _ in rutas if zonas[r] == zona])
        print(f"  {zona:20} {n:3} rutas")
    return 0 if not errores else 1


if __name__ == "__main__":
    raise SystemExit(main())
