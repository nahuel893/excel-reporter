"""Apertura por ruta Y POR MARCA de los cupos de BRANCA (fuerza de ventas 4).

Sucesor de `apertura_rutas_branca_casa_central.py`, que abria el cupo en cuatro
categorias gruesas. Aca cada bloque se abre en las MARCAS que lo componen,
igual que cerveza se abre en SALTA / HEINEKEN / IMPERIAL / MILLER.

Como se reparte
---------------
El cupo llega por bloque y por zona. Cada par (bloque, zona) se reparte en UNA
sola pasada sobre los pares (ruta x etiqueta) de esa zona, proporcional a la
historia del mes anterior. Una sola pasada — y no una por marca — es lo que
garantiza que la suma de las etiquetas de un bloque de EXACTO el cupo del
bloque, sin arrastrar el redondeo de cada marca.

La marca que no vendio nada en la ventana no puede entrar a ese reparto: su
peso es cero y se quedaria en cero para siempre. Cobra por PISO, y el piso sale
del mismo cupo del bloque, asi que el bloque cierra igual en el numero pedido.

Bloques
-------
- BRANCA      = generico FRATELLI B, abierto en sus marcas: las que tienen
                historia mas las que ya venian con cupo el mes anterior.
- VINOS       = generico VINOS, solo las cuatro marcas del objetivo.
- VINOS FINOS = generico VINOS FINOS EXCLUYENDO QUARA. Etiqueta unica: el
                objetivo llega como una sola linea, no abierto por bodega.
- QUARA       = la marca QUARA, que lleva cupo propio fuera de VINOS FINOS.
- JUGOS       = sin cupo. Se emite igual en cero para no perder la etiqueta que
                ya existe en la base desde abril.

La ruta sale SIEMPRE de `dim_cliente.id_ruta_fv4`, nunca de `fact_ventas`: el
fact no tiene ruta, y la ruta del vendedor que facturo no es la ruta a la que
esta asignado el cliente hoy.

Zonas
-----
La zona es una propiedad geografica del CLIENTE, y las rutas de fv4 no calzan
con las de fv1. El mapa NO se hardcodea: se deduce cruzando cada cliente con su
ruta de cervezas, cuyo esquema de zonas ya vive en config/settings.py. Cada
ruta de fv4 cae en la zona donde esta la mayoria de sus clientes.

Formato de salida
-----------------
La hoja `LISTA` reproduce el contrato del loader de medallion-etl
(`bronze/loaders/cupos_loader.py::_read_branca`), que lee `cupo_branca.xlsx`
POR POSICION desde la fila 2:

    row[0] Codigo -> id_ruta      row[3] GENERICO -> generico
    row[1] RUTA   -> descripcion  row[4] Cupo     -> cupo
    row[2] PREVENTISTA            proveedor='BRANCA', desagregado=NULL fijos

No tiene filtro de NIVEL: TODA fila con codigo numerico y generico no nulo se
carga. Por eso `LISTA` no lleva fila de totales — una fila TOTAL entraria a la
base como si fuera una ruta. Los totales van en las otras hojas.

Uso
---
    python scripts/apertura_rutas_branca_por_marca.py [--force]
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

# --- Periodo y destino ------------------------------------------------------
PERIODO = "2026-08"
PERIODO_ANTERIOR = "2026-07"          # de donde se heredan las etiquetas
# UN mes, igual que cerveza. La ventana larga no es neutral: promedia regimenes
# distintos y termina premiando lo que se muere. TORO paso de 3.455 bultos en
# febrero a 293 en julio mientras ARIZU saltaba de 241 a 1.904; con seis meses
# el objetivo daba TORO 1.671 contra ARIZU 1.042, al reves de lo que pasa. Con
# uno queda ARIZU 2.274 / TORO 349, que es el regimen vigente.
#
# El costo de la ventana corta —nueve etiquetas de veintiuna sin historia— lo
# absorbe el piso, que existe justamente para eso: son 81 cajas sobre 9.900.
# Sin el piso esta ventana no serviria, porque esas nueve quedaban en cero.
HIST_DESDE, HIST_HASTA = "2026-07-01", "2026-08-01"
SALIDA = ROOT / "data/output/cupos" / PERIODO / (
    "CUPO BRANCA POR RUTA Y MARCA - AGOSTO 2026.xlsx")

ID_SUCURSAL = 1

# Piso para la marca que no vendio NADA en la ventana, como fraccion del cupo
# del bloque en esa zona. 0,15% de 5.000 son 7,5 cajas, el mismo orden que el
# piso que ya trae la carga de julio (~8,3 por marca muerta).
PISO_MARCA_PCT = 0.0015
# Tope de seguridad: el piso nunca puede comerse mas de esto del bloque.
PISO_MAX_SHARE = 0.25

# --- Zonas ------------------------------------------------------------------
# Rutas de CERVEZAS (fv1) que definen cada zona — ver config/settings.py.
RUTAS_VALLE_FV1 = {81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92,
                   118, 119, 120, 122}
RUTAS_SUBD_FV1 = {93}
ZONAS = ["CASA CENTRAL", "VALLE SALTA", "SUB DISTRIBUIDORES"]

# DIRECTA VINOS no es ruta de preventa: absorberia cupo que tienen que
# perseguir los preventistas. Mismo criterio que DIRECTA en fv1.
VENDEDORES_SIN_CUPO = {"DIRECTA VINOS"}

# --- Bloques ----------------------------------------------------------------
BRANCA, VINOS, VINOS_FINOS, QUARA, JUGOS = (
    "BRANCA", "VINOS", "VINOS FINOS", "QUARA", "JUGOS")

# Las cuatro del objetivo. RESERO tambien es marca del generico VINOS pero
# quedo fuera del cupo, asi que tampoco pesa en el reparto.
MARCAS_VINOS = ["ARIZU", "TORO", "ANIMANA1", "CANCILLER"]

CUPOS: dict[str, dict[str, float]] = {
    #                    BRANCA  VINOS  VINOS FINOS  QUARA  JUGOS
    "CASA CENTRAL":       {BRANCA: 5000, VINOS: 2500, VINOS_FINOS: 370, QUARA: 450, JUGOS: 0},
    "VALLE SALTA":        {BRANCA:  800, VINOS:  500, VINOS_FINOS:  30, QUARA:  50, JUGOS: 0},
    "SUB DISTRIBUIDORES": {BRANCA:  200, VINOS:    0, VINOS_FINOS:   0, QUARA:   0, JUGOS: 0},
}


def _texto(valor) -> str:
    """Normaliza a mayusculas tolerando NULL/NaN de gold."""
    if valor is None or (isinstance(valor, float) and valor != valor):
        return ""
    return str(valor).strip().upper()


def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`, con el residuo al de MAYOR historia.

    Los pesos negativos (notas de credito) se tratan como cero: una devolucion
    no puede restar cupo. Si nadie tiene historia se reparte en partes iguales.
    """
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, 2)] * (n - 1)
        partes.append(round(total - sum(partes), 2))
        return partes
    partes = [round(total * p / suma, 2) for p in positivos]
    residuo = round(total - sum(partes), 2)
    if residuo:
        mayor = max(range(n), key=lambda i: positivos[i])
        partes[mayor] = round(partes[mayor] + residuo, 2)
    return partes


# --- Datos ------------------------------------------------------------------
def deducir_zonas(dl: DataLoader) -> tuple[dict[int, str], list[tuple[int, str, str]]]:
    """Deduce la zona de cada ruta de fv4 desde las rutas de cervezas (fv1).

    Devuelve ({id_ruta_fv4: zona}, [(ruta, des_ruta, preventista)]).
    """
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv4 AS ruta, dc.id_ruta_fv1 AS ruta_cervezas,
               MIN(dc.des_ruta_fv4) AS des_ruta,
               MIN(dc.des_personal_fv4) AS preventista,
               COUNT(*) AS clientes
        FROM gold.dim_cliente dc
        WHERE dc.id_sucursal = :s
          AND COALESCE(dc.anulado, false) = false
          AND dc.id_ruta_fv4 IS NOT NULL
        GROUP BY dc.id_ruta_fv4, dc.id_ruta_fv1
    """, {"s": ID_SUCURSAL})

    conteo: dict[int, dict[str, int]] = {}
    meta: dict[int, tuple[str, str]] = {}
    for f in df.itertuples(index=False):
        ruta = int(f.ruta)
        r1 = f.ruta_cervezas
        if r1 is None or (isinstance(r1, float) and r1 != r1):
            zona = "CASA CENTRAL"
        elif int(r1) in RUTAS_SUBD_FV1:
            zona = "SUB DISTRIBUIDORES"
        elif int(r1) in RUTAS_VALLE_FV1:
            zona = "VALLE SALTA"
        else:
            zona = "CASA CENTRAL"
        conteo.setdefault(ruta, {z: 0 for z in ZONAS})[zona] += int(f.clientes)
        meta.setdefault(ruta, (_texto(f.des_ruta) or f"RUTA {ruta}",
                               _texto(f.preventista) or "SIN PREVENTISTA"))

    zonas = {ruta: max(c, key=c.get) for ruta, c in conteo.items()}
    rutas = [(ruta, *meta[ruta]) for ruta in sorted(conteo)
             if meta[ruta][1] not in VENDEDORES_SIN_CUPO]
    return zonas, rutas


def cargar_historia(dl: DataLoader):
    """Historia del mes anterior por (ruta_fv4, generico, marca).

    La ruta sale de `dim_cliente`, no del fact: es la asignacion vigente del
    cliente, que es sobre quien se va a perseguir el objetivo.
    """
    return dl.execute_query("""
        SELECT dc.id_ruta_fv4 AS ruta, da.generico, da.marca,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.id_sucursal = :s AND fv.anulado = false
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv4 IS NOT NULL
          AND da.generico IN ('FRATELLI B', 'VINOS', 'VINOS FINOS')
        GROUP BY dc.id_ruta_fv4, da.generico, da.marca
    """, {"s": ID_SUCURSAL, "d": HIST_DESDE, "h": HIST_HASTA})


def marcas_heredadas(dl: DataLoader) -> set[str]:
    """Marcas de FRATELLI B que ya venian con cupo en el periodo anterior.

    Se heredan para que una marca no desaparezca del archivo solo porque tuvo
    un semestre flojo — Branca sigue esperando ver la etiqueta. Se cruza contra
    `dim_articulo` a proposito: la carga de julio trae etiquetas que NO son
    marcas (FERNET, JUGOS, VINOS FINOS) y esas se manejan como bloque, no como
    marca. El cruce las descarta solo.
    """
    df = dl.execute_query("""
        SELECT DISTINCT UPPER(TRIM(fc.generico)) AS marca
        FROM gold.fact_cupos fc
        JOIN gold.dim_articulo da
          ON UPPER(TRIM(da.marca)) = UPPER(TRIM(fc.generico))
         AND da.generico = 'FRATELLI B'
        WHERE fc.periodo = :p AND fc.proveedor = 'BRANCA'
    """, {"p": PERIODO_ANTERIOR})
    return {_texto(f.marca) for f in df.itertuples(index=False)}


def armar_bloques(df, heredadas: set[str]):
    """Traduce la historia a etiquetas de cupo.

    Devuelve ({bloque: [etiquetas]}, {(ruta, etiqueta): qty}, volumen_descartado).

    Las marcas de BRANCA salen de los datos mas las heredadas del mes anterior.
    Las que no tienen historia igual llevan etiqueta: cobran por piso, no por
    reparto proporcional — un peso cero no puede repartir nada.
    """
    historia: dict[tuple[int, str], float] = {}
    volumen: dict[str, float] = {}
    descartado = 0.0

    for f in df.itertuples(index=False):
        g, m, qty = _texto(f.generico), _texto(f.marca), float(f.qty or 0)
        ruta = int(f.ruta)
        if g == "FRATELLI B":
            # Sin marca no hay etiqueta posible; inventar una meteria en la
            # base un rotulo que Branca no reconoce.
            if not m:
                descartado += qty
                continue
            etiqueta = m
        elif g == "VINOS":
            if m not in MARCAS_VINOS:
                descartado += qty
                continue
            etiqueta = m
        elif g == "VINOS FINOS":
            etiqueta = QUARA if m == "QUARA" else VINOS_FINOS
        else:
            continue
        historia[(ruta, etiqueta)] = historia.get((ruta, etiqueta), 0.0) + qty
        volumen[etiqueta] = volumen.get(etiqueta, 0.0) + qty

    con_historia = {e for e, v in volumen.items()
                    if v > 0 and e not in MARCAS_VINOS
                    and e not in (VINOS_FINOS, QUARA)}
    marcas_branca = sorted(con_historia | heredadas,
                           key=lambda e: (-volumen.get(e, 0.0), e))

    bloques = {
        BRANCA: marcas_branca,
        VINOS: [m for m in MARCAS_VINOS],
        VINOS_FINOS: [VINOS_FINOS],
        QUARA: [QUARA],
        JUGOS: [JUGOS],
    }
    return bloques, historia, descartado


def repartir_bloque(cupo, rutas_zona, etiquetas, historia):
    """Reparte el cupo de UN bloque en UNA zona sobre los pares ruta x etiqueta.

    Dos poblaciones distintas, porque no se pueden mezclar:

    - La etiqueta que vendio algo en la zona reparte por historia. Todas juntas,
      en una sola pasada sobre sus pares, para que el bloque cierre exacto sin
      arrastrar el redondeo de cada marca.
    - La etiqueta que no vendio NADA cobra un piso fijo. No puede entrar al
      reparto proporcional: su peso es cero y se quedaria en cero para siempre.

    El piso NO es plata extra: sale del mismo cupo del bloque. Los que si tienen
    venta reparten `cupo - piso_total`, asi que el bloque sigue cerrando en el
    numero que pidio el objetivo.
    """
    partes: dict[tuple[int, str], float] = {}
    if not rutas_zona or not etiquetas or cupo <= 0:
        return {(r, e): 0.0 for r in rutas_zona for e in etiquetas}

    vivas = [e for e in etiquetas
             if sum(historia.get((r, e), 0.0) for r in rutas_zona) > 0]
    muertas = [e for e in etiquetas if e not in vivas]

    # Sin ninguna etiqueta viva no hay historia que repartir: parte igual.
    if not vivas:
        pares = [(r, e) for r in rutas_zona for e in etiquetas]
        return dict(zip(pares, split_proporcional(cupo, [0.0] * len(pares))))

    piso = round(cupo * PISO_MARCA_PCT, 2) if muertas else 0.0
    if piso * len(muertas) > cupo * PISO_MAX_SHARE:
        piso = round(cupo * PISO_MAX_SHARE / len(muertas), 2)
    resto = round(cupo - piso * len(muertas), 2)

    pares_vivos = [(r, e) for r in rutas_zona for e in vivas]
    for par, parte in zip(pares_vivos,
                          split_proporcional(resto, [historia.get(p, 0.0)
                                                     for p in pares_vivos])):
        partes[par] = parte

    # El piso de una marca muerta se reparte entre las rutas segun el peso que
    # cada una tiene en el bloque: la ruta que mas mueve la categoria es la que
    # mejor puede colocar una marca que hoy no vende.
    peso_ruta = [sum(historia.get((r, e), 0.0) for e in vivas) for r in rutas_zona]
    for e in muertas:
        for r, parte in zip(rutas_zona, split_proporcional(piso, peso_ruta)):
            partes[(r, e)] = parte
    return partes


def repartir(rutas, zonas, bloques, historia) -> dict[int, dict[str, float]]:
    """Abre el cupo de cada (bloque, zona) entre los pares ruta x etiqueta."""
    etiquetas = [e for lista in bloques.values() for e in lista]
    valores = {ruta: {e: 0.0 for e in etiquetas} for ruta, _, _ in rutas}

    for zona in ZONAS:
        de_zona = [r for r, _, _ in rutas if zonas[r] == zona]
        if not de_zona:
            continue
        for bloque, lista in bloques.items():
            partes = repartir_bloque(float(CUPOS[zona].get(bloque, 0)),
                                     de_zona, lista, historia)
            for (ruta, e), parte in partes.items():
                valores[ruta][e] = parte
    return valores


def validar(rutas, zonas, bloques, valores) -> int:
    """Cada (bloque, zona) tiene que sumar exactamente su cupo."""
    print("Validacion suma rutas == cupo del bloque en la zona")
    errores = 0
    for zona in ZONAS:
        de_zona = [r for r, _, _ in rutas if zonas[r] == zona]
        for bloque, lista in bloques.items():
            total = round(sum(valores[r][e] for r in de_zona for e in lista), 2)
            esperado = float(CUPOS[zona].get(bloque, 0))
            if abs(total - esperado) > 0.01:
                print(f"  !! {zona:20} {bloque:12} {total} != {esperado}")
                errores += 1
    print(f"  {'OK — todo cierra' if not errores else f'{errores} diferencias'}")
    return errores


# --- Excel ------------------------------------------------------------------
HDR = PatternFill("solid", fgColor="1F4E78")
SUB = PatternFill("solid", fgColor="D9E1F2")
TOT = PatternFill("solid", fgColor="FCE4D6")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
BOLD = Font(bold=True)


def escribir_hoja(ws, headers, filas, etiqueta_idx, primera_num, formatos=None):
    formatos = formatos or {}
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = HDR, Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for fila in filas:
        ws.append(fila)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        etiqueta = str(fila[etiqueta_idx].value or "")
        es_total = etiqueta == "TOTAL GENERAL"
        es_sub = etiqueta.startswith("TOTAL ") and not es_total
        for i, c in enumerate(fila):
            c.border = BORDE
            if i >= primera_num and isinstance(c.value, (int, float)):
                c.number_format = formatos.get(i, "#,##0.00")
            if es_sub:
                c.font, c.fill = BOLD, SUB
            elif es_total:
                c.font, c.fill = BOLD, TOT
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(20, len(str(h)) + 3))
    ws.freeze_panes = ws.cell(row=2, column=primera_num + 1).coordinate


def construir_workbook(rutas, zonas, bloques, valores) -> Workbook:
    wb = Workbook()
    etiquetas = [e for lista in bloques.values() for e in lista]

    # --- Hoja 1: LISTA — el formato que lee el ETL, sin fila de totales ------
    ws = wb.active
    ws.title = "LISTA"
    filas = []
    for zona in ZONAS:
        for ruta, des_ruta, preventista in [r for r in rutas if zonas[r[0]] == zona]:
            for e in etiquetas:
                filas.append([ruta, des_ruta, preventista, e, valores[ruta][e]])
    escribir_hoja(ws, ["Código", "RUTA", "PREVENTISTA", "GENERICO", "Cupo"],
                  filas, etiqueta_idx=2, primera_num=0,
                  formatos={0: "0", 4: "#,##0.00"})

    # --- Hoja 2: resumen por vendedor fv4, marcas y genericos en columnas ----
    ws2 = wb.create_sheet("Resumen Vendedor")
    headers2, cols = ["PREVENTISTA"], []
    for bloque, lista in bloques.items():
        if len(lista) > 1:                       # bloque abierto: total + detalle
            headers2.append(f"TOTAL {bloque}")
            cols.append(("bloque", lista))
            for e in lista:
                headers2.append(e)
                cols.append(("etiqueta", [e]))
        else:                                    # etiqueta unica: una columna
            headers2.append(lista[0])
            cols.append(("etiqueta", lista))
    headers2.append("TOTAL GENERAL")
    cols.append(("etiqueta", etiquetas))

    por_prev: dict[str, dict[str, float]] = {}
    for ruta, _, preventista in rutas:
        acc = por_prev.setdefault(preventista, {e: 0.0 for e in etiquetas})
        for e in etiquetas:
            acc[e] += valores[ruta][e]

    filas2 = [[prev] + [round(sum(vals[e] for e in lista), 2) for _, lista in cols]
              for prev, vals in sorted(por_prev.items())]
    filas2.append(["TOTAL GENERAL"] + [round(sum(f[i] for f in filas2), 2)
                                       for i in range(1, len(headers2))])
    escribir_hoja(ws2, headers2, filas2, etiqueta_idx=0, primera_num=1)

    # --- Hoja 3: detalle por ruta, para revisar antes de cargar --------------
    ws3 = wb.create_sheet("Cupo Ruta")
    filas3 = []
    for zona in ZONAS:
        de_zona = [r for r in rutas if zonas[r[0]] == zona]
        for ruta, des_ruta, preventista in de_zona:
            filas3.append([zona, preventista, ruta, des_ruta]
                          + [valores[ruta][e] for e in etiquetas])
        filas3.append([zona, f"TOTAL {zona}", "", ""]
                      + [round(sum(valores[r][e] for r, _, _ in de_zona), 2)
                         for e in etiquetas])
    filas3.append(["", "TOTAL GENERAL", "", ""]
                  + [round(sum(valores[r][e] for r, _, _ in rutas), 2)
                     for e in etiquetas])
    escribir_hoja(ws3, ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA"] + etiquetas,
                  filas3, etiqueta_idx=1, primera_num=2, formatos={2: "0"})

    # --- Hoja 4: el cupo por bloque y zona, contra el objetivo original ------
    ws4 = wb.create_sheet("Cupo por Bloque")
    filas4 = []
    for bloque, lista in bloques.items():
        por_zona = [round(sum(valores[r][e] for r, _, _ in rutas
                              if zonas[r] == zona for e in lista), 2)
                    for zona in ZONAS]
        filas4.append([bloque, *por_zona, round(sum(por_zona), 2)])
    filas4.append(["TOTAL GENERAL", *[round(sum(f[i] for f in filas4), 2)
                                      for i in range(1, len(ZONAS) + 2)]])
    escribir_hoja(ws4, ["BLOQUE", *ZONAS, "TOTAL"], filas4,
                  etiqueta_idx=0, primera_num=1)

    # --- Hoja 5: de donde salio la zona de cada ruta, para auditarlo ---------
    ws5 = wb.create_sheet("Mapa Zonas")
    filas5 = [[ruta, des_ruta, preventista, zonas[ruta]]
              for ruta, des_ruta, preventista in rutas]
    escribir_hoja(ws5, ["CÓDIGO", "RUTA", "PREVENTISTA", "ZONA DEDUCIDA"],
                  filas5, etiqueta_idx=1, primera_num=0, formatos={0: "0"})
    return wb


def main() -> int:
    # El xlsx se ajusta a mano despues de generarlo. Regenerar encima destruye
    # ese trabajo sin aviso, asi que hay que pedirlo explicito con --force.
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"El archivo ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nPuede tener ajustes hechos a mano. Para regenerarlo igual, "
              "guarda una copia y corre con --force.")
        return 1

    dl = DataLoader()
    zonas, rutas = deducir_zonas(dl)
    bloques, historia, descartado = armar_bloques(cargar_historia(dl),
                                                  marcas_heredadas(dl))
    valores = repartir(rutas, zonas, bloques, historia)
    errores = validar(rutas, zonas, bloques, valores)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(rutas, zonas, bloques, valores).save(SALIDA)

    print(f"\nHistoria de referencia: {HIST_DESDE} a {HIST_HASTA} (1 mes)")
    print(f"Guardado: {SALIDA}")
    print(f"Rutas: {len(rutas)} | preventistas: {len({p for _, _, p in rutas})}")
    for bloque, lista in bloques.items():
        sin_venta = [e for e in lista
                     if sum(historia.get((r, e), 0.0) for r, _, _ in rutas) <= 0]
        marca = f"  (por piso: {', '.join(sin_venta)})" if sin_venta else ""
        print(f"  {bloque:12} {len(lista):2} etiqueta(s){marca}")
    if descartado:
        print(f"\nVolumen fuera del cupo (no pesa en el reparto): {descartado:,.1f} bultos")
    return 0 if not errores else 1


if __name__ == "__main__":
    raise SystemExit(main())
