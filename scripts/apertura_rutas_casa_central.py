"""Apertura por ruta de los cupos de CASA CENTRAL (sucursal 1).

A diferencia del interior, en sucursal 1 el cupo NO viene por preventista sino
por ZONA (CASA CENTRAL / VALLE SALTA / SUB DISTRIBUIDORES). Cada zona reparte
su cupo entre SUS rutas, proporcional a la historia del mes anterior.

CERVEZAS se abre en 7 categorias de marca en una sola pasada: el cupo de
cerveza de la zona se reparte entre todos los pares (ruta x marca), asi la
suma de las 7 columnas da exactamente el cupo de la zona.

Reglas fijas (definidas por Nahuel, 2026-08-04)
-----------------------------------------------
- SALTA es SOLO la marca SALTA. SCHNEIDER y NORTE van a MULTICERVEZAS y
  SALTA CAUTIVA1 es categoria propia. (En el interior SALTA agrupa las tres:
  son criterios distintos que conviven.)
- IMPORTADAS = BLUE MOON + KUNSTMAN.
- El bloque Branca (BRANCA / ARIZU / VINOS FINOS / QUARA) NO se abre para
  esta fuerza de ventas: lo trabaja otro grupo.
- DIRECTA, VENDEDOR CHOPERAS y los preventistas dados de baja no reciben
  cupo: absorberian volumen que tienen que perseguir los activos.

Uso
---
    python scripts/apertura_rutas_casa_central.py
"""
from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

# --- Periodo ---------------------------------------------------------------
PERIODO = "2026-09"
HIST_DESDE, HIST_HASTA = "2026-08-01", "2026-09-01"

# --- Identificacion para la carga en gold.fact_cupos -------------------------
PROVEEDOR = "CCU"
ID_SUCURSAL, SUCURSAL = 1, "1 - CASA CENTRAL"
# categoria de la apertura -> (generico, desagregado) de gold.fact_cupos.
# Tomado de como esta cargado JULIO en la sucursal 1, que es la referencia:
#   `generico`    es la MARCA de dim_articulo (SALTA, AMSTEL, BLUE MOON...)
#   `desagregado` es la CATEGORIA de cupo a la que se imputa
# Cuando la marca ES la categoria (SALTA, HEINEKEN) las dos coinciden.
#
# CERVEZAS va como fila propia con el total: la base SI guarda el agregado.
# En la base la categoria es AGUAS DANONE (con S) y MULTICCU va junto.
#
# MULTICERVEZAS e IMPORTADAS se cargan con el nombre de la categoria en las
# dos columnas. Julio los tiene abiertos por marca (AMSTEL, GROLSCH,
# WARSTEINER, SCHNEIDER / BLUE MOON, KUNSTMANN), pero esta apertura los
# calcula como un solo numero y repartirlos seria inventar cupos.
MAPEO_BD: dict[str, tuple[str, str]] = {
    "CERVEZAS": ("CERVEZAS", "CERVEZAS"),
    "SALTA": ("SALTA", "SALTA"),
    "HEINEKEN": ("HEINEKEN", "HEINEKEN"),
    "IMPERIAL": ("IMPERIAL", "IMPERIAL"),
    "MILLER": ("MILLER", "MILLER"),
    "MULTICERVEZAS": ("MULTICERVEZAS", "MULTICERVEZAS"),
    "IMPORTADAS": ("IMPORTADAS", "IMPORTADAS"),
    "AGUA DANONE": ("AGUAS DANONE", "AGUAS DANONE"),
    "VINOS CCU": ("VINOS CCU", "MULTICCU"),
    "SIDRAS Y LICORES": ("SIDRAS Y LICORES", "MULTICCU"),
    "PERNOD RICARD": ("PERNOD RICARD", "MULTICCU"),
}
CAMPOS_BD = ["periodo", "proveedor", "id_sucursal", "sucursal", "id_ruta",
             "descripcion", "preventista", "generico", "desagregado", "cupo"]
# Todos los informes de cupos van a data/output/cupos/{mes}/.
SALIDA = ROOT / "data/output/cupos" / PERIODO / (
    "CUPO DESAGREGADO POR RUTA CASA CENTRAL - SEPTIEMBRE 2026.xlsx")

# --- Zonas virtuales de CASA CENTRAL (ver config/settings.py) ---------------
RUTAS_VALLE = {81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 118, 119, 120, 122}
RUTAS_SUBD = {93}
ZONAS = ["CASA CENTRAL", "VALLE SALTA", "SUB DISTRIBUIDORES"]

# Vendedores que NO reciben cupo. DIRECTA y VENDEDOR CHOPERAS no son rutas de
# preventa; RUIZ MARCELO esta de baja (217 clientes asignados, 4 activos).
# SUB DISTRIBUIDOR se queda: es una zona con cupo propio.
VENDEDORES_SIN_CUPO = {"DIRECTA", "VENDEDOR CHOPERAS", "RUIZ MARCELO"}
RUTAS_SIN_CUPO = {77, 999}   # ANULADA / RUTA DE VENTA 999

# --- Categorias ------------------------------------------------------------
CERVEZA_CATS = ["SALTA", "HEINEKEN", "IMPERIAL",
                "MILLER", "MULTICERVEZAS", "IMPORTADAS"]
MULTICCU_CATS = ["VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]
OTRAS_CATS = ["AGUA DANONE", *MULTICCU_CATS]
# Orden de lectura: el AGREGADO primero y su desagregado inmediatamente
# despues, igual que el objetivo del interior (CERVEZAS | SALTA | HEINEKEN |
# ...). Antes "multi ccu" iba al final, detras de sus partes: quedaba al reves
# que CERVEZAS y no se podia leer el bloque de un vistazo.
COLUMNAS = ["CERVEZAS", *CERVEZA_CATS, "AGUA DANONE", "multi ccu", *MULTICCU_CATS]
# Filas de la tabla vertical: el generico/agregado primero y sus partes debajo.
JERARQUIA = [("CERVEZAS", True), *[(c, False) for c in CERVEZA_CATS],
             ("AGUA DANONE", True),
             ("multi ccu", True), *[(c, False) for c in MULTICCU_CATS]]

# Formato largo. CERVEZAS y "multi ccu" van como filas, pero marcadas
# NIVEL='AGREGADO': son la suma de sus partes, asi que sumar la columna CUPO
# sin filtrar las cuenta dos veces. Para el total real: NIVEL='DETALLE'.
GRUPO_DE = {**{c: "CERVEZAS" for c in CERVEZA_CATS},
            "AGUA DANONE": "AGUA DANONE",
            **{c: "MULTI CCU" for c in MULTICCU_CATS}}
# Filas de la hoja Base Pivot: (clave_interna, GRUPO, CATEGORIA, NIVEL).
#
# El ETL de medallion-etl mapea **GRUPO -> generico** y **CATEGORIA ->
# desagregado** (data/cupos_config.json, fix ddb655c). `generico` tiene que
# coincidir con `dim_articulo.generico` para poder cruzarlo contra ventas.
#
# De ahi sale una asimetria que parece un error y no lo es:
# - En cerveza el generico real es CERVEZAS y las marcas son el detalle,
#   asi que GRUPO=CERVEZAS y CATEGORIA=SALTA / HEINEKEN / ...
# - En el bloque multi es al reves: VINOS CCU, SIDRAS Y LICORES y PERNOD
#   RICARD SON genericos de dim_articulo, y "MULTICCU" es una agrupacion que
#   no existe ahi. Por eso GRUPO lleva el generico y CATEGORIA el grupo.
FILAS_LARGO = [
    # CERVEZAS es el UNICO que lleva CERVEZAS en GRUPO. Si las marcas tambien
    # lo llevaran, `generico='CERVEZAS'` sumaria el total Y sus partes.
    ("CERVEZAS", "CERVEZAS", "CERVEZAS", "DETALLE"),
    *[(c, c, c, "DETALLE") for c in CERVEZA_CATS],
    ("AGUA DANONE", "AGUAS DANONE", "AGUAS DANONE", "DETALLE"),
    *[(c, c, "MULTICCU", "DETALLE") for c in MULTICCU_CATS],
    # El total del multi no se carga: el par MULTICCU/MULTICCU no existe en la
    # base. Va etiquetado TOTAL para que no colisione con sus tres partes, que
    # llevan CATEGORIA=MULTICCU — si no, una dinamica que no filtre NIVEL
    # contaria el bloque dos veces.
    ("multi ccu", "TOTAL MULTICCU", "TOTAL MULTICCU", "AGREGADO"),
]

MARCAS_ARIZU = {"ARIZU", "CANCILLER", "ANIMANA1", "TORO"}
MARCAS_IMPORTADAS = {"BLUE MOON", "KUNSTMAN"}


# --- Redistribuciones puntuales del mes -------------------------------------
@dataclass(frozen=True)
class RedistribucionPuntual:
    """Excepcion de UN mes, no una regla del informe.

    Una venta puntual (no recurrente) le infla el cupo al preventista que la
    hizo, porque el reparto es proporcional a la historia. Esto le saca ese
    cupo y lo reparte en partes iguales entre los preventistas de su misma
    zona.

    Se apaga poniendo `activo=False`: el informe vuelve al reparto puro por
    historia sin tocar nada mas. Revisar esta lista cada mes — lo que fue
    excepcion en agosto no tiene por que serlo en septiembre.

    `marca_origen` tiene que ser una marca del generico CERVEZAS: el monto a
    mover se calcula con el cupo de cerveza de la zona.
    """

    activo: bool
    preventista: str
    marca_origen: str
    categoria: str
    motivo: str


REDISTRIBUCIONES = [
    RedistribucionPuntual(
        activo=True,
        preventista="NORMA CACHARI",
        marca_origen="SALTA CAUTIVA1",
        categoria="MULTICERVEZAS",
        motivo="Venta puntual de 65 bultos de Salta Cautiva en la ruta 76 "
               "(jul-26). Le genera ~74 de cupo en agosto sin ser demanda "
               "recurrente: se reparte parejo entre la zona.",
    ),
]

# Marcas cuya historia hay que seguir aparte para poder redistribuirlas.
MARCAS_SEGUIDAS = {r.marca_origen for r in REDISTRIBUCIONES if r.activo}

# --- Cupos del mes, por zona ------------------------------------------------
# SEPTIEMBRE 2026. MULTI CCU cierra en 7.004 = 504 sidras + 5.000 vinos + 1.500 PR.
CUPOS: dict[str, dict[str, float]] = {
    "CASA CENTRAL": {"CERVEZAS": 81000, "AGUA DANONE": 30000,
                     "VINOS CCU": 3300, "SIDRAS Y LICORES": 504,
                     "PERNOD RICARD": 1000},
    "VALLE SALTA": {"CERVEZAS": 12000, "AGUA DANONE": 3000,
                    "VINOS CCU": 1000, "SIDRAS Y LICORES": 0,
                    "PERNOD RICARD": 200},
    "SUB DISTRIBUIDORES": {"CERVEZAS": 12000, "AGUA DANONE": 2000,
                           "VINOS CCU": 700, "SIDRAS Y LICORES": 0,
                           "PERNOD RICARD": 300},
}

# Cupo FIJO por preventista, no proporcional a la historia. Septiembre 2026:
# "sidras y licores 504, 18 por preventista preventa" — y CASA CENTRAL tiene
# exactamente 28 preventistas, asi que 18 x 28 = 504 cierra al bulto.
# Repartir esos 504 por historia daria muy distinto: los preventistas que hoy
# no venden sidra quedarian en cero, que es justo lo contrario de lo que pide
# un objetivo de 18 parejo. Dentro de cada preventista SI se abre por historia,
# para que el numero baje a ruta.
CUPO_FIJO_POR_PREVENTISTA: dict[tuple[str, str], float] = {
    ("CASA CENTRAL", "SIDRAS Y LICORES"): 18.0,
}


def _texto(valor) -> str:
    """Normaliza a mayusculas tolerando NULL/NaN de gold."""
    if valor is None or (isinstance(valor, float) and valor != valor):
        return ""
    return str(valor).strip().upper()


def clasificar(generico, marca) -> str | None:
    """Mapea una venta a una categoria de cupo de CASA CENTRAL."""
    g, m = _texto(generico), _texto(marca)
    if g == "CERVEZAS":
        if m == "SALTA":
            return "SALTA"
        if m in ("HEINEKEN", "IMPERIAL", "MILLER"):
            return m
        if m in MARCAS_IMPORTADAS:
            return "IMPORTADAS"
        # Todo el resto: SCHNEIDER, NORTE y SALTA CAUTIVA1 incluidos.
        return "MULTICERVEZAS"
    if g == "AGUAS DANONE":
        return "AGUA DANONE"
    if g in ("VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"):
        return g
    return None                      # el bloque Branca no se abre aca


def zona_de(ruta: int) -> str:
    if ruta in RUTAS_SUBD:
        return "SUB DISTRIBUIDORES"
    return "VALLE SALTA" if ruta in RUTAS_VALLE else "CASA CENTRAL"


def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`, con el residuo al de MAYOR historia.

    Distinto del reparto del interior (que se lo da al ultimo): aca los pares
    (ruta x marca) incluyen combinaciones sin historia, y darle el residuo a
    una de esas produce cupos negativos.
    """
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, 2)] * (n - 1)
        partes.append(round(total - sum(partes), 2))
        return partes
    partes = [round(total * p / suma, 2) for p in positivos]
    residuo = round(total - sum(partes), 2)
    if residuo:
        mayor = max(range(n), key=lambda i: positivos[i])
        partes[mayor] = round(partes[mayor] + residuo, 2)
    return partes


# --- Datos ------------------------------------------------------------------
def cargar_rutas(dl: DataLoader) -> list[tuple[int, str, str]]:
    """[(id_ruta, des_ruta, preventista)] de las rutas que reciben cupo."""
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv1 AS ruta,
               MIN(dc.des_ruta_fv1) AS des_ruta,
               MIN(dv.des_vendedor) AS vendedor
        FROM gold.dim_cliente dc
        LEFT JOIN gold.dim_vendedor dv ON dv.id_vendedor = dc.id_personal_fv1
                                      AND dv.id_sucursal = dc.id_sucursal
        WHERE dc.id_sucursal = 1
          AND COALESCE(dc.anulado, false) = false
          AND dc.id_ruta_fv1 IS NOT NULL
          AND dc.id_personal_fv1 IS NOT NULL
        GROUP BY dc.id_ruta_fv1
        ORDER BY dc.id_ruta_fv1
    """)
    rutas = []
    for f in df.itertuples(index=False):
        ruta, vendedor = int(f.ruta), _texto(f.vendedor) or "SIN PREVENTISTA"
        if ruta in RUTAS_SIN_CUPO or vendedor in VENDEDORES_SIN_CUPO:
            continue
        rutas.append((ruta, _texto(f.des_ruta) or f"RUTA {ruta}", vendedor))
    return rutas


def cargar_historia(
    dl: DataLoader,
) -> tuple[dict[tuple[int, str], float], dict[tuple[int, str], float]]:
    """Devuelve (historia por categoria, historia por marca seguida).

    La segunda solo trae las marcas de `MARCAS_SEGUIDAS`: se necesitan aparte
    porque ya quedaron fundidas dentro de su categoria.
    """
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv1 AS ruta, da.generico, da.marca,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.id_sucursal = 1 AND fv.anulado = false
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv1 IS NOT NULL
        GROUP BY dc.id_ruta_fv1, da.generico, da.marca
    """, {"d": HIST_DESDE, "h": HIST_HASTA})
    historia: dict[tuple[int, str], float] = {}
    por_marca: dict[tuple[int, str], float] = {}
    for f in df.itertuples(index=False):
        cat = clasificar(f.generico, f.marca)
        if cat is None:
            continue
        ruta, qty = int(f.ruta), float(f.qty or 0)
        historia[(ruta, cat)] = historia.get((ruta, cat), 0.0) + qty
        marca = _texto(f.marca)
        if marca in MARCAS_SEGUIDAS:
            por_marca[(ruta, marca)] = por_marca.get((ruta, marca), 0.0) + qty
    return historia, por_marca


def repartir(rutas, historia) -> dict[int, dict[str, float]]:
    """Abre el cupo de cada zona entre sus rutas."""
    valores = {ruta: {c: 0.0 for c in COLUMNAS} for ruta, _, _ in rutas}

    for zona in ZONAS:
        del_zona = [r for r in rutas if zona_de(r[0]) == zona]
        if not del_zona:
            continue

        # CERVEZAS: una sola pasada sobre los pares (ruta x marca), asi la
        # suma de las 7 columnas da exacto el cupo de cerveza de la zona.
        pares = [(ruta, cat) for ruta, _, _ in del_zona for cat in CERVEZA_CATS]
        pesos = [historia.get(par, 0.0) for par in pares]
        partes = split_proporcional(float(CUPOS[zona]["CERVEZAS"]), pesos)
        for (ruta, cat), parte in zip(pares, partes):
            valores[ruta][cat] = parte

        for cat in OTRAS_CATS:
            fijo = CUPO_FIJO_POR_PREVENTISTA.get((zona, cat))
            if fijo is not None:
                # Cada preventista recibe el mismo cupo; adentro se abre por
                # historia entre SUS rutas.
                por_prev: dict[str, list] = {}
                for ruta, _, prev in del_zona:
                    por_prev.setdefault(prev, []).append(ruta)
                for prev, sus in sorted(por_prev.items()):
                    pesos = [historia.get((ruta, cat), 0.0) for ruta in sus]
                    for ruta, parte in zip(sus, split_proporcional(fijo, pesos)):
                        valores[ruta][cat] = parte
                continue
            pesos = [historia.get((ruta, cat), 0.0) for ruta, _, _ in del_zona]
            partes = split_proporcional(float(CUPOS[zona].get(cat, 0)), pesos)
            for (ruta, _, _), parte in zip(del_zona, partes):
                valores[ruta][cat] = parte

    for ruta, _, _ in rutas:
        v = valores[ruta]
        v["CERVEZAS"] = round(sum(v[c] for c in CERVEZA_CATS), 2)
        v["multi ccu"] = round(sum(v[c] for c in MULTICCU_CATS), 2)
    return valores


def aplicar_redistribuciones(rutas, valores, historia, por_marca) -> None:
    """Mueve el cupo de las ventas puntuales al resto de la zona.

    Se trabaja zona por zona a proposito: el cupo de cada zona lo fijo CCU y
    tiene que seguir cerrando igual despues del ajuste.
    """
    for regla in REDISTRIBUCIONES:
        if not regla.activo:
            continue
        for zona in ZONAS:
            de_zona = [r for r in rutas if zona_de(r[0]) == zona]
            origen = [r for r in de_zona if r[2] == regla.preventista]
            if not origen:
                continue

            # Cupo que esa marca le genero al preventista: su historia sobre
            # la historia total de cerveza de la zona, por el cupo de la zona.
            peso_marca = sum(por_marca.get((ruta, regla.marca_origen), 0.0)
                             for ruta, _, _ in origen)
            peso_total = sum(historia.get((ruta, cat), 0.0)
                             for ruta, _, _ in de_zona for cat in CERVEZA_CATS)
            if peso_marca <= 0 or peso_total <= 0:
                continue
            monto = round(CUPOS[zona]["CERVEZAS"] * peso_marca / peso_total, 2)

            # Se lo saco a sus rutas, proporcional a lo que cada una aporto.
            pesos = [por_marca.get((ruta, regla.marca_origen), 0.0)
                     for ruta, _, _ in origen]
            for (ruta, _, _), parte in zip(origen, split_proporcional(monto, pesos)):
                valores[ruta][regla.categoria] = round(
                    valores[ruta][regla.categoria] - parte, 2)

            # Y lo reparto parejo entre los preventistas de la zona (el de
            # origen incluido) y, dentro de cada uno, entre sus rutas.
            preventistas = sorted({v for _, _, v in de_zona}
                                  - {"SUB DISTRIBUIDOR"})
            if not preventistas:
                continue
            for preventista, cuota in zip(
                    preventistas,
                    split_proporcional(monto, [1.0] * len(preventistas))):
                suyas = [ruta for ruta, _, v in de_zona if v == preventista]
                for ruta, parte in zip(suyas, split_proporcional(
                        cuota, [1.0] * len(suyas))):
                    valores[ruta][regla.categoria] = round(
                        valores[ruta][regla.categoria] + parte, 2)

            # CERVEZAS es derivada: hay que recalcularla despues del ajuste.
            for ruta, _, _ in de_zona:
                valores[ruta]["CERVEZAS"] = round(
                    sum(valores[ruta][c] for c in CERVEZA_CATS), 2)

            print(f"  Redistribuido: {monto} de {regla.preventista} "
                  f"({regla.marca_origen}) -> {regla.categoria} de "
                  f"{len(preventistas)} preventistas de {zona}")


def validar(rutas, valores) -> int:
    """Imprime y cuenta las zonas/categorias donde el reparto no cierra."""
    print("Validacion suma rutas == cupo de la zona")
    errores = 0
    for zona in ZONAS:
        del_zona = [r for r, _, _ in rutas if zona_de(r) == zona]
        for cat in ["CERVEZAS", *OTRAS_CATS]:
            total = round(sum(valores[r][cat] for r in del_zona), 2)
            esperado = float(CUPOS[zona].get(cat, 0))
            if abs(total - esperado) > 0.01:
                print(f"  !! {zona:20} {cat:18} {total} != {esperado}")
                errores += 1
    print(f"  {'OK — todo cierra' if not errores else f'{errores} diferencias'}")
    return errores


# --- Excel ------------------------------------------------------------------
HDR = PatternFill("solid", fgColor="1F4E78")
SUB = PatternFill("solid", fgColor="D9E1F2")
TOT = PatternFill("solid", fgColor="FCE4D6")
AGREGADO = PatternFill("solid", fgColor="FFFF00")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
BOLD = Font(bold=True)


def escribir_hoja(ws, headers, filas, etiqueta_idx, primera_num,
                  formatos=None, filas_agregado=(), idx_agregado=None):
    """Vuelca headers + filas y aplica el formato de tabla.

    `filas_agregado` son los valores que se pintan como agregado; se buscan en
    la columna `idx_agregado` (por defecto, la misma de la etiqueta).
    """
    formatos = formatos or {}
    idx_agregado = etiqueta_idx if idx_agregado is None else idx_agregado
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = HDR, Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for fila in filas:
        ws.append(fila)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        etiqueta = str(fila[etiqueta_idx].value or "")
        es_total = etiqueta == "TOTAL GENERAL"
        es_sub = etiqueta.startswith("TOTAL ") and not es_total
        es_agregado = str(fila[idx_agregado].value or "").strip() in filas_agregado
        for i, c in enumerate(fila):
            c.border = BORDE
            if i >= primera_num and isinstance(c.value, (int, float)):
                c.number_format = formatos.get(i, "#,##0.00")
            if es_sub:
                c.font, c.fill = BOLD, SUB
            elif es_total:
                c.font, c.fill = BOLD, TOT
            elif es_agregado:
                c.font, c.fill = BOLD, AGREGADO
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(22, len(str(h)) + 3))
    ws.freeze_panes = ws.cell(row=2, column=primera_num + 1).coordinate


def construir_workbook(rutas, valores) -> Workbook:
    wb = Workbook()
    total_general = [round(sum(valores[r][c] for r, _, _ in rutas), 2)
                     for c in COLUMNAS]

    # Hoja 1 — detalle por ruta
    ws = wb.active
    ws.title = "Cupo Ruta"
    filas = []
    for zona in ZONAS:
        del_zona = [r for r in rutas if zona_de(r[0]) == zona]
        for ruta, des_ruta, vendedor in del_zona:
            filas.append([zona, vendedor, ruta, des_ruta]
                         + [valores[ruta][c] for c in COLUMNAS])
        subtotal = [round(sum(valores[r][c] for r, _, _ in del_zona), 2)
                    for c in COLUMNAS]
        filas.append([zona, f"TOTAL {zona}", "", "", *subtotal])
    filas.append(["", "TOTAL GENERAL", "", "", *total_general])
    escribir_hoja(ws, ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA"] + COLUMNAS,
                  filas, etiqueta_idx=1, primera_num=2, formatos={2: "0"})

    # Hoja 2 — todo en filas: marcas, genericos y categorias contra las zonas
    ws2 = wb.create_sheet("Cupo por Categoria")
    filas2 = []
    for etiqueta, es_agregado in JERARQUIA:
        por_zona = [round(sum(valores[r][etiqueta] for r, _, _ in rutas
                              if zona_de(r) == zona), 2) for zona in ZONAS]
        filas2.append([etiqueta if es_agregado else f"    {etiqueta}",
                       *por_zona, round(sum(por_zona), 2)])
    escribir_hoja(ws2, ["CATEGORIA", *ZONAS, "TOTAL"], filas2,
                  etiqueta_idx=0, primera_num=1,
                  filas_agregado={"CERVEZAS", "AGUA DANONE", "multi ccu"})

    # Hoja 3 — formato largo: una fila por preventista x ruta x categoria.
    # Incluye los agregados (CERVEZAS, multi ccu) marcados con NIVEL para que
    # una dinamica pueda filtrarlos y no contar dos veces.
    ws_largo = wb.create_sheet("Base Pivot")
    filas_largo = []
    for zona in ZONAS:
        for ruta, des_ruta, vendedor in [r for r in rutas if zona_de(r[0]) == zona]:
            for clave, grupo, categoria, nivel in FILAS_LARGO:
                filas_largo.append([zona, vendedor, ruta, des_ruta,
                                    nivel, grupo, categoria, valores[ruta][clave]])
    escribir_hoja(ws_largo,
                  ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA", "NIVEL", "GRUPO",
                   "CATEGORIA", "CUPO"],
                  filas_largo, etiqueta_idx=1, primera_num=2,
                  formatos={2: "0"},
                  filas_agregado={"AGREGADO"}, idx_agregado=4)

    # Hoja 4 — exactamente las columnas de gold.fact_cupos, lista para cargar.
    # Sin filas de agregado: la base guarda solo las hojas y reconstruye los
    # totales agrupando por `desagregado`. Es la contraparte de "Base Pivot",
    # que si los lleva porque es para leer, no para cargar.
    ws_bd = wb.create_sheet("Formato BD")
    filas_bd = []
    for zona in ZONAS:
        for ruta, des_ruta, preventista in [r for r in rutas if zona_de(r[0]) == zona]:
            for categoria, (generico, desagregado) in MAPEO_BD.items():
                filas_bd.append([PERIODO, PROVEEDOR, ID_SUCURSAL, SUCURSAL,
                                 ruta, des_ruta, preventista,
                                 generico, desagregado, valores[ruta][categoria]])
    escribir_hoja(ws_bd, CAMPOS_BD, filas_bd, etiqueta_idx=6, primera_num=9,
                  formatos={2: "0", 4: "0"})

    # Hoja 5 — agregado por preventista (uno puede trabajar en dos zonas)
    ws3 = wb.create_sheet("Resumen Preventista")
    por_vendedor: dict[tuple[str, str], dict[str, float]] = {}
    for ruta, _, vendedor in rutas:
        acc = por_vendedor.setdefault((zona_de(ruta), vendedor),
                                      {c: 0.0 for c in COLUMNAS})
        for c in COLUMNAS:
            acc[c] += valores[ruta][c]
    filas3 = [[zona, vendedor] + [round(vals[c], 2) for c in COLUMNAS]
              for (zona, vendedor), vals in sorted(por_vendedor.items())]
    filas3.append(["", "TOTAL GENERAL", *total_general])
    escribir_hoja(ws3, ["ZONA", "PREVENTISTA"] + COLUMNAS, filas3,
                  etiqueta_idx=1, primera_num=2)
    return wb


def main() -> int:
    # El xlsx se edita a mano despues de generarlo (ajustes que no viven en el
    # codigo). Regenerar encima destruye ese trabajo sin aviso, asi que si el
    # archivo ya existe hay que pedirlo explicito con --force.
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"El archivo ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nPuede tener ajustes hechos a mano. Para regenerarlo igual, "
              "guarda una copia y corre con --force.")
        return 1

    dl = DataLoader()
    rutas = cargar_rutas(dl)
    historia, por_marca = cargar_historia(dl)
    valores = repartir(rutas, historia)
    if any(r.activo for r in REDISTRIBUCIONES):
        print("Redistribuciones puntuales del mes")
        aplicar_redistribuciones(rutas, valores, historia, por_marca)
    errores = validar(rutas, valores)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(rutas, valores).save(SALIDA)

    preventistas = {v for _, _, v in rutas} - {"SUB DISTRIBUIDOR"}
    print(f"\nGuardado: {SALIDA}")
    print(f"Rutas con cupo: {len(rutas)}  |  preventistas: {len(preventistas)}")
    for zona in ZONAS:
        n = len([r for r, _, _ in rutas if zona_de(r) == zona])
        print(f"  {zona:20} {n:3} rutas")
    return 0 if not errores else 1


if __name__ == "__main__":
    raise SystemExit(main())
