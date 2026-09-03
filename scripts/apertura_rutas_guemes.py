"""Apertura por ruta de los cupos de SUCURSAL GUEMES (sucursal 16).

Hermano de `apertura_rutas_casa_central.py`, pero mas simple: GUEMES no tiene
zonas virtuales, asi que el cupo de la sucursal se reparte directo entre sus
rutas, proporcional a la venta del mes anterior.

Dos diferencias con CASA CENTRAL que NO son un descuido
-------------------------------------------------------
1. El esquema de cerveza es el mismo que CASA CENTRAL — SALTA es solo la
   marca SALTA, y NORTE y SCHNEIDER van a MULTICERVEZAS — con UNA excepcion:
   SALTA CAUTIVA1 sigue siendo categoria propia, como esta cargado julio en
   la sucursal 16. En CASA CENTRAL se fundio en MULTICERVEZAS desde agosto.
2. La columna ZONA lleva la sucursal, no una zona. El loader del ETL usa
   `sucursal_value` del config cuando la hoja es de una sola sucursal, o la
   columna ZONA cuando no lo tiene. Para que esta hoja entre hay que agregar
   su entrada en `medallion-etl/data/cupos_config.json` SIN `sucursal_value`.

Uso
---
    python scripts/apertura_rutas_guemes.py
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402

# --- Periodo y destino ------------------------------------------------------
PERIODO = "2026-09"
HIST_DESDE, HIST_HASTA = "2026-08-01", "2026-09-01"
SALIDA = ROOT / "data/output/cupos" / PERIODO / (
    "CUPO DESAGREGADO POR RUTA GUEMES - SEPTIEMBRE 2026.xlsx")
HOJA = "Base Pivot GUEMES"

ID_SUCURSAL = 16
SUCURSAL = "16 - SUCURSAL GUEMES"     # va en la columna ZONA

# DIRECTA no es ruta de preventa: absorberia cupo de los preventistas.
VENDEDORES_SIN_CUPO = {"DIRECTA"}

# --- Categorias -------------------------------------------------------------
CERVEZA_CATS = ["SALTA", "SALTA CAUTIVA1", "HEINEKEN", "IMPERIAL",
                "MILLER", "MULTICERVEZAS", "IMPORTADAS"]
MULTICCU_CATS = ["VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]
OTRAS_CATS = ["AGUA DANONE", *MULTICCU_CATS]

# SALTA es SOLO la marca SALTA. NORTE y SCHNEIDER caen en MULTICERVEZAS por
# el default de `clasificar`, igual que en CASA CENTRAL.
MARCAS_SALTA = {"SALTA"}
MARCAS_IMPORTADAS = {"BLUE MOON", "KUNSTMAN", "KUNSTMANN"}
MARCAS_PROPIAS = {"HEINEKEN", "IMPERIAL", "MILLER"}

# Cupos de agosto 2026. MULTICCU = 510 = 400 + 10 + 100.
# SEPTIEMBRE 2026. MULTI CCU cierra en 260 = 150 vinos + 30 sidras + 80 PR.
CUPOS: dict[str, float] = {
    "CERVEZAS": 4700,
    "AGUA DANONE": 4000,
    "VINOS CCU": 150,
    "SIDRAS Y LICORES": 30,
    "PERNOD RICARD": 80,
}

# (clave_interna, GRUPO, CATEGORIA, NIVEL). GRUPO -> generico y
# CATEGORIA -> desagregado, segun medallion-etl/data/cupos_config.json.
# Convencion vigente (la de `CORRECION MULTI`, elegida por Nahuel el
# 2026-08-18): GRUPO = CATEGORIA = la etiqueta, salvo los tres genericos de
# MULTI CCU, que llevan CATEGORIA=MULTICCU para poder agruparlos.
# CERVEZAS va como DETALLE y por lo tanto SE CARGA: en gold es el TOTAL y
# convive con sus marcas. Sumar todos los genericos cuenta la cerveza dos
# veces; hay que elegir el total O el detalle.
FILAS_LARGO = [
    ("CERVEZAS", "CERVEZAS", "CERVEZAS", "DETALLE"),
    *[(c, c, c, "DETALLE") for c in CERVEZA_CATS],
    ("AGUA DANONE", "AGUAS DANONE", "AGUAS DANONE", "DETALLE"),
    *[(c, c, "MULTICCU", "DETALLE") for c in MULTICCU_CATS],
    ("multi ccu", "TOTAL MULTICCU", "TOTAL MULTICCU", "AGREGADO"),
]
COLUMNAS = ["CERVEZAS", *CERVEZA_CATS, "AGUA DANONE", "multi ccu", *MULTICCU_CATS]


def _texto(valor) -> str:
    if valor is None or (isinstance(valor, float) and valor != valor):
        return ""
    return str(valor).strip().upper()


def clasificar(generico, marca) -> str | None:
    """Mapea una venta a una categoria de cupo de GUEMES."""
    g, m = _texto(generico), _texto(marca)
    if g == "CERVEZAS":
        if m in MARCAS_SALTA:
            return "SALTA"
        if m == "SALTA CAUTIVA1":
            return "SALTA CAUTIVA1"
        if m in MARCAS_PROPIAS:
            return m
        if m in MARCAS_IMPORTADAS:
            return "IMPORTADAS"
        # SCHNEIDER, NORTE, AMSTEL, GROLSCH, IGUANA, WARSTEINER...
        return "MULTICERVEZAS"
    if g == "AGUAS DANONE":
        return "AGUA DANONE"
    if g in ("VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"):
        return g
    return None


def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`, con el residuo al de MAYOR historia."""
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, 2)] * (n - 1)
        partes.append(round(total - sum(partes), 2))
        return partes
    partes = [round(total * p / suma, 2) for p in positivos]
    residuo = round(total - sum(partes), 2)
    if residuo:
        partes[max(range(n), key=lambda i: positivos[i])] = round(
            partes[max(range(n), key=lambda i: positivos[i])] + residuo, 2)
    return partes


def cargar_rutas(dl: DataLoader) -> list[tuple[int, str, str]]:
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv1 AS ruta, MIN(dc.des_ruta_fv1) AS des_ruta,
               MIN(dv.des_vendedor) AS preventista
        FROM gold.dim_cliente dc
        LEFT JOIN gold.dim_vendedor dv ON dv.id_vendedor = dc.id_personal_fv1
                                      AND dv.id_sucursal = dc.id_sucursal
        WHERE dc.id_sucursal = :s AND COALESCE(dc.anulado, false) = false
          AND dc.id_ruta_fv1 IS NOT NULL AND dc.id_personal_fv1 IS NOT NULL
        GROUP BY 1 ORDER BY 1
    """, {"s": ID_SUCURSAL})
    return [(int(f.ruta), _texto(f.des_ruta) or f"RUTA {int(f.ruta)}",
             _texto(f.preventista) or "SIN PREVENTISTA")
            for f in df.itertuples(index=False)
            if _texto(f.preventista) not in VENDEDORES_SIN_CUPO]


def cargar_historia(dl: DataLoader) -> dict[tuple[int, str], float]:
    df = dl.execute_query("""
        SELECT dc.id_ruta_fv1 AS ruta, da.generico, da.marca,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.id_sucursal = :s AND fv.anulado = false
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv1 IS NOT NULL
        GROUP BY 1, 2, 3
    """, {"s": ID_SUCURSAL, "d": HIST_DESDE, "h": HIST_HASTA})
    historia: dict[tuple[int, str], float] = {}
    for f in df.itertuples(index=False):
        cat = clasificar(f.generico, f.marca)
        if cat is None:
            continue
        clave = (int(f.ruta), cat)
        historia[clave] = historia.get(clave, 0.0) + float(f.qty or 0)
    return historia


def repartir(rutas, historia) -> dict[int, dict[str, float]]:
    valores = {r: {c: 0.0 for c in COLUMNAS} for r, _, _ in rutas}
    # CERVEZAS: una pasada sobre los pares (ruta x marca) para que la suma de
    # las 7 columnas de exacto el cupo de cerveza.
    pares = [(r, c) for r, _, _ in rutas for c in CERVEZA_CATS]
    pesos = [historia.get(p, 0.0) for p in pares]
    for (r, c), parte in zip(pares, split_proporcional(float(CUPOS["CERVEZAS"]), pesos)):
        valores[r][c] = parte
    for cat in OTRAS_CATS:
        pesos = [historia.get((r, cat), 0.0) for r, _, _ in rutas]
        for (r, _, _), parte in zip(rutas, split_proporcional(float(CUPOS.get(cat, 0)), pesos)):
            valores[r][cat] = parte
    for r, _, _ in rutas:
        valores[r]["CERVEZAS"] = round(sum(valores[r][c] for c in CERVEZA_CATS), 2)
        valores[r]["multi ccu"] = round(sum(valores[r][c] for c in MULTICCU_CATS), 2)
    return valores


def validar(rutas, valores) -> int:
    print("Validacion suma rutas == cupo de la sucursal")
    errores = 0
    for cat in ["CERVEZAS", *OTRAS_CATS]:
        total = round(sum(valores[r][cat] for r, _, _ in rutas), 2)
        esperado = float(CUPOS.get(cat, 0))
        if abs(total - esperado) > 0.01:
            print(f"  !! {cat:18} {total} != {esperado}")
            errores += 1
    print(f"  {'OK — todo cierra' if not errores else f'{errores} diferencias'}")
    return errores


def construir_workbook(rutas, valores) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    headers = ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA", "NIVEL", "GRUPO",
               "CATEGORIA", "CUPO"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    for ruta, des_ruta, preventista in rutas:
        for clave, grupo, categoria, nivel in FILAS_LARGO:
            ws.append([SUCURSAL, preventista, ruta, des_ruta, nivel, grupo,
                       categoria, valores[ruta][clave]])
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(fila):
            c.border = borde
            if i == 2:
                c.number_format = "0"
            if i == 7:
                c.number_format = "#,##0.00"
    for col, w in zip("ABCDEFGH", (22, 20, 9, 22, 11, 18, 18, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return wb


def main() -> int:
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"El archivo ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nPara regenerarlo igual, guarda una copia y corre con --force.")
        return 1

    dl = DataLoader()
    rutas = cargar_rutas(dl)
    historia = cargar_historia(dl)
    valores = repartir(rutas, historia)
    errores = validar(rutas, valores)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(rutas, valores).save(SALIDA)
    print(f"\nGuardado: {SALIDA}")
    print(f"Rutas: {len(rutas)} | preventistas: {len({p for _, _, p in rutas})}")
    return 0 if not errores else 1


if __name__ == "__main__":
    raise SystemExit(main())
