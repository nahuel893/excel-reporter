"""Apertura por ruta del cupo de las SUCURSALES (interior + Guemes), fv1.

El objetivo llega de CCU por preventista en `Objetivo <MES> Badie`. Este script
lo baja a ruta, proporcional a la venta del mes anterior, y lo emite en el
mismo formato `Base Pivot` que Casa Central y Guemes — para que el ETL lo lea
sin tratarlo distinto.

La marca individual depende de la PROVINCIA
-------------------------------------------
No es lo mismo Salta que Jujuy, y el objetivo viene con una sola columna para
las dos realidades:

- JUJUY (Perico, Libertador, Maimara, Humahuaca, Abra Pampa, La Quiaca,
  San Pedro): la columna "SALTA" del objetivo es el cupo de NORTE. NORTE va
  como categoria propia; SALTA y SCHNEIDER caen en MULTICERVEZA.
- SALTA (Cafayate, JVG, Metan, Oran, Tartagal, Guemes): al reves. SALTA es la
  categoria propia; NORTE y SCHNEIDER caen en MULTICERVEZA.

Es la diferencia entre pedirle a Libertador 10.473 de NORTE —que vende 6.286—
o de SALTA, que vende 965.

MULTI CCU viene por SUCURSAL, no por preventista (hoja `Sheet1`), asi que se
reparte entre todas las rutas de la sucursal.

Uso
---
    python scripts/apertura_rutas_sucursales.py [--force]
"""
from __future__ import annotations

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402


def split_proporcional(total: float, pesos: list[float], decimales: int) -> list[float]:
    """Reparte `total` segun `pesos`. Misma cuenta que el resto de las aperturas.

    - El residuo de redondeo va al de MAYOR peso, nunca al ultimo por posicion:
      darselo al ultimo produce cupos negativos cuando ese par no tiene historia.
    - Los pesos negativos (notas de credito) cuentan como cero.
    - Sin historia, partes iguales y el residuo al ultimo — asi lo resuelven los
      generadores viejos y hay que reproducirlos.
    """
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    positivos = [p if p > 0 else 0.0 for p in pesos]
    suma = sum(positivos)
    if suma <= 0:
        partes = [round(total / n, decimales)] * (n - 1)
        partes.append(round(total - sum(partes), decimales))
        return partes
    partes = [round(total * p / suma, decimales) for p in positivos]
    residuo = round(total - sum(partes), decimales)
    if residuo:
        mayor = max(range(n), key=lambda i: positivos[i])
        partes[mayor] = round(partes[mayor] + residuo, decimales)
    return partes

PERIODO = "2026-08"
HIST_DESDE, HIST_HASTA = "2026-07-01", "2026-08-01"
FUENTE = ROOT / "data/output/cupos" / PERIODO / "Objetivo AGOSTO 2026 Badie - Nahuel-2.xlsx"
SALIDA = ROOT / "data/output/cupos" / PERIODO / "CUPO DESAGREGADO POR RUTA SUCURSALES - AGOSTO 2026.xlsx"
HOJA_SALIDA = "Base Pivot SUCURSALES"
AMARILLO = "FFFF00"   # pestana de las hojas que se cargan en la base
DECIMALES = 2

PROV_JUJUY = {9, 10, 11, 12, 13, 14, 15}

# El objetivo nombra sucursales corto; dim_sucursal las nombra largo.
ALIAS_SUCURSAL = {"JVG": "JOAQUIN V GONZALEZ"}
# Altas/bajas que el objetivo todavia no refleja.
ALIAS_PREVENTISTA = {
    ("CRUZ GABRIEL ARNALDO", "LA QUIACA"): "CRUZ GABRIEL",
    # Figura en ABRA PAMPA (cerrada); sus rutas viven en LA QUIACA.
    ("LAMAS SEBASTIAN", "ABRA PAMPA"): ("LAMAS SEBASTIAN LQ", "LA QUIACA"),
}
# No es ruta de preventa: absorberia cupo que nadie persigue.
SIN_CUPO = {"DIRECTA"}

# Objetivo del mes anterior, usado SOLO como pesos de reparto entre preventistas
# (no como cupo). Ver `redistribuir_por_pesos`.
FUENTE_PESOS = ROOT / "data/output/cupos" / PERIODO / "OBJETIVOS JUL26 ORAN.xlsx"
# ORAN: AQUINO GUSTAVO ya no trabaja y no tiene rutas en dim_cliente. Sin esto
# su cupo cae en `huerfanos` y se PIERDE — la sucursal queda 2.393,76 bultos de
# CERVEZAS por debajo de su objetivo, y lo mismo en cada categoria. Con esto el
# TOTAL de cada categoria de la sucursal se vuelve a repartir entre los que
# siguen, con los pesos de julio.
SUCURSALES_REDISTRIBUIR = {"ORAN"}
# GUEMES tiene su propio archivo de cupo (CUPO DESAGREGADO POR RUTA GUEMES).
# Si entrara aca tambien, el mismo objetivo se cargaria dos veces en la base.
SUCURSALES_EXCLUIDAS = {16}
NOMBRES_EXCLUIDOS = {"GUEMES"}

INDIVIDUAL = "__INDIVIDUAL__"     # SALTA en Salta, NORTE en Jujuy
# Plural, como Casa Central y Guemes. En singular la base termina con dos
# valores para la misma categoria y el grano se mezcla.
MULTICERV = "MULTICERVEZAS"
IMPORTADAS = "IMPORTADAS"
MULTICCU = "MULTICCU"
# El objetivo trae MULTI CCU como un solo total por sucursal, pero Casa Central
# y Guemes lo cargan abierto en sus tres genericos. Se abre por historia para
# que el grano coincida; si no, bronze recibe dos granos distintos.
GEN_MULTICCU = ["VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]
MARCAS_IMPORTADAS = {"BLUE MOON", "KUNSTMAN", "KUNSTMANN"}
CERVEZA_CATS = [INDIVIDUAL, "HEINEKEN", "IMPERIAL", "MILLER", MULTICERV, IMPORTADAS]


def _txt(v) -> str:
    if v is None or (isinstance(v, float) and v != v):
        return ""
    s = unicodedata.normalize("NFKD", str(v).strip().upper())
    return "".join(c for c in s if not unicodedata.combining(c))


def leer_objetivo(path: Path):
    """(preventista, sucursal) -> {categoria: cupo}, y MULTI CCU por sucursal."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in ws[1]]
    # Patron `<CATEGORIA> | CUPO | %CUPO`: la etiqueta manda, el cupo es la de al lado.
    cols = {_txt(v): i + 1 for i, v in enumerate(hdr)
            if v and _txt(v) not in ("PREVENTISTA", "SUCURSAL", "CUPO", "%CUPO")}
    cupos: dict[tuple[str, str], dict[str, float]] = {}
    for r in ws.iter_rows(min_row=2, max_row=44, values_only=True):
        prev, suc = _txt(r[0]), _txt(r[1])
        if not prev or not suc or suc in ("SUCURSAL", "TOTAL"):
            continue
        d = cupos.setdefault((prev, suc), {})
        for cat, ci in cols.items():
            v = r[ci]
            if isinstance(v, (int, float)):
                d[cat] = d.get(cat, 0.0) + float(v)
    multiccu = {}
    if "Sheet1" in wb.sheetnames:
        for r in wb["Sheet1"].iter_rows(values_only=True):
            if r[0] and isinstance(r[1], (int, float)):
                multiccu[_txt(r[0])] = float(r[1])
    wb.close()
    return cupos, multiccu


def cargar_rutas(dl: DataLoader):
    df = dl.execute_query("""
        SELECT dc.id_sucursal,
               REPLACE(ds.descripcion, 'SUCURSAL ', '') AS suc_corta,
               ds.descripcion AS suc_larga,
               dc.id_ruta_fv1 AS ruta,
               MIN(dc.des_ruta_fv1) AS des_ruta,
               UPPER(TRIM(dc.des_personal_fv1)) AS preventista
        FROM gold.dim_cliente dc
        JOIN gold.dim_sucursal ds ON ds.id_sucursal = dc.id_sucursal
        WHERE COALESCE(dc.anulado, false) = false
          AND dc.id_sucursal NOT IN :excl
          AND dc.id_ruta_fv1 IS NOT NULL AND dc.des_personal_fv1 IS NOT NULL
        GROUP BY 1, 2, 3, 4, 6
    """, {"excl": tuple({1} | SUCURSALES_EXCLUIDAS)})
    rutas: dict[tuple[str, str], list] = {}
    meta: dict[int, tuple] = {}
    for f in df.itertuples(index=False):
        prev = _txt(f.preventista)
        if prev in SIN_CUPO:
            continue
        clave = (prev, _txt(f.suc_corta))
        # REGLA DE ORO: id_ruta se REUSA entre sucursales. La clave es
        # (id_sucursal, id_ruta); keyear por ruta sola colapsa la ruta 1 de
        # Oran con la 1 de Metan y se pierden dos tercios del reparto.
        rk = (int(f.id_sucursal), int(f.ruta))
        rutas.setdefault(clave, []).append(
            (rk, _txt(f.des_ruta) or f"RUTA {int(f.ruta)}"))
        meta[rk] = (int(f.id_sucursal), f.suc_larga)
    for v in rutas.values():
        v.sort()
    return rutas, meta


def cargar_historia(dl: DataLoader):
    """(ruta, categoria) -> bultos. La categoria depende de la provincia."""
    df = dl.execute_query("""
        SELECT dc.id_sucursal, dc.id_ruta_fv1 AS ruta, da.generico, da.marca,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.anulado = false AND fv.id_sucursal NOT IN :excl
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv1 IS NOT NULL
        GROUP BY 1, 2, 3, 4
    """, {"d": HIST_DESDE, "h": HIST_HASTA,
           "excl": tuple({1} | SUCURSALES_EXCLUIDAS)})
    hist: dict[tuple[int, str], float] = {}
    for f in df.itertuples(index=False):
        cat = clasificar(int(f.id_sucursal), f.generico, f.marca)
        if cat is None:
            continue
        k = (int(f.id_sucursal), int(f.ruta), cat)
        hist[k] = hist.get(k, 0.0) + float(f.qty or 0)
    return hist


def clasificar(id_sucursal: int, generico, marca) -> str | None:
    g, m = _txt(generico), _txt(marca)
    if g == "CERVEZAS":
        propia = "NORTE" if id_sucursal in PROV_JUJUY else "SALTA"
        if m == propia:
            return INDIVIDUAL
        if m in ("HEINEKEN", "IMPERIAL", "MILLER"):
            return m
        if m in MARCAS_IMPORTADAS:
            return IMPORTADAS
        return MULTICERV          # SCHNEIDER, la otra grande, y todo el resto
    if g == "AGUAS DANONE":
        return "AGUA DANONE"
    if g in GEN_MULTICCU:
        return g                  # cada generico guarda su propia historia
    return None


def etiqueta_individual(id_sucursal: int) -> str:
    return "NORTE" if id_sucursal in PROV_JUJUY else "SALTA"


def resolver(prev: str, suc: str, rutas: dict):
    """Aplica los alias del objetivo contra los nombres de dim_cliente."""
    alias = ALIAS_PREVENTISTA.get((prev, suc))
    if isinstance(alias, tuple):
        prev, suc = alias
    elif alias:
        prev = alias
    suc = ALIAS_SUCURSAL.get(suc, suc)
    return (prev, suc) if (prev, suc) in rutas else None


def redistribuir_por_pesos(cupos, rutas, pesos, sucursales):
    """Reparte el TOTAL de cada categoria de la sucursal entre los preventistas
    que siguen dados de alta, ponderando por `pesos`.

    Muta `cupos`. El total de la sucursal se conserva EXACTO: es el mismo numero
    que trae el objetivo, solo cambia de manos. Los preventistas que ya no tienen
    rutas se borran del dict — si quedaran, `repartir` los mandaria a huerfanos y
    su cupo se perderia por segunda vez.
    """
    informe = []
    for suc in sorted(sucursales):
        de_suc = {p: c for (p, s), c in cupos.items() if s == suc}
        if not de_suc:
            continue
        vivos = sorted(p for p in de_suc if resolver(p, suc, rutas) is not None)
        bajas = sorted(p for p in de_suc if p not in vivos)
        if not vivos:
            continue
        totales = {}
        for cat in sorted({c for d in de_suc.values() for c in d}):
            total = round(sum(d.get(cat, 0.0) for d in de_suc.values()), DECIMALES)
            w = [pesos.get((p, suc), {}).get(cat, 0.0) for p in vivos]
            totales[cat] = (total, sum(w))
            for prev, parte in zip(vivos, split_proporcional(total, w, DECIMALES)):
                cupos[(prev, suc)][cat] = parte
        for prev in bajas:
            del cupos[(prev, suc)]
        informe.append((suc, vivos, bajas, totales))
    return informe


def repartir(cupos, multiccu, rutas, meta, hist):
    """cupo del preventista -> sus rutas, proporcional a la historia."""
    valores: dict[tuple, dict[str, float]] = {}
    huerfanos, avisos, excluidos = [], [], []

    for (prev, suc), cats in sorted(cupos.items()):
        if suc in NOMBRES_EXCLUIDOS:
            excluidos.append((prev, suc,
                              sum(v for k, v in cats.items() if k != "CERVEZAS")))
            continue
        clave = resolver(prev, suc, rutas)
        if clave is None:
            total = sum(v for k, v in cats.items() if k != "CERVEZAS")
            huerfanos.append((prev, suc, total))
            continue
        if clave != (prev, ALIAS_SUCURSAL.get(suc, suc)):
            avisos.append(f"{prev} ({suc}) -> {clave[0]} ({clave[1]})")
        de_prev = rutas[clave]
        for cat, cupo in cats.items():
            # CERVEZAS es el total; sale de la suma de sus categorias.
            if cat == "CERVEZAS" or not cupo:
                continue
            # El objetivo de CCU dice MULTICERVEZA en singular; la base usa el
            # plural. Se normaliza aca y no en la salida: si no, la categoria
            # queda en cero y CERVEZAS pierde ese volumen sin avisar.
            interno = {"SALTA": INDIVIDUAL,
                       "MULTICERVEZA": MULTICERV}.get(cat, cat)
            pesos = [hist.get((s, r, interno), 0.0) for (s, r), _ in de_prev]
            partes = split_proporcional(float(cupo), pesos, DECIMALES)
            for (rk, _), parte in zip(de_prev, partes):
                valores.setdefault(rk, {})[interno] = parte

    # MULTI CCU llega por sucursal: se abre entre TODAS las rutas de esa sucursal.
    for suc_obj, cupo in multiccu.items():
        if suc_obj in NOMBRES_EXCLUIDOS:
            excluidos.append(("(MULTI CCU)", suc_obj, cupo))
            continue
        suc = ALIAS_SUCURSAL.get(suc_obj, suc_obj)
        de_suc = sorted({rk for (p, s), rs in rutas.items() if s == suc for rk, _ in rs})
        if not de_suc:
            huerfanos.append(("(MULTI CCU)", suc_obj, cupo))
            continue
        # Una sola pasada sobre los pares (ruta x generico): asi la suma de los
        # tres da EXACTO el total de la sucursal, sin arrastrar el redondeo de
        # cada generico por separado.
        pares = [(rk, gen) for rk in de_suc for gen in GEN_MULTICCU]
        pesos = [hist.get((rk[0], rk[1], gen), 0.0) for rk, gen in pares]
        for (rk, gen), parte in zip(pares, split_proporcional(float(cupo), pesos, DECIMALES)):
            valores.setdefault(rk, {})[gen] = parte
    return valores, huerfanos, avisos, excluidos


def construir_workbook(valores, rutas, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_SALIDA
    ws.append(["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA", "NIVEL", "GRUPO",
               "CATEGORIA", "CUPO"])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    ruta_prev = {rk: p for (p, s), rs in rutas.items() for rk, _ in rs}
    ruta_des = {rk: d for rs in rutas.values() for rk, d in rs}
    for rk in sorted(valores, key=lambda k: (k[0], k[1])):
        id_suc, suc_larga = meta[rk]
        ruta = rk[1]
        zona = f"{id_suc} - {suc_larga}"
        v = valores[rk]
        indiv = etiqueta_individual(id_suc)
        # Convencion vigente (la de `CORRECION MULTI`, elegida por Nahuel el
        # 2026-08-18): GRUPO = CATEGORIA = la etiqueta, salvo los tres genericos
        # de MULTI CCU, que llevan CATEGORIA=MULTICCU para poder agruparlos.
        #
        # CERVEZAS va como DETALLE, o sea que SI se carga: en gold el
        # `generico='CERVEZAS'` es el TOTAL y convive con sus marcas. Sumar
        # todos los genericos de una ruta cuenta la cerveza dos veces; hay que
        # elegir el total O el detalle, nunca los dos.
        cerv = round(sum(v.get(c, 0.0) for c in CERVEZA_CATS), DECIMALES)
        filas = [("DETALLE", "CERVEZAS", "CERVEZAS", cerv)]
        for cat in CERVEZA_CATS:
            etq = indiv if cat == INDIVIDUAL else cat
            filas.append(("DETALLE", etq, etq, v.get(cat, 0.0)))
        filas.append(("DETALLE", "AGUAS DANONE", "AGUAS DANONE", v.get("AGUA DANONE", 0.0)))
        total_multi = round(sum(v.get(g, 0.0) for g in GEN_MULTICCU), DECIMALES)
        filas.append(("AGREGADO", "TOTAL MULTICCU", "TOTAL MULTICCU", total_multi))
        for gen in GEN_MULTICCU:
            filas.append(("DETALLE", gen, MULTICCU, v.get(gen, 0.0)))
        for nivel, grupo, categoria, cupo in filas:
            ws.append([zona, ruta_prev.get(rk, ""), ruta, ruta_des.get(rk, ""),
                       nivel, grupo, categoria, cupo])

    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(fila):
            c.border = borde
            if i == 2:
                c.number_format = "0"
            if i == 7:
                c.number_format = "#,##0.00"
    for col, w in zip("ABCDEFGH", (26, 24, 9, 24, 11, 16, 16, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    # Amarillo = esta hoja se carga en la base. Es la senal para no confundirla
    # con las satelite, que existen solo para leer.
    ws.sheet_properties.tabColor = AMARILLO

    _hoja_matriz(wb, "Cupo Preventista", valores, rutas, meta, por_ruta=False)
    _hoja_matriz(wb, "Cupo Ruta", valores, rutas, meta, por_ruta=True)
    return wb


def _hoja_matriz(wb, titulo, valores, rutas, meta, por_ruta: bool):
    """Categorias en columnas; vendedores (o rutas) en filas."""
    ruta_prev = {rk: p for (p, s), rs in rutas.items() for rk, _ in rs}
    ruta_des = {rk: d for rs in rutas.values() for rk, d in rs}

    filas: dict[tuple, dict[str, float]] = {}
    etiquetas: dict[tuple, list] = {}
    for rk, v in valores.items():
        id_suc, suc_larga = meta[rk]
        zona = f"{id_suc} - {suc_larga}"
        prev = ruta_prev.get(rk, "")
        clave = (id_suc, rk[1]) if por_ruta else (id_suc, prev)
        acc = filas.setdefault(clave, {})
        etiquetas[clave] = ([zona, prev, rk[1], ruta_des.get(rk, "")]
                            if por_ruta else [zona, prev])
        indiv = etiqueta_individual(id_suc)
        for cat in CERVEZA_CATS + ["AGUA DANONE"] + GEN_MULTICCU:
            etq = indiv if cat == INDIVIDUAL else cat
            acc[etq] = acc.get(etq, 0.0) + v.get(cat, 0.0)

    cols = ["CERVEZAS", "SALTA", "NORTE", "HEINEKEN", "IMPERIAL", "MILLER",
            MULTICERV, IMPORTADAS, "AGUA DANONE"] + GEN_MULTICCU + ["TOTAL MULTICCU"]
    cabecera = (["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA"] if por_ruta
                else ["ZONA", "PREVENTISTA"])
    ws = wb.create_sheet(titulo)
    ws.append(cabecera + cols)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    total = {c: 0.0 for c in cols}
    for clave in sorted(filas, key=lambda k: (k[0], k[1])):
        d = filas[clave]
        # CERVEZAS es la suma de sus cinco, no un dato aparte.
        d["CERVEZAS"] = round(sum(d.get(c, 0.0) for c in
                                  ("SALTA", "NORTE", "HEINEKEN", "IMPERIAL",
                                   "MILLER", MULTICERV, IMPORTADAS)), DECIMALES)
        d["TOTAL MULTICCU"] = round(sum(d.get(g, 0.0) for g in GEN_MULTICCU), DECIMALES)
        ws.append(etiquetas[clave] + [round(d.get(c, 0.0), DECIMALES) for c in cols])
        for c in cols:
            total[c] += d.get(c, 0.0)
    ws.append(["TOTAL GENERAL"] + [""] * (len(cabecera) - 1)
              + [round(total[c], DECIMALES) for c in cols])

    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    n = len(cabecera)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        es_total = fila[0].value == "TOTAL GENERAL"
        for i, c in enumerate(fila):
            c.border = borde
            if por_ruta and i == 2:
                c.number_format = "0"
            elif i >= n:
                c.number_format = "#,##0.00"
            if es_total:
                c.font, c.fill = Font(bold=True), PatternFill("solid", fgColor="FCE4D6")
    anchos = ([26, 24, 9, 24] if por_ruta else [26, 24]) + [13] * len(cols)
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=2, column=n + 1).coordinate
    return ws


def main() -> int:
    if SALIDA.exists() and "--force" not in sys.argv:
        print(f"Ya existe y NO se regenera:\n  {SALIDA}")
        print(f"  modificado: {datetime.fromtimestamp(SALIDA.stat().st_mtime):%Y-%m-%d %H:%M}")
        print("\nGuarda una copia y corre con --force.")
        return 1

    cupos, multiccu = leer_objetivo(FUENTE)
    dl = DataLoader()
    rutas, meta = cargar_rutas(dl)
    hist = cargar_historia(dl)
    if SUCURSALES_REDISTRIBUIR:
        if not FUENTE_PESOS.exists():
            print(f"No existe el objetivo de referencia: {FUENTE_PESOS}")
            return 3
        pesos, _ = leer_objetivo(FUENTE_PESOS)
        for suc, vivos, bajas, totales in redistribuir_por_pesos(
                cupos, rutas, pesos, SUCURSALES_REDISTRIBUIR):
            print(f"Redistribucion {suc} con pesos de {FUENTE_PESOS.name}")
            print(f"  entre {len(vivos)}: {', '.join(vivos)}")
            if bajas:
                print(f"  BAJAS (su cupo se reparte): {', '.join(bajas)}")
            for cat, (total, sw) in sorted(totales.items()):
                aviso = "  <-- sin pesos, reparto parejo" if sw <= 0 else ""
                print(f"    {cat:<16} total {total:>12,.2f}{aviso}")
    valores, huerfanos, avisos, excluidos = repartir(cupos, multiccu, rutas, meta, hist)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(valores, rutas, meta).save(SALIDA)

    print(f"Fuente : {FUENTE.name}")
    print(f"Guardado: {SALIDA.name}")
    print(f"  {len(valores)} rutas | {len(cupos)} preventistas en el objetivo")
    if avisos:
        print("\n  Nombres reasignados:")
        for a in avisos:
            print(f"    {a}")
    if excluidos:
        tot = sum(t for _, _, t in excluidos)
        print(f"\n  EXCLUIDOS a proposito ({tot:,.0f} bultos) — tienen archivo propio:")
        for p_, s_, t_ in excluidos:
            print(f"    {p_} ({s_}): {t_:,.0f}")
    if huerfanos:
        perdido = sum(t for _, _, t in huerfanos)
        print(f"\n  SIN RUTAS — su cupo NO se reparte ({perdido:,.0f} bultos):")
        for p, s, t in huerfanos:
            print(f"    {p} ({s}): {t:,.0f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
