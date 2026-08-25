"""Genera una calculadora de costo de tokens por usuario, en Excel.

Que es
------
Un libro con formulas VIVAS: se tipea en la hoja `Calculadora` (tokens por
tarea, tareas por usuario, cantidad de usuarios) y el costo se recalcula solo.
No es un informe con numeros congelados.

Los precios salen de la API publica de OpenRouter, que publica el precio por
token de cada modelo. Se traen una vez al generar el archivo y quedan escritos
en la hoja `Precios` con la fecha de captura: los precios cambian, y un numero
sin fecha al lado no se puede auditar. Volver a correr este script los refresca.

    python scripts/calculadora_costos_tokens.py
    python scripts/calculadora_costos_tokens.py --salida /ruta/al.xlsx
    python scripts/calculadora_costos_tokens.py --sin-red   # usa precios de respaldo

Por que OpenRouter y no la API de cada proveedor: ni Anthropic ni OpenAI
publican un endpoint de precios. OpenRouter mantiene el catalogo de todos y su
endpoint /models es publico y no pide credenciales.
"""
from __future__ import annotations

import argparse
import json
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

API_MODELOS = "https://openrouter.ai/api/v1/models"
TIMEOUT_SEG = 25

# --- Paleta por rol (skills/formato-excel-badie) -----------------------------
HEADER_FILL = "2E75B6"
BANDA_FILL = "DDEBF7"
INPUT_FILL = "FFF2CC"       # celda que el usuario TIPEA
RESULTADO_FILL = "FFE08A"   # el numero que se viene a buscar
ALERTA_FILL = "FFC7CE"
ZEBRA_FILL = "F7F9FC"
FUENTE_HEADER = "FFFFFF"
FUENTE_SUBTITULO = "546E7A"
BORDE = Side(style="thin", color="D9D9D9")
BORDE_FUERTE = Side(style="medium", color="8EA9DB")

FMT_TOKENS = "#,##0"
FMT_USD = '"US$" #,##0.0000'
FMT_USD_GRANDE = '"US$" #,##0.00'
FMT_ARS = '"$" #,##0.00'
FMT_ENTERO = "#,##0"

# Familias que entran en la hoja de precios. El catalogo completo son 400+
# modelos y una lista asi no se puede usar en un desplegable.
FAMILIAS = (
    "anthropic/",
    "openai/gpt-",
    "openai/o1",
    "openai/o3",
    "openai/o4",
    "google/gemini-",
    "deepseek/deepseek-",
    "meta-llama/llama-3",
    "meta-llama/llama-4",
    "mistralai/mistral-",
    "x-ai/grok-",
)

# Respaldo por si no hay red. Se marcan como tales en la hoja para que nadie
# los confunda con precios frescos.
PRECIOS_RESPALDO: list[tuple[str, str, float, float]] = [
    ("anthropic/claude-opus-4.8", "Claude Opus 4.8", 5.0, 25.0),
    ("anthropic/claude-sonnet-4.6", "Claude Sonnet 4.6", 3.0, 15.0),
    ("openai/gpt-4o", "GPT-4o", 2.5, 10.0),
    ("google/gemini-2.0-flash-lite-001", "Gemini 2.0 Flash Lite", 0.075, 0.30),
]


def traer_precios(sin_red: bool = False) -> tuple[list[tuple[str, str, float, float]], str]:
    """Precios por millon de tokens: [(id, nombre, usd_in, usd_out)].

    Returns:
        ``(precios, fuente)``, donde `fuente` describe de donde salieron para
        poder escribirlo en la hoja.
    """
    if sin_red:
        return PRECIOS_RESPALDO, "RESPALDO local (--sin-red): pueden estar vencidos"

    try:
        with urllib.request.urlopen(API_MODELOS, timeout=TIMEOUT_SEG) as resp:
            catalogo = json.loads(resp.read())["data"]
    except Exception as exc:  # red caida, DNS, timeout, formato inesperado
        print(f"  ! no se pudo leer {API_MODELOS}: {exc}")
        print("  ! se usan los precios de respaldo, marcados como tales en la hoja")
        return PRECIOS_RESPALDO, f"RESPALDO local — la API fallo ({type(exc).__name__})"

    filas: list[tuple[str, str, float, float]] = []
    for modelo in catalogo:
        mid = modelo.get("id", "")
        if not mid.startswith(FAMILIAS):
            continue
        precios = modelo.get("pricing") or {}
        try:
            usd_in = float(precios.get("prompt", 0)) * 1_000_000
            usd_out = float(precios.get("completion", 0)) * 1_000_000
        except (TypeError, ValueError):
            continue
        # Precio 0 = modelo gratuito o sin precio publicado. No sirve para
        # presupuestar y ensucia el desplegable.
        if usd_in <= 0 and usd_out <= 0:
            continue
        filas.append((mid, modelo.get("name") or mid, usd_in, usd_out))

    filas.sort(key=lambda f: (f[0].split("/")[0], -f[3]))
    return filas, f"{API_MODELOS} (publica, sin credenciales)"


def _titulo(ws, texto: str, subtitulo: str) -> None:
    ws.cell(1, 1, texto).font = Font(bold=True, size=14)
    c = ws.cell(2, 1, subtitulo)
    c.font = Font(italic=True, size=10, color=FUENTE_SUBTITULO)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _borde(todo: Side = BORDE) -> Border:
    return Border(left=todo, right=todo, top=todo, bottom=todo)


def hoja_precios(wb: Workbook, precios: list, fuente: str, capturado: str):
    ws = wb.create_sheet("Precios")
    _titulo(
        ws,
        "Precios por millon de tokens (USD)",
        f"Fuente: {fuente} | Capturado: {capturado} | "
        f"Para refrescar: python scripts/calculadora_costos_tokens.py",
    )

    encabezados = ["ID del modelo", "Nombre", "USD / 1M entrada", "USD / 1M salida"]
    for j, texto in enumerate(encabezados, start=1):
        c = ws.cell(4, j, texto)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.font = Font(bold=True, color=FUENTE_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()

    for i, (mid, nombre, usd_in, usd_out) in enumerate(precios):
        r = 5 + i
        for j, valor in enumerate((mid, nombre, usd_in, usd_out), start=1):
            c = ws.cell(r, j, valor)
            # El valor se guarda con todos sus decimales; el formato es de
            # presentacion. Un precio de 0,075 no se puede redondear a 0.
            if j >= 3:
                c.number_format = FMT_USD
            if i % 2:
                c.fill = PatternFill("solid", fgColor=ZEBRA_FILL)
            c.border = _borde()

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:D{4 + len(precios)}"
    return ws


def hoja_calculadora(wb: Workbook, n_precios: int, capturado: str):
    """La hoja que se tipea. Todo lo de abajo son formulas vivas."""
    ws = wb.create_sheet("Calculadora", 0)
    _titulo(
        ws,
        "Calculadora de costo de tokens por usuario",
        "Tipea SOLO las celdas en color crema. Todo lo demas son formulas: "
        "se recalcula solo al cambiar cualquier dato.",
    )

    ultima = 4 + n_precios

    def etiqueta(r: int, texto: str, nota: str = "") -> None:
        c = ws.cell(r, 1, texto)
        c.font = Font(bold=True)
        c.border = _borde()
        if nota:
            n = ws.cell(r, 4, nota)
            n.font = Font(italic=True, size=9, color=FUENTE_SUBTITULO)

    def entrada(r: int, valor, fmt: str) -> None:
        c = ws.cell(r, 2, valor)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=INPUT_FILL)
        c.font = Font(bold=True)
        c.border = Border(left=BORDE_FUERTE, right=BORDE_FUERTE, top=BORDE_FUERTE, bottom=BORDE_FUERTE)

    def formula(r: int, expr: str, fmt: str, destacar: bool = False) -> None:
        c = ws.cell(r, 2, expr)
        c.number_format = fmt
        c.border = _borde()
        if destacar:
            c.fill = PatternFill("solid", fgColor=RESULTADO_FILL)
            c.font = Font(bold=True)

    def banda(r: int, texto: str) -> None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1, texto)
        c.fill = PatternFill("solid", fgColor=BANDA_FILL)
        c.font = Font(bold=True)
        for j in range(1, 5):
            ws.cell(r, j).border = Border(top=BORDE_FUERTE, bottom=BORDE, left=BORDE, right=BORDE)

    # --- entradas ---
    banda(4, "LO QUE TIPEAS")

    etiqueta(5, "Modelo", "elegir de la lista (hoja Precios)")
    entrada(5, "anthropic/claude-sonnet-4.6", "@")
    dv = DataValidation(type="list", formula1=f"=Precios!$A$5:$A${ultima}", allow_blank=False)
    dv.error = "Tiene que ser un modelo de la hoja Precios"
    dv.errorTitle = "Modelo desconocido"
    ws.add_data_validation(dv)
    dv.add(ws.cell(5, 2))

    etiqueta(6, "Tokens de ENTRADA por tarea", "prompt + contexto + archivos")
    entrada(6, 12000, FMT_TOKENS)

    etiqueta(7, "Tokens de SALIDA por tarea", "lo que responde el modelo")
    entrada(7, 2500, FMT_TOKENS)

    etiqueta(8, "Tareas por usuario por mes", "consultas, informes, etc.")
    entrada(8, 60, FMT_ENTERO)

    etiqueta(9, "Cantidad de usuarios", "")
    entrada(9, 8, FMT_ENTERO)

    etiqueta(10, "Cotizacion USD -> ARS", "para ver el costo en pesos")
    entrada(10, 1000, FMT_ARS)

    # --- precio del modelo elegido ---
    banda(12, "PRECIO DEL MODELO ELEGIDO (se busca solo en la hoja Precios)")
    etiqueta(13, "USD por 1M de tokens de entrada")
    formula(13, f'=IFERROR(VLOOKUP($B$5,Precios!$A$5:$D${ultima},3,FALSE),"modelo no encontrado")', FMT_USD)
    etiqueta(14, "USD por 1M de tokens de salida")
    formula(14, f'=IFERROR(VLOOKUP($B$5,Precios!$A$5:$D${ultima},4,FALSE),"modelo no encontrado")', FMT_USD)
    etiqueta(15, "Nombre comercial")
    ws.cell(15, 2, f'=IFERROR(VLOOKUP($B$5,Precios!$A$5:$D${ultima},2,FALSE),"— elegi un modelo de la lista —")').border = _borde()

    # --- resultados ---
    banda(17, "COSTO POR TAREA")
    etiqueta(18, "Costo de entrada")
    formula(18, "=B6/1000000*B13", FMT_USD)
    etiqueta(19, "Costo de salida")
    formula(19, "=B7/1000000*B14", FMT_USD)
    etiqueta(20, "COSTO POR TAREA (USD)")
    formula(20, "=B18+B19", FMT_USD, destacar=True)
    etiqueta(21, "COSTO POR TAREA (ARS)")
    formula(21, "=B20*B10", FMT_ARS, destacar=True)

    banda(23, "COSTO POR USUARIO")
    etiqueta(24, "Tokens de entrada por usuario / mes")
    formula(24, "=B6*B8", FMT_TOKENS)
    etiqueta(25, "Tokens de salida por usuario / mes")
    formula(25, "=B7*B8", FMT_TOKENS)
    etiqueta(26, "COSTO POR USUARIO / MES (USD)")
    formula(26, "=B20*B8", FMT_USD_GRANDE, destacar=True)
    etiqueta(27, "COSTO POR USUARIO / MES (ARS)")
    formula(27, "=B26*B10", FMT_ARS, destacar=True)

    banda(29, "COSTO TOTAL")
    etiqueta(30, "Tareas totales / mes")
    formula(30, "=B8*B9", FMT_ENTERO)
    etiqueta(31, "Tokens totales / mes")
    formula(31, "=(B6+B7)*B30", FMT_TOKENS)
    etiqueta(32, "COSTO TOTAL / MES (USD)")
    formula(32, "=B26*B9", FMT_USD_GRANDE, destacar=True)
    etiqueta(33, "COSTO TOTAL / MES (ARS)")
    formula(33, "=B32*B10", FMT_ARS, destacar=True)
    etiqueta(34, "COSTO TOTAL / ANIO (USD)")
    formula(34, "=B32*12", FMT_USD_GRANDE, destacar=True)
    etiqueta(35, "COSTO TOTAL / ANIO (ARS)")
    formula(35, "=B34*B10", FMT_ARS, destacar=True)

    # --- aviso de vencimiento de los precios ---
    c = ws.cell(37, 1, f"Precios capturados el {capturado}. Los proveedores los cambian: "
                       f"si esta fecha quedo vieja, volve a correr el script para refrescarlos.")
    c.font = Font(italic=True, size=9, color="9C0006")
    c.fill = PatternFill("solid", fgColor=ALERTA_FILL)
    ws.merge_cells(start_row=37, start_column=1, end_row=37, end_column=4)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 40
    return ws


def hoja_sensibilidad(wb: Workbook):
    """Usuarios x tareas, con formulas que apuntan a la Calculadora.

    Sirve para la pregunta que sigue siempre a la primera: "y si en vez de 8
    usuarios son 20". Cambiando el modelo o los tokens, la matriz entera se
    mueve sola.
    """
    ws = wb.create_sheet("Sensibilidad")
    _titulo(
        ws,
        "Costo total por mes (USD) segun usuarios y tareas",
        "Usa el modelo y los tokens por tarea de la hoja Calculadora. "
        "Cambia cualquiera de los dos y esta matriz se recalcula sola.",
    )

    usuarios = [1, 3, 5, 8, 10, 15, 20, 30, 50]
    tareas = [10, 20, 40, 60, 80, 120, 200, 300]

    c = ws.cell(4, 1, "Usuarios \\ Tareas por usuario/mes")
    c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    c.font = Font(bold=True, color=FUENTE_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _borde()

    for j, t in enumerate(tareas, start=2):
        c = ws.cell(4, j, t)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.font = Font(bold=True, color=FUENTE_HEADER)
        c.alignment = Alignment(horizontal="center")
        c.number_format = FMT_ENTERO
        c.border = _borde()

    for i, u in enumerate(usuarios, start=5):
        c = ws.cell(i, 1, u)
        c.fill = PatternFill("solid", fgColor=BANDA_FILL)
        c.font = Font(bold=True)
        c.number_format = FMT_ENTERO
        c.border = _borde()
        for j, _ in enumerate(tareas, start=2):
            col = get_column_letter(j)
            # costo por tarea (Calculadora!B20) x tareas x usuarios
            c = ws.cell(i, j, f"=Calculadora!$B$20*{col}$4*$A{i}")
            c.number_format = FMT_USD_GRANDE
            c.border = _borde()
            if (i - 5) % 2:
                c.fill = PatternFill("solid", fgColor=ZEBRA_FILL)

    ws.column_dimensions["A"].width = 32
    for j in range(2, len(tareas) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = "B5"
    return ws


def hoja_por_tarea(wb: Workbook, n_precios: int):
    """Detalle tipeable: una fila por tipo de tarea.

    La calculadora de la primera hoja asume que todas las tareas cuestan lo
    mismo. En la practica no: una consulta corta y un informe de 30 hojas no
    gastan igual. Aca se carga cada tipo con su propio consumo.
    """
    ws = wb.create_sheet("Por Tarea")
    _titulo(
        ws,
        "Costo por tipo de tarea",
        "Tipea las columnas en crema. El modelo se elige por fila: podes "
        "presupuestar tareas baratas con un modelo chico y las pesadas con uno grande.",
    )

    ultima = 4 + n_precios
    encabezados = [
        ("Tipo de tarea", 30, None),
        ("Modelo", 34, None),
        ("Tokens entrada", 15, FMT_TOKENS),
        ("Tokens salida", 15, FMT_TOKENS),
        ("Veces / mes", 13, FMT_ENTERO),
        ("Usuarios", 11, FMT_ENTERO),
        ("USD / tarea", 15, FMT_USD),
        ("USD / mes", 15, FMT_USD_GRANDE),
        ("ARS / mes", 16, FMT_ARS),
    ]
    for j, (texto, ancho, _) in enumerate(encabezados, start=1):
        c = ws.cell(4, j, texto)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.font = Font(bold=True, color=FUENTE_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()
        ws.column_dimensions[get_column_letter(j)].width = ancho

    dv = DataValidation(type="list", formula1=f"=Precios!$A$5:$A${ultima}", allow_blank=True)
    ws.add_data_validation(dv)

    ejemplos = [
        ("Consulta corta por WhatsApp", "google/gemini-2.5-flash-lite", 2000, 500, 400, 8),
        ("Informe mensual", "anthropic/claude-sonnet-4.6", 40000, 8000, 20, 3),
        ("Analisis profundo", "anthropic/claude-opus-4.8", 120000, 15000, 4, 2),
    ]
    PRIMERA, ULTIMA_FILA = 5, 34
    for i in range(PRIMERA, ULTIMA_FILA + 1):
        datos = ejemplos[i - PRIMERA] if i - PRIMERA < len(ejemplos) else ("", "", None, None, None, None)
        for j, valor in enumerate(datos, start=1):
            c = ws.cell(i, j, valor if valor != "" else None)
            c.fill = PatternFill("solid", fgColor=INPUT_FILL)
            c.border = _borde()
            if j in (3, 4):
                c.number_format = FMT_TOKENS
            elif j in (5, 6):
                c.number_format = FMT_ENTERO
        dv.add(ws.cell(i, 2))

        # USD por tarea: solo si la fila tiene modelo, para no llenar la hoja
        # de #N/A en las filas vacias.
        # IFERROR devuelve TEXTO, no cero, cuando el modelo no esta en la lista:
        # SUM ignora el texto, asi que el TOTAL sigue siendo valido y la fila
        # rota se ve. Con #N/A el error se propagaba y el archivo entero
        # parecia roto; con 0 la fila desaparecia del total en silencio.
        ws.cell(i, 7, (
            f'=IF($B{i}="","",IFERROR('
            f'C{i}/1000000*VLOOKUP($B{i},Precios!$A$5:$D${ultima},3,FALSE)'
            f'+D{i}/1000000*VLOOKUP($B{i},Precios!$A$5:$D${ultima},4,FALSE),'
            f'"modelo no esta en la hoja Precios"))'
        )).number_format = FMT_USD
        ws.cell(i, 8, f'=IF(ISNUMBER(G{i}),G{i}*E{i}*F{i},"")').number_format = FMT_USD_GRANDE
        ws.cell(i, 9, f'=IF(ISNUMBER(H{i}),H{i}*Calculadora!$B$10,"")').number_format = FMT_ARS
        for j in (7, 8, 9):
            ws.cell(i, j).border = _borde()

    # TOTAL GENERAL: todo informe lleva su fila de totales.
    r = ULTIMA_FILA + 1
    c = ws.cell(r, 1, "TOTAL GENERAL")
    for j in range(1, 10):
        cel = ws.cell(r, j)
        cel.fill = PatternFill("solid", fgColor=RESULTADO_FILL)
        cel.font = Font(bold=True)
        cel.border = Border(left=BORDE, right=BORDE, top=BORDE_FUERTE, bottom=BORDE_FUERTE)
    ws.cell(r, 5, f"=SUM(E{PRIMERA}:E{ULTIMA_FILA})").number_format = FMT_ENTERO
    ws.cell(r, 8, f"=SUM(H{PRIMERA}:H{ULTIMA_FILA})").number_format = FMT_USD_GRANDE
    ws.cell(r, 9, f"=SUM(I{PRIMERA}:I{ULTIMA_FILA})").number_format = FMT_ARS

    # Los usuarios NO se suman: el mismo usuario hace varios tipos de tarea.
    c = ws.cell(r + 2, 1, "La columna Usuarios no se totaliza a proposito: el mismo "
                          "usuario hace varios tipos de tarea y sumarla lo contaria "
                          "una vez por fila.")
    c.font = Font(italic=True, size=9, color=FUENTE_SUBTITULO)
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=9)

    ws.freeze_panes = "A5"
    return ws


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--salida", default="calculadora_costos_tokens.xlsx")
    parser.add_argument("--sin-red", action="store_true",
                        help="no consulta la API; usa los precios de respaldo")
    args = parser.parse_args()

    print(f"Trayendo precios de {API_MODELOS} ...")
    precios, fuente = traer_precios(args.sin_red)
    capturado = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"  {len(precios)} modelos con precio publicado")

    wb = Workbook()
    wb.remove(wb.active)
    hoja_precios(wb, precios, fuente, capturado)
    hoja_calculadora(wb, len(precios), capturado)
    hoja_por_tarea(wb, len(precios))
    hoja_sensibilidad(wb)
    # La Calculadora se inserto en la posicion 0; el orden final queda
    # Calculadora, Precios, Por Tarea, Sensibilidad.
    wb.move_sheet("Por Tarea", offset=-1)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))
    print(f"OK -> {salida}")
    print(f"   hojas: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
