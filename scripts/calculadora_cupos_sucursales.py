"""Calculadora de cupos del INTERIOR: se tipea el cupo por sucursal y categoria,
y el objetivo por preventista y por ruta sale con formulas de Excel.

Invierte el flujo que habia hasta agosto 2026. Antes CCU pasaba el objetivo ya
abierto por preventista y `apertura_rutas_sucursales.py` lo bajaba a ruta. Ahora
el objetivo llega por SUCURSAL y la apertura a preventista tambien la hace este
libro — en vivo, con formulas, para poder tantear un cupo y ver el reparto sin
volver a correr nada.

Reglas de negocio (las mismas que el script de apertura)
--------------------------------------------------------
- La ruta sale de `dim_cliente.id_ruta_fv1`, NUNCA del vendedor que facturo en
  `fact_ventas`: ese es quien emitio la factura, no quien persigue el objetivo.
- La clave es COMPUESTA `(id_sucursal, id_ruta)`. `id_ruta` se reusa entre
  sucursales; keyear por ruta sola colapsa la ruta 1 de Oran con la de Metan.
- El reparto es en cascada: SUCURSAL -> PREVENTISTA -> RUTA. La ruta reparte el
  objetivo de SU preventista, no el de la sucursal, asi la suma de las rutas de
  un preventista da exactamente su objetivo.
- Las ventas negativas (notas de credito) pesan CERO, nunca restan cupo.
- Si nadie tiene historia en una sucursal-categoria, el reparto es parejo.
- NO se redondea en ninguna formula. La suma de las partes da exacto el cupo
  tipeado; los decimales se esconden con formato de celda, no con ROUND.

Categorias (cambio de SEPTIEMBRE 2026)
--------------------------------------
SALTA, SCHNEIDER y NORTE pasan a ser marca individual en TODAS las sucursales.
Antes la propia de la provincia iba individual y las otras dos caian en
MULTICERVEZAS (NORTE individual en Jujuy, SALTA individual en Salta). Eso ya no
existe: las tres tienen columna propia y MULTICERVEZAS queda con el resto.

MULTI CCU se tipea como UN solo cupo por sucursal y se abre en sus tres
genericos por historia, como venia. Casa Central y Guemes los cargan abiertos y
si el grano no coincide bronze recibe dos granos distintos.
"""
from __future__ import annotations

import sys
import unicodedata
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.data_loader import DataLoader  # noqa: E402
from src.core.output_paths import service_output_dir  # noqa: E402

MESES_HISTORIA = 3
# Casa Central (1) y Guemes (16) tienen su propio archivo de cupo. Si entraran
# aca tambien, el mismo objetivo se cargaria dos veces en la base.
SUCURSALES_EXCLUIDAS = {1, 16}
# No es ruta de preventa: absorberia cupo que nadie persigue.
SIN_CUPO = {"DIRECTA", "RUIZ MARCELO", "VENDEDOR CHOPERAS"}

# --- Categorias -------------------------------------------------------------
MULTICERV = "MULTICERVEZAS"
IMPORTADAS = "IMPORTADAS"
MULTICCU = "MULTI CCU"
GEN_MULTICCU = ["VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]
MARCAS_IMPORTADAS = {"BLUE MOON", "KUNSTMAN", "KUNSTMANN"}
CERVEZA_INDIVIDUALES = ["SALTA", "SCHNEIDER", "NORTE", "HEINEKEN", "IMPERIAL", "MILLER"]
# El orden manda: es el de las filas del libro y el de los cuadros.
CERVEZA_CATS = [*CERVEZA_INDIVIDUALES, MULTICERV, IMPORTADAS]
CATEGORIAS = [*CERVEZA_CATS, "AGUAS DANONE", MULTICCU]

# Colores
AZUL = "1F4E78"
AMARILLO = "FFFF00"      # pestana de la hoja que se carga en la base
VERDE = "C6EFCE"         # celdas que se tipean a mano
GRIS = "F2F2F2"
GRIS_BANDA = "808080"    # banda de los bloques de solo lectura
VERDE_BANDA = "375623"   # banda del bloque que se tipea
NARANJA = "FCE4D6"       # TOTAL GENERAL
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
# Sin decimales: son bultos, y el libro se lee de un pantallazo. El VALOR
# sigue con todos sus decimales — esto es formato de celda, no ROUND, asi la
# suma de las partes cierra exacto contra el cupo tipeado.
FMT = "#,##0"
FMT_PCT = "0.0%"


def _txt(v) -> str:
    if v is None or (isinstance(v, float) and v != v):
        return ""
    s = unicodedata.normalize("NFKD", str(v).strip().upper())
    return "".join(c for c in s if not unicodedata.combining(c))


def meses_atras(hoy: date, n: int) -> list[str]:
    """Los n meses cerrados anteriores a `hoy`, mas viejo primero."""
    y, m = hoy.year, hoy.month
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            y, m = y - 1, 12
        out.append(f"{y}-{m:02d}")
    return list(reversed(out))


def clasificar(generico, marca) -> str | None:
    """marca de gold -> categoria del libro. None = fuera del universo."""
    g, m = _txt(generico), _txt(marca)
    if g == "CERVEZAS":
        if m in CERVEZA_INDIVIDUALES:
            return m
        if m in MARCAS_IMPORTADAS:
            return IMPORTADAS
        return MULTICERV
    if g == "AGUAS DANONE":
        return "AGUAS DANONE"
    if g in GEN_MULTICCU:
        return g          # se agrupa en MULTI CCU, pero guarda su propia historia
    return None


def categoria_padre(cat: str) -> str:
    return MULTICCU if cat in GEN_MULTICCU else cat


def cargar_rutas(dl: DataLoader):
    """(sucursal, preventista) -> [(id_sucursal, id_ruta, des_ruta)]."""
    df = dl.execute_query("""
        SELECT ds.descripcion AS sucursal, dc.id_sucursal,
               dc.id_ruta_fv1 AS ruta,
               MIN(dc.des_ruta_fv1) AS des_ruta,
               UPPER(TRIM(dc.des_personal_fv1)) AS preventista
        FROM gold.dim_cliente dc
        JOIN gold.dim_sucursal ds ON ds.id_sucursal = dc.id_sucursal
        WHERE COALESCE(dc.anulado, false) = false
          AND dc.id_sucursal NOT IN :excl
          AND dc.id_ruta_fv1 IS NOT NULL AND dc.des_personal_fv1 IS NOT NULL
        GROUP BY 1, 2, 3, 5
    """, {"excl": tuple(SUCURSALES_EXCLUIDAS)})
    rutas: dict[tuple[str, str], list] = {}
    for f in df.itertuples(index=False):
        prev = _txt(f.preventista)
        if prev in SIN_CUPO:
            continue
        rutas.setdefault((_txt(f.sucursal), prev), []).append(
            (int(f.id_sucursal), int(f.ruta),
             _txt(f.des_ruta) or f"RUTA {int(f.ruta)}"))
    for v in rutas.values():
        v.sort()
    return rutas


def cargar_historia(dl: DataLoader, meses: list[str]):
    """Ventas por (id_sucursal, id_ruta, categoria, mes). Ruta de dim_cliente."""
    desde = f"{meses[0]}-01"
    y, m = int(meses[-1][:4]), int(meses[-1][5:])
    hasta = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"
    df = dl.execute_query("""
        SELECT dc.id_sucursal, dc.id_ruta_fv1 AS ruta, da.generico, da.marca,
               to_char(fv.fecha_comprobante, 'YYYY-MM') AS mes,
               SUM(fv.cantidades_total) AS qty
        FROM gold.fact_ventas fv
        JOIN gold.dim_cliente dc ON dc.id_cliente = fv.id_cliente
                                AND dc.id_sucursal = fv.id_sucursal
        JOIN gold.dim_articulo da ON da.id_articulo = fv.id_articulo
        WHERE fv.anulado = false
          AND dc.id_sucursal NOT IN :excl
          AND fv.fecha_comprobante >= :d AND fv.fecha_comprobante < :h
          AND dc.id_ruta_fv1 IS NOT NULL
        GROUP BY 1, 2, 3, 4, 5
    """, {"excl": tuple(SUCURSALES_EXCLUIDAS), "d": desde, "h": hasta})
    hist: dict[tuple[int, int, str, str], float] = {}
    for f in df.itertuples(index=False):
        cat = clasificar(f.generico, f.marca)
        if cat is None:
            continue
        k = (int(f.id_sucursal), int(f.ruta), cat, f.mes)
        hist[k] = hist.get(k, 0.0) + float(f.qty or 0)
    return hist


def _banda(ws, r: int, ancho: int, texto: str, color: str):
    """Fila-titulo de un bloque de la matriz."""
    ws.cell(r, 1, texto).font = Font(bold=True, color="FFFFFF")
    for c in range(1, ancho + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=color)
    return r


def _bloque_matriz(ws, r0: int, titulo: str, color_banda: str, sucursales: list[str],
                   cats: list[str], valor=None, editable: bool = False,
                   fmt: str = None, total=None):
    """Un bloque `sucursal x categoria` con su banda de titulo y sus totales.

    `valor(suc, cat)` devuelve lo que va en cada celda — un numero o una formula.
    Devuelve (fila_encabezado, primera_fila_datos, ultima_fila_datos).
    """
    ancho = len(cats) + 2                      # SUCURSAL + categorias + TOTAL
    _banda(ws, r0, ancho, titulo, color_banda)
    rh = r0 + 1
    ws.cell(rh, 1, "SUCURSAL")
    for j, cat in enumerate(cats):
        ws.cell(rh, 2 + j, cat)
    ws.cell(rh, ancho, "TOTAL")
    for c in range(1, ancho + 1):
        cel = ws.cell(rh, c)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
    r = rh + 1
    for suc in sucursales:
        ws.cell(r, 1, suc)
        for j, cat in enumerate(cats):
            cel = ws.cell(r, 2 + j)
            if valor is not None:
                cel.value = valor(suc, cat)
            if editable:
                cel.fill = PatternFill("solid", fgColor=VERDE)
        c0, c1 = get_column_letter(2), get_column_letter(1 + len(cats))
        # `total` existe porque un bloque de ratios no se totaliza sumando:
        # la suma de nueve porcentajes no significa nada.
        ws.cell(r, ancho, total("fila", r, None) if total
                else f"=SUM({c0}{r}:{c1}{r})")
        r += 1
    ultima = r - 1
    ws.cell(r, 1, "TOTAL")
    for c in range(2, ancho + 1):
        L = get_column_letter(c)
        ws.cell(r, c, total("col", r, c) if total
                else f"=SUM({L}{rh + 1}:{L}{ultima})")
    for c in range(1, ancho + 1):
        cel = ws.cell(r, c)
        cel.font = Font(bold=True)
        cel.fill = PatternFill("solid", fgColor=NARANJA)
    for fila in ws.iter_rows(min_row=rh + 1, max_row=r, min_col=1, max_col=ancho):
        for i, cel in enumerate(fila, start=1):
            cel.border = BORDE
            if i >= 2:
                cel.number_format = fmt or FMT
    return rh, rh + 1, ultima


def _encabezado(ws, headers: list[str], anchos: list[int]):
    ws.append(headers)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"


def _cerrar(ws, primera_num: int, filas_total: list[int] = (),
            cols_pct: tuple[int, ...] = ()):
    """Bordes, formato numerico y pintado de las filas de total.

    `cols_pct` va aparte porque este barrido pisa cualquier `number_format`
    puesto celda por celda antes: sin esto las columnas de PESO y % PART
    quedaban con formato de bultos y se leian como 0.
    """
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(fila, start=1):
            c.border = BORDE
            if i in cols_pct:
                c.number_format = FMT_PCT
            elif i >= primera_num:
                c.number_format = FMT
    for r in filas_total:
        for c in ws[r]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=NARANJA)


def construir(rutas, hist, meses):
    wb = Workbook()

    # ------------------------------------------------------------------ Cupos
    # Una fila por sucursal-categoria. Es mas largo que un grid, pero se tipea
    # bajando y las otras hojas apuntan a UN solo rango con SUMIFS.
    ws = wb.active
    ws.title = "Cupos"
    ws.sheet_properties.tabColor = VERDE

    sucursales = sorted({s for s, _ in rutas})
    # (sucursal, categoria, mes) -> bultos, ya agrupado al padre
    venta_suc: dict[tuple[str, str, str], float] = {}
    for (suc, prev), rs in rutas.items():
        for id_suc, ruta, _ in rs:
            for cat in CATEGORIAS + GEN_MULTICCU:
                for mes in meses:
                    q = hist.get((id_suc, ruta, cat, mes), 0.0)
                    if q:
                        k = (suc, categoria_padre(cat), mes)
                        venta_suc[k] = venta_suc.get(k, 0.0) + q

    ancho = len(CATEGORIAS) + 2
    ws.column_dimensions["A"].width = 30
    for j in range(len(CATEGORIAS) + 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13

    r = 1
    # Un bloque por mes: se ve la tendencia sin salir de la hoja donde se tipea.
    for mes in meses:
        _, _, _ = _bloque_matriz(
            ws, r, f"VENTA {mes}", GRIS_BANDA, sucursales, CATEGORIAS,
            valor=lambda suc, cat, _m=mes: venta_suc.get((suc, cat, _m), 0.0))
        r += len(sucursales) + 4
    _, f_v3m0, f_v3m1 = _bloque_matriz(
        ws, r, "VENTA TOTAL 3 MESES", GRIS_BANDA, sucursales, CATEGORIAS,
        valor=lambda suc, cat: sum(venta_suc.get((suc, cat, m), 0.0) for m in meses))
    r += len(sucursales) + 4

    # El bloque que se tipea. Es el unico editable de todo el libro.
    h_cupo, f_cupo0, f_cupo1 = _bloque_matriz(
        ws, r, "CUPO A TIPEAR  (unico bloque editable)", VERDE_BANDA,
        sucursales, CATEGORIAS, editable=True)
    r += len(sucursales) + 4

    # Control de razonabilidad: cuanto pide el cupo contra lo que se vendio.
    # Un 400% avisa de un cero mal tipeado antes de que baje a las rutas.
    col_ini, col_fin = get_column_letter(2), get_column_letter(1 + len(CATEGORIAS))
    ancho_m = len(CATEGORIAS) + 2

    def _ratio_celda(suc, cat):
        L = get_column_letter(2 + CATEGORIAS.index(cat))
        i = sucursales.index(suc)
        return (f"=IFERROR({L}{f_cupo0 + i}/{L}{f_v3m0 + i},\"\")")

    def _ratio_total(clase, r_, c_):
        # El total del bloque es el ratio de los TOTALES, no la suma de ratios.
        if clase == "fila":
            i = r_ - (r + 2)                       # offset dentro del bloque
            L = get_column_letter(ancho_m)
            return f"=IFERROR({L}{f_cupo0 + i}/{L}{f_v3m0 + i},\"\")"
        L = get_column_letter(c_)
        return f"=IFERROR({L}{f_cupo1 + 1}/{L}{f_v3m1 + 1},\"\")"

    _bloque_matriz(
        ws, r, "CUPO / VENTA 3M", GRIS_BANDA, sucursales, CATEGORIAS,
        valor=_ratio_celda, fmt=FMT_PCT, total=_ratio_total)

    ws.freeze_panes = "B2"
    # Coordenadas del bloque de cupo, para el INDEX/MATCH de las otras hojas.
    rango_cupo = (f"Cupos!${col_ini}${f_cupo0}:${col_fin}${f_cupo1}",
                  f"Cupos!$A${f_cupo0}:$A${f_cupo1}",
                  f"Cupos!${col_ini}${h_cupo}:${col_fin}${h_cupo}")
    total_cupos = len(sucursales) * len(CATEGORIAS)

    # ------------------------------------------------------ Cupo Preventista
    wsp = wb.create_sheet("Cupo Preventista")
    # La venta va ABIERTA POR MES, no solo el acumulado: sirve para ver si un
    # preventista viene cayendo o creciendo antes de aceptar el reparto.
    _encabezado(wsp, ["SUCURSAL", "PREVENTISTA", "CATEGORIA", *meses,
                      "TOTAL 3M", "PESO", "OBJETIVO"],
                [30, 26, 16, 12, 12, 12, 13, 10, 14])
    n_mes = len(meses)
    c_m0 = get_column_letter(4)                  # primer mes
    c_m1 = get_column_letter(3 + n_mes)          # ultimo mes
    c_tot = get_column_letter(4 + n_mes)         # TOTAL 3M
    c_peso = get_column_letter(5 + n_mes)
    c_obj = get_column_letter(6 + n_mes)
    fila_prev: dict[tuple[str, str, str], int] = {}
    r = 2
    for (suc, prev) in sorted(rutas):
        for cat in CATEGORIAS:
            por_mes = {mes: 0.0 for mes in meses}
            for id_suc, ruta, _ in rutas[(suc, prev)]:
                for c in (GEN_MULTICCU if cat == MULTICCU else [cat]):
                    for mes in meses:
                        por_mes[mes] += hist.get((id_suc, ruta, c, mes), 0.0)
            wsp.cell(r, 1, suc)
            wsp.cell(r, 2, prev)
            wsp.cell(r, 3, cat)
            for i, mes in enumerate(meses):
                wsp.cell(r, 4 + i, por_mes[mes])
            # Las notas de credito no restan cupo: el peso arranca en cero. Un
            # mes puede quedar negativo, el TOTAL que pondera nunca.
            wsp.cell(r, 4 + n_mes, f"=MAX(0,SUM({c_m0}{r}:{c_m1}{r}))")
            base = f'SUMIFS(${c_tot}:${c_tot},$A:$A,$A{r},$C:$C,$C{r})'
            wsp.cell(r, 5 + n_mes, f'=IFERROR(${c_tot}{r}/{base},0)')
            # La hoja Cupos es una matriz: la celda se ubica cruzando sucursal
            # (filas) con categoria (columnas), no con SUMIFS sobre una columna.
            cupo_ref = (f"IFERROR(INDEX({rango_cupo[0]},"
                        f"MATCH($A{r},{rango_cupo[1]},0),"
                        f"MATCH($C{r},{rango_cupo[2]},0)),0)")
            # Sin historia en toda la sucursal-categoria, reparto parejo.
            wsp.cell(r, 6 + n_mes,
                     f'=IF({base}=0,'
                     f'IFERROR({cupo_ref}/COUNTIFS($A:$A,$A{r},$C:$C,$C{r}),0),'
                     f'${c_peso}{r}*{cupo_ref})')
            fila_prev[(suc, prev, cat)] = r
            r += 1
    wsp.cell(r, 1, "TOTAL GENERAL")
    for col in (*range(4, 4 + n_mes), 4 + n_mes, 6 + n_mes):
        L = get_column_letter(col)
        wsp.cell(r, col, f"=SUM({L}2:{L}{r - 1})")
    _cerrar(wsp, primera_num=4, filas_total=[r], cols_pct=(5 + n_mes,))
    wsp.auto_filter.ref = f"A1:{c_obj}{r - 1}"

    # ------------------------------------------------------------- Cupo Ruta
    wsr = wb.create_sheet("Cupo Ruta")
    # Tambien abierta por mes, igual que la de preventista.
    _encabezado(wsr, ["SUCURSAL", "PREVENTISTA", "CÓDIGO", "RUTA", "GRUPO",
                      "CATEGORIA", *meses, "TOTAL 3M", "PESO", "OBJETIVO"],
                [30, 26, 9, 22, 18, 16, 12, 12, 12, 13, 10, 14])
    r_m0 = get_column_letter(7)                  # primer mes
    r_m1 = get_column_letter(6 + n_mes)          # ultimo mes
    r_tot = get_column_letter(7 + n_mes)         # TOTAL 3M
    r_peso = get_column_letter(8 + n_mes)
    r_obj = get_column_letter(9 + n_mes)
    col_obj_ruta = 9 + n_mes
    filas_ruta: list[tuple[int, int, int, str, str, str]] = []
    r = 2
    for (suc, prev) in sorted(rutas):
        for id_suc, ruta, des_ruta in rutas[(suc, prev)]:
            for cat in CATEGORIAS:
                # MULTI CCU baja abierto en sus tres genericos; el resto va uno a uno.
                for gen in (GEN_MULTICCU if cat == MULTICCU else [cat]):
                    wsr.cell(r, 1, suc)
                    wsr.cell(r, 2, prev)
                    wsr.cell(r, 3, ruta)
                    wsr.cell(r, 4, des_ruta)
                    wsr.cell(r, 5, gen)
                    wsr.cell(r, 6, cat)
                    for i, mes in enumerate(meses):
                        wsr.cell(r, 7 + i, hist.get((id_suc, ruta, gen, mes), 0.0))
                    wsr.cell(r, 7 + n_mes, f"=MAX(0,SUM({r_m0}{r}:{r_m1}{r}))")
                    # El peso es dentro del PREVENTISTA, no de la sucursal: asi la
                    # suma de sus rutas da exacto su objetivo.
                    base = (f'SUMIFS(${r_tot}:${r_tot},$A:$A,$A{r},'
                            f'$B:$B,$B{r},$F:$F,$F{r})')
                    wsr.cell(r, 8 + n_mes, f'=IFERROR(${r_tot}{r}/{base},0)')
                    obj = (f"SUMIFS('Cupo Preventista'!${c_obj}:${c_obj},"
                           f"'Cupo Preventista'!$A:$A,$A{r},"
                           f"'Cupo Preventista'!$B:$B,$B{r},"
                           f"'Cupo Preventista'!$C:$C,$F{r})")
                    wsr.cell(r, col_obj_ruta,
                             f'=IF({base}=0,'
                             f'IFERROR({obj}/COUNTIFS($A:$A,$A{r},$B:$B,$B{r},$F:$F,$F{r}),0),'
                             f'${r_peso}{r}*{obj})')
                    filas_ruta.append((r, id_suc, ruta, des_ruta, gen, cat))
                    r += 1
    wsr.cell(r, 1, "TOTAL GENERAL")
    for col in (*range(7, 7 + n_mes), 7 + n_mes, col_obj_ruta):
        L = get_column_letter(col)
        wsr.cell(r, col, f"=SUM({L}2:{L}{r - 1})")
    _cerrar(wsr, primera_num=7, filas_total=[r], cols_pct=(8 + n_mes,))
    wsr.auto_filter.ref = f"A1:{r_obj}{r - 1}"

    # -------------------------------------------------- Base Pivot SUCURSALES
    # La hoja que carga el ETL. GRUPO = CATEGORIA = la etiqueta, salvo los tres
    # genericos de MULTI CCU, que llevan CATEGORIA=MULTICCU para poder agruparlos.
    # CERVEZAS va como DETALLE y por lo tanto SE CARGA: en gold es el TOTAL y
    # convive con sus marcas.
    wsb = wb.create_sheet("Base Pivot SUCURSALES")
    wsb.sheet_properties.tabColor = AMARILLO
    _encabezado(wsb, ["ZONA", "PREVENTISTA", "CÓDIGO", "RUTA", "NIVEL", "GRUPO",
                      "CATEGORIA", "CUPO"], [30, 26, 9, 22, 10, 18, 18, 14])
    por_ruta: dict[tuple[int, int], list] = {}
    for fr, id_suc, ruta, des_ruta, gen, cat in filas_ruta:
        por_ruta.setdefault((id_suc, ruta), []).append((fr, des_ruta, gen, cat))
    zona = {}
    prevs = {}
    for (suc, prev), rs in rutas.items():
        for id_suc, ruta, _ in rs:
            zona[(id_suc, ruta)] = f"{id_suc} - {suc}"
            prevs[(id_suc, ruta)] = prev
    r = 2
    for clave in sorted(por_ruta):
        items = por_ruta[clave]
        z, prev = zona[clave], prevs[clave]
        des_ruta = items[0][1]
        refs = {gen: fr for fr, _, gen, _ in items}
        # CERVEZAS: el total, suma de sus ocho.
        suma = "+".join(f"'Cupo Ruta'!${r_obj}${refs[c]}" for c in CERVEZA_CATS if c in refs)
        wsb.append([z, prev, clave[1], des_ruta, "DETALLE", "CERVEZAS", "CERVEZAS",
                    f"={suma}" if suma else 0])
        r += 1
        for cat in CERVEZA_CATS:
            if cat in refs:
                wsb.append([z, prev, clave[1], des_ruta, "DETALLE", cat, cat,
                            f"='Cupo Ruta'!${r_obj}${refs[cat]}"])
                r += 1
        if "AGUAS DANONE" in refs:
            wsb.append([z, prev, clave[1], des_ruta, "DETALLE", "AGUAS DANONE",
                        "AGUAS DANONE", f"='Cupo Ruta'!${r_obj}${refs['AGUAS DANONE']}"])
            r += 1
        suma_m = "+".join(f"'Cupo Ruta'!${r_obj}${refs[g]}" for g in GEN_MULTICCU if g in refs)
        if suma_m:
            wsb.append([z, prev, clave[1], des_ruta, "AGREGADO", "TOTAL MULTICCU",
                        "TOTAL MULTICCU", f"={suma_m}"])
            r += 1
            for gen in GEN_MULTICCU:
                if gen in refs:
                    wsb.append([z, prev, clave[1], des_ruta, "DETALLE", gen,
                                "MULTICCU", f"='Cupo Ruta'!${r_obj}${refs[gen]}"])
                    r += 1
    _cerrar(wsb, primera_num=8)
    wsb.auto_filter.ref = f"A1:H{r - 1}"

    return wb, total_cupos, len(filas_ruta)


def main() -> int:
    meses = meses_atras(date.today(), MESES_HISTORIA)
    dl = DataLoader()
    rutas = cargar_rutas(dl)
    hist = cargar_historia(dl, meses)

    sucursales = sorted({s for s, _ in rutas})
    print(f"Historia: {', '.join(meses)}")
    print(f"{len(sucursales)} sucursales | {len(rutas)} preventistas | "
          f"{sum(len(v) for v in rutas.values())} rutas")

    wb, filas_cupos, filas_ruta = construir(rutas, hist, meses)

    # El archivo vive en la carpeta del mes del CUPO (el que se esta armando),
    # no en la del ultimo mes de historia.
    hoy = date.today()
    periodo = f"{hoy.year}-{hoy.month:02d}"
    salida = service_output_dir("cupos", f"{periodo}-01", "month") / \
        f"CALCULADORA CUPOS SUCURSALES - {periodo}.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)
    if salida.exists() and "--force" not in sys.argv:
        print(f"\nYa existe y NO se sobrescribe:\n  {salida}\n\nCorre con --force.")
        return 1
    wb.save(salida)
    print(f"\nGuardado: {salida}")
    print(f"  Cupos             {filas_cupos} celdas para tipear (matriz)")
    print(f"  Cupo Preventista  {len(rutas) * len(CATEGORIAS)} filas")
    print(f"  Cupo Ruta         {filas_ruta} filas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
