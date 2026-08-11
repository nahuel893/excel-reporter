"""Reacomoda una apertura de cupos por ruta al formato de `gold.fact_cupos`.

NO toca ningun numero. Lee los valores tal como estan en el xlsx — incluido el
resultado ya calculado de las formulas que se hayan puesto a mano — y solo
cambia la FORMA: de una tabla ancha (una columna por categoria) a la tabla
larga que usa la base.

Formato destino (`gold.fact_cupos`)
-----------------------------------
    periodo | proveedor | id_sucursal | sucursal | id_ruta | descripcion |
    preventista | generico | desagregado | cupo

Ojo con los nombres, que son contraintuitivos:
- `generico`    es el ITEM  (SALTA, HEINEKEN, VINOS CCU, ...)
- `desagregado` es el GRUPO al que pertenece (CERVEZAS, MULTI CCU, ...)

Y **no se cargan filas de agregado**: no existe una fila con generico='CERVEZAS'
sumando sus marcas. La base guarda solo las hojas; los totales se reconstruyen
agrupando por `desagregado`. Si se cargaran los agregados, cualquier SUM(cupo)
contaria doble.

Uso
---
    python scripts/cupos_a_formato_bd.py <archivo.xlsx> --periodo 2026-08 \\
        --proveedor CCU [--hoja "Cupo Ruta"] [--salida <archivo.csv>]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# generico -> desagregado, tal como esta cargado en gold.fact_cupos.
# Las claves son los encabezados que usa la apertura; el valor es el par
# (generico_bd, desagregado_bd). None en desagregado = columna NULL.
MAPEO_CCU: dict[str, tuple[str, str | None]] = {
    # CERVEZAS es el total y va como fila propia: la base guarda el agregado.
    "CERVEZAS": ("CERVEZAS", "CERVEZAS"),
    "SALTA": ("SALTA", "SALTA"),
    "HEINEKEN": ("HEINEKEN", "HEINEKEN"),
    "IMPERIAL": ("IMPERIAL", "IMPERIAL"),
    "MILLER": ("MILLER", "MILLER"),
    # Julio los tiene abiertos por marca; esta apertura los calcula como un
    # solo numero, asi que van con el nombre de la categoria en las dos.
    "MULTICERVEZAS": ("MULTICERVEZAS", "MULTICERVEZAS"),
    "IMPORTADAS": ("IMPORTADAS", "IMPORTADAS"),
    # En la base es AGUAS DANONE (con S) y MULTICCU va junto, sin espacio.
    "AGUA DANONE": ("AGUAS DANONE", "AGUAS DANONE"),
    "VINOS CCU": ("VINOS CCU", "MULTICCU"),
    "SIDRAS Y LICORES": ("SIDRAS Y LICORES", "MULTICCU"),
    "PERNOD RICARD": ("PERNOD RICARD", "MULTICCU"),
}

# BRANCA se carga a nivel marca y con desagregado NULL.
MAPEO_BRANCA: dict[str, tuple[str, str | None]] = {
    "BRANCA": ("FRATELLI BRANCA", None),
    "ARIZU": ("ARIZU", None),
    "VINOS FINOS": ("VINOS FINOS", None),
    "QUARA": ("QUARA", None),
}

MAPEOS = {"CCU": MAPEO_CCU, "BRANCA": MAPEO_BRANCA}

# Columnas que NO son categorias: identifican la ruta.
COL_PREVENTISTA, COL_CODIGO, COL_RUTA = "PREVENTISTA", "CÓDIGO", "RUTA"

CAMPOS = ["periodo", "proveedor", "id_sucursal", "sucursal", "id_ruta",
          "descripcion", "preventista", "generico", "desagregado", "cupo"]


def localizar_encabezados(ws) -> tuple[int, dict[str, list[int]]]:
    """Encuentra la fila de encabezados y TODAS las columnas de cada nombre.

    No se asume que sea la fila 1: al editar el archivo a mano es comun
    insertar una fila arriba para anotar. Se busca la primera fila que tenga
    a la vez PREVENTISTA y CÓDIGO.

    Devuelve una LISTA de columnas por nombre, no una sola, porque los
    encabezados se repiten cuando se agrega un bloque corregido al lado del
    viejo. Quedarse en silencio con una de las dos es como se cargan cupos
    equivocados sin que nadie lo note.
    """
    for fila in range(1, min(ws.max_row, 10) + 1):
        encabezados: dict[str, list[int]] = {}
        for c in range(1, ws.max_column + 1):
            valor = ws.cell(row=fila, column=c).value
            if valor is not None and str(valor).strip():
                encabezados.setdefault(str(valor).strip(), []).append(c)
        if COL_PREVENTISTA in encabezados and COL_CODIGO in encabezados:
            return fila, encabezados
    raise ValueError(
        f"No encontre la fila de encabezados en la hoja {ws.title!r}: "
        f"tiene que haber una fila con {COL_PREVENTISTA!r} y {COL_CODIGO!r}."
    )


def resolver_columna(nombre: str, columnas: list[int], usar: str) -> int:
    """Elige la columna cuando el encabezado esta repetido.

    `usar` decide si vale la primera o la ultima aparicion. Se avisa siempre:
    una eleccion silenciosa entre dos bloques con el mismo nombre es
    exactamente el error que carga el cupo viejo como si fuera el nuevo.
    """
    if len(columnas) == 1:
        return columnas[0]
    letras = ", ".join(get_column_letter(c) for c in columnas)
    elegida = columnas[-1] if usar == "ultima" else columnas[0]
    print(f"  AVISO: el encabezado {nombre!r} aparece en {len(columnas)} "
          f"columnas ({letras}). Tomo la {usar}: "
          f"{get_column_letter(elegida)}.")
    return elegida


def convertir(path: Path, hoja: str, periodo: str, proveedor: str,
              id_sucursal: int, sucursal: str,
              usar_columna: str = "ultima") -> list[dict]:
    """Devuelve las filas en formato fact_cupos. Los cupos van tal cual."""
    mapeo = MAPEOS[proveedor]
    # data_only=True devuelve el valor YA CALCULADO de las formulas, no la
    # formula. Es lo que permite copiar los ajustes hechos a mano sin
    # recalcular nada.
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if hoja not in wb.sheetnames:
            raise ValueError(f"La hoja {hoja!r} no existe. Hay: {wb.sheetnames}")
        ws = wb[hoja]
        fila_hdr, encabezados = localizar_encabezados(ws)

        faltantes = [c for c in mapeo if c not in encabezados]
        if faltantes:
            raise ValueError(
                f"Faltan columnas de {proveedor} en la hoja: {faltantes}. "
                f"Encontre: {[h for h in encabezados if h]}")

        # Resolver una sola vez que columna se usa para cada nombre.
        col = {n: resolver_columna(n, encabezados[n], usar_columna)
               for n in (*mapeo, COL_CODIGO, COL_PREVENTISTA, COL_RUTA)}

        filas: list[dict] = []
        sin_valor: list[str] = []
        for r in range(fila_hdr + 1, ws.max_row + 1):
            codigo = ws.cell(row=r, column=col[COL_CODIGO]).value
            if not isinstance(codigo, (int, float)):
                continue          # subtotales, TOTAL GENERAL, filas vacias
            preventista = ws.cell(row=r, column=col[COL_PREVENTISTA]).value
            descripcion = ws.cell(row=r, column=col[COL_RUTA]).value
            for columna, (generico, desagregado) in mapeo.items():
                cupo = ws.cell(row=r, column=col[columna]).value
                if cupo is None:
                    # Formula sin valor cacheado: el archivo nunca se abrio en
                    # Excel/WPS despues de escribirla. Inventar un 0 aca seria
                    # cargar un cupo falso.
                    sin_valor.append(f"ruta {int(codigo)} / {columna}")
                    continue
                filas.append({
                    "periodo": periodo, "proveedor": proveedor,
                    "id_sucursal": id_sucursal, "sucursal": sucursal,
                    "id_ruta": int(codigo),
                    "descripcion": str(descripcion or "").strip(),
                    "preventista": str(preventista or "").strip(),
                    "generico": generico, "desagregado": desagregado,
                    "cupo": cupo,
                })
        if sin_valor:
            raise ValueError(
                f"{len(sin_valor)} celdas son formulas sin valor calculado "
                f"(ej: {sin_valor[:3]}). Abri el archivo en Excel/WPS, guardalo "
                f"para que queden los valores, y volve a correr.")
        return filas
    finally:
        wb.close()


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Reacomoda una apertura de cupos al formato de gold.fact_cupos.")
    p.add_argument("archivo", type=Path, help="xlsx de la apertura por ruta.")
    p.add_argument("--periodo", required=True, help="YYYY-MM (ej. 2026-08).")
    p.add_argument("--proveedor", required=True, choices=sorted(MAPEOS))
    p.add_argument("--hoja", default="Cupo Ruta")
    p.add_argument("--id-sucursal", type=int, default=1)
    p.add_argument("--sucursal", default="1 - CASA CENTRAL")
    p.add_argument("--usar-columna", choices=("primera", "ultima"),
                   default="ultima",
                   help="Con encabezados repetidos, cual vale. Default: ultima.")
    p.add_argument("--salida", type=Path, default=None,
                   help="csv de salida. Default: <archivo> - FORMATO BD.csv")
    return p.parse_args(argv)


def main(argv=None) -> int:
    import csv

    args = parse_args(argv)
    if not args.archivo.exists():
        print(f"Error: no existe {args.archivo}")
        return 1

    try:
        filas = convertir(args.archivo, args.hoja, args.periodo,
                          args.proveedor, args.id_sucursal, args.sucursal,
                          args.usar_columna)
    except ValueError as exc:
        print(f"Error: {exc}")
        return 1

    salida = args.salida or args.archivo.with_name(
        f"{args.archivo.stem} - FORMATO BD.csv")
    with open(salida, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=CAMPOS)
        w.writeheader()
        w.writerows(filas)

    rutas = len({f["id_ruta"] for f in filas})
    print(f"Leido:   {args.archivo.name}  (hoja {args.hoja!r})")
    print(f"Salida:  {salida}")
    print(f"Filas:   {len(filas)}  =  {rutas} rutas x {len(MAPEOS[args.proveedor])} genericos")
    print("\nTotal por generico (copiado tal cual, sin redondear):")
    por_gen: dict[str, float] = {}
    for f in filas:
        por_gen[f["generico"]] = por_gen.get(f["generico"], 0.0) + float(f["cupo"])
    for gen, tot in sorted(por_gen.items()):
        print(f"   {gen:<18} {tot:>14,.4f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
