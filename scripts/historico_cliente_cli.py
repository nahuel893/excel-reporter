#!/usr/bin/env python
"""Generate the historico-cliente report from one command.

Exists so a caller — an agent, a cron job, a person in a hurry — can ask for
this report without hand-writing a JSON config. Every setting that is wrong
*silently* when composed by hand is decided here instead:

  - the CCU marca universe (get it wrong and the coverage gaps vanish),
  - the composite client key (id_cliente alone is not unique across sucursales),
  - the capture range (a stale hardcoded range crops the TOTAL GENERAL row out
    of the PNG while the xlsx stays correct and nothing raises).

Prints a JSON object on stdout so the caller can parse the result:

    {"ok": true, "cliente": "...", "id_cliente": 7255, "id_sucursal": 1,
     "xlsx": "/abs/path.xlsx", "png": "/abs/path.png",
     "rango": "A1:P47", "total_general": 75.41666667,
     "desde": "2025-09-01", "hasta": "2026-08-11", "solo_con_cargo": true}

On failure it prints {"ok": false, "error": "..."} and exits non-zero.

Usage:
    python scripts/historico_cliente_cli.py --cliente 7255 [--sucursal 1]
        [--meses 12] [--solo-con-cargo] [--sin-imagen]
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.services.historico_cliente.service import (  # noqa: E402
    HistoricoClienteConfig,
    columnas_desbordadas,
    rango_captura,
)

# The four CCU genericos whose full marca set defines the coverage universe.
# Marcas the client never bought show as 0, which is the point of the report.
GENERICOS_UNIVERSO = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES"]

_DPI = 300


class ClienteNoEncontrado(Exception):
    """No row in gold.dim_cliente for that id."""


class ClienteAmbiguo(Exception):
    """The id exists in more than one sucursal; the caller must pick one."""


class CapturaInvalida(Exception):
    """The sheet would render with ### in at least one column."""


# ── Ventana ──────────────────────────────────────────────────────────────────

def ventana(
    meses: int | None = None,
    *,
    desde: str | None = None,
    hasta: str | None = None,
    anios: list[int] | None = None,
    hoy: date | None = None,
) -> tuple[str, str]:
    """Return (desde, hasta) for the requested window.

    Three ways to ask, in order of precedence:

    - ``desde`` / ``hasta`` explicit;
    - ``anios``: whole calendar years, from January 1st of the earliest;
    - ``meses``: the last N months, current one included and partial.

    ``hasta`` never runs past today: projecting an in-progress year to
    December 31st would show empty months as if they were zero sales.
    """
    hoy = hoy or date.today()

    if desde and hasta:
        return desde, hasta

    if anios:
        primero, ultimo = min(anios), max(anios)
        fin = date(ultimo, 12, 31)
        return (
            date(primero, 1, 1).isoformat(),
            min(fin, hoy).isoformat(),
        )

    total = (hoy.year * 12 + hoy.month - 1) - ((meses or 12) - 1)
    anio, mes = divmod(total, 12)
    return f"{anio:04d}-{mes + 1:02d}-01", hoy.isoformat()


def parse_cliente(txt: str) -> tuple[int, int | None]:
    """Parse ``"30158"`` or ``"30158:4"`` into (id_cliente, id_sucursal|None).

    The suffix pins the sucursal for that code only, so one ambiguous client
    in a batch does not force the caller to pin every other one.
    """
    txt = str(txt).strip()
    codigo, _, suc = txt.partition(":")
    if not codigo.isdigit() or (suc and not suc.isdigit()):
        raise ValueError(
            f"Código de cliente inválido: {txt!r}. Se espera 'ID' o 'ID:SUCURSAL'."
        )
    return int(codigo), int(suc) if suc else None


# ── Cliente ──────────────────────────────────────────────────────────────────

def resolver_cliente(
    loader, id_cliente: int, id_sucursal: int | None = None
) -> tuple[int, int, str]:
    """Resolve (id_cliente, id_sucursal, nombre) against gold.dim_cliente.

    ``id_cliente`` is NOT unique — the same code is reused across sucursales.
    With several matches and no ``id_sucursal`` this raises instead of picking
    one, because guessing silently reports another branch's client.

    ``fantasia`` may be an empty string rather than NULL, so plain COALESCE is
    not enough to fall back to razon_social.
    """
    filas = loader.execute_query(
        """
        SELECT id_cliente, id_sucursal,
               COALESCE(
                   NULLIF(TRIM(fantasia), ''),
                   NULLIF(TRIM(razon_social), ''),
                   CAST(id_cliente AS TEXT)
               ) AS nombre
        FROM gold.dim_cliente
        WHERE id_cliente = :id_cliente
        ORDER BY id_sucursal
        """,
        {"id_cliente": id_cliente},
    )
    if filas is None or filas.empty:
        raise ClienteNoEncontrado(f"No existe el cliente {id_cliente} en dim_cliente.")

    if id_sucursal is not None:
        filas = filas[filas["id_sucursal"] == id_sucursal]
        if filas.empty:
            raise ClienteNoEncontrado(
                f"El cliente {id_cliente} no existe en la sucursal {id_sucursal}."
            )

    if len(filas) > 1:
        opciones = ", ".join(
            f"{int(r.id_sucursal)} ({r.nombre})" for r in filas.itertuples()
        )
        raise ClienteAmbiguo(
            f"El cliente {id_cliente} existe en varias sucursales: {opciones}. "
            f"Volvé a pedirlo con --sucursal."
        )

    fila = filas.iloc[0]
    return int(fila["id_cliente"]), int(fila["id_sucursal"]), str(fila["nombre"])


# ── Config ───────────────────────────────────────────────────────────────────

def construir_config(
    *, clientes: list[tuple[int, int]], desde: str, hasta: str,
    solo_con_cargo: bool, nombre: str,
) -> HistoricoClienteConfig:
    """Build the config with the settings this report only works with.

    ``clientes`` is an ordered list of (id_cliente, id_sucursal); the service
    writes one sheet per entry, in this order.
    """
    return HistoricoClienteConfig(
        fecha_desde=desde,
        fecha_hasta=hasta,
        clientes=[
            {"id_cliente": c, "id_sucursal": s} for c, s in clientes
        ],
        agrupar_por_generico=True,
        marcas_completas=True,
        genericos_universo=list(GENERICOS_UNIVERSO),
        solo_con_cargo=solo_con_cargo,
        nombre_archivo=nombre,
    )


# ── Captura ──────────────────────────────────────────────────────────────────

def rango_de(ws) -> str:
    """Capture range for ``ws``, always derived from the sheet."""
    return rango_captura(ws)


def validar_captura(ws) -> None:
    """Abort if any column would render as ### instead of its number."""
    fuera = columnas_desbordadas(ws)
    if fuera:
        raise CapturaInvalida(
            "Estas columnas no entran y LibreOffice las renderiza como ###: "
            + ", ".join(fuera)
        )


# ── Orquestación ─────────────────────────────────────────────────────────────

def _total_general(ws) -> float | None:
    """Value of the Total column on the TOTAL GENERAL row, if present."""
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=2).value == "TOTAL GENERAL":
            return ws.cell(row=r, column=ws.max_column).value
    return None


def generar(
    *, clientes: list[str | tuple[int, int | None]],
    meses: int | None = None, desde: str | None = None, hasta: str | None = None,
    anios: list[int] | None = None,
    solo_con_cargo: bool = False, con_imagen: bool = True,
    nombre_archivo: str | None = None, hoy: date | None = None,
) -> dict:
    """Resolve every client, generate one workbook with a sheet each, render.

    A single workbook — not one file per client — so the batch travels as one
    attachment and the sheets can be compared side by side.
    """
    from openpyxl import load_workbook

    from src.core.data_loader import DataLoader
    from src.services.historico_cliente.service import HistoricoClienteService

    loader = DataLoader()

    resueltos: list[tuple[int, int, str]] = []
    for entrada in clientes:
        cod, suc = parse_cliente(entrada) if isinstance(entrada, str) else entrada
        resueltos.append(resolver_cliente(loader, cod, suc))

    desde, hasta = ventana(meses, desde=desde, hasta=hasta, anios=anios, hoy=hoy)
    sufijo = " (con cargo)" if solo_con_cargo else ""

    if nombre_archivo is None:
        if len(resueltos) == 1:
            nombre_archivo = f"Historico Ventas - {resueltos[0][2]} {resueltos[0][0]}{sufijo}"
        else:
            nombre_archivo = (
                f"Historico Ventas - {len(resueltos)} clientes "
                f"{desde[:7]} a {hasta[:7]}{sufijo}"
            )

    config = construir_config(
        clientes=[(c, s) for c, s, _ in resueltos],
        desde=desde, hasta=hasta,
        solo_con_cargo=solo_con_cargo, nombre=nombre_archivo,
    )

    resultado = HistoricoClienteService(data_loader=loader).generar_reporte(config)
    if not resultado.sheets_generated:
        raise ClienteNoEncontrado(
            f"Ningún cliente pedido tiene ventas entre {desde} y {hasta}."
        )

    wb = load_workbook(resultado.ruta_archivo)
    renderer = None
    if con_imagen:
        from src.core.excel_renderers import get_renderer

        renderer = get_renderer("libreoffice")

    hojas: list[dict] = []
    for hoja in resultado.sheets_generated:
        ws = wb[hoja]
        validar_captura(ws)
        rango = rango_de(ws)
        png = None
        if renderer is not None:
            png = renderer.render(
                resultado.ruta_archivo, hoja, rango,
                resultado.ruta_archivo.parent, dpi=_DPI, crop=True,
            )
        hojas.append({
            "hoja": hoja,
            "rango": rango,
            "png": str(png) if png else None,
            "total_general": _total_general(ws),
        })

    # Un cliente pedido sin ventas en la ventana no genera hoja: nombrarlo, en
    # vez de dejar que el faltante pase inadvertido en un lote grande.
    sin_datos = [
        f"{nom} ({cod})" for cod, _, nom in resueltos
        if nom[:31] not in resultado.sheets_generated
    ]

    return {
        "ok": True,
        "xlsx": str(resultado.ruta_archivo),
        "hojas": hojas,
        "sin_datos": sin_datos,
        "clientes": [
            {"id_cliente": c, "id_sucursal": s, "nombre": n} for c, s, n in resueltos
        ],
        "desde": desde,
        "hasta": hasta,
        "solo_con_cargo": solo_con_cargo,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Genera el histórico de compras por marca de un cliente."
    )
    parser.add_argument(
        "--cliente", nargs="+", required=True, metavar="ID[:SUC]",
        help="Uno o más códigos. Una hoja por código. Sufijo ':N' fija la sucursal.",
    )
    parser.add_argument("--meses", type=int, default=None, help="Meses hacia atrás (default 12)")
    parser.add_argument("--desde", default=None, help="Fecha inicio YYYY-MM-DD")
    parser.add_argument("--hasta", default=None, help="Fecha fin YYYY-MM-DD")
    parser.add_argument(
        "--anios", nargs="+", type=int, default=None, metavar="AAAA",
        help="Años calendario completos, ej: --anios 2024 2025 2026",
    )
    parser.add_argument("--nombre", default=None, help="Nombre del archivo de salida")
    parser.add_argument(
        "--solo-con-cargo", action="store_true",
        help="Contar solo unidades facturadas; excluye bonificación 100%%.",
    )
    parser.add_argument(
        "--sin-imagen", action="store_true",
        help="No renderizar el PNG (más rápido; el render tarda 30-60s).",
    )
    args = parser.parse_args(argv)

    try:
        salida = generar(
            clientes=args.cliente,
            meses=args.meses,
            desde=args.desde,
            hasta=args.hasta,
            anios=args.anios,
            solo_con_cargo=args.solo_con_cargo,
            con_imagen=not args.sin_imagen,
            nombre_archivo=args.nombre,
        )
    except (ClienteNoEncontrado, ClienteAmbiguo, CapturaInvalida, ValueError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False))
        return 1
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": f"{type(exc).__name__}: {exc}"}, ensure_ascii=False))
        return 1

    print(json.dumps(salida, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
