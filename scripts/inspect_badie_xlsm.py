"""
One-shot inspection script for AVANCE BADIE *.xlsm.

Do NOT run in CI. Execute manually to determine sheet names and header columns
before populating SHEET_CONFIGS_BADIE in src/services/avances/service.py.

Usage:
    python scripts/inspect_badie_xlsm.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not found — activate the project venv first.")

XLSM_PATH = Path("/home/nahuel/VM shared/AVANCE BADIE MAYO 2026.xlsm")
TARGET_SHEETS = ["aexcel", "pivot_python", "cober_gen", "cober_marca"]


def inspect_sheet(ws, sheet_name: str) -> None:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name!r}")
    print(f"{'='*60}")

    # Detect headers — try row 1 and row 2
    headers_row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers_row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    # Count non-None headers in each row
    row1_count = sum(1 for h in headers_row1 if h is not None)
    row2_count = sum(1 for h in headers_row2 if h is not None)

    detected_header_row = 1 if row1_count >= row2_count else 2
    headers = headers_row1 if detected_header_row == 1 else headers_row2

    print(f"Max column (ws.max_column):  {ws.max_column}")
    print(f"Max row    (ws.max_row):     {ws.max_row}")
    print(f"Detected header row:         {detected_header_row}")
    print(f"Non-None headers:            {sum(1 for h in headers if h is not None)}")
    print()
    print("Column mapping (letter → header name):")
    for col_idx, h in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        print(f"  {letter:3s}  col {col_idx:3d}  →  {h!r}")

    print()
    print("First 3 data rows:")
    data_start = detected_header_row + 1
    for row_idx in range(data_start, min(data_start + 3, ws.max_row + 1)):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        print(f"  Row {row_idx}: {row_values}")


def main() -> None:
    print(f"Opening: {XLSM_PATH}")
    if not XLSM_PATH.exists():
        sys.exit(f"ERROR: File not found — {XLSM_PATH}")

    # read_only=True + keep_vba=False avoids executing macros
    wb = load_workbook(str(XLSM_PATH), read_only=True, keep_vba=False, data_only=True)

    print(f"\nAll sheet names ({len(wb.sheetnames)}):")
    for name in wb.sheetnames:
        print(f"  - {name!r}")

    for sheet_name in TARGET_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            inspect_sheet(ws, sheet_name)
        else:
            print(f"\n[MISSING] Sheet {sheet_name!r} not found in workbook")

    wb.close()
    print("\n\nDone.")


if __name__ == "__main__":
    main()
