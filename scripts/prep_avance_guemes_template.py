"""One-time prep of the AVANCE GUEMES base template.

Workflow: the user saves the Avance Badie Guemes .xlsm as
``data/input/avances/AVANCE GUEMES.xlsx`` from Excel (for pivot/formula
fidelity). This script then applies the two HEADER-ONLY edits the GUEMES loader
requires. It NEVER inserts, deletes, or moves a column, so the workbook's ~138
cross-sheet formulas (which hardcode absolute column letters) keep resolving:

1. CuposVolumen: rename the duplicate legend header at col P ("GENERICO") to
   "GENERICO_LEGEND". Otherwise replace_sheet_data's last-wins col_map maps
   "GENERICO" to col P (the legend) and writes data there, leaving the real
   data col C empty (silent corruption).
2. Create the CuposCoberGen sheet (absent in the GUEMES template) with the badie
   headers: blank spacer col A + Ruta/Preventista/Generico/ZONA/"CUPO " at B..F.

Idempotent: safe to re-run. Backs up the file first (data/ is gitignored).

Usage:
    python scripts/prep_avance_guemes_template.py "data/input/avances/AVANCE GUEMES.xlsx"
"""
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

CUPOSVOL = "CuposVolumen"
COBERGEN = "CuposCoberGen"
# badie CuposCoberGen layout: blank spacer col A, data at B..F
COBERGEN_HEADERS = {2: "Ruta", 3: "Preventista", 4: "Generico", 5: "ZONA", 6: "CUPO "}
LEGEND_COL = 16  # col P — the static GENERICO legend to disambiguate
DATA_COL = 3     # col C — the real GENERICO data column


def main(path_str: str) -> int:
    path = Path(path_str)
    if not path.exists():
        print(f"ERROR: no existe {path}\n"
              f"Primero guardá el .xlsm como este .xlsx desde Excel (Guardar como).")
        return 1

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = path.with_name(f"{path.stem}.prebak-{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    print(f"Backup: {backup.name}")

    wb = load_workbook(str(path))

    if CUPOSVOL not in wb.sheetnames:
        print(f"ERROR: falta la hoja {CUPOSVOL}")
        return 1
    ws = wb[CUPOSVOL]

    # Guard: the data col C must be the GENERICO data column
    col_c = ws.cell(row=1, column=DATA_COL).value
    if col_c != "GENERICO":
        print(f"ABORT: {CUPOSVOL}!C1 esperaba 'GENERICO' (col de datos), "
              f"encontró {col_c!r}. Revisar el template antes de continuar.")
        return 1

    # 1. Rename the legend header at col P
    cur = ws.cell(row=1, column=LEGEND_COL).value
    if cur == "GENERICO":
        ws.cell(row=1, column=LEGEND_COL, value="GENERICO_LEGEND")
        print(f"{CUPOSVOL}!P1: 'GENERICO' -> 'GENERICO_LEGEND'")
    elif cur == "GENERICO_LEGEND":
        print(f"{CUPOSVOL}!P1 ya es 'GENERICO_LEGEND' (idempotente)")
    else:
        print(f"ABORT: {CUPOSVOL}!P1 esperaba 'GENERICO', encontró {cur!r}. "
              f"El template no tiene la estructura esperada.")
        return 1

    # 2. Create CuposCoberGen if absent
    if COBERGEN in wb.sheetnames:
        print(f"{COBERGEN} ya existe — no se recrea")
    else:
        ws2 = wb.create_sheet(COBERGEN)
        for col, header in COBERGEN_HEADERS.items():
            ws2.cell(row=1, column=col, value=header)
        print(f"Creada hoja {COBERGEN} con headers Ruta/Preventista/Generico/ZONA/'CUPO ' en B..F")

    wb.save(str(path))
    print(f"Guardado: {path}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: python scripts/prep_avance_guemes_template.py "<AVANCE GUEMES.xlsx>"')
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
