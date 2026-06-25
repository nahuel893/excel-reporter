#!/usr/bin/env python3
"""Re-match the Coca-Cola provider article file against the SURIA database and
(re)freeze configs/stock_suria_articulos.json.

Run this when the provider sends an UPDATED active-articles file. The daily
``stock-suria`` report reads the frozen JSON (it does NOT re-match on its own),
so after running this the next report run reflects the new article set while the
stock itself refreshes daily.

Matching strategy (deterministic candidates + description validation):
  For each provider code C (column A, marked active with 'x' in column C):
    candidates = [ int("40"+C), int("400"+C), int(C) ]   # 3 schemes: 40 / 400 / bare
  Keep the candidates that exist in gold.dim_articulo, then pick the one whose
  description best matches the provider description (token-overlap / Jaccard).
  Accept if overlap > 0 OR the SURIA description contains "DUAL" (dualpacks have
  0 token overlap but are valid). This disambiguates collisions (e.g. provider
  6606 "Rabieta": 40+6606=AYBAL collision vs bare 6606=RABIETA wins by description).

Usage:
    python scripts/rematch_stock_suria.py                 # re-freeze from the default provider file
    python scripts/rematch_stock_suria.py --provider PATH # use a different provider xlsx
    python scripts/rematch_stock_suria.py --dry-run       # compute + verify, do NOT overwrite the config
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_PROVIDER = Path("/home/nahuel/VM shared/archivos_diarios/articulos-coca/articulos_coca.xlsx")
CONFIG_PATH = ROOT / "configs" / "stock_suria_articulos.json"


def _suria_engine():
    """SQLAlchemy engine for the SURIA database (medallion_db_suria)."""
    try:
        from dotenv import load_dotenv
        env = ROOT / ".env"
        if env.exists():
            load_dotenv(str(env), override=False)
    except ImportError:
        pass
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    host = os.getenv("DB_HOST", "localhost")
    port = os.getenv("DB_PORT", "5432")
    db = os.getenv("DB_NAME_SURIA")
    if not all([user, password, db]):
        raise SystemExit("ERROR: DB_USER / DB_PASSWORD / DB_NAME_SURIA must be set (.env).")
    return create_engine(
        f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{db}",
        connect_args={"connect_timeout": 30},
    )


def _tokens(s: str) -> set[str]:
    """Significant tokens of a description (uppercase, Ñ→N, alphanumeric, len>=2)."""
    s = re.sub(r"[^A-Z0-9]", " ", (s or "").upper().replace("Ñ", "N"))
    return {w for w in s.split() if len(w) >= 2}


def _candidates(code: int) -> list[tuple[int, str]]:
    """The 3 deterministic id_articulo candidates for a provider code."""
    s = str(code)
    return [(int("40" + s), "40"), (int("400" + s), "400"), (int(s), "pelado")]


def _read_active(provider_path: Path) -> list[tuple[int, str]]:
    """Active provider articles: (cod_basis, descripcion) where column C == 'x'."""
    ws = load_workbook(provider_path, data_only=True)["Hoja1"]
    out = []
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        desc = ws.cell(row=r, column=2).value
        x = ws.cell(row=r, column=3).value
        if cod is not None and x and str(x).strip().lower() == "x":
            out.append((int(cod), str(desc or "").strip()))
    return out


def rematch(provider_path: Path) -> dict:
    """Recompute the provider→SURIA match and build the frozen config dict."""
    active = _read_active(provider_path)
    eng = _suria_engine()
    with eng.connect() as con:
        all_ids = {i for c, _ in active for i, _ in _candidates(c)}
        rows = con.execute(
            text(
                "SELECT id_articulo, des_articulo, marca, generico "
                "FROM gold.dim_articulo WHERE id_articulo = ANY(:ids)"
            ),
            {"ids": list(all_ids)},
        ).fetchall()
        sur = {r.id_articulo: r for r in rows}

        matched, unmatched = [], []
        for c, d in active:
            opts = []
            for i, scheme in _candidates(c):
                if i in sur:
                    sd = sur[i].des_articulo or ""
                    ov = (
                        len(_tokens(d) & _tokens(sd)) / len(_tokens(d) | _tokens(sd))
                        if (_tokens(d) | _tokens(sd))
                        else 0
                    )
                    opts.append((ov, i, scheme, sd))
            if not opts:
                unmatched.append({"cod_prov": c, "desc_prov": d})
                continue
            opts.sort(reverse=True)
            ov, i, scheme, sd = opts[0]
            if ov > 0 or "DUAL" in sd.upper():
                matched.append(
                    {
                        "cod_prov": c,
                        "id_articulo": i,
                        "esquema": scheme,
                        "desc_prov": d,
                        "desc_suria": sur[i].des_articulo,
                        "marca": sur[i].marca,
                    }
                )
            else:
                unmatched.append({"cod_prov": c, "desc_prov": d})

        # Enrich each unmatched with its closest SURIA article (any scheme/brand).
        universe = [
            (a.id_articulo, a.des_articulo, _tokens(a.des_articulo))
            for a in con.execute(
                text("SELECT id_articulo, des_articulo FROM gold.dim_articulo WHERE des_articulo IS NOT NULL")
            ).fetchall()
        ]
        for sm in unmatched:
            T = _tokens(sm["desc_prov"])
            best, best_score = None, 0.0
            for iid, des, ts in universe:
                j = len(T & ts) / len(T | ts) if (T | ts) else 0
                if j > best_score:
                    best_score, best = j, (iid, des)
            sm["closest_id"] = best[0] if best else None
            sm["closest_desc"] = best[1] if best else None
            sm["closest_sim"] = round(best_score, 4)

    from collections import Counter

    by_scheme = Counter(m["esquema"] for m in matched)
    return {
        "generado": date.today().isoformat(),
        "archivo_proveedor": provider_path.name,
        "db": os.getenv("DB_NAME_SURIA"),
        "metodo": "candidatos 40/400/pelado + validacion por descripcion",
        "resumen": {
            "total_activos": len(active),
            "matched": len(matched),
            "sin_match": len(unmatched),
            "por_esquema": {k: by_scheme.get(k, 0) for k in ("40", "400", "pelado")},
        },
        "articulos": sorted(matched, key=lambda m: m["cod_prov"]),
        "sin_match": sorted(unmatched, key=lambda m: m["cod_prov"]),
    }


def verify(cfg: dict) -> bool:
    """Self-check the frozen config against the live SURIA DB. Returns True if OK."""
    total = cfg["resumen"]["total_activos"]
    n_match = len(cfg["articulos"])
    n_sin = len(cfg["sin_match"])
    c1 = (n_match + n_sin == total)
    ids = [a["id_articulo"] for a in cfg["articulos"]]
    c2_dups = (len(set(ids)) == len(ids))
    eng = _suria_engine()
    with eng.connect() as con:
        live = {
            r.id_articulo: r.des_articulo
            for r in con.execute(
                text("SELECT id_articulo, des_articulo FROM gold.dim_articulo WHERE id_articulo = ANY(:i)"),
                {"i": ids},
            ).fetchall()
        }
    c3 = all(a["id_articulo"] in live and live[a["id_articulo"]] == a["desc_suria"] for a in cfg["articulos"])
    print(f"  conteos {n_match}+{n_sin}={total}: {'OK' if c1 else 'FAIL'}")
    print(f"  sin id duplicados: {'OK' if c2_dups else 'FAIL'}")
    print(f"  cada id existe en SURIA y descripcion coincide: {'OK' if c3 else 'FAIL'}")
    return c1 and c2_dups and c3


def main() -> int:
    ap = argparse.ArgumentParser(description="Re-match the provider file and re-freeze the SURIA article list.")
    ap.add_argument("--provider", type=Path, default=DEFAULT_PROVIDER, help="provider xlsx (col A=cod, col C='x'=active)")
    ap.add_argument("--dry-run", action="store_true", help="compute + verify but do NOT overwrite the config")
    args = ap.parse_args()

    if not args.provider.exists():
        raise SystemExit(f"ERROR: provider file not found: {args.provider}")

    print(f"Re-matching from: {args.provider}")
    cfg = rematch(args.provider)
    r = cfg["resumen"]
    print(f"  matched={r['matched']} sin_match={r['sin_match']} total={r['total_activos']} | por_esquema={r['por_esquema']}")

    print("Verifying...")
    ok = verify(cfg)
    if not ok:
        raise SystemExit("ERROR: verification FAILED — config not written.")

    if args.dry_run:
        print("Dry run — config NOT written.")
        return 0

    CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2))
    print(f"OK — frozen list written to {CONFIG_PATH}")
    print("The next stock-suria report run will use this list.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
