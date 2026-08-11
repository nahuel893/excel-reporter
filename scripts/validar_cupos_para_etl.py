"""Valida un xlsx de cupos contra el contrato del loader de medallion-etl.

Existe para no tener que comparar formatos a mano. Chequea lo mismo que
`cupos_loader._validate_sheet_headers` y `_row_passes_filters`, pero ANTES de
correr el ETL, y con el detalle de que fila falla.

Contrato (medallion-etl, `bronze/loaders/cupos_loader.py`)
----------------------------------------------------------
La hoja va en formato ORIGEN, no en el de la base. El ETL hace el mapeo:

    CÓDIGO      -> id_ruta        RUTA        -> descripcion
    PREVENTISTA -> preventista    CATEGORIA   -> generico
    GRUPO       -> desagregado    ZONA        -> zona
    CUPO        -> cupo

y filtra `NIVEL = DETALLE`: las filas AGREGADO existen para leer, no se cargan.

Uso
---
    python scripts/validar_cupos_para_etl.py <archivo.xlsx> [--hoja <nombre>]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

HOJA_DEFAULT = "Base Pivot CASA CENTRAL"

# Encabezados que el loader mapea, en el orden en que los escribe el informe.
COLUMNAS_MAPEADAS = {
    "CÓDIGO": "id_ruta", "RUTA": "descripcion", "PREVENTISTA": "preventista",
    "CATEGORIA": "generico", "GRUPO": "desagregado", "ZONA": "zona",
    "CUPO": "cupo",
}
COLUMNA_FILTRO, VALOR_FILTRO = "NIVEL", "DETALLE"


def validar(path: Path, hoja: str) -> list[str]:
    """Devuelve la lista de problemas. Vacia = el archivo carga."""
    problemas: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if hoja not in wb.sheetnames:
            return [f"No existe la hoja {hoja!r}. Hay: {wb.sheetnames}"]
        ws = wb[hoja]

        encabezados = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None and str(v).strip():
                encabezados.setdefault(str(v).strip(), []).append(c)

        # 1. Columnas que el loader necesita. Si falta una, falla loud.
        for col, canonica in {**COLUMNAS_MAPEADAS,
                              COLUMNA_FILTRO: "(filtro)"}.items():
            if col not in encabezados:
                problemas.append(
                    f"Falta la columna {col!r} (el ETL la mapea a {canonica}). "
                    f"Hay: {sorted(encabezados)}")
            elif len(encabezados[col]) > 1:
                problemas.append(
                    f"La columna {col!r} esta repetida {len(encabezados[col])} "
                    f"veces: el loader tomaria una sola, en silencio.")
        if problemas:
            return problemas

        idx = {c: cols[0] for c, cols in encabezados.items()}

        # 2. Filas que el filtro deja pasar, y su integridad.
        detalle = agregado = 0
        sin_cupo: list[int] = []
        sin_categoria: list[int] = []
        for r in range(2, ws.max_row + 1):
            nivel = ws.cell(row=r, column=idx[COLUMNA_FILTRO]).value
            if nivel is None:
                continue
            if str(nivel).strip() != VALOR_FILTRO:
                agregado += 1
                continue
            detalle += 1
            if ws.cell(row=r, column=idx["CUPO"]).value is None:
                sin_cupo.append(r)
            if not ws.cell(row=r, column=idx["CATEGORIA"]).value:
                sin_categoria.append(r)

        if detalle == 0:
            problemas.append(
                f"Ninguna fila tiene {COLUMNA_FILTRO}={VALOR_FILTRO!r}: el ETL "
                f"cargaria CERO filas sin avisar.")
        if sin_cupo:
            problemas.append(
                f"{len(sin_cupo)} filas DETALLE sin valor de CUPO "
                f"(filas {sin_cupo[:5]}...). Suele pasar cuando el archivo "
                f"tiene formulas y no se abrio en Excel/WPS para recalcular.")
        if sin_categoria:
            problemas.append(
                f"{len(sin_categoria)} filas DETALLE sin CATEGORIA "
                f"(filas {sin_categoria[:5]}...).")

        print(f"Hoja      : {hoja}")
        print(f"Columnas  : {[c for c in ws[1] if c.value] and [str(c.value).strip() for c in ws[1] if c.value]}")
        print(f"Filas     : {detalle} DETALLE (se cargan) + {agregado} AGREGADO (se descartan)")
        return problemas
    finally:
        wb.close()


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Valida un xlsx de cupos contra el contrato del ETL.")
    p.add_argument("archivo", type=Path)
    p.add_argument("--hoja", default=HOJA_DEFAULT)
    args = p.parse_args(argv)

    if not args.archivo.exists():
        print(f"No existe: {args.archivo}")
        return 1

    print(f"Archivo   : {args.archivo.name}")
    problemas = validar(args.archivo, args.hoja)
    print()
    if problemas:
        print(f"NO CARGA — {len(problemas)} problema(s):")
        for i, p_ in enumerate(problemas, 1):
            print(f"  {i}. {p_}")
        return 1
    print("OK — el archivo cumple el contrato del ETL y se puede cargar.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
