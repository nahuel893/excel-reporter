"""Vinos Danielito: volumen y cobertura por mes, dos anios enfrentados.

Informe ad-hoc de los 7 vinos de Danielito. Una fila por articulo, doce
columnas de mes y el total del anio; abajo de cada bloque, la fila de TOTAL
BULTOS y la de COBERTURA.

Por que la lista de articulos esta escrita a mano
-------------------------------------------------
Los 7 ids NO se pueden derivar de una marca ni de un generico. Los articulos
son generico VINOS FINOS y marcas PEQUENOS PARCELEROS / TROPICO SUR / FINCA
OLLANTAY, pero VINOS FINOS tiene muchos mas articulos, y hay tres PEQUENOS
PARCELEROS (500003, 500011, 500012) que quedan deliberadamente afuera.
El articulo 500013 ni siquiera tiene marca ni generico cargados en
dim_articulo. La unica definicion valida del universo es la lista de ids.

Criterios
---------
- Universo: TODA la empresa. Sin filtro de sucursal ni de fuerza de ventas.
- Volumen: SUM(cantidades_total) de gold.fact_ventas, en bultos. Se guarda el
  valor exacto de la base; los decimales se ocultan con number_format, nunca
  se redondea el dato.
- Cobertura: clientes distintos. Se totaliza al grano (id_cliente,
  id_sucursal) DENTRO del corte y recien despues se filtra > 0. El total del
  anio es la UNION del anio, NO la suma de los meses: la cobertura no es
  aditiva entre periodos.
- La ventana no tiene tope superior por defecto: el informe llega hasta el
  ultimo dato cargado en la base. `--hasta` lo cierra, que es lo que hace falta
  cuando el cuadro va a un deck de un mes ya cerrado: un deck de JULIO con tres
  dias de AGOSTO se lee como una caida de la venta.

Uso
---
    python scripts/vinos_danielito.py                    # periodo = mes actual
    python scripts/vinos_danielito.py --periodo 2026-08
    python scripts/vinos_danielito.py --anios 2025 2026
    python scripts/vinos_danielito.py --periodo 2026-07 --hasta 2026-07-31
    python scripts/vinos_danielito.py --force            # sobrescribe la salida
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(ROOT / ".env")

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from src.core.data_loader import DataLoader  # noqa: E402
from src.core.output_paths import service_output_dir  # noqa: E402

SLUG = "vinos-danielito"
NOMBRE_ARCHIVO = "Vinos Danielito - Volumen por mes.xlsx"
HOJA = "Volumen por mes"
RANGO_CAPTURA = "A1:O28"

# Ver el docstring del modulo: esta lista ES la definicion del universo.
IDS = [500000, 500001, 500002, 500005, 500006, 500010, 500013]

DESDE = "2025-01-01"
MES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

TINTA = "1F3A5F"
BANDA = "2E5C8A"
CEBRA = "F5F8FB"
TOTALF = "E8EDF3"
ACENTO = "E2EFDA"
GRIS = "B4B4B4"
LINEA = "D6DCE4"
COBF = "FFF2CC"


def _cargar(dl: DataLoader, hasta: str | None = None) -> tuple:
    """Trae volumen por articulo/mes y las dos vistas de cobertura.

    `hasta` cierra la ventana por arriba. Va a las TRES consultas de ventas o a
    ninguna: si solo se topara el volumen, un cliente quedaria contado en un mes
    cuyos bultos no estan en el mismo cuadro.
    """
    lista = ",".join(map(str, IDS))
    tope = f" AND fv.fecha_comprobante <= '{hasta}'" if hasta else ""

    raw = dl.execute_query(
        f"""SELECT EXTRACT(YEAR FROM fv.fecha_comprobante)::int anio,
          EXTRACT(MONTH FROM fv.fecha_comprobante)::int nm, fv.id_articulo,
          SUM(fv.cantidades_total) cantidad
          FROM gold.fact_ventas fv WHERE fv.id_articulo IN ({lista})
          AND fv.fecha_comprobante >= '{DESDE}'{tope} GROUP BY 1,2,3"""
    )
    desc = dict(
        dl.execute_query(
            f"""SELECT id_articulo, des_articulo FROM gold.dim_articulo
              WHERE id_articulo IN ({lista})"""
        ).values
    )

    # La cobertura se cuenta al grano cliente dentro del corte. Agrupar antes
    # de filtrar: quien compra y devuelve dentro del mismo mes no esta cubierto.
    cob_mes = {
        (r.anio, r.nm): int(r.clientes)
        for r in dl.execute_query(
            f"""WITH v AS (
              SELECT fv.id_cliente, fv.id_sucursal,
                EXTRACT(YEAR FROM fv.fecha_comprobante)::int anio,
                EXTRACT(MONTH FROM fv.fecha_comprobante)::int nm,
                SUM(fv.cantidades_total) b
              FROM gold.fact_ventas fv WHERE fv.id_articulo IN ({lista})
                AND fv.fecha_comprobante >= '{DESDE}'{tope} GROUP BY 1,2,3,4)
            SELECT anio, nm, COUNT(*) clientes FROM v WHERE b>0 GROUP BY 1,2"""
        ).itertuples()
    }
    # El anual se recalcula sobre la ventana entera, no se suman los meses.
    cob_anio = {
        int(r.anio): int(r.clientes)
        for r in dl.execute_query(
            f"""WITH v AS (
              SELECT fv.id_cliente, fv.id_sucursal,
                EXTRACT(YEAR FROM fv.fecha_comprobante)::int anio,
                SUM(fv.cantidades_total) b
              FROM gold.fact_ventas fv WHERE fv.id_articulo IN ({lista})
                AND fv.fecha_comprobante >= '{DESDE}'{tope} GROUP BY 1,2,3)
            SELECT anio, COUNT(*) clientes FROM v WHERE b>0 GROUP BY 1"""
        ).itertuples()
    }
    return raw, desc, cob_mes, cob_anio


def construir(raw, desc, cob_mes, cob_anio, anios: list[int]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    borde = Border(*[Side(style="thin", color=LINEA)] * 4)

    ws["A1"] = "Vinos Danielito — volumen y cobertura por mes"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=TINTA)
    ws["A2"] = (
        "Bultos (cantidades_total), sin decimales a la vista; el valor guardado es el exacto de la base · "
        "un 0 en GRIS es sin ventas, un 0 en NEGRO son ventas menores a medio bulto · "
        "la COBERTURA son clientes distintos: no se suma entre meses, el anual es la unión del año"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="7F7F7F")

    fila = 4
    for anio in anios:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=15)
        c = ws.cell(fila, 1, f"AÑO {anio}")
        c.fill = PatternFill("solid", start_color=BANDA, end_color=BANDA)
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", indent=1)
        fila += 1

        for j, t in enumerate(["Código", "Artículo", *MES, "Total"], 1):
            c = ws.cell(fila, j, t)
            c.fill = PatternFill("solid", start_color=TINTA, end_color=TINTA)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.alignment = Alignment(
                horizontal="center" if j > 2 else "left", indent=0 if j > 2 else 1
            )
            c.border = borde
        fila += 1

        for k, aid in enumerate(IDS):
            cebra = (
                PatternFill("solid", start_color=CEBRA, end_color=CEBRA) if k % 2 else None
            )
            vals = [
                raw[(raw.anio == anio) & (raw.nm == m) & (raw.id_articulo == aid)].cantidad.sum()
                for m in range(1, 13)
            ]
            for j, v in enumerate([aid, desc.get(aid, "(sin descripción)"), *vals, sum(vals)], 1):
                c = ws.cell(fila, j, v)
                c.border = borde
                c.font = Font(name="Calibri", size=10)
                if cebra:
                    c.fill = cebra
                if j == 1:
                    c.number_format = "0"
                    c.alignment = Alignment(horizontal="left", indent=1)
                elif j == 2:
                    c.alignment = Alignment(horizontal="left", indent=1)
                else:
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="right", indent=1)
                    if not v:
                        c.font = Font(name="Calibri", size=10, color=GRIS)
                if j == 15:
                    c.fill = PatternFill("solid", start_color=ACENTO, end_color=ACENTO)
                    c.font = Font(name="Calibri", size=10, bold=True)
            fila += 1

        tot = [raw[(raw.anio == anio) & (raw.nm == m)].cantidad.sum() for m in range(1, 13)]
        for j, v in enumerate(["", "TOTAL BULTOS", *tot, sum(tot)], 1):
            c = ws.cell(fila, j, v)
            c.fill = PatternFill("solid", start_color=TOTALF, end_color=TOTALF)
            c.font = Font(name="Calibri", size=10, bold=True, color=TINTA)
            c.border = Border(
                top=Side(style="medium", color=TINTA),
                bottom=Side(style="thin", color=LINEA),
                left=Side(style="thin", color=LINEA),
                right=Side(style="thin", color=LINEA),
            )
            c.alignment = Alignment(horizontal="right" if j > 2 else "left", indent=1)
            if j > 2:
                c.number_format = "#,##0"
        fila += 1

        # Otra metrica, otro color: que a nadie se le ocurra sumarla con la de arriba.
        cm = [cob_mes.get((anio, m), 0) for m in range(1, 13)]
        for j, v in enumerate(["", "COBERTURA (clientes únicos)", *cm, cob_anio.get(anio, 0)], 1):
            c = ws.cell(fila, j, v)
            c.fill = PatternFill("solid", start_color=COBF, end_color=COBF)
            c.font = Font(name="Calibri", size=10, bold=True, color="7F6000")
            c.border = borde
            c.alignment = Alignment(horizontal="right" if j > 2 else "left", indent=1)
            if j > 2:
                c.number_format = "#,##0"
                if not v:
                    c.font = Font(name="Calibri", size=10, bold=True, color=GRIS)
        fila += 3

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 44
    for j in range(3, 16):
        ws.column_dimensions[get_column_letter(j)].width = 7.5
    ws.column_dimensions["O"].width = 9
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C6"
    return wb


def main() -> int:
    hoy = date.today()
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--periodo", default=f"{hoy:%Y-%m}", help="carpeta de salida YYYY-MM")
    ap.add_argument(
        "--anios", nargs="+", type=int, default=None,
        help="anios a mostrar (default: el anterior y el del periodo)",
    )
    ap.add_argument(
        "--hasta", default=None,
        help="ultimo dia a incluir (YYYY-MM-DD); por defecto llega al ultimo dato cargado",
    )
    ap.add_argument("--force", action="store_true", help="sobrescribir la salida si ya existe")
    ap.add_argument("--sin-png", action="store_true", help="no renderizar la imagen")
    args = ap.parse_args()

    anio_periodo = int(args.periodo[:4])
    anios = args.anios or [anio_periodo - 1, anio_periodo]

    out = service_output_dir(SLUG, f"{args.periodo}-01", granularity="month")
    out.mkdir(parents=True, exist_ok=True)
    ruta = out / NOMBRE_ARCHIVO
    if ruta.exists() and not args.force:
        print(f"Ya existe: {ruta}\nUsar --force para sobrescribirlo.")
        return 1

    dl = DataLoader()
    raw, desc, cob_mes, cob_anio = _cargar(dl, hasta=args.hasta)
    construir(raw, desc, cob_mes, cob_anio, anios).save(ruta)
    print(f"archivo: {ruta}")

    if not args.sin_png:
        from src.core.excel_renderers import get_renderer

        png = get_renderer("libreoffice").render(
            xlsx_path=ruta, sheet=HOJA, range_addr=RANGO_CAPTURA, output_dir=out, crop=False
        )
        print(f"PNG: {png.name}")

    for anio in anios:
        print(
            f"cobertura {anio}: {[cob_mes.get((anio, m), 0) for m in range(1, 13)]} "
            f"-> anual {cob_anio.get(anio, 0)}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
