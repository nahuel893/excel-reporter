"""
ExcelManager - Operaciones sobre archivos Excel existentes.

A diferencia de ExcelWriter (que crea archivos), ExcelManager opera sobre
archivos .xlsx ya guardados en disco: captura rangos como imagen PNG,
lee rangos como DataFrame, etc.

Requiere:
  - Pillow (PIL) para renderizado de imagen
  - openpyxl para lectura de celdas y estilos
  - LibreOffice (soffice) para recalcular formulas antes del renderizado
"""
import logging
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from config.settings import DATA_OUTPUT

logger = logging.getLogger(__name__)

# ── Constantes de renderizado ────────────────────────────────────────────────

_DEFAULT_COL_WIDTH = 8.43
_DEFAULT_ROW_HEIGHT = 15.0
_CHAR_TO_PX = 7.5          # 1 caracter Excel ≈ 7.5 px
_PT_TO_PX = 96.0 / 72.0    # 1 punto ≈ 1.333 px
_PADDING_X = 4              # padding horizontal en celdas
_PADDING_Y = 2              # padding vertical en celdas
_DEFAULT_FONT_SIZE = 11


def _col_letter_to_index(letter: str) -> int:
    """Convierte letra de columna Excel a indice 1-based (A->1, Z->26, AA->27)."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def _parse_range(range_addr: str) -> tuple[int, int, int, int]:
    """Parsea rango Excel 'A1:H20' a (col_inicio, fila_inicio, col_fin, fila_fin) 1-based."""
    parts = range_addr.upper().split(":")
    if len(parts) != 2:
        raise ValueError(f"Rango invalido: {range_addr!r}. Formato esperado: 'A1:H20'")

    def parse_cell(cell: str) -> tuple[int, int]:
        col_str = "".join(c for c in cell if c.isalpha())
        row_str = "".join(c for c in cell if c.isdigit())
        return _col_letter_to_index(col_str), int(row_str)

    col1, row1 = parse_cell(parts[0])
    col2, row2 = parse_cell(parts[1])
    return col1, row1, col2, row2


def _openpyxl_color_to_rgb(color, default: str = "#000000") -> str:
    """Convierte un openpyxl Color a hex RGB '#RRGGBB'."""
    if color is None:
        return default
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        rgb = color.rgb
        # openpyxl usa ARGB (8 chars) o RGB (6 chars)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        return f"#{rgb}"
    if color.type == "theme":
        # Temas comunes: 0=blanco, 1=negro, etc.
        theme_map = {0: "#FFFFFF", 1: "#000000", 2: "#44546A", 3: "#E7E6E6",
                     4: "#4472C4", 5: "#ED7D31", 6: "#A5A5A5", 7: "#FFC000",
                     8: "#5B9BD5", 9: "#70AD47"}
        base = theme_map.get(color.theme, default)
        if color.tint and color.tint != 0:
            return _apply_tint(base, color.tint)
        return base
    if color.type == "indexed" and color.indexed is not None:
        # Colores indexados — solo los mas comunes
        if color.indexed == 64:  # system foreground (negro)
            return "#000000"
        if color.indexed == 65:  # system background (blanco)
            return "#FFFFFF"
    return default


def _apply_tint(hex_color: str, tint: float) -> str:
    """Aplica tint de Excel a un color hex."""
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    else:
        factor = 1 + tint
        r = int(r * factor)
        g = int(g * factor)
        b = int(b * factor)
    r, g, b = min(255, max(0, r)), min(255, max(0, g)), min(255, max(0, b))
    return f"#{r:02X}{g:02X}{b:02X}"


def _hex_to_tuple(hex_color: str) -> tuple[int, int, int]:
    """'#RRGGBB' -> (R, G, B)."""
    h = hex_color.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _border_width(style: str | None) -> int:
    """Devuelve ancho en px para un estilo de borde de openpyxl."""
    if style is None:
        return 0
    return {"thin": 1, "medium": 2, "thick": 3, "double": 2,
            "hair": 1, "mediumDashed": 2, "dashDot": 1,
            "dashed": 1, "dotted": 1}.get(style, 1)


class ExcelManager:
    """Operaciones sobre un archivo .xlsx existente en disco."""

    def __init__(self, ruta_excel: Path):
        if not Path(ruta_excel).exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_excel}")
        self.ruta_excel = Path(ruta_excel)

    # ── LibreOffice helpers ───────────────────────────────────────────────

    @staticmethod
    def _find_soffice() -> str | None:
        """Devuelve la ruta a soffice/libreoffice, o None si no esta disponible."""
        for candidate in ("soffice", "libreoffice"):
            path = shutil.which(candidate)
            if path:
                return path
        return None

    @staticmethod
    def _recalc_with_libreoffice(source: Path, work_dir: Path, soffice: str) -> Path:
        """
        Copia el xlsx a work_dir, ejecuta LibreOffice para recalcular formulas
        y retorna el Path al archivo recalculado.

        Args:
            source:   Path al xlsx original
            work_dir: Directorio temporal de trabajo
            soffice:  Ruta al ejecutable soffice/libreoffice

        Returns:
            Path al xlsx recalculado dentro de work_dir

        Raises:
            RuntimeError: Si soffice retorna error o el archivo de salida no existe
        """
        # Usar un nombre distinto para la salida para poder validar que LibreOffice generó algo
        out_dir = Path(work_dir) / "out"
        out_dir.mkdir(exist_ok=True)

        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(source)],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice fallo al recalcular '{source.name}' "
                f"(exit {result.returncode}): {result.stderr.strip()}"
            )

        recalculated = out_dir / source.name
        if not recalculated.exists():
            raise RuntimeError(
                f"LibreOffice no genero el archivo esperado: {recalculated}"
            )
        return recalculated

    # ── Captura de rango como imagen ─────────────────────────────────────

    def capture_range(
        self,
        sheet_name: str,
        range_addr: str,
        output_dir: Path | None = None,
        dpi: int = 300,
    ) -> Path:
        """
        Captura un rango de una hoja como imagen PNG de alta calidad.

        Proceso:
        1. LibreOffice recalcula formulas y re-guarda el xlsx
        2. Se crea un xlsx temporal con SOLO la hoja target activada, fit-to-page
        3. LibreOffice exporta a PDF (escala vectorial)
        4. pdftoppm rasteriza el PDF a PNG en el DPI pedido (default 300)
        5. Pillow recorta bordes blancos alrededor del contenido

        Requiere `pdftoppm` (paquete poppler-utils) en el PATH.

        Args:
            sheet_name: Nombre de la hoja del Excel
            range_addr: Rango en formato 'A1:H20' o 'auto' para auto-deteccion
            output_dir: Directorio de salida (por defecto DATA_OUTPUT)
            dpi: Resolucion de salida en puntos por pulgada (default 300).
                 Valores tipicos: 150 draft, 300 calidad de email, 600 impresion.

        Returns:
            Path al archivo PNG generado
        """
        # Validar hoja
        wb_check = openpyxl.load_workbook(self.ruta_excel, data_only=True)
        if sheet_name not in wb_check.sheetnames:
            wb_check.close()
            raise ValueError(
                f"Hoja '{sheet_name}' no encontrada en {self.ruta_excel.name}. "
                f"Hojas disponibles: {wb_check.sheetnames}"
            )
        wb_check.close()

        # Auto-deteccion de bordes (informativo; el filename lo usa)
        if range_addr == "auto":
            detected = self.detect_bordered_range(sheet_name)
            if detected is None:
                raise ValueError(
                    f"No se detectaron bordes gruesos en la hoja '{sheet_name}'"
                )
            range_addr = detected
            logger.info("Auto-detected range: %s", range_addr)

        # Verificar LibreOffice
        soffice = self._find_soffice()
        if soffice is None:
            raise RuntimeError(
                "LibreOffice no encontrado. Instalar con: sudo pacman -S libreoffice-fresh"
            )

        # Verificar pdftoppm
        pdftoppm = shutil.which("pdftoppm")
        if pdftoppm is None:
            raise RuntimeError(
                "pdftoppm no encontrado. Instalar poppler: sudo pacman -S poppler"
            )

        from PIL import Image

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_dir = Path(tmp_dir)

            # Paso 1: Recalcular formulas con LibreOffice
            recalc_path = self._recalc_with_libreoffice(
                self.ruta_excel, tmp_dir, soffice
            )

            # Paso 2: Preparar xlsx con la hoja target activa y las demas ocultas
            export_path = self._prepare_sheet_for_export(
                recalc_path, sheet_name, tmp_dir
            )

            # Paso 3: LibreOffice exporta a PDF (escala vectorial)
            pdf_dir = tmp_dir / "pdf"
            pdf_dir.mkdir()
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(pdf_dir), str(export_path)],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice fallo al exportar PDF: {result.stderr.strip()}"
                )

            pdf_candidates = list(pdf_dir.glob("*.pdf"))
            if not pdf_candidates:
                raise RuntimeError("LibreOffice no genero ningun archivo PDF")
            pdf_path = pdf_candidates[0]

            # Paso 4: pdftoppm rasteriza PDF → PNG a DPI alta
            png_dir = tmp_dir / "png"
            png_dir.mkdir()
            png_prefix = png_dir / "page"
            result = subprocess.run(
                [pdftoppm, "-r", str(dpi), "-png", "-f", "1", "-l", "1",
                 str(pdf_path), str(png_prefix)],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode != 0:
                raise RuntimeError(
                    f"pdftoppm fallo al rasterizar PDF: {result.stderr.strip()}"
                )

            png_candidates = list(png_dir.glob("page-*.png"))
            if not png_candidates:
                raise RuntimeError("pdftoppm no genero ningun archivo PNG")

            img = Image.open(png_candidates[0])

            # Paso 5: Recortar bordes blancos alrededor del contenido
            import numpy as np
            arr = np.array(img.convert("RGB"))
            non_white = np.any(arr < 250, axis=2)
            rows = np.any(non_white, axis=1)
            cols = np.any(non_white, axis=0)
            if rows.any() and cols.any():
                rmin, rmax = np.where(rows)[0][[0, -1]]
                cmin, cmax = np.where(cols)[0][[0, -1]]
                cropped = img.crop((cmin, rmin, cmax + 1, rmax + 1))
            else:
                cropped = img

        # Guardar
        output_dir = Path(output_dir) if output_dir else DATA_OUTPUT
        output_dir.mkdir(parents=True, exist_ok=True)
        range_slug = range_addr.replace(":", "_")
        out_path = output_dir / f"{self.ruta_excel.stem}_{sheet_name}_{range_slug}.png"
        cropped.save(out_path, "PNG", dpi=(dpi, dpi))

        logger.info(
            "Imagen capturada: %s (%dx%d px @ %d DPI)",
            out_path.name, cropped.width, cropped.height, dpi,
        )
        return out_path

    @staticmethod
    def _prepare_sheet_for_export(source: Path, sheet_name: str, work_dir: Path) -> Path:
        """Prepara un xlsx donde la hoja target es la activa y las demas estan ocultas.

        Configura el page setup para que toda la tabla quepa en una sola pagina,
        evitando que LibreOffice corte la imagen al ancho de A4.
        """
        wb = openpyxl.load_workbook(source)
        ws = wb[sheet_name]

        # Activar la hoja target
        wb.active = wb.sheetnames.index(sheet_name)
        # Ocultar las demas
        for s in wb.sheetnames:
            if s != sheet_name:
                wb[s].sheet_state = "hidden"

        # Configurar page setup: toda la tabla en 1 sola pagina
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # Margenes minimas para maximizar el contenido
        ws.page_margins.left = 0
        ws.page_margins.right = 0
        ws.page_margins.top = 0
        ws.page_margins.bottom = 0
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        out_path = work_dir / f"export_{source.name}"
        wb.save(out_path)
        wb.close()
        return out_path


    # ── Helpers internos ─────────────────────────────────────────────────

    @staticmethod
    def _load_font(bold: bool = False, size: int = _DEFAULT_FONT_SIZE):
        """Carga una fuente TrueType o devuelve la fuente por defecto de Pillow."""
        from PIL import ImageFont

        px_size = int(size * _PT_TO_PX)
        # Intentar fuentes comunes del sistema
        candidates = (
            ["/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
             "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
             "/usr/share/fonts/dejavu-sans-fonts/DejaVuSans-Bold.ttf"]
            if bold else
            ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
             "/usr/share/fonts/TTF/DejaVuSans.ttf",
             "/usr/share/fonts/dejavu-sans-fonts/DejaVuSans.ttf"]
        )
        for path in candidates:
            if Path(path).exists():
                return ImageFont.truetype(path, px_size)
        try:
            return ImageFont.truetype("DejaVuSans.ttf", px_size)
        except OSError:
            return ImageFont.load_default()

    @staticmethod
    def _format_value(value, number_format: str | None = None) -> str:
        """Formatea un valor de celda para mostrar como texto."""
        if value is None:
            return ""
        if isinstance(value, float):
            if number_format and ("%" in number_format):
                return f"{value * 100:.1f}%"
            if number_format and ("#,##0" in number_format):
                if value == int(value):
                    return f"{int(value):,}".replace(",", ".")
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            if value == int(value):
                return str(int(value))
            return f"{value:.2f}"
        if isinstance(value, int):
            if number_format and ("#,##0" in number_format):
                return f"{value:,}".replace(",", ".")
            return str(value)
        return str(value)

    def _detect_bordered_range(
        self,
        ws,
        border_styles: tuple[str, ...] = ("medium", "thick", "double"),
    ) -> str | None:
        """Detecta el rectangulo definido por bordes gruesos en una hoja."""
        min_r = min_c = float("inf")
        max_r = max_c = 0

        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column,
        ):
            for cell in row:
                border = cell.border
                if border is None:
                    continue
                has_border = any(
                    getattr(side, "style", None) in border_styles
                    for side in (border.top, border.bottom, border.left, border.right)
                    if side is not None
                )
                if has_border:
                    min_r = min(min_r, cell.row)
                    max_r = max(max_r, cell.row)
                    min_c = min(min_c, cell.column)
                    max_c = max(max_c, cell.column)

        if max_r == 0:
            return None

        return f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"

    def detect_bordered_range(
        self,
        sheet_name: str,
        border_styles: tuple[str, ...] = ("medium", "thick", "double"),
    ) -> str | None:
        """Detecta el rectangulo definido por bordes gruesos (API publica)."""
        wb = openpyxl.load_workbook(self.ruta_excel, data_only=False)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        result = self._detect_bordered_range(ws, border_styles)
        wb.close()
        return result

    # ── Lectura de rango como DataFrame ──────────────────────────────────

    def get_dataframe(self, sheet_name: str, range_addr: str) -> pd.DataFrame:
        """
        Lee un rango de una hoja como DataFrame.
        La primera fila del rango se usa como encabezados.
        """
        wb = openpyxl.load_workbook(self.ruta_excel, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Hoja '{sheet_name}' no encontrada en {self.ruta_excel.name}. "
                f"Hojas disponibles: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        col1, row1, col2, row2 = _parse_range(range_addr)

        rows = []
        for r in range(row1, row2 + 1):
            row_data = []
            for c in range(col1, col2 + 1):
                cell = ws.cell(row=r, column=c)
                row_data.append(cell.value)
            rows.append(row_data)

        if not rows:
            return pd.DataFrame()

        headers = rows[0]
        data = rows[1:]
        return pd.DataFrame(data, columns=headers)
