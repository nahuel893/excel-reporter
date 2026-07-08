"""HTML + Playwright renderer (no LibreOffice required).

Pipeline:
    xlsx → xlsx2html (HTML with inline styles) → Playwright Chromium
    headless → screenshot full_page → PNG.

Limitations (Phase 1):
    - `range_addr` is not enforced by xlsx2html; the whole sheet is rendered.
      The range is retained in the output filename for consistency.
    - xlsx2html may drop fidelity on complex conditional formatting,
      irregular merged cell shapes, or embedded charts. Prefer the
      libreoffice renderer for those cases until validated.
"""
from __future__ import annotations

import importlib
import logging
import tempfile
from pathlib import Path

import openpyxl


logger = logging.getLogger(__name__)


def _require(module_name: str, install_hint: str):
    """Import a module or raise RuntimeError with an install hint."""
    try:
        return importlib.import_module(module_name)
    except ImportError as exc:
        raise RuntimeError(
            f"Dependency '{module_name}' not available. Install: {install_hint}"
        ) from exc


class HtmlPlaywrightRenderer:
    """Renderer backed by xlsx2html + Chromium headless via Playwright."""

    name = "html_playwright"

    def render(
        self,
        xlsx_path: Path,
        sheet: str,
        range_addr: str,
        output_dir: Path,
        dpi: int = 300,
        crop: bool = False,
    ) -> Path:
        # `crop` is accepted for Protocol consistency but ignored: xlsx2html
        # renders the whole sheet and cannot restrict to a range/print_area.
        xlsx_path = Path(xlsx_path)
        output_dir = Path(output_dir)

        # Validate sheet exists before touching any heavy dep.
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"Sheet {sheet!r} not in {xlsx_path.name}. "
                    f"Available: {wb.sheetnames}"
                )
        finally:
            wb.close()

        # Lazy-import heavy deps so missing deps fail only here.
        xlsx2html_mod = _require("xlsx2html", "pip install xlsx2html")
        playwright_mod = _require(
            "playwright.sync_api",
            "pip install playwright && playwright install chromium",
        )

        output_dir.mkdir(parents=True, exist_ok=True)
        range_slug = range_addr.replace(":", "_")
        out_png = output_dir / f"{xlsx_path.stem}_{sheet}_{range_slug}.png"

        # Scale relative to CSS baseline 96 DPI.
        device_scale = max(1.0, dpi / 96.0)

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_dir = Path(tmp_dir)
            html_path = tmp_dir / f"{sheet}.html"

            xlsx2html_mod.xlsx2html(str(xlsx_path), str(html_path), sheet=sheet)

            with playwright_mod.sync_playwright() as p:
                browser = p.chromium.launch()
                try:
                    ctx = browser.new_context(device_scale_factor=device_scale)
                    page = ctx.new_page()
                    page.goto(f"file://{html_path.absolute()}")
                    page.screenshot(path=str(out_png), full_page=True)
                finally:
                    browser.close()

        logger.info(
            "HTML+Playwright capture: %s → %s (%d bytes)",
            sheet, out_png.name, out_png.stat().st_size,
        )
        return out_png
