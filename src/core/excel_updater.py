"""
excel_updater - Reemplazo de datos en hojas Excel existentes.

Proporciona funcionalidad para actualizar datos en worksheets existentes
preservando columnas de formulas y definiciones de tablas Excel.
"""
import logging
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

logger = logging.getLogger(__name__)


def _adjust_formula_row(formula: str, delta: int) -> str:
    """Increment relative row references in a formula string by delta rows.

    Absolute row references (prefixed with $) are left unchanged.
    Examples:
        _adjust_formula_row("=BUSCARV(A2,$B$1:$D$100,3,0)", 1) -> "=BUSCARV(A3,$B$1:$D$100,3,0)"
        _adjust_formula_row("=SUMA($A$1:A2)", 2)               -> "=SUMA($A$1:A4)"
    """
    def _replacer(m: re.Match) -> str:
        col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), m.group(4)
        if row_abs == "$":
            return f"{col_abs}{col}{row_abs}{row}"  # absolute row — keep
        return f"{col_abs}{col}{row_abs}{int(row) + delta}"  # relative — adjust

    return re.sub(r"(\$?)([A-Z]+)(\$?)(\d+)", _replacer, formula)


def _extend_formula_columns(
    ws,
    header_row: int,
    rows_written: int,
    data_col_indices: set[int],
) -> int:
    """Extend formula columns (adjacent to data) to cover all written data rows.

    Scans the first data row for formula cells that are NOT in data_col_indices,
    then copies each formula to all subsequent data rows, adjusting relative row
    references. Idempotent — safe to call on reruns.

    Returns the number of formula columns extended.
    """
    if rows_written <= 1:
        return 0

    first_data_row = header_row + 1
    last_data_row = header_row + rows_written
    extended = 0

    for col_idx in range(1, ws.max_column + 1):
        if col_idx in data_col_indices:
            continue
        template = ws.cell(row=first_data_row, column=col_idx).value
        if not isinstance(template, str) or not template.startswith("="):
            continue
        for row in range(first_data_row + 1, last_data_row + 1):
            ws.cell(row=row, column=col_idx).value = _adjust_formula_row(
                template, row - first_data_row
            )
        extended += 1

    if extended:
        logger.debug("Columnas de formula extendidas hasta fila %d: %d", last_data_row, extended)

    return extended


def _coerce_value(val):
    """Convierte tipos pandas/numpy a Python nativo para openpyxl."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        if np.isnan(val):
            return None
        return float(val)
    if isinstance(val, float) and np.isnan(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if val is pd.NaT:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        return val
    return val


def replace_sheet_data(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    data_columns: list[str],
    header_row: int = 1,
    numeric_columns: list[str] | None = None,
) -> int:
    """
    Reemplaza datos en columnas especificas de una hoja Excel existente,
    preservando columnas de formulas y definiciones de tablas.

    Args:
        wb: Workbook de openpyxl (ya cargado)
        sheet_name: Nombre de la hoja a actualizar
        df: DataFrame con los nuevos datos
        data_columns: Lista de nombres de columnas a reemplazar
        header_row: Fila del encabezado (1-based, default: 1)
        numeric_columns: Columnas (subconjunto de data_columns) cuyas celdas
            deben forzarse a formato numerico ("0") al escribirse. Necesario
            para columnas-clave enteras (p.ej. id_ruta) que la plantilla trae
            con formato de FECHA heredado: sin esto, un round-trip de openpyxl
            reserializa el entero como fecha y el bug del anio bisiesto 1900
            corre los seriales >= 60 un dia, rompiendo los VLOOKUP de coincidencia
            exacta rio abajo (ver CuposVolumen!Código -> AvanceR).

    Returns:
        Numero de filas escritas

    Raises:
        KeyError: Si la hoja no existe en el workbook
        ValueError: Si alguna columna de data_columns no existe en el header o en el DataFrame
    """
    numeric_set = set(numeric_columns or [])
    # 1. Obtener worksheet
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Hoja '{sheet_name}' no encontrada en el workbook")
    ws = wb[sheet_name]

    # 2. Leer header y construir mapa: nombre_columna -> col_index (1-based)
    header_cells = ws[header_row]
    col_map: dict[str, int] = {}
    for cell in header_cells:
        if cell.value is not None:
            col_map[str(cell.value)] = cell.column

    # 3. Validar que todas las data_columns existen en header Y en DataFrame
    missing_in_header = [c for c in data_columns if c not in col_map]
    missing_in_df = [c for c in data_columns if c not in df.columns]

    errors = []
    if missing_in_header:
        errors.append(f"Columnas no encontradas en el header de la hoja: {missing_in_header}")
    if missing_in_df:
        errors.append(f"Columnas no encontradas en el DataFrame: {missing_in_df}")
    if errors:
        raise ValueError(" | ".join(errors))

    # 4. Limpiar datos existentes (solo celdas en data_columns)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        for col_name in data_columns:
            col_idx = col_map[col_name]
            ws.cell(row=row_idx, column=col_idx).value = None

    # 5. Escribir nuevos datos fila por fila
    rows_written = 0
    for df_row_idx, (_, row) in enumerate(df.iterrows()):
        data_row = header_row + 1 + df_row_idx
        for col_name in data_columns:
            col_idx = col_map[col_name]
            raw_val = row[col_name]
            cell = ws.cell(row=data_row, column=col_idx, value=_coerce_value(raw_val))
            # Force numeric format on integer key columns so a later openpyxl
            # round-trip does not reserialize the value as a date (would shift
            # serials >= 60 via the 1900 leap-year bug and break exact VLOOKUPs).
            if col_name in numeric_set:
                cell.number_format = "0"
        rows_written += 1

    # 5b. Extender columnas de fórmulas adyacentes (BUSCARVs, etc.)
    data_col_indices = {col_map[c] for c in data_columns}
    _extend_formula_columns(ws, header_row, rows_written, data_col_indices)

    # 6. Redimensionar ref de tabla Excel si existe
    tables = list(ws.tables.values())
    if len(tables) > 1:
        logger.warning(
            "Hoja '%s' tiene %d tablas; se actualizara solo la primera: '%s'",
            sheet_name,
            len(tables),
            tables[0].displayName,
        )

    if tables:
        table = tables[0]
        current_ref = table.ref  # ej: "A3:Z100"

        # Extraer letras de columna del ref existente (preservar span completo)
        ref_start, ref_end = current_ref.split(":")
        # Separar letra de numero en cada extremo
        start_col_letter = "".join(c for c in ref_start if c.isalpha())
        end_col_letter = "".join(c for c in ref_end if c.isalpha())

        new_end_row = header_row + rows_written
        table.ref = f"{start_col_letter}{header_row}:{end_col_letter}{new_end_row}"
        logger.debug(
            "Tabla '%s' ref actualizada: %s -> %s",
            table.displayName,
            current_ref,
            table.ref,
        )

    return rows_written


def import_xlsx_as_sheet(
    target_wb: Workbook,
    source_path: Path,
    target_sheet_name: str,
) -> int:
    """Read source xlsx (first sheet, values only) into target_wb under target_sheet_name.

    If target_sheet_name already exists in target_wb, it is removed first.
    Source is opened with data_only=True so formulas resolve to last calculated values.

    Args:
        target_wb: Destination Workbook instance (already loaded or fresh).
        source_path: Path to the source xlsx file (must exist).
        target_sheet_name: Name to give the imported sheet in target_wb.

    Returns:
        Number of data rows written (excluding the header row). 0 if source has no rows
        beyond the header.

    Raises:
        FileNotFoundError: If source_path does not exist.
    """
    if not source_path.exists():
        raise FileNotFoundError(f"Source xlsx not found: {source_path}")

    src_wb = load_workbook(str(source_path), data_only=True, read_only=True)
    try:
        src_ws = src_wb.worksheets[0]  # first sheet only
        if target_sheet_name in target_wb.sheetnames:
            del target_wb[target_sheet_name]
        dst_ws = target_wb.create_sheet(title=target_sheet_name)
        total_rows = 0
        for row_idx, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                dst_ws.cell(row=row_idx, column=col_idx, value=value)
            total_rows = row_idx
        # Data rows = total rows - 1 header row (0 if source is empty or header-only)
        return max(0, total_rows - 1) if total_rows > 0 else 0
    finally:
        src_wb.close()
