"""
ExcelWriter - Generador de archivos Excel.

Proporciona funcionalidad para generar reportes Excel
con formato de tabla profesional y sistema de formatos modular.
"""
from pathlib import Path
from dataclasses import dataclass, field
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.settings import DATA_OUTPUT


@dataclass
class ColumnFormat:
    """Formato configurable para una columna."""
    number_format: str | None = None
    alignment: str | None = None
    font_bold: bool = False
    width: int | None = None


@dataclass
class ColumnGroup:
    """Define un grupo de columnas colapsable.

    Attributes:
        start_col: Nombre de columna donde inicia el grupo
        end_col: Nombre de columna donde termina el grupo
        collapsed: Si el grupo inicia colapsado
        hidden: Si las columnas estan ocultas
    """
    start_col: str
    end_col: str
    collapsed: bool = True
    hidden: bool = False


@dataclass
class SheetStyle:
    """Estilos configurables para una hoja Excel.

    Attributes:
        header_bold: Encabezados en negrita
        header_center: Encabezados centrados
        auto_width: Ajustar ancho automaticamente
        numeric_format: Formato para columnas numericas (ej: '#,##0' sin decimales)
        column_formats: Formatos especificos por nombre de columna
        column_groups: Lista de grupos de columnas colapsables
        summary_rows: Filas de resumen al inicio {etiqueta: valor}
        as_table: Convertir datos a tabla Excel nativa
        table_style: Estilo de tabla Excel (ej: 'TableStyleMedium9')
    """
    header_bold: bool = True
    header_center: bool = True
    auto_width: bool = True
    numeric_format: str = "#,##0"
    column_formats: dict[str, ColumnFormat] = field(default_factory=dict)
    column_groups: list[ColumnGroup] = field(default_factory=list)
    summary_rows: dict[str, int | float | str] = field(default_factory=dict)
    as_table: bool = True
    table_style: str = "TableStyleMedium9"


# Estilo por defecto
DEFAULT_STYLE = SheetStyle()


def _detect_numeric_columns(df: pd.DataFrame) -> list[str]:
    """Detecta columnas numericas en un DataFrame."""
    return [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]


def _apply_cell_format(cell, col_name: str, style: SheetStyle, is_header: bool = False):
    """
    Aplica formato a una celda segun el estilo y tipo de columna.

    Args:
        cell: Celda de openpyxl
        col_name: Nombre de la columna
        style: Estilo de la hoja
        is_header: Si es celda de encabezado
    """
    if is_header:
        # Texto blanco, negrita, centrado y distribuido
        cell.font = Font(bold=style.header_bold, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        return

    # Formato especifico por columna
    if col_name in style.column_formats:
        fmt = style.column_formats[col_name]
        if fmt.number_format:
            cell.number_format = fmt.number_format
        if fmt.alignment:
            cell.alignment = Alignment(horizontal=fmt.alignment)
        if fmt.font_bold:
            cell.font = Font(bold=True)
        return

    # Formato generico para numericos
    if isinstance(cell.value, (int, float)) and cell.value is not None:
        cell.number_format = style.numeric_format


def _auto_fit_columns(ws, style: SheetStyle, header_row: int = 1):
    """Ajusta el ancho de columnas automaticamente o segun configuracion."""
    if not style.auto_width:
        return

    headers = [cell.value for cell in ws[header_row]]

    for col_idx, column in enumerate(ws.columns):
        col_name = headers[col_idx] if col_idx < len(headers) else None
        column_letter = column[0].column_letter

        # Ancho fijo si esta configurado
        if col_name and col_name in style.column_formats:
            fmt = style.column_formats[col_name]
            if fmt.width:
                ws.column_dimensions[column_letter].width = fmt.width
                continue

        # Auto-fit
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2


def _apply_column_groups(ws, df: pd.DataFrame, style: SheetStyle):
    """Aplica grupos de columnas colapsables."""
    if not style.column_groups:
        return

    headers = list(df.columns)
    col_to_idx = {name: idx + 1 for idx, name in enumerate(headers)}

    for group in style.column_groups:
        start_idx = col_to_idx.get(group.start_col)
        end_idx = col_to_idx.get(group.end_col)

        if start_idx and end_idx and start_idx <= end_idx:
            # Agrupar columnas
            ws.column_dimensions.group(
                get_column_letter(start_idx),
                get_column_letter(end_idx),
                hidden=group.hidden
            )

    # Configurar para que el boton de grupo este arriba
    ws.sheet_properties.outlinePr.summaryRight = False


def _write_summary_rows(ws, style: SheetStyle) -> int:
    """Escribe filas de resumen al inicio de la hoja.

    Returns:
        Numero de filas escritas (offset para los datos)
    """
    if not style.summary_rows:
        return 0

    for r_idx, (label, value) in enumerate(style.summary_rows.items(), 1):
        # Columna A: etiqueta en negrita
        cell_label = ws.cell(row=r_idx, column=1, value=label)
        cell_label.font = Font(bold=True)

        # Columna B: valor
        cell_value = ws.cell(row=r_idx, column=2, value=value)

    # Fila vacia de separacion
    return len(style.summary_rows) + 1


def _apply_table_format(ws, df: pd.DataFrame, style: SheetStyle, header_row: int):
    """Aplica formato de tabla Excel nativa al rango de datos.

    Args:
        ws: Worksheet de openpyxl
        df: DataFrame con los datos
        style: Estilo de la hoja
        header_row: Fila donde inicia el encabezado de la tabla
    """
    if not style.as_table or df.empty:
        return

    # Calcular rango de la tabla
    num_rows = len(df) + 1  # +1 por el header
    num_cols = len(df.columns)
    start_cell = f"A{header_row}"
    end_cell = f"{get_column_letter(num_cols)}{header_row + num_rows - 1}"
    table_range = f"{start_cell}:{end_cell}"

    # Crear tabla con nombre unico
    table_name = f"Tabla_{ws.title.replace(' ', '_')}"
    table = Table(displayName=table_name, ref=table_range)

    # Aplicar estilo
    table_style = TableStyleInfo(
        name=style.table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = table_style

    ws.add_table(table)


def _write_sheet(ws, df: pd.DataFrame, style: SheetStyle):
    """Escribe datos y aplica formato a una hoja."""
    headers = list(df.columns)

    # Escribir filas de resumen primero
    row_offset = _write_summary_rows(ws, style)

    # Escribir datos con offset
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        actual_row = r_idx + row_offset
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=actual_row, column=c_idx, value=value)
            col_name = headers[c_idx - 1] if c_idx <= len(headers) else ""
            _apply_cell_format(cell, col_name, style, is_header=(r_idx == 1))

    header_row = row_offset + 1
    _auto_fit_columns(ws, style, header_row=header_row)
    _apply_column_groups(ws, df, style)
    _apply_table_format(ws, df, style, header_row=header_row)


def generar_excel(
    df: pd.DataFrame,
    nombre_archivo: str,
    sheet_name: str = "Datos",
    style: SheetStyle | None = None
) -> Path:
    """
    Genera archivo Excel con formato de tabla.

    Args:
        df: DataFrame procesado con el formato de reporte
        nombre_archivo: Nombre del archivo sin extension
        sheet_name: Nombre de la hoja (default: "Datos")
        style: Estilo personalizado. Si es None, usa DEFAULT_STYLE.

    Returns:
        Path del archivo generado
    """
    style = style or DEFAULT_STYLE
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_sheet(ws, df, style)

    # Guardar archivo
    DATA_OUTPUT.mkdir(parents=True, exist_ok=True)
    ruta_archivo = DATA_OUTPUT / f"{nombre_archivo}.xlsx"
    wb.save(ruta_archivo)

    return ruta_archivo


class ExcelWriter:
    """
    Clase para generar Excel con mayor control.

    Permite configurar estilos, multiples hojas, etc.
    """

    def __init__(self, nombre_archivo: str, output_dir: Path | None = None, style: SheetStyle | None = None):
        self.nombre_archivo = nombre_archivo
        self.output_dir = output_dir or DATA_OUTPUT
        self.default_style = style or DEFAULT_STYLE
        self.workbook = Workbook()
        self._first_sheet = True

    def add_sheet(self, df: pd.DataFrame, sheet_name: str = "Datos", style: SheetStyle | None = None) -> None:
        """Agrega una hoja al workbook con formato."""
        style = style or self.default_style

        if self._first_sheet:
            ws = self.workbook.active
            ws.title = sheet_name
            self._first_sheet = False
        else:
            ws = self.workbook.create_sheet(title=sheet_name)

        _write_sheet(ws, df, style)

    def save(self) -> Path:
        """Guarda el archivo Excel."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        ruta_archivo = self.output_dir / f"{self.nombre_archivo}.xlsx"
        self.workbook.save(ruta_archivo)
        return ruta_archivo
