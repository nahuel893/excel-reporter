"""
RangeRecognizer — detects maximal rectangular regions bounded by
medium-weight openpyxl borders in a worksheet.

Used to auto-detect "card" regions (e.g. per-supervisor or per-zone blocks)
in report templates without hand-maintaining a list of fixed A1 ranges.

Algorithm summary
------------------
1. Build two boolean wall grids from the sheet's borders:
   - ``h_wall[r][c]``: True if there is a wall segment ABOVE row ``r`` at
     column ``c`` (i.e. between row r-1 and row r).
   - ``v_wall[r][c]``: True if there is a wall segment LEFT of column ``c``
     at row ``r`` (i.e. between column c-1 and column c).
   Each side is resolved with the "adjacency-OR" rule: a wall may be
   declared on EITHER of the two neighboring cells that share that edge
   (matches how ``ExcelManager.detect_bordered_range`` reads borders).
2. Enumerate every rectangle whose four sides are fully continuous walls
   ("closed" rectangles), anchored at every wall-corner in the sheet.
3. Reduce to MAXIMAL rectangles only: a closed rectangle is discarded if it
   is properly contained within any OTHER strictly larger closed rectangle
   from the same candidate set. This keeps outer card outlines while
   discarding nested sub-rectangles formed by interior gridlines/mini-widget
   grids — however many levels or an irregular shape they're subdivided
   into — since the outer block is itself always found as its own closed
   rectangle and strictly contains every one of its interior pieces.
   (See "Known limitation" below for the one case this simplification does
   not perfectly resolve.)
4. Sort surviving rectangles in reading order: top row ascending, then left
   column ascending.

Known limitation
-----------------
Containment-based maximality cannot always distinguish "two independent,
touching regions whose union happens to also be a valid closed rectangle"
from "one region internally subdivided by gridlines/mini-widgets" — both
produce a larger closed rectangle strictly containing smaller closed
rectangles, and by design the larger one always wins. In the target report
templates this matches the desired behavior (internal card substructure is
always meant to collapse into its outer card), and genuinely distinct
sibling cards are always separated by a gap column/row (breaking wall
continuity, so no accidental union can close). If a future template ever
places two DISTINCT regions directly touching with no gap AND their
combined bounding box also happens to be a valid closed rectangle, this
recognizer will report the union instead of the two pieces — a known,
documented edge case, not silently swallowed.

Requires a NON read-only ``openpyxl`` load — ``read_only=True`` does not
reliably expose ``cell.border``.
"""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_BORDER_STYLES = ("medium", "thick", "double")

# (row, col) -> (min_row, min_col, max_row, max_col) of the merged range
# that covers that cell.
MergeIndex = dict[tuple[int, int], tuple[int, int, int, int]]

# 1-based inclusive rectangle: (row_start, row_end, col_start, col_end)
RectTuple = tuple[int, int, int, int]


def _is_wall(side, border_styles: tuple[str, ...]) -> bool:
    """True if an openpyxl Border Side is a wall-worthy border weight."""
    if side is None:
        return False
    return getattr(side, "style", None) in border_styles


def _build_merge_index(ws) -> MergeIndex:
    """Maps every (row, col) covered by a merged range to the range's
    (min_row, min_col, max_row, max_col) bounds."""
    index: MergeIndex = {}
    for merged_range in ws.merged_cells.ranges:
        bounds = (
            merged_range.min_row, merged_range.min_col,
            merged_range.max_row, merged_range.max_col,
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                index[(row, col)] = bounds
    return index


def _effective_side(
    ws, merge_index: MergeIndex, row: int, col: int, side_name: str,
    border_styles: tuple[str, ...],
) -> bool:
    """Resolves whether the given side of cell (row, col) is a wall.

    Routes through the merge index when the cell belongs to a merged range:
    if the cell itself carries no border for that side, but the side is on
    the OUTER boundary of its merge, the merge's top-left anchor cell's
    border is used instead (Excel/openpyxl commonly stores the border only
    on the merge's anchor cell)."""
    cell = ws.cell(row=row, column=col)
    if _is_wall(getattr(cell.border, side_name, None), border_styles):
        return True

    bounds = merge_index.get((row, col))
    if bounds is None:
        return False

    min_row, min_col, max_row, max_col = bounds
    is_boundary = (
        (side_name == "top" and row == min_row)
        or (side_name == "bottom" and row == max_row)
        or (side_name == "left" and col == min_col)
        or (side_name == "right" and col == max_col)
    )
    if not is_boundary:
        return False

    anchor = ws.cell(row=min_row, column=min_col)
    return _is_wall(getattr(anchor.border, side_name, None), border_styles)


def _to_a1(rect: RectTuple) -> str:
    r1, r2, c1, c2 = rect
    return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"


class RangeRecognizer:
    """Detects maximal medium-bordered rectangular regions in a worksheet."""

    def __init__(self, xlsx_path: Path, border_styles: tuple[str, ...] = DEFAULT_BORDER_STYLES):
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"File not found: {xlsx_path}")
        self.xlsx_path = xlsx_path
        self.border_styles = border_styles
        self._workbook = None  # lazily loaded and cached — non-read_only loads
        # of large multi-sheet workbooks are expensive; reuse it across calls
        # on the same instance instead of reloading per detect_ranges() call.

    def _get_workbook(self):
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(self.xlsx_path, read_only=False, data_only=False)
        return self._workbook

    def close(self) -> None:
        """Releases the cached workbook, if one was loaded."""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def __enter__(self) -> "RangeRecognizer":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def detect_ranges(self, sheet: str) -> list[str]:
        """Returns the maximal bordered regions of ``sheet`` as A1 range
        strings, in reading order (top-to-bottom, then left-to-right)."""
        wb = self._get_workbook()
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {self.xlsx_path.name}")
        return self._detect_ranges_in_sheet(wb[sheet])

    def detect_all(self) -> dict[str, list[str]]:
        """Returns detected regions for every sheet in the workbook."""
        wb = self._get_workbook()
        return {name: self._detect_ranges_in_sheet(wb[name]) for name in wb.sheetnames}

    # ── Core algorithm ────────────────────────────────────────────────────

    def _detect_ranges_in_sheet(self, ws) -> list[str]:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return []

        merge_index = _build_merge_index(ws)
        h_wall = self._build_h_wall(ws, merge_index, max_row, max_col)
        v_wall = self._build_v_wall(ws, merge_index, max_row, max_col)

        h_reach = self._compute_h_reach(h_wall, max_row, max_col)
        v_reach = self._compute_v_reach(v_wall, max_row, max_col)

        closed_rects = self._find_closed_rectangles(h_wall, v_wall, h_reach, v_reach, max_row, max_col)
        maximal_rects = self._filter_maximal(closed_rects)
        maximal_rects.sort(key=lambda rect: (rect[0], rect[2]))  # (row_start, col_start)
        return [_to_a1(rect) for rect in maximal_rects]

    def _build_h_wall(self, ws, merge_index: MergeIndex, max_row: int, max_col: int) -> dict[int, dict[int, bool]]:
        """h_wall[r][c]: wall segment ABOVE row r at column c, for r in
        1..max_row+1 (r = max_row+1 represents the boundary below the last
        row)."""
        h_wall: dict[int, dict[int, bool]] = {}
        for r in range(1, max_row + 2):
            row_walls: dict[int, bool] = {}
            for c in range(1, max_col + 1):
                wall = False
                if r <= max_row:
                    wall = wall or _effective_side(ws, merge_index, r, c, "top", self.border_styles)
                if r > 1:
                    wall = wall or _effective_side(ws, merge_index, r - 1, c, "bottom", self.border_styles)
                row_walls[c] = wall
            h_wall[r] = row_walls
        return h_wall

    def _build_v_wall(self, ws, merge_index: MergeIndex, max_row: int, max_col: int) -> dict[int, dict[int, bool]]:
        """v_wall[r][c]: wall segment LEFT of column c at row r, for c in
        1..max_col+1 (c = max_col+1 represents the boundary right of the
        last column)."""
        v_wall: dict[int, dict[int, bool]] = {}
        for r in range(1, max_row + 1):
            row_walls: dict[int, bool] = {}
            for c in range(1, max_col + 2):
                wall = False
                if c <= max_col:
                    wall = wall or _effective_side(ws, merge_index, r, c, "left", self.border_styles)
                if c > 1:
                    wall = wall or _effective_side(ws, merge_index, r, c - 1, "right", self.border_styles)
                row_walls[c] = wall
            v_wall[r] = row_walls
        return v_wall

    @staticmethod
    def _compute_h_reach(h_wall: dict[int, dict[int, bool]], max_row: int, max_col: int) -> dict[int, dict[int, int]]:
        """h_reach[r][c] = furthest column c2 >= c such that h_wall[r][k] is
        True for every k in [c, c2]. If h_wall[r][c] is False, h_reach[r][c]
        is set to c-1 (an invalid/empty marker)."""
        h_reach: dict[int, dict[int, int]] = {}
        for r in range(1, max_row + 2):
            row_wall = h_wall[r]
            row_reach: dict[int, int] = {}
            for c in range(max_col, 0, -1):
                if row_wall.get(c, False):
                    if c < max_col and row_wall.get(c + 1, False):
                        row_reach[c] = row_reach[c + 1]
                    else:
                        row_reach[c] = c
                else:
                    row_reach[c] = c - 1
            h_reach[r] = row_reach
        return h_reach

    @staticmethod
    def _compute_v_reach(v_wall: dict[int, dict[int, bool]], max_row: int, max_col: int) -> dict[int, dict[int, int]]:
        """v_reach[c][r] = furthest row r2 >= r such that v_wall[k][c] is
        True for every k in [r, r2]. If v_wall[r][c] is False, v_reach[c][r]
        is set to r-1 (an invalid/empty marker)."""
        v_reach: dict[int, dict[int, int]] = {}
        for c in range(1, max_col + 2):
            col_reach: dict[int, int] = {}
            for r in range(max_row, 0, -1):
                if v_wall[r].get(c, False):
                    if r < max_row and v_wall[r + 1].get(c, False):
                        col_reach[r] = col_reach[r + 1]
                    else:
                        col_reach[r] = r
                else:
                    col_reach[r] = r - 1
            v_reach[c] = col_reach
        return v_reach

    @staticmethod
    def _find_closed_rectangles(
        h_wall: dict[int, dict[int, bool]], v_wall: dict[int, dict[int, bool]],
        h_reach: dict[int, dict[int, int]], v_reach: dict[int, dict[int, int]],
        max_row: int, max_col: int,
    ) -> list[RectTuple]:
        """Enumerates every rectangle whose four sides are fully continuous
        walls, anchored at every valid top-left wall-corner in the grid."""
        rects: list[RectTuple] = []
        for r1 in range(1, max_row + 1):
            for c1 in range(1, max_col + 1):
                if not h_wall[r1].get(c1, False) or not v_wall[r1].get(c1, False):
                    continue  # not a top-left wall corner

                max_c2 = h_reach[r1].get(c1, c1 - 1)
                max_r2 = v_reach[c1].get(r1, r1 - 1)
                if max_c2 < c1 or max_r2 < r1:
                    continue

                for r2 in range(r1, max_r2 + 1):
                    bottom_reach = h_reach.get(r2 + 1, {}).get(c1, c1 - 1)
                    upper_c2 = min(max_c2, bottom_reach)
                    if upper_c2 < c1:
                        continue
                    for c2 in range(c1, upper_c2 + 1):
                        right_reach = v_reach.get(c2 + 1, {}).get(r1, r1 - 1)
                        if right_reach >= r2:
                            rects.append((r1, r2, c1, c2))
        return rects

    @staticmethod
    def _filter_maximal(rects: list[RectTuple]) -> list[RectTuple]:
        """Discards any closed rectangle that is properly contained within
        another, strictly larger closed rectangle from the same candidate
        set. A rectangle survives only if no other candidate's bounding box
        fully covers it."""
        maximal: list[RectTuple] = []
        for rect in rects:
            r1, r2, c1, c2 = rect
            contained_in_larger = any(
                other != rect
                and other[0] <= r1 and other[1] >= r2
                and other[2] <= c1 and other[3] >= c2
                for other in rects
            )
            if not contained_in_larger:
                maximal.append(rect)
        return maximal
