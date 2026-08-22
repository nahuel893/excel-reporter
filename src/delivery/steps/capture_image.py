"""
CaptureImageStep - Paso de captura de rango Excel como imagen PNG.
"""
import logging
from pathlib import Path

from src.delivery.pipeline import CaptureConfig, DeliveryConfig, DeliveryStep, ReportArtifact, StepResult

AUTO_BORDES_SENTINEL = "auto:bordes"
AUTO_HOJA_SENTINEL = "auto:hoja"
AUTO_BORDES_RENDERER = "libreoffice"  # only backend that supports print_area cropping


class CaptureImageStep(DeliveryStep):
    """Captura N rangos de hojas Excel como imagenes PNG (LibreOffice + Pillow).

    Cada captura es aislada: si una falla, las otras se intentan igual.
    Poblar artifact.rutas_imagenes con la lista de PNGs generados.

    Soporta el sentinel ``rango: "auto:bordes"``: se expande, ANTES del loop
    de render, en N CaptureConfig concretos (uno por region detectada por
    RangeRecognizer), cada uno renderizado con crop=True. Nunca muta
    `config`/`DeliveryConfig` — la expansion arma una lista local nueva.

    Las entradas con rango fijo (no ``auto:bordes``) se renderizan con
    crop=``cfg.recortar`` (default False, comportamiento historico de
    branca/schneider preservado); las expandidas de ``auto:bordes`` siempre
    usan crop=True.

    Gate de consumo: si ningun canal de entrega va a CONSUMIR las imagenes
    (ver ``_images_consumed``), el paso se omite inmediatamente, ANTES de
    ``_expand_auto_bordes`` (la costosa carga de workbook de RangeRecognizer)
    y antes de cualquier render — evita gastar tiempo de render en imagenes
    que nadie va a ver.

    Ruteo por lote (``render_many``): si la lista expandida tiene 2+
    entradas, TODAS con renderer 'libreoffice', y ese renderer expone
    ``render_many``, se llama UNA sola vez con todos los specs en lugar de
    iterar `render()` por-item — recalcula el workbook una sola vez (ver
    ``ExcelManager.capture_ranges``), probado pixel-identico al loop
    per-item. Una unica captura 'libreoffice', renderers mixtos (p.ej.
    'html_playwright'), o un renderer sin `render_many`, siguen usando el
    loop per-item original sin cambios.
    """

    def execute(
        self,
        artifact: ReportArtifact,
        config: DeliveryConfig,
        logger: logging.Logger,
    ) -> StepResult:
        # Support either plural capture_images or legacy singular capture_image.
        captures = list(config.capture_images or [])
        if not captures and config.capture_image is not None:
            captures = [config.capture_image]

        if not captures:
            return StepResult(
                status="skipped",
                step_name="CaptureImageStep",
                message="capture_images no configurado",
            )

        if not self._images_consumed(config):
            return StepResult(
                status="skipped",
                step_name="CaptureImageStep",
                message=(
                    "Sin canal que consuma imagenes (whatsapp off, email sin "
                    "adjunto imagen); captura omitida"
                ),
            )

        from src.core.excel_renderers import get_renderer

        errores: list[str] = []
        expanded = self._expand_auto_bordes(artifact, captures, errores, logger)

        if not expanded:
            if errores:
                return StepResult(
                    status="error",
                    step_name="CaptureImageStep",
                    message=f"Todas las capturas fallaron: {'; '.join(errores)}",
                )
            # Reaching here means captures WERE configured but every auto:bordes
            # sentinel detected zero regions (e.g. the template lost its border
            # styling). Report it distinctly — naming the sheet(s) — instead of
            # the misleading "no configurado" so an operator can investigate.
            auto_sheets = [c.hoja for c in captures if c.rango == AUTO_BORDES_SENTINEL]
            if auto_sheets:
                sheets_str = ", ".join(f"'{s}'" for s in auto_sheets)
                return StepResult(
                    status="error",
                    step_name="CaptureImageStep",
                    message=f"auto:bordes: no bordered regions detected in sheet {sheets_str}",
                )
            return StepResult(
                status="skipped",
                step_name="CaptureImageStep",
                message="capture_images no configurado",
            )

        produced: list[str] = []

        batch_renderer = self._resolve_batch_renderer(expanded, get_renderer)

        if batch_renderer is not None:
            renderer_name = "libreoffice"
            specs = [(cfg.hoja, cfg.rango, crop) for cfg, crop in expanded]
            try:
                results = batch_renderer.render_many(
                    xlsx_path=artifact.ruta_excel,
                    specs=specs,
                    output_dir=artifact.ruta_excel.parent,
                )
            except ImportError as exc:
                # A genuinely missing renderer dependency (e.g. Pillow) affects
                # every spec equally — there is no point continuing.
                logger.warning(
                    "Dependencia faltante en renderer '%s' (lote), omitiendo capturas: %s",
                    renderer_name, exc,
                )
                return StepResult(
                    status="skipped",
                    step_name="CaptureImageStep",
                    message=str(exc),
                )
            except Exception as exc:
                # A batch-wide failure (e.g. the shared recalc itself failed)
                # affects every spec in this batch equally — mirror what would
                # happen if each spec's own recalc failed individually in the
                # per-item path (every spec ends up as its own error entry).
                logger.error(
                    "render_many fallo antes de producir resultados: %s", exc,
                )
                for cfg, _crop in expanded:
                    errores.append(f"{cfg.hoja}[{cfg.rango}]/{renderer_name}: {exc}")
            else:
                for (cfg, _crop), result in zip(expanded, results):
                    if isinstance(result, Exception):
                        logger.error(
                            "Captura fallida [%s %s/%s]: %s",
                            cfg.hoja, cfg.rango, renderer_name, result,
                        )
                        errores.append(f"{cfg.hoja}[{cfg.rango}]/{renderer_name}: {result}")
                    else:
                        artifact.rutas_imagenes.append(result)
                        artifact.nombres_hojas.append(cfg.caption or cfg.hoja)
                        produced.append(f"{cfg.hoja}[{renderer_name}]:{result.name}")
        else:
            for cfg, crop in expanded:
                renderer_name = getattr(cfg, "renderer", "libreoffice")
                try:
                    renderer = get_renderer(renderer_name)
                    png_path = renderer.render(
                        xlsx_path=artifact.ruta_excel,
                        sheet=cfg.hoja,
                        range_addr=cfg.rango,
                        output_dir=artifact.ruta_excel.parent,
                        crop=crop,
                    )
                    artifact.rutas_imagenes.append(png_path)
                    artifact.nombres_hojas.append(cfg.caption or cfg.hoja)
                    produced.append(f"{cfg.hoja}[{renderer_name}]:{png_path.name}")
                except ImportError as exc:
                    # A genuinely missing renderer dependency (e.g. Pillow) affects
                    # every region equally — there is no point continuing the loop.
                    logger.warning(
                        "Dependencia faltante en renderer '%s', omitiendo capturas: %s",
                        renderer_name, exc,
                    )
                    return StepResult(
                        status="skipped",
                        step_name="CaptureImageStep",
                        message=str(exc),
                    )
                except Exception as exc:  # per-region failure (incl. RuntimeError)
                    # RuntimeError from soffice/pdftoppm (non-zero exit, missing
                    # binary) is per-render: isolate it so one bad region does not
                    # abort the remaining regions. Record it and keep going.
                    logger.error(
                        "Captura fallida [%s %s/%s]: %s",
                        cfg.hoja, cfg.rango, renderer_name, exc,
                    )
                    errores.append(f"{cfg.hoja}[{cfg.rango}]/{renderer_name}: {exc}")

        if produced and not errores:
            return StepResult(
                status="success",
                step_name="CaptureImageStep",
                message=f"Imagenes generadas: {', '.join(produced)}",
                artifact_path=artifact.rutas_imagenes[0] if artifact.rutas_imagenes else None,
            )
        if produced and errores:
            return StepResult(
                status="partial",
                step_name="CaptureImageStep",
                message=(
                    f"Exito: {', '.join(produced)} | "
                    f"Fallos: {'; '.join(errores)}"
                ),
                artifact_path=artifact.rutas_imagenes[0] if artifact.rutas_imagenes else None,
            )
        return StepResult(
            status="error",
            step_name="CaptureImageStep",
            message=f"Todas las capturas fallaron: {'; '.join(errores)}",
        )

    @staticmethod
    def _images_consumed(config: DeliveryConfig) -> bool:
        """True iff at least one delivery channel will actually consume the
        rendered images:
        - WhatsApp is configured AND enviar_como is 'imagen' or 'ambos', OR
        - Email is configured AND 'imagen' is in adjuntos.

        Used to gate the expensive render/expand path (RangeRecognizer
        workbook load + LibreOffice renders per region) so it never runs
        when nothing downstream will consume the output — e.g. WhatsApp
        disabled/archivo-only and email attaching only the excel file."""
        whatsapp_consumes = (
            config.whatsapp is not None
            and config.whatsapp.enviar_como in ("imagen", "ambos")
        )
        email_consumes = (
            config.email is not None
            and "imagen" in config.email.adjuntos
        )
        return whatsapp_consumes or email_consumes

    @staticmethod
    def _resolve_batch_renderer(expanded: list[tuple[CaptureConfig, bool]], get_renderer):
        """Returns the renderer instance to use for a SINGLE batched
        `render_many` call, or None to fall back to the per-item loop.

        Eligible only when: there are 2+ expanded entries (batching a
        single item has no recalc-amortization benefit — see
        ExcelManager.capture_ranges), EVERY entry uses renderer
        'libreoffice', and the resolved renderer actually exposes
        `render_many`. Anything else (mixed renderer types, a single
        capture, or a renderer without `render_many`) returns None."""
        if len(expanded) <= 1:
            return None
        renderer_names = {getattr(cfg, "renderer", "libreoffice") for cfg, _crop in expanded}
        if renderer_names != {"libreoffice"}:
            return None
        candidate = get_renderer("libreoffice")
        if not hasattr(candidate, "render_many"):
            return None
        return candidate

    @staticmethod
    def _rango_usado(
        artifact: ReportArtifact,
        hoja: str,
        errores: list[str],
        logger: logging.Logger,
    ) -> str | None:
        """Rango A1 que ocupa la hoja entera, o ``None`` si no se pudo leer.

        Para informes cuyo alto NO se conoce al escribir el config: el cuadro
        de cobertura por calibre crece o se achica con los envases que se
        vendieron ese mes, asi que un rango fijo recorta filas algunos meses y
        deja franjas vacias otros.

        Es mucho mas barato que ``auto:bordes``: abre el workbook en modo
        read_only y lee el rango usado, sin recorrer bordes celda por celda.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(artifact.ruta_excel, read_only=True)
            try:
                if hoja not in wb.sheetnames:
                    raise KeyError(f"la hoja '{hoja}' no existe en el archivo")
                return wb[hoja].calculate_dimension()
            finally:
                wb.close()
        except Exception as exc:
            logger.error(
                "Fallo resolviendo '%s' en hoja '%s': %s",
                AUTO_HOJA_SENTINEL, hoja, exc,
            )
            errores.append(f"{hoja}/{AUTO_HOJA_SENTINEL}: {exc}")
            return None

    @staticmethod
    def _expand_auto_bordes(
        artifact: ReportArtifact,
        captures: list[CaptureConfig],
        errores: list[str],
        logger: logging.Logger,
    ) -> list[tuple[CaptureConfig, bool]]:
        """Expands any ``rango == "auto:bordes"`` entry into one concrete
        CaptureConfig per detected region (crop=True), leaving non-sentinel
        entries untouched (crop=``cfg.recortar``, default False). Builds a
        NEW local list — `captures` (and therefore the shared DeliveryConfig
        it came from) is never mutated. Reuses a SINGLE RangeRecognizer
        instance across every auto:bordes entry in this run (the underlying
        workbook load is expensive, ~110-125s on real templates — see
        range_recognizer.py)."""
        expanded: list[tuple[CaptureConfig, bool]] = []
        recognizer = None
        try:
            for cfg in captures:
                if cfg.rango == AUTO_HOJA_SENTINEL:
                    rango = CaptureImageStep._rango_usado(
                        artifact, cfg.hoja, errores, logger
                    )
                    if rango is not None:
                        expanded.append((cfg.model_copy(update={"rango": rango}), True))
                    continue

                if cfg.rango != AUTO_BORDES_SENTINEL:
                    expanded.append((cfg, cfg.recortar))
                    continue

                renderer_name = getattr(cfg, "renderer", "libreoffice")
                if renderer_name != AUTO_BORDES_RENDERER:
                    msg = (
                        f"'{AUTO_BORDES_SENTINEL}' solo soportado con "
                        f"renderer='{AUTO_BORDES_RENDERER}', recibido "
                        f"'{renderer_name}' para hoja '{cfg.hoja}'"
                    )
                    logger.error(msg)
                    errores.append(f"{cfg.hoja}/{renderer_name}: {msg}")
                    continue

                if recognizer is None:
                    from src.core.range_recognizer import RangeRecognizer
                    recognizer = RangeRecognizer(artifact.ruta_excel)

                try:
                    regions = recognizer.detect_ranges_with_captions(
                        cfg.hoja, caption_anchor=cfg.caption_anchor, caption_header=cfg.caption_header,
                    )
                except Exception as exc:
                    logger.error(
                        "Fallo detectando regiones '%s' en hoja '%s': %s",
                        AUTO_BORDES_SENTINEL, cfg.hoja, exc,
                    )
                    errores.append(f"{cfg.hoja}/{AUTO_BORDES_SENTINEL}: {exc}")
                    continue

                for a1_range, region_caption in regions:
                    concrete = cfg.model_copy(update={
                        "rango": a1_range,
                        "caption": region_caption or cfg.caption,
                    })
                    expanded.append((concrete, True))
        finally:
            if recognizer is not None:
                recognizer.close()

        return expanded
