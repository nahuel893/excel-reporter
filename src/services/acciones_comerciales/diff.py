"""Parallel diff harness (RF-12, Decision 1 + Decision 14).

During the Phase-1 parallel-run window (Decision 7) this compares the new
Python BASE pivot frames against the manual backup engine/informe pivots
(``data/backups/acciones-comerciales-<fecha>/``) and validates the RF-01
terna->precio pick rule empirically against the real ``aexcel.xlsx`` file.

Three guarantees, each mirrored by a test:

  1. **Like-for-like period scope** — the backup engine table accumulates
     rows across a wider window than the BASE run; ``filter_period_scope``
     drops out-of-scope rows (by their ``Fecha`` field) so they are neither
     compared nor reported as missing/extra (RF-12).
  2. **Exact $0.01 comparison** — a delta of a cent or less is a match;
     anything greater is surfaced. The tolerance is a COMPARISON mechanism
     only: no underlying float is ever rounded/truncated (RF-23).
  3. **Classified deltas** — every surfaced delta is tagged EXACTLY one of
     ``baseline-defect`` (attributable to a specific known manual-flow bug —
     stale BD:BE / BG:BH snapshots, the 6-row "es CCU?" map, tabla_control
     AZ/AX drift, evidence supplied via ``KnownDefectContext``) or
     ``real-divergence`` (unexplained — surfaced, never absorbed by a
     tolerance margin).

Classification is EVIDENCE-DRIVEN, not guessed: the ``KnownDefectContext``
carries the sets of clients/ternas/genéricos/columns proven during the
parallel run to sit in a known-defect zone. Anything a predicate cannot
explain stays ``real-divergence`` (Decision 1). The context is optionally
loaded from a ``known_defects.json`` file next to the backups, so the analyst
grows the evidence set across the sign-off month without a code change.

The terna->precio validator (``validate_terna_precios``) compares the
generated pick against the real aexcel file and SURFACES any precio/Bonific
mismatch (Decision 14 — the pick rule is not assumed correct a priori; it is
adjusted from this evidence).

Outputs (``write_diff_report``): a machine-readable JSON, an xlsx workbook
(each sheet ending in a distinctly-styled TOTAL GENERAL row per the project
rule), and a human-readable ``.txt`` summary for sign-off review.
"""
from __future__ import annotations

import json
import logging
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.services.acciones_comerciales.constants import (
    ACC_GEN_ROW_FIELDS,
    ART_ACCION_ROW_FIELDS,
    CLIENTE_FECHA_ROW_FIELDS,
    FACT_NET_ROW_FIELDS,
    GENERICOS_ORDER,
    LABEL_SUMA_CAMPO1,
    LABEL_SUMA_DESCUENTO,
    LABEL_SUMA_DESCUENTOS,
    LABEL_SUMA_FACT_NETA,
)

logger = logging.getLogger(__name__)

DEFAULT_TOLERANCE = 0.01  # exact-to-$0.01 comparison discipline (RF-12)
TOTAL_GENERAL_LABEL = "TOTAL GENERAL"
_TOTAL_FILL_COLOR = "FFE08A"  # amber — project convention for totals rows

BASELINE_DEFECT = "baseline-defect"
REAL_DIVERGENCE = "real-divergence"

# Row-field (key) tuple per pivot — matches pivots.py output column order.
KEY_FIELDS_BY_PIVOT: dict[str, list[str]] = {
    "FACT_NET": FACT_NET_ROW_FIELDS,
    "ART-ACCION": ART_ACCION_ROW_FIELDS,
    "CLIENTE-FECHA": CLIENTE_FECHA_ROW_FIELDS,
    "ACC-GEN": ACC_GEN_ROW_FIELDS,
}

# Value/measure columns compared per pivot.
VALUE_COLUMNS_BY_PIVOT: dict[str, list[str]] = {
    "FACT_NET": [LABEL_SUMA_FACT_NETA, LABEL_SUMA_DESCUENTOS, LABEL_SUMA_CAMPO1],
    "ART-ACCION": [LABEL_SUMA_DESCUENTO],
    "CLIENTE-FECHA": [LABEL_SUMA_DESCUENTO],
    "ACC-GEN": GENERICOS_ORDER,
}

# The field carrying a per-row date, when the pivot has one (period scope).
_FECHA_FIELD = "Fecha"

# Canonical terna columns of the generated / aexcel frames.
_TERNA_FECHA = "Descripción Período"
_TERNA_CLIENTE = "Cod. Cliente"
_TERNA_ARTICULO = "Código"
_TERNA_PRECIO = "Precio"
_TERNA_BONIFIC = "Bonific"
_AEXCEL_EXPORT_COLUMNS = [_TERNA_FECHA, _TERNA_CLIENTE, _TERNA_ARTICULO, _TERNA_PRECIO, _TERNA_BONIFIC]

# ── reason strings — each NAMES the specific known manual-flow bug ──
_REASON_BG_BH = (
    "baseline-defect: stale BG:BH SUCURSAL snapshot (manual map froze at row "
    "72,759); client added after the freeze is not covered by the backup"
)
_REASON_BD_BE = (
    "baseline-defect: stale BD:BE PRECIO snapshot (#N/A after freeze); terna "
    "priced after the manual snapshot, so the backup Descuento is wrong"
)
_REASON_ES_CCU = (
    "baseline-defect: 6-row 'es CCU?' map gap; genérico/marca not classified "
    "by the manual map, so the backup genérico column is wrong"
)
_REASON_AZ_AX = (
    "baseline-defect: tabla_control AZ/AX column drift (backup summed the "
    "unit-price / es-CCU? text-flag column instead of Facturacion Neta / "
    "Descuentos)"
)
_REASON_REAL = "real-divergence: unexplained delta — no known baseline-bug matches; needs investigation"


# ═════════════════════════════════════════════════════════════════════════
# dataclasses
# ═════════════════════════════════════════════════════════════════════════


@dataclass(frozen=True)
class KnownDefectContext:
    """Evidence sets that let a delta be attributed to a specific known
    manual-flow bug. Empty by default — with no evidence, every delta stays
    ``real-divergence`` (Decision 1). The analyst grows these sets across the
    parallel-run month (via ``known_defects.json``)."""

    stale_sucursal_clients: frozenset = frozenset()  # BG:BH snapshot gap
    stale_precio_clients: frozenset = frozenset()  # BD:BE snapshot gap (by client)
    stale_precio_ternas: frozenset = frozenset()  # BD:BE snapshot gap (by terna)
    es_ccu_defect_generics: frozenset = frozenset()  # 6-row 'es CCU?' map gap
    az_ax_drift_columns: frozenset = frozenset()  # tabla_control AZ/AX drift


@dataclass
class DiffRow:
    """One surfaced per-cell delta (base vs backup) with its classification."""

    pivot: str
    key: tuple
    key_fields: tuple
    column: str
    base_value: float | None
    backup_value: float | None
    delta: float
    presence: str  # 'both' | 'base-only' | 'backup-only'
    classification: str
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "pivot": self.pivot,
            "key": [_json_safe(v) for v in self.key],
            "key_fields": list(self.key_fields),
            "column": self.column,
            "base_value": _json_safe(self.base_value),
            "backup_value": _json_safe(self.backup_value),
            "delta": _json_safe(self.delta),
            "presence": self.presence,
            "classification": self.classification,
            "reason": self.reason,
        }


@dataclass
class TernaPriceMismatch:
    """A generated terna->precio (or ->Bonific) pick that disagrees with the
    real aexcel file (Decision 14) — surfaced, never silently accepted."""

    fecha: Any
    id_cliente: Any
    id_articulo: Any
    generated_precio: float | None
    aexcel_precio: float | None
    generated_bonific: float | None
    aexcel_bonific: float | None
    kind: str  # 'precio' | 'bonific' | 'missing-in-aexcel'

    def to_dict(self) -> dict[str, Any]:
        return {k: _json_safe(v) for k, v in asdict(self).items()}


@dataclass
class DiffReport:
    """Full diff result: classified deltas + terna->precio mismatches."""

    rows: list[DiffRow] = field(default_factory=list)
    terna_mismatches: list[TernaPriceMismatch] = field(default_factory=list)
    fecha_desde: str | None = None
    fecha_hasta: str | None = None
    tolerance: float = DEFAULT_TOLERANCE

    @property
    def baseline_defects(self) -> list[DiffRow]:
        return [r for r in self.rows if r.classification == BASELINE_DEFECT]

    @property
    def real_divergences(self) -> list[DiffRow]:
        return [r for r in self.rows if r.classification == REAL_DIVERGENCE]

    def summary_counts(self) -> dict[str, int]:
        return {
            "total_deltas": len(self.rows),
            "baseline_defect": len(self.baseline_defects),
            "real_divergence": len(self.real_divergences),
            "terna_mismatches": len(self.terna_mismatches),
        }

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "fecha_desde": self.fecha_desde,
            "fecha_hasta": self.fecha_hasta,
            "tolerance": self.tolerance,
            "summary": self.summary_counts(),
            "rows": [r.to_dict() for r in self.rows],
            "terna_mismatches": [m.to_dict() for m in self.terna_mismatches],
        }

    def to_summary_text(self) -> str:
        counts = self.summary_counts()
        lines = [
            "ACCIONES COMERCIALES — parallel diff summary",
            f"period scope: {self.fecha_desde} .. {self.fecha_hasta}  (tolerance ${self.tolerance})",
            "",
            f"total surfaced deltas : {counts['total_deltas']}",
            f"  baseline-defect     : {counts['baseline_defect']}  (explained by a known manual-flow bug)",
            f"  real-divergence     : {counts['real_divergence']}  (UNEXPLAINED — investigate before sign-off)",
            f"terna->precio picks disagreeing with the real aexcel: {counts['terna_mismatches']}",
        ]
        if self.real_divergences:
            lines += ["", "REAL DIVERGENCES (investigate):"]
            for r in self.real_divergences:
                lines.append(f"  [{r.pivot}] {r.key} · {r.column}: base={r.base_value} backup={r.backup_value} Δ={r.delta}")
        if self.terna_mismatches:
            lines += ["", "TERNA PICK MISMATCHES vs real aexcel:"]
            for m in self.terna_mismatches:
                lines.append(
                    f"  ({m.fecha}, {m.id_cliente}, {m.id_articulo}) [{m.kind}] "
                    f"gen precio={m.generated_precio}/bonif={m.generated_bonific} "
                    f"aexcel precio={m.aexcel_precio}/bonif={m.aexcel_bonific}"
                )
        return "\n".join(lines) + "\n"


# ═════════════════════════════════════════════════════════════════════════
# small helpers
# ═════════════════════════════════════════════════════════════════════════


def _json_safe(value: Any) -> Any:
    """Coerce a value to a JSON-serialisable primitive WITHOUT rounding: NaN
    and None become ``null``; numpy scalars become native floats via a
    (loss-free) float cast; everything else passes through as-is."""
    if value is None:
        return None
    if isinstance(value, str):
        return value
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        return value
    if isinstance(value, (int,)) and not isinstance(value, bool):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _cell_number(value: Any) -> float | None:
    """A cell's numeric value, or None for a blank/non-numeric cell. Never
    rounds — returns the value verbatim as a float."""
    if value is None or value == "":
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _iso_fecha(series: pd.Series) -> pd.Series:
    """Normalize a date-ish column to ISO ``YYYY-MM-DD`` strings for stable
    terna/period comparison across dtypes."""
    return pd.to_datetime(series, errors="coerce").dt.strftime("%Y-%m-%d")


def strip_total_general(df: pd.DataFrame) -> pd.DataFrame:
    """Drop a pivots.py-style TOTAL GENERAL grand-total row (identified by the
    label in the first column) so it never contaminates the key set."""
    if df.empty:
        return df
    first_col = df.columns[0]
    mask = df[first_col].astype(str) != TOTAL_GENERAL_LABEL
    return df[mask].reset_index(drop=True)


def filter_period_scope(
    df: pd.DataFrame, fecha_col: str, fecha_desde: str | None, fecha_hasta: str | None
) -> pd.DataFrame:
    """Keep only rows whose ``fecha_col`` falls within
    ``[fecha_desde, fecha_hasta]`` (inclusive). Out-of-scope accumulated rows
    are dropped for a like-for-like comparison (RF-12). No bound -> no filter."""
    if fecha_col not in df.columns or fecha_desde is None or fecha_hasta is None:
        return df
    iso = _iso_fecha(df[fecha_col])
    mask = (iso >= fecha_desde) & (iso <= fecha_hasta)
    return df[mask.fillna(False)].reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════
# classification — evidence-driven, first matching predicate wins
# ═════════════════════════════════════════════════════════════════════════


def _terna_of(row: DiffRow) -> tuple | None:
    """The (fecha, cliente, articulo) terna extractable from a diff key, if
    the pivot carries those fields (CLIENTE-FECHA); else None."""
    fields = KEY_FIELDS_BY_PIVOT[row.pivot]
    needed = (_FECHA_FIELD, _TERNA_CLIENTE, "Artículo Distribuidora")
    if all(n in fields for n in needed):
        return tuple(row.key[fields.index(n)] for n in needed)
    return None


def _p_stale_precio(row: DiffRow, ctx: KnownDefectContext) -> str | None:
    if ctx.stale_precio_clients and any(v in ctx.stale_precio_clients for v in row.key):
        return _REASON_BD_BE
    terna = _terna_of(row)
    if terna is not None and terna in ctx.stale_precio_ternas:
        return _REASON_BD_BE
    return None


def _p_stale_sucursal(row: DiffRow, ctx: KnownDefectContext) -> str | None:
    if ctx.stale_sucursal_clients and any(v in ctx.stale_sucursal_clients for v in row.key):
        return _REASON_BG_BH
    return None


def _p_es_ccu(row: DiffRow, ctx: KnownDefectContext) -> str | None:
    if not ctx.es_ccu_defect_generics:
        return None
    if row.column in ctx.es_ccu_defect_generics:
        return _REASON_ES_CCU
    if any(v in ctx.es_ccu_defect_generics for v in row.key):
        return _REASON_ES_CCU
    return None


def _p_az_ax(row: DiffRow, ctx: KnownDefectContext) -> str | None:
    if ctx.az_ax_drift_columns and row.column in ctx.az_ax_drift_columns:
        return _REASON_AZ_AX
    return None


# order fixed & documented: first matching known-bug predicate wins.
_DEFECT_PREDICATES = (_p_stale_precio, _p_stale_sucursal, _p_es_ccu, _p_az_ax)


def classify_delta(row: DiffRow, context: KnownDefectContext) -> tuple[str, str]:
    """Classify a delta as ``baseline-defect`` (first matching known-bug
    predicate) or ``real-divergence`` (no predicate matched)."""
    for predicate in _DEFECT_PREDICATES:
        reason = predicate(row, context)
        if reason:
            return BASELINE_DEFECT, reason
    return REAL_DIVERGENCE, _REASON_REAL


# ═════════════════════════════════════════════════════════════════════════
# per-pivot diff
# ═════════════════════════════════════════════════════════════════════════


def _keyed_values(df: pd.DataFrame, key_fields: list[str], value_cols: list[str]) -> dict[tuple, dict[str, float | None]]:
    present_values = [c for c in value_cols if c in df.columns]
    out: dict[tuple, dict[str, float | None]] = {}
    for _, r in df.iterrows():
        key = tuple(r[f] for f in key_fields)
        out[key] = {c: _cell_number(r[c]) for c in present_values}
    return out


def diff_pivot(
    pivot: str,
    base: pd.DataFrame,
    backup: pd.DataFrame,
    *,
    fecha_desde: str | None,
    fecha_hasta: str | None,
    context: KnownDefectContext,
    tolerance: float = DEFAULT_TOLERANCE,
) -> list[DiffRow]:
    """Key-based diff of one pivot's BASE frame vs its backup frame.

    Strips the TOTAL GENERAL row, applies like-for-like period-scope
    filtering (when the pivot carries a ``Fecha`` field), then surfaces every
    value cell whose absolute delta exceeds ``tolerance`` — classified per
    ``context``. Missing keys are surfaced as base-only / backup-only deltas."""
    key_fields = KEY_FIELDS_BY_PIVOT[pivot]
    value_cols = VALUE_COLUMNS_BY_PIVOT[pivot]

    base = strip_total_general(base)
    backup = strip_total_general(backup)
    if _FECHA_FIELD in key_fields:
        base = filter_period_scope(base, _FECHA_FIELD, fecha_desde, fecha_hasta)
        backup = filter_period_scope(backup, _FECHA_FIELD, fecha_desde, fecha_hasta)

    base_map = _keyed_values(base, key_fields, value_cols)
    backup_map = _keyed_values(backup, key_fields, value_cols)

    rows: list[DiffRow] = []
    for key in list(base_map.keys()) + [k for k in backup_map if k not in base_map]:
        base_vals = base_map.get(key, {})
        backup_vals = backup_map.get(key, {})
        if key in base_map and key in backup_map:
            presence = "both"
        elif key in base_map:
            presence = "base-only"
        else:
            presence = "backup-only"

        for col in value_cols:
            bv = base_vals.get(col)
            kv = backup_vals.get(col)
            delta = (bv if bv is not None else 0.0) - (kv if kv is not None else 0.0)
            if abs(delta) <= tolerance:
                continue
            row = DiffRow(
                pivot=pivot,
                key=key,
                key_fields=tuple(key_fields),
                column=col,
                base_value=bv,
                backup_value=kv,
                delta=delta,
                presence=presence,
                classification="",
                reason="",
            )
            row.classification, row.reason = classify_delta(row, context)
            rows.append(row)
    return rows


# ═════════════════════════════════════════════════════════════════════════
# terna->precio empirical validator (Decision 14)
# ═════════════════════════════════════════════════════════════════════════


def read_aexcel_export(path: str | Path) -> pd.DataFrame:
    """Read the real ``aexcel.xlsx`` file into the canonical terna->precio
    frame (``Descripción Período`` / ``Cod. Cliente`` / ``Código`` /
    ``Precio`` / ``Bonific``).

    NOTE (open item, design §Open Questions): the real export's exact header
    row / column labels are confirmed during the parallel run; the default
    contract here mirrors the aexcel-equivalent naming so a fixture (and the
    first real export shaped to it) reads cleanly. Adjust the mapping here
    once the real header layout is pinned."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"aexcel.xlsx no encontrado: {path}")
    df = pd.read_excel(path, engine="openpyxl")
    missing = [c for c in _AEXCEL_EXPORT_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"aexcel export missing expected columns: {missing}")
    return df[_AEXCEL_EXPORT_COLUMNS].copy()


def validate_terna_precios(
    generated: pd.DataFrame, aexcel: pd.DataFrame, *, tolerance: float = DEFAULT_TOLERANCE
) -> list[TernaPriceMismatch]:
    """Compare the generated terna->precio/Bonific pick against the real
    aexcel export; surface every disagreement (Decision 14). One mismatch per
    terna, first disagreement wins: missing-in-aexcel > precio > bonific."""
    if generated is None or generated.empty:
        return []

    gen = generated.copy()
    ax = aexcel.copy()
    gen[_TERNA_FECHA] = _iso_fecha(gen[_TERNA_FECHA])
    ax[_TERNA_FECHA] = _iso_fecha(ax[_TERNA_FECHA])

    ax_lookup: dict[tuple, tuple[float | None, float | None]] = {}
    for _, r in ax.iterrows():
        ax_lookup[(r[_TERNA_FECHA], r[_TERNA_CLIENTE], r[_TERNA_ARTICULO])] = (
            _cell_number(r[_TERNA_PRECIO]),
            _cell_number(r[_TERNA_BONIFIC]),
        )

    mismatches: list[TernaPriceMismatch] = []
    for _, r in gen.iterrows():
        key = (r[_TERNA_FECHA], r[_TERNA_CLIENTE], r[_TERNA_ARTICULO])
        gen_precio = _cell_number(r[_TERNA_PRECIO])
        gen_bonific = _cell_number(r[_TERNA_BONIFIC]) if _TERNA_BONIFIC in gen.columns else None

        if key not in ax_lookup:
            mismatches.append(
                TernaPriceMismatch(key[0], key[1], key[2], gen_precio, None, gen_bonific, None, "missing-in-aexcel")
            )
            continue

        ax_precio, ax_bonific = ax_lookup[key]
        if _abs_diff(gen_precio, ax_precio) > tolerance:
            mismatches.append(
                TernaPriceMismatch(key[0], key[1], key[2], gen_precio, ax_precio, gen_bonific, ax_bonific, "precio")
            )
        elif _abs_diff(gen_bonific, ax_bonific) > tolerance:
            mismatches.append(
                TernaPriceMismatch(key[0], key[1], key[2], gen_precio, ax_precio, gen_bonific, ax_bonific, "bonific")
            )
    return mismatches


def _abs_diff(a: float | None, b: float | None) -> float:
    """Absolute difference treating a missing value as 0 (never rounds)."""
    return abs((a if a is not None else 0.0) - (b if b is not None else 0.0))


# ═════════════════════════════════════════════════════════════════════════
# orchestration
# ═════════════════════════════════════════════════════════════════════════


def run_diff(
    base_frames: Mapping[str, pd.DataFrame],
    backup_frames: Mapping[str, pd.DataFrame],
    *,
    fecha_desde: str | None,
    fecha_hasta: str | None,
    context: KnownDefectContext | None = None,
    generated_ternas: pd.DataFrame | None = None,
    aexcel: pd.DataFrame | None = None,
    tolerance: float = DEFAULT_TOLERANCE,
) -> DiffReport:
    """Diff every pivot present in BOTH frame maps and (optionally) validate
    the terna->precio pick against the real aexcel."""
    ctx = context or KnownDefectContext()
    report = DiffReport(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, tolerance=tolerance)

    for pivot in KEY_FIELDS_BY_PIVOT:
        if pivot in base_frames and pivot in backup_frames:
            report.rows.extend(
                diff_pivot(
                    pivot,
                    base_frames[pivot],
                    backup_frames[pivot],
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    context=ctx,
                    tolerance=tolerance,
                )
            )

    if generated_ternas is not None and aexcel is not None:
        report.terna_mismatches = validate_terna_precios(generated_ternas, aexcel, tolerance=tolerance)

    return report


# ═════════════════════════════════════════════════════════════════════════
# report writers (JSON + xlsx + human summary)
# ═════════════════════════════════════════════════════════════════════════


def _style_total_general_row(ws: Worksheet, row: int, num_cols: int) -> None:
    fill = PatternFill(start_color=_TOTAL_FILL_COLOR, end_color=_TOTAL_FILL_COLOR, fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


def _write_sheet(ws: Worksheet, headers: list[str], data_rows: list[list[Any]], total_row: list[Any]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
    r = 2
    for data_row in data_rows:
        for c, v in enumerate(data_row, 1):
            ws.cell(r, c, _json_safe(v))
        r += 1
    for c, v in enumerate(total_row, 1):
        ws.cell(r, c, _json_safe(v))
    _style_total_general_row(ws, r, len(headers))


def _diferencias_rows(report: DiffReport) -> tuple[list[list[Any]], list[Any]]:
    headers_len = 9
    rows = [
        [r.pivot, str(r.key), r.column, r.base_value, r.backup_value, r.delta, r.presence, r.classification, r.reason]
        for r in report.rows
    ]
    total_delta = 0.0
    for r in report.rows:
        total_delta += r.delta
    total = [TOTAL_GENERAL_LABEL, "", "", "", "", total_delta, f"{len(report.rows)} deltas", "", ""]
    assert len(total) == headers_len
    return rows, total


def write_diff_report(report: DiffReport, output_dir: str | Path, *, stem: str = "diff_acciones_comerciales") -> dict[str, Path]:
    """Emit the JSON, xlsx and human-readable ``.txt`` summary next to the
    BASE output. Every xlsx sheet ends in a distinctly-styled TOTAL GENERAL
    row (project rule)."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    json_path = output_dir / f"{stem}.json"
    xlsx_path = output_dir / f"{stem}.xlsx"
    summary_path = output_dir / f"{stem}.txt"

    json_path.write_text(json.dumps(report.to_json_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    summary_path.write_text(report.to_summary_text(), encoding="utf-8")

    wb = Workbook()

    # Resumen sheet.
    ws_res = wb.active
    ws_res.title = "Resumen"
    counts = report.summary_counts()
    res_rows = [
        ["baseline-defect", counts["baseline_defect"]],
        ["real-divergence", counts["real_divergence"]],
        ["terna->precio mismatches", counts["terna_mismatches"]],
    ]
    _write_sheet(ws_res, ["Metrica", "Cantidad"], res_rows, [TOTAL_GENERAL_LABEL, counts["total_deltas"]])

    # Diferencias sheet.
    ws_dif = wb.create_sheet("Diferencias")
    dif_headers = ["Pivot", "Key", "Columna", "Base", "Backup", "Delta", "Presencia", "Clasificacion", "Motivo"]
    dif_rows, dif_total = _diferencias_rows(report)
    _write_sheet(ws_dif, dif_headers, dif_rows, dif_total)

    # Validacion Ternas sheet.
    ws_ter = wb.create_sheet("Validacion Ternas")
    ter_headers = ["Fecha", "Cod. Cliente", "Codigo", "Precio generado", "Precio aexcel", "Bonif generado", "Bonif aexcel", "Tipo"]
    ter_rows = [
        [m.fecha, m.id_cliente, m.id_articulo, m.generated_precio, m.aexcel_precio, m.generated_bonific, m.aexcel_bonific, m.kind]
        for m in report.terna_mismatches
    ]
    _write_sheet(ws_ter, ter_headers, ter_rows, [TOTAL_GENERAL_LABEL, "", "", "", "", "", "", f"{len(report.terna_mismatches)}"])

    wb.save(xlsx_path)

    return {"json": json_path, "xlsx": xlsx_path, "summary": summary_path}


# ═════════════════════════════════════════════════════════════════════════
# backup / context loading + optional service-CLI step (S4.3)
# ═════════════════════════════════════════════════════════════════════════

_BACKUP_SHEET_NAMES = list(KEY_FIELDS_BY_PIVOT.keys())


def load_pivots_from_workbook(path: str | Path) -> dict[str, pd.DataFrame]:
    """Read the backup workbook's pivot sheets (FACT_NET / ART-ACCION /
    CLIENTE-FECHA / ACC-GEN) into frames, header in row 1 (BASE-control
    shape). Sheets absent from the workbook are simply skipped — the manual
    backup may only export a subset."""
    path = Path(path)
    xls = pd.ExcelFile(path, engine="openpyxl")
    frames: dict[str, pd.DataFrame] = {}
    for sheet in _BACKUP_SHEET_NAMES:
        if sheet in xls.sheet_names:
            frames[sheet] = xls.parse(sheet)
    return frames


def load_defect_context(path: str | Path) -> KnownDefectContext:
    """Load the evidence sets from a ``known_defects.json`` file (lists ->
    frozensets). A missing file yields an empty context (everything stays
    ``real-divergence`` until evidence is gathered)."""
    path = Path(path)
    if not path.exists():
        return KnownDefectContext()
    payload = json.loads(path.read_text(encoding="utf-8"))

    def _fs(key: str) -> frozenset:
        raw = payload.get(key, [])
        return frozenset(tuple(v) if isinstance(v, list) else v for v in raw)

    return KnownDefectContext(
        stale_sucursal_clients=_fs("stale_sucursal_clients"),
        stale_precio_clients=_fs("stale_precio_clients"),
        stale_precio_ternas=_fs("stale_precio_ternas"),
        es_ccu_defect_generics=_fs("es_ccu_defect_generics"),
        az_ax_drift_columns=_fs("az_ax_drift_columns"),
    )


def _find_backup_workbook(backup_dir: Path) -> Path | None:
    """Pick the backup workbook: prefer ``backup.xlsx``, else the first
    ``*.xlsx`` in the directory (deterministic — sorted)."""
    preferred = backup_dir / "backup.xlsx"
    if preferred.exists():
        return preferred
    candidates = sorted(backup_dir.glob("*.xlsx"))
    return candidates[0] if candidates else None


def run_diff_step(
    *,
    base_frames: Mapping[str, pd.DataFrame],
    backup_dir: str | Path,
    aexcel_path: str | Path | None,
    generated_ternas: pd.DataFrame | None,
    fecha_desde: str | None,
    fecha_hasta: str | None,
    output_dir: str | Path,
    context: KnownDefectContext | None = None,
    stem: str = "diff_acciones_comerciales",
) -> dict[str, Path] | None:
    """Optional parallel-diff step wired behind ``config.backup_dir`` (S4.3).

    Loads the backup pivots + (optional) real aexcel + (optional)
    ``known_defects.json`` from ``backup_dir``, runs the diff, and writes the
    report next to the BASE output. Returns the report paths, or ``None`` when
    no backup workbook is present (the step is a no-op then — never fails the
    BASE run)."""
    backup_dir = Path(backup_dir)
    if not backup_dir.exists():
        logger.warning("diff step: backup_dir %s does not exist — skipping", backup_dir)
        return None

    backup_path = _find_backup_workbook(backup_dir)
    if backup_path is None:
        logger.warning("diff step: no *.xlsx backup workbook under %s — skipping", backup_dir)
        return None

    backup_frames = load_pivots_from_workbook(backup_path)
    ctx = context or load_defect_context(backup_dir / "known_defects.json")

    aexcel = None
    if aexcel_path is not None and Path(aexcel_path).exists():
        aexcel = read_aexcel_export(aexcel_path)
    elif aexcel_path is not None:
        logger.warning("diff step: aexcel_path %s not found — terna validation skipped", aexcel_path)

    report = run_diff(
        base_frames,
        backup_frames,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        context=ctx,
        generated_ternas=generated_ternas,
        aexcel=aexcel,
    )
    paths = write_diff_report(report, output_dir, stem=stem)
    logger.info("diff step: report written to %s (%s)", output_dir, report.summary_counts())
    return paths
