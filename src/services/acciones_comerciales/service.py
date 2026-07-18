"""AccionesComercialesService — Phase-1 skeleton (RF-22).

Generates the BASE control workbook for acciones-comerciales into
``data/output/acciones-comerciales/{YYYY-MM}/`` (service_output_dir
convention). This slice (S1) is intentionally a SKELETON: it proves the
config -> service -> CLI plumbing end-to-end (output path convention, the
Phase-2 escribir_informe flag defaulting OFF, zero external-file writes)
without yet wiring the gold datasource / wapi / compras / pivots pipeline.

S3 ("Wire full Phase-1 orchestration into service.py") replaces the
placeholder sheet built here with the real 6-sheet BASE control workbook
(4 pivots + wapi-derived table + reconciliation, per RF-10/RF-11) once
gold_source (S1, done), processor (S2) and pivots (S2) exist.

Phase 2 (informe write, captures, delivery) stays behind
``config.escribir_informe`` (default False, RF-13) — this skeleton never
touches ``config.informe_path``.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.core.excel_writer import ExcelWriter
from src.services.acciones_comerciales.config import AccionesComercialesConfig
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)

# Distinct styling for the TOTAL GENERAL row (project convention — matches
# the amber fill used across other services, e.g. ventas_marca/service.py).
_TOTAL_FILL = "FFE08A"


@dataclass
class AccionesComercialesResult:
    """Resultado de la generacion del reporte de acciones comerciales."""

    ruta_archivo: Path
    registros_procesados: int


def _style_total_general_row(ws, row: int, num_cols: int) -> None:
    """Apply distinct bold+fill styling to a TOTAL GENERAL row (project
    rule: every generated sheet ends with a distinctly-styled totals row)."""
    fill = PatternFill(start_color=_TOTAL_FILL, end_color=_TOTAL_FILL, fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


class AccionesComercialesService(BaseService):
    """Genera el reporte de acciones comerciales (Phase 1: BASE control)."""

    SERVICE_SLUG = "acciones-comerciales"
    GRANULARITY = "month"

    def generar_reporte(self, config: AccionesComercialesConfig) -> AccionesComercialesResult:
        if not self.validar_fechas(config.fecha_desde, config.fecha_hasta):
            raise ValueError(
                "fecha_desde/fecha_hasta invalidas — formato esperado YYYY-MM-DD"
            )

        output_dir = (
            Path(config.output_dir) if config.output_dir else self._output_dir(config.fecha_desde)
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        nombre = config.nombre_archivo or "BASE control Acciones Comerciales"

        # Placeholder BASE control sheet — replaced by the real 6-sheet
        # writer (writers/base_control.py) in S3. Honest zero-row content
        # (no gold/wapi/compras data is read at this slice) with a
        # TOTAL GENERAL row so the project's totals-row rule holds even in
        # skeleton form.
        df = pd.DataFrame(
            [
                {"Periodo": f"{config.fecha_desde} a {config.fecha_hasta}", "Registros": 0},
                {"Periodo": "TOTAL GENERAL", "Registros": 0},
            ]
        )

        writer = ExcelWriter(nombre, output_dir=output_dir)
        ws = writer.add_sheet(df, sheet_name="BASE control")
        _style_total_general_row(ws, row=ws.max_row, num_cols=len(df.columns))
        ruta = writer.save()

        logger.info("Acciones Comerciales BASE control generado: %s", ruta)

        return AccionesComercialesResult(ruta_archivo=ruta, registros_procesados=0)
