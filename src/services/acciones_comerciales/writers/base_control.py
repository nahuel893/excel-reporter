"""BASE control workbook writer (RF-10, RF-11) — 6 sheets: 4 pivots + the
wapi-derived enriched table + a reconciliation sheet, EVERY sheet ending in
a distinctly-styled (bold + amber fill) ``TOTAL GENERAL`` row.

Uses the core ``ExcelWriter`` (``src/core/excel_writer.py``) for the 5
uniform, single-DataFrame sheets (4 pivots + wapi). The reconciliation sheet
is heterogeneous (several stacked audit sections, not one flat table), so it
is written directly on the SAME ``ExcelWriter.workbook`` via plain openpyxl
calls, then the whole file is saved through ``ExcelWriter.save()`` — still
"one writer, one file", just with a hand-built sheet for the one sheet whose
shape a single DataFrame cannot express.

Reconciliation sheet section order (RF-11), TOP to BOTTOM:
  1. RANGO DE FECHAS (MIN/MAX) — per-source date coverage vs the configured
     period.
  2. TERNAS MULTI-PRECIO — the RF-01 grain-collapse audit trail (terna key,
     candidate prices/Bonific, the picked value, the pick reason) that the
     RF-12 diff harness compares against the real aexcel file.
  3. FILAS NO RESUELTAS — SUCURSAL (RF-04) — unresolved dim_cliente lookups.
  4. FILAS NO RESUELTAS — PRECIO (RF-05) — unresolved terna price lookups.
  5. RECONCILIACION FACTURACION / DESCUENTOS — per-SUCURSAL comparison of
     the aexcel-side ``Facturacion Neta``/``Descuentos`` sums against the
     wapi-side ``Descuento`` sum, by their CORRECT named pandas columns
     (never a column-letter/AZ-AX-style drift bug — there IS no column
     letter here, only named columns). This section is placed LAST so its
     TOTAL GENERAL row is the sheet's final row (RF-10: every sheet MUST
     END with the totals row).

No value is ever rounded / truncated / cast to int here (RF-23) — the
$0.01 tolerance flag is a comparison-only classification, never a stored
value transform.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.excel_writer import ExcelWriter
from src.services.acciones_comerciales.constants import (
    COL_DESCUENTO,
    COL_SUCURSAL,
    FACT_NET_DESCUENTOS_SRC,
    FACT_NET_FACT_NETA_SRC,
    FACT_NET_ROW_FIELDS,
)
from src.services.acciones_comerciales.gold_source import MultiPriceTerna

TOTAL_GENERAL_LABEL = "TOTAL GENERAL"

SHEET_FACT_NET = "FACT_NET"
SHEET_ART_ACCION = "ART-ACCION"
SHEET_CLIENTE_FECHA = "CLIENTE-FECHA"
SHEET_ACC_GEN = "ACC-GEN"
SHEET_WAPI = "wapi"
SHEET_RECONCILIACION = "Reconciliacion"

_TOTAL_FILL_COLOR = "FFE08A"  # amber — project convention for totals rows
_DEFAULT_TOLERANCE = 0.01  # RF-12 exact-to-$0.01 comparison discipline

# Aexcel-side terna-grain frame column carrying the SUCURSAL label + date.
_AEXCEL_SUCURSAL_COL = FACT_NET_ROW_FIELDS[0]  # "Sucursal"
_AEXCEL_FECHA_COL = "Descripción Período"
_WAPI_FECHA_COL = "Fecha"

# Numeric measure columns actually summed into the wapi sheet's TOTAL
# GENERAL row — deliberately NOT "every numeric column" (Cod. Cliente,
# Comprobante, per-unit prices and percentage columns are numeric dtype but
# summing them is meaningless); this is the sensible-measures subset.
_WAPI_SUMMABLE_COLUMNS: list[str] = [
    "Cantidad",
    "Total",
    "Cantidad Sin Cargo",
    "Descuento $ sobre PN SF",
    "Participación CMQ",
    "Monto A Acreditar",
    "Total2",
    COL_DESCUENTO,
]

_RECONCILIACION_HEADERS = [
    "SUCURSAL",
    "Facturacion Neta (aexcel)",
    "Descuentos (aexcel)",
    "Descuento (wapi)",
    "Delta (Descuentos aexcel - Descuento wapi)",
    "Dentro de tolerancia (0.01)",
]

_MULTI_PRECIO_HEADERS = [
    "Fecha",
    "Cod. Cliente",
    "Codigo Articulo",
    "Precios candidatos",
    "Bonific candidatos",
    "Precio elegido",
    "Bonific elegido",
    "Motivo de eleccion",
]

_UNRESOLVED_COLUMNS = ["Cod. Cliente", "Fecha", "Artículo Distribuidora"]

_FECHA_RANGO_HEADERS = [
    "Fuente",
    "Fecha Minima",
    "Fecha Maxima",
    "Fecha Desde (config)",
    "Fecha Hasta (config)",
    "Dentro de rango configurado",
]


@dataclass
class ReconciliationInputs:
    """Inputs the reconciliation sheet is built from (RF-11).

    ``aexcel_data`` — the terna-grain aexcel-equivalent frame (RF-01),
        carrying ``Sucursal`` / ``Descripción Período`` / ``Facturacion
        Neta`` / ``Descuentos``.
    ``wapi_enriched`` — the enriched wapi table (RF-04..RF-08), carrying
        ``SUCURSAL`` / ``Fecha`` / ``Descuento``.
    ``multi_price_ternas`` — the RF-01 grain-collapse audit trail
        (``gold_source.AexcelEquivalentResult.multi_price_ternas``).
    ``unresolved_sucursal`` / ``unresolved_precio`` — the RF-04/RF-05
        flagged subsets (``processor.EnrichedWapiResult``).
    ``fecha_desde`` / ``fecha_hasta`` — the configured period, used as the
        expected date-range bound for the MIN/MAX checks.
    ``tolerance`` — the $0.01 comparison tolerance (RF-12 discipline);
        comparison-only, never applied to any stored value.
    """

    aexcel_data: pd.DataFrame
    wapi_enriched: pd.DataFrame
    multi_price_ternas: list[MultiPriceTerna] = field(default_factory=list)
    unresolved_sucursal: pd.DataFrame = field(default_factory=pd.DataFrame)
    unresolved_precio: pd.DataFrame = field(default_factory=pd.DataFrame)
    fecha_desde: str | None = None
    fecha_hasta: str | None = None
    tolerance: float = _DEFAULT_TOLERANCE


# ─────────────────────────────────────────────────────────────────────────
# shared styling helpers
# ─────────────────────────────────────────────────────────────────────────


def _style_total_general_row(ws: Worksheet, row: int, num_cols: int) -> None:
    """Distinct bold+amber-fill styling for a TOTAL GENERAL row (RF-10)."""
    fill = PatternFill(start_color=_TOTAL_FILL_COLOR, end_color=_TOTAL_FILL_COLOR, fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


def _style_section_header(ws: Worksheet, row: int) -> None:
    ws.cell(row, 1).font = Font(bold=True, size=12)


def _style_header_row(ws: Worksheet, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        ws.cell(row, col).font = Font(bold=True)


# ─────────────────────────────────────────────────────────────────────────
# wapi-derived table sheet — TOTAL GENERAL over the sensible measures
# ─────────────────────────────────────────────────────────────────────────


def _append_wapi_total_general(wapi_enriched: pd.DataFrame) -> pd.DataFrame:
    label_col = wapi_enriched.columns[0]
    total_row: dict[str, Any] = {}
    for col in wapi_enriched.columns:
        if col == label_col:
            total_row[col] = TOTAL_GENERAL_LABEL
        elif col in _WAPI_SUMMABLE_COLUMNS:
            total_row[col] = wapi_enriched[col].sum()
        else:
            total_row[col] = ""
    total_df = pd.DataFrame([total_row], columns=wapi_enriched.columns)
    return pd.concat([wapi_enriched, total_df], ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────
# reconciliation — date-range block
# ─────────────────────────────────────────────────────────────────────────


def _min_max_dates(df: pd.DataFrame, col: str) -> tuple[str | None, str | None]:
    if df.empty or col not in df.columns:
        return (None, None)
    parsed = pd.to_datetime(df[col], errors="coerce").dropna()
    if parsed.empty:
        return (None, None)
    return (parsed.min().strftime("%Y-%m-%d"), parsed.max().strftime("%Y-%m-%d"))


def _within_configured_range(
    fecha_min: str | None, fecha_max: str | None, fecha_desde: str | None, fecha_hasta: str | None
) -> str:
    if not fecha_min or not fecha_max or not fecha_desde or not fecha_hasta:
        return "N/D"
    return "SI" if (fecha_min >= fecha_desde and fecha_max <= fecha_hasta) else "NO"


def _build_date_range_rows(inputs: ReconciliationInputs) -> list[list[Any]]:
    ax_min, ax_max = _min_max_dates(inputs.aexcel_data, _AEXCEL_FECHA_COL)
    wapi_min, wapi_max = _min_max_dates(inputs.wapi_enriched, _WAPI_FECHA_COL)
    return [
        [
            "aexcel (gold)",
            ax_min,
            ax_max,
            inputs.fecha_desde,
            inputs.fecha_hasta,
            _within_configured_range(ax_min, ax_max, inputs.fecha_desde, inputs.fecha_hasta),
        ],
        [
            "wapi",
            wapi_min,
            wapi_max,
            inputs.fecha_desde,
            inputs.fecha_hasta,
            _within_configured_range(wapi_min, wapi_max, inputs.fecha_desde, inputs.fecha_hasta),
        ],
    ]


# ─────────────────────────────────────────────────────────────────────────
# reconciliation — multi-price terna audit section (RF-11)
# ─────────────────────────────────────────────────────────────────────────


def _build_multi_price_rows(multi_price_ternas: list[MultiPriceTerna]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for terna in multi_price_ternas:
        rows.append(
            [
                terna.fecha,
                terna.id_cliente,
                terna.id_articulo,
                ", ".join(str(p) for p in terna.candidate_precios),
                ", ".join(str(b) for b in terna.candidate_bonific),
                terna.picked_precio,
                terna.picked_bonific,
                terna.pick_reason,
            ]
        )
    return rows


# ─────────────────────────────────────────────────────────────────────────
# reconciliation — unresolved-rows sections (RF-04/RF-05)
# ─────────────────────────────────────────────────────────────────────────


def _build_unresolved_rows(df: pd.DataFrame) -> list[list[Any]]:
    if df.empty:
        return []
    present = [c for c in _UNRESOLVED_COLUMNS if c in df.columns]
    return df[present].reindex(columns=_UNRESOLVED_COLUMNS).values.tolist()


# ─────────────────────────────────────────────────────────────────────────
# reconciliation — FACTURACION/DESCUENTOS delta section (RF-11, ends sheet)
# ─────────────────────────────────────────────────────────────────────────


_SUCURSAL_PREFIX_RE = re.compile(r"^\d+\s*-\s*")


def _strip_sucursal_prefix(label: Any) -> Any:
    """Normalize a ``"{id} - {DESC}"`` Sucursal label to the bare
    description for like-for-like reconciliation. Both the aexcel-side
    Sucursal (FACT_NET, ``get_aexcel_equivalent``) and the wapi-side
    SUCURSAL (RF-04) now carry the id prefix, so both are stripped to the
    bare name before their per-sucursal totals are compared."""
    if not isinstance(label, str):
        return label
    return _SUCURSAL_PREFIX_RE.sub("", label).strip()


def _group_sum(df: pd.DataFrame, group_col: str, value_col: str) -> dict[Any, float]:
    if df.empty or group_col not in df.columns or value_col not in df.columns:
        return {}
    return df.groupby(group_col)[value_col].sum().to_dict()


def _build_reconciliacion_rows_and_total(
    inputs: ReconciliationInputs,
) -> tuple[list[list[Any]], list[Any]]:
    aexcel_normalized = inputs.aexcel_data.copy()
    if not aexcel_normalized.empty and _AEXCEL_SUCURSAL_COL in aexcel_normalized.columns:
        aexcel_normalized[_AEXCEL_SUCURSAL_COL] = aexcel_normalized[_AEXCEL_SUCURSAL_COL].map(
            _strip_sucursal_prefix
        )

    wapi_normalized = inputs.wapi_enriched.copy()
    if not wapi_normalized.empty and COL_SUCURSAL in wapi_normalized.columns:
        wapi_normalized[COL_SUCURSAL] = wapi_normalized[COL_SUCURSAL].map(_strip_sucursal_prefix)

    fn_by_suc = _group_sum(aexcel_normalized, _AEXCEL_SUCURSAL_COL, FACT_NET_FACT_NETA_SRC)
    desc_ax_by_suc = _group_sum(aexcel_normalized, _AEXCEL_SUCURSAL_COL, FACT_NET_DESCUENTOS_SRC)
    desc_wapi_by_suc = _group_sum(wapi_normalized, COL_SUCURSAL, COL_DESCUENTO)

    sucursales = sorted(set(fn_by_suc) | set(desc_ax_by_suc) | set(desc_wapi_by_suc), key=str)

    rows: list[list[Any]] = []
    total_fn = total_desc_ax = total_desc_wapi = total_delta = 0.0
    for suc in sucursales:
        fn = fn_by_suc.get(suc, 0.0)
        desc_ax = desc_ax_by_suc.get(suc, 0.0)
        desc_wapi = desc_wapi_by_suc.get(suc, 0.0)
        delta = desc_ax - desc_wapi
        dentro = "SI" if abs(delta) <= inputs.tolerance else "NO"
        rows.append([suc, fn, desc_ax, desc_wapi, delta, dentro])
        total_fn += fn
        total_desc_ax += desc_ax
        total_desc_wapi += desc_wapi
        total_delta += delta

    overall = "SI" if abs(total_delta) <= inputs.tolerance else "NO"
    total_row = [TOTAL_GENERAL_LABEL, total_fn, total_desc_ax, total_desc_wapi, total_delta, overall]
    return rows, total_row


# ─────────────────────────────────────────────────────────────────────────
# reconciliation sheet assembly
# ─────────────────────────────────────────────────────────────────────────


def _write_block(ws: Worksheet, row: int, title: str, headers: list[str], rows: list[list[Any]]) -> int:
    """Write one stacked section (bold title, bold header row, data rows,
    trailing blank separator). Returns the next free row."""
    ws.cell(row, 1, title)
    _style_section_header(ws, row)
    row += 1

    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row, c_idx, header)
    _style_header_row(ws, row, len(headers))
    header_row = row
    row += 1

    if not rows:
        ws.cell(row, 1, "Sin filas para este periodo")
        row += 1
    else:
        for data_row in rows:
            for c_idx, value in enumerate(data_row, start=1):
                ws.cell(row, c_idx, value)
            row += 1

    return row + 1  # trailing blank separator


def _write_reconciliation_sheet(wb: Workbook, inputs: ReconciliationInputs) -> Worksheet:
    ws = wb.create_sheet(title=SHEET_RECONCILIACION)
    row = 1

    row = _write_block(
        ws, row, "RANGO DE FECHAS (MIN/MAX)", _FECHA_RANGO_HEADERS, _build_date_range_rows(inputs)
    )
    row = _write_block(
        ws,
        row,
        "TERNAS MULTI-PRECIO (RF-11)",
        _MULTI_PRECIO_HEADERS,
        _build_multi_price_rows(inputs.multi_price_ternas),
    )
    row = _write_block(
        ws,
        row,
        "FILAS NO RESUELTAS - SUCURSAL (RF-04)",
        _UNRESOLVED_COLUMNS,
        _build_unresolved_rows(inputs.unresolved_sucursal),
    )
    row = _write_block(
        ws,
        row,
        "FILAS NO RESUELTAS - PRECIO (RF-05)",
        _UNRESOLVED_COLUMNS,
        _build_unresolved_rows(inputs.unresolved_precio),
    )

    # Last section: the sheet MUST end with the TOTAL GENERAL row (RF-10).
    ws.cell(row, 1, "RECONCILIACION FACTURACION / DESCUENTOS (RF-11)")
    _style_section_header(ws, row)
    row += 1
    for c_idx, header in enumerate(_RECONCILIACION_HEADERS, start=1):
        ws.cell(row, c_idx, header)
    _style_header_row(ws, row, len(_RECONCILIACION_HEADERS))
    row += 1

    delta_rows, total_row = _build_reconciliacion_rows_and_total(inputs)
    for data_row in delta_rows:
        for c_idx, value in enumerate(data_row, start=1):
            ws.cell(row, c_idx, value)
        row += 1

    for c_idx, value in enumerate(total_row, start=1):
        ws.cell(row, c_idx, value)
    _style_total_general_row(ws, row, len(total_row))

    return ws


# ─────────────────────────────────────────────────────────────────────────
# public entry point
# ─────────────────────────────────────────────────────────────────────────


def build_base_control_workbook(
    *,
    nombre_archivo: str,
    output_dir: Path,
    fact_net: pd.DataFrame,
    art_accion: pd.DataFrame,
    cliente_fecha: pd.DataFrame,
    acc_gen: pd.DataFrame,
    wapi_enriched: pd.DataFrame,
    reconciliation: ReconciliationInputs,
) -> Path:
    """Build the 6-sheet BASE control workbook (RF-10, RF-11).

    The 4 pivot frames (``fact_net``/``art_accion``/``cliente_fecha``/
    ``acc_gen``) are expected to already carry their own TOTAL GENERAL row
    (``pivots.py``'s ``_append_total_general``) — this function only styles
    it. The wapi-derived table and the reconciliation sheet get their
    TOTAL GENERAL row appended here.
    """
    writer = ExcelWriter(nombre_archivo, output_dir=output_dir)

    ws = writer.add_sheet(fact_net, sheet_name=SHEET_FACT_NET)
    _style_total_general_row(ws, ws.max_row, len(fact_net.columns))

    ws = writer.add_sheet(art_accion, sheet_name=SHEET_ART_ACCION)
    _style_total_general_row(ws, ws.max_row, len(art_accion.columns))

    ws = writer.add_sheet(cliente_fecha, sheet_name=SHEET_CLIENTE_FECHA)
    _style_total_general_row(ws, ws.max_row, len(cliente_fecha.columns))

    ws = writer.add_sheet(acc_gen, sheet_name=SHEET_ACC_GEN)
    _style_total_general_row(ws, ws.max_row, len(acc_gen.columns))

    wapi_with_total = _append_wapi_total_general(wapi_enriched)
    ws = writer.add_sheet(wapi_with_total, sheet_name=SHEET_WAPI)
    _style_total_general_row(ws, ws.max_row, len(wapi_with_total.columns))

    _write_reconciliation_sheet(writer.workbook, reconciliation)

    return writer.save()
