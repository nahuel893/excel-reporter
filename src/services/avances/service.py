"""
AvancesService - Generates avances Excel reports from a base template,
preserving formulas and user-added sheets.

Workflow:
- The base template is resolved automatically: first looks for the previous
  month's output (so customizations carry forward), then falls back to
  ``archivo_plantilla`` from config (for first-time runs).
- Output is written to ``data/output/avances/{YYYY-MM}/{nombre_archivo}.xlsx``.
- Re-running the same period UPDATES the existing output (preserving user
  customizations).
"""

import logging
import shutil
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from src.core.excel_updater import replace_sheet_data
from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)


@dataclass
class SheetConfig:
    """Maps an Excel sheet to its DB query and writable columns.

    column_rename: optional mapping {df_column_name: excel_header_name}.
    Applied before replace_sheet_data when DataFrame column names differ
    from the Excel header row. Only the renamed columns participate in the
    write; columns absent from both the rename map and data_columns are not
    touched.  data_columns must use the Excel header names (post-rename).
    """

    sheet_name: str
    query_method: str  # DataLoader method name
    query_params: list[str] = field(default_factory=list)  # param names from AvancesConfig
    data_columns: list[str] = field(default_factory=list)
    header_row: int = 1
    column_rename: dict[str, str] = field(default_factory=dict)  # df_col -> excel_header


SHEET_CONFIGS_BRANCA: list[SheetConfig] = [
    SheetConfig(
        sheet_name="gold fact_ventas",
        query_method="get_fact_ventas_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_sucursal"],
        data_columns=[
            "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
            "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
            "anulado", "cantidades_total", "bonificacion",
        ],
        header_row=1,
    ),
    SheetConfig(
        sheet_name="gold dim_articulo",
        query_method="get_dim_articulo_raw",
        data_columns=[
            "id_articulo", "des_articulo", "marca", "generico", "calibre",
            "proveedor", "unidad_negocio", "factor_hectolitros",
        ],
        header_row=1,
    ),
    SheetConfig(
        sheet_name="gold dim_cliente",
        query_method="get_dim_cliente_raw",
        query_params=["id_sucursal"],
        data_columns=[
            "id_cliente", "fantasia", "razon_social", "des_sucursal", "id_sucursal",
            "id_ruta_fv1", "des_personal_fv1", "id_ruta_fv4", "des_personal_fv4",
        ],
        header_row=2,
    ),
    SheetConfig(
        sheet_name="gold cob_preventista_generico",
        query_method="get_cob_preventista_generico_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_fuerza_ventas", "id_sucursal"],
        data_columns=[
            "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
            "id_sucursal", "ds_sucursal", "generico", "clientes_compradores",
            "volumen_total",
        ],
        header_row=1,
    ),
    SheetConfig(
        sheet_name="gold cob_preventista_marca",
        query_method="get_cob_preventista_marca_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_fuerza_ventas", "id_sucursal"],
        data_columns=[
            "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
            "id_sucursal", "ds_sucursal", "marca", "clientes_compradores",
            "volumen_total",
        ],
        header_row=1,
    ),
]

# Populated in PR #2 from scripts/inspect_badie_xlsm.py output.
#
# Inspection results (2026-06-05):
#   pivot_python  — 12 cols, 26443 rows, header_row=1
#     Excel headers: Sucursal, Descripcion Período, Descripcion Vendedor, Ruta,
#       Descripcion_Ruta, Descripcion_Marca, GENERICO, Código_Articulo,
#       Descripcion_Articulo, Cantidades Totales, CATEGORIA, Columna1
#     DataLoader (get_fact_ventas_raw): id_cliente, id_articulo, id_vendedor,
#       id_sucursal, fecha_comprobante, id_documento, letra, serie, nro_doc,
#       anulado, cantidades_total, bonificacion
#     Decision: use pivot_python — 12-col shape matches get_fact_ventas_raw.
#       column_rename maps DB names → Excel headers for the overlapping fields.
#
#   cober_gen     — 7 cols, 598 rows, header_row=1
#     Excel headers: Column1, Sucursal, Descripcion Vendedor, Ruta,
#       GENERICO, Numero_Clientes, Columna1
#     DataLoader (get_cob_preventista_generico_raw): id, periodo,
#       id_fuerza_ventas, id_vendedor, id_ruta, id_sucursal, ds_sucursal,
#       generico, clientes_compradores, volumen_total
#
#   cober_marca   — 6 cols, 2140 rows, header_row=1
#     Excel headers: Column1, Sucursal, Descripcion Vendedor, Ruta,
#       Descripcion_Marca, Numero_Clientes
#     DataLoader (get_cob_preventista_marca_raw): same shape as cober_gen
#       but with marca instead of generico.
SHEET_CONFIGS_BADIE: list[SheetConfig] = [
    # 1. Ventas — pivot_python sheet (12 cols, header_row=1)
    #    DataLoader columns renamed to match Excel headers where semantically aligned.
    #    Columns written: Sucursal, Descripcion Período, Descripcion Vendedor,
    #      Ruta, Código_Articulo, Cantidades Totales
    #    Excel-only cols (Descripcion_Ruta, Descripcion_Marca, GENERICO,
    #      Descripcion_Articulo, CATEGORIA, Columna1) are left untouched.
    SheetConfig(
        sheet_name="pivot_python",
        query_method="get_fact_ventas_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_sucursal"],
        column_rename={
            "id_sucursal": "Sucursal",
            "fecha_comprobante": "Descripcion Período",
            "id_vendedor": "Descripcion Vendedor",
            "nro_doc": "Ruta",
            "id_articulo": "Código_Articulo",
            "cantidades_total": "Cantidades Totales",
        },
        data_columns=[
            "Sucursal",
            "Descripcion Período",
            "Descripcion Vendedor",
            "Ruta",
            "Código_Articulo",
            "Cantidades Totales",
        ],
        header_row=1,
    ),
    # 2. Cobertura por genérico — cober_gen sheet (7 cols, header_row=1)
    #    Columns written: Column1, Sucursal, Descripcion Vendedor, Ruta, GENERICO, Numero_Clientes
    #    Columna1 (col G, extra display column) is left untouched.
    SheetConfig(
        sheet_name="cober_gen",
        query_method="get_cob_preventista_generico_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_fuerza_ventas", "id_sucursal"],
        column_rename={
            "id": "Column1",
            "id_sucursal": "Sucursal",
            "id_vendedor": "Descripcion Vendedor",
            "id_ruta": "Ruta",
            "generico": "GENERICO",
            "clientes_compradores": "Numero_Clientes",
        },
        data_columns=[
            "Column1",
            "Sucursal",
            "Descripcion Vendedor",
            "Ruta",
            "GENERICO",
            "Numero_Clientes",
        ],
        header_row=1,
    ),
    # 3. Cobertura por marca — cober_marca sheet (6 cols, header_row=1)
    #    Columns written: Column1, Sucursal, Descripcion Vendedor, Ruta,
    #      Descripcion_Marca, Numero_Clientes
    SheetConfig(
        sheet_name="cober_marca",
        query_method="get_cob_preventista_marca_raw",
        query_params=["fecha_desde", "fecha_hasta", "id_fuerza_ventas", "id_sucursal"],
        column_rename={
            "id": "Column1",
            "id_sucursal": "Sucursal",
            "id_vendedor": "Descripcion Vendedor",
            "id_ruta": "Ruta",
            "marca": "Descripcion_Marca",
            "clientes_compradores": "Numero_Clientes",
        },
        data_columns=[
            "Column1",
            "Sucursal",
            "Descripcion Vendedor",
            "Ruta",
            "Descripcion_Marca",
            "Numero_Clientes",
        ],
        header_row=1,
    ),
]

PLANTILLA_SHEET_CONFIGS: dict[str, list[SheetConfig]] = {
    "branca": SHEET_CONFIGS_BRANCA,
    "badie": SHEET_CONFIGS_BADIE,
}


@dataclass
class AvancesConfig:
    """Configuracion para el reporte de avances.

    Filtros de datos aplicados a las queries crudas:
      - id_sucursal: filtra fact_ventas, dim_cliente, cob_preventista_*
      - id_fuerza_ventas: filtra cob_preventista_*
      - dim_articulo se filtra por articulos vendidos en el rango + sucursal

    Template resolution:
      - If archivo_plantilla is set and the previous month's output exists,
        the previous month's output is used as base (customizations carry forward).
      - If there is no previous month's output, archivo_plantilla is used.
      - archivo_plantilla is now optional — if omitted, only previous-month
        resolution is attempted.
    """

    fecha_desde: str
    fecha_hasta: str
    tipo_plantilla: Literal["branca", "badie"] = "branca"
    archivo_plantilla: str | None = None  # path to BASE template (fallback if no prev output)
    id_sucursal: int = 1
    id_fuerza_ventas: int = 1
    nombre_archivo: str | None = None  # output filename (no extension)
    output_dir: Path | None = None  # override; if None, derived from fecha_desde


@dataclass
class AvancesResult:
    """Resultado de la generacion del reporte de avances."""

    ruta_archivo: Path
    registros_por_hoja: dict[str, int]


class AvancesService(BaseService):
    """Generates avances Excel reports from a base template into per-period folders.

    Template resolution order (first match wins):
    1. Previous month's output in data/output/avances/{YYYY-MM}/
    2. archivo_plantilla from config (fallback for first-time runs)
    """

    SERVICE_SLUG = "avances"
    GRANULARITY = "month"

    def _resolve_base(self, config: AvancesConfig, output_dir: Path) -> Path | None:
        """Find the base template: previous month's output > archivo_plantilla.

        The previous month is derived from fecha_desde. When multiple reports
        share the output directory (e.g. Branca and Badie), we filter candidates
        by the name prefix extracted from nombre_archivo to avoid picking the
        wrong report's output.
        """
        # Extract name prefix for filtering (e.g. "AVANCE BRANCA" from "AVANCE BRANCA - MAYO 2026")
        name_prefix = ""
        if config.nombre_archivo:
            # Split on " - " and take the first part(s) before the date
            parts = config.nombre_archivo.split(" - ")
            if len(parts) > 1:
                name_prefix = parts[0].strip()  # e.g. "AVANCE BRANCA"

        # 1. Try previous month's output
        desde = datetime.strptime(config.fecha_desde, "%Y-%m-%d")
        prev_year = desde.year if desde.month > 1 else desde.year - 1
        prev_month = desde.month - 1 if desde.month > 1 else 12
        prev_period = f"{prev_year}-{prev_month:02d}"
        prev_dir = service_output_dir(self.SERVICE_SLUG, f"{prev_period}-01", "month")

        if prev_dir.is_dir():
            candidates = sorted(
                p for p in prev_dir.glob("*.xlsx")
                if "_backup" not in p.stem
            )
            # Filter by name prefix to avoid picking wrong report's output
            if name_prefix and candidates:
                matching = [c for c in candidates if c.stem.startswith(name_prefix)]
                if matching:
                    logger.info("Usando output del mes anterior como base: %s", matching[0])
                    return matching[0]
            # Fallback: if no name match or no prefix, use first candidate
            if candidates:
                logger.info("Usando output del mes anterior como base: %s", candidates[0])
                return candidates[0]

        # 2. Fall back to archivo_plantilla from config
        if config.archivo_plantilla:
            base_path = Path(config.archivo_plantilla)
            if base_path.exists():
                logger.info("Usando archivo_plantilla del config: %s", base_path)
                return base_path

        return None

    def generar_reporte(self, config: AvancesConfig) -> AvancesResult:
        if not config.nombre_archivo:
            raise ValueError(
                "nombre_archivo is required — used as the output filename "
                "(without extension)"
            )

        output_dir = (
            Path(config.output_dir)
            if config.output_dir
            else service_output_dir(self.SERVICE_SLUG, config.fecha_desde, "month")
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{config.nombre_archivo}.xlsx"

        base_path = self._resolve_base(config, output_dir)
        if base_path is None:
            raise FileNotFoundError(
                "No se encontro plantilla base ni output del mes anterior. "
                "Proporcione archivo_plantilla en el config."
            )

        # If output already exists, update in-place (preserving user customizations).
        # Otherwise, seed from the resolved base.
        if output_path.exists():
            logger.info("Archivo output existente — actualizando in-place: %s", output_path)
        else:
            shutil.copy2(str(base_path), str(output_path))
            logger.info("Archivo output creado desde base: %s", output_path)

        logger.info("Cargando workbook %s ...", output_path.name)
        t0 = time.perf_counter()
        wb = load_workbook(str(output_path), data_only=False, keep_links=False)
        logger.info("Workbook cargado en %.1fs", time.perf_counter() - t0)

        registros = {}

        for sc in PLANTILLA_SHEET_CONFIGS[config.tipo_plantilla]:
            if sc.sheet_name not in wb.sheetnames:
                logger.info("Sheet '%s' not found, creating", sc.sheet_name)
                ws = wb.create_sheet(sc.sheet_name)
                for col_idx, col_name in enumerate(sc.data_columns, 1):
                    ws.cell(row=sc.header_row, column=col_idx, value=col_name)

            params = {p: getattr(config, p) for p in sc.query_params}
            method = getattr(self.data_loader, sc.query_method)

            logger.info("Query %s(%s) ...", sc.query_method, params or "")
            t1 = time.perf_counter()
            df = method(**params) if params else method()
            logger.info("Query %s: %d filas en %.1fs", sc.query_method, len(df), time.perf_counter() - t1)

            # Apply column rename if the template uses different header names
            if sc.column_rename:
                df = df.rename(columns=sc.column_rename)

            t2 = time.perf_counter()
            rows = replace_sheet_data(
                wb, sc.sheet_name, df, sc.data_columns, sc.header_row
            )
            logger.info("Sheet '%s': %d filas escritas en %.1fs", sc.sheet_name, rows, time.perf_counter() - t2)
            registros[sc.sheet_name] = rows

        logger.info("Guardando workbook ...")
        t3 = time.perf_counter()
        wb.save(str(output_path))
        wb.close()
        logger.info("Guardado en %.1fs -> %s", time.perf_counter() - t3, output_path)

        return AvancesResult(ruta_archivo=output_path, registros_por_hoja=registros)
