"""
CoberturaService - Servicio para generacion de reportes de cobertura.

Orquesta el flujo completo: extraccion, procesamiento y generacion de Excel
para reportes de cobertura de preventistas por ruta.

Compara N periodos lado a lado. Los periodos se DERIVAN de `fecha_desde` con
offsets en meses (ver src.core.periodos), nunca se escriben en el config: el
daily patchea fechas pero no el resto del JSON, asi que un mes hardcodeado se
desincroniza solo al cambiar de mes.
"""
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.core.excel_slicers import agregar_slicers, slicers_disponibles
from src.core.excel_writer import ColumnFormat, ExcelWriter, SheetStyle
from src.core.output_paths import service_output_dir
from src.core.periodos import periodo_meses_atras
from src.services.base_service import BaseService
from src.services.cobertura.processor import procesar_cobertura

# Tipos de reporte de cobertura
TIPO_PREVENTISTA_GENERICO = "preventista_generico"
TIPO_PREVENTISTA_MARCA = "preventista_marca"
TIPO_SUCURSAL_MARCA = "sucursal_marca"

TIPOS_VALIDOS = (
    TIPO_PREVENTISTA_GENERICO,
    TIPO_PREVENTISTA_MARCA,
    TIPO_SUCURSAL_MARCA,
)

# Offsets en meses respecto del mes de `fecha_desde`. 13 = mismo mes del año
# anterior al mes cerrado; 1 = el mes cerrado. Es el caso de uso por defecto:
# cobertura del mes pasado contra la del mismo mes un año atras.
MESES_ATRAS_DEFAULT = (13, 1)

TOTAL_FILL = "FFE08A"  # ámbar — fila TOTAL GENERAL
AVISO_TOTAL = "suma de coberturas — no son clientes únicos"

# Columnas index por tipo (las que no son periodos)
_INDEX_COLUMNS = {
    TIPO_PREVENTISTA_GENERICO: ["sucursal", "vendedor", "id_ruta", "generico"],
    TIPO_PREVENTISTA_MARCA: ["sucursal", "vendedor", "id_ruta", "marca"],
    TIPO_SUCURSAL_MARCA: ["sucursal", "marca"],
}

# Columnas para slicers por tipo
_SLICER_COLUMNS = {
    TIPO_PREVENTISTA_GENERICO: ["sucursal", "vendedor", "generico"],
    TIPO_PREVENTISTA_MARCA: ["sucursal", "vendedor", "marca"],
    TIPO_SUCURSAL_MARCA: ["sucursal", "marca"],
}

# Nombres de hoja por tipo
_SHEET_NAMES = {
    TIPO_PREVENTISTA_GENERICO: "Cob Prev Generico",
    TIPO_PREVENTISTA_MARCA: "Cob Prev Marca",
    TIPO_SUCURSAL_MARCA: "Cob Sucursal Marca",
}


@dataclass
class ReporteCoberturaConfig:
    """Configuracion para generar un reporte de cobertura.

    Args:
        fecha_desde: Dia del que se derivan los periodos (YYYY-MM-DD).
        tipo: Apertura del informe (ver TIPOS_VALIDOS).
        meses_atras: Offsets en meses desde el mes de `fecha_desde`. Se ordenan
            de mayor a menor para que las columnas queden cronologicas.
        periodos: Escape hatch para uso ad-hoc. Si se pasa, gana sobre la
            derivacion. Los configs del daily NO deben usarlo.
        sucursales: Sucursales a filtrar (descripcion, no ID).
        nombre_archivo: Nombre de salida sin extension.
        con_slicers: Solo tiene efecto en Windows con Excel instalado.
    """
    fecha_desde: str
    tipo: str = TIPO_PREVENTISTA_GENERICO
    meses_atras: list[int] = field(default_factory=lambda: list(MESES_ATRAS_DEFAULT))
    periodos: list[str] | None = None
    sucursales: list[str] | None = None
    nombre_archivo: str | None = None
    con_slicers: bool = True

    def __post_init__(self):
        if self.tipo not in TIPOS_VALIDOS:
            raise ValueError(
                f"tipo de reporte no valido: {self.tipo!r}. Validos: {list(TIPOS_VALIDOS)}"
            )
        if self.periodos is None:
            self.periodos = self._derivar_periodos()
        if self.nombre_archivo is None:
            self.nombre_archivo = f"cobertura_{self.tipo}"

    def _derivar_periodos(self) -> list[str]:
        """Traduce los offsets en meses a periodos, en orden cronologico.

        Se deduplica: un periodo repetido haria que el pivot sume un mes sobre
        si mismo, y la cobertura no es aditiva entre periodos.
        """
        if not self.meses_atras:
            raise ValueError("meses_atras no puede estar vacio")
        offsets = sorted(set(self.meses_atras), reverse=True)
        return [periodo_meses_atras(self.fecha_desde, m) for m in offsets]


@dataclass
class ReporteCoberturaResult:
    """Resultado de la generacion de un reporte de cobertura."""
    ruta_archivo: Path
    registros_raw: int
    registros_procesados: int
    tipo: str
    periodos: list[str] = field(default_factory=list)
    slicers_agregados: bool = False


class CoberturaService(BaseService):
    """
    Servicio para generacion de reportes de cobertura.

    Soporta tres aperturas (ver TIPOS_VALIDOS):
    - preventista_generico: Cobertura por vendedor, ruta y generico
    - preventista_marca: Cobertura por vendedor, ruta y marca
    - sucursal_marca: Cobertura agregada por sucursal y marca
    """

    SERVICE_SLUG = "cobertura"
    GRANULARITY = "month"

    def generar_reporte(self, config: ReporteCoberturaConfig) -> ReporteCoberturaResult:
        """Genera un reporte de cobertura."""
        # 1. Extraer datos con joins
        df_raw = self._extraer_datos(config)

        # 2. Pivotear periodos como columnas
        index_cols = _INDEX_COLUMNS[config.tipo]
        df_procesado = procesar_cobertura(df_raw, index_cols)

        # 3. Escribir Excel bajo data/output/{slug}/{YYYY-MM}/
        sheet_name = _SHEET_NAMES[config.tipo]
        out_dir = service_output_dir(
            self.SERVICE_SLUG, config.fecha_desde, granularity=self.GRANULARITY
        )
        out_dir.mkdir(parents=True, exist_ok=True)

        style = self._crear_estilo(df_procesado, config.tipo)
        writer = ExcelWriter(config.nombre_archivo, output_dir=out_dir, style=style)
        ws = writer.add_sheet(df_procesado, sheet_name=sheet_name, style=style)
        self._escribir_total_general(ws, df_procesado, index_cols)
        ruta = writer.save()

        # 4. Agregar slicers
        slicers_ok = False
        if config.con_slicers and slicers_disponibles():
            nombre_tabla = f"Tabla_{sheet_name.replace(' ', '_')}"
            columnas_slicer = _SLICER_COLUMNS.get(config.tipo, [])
            slicers_ok = agregar_slicers(ruta, nombre_tabla, columnas_slicer)

        return ReporteCoberturaResult(
            ruta_archivo=ruta,
            registros_raw=len(df_raw),
            registros_procesados=len(df_procesado),
            tipo=config.tipo,
            periodos=list(config.periodos),
            slicers_agregados=slicers_ok
        )

    def _extraer_datos(self, config: ReporteCoberturaConfig):
        """Extrae datos de cobertura segun el tipo."""
        if config.tipo == TIPO_PREVENTISTA_GENERICO:
            return self.data_loader.get_cobertura_preventista_generico(
                periodos=config.periodos, sucursales=config.sucursales
            )
        elif config.tipo == TIPO_PREVENTISTA_MARCA:
            return self.data_loader.get_cobertura_preventista_marca(
                periodos=config.periodos, sucursales=config.sucursales
            )
        elif config.tipo == TIPO_SUCURSAL_MARCA:
            return self.data_loader.get_cobertura_sucursal_marca(
                periodos=config.periodos, sucursales=config.sucursales
            )
        else:
            raise ValueError(f"Tipo de reporte no valido: {config.tipo}")

    @staticmethod
    def _columnas_periodos(df: pd.DataFrame, index_cols: list[str]) -> list[str]:
        """Columnas de periodos = todas las que no son index."""
        return [c for c in df.columns if c not in index_cols]

    def _crear_estilo(self, df, tipo: str) -> SheetStyle:
        """Crea estilo Excel para el reporte de cobertura."""
        column_formats = {
            col: ColumnFormat(number_format='#,##0')
            for col in self._columnas_periodos(df, _INDEX_COLUMNS[tipo])
        }
        return SheetStyle(
            column_formats=column_formats,
            as_table=True,
            table_style="TableStyleMedium9"
        )

    def _escribir_total_general(
        self, ws: Worksheet, df: pd.DataFrame, index_cols: list[str]
    ) -> None:
        """Fila TOTAL GENERAL, una fila debajo de la tabla.

        Va FUERA del rango de la tabla: adentro, el autofiltro y el ordenamiento
        de Excel la arrastrarian como una fila de datos mas.

        El total cruza genericos/marcas, y la cobertura NO es aditiva en ese eje
        (el mismo cliente compra varias marcas). Es una suma de coberturas, no un
        conteo de clientes distintos, y la fila lo dice: contar clientes unicos
        exigiria volver al grano de `fact_ventas`, que estas tablas ya agregaron.
        """
        if df.empty:
            return

        periodos = self._columnas_periodos(df, index_cols)
        fila = ws.max_row + 1
        fill = PatternFill("solid", fgColor=TOTAL_FILL)
        bold = Font(bold=True)

        etiquetas = ["TOTAL GENERAL"] + [AVISO_TOTAL] + [""] * (len(index_cols) - 2)
        for col, texto in enumerate(etiquetas, 1):
            celda = ws.cell(fila, col, texto or None)
            celda.font = bold
            celda.fill = fill

        for offset, periodo in enumerate(periodos):
            # `.item()` desempaqueta el escalar de numpy a int/float de Python
            # sin tocar el valor — openpyxl no serializa np.int64.
            celda = ws.cell(fila, len(index_cols) + offset + 1, df[periodo].sum().item())
            celda.number_format = '#,##0'
            celda.font = bold
            celda.fill = fill

    def listar_sucursales(self) -> list[str]:
        """Obtiene lista de sucursales disponibles."""
        df = self.data_loader.get_sucursales()
        return df["sucursal"].tolist()
