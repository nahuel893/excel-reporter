"""CoberturaAguasService — cobertura de aguas por sucursal y marca.

Abre los ultimos N meses (2 por defecto) con su cobertura mensual, el acumulado
de la ventana, y dos pesos: cuanto pesa cada marca sobre el acumulado de aguas y
cuanto pesa sobre el padron de clientes no anulados.

Las ventanas se DERIVAN de `fecha`, nunca se escriben en el config: el daily
patchea fechas pero no el resto del JSON, asi que un mes a mano se desincroniza
solo al cambiar de mes (ver src.core.periodos).
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.output_paths import service_output_dir
from src.core.periodos import etiqueta_mes, periodo_meses_atras
from src.services.base_service import BaseService

from .constants import (
    ID_FUERZA_VENTAS_PREVENTA,
    MARCAS_AGUAS,
    RUTAS_EXCLUIDAS,
    TOTAL_AGUAS,
)
from .processor import COLUMNAS_VENTAS, construir_tabla

logger = logging.getLogger(__name__)

HEADER_FILL = "2E75B6"
HEADER_FONT = "FFFFFF"
SUC_FILL = "DDEBF7"      # cabecera de cada sucursal
GRUPO_FILL = "F2F2F2"    # AGUA MINERAL / AGUA SABORIZADA
TOTAL_SUC_FILL = "FFF2CC"  # TOTAL AGUAS de la sucursal
TOTAL_GRAL_FILL = "FFE08A"  # TOTAL GENERAL


@dataclass
class CoberturaAguasConfig:
    """Config del informe.

    Args:
        fecha: dia de corte. El informe cubre el mes de esta fecha (parcial hasta
            el dia) y los `meses - 1` meses anteriores completos.
        meses: cantidad de meses a abrir. 2 = mes anterior + mes en curso.
        umbral: piso de bultos para dar por cubierto a un cliente. Por defecto
            `> 0`, que es el criterio general y el que usan las tablas del ETL.
        nombre_archivo: nombre de salida sin extension.
    """
    fecha: str
    meses: int = 2
    umbral: float = 0.0
    nombre_archivo: str | None = None


@dataclass
class CoberturaAguasResult:
    ruta_archivo: Path
    sucursales: int
    meses: list[str]
    cobertura_acumulada: int
    padron: int


def _thin() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class CoberturaAguasService(BaseService):
    """Genera el informe de cobertura de aguas."""

    SERVICE_SLUG = "cobertura-aguas"
    GRANULARITY = "month"

    def _meses(self, config: CoberturaAguasConfig) -> list[str]:
        """Los meses de la ventana como 'YYYY-MM', del mas viejo al mas nuevo."""
        if config.meses < 1:
            raise ValueError(f"meses debe ser >= 1, recibido {config.meses}")
        return [
            periodo_meses_atras(config.fecha, i)[:7]
            for i in reversed(range(config.meses))
        ]

    def _fetch(self, config: CoberturaAguasConfig, meses: list[str]) -> pd.DataFrame:
        desde = f"{meses[0]}-01"
        ventas = self.data_loader.get_ventas_cliente_marca_mes(
            marcas=list(MARCAS_AGUAS),
            fecha_desde=desde,
            fecha_hasta=config.fecha,
            id_fuerza_ventas=ID_FUERZA_VENTAS_PREVENTA,
            rutas_excluidas=list(RUTAS_EXCLUIDAS),
        )
        if ventas is None or ventas.empty:
            logger.warning("Sin ventas de aguas entre %s y %s", desde, config.fecha)
            return pd.DataFrame(columns=COLUMNAS_VENTAS)
        return ventas

    def generar_reporte(self, config: CoberturaAguasConfig) -> CoberturaAguasResult:
        meses = self._meses(config)
        ventas = self._fetch(config, meses)
        # El MISMO conjunto de rutas que en `_fetch`: numerador y denominador
        # del peso sobre padron tienen que salir del mismo universo.
        padron = self.data_loader.get_padron_activo(
            rutas_excluidas=list(RUTAS_EXCLUIDAS)
        )
        if padron is None or padron.empty:
            raise RuntimeError("El padron vino vacio: sin denominador no hay informe")

        tabla = construir_tabla(ventas, padron, meses, umbral=config.umbral)

        nombre = config.nombre_archivo or f"Cobertura Aguas - {config.fecha}"
        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(config, meses, tabla, ruta)

        total = tabla[tabla.es_total_general & (tabla.fila == TOTAL_AGUAS)].iloc[0]
        return CoberturaAguasResult(
            ruta_archivo=ruta,
            sucursales=int(padron["id_sucursal"].nunique()),
            meses=meses,
            cobertura_acumulada=int(total["cob_acum"]),
            padron=int(padron["padron"].sum()),
        )

    # --- Excel --------------------------------------------------------------

    def _build_workbook(
        self,
        config: CoberturaAguasConfig,
        meses: list[str],
        tabla: pd.DataFrame,
        ruta: Path,
    ) -> None:
        wb = Workbook()
        # El resumen va primero porque es la hoja que se fotografia: el detalle
        # son 120 filas y en una imagen de WhatsApp no se lee.
        self._hoja_resumen(wb.active, config, meses, tabla)
        self._hoja_detalle(wb.create_sheet("Detalle"), config, meses, tabla)
        self._hoja_criterio(wb.create_sheet("Criterio"), config, meses)
        wb.save(ruta)

    def _hoja_resumen(
        self,
        ws: Worksheet,
        config: CoberturaAguasConfig,
        meses: list[str],
        tabla: pd.DataFrame,
    ) -> None:
        """Una fila por sucursal arriba, una fila por marca abajo. Para leer de un vistazo."""
        ws.title = "Resumen"
        border = _thin()
        etiquetas = [etiqueta_mes(f"{m}-01") for m in meses]
        cob_cols = [f"cob_{m}" for m in meses]

        ws["A1"] = "Cobertura de Aguas — Resumen"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            f"Clientes distintos con compra neta > {config.umbral:g}  |  "
            f"Corte al {config.fecha}  |  Acumulado = neto de la ventana completa"
        )
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        def encabezado(fila: int, headers: list[str]) -> None:
            for j, h in enumerate(headers, 1):
                c = ws.cell(fila, j, h)
                c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
                c.font = Font(bold=True, color=HEADER_FONT)
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                c.border = border

        def pintar(fila: int, n: int, relleno: str | None, negrita: bool) -> None:
            for j in range(1, n + 1):
                c = ws.cell(fila, j)
                c.border = border
                if j >= 2:
                    c.alignment = Alignment(horizontal="right")
                if negrita:
                    c.font = Font(bold=True)
                if relleno:
                    c.fill = PatternFill("solid", start_color=relleno, end_color=relleno)

        # --- Bloque 1: por sucursal -----------------------------------------
        heads1 = (
            ["Sucursal", "Padron"] + etiquetas
            + ["ACUM AGUAS", "% s/ Padron",
               "AGUA MINERAL", "% s/ Aguas", "AGUA SABORIZADA", "% s/ Aguas"]
        )
        r = 4
        encabezado(r, heads1)
        r += 1
        for suc in tabla["sucursal"].unique():
            b = tabla[tabla["sucursal"] == suc].set_index("fila")
            ag, mi, sa = b.loc[TOTAL_AGUAS], b.loc["AGUA MINERAL"], b.loc["AGUA SABORIZADA"]
            valores = (
                [suc, int(ag["padron"])]
                + [int(ag[c]) for c in cob_cols]
                + [int(ag["cob_acum"]), float(ag["pct_padron"]),
                   int(mi["cob_acum"]), float(mi["pct_acum"]),
                   int(sa["cob_acum"]), float(sa["pct_acum"])]
            )
            for j, v in enumerate(valores, 1):
                c = ws.cell(r, j, v)
                if isinstance(v, float):
                    c.number_format = "0.0%"
                elif isinstance(v, int):
                    c.number_format = "#,##0"
            es_tg = suc == "TOTAL GENERAL"
            pintar(r, len(heads1), TOTAL_GRAL_FILL if es_tg else None, es_tg)
            r += 1

        # --- Bloque 2: por marca, consolidado -------------------------------
        r += 2
        ws.cell(r, 1, "Consolidado por marca").font = Font(bold=True, size=12)
        r += 1
        heads2 = ["Marca"] + etiquetas + ["ACUM", "% s/ Aguas", "% s/ Padron"]
        encabezado(r, heads2)
        r += 1
        for _, f in tabla[tabla.es_total_general].iterrows():
            valores = (
                [f["fila"]] + [int(f[c]) for c in cob_cols]
                + [int(f["cob_acum"]), float(f["pct_acum"]), float(f["pct_padron"])]
            )
            for j, v in enumerate(valores, 1):
                c = ws.cell(r, j, v)
                if isinstance(v, float):
                    c.number_format = "0.0%"
                elif isinstance(v, int):
                    c.number_format = "#,##0"
            tipo = f["tipo"]
            relleno = TOTAL_GRAL_FILL if tipo == "total" else (
                GRUPO_FILL if tipo == "grupo" else None
            )
            pintar(r, len(heads2), relleno, tipo in ("grupo", "total"))
            r += 1

        ws.column_dimensions["A"].width = 30
        for j in range(2, len(heads1) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 15

    def _hoja_detalle(
        self,
        ws: Worksheet,
        config: CoberturaAguasConfig,
        meses: list[str],
        tabla: pd.DataFrame,
    ) -> None:
        border = _thin()
        etiquetas = [etiqueta_mes(f"{m}-01") for m in meses]
        acum_txt = f"ACUM {etiquetas[0].split()[0]}-{etiquetas[-1]}" if len(meses) > 1 \
            else f"ACUM {etiquetas[0]}"

        # (encabezado, clave en la tabla, formato). Explicito y no por posicion:
        # calcular que columna es porcentaje con aritmetica de indices se rompe
        # solo en cuanto se agrega una columna al medio.
        columnas: list[tuple[str, str, str]] = (
            [(e, f"cob_{m}", "#,##0") for e, m in zip(etiquetas, meses)]
            + [
                (acum_txt, "cob_acum", "#,##0"),
                # El denominador de "% s/ Aguas", al lado del porcentaje, igual
                # que Padron. Sin el hay que ir a buscar la fila TOTAL AGUAS.
                ("Total Aguas", "base_aguas", "#,##0"),
                ("% s/ Aguas", "pct_acum", "0.0%"),
                ("Padron", "padron", "#,##0"),
                ("% s/ Padron", "pct_padron", "0.0%"),
            ]
        )
        headers = ["Sucursal", "Marca"] + [h for h, _, _ in columnas]
        n_cols = len(headers)

        ws["A1"] = "Cobertura de Aguas por Sucursal y Marca"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            f"Cobertura = clientes distintos con compra neta > {config.umbral:g} "
            f"en el corte  |  Corte al {config.fecha}  |  "
            f"Padron = dim_cliente no anulados  |  "
            f"Los grupos son la UNION de sus marcas, no la suma"
        )
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        r = 4
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h)
            c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = border
        fila_header = r
        r += 1

        cob_cols = [f"cob_{m}" for m in meses]
        suc_previa = object()
        for _, f in tabla.iterrows():
            es_tg = bool(f["es_total_general"])
            # La sucursal se escribe una sola vez por bloque: repetirla en las
            # ocho filas convierte la columna en ruido visual.
            if f["sucursal"] != suc_previa:
                ws.cell(r, 1, f["sucursal"])
                suc_previa = f["sucursal"]
            ws.cell(r, 2, f["fila"])
            for j, (_, clave, fmt) in enumerate(columnas, 3):
                ws.cell(r, j, float(f[clave]) if fmt.endswith("%") else int(f[clave]))

            relleno = None
            if es_tg:
                relleno = TOTAL_GRAL_FILL
            elif f["tipo"] == "total":
                relleno = TOTAL_SUC_FILL
            elif f["tipo"] == "grupo":
                relleno = GRUPO_FILL

            negrita = es_tg or f["tipo"] in ("grupo", "total")
            for j in range(1, n_cols + 1):
                c = ws.cell(r, j)
                c.border = border
                if j >= 3:
                    c.number_format = columnas[j - 3][2]
                    c.alignment = Alignment(horizontal="right")
                if negrita:
                    c.font = Font(bold=True)
                if relleno:
                    c.fill = PatternFill("solid", start_color=relleno, end_color=relleno)
            r += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20
        for j in range(3, n_cols + 1):
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.freeze_panes = ws.cell(row=fila_header + 1, column=3).coordinate
        ws.auto_filter.ref = f"A{fila_header}:{get_column_letter(n_cols)}{r - 1}"

    def _hoja_criterio(
        self, ws: Worksheet, config: CoberturaAguasConfig, meses: list[str]
    ) -> None:
        """Deja escrito como se conto, para que el numero se pueda auditar."""
        etiquetas = " y ".join(etiqueta_mes(f"{m}-01") for m in meses)
        lineas = [
            ("Criterio de calculo", True),
            ("", False),
            ("Cobertura = clientes DISTINTOS que compraron, en cuatro pasos:", False),
            ("  1. definir el corte (sucursal, marca o grupo, mes)", False),
            ("  2. totalizar por cliente DENTRO del corte", False),
            (f"  3. filtrar los que superan el umbral (> {config.umbral:g})", False),
            ("  4. contar los que quedan", False),
            ("", False),
            ("Agrupar antes de filtrar es lo que evita contar al cliente que", False),
            ("devolvio todo lo que compro, y evita descartar al que llego al", False),
            ("umbral sumando varias compras chicas.", False),
            ("", False),
            ("Clave del cliente: (id_cliente, id_sucursal). El id se reusa entre", False),
            ("sucursales, asi que solo no alcanza para identificarlo.", False),
            ("", False),
            ("Aditividad", True),
            ("  SI se suma  : entre sucursales, rutas y preventistas", False),
            ("  NO se suma  : entre marcas, grupos ni meses", False),
            ("Un grupo (AGUA MINERAL, AGUA SABORIZADA) es la UNION de los", False),
            ("clientes de sus marcas: el que compra VILLA DEL SUR y", False),
            ("VILLAVICENCIO cuenta una sola vez.", False),
            ("", False),
            ("Acumulado", True),
            (f"  Ventana: {etiquetas} (el ultimo mes va hasta el {config.fecha})", False),
            ("  El acumulado totaliza el neto sobre la ventana COMPLETA y recien", False),
            ("  ahi filtra. No es la union de los conjuntos mensuales: el cliente", False),
            ("  que compra 5 en un mes y devuelve 5 en el otro queda cubierto en", False),
            ("  el primero y fuera del acumulado.", False),
            ("", False),
            ("Fuente", True),
            ("  gold.fact_ventas con cantidades_total (incluye bonificados: si el", False),
            ("  producto llego al pdv, ese pdv esta cubierto), anulado = false y", False),
            (f"  fuerza de ventas {ID_FUERZA_VENTAS_PREVENTA} (preventa).", False),
            ("  Con ese filtro el conteo reproduce gold.cob_sucursal_marca exacto:", False),
            ("  julio-2026, aguas, 23.748 contra 23.748 — 0 filas de 65 con", False),
            ("  diferencia. El acumulado no se puede leer de ninguna tabla cob_*", False),
            ("  porque son mensuales y la cobertura no se suma entre meses.", False),
            ("", False),
            ("Padron", True),
            ("  gold.dim_cliente con anulado = false. Es SCD tipo 1: guarda el", False),
            ("  estado de HOY, no el que tenia el cliente durante el mes medido.", False),
            ("", False),
            ("Marcas", True),
            ("  AGUA MINERAL    : VILLA DEL SUR, VILLAVICENCIO", False),
            ("  AGUA SABORIZADA : LEVITE, BRIO", False),
            ("  FULL SPORT entra en TOTAL AGUAS pero en ningun grupo: es", False),
            ("  isotonica, no agua saborizada.", False),
        ]
        for i, (texto, es_titulo) in enumerate(lineas, 1):
            c = ws.cell(i, 1, texto)
            if es_titulo:
                c.font = Font(bold=True, size=12, color=HEADER_FILL)
        ws.column_dimensions["A"].width = 78

    def run(self, config: CoberturaAguasConfig) -> CoberturaAguasResult:
        return self.generar_reporte(config)
