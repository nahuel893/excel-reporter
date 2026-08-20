"""CoberturaCuposService — cobertura por generico y marca para repartir cupos.

Una hoja por zona. Dentro de cada hoja, un bloque por generico CCU: sus marcas
ordenadas de mayor a menor cobertura, y al pie la fila TOTAL del generico.

Ese total NO es la suma de las marcas de arriba. Es una consulta aparte contra
el grano de generico, porque la cobertura son clientes DISTINTOS y el mismo
cliente compra varias marcas: sumar las filas de PERNOD en julio-2026 da 1083
contra los 721 reales. La columna Cupo queda vacia, para cargar a mano.

Las dos ventanas se DERIVAN de `fecha_desde` (mes anterior cerrado y ese mismo
mes un anio atras), nunca se escriben en el config: el daily patchea fechas pero
no el resto del JSON, asi que un mes a mano se desincroniza al cambiar de mes.
"""
import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.output_paths import service_output_dir
from src.core.periodos import etiqueta_mes, periodo_meses_atras
from src.services.base_service import BaseService
from src.services.cobertura_cupos.constants import (
    GENERICOS_CCU,
    Zona,
    zonas_desde_sucursales,
    zonas_por_defecto,
)

logger = logging.getLogger(__name__)

HEADER_FILL, HEADER_FONT = "4472C4", "FFFFFF"
GENERICO_FILL, GENERICO_FONT = "D9E1F2", "1F3864"  # banda de titulo del bloque
TOTAL_FILL = "FFE08A"                              # ambar — fila de total
CUPO_FILL = "FFF9E6"                               # columna a completar a mano

# Offsets en meses respecto del mes de `fecha_desde`, en el ORDEN en que se
# leen las columnas — que no es el cronologico:
#   1  -> MES ANTERIOR cerrado: donde se para el analisis.
#   12 -> MISMO MES DEL ANIO ANTERIOR: el mes EN CURSO un anio atras. Para una
#         corrida de agosto-2026 es agosto-2025, la referencia estacional del
#         mes que se esta cupificando. NO es 13 (que seria julio-2025, el
#         interanual del mes cerrado): el cupo se reparte contra el agosto que
#         realmente paso, no contra el julio anterior.
MESES_ATRAS_DEFAULT: tuple[int, ...] = (1, 12)

_ANCHO_MARCA, _ANCHO_PERIODO, _ANCHO_CUPO = 34, 20, 14


@dataclass
class CoberturaCuposConfig:
    """Config del informe.

    Args:
        fecha_desde: Dia del que se derivan las ventanas (YYYY-MM-DD).
        meses_atras: Offsets en meses desde el mes de `fecha_desde`, una columna
            por offset. Se respeta el ORDEN de la lista: aca el orden es una
            decision de lectura, no cronologica, asi que NO se reordena.
        genericos: Genericos a incluir. None -> los 5 CCU.
        zonas: Zonas a abrir, ya resueltas. Gana sobre `sucursales`.
        sucursales: Nombres de sucursal, una hoja entera cada una. El id se
            resuelve contra la BD en el servicio. Sirve para el interior, que
            no tiene zonas virtuales. None -> las tres zonas por defecto.
        nombre_archivo: Nombre de salida sin extension.
    """
    fecha_desde: str
    meses_atras: list[int] | None = None
    genericos: list[str] | None = None
    zonas: list[Zona] | None = None
    sucursales: list[str] | None = None
    nombre_archivo: str | None = None

    def __post_init__(self):
        if self.meses_atras is None:
            self.meses_atras = list(MESES_ATRAS_DEFAULT)
        if not self.meses_atras:
            raise ValueError("meses_atras no puede estar vacio")
        if len(set(self.meses_atras)) != len(self.meses_atras):
            raise ValueError(f"meses_atras tiene offsets repetidos: {self.meses_atras}")
        if not self.genericos:
            self.genericos = list(GENERICOS_CCU)
        # Con `sucursales` las zonas se resuelven en el servicio (necesita BD
        # para el id_sucursal). Dejarlas en None es la senal de "pendiente".
        if not self.zonas and not self.sucursales:
            self.zonas = zonas_por_defecto()
        if self.nombre_archivo is None:
            self.nombre_archivo = f"Cobertura y Cupos {etiqueta_mes(self.periodo_principal)}"

    @property
    def periodos(self) -> list[str]:
        """Un periodo por offset, en el orden de `meses_atras`."""
        return [periodo_meses_atras(self.fecha_desde, m) for m in self.meses_atras]

    @property
    def periodo_principal(self) -> str:
        """El primero — el mes sobre el que se lee el informe."""
        return self.periodos[0]


@dataclass
class BloqueGenerico:
    """Un generico con sus marcas y su total propio.

    `totales` tiene un valor por periodo, en el mismo orden que las columnas
    `cob_0..cob_N` de `marcas`. Cada uno sale de SU consulta contra el grano de
    generico — nunca de sumar la columna de arriba.
    """
    generico: str
    marcas: pd.DataFrame          # [marca, cob_0, cob_1, ...]
    totales: list[int]


@dataclass
class CoberturaCuposResult:
    ruta_archivo: Path
    periodos: list[str] = field(default_factory=list)
    zonas: list[str] = field(default_factory=list)
    filas_marca: int = 0

    @property
    def periodo_principal(self) -> str:
        return self.periodos[0] if self.periodos else ""


class CoberturaCuposService(BaseService):
    SERVICE_SLUG = "cobertura-cupos"
    GRANULARITY = "month"

    def _resolver_zonas(self, config: CoberturaCuposConfig) -> list[Zona]:
        """Zonas explicitas, o una por sucursal resuelta contra la BD."""
        if config.zonas:
            return config.zonas
        mapa = self.data_loader.get_mapa_sucursales()
        return zonas_desde_sucursales(config.sucursales, mapa)

    def generar_reporte(self, config: CoberturaCuposConfig) -> CoberturaCuposResult:
        config.zonas = self._resolver_zonas(config)
        bloques_por_zona = {
            zona.nombre: [self._bloque(config, zona, g) for g in config.genericos]
            for zona in config.zonas
        }

        out_dir = service_output_dir(
            self.SERVICE_SLUG, config.fecha_desde, granularity=self.GRANULARITY
        )
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{config.nombre_archivo}.xlsx"
        self._build_workbook(config, bloques_por_zona, ruta)

        return CoberturaCuposResult(
            ruta_archivo=ruta,
            periodos=list(config.periodos),
            zonas=[z.nombre for z in config.zonas],
            filas_marca=sum(
                len(b.marcas) for bloques in bloques_por_zona.values() for b in bloques
            ),
        )

    def _bloque(
        self, config: CoberturaCuposConfig, zona: Zona, generico: str
    ) -> BloqueGenerico:
        """Trae las N ventanas de un generico en una zona y las combina.

        El total del generico sale de su propia consulta contra el grano de
        generico, NO de sumar `marcas`.
        """
        kwargs = dict(
            id_sucursal=zona.id_sucursal,
            rutas_incluidas=list(zona.rutas_incluidas) if zona.rutas_incluidas is not None else None,
            rutas_excluidas=list(zona.rutas_excluidas),
        )
        periodos = config.periodos
        universo = self.data_loader.get_marcas_de_generico(generico)
        return BloqueGenerico(
            generico=generico,
            marcas=self._combinar(
                [
                    self.data_loader.get_cobertura_marca_de_generico_zona(
                        generico=generico, periodo=p, **kwargs
                    )
                    for p in periodos
                ],
                universo=universo,
            ),
            totales=[
                self.data_loader.get_cobertura_generico_zona(
                    generico=generico, periodo=p, **kwargs
                )
                for p in periodos
            ],
        )

    @staticmethod
    def _combinar(
        frames: list[pd.DataFrame], universo: pd.DataFrame | None = None
    ) -> pd.DataFrame:
        """Cruza las N ventanas por marca, una columna `cob_i` por periodo.

        El eje de filas es el UNIVERSO de marcas del generico, no las que
        tuvieron cobertura: la hoja de cada zona lleva la lista completa, con 0
        donde no vendio. Asi las tres zonas son comparables fila a fila y se le
        puede asignar cupo a una marca que hoy no vende en esa zona.

        Las columnas se nombran por POSICION, no por etiqueta de mes: dos
        periodos podrian resolver a la misma etiqueta y una colision
        sobreescribiria una columna en silencio.
        """
        vacio = pd.DataFrame(columns=["marca", "cobertura"])
        columnas = [f"cob_{i}" for i in range(len(frames))]

        base = (
            universo[["marca"]].drop_duplicates()
            if universo is not None and not universo.empty
            else None
        )
        wide = base
        for i, f in enumerate(frames):
            usable = f if f is not None and not f.empty else vacio
            parcial = usable[["marca", "cobertura"]].rename(columns={"cobertura": columnas[i]})
            # Outer, no left: el universo AMPLIA el eje de filas, nunca lo
            # recorta. Si el maestro no tuviera una marca que igual vendio, esa
            # cobertura no puede desaparecer del informe.
            wide = parcial if wide is None else wide.merge(parcial, on="marca", how="outer")

        assert wide is not None  # `frames` nunca viene vacio: meses_atras lo garantiza
        if wide.empty:
            return wide
        wide[columnas] = wide[columnas].fillna(0)
        # Orden por el primer periodo: es el mes sobre el que se lee el informe.
        # Las marcas sin cobertura caen al final, alfabeticas.
        return wide.sort_values(
            [columnas[0], "marca"], ascending=[False, True], kind="mergesort"
        ).reset_index(drop=True)

    # ------------------------------------------------------------------ Excel

    def _build_workbook(self, config, bloques_por_zona: dict, ruta: Path) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        for nombre_zona, bloques in bloques_por_zona.items():
            self._escribir_hoja(wb, config, nombre_zona, bloques)
        wb.save(ruta)

    def _escribir_hoja(self, wb: Workbook, config, nombre_zona: str, bloques: list) -> None:
        # El nombre de hoja de Excel admite 31 caracteres.
        ws = wb.create_sheet(title=nombre_zona[:31])
        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        etiquetas = [etiqueta_mes(p) for p in config.periodos]

        ws["A1"] = f"Cobertura por Genérico y Marca — {nombre_zona}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (
            f"{' vs '.join(etiquetas)}  |  Cobertura = clientes compradores  |  "
            "El total del genérico NO es la suma de sus marcas"
        )
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        encabezados = ["Marca"] + [f"Cob {e}" for e in etiquetas] + ["Cupo"]
        fila = 4
        for bloque in bloques:
            fila = self._escribir_bloque(ws, bloque, encabezados, fila, border)

        anchos = [_ANCHO_MARCA] + [_ANCHO_PERIODO] * len(etiquetas) + [_ANCHO_CUPO]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
        ws.freeze_panes = "A4"

    def _escribir_bloque(
        self, ws, bloque: BloqueGenerico, encabezados: list[str], fila: int, border
    ) -> int:
        """Escribe un generico: banda, header, marcas y su fila de total."""
        # Banda con el nombre del generico
        ws.cell(fila, 1, bloque.generico).font = Font(bold=True, size=12, color=GENERICO_FONT)
        for col in range(1, len(encabezados) + 1):
            ws.cell(fila, col).fill = PatternFill("solid", fgColor=GENERICO_FILL)
            ws.cell(fila, col).border = border
        fila += 1

        for col, texto in enumerate(encabezados, 1):
            celda = ws.cell(fila, col, texto)
            celda.fill = PatternFill("solid", fgColor=HEADER_FILL)
            celda.font = Font(bold=True, color=HEADER_FONT)
            celda.alignment = Alignment(horizontal="center")
            celda.border = border
        fila += 1

        col_cupo = len(encabezados)
        for _, marca in bloque.marcas.iterrows():
            ws.cell(fila, 1, marca["marca"]).border = border
            for i in range(len(bloque.totales)):
                self._celda_num(ws, fila, 2 + i, marca[f"cob_{i}"], border)
            # Columna de cupo: vacía a propósito, se carga a mano.
            celda_cupo = ws.cell(fila, col_cupo)
            celda_cupo.fill = PatternFill("solid", fgColor=CUPO_FILL)
            celda_cupo.border = border
            fila += 1

        # Total del generico — de consulta propia, no de sumar las filas de arriba.
        ws.cell(fila, 1, f"TOTAL {bloque.generico}").font = Font(bold=True)
        for i, total in enumerate(bloque.totales):
            self._celda_num(ws, fila, 2 + i, total, border, bold=True)
        for col in range(1, len(encabezados) + 1):
            ws.cell(fila, col).fill = PatternFill("solid", fgColor=TOTAL_FILL)
            ws.cell(fila, col).border = border
            ws.cell(fila, col).font = Font(bold=True)
        return fila + 2  # una fila en blanco entre bloques

    @staticmethod
    def _celda_num(ws, fila: int, col: int, valor, border, bold: bool = False) -> None:
        # El valor va tal cual: nunca se trunca con int(), el formato lo pone
        # `number_format`. `.item()` solo desempaqueta el escalar de numpy a uno
        # de Python (openpyxl no serializa np.int64), sin tocar el valor.
        celda = ws.cell(fila, col, valor.item() if hasattr(valor, "item") else valor)
        celda.number_format = "#,##0"
        celda.alignment = Alignment(horizontal="right")
        celda.border = border
        if bold:
            celda.font = Font(bold=True)

    def run(self, config):
        return self.generar_reporte(config)
