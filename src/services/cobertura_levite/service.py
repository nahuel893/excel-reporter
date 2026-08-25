"""CoberturaLeviteService — cobertura abierta por calibre y marca.

Una hoja por generico (AGUAS DANONE y CERVEZAS), y en cada una el mismo cuadro
repetido para TRES periodos, de arriba hacia abajo:

1. la ventana del config (el mes en curso, casi siempre parcial),
2. el mes anterior completo,
3. el mismo mes del anio anterior, completo.

Los dos historicos se DERIVAN de `fecha_hasta`, nunca se escriben en el config.
Van como meses calendario enteros: son periodos cerrados y quedan estables toda
la corrida del mes, a diferencia del primero, que crece dia a dia. El encabezado
de cada cuadro dice hasta que dia llega para que la diferencia no se lea como
una caida del negocio.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.data_loader import DataLoader
from src.core.output_paths import service_output_dir
from src.core.periodos import etiqueta_mes, periodo_meses_atras, rango_mes
from src.services.base_service import BaseService

from .processor import (
    CUADROS,
    Cuadro,
    extraer_calibre,
    matriz_calibre_marca,
    ordenar_calibres,
    procesar_cobertura_sucursal_calibre,
)

logger = logging.getLogger(__name__)

# Paleta de colores institucional
HEADER_FILL = "1F4E78"       # Azul corporativo oscuro
HEADER_FONT = "FFFFFF"       # Blanco
SUC_FILL = "DDEBF7"          # Azul claro
SUBTOTAL_FILL = "F2F2F2"     # Gris suave
TOTAL_SUC_FILL = "FFF2CC"    # Amarillo pastel
TOTAL_GRAL_FILL = "FFE08A"   # Dorado / Ambar para totales generales
ACCENT_FILL = "E2EFDA"       # Verde pastel para totales de fila


@dataclass
class CoberturaLeviteConfig:
    """Configuracion del reporte de cobertura de Levite por calibre."""
    fecha_desde: str
    fecha_hasta: str
    umbral: float = 0.0
    nombre_archivo: str | None = None
    # Sucursales del informe, por id. None = todas. El padron se filtra con el
    # MISMO criterio: es el denominador del % y tiene que salir del mismo
    # universo que las ventas.
    sucursales: list[int] | None = None


@dataclass
class CoberturaLeviteResult:
    """Resultado de la generacion del reporte."""
    ruta_archivo: Path
    sucursales: int
    clientes_compradores: int
    volumen_total: float
    padron_total: int


def _thin_border() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class CoberturaLeviteService(BaseService):
    """Servicio para generar el reporte de cobertura de Levite por calibre."""

    SERVICE_SLUG = "cobertura-levite"
    GRANULARITY = "month"

    def _fetch_ventas(
        self, fecha_desde: str, fecha_hasta: str,
        genericos: list[str] | None = None, sucursales: list[int] | None = None,
    ) -> pd.DataFrame:
        """Ventas con joins compuestos y clasificacion por calibre.

        El recorte es por GENERICO, no por marca: el total de cada cuadro es la
        cobertura del generico entero, incluidas las marcas que no tienen
        columna propia (SCHNEIDER no es una de las cuatro principales pero es
        CERVEZAS). Cada cuadro filtra despues lo suyo sobre el mismo DataFrame:
        una consulta por marca seria el mismo trabajo repetido, y podrian caer
        en momentos distintos y no cerrar entre si.
        """
        query = """
        SELECT
            fv.id_sucursal,
            ds.descripcion AS sucursal,
            fv.id_cliente,
            COALESCE(
                NULLIF(TRIM(dc.fantasia), ''),
                NULLIF(TRIM(dc.razon_social), ''),
                CONCAT('CLIENTE ', fv.id_cliente)
            ) AS cliente,
            COALESCE(dc.id_ruta_fv1, 0) AS id_ruta,
            COALESCE(dv.des_vendedor, 'SIN VENDEDOR') AS vendedor,
            fv.id_articulo,
            da.des_articulo,
            da.generico,
            da.marca,
            SUM(fv.cantidades_total) AS bultos
        FROM gold.fact_ventas fv
        JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
        JOIN gold.dim_sucursal ds ON fv.id_sucursal = ds.id_sucursal
        JOIN gold.dim_vendedor dv ON fv.id_vendedor = dv.id_vendedor AND fv.id_sucursal = dv.id_sucursal
        LEFT JOIN gold.dim_cliente dc ON fv.id_cliente = dc.id_cliente AND fv.id_sucursal = dc.id_sucursal
        WHERE da.generico = ANY(:genericos)
          {filtro_suc}
          AND fv.anulado = false
          AND dv.id_fuerza_ventas = 1
          AND fv.fecha_comprobante >= :desde
          AND fv.fecha_comprobante <= :hasta
        GROUP BY
            fv.id_sucursal, ds.descripcion, fv.id_cliente,
            COALESCE(
                NULLIF(TRIM(dc.fantasia), ''),
                NULLIF(TRIM(dc.razon_social), ''),
                CONCAT('CLIENTE ', fv.id_cliente)
            ),
            COALESCE(dc.id_ruta_fv1, 0), COALESCE(dv.des_vendedor, 'SIN VENDEDOR'),
            fv.id_articulo, da.des_articulo, da.generico, da.marca
        """
        params = {
            "desde": fecha_desde, "hasta": fecha_hasta,
            "genericos": list(genericos or [c.generico for c in CUADROS]),
        }
        if sucursales:
            query = query.replace("{filtro_suc}", "AND fv.id_sucursal = ANY(:sucursales)")
            params["sucursales"] = list(sucursales)
        else:
            query = query.replace("{filtro_suc}", "")
        df = self.data_loader.execute_query(query, params)
        if not df.empty:
            df["calibre"] = df["des_articulo"].apply(extraer_calibre)
        else:
            df["calibre"] = pd.Series(dtype=str)
        return df

    def _fetch_padron(self, sucursales: list[int] | None = None) -> pd.DataFrame:
        """Obtiene el padron de clientes activos (no anulados) por sucursal."""
        query = """
        SELECT 
            dc.id_sucursal,
            ds.descripcion AS sucursal,
            COUNT(DISTINCT dc.id_cliente) AS padron
        FROM gold.dim_cliente dc
        JOIN gold.dim_sucursal ds ON dc.id_sucursal = ds.id_sucursal
        WHERE dc.anulado = false
          {filtro_suc}
        GROUP BY dc.id_sucursal, ds.descripcion
        ORDER BY ds.descripcion
        """
        # El MISMO recorte que las ventas: el padron es el denominador del % y
        # con otro universo el porcentaje compara cosas distintas.
        params: dict = {}
        if sucursales:
            query = query.replace("{filtro_suc}", "AND dc.id_sucursal = ANY(:sucursales)")
            params["sucursales"] = list(sucursales)
        else:
            query = query.replace("{filtro_suc}", "")
        return self.data_loader.execute_query(query, params)

    def _ventanas(self, config: CoberturaLeviteConfig) -> list[tuple[str, str, str]]:
        """Las tres ventanas del informe: la del config, el mes anterior y el
        mismo mes del anio anterior, en ese orden.

        Los dos historicos son MESES CALENDARIO COMPLETOS y se DERIVAN de
        `fecha_hasta`; nunca se escriben en el config. El daily parchea las
        fechas en cada corrida y deja el resto quieto, asi que un mes escrito a
        mano envejece en silencio (es el bug de la captura de schneider-710).
        """
        hasta = config.fecha_hasta
        return [
            (config.fecha_desde, hasta, self._etiqueta_ventana(config.fecha_desde, hasta)),
            (*rango_mes(periodo_meses_atras(hasta, 1)), ""),
            (*rango_mes(periodo_meses_atras(hasta, 12)), ""),
        ]

    @staticmethod
    def _etiqueta_ventana(desde: str, hasta: str) -> str:
        """`AGOSTO 2026 (01 al 21)` cuando el mes esta a medias, `(mes completo)`
        cuando esta entero. El cuadro del mes en curso se compara contra meses
        cerrados: si el encabezado no dice hasta que dia llega, la caida es una
        caida del negocio y no lo que es, un mes con menos dias."""
        _, ultimo = rango_mes(hasta)
        if desde == rango_mes(desde)[0] and hasta == ultimo:
            return f"{etiqueta_mes(hasta)} (mes completo)"
        return f"{etiqueta_mes(hasta)} ({desde[-2:]} al {hasta[-2:]})"

    def generar_reporte(self, config: CoberturaLeviteConfig) -> CoberturaLeviteResult:
        """Ejecuta la extraccion, transformacion y generacion del workbook Excel."""
        logger.info(
            "Generando reporte Cobertura por Calibre y Marca (%s a %s)...",
            config.fecha_desde, config.fecha_hasta,
        )

        ventanas = self._ventanas(config)
        for desde, hasta, _ in ventanas[1:]:
            logger.info("Ventana historica: %s a %s (mes completo)", desde, hasta)

        # Una consulta por ventana con TODO el universo; cada cuadro filtra lo
        # suyo sobre el mismo DataFrame.
        genericos = [c.generico for c in CUADROS]
        datos = [
            (
                etiqueta or self._etiqueta_ventana(desde, hasta),
                self._fetch_ventas(desde, hasta, genericos, config.sucursales),
            )
            for desde, hasta, etiqueta in ventanas
        ]

        df_actual = datos[0][1]
        df_padron = self._fetch_padron(config.sucursales)

        # Estadisticas del resultado: siguen saliendo de LEVITE en la ventana
        # actual, que es lo que reporta el CLI.
        df_levite = df_actual[df_actual["marca"] == "LEVITE"].copy()
        calibres_presentes = (
            ordenar_calibres(list(df_levite["calibre"].unique()))
            if not df_levite.empty
            else ["500cc", "1500cc", "2250cc"]
        )
        df_matriz, _ = procesar_cobertura_sucursal_calibre(
            df_levite, df_padron, [c for c in calibres_presentes if c != "OTRO"]
        )

        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha_hasta, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)

        nombre = config.nombre_archivo or f"Cobertura por Calibre y Marca - {config.fecha_hasta}"
        ruta_archivo = out_dir / f"{nombre}.xlsx"

        # Una hoja por generico, con los tres periodos apilados. Las hojas de
        # matriz por sucursal y de detalle de clientes se sacaron el 2026-08-19
        # a pedido de Nahuel: lo que se mira es la apertura calibre x marca.
        wb = Workbook()
        wb.remove(wb.active)
        escritas = 0
        for cuadro in CUADROS:
            if self._build_hoja_generico(wb, cuadro, datos, config):
                escritas += 1

        if not escritas:
            raise ValueError(
                f"Sin ventas de {' ni '.join(genericos)} entre {config.fecha_desde} "
                f"y {config.fecha_hasta} para las sucursales {config.sucursales or 'todas'}"
            )

        wb.save(ruta_archivo)
        logger.info("Reporte guardado en %s", ruta_archivo)

        total_gen = df_matriz[df_matriz["es_total_general"]].iloc[0] if not df_matriz.empty else None

        return CoberturaLeviteResult(
            ruta_archivo=ruta_archivo,
            sucursales=len(df_padron),
            clientes_compradores=int(total_gen["cob_total"]) if total_gen is not None else 0,
            volumen_total=float(total_gen["vol_total"]) if total_gen is not None else 0.0,
            padron_total=int(df_padron["padron"].sum()),
        )

    def _build_hoja_generico(
        self,
        wb: Workbook,
        cuadro: Cuadro,
        datos: list[tuple[str, pd.DataFrame]],
        config: CoberturaLeviteConfig,
    ) -> bool:
        """Escribe la hoja de un generico con sus tres periodos apilados.

        Las columnas y las filas se calculan sobre la UNION de los tres
        periodos y se fuerzan en los tres cuadros. Sin eso, un calibre que
        vendio en julio y no en agosto correria las filas y los cuadros
        dejarian de leerse uno debajo del otro.

        Returns:
            False si el generico no tuvo ventas en ninguna de las tres
            ventanas — no se crea una hoja vacia.
        """
        recortes = [(etq, self._recortar(df, cuadro)) for etq, df in datos]
        if all(df.empty for _, df in recortes):
            logger.warning("Sin ventas de %s en ninguna ventana; hoja omitida", cuadro.generico)
            return False

        bloques, calibres = self._ejes(recortes, cuadro)

        ws = wb.create_sheet(cuadro.hoja)
        ws.cell(1, 1, f"Cobertura por Calibre y Marca — {cuadro.generico}").font = Font(bold=True, size=14)
        # Corta: la hoja de cervezas tiene 6 columnas y una nota larga se sale
        # del rango que se captura como imagen.
        ws.cell(
            2, 1,
            "Clientes distintos con compra neta > 0. Los totales se recalculan, no se suman.",
        ).font = Font(italic=True, size=9, color="546E7A")

        fila = 4
        for etiqueta, df in recortes:
            df_cuadro, _ = matriz_calibre_marca(
                df, cuadro.categorias, config.umbral,
                total_label=cuadro.total_label,
                con_subtotales=cuadro.con_subtotales,
                bloques=bloques, calibres=calibres,
            )
            fila = self._escribir_cuadro(ws, fila, etiqueta, df_cuadro, bloques, cuadro) + 1

        columnas = self._columnas(bloques, cuadro)
        ws.column_dimensions["A"].width = 12
        for j in range(2, len(columnas) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 15
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        return True

    @staticmethod
    def _ejes(
        recortes: list[tuple[str, pd.DataFrame]], cuadro: Cuadro,
    ) -> tuple[list[tuple[str, list[str]]], list[str]]:
        """Columnas y filas comunes a los tres periodos.

        Las columnas son las marcas del cuadro que vendieron en ALGUNA de las
        tres ventanas: una columna en cero en las tres es ruido, en una sola no.

        Las filas siguen a las columnas. Un calibre que solo vendio una marca
        sin columna propia no genera fila —saldria entera en cero salvo el
        total—, pero sus clientes SIGUEN contando en la columna del total, que
        es la cobertura del generico completo. KUNSTMAN 470cc es el caso: es
        CERVEZAS, no es una de las cuatro principales.
        """
        marcas_columna = {m for _, marcas in cuadro.categorias for m in marcas}
        marcas_vendidas: set[str] = set()
        calibres_vendidos: set[str] = set()
        for _, df in recortes:
            if df.empty:
                continue
            marcas_vendidas |= set(df["marca"].unique())
            con_columna = df[df["marca"].isin(marcas_columna)]
            calibres_vendidos |= {c for c in con_columna["calibre"].unique() if c != "OTRO"}

        bloques = [
            (etiqueta, cols)
            for etiqueta, marcas in cuadro.categorias
            if (cols := [m for m in marcas if m in marcas_vendidas])
        ]
        return bloques, ordenar_calibres(list(calibres_vendidos))

    @staticmethod
    def _recortar(df: pd.DataFrame, cuadro: Cuadro) -> pd.DataFrame:
        """El universo del cuadro: su generico y, si lo tiene, sus marcas."""
        if df.empty:
            return df
        recorte = df[df["generico"] == cuadro.generico]
        if cuadro.marcas_total is not None:
            recorte = recorte[recorte["marca"].isin(cuadro.marcas_total)]
        return recorte

    @staticmethod
    def _columnas(bloques: list[tuple[str, list[str]]], cuadro: Cuadro) -> list[str]:
        columnas: list[str] = []
        for etiqueta, marcas in bloques:
            columnas += list(marcas)
            if cuadro.con_subtotales:
                columnas.append(f"TOTAL {etiqueta}")
        columnas.append(cuadro.total_label)
        return columnas

    def _escribir_cuadro(
        self, ws: Worksheet, fila: int, etiqueta: str,
        df: pd.DataFrame, bloques: list[tuple[str, list[str]]], cuadro: Cuadro,
    ) -> int:
        """Escribe UN periodo a partir de `fila`. Devuelve la ultima fila usada.

        La celda es cobertura: clientes DISTINTOS que compraron ese calibre de
        esa marca. Ningun total sale de sumar la fila o la columna — todos se
        recalculan en el processor.
        """
        columnas = self._columnas(bloques, cuadro)
        borde = _thin_border()
        ancho_total = len(columnas) + 1

        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho_total)
        c = ws.cell(fila, 1, etiqueta)
        c.font = Font(bold=True, size=11, color=HEADER_FONT)
        c.alignment = Alignment(horizontal="center")
        # El merge solo pinta la celda de arriba a la izquierda: hay que
        # recorrer las demas o la banda sale a medio colorear.
        for j in range(1, ancho_total + 1):
            ws.cell(fila, j).fill = PatternFill("solid", fgColor=HEADER_FILL)
            ws.cell(fila, j).border = borde

        header = fila + 1
        if cuadro.con_subtotales:
            banda = header
            header += 1
            col = 2
            for etiqueta_cat, marcas in bloques:
                ancho = len(marcas) + 1
                ws.merge_cells(start_row=banda, start_column=col, end_row=banda, end_column=col + ancho - 1)
                cc = ws.cell(banda, col, etiqueta_cat)
                cc.fill = PatternFill("solid", fgColor=SUC_FILL)
                cc.font = Font(bold=True, size=10)
                cc.alignment = Alignment(horizontal="center")
                for j in range(col, col + ancho):
                    ws.cell(banda, j).border = borde
                col += ancho
            ws.cell(banda, 1).border = borde
            ws.cell(banda, col).border = borde

        for j, nombre in enumerate(["Calibre"] + columnas, start=1):
            c = ws.cell(header, j, nombre)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde

        r = header
        for _, valores in df.iterrows():
            r += 1
            es_total = valores["calibre"] == "TOTAL"
            for j, nombre in enumerate(["calibre"] + columnas, start=1):
                clave = "calibre" if j == 1 else nombre
                c = ws.cell(r, j, valores.get(clave))
                if j > 1:
                    c.number_format = "#,##0"
                c.border = borde
                if es_total:
                    c.fill = PatternFill("solid", fgColor=TOTAL_GRAL_FILL)
                    c.font = Font(bold=True)
                elif nombre.startswith("TOTAL"):
                    c.fill = PatternFill("solid", fgColor=ACCENT_FILL)
                    c.font = Font(bold=True)
        return r

