"""ComparativoSaltaService — cobertura de la marca SALTA abierta por calibre.

Reconstruye el informe ad-hoc de junio 2026 ("Comparativo Cobertura SALTA"):
cuantos clientes distintos compraron SALTA en el periodo, abierto por tamano de
envase (1000CC, 1200CC, 473CC, ...), contra el mes anterior y contra el mismo mes
del ano anterior.
"""
import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import ZONAS_VIRTUALES
from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService
from src.services.comparativo_salta.processor import (
    MARCA_TOTAL,
    asignar_zona,
    calibres_ordenados,
    combos_sabor_calibre,
    construir_cobertura_mensual,
    construir_cobertura_vendedor,
    construir_cobertura_vendedor_bloques,
    construir_detalle_clientes,
    construir_resumen,
    construir_resumen_por_zona,
)

logger = logging.getLogger(__name__)

MARCA = "SALTA"
HEADER_FILL, HEADER_FONT = "1F4E78", "FFFFFF"
TOTAL_FILL = "FFE08A"          # ámbar — fila de la marca (el total del bloque)
ZONA_FILL = "D6E0F0"           # celeste — primera fila de cada zona
# Fuente NEGRA, a pedido de Nahuel (2026-08-18). OJO: en las columnas Var.
# estas celdas NO llevan relleno, asi que el color era el unico realce del
# signo. La informacion no se pierde —el numero negativo ya muestra el menos—
# pero deja de saltar a la vista. Si hace falta el realce, va un relleno.
POS_FONT = NEG_FONT = "000000"

_MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


@dataclass
class ComparativoSaltaConfig:
    """Config del comparativo de cobertura SALTA.

    Args:
        fecha_desde: primer dia del periodo a medir ('YYYY-MM-DD').
        fecha_hasta: ultimo dia del periodo a medir ('YYYY-MM-DD').
        con_detalle_clientes: si False, omite la hoja de volumen por cliente
            (~15k filas). Util para envios diarios livianos.
        nombre_archivo: nombre de salida sin extension.
    """
    fecha_desde: str
    fecha_hasta: str
    con_detalle_clientes: bool = True
    # Años para la hoja de cobertura mensual (sucursal x sabor x calibre x mes).
    anios_mensual: list[int] | None = None
    # Sucursal cuyos años se apilan uno debajo del otro para comparar.
    sucursal_comparativa: str | None = None
    # Hoja de cobertura por preventista: meses 'YYYY-MM' a mostrar como columnas.
    meses_vendedor: list[str] | None = None
    # Variante con bloques definidos a mano. Cada bloque: sabor, calibre, meses,
    # y opcionalmente cupo (objetivo total) y grupo (rotulo que se fusiona arriba).
    bloques_vendedor: list[dict] | None = None
    # id_sucursal al que se acota esa hoja. None = todas.
    id_sucursal_vendedor: int | None = None
    # Preventistas dados de baja que no deben figurar. Se filtran DESDE EL ORIGEN
    # (no solo la fila) para que la suma de vendedores siga cerrando contra el
    # total; el costo es que sus clientes tampoco cuentan en el total.
    excluir_vendedores: list[str] | None = None
    nombre_archivo: str | None = None


@dataclass
class ComparativoSaltaResult:
    """Resultado del informe."""
    ruta_archivo: Path
    calibres: list[str]
    cobertura_total: int
    fecha_desde: str
    fecha_hasta: str


def _etiqueta(fecha: str) -> str:
    d = pd.to_datetime(fecha)
    return f"{_MESES[d.month - 1]}-{d.year}"


def _periodo_desplazado(desde: str, hasta: str, *, meses: int) -> tuple[str, str]:
    """Same day-of-month window shifted back N months.

    The end date is clamped to the shifted month's length so a month-to-date
    window (1..29 of July) maps to 1..29 of the comparison month instead of
    overflowing into the next one.
    """
    d, h = pd.to_datetime(desde), pd.to_datetime(hasta)
    d2 = d - pd.DateOffset(months=meses)
    h2 = h - pd.DateOffset(months=meses)
    fin_de_mes = (d2 + pd.offsets.MonthEnd(0)).day
    h2 = h2.replace(day=min(h.day, fin_de_mes))
    return d2.strftime("%Y-%m-%d"), h2.strftime("%Y-%m-%d")


def _thin() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class ComparativoSaltaService(BaseService):
    """Genera el comparativo de cobertura SALTA por calibre."""

    SERVICE_SLUG = "comparativo-salta"
    GRANULARITY = "month"

    def generar_reporte(self, config: ComparativoSaltaConfig) -> ComparativoSaltaResult:
        desde, hasta = config.fecha_desde, config.fecha_hasta
        prev_d, prev_h = _periodo_desplazado(desde, hasta, meses=1)
        mmaa_d, mmaa_h = _periodo_desplazado(desde, hasta, meses=12)

        logger.info("Comparativo SALTA: %s..%s vs %s..%s vs %s..%s",
                    desde, hasta, prev_d, prev_h, mmaa_d, mmaa_h)

        def _traer(d: str, h: str) -> pd.DataFrame:
            df = self.data_loader.get_ventas_cliente_calibre(MARCA, d, h)
            return asignar_zona(df, ZONAS_VIRTUALES)

        actual, anterior, mmaa = _traer(desde, hasta), _traer(prev_d, prev_h), _traer(mmaa_d, mmaa_h)

        etiquetas = {
            "actual": f"Cob. {_etiqueta(desde)}",
            "anterior": f"Cob. {_etiqueta(prev_d)}",
            "mmaa": f"Cob. {_etiqueta(mmaa_d)}",
        }
        resumen = construir_resumen(actual, anterior, mmaa, etiquetas)
        por_zona = construir_resumen_por_zona(actual, anterior, mmaa, etiquetas)
        calibres = calibres_ordenados(actual, anterior, mmaa)

        wb = Workbook()
        wb.remove(wb.active)
        subtitulo = (
            f"Clientes distintos con bultos > 0 · {desde} al {hasta} · "
            f"la cobertura NO es aditiva: los calibres no suman el total de la marca "
            f"(quien compra 1000 y 1200 cuenta una vez en el total y una en cada calibre)"
        )
        self._hoja_resumen(wb, "Cobertura", resumen,
                           f"Cobertura SALTA por calibre — {_etiqueta(desde)}", subtitulo)
        self._hoja_resumen(wb, "Cobertura por Zona", por_zona,
                           f"Cobertura SALTA por calibre y zona — {_etiqueta(desde)}",
                           subtitulo, col_zona=True)

        mensual_por_anio = {}
        for anio in config.anios_mensual or []:
            # Sucursal CRUDA a proposito: esta hoja no aplica zonas virtuales.
            # CASA CENTRAL va entera (incluye VALLE SALTA y SUB DISTRIBUIDORES),
            # que es como se venia leyendo historicamente este cuadro.
            df_anio = self.data_loader.get_ventas_cliente_sabor_mes(
                MARCA, f"{anio}-01-01", f"{anio}-12-31"
            )
            mensual_por_anio[anio] = df_anio
            self._hoja_mensual(wb, construir_cobertura_mensual(df_anio, anio), anio)

        if config.sucursal_comparativa and mensual_por_anio:
            suc = config.sucursal_comparativa
            bloques = []
            for anio, df in sorted(mensual_por_anio.items()):
                bloque = construir_cobertura_mensual(
                    df[df["sucursal"] == suc], anio, forzar_12_meses=True
                )
                # Con una sola sucursal el TOTAL GENERAL repite al TOTAL de la
                # sucursal: dos filas idénticas que no aportan nada.
                bloque = bloque[bloque["Sucursal"] != "TOTAL GENERAL"]
                bloques.append((anio, bloque))
            self._hoja_comparativo_sucursal(wb, bloques, suc)

        if config.bloques_vendedor:
            meses_pedidos = sorted({m for b in config.bloques_vendedor for m in b["meses"]})
            frames = {}
            for mes in meses_pedidos:
                inicio = f"{mes}-01"
                fin = (pd.to_datetime(inicio) + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                df_mes = self.data_loader.get_ventas_cliente_sabor_mes(MARCA, inicio, fin)
                if config.id_sucursal_vendedor is not None:
                    df_mes = df_mes[df_mes["id_sucursal"] == config.id_sucursal_vendedor]
                if config.excluir_vendedores:
                    excluidos = {v.strip().upper() for v in config.excluir_vendedores}
                    antes = df_mes.drop_duplicates(["id_cliente", "id_sucursal"]).shape[0]
                    # El preventista NULL se compara como "(SIN ASIGNAR)", que es la
                    # etiqueta con la que despues aparece en el cuadro: si se comparara
                    # contra "" no habria forma de excluirlo desde el config.
                    nombres = (
                        df_mes["preventista"].fillna("(sin asignar)").str.strip().str.upper()
                    )
                    df_mes = df_mes[~nombres.isin(excluidos)]
                    quitados = antes - df_mes.drop_duplicates(["id_cliente", "id_sucursal"]).shape[0]
                    if quitados:
                        logger.info(
                            "%s: %d cliente(s) fuera del cuadro por vendedor dado de baja (%s)",
                            mes, quitados, ", ".join(sorted(excluidos)),
                        )
                frames[mes] = df_mes

            tabla = construir_cobertura_vendedor_bloques(frames, config.bloques_vendedor)
            sucursal_unica = None
            if config.id_sucursal_vendedor is not None and not tabla.empty:
                nombres = [s for s in tabla["Sucursal"].unique() if not str(s).startswith("TOTAL")]
                sucursal_unica = nombres[0] if len(nombres) == 1 else None
            self._hoja_vendedor_bloques(wb, tabla, config.bloques_vendedor, sucursal_unica)

        elif config.meses_vendedor:
            frames = {}
            for mes in config.meses_vendedor:
                inicio = f"{mes}-01"
                fin = (pd.to_datetime(inicio) + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                df_mes = self.data_loader.get_ventas_cliente_sabor_mes(MARCA, inicio, fin)
                if config.id_sucursal_vendedor is not None:
                    df_mes = df_mes[df_mes["id_sucursal"] == config.id_sucursal_vendedor]
                frames[mes] = df_mes

            combos = combos_sabor_calibre(frames)
            tabla = construir_cobertura_vendedor(frames, config.meses_vendedor, combos)
            sucursal_unica = None
            if config.id_sucursal_vendedor is not None and not tabla.empty:
                nombres = [s for s in tabla["Sucursal"].unique() if not str(s).startswith("TOTAL")]
                sucursal_unica = nombres[0] if len(nombres) == 1 else None
            self._hoja_vendedor(wb, tabla, config.meses_vendedor, combos, sucursal_unica)

        if config.con_detalle_clientes:
            detalle = construir_detalle_clientes(actual, anterior, calibres)
            self._hoja_detalle(wb, detalle, calibres, etiquetas)

        nombre = config.nombre_archivo or f"Comparativo Cobertura SALTA - {_etiqueta(desde)}"
        out_dir = service_output_dir(self.SERVICE_SLUG, desde, granularity=self.GRANULARITY)
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"
        wb.save(ruta)

        total = int(resumen.iloc[0][etiquetas["actual"]]) if not resumen.empty else 0
        return ComparativoSaltaResult(
            ruta_archivo=ruta, calibres=calibres, cobertura_total=total,
            fecha_desde=desde, fecha_hasta=hasta,
        )

    # ── hojas ────────────────────────────────────────────────────────────

    def _hoja_resumen(self, wb, nombre, df, titulo, subtitulo, col_zona=False):
        ws = wb.create_sheet(nombre)
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = subtitulo
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        if df.empty:
            ws["A4"] = "Sin datos para el periodo."
            return

        head = 4
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(row=head, column=j, value=col)
            c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = _thin()

        detalle_col = list(df.columns).index("Detalle")
        for i, row in enumerate(df.itertuples(index=False), head + 1):
            es_total = row[detalle_col] == MARCA_TOTAL
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.border = _thin()
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                if es_total:
                    fill = ZONA_FILL if col_zona else TOTAL_FILL
                    c.fill = PatternFill("solid", start_color=fill, end_color=fill)
                    c.font = Font(bold=True)
                if df.columns[j - 1].startswith("Var.") and isinstance(val, (int, float)) and val:
                    c.font = Font(bold=es_total, color=POS_FONT if val > 0 else NEG_FONT)

        ws.column_dimensions["A"].width = 24 if col_zona else 26
        for j in range(2, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 26 if (col_zona and j == 2) else 15
        ws.freeze_panes = ws.cell(row=head + 1, column=1)

    def _hoja_mensual(self, wb, df, anio: int):
        """Cobertura mes a mes por sucursal / sabor / calibre."""
        ws = wb.create_sheet(f"Cobertura Mensual {anio}")
        ws["A1"] = f"Cobertura SALTA por sabor y calibre — {anio}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (
            "Clientes distintos con bultos > 0 · el anual NO es la suma de los meses "
            "(quien compra todos los meses cuenta una sola vez), y el TOTAL de la "
            "sucursal tampoco es la suma de sus filas"
        )
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        if df.empty:
            ws["A4"] = "Sin datos para el periodo."
            return

        head = 4
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(row=head, column=j, value=col)
            c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin()

        anio_col = df.columns[-1]
        for i, row in enumerate(df.itertuples(index=False), head + 1):
            etiqueta = str(row[0])
            es_total = etiqueta.startswith("TOTAL")
            es_general = etiqueta == "TOTAL GENERAL"
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.border = _thin()
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                if es_total:
                    fill = TOTAL_FILL if es_general else ZONA_FILL
                    c.fill = PatternFill("solid", start_color=fill, end_color=fill)
                    c.font = Font(bold=True)
                elif df.columns[j - 1] == anio_col:
                    c.fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
                    c.font = Font(bold=True)

        for j, col in enumerate(df.columns, 1):
            ancho = 22 if j == 1 else (16 if j == 2 else (9 if j == 3 else 10))
            ws.column_dimensions[get_column_letter(j)].width = ancho
        ws.freeze_panes = ws.cell(row=head + 1, column=4)

    def _hoja_comparativo_sucursal(self, wb, bloques, sucursal: str):
        """Un año arriba del otro, misma sucursal, columnas alineadas.

        `bloques` es [(anio, frame)] ya filtrado a `sucursal`. Se apilan con tres
        filas en blanco entre medio para que se lean como dos cuadros y no como
        una tabla sola.
        """
        ws = wb.create_sheet(f"Comparativo {sucursal}"[:31])
        ws["A1"] = f"Cobertura SALTA — {sucursal} — {' vs '.join(str(a) for a, _ in bloques)}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (
            "Clientes distintos con bultos > 0 · el anual NO es la suma de los meses · "
            "los meses sin ventas quedan en 0 para que las columnas de los dos años coincidan"
        )
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        fila = 4
        n_cols = 0
        for anio, df in bloques:
            if df.empty:
                continue
            n_cols = max(n_cols, len(df.columns))
            for j, col in enumerate(df.columns, 1):
                c = ws.cell(row=fila, column=j, value=col)
                c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
                c.font = Font(bold=True, color=HEADER_FONT)
                c.alignment = Alignment(horizontal="center")
                c.border = _thin()

            anio_col = df.columns[-1]
            for i, row in enumerate(df.itertuples(index=False), fila + 1):
                es_total = str(row[0]).startswith("TOTAL")
                for j, val in enumerate(row, 1):
                    c = ws.cell(row=i, column=j, value=val)
                    c.border = _thin()
                    if isinstance(val, (int, float)):
                        c.number_format = "#,##0"
                    if es_total:
                        c.fill = PatternFill("solid", start_color=ZONA_FILL, end_color=ZONA_FILL)
                        c.font = Font(bold=True)
                    elif df.columns[j - 1] == anio_col:
                        c.fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
                        c.font = Font(bold=True)

            fila += len(df) + 4  # bloque + header + 3 en blanco

        for j in range(1, max(n_cols, 1) + 1):
            ws.column_dimensions[get_column_letter(j)].width = (
                22 if j == 1 else (16 if j == 2 else (9 if j == 3 else 10))
            )

    def _hoja_vendedor_bloques(self, wb, df, bloques, sucursal_unica=None):
        """Cobertura por preventista con bloques definidos: grupo / sabor / calibre / mes."""
        ws = wb.create_sheet("Cobertura x Vendedor")
        titulo = "Cobertura SALTA por ruta — incentivo preventa"
        if sucursal_unica:
            titulo += f" — {sucursal_unica}"
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=13)
        # La aclaracion territorial no es cosmetica: sin ella el cuadro se lee como
        # performance personal, y los meses historicos NO lo son. La atribucion usa
        # la asignacion de HOY (dim_cliente.des_personal_fv1) aplicada hacia atras,
        # que es lo correcto para un incentivo por ruta — es la unica forma de que
        # un preventista que acaba de tomar la ruta tenga un baseline medible.
        ws["A2"] = (
            "Clientes distintos con bultos > 0 (pdv compradores) · cada fila es la RUTA que "
            "hoy atiende ese preventista, no su desempeño personal: los meses de 2025 son lo "
            "que vendió la ruta, aunque en ese momento la tuviera otra persona · sumar entre "
            "filas es válido, sumar calibres o meses NO · el Cupo está repartido en proporción "
            "al mes histórico de cada bloque y cierra exacto contra el objetivo del TOTAL"
        )
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        if df.empty:
            ws["A4"] = "Sin datos para el periodo."
            return

        if sucursal_unica:
            df = df.copy()
            es_total = df["Sucursal"].astype(str).str.startswith("TOTAL")
            df.loc[es_total, "Vendedor"] = df.loc[es_total, "Sucursal"]
            df = df.drop(columns=["Sucursal"])

        f_grupo, f_sabor, f_cal, f_mes = 4, 5, 6, 7
        relleno = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
        fuente = Font(bold=True, color=HEADER_FONT)
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fijas = ["Vendedor"] if sucursal_unica else ["Sucursal", "Vendedor"]
        for col, rotulo in enumerate(fijas, 1):
            ws.merge_cells(start_row=f_grupo, start_column=col, end_row=f_mes, end_column=col)
            c = ws.cell(row=f_grupo, column=col, value=rotulo)
            c.fill, c.font, c.alignment, c.border = relleno, fuente, centro, _thin()

        primera_valor = len(fijas) + 1
        col = primera_valor
        # Los bloques consecutivos que comparten `grupo` se fusionan arriba.
        tramos_grupo: list[tuple[str, int, int]] = []
        for b in bloques:
            ancho = len(b["meses"]) + 1
            grupo = b.get("grupo", "")
            if tramos_grupo and tramos_grupo[-1][0] == grupo:
                tramos_grupo[-1] = (grupo, tramos_grupo[-1][1], col + ancho - 1)
            else:
                tramos_grupo.append((grupo, col, col + ancho - 1))

            etiqueta = "SALTA RUBIA" if "rubia" in b["sabor"].lower() else f"SALTA {b['sabor']}"
            for fila, texto in ((f_sabor, etiqueta), (f_cal, f"{b['calibre']} cc")):
                ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + ancho - 1)
                c = ws.cell(row=fila, column=col, value=texto)
                c.fill, c.font, c.alignment = relleno, fuente, centro

            for i, mes in enumerate([*b["meses"], "Cupo"]):
                texto = "Cupo" if mes == "Cupo" else _etiqueta(f"{mes}-01")
                c = ws.cell(row=f_mes, column=col + i, value=texto)
                c.fill, c.font, c.alignment, c.border = relleno, fuente, centro, _thin()
            col += ancho

        for grupo, desde, hasta in tramos_grupo:
            ws.merge_cells(start_row=f_grupo, start_column=desde, end_row=f_grupo, end_column=hasta)
            c = ws.cell(row=f_grupo, column=desde, value=grupo)
            c.fill, c.font, c.alignment = relleno, fuente, centro

        for fila in (f_grupo, f_sabor, f_cal):
            for j in range(primera_valor, col):
                ws.cell(row=fila, column=j).border = _thin()

        columnas = list(df.columns)
        for i, row in enumerate(df.itertuples(index=False), f_mes + 1):
            es_total = str(row[0]).startswith("TOTAL")
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.border = _thin()
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                if es_total:
                    c.fill = PatternFill("solid", start_color=TOTAL_FILL, end_color=TOTAL_FILL)
                    c.font = Font(bold=True)
                elif columnas[j - 1].endswith("|Cupo"):
                    c.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

        anchos = [26] if sucursal_unica else [22, 26]
        for j, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = ancho
        for j in range(primera_valor, col):
            ws.column_dimensions[get_column_letter(j)].width = 10
        ws.freeze_panes = ws.cell(row=f_mes + 1, column=primera_valor)

    def _hoja_vendedor(self, wb, df, meses, combos, sucursal_unica=None):
        """Cobertura por preventista con encabezado de 3 niveles: sabor / calibre / mes.

        Si `sucursal_unica` viene dado, la hoja es de esa sola sucursal: el nombre
        va al titulo y la columna Sucursal no se dibuja, porque repetir el mismo
        valor en 87 filas no aporta nada.
        """
        ws = wb.create_sheet("Cobertura x Vendedor")
        titulo = "Cobertura SALTA por vendedor — sabor, calibre y mes"
        if sucursal_unica:
            titulo += f" — {sucursal_unica}"
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (
            "Clientes distintos con bultos > 0 · el TOTAL se recuenta sobre toda la "
            "sucursal; coincide con la suma de la columna porque cada cliente tiene un "
            "solo preventista asignado · lo que NO se puede sumar son los calibres "
            "entre sí ni los meses: ahí el mismo cliente se repite · "
            "la columna Obj queda vacía para completar a mano"
        )
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        if df.empty:
            ws["A4"] = "Sin datos para el periodo."
            return

        if sucursal_unica:
            # El rótulo del total vive en Sucursal; si se descarta la columna sin
            # moverlo, la fila de totales queda sin nombre.
            df = df.copy()
            es_total = df["Sucursal"].astype(str).str.startswith("TOTAL")
            df.loc[es_total, "Vendedor"] = df.loc[es_total, "Sucursal"]
            df = df.drop(columns=["Sucursal"])

        f_sabor, f_cal, f_mes = 4, 5, 6          # las tres filas del encabezado
        relleno = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
        fuente = Font(bold=True, color=HEADER_FONT)
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fijas = ["Vendedor"] if sucursal_unica else ["Sucursal", "Vendedor"]
        for col, rotulo in enumerate(fijas, 1):
            ws.merge_cells(start_row=f_sabor, start_column=col, end_row=f_mes, end_column=col)
            c = ws.cell(row=f_sabor, column=col, value=rotulo)
            c.fill, c.font, c.alignment, c.border = relleno, fuente, centro, _thin()

        # Cada combo ocupa len(meses) + 1 columnas: los meses y el objetivo.
        ancho_bloque = len(meses) + 1
        primera_valor = len(fijas) + 1
        col = primera_valor
        for sabor, calibre in combos:
            etiqueta_sabor = "SALTA RUBIA" if "rubia" in sabor.lower() else f"SALTA {sabor}"

            ws.merge_cells(start_row=f_sabor, start_column=col,
                           end_row=f_sabor, end_column=col + ancho_bloque - 1)
            c = ws.cell(row=f_sabor, column=col, value=etiqueta_sabor)
            c.fill, c.font, c.alignment = relleno, fuente, centro

            ws.merge_cells(start_row=f_cal, start_column=col,
                           end_row=f_cal, end_column=col + ancho_bloque - 1)
            c = ws.cell(row=f_cal, column=col, value=f"{calibre} cc")
            c.fill, c.font, c.alignment = relleno, fuente, centro

            for i, mes in enumerate([*meses, "Objetivo"]):
                titulo = "Obj" if mes == "Objetivo" else _etiqueta(f"{mes}-01")
                c = ws.cell(row=f_mes, column=col + i, value=titulo)
                c.fill, c.font, c.alignment, c.border = relleno, fuente, centro, _thin()
            col += ancho_bloque

        # Bordes del encabezado que quedaron en celdas fusionadas
        for fila in (f_sabor, f_cal):
            for j in range(primera_valor, col):
                ws.cell(row=fila, column=j).border = _thin()

        columnas = list(df.columns)
        for i, row in enumerate(df.itertuples(index=False), f_mes + 1):
            es_total = str(row[0]).startswith("TOTAL")
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.border = _thin()
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                if es_total:
                    c.fill = PatternFill("solid", start_color=TOTAL_FILL, end_color=TOTAL_FILL)
                    c.font = Font(bold=True)
                elif columnas[j - 1].endswith("|Objetivo"):
                    # Celda de carga manual: se pinta para que se vea dónde escribir.
                    c.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

        anchos = [26] if sucursal_unica else [22, 26]
        for j, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = ancho
        for j in range(primera_valor, col):
            ws.column_dimensions[get_column_letter(j)].width = 9
        ws.freeze_panes = ws.cell(row=f_mes + 1, column=primera_valor)

    def _hoja_detalle(self, wb, df, calibres, etiquetas):
        ws = wb.create_sheet("Volumen por Cliente")
        ws["A1"] = "Volumen SALTA (bultos) por cliente"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (
            f"Incluye clientes con compra en cualquiera de los dos periodos "
            f"({etiquetas['actual'].replace('Cob. ', '')} o "
            f"{etiquetas['anterior'].replace('Cob. ', '')}), por eso hay filas en 0."
        )
        ws["A2"].font = Font(italic=True, size=9, color="595959")

        if df.empty:
            ws["A4"] = "Sin datos para el periodo."
            return

        fijas = ["sucursal", "razon_social", "fantasia", "preventista", "id_ruta"]
        rotulos = {"sucursal": "Zona", "razon_social": "Razón Social", "fantasia": "Fantasía",
                   "preventista": "Preventista", "id_ruta": "Ruta"}
        head = 4
        for j, col in enumerate(df.columns, 1):
            if col in rotulos:
                etiqueta = rotulos[col]
            else:
                calibre, periodo = col.rsplit("_", 1)
                sufijo = etiquetas["actual"] if periodo == "actual" else etiquetas["anterior"]
                etiqueta = f"SALTA {calibre.replace('CC', '')}\n{sufijo.replace('Cob. ', '')}"
            c = ws.cell(row=head, column=j, value=etiqueta)
            c.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = _thin()

        for i, row in enumerate(df.itertuples(index=False), head + 1):
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.border = _thin()
                if df.columns[j - 1] == "id_ruta":
                    # Entero explícito: con formato decimal los VLOOKUP de abajo
                    # leen "6,00" en vez de "6" y no matchean.
                    c.number_format = "0"
                elif isinstance(val, (int, float)):
                    c.number_format = "#,##0.##"

        anchos = {"sucursal": 20, "razon_social": 34, "fantasia": 28, "preventista": 24, "id_ruta": 8}
        for j, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(j)].width = anchos.get(col, 14)
        ws.freeze_panes = ws.cell(row=head + 1, column=len(fijas) + 1)
