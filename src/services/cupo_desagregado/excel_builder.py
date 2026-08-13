"""Construccion del workbook de Cupo Desagregado Por Ruta.

Tres hojas: el detalle por ruta, el cupo original por preventista y el
totalizado por sucursal. Cada hoja cierra con una fila TOTAL GENERAL.
"""
from __future__ import annotations

from itertools import groupby
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.services.cupo_desagregado.constants import (
    CATEGORIAS,
    ETIQUETA_TOTAL_GENERAL,
)
from src.services.cupo_desagregado.processor import Vendedor

_FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
_FILL_SUBTOTAL = PatternFill("solid", fgColor="D9E1F2")
_FILL_TOTAL = PatternFill("solid", fgColor="FCE4D6")
_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FONT_BOLD = Font(bold=True)
_BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
_FORMATO_NUMERICO = "#,##0.00"
# El codigo de ruta es un identificador, no una medida: sin decimales ni
# separador de miles, o los joins downstream leen "1,00" en vez de 1.
_FORMATO_CODIGO = "0"


def construir_workbook(filas: list[dict], vendedores: list[Vendedor]) -> Workbook:
    """Arma el workbook completo en memoria (no toca disco)."""
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_cupo_ruta(wb, filas)
    _hoja_cupo_preventa(wb, vendedores)
    _hoja_resumen_sucursal(wb, vendedores)
    return wb


def escribir_excel(
    filas: list[dict], vendedores: list[Vendedor], ruta: Path
) -> Path:
    """Persiste el workbook en `ruta` (crea el directorio si falta)."""
    ruta.parent.mkdir(parents=True, exist_ok=True)
    construir_workbook(filas, vendedores).save(ruta)
    return ruta


def _hoja_cupo_ruta(wb: Workbook, filas: list[dict]) -> None:
    """Detalle por ruta, con subtotal por preventista.

    OJO: el CÓDIGO de ruta se REUSA entre sucursales (el 1 es DIAS LU-JU en
    CAFAYATE y VILLA LU-JU en JVG). La clave unica de cada fila es
    (SUCURSAL, CÓDIGO). No usar el codigo solo como PK aguas abajo.
    """
    ws = wb.create_sheet("Cupo Ruta")
    filas_datos: list[list] = []
    for (sucursal, vendedor), grupo in groupby(
        filas, key=lambda f: (f["sucursal"], f["vendedor"])
    ):
        rutas = list(grupo)
        for fila in rutas:
            filas_datos.append(
                [fila["sucursal"], fila["vendedor"], fila["codigo"], fila["ruta"]]
                + [fila["vals"][c] for c in CATEGORIAS]
            )
        subtotal = [round(sum(f["vals"][c] for f in rutas), 2) for c in CATEGORIAS]
        filas_datos.append([sucursal, f"TOTAL {vendedor}", "", "", *subtotal])

    _escribir_hoja(
        ws,
        headers=["SUCURSAL", "PREVENTISTA", "CÓDIGO", "RUTA"] + CATEGORIAS,
        filas=filas_datos,
        etiqueta_idx=1,
        primera_col_numerica=2,
        formatos_por_columna={2: _FORMATO_CODIGO},
        total_general=_total_general(filas, cols_texto=4, etiqueta_idx=1),
    )


def _hoja_cupo_preventa(wb: Workbook, vendedores: list[Vendedor]) -> None:
    """Cupo original por preventista, tal como viene del archivo fuente."""
    ws = wb.create_sheet("Cupo Preventa")
    filas = [[v.sucursal, v.nombre] + [v.cupos[c] for c in CATEGORIAS]
             for v in vendedores]
    totales = [round(sum(v.cupos[c] for v in vendedores), 2) for c in CATEGORIAS]
    _escribir_hoja(
        ws,
        headers=["SUCURSAL", "PREVENTISTA"] + CATEGORIAS,
        filas=filas,
        etiqueta_idx=1,
        primera_col_numerica=2,
        total_general=["", ETIQUETA_TOTAL_GENERAL, *totales],
    )


def _hoja_resumen_sucursal(wb: Workbook, vendedores: list[Vendedor]) -> None:
    """Cupo agregado por sucursal."""
    ws = wb.create_sheet("Resumen Sucursal")
    acumulado: dict[str, dict[str, float]] = {}
    for v in vendedores:
        sucursal = acumulado.setdefault(v.sucursal, {c: 0.0 for c in CATEGORIAS})
        for categoria in CATEGORIAS:
            sucursal[categoria] += v.cupos[categoria]

    filas = [[nombre] + [round(vals[c], 2) for c in CATEGORIAS]
             for nombre, vals in acumulado.items()]
    totales = [round(sum(vals[c] for vals in acumulado.values()), 2)
               for c in CATEGORIAS]
    _escribir_hoja(
        ws,
        headers=["SUCURSAL"] + CATEGORIAS,
        filas=filas,
        etiqueta_idx=0,
        primera_col_numerica=1,
        total_general=[ETIQUETA_TOTAL_GENERAL, *totales],
    )


def _total_general(filas: list[dict], cols_texto: int, etiqueta_idx: int) -> list:
    """Fila TOTAL GENERAL de la hoja de detalle (suma solo las rutas).

    `cols_texto` es la cantidad de columnas de texto antes de las categorias
    y `etiqueta_idx` la posicion 0-based donde va la etiqueta.
    """
    totales = [round(sum(f["vals"][c] for f in filas), 2) for c in CATEGORIAS]
    encabezado = [""] * cols_texto
    encabezado[etiqueta_idx] = ETIQUETA_TOTAL_GENERAL
    return encabezado + totales


def _escribir_hoja(
    ws: Worksheet,
    headers: list[str],
    filas: list[list],
    etiqueta_idx: int,
    primera_col_numerica: int,
    total_general: list,
    formatos_por_columna: dict[int, str] | None = None,
) -> None:
    """Vuelca headers + filas + TOTAL GENERAL y aplica el formato de tabla."""
    formatos_por_columna = formatos_por_columna or {}

    ws.append(headers)
    for celda in ws[1]:
        celda.fill = _FILL_HEADER
        celda.font = _FONT_HEADER
        celda.alignment = Alignment(horizontal="center")

    for fila in filas:
        ws.append(fila)
    ws.append(total_general)

    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        etiqueta = str(fila[etiqueta_idx].value or "")
        es_total = etiqueta == ETIQUETA_TOTAL_GENERAL
        es_subtotal = etiqueta.startswith("TOTAL ") and not es_total
        for i, celda in enumerate(fila):
            celda.border = _BORDE
            if i >= primera_col_numerica and isinstance(celda.value, (int, float)):
                celda.number_format = formatos_por_columna.get(i, _FORMATO_NUMERICO)
            if es_subtotal:
                celda.font = _FONT_BOLD
                celda.fill = _FILL_SUBTOTAL
            elif es_total:
                celda.font = _FONT_BOLD
                celda.fill = _FILL_TOTAL

    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            13, min(32, len(str(header)) + 6))
    ws.freeze_panes = "A2"
