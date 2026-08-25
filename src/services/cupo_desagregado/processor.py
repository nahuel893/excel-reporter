"""Logica pura del reporte Cupo Desagregado Por Ruta.

Sin acceso a base de datos ni escritura de archivos: recibe DataFrames /
dicts y devuelve estructuras en memoria. Todo lo testeable vive aca.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

from src.services.cupo_desagregado.constants import (
    CATEGORIAS,
    CATEGORIAS_DISTRIBUIBLES,
    CERVEZA_MARCAS,
    ETIQUETA_SIN_RUTA,
    MARCAS_MULTICERVEZA,
    MARCAS_PROPIAS,
    MARCAS_SALTA,
    MESES_ES,
    NOMBRE_OVERRIDES,
    RUTAS_OVERRIDE,
    SRC_COLS,
    SRC_ROW_FIN,
    SRC_ROW_INI,
    SUCURSAL_IDS,
)


@dataclass
class Vendedor:
    """Un preventista del archivo fuente con su cupo por categoria."""

    nombre: str
    sucursal: str
    id_sucursal: int
    cupos: dict[str, float] = field(default_factory=dict)


@dataclass
class Distribucion:
    """Resultado de repartir los cupos entre las rutas.

    filas: un dict por (vendedor, ruta) con claves sucursal / vendedor /
    codigo / ruta / vals. `sin_ruta` y `sin_historia` son diagnosticos para
    que el operador revise el mes.
    """

    filas: list[dict] = field(default_factory=list)
    sin_ruta: list[str] = field(default_factory=list)
    sin_historia: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Periodo
# ---------------------------------------------------------------------------
def hoja_del_mes(fecha_desde: str) -> str:
    """'2026-07-01' -> 'JULIO' (nombre de hoja del archivo de objetivos)."""
    return MESES_ES[date.fromisoformat(fecha_desde[:10]).month]


def periodo_historia(fecha_desde: str) -> tuple[date, date]:
    """Mes completo anterior al periodo del cupo, como [desde, hasta).

    Los cupos de julio se abren con la historia de junio; si el operador
    quiere otra ventana la pasa explicita por config.
    """
    inicio_cupo = date.fromisoformat(fecha_desde[:10]).replace(day=1)
    if inicio_cupo.month == 1:
        inicio_hist = date(inicio_cupo.year - 1, 12, 1)
    else:
        inicio_hist = date(inicio_cupo.year, inicio_cupo.month - 1, 1)
    return inicio_hist, inicio_cupo


# ---------------------------------------------------------------------------
# Clasificacion de ventas
# ---------------------------------------------------------------------------
def _texto(valor) -> str:
    """Normaliza un valor de gold a texto en mayusculas.

    Los campos de dim_articulo llegan como NaN cuando son NULL, asi que no
    alcanza con `valor or ""`: NaN es truthy.
    """
    if valor is None or (isinstance(valor, float) and valor != valor):
        return ""
    return str(valor).strip().upper()


def clasificar_categoria(
    generico: str | None, marca: str | None, es_fernet: bool
) -> str | None:
    """Mapea una venta a una categoria del cupo. None = no participa.

    El orden importa: la primera coincidencia gana. Las cervezas se resuelven
    antes que FERNET porque `es_fernet` mira el texto del articulo, no la
    marca (no existe marca 'FERNET' en dim_articulo: el fernet aparece como
    FRATELLI BRANCA y VITTONE).
    """
    generico = _texto(generico)
    marca = _texto(marca)

    if generico == "CERVEZAS":
        if marca in MARCAS_SALTA:
            return "SALTA"
        if marca in MARCAS_PROPIAS:
            return marca
        if marca in MARCAS_MULTICERVEZA:
            return "MULTICERVEZA"
        return None
    if generico == "AGUAS DANONE":
        return "AGUA DANONE"
    if es_fernet:
        return "FERNET"
    if generico == "VINOS":
        # La marca 'VINOS' no registra ventas: clasifica el generico.
        return "VINOS"
    if marca == "R2":
        return "R2"
    return None


def agregar_historia(df: pd.DataFrame) -> dict[tuple[int, int, str], float]:
    """(id_sucursal, id_ruta, categoria) -> cantidad vendida en el periodo.

    Espera las columnas id_sucursal, id_ruta, generico, marca, es_fernet,
    cantidad tal como las devuelve DataLoader.get_ventas_por_ruta_categoria.
    """
    historia: dict[tuple[int, int, str], float] = {}
    for fila in df.itertuples(index=False):
        # Venta de un cliente sin ruta: no se puede imputar a ninguna.
        if fila.id_ruta is None or pd.isna(fila.id_ruta):
            continue
        categoria = clasificar_categoria(
            fila.generico, fila.marca, bool(fila.es_fernet)
        )
        if categoria is None:
            continue
        cantidad = (0.0 if fila.cantidad is None or pd.isna(fila.cantidad)
                    else float(fila.cantidad))
        clave = (int(fila.id_sucursal), int(fila.id_ruta), categoria)
        historia[clave] = historia.get(clave, 0.0) + cantidad
    return historia


# ---------------------------------------------------------------------------
# Lookups desde gold
# ---------------------------------------------------------------------------
def construir_mapa_vendedores(df: pd.DataFrame) -> dict[tuple[str, int], int]:
    """(nombre_normalizado, id_sucursal) -> id_vendedor.

    REGLA DE ORO: id_vendedor se reusa entre sucursales, por eso la clave
    lleva id_sucursal.
    """
    return {
        (_texto(f.des_vendedor), int(f.id_sucursal)): int(f.id_vendedor)
        for f in df.itertuples(index=False)
    }


def construir_rutas_por_vendedor(
    df: pd.DataFrame,
) -> dict[tuple[int, int], list[tuple[int, str]]]:
    """(id_vendedor, id_sucursal) -> [(id_ruta, des_ruta)].

    Igual que arriba: id_ruta tampoco es unico global (el codigo 1 existe en
    10+ sucursales), por eso la sucursal viaja en la clave.
    """
    rutas: dict[tuple[int, int], list[tuple[int, str]]] = {}
    for f in df.itertuples(index=False):
        if f.id_ruta is None or pd.isna(f.id_ruta):
            continue
        id_ruta = int(f.id_ruta)
        descripcion = (str(f.des_ruta).strip() if f.des_ruta is not None
                       and not pd.isna(f.des_ruta) else f"RUTA {id_ruta}")
        rutas.setdefault((int(f.id_vendedor), int(f.id_sucursal)), []).append(
            (id_ruta, descripcion)
        )
    return rutas


# ---------------------------------------------------------------------------
# Lectura del archivo fuente
# ---------------------------------------------------------------------------
def leer_cupos(path: Path, hoja: str) -> list[Vendedor]:
    """Lee el bloque 'Objetivo' del archivo de cupos del mes."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if hoja not in wb.sheetnames:
            raise ValueError(
                f"La hoja {hoja!r} no existe en {path.name}. "
                f"Hojas disponibles: {wb.sheetnames}"
            )
        ws = wb[hoja]
        vendedores: list[Vendedor] = []
        for row in range(SRC_ROW_INI, SRC_ROW_FIN + 1):
            nombre = ws[f"A{row}"].value
            sucursal = ws[f"B{row}"].value
            if not nombre or not sucursal:
                continue
            nombre, sucursal = str(nombre).strip(), str(sucursal).strip()
            if nombre.upper() == "TOTAL" or sucursal.upper() == "TOTAL":
                continue
            if sucursal not in SUCURSAL_IDS:
                raise ValueError(
                    f"Sucursal desconocida {sucursal!r} en la fila {row} de "
                    f"{path.name}. Agregala a SUCURSAL_IDS en constants.py."
                )
            vendedores.append(Vendedor(
                nombre=nombre,
                sucursal=sucursal,
                id_sucursal=SUCURSAL_IDS[sucursal],
                cupos={cat: float(ws[f"{col}{row}"].value or 0)
                       for cat, col in SRC_COLS.items()},
            ))
        return vendedores
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Distribucion
# ---------------------------------------------------------------------------
def split_proporcional(total: float, pesos: list[float]) -> list[float]:
    """Reparte `total` segun `pesos`, redondeando a 2 decimales.

    La ultima ruta absorbe el residuo, asi la suma de las partes es
    exactamente `total`. Sin pesos positivos reparte parejo.
    """
    n = len(pesos)
    if n == 0 or total == 0:
        return [0.0] * n
    suma = sum(pesos)
    if suma <= 0:
        partes = [round(total / n, 2)] * (n - 1)
    else:
        partes = [round(total * p / suma, 2) for p in pesos[:-1]]
    partes.append(round(total - sum(partes), 2))
    return partes


def distribuir_cupos(
    vendedores: list[Vendedor],
    rutas_de_vendedor: dict[tuple[int, int], list[tuple[int, str]]],
    mapa_vendedores: dict[tuple[str, int], int],
    historia: dict[tuple[int, int, str], float],
    nombre_overrides: dict[tuple[str, int], int] | None = None,
    rutas_override: dict[tuple[str, int], list[tuple[int, int, str]]] | None = None,
) -> Distribucion:
    """Abre el cupo de cada vendedor entre sus rutas segun la historia."""
    nombre_overrides = (NOMBRE_OVERRIDES if nombre_overrides is None
                        else nombre_overrides)
    rutas_override = (RUTAS_OVERRIDE if rutas_override is None
                      else rutas_override)
    resultado = Distribucion()

    for vendedor in vendedores:
        sid = vendedor.id_sucursal
        clave = (vendedor.nombre, sid)

        # Vendedor migrado: su historia vive en las rutas de otra sucursal.
        if clave in rutas_override:
            destinos = [(s_hist, rid, etiqueta)
                        for s_hist, rid, etiqueta in rutas_override[clave]]
            _emitir_filas(vendedor, destinos, historia, resultado)
            continue

        id_vendedor = nombre_overrides.get(clave) or mapa_vendedores.get(
            (vendedor.nombre.upper(), sid))
        rutas = (rutas_de_vendedor.get((id_vendedor, sid), [])
                 if id_vendedor is not None else [])

        if not rutas:
            # Sin clientes asignados: el cupo entero va a una fila testigo.
            resultado.sin_ruta.append(vendedor.nombre)
            resultado.filas.append({
                "sucursal": vendedor.sucursal, "vendedor": vendedor.nombre,
                "codigo": None, "ruta": ETIQUETA_SIN_RUTA,
                "vals": dict(vendedor.cupos),
            })
            continue

        _emitir_filas(vendedor,
                      [(sid, rid, descripcion) for rid, descripcion in rutas],
                      historia, resultado)

    return resultado


def _emitir_filas(
    vendedor: Vendedor,
    destinos: list[tuple[int, int, str]],
    historia: dict[tuple[int, int, str], float],
    resultado: Distribucion,
) -> None:
    """Reparte el cupo del vendedor entre `destinos` y agrega las filas.

    `destinos` es [(id_sucursal_historia, id_ruta, etiqueta)] — la sucursal
    puede diferir de la del cupo cuando el vendedor migro.
    """
    pesos: dict[str, list[float]] = {}
    for categoria in CATEGORIAS_DISTRIBUIBLES:
        pesos[categoria] = [historia.get((s_hist, rid, categoria), 0.0)
                            for s_hist, rid, _ in destinos]
        if vendedor.cupos[categoria] > 0 and sum(pesos[categoria]) <= 0:
            resultado.sin_historia.append(f"{vendedor.nombre} / {categoria}")

    reparto = {categoria: split_proporcional(vendedor.cupos[categoria],
                                             pesos[categoria])
               for categoria in CATEGORIAS_DISTRIBUIBLES}

    for i, (_, id_ruta, etiqueta) in enumerate(destinos):
        vals = {categoria: reparto[categoria][i]
                for categoria in CATEGORIAS_DISTRIBUIBLES}
        # Doble apertura: CERVEZAS es la suma de sus 5 marcas en esta ruta.
        vals["CERVEZAS"] = round(sum(vals[m] for m in CERVEZA_MARCAS), 2)
        resultado.filas.append({
            "sucursal": vendedor.sucursal, "vendedor": vendedor.nombre,
            "codigo": id_ruta, "ruta": etiqueta, "vals": vals,
        })


def validar(
    filas: list[dict], vendedores: list[Vendedor], tolerancia: float = 0.02
) -> dict[str, float]:
    """Chequea suma_rutas == cupo por (vendedor, categoria).

    Devuelve {"<vendedor>/<categoria>": diferencia} solo con las que fallan;
    dict vacio significa que el reparto cierra.
    """
    sumas: dict[tuple[str, str], dict[str, float]] = {}
    for fila in filas:
        acumulado = sumas.setdefault(
            (fila["sucursal"], fila["vendedor"]), {c: 0.0 for c in CATEGORIAS})
        for categoria in CATEGORIAS:
            acumulado[categoria] += fila["vals"][categoria]

    errores: dict[str, float] = {}
    for vendedor in vendedores:
        acumulado = sumas.get((vendedor.sucursal, vendedor.nombre), {})
        for categoria in CATEGORIAS:
            # CERVEZAS del archivo puede no cerrar con sus marcas; la
            # referencia real es la suma de las 5 marcas.
            esperado = (vendedor.cupos[categoria] if categoria != "CERVEZAS"
                        else round(sum(vendedor.cupos[m] for m in CERVEZA_MARCAS), 2))
            diferencia = abs(acumulado.get(categoria, 0.0) - esperado)
            if diferencia > tolerancia:
                errores[f"{vendedor.nombre}/{categoria}"] = round(diferencia, 4)
    return errores
