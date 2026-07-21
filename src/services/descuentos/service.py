"""Descuentos CCU service.

Migra el Excel manual "DESCUENTOS CCU.xlsx" (export del ERP) a generación desde
el Data Warehouse (gold). Produce dos hojas con tablas de descuentos
materializadas como valores (no tablas dinámicas):

- "normal":       3 tablas → por Sucursal, Sucursal×Genérico, Sucursal×Genérico×Marca
- "lista_precio": las mismas aperturas pero abiertas además por Lista de Precio

Medidas (mapeo gold, ver engram decision descuentos-ccu/data-mapping):
- ImporteNetoSinDesc = SUM(subtotal_neto + descuentos)   (importe antes del descuento comercial)
- Bonificacion$      = SUM(descuentos)                    (descuento en pesos)
- % Desc             = Bonificacion$ / ImporteNetoSinDesc

NO se filtra `anulado` (regla del proyecto). ENVASES CCU se excluye explícitamente.
CASA CENTRAL se splitea en VALLE SALTA según id_ruta (zona virtual de ventas).
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.core.output_paths import service_output_dir
from src.services.descuentos.constants import label_lista_precio

logger = logging.getLogger(__name__)

# Genéricos CCU incluidos (ENVASES CCU se excluye a propósito — suma en la BD
# pero no debe ir en el reporte).
GENERICOS_CCU: list[str] = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]

# Zona virtual: rutas de CASA CENTRAL (id_sucursal 1) que se renombran a VALLE SALTA.
# Espeja _ZONAS_VIRTUALES_VENTAS de src/services/ventas/service.py.
VALLE_SALTA_RUTAS: set[int] = {81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 118, 119, 120, 122}

# ── Estilos (espeja el look del Excel original) ──────────────────────────────
_TITLE_FONT = Font(bold=True, size=13)
_SUBTITLE_FONT = Font(italic=True, color="546E7A")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="A5D6A7")
_TOTAL_FONT = Font(bold=True)
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
_MONEY_FMT = "#,##0"
_PCT_FMT = "0.0%"

_MEASURES = ["ImporteNetoSinDesc", "Bonificacion$", "% Desc"]


@dataclass
class DescuentosConfig:
    fecha_desde: str
    fecha_hasta: str
    nombre_archivo: str | None = None
    genericos: list[str] = field(default_factory=lambda: list(GENERICOS_CCU))
    # Si es False, NO se genera la hoja "lista_precio" (solo la hoja "normal").
    con_lista_precio: bool = True


@dataclass
class DescuentosResult:
    ruta_archivo: Path
    registros_procesados: int
    sucursales: int


class DescuentosService:
    def __init__(self, data_loader: DataLoader | None = None) -> None:
        self.data_loader = data_loader or DataLoader()

    # ── Data ─────────────────────────────────────────────────────────────────
    def _fetch(self, config: DescuentosConfig) -> pd.DataFrame:
        genericos_sql = ", ".join(f"'{g}'" for g in config.genericos)
        query = f"""
            SELECT
                fv.id_sucursal,
                ds.descripcion                       AS sucursal_desc,
                dc.id_ruta_fv1                        AS id_ruta,
                da.generico                           AS generico,
                da.marca                              AS marca,
                dc.id_lista_precio                    AS id_lista_precio,
                SUM(fv.subtotal_neto + fv.descuentos) AS importe_neto_sin_desc,
                SUM(fv.descuentos)                    AS bonificacion_pesos
            FROM gold.fact_ventas fv
            JOIN gold.dim_articulo da ON fv.id_articulo = da.id_articulo
            JOIN gold.dim_cliente  dc ON fv.id_cliente  = dc.id_cliente
                                     AND fv.id_sucursal = dc.id_sucursal
            JOIN gold.dim_sucursal ds ON fv.id_sucursal = ds.id_sucursal
            WHERE fv.fecha_comprobante BETWEEN :desde AND :hasta
              AND da.generico IN ({genericos_sql})
            GROUP BY fv.id_sucursal, ds.descripcion, dc.id_ruta_fv1,
                     da.generico, da.marca, dc.id_lista_precio
        """
        df = self.data_loader.execute_query(
            query, {"desde": config.fecha_desde, "hasta": config.fecha_hasta}
        )
        if df.empty:
            return df

        # Etiqueta de sucursal "<id> - <descripcion>", con split de zona virtual:
        # CASA CENTRAL (id 1) + ruta en VALLE_SALTA_RUTAS → "VALLE SALTA".
        def _label(row) -> str:
            if row["id_sucursal"] == 1 and row["id_ruta"] in VALLE_SALTA_RUTAS:
                return "VALLE SALTA"
            return f"{row['id_sucursal']} - {row['sucursal_desc']}"

        df["sucursal"] = df.apply(_label, axis=1)
        df["lista_precio"] = df["id_lista_precio"].map(label_lista_precio)
        df = df.drop(columns=["id_sucursal", "sucursal_desc", "id_ruta", "id_lista_precio"])
        # Reagrupar tras el split (varias rutas colapsan en la misma sucursal/zona).
        df = df.groupby(
            ["sucursal", "generico", "marca", "lista_precio"], as_index=False, dropna=False
        ).sum()
        return df

    # ── Excel ──────────────────────────────────────────────────────────────────
    def _write_table(self, ws, start_row: int, title: str, df: pd.DataFrame, group_cols: list[str]) -> int:
        """Escribe una tabla agrupada (suma de medidas) y devuelve la fila siguiente."""
        agg = (
            df.groupby(group_cols, as_index=False, dropna=False)[
                ["importe_neto_sin_desc", "bonificacion_pesos"]
            ]
            .sum()
            .sort_values(group_cols)
        )

        ws.cell(row=start_row, column=1, value=title).font = _TITLE_FONT
        header_row = start_row + 1
        headers = group_cols + _MEASURES
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=header_row, column=col_idx, value=h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.border = _THIN
            c.alignment = Alignment(horizontal="center")

        r = header_row + 1
        n_keys = len(group_cols)
        for _, row in agg.iterrows():
            for j, gc in enumerate(group_cols, 1):
                c = ws.cell(row=r, column=j, value=row[gc])
                c.border = _THIN
            insd = float(row["importe_neto_sin_desc"])
            bonif = float(row["bonificacion_pesos"])
            pct = (bonif / insd) if insd else 0.0
            for off, (val, fmt) in enumerate(
                [(insd, _MONEY_FMT), (bonif, _MONEY_FMT), (pct, _PCT_FMT)]
            ):
                c = ws.cell(row=r, column=n_keys + 1 + off, value=val)
                c.number_format = fmt
                c.border = _THIN
                c.alignment = Alignment(horizontal="right")
            r += 1

        # Total general
        tot_insd = float(agg["importe_neto_sin_desc"].sum())
        tot_bonif = float(agg["bonificacion_pesos"].sum())
        tot_pct = (tot_bonif / tot_insd) if tot_insd else 0.0
        c = ws.cell(row=r, column=1, value="Total general")
        c.fill = _TOTAL_FILL
        c.font = _TOTAL_FONT
        c.border = _THIN
        for j in range(2, n_keys + 1):
            cc = ws.cell(row=r, column=j)
            cc.fill = _TOTAL_FILL
            cc.border = _THIN
        for off, (val, fmt) in enumerate(
            [(tot_insd, _MONEY_FMT), (tot_bonif, _MONEY_FMT), (tot_pct, _PCT_FMT)]
        ):
            cc = ws.cell(row=r, column=n_keys + 1 + off, value=val)
            cc.number_format = fmt
            cc.fill = _TOTAL_FILL
            cc.font = _TOTAL_FONT
            cc.border = _THIN
            cc.alignment = Alignment(horizontal="right")
        return r + 3  # spacer entre tablas

    def _autosize(self, ws) -> None:
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)

    def _build_workbook(
        self, df: pd.DataFrame, output_path: Path, periodo: str, con_lista_precio: bool = True
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        # Hoja "normal": aperturas sin lista de precio
        ws_n = wb.create_sheet("normal")
        ws_n.cell(row=1, column=1, value="Descuentos CCU").font = Font(bold=True, size=14)
        ws_n.cell(row=2, column=1, value=f"Período: {periodo}").font = _SUBTITLE_FONT
        nxt = 4
        nxt = self._write_table(ws_n, nxt, "Por Sucursal", df, ["sucursal"])
        nxt = self._write_table(ws_n, nxt, "Por Sucursal y Genérico", df, ["sucursal", "generico"])
        nxt = self._write_table(ws_n, nxt, "Por Sucursal, Genérico y Marca", df, ["sucursal", "generico", "marca"])
        self._autosize(ws_n)

        # Hoja "lista_precio": mismas aperturas + Lista de Precio como nivel adicional.
        # Se omite cuando con_lista_precio=False (ej. reporte de Walter Vilte).
        if con_lista_precio:
            ws_l = wb.create_sheet("lista_precio")
            ws_l.cell(row=1, column=1, value="Descuentos CCU — apertura por Lista de Precio").font = Font(bold=True, size=14)
            ws_l.cell(row=2, column=1, value=f"Período: {periodo}").font = _SUBTITLE_FONT
            nxt = 4
            nxt = self._write_table(ws_l, nxt, "Por Sucursal y Lista de Precio", df, ["sucursal", "lista_precio"])
            nxt = self._write_table(ws_l, nxt, "Por Sucursal, Lista de Precio y Genérico", df, ["sucursal", "lista_precio", "generico"])
            nxt = self._write_table(ws_l, nxt, "Por Sucursal, Lista de Precio, Genérico y Marca", df, ["sucursal", "lista_precio", "generico", "marca"])
            self._autosize(ws_l)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

    # ── Entrada principal ────────────────────────────────────────────────────
    def generar_reporte(self, config: DescuentosConfig) -> DescuentosResult:
        df = self._fetch(config)
        nombre = config.nombre_archivo or "Descuentos CCU"
        out_dir = service_output_dir("descuentos-ccu", config.fecha_desde, "month")
        output_path = out_dir / f"{nombre}.xlsx"
        periodo = f"{config.fecha_desde} a {config.fecha_hasta}"
        self._build_workbook(df, output_path, periodo, con_lista_precio=config.con_lista_precio)
        logger.info("Descuentos CCU generado: %s (%d filas)", output_path.name, len(df))
        return DescuentosResult(
            ruta_archivo=output_path,
            registros_procesados=len(df),
            sucursales=df["sucursal"].nunique() if not df.empty else 0,
        )
