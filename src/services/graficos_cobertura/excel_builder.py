"""Build resumen.xlsx summary workbook.

Receives already-processed DataFrames (one per sheet) and writes them via
openpyxl. Separating write from processing keeps this module testable
without needing the full data pipeline.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet with a styled header row."""
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in df.itertuples(index=False, name=None):
        ws.append([(v if v != 0 else "") if isinstance(v, (int, float)) else v for v in row])

    # Auto-width (lightweight)
    for col_idx, col_name in enumerate(df.columns, 1):
        width = max(len(str(col_name)), 8) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _safe_sheet_name(name: str) -> str:
    """Excel caps sheet names at 31 chars."""
    return name[:31]


def build_resumen_xlsx(
    output_path: Path,
    sheets_por_generico: dict[str, pd.DataFrame],
    sheets_mensuales: dict[str, pd.DataFrame],
    sheet_comparativo: pd.DataFrame | None,
) -> Path:
    """Write resumen.xlsx with three sections: per-generico, mensual, comparativo.

    Args:
        output_path: full path (including filename) where to write the xlsx.
        sheets_por_generico: mapping generico name -> DataFrame to write.
        sheets_mensuales: mapping sheet name (e.g. "Marzo 2026") -> DataFrame.
        sheet_comparativo: optional comparativo DataFrame. If None, no sheet.

    Returns:
        The output_path (unchanged).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Drop the default empty sheet — we'll add our own
    default_sheet = wb.active
    wb.remove(default_sheet)

    for generico, df in sheets_por_generico.items():
        ws = wb.create_sheet(title=_safe_sheet_name(generico))
        _write_sheet(ws, df)

    for name, df in sheets_mensuales.items():
        ws = wb.create_sheet(title=_safe_sheet_name(name))
        _write_sheet(ws, df)

    if sheet_comparativo is not None and not sheet_comparativo.empty:
        ws = wb.create_sheet(title=_safe_sheet_name("Comparativo"))
        _write_sheet(ws, sheet_comparativo)

    # If no sheets at all, add a placeholder (openpyxl can't save empty wb)
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Vacio")

    wb.save(output_path)
    return output_path
