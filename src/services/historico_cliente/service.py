"""
HistoricoClienteService - Ventas historicas por cliente, una hoja por cliente,
filas = articulos o marcas (segun filtro en config), columnas = meses.
"""
import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from src.core.excel_writer import ExcelWriter, SheetStyle, ColumnFormat
from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)


@dataclass
class HistoricoClienteConfig:
    """Configuracion para el reporte historico por cliente."""
    fecha_desde: str
    fecha_hasta: str
    clientes: list[dict]                  # {"id_cliente": int, "id_sucursal": int}
    articulos: list[int] | None = None    # mutually exclusive with marcas
    marcas: list[str] | None = None       # mutually exclusive with articulos
    # When True: show ALL marcas grouped by generico, with a subtotal row per
    # generico and a grand-total row. No marca/articulo filter required.
    agrupar_por_generico: bool = False
    # When True (requires agrupar_por_generico): fill the FULL marca universe of
    # `genericos_universo` (from dim_articulo), showing 0 for marcas the client
    # did not buy — highlights coverage gaps.
    marcas_completas: bool = False
    genericos_universo: list[str] | None = None
    # When True, only billed units are counted (cantidades_con_cargo): 100%-discount
    # lines (gifts) are excluded from every month and from the totals.
    solo_con_cargo: bool = False
    nombre_archivo: str | None = None


@dataclass
class HistoricoClienteResult:
    """Resultado del reporte historico por cliente."""
    ruta_archivo: Path
    sheets_generated: list[str]
    registros_procesados: int


_STYLE = SheetStyle(
    numeric_format="#,##0.##",
    column_formats={},
    as_table=True,
    table_style="TableStyleMedium9",
)

# Grouped mode uses a plain (non-table) sheet so subtotal/grand-total rows can be
# interleaved. Header styling is applied by ExcelWriter; subtotal/grand rows are
# post-styled below.
#
# auto_width is off on purpose: it pads every month column to the header width
# plus slack, which on a 12-month sheet wastes most of the horizontal space in
# the WhatsApp capture. `_apply_compact_layout` sets every width explicitly.
_GROUPED_STYLE = SheetStyle(
    numeric_format="#,##0.##",
    column_formats={},
    as_table=False,
    auto_width=False,
)

# The grouped sheet is consumed as an image on a phone, not as a spreadsheet:
# bigger type, and columns cut down to what their content actually needs.
_FONT_SIZE = 14
_FONT_SIZE_TITULO = 18
_FONT_SIZE_SUBTITULO = 10

# Arial and not the openpyxl default: LibreOffice on the render host maps Arial
# to Liberation Sans, which is metric-identical, while Excel on Windows uses
# real Arial — so a width computed here holds in both. Calibri instead falls
# back to Noto Sans here, whose metrics do NOT match, which is what made the
# first pass at these widths clip.
_FONT_NAME = "Arial"

# Excel width units track the workbook's 11pt Calibri normal style, so a string
# set at _FONT_SIZE in Arial needs this many units per character.
_CHAR_W = (_FONT_SIZE / 11) * 1.10

# Floors: a column never gets narrower than this even with tiny content.
_WIDTH_GENERICO_MIN = 16
_WIDTH_MARCA_MIN = 20
_WIDTH_MES_MIN = 8
_WIDTH_TOTAL_MIN = 10

# Corporate blue palette. The grand total is the darkest block on the sheet and
# the only one with reversed (white-on-navy) type, so it reads as the endpoint.
_HEADER_FILL = "1F3864"     # navy — column header
_HEADER_FONT = "FFFFFF"
_SUBTOTAL_FILL = "D9E2F3"   # light blue — per-generico subtotal
_SUBTOTAL_FONT = "1F3864"
_GRAND_FILL = "1F3864"      # navy — grand total
_GRAND_FONT = "FFFFFF"
_TOTAL_COL_FILL = "EDF2FA"  # faint blue wash down the Total column
_BANDA_FILL = "F7F9FC"      # zebra banding on alternating data rows
_ZERO_FILL = "F4F6FA"       # marca sin venta (hueco de compra)
_ZERO_FONT = "A3B0C4"
_TITULO_FONT = "1F3864"
_SUBTITULO_FONT = "7F7F7F"
_BORDE_COLOR = "B4C6E7"     # light blue grid
_BORDE_FUERTE = "1F3864"    # navy rule under the header / above the grand total

_GENERICO_COL = "Genérico"
_MARCA_COL = "Marca"
_TOTAL_COL = "Total"

# Rows inserted above the table: title, subtitle, spacer.
_FILAS_TITULO = 3


class HistoricoClienteService(BaseService):
    """Genera Excel con historico de ventas por cliente, mes a mes."""

    SERVICE_SLUG = "historico-cliente"
    GRANULARITY = "month"

    def generar_reporte(self, config: HistoricoClienteConfig) -> HistoricoClienteResult:
        # 1. Validate filter combination
        if config.articulos and config.marcas:
            raise ValueError("Config tiene 'articulos' Y 'marcas'. Especifica solo uno.")
        # Grouped mode shows all marcas grouped by generico → no filter required.
        if not config.agrupar_por_generico and not config.articulos and not config.marcas:
            raise ValueError(
                "Config debe tener 'articulos' O 'marcas' (o usar 'agrupar_por_generico')."
            )

        # 2. Fetch data
        df = self.data_loader.get_ventas_historico_cliente(
            fecha_desde=config.fecha_desde,
            fecha_hasta=config.fecha_hasta,
            clientes=config.clientes,
            articulos=config.articulos,
            marcas=config.marcas,
            agrupar_por_generico=config.agrupar_por_generico,
            solo_con_cargo=config.solo_con_cargo,
        )

        # 2b. Optional: full marca universe (for the "marcas completas" mode)
        universe_df = None
        if config.agrupar_por_generico and config.marcas_completas:
            universe_df = self.data_loader.get_marca_universe(
                config.genericos_universo or []
            )

        # 3. Build full month range from config
        meses = (
            pd.date_range(
                start=config.fecha_desde, end=config.fecha_hasta, freq="MS"
            )
            .strftime("%Y-%m")
            .tolist()
        )

        # 4. Build Excel
        nombre = config.nombre_archivo or "Historico Cliente"
        output_dir = service_output_dir("historico-cliente", config.fecha_desde, granularity="month")
        writer = ExcelWriter(nombre, output_dir=output_dir)

        sheets_generated: list[str] = []
        total_registros = 0

        for cliente_cfg in config.clientes:
            id_cli = cliente_cfg["id_cliente"]
            id_suc = cliente_cfg["id_sucursal"]

            df_cli = df[(df["id_cliente"] == id_cli) & (df["id_sucursal"] == id_suc)]
            if df_cli.empty:
                logger.warning(
                    "Cliente id_cliente=%d id_sucursal=%d sin datos, se omite la hoja.",
                    id_cli, id_suc,
                )
                continue

            # Sheet name: nombre_cliente or fallback, truncated to 31 chars
            nombre_cliente = str(df_cli["nombre_cliente"].iloc[0])
            sheet_name = (nombre_cliente or f"{id_cli}-{id_suc}")[:31]
            # Avoid duplicates if truncated name collides
            base = sheet_name
            counter = 1
            while sheet_name in sheets_generated:
                suffix = f" ({counter})"
                sheet_name = base[: 31 - len(suffix)] + suffix
                counter += 1

            if config.agrupar_por_generico:
                sheet_df, subtotal_rows, grand_row, zero_rows = _build_grouped_frame(
                    df_cli, meses, universe_df
                )
                writer.add_sheet(sheet_df, sheet_name=sheet_name, style=_GROUPED_STYLE)
                ws = writer.workbook[sheet_name]
                # Layout first: it rewrites every font, so the summary-row
                # styling below must run after or it would be overwritten.
                _apply_corporate_layout(
                    ws,
                    list(sheet_df.columns),
                    titulo=f"{nombre_cliente} — histórico de compras por marca",
                    subtitulo=(
                        f"Bultos por mes · {config.fecha_desde} a {config.fecha_hasta} · "
                        f"cliente {id_cli} · sucursal {id_suc}"
                        # The basis has to travel with the image: gifts can be a
                        # double-digit share of a month, so a reader who only
                        # sees the capture must know which total this is.
                        + (" · SOLO unidades CON CARGO (excluye bonificación 100%, sin cargo)"
                           if config.solo_con_cargo else "")
                        + (" · marcas en gris: sin compra en el período"
                           if config.marcas_completas else "")
                    ),
                )
                _style_summary_rows(
                    ws, subtotal_rows, grand_row, zero_rows,
                    n_cols=len(sheet_df.columns),
                )
                total_registros += len(sheet_df)
            else:
                # Pivot: rows = row_key, cols = mes, values = bultos
                pivot = df_cli.pivot_table(
                    index="row_key",
                    columns="mes",
                    values="bultos",
                    aggfunc="sum",
                    fill_value=0,
                ).reindex(columns=meses, fill_value=0)

                pivot = pivot.reset_index().rename(
                    columns={"row_key": "Marca" if config.marcas else "Articulo"}
                )
                pivot["Total"] = pivot[meses].sum(axis=1)

                writer.add_sheet(pivot, sheet_name=sheet_name, style=_STYLE)
                total_registros += len(pivot)

            sheets_generated.append(sheet_name)

        ruta = writer.save()

        return HistoricoClienteResult(
            ruta_archivo=ruta,
            sheets_generated=sheets_generated,
            registros_procesados=total_registros,
        )


def _build_grouped_frame(
    df_cli: pd.DataFrame, meses: list[str], universe_df: pd.DataFrame | None = None
) -> tuple[pd.DataFrame, list[int], int, list[int]]:
    """Build the grouped-by-generico sheet frame with interleaved summary rows.

    Rows are: for each generico (ordered by descending total) its marcas (also by
    descending total), followed by a ``TOTAL {generico}`` subtotal row; finally a
    ``TOTAL GENERAL`` grand-total row.

    When ``universe_df`` (columns generico, marca) is provided, the frame is
    reindexed onto the full marca universe so marcas the client never bought
    appear with 0 across every month (coverage-gap view).

    Returns:
        (frame, subtotal_row_indices, grand_row_index, zero_row_indices) where the
        row indices are 0-based positions within the frame's data rows (excluding
        the header). ``zero_row_indices`` are marca rows whose Total is 0.
    """
    # Pivot at (generico, marca) grain, then reindex months to the full range.
    pivot = df_cli.pivot_table(
        index=["generico", "row_key"],
        columns="mes",
        values="bultos",
        aggfunc="sum",
        fill_value=0,
    ).reindex(columns=meses, fill_value=0)

    # Fill the full marca universe with 0 for marcas the client did not buy.
    if universe_df is not None and not universe_df.empty:
        uni_index = pd.MultiIndex.from_frame(
            universe_df.rename(columns={"marca": "row_key"})[["generico", "row_key"]]
        )
        pivot = pivot.reindex(uni_index.union(pivot.index), fill_value=0)

    pivot["Total"] = pivot[meses].sum(axis=1)

    gen_totales = (
        pivot["Total"].groupby(level="generico").sum().sort_values(ascending=False)
    )

    records: list[dict] = []
    subtotal_rows: list[int] = []
    zero_rows: list[int] = []
    for generico in gen_totales.index:
        block = pivot.xs(generico, level="generico").sort_values(
            "Total", ascending=False
        )
        for marca, row in block.iterrows():
            if row[_TOTAL_COL] == 0:
                zero_rows.append(len(records))
            records.append(
                {_GENERICO_COL: generico, _MARCA_COL: marca,
                 **{m: row[m] for m in meses}, _TOTAL_COL: row[_TOTAL_COL]}
            )
        # Subtotal row for this generico
        subtotal_rows.append(len(records))
        records.append(
            {_GENERICO_COL: "", _MARCA_COL: f"TOTAL {generico}",
             **{m: block[m].sum() for m in meses}, _TOTAL_COL: block[_TOTAL_COL].sum()}
        )

    # Grand total
    grand_row = len(records)
    records.append(
        {_GENERICO_COL: "", _MARCA_COL: "TOTAL GENERAL",
         **{m: pivot[m].sum() for m in meses}, _TOTAL_COL: pivot[_TOTAL_COL].sum()}
    )

    frame = pd.DataFrame(records, columns=[_GENERICO_COL, _MARCA_COL, *meses, _TOTAL_COL])
    # Twelve month columns drive the sheet width, so they carry the short label.
    # Only this grouped view is relabelled; the marca/articulo sheets keep the
    # full YYYY-MM, which is unambiguous and sorts lexicographically.
    frame = frame.rename(columns={m: _mes_label(m) for m in meses})
    return frame, subtotal_rows, grand_row, zero_rows


def _mes_label(mes: str) -> str:
    """'2026-01' -> '01/26'. Two characters shorter, same information."""
    anio, mm = mes.split("-")
    return f"{mm}/{anio[2:]}"


def _apply_corporate_layout(
    ws, columnas: list[str], titulo: str, subtitulo: str
) -> None:
    """Turn the raw grid into the corporate blue sheet.

    Inserts the title band, repaints the header, draws the grid, bands the data
    rows and sizes every column to its own content. Runs BEFORE
    ``_style_summary_rows``, which then overrides the subtotal/grand rows.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    n_cols = len(columnas)
    ws.insert_rows(1, _FILAS_TITULO)
    fila_hdr = _FILAS_TITULO + 1

    # ── Title band ────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value=titulo).font = Font(
        name=_FONT_NAME, size=_FONT_SIZE_TITULO, bold=True, color=_TITULO_FONT
    )
    ws.cell(row=2, column=1, value=subtitulo).font = Font(
        name=_FONT_NAME, size=_FONT_SIZE_SUBTITULO, italic=True, color=_SUBTITULO_FONT
    )
    for fila in (1, 2, 3):
        for col in range(1, n_cols + 1):
            celda = ws.cell(row=fila, column=col)
            if celda.font.name != _FONT_NAME:
                celda.font = Font(name=_FONT_NAME, size=_FONT_SIZE)

    # ── Grid ──────────────────────────────────────────────────────────────
    fino = Side(style="thin", color=_BORDE_COLOR)
    grueso = Side(style="medium", color=_BORDE_FUERTE)
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)

    idx_total = columnas.index(_TOTAL_COL) + 1 if _TOTAL_COL in columnas else None
    banda = PatternFill(start_color=_BANDA_FILL, end_color=_BANDA_FILL, fill_type="solid")
    wash = PatternFill(start_color=_TOTAL_COL_FILL, end_color=_TOTAL_COL_FILL, fill_type="solid")

    for fila in range(fila_hdr, ws.max_row + 1):
        es_header = fila == fila_hdr
        # Zebra on data rows only; summary rows get repainted right after.
        raya = (not es_header) and ((fila - fila_hdr) % 2 == 0)
        for col in range(1, n_cols + 1):
            celda = ws.cell(row=fila, column=col)
            celda.border = borde
            if es_header:
                celda.fill = PatternFill(
                    start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type="solid"
                )
                celda.font = Font(
                    name=_FONT_NAME, size=_FONT_SIZE, bold=True, color=_HEADER_FONT
                )
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                celda.font = Font(name=_FONT_NAME, size=_FONT_SIZE)
                if col == idx_total:
                    celda.fill = wash
                elif raya:
                    celda.fill = banda
                if col > 2:
                    celda.alignment = Alignment(horizontal="center")

    # Navy rule closing the header.
    for col in range(1, n_cols + 1):
        c = ws.cell(row=fila_hdr, column=col)
        c.border = Border(left=fino, right=fino, top=fino, bottom=grueso)

    # ── Widths ────────────────────────────────────────────────────────────
    pisos = {
        _GENERICO_COL: _WIDTH_GENERICO_MIN,
        _MARCA_COL: _WIDTH_MARCA_MIN,
        _TOTAL_COL: _WIDTH_TOTAL_MIN,
    }
    for idx, nombre in enumerate(columnas, start=1):
        piso = pisos.get(nombre, _WIDTH_MES_MIN)
        ws.column_dimensions[get_column_letter(idx)].width = max(
            piso, _width_for_column(ws, idx, desde=fila_hdr)
        )


def rango_captura(ws) -> str:
    """A1 range covering the whole used area of ``ws``.

    Public on purpose: every caller that captures this sheet as an image must
    derive the range instead of hardcoding it. A stale hardcoded range crops
    the image — typically eating the TOTAL GENERAL row — while the xlsx stays
    correct and nothing raises, so the loss is invisible until someone reads
    the number off the picture.
    """
    from openpyxl.utils import get_column_letter

    return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def columnas_desbordadas(ws) -> list[str]:
    """Column letters whose width cannot fit their own widest value.

    LibreOffice renders a number that does not fit as ``###``, so the figure
    vanishes from the capture. Call this before sending the image.
    """
    from openpyxl.utils import get_column_letter

    fuera: list[str] = []
    for idx in range(1, ws.max_column + 1):
        letra = get_column_letter(idx)
        dim = ws.column_dimensions.get(letra)
        ancho = getattr(dim, "width", None)
        if not ancho:
            continue
        necesita = max(
            (len(_formato_visible(ws.cell(row=r, column=idx).value))
             for r in range(_FILAS_TITULO + 1, ws.max_row + 1)),
            default=0,
        ) * _CHAR_W
        if ancho < necesita:
            fuera.append(letra)
    return fuera


def _width_for_column(ws, idx: int, desde: int = 1) -> float:
    """Width needed by the widest value in column ``idx`` at ``_FONT_SIZE``.

    A fixed width is not safe here: LibreOffice renders an overflowing number
    as ``###`` in the capture, so a month that happens to reach three digits
    would drop out of the image while the xlsx still read fine.
    """
    mas_largo = max(
        (len(_formato_visible(ws.cell(row=r, column=idx).value))
         for r in range(desde, ws.max_row + 1)),
        default=0,
    )
    return mas_largo * _CHAR_W + 1.5  # slack for the bold summary rows


def _formato_visible(value) -> str:
    """Render a cell value the way the '#,##0.##' number format shows it."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        texto = f"{value:,.2f}"
        return texto.rstrip("0").rstrip(".") if "." in texto else texto
    return str(value)


def _style_summary_rows(
    ws, subtotal_rows: list[int], grand_row: int, zero_rows: list[int], n_cols: int
) -> None:
    """Fill the subtotal, grand-total and zero-sale rows of the grouped sheet.

    Row indices are 0-based within the data rows. The title band occupies the
    first ``_FILAS_TITULO`` rows and the column header the next one, so data row
    ``i`` maps to worksheet row ``i + _FILAS_TITULO + 2``.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fino = Side(style="thin", color=_BORDE_COLOR)
    grueso = Side(style="medium", color=_BORDE_FUERTE)

    def _paint(
        data_idx: int, color: str, *, bold: bool = True,
        font_color: str | None = None, regla_arriba: bool = False,
    ) -> None:
        excel_row = data_idx + _FILAS_TITULO + 2
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        borde = Border(
            left=fino, right=fino, bottom=fino,
            top=grueso if regla_arriba else fino,
        )
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = fill
            # name/size are respecified: rebuilding the Font would otherwise drop
            # back to the 11pt Calibri default set before the layout pass.
            cell.font = Font(
                name=_FONT_NAME, size=_FONT_SIZE, bold=bold, color=font_color
            )
            cell.border = borde
            if col > 2:
                cell.alignment = Alignment(horizontal="center")

    # Zero-sale rows first (subtotal/grand override if they ever collided).
    for idx in zero_rows:
        _paint(idx, _ZERO_FILL, bold=False, font_color=_ZERO_FONT)
    for idx in subtotal_rows:
        _paint(idx, _SUBTOTAL_FILL, font_color=_SUBTOTAL_FONT)
    _paint(grand_row, _GRAND_FILL, font_color=_GRAND_FONT, regla_arriba=True)
