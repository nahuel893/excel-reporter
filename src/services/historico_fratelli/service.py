"""
HistoricoFratelliService - Genera Excel con historico de ventas de FRATELLI B.

Crea un archivo Excel con 3 secciones:
1. Totales por mes (todos los datos)
2. Desglose por marca
3. Desglose por lista de precio
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────

MESES = [
    "Ene",
    "Feb",
    "Mar",
    "Abr",
    "May",
    "Jun",
    "Jul",
    "Ago",
    "Sep",
    "Oct",
    "Nov",
    "Dic",
]
ANIOS = [2024, 2025, 2026]

# ── Styles ─────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True)
_YEAR_FONT = Font(bold=True, color="FFFFFF")
_YEAR_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_SECTION_FONT = Font(bold=True, size=13)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUMBER_FORMAT = "#,##0"


# ── Pure helper functions ───────────────────────────────────


def _pivot_por_mes(df: pd.DataFrame, val_col: str = "cantidad") -> dict:
    """
    Aggregate by (anio, mes), return {anio: [12 values]}.
    Missing months filled with 0.
    """
    result = {}
    for anio in ANIOS:
        values = [0] * 12
        if not df.empty:
            df_anio = df[df["anio"] == anio]
            for _, row in df_anio.iterrows():
                mes_idx = int(row["mes"]) - 1
                if 0 <= mes_idx < 12:
                    values[mes_idx] += int(row[val_col])
        result[anio] = values
    return result


def _pivot_por_dimension(
    df: pd.DataFrame, dim_col: str, val_col: str = "cantidad"
) -> dict:
    """
    Aggregate by (anio, mes, dim_col), pivot to DataFrame per year.
    Returns {anio: DataFrame with dim_col + 12 month columns}.
    Missing months filled with 0.
    """
    result = {}
    for anio in ANIOS:
        if df.empty:
            result[anio] = pd.DataFrame()
            continue
        df_anio = df[df["anio"] == anio]
        if df_anio.empty:
            result[anio] = pd.DataFrame()
            continue
        # Aggregate by dim_col and mes
        agg = df_anio.groupby([dim_col, "mes"])[val_col].sum().reset_index()
        # Get all unique dimension values
        dim_values = sorted(
            agg[dim_col].unique().tolist(),
            key=lambda x: (x is None, str(x) if x is not None else ""),
        )
        rows = []
        for val in dim_values:
            row = {dim_col: val}
            df_val = agg[agg[dim_col] == val]
            for m in range(1, 13):
                match = df_val[df_val["mes"] == m]
                row[m] = int(match[val_col].iloc[0]) if not match.empty else 0
            rows.append(row)
        pivot_df = pd.DataFrame(rows, columns=[dim_col] + list(range(1, 13)))
        result[anio] = pivot_df
    return result


def _write_section_title(ws, row: int, title: str) -> int:
    """Write bold section title. Returns next row."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    return row + 1


def _write_year_block_simple(
    ws, row: int, anio: int, values: list, with_totals: bool = False
) -> int:
    """
    Write year label (merged, styled) + month headers + 1 data row.
    Returns next row (with spacing).
    """
    end_col = 13 if with_totals else 12
    # Year label row — merge across months (col 1 to end_col)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=anio)
    cell.font = _YEAR_FONT
    cell.fill = _YEAR_FILL
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # Month headers
    for i, mes in enumerate(MESES):
        cell = ws.cell(row=row, column=i + 1, value=mes)
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    if with_totals:
        total_header = ws.cell(row=row, column=13, value="Total")
        total_header.font = _HEADER_FONT
        total_header.border = _THIN_BORDER
        total_header.alignment = Alignment(horizontal="center")
    row += 1

    # Data row
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=i + 1, value=val)
        cell.number_format = _NUMBER_FORMAT
        cell.border = _THIN_BORDER

    if with_totals:
        total_cell = ws.cell(row=row, column=13, value=sum(values))
        total_cell.number_format = _NUMBER_FORMAT
        total_cell.font = _HEADER_FONT
        total_cell.border = _THIN_BORDER
    row += 1

    # Spacing
    return row + 1


def _write_year_block_with_rows(
    ws, row: int, anio: int, df: pd.DataFrame, dim_col: str, with_totals: bool = False
) -> int:
    """
    Write year label + month headers + N data rows.
    Handles empty df. Returns next row (with spacing).
    """
    end_col = 14 if with_totals else 13
    # Year label — merge across dim_col + months (+ total)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=anio)
    cell.font = _YEAR_FONT
    cell.fill = _YEAR_FILL
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # Month headers — col 1 = dim label, cols 2-13 = months
    header_cell = ws.cell(row=row, column=1, value=dim_col.replace("_", " ").title())
    header_cell.font = _HEADER_FONT
    header_cell.border = _THIN_BORDER
    for i, mes in enumerate(MESES):
        cell = ws.cell(row=row, column=i + 2, value=mes)
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    if with_totals:
        total_header = ws.cell(row=row, column=14, value="Total")
        total_header.font = _HEADER_FONT
        total_header.border = _THIN_BORDER
        total_header.alignment = Alignment(horizontal="center")
    row += 1

    if df is None or df.empty:
        return row + 1

    # Data rows
    for _, data_row in df.iterrows():
        ws.cell(row=row, column=1, value=data_row[dim_col])
        row_sum = 0
        for m in range(1, 13):
            val = data_row.get(m, 0)
            cell = ws.cell(row=row, column=m + 1, value=val)
            cell.number_format = _NUMBER_FORMAT
            row_sum += val

        if with_totals:
            total_cell = ws.cell(row=row, column=14, value=row_sum)
            total_cell.number_format = _NUMBER_FORMAT
            total_cell.font = _HEADER_FONT
            total_cell.border = _THIN_BORDER
        row += 1

    # Column totals row
    if with_totals:
        total_label = ws.cell(row=row, column=1, value="Total")
        total_label.font = _HEADER_FONT
        total_label.border = _THIN_BORDER
        grand_total = 0
        for m in range(1, 13):
            col_total = df[m].sum() if m in df.columns else 0
            cell = ws.cell(row=row, column=m + 1, value=col_total)
            cell.number_format = _NUMBER_FORMAT
            cell.font = _HEADER_FONT
            cell.border = _THIN_BORDER
            grand_total += col_total

        gt_cell = ws.cell(row=row, column=14, value=grand_total)
        gt_cell.number_format = _NUMBER_FORMAT
        gt_cell.font = _HEADER_FONT
        gt_cell.border = _THIN_BORDER
        row += 1

    # Spacing
    return row + 1


# ── Dataclasses ────────────────────────────────────────────


@dataclass
class HistoricoFratelliConfig:
    """Configuracion para el informe Historico FRATELLI B."""

    nombre_archivo: str | None = None


@dataclass
class HistoricoFratelliResult:
    """Resultado de la generacion del informe Historico FRATELLI B."""

    ruta_archivo: Path
    registros_procesados: int
    hojas: list = field(default_factory=list)


# ── Service ────────────────────────────────────────────────


class HistoricoFratelliService(BaseService):
    """Servicio para generar el informe historico de FRATELLI B."""

    SERVICE_SLUG = "historico-fratelli"
    GRANULARITY = "month"

    def generar_reporte(
        self, config: HistoricoFratelliConfig
    ) -> HistoricoFratelliResult:
        """Genera el Excel con 3 secciones de datos historicos."""
        df = self.data_loader.get_ventas_historico_fratelli()

        registros_procesados = len(df)

        # Pivot data for each section
        pivot_mes = _pivot_por_mes(df)
        pivot_marca = _pivot_por_dimension(df, "marca")
        pivot_lista = _pivot_por_dimension(df, "id_lista_precio")

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historico FRATELLI B"

        row = 1

        # Section 1: Totales por mes
        row = _write_section_title(ws, row, "Totales por Mes")
        for anio in ANIOS:
            row = _write_year_block_simple(
                ws, row, anio, pivot_mes[anio], with_totals=True
            )

        # Section 2: Por lista de precio
        row = _write_section_title(ws, row, "Por Lista de Precio")
        for anio in ANIOS:
            row = _write_year_block_with_rows(
                ws, row, anio, pivot_lista[anio], "id_lista_precio", with_totals=True
            )

        # Section 3: Por marca
        row = _write_section_title(ws, row, "Por Marca")
        for anio in ANIOS:
            row = _write_year_block_with_rows(
                ws, row, anio, pivot_marca[anio], "marca", with_totals=True
            )

        hojas = [ws.title]

        # ── Hoja de Descuentos ──
        ws_desc = wb.create_sheet("Descuentos FRATELLI B")
        pivot_mes_desc = _pivot_por_mes(df, val_col="descuentos")
        pivot_lista_desc = _pivot_por_dimension(
            df, "id_lista_precio", val_col="descuentos"
        )
        pivot_marca_desc = _pivot_por_dimension(df, "marca", val_col="descuentos")

        row = 1
        row = _write_section_title(ws_desc, row, "Descuentos por Mes")
        for anio in ANIOS:
            row = _write_year_block_simple(
                ws_desc, row, anio, pivot_mes_desc[anio], with_totals=True
            )

        row = _write_section_title(ws_desc, row, "Descuentos por Lista de Precio")
        for anio in ANIOS:
            row = _write_year_block_with_rows(
                ws_desc,
                row,
                anio,
                pivot_lista_desc[anio],
                "id_lista_precio",
                with_totals=True,
            )

        row = _write_section_title(ws_desc, row, "Descuentos por Marca")
        for anio in ANIOS:
            row = _write_year_block_with_rows(
                ws_desc, row, anio, pivot_marca_desc[anio], "marca", with_totals=True
            )

        hojas.append(ws_desc.title)

        # ── Hoja de Cobertura ──
        cob_data = self.data_loader.get_cobertura_historico_fratelli()

        if any(not df.empty for df in cob_data.values()):
            ws_cob = wb.create_sheet("Cobertura FRATELLI B")
            ws_cob.column_dimensions["A"].width = 18
            for i in range(13):
                ws_cob.column_dimensions[get_column_letter(i + 2)].width = 12

            cob_row = 1

            # Cob Section 1: Total por mes
            _val = "clientes_compradores"
            cob_pivot_mes = (
                _pivot_por_mes(cob_data["total"], val_col=_val)
                if not cob_data["total"].empty
                else {a: [0] * 12 for a in ANIOS}
            )
            cob_row = _write_section_title(ws_cob, cob_row, "Cobertura Total por Mes")
            for anio in ANIOS:
                cob_row = _write_year_block_simple(
                    ws_cob, cob_row, anio, cob_pivot_mes[anio]
                )

            # Cob Section 2: Por lista de precio
            cob_pivot_lista = (
                _pivot_por_dimension(cob_data["lista"], "id_lista_precio", val_col=_val)
                if not cob_data["lista"].empty
                else {a: pd.DataFrame() for a in ANIOS}
            )
            cob_row = _write_section_title(
                ws_cob, cob_row, "Cobertura por Lista de Precio"
            )
            for anio in ANIOS:
                cob_row = _write_year_block_with_rows(
                    ws_cob, cob_row, anio, cob_pivot_lista[anio], "id_lista_precio"
                )

            # Cob Section 3: Por marca
            cob_pivot_marca = (
                _pivot_por_dimension(cob_data["marca"], "marca", val_col=_val)
                if not cob_data["marca"].empty
                else {a: pd.DataFrame() for a in ANIOS}
            )
            cob_row = _write_section_title(ws_cob, cob_row, "Cobertura por Marca")
            for anio in ANIOS:
                cob_row = _write_year_block_with_rows(
                    ws_cob, cob_row, anio, cob_pivot_marca[anio], "marca"
                )

            hojas.append(ws_cob.title)

        # Save file
        out = self._output_dir(None)
        out.mkdir(parents=True, exist_ok=True)
        nombre = config.nombre_archivo or "Historico FRATELLI B"
        ruta = out / f"{nombre}.xlsx"
        wb.save(ruta)

        return HistoricoFratelliResult(
            ruta_archivo=ruta,
            registros_procesados=registros_procesados,
            hojas=hojas,
        )
