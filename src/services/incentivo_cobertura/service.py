"""IncentivoCoberturaService — reporte de cobertura para incentivo ON PREMISE.

Cuenta clientes únicos por vendedor que compraron de cada marca-target
(O-61, LA CELIA, COLON DULCES, FULL SPORT) dentro de la lista de precio
ON PREMISE (id_lista_precio=4) en CASA CENTRAL.

Resultado: Excel con 2 secciones (vendedor y supervisor), filas por
(vendedor/supervisor) y columnas por marca con [Cantidad, Objetivo, %].
"""
import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.services.base_service import BaseService, BaseReporteConfig, BaseReporteResult
from src.services.incentivo_cobertura.constants import (
    ID_LISTA_PRECIO_ON_PREMISE,
    ID_SUCURSAL_CASA_CENTRAL,
    INCENTIVO_TARGETS,
)
from src.services.rebotes.constants import SUPERVISOR_VENDOR_MAP

logger = logging.getLogger(__name__)

# Paleta pastel (consistente con feat/pastel-semaforo-styles).
ROJO = "FFCDD2"
AMARILLO = "FFF59D"
VERDE = "C8E6C9"
ROJO_FONT = "B71C1C"
AMARILLO_FONT = "5D4037"
VERDE_FONT = "1B5E20"

HEADER_FILL = "90A4AE"  # Blue Grey 300
HEADER_FONT = "FFFFFF"
SECTION_FILL = "ECEFF1"  # Blue Grey 50


def _thin_border() -> Border:
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _semaforo_for_pct(pct: float) -> tuple[str, str]:
    """Return (fill, font) for completion %.

    - >= 100%: verde (objetivo cumplido)
    - >= 50%:  amarillo (en camino)
    - <  50%:  rojo (lejos)
    """
    if pct >= 1.0:
        return VERDE, VERDE_FONT
    if pct >= 0.5:
        return AMARILLO, AMARILLO_FONT
    return ROJO, ROJO_FONT


@dataclass
class IncentivoCoberturaConfig(BaseReporteConfig):
    """Config para el reporte de incentivo cobertura ON PREMISE.

    El reporte usa constantes hardcoded (targets, sucursal, lista_precio)
    porque el incentivo es puntual y específico — no se quiere
    parametrizar para evitar uso incorrecto.

    `solo_supervisor`: si se setea, filtra el reporte a los preventistas
    de ese supervisor (key de SUPERVISOR_VENDOR_MAP — ej. "VCHAPUR").
    En ese caso GFARAH = suma de los preventistas del supervisor.

    `vendedores_excluidos`: nombres (case-insensitive) que no participan
    del incentivo. Se filtran del universo y de los totales/objetivos.
    """
    solo_supervisor: str | None = None
    vendedores_excluidos: list[str] | None = None


@dataclass
class IncentivoCoberturaResult(BaseReporteResult):
    """Resultado del reporte."""
    vendedores: int = 0
    supervisores: list[str] = None
    targets: list[str] = None

    def __post_init__(self):
        if self.supervisores is None:
            self.supervisores = []
        if self.targets is None:
            self.targets = []


GFARAH_SUPERVISOR = "GFARAH"  # gerente comercial — total general


def _build_vendor_to_supervisor_map() -> dict[str, str]:
    """Invierte SUPERVISOR_VENDOR_MAP a vendor_upper → supervisor.

    Excluye GFARAH del mapeo porque sus 'vendedores' son códigos de supervisores
    (no preventistas reales). En el reporte GFARAH = total general aplicado aparte.
    """
    result: dict[str, str] = {}
    for sup, vendors in SUPERVISOR_VENDOR_MAP.items():
        if sup == GFARAH_SUPERVISOR:
            continue
        for v in vendors:
            result[v.upper()] = sup
    return result


def _supervisores_ordenados() -> list[str]:
    """Lista de supervisores reales (sin GFARAH) ordenados alfabéticamente."""
    return sorted(s for s in SUPERVISOR_VENDOR_MAP.keys() if s != GFARAH_SUPERVISOR)


class IncentivoCoberturaService(BaseService):
    """Genera el reporte de incentivo cobertura ON PREMISE para CASA CENTRAL."""

    SERVICE_SLUG = "incentivo-cobertura-on-premise"
    GRANULARITY = "month"

    def _fetch_data(self, config: IncentivoCoberturaConfig) -> pd.DataFrame:
        """Devuelve DataFrame raw del query."""
        return self.data_loader.get_incentivo_cobertura_on_premise(
            fecha_desde=config.fecha_desde,
            fecha_hasta=config.fecha_hasta,
            id_sucursal=ID_SUCURSAL_CASA_CENTRAL,
            id_lista_precio=ID_LISTA_PRECIO_ON_PREMISE,
            target_specs=INCENTIVO_TARGETS,
        )

    def _pivot_por_actor(
        self, df_raw: pd.DataFrame, actor_col: str
    ) -> pd.DataFrame:
        """Pivot df_raw a wide: una fila por actor, una col por marca.

        Filas con clientes_compradores = 0 NO aparecen en raw — se rellenan con 0
        post-pivot. Ordena las columnas según el orden de INCENTIVO_TARGETS.
        """
        labels_ordered = [t["label"] for t in INCENTIVO_TARGETS]
        if df_raw.empty:
            return pd.DataFrame(columns=[actor_col] + labels_ordered).set_index(actor_col)
        pivot = (
            df_raw.pivot_table(
                index=actor_col, columns="marca_label",
                values="clientes_compradores", aggfunc="sum", fill_value=0,
            )
        )
        # Asegurar que TODAS las marcas target aparezcan como columnas
        for label in labels_ordered:
            if label not in pivot.columns:
                pivot[label] = 0
        pivot = pivot[labels_ordered]
        return pivot

    def _write_actor_section(
        self,
        ws,
        start_row: int,
        title: str,
        pivot: pd.DataFrame,
        actor_label: str,
        gfarah_row: "pd.Series | None" = None,
        objetivos_por_actor: "dict[str, dict[str, int]] | None" = None,
    ) -> int:
        """Escribe una sección (vendedor o supervisor) y devuelve la próxima fila libre.

        - `gfarah_row`: si viene, se agrega al final con fila vacía como separador.
        - `objetivos_por_actor`: si viene `{actor: {marca: obj_actor}}`, se usa por fila
          en lugar del objetivo individual. Útil para supervisores donde el objetivo
          se multiplica por la cantidad de preventistas que tienen.
        """
        border = _thin_border()
        labels = [t["label"] for t in INCENTIVO_TARGETS]
        objetivos = {t["label"]: t["objetivo"] for t in INCENTIVO_TARGETS}

        # Section title
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = Font(bold=True, size=12, color="263238")
        cell.fill = PatternFill(start_color=SECTION_FILL, end_color=SECTION_FILL, fill_type="solid")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=1 + len(labels) * 3)
        start_row += 1

        # Header row 1: actor + grupos de marca (merged)
        ws.cell(row=start_row, column=1, value=actor_label).font = Font(bold=True, color=HEADER_FONT)
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        ws.cell(row=start_row, column=1).border = border
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)
        col = 2
        for label in labels:
            for offset in range(3):
                c = ws.cell(row=start_row, column=col + offset)
                c.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
                c.border = border
            c = ws.cell(row=start_row, column=col, value=label)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 2)
            col += 3

        # Header row 2: Cant | Obj | %
        col = 2
        for _ in labels:
            for sub in ("Cant.", "Obj.", "%"):
                c = ws.cell(row=start_row + 1, column=col, value=sub)
                c.font = Font(bold=True, color=HEADER_FONT)
                c.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
                c.border = border
                c.alignment = Alignment(horizontal="center")
                col += 1
        current_row = start_row + 2

        def _write_data_row(actor: str, row_data, bold_actor: bool = False) -> None:
            """Helper: escribe 1 fila completa con cant/obj/% por marca."""
            nonlocal current_row
            c_actor = ws.cell(row=current_row, column=1, value=actor)
            c_actor.border = border
            c_actor.font = Font(bold=True if not bold_actor else True, size=11 if bold_actor else 11)
            col = 2
            for label in labels:
                cant = int(row_data.get(label, 0))
                if objetivos_por_actor is not None and actor in objetivos_por_actor:
                    obj = objetivos_por_actor[actor].get(label, objetivos[label])
                else:
                    obj = objetivos[label]
                pct = (cant / obj) if obj > 0 else 0.0

                c_cant = ws.cell(row=current_row, column=col, value=cant)
                c_cant.number_format = "#,##0"
                c_cant.border = border
                c_cant.alignment = Alignment(horizontal="center")

                c_obj = ws.cell(row=current_row, column=col + 1, value=obj)
                c_obj.number_format = "#,##0"
                c_obj.border = border
                c_obj.alignment = Alignment(horizontal="center")
                c_obj.font = Font(italic=True, color="78909C")

                c_pct = ws.cell(row=current_row, column=col + 2, value=pct)
                c_pct.number_format = "0%"
                c_pct.border = border
                c_pct.alignment = Alignment(horizontal="center")
                fill_color, font_color = _semaforo_for_pct(pct)
                c_pct.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c_pct.font = Font(color=font_color, bold=True)
                col += 3
            current_row += 1

        # Body — supervisores/vendedores
        for actor, row_pivot in pivot.iterrows():
            _write_data_row(actor, row_pivot)

        # GFARAH (total general) — fila vacía como separador + fila de GFARAH
        if gfarah_row is not None:
            current_row += 1  # separador visual
            _write_data_row(GFARAH_SUPERVISOR, gfarah_row, bold_actor=True)

        return current_row + 1  # 1-row spacer

    def _build_workbook(
        self, df_raw: pd.DataFrame, output_path: Path, fecha_desde: str, fecha_hasta: str
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Incentivo Cobertura ON PREMISE")

        # Anchos
        ws.column_dimensions["A"].width = 28
        for i, _ in enumerate([t for t in INCENTIVO_TARGETS]):
            base = 2 + i * 3
            ws.column_dimensions[get_column_letter(base)].width = 8
            ws.column_dimensions[get_column_letter(base + 1)].width = 6
            ws.column_dimensions[get_column_letter(base + 2)].width = 7

        # Header global
        scope_text = f" — Supervisor {self._solo_supervisor}" if getattr(self, "_solo_supervisor", None) else ""
        ws.cell(row=1, column=1, value=f"INCENTIVO COBERTURA ON PREMISE — Casa Central{scope_text}").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Período: {fecha_desde} a {fecha_hasta}").font = Font(italic=True, color="546E7A")

        # Map vendor → supervisor
        vendor_to_sup = _build_vendor_to_supervisor_map()

        df_with_sup = df_raw.copy()
        if df_with_sup.empty:
            df_with_sup["Supervisor"] = []
        else:
            df_with_sup["vendedor_upper"] = df_with_sup["vendedor"].fillna("").str.upper()
            df_with_sup["Supervisor"] = df_with_sup["vendedor_upper"].map(vendor_to_sup).fillna("Sin Supervisor")
            df_with_sup = df_with_sup.drop(columns=["vendedor_upper"])

        # Filtro por supervisor: acotamos raw a los preventistas del supervisor
        solo_sup = getattr(self, "_solo_supervisor", None)
        if solo_sup and not df_with_sup.empty:
            sup_vendors_upper = {v.upper() for v in SUPERVISOR_VENDOR_MAP.get(solo_sup, [])}
            df_with_sup = df_with_sup[df_with_sup["vendedor"].fillna("").str.upper().isin(sup_vendors_upper)]

        # Filtro de vendedores excluidos (no participan del incentivo)
        excluidos_upper = getattr(self, "_vendedores_excluidos", set())
        if excluidos_upper and not df_with_sup.empty:
            df_with_sup = df_with_sup[~df_with_sup["vendedor"].fillna("").str.upper().isin(excluidos_upper)]

        # Pivot vendedor — reindex con el universo completo (rellena 0 los que faltan).
        # df_raw incluye DIRECTA; el universo excluye DIRECTA, así que cae fuera
        # de la tabla vendedor — lo cual es correcto (DIRECTA no es preventista,
        # solo suma a GFARAH).
        pivot_v = self._pivot_por_actor(df_with_sup, "vendedor")
        if hasattr(self, "_universo_vendedores") and self._universo_vendedores:
            pivot_v = pivot_v.reindex(self._universo_vendedores, fill_value=0)
        pivot_v = pivot_v.sort_index()

        # Pivot supervisor:
        # - Supervisores reales (sin GFARAH, sin "Sin Supervisor"): sumar solo
        #   los vendedores mapeados.
        # - GFARAH = total general → sumar TODO el raw (incluye DIRECTA,
        #   incluye preventistas sin mapping, todo).
        # Render order: supervisores reales (alfa), blank row, GFARAH.
        labels = [t["label"] for t in INCENTIVO_TARGETS]
        solo_sup = getattr(self, "_solo_supervisor", None)
        # Si filtro por supervisor, solo lo mostramos a él
        supervisores_reales = [solo_sup] if solo_sup else _supervisores_ordenados()

        if df_with_sup.empty:
            pivot_s_reales = pd.DataFrame(0, index=supervisores_reales, columns=labels)
            gfarah_row = pd.Series(0, index=labels, name=GFARAH_SUPERVISOR)
        else:
            # Filtramos: solo rows con supervisor mapeado real (no "Sin Supervisor")
            mask_real = df_with_sup["Supervisor"].isin(supervisores_reales)
            agg_sup = (
                df_with_sup.loc[mask_real]
                .groupby(["Supervisor", "marca_label"], as_index=False)["clientes_compradores"]
                .sum()
                .rename(columns={"Supervisor": "supervisor"})
            )
            pivot_s_reales = self._pivot_por_actor(agg_sup, "supervisor")
            pivot_s_reales = pivot_s_reales.reindex(supervisores_reales, fill_value=0)

            # GFARAH = todo el raw agregado (sin filtrar por supervisor)
            gfarah_totals = (
                df_with_sup.groupby("marca_label")["clientes_compradores"].sum()
            )
            gfarah_row = pd.Series(
                {label: int(gfarah_totals.get(label, 0)) for label in labels},
                name=GFARAH_SUPERVISOR,
            )

        # Objetivos por supervisor:
        # cada supervisor representa N preventistas → su objetivo es N × obj_individual.
        # Solo cuentan preventistas que existen en el universo (clientes ON PREMISE).
        objetivos_individuales = {t["label"]: t["objetivo"] for t in INCENTIVO_TARGETS}
        universo_set = set(self._universo_vendedores) if hasattr(self, "_universo_vendedores") else set()
        objetivos_por_actor: dict[str, dict[str, int]] = {}
        for sup in supervisores_reales:
            vendedores_mapeados = SUPERVISOR_VENDOR_MAP.get(sup, [])
            n_prev = sum(1 for v in vendedores_mapeados if v in universo_set)
            objetivos_por_actor[sup] = {
                label: n_prev * obj for label, obj in objetivos_individuales.items()
            }
        # GFARAH: total preventistas del universo (todos)
        objetivos_por_actor[GFARAH_SUPERVISOR] = {
            label: len(universo_set) * obj for label, obj in objetivos_individuales.items()
        }

        # Render. Cuando filtramos por un supervisor puntual, la sección
        # "Por Supervisor" (y GFARAH) es redundante con la tabla de vendedores
        # → se omite y queda solo la lista de preventistas.
        next_row = 4
        next_row = self._write_actor_section(ws, next_row, "Por Vendedor", pivot_v, "Vendedor")
        if not solo_sup:
            next_row = self._write_actor_section(
                ws, next_row, "Por Supervisor (totales)", pivot_s_reales, "Supervisor",
                gfarah_row=gfarah_row,
                objetivos_por_actor=objetivos_por_actor,
            )

        wb.save(output_path)

    def generar_reporte(self, config: IncentivoCoberturaConfig) -> IncentivoCoberturaResult:
        df_raw = self._fetch_data(config)

        if df_raw.empty:
            logger.warning("No se encontraron datos para el rango %s-%s", config.fecha_desde, config.fecha_hasta)

        # Universo de vendedores ON PREMISE en Casa Central (para mostrar también
        # los que tienen 0 cobertura en todas las marcas — sino quedan ocultos).
        df_universo_v = self.data_loader.get_vendedores_on_premise_universo(
            id_sucursal=ID_SUCURSAL_CASA_CENTRAL,
            id_lista_precio=ID_LISTA_PRECIO_ON_PREMISE,
        )
        universo_vendedores = sorted(df_universo_v["vendedor"].dropna().unique().tolist())

        # Filtro opcional por supervisor: acotamos universo a sus preventistas mapeados
        if config.solo_supervisor:
            sup_vendors = set(SUPERVISOR_VENDOR_MAP.get(config.solo_supervisor, []))
            universo_vendedores = [v for v in universo_vendedores if v in sup_vendors]
            # En este modo el dataset raw también debe filtrarse — se hace abajo
            # al pivotar (df_with_sup queda acotado)

        # Filtro opcional de vendedores que no participan del incentivo
        excluidos_upper = {v.upper() for v in (config.vendedores_excluidos or [])}
        if excluidos_upper:
            universo_vendedores = [v for v in universo_vendedores if v.upper() not in excluidos_upper]

        # Supervisor map (todos los supervisores del SUPERVISOR_VENDOR_MAP aparecen
        # — el usuario quiere ver todos, no solo los que tuvieron actividad).
        vendor_to_sup = _build_vendor_to_supervisor_map()
        universo_supervisores = sorted(SUPERVISOR_VENDOR_MAP.keys())
        # No incluimos "Sin Supervisor" salvo que haya vendedores en el universo
        # sin mapping (típicamente DIRECTA queda fuera del universo, así que es raro).
        if any(vendor_to_sup.get(v.upper()) is None for v in universo_vendedores):
            universo_supervisores = sorted(set(universo_supervisores) | {"Sin Supervisor"})
        # Pass para _build_workbook
        self._universo_vendedores = universo_vendedores
        self._universo_supervisores = universo_supervisores
        self._solo_supervisor = config.solo_supervisor
        self._vendedores_excluidos = excluidos_upper

        nombre = config.nombre_archivo or f"Incentivo Cobertura ON PREMISE - {config.fecha_desde} al {config.fecha_hasta}"
        out_dir = self._output_dir(config.fecha_desde)
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(df_raw, ruta, config.fecha_desde, config.fecha_hasta)

        return IncentivoCoberturaResult(
            ruta_archivo=ruta,
            registros_procesados=len(df_raw),
            vendedores=len(universo_vendedores),
            supervisores=universo_supervisores,
            targets=[t["label"] for t in INCENTIVO_TARGETS],
        )

    def run(self, config: IncentivoCoberturaConfig) -> IncentivoCoberturaResult:
        return self.generar_reporte(config)
