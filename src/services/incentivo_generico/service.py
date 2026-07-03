"""IncentivoGenericoService — incentivo de cobertura por generico.

Cuenta clientes compradores por vendedor (y por supervisor) de un unico
generico, leyendo la cobertura pre-calculada de ``gold.cob_preventista_generico``.
Pensado para genericos donde el objetivo aun no esta definido: cuando
``objetivo`` es None, se ocultan las columnas Obj./%; cuando se setea, se
muestran con semaforo.

Estructura del Excel (una hoja):
  - Seccion "Por Vendedor": fila por vendedor, columna del generico.
  - Seccion "Por Supervisor": fila por supervisor, agregando sus vendedores.
"""
import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService
from src.services.rebotes.constants import SUPERVISOR_VENDOR_MAP

logger = logging.getLogger(__name__)

# Gerente comercial — sus "vendedores" en el mapa son códigos de supervisores,
# no preventistas reales; se excluye al invertir el mapa.
GERENTE_KEY = "GFARAH"

# Vendedores que no participan del incentivo (no son preventistas reales).
VENDEDORES_EXCLUIDOS = {"DIRECTA"}


def _vendor_to_supervisor() -> dict[str, str]:
    """Invierte el mapeo curado SUPERVISOR_VENDOR_MAP → {vendedor_upper: supervisor}.

    Es la fuente de verdad del match vendedor↔supervisor (dim_vendedor.supervisor
    no es confiable). Excluye al gerente (GFARAH).
    """
    result: dict[str, str] = {}
    for sup, vendors in SUPERVISOR_VENDOR_MAP.items():
        if sup == GERENTE_KEY:
            continue
        for v in vendors:
            result[v.upper()] = sup
    return result

# Paleta (consistente con incentivo ON PREMISE).
ROJO, AMARILLO, VERDE = "FFCDD2", "FFF59D", "C8E6C9"
ROJO_FONT, AMARILLO_FONT, VERDE_FONT = "B71C1C", "5D4037", "1B5E20"
HEADER_FILL, HEADER_FONT = "90A4AE", "FFFFFF"
SECTION_FILL = "ECEFF1"
TOTAL_FILL = "FFE08A"  # ámbar — fila TOTAL GENERAL

# CASA CENTRAL / fuerza de ventas con todos los preventistas.
ID_SUCURSAL_CASA_CENTRAL = 1
ID_FUERZA_VENTAS_TODOS = 1


@dataclass
class IncentivoGenericoConfig:
    """Config del incentivo de cobertura por generico.

    Args:
        generico: Nombre exacto del generico (ej. 'PERNOD RICARD').
        fecha: Dia desde (YYYY-MM-DD). La cobertura se calcula al grano diario.
        fecha_hasta: Dia hasta (YYYY-MM-DD). None → mismo dia que `fecha`.
        objetivo: Clientes unicos objetivo por vendedor. None → sin objetivo
            (se ocultan Obj./% y se muestra solo la cantidad).
        id_sucursal / id_fuerza_ventas: universo (default CASA CENTRAL / todos).
        nombre_archivo: nombre de salida (sin extension).
    """
    generico: str
    fecha: str
    fecha_hasta: str | None = None
    objetivo: int | None = None
    id_sucursal: int = ID_SUCURSAL_CASA_CENTRAL
    id_fuerza_ventas: int = ID_FUERZA_VENTAS_TODOS
    nombre_archivo: str | None = None


@dataclass
class IncentivoGenericoResult:
    """Resultado del reporte."""
    ruta_archivo: Path
    registros_procesados: int
    vendedores: int
    supervisores: list[str]
    fecha_desde: str
    fecha_hasta: str


def _thin_border() -> Border:
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _semaforo_for_pct(pct: float) -> tuple[str, str]:
    if pct >= 1.0:
        return VERDE, VERDE_FONT
    if pct >= 0.5:
        return AMARILLO, AMARILLO_FONT
    return ROJO, ROJO_FONT


class IncentivoGenericoService(BaseService):
    """Genera el incentivo de cobertura por generico para una sucursal."""

    SERVICE_SLUG = "incentivo-generico"
    GRANULARITY = "month"

    def generar_reporte(self, config: IncentivoGenericoConfig) -> IncentivoGenericoResult:
        fecha_desde = config.fecha
        fecha_hasta = config.fecha_hasta or config.fecha

        df = self.data_loader.get_cobertura_generico_por_vendedor(
            generico=config.generico,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            id_sucursal=config.id_sucursal,
            id_fuerza_ventas=config.id_fuerza_ventas,
        )
        if df.empty:
            logger.warning(
                "Sin cobertura para generico=%s fechas=%s..%s suc=%s fuerza=%s",
                config.generico, fecha_desde, fecha_hasta, config.id_sucursal, config.id_fuerza_ventas,
            )
            df = pd.DataFrame(columns=["vendedor", "clientes"])

        df = df.copy()
        df["vendedor"] = df["vendedor"].fillna("(sin vendedor)")
        df["clientes"] = df["clientes"].fillna(0).astype(int)

        # Excluir vendedores que no participan (ej. DIRECTA).
        df = df[~df["vendedor"].str.upper().isin(VENDEDORES_EXCLUIDOS)]

        # Match vendedor↔supervisor via el mapeo curado (no dim_vendedor).
        vendor_to_sup = _vendor_to_supervisor()
        df["supervisor"] = df["vendedor"].str.upper().map(vendor_to_sup).fillna("Sin Supervisor")

        por_vendedor = df.sort_values(
            ["supervisor", "clientes", "vendedor"], ascending=[True, False, True]
        )[["vendedor", "supervisor", "clientes"]]

        por_supervisor = (
            df.groupby("supervisor", as_index=False)["clientes"].sum()
            .sort_values(["clientes", "supervisor"], ascending=[False, True])
        )

        nombre = config.nombre_archivo or f"Incentivo Cobertura {config.generico} - {config.fecha}"
        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(config, fecha_desde, fecha_hasta, por_vendedor, por_supervisor, ruta)

        return IncentivoGenericoResult(
            ruta_archivo=ruta,
            registros_procesados=len(df),
            vendedores=len(por_vendedor),
            supervisores=sorted(df["supervisor"].unique().tolist()),
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )

    def _build_workbook(
        self,
        config: IncentivoGenericoConfig,
        fecha_desde: str,
        fecha_hasta: str,
        por_vendedor: pd.DataFrame,
        por_supervisor: pd.DataFrame,
        ruta: Path,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Incentivo Cobertura")
        border = _thin_border()
        con_objetivo = config.objetivo is not None
        # Columnas: Actor | Clientes [| Obj. | %]
        n_cols = 4 if con_objetivo else 2

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        if con_objetivo:
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 8

        ws.cell(row=1, column=1,
                value=f"INCENTIVO COBERTURA — {config.generico} — Casa Central").font = Font(bold=True, size=14)
        fecha_txt = fecha_desde if fecha_desde == fecha_hasta else f"{fecha_desde} a {fecha_hasta}"
        ws.cell(row=2, column=1,
                value=f"Fecha: {fecha_txt}").font = Font(italic=True, color="546E7A")
        if not con_objetivo:
            ws.cell(row=3, column=1,
                    value="Objetivo aún no definido — se muestra cobertura (clientes) por preventista.").font = (
                Font(italic=True, size=9, color="90A4AE"))

        next_row = 5
        next_row = self._write_section(ws, next_row, "Por Vendedor", "Vendedor",
                                       por_vendedor, "vendedor", config, border, n_cols)
        next_row += 1
        self._write_section(ws, next_row, "Por Supervisor (totales)", "Supervisor",
                            por_supervisor, "supervisor", config, border, n_cols)

        wb.save(ruta)

    def _write_section(
        self, ws, start_row: int, title: str, actor_label: str,
        data: pd.DataFrame, actor_col: str, config: IncentivoGenericoConfig,
        border: Border, n_cols: int,
    ) -> int:
        con_objetivo = config.objetivo is not None

        # Section title
        c = ws.cell(row=start_row, column=1, value=title)
        c.font = Font(bold=True, size=12, color="263238")
        c.fill = PatternFill(start_color=SECTION_FILL, end_color=SECTION_FILL, fill_type="solid")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
        start_row += 1

        # Header row. The count column is "Clientes" (the generico is already in
        # the title/section); keeps the header from overflowing its column.
        headers = [actor_label, "Clientes"]
        if con_objetivo:
            headers += ["Obj.", "%"]
        for j, h in enumerate(headers, 1):
            hc = ws.cell(row=start_row, column=j, value=h)
            hc.font = Font(bold=True, color=HEADER_FONT)
            hc.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
            hc.border = border
            hc.alignment = Alignment(horizontal="center", vertical="center")
        row = start_row + 1

        for _, r in data.iterrows():
            actor = r[actor_col]
            clientes = int(r["clientes"])
            ca = ws.cell(row=row, column=1, value=actor)
            ca.border = border
            ca.font = Font(bold=True)
            cc = ws.cell(row=row, column=2, value=clientes)
            cc.number_format = "#,##0"
            cc.border = border
            cc.alignment = Alignment(horizontal="center")
            if con_objetivo:
                obj = config.objetivo
                pct = (clientes / obj) if obj else 0.0
                co = ws.cell(row=row, column=3, value=obj)
                co.number_format = "#,##0"
                co.border = border
                co.alignment = Alignment(horizontal="center")
                co.font = Font(italic=True, color="78909C")
                cp = ws.cell(row=row, column=4, value=pct)
                cp.number_format = "0%"
                cp.border = border
                cp.alignment = Alignment(horizontal="center")
                fill_c, font_c = _semaforo_for_pct(pct)
                cp.fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type="solid")
                cp.font = Font(color=font_c, bold=True)
            row += 1

        # TOTAL GENERAL — suma de la sección.
        total_clientes = int(data["clientes"].sum()) if not data.empty else 0
        tot_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")
        ct = ws.cell(row=row, column=1, value="TOTAL GENERAL")
        ct.font = Font(bold=True)
        ct.fill = tot_fill
        ct.border = border
        cv = ws.cell(row=row, column=2, value=total_clientes)
        cv.number_format = "#,##0"
        cv.font = Font(bold=True)
        cv.fill = tot_fill
        cv.border = border
        cv.alignment = Alignment(horizontal="center")
        if con_objetivo:
            obj_total = config.objetivo * len(data)
            pct_total = (total_clientes / obj_total) if obj_total else 0.0
            cot = ws.cell(row=row, column=3, value=obj_total)
            cot.number_format = "#,##0"
            cot.font = Font(bold=True)
            cot.fill = tot_fill
            cot.border = border
            cot.alignment = Alignment(horizontal="center")
            cpt = ws.cell(row=row, column=4, value=pct_total)
            cpt.number_format = "0%"
            cpt.font = Font(bold=True)
            cpt.fill = tot_fill
            cpt.border = border
            cpt.alignment = Alignment(horizontal="center")
        row += 1

        return row

    def run(self, config: IncentivoGenericoConfig) -> IncentivoGenericoResult:
        return self.generar_reporte(config)
