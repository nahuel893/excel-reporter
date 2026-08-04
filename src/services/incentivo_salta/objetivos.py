"""Lectura del archivo de objetivos del incentivo preventa SALTA.

El xlsx es la UNICA fuente de verdad del incentivo: define los bloques (grupo,
sabor, calibre, mes) y el cupo fijo de cada preventista. Se lee, nunca se
recalcula — un cupo que se recalcula todos los dias deja de ser un objetivo.

Layout esperado (`configs/objetivos_incentivo_salta.xlsx`):

    fila 4    grupo      INCENTIVO AGOSTO            INCENTIVO SEPTIEMBRE
    fila 5    sabor      SALTA NEGRA   SALTA RUBIA   ...
    fila 6    calibre    1000 cc       1200 cc       ...
    fila 7    medidas    Cupo | <fecha> | %          ...
    fila 8+   datos      PREVENTISTA | cupo | ...
    ultima    TOTAL ...

Los bloques se descubren por las celdas "Cupo" de la fila de medidas, y el mes
sale de la fecha que esta a su derecha. Nada de eso se hardcodea: agregar un
bloque al xlsx alcanza para que el informe lo tome.
"""
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

# "SALTA RUBIA" es como lo llama el negocio; en dim_articulo el sabor derivado
# es "BLANCA (rubia)". El mapeo vive aca y no en el servicio.
_SABORES = {"NEGRA": "NEGRA", "RUBIA": "BLANCA (rubia)", "BLANCA": "BLANCA (rubia)"}


@dataclass(frozen=True)
class BloqueIncentivo:
    """Un bloque del incentivo: que se mide, en que mes, y con que cupo."""
    grupo: str
    sabor: str
    calibre: str
    mes: str                    # 'YYYY-MM'
    cupos: dict[str, float]     # preventista -> cupo fijo

    @property
    def cupo_total(self) -> float:
        return float(sum(self.cupos.values()))


def _hacia_izquierda(ws, fila: int, col: int):
    """Valor de una celda combinada: se busca a la izquierda hasta encontrarlo."""
    for c in range(col, 0, -1):
        v = ws.cell(fila, c).value
        if v is not None:
            return v
    return None


def _sabor(texto: str) -> str:
    for clave, valor in _SABORES.items():
        if clave in str(texto).upper():
            return valor
    raise ValueError(f"No se reconoce el sabor en '{texto}'")


def _calibre(texto: str) -> str:
    """'1200 cc' -> '1200'. El calibre viaja como texto, igual que en el fact."""
    return str(texto).lower().replace("cc", "").strip()


def leer_objetivos(ruta: str | Path) -> list[BloqueIncentivo]:
    """Devuelve los bloques del incentivo con sus cupos por preventista.

    Raises:
        ValueError: si el archivo no tiene la fila de medidas con "Cupo", si un
            bloque no trae la fecha del mes, o si un bloque queda sin cupos. Se
            falla ruidosamente a proposito: un incentivo con cupos en cero se
            veria como que nadie llego al objetivo.
    """
    ws = load_workbook(ruta, data_only=True).active

    fila_medidas = next(
        (r for r in range(1, min(ws.max_row, 30) + 1)
         if any(str(ws.cell(r, c).value).strip().lower() == "cupo"
                for c in range(1, ws.max_column + 1))),
        None,
    )
    if fila_medidas is None:
        raise ValueError(f"{ruta}: no se encontro la fila de medidas con 'Cupo'")

    cols_cupo = [c for c in range(1, ws.max_column + 1)
                 if str(ws.cell(fila_medidas, c).value).strip().lower() == "cupo"]

    fila_ini = fila_medidas + 1
    fila_fin = next(
        (r for r in range(fila_ini, ws.max_row + 1)
         if str(ws.cell(r, 1).value or "").upper().startswith("TOTAL")),
        ws.max_row + 1,
    )

    bloques: list[BloqueIncentivo] = []
    for col in cols_cupo:
        fecha = ws.cell(fila_medidas, col + 1).value
        if not hasattr(fecha, "strftime"):
            raise ValueError(
                f"{ruta}: el bloque de la columna {col} no tiene fecha de mes "
                f"a su derecha (encontrado: {fecha!r})"
            )
        cupos = {
            str(ws.cell(r, 1).value).strip(): float(ws.cell(r, col).value)
            for r in range(fila_ini, fila_fin)
            if ws.cell(r, 1).value and ws.cell(r, col).value is not None
        }
        if not cupos:
            raise ValueError(f"{ruta}: el bloque de la columna {col} no tiene cupos")
        bloques.append(BloqueIncentivo(
            grupo=str(_hacia_izquierda(ws, fila_medidas - 3, col) or "").strip(),
            sabor=_sabor(_hacia_izquierda(ws, fila_medidas - 2, col)),
            calibre=_calibre(_hacia_izquierda(ws, fila_medidas - 1, col)),
            mes=fecha.strftime("%Y-%m"),
            cupos=cupos,
        ))
    return bloques
