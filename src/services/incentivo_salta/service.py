"""IncentivoSaltaService — cuadro de avance del incentivo preventa SALTA.

Una fila por preventista de CASA CENTRAL y, por cada bloque del incentivo,
tres columnas: el cupo fijo, la cobertura del mes de ese bloque, y el % de
avance con semaforo.

Dos decisiones que no son obvias:

- **Los cupos se LEEN** de `configs/objetivos_incentivo_salta.xlsx`, que ademas
  define los bloques (grupo, sabor, calibre, mes). Nunca se recalculan: un cupo
  que se mueve todos los dias deja de ser un objetivo. Para cambiar el
  incentivo se edita ese archivo, no el codigo.
- **Cada bloque mide SU propio mes.** El de agosto carga agosto, el de
  septiembre carga septiembre. Un bloque cuyo mes todavia no empezo queda en
  blanco — no en cero, que se leeria como fracaso y pintaria el semaforo rojo
  antes de tiempo.

La cobertura cuenta el cliente con neto POSITIVO en el corte (`> 0`). Hasta el
2026-08-19 usaba umbral 0.5 bultos; el negocio cambio el criterio.
"""
import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService
from src.services.incentivo_salta.objetivos import BloqueIncentivo, leer_objetivos
from src.services.incentivo_salta.processor import (
    contar_cobertura,
    cobertura_total,
    mes_ya_empezo,
    ventana_del_mes,
)

logger = logging.getLogger(__name__)

MARCA = "SALTA"
ID_SUCURSAL_CASA_CENTRAL = 1

HEADER_FILL, HEADER_FONT = "1F4E78", "FFFFFF"
TOTAL_FILL = "FFE08A"
CUPO_FILL = "FFF2CC"
_SHEET_TITLE = "Incentivo Preventa"
# Umbrales FIJOS del semaforo: verde al 100%, amarillo desde 50%, rojo debajo.
# `num` y no percentiles — el criterio es del negocio, no relativo al rango.
_SEMAFORO = IconSetRule("3TrafficLights1", "num", [0, 0.5, 1], showValue=True)

_MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


@dataclass
class IncentivoSaltaConfig:
    """Config del informe.

    Args:
        fecha_hasta: ultimo dia con datos. Recorta el mes en curso y decide que
            bloques ya arrancaron.
        objetivos_path: xlsx con los bloques y los cupos fijos.
        id_sucursal: sucursal del incentivo.
        excluir_vendedores: preventistas que no participan (DIRECTA,
            SUB DISTRIBUIDOR, etc). Se filtran del origen para que la suma de
            filas siga cerrando contra el total.
        nombre_archivo: nombre de salida sin extension.
    """
    fecha_hasta: str
    objetivos_path: str
    id_sucursal: int = ID_SUCURSAL_CASA_CENTRAL
    excluir_vendedores: list[str] = field(default_factory=list)
    nombre_archivo: str | None = None


@dataclass
class IncentivoSaltaResult:
    ruta_archivo: Path
    bloques: int
    preventistas: int
    bloques_activos: list[str]
    fecha_hasta: str


def _etiqueta_mes(mes: str) -> str:
    anio, m = (int(x) for x in mes.split("-"))
    return f"{_MESES[m - 1]}-{anio}"


def _thin() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class IncentivoSaltaService(BaseService):
    SERVICE_SLUG = "incentivo-salta"
    GRANULARITY = "month"

    def generar_reporte(self, config: IncentivoSaltaConfig) -> IncentivoSaltaResult:
        bloques = leer_objetivos(config.objetivos_path)
        excluidos = {v.strip().upper() for v in config.excluir_vendedores}

        # Un solo query por mes distinto: varios bloques comparten mes.
        coberturas: dict[str, dict[str, int]] = {}
        totales: dict[str, int] = {}
        activos: list[str] = []
        for b in bloques:
            clave = f"{b.mes}|{b.sabor}|{b.calibre}"
            if not mes_ya_empezo(b.mes, config.fecha_hasta):
                continue
            activos.append(clave)
            desde, hasta = ventana_del_mes(b.mes, config.fecha_hasta)
            df = self._traer(config, desde, hasta)
            coberturas[clave] = contar_cobertura(df, b.sabor, b.calibre)
            totales[clave] = cobertura_total(df, b.sabor, b.calibre)
            if excluidos:
                # El total se recalcula sin los excluidos para que cierre contra
                # la suma de las filas visibles.
                totales[clave] = sum(
                    v for k, v in coberturas[clave].items()
                    if str(k).strip().upper() not in excluidos
                )

        preventistas = self._preventistas(bloques, excluidos)
        nombre = config.nombre_archivo or "Incentivo Preventa SALTA"
        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha_hasta, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(config, bloques, preventistas, coberturas, totales, ruta)

        return IncentivoSaltaResult(
            ruta_archivo=ruta,
            bloques=len(bloques),
            preventistas=len(preventistas),
            bloques_activos=activos,
            fecha_hasta=config.fecha_hasta,
        )

    def _traer(self, config: IncentivoSaltaConfig, desde: str, hasta: str) -> pd.DataFrame:
        df = self.data_loader.get_ventas_cliente_sabor_mes(
            marca=MARCA, fecha_desde=desde, fecha_hasta=hasta
        )
        if df is None or df.empty:
            logger.warning("Sin ventas de %s entre %s y %s", MARCA, desde, hasta)
            return pd.DataFrame(columns=["id_cliente", "id_sucursal", "preventista",
                                         "sabor", "calibre", "bultos"])
        return df[df["id_sucursal"] == config.id_sucursal]

    @staticmethod
    def _preventistas(bloques: list[BloqueIncentivo], excluidos: set[str]) -> list[str]:
        """Filas del cuadro: los preventistas del archivo de objetivos.

        El archivo manda. Un preventista que vendio pero no tiene cupo no entra
        al incentivo, y uno con cupo que no vendio tiene que aparecer en cero.
        """
        vistos: dict[str, None] = {}
        for b in bloques:
            for v in b.cupos:
                if v.strip().upper() not in excluidos:
                    vistos.setdefault(v, None)
        return sorted(vistos)

    def _build_workbook(self, config, bloques, preventistas, coberturas, totales, ruta):
        wb = Workbook()
        ws = wb.active
        ws.title = _SHEET_TITLE
        borde = _thin()
        relleno = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
        fuente = Font(bold=True, color=HEADER_FONT)
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws["A1"] = "Cobertura SALTA por ruta — incentivo preventa — CASA CENTRAL"
        ws["A1"].font = Font(bold=True, size=13)

        F_GRUPO, F_SABOR, F_CAL, F_MED = 3, 4, 5, 6
        ANCHO = 3                                   # Cupo | cobertura | % Avance
        ultima_col = 1 + ANCHO * len(bloques)

        ws.merge_cells(start_row=F_GRUPO, start_column=1, end_row=F_MED, end_column=1)
        c = ws.cell(F_GRUPO, 1, "Vendedor")
        c.fill, c.font, c.alignment, c.border = relleno, fuente, centro, borde

        # Encabezado: los bloques consecutivos con el mismo grupo se fusionan.
        tramos: list[list] = []
        for i, b in enumerate(bloques):
            col = 2 + ANCHO * i
            if tramos and tramos[-1][0] == b.grupo:
                tramos[-1][2] = col + ANCHO - 1
            else:
                tramos.append([b.grupo, col, col + ANCHO - 1])

            etiqueta = "SALTA RUBIA" if "rubia" in b.sabor.lower() else f"SALTA {b.sabor}"
            for fila, texto in ((F_SABOR, etiqueta), (F_CAL, f"{b.calibre} cc")):
                ws.merge_cells(start_row=fila, start_column=col,
                               end_row=fila, end_column=col + ANCHO - 1)
                cc = ws.cell(fila, col, texto)
                cc.fill, cc.font, cc.alignment = relleno, fuente, centro
            for j, texto in enumerate(("Cupo", _etiqueta_mes(b.mes), "% Avance")):
                cc = ws.cell(F_MED, col + j, texto)
                cc.fill, cc.font, cc.alignment, cc.border = relleno, fuente, centro, borde

        for grupo, desde, hasta in tramos:
            ws.merge_cells(start_row=F_GRUPO, start_column=desde,
                           end_row=F_GRUPO, end_column=hasta)
            cc = ws.cell(F_GRUPO, desde, grupo)
            cc.fill, cc.font, cc.alignment = relleno, fuente, centro
        for fila in (F_GRUPO, F_SABOR, F_CAL):
            for j in range(2, ultima_col + 1):
                ws.cell(fila, j).border = borde

        fila_ini = F_MED + 1
        for r, vendedor in enumerate(preventistas, fila_ini):
            cv = ws.cell(r, 1, vendedor)
            cv.border = borde
            for i, b in enumerate(bloques):
                col = 2 + ANCHO * i
                clave = f"{b.mes}|{b.sabor}|{b.calibre}"
                cupo = b.cupos.get(vendedor)
                cc = ws.cell(r, col, cupo)
                cc.number_format = "#,##0"; cc.border = borde
                cc.fill = PatternFill("solid", start_color=CUPO_FILL, end_color=CUPO_FILL)
                # Bloque cuyo mes no empezo: se deja en blanco a proposito.
                valor = coberturas.get(clave, {}).get(vendedor, 0) if clave in coberturas else None
                cb = ws.cell(r, col + 1, valor)
                cb.number_format = "#,##0"; cb.border = borde
                cp = ws.cell(r, col + 2)
                cp.border = borde
                if clave in coberturas:
                    cp.value = (f'=IF({get_column_letter(col)}{r}=0,"",'
                                f'{get_column_letter(col + 1)}{r}/{get_column_letter(col)}{r})')
                    cp.number_format = "0.0%"
                    cp.alignment = Alignment(horizontal="right")

        fila_fin = fila_ini + len(preventistas) - 1
        r_total = fila_fin + 1
        r_avance = r_total + 1
        total_fill = PatternFill("solid", start_color=TOTAL_FILL, end_color=TOTAL_FILL)
        for etiqueta, fila in (("TOTAL GENERAL", r_total), ("% AVANCE", r_avance)):
            ws.cell(fila, 1, etiqueta).font = Font(bold=True)
            for j in range(1, ultima_col + 1):
                ws.cell(fila, j).fill = total_fill
                ws.cell(fila, j).border = borde
                ws.cell(fila, j).font = Font(bold=True)

        for i, b in enumerate(bloques):
            col = 2 + ANCHO * i
            clave = f"{b.mes}|{b.sabor}|{b.calibre}"
            L = get_column_letter
            ct = ws.cell(r_total, col, f"=SUM({L(col)}{fila_ini}:{L(col)}{fila_fin})")
            ct.number_format = "#,##0"
            if clave in totales:
                # El total de cobertura NO es la suma de las filas: se cuenta
                # desde el grano cliente.
                cc = ws.cell(r_total, col + 1, totales[clave])
                cc.number_format = "#,##0"
                ca = ws.cell(r_avance, col + 2,
                             f'=IF({L(col)}{r_total}=0,"",'
                             f'{L(col + 1)}{r_total}/{L(col)}{r_total})')
                ca.number_format = "0.0%"
                ca.alignment = Alignment(horizontal="right")
            ws.conditional_formatting.add(f"{L(col + 2)}{fila_ini}:{L(col + 2)}{r_avance}",
                                          _SEMAFORO)

        ws.column_dimensions["A"].width = 26
        for j in range(2, ultima_col + 1):
            ws.column_dimensions[get_column_letter(j)].width = 11
        ws.freeze_panes = ws.cell(fila_ini, 2)
        wb.save(ruta)

    def run(self, config: IncentivoSaltaConfig) -> IncentivoSaltaResult:
        return self.generar_reporte(config)
