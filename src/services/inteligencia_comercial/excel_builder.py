"""Construccion del libro Excel de Inteligencia Comercial.

El libro se arma de forma declarativa: cada hoja es un `SheetSpec` que dice de
que analisis y de que tabla sale, y opcionalmente una funcion `decorate` que le
agrega graficos y formato condicional una vez escrita la tabla.

Ese desacople es a proposito. Los modulos de analisis no saben nada de openpyxl,
y esta capa no sabe nada de SQL ni de estadistica: cada una se puede testear sola.
"""
from __future__ import annotations

import logging
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.services.inteligencia_comercial import excel_style as st
from src.services.inteligencia_comercial import sheet_writer as sw
from src.services.inteligencia_comercial.contracts import AnalysisResult

logger = logging.getLogger(__name__)

# Excel corta los nombres de hoja en 31 caracteres.
MAX_SHEET_NAME = 31

SEVERITY_ORDER = {"critica": 0, "alta": 1, "media": 2, "info": 3}
SEVERITY_TONE = {"critica": "bad", "alta": "bad", "media": "warn", "info": "info"}


@dataclass
class SheetSpec:
    """Declaracion de una hoja de datos.

    Attributes:
        sheet: nombre de la hoja (se recorta a 31 caracteres).
        analysis: clave del AnalysisResult de origen.
        table: clave de la tabla dentro de ese resultado.
        title: titulo del banner.
        subtitle: linea de contexto bajo el titulo.
        max_rows: tope de filas. Lo que se corta se informa en la hoja.
        formats: overrides de formato numerico por columna.
        decorate: callback (ws, block, result) para graficos y formato condicional.
        tab_color: color de la solapa.
        sin_total_porque: motivo por el que esta hoja NO lleva fila de total.
            Todo informe de la casa lleva TOTAL GENERAL, asi que la ausencia hay
            que justificarla en la hoja: sumar coeficientes de Gini o meses de un
            pronostico no significa nada, y un total sin sentido es peor que ninguno.
    """

    sheet: str
    analysis: str
    table: str
    title: str
    subtitle: str = ""
    max_rows: int | None = None
    formats: dict[str, str] = field(default_factory=dict)
    decorate: Callable[[Worksheet, dict, AnalysisResult], None] | None = None
    tab_color: str = "brand"
    sin_total_porque: str = ""


def _safe_name(name: str) -> str:
    return name[:MAX_SHEET_NAME]


def build_workbook(
    results: dict[str, AnalysisResult],
    specs: list[SheetSpec],
    output_path: Path,
    periodo: str,
    generado: str,
    reconciliaciones: list[str] | None = None,
) -> Path:
    """Arma y guarda el libro completo.

    Args:
        results: {clave de analisis: AnalysisResult}.
        specs: hojas de datos a construir, en orden.
        output_path: destino del .xlsx. El directorio se crea si falta.
        periodo: descripcion de la ventana analizada, para la portada.
        generado: timestamp de generacion, ya formateado.
        reconciliaciones: diferencias conocidas entre analisis, explicadas. Dos
            numeros distintos para lo mismo hunden la credibilidad de un informe
            entero, asi que se declaran arriba de todo en Metodologia.

    Returns:
        La ruta escrita.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Las hojas se crean en orden; la portada se llena al final porque su indice
    # necesita saber que hojas existen realmente (una tabla vacia no genera hoja).
    portada = workbook.create_sheet("Portada")
    escritas: list[tuple[str, str, str]] = []

    for spec in specs:
        result = results.get(spec.analysis)
        if result is None:
            logger.warning("hoja %s omitida: falta el analisis %s", spec.sheet, spec.analysis)
            continue
        frame = result.table(spec.table)
        if frame is None or frame.empty:
            logger.warning(
                "hoja %s omitida: la tabla %s.%s vino vacia", spec.sheet, spec.analysis, spec.table
            )
            continue

        ws = workbook.create_sheet(_safe_name(spec.sheet))
        _build_data_sheet(ws, spec, frame, result)
        escritas.append((_safe_name(spec.sheet), spec.title, spec.subtitle))

    metodologia = workbook.create_sheet("Metodologia")
    _build_methodology_sheet(metodologia, results, periodo, generado, reconciliaciones or [])
    escritas.append(
        ("Metodologia", "Metodologia y advertencias",
         "Como se calculo cada numero y que no hay que leer de mas.")
    )

    _build_cover_sheet(portada, results, escritas, periodo, generado)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info("inteligencia-comercial: libro escrito en %s (%d hojas)",
                output_path, len(workbook.sheetnames))
    return output_path


def _build_data_sheet(
    ws: Worksheet, spec: SheetSpec, frame: pd.DataFrame, result: AnalysisResult
) -> None:
    row = st.write_title(ws, spec.title, spec.subtitle)
    block = sw.write_dataframe(
        ws, frame, row=row, formats=spec.formats, max_rows=spec.max_rows
    )

    nota_row = block["last_row"] + 2
    if block["dropped"]:
        # Nunca truncar en silencio: una tabla cortada se lee como completa.
        mostradas = f"{block['written']:,}".replace(",", ".")
        totales = f"{block['written'] + block['dropped']:,}".replace(",", ".")
        nota_row = st.write_note(
            ws,
            f"Se muestran {mostradas} filas de {totales}. El resto quedo fuera por "
            f"tamano, no por falta de datos. La fila TOTAL GENERAL sigue siendo la del "
            f"universo completo, no la del recorte.",
            nota_row,
            width=max(block["last_col"], 6),
            tone="warn",
        )

    if spec.sin_total_porque:
        st.write_note(
            ws,
            f"Esta hoja no lleva fila TOTAL GENERAL: {spec.sin_total_porque}",
            nota_row,
            width=max(block["last_col"], 6),
            tone="info",
        )

    st.finish_sheet(ws, freeze=f"A{block['first_row']}", tab_color=spec.tab_color)

    if spec.decorate is not None:
        try:
            spec.decorate(ws, block, result)
        except Exception:  # noqa: BLE001
            # Un grafico que falla no puede tumbar el reporte entero: la tabla
            # ya esta escrita y es el dato que importa.
            logger.exception("fallo la decoracion de la hoja %s", spec.sheet)


def _build_cover_sheet(
    ws: Worksheet,
    results: dict[str, AnalysisResult],
    escritas: list[tuple[str, str, str]],
    periodo: str,
    generado: str,
) -> None:
    # Las 12 columnas van iguales: cada tarjeta ocupa 3, y si la primera fuera
    # mas ancha la grilla de KPIs quedaria desalineada.
    st.set_widths(ws, {i: 15 for i in range(1, 13)})

    row = st.write_title(
        ws,
        "BADIE  ·  Inteligencia Comercial",
        f"{periodo}   |   generado {generado}",
    )

    headlines = [h for result in results.values() for h in result.headlines]
    if headlines:
        row = st.write_section(ws, "Indicadores clave", row)
        row = st.write_kpi_cards(
            ws,
            [
                st.Kpi(
                    label=h.label,
                    value=h.value,
                    number_format=h.number_format,
                    delta=h.delta,
                    note=h.note,
                    tone=h.tone,
                    higher_is_better=h.higher_is_better,
                )
                for h in headlines
            ],
            row,
        )

    alerts = [a for result in results.values() for a in result.alerts]
    alerts.sort(key=lambda a: (SEVERITY_ORDER.get(a.severity, 9), -(a.amount or 0)))
    if alerts:
        row = st.write_section(ws, "Hallazgos que requieren decision", row)
        for alert in alerts:
            tone = SEVERITY_TONE.get(alert.severity, "info")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            tag = ws.cell(row=row, column=1, value=alert.severity.upper())
            tag.font = Font(bold=True, size=9, color=st.PALETTE["paper"])
            tag.fill = st.fill(tone)
            tag.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
            body = ws.cell(row=row, column=3, value=f"{alert.title} — {alert.detail}")
            body.font = Font(size=10, color=st.PALETTE["ink"])
            body.fill = st.fill(f"{tone}_light")
            body.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            ws.row_dimensions[row].height = max(20, 13 * (1 + len(alert.detail) // 110))
            row += 1
        row += 1

    row = st.write_section(ws, "Contenido", row)
    st.write_index(ws, escritas, row)

    failed = [r.name for r in results.values() if r.failed]
    if failed:
        st.write_note(
            ws,
            "Analisis que no pudieron ejecutarse: " + ", ".join(failed)
            + ". El detalle del motivo esta en la hoja Metodologia.",
            ws.max_row + 2,
            tone="warn",
        )

    st.finish_sheet(ws, freeze=None, tab_color="ink")


def _build_methodology_sheet(
    ws: Worksheet,
    results: dict[str, AnalysisResult],
    periodo: str,
    generado: str,
    reconciliaciones: list[str],
) -> None:
    st.set_widths(ws, {1: 26, **{i: 12 for i in range(2, 13)}})
    row = st.write_title(
        ws,
        "Metodologia y advertencias",
        "Cada numero de este libro sale de uno de estos metodos. Las advertencias no son letra chica.",
    )

    row = st.write_section(ws, "Alcance", row)
    for label, value in (
        ("Ventana analizada", periodo),
        ("Generado", generado),
        ("Origen", "PostgreSQL, esquema gold (capa Gold del Data Warehouse). Solo lectura."),
        ("Herramientas", "numpy y pandas. Toda la estadistica esta implementada en el "
                         "proyecto (src/services/inteligencia_comercial/stats.py) y cubierta "
                         "por tests contra valores analiticos conocidos."),
    ):
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=10, color=st.PALETTE["ink"])
        cell.alignment = Alignment(vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        body = ws.cell(row=row, column=2, value=value)
        body.font = Font(size=10, color=st.PALETTE["ink_soft"])
        body.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = max(18, 13 * (1 + len(value) // 120))
        row += 1
    row += 1

    if reconciliaciones:
        row = st.write_section(ws, "Conciliacion entre analisis", row)
        row = st.write_note(
            ws,
            "Distintos analisis miden universos y ventanas distintos a proposito. "
            "Cuando dos cifras parecidas no coinciden, la diferencia esta explicada aca. "
            "Ninguna es un error de la otra.",
            row, tone="warn",
        )
        for texto in reconciliaciones:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            cell = ws.cell(row=row, column=1, value=f"·  {texto}")
            cell.font = Font(size=10, color=st.PALETTE["ink"])
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            ws.row_dimensions[row].height = max(16, 12 * (1 + len(texto) // 125))
            row += 1
        row += 1

    for result in results.values():
        if not result.notes and not result.failed:
            continue
        row = st.write_section(ws, result.name, row)
        if result.failed:
            row = st.write_note(
                ws, "Este analisis no pudo ejecutarse. Motivo abajo.", row, tone="bad"
            )
        for note in result.notes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            cell = ws.cell(row=row, column=1, value=f"·  {note}")
            cell.font = Font(size=10, color=st.PALETTE["ink"])
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            ws.row_dimensions[row].height = max(16, 12 * (1 + len(note) // 125))
            row += 1
        row += 1

    st.finish_sheet(ws, freeze=None, tab_color="ink_soft")
