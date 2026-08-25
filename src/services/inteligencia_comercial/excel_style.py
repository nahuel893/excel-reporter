"""Visual layer for the Inteligencia Comercial workbook.

The house `ExcelWriter` renders a data table with one fixed look. This module adds
what a decision-making workbook needs on top of that: a colour system, KPI cards,
conditional formatting and native Excel charts.

Charts are built with `openpyxl.chart` rather than pasted as images wherever the
chart type allows it, so the series stay linked to the cells and the reader can
filter, hover and re-scale them. Matplotlib is used only for the figures Excel
has no native equivalent for (see charts.py).
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colour system
# ---------------------------------------------------------------------------

# Anchored on A92C1F, the red the rest of the project's reports already use for
# headers, so this workbook reads as part of the same family.
PALETTE = {
    "brand": "A92C1F",
    "brand_dark": "7D1F16",
    "brand_light": "E8C4C0",
    "ink": "1F2430",
    "ink_soft": "5A6472",
    "paper": "FFFFFF",
    "paper_alt": "F4F6F8",
    "rule": "D6DBE1",
    # Semantic states
    "good": "1E7B4F",
    "good_light": "D6EFE1",
    "warn": "B8860B",
    "warn_light": "FBF0D0",
    "bad": "B3261E",
    "bad_light": "F8D7D4",
    "info": "1F5C8B",
    "info_light": "D6E6F2",
    "neutral": "6B7280",
    "neutral_light": "E9ECEF",
}

# Qualitative series colours for charts, ordered for maximum separation.
SERIES_COLORS = [
    "A92C1F", "1F5C8B", "1E7B4F", "B8860B", "6B3FA0",
    "C2571A", "2E8B94", "8B1A4A", "4A5568", "7A9A01",
]

# Diverging scale for heatmaps: bad -> neutral -> good.
SCALE_BAD, SCALE_MID, SCALE_GOOD = "F8696B", "FFEB84", "63BE7B"

THIN = Side(style="thin", color=PALETTE["rule"])
BORDER_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color_key: str) -> PatternFill:
    """Solid fill from a palette key, or from a literal hex if not a known key."""
    color = PALETTE.get(color_key, color_key)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

FMT_INT = "#,##0"
FMT_DEC1 = "#,##0.0"
FMT_DEC2 = "#,##0.00"
FMT_MONEY = '"$" #,##0'
FMT_MONEY_M = '"$" #,##0,,"M"'
FMT_PCT1 = "0.0%"
FMT_PCT2 = "0.00%"
FMT_DATE = "dd/mm/yyyy"
FMT_MONTH = "mmm-yy"
# Integer key columns must never carry a thousands separator: downstream VLOOKUPs
# read "1,00" instead of "1" and silently fail to match.
FMT_KEY = "0"


# ---------------------------------------------------------------------------
# Titles and KPI cards
# ---------------------------------------------------------------------------


@dataclass
class Kpi:
    """One KPI card.

    Attributes:
        label: what the number is.
        value: the number itself, already in display units.
        number_format: Excel format string.
        delta: optional variation vs the comparison period, as a fraction (0.12 = +12%).
        note: one line of context under the value.
        tone: 'good' | 'bad' | 'warn' | 'info' | 'neutral', or None to derive from delta.
        higher_is_better: how to colour the delta when tone is not forced.
    """

    label: str
    value: float | int | str
    number_format: str = FMT_INT
    delta: float | None = None
    note: str = ""
    tone: str | None = None
    higher_is_better: bool = True

    def resolved_tone(self) -> str:
        if self.tone:
            return self.tone
        if self.delta is None:
            return "info"
        improving = self.delta >= 0 if self.higher_is_better else self.delta <= 0
        if abs(self.delta) < 0.01:
            return "neutral"
        return "good" if improving else "bad"


def write_title(
    ws: Worksheet,
    title: str,
    subtitle: str = "",
    row: int = 1,
    width: int = 12,
) -> int:
    """Write a banner title block. Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=width)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=18, color=PALETTE["paper"])
    cell.fill = fill("brand")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    ws.row_dimensions[row + 1].height = 12

    next_row = row + 2
    if subtitle:
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=width)
        sub = ws.cell(row=next_row, column=1, value=subtitle)
        sub.font = Font(italic=True, size=10, color=PALETTE["ink_soft"])
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[next_row].height = 18
        next_row += 1
    return next_row + 1


def write_kpi_cards(
    ws: Worksheet,
    kpis: list[Kpi],
    row: int,
    per_row: int = 4,
    card_width: int = 3,
) -> int:
    """Lay out KPI cards in a grid. Returns the next free row.

    Each card is a merged block: label on top, big value in the middle, and a
    delta/note line at the bottom, tinted by whether the movement is good or bad.
    """
    if not kpis:
        return row

    current = row
    for start in range(0, len(kpis), per_row):
        chunk = kpis[start : start + per_row]
        for offset, kpi in enumerate(chunk):
            col = 1 + offset * card_width
            tone = kpi.resolved_tone()
            _write_single_card(ws, kpi, current, col, card_width, tone)
        ws.row_dimensions[current].height = 16
        ws.row_dimensions[current + 1].height = 30
        ws.row_dimensions[current + 2].height = 16
        current += 4
    return current


def _write_single_card(
    ws: Worksheet, kpi: Kpi, row: int, col: int, card_width: int, tone: str
) -> None:
    end_col = col + card_width - 1

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    label = ws.cell(row=row, column=col, value=kpi.label.upper())
    label.font = Font(bold=True, size=9, color=PALETTE["paper"])
    label.fill = fill("brand_dark")
    label.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=end_col)
    value = ws.cell(row=row + 1, column=col, value=kpi.value)
    value.font = Font(bold=True, size=16, color=PALETTE["ink"])
    value.fill = fill("paper_alt")
    value.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(kpi.value, (int, float)):
        value.number_format = kpi.number_format

    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=end_col)
    parts = []
    if kpi.delta is not None:
        arrow = "▲" if kpi.delta > 0 else ("▼" if kpi.delta < 0 else "＝")
        parts.append(f"{arrow} {kpi.delta:+.1%}")
    if kpi.note:
        parts.append(kpi.note)
    foot = ws.cell(row=row + 2, column=col, value="  ·  ".join(parts))
    foot.font = Font(bold=True, size=9, color=PALETTE[tone])
    foot.fill = fill(f"{tone}_light" if f"{tone}_light" in PALETTE else "paper_alt")
    foot.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(row, row + 3):
        for c in range(col, end_col + 1):
            ws.cell(row=r, column=c).border = BORDER_BOX


def write_note(ws: Worksheet, text: str, row: int, width: int = 12, tone: str = "info") -> int:
    """Write a full-width contextual note. Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(size=9, italic=True, color=PALETTE[tone])
    cell.fill = fill(f"{tone}_light" if f"{tone}_light" in PALETTE else "paper_alt")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = max(18, 13 * (1 + len(text) // 140))
    return row + 2


def write_section(ws: Worksheet, text: str, row: int, width: int = 12) -> int:
    """Write a section divider. Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text.upper())
    cell.font = Font(bold=True, size=11, color=PALETTE["paper"])
    cell.fill = fill("ink")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 2


# ---------------------------------------------------------------------------
# Conditional formatting
# ---------------------------------------------------------------------------


def color_scale(
    ws: Worksheet,
    cell_range: str,
    low: str = SCALE_BAD,
    mid: str = SCALE_MID,
    high: str = SCALE_GOOD,
    reverse: bool = False,
) -> None:
    """Three-colour scale across a range (heatmap)."""
    if reverse:
        low, high = high, low
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="percentile", start_value=5, start_color=low,
            mid_type="percentile", mid_value=50, mid_color=mid,
            end_type="percentile", end_value=95, end_color=high,
        ),
    )


def data_bars(ws: Worksheet, cell_range: str, color: str = "638EC6") -> None:
    """In-cell proportional bars — a sparkline for magnitude."""
    ws.conditional_formatting.add(
        cell_range,
        DataBarRule(start_type="num", start_value=0, end_type="max", color=color, showValue=True),
    )


def highlight_threshold(
    ws: Worksheet,
    cell_range: str,
    operator: str,
    formula: list[str],
    tone: str = "bad",
) -> None:
    """Tint cells that breach a threshold (e.g. negative margin, overdue clients)."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator=operator,
            formula=formula,
            fill=fill(f"{tone}_light" if f"{tone}_light" in PALETTE else tone),
            font=Font(bold=True, color=PALETTE.get(tone, tone)),
        ),
    )


# ---------------------------------------------------------------------------
# Native Excel charts
# ---------------------------------------------------------------------------


def _style_axes(chart, x_title: str, y_title: str) -> None:
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorGridlines = None


def bar_chart(
    ws: Worksheet,
    anchor: str,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    x_title: str = "",
    y_title: str = "",
    horizontal: bool = False,
    stacked: bool = False,
    width: float = 20,
    height: float = 9,
    show_values: bool = False,
    titles_from_data: bool = True,
) -> BarChart:
    """Native clustered/stacked bar chart anchored at a cell."""
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100
    chart.title = title
    chart.style = 2
    chart.add_data(data_ref, titles_from_data=titles_from_data)
    chart.set_categories(cats_ref)
    _style_axes(chart, x_title, y_title)
    chart.width, chart.height = width, height
    if show_values:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
    for idx, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = SERIES_COLORS[idx % len(SERIES_COLORS)]
        series.graphicalProperties.line.noFill = True
    ws.add_chart(chart, anchor)
    return chart


def line_chart(
    ws: Worksheet,
    anchor: str,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    x_title: str = "",
    y_title: str = "",
    width: float = 24,
    height: float = 10,
    dashed_series: tuple[int, ...] = (),
    smooth: bool = False,
) -> LineChart:
    """Native line chart. `dashed_series` indexes series to render dashed.

    Used for forecast bands: the actuals stay solid, the projected values and
    their confidence limits are dashed so nobody mistakes a model output for a
    measurement.
    """
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    _style_axes(chart, x_title, y_title)
    chart.width, chart.height = width, height
    for idx, series in enumerate(chart.series):
        color = SERIES_COLORS[idx % len(SERIES_COLORS)]
        series.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
        series.smooth = smooth
        if idx in dashed_series:
            series.graphicalProperties.line.dashStyle = "dash"
            series.graphicalProperties.line.w = 14000
    ws.add_chart(chart, anchor)
    return chart


def scatter_chart(
    ws: Worksheet,
    anchor: str,
    title: str,
    x_ref: Reference,
    y_ref: Reference,
    x_title: str = "",
    y_title: str = "",
    width: float = 20,
    height: float = 12,
    with_trendline: bool = False,
) -> ScatterChart:
    """Native scatter (markers only), optionally with a linear trendline."""
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    series = Series(y_ref, x_ref, title_from_data=True)
    series.marker = Marker(symbol="circle", size=6)
    series.graphicalProperties.line.noFill = True
    series.marker.graphicalProperties.solidFill = SERIES_COLORS[1]
    if with_trendline:
        series.trendline = Trendline(trendlineType="linear")
    chart.series.append(series)
    chart.width, chart.height = width, height
    ws.add_chart(chart, anchor)
    return chart


def waterfall_chart(
    ws: Worksheet,
    anchor: str,
    title: str,
    base_ref: Reference,
    delta_ref: Reference,
    cats_ref: Reference,
    colors: list[str],
    y_title: str = "",
    width: float = 22,
    height: float = 11,
) -> BarChart:
    """Waterfall (bridge) chart.

    Excel has no native waterfall type reachable through openpyxl, so it is built
    the standard way: a stacked column chart whose first series is an invisible
    riser holding each bar at its starting height, and whose second series is the
    visible movement. `colors` tints each visible bar individually, which is what
    makes a bridge readable — increases and decreases must not share a colour.

    Args:
        base_ref: the invisible riser series (one value per category).
        delta_ref: the visible movement series.
        colors: one hex colour per category, same length as the delta series.
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.add_data(base_ref, titles_from_data=True)
    chart.add_data(delta_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    _style_axes(chart, "", y_title)
    chart.width, chart.height = width, height

    riser, movement = chart.series[0], chart.series[1]
    riser.graphicalProperties.noFill = True
    riser.graphicalProperties.line.noFill = True

    # Per-point colouring: each data point carries its own shape properties.
    # DataPoint takes the raw `spPr` field — `graphicalProperties` is a read-only
    # alias on the instance and is rejected as a constructor keyword.
    movement.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=color))
        for i, color in enumerate(colors)
    ]
    movement.graphicalProperties.line.noFill = True

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.legend = None
    ws.add_chart(chart, anchor)
    return chart


def pie_chart(
    ws: Worksheet,
    anchor: str,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    width: float = 12,
    height: float = 9,
) -> PieChart:
    """Native pie chart with percentage labels."""
    chart = PieChart()
    chart.title = title
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width, chart.height = width, height
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    ws.add_chart(chart, anchor)
    return chart


# ---------------------------------------------------------------------------
# Sheet-level helpers
# ---------------------------------------------------------------------------


def finish_sheet(
    ws: Worksheet,
    freeze: str | None = None,
    tab_color: str = "brand",
    hide_gridlines: bool = True,
    fit_to_width: bool = True,
) -> None:
    """Apply the finishing touches every sheet in this workbook shares."""
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = PALETTE.get(tab_color, tab_color)
    if hide_gridlines:
        ws.sheet_view.showGridLines = False
    if fit_to_width:
        # Estas tablas son anchas: sin esto, imprimir parte una fila en tres
        # paginas y el informe se vuelve ilegible en papel.
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"


def set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    """Set explicit column widths by 1-based index."""
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_index(
    ws: Worksheet,
    entries: list[tuple[str, str, str]],
    row: int,
    width: int = 12,
) -> int:
    """Write a clickable table of contents.

    Args:
        entries: (sheet name, display label, one-line description).
    """
    header = ["Hoja", "Contenido"]
    for offset, text in enumerate(header):
        cell = ws.cell(row=row, column=1 + offset * 3, value=text)
        cell.font = Font(bold=True, size=10, color=PALETTE["paper"])
        cell.fill = fill("ink")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(
            start_row=row, start_column=1 + offset * 3,
            end_row=row, end_column=(3 if offset == 0 else width),
        )
    row += 1

    for sheet_name, label, description in entries:
        link = ws.cell(row=row, column=1, value=label)
        # Must be a Hyperlink with `location`, not a string: assigning a string
        # populates `target` (an EXTERNAL link) and Excel will not navigate to a
        # sheet in the same workbook. Quoting the name keeps it valid with spaces.
        link.hyperlink = Hyperlink(ref=link.coordinate, location=f"'{sheet_name}'!A1")
        link.font = Font(color=PALETTE["info"], underline="single", bold=True, size=10)
        link.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        desc = ws.cell(row=row, column=4, value=description)
        desc.font = Font(size=9, color=PALETTE["ink_soft"])
        desc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=width)
        ws.row_dimensions[row].height = 18
        row += 1
    return row + 1
