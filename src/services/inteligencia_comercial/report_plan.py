"""Plan de hojas del libro: que tabla va en cada hoja, con que formato y que grafico.

Separado del builder a proposito: aca vive todo lo que hay que tocar para agregar,
sacar o reordenar una hoja, sin entrar en la mecanica de openpyxl.

Sobre los formatos explicitos: `sheet_writer.infer_format` acierta en la mayoria
de las columnas, pero no puede adivinar que 'OTIF (<=1 dia)', 'Gini' o 'M0' son
porcentajes o coeficientes. Una columna de porcentaje mostrada como entero se lee
como 0 o 1 y es directamente un numero falso, asi que todo lo dudoso va explicito.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.chart import Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.services.inteligencia_comercial import excel_style as st
from src.services.inteligencia_comercial import sheet_writer as sw
from src.services.inteligencia_comercial.contracts import AnalysisResult
from src.services.inteligencia_comercial.excel_builder import SheetSpec

PCT = st.FMT_PCT1
PCT2 = st.FMT_PCT2
DEC1 = st.FMT_DEC1
DEC2 = st.FMT_DEC2
MONEY = st.FMT_MONEY
INT = st.FMT_INT

# Columna libre a la derecha de cualquier tabla, para los bloques auxiliares que
# alimentan los graficos (una cascada necesita series que no existen en el dato).
AUX_GAP = 2


def _aux_col(block: dict) -> int:
    return block["last_col"] + AUX_GAP


def _write_aux(ws: Worksheet, row: int, col: int, headers: list[str], rows: list[list]) -> dict:
    """Escribe un bloque auxiliar oculto para graficos. Devuelve sus coordenadas."""
    for offset, text in enumerate(headers):
        cell = ws.cell(row=row, column=col + offset, value=text)
        cell.font = Font(bold=True, size=8, color=st.PALETTE["ink_soft"])
        cell.alignment = Alignment(horizontal="center")
    for r_offset, record in enumerate(rows, start=1):
        for c_offset, value in enumerate(record):
            ws.cell(row=row + r_offset, column=col + c_offset, value=value)
    return {"header_row": row, "first_row": row + 1, "last_row": row + len(rows), "col": col}


def _ref(ws, col: int, first: int, last: int) -> Reference:
    return Reference(ws, min_col=col, max_col=col, min_row=first, max_row=last)


def _col_ref(ws, block: dict, name: str, with_header: bool = True) -> Reference:
    col = block["columns"][name]
    first = block["header_row"] if with_header else block["first_row"]
    return Reference(ws, min_col=col, max_col=col, min_row=first, max_row=block["last_row"])


def _anchor(block: dict, offset_cols: int = 0, row: int | None = None) -> str:
    """Ancla una figura debajo o al costado de la tabla."""
    col = block["last_col"] + AUX_GAP + offset_cols
    return f"{get_column_letter(col)}{row or block['header_row']}"


def _sin_total(df: pd.DataFrame) -> pd.DataFrame:
    first = df.columns[0]
    mask = df[first].astype(str).str.upper().str.startswith("TOTAL")
    return df[~mask]


# ---------------------------------------------------------------------------
# Decoradores por hoja
# ---------------------------------------------------------------------------


def _dec_serie_mensual(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Evolucion mensual: el total y los tres genericos que mandan."""
    series = ["TOTAL GENERAL - Bultos", "CERVEZAS - Bultos", "FRATELLI B - Bultos",
              "AGUAS DANONE - Bultos"]
    presentes = [s for s in series if s in block["columns"]]
    if not presentes:
        return
    last = block["last_row"] - 1  # la ultima fila es TOTAL GENERAL, no es un mes
    cats = Reference(ws, min_col=block["columns"]["mes"], min_row=block["first_row"], max_row=last)
    for idx, name in enumerate(presentes):
        col = block["columns"][name]
        data = Reference(ws, min_col=col, max_col=col, min_row=block["header_row"], max_row=last)
        if idx == 0:
            chart = st.line_chart(
                ws, _anchor(block), "Volumen mensual en bultos (serie real, no pesos)",
                data, cats, y_title="bultos", width=30, height=12,
            )
        else:
            chart.add_data(data, titles_from_data=True)
    st.write_note(
        ws,
        "Serie en BULTOS a proposito. En pesos nominales la misma serie crece +41,8% "
        "y en volumen real +10,6%: la diferencia es inflacion, no negocio.",
        block["last_row"] + 2, width=8, tone="info",
    )


def _dec_rfm_resumen(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    cats = _col_ref(ws, block, "Segmento", with_header=False)
    st.bar_chart(
        ws, _anchor(block), "Neto 12m por segmento (nominal $)",
        _col_ref(ws, block, "Neto 12m ($)") if "Neto 12m ($)" in block["columns"]
        else _col_ref(ws, block, [c for c in block["columns"] if "Neto" in c][0]),
        cats, y_title="$ neto", width=22, height=11,
    )
    st.bar_chart(
        ws, _anchor(block, offset_cols=12), "Clientes por segmento",
        _col_ref(ws, block, "Clientes"), cats, y_title="clientes",
        width=18, height=11, horizontal=True,
    )


def _dec_cohortes(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Mapa de calor de retencion: verde = sobrevive, rojo = se fue."""
    columnas = [c for c in block["columns"] if str(c).startswith("M")]
    if not columnas:
        return
    first = block["columns"][columnas[0]]
    last = block["columns"][columnas[-1]]
    rango = (f"{get_column_letter(first)}{block['first_row']}:"
             f"{get_column_letter(last)}{block['last_row']}")
    st.color_scale(ws, rango, low=st.SCALE_BAD, mid=st.SCALE_MID, high=st.SCALE_GOOD)


def _dec_puente(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Cascada del crecimiento en hectolitros: de donde salio y donde se perdio."""
    df = _sin_total(result.table("puente"))
    if df.empty or "Delta htl" not in df.columns:
        return

    filas, base, acumulado = [], [], 0.0
    for _, record in df.iterrows():
        delta = float(record["Delta htl"])
        # El riser arranca en el piso del tramo: para una caida, en el valor final.
        base.append(acumulado + min(delta, 0.0))
        filas.append([str(record["Movimiento"]), acumulado + min(delta, 0.0), abs(delta)])
        acumulado += delta
    filas.append(["NETO DEL PERIODO", 0.0, acumulado])

    aux = _write_aux(ws, block["header_row"], _aux_col(block),
                     ["Movimiento", "Base", "Delta"], filas)
    colores = [
        st.PALETTE["good"] if float(v[2]) >= 0 and "Perdid" not in v[0] and "Downsell" not in v[0]
        else st.PALETTE["bad"]
        for v in filas[:-1]
    ] + [st.PALETTE["info"]]

    st.waterfall_chart(
        ws, _anchor(block, offset_cols=4),
        "Puente de crecimiento en hectolitros (volumen real)",
        _ref(ws, aux["col"] + 1, aux["header_row"], aux["last_row"]),
        _ref(ws, aux["col"] + 2, aux["header_row"], aux["last_row"]),
        _ref(ws, aux["col"], aux["first_row"], aux["last_row"]),
        colors=colores, y_title="hectolitros", width=26, height=12,
    )
    for col in range(aux["col"], aux["col"] + 3):
        ws.column_dimensions[get_column_letter(col)].hidden = True


def _dec_lorenz(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    cats = _col_ref(ws, block, "% acumulado de la poblacion", with_header=False)
    nombres = ["Igualdad perfecta", "% acumulado del neto — clientes",
               "% acumulado del neto — articulos"]
    chart = None
    for name in nombres:
        if name not in block["columns"]:
            continue
        data = _col_ref(ws, block, name)
        if chart is None:
            chart = st.line_chart(
                ws, _anchor(block),
                "Curva de Lorenz: cuanto se concentra la facturacion",
                data, cats, x_title="% acumulado de entidades",
                y_title="% acumulado del neto", width=20, height=14, dashed_series=(0,),
            )
        else:
            chart.add_data(data, titles_from_data=True)
    st.write_note(
        ws,
        "Cuanto mas se hunde la curva bajo la diagonal, mas concentrado esta. "
        "La curva de articulos se hunde muchisimo mas que la de clientes: el riesgo "
        "de BADIE esta en el producto, no en la cartera.",
        block["last_row"] + 2, width=6, tone="info",
    )


def _dec_top_articulos(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    last = block["last_row"] - 1
    cats = Reference(ws, min_col=block["columns"]["Articulo"],
                     min_row=block["first_row"], max_row=last)
    col = block["columns"]["Neto (12m)"]
    st.bar_chart(
        ws, _anchor(block), "Top 25 SKUs por neto (nominal $)",
        Reference(ws, min_col=col, max_col=col, min_row=block["header_row"], max_row=last),
        cats, y_title="$ neto", width=30, height=13, horizontal=True,
    )


def _dec_abc_box(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    cats = _col_ref(ws, block, "Celda", with_header=False)
    st.bar_chart(
        ws, _anchor(block), "Neto por celda ABC-XYZ", _col_ref(ws, block, "Neto 12m ($)"),
        cats, y_title="$ neto", width=20, height=10, show_values=True,
    )
    st.bar_chart(
        ws, _anchor(block, offset_cols=11), "SKUs por celda ABC-XYZ",
        _col_ref(ws, block, "SKUs"), cats, y_title="SKUs", width=20, height=10,
    )


def _dec_residuos(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Residuos chi-cuadrado: rojo = compra mucho menos de lo esperable, verde = mas."""
    columnas = [c for c in block["columns"] if c != "Subcanal"]
    if not columnas:
        return
    first = block["columns"][columnas[0]]
    last = block["columns"][columnas[-1]]
    rango = (f"{get_column_letter(first)}{block['first_row']}:"
             f"{get_column_letter(last)}{block['last_row'] - 1}")
    st.color_scale(ws, rango)
    st.write_note(
        ws,
        "Residuo estandarizado = (observado - esperado) / raiz(esperado). Por encima de "
        "|2| la celda se aparta de lo que su tamano haria esperar. El agujero mas grande "
        "de la matriz es MAYORISTAS x AGUAS DANONE.",
        block["last_row"] + 2, width=len(columnas) + 1, tone="info",
    )


def _dec_canal_yoy(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    last = block["last_row"] - 1
    cats = Reference(ws, min_col=block["columns"]["Subcanal"],
                     min_row=block["first_row"], max_row=last)
    for name, anchor_offset, title in (
        ("Delta % Bultos", 0, "Variacion de volumen por subcanal (12m vs 12m previos)"),
        ("Delta % Clientes", 12, "Variacion de clientes por subcanal"),
    ):
        if name not in block["columns"]:
            continue
        col = block["columns"][name]
        st.bar_chart(
            ws, _anchor(block, offset_cols=anchor_offset), title,
            Reference(ws, min_col=col, max_col=col, min_row=block["header_row"], max_row=last),
            cats, y_title="variacion", width=22, height=12, horizontal=True,
        )
    st.write_note(
        ws,
        "Leer las dos juntas: un subcanal que sube en bultos y baja en clientes esta "
        "ganando profundidad y perdiendo ancho. Es lo que le pasa a SUBCANAL ALMACEN.",
        block["last_row"] + 2, width=8, tone="warn",
    )


def _dec_cascada(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Cascada de precio: del bruto de lista al neto que entra."""
    df = result.table("cascada")
    df = df[(df["ambito"] == "TOTAL GENERAL") & (~df["concepto"].str.startswith("Memo"))]
    if df.empty:
        return

    filas, acumulado = [], 0.0
    for _, record in df.iterrows():
        monto = float(record["monto"])
        concepto = str(record["concepto"])
        if concepto.startswith("Bruto"):
            filas.append([concepto, 0.0, monto])
            acumulado = monto
        elif concepto.startswith("Neto"):
            filas.append([concepto, 0.0, monto])
        elif bool(record["es_resta"]):
            acumulado -= monto
            filas.append([concepto, acumulado, monto])
        else:
            filas.append([concepto, acumulado, monto])
            acumulado += monto

    aux = _write_aux(ws, block["header_row"], _aux_col(block),
                     ["Concepto", "Base", "Delta"], filas)
    colores = []
    for record in filas:
        nombre = record[0]
        if nombre.startswith("Bruto"):
            colores.append(st.PALETTE["info"])
        elif nombre.startswith("Neto"):
            colores.append(st.PALETTE["good"])
        else:
            colores.append(st.PALETTE["bad"])

    st.waterfall_chart(
        ws, _anchor(block, offset_cols=4),
        "Del bruto de lista al neto realizado (12m, nominal $)",
        _ref(ws, aux["col"] + 1, aux["header_row"], aux["last_row"]),
        _ref(ws, aux["col"] + 2, aux["header_row"], aux["last_row"]),
        _ref(ws, aux["col"], aux["first_row"], aux["last_row"]),
        colors=colores, y_title="$ nominales", width=26, height=13,
    )
    for col in range(aux["col"], aux["col"] + 3):
        ws.column_dimensions[get_column_letter(col)].hidden = True


def _dec_estacionalidad(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Los indices de los genericos grandes, mes a mes."""
    meses = ["indice_ene", "indice_feb", "indice_mar", "indice_abr", "indice_may",
             "indice_jun", "indice_jul", "indice_ago", "indice_sep", "indice_oct",
             "indice_nov", "indice_dic"]
    if not all(m in block["columns"] for m in meses):
        return
    df = result.table("estacionalidad")
    objetivo = ["CERVEZAS", "FRATELLI B", "AGUAS DANONE", "VINOS CCU"]
    filas = [["Mes"] + objetivo]
    etiquetas = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                 "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    for idx, etiqueta in enumerate(etiquetas):
        fila = [etiqueta]
        for generico in objetivo:
            match = df[df["generico"] == generico]
            fila.append(float(match.iloc[0][meses[idx]]) if not match.empty else None)
        filas.append(fila)

    aux = _write_aux(ws, block["last_row"] + 3, 1, filas[0], filas[1:])
    cats = _ref(ws, aux["col"], aux["first_row"], aux["last_row"])
    chart = None
    for offset in range(1, len(objetivo) + 1):
        data = _ref(ws, aux["col"] + offset, aux["header_row"], aux["last_row"])
        if chart is None:
            chart = st.line_chart(
                ws, f"A{aux['last_row'] + 3}",
                "Indice estacional por mes (1,00 = mes promedio)",
                data, cats, x_title="mes", y_title="indice", width=28, height=13,
            )
        else:
            chart.add_data(data, titles_from_data=True)
    st.write_note(
        ws,
        "CERVEZAS y FRATELLI B son contra-estacionales: una pica en diciembre y la otra "
        "en julio. Eso es un hecho de deposito y de caja, no una curiosidad.",
        aux["last_row"] + 1, width=6, tone="warn",
    )


def _dec_pronostico(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    """Historico y proyeccion del generico principal, con su banda."""
    df = result.table("pronostico")
    if df.empty:
        return
    generico = "CERVEZAS" if (df["generico"] == "CERVEZAS").any() else df["generico"].iloc[0]
    serie = df[df["generico"] == generico].copy()
    if serie.empty:
        return

    filas = [
        [
            str(record["mes"])[:7],
            float(record["bultos"]) if record["tipo"] == "Historico" else None,
            float(record["bultos"]) if record["tipo"] == "Pronostico" else None,
            None if pd.isna(record["limite_inferior"]) else float(record["limite_inferior"]),
            None if pd.isna(record["limite_superior"]) else float(record["limite_superior"]),
        ]
        for _, record in serie.iterrows()
    ]
    aux = _write_aux(
        ws, block["header_row"], _aux_col(block),
        ["Mes", "Historico", "Pronostico", "Limite inferior", "Limite superior"], filas,
    )
    cats = _ref(ws, aux["col"], aux["first_row"], aux["last_row"])
    chart = None
    for offset in range(1, 5):
        data = _ref(ws, aux["col"] + offset, aux["header_row"], aux["last_row"])
        if chart is None:
            chart = st.line_chart(
                ws, _anchor(block, offset_cols=7),
                f"{generico}: historico y pronostico a 6 meses con banda del 95%",
                data, cats, y_title="bultos", width=30, height=13, dashed_series=(1, 2, 3),
            )
        else:
            chart.add_data(data, titles_from_data=True)
    for col in range(aux["col"], aux["col"] + 5):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    st.write_note(
        ws,
        "El modelo elegido por serie sale de un backtest de origen movil a 1 mes: se corren "
        "Holt-Winters y una linea base estacional-ingenua y gana la de menor MAPE. En las "
        "series grandes gana la ingenua, y eso esta en la tabla para que se pueda auditar.",
        block["last_row"] + 2, width=9, tone="info",
    )


def _dec_sla(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    df = result.table("sla")
    sucursales = df[df["Nivel"] == "Sucursal"] if "Nivel" in df.columns else df
    if sucursales.empty:
        return
    first = block["first_row"]
    last = first + len(sucursales) - 1
    cats = Reference(ws, min_col=block["columns"]["Entidad"], min_row=first, max_row=last)
    col = block["columns"]["OTIF (<=1 dia)"]
    st.bar_chart(
        ws, _anchor(block), "OTIF por sucursal (entregas en 1 dia o menos)",
        Reference(ws, min_col=col, max_col=col, min_row=block["header_row"], max_row=last),
        cats, y_title="OTIF", width=24, height=11, horizontal=True,
    )
    st.highlight_threshold(ws, sw.data_range(block, "Brecha vs Red"), "lessThan", ["-0.05"], "bad")
    st.highlight_threshold(ws, sw.data_range(block, "Brecha vs Red"), "greaterThan", ["0.05"], "good")


def _dec_rechazos(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    st.data_bars(ws, sw.data_range(block, "Tasa Rechazo"), color="C00000")
    st.write_note(
        ws,
        "Hallazgo del analisis: cantidades_rechazo solo esta poblado en notas de credito "
        "(DVVTA). En este esquema rechazo y devolucion son el mismo evento. El outlier real "
        "de la red es TARTAGAL, no ABRA PAMPA (que ademas es una sucursal cerrada).",
        block["last_row"] + 2, width=10, tone="warn",
    )


def _dec_stock(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    st.color_scale(ws, sw.data_range(block, "Dias de Cobertura"),
                   low=st.SCALE_BAD, mid=st.SCALE_GOOD, high=st.SCALE_MID)
    st.highlight_threshold(ws, sw.data_range(block, "Dias de Cobertura"),
                           "lessThan", ["7"], "bad")
    st.write_note(
        ws,
        "Dias de cobertura = stock actual / velocidad diaria de los ultimos 60 dias. "
        "Rojo abajo de 7 dias (riesgo de quiebre) y amarillo arriba de 90 (capital dormido). "
        "fact_stock arranca en 2026-02: no hay historia para calcular rotacion ni antiguedad.",
        block["last_row"] + 2, width=10, tone="info",
    )


def _dec_rutas(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    last = block["last_row"] - 1
    x = Reference(ws, min_col=block["columns"]["Visitas Facturadas"],
                  min_row=block["first_row"], max_row=last)
    col = block["columns"]["Drop Size (bultos/visita)"]
    y = Reference(ws, min_col=col, max_col=col, min_row=block["header_row"], max_row=last)
    st.scatter_chart(
        ws, _anchor(block), "Densidad de reparto: visitas vs bultos por visita",
        x, y, x_title="visitas facturadas (12m)", y_title="bultos por visita",
        width=22, height=13,
    )
    st.write_note(
        ws,
        "Abajo a la derecha estan las rutas caras: muchas visitas y poco volumen por visita. "
        "Ahi es donde el costo de servir se come el margen.",
        block["last_row"] + 2, width=10, tone="warn",
    )


def _dec_top_cliente_riesgo(ws: Worksheet, block: dict, result: AnalysisResult) -> None:
    if "neto_12m" in block["columns"]:
        st.data_bars(ws, sw.data_range(block, "neto_12m"))
    for name in ("Neto 12m ($)", "Neto 12m (Nominal $)"):
        if name in block["columns"]:
            st.data_bars(ws, sw.data_range(block, name))
            break


# ---------------------------------------------------------------------------
# Plan
# ---------------------------------------------------------------------------

PLAN: list[SheetSpec] = [
    SheetSpec(
        sheet="Volumen Mensual", analysis="demanda", table="serie_mensual",
        title="Volumen mensual por generico",
        subtitle="Bultos y hectolitros, historia completa. La serie real del negocio, sin inflacion.",
        decorate=_dec_serie_mensual, tab_color="info",
    ),
    SheetSpec(
        sheet="RFM Resumen", analysis="clientes", table="rfm_resumen",
        title="Segmentacion RFM — resumen por segmento",
        subtitle="Recencia, Frecuencia y Monto en quintiles. Cada segmento trae la accion que le corresponde.",
        formats={"% Clientes": PCT, "% Neto": PCT},
        decorate=_dec_rfm_resumen, tab_color="brand",
    ),
    SheetSpec(
        sheet="Segmentacion RFM", analysis="clientes", table="rfm",
        title="Segmentacion RFM — detalle por cliente",
        subtitle="Un cliente por fila con su score, su celda RFM y la accion sugerida. Ordenado por neto.",
        max_rows=8000, tab_color="brand",
    ),
    SheetSpec(
        sheet="Riesgo de Fuga", analysis="clientes", table="fuga",
        title="Riesgo de fuga contra el ritmo propio de cada cliente",
        subtitle="Un cliente esta atrasado cuando su silencio supera SU p90 historico, no un umbral fijo de 60 dias.",
        max_rows=4000, decorate=_dec_top_cliente_riesgo, tab_color="bad",
    ),
    SheetSpec(
        sheet="Cohortes", analysis="clientes", table="cohortes",
        title="Retencion por cohorte mensual",
        subtitle="Cohortes desde que la red esta completa (2024-07). Antes de esa fecha solo existe CASA CENTRAL en fact_ventas.",
        formats={f"M{i}": PCT for i in range(25)},
        decorate=_dec_cohortes, tab_color="info",
    ),
    SheetSpec(
        sheet="Puente de Crecimiento", analysis="clientes", table="puente",
        title="Puente de crecimiento: de donde salio y donde se perdio el volumen",
        subtitle="Descomposicion en hectolitros (real) y en pesos nominales. Los buckets cierran exactamente contra el delta total.",
        formats={"% del delta htl": PCT, "% del delta neto": PCT,
                 "Htl previo": DEC1, "Htl actual": DEC1, "Delta htl": DEC1},
        decorate=_dec_puente, tab_color="good",
    ),
    SheetSpec(
        sheet="Concentracion", analysis="clientes", table="concentracion",
        title="Riesgo de concentracion",
        subtitle="Gini, HHI y Pareto sobre clientes y sobre articulos. Los dos universos dan resultados opuestos.",
        formats={"Gini": DEC2, "HHI": DEC1, "N efectivo": DEC1, "% que hace el 80%": PCT,
                 **{f"Top {n}": PCT for n in (1, 5, 10, 20, 50, 100)}},
        tab_color="warn",
        sin_total_porque=(
            "sumar coeficientes de Gini y HHI no significa nada. Los totales en pesos y volumen estan en la hoja Top Articulos."
        ),
    ),
    SheetSpec(
        sheet="Curva de Lorenz", analysis="clientes", table="lorenz",
        title="Curva de Lorenz",
        subtitle="Cuanto se aparta la distribucion real de un reparto parejo.",
        formats={c: PCT for c in ("% acumulado de la poblacion", "Igualdad perfecta",
                                  "% acumulado del neto — clientes",
                                  "% acumulado del neto — articulos")},
        decorate=_dec_lorenz, tab_color="warn",
        sin_total_porque=(
            "es una curva remuestreada sobre una grilla de 0 a 100%, no un conjunto de medidas sumables."
        ),
    ),
    SheetSpec(
        sheet="Top Articulos", analysis="clientes", table="top_articulos",
        title="Top 25 SKUs por facturacion neta",
        subtitle="Con participacion y acumulado. Aca se ve el tamano real del riesgo de portafolio.",
        formats={"% del neto": PCT, "% acumulado": PCT, "Hectolitros (12m)": DEC1},
        decorate=_dec_top_articulos, tab_color="warn",
    ),
    SheetSpec(
        sheet="Portafolio ABC-XYZ", analysis="portafolio", table="abc_xyz",
        title="Portafolio ABC-XYZ",
        subtitle="ABC por aporte al neto; XYZ por coeficiente de variacion de la demanda mensual.",
        formats={"% Neto": PCT, "% Neto Acumulado": PCT, "CV Demanda": DEC2},
        max_rows=800, tab_color="info",
    ),
    SheetSpec(
        sheet="ABC-XYZ 9-Box", analysis="portafolio", table="abc_xyz_box",
        title="Matriz ABC-XYZ",
        subtitle="Cuantos SKUs y cuanta plata cae en cada celda. La celda AZ vacia es una buena noticia.",
        formats={"% SKUs": PCT, "% Neto": PCT},
        decorate=_dec_abc_box, tab_color="info",
    ),
    SheetSpec(
        sheet="Canal x Generico", analysis="portafolio", table="canal_generico",
        title="Mix de volumen por subcanal y generico",
        subtitle="Tabla de contingencia en bultos, 12 meses.",
        tab_color="info",
    ),
    SheetSpec(
        sheet="Residuos Chi2", analysis="portafolio", table="canal_generico_residuos",
        title="Residuos estandarizados chi-cuadrado",
        subtitle="Que subcanal compra mucho mas (verde) o mucho menos (rojo) de lo que su tamano haria esperar.",
        formats={}, decorate=_dec_residuos, tab_color="warn",
    ),
    SheetSpec(
        sheet="Subcanal YoY", analysis="portafolio", table="canal_yoy",
        title="Subcanales: volumen y ancho de cartera, 12m vs 12m previos",
        subtitle="Comparacion en bultos, no en pesos: en pesos nominales todo crece por inflacion.",
        formats={"Delta % Bultos": PCT, "Delta % Clientes": PCT, "Delta % Bultos/Cliente": PCT,
                 "Bultos/Cliente Actual": DEC1, "Bultos/Cliente Previo": DEC1},
        decorate=_dec_canal_yoy, tab_color="good",
    ),
    SheetSpec(
        sheet="Cross-sell Oportunidad", analysis="portafolio", table="cross_sell",
        title="Espacio en blanco por subcanal y generico",
        subtitle="Penetracion entre pares y tamano de la oportunidad. Es una COTA SUPERIOR, no un pronostico.",
        formats={"Penetracion": PCT}, tab_color="good",
    ),
    SheetSpec(
        sheet="Cross-sell Clientes", analysis="portafolio", table="cross_sell_clientes",
        title="Cross-sell: lista de clientes accionable",
        subtitle="Clientes que no compran un generico que sus pares del mismo subcanal si compran.",
        tab_color="good",
    ),
    SheetSpec(
        sheet="Reglas de Asociacion", analysis="portafolio", table="reglas",
        title="Reglas de asociacion (market basket)",
        subtitle="Calculadas DENTRO de cada fuerza de ventas: una factura nunca mezcla preventa con autoventa.",
        formats={"Soporte": PCT2, "Confianza": PCT, "Lift": DEC2,
                 "Leverage": DEC2, "Conviccion": DEC2},
        tab_color="good",
    ),
    SheetSpec(
        sheet="Ciclo de Vida SKU", analysis="portafolio", table="ciclo_vida",
        title="Ciclo de vida de SKUs",
        subtitle="Lanzamientos de los ultimos 12 meses y articulos sin venta que siguen ocupando deposito.",
        tab_color="info",
    ),
    SheetSpec(
        sheet="Margen", analysis="rentabilidad", table="margen",
        title="Margen bruto ponderado por volumen",
        subtitle="ATENCION: esta tabla llega hasta 2026-05-05, el corte del ETL contable. No es contemporanea con el resto del libro.",
        formats={"Margen Ponderado %": PCT, "Margen Panel Constante %": PCT,
                 "Margen p05 Linea %": PCT, "Margen p25 Linea %": PCT,
                 "Margen Mediano Linea %": PCT, "Margen p75 Linea %": PCT,
                 "Margen p95 Linea %": PCT, "Rango Intercuartil %": PCT,
                 "MAD Margen Linea %": PCT, "% Lineas Bajo Costo": PCT},
        tab_color="warn",
    ),
    SheetSpec(
        sheet="Venta Bajo Costo", analysis="rentabilidad", table="bajo_costo",
        title="Ventas por debajo del costo",
        subtitle="Buena parte de esto es un costo mal cargado, no politica comercial. Verificar antes de confrontar a nadie.",
        formats={"Margen Ponderado %": PCT}, tab_color="bad",
    ),
    SheetSpec(
        sheet="Cascada de Descuentos", analysis="rentabilidad", table="cascada",
        title="Del bruto de lista al neto realizado",
        subtitle="La fuga completa, abierta por sucursal y por generico. Todo en pesos nominales de la ventana.",
        formats={"monto": MONEY, "base_acumulada": MONEY, "pct_bruto": PCT},
        decorate=_dec_cascada, tab_color="bad",
        sin_total_porque=(
            "el total ya es la linea 'Neto realizado'. Sumar los tramos contaria el bruto dos veces."
        ),
    ),
    SheetSpec(
        sheet="Fuga de Descuentos", analysis="rentabilidad", table="fuga_outliers",
        title="Clientes y preventistas con descuento atipico",
        subtitle="z robusto sobre mediana y MAD: no lo mueve un puñado de extremos, a diferencia de la media y el desvio.",
        formats={"tasa": PCT2, "z_robusto": DEC2}, tab_color="bad",
    ),
    SheetSpec(
        sheet="Dispersion de Precios", analysis="rentabilidad", table="dispersion",
        title="Dispersion del precio realizado dentro del mes",
        subtitle="Con inflacion mensual de 2-3% el piso de ruido es un CV de ~3%. Debajo de eso NO hay problema de control de precios.",
        formats={"cv_ponderado_pct": PCT2, "ratio_p90_p10": DEC2}, tab_color="warn",
    ),
    SheetSpec(
        sheet="Estacionalidad", analysis="demanda", table="estacionalidad",
        title="Indices estacionales por generico",
        subtitle="Descomposicion multiplicativa clasica. 1,00 es el mes promedio del ano.",
        formats={**{f"indice_{m}": DEC2 for m in
                    ("ene", "feb", "mar", "abr", "may", "jun",
                     "jul", "ago", "sep", "oct", "nov", "dic")},
                 "fuerza_estacional": PCT, "indice_pico": DEC2, "indice_valle": DEC2,
                 "amplitud_pico_valle": DEC2, "desvio_residual": DEC2},
        decorate=_dec_estacionalidad, tab_color="info",
    ),
    SheetSpec(
        sheet="Pronostico", analysis="demanda", table="pronostico",
        title="Pronostico a 6 meses con backtest honesto",
        subtitle="Se corren Holt-Winters y una linea base estacional-ingenua; gana la de menor MAPE y las dos quedan en la tabla.",
        formats={"mape_hw": DEC1, "mape_naive": DEC1},
        decorate=_dec_pronostico, tab_color="info",
        sin_total_porque=(
            "mezcla meses historicos con meses proyectados de dos modelos distintos; un total agregaria manzanas con peras."
        ),
    ),
    SheetSpec(
        sheet="Alerta Estacional", analysis="demanda", table="estacionalidad_alerta",
        title="Que se viene en los proximos 3 meses",
        subtitle="Lectura accionable del indice estacional contra el mes de referencia.",
        formats={"indice_referencia": DEC2, "indice": DEC2, "variacion_vs_referencia": PCT},
        tab_color="good",
    ),
    SheetSpec(
        sheet="Anomalias SPC", analysis="demanda", table="anomalias",
        title="Dias fuera de control estadistico",
        subtitle="Limites robustos sobre mediana y MAD, en escala logaritmica y excluyendo domingos.",
        tab_color="warn",
    ),
    SheetSpec(
        sheet="Eventos Simultaneos", analysis="demanda", table="anomalias_dias",
        title="Fechas con quiebre simultaneo en varias sucursales",
        subtitle="Cuando tres o mas sucursales rompen el mismo dia, es un evento real y no ruido.",
        tab_color="warn",
    ),
    SheetSpec(
        sheet="Logistica SLA", analysis="logistica", table="sla",
        title="Nivel de servicio a grano factura",
        subtitle="Una factura de 40 lineas es UNA entrega. Solo se acusa a quien pasa el z-test contra la red.",
        formats={"OTIF (<=1 dia)": PCT, "Tasa Red": PCT, "Brecha vs Red": PCT,
                 "z": DEC2, "p-valor": DEC2, "Lead Medio (dias)": DEC2,
                 "Lead p50 (dias)": DEC1, "Lead p90 (dias)": DEC1},
        decorate=_dec_sla, tab_color="info",
    ),
    SheetSpec(
        sheet="Rechazos", analysis="logistica", table="rechazos",
        title="Rechazos de entrega",
        subtitle="Tasa y valor por sucursal, fletero, cliente, articulo, subcanal y dia de semana.",
        formats={"Tasa Rechazo": PCT2, "Tasa Red": PCT2, "Brecha vs Red": PCT2,
                 "z": DEC2, "p-valor": DEC2, "Share del Valor Rechazado": PCT,
                 "Pareto Acumulado": PCT},
        decorate=_dec_rechazos, tab_color="bad",
    ),
    SheetSpec(
        sheet="Devoluciones", analysis="logistica", table="devoluciones",
        title="Devoluciones sobre la venta propia de cada producto",
        subtitle="Medido contra lo que vende cada producto, no en absoluto: en absoluto solo rankea a los mas grandes.",
        formats={"Tasa Devolucion (bultos)": PCT2, "Tasa Devolucion (valor)": PCT2,
                 "z Robusto": DEC2},
        tab_color="bad",
    ),
    SheetSpec(
        sheet="Stock y Cobertura", analysis="logistica", table="stock",
        title="Cobertura de stock y capital inmovilizado",
        subtitle="Dias de cobertura contra la velocidad real de venta, por deposito y articulo.",
        formats={"Dias de Cobertura": DEC1, "Velocidad (bultos/dia)": DEC2},
        decorate=_dec_stock, tab_color="bad",
    ),
    SheetSpec(
        sheet="Rutas", analysis="logistica", table="rutas",
        title="Economia de rutas",
        subtitle="Drop size por ruta. La clave es (sucursal, ruta): el codigo de ruta NO es unico global.",
        formats={"Drop Size (bultos/visita)": DEC2, "Drop Mediano (bultos/visita)": DEC2,
                 "Lineas por Visita": DEC2},
        decorate=_dec_rutas, tab_color="warn",
    ),
]
