"""DataFrame -> formatted worksheet.

The house `ExcelWriter` writes a table starting at row 1 and sets every data cell
to bold, which is right for the operational reports but wrong here: these sheets
carry KPI cards and titles above the table, and a 15,000-row client list in bold
is unreadable. This module follows the `comparativo_salta` precedent of building
formatted analytical sheets on raw openpyxl.

Number formats are inferred from the column NAME so the analysis modules do not
have to carry presentation concerns, with explicit overrides always available.
"""
from __future__ import annotations

import re

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.services.inteligencia_comercial.excel_style import (
    BORDER_BOX,
    FMT_DATE,
    FMT_DEC1,
    FMT_DEC2,
    FMT_INT,
    FMT_KEY,
    FMT_MONEY,
    FMT_MONTH,
    FMT_PCT1,
    PALETTE,
    fill,
)

# Label used for the mandatory grand-total row every ranking table carries.
TOTAL_LABEL = "TOTAL GENERAL"

# Inference rules, first match wins. Matched against the lower-cased column name.
_FORMAT_RULES: tuple[tuple[str, str], ...] = (
    # Identifier-like columns must stay separator-free: a thousands separator
    # turns 1 into "1,00" and every downstream VLOOKUP silently stops matching.
    (r"^(id_|.*_id$|codigo|cod_)", FMT_KEY),
    (r"(^|_)(pct|porc|tasa|share|penetracion|participacion|retencion|margen_pct)", FMT_PCT1),
    (r"%", FMT_PCT1),
    (r"(neto|bruto|monto|importe|valor|pesos|ars|facturacion|descuento|exceso|oportunidad|capital|costo|precio|ticket)", FMT_MONEY),
    (r"(fecha|mes|periodo)", FMT_DATE),
    (r"(bultos|unidades|cantidad|clientes|skus?|lineas|entregas|visitas|dias_compra|frecuencia|n_|cant_)", FMT_INT),
    (r"(dias|lead|gap|recencia|cobertura|ratio|lift|soporte|confianza|conviccion|leverage|cv|z_|zscore|indice|mape)", FMT_DEC2),
    (r"(htl|hectolitros|volumen)", FMT_DEC1),
)


def infer_format(column: str) -> str:
    """Pick an Excel number format from a column name."""
    name = str(column).strip().lower()
    for pattern, fmt in _FORMAT_RULES:
        if re.search(pattern, name):
            return fmt
    return FMT_INT


def _is_total_row(value) -> bool:
    return isinstance(value, str) and value.strip().upper().startswith("TOTAL")


def write_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    row: int,
    col: int = 1,
    formats: dict[str, str] | None = None,
    max_rows: int | None = None,
    autofilter: bool = True,
    zebra: bool = True,
) -> dict:
    """Write a DataFrame as a formatted block.

    Args:
        row: 1-based row of the header.
        formats: explicit {column: number_format} overrides on top of inference.
        max_rows: cap the rows written. The caller is told what was dropped in
            the return value so it can say so on the sheet — a silently truncated
            table reads as a complete one.
        zebra: alternate row shading for legibility on long tables.

    Returns:
        {header_row, first_row, last_row, first_col, last_col, written, dropped,
         columns: {name: 1-based column index}}
    """
    if df is None or df.empty:
        note = ws.cell(row=row, column=col, value="Sin datos para el periodo analizado.")
        note.font = Font(italic=True, color=PALETTE["ink_soft"])
        return {
            "header_row": row, "first_row": row, "last_row": row,
            "first_col": col, "last_col": col, "written": 0, "dropped": 0, "columns": {},
        }

    formats = formats or {}
    total = len(df)
    if max_rows is None or total <= max_rows:
        body = df
        dropped = 0
    else:
        # La fila de total es la ultima, asi que un head() se la come justo en las
        # tablas mas largas — que son las que mas necesitan un total. Se la vuelve
        # a pegar, y sigue siendo el total del universo COMPLETO, no del recorte.
        body = df.head(max_rows)
        dropped = total - max_rows
        cola = df.tail(1)
        if _is_total_row(cola.iloc[0, 0]):
            body = pd.concat([body, cola], ignore_index=True)
            dropped -= 1

    columns = list(body.columns)
    col_index = {name: col + i for i, name in enumerate(columns)}
    resolved = {name: formats.get(name, infer_format(name)) for name in columns}

    for offset, name in enumerate(columns):
        cell = ws.cell(row=row, column=col + offset, value=str(name))
        cell.font = Font(bold=True, size=9, color=PALETTE["paper"])
        cell.fill = fill("brand")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_BOX
    ws.row_dimensions[row].height = 30

    zebra_fill = PatternFill(
        start_color=PALETTE["paper_alt"], end_color=PALETTE["paper_alt"], fill_type="solid"
    )
    total_fill = fill("ink")

    for r_offset, (_, record) in enumerate(body.iterrows(), start=1):
        target_row = row + r_offset
        is_total = _is_total_row(record.iloc[0])
        for c_offset, name in enumerate(columns):
            value = record[name]
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif hasattr(value, "item"):
                # numpy scalars must be unboxed or openpyxl writes them as strings
                value = value.item()

            cell = ws.cell(row=target_row, column=col + c_offset, value=value)
            cell.number_format = resolved[name]
            cell.alignment = Alignment(
                horizontal="left" if isinstance(value, str) else "right", vertical="center"
            )
            if is_total:
                cell.font = Font(bold=True, size=10, color=PALETTE["paper"])
                cell.fill = total_fill
            elif zebra and r_offset % 2 == 0:
                cell.fill = zebra_fill

    last_row = row + len(body)
    last_col = col + len(columns) - 1

    if autofilter and len(body) > 1:
        ws.auto_filter.ref = (
            f"{get_column_letter(col)}{row}:{get_column_letter(last_col)}{last_row}"
        )

    _fit_widths(ws, body, col)

    return {
        "header_row": row,
        "first_row": row + 1,
        "last_row": last_row,
        "first_col": col,
        "last_col": last_col,
        "written": len(body),
        "dropped": dropped,
        "columns": col_index,
    }


def _fit_widths(ws: Worksheet, df: pd.DataFrame, col: int, cap: float = 42.0) -> None:
    """Size columns to their content, sampling the head for long frames."""
    sample = df.head(400)
    for offset, name in enumerate(df.columns):
        header_len = max(len(part) for part in str(name).split()) if str(name) else 8
        series = sample[name]
        if pd.api.types.is_numeric_dtype(series):
            content_len = 13
        else:
            lengths = series.astype(str).str.len()
            content_len = int(lengths.max()) if len(lengths) and lengths.notna().any() else 10
        width = min(max(header_len + 3, content_len + 2, 9), cap)
        ws.column_dimensions[get_column_letter(col + offset)].width = width


def column_letter(block: dict, name: str) -> str:
    """Excel column letter for a named column of a written block."""
    return get_column_letter(block["columns"][name])


def data_range(block: dict, name: str, include_header: bool = False) -> str:
    """A1-style range for one column of a written block."""
    letter = column_letter(block, name)
    start = block["header_row"] if include_header else block["first_row"]
    return f"{letter}{start}:{letter}{block['last_row']}"


def add_total_row(
    df: pd.DataFrame,
    label_col: str | None = None,
    sum_cols: list[str] | None = None,
    mean_cols: list[str] | None = None,
    recompute: dict | None = None,
) -> pd.DataFrame:
    """Append the mandatory TOTAL GENERAL row.

    Sums are the default for measures; rates must be RECOMPUTED from their
    components, never averaged — the mean of per-branch discount rates is not
    the company discount rate unless every branch is the same size.

    Args:
        recompute: {column: callable(df) -> value} for ratio columns.
    """
    if df is None or df.empty:
        return df

    label_col = label_col or df.columns[0]
    numeric = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    sum_cols = numeric if sum_cols is None else sum_cols
    mean_cols = mean_cols or []
    recompute = recompute or {}

    total = {}
    for column in df.columns:
        if column == label_col:
            total[column] = TOTAL_LABEL
        elif column in recompute:
            total[column] = recompute[column](df)
        elif column in mean_cols:
            total[column] = df[column].mean()
        elif column in sum_cols:
            total[column] = df[column].sum()
        else:
            total[column] = None

    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)
