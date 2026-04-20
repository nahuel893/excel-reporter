"""
MisionImposibleService - Genera Excel con datos de cobertura para formulas manuales.

Crea un archivo Excel con hojas de datos crudos (cobertura por preventista y sucursal).
El usuario agrega formulas y vistas manualmente sobre estos datos.
"""

import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook

from config.settings import ZONAS_VIRTUALES
from src.core.data_loader import DataLoader
from src.core.excel_writer import ExcelWriter, SheetStyle, ColumnFormat, _write_sheet
from src.core.zonas import aplicar_zonas_virtuales
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)


SHEET_INFO = "INFO"

# Sheet names
SHEET_COB_PREV_GENERICO = "Cob Preventista Generico"
SHEET_COB_PREV_MARCA = "Cob Preventista Marca"
SHEET_COB_SUC_GENERICO = "Cob Sucursal Generico"
SHEET_COB_SUC_MARCA = "Cob Sucursal Marca"

# Column display names
_COB_PREV_GEN_COLUMNS = {
    "periodo": "Periodo",
    "sucursal": "Sucursal",
    "vendedor": "Vendedor",
    "generico": "Generico",
    "clientes_compradores": "Clientes Compradores",
    "volumen_total": "Volumen Total",
}

_COB_PREV_MARCA_COLUMNS = {
    "periodo": "Periodo",
    "sucursal": "Sucursal",
    "vendedor": "Vendedor",
    "marca": "Marca",
    "clientes_compradores": "Clientes Compradores",
    "volumen_total": "Volumen Total",
}

_COB_SUC_GEN_COLUMNS = {
    "periodo": "Periodo",
    "sucursal": "Sucursal",
    "generico": "Generico",
    "clientes_compradores": "Clientes Compradores",
    "volumen_total": "Volumen Total",
}

_COB_SUC_MARCA_COLUMNS = {
    "periodo": "Periodo",
    "sucursal": "Sucursal",
    "marca": "Marca",
    "clientes_compradores": "Clientes Compradores",
    "volumen_total": "Volumen Total",
}

_DATA_STYLE = SheetStyle(
    column_formats={
        "Clientes Compradores": ColumnFormat(number_format="#,##0", width=20),
        "Volumen Total": ColumnFormat(number_format="#,##0.00", width=15),
    },
    as_table=True,
    table_style="TableStyleMedium9",
)


def _fechas_a_periodos(fecha_desde: str, fecha_hasta: str) -> list[str]:
    """Convierte un rango de fechas en lista de primeros dias de mes cubiertos."""
    desde = pd.to_datetime(fecha_desde).replace(day=1)
    hasta = pd.to_datetime(fecha_hasta)
    periodos = pd.date_range(desde, hasta, freq="MS")
    return [p.strftime("%Y-%m-%d") for p in periodos]


def _preparar_hoja(
    df: pd.DataFrame | None,
    sort_cols: list[str],
    rename_map: dict[str, str],
) -> pd.DataFrame | None:
    """Ordena, selecciona y renombra columnas para una hoja de datos."""
    if df is None or df.empty:
        return None
    df = df.sort_values(sort_cols).reset_index(drop=True)
    df = df[list(rename_map.keys())]
    return df.rename(columns=rename_map)


def _nombre_tabla(cat_name: str) -> str:
    """Genera nombre de tabla Excel valido desde el nombre de categoria.

    Excel tables: no espacios, no caracteres especiales, max 255 chars.
    Ej: 'SCHNEIDER 710' → 'Tbl_SCHNEIDER_710'
    """
    import re
    clean = re.sub(r"[^a-zA-Z0-9_]", "_", cat_name)
    return f"Tbl_{clean}"[:255]


def _aplicar_zonas_categoria(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica zonas virtuales y prepara columnas para pivot de categoria."""
    df = df.copy()
    for zona_nombre, zona_config in ZONAS_VIRTUALES.items():
        mask = (df["sucursal"] == zona_config["sucursal_real"]) & (
            df["id_ruta"].isin(zona_config["rutas"])
        )
        df.loc[mask, "sucursal"] = zona_nombre

    df = df.rename(columns={"sucursal": "zona_virtual"})
    df["articulo_desc"] = df["id_articulo"].astype(str) + " - " + df["des_articulo"]
    return df


def _pivotear_categoria(df: pd.DataFrame) -> pd.DataFrame:
    """Crea pivot table con marca/articulo en columnas y totales por marca."""
    idx_cols = [
        "zona_virtual",
        "id_ruta",
        "id_vendedor",
        "vendedor",
        "id_cliente",
        "cliente",
    ]
    df_pivot = df.pivot_table(
        index=idx_cols,
        columns=["marca", "articulo_desc"],
        values="cantidad",
        aggfunc="sum",
        fill_value=0,
    )

    # Subtotales por marca
    totales_marca = df_pivot.T.groupby(level="marca").sum().T
    totales_marca.columns = pd.MultiIndex.from_product(
        [totales_marca.columns, ["Total"]]
    )

    df_final = pd.concat([df_pivot, totales_marca], axis=1)
    df_final = df_final.sort_index(axis=1, level=0)

    # Aplanar MultiIndex a "Marca - Articulo" para Excel (sin \n para compatibilidad con tablas)
    df_final.columns = [f"{col[0]} - {col[1]}" for col in df_final.columns]
    df_final = df_final.reset_index()

    return df_final


def _parse_operador(formula_str: str) -> tuple[str, float]:
    """Parsea '>= 0.25' en ('>=', 0.25)."""
    formula_str = formula_str.strip()
    for op in (">=", "<=", ">", "<", "==", "!="):
        if formula_str.startswith(op):
            return op, float(formula_str[len(op):].strip())
    raise ValueError(f"Formula no reconocida: {formula_str}")


def _evaluar_condicion(valor, op: str, umbral: float) -> bool:
    """Evalua una condicion: valor op umbral."""
    if op == ">=":
        return valor >= umbral
    if op == ">":
        return valor > umbral
    if op == "<=":
        return valor <= umbral
    if op == "<":
        return valor < umbral
    if op == "==":
        return valor == umbral
    if op == "!=":
        return valor != umbral
    return False


def _aplicar_formula(
    df_pivot: pd.DataFrame,
    df_raw: pd.DataFrame,
    formula: str | dict | None,
) -> pd.DataFrame:
    """
    Agrega columna 'CUMPLE' al pivot basandose en la formula de la categoria.

    Tipos de formula:
    - ">= X": total general (suma de todos los Total de marca) cumple condicion
    - "todas_marcas >= X": CADA Total de marca cumple la condicion
    - {"por_calibre": {"grupo": {"calibres": [...], "criterio": ">= X"}, ...}}:
      total por grupo de calibres cumple su criterio individual
    """
    if formula is None:
        return df_pivot

    idx_cols = [
        "zona_virtual", "id_ruta", "id_vendedor", "vendedor",
        "id_cliente", "cliente",
    ]
    total_cols = [c for c in df_pivot.columns if c.endswith(" - Total")]

    if isinstance(formula, str) and formula.startswith("todas_marcas"):
        # "todas_marcas >= 0.25" → cada marca Total >= 0.25
        criterio_str = formula.replace("todas_marcas", "").strip()
        op, umbral = _parse_operador(criterio_str)
        df_pivot["CUMPLE"] = df_pivot[total_cols].apply(
            lambda row: int(all(_evaluar_condicion(v, op, umbral) for v in row)),
            axis=1,
        )

    elif isinstance(formula, str):
        # ">= 0.25" → total general cumple condicion
        op, umbral = _parse_operador(formula)
        gran_total = df_pivot[total_cols].sum(axis=1)
        df_pivot["CUMPLE"] = gran_total.apply(
            lambda v: int(_evaluar_condicion(v, op, umbral))
        )

    elif isinstance(formula, dict) and "por_calibre" in formula:
        # Por calibre: necesitamos calcular totales por grupo de calibres
        # desde df_raw que tiene la columna 'calibre'
        grupos = formula["por_calibre"]

        # Construir totales por calibre-grupo y cliente
        totales_por_cliente = {}
        for grupo_nombre, grupo_cfg in grupos.items():
            calibres = grupo_cfg["calibres"]
            df_grupo = df_raw[df_raw["calibre"].isin(calibres)]
            if df_grupo.empty:
                continue
            totales = df_grupo.groupby(idx_cols)["cantidad"].sum()
            totales_por_cliente[grupo_nombre] = totales

        # Evaluar: TODOS los grupos deben cumplir su criterio
        def evaluar_fila(row):
            key = tuple(row[c] for c in idx_cols)
            for grupo_nombre, grupo_cfg in grupos.items():
                op, umbral = _parse_operador(grupo_cfg["criterio"])
                total_grupo = totales_por_cliente.get(grupo_nombre)
                if total_grupo is None:
                    return 0
                val = total_grupo.get(key, 0)
                if not _evaluar_condicion(val, op, umbral):
                    return 0
            return 1

        df_pivot["CUMPLE"] = df_pivot.apply(evaluar_fila, axis=1)

    return df_pivot


def _escribir_hoja_info(
    wb,
    categorias_con_cumple: list[tuple[str, str]],
    sucursales: list[str],
):
    """
    Escribe/actualiza la hoja INFO con formulas SUMIFS para cada categoria.

    Args:
        wb: Workbook de openpyxl
        categorias_con_cumple: lista de (nombre_categoria, nombre_tabla)
        sucursales: lista de nombres de sucursal/zona_virtual
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Borrar hoja INFO si existe (es gestionada)
    if SHEET_INFO in wb.sheetnames:
        del wb[SHEET_INFO]

    ws = wb.create_sheet(title=SHEET_INFO, index=0)

    if not categorias_con_cumple or not sucursales:
        ws["A1"] = "Sin datos para generar resumen"
        return

    # Layout: col A vacia, col B = Sucursal, luego por cada categoria: Real
    # Row 1: headers de categoria
    # Row 2: sub-header "Clientes"
    # Row 3+: sucursales con formulas

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Encabezado Sucursal
    ws.cell(1, 2, "Sucursal").font = header_font_white
    ws.cell(1, 2).fill = header_fill
    ws.cell(1, 2).alignment = Alignment(horizontal="center")
    ws.column_dimensions["B"].width = 25

    # Escribir headers de categorias
    col = 3
    for cat_name, table_name in categorias_con_cumple:
        cell = ws.cell(1, col, cat_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18
        col += 1

    # Escribir sucursales y formulas
    for row_idx, sucursal in enumerate(sucursales, start=2):
        ws.cell(row_idx, 2, sucursal)

        col = 3
        for cat_name, table_name in categorias_con_cumple:
            # =SUMIFS(Tbl_XXX[CUMPLE], Tbl_XXX[zona_virtual], $B{row})
            formula = f'=SUMIFS({table_name}[CUMPLE],{table_name}[zona_virtual],$B{row_idx})'
            ws.cell(row_idx, col, formula)
            ws.cell(row_idx, col).alignment = Alignment(horizontal="center")
            col += 1

    # Fila de totales al final
    total_row = len(sucursales) + 2
    ws.cell(total_row, 2, "TOTAL").font = Font(bold=True)
    col = 3
    for _ in categorias_con_cumple:
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'
        cell = ws.cell(total_row, col, formula)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        col += 1


@dataclass
class CategoriaConfig:
    """Configuracion de una categoria individual."""

    articulos: list[int]
    formula: str | dict | None = None

    @staticmethod
    def from_raw(raw) -> "CategoriaConfig":
        """Parsea desde JSON: acepta lista de IDs o dict con articulos+formula."""
        if isinstance(raw, list):
            return CategoriaConfig(articulos=raw)
        return CategoriaConfig(
            articulos=raw["articulos"],
            formula=raw.get("formula"),
        )


@dataclass
class MisionImposibleConfig:
    """Configuracion para el informe Mision Imposible."""

    fecha_desde: str
    fecha_hasta: str
    genericos: list[str] | None = None
    categorias: dict[str, any] | None = None
    nombre_archivo: str | None = None

    def get_categorias(self) -> dict[str, CategoriaConfig]:
        """Retorna categorias parseadas como CategoriaConfig."""
        if not self.categorias:
            return {}
        return {
            name: CategoriaConfig.from_raw(raw)
            for name, raw in self.categorias.items()
        }


@dataclass
class MisionImposibleResult:
    """Resultado de la generacion del informe Mision Imposible."""

    ruta_archivo: Path
    registros_procesados: int
    sucursales: int
    genericos_incluidos: list[str]
    hojas: list[str]


class MisionImposibleService(BaseService):
    """Servicio para generar el informe Mision Imposible con datos de cobertura."""

    SERVICE_SLUG = "mision-imposible"
    GRANULARITY = "month"

    # Prefijos/nombres de hojas gestionadas por el servicio
    _MANAGED_PREFIXES = ("Cob ", "Cat ", SHEET_INFO)

    def generar_reporte(self, config: MisionImposibleConfig) -> MisionImposibleResult:
        """Genera el Excel con hojas de cobertura y categorias.

        Si el archivo ya existe, actualiza solo las hojas gestionadas
        (Cob * y Cat *) preservando las hojas manuales del usuario.
        """
        periodos = _fechas_a_periodos(config.fecha_desde, config.fecha_hasta)

        # Fetch all cobertura data
        df_prev_gen = self._fetch_preventista_generico(periodos)
        df_prev_marca = self._fetch_preventista_marca(periodos)
        df_suc_gen = self._fetch_sucursal_generico(periodos)
        df_suc_marca = self._fetch_sucursal_marca(periodos)

        # Filter by genericos if specified
        if config.genericos:
            df_prev_gen = self._filtrar_genericos(
                df_prev_gen, config.genericos, "generico"
            )
            df_suc_gen = self._filtrar_genericos(
                df_suc_gen, config.genericos, "generico"
            )

        nombre = config.nombre_archivo or f"Mision Imposible {config.fecha_hasta}"
        output_dir = self._output_dir(config.fecha_desde)
        output_dir.mkdir(parents=True, exist_ok=True)
        ruta_archivo = output_dir / f"{nombre}.xlsx"

        # Si el archivo existe, cargarlo y borrar solo hojas gestionadas
        if ruta_archivo.exists():
            wb = load_workbook(str(ruta_archivo))
            hojas_borradas = []
            for sheet_name in wb.sheetnames[:]:
                if sheet_name.startswith(("Cob ", "Cat ")) or sheet_name == SHEET_INFO:
                    del wb[sheet_name]
                    hojas_borradas.append(sheet_name)
            if hojas_borradas:
                logger.info("Hojas actualizadas: %s", ", ".join(hojas_borradas))
            _new_file = False
        else:
            from openpyxl import Workbook
            wb = Workbook()
            _new_file = True

        hojas = []
        total = 0

        # Escribir hojas base de cobertura
        sheets = [
            (
                df_prev_gen,
                ["sucursal", "vendedor", "generico"],
                _COB_PREV_GEN_COLUMNS,
                SHEET_COB_PREV_GENERICO,
            ),
            (
                df_prev_marca,
                ["sucursal", "vendedor", "marca"],
                _COB_PREV_MARCA_COLUMNS,
                SHEET_COB_PREV_MARCA,
            ),
            (
                df_suc_gen,
                ["sucursal", "generico"],
                _COB_SUC_GEN_COLUMNS,
                SHEET_COB_SUC_GENERICO,
            ),
            (
                df_suc_marca,
                ["sucursal", "marca"],
                _COB_SUC_MARCA_COLUMNS,
                SHEET_COB_SUC_MARCA,
            ),
        ]

        def _add_ws(wb, name, _new_file_holder=[_new_file]):
            """Crea hoja; en archivo nuevo la primera reusa la hoja default."""
            if _new_file_holder[0]:
                ws = wb.active
                ws.title = name
                _new_file_holder[0] = False
            else:
                ws = wb.create_sheet(title=name)
            return ws

        for df, sort_cols, rename_map, sheet_name in sheets:
            df_sheet = _preparar_hoja(df, sort_cols, rename_map)
            if df_sheet is not None:
                ws = _add_ws(wb, sheet_name)
                _write_sheet(ws, df_sheet, _DATA_STYLE)
                hojas.append(sheet_name)
                total += len(df_sheet)

        # Escribir hojas de categorias
        categorias_con_cumple: list[tuple[str, str]] = []  # (cat_name, table_name)
        all_sucursales: set[str] = set()

        categorias = config.get_categorias()
        if categorias:
            for cat_name, cat_cfg in categorias.items():
                try:
                    df_cat = self._procesar_categoria(
                        config.fecha_desde, config.fecha_hasta,
                        cat_cfg.articulos, cat_cfg.formula,
                    )
                except Exception as exc:
                    logger.error("Error procesando categoria '%s': %s", cat_name, exc)
                    continue
                if df_cat is not None and not df_cat.empty:
                    sheet_name = f"Cat {cat_name}"[:31]
                    table_name = _nombre_tabla(cat_name)
                    ws = _add_ws(wb, sheet_name)
                    _write_sheet(ws, df_cat, SheetStyle(
                        as_table=True,
                        table_style="TableStyleMedium9",
                        table_name=table_name,
                    ))
                    hojas.append(sheet_name)
                    total += len(df_cat)

                    # Rastrear categorias con formula para hoja INFO
                    if "CUMPLE" in df_cat.columns:
                        categorias_con_cumple.append((cat_name, table_name))
                        all_sucursales.update(df_cat["zona_virtual"].unique())

        # Generar hoja INFO con formulas SUMIFS
        if categorias_con_cumple:
            sucursales_ordenadas = sorted(all_sucursales)
            _escribir_hoja_info(wb, categorias_con_cumple, sucursales_ordenadas)
            hojas.insert(0, SHEET_INFO)

        wb.save(str(ruta_archivo))

        # Collect genericos from data
        genericos = []
        if df_prev_gen is not None and not df_prev_gen.empty:
            genericos = sorted(df_prev_gen["generico"].unique().tolist())

        # Count unique sucursales
        sucursales = 0
        if df_prev_gen is not None and not df_prev_gen.empty:
            sucursales = df_prev_gen["sucursal"].nunique()

        return MisionImposibleResult(
            ruta_archivo=ruta_archivo,
            registros_procesados=total,
            sucursales=sucursales,
            genericos_incluidos=genericos,
            hojas=hojas,
        )

    def _fetch_preventista_generico(self, periodos: list[str]) -> pd.DataFrame | None:
        """Fetch cobertura preventista generico with zonas virtuales."""
        try:
            df = self.data_loader.get_cobertura_preventista_generico(periodos=periodos)
            if not df.empty:
                return aplicar_zonas_virtuales(df)
        except Exception:
            pass
        return None

    def _fetch_preventista_marca(self, periodos: list[str]) -> pd.DataFrame | None:
        """Fetch cobertura preventista marca with zonas virtuales."""
        try:
            df = self.data_loader.get_cobertura_preventista_marca(periodos=periodos)
            if not df.empty:
                return aplicar_zonas_virtuales(df)
        except Exception:
            pass
        return None

    def _fetch_sucursal_generico(self, periodos: list[str]) -> pd.DataFrame | None:
        """Fetch cobertura sucursal generico."""
        try:
            df = self.data_loader.get_cobertura_sucursal_generico(periodos=periodos)
            if not df.empty:
                return df
        except Exception:
            pass
        return None

    def _fetch_sucursal_marca(self, periodos: list[str]) -> pd.DataFrame | None:
        """Fetch cobertura sucursal marca."""
        try:
            df = self.data_loader.get_cobertura_sucursal_marca(periodos=periodos)
            if not df.empty:
                return df
        except Exception:
            pass
        return None

    def _procesar_categoria(
        self, fecha_desde: str, fecha_hasta: str,
        art_ids: list[int], formula: str | dict | None = None,
    ) -> pd.DataFrame | None:
        """Obtiene y pivotea las ventas de una categoria, aplicando formula si existe."""
        df = self.data_loader.get_ventas_mision_imposible_categorias(
            fecha_desde, fecha_hasta, art_ids
        )
        if df.empty:
            return None

        df = _aplicar_zonas_categoria(df)
        df_pivot = _pivotear_categoria(df)
        return _aplicar_formula(df_pivot, df, formula)

    @staticmethod
    def _filtrar_genericos(
        df: pd.DataFrame | None, genericos: list[str], col: str
    ) -> pd.DataFrame | None:
        """Filter DataFrame by genericos list."""
        if df is None or df.empty:
            return df
        return df[df[col].isin(genericos)]
