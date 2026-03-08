"""
MisionPosibleService - Servicio para generacion de reportes de cobertura.

Genera un reporte Excel con una hoja por marca, mostrando tablas de cobertura
por sucursal y por vendedor con objetivos, faltantes y porcentajes.
"""
from dataclasses import dataclass, field
from datetime import date
from math import ceil
from pathlib import Path

import pandas as pd

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from openpyxl import Workbook

from config.settings import DATA_OUTPUT
from src.core.data_loader import DataLoader
from src.core.zonas import aplicar_zonas_virtuales, expandir_sucursales
from src.services.base_service import BaseService
from src.services.mision_posible.processor import (
    procesar_cobertura_sucursal,
    procesar_cobertura_vendedor,
)


@dataclass
class MisionPosibleConfig:
    """Configuracion para el reporte Mision Posible."""
    periodo: str                                # "YYYY-MM-DD", primer dia del mes
    marcas: list[str]                           # ["Imperial", "Levite", "Villa del Sur"]
    objetivos: dict[str, int] = field(default_factory=dict)
    porcentajes_sucursal: dict[str, float] = field(default_factory=dict)
    nombre_archivo: str | None = None
    supervisores: dict[str, list[str]] | None = None


@dataclass
class MisionPosibleResult:
    """Resultado de la generacion de un reporte Mision Posible."""
    ruta_archivos: list[Path]
    marcas_incluidas: list[str]
    hojas: list[str]
    supervisor: str | None = None


def _normalizar_periodo(periodo: str) -> str:
    """Normaliza periodo al primer dia del mes. Imprime warning si difiere."""
    d = pd.to_datetime(periodo)
    normalizado = d.replace(day=1).strftime("%Y-%m-%d")
    if normalizado != periodo:
        print(f"⚠ Periodo normalizado de {periodo} a {normalizado}")
    return normalizado


def _nombre_reporte(periodo: str, supervisor: str | None = None) -> str:
    """Genera nombre de archivo: 'Mision Posible [supervisor] MM-YYYY'."""
    d = pd.to_datetime(periodo)
    mm_yyyy = d.strftime("%m-%Y")
    if supervisor:
        return f"Mision Posible {supervisor} {mm_yyyy}"
    return f"Mision Posible {mm_yyyy}"


_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Layout constants for the "Sucursales" sheet grid
MARCAS_POR_FILA = 4
COLS_POR_TABLA = 5       # Sucursal, Cobertura, Objetivo, Faltante, %
COL_SEPARADOR_ANCHO = 2  # Excel units
FILA_SEPARADOR_ALTO = 6  # points
ANCHOS_TABLA = [25, 12, 12, 12, 10]

# Header/title style constants
_HEADER_FILL = PatternFill(start_color="A92C1F", end_color="A92C1F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITULO_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_DATA_FONT = Font(bold=True)


def _aplicar_formato_condicional(ws, col_letter: str, first_row: int, last_row: int):
    """Aplica formato condicional rojo/amarillo/verde a la columna de porcentaje."""
    rango = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="greaterThanOrEqual", formula=["0.80"], fill=_FILL_GREEN)
    )
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="lessThan", formula=["0.40"], fill=_FILL_RED)
    )
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="between", formula=["0.40", "0.799"], fill=_FILL_YELLOW)
    )


class MisionPosibleService(BaseService):
    """Servicio para generar reportes de cobertura Mision Posible."""

    def generar_reporte(self, config: MisionPosibleConfig) -> MisionPosibleResult:
        """Genera un unico archivo con hojas 'Sucursales' y 'Por Vendedor'."""
        if not config.marcas:
            raise ValueError("La lista de marcas no puede estar vacia.")

        periodo = _normalizar_periodo(config.periodo)
        df_cob, ultima_fecha = self._fetch_data(periodo)

        nombre = config.nombre_archivo or _nombre_reporte(periodo)

        wb = Workbook()
        ws_suc = wb.active
        ws_suc.title = "Sucursales"

        tablas_suc = []
        tablas_vend = []
        for marca in config.marcas:
            objetivo_total = config.objetivos.get(marca)
            df_suc = procesar_cobertura_sucursal(
                df_cob, marca, objetivo_total, config.porcentajes_sucursal
            )
            df_vend = procesar_cobertura_vendedor(
                df_cob, marca, objetivo_total, config.porcentajes_sucursal
            )
            tablas_suc.append((marca, df_suc))
            tablas_vend.append((marca, df_vend))

        self._escribir_hoja_sucursales(ws_suc, tablas_suc, ultima_fecha)

        ws_vend = wb.create_sheet("Por Vendedor")
        self._escribir_hoja_vendedores(ws_vend, tablas_vend, ultima_fecha)

        DATA_OUTPUT.mkdir(parents=True, exist_ok=True)
        ruta = DATA_OUTPUT / f"{nombre}.xlsx"
        wb.save(ruta)

        return MisionPosibleResult(
            ruta_archivos=[ruta],
            marcas_incluidas=list(config.marcas),
            hojas=["Sucursales", "Por Vendedor"],
        )

    def generar_reporte_supervisores(
        self,
        config: MisionPosibleConfig,
        supervisores: dict[str, list[str]],
    ) -> list[MisionPosibleResult]:
        """Genera un archivo por supervisor con una sola consulta a BD."""
        if not config.marcas:
            raise ValueError("La lista de marcas no puede estar vacia.")

        periodo = _normalizar_periodo(config.periodo)
        df_cob, ultima_fecha = self._fetch_data(periodo)

        DATA_OUTPUT.mkdir(parents=True, exist_ok=True)
        results = []
        for supervisor, sucursales in supervisores.items():
            sucursales_exp = expandir_sucursales(sucursales)
            df_cob_sup = df_cob[df_cob["sucursal"].isin(sucursales_exp)] if not df_cob.empty else df_cob

            nombre = config.nombre_archivo or _nombre_reporte(periodo, supervisor)

            wb = Workbook()
            ws_suc = wb.active
            ws_suc.title = "Sucursales"

            tablas_suc = []
            tablas_vend = []
            for marca in config.marcas:
                objetivo_total = config.objetivos.get(marca)
                df_suc = procesar_cobertura_sucursal(
                    df_cob_sup, marca, objetivo_total, config.porcentajes_sucursal
                )
                df_vend = procesar_cobertura_vendedor(
                    df_cob_sup, marca, objetivo_total, config.porcentajes_sucursal
                )
                tablas_suc.append((marca, df_suc))
                tablas_vend.append((marca, df_vend))

            self._escribir_hoja_sucursales(ws_suc, tablas_suc, ultima_fecha)

            ws_vend = wb.create_sheet("Por Vendedor")
            self._escribir_hoja_vendedores(ws_vend, tablas_vend, ultima_fecha)

            ruta = DATA_OUTPUT / f"{nombre}.xlsx"
            wb.save(ruta)

            results.append(MisionPosibleResult(
                ruta_archivos=[ruta],
                marcas_incluidas=list(config.marcas),
                hojas=["Sucursales", "Por Vendedor"],
                supervisor=supervisor,
            ))

        return results

    def _escribir_hoja_sucursales(
        self,
        ws,
        tablas: list[tuple[str, pd.DataFrame]],
        ultima_fecha: date | None,
    ) -> None:
        """Escribe la grilla de tablas de sucursales en la hoja 'Sucursales'."""
        # --- Summary row ---
        if ultima_fecha is not None:
            cell_label = ws.cell(row=1, column=1, value="Ult. Actualizacion")
            cell_label.font = Font(bold=True)
            ws.cell(row=1, column=2, value=ultima_fecha.strftime("%d/%m/%Y"))
            fila_inicio_base = 3
        else:
            fila_inicio_base = 1

        # --- Column widths (applied once for all groups) ---
        for col_grupo in range(MARCAS_POR_FILA):
            col_base = 1 + col_grupo * (COLS_POR_TABLA + 1)
            for offset, ancho in enumerate(ANCHOS_TABLA):
                ws.column_dimensions[get_column_letter(col_base + offset)].width = ancho
            # separator column
            ws.column_dimensions[get_column_letter(col_base + COLS_POR_TABLA)].width = COL_SEPARADOR_ANCHO

        # --- Calculate fila_inicio per group ---
        num_grupos = ceil(len(tablas) / MARCAS_POR_FILA)

        # max_filas_en_grupo[g] = titulo(1) + encabezado(1) + max(len(df_suc)) in group g
        max_filas_en_grupo: list[int] = []
        for g in range(num_grupos):
            tablas_grupo = tablas[g * MARCAS_POR_FILA: (g + 1) * MARCAS_POR_FILA]
            max_datos = max(len(df_suc) for _, df_suc in tablas_grupo)
            max_filas_en_grupo.append(2 + max_datos)  # titulo + encabezado + datos

        fila_inicio_grupo: list[int] = [0] * num_grupos
        fila_inicio_grupo[0] = fila_inicio_base
        for g in range(1, num_grupos):
            fila_inicio_grupo[g] = fila_inicio_grupo[g - 1] + max_filas_en_grupo[g - 1] + 1

        # --- Separator row heights (only between groups, not after the last) ---
        for g in range(num_grupos - 1):
            fila_sep = fila_inicio_grupo[g + 1] - 1
            ws.row_dimensions[fila_sep].height = FILA_SEPARADOR_ALTO

        # --- Write each tabla ---
        headers = ["Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]
        for i, (marca, df_suc) in enumerate(tablas):
            fila_grupo = i // MARCAS_POR_FILA
            col_grupo = i % MARCAS_POR_FILA
            col_inicio = 1 + col_grupo * (COLS_POR_TABLA + 1)
            fila_inicio = fila_inicio_grupo[fila_grupo]

            # Title row (merged across 5 columns)
            ws.merge_cells(
                start_row=fila_inicio, start_column=col_inicio,
                end_row=fila_inicio, end_column=col_inicio + COLS_POR_TABLA - 1,
            )
            cell_titulo = ws.cell(row=fila_inicio, column=col_inicio, value=marca)
            cell_titulo.fill = _HEADER_FILL
            cell_titulo.font = _HEADER_FONT
            cell_titulo.alignment = _TITULO_ALIGNMENT

            # Header row
            for offset, header in enumerate(headers):
                cell_h = ws.cell(row=fila_inicio + 1, column=col_inicio + offset, value=header)
                cell_h.fill = _HEADER_FILL
                cell_h.font = _HEADER_FONT
                cell_h.alignment = _HEADER_ALIGNMENT

            # Data rows
            col_names = list(df_suc.columns)
            for row_idx, row_data in enumerate(df_suc.itertuples(index=False)):
                fila_dato = fila_inicio + 2 + row_idx
                for col_offset, value in enumerate(row_data):
                    cell = ws.cell(row=fila_dato, column=col_inicio + col_offset, value=value)
                    cell.font = _DATA_FONT
                    # Number formats
                    col_name = col_names[col_offset] if col_offset < len(col_names) else ""
                    if col_name == "%":
                        cell.number_format = "0.00%"
                    elif col_name in ("Cobertura", "Objetivo", "Faltante"):
                        cell.number_format = "#,##0"

            # Conditional formatting on % column
            col_pct_letter = get_column_letter(col_inicio + 4)
            first_row = fila_inicio + 2
            last_row = fila_inicio + 1 + len(df_suc)
            if len(df_suc) > 0:
                _aplicar_formato_condicional(ws, col_pct_letter, first_row, last_row)

    def _escribir_hoja_vendedores(
        self,
        ws,
        tablas: list[tuple[str, pd.DataFrame]],
        ultima_fecha: date | None,
    ) -> None:
        """Escribe los bloques de vendedores apilados en la hoja 'Por Vendedor'."""
        col_widths = [15, 25, 12, 12, 12, 10]
        col_headers = ["Vendedor", "Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]
        num_cols = len(col_headers)

        # Apply column widths
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        current_row = 1

        # Write Ult. Actualizacion in row 1 if available; data starts at row 3
        if ultima_fecha is not None:
            label_cell = ws.cell(row=1, column=1, value="Ult. Actualizacion")
            label_cell.font = _DATA_FONT
            ws.cell(row=1, column=2, value=ultima_fecha.strftime("%d/%m/%Y"))
            current_row = 3

        for marca, df_vend in tablas:
            # Title row: merge across all 6 columns, burdeo fill, white bold text
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=num_cols,
            )
            title_cell = ws.cell(row=current_row, column=1, value=marca)
            title_cell.fill = _HEADER_FILL
            title_cell.font = _HEADER_FONT
            title_cell.alignment = _TITULO_ALIGNMENT
            current_row += 1

            # Header row
            for col_idx, header in enumerate(col_headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _HEADER_ALIGNMENT
            current_row += 1

            # Data rows
            first_data_row = current_row
            for _, row in df_vend.iterrows():
                values = list(row)
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    # cols 3-5 = Cobertura/Objetivo/Faltante, col 6 = %
                    if col_idx in (3, 4, 5):
                        cell.number_format = "#,##0"
                    elif col_idx == 6:
                        cell.number_format = "0.00%"
                current_row += 1
            last_data_row = current_row - 1

            # Conditional formatting on % column (col 6) only if there are data rows
            if len(df_vend) > 0:
                col_pct_letter = get_column_letter(6)
                _aplicar_formato_condicional(ws, col_pct_letter, first_data_row, last_data_row)

            # Empty separator row between marca blocks
            current_row += 1

    def _fetch_data(self, periodo: str) -> tuple[pd.DataFrame, date | None]:
        """Fetches cobertura data and ultima fecha venta.

        Returns:
            Tuple of (df_cob post zonas virtuales, ultima_fecha_venta).
        """
        df_cob = pd.DataFrame()
        try:
            df_cob_raw = self.data_loader.get_cobertura_preventista_marca(
                periodos=[periodo]
            )
            if not df_cob_raw.empty:
                df_cob = aplicar_zonas_virtuales(df_cob_raw)
        except Exception:
            pass

        ultima_fecha = None
        try:
            ultima_fecha = self.data_loader.get_ultima_fecha_venta()
        except Exception:
            pass

        return df_cob, ultima_fecha
