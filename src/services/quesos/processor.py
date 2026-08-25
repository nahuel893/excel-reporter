"""Aritmetica del informe de quesos. Sin base y sin Excel.

Tres medidas por mes, y cada una se agrega distinto:

- **Bultos**: se suman.
- **Kg**: se suman, pero salen de multiplicar por un factor POR ARTICULO, asi
  que hay que convertir antes de agregar, nunca despues.
- **Cobertura**: NO se suma. El total del anio es el conteo de clientes
  distintos sobre el anio entero, no la suma de los doce meses.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .constants import COL_CODIGO, COL_PESO_KG, MESES_CORTOS

COLUMNAS_VENTAS = ["mes", "id_articulo", "id_cliente", "id_sucursal", "bultos"]
CLAVE_CLIENTE = ["id_cliente", "id_sucursal"]


def leer_factores(ruta: str | Path) -> dict[int, float]:
    """id_articulo -> kg por unidad, desde el xlsx exportado del proveedor.

    Se toleran encabezados y filas sueltas: solo entran las filas donde las dos
    celdas son numeros. El archivo trae los encabezados repetidos en el medio.

    Raises:
        FileNotFoundError: si el archivo no esta. Sin factores no hay kg, y un
            informe con la columna en cero se lee como "no vendimos".
        ValueError: si no se pudo leer ningun factor.
    """
    ruta = Path(ruta)
    if not ruta.exists():
        raise FileNotFoundError(f"falta el archivo de factores de quesos: {ruta}")

    ws = load_workbook(ruta, data_only=True).active
    factores: dict[int, float] = {}
    for r in range(1, ws.max_row + 1):
        cod, peso = ws.cell(r, COL_CODIGO).value, ws.cell(r, COL_PESO_KG).value
        if isinstance(cod, (int, float)) and isinstance(peso, (int, float)):
            factores[int(cod)] = float(peso)
    if not factores:
        raise ValueError(f"{ruta}: no se leyo ningun factor de conversion")
    return factores


def articulos_sin_factor(ventas: pd.DataFrame, factores: dict[int, float]) -> list[int]:
    """Articulos con venta que no tienen kg. Se reportan, nunca se ocultan.

    Un articulo nuevo sin factor suma bultos y no suma kg, y la fila queda
    coherente a la vista pero con los kilos cortos. Paso de verdad: el archivo
    de la hoja 'queso' del avance branca se quedo sin tres articulos y desde
    mayo-2026 los kg venian por debajo.
    """
    if ventas.empty:
        return []
    con_venta = ventas.groupby("id_articulo")["bultos"].sum()
    return sorted(int(a) for a in con_venta[con_venta != 0].index if int(a) not in factores)


def _cobertura(ventas: pd.DataFrame, umbral: float = 0.0) -> int:
    """Clientes distintos con neto > umbral. Agrupa ANTES de filtrar."""
    if ventas.empty:
        return 0
    neto = ventas.groupby(CLAVE_CLIENTE, sort=False)["bultos"].sum()
    return int((neto > umbral).sum())


def construir_anio(
    ventas: pd.DataFrame,
    factores: dict[int, float],
    anio: int,
    umbral: float = 0.0,
) -> pd.DataFrame:
    """Una fila por medida y una columna por mes, mas el total del anio.

    Returns:
        DataFrame indexado por medida (``Bultos``/``Kg``/``Coberturas``) con
        columnas ``ene``..``dic`` y ``TOTAL``.
    """
    del_anio = ventas[ventas["mes"].str.startswith(str(anio))] if not ventas.empty else ventas

    filas: dict[str, list[float]] = {"Bultos": [], "Kg": [], "Coberturas": []}
    for i, _ in enumerate(MESES_CORTOS, start=1):
        mes = f"{anio}-{i:02d}"
        del_mes = del_anio[del_anio["mes"] == mes] if not del_anio.empty else del_anio
        filas["Bultos"].append(float(del_mes["bultos"].sum()) if not del_mes.empty else 0.0)
        # El factor es POR ARTICULO: se convierte fila por fila y despues se suma.
        filas["Kg"].append(
            float((del_mes["bultos"] * del_mes["id_articulo"].map(factores).fillna(0.0)).sum())
            if not del_mes.empty else 0.0
        )
        filas["Coberturas"].append(_cobertura(del_mes, umbral))

    # Bultos y Kg se suman; la cobertura del anio se cuenta desde el grano.
    filas["Bultos"].append(sum(filas["Bultos"]))
    filas["Kg"].append(sum(filas["Kg"]))
    filas["Coberturas"].append(_cobertura(del_anio, umbral))

    return pd.DataFrame(filas, index=[*MESES_CORTOS, "TOTAL"]).T
