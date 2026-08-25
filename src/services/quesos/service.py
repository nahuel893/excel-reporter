"""QuesosService — volumen y cobertura de quesos LA HUERTA, por mes.

Un bloque por anio con tres filas: Bultos, Kg y Coberturas. Los kilos salen de
un factor POR ARTICULO que se exporta a mano del proveedor.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService

from .constants import MARCA, MESES_CORTOS
from .processor import articulos_sin_factor, construir_anio, leer_factores

logger = logging.getLogger(__name__)

HDR = "2E75B6"
CLARO = "DDEBF7"
TOTAL = "FFE08A"
ALERTA = "FFC7CE"


@dataclass
class QuesosConfig:
    """Config del informe.

    Args:
        anios: anios a mostrar, uno debajo del otro.
        factores_path: xlsx con id_articulo -> kg por unidad.
        umbral: piso de unidades para dar por cubierto a un cliente (`> 0`).
        nombre_archivo: nombre de salida sin extension.
    """
    anios: list[int] = field(default_factory=lambda: [2025, 2026])
    factores_path: str = "factor_conversion_quesos.xlsx"
    umbral: float = 0.0
    nombre_archivo: str | None = None


@dataclass
class QuesosResult:
    ruta_archivo: Path
    anios: list[int]
    bultos: float
    kg: float
    sin_factor: list[int]


def _borde() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class QuesosService(BaseService):
    SERVICE_SLUG = "quesos"
    GRANULARITY = "month"

    def generar_reporte(self, config: QuesosConfig) -> QuesosResult:
        factores = leer_factores(config.factores_path)
        anios = sorted(config.anios)

        ventas = self.data_loader.get_ventas_articulo_cliente_mes(
            marca=MARCA,
            fecha_desde=f"{anios[0]}-01-01",
            fecha_hasta=f"{anios[-1]}-12-31",
        )
        if ventas is None or ventas.empty:
            raise RuntimeError(f"sin ventas de {MARCA} entre {anios[0]} y {anios[-1]}")

        faltantes = articulos_sin_factor(ventas, factores)
        if faltantes:
            # No se corta la corrida: se avisa. Un articulo nuevo sin factor
            # suma bultos y no suma kg, y el informe queda corto en silencio.
            logger.warning(
                "Articulos con venta y SIN factor de conversion: %s. "
                "Sus bultos entran pero sus kilos NO.", faltantes,
            )

        bloques = {a: construir_anio(ventas, factores, a, config.umbral) for a in anios}

        nombre = config.nombre_archivo or "Quesos - volumen y cobertura por mes"
        out_dir = service_output_dir(self.SERVICE_SLUG, f"{anios[-1]}-12-01", "month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"
        self._build_workbook(config, bloques, faltantes, ruta)

        return QuesosResult(
            ruta_archivo=ruta,
            anios=anios,
            bultos=float(sum(b.loc["Bultos", "TOTAL"] for b in bloques.values())),
            kg=float(sum(b.loc["Kg", "TOTAL"] for b in bloques.values())),
            sin_factor=faltantes,
        )

    def _build_workbook(
        self,
        config: QuesosConfig,
        bloques: dict[int, pd.DataFrame],
        faltantes: list[int],
        ruta: Path,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quesos"
        borde = _borde()

        ws["A1"] = "Volumen y cobertura por mes — QUESOS LA HUERTA"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            f"Kg = unidades x peso promedio por unidad (factor por articulo) · "
            f"Cobertura = clientes distintos con compra neta > {config.umbral:g} · "
            f"el TOTAL de coberturas NO es la suma de los meses"
        )
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        r = 4
        for anio, bloque in bloques.items():
            ws.cell(r, 1, f"Año {anio}").font = Font(bold=True, size=12, color=HDR)
            r += 1
            self._escribir_bloque(ws, r, bloque, borde)
            r += len(bloque) + 3

        if faltantes:
            c = ws.cell(r, 1, "Articulos con venta y SIN factor de conversion: "
                              + ", ".join(str(a) for a in faltantes)
                              + ". Sus bultos entran, sus kilos NO.")
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", start_color=ALERTA, end_color=ALERTA)

        ws.column_dimensions["A"].width = 24
        for j in range(2, len(MESES_CORTOS) + 3):
            ws.column_dimensions[get_column_letter(j)].width = 11
        wb.save(ruta)

    @staticmethod
    def _escribir_bloque(ws: Worksheet, fila: int, bloque: pd.DataFrame, borde: Border) -> None:
        headers = ["Indicador", *[c for c in bloque.columns]]
        for j, h in enumerate(headers, 1):
            c = ws.cell(fila, j, h.upper() if h == "TOTAL" else h)
            c.fill = PatternFill("solid", start_color=HDR, end_color=HDR)
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border = borde

        for i, (medida, valores) in enumerate(bloque.iterrows(), 1):
            r = fila + i
            ec = ws.cell(r, 1, medida)
            ec.font = Font(bold=True)
            ec.border = borde
            ec.fill = PatternFill("solid", start_color=CLARO, end_color=CLARO)
            for j, col in enumerate(bloque.columns, 2):
                v = float(valores[col])
                c = ws.cell(r, j, int(v) if medida == "Coberturas" else v)
                c.number_format = "#,##0" if medida in ("Bultos", "Coberturas") else "#,##0.00"
                c.border = borde
                c.alignment = Alignment(horizontal="right")
                if col == "TOTAL":
                    c.fill = PatternFill("solid", start_color=TOTAL, end_color=TOTAL)
                    c.font = Font(bold=True)

    def run(self, config: QuesosConfig) -> QuesosResult:
        return self.generar_reporte(config)
