"""
RebotesService - Reporte de bounces/rechazos por vendedor, supervisor y cliente.

Genera un archivo Excel con 3 hojas:
  1. Rebotes: Vendedor | Bultos Vendidos | Bultos Rechazados | % Rechazo | Supervisor
     [vendedores agrupados por supervisor]
     TOTALES | blank | subtotales supervisor
  2. Ventas por Cliente: fantasia | razon_social | [generico: bultos_vendidos | bultos_rechazados | %]
  3. Rechazos por Cliente: fantasia | razon_social | [generico: bultos_rechazados]
"""
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.services.base_service import BaseService, BaseReporteConfig, BaseReporteResult
from src.services.rebotes.constants import SUPERVISOR_VENDOR_MAP
from src.services.rebotes.processor import (
    agregar_totales_supervisor,
    calcular_rebotes_vendedor,
    pivot_rebotes_por_generico,
    pivot_rebotes_por_generico_supervisor,
)

logger = __import__("logging").getLogger(__name__)

CCU_GENERICOS = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES"]

# Colores semaforo
ROJO = "FF6366"
AMARILLO = "FFEB9C"
VERDE = "C6EFCE"
ROJO_FONT = "9C0006"
AMARILLO_FONT = "9C5700"
VERDE_FONT = "006100"


@dataclass
class RebotesConfig(BaseReporteConfig):
    """Configuracion para el reporte de rebotes."""
    genericos: list[str] | None = None


@dataclass
class RebotesResult(BaseReporteResult):
    """Resultado del reporte de rebotes."""
    vendedores: int = 0
    supervisores: list[str] = None
    clientes: int = 0

    def __post_init__(self):
        if self.supervisores is None:
            self.supervisores = []


class RebotesService(BaseService):
    """Genera el reporte de rebotes (bounces/rechazos por vendedor)."""

    SERVICE_SLUG = "reporte-rebotes"
    GRANULARITY = "month"

    def _thin_border(self) -> Border:
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_row(self, ws, row_idx: int, values: list, bold: bool = False, pct: float = None) -> None:
        border = self._thin_border()
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if bold:
                cell.font = Font(bold=True)
            if col_idx in (2, 3):
                cell.number_format = "#,##0"
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = "0.0%"
                if pct is not None:
                    if pct < 0.03:
                        cell.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
                        cell.font = Font(color=VERDE_FONT, bold=bold)
                    elif pct < 0.05:
                        cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
                        cell.font = Font(color=AMARILLO_FONT, bold=bold)
                    else:
                        cell.fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")
                        cell.font = Font(color=ROJO_FONT, bold=bold)

    def _build_sheet_vendedor(
        self, wb, df_vendedor: pd.DataFrame, df_supervisor: pd.DataFrame
    ) -> None:
        ws = wb.create_sheet(title="Rebotes")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 18

        current_row = 1
        self._write_row(ws, current_row, ["Vendedor", "Bultos Vendidos", "Bultos Rechazados", "% Rechazo", "Supervisor"], bold=True)
        current_row += 1

        for _, row in df_vendedor.iterrows():
            pct = row.get("% Rechazo", 0)
            self._write_row(ws, current_row, [
                row["vendedor"],
                row["bultos_vendidos"],
                row["bultos_rechazados"],
                pct,
                row.get("Supervisor", ""),
            ], pct=pct)
            if row["vendedor"] == "DIRECTA":
                ws.row_dimensions[current_row].hidden = True
            current_row += 1

        total_vendidos = df_vendedor["bultos_vendidos"].sum()
        total_rechazados = df_vendedor["bultos_rechazados"].sum()
        total_pct = (total_rechazados / total_vendidos) if total_vendidos > 0 else 0
        self._write_row(ws, current_row, ["TOTALES", total_vendidos, total_rechazados, total_pct, ""], bold=True, pct=total_pct)
        current_row += 1

        self._write_row(ws, current_row, ["", "", "", "", ""])
        current_row += 1

        for _, srow in df_supervisor.iterrows():
            pct = srow.get("% Rechazo", 0)
            self._write_row(ws, current_row, [
                srow["Supervisor"],
                srow["Bultos Vendidos"],
                srow["Bultos Rechazados"],
                pct,
                "",
            ], bold=True, pct=pct)
            current_row += 1

    def _build_sheet_ventas_cliente(
        self, wb, df_raw: pd.DataFrame, genericos: list[str]
    ) -> None:
        ws = wb.create_sheet(title="Ventas por Cliente")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 35
        n_gen = len(genericos)
        for i in range(n_gen):
            base = 4 + i * 3
            ws.column_dimensions[get_column_letter(base)].width = 16
            ws.column_dimensions[get_column_letter(base + 1)].width = 16
            ws.column_dimensions[get_column_letter(base + 2)].width = 12

        border = self._thin_border()
        sub_headers = ["Bultos Vendidos", "Bultos Rechazados", "% Rechazo"]

        col_start = 4
        for g in genericos:
            col_end = col_start + 2
            cell = ws.cell(row=1, column=col_start, value=g)
            cell.font = Font(bold=True, size=11)
            cell.border = border
            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            col_start = col_end + 1

        col = 1
        for h in ["Codigo", "Fantasia", "Razon Social"]:
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True)
            cell.border = border
            col += 1
        for g in genericos:
            for sh in sub_headers:
                cell = ws.cell(row=2, column=col, value=sh)
                cell.font = Font(bold=True)
                cell.border = border
                col += 1

        current_row = 3

        for (id_cliente, fantasia, razon_social), group in df_raw.groupby(["id_cliente", "fantasia", "razon_social"]):
            col = 1
            for val in (id_cliente, fantasia, razon_social):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = border
                col += 1
            for g in genericos:
                g_data = group[group["generico"] == g]
                vendidos = g_data["bultos_vendidos"].sum() if len(g_data) > 0 else 0
                rechazados = g_data["bultos_rechazados"].sum() if len(g_data) > 0 else 0
                pct = (rechazados / vendidos) if vendidos > 0 else 0.0

                for val, nf in [(vendidos, "#,##0"), (rechazados, "#,##0"), (pct, "0.0%")]:
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.number_format = nf
                    cell.border = border
                    if nf == "0.0%":
                        if pct < 0.03:
                            cell.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
                            cell.font = Font(color=VERDE_FONT)
                        elif pct < 0.05:
                            cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
                            cell.font = Font(color=AMARILLO_FONT)
                        else:
                            cell.fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")
                            cell.font = Font(color=ROJO_FONT)
                    col += 1
            current_row += 1

        col = 1
        for val in ("TOTALES", "", ""):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = Font(bold=True)
            cell.border = border
            col += 1
        for g in genericos:
            g_data = df_raw[df_raw["generico"] == g]
            vendidos = g_data["bultos_vendidos"].sum()
            rechazados = g_data["bultos_rechazados"].sum()
            pct = (rechazados / vendidos) if vendidos > 0 else 0.0
            for val, nf in [(vendidos, "#,##0"), (rechazados, "#,##0"), (pct, "0.0%")]:
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = Font(bold=True)
                cell.number_format = nf
                cell.border = border
                if nf == "0.0%":
                    if pct < 0.03:
                        cell.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
                    elif pct < 0.05:
                        cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")
                col += 1

    def _build_sheet_rechazos_cliente(
        self, wb, df_rechazos: pd.DataFrame, genericos: list[str]
    ) -> None:
        ws = wb.create_sheet(title="Rechazos por Cliente")

        headers = ["Codigo", "Fantasia", "Razon Social"] + genericos
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.border = self._thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 35

        current_row = 2
        for (id_cliente, fantasia, razon_social), group in df_rechazos.groupby(["id_cliente", "fantasia", "razon_social"]):
            row_vals = [id_cliente, fantasia, razon_social]
            for g in genericos:
                g_data = group[group["generico"] == g]
                row_vals.append(g_data["bultos_rechazados"].sum() if len(g_data) > 0 else 0)
            border = self._thin_border()
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = border
                if col_idx > 3:
                    cell.number_format = "#,##0"
            current_row += 1

        row_vals = ["TOTALES", "", ""]
        for g in genericos:
            val = df_rechazos[df_rechazos["generico"] == g]["bultos_rechazados"].sum()
            row_vals.append(val)
        border = self._thin_border()
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = border
            cell.font = Font(bold=True)
            if col_idx > 3:
                cell.number_format = "#,##0"

    def _build_sheet_rebotes_por_generico(
        self, wb, df_pivot: pd.DataFrame, df_supervisor_pivot: pd.DataFrame
    ) -> None:
        ws = wb.create_sheet(title="% Rebotes x Generico")

        border = self._thin_border()
        headers = list(df_pivot.columns)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 18
        col = 3
        for _ in ["CERVEZAS", "AGUAS DANONE", "MULTICCU"]:
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1

        ws.row_dimensions[1].height = 35
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="JetBrains Mono", bold=True, color="000000")
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for r_idx, (_, row) in enumerate(df_pivot.iterrows(), 2):
            for col_idx, col_name in enumerate(headers, 1):
                val = row[col_name]
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.border = border
                cell.font = Font(name="JetBrains Mono", color="000000")
                if col_name.startswith("%"):
                    cell.number_format = "0.0%"
                    pct = val or 0.0
                    if pct < 0.03:
                        cell.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
                    elif pct < 0.05:
                        cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")
                elif "Bultos" in col_name or "Rechazados" in col_name:
                    cell.number_format = "#,##0"

        start_row = len(df_pivot) + 3
        for _, srow in df_supervisor_pivot.iterrows():
            for col_idx, col_name in enumerate(headers, 1):
                val = srow.get(col_name, "") if col_name in df_supervisor_pivot.columns else ""
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.font = Font(name="JetBrains Mono", bold=True, color="000000")
                cell.border = border
                if col_name.startswith("%"):
                    cell.number_format = "0.0%"
                    pct = val or 0.0
                    if pct < 0.03:
                        cell.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
                    elif pct < 0.05:
                        cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")
                elif "Bultos" in col_name or "Rechazados" in col_name:
                    cell.number_format = "#,##0"
            start_row += 1

    def _build_excel(
        self,
        df_vendedor: pd.DataFrame,
        df_supervisor: pd.DataFrame,
        df_cliente_raw: pd.DataFrame,
        df_rechazos: pd.DataFrame,
        df_pivot_generico: pd.DataFrame,
        df_supervisor_pivot: pd.DataFrame,
        nombre: str,
        output_dir: Path,
        genericos: list[str],
    ) -> Path:
        wb = Workbook()
        wb.remove(wb.active)
        self._build_sheet_vendedor(wb, df_vendedor, df_supervisor)
        self._build_sheet_ventas_cliente(wb, df_cliente_raw, genericos)
        self._build_sheet_rechazos_cliente(wb, df_rechazos, genericos)
        self._build_sheet_rebotes_por_generico(wb, df_pivot_generico, df_supervisor_pivot)

        ruta = output_dir / f"{nombre}.xlsx"
        wb.save(ruta)
        return ruta

    def generar_reporte(self, config: RebotesConfig) -> RebotesResult:
        genericos = config.genericos if config.genericos else CCU_GENERICOS

        # 1. Datos de vendedores
        df_raw = self.data_loader.get_rebotes_vendedor(
            config.fecha_desde, config.fecha_hasta, genericos
        )
        if df_raw.empty:
            logger.warning("No se encontraron datos de rebotes para el periodo %s - %s", config.fecha_desde, config.fecha_hasta)

        df_con_pct = calcular_rebotes_vendedor(df_raw)
        df_vendedor, df_supervisor = agregar_totales_supervisor(df_con_pct, SUPERVISOR_VENDOR_MAP)

        # 2. Datos de clientes (ventas)
        df_cliente_raw = self.data_loader.get_ventas_por_cliente(
            config.fecha_desde, config.fecha_hasta, genericos
        )

        # 3. Datos de clientes (rechazos)
        df_rechazos = self.data_loader.get_rechazos_por_cliente(
            config.fecha_desde, config.fecha_hasta, genericos
        )

        # 4. Datos % rechazo por vendedor x generico
        df_gen_raw = self.data_loader.get_rebotes_vendedor_por_generico(
            config.fecha_desde, config.fecha_hasta, genericos
        )
        df_pivot_generico = pivot_rebotes_por_generico(df_gen_raw, SUPERVISOR_VENDOR_MAP)
        df_supervisor_pivot = pivot_rebotes_por_generico_supervisor(df_gen_raw, SUPERVISOR_VENDOR_MAP)

        nombre = config.nombre_archivo or f"Rebotes {config.fecha_desde} {config.fecha_hasta}"
        out_dir = self._output_dir(config.fecha_desde)
        out_dir.mkdir(parents=True, exist_ok=True)

        ruta = self._build_excel(df_vendedor, df_supervisor, df_cliente_raw, df_rechazos, df_pivot_generico, df_supervisor_pivot, nombre, out_dir, genericos)

        return RebotesResult(
            ruta_archivo=ruta,
            registros_procesados=len(df_con_pct),
            vendedores=len(df_vendedor),
            supervisores=sorted(df_supervisor["Supervisor"].unique().tolist()) if not df_supervisor.empty else [],
            clientes=len(df_cliente_raw["fantasia"].unique()) if not df_cliente_raw.empty else 0,
        )

    def run(self, config: RebotesConfig) -> RebotesResult:
        return self.generar_reporte(config)