"""
Processor for Reporte General Badie.

Builds an openpyxl Workbook with 4 sheets:
  1. Reporte  — formula-driven pivot (visible, index 0)
  2. VentasCCU — raw monthly sales (visible, Table named TblVentasCCU)
  3. CoberturaCCU — raw client coverage (visible, Table named TblCoberturaCCU)
  4. _Trimestres — hidden list of "YYYY-Q[1-4]" strings for the Reporte dropdown

Quarter selector in B2 is an openpyxl DataValidation dropdown referencing _Trimestres!$A$1:$A$N.
All data columns use formula-only approach: no Python-side aggregation, Excel does it via
SUMPRODUCT referencing structured table columns.
"""

from __future__ import annotations

from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------------------------------------------------------
# Header styling constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column widths for the Reporte sheet (col letter → width)
_COL_WIDTHS: dict[str, float] = {
    "A": 26,   # Sucursal (long names like "SUCURSAL ABRA PAMPA")
    "B": 12,   # Total CCU
    "C": 16,   # Total CCU Año Anterior
    "D": 12,   # AA vs MMAA
    "E": 11,   # % Cerveza
    "F": 14,   # % Aguas Danone
    "G": 12,   # % Multi CCU
    "H": 13,   # Cobertura Normal
    "I": 16,   # Cobertura ≥3 Bultos
    "J": 18,   # Cobertura Prom. ≥1 Bulto/Mes
    "K": 13,   # Cobertura Normal (s/regalos)
    "L": 16,   # Cobertura ≥3 Bultos (s/regalos)
    "M": 18,   # Cobertura Prom. ≥1 Bulto/Mes (s/regalos)
    "N": 16,   # Cob. <1 Bulto (s/regalos)
    "O": 14,   # Cob. <1 Bulto (con regalos)
    "P": 16,   # AGUAS DANONE ≥3 Bultos
    "Q": 18,   # AGUAS DANONE ≥3 Bultos (s/regalos)
}

_REPORTE_HEADERS = [
    "Sucursal",
    "Total CCU",
    "Total CCU Año Anterior",
    "AA vs MMAA",
    "% Cerveza",
    "% Aguas Danone",
    "% Multi CCU",
    "Cobertura Normal",
    "Cobertura ≥3 Bultos",
    "Cobertura Prom. ≥1 Bulto/Mes",
    "Cob. Normal (s/regalos)",
    "Cob. ≥3 Bultos (s/regalos)",
    "Cob. Prom. ≥1 Bulto/Mes (s/regalos)",
    "Cob. <1 Bulto (s/regalos)",
    "Cob. <1 Bulto (c/regalos)",
    "AGUAS DANONE ≥3 Bultos",
    "AGUAS DANONE ≥3 Bultos (s/regalos)",
]

# Number formats per column letter (A=1, B=2, ...)
_COL_FORMATS: dict[str, str] = {
    "B": "#,##0",
    "C": "#,##0",
    "D": "0.0%",
    "E": "0.0%",
    "F": "0.0%",
    "G": "0.0%",
    "H": "#,##0",
    "I": "#,##0",
    "J": "#,##0",
    "K": "#,##0",
    "L": "#,##0",
    "M": "#,##0",
    "N": "#,##0",
    "O": "#,##0",
    "P": "#,##0",
    "Q": "#,##0",
}

# Group headers: row 3 (col_start, col_end, title)
_GROUP_HEADERS: list[tuple[str, str, str]] = [
    ("B", "D", "Volumen Trimestre"),
    ("E", "G", "Mix Volumen"),
    ("H", "J", "Cobertura"),
    ("K", "M", "Cobertura sin Regalos"),
    ("N", "O", "Auditoría"),
    ("P", "Q", "Cobertura Aguas Danone"),
]
_GROUP_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_workbook(
    sucursales: list[str],
    df_ventas: pd.DataFrame,
    df_cob: pd.DataFrame,
    trimestres: list[str],
) -> Workbook:
    """Build and return the complete workbook."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    ws_reporte = wb.create_sheet("Reporte")
    ws_ventas = wb.create_sheet("VentasCCU")
    ws_cob = wb.create_sheet("CoberturaCCU")
    ws_trim = wb.create_sheet("_Trimestres")

    _build_trimestres_sheet(ws_trim, trimestres)
    _build_ventas_sheet(ws_ventas, df_ventas)
    _build_cobertura_sheet(ws_cob, df_cob)
    _build_reporte_sheet(ws_reporte, sucursales, trimestres)

    return wb


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_trimestres_sheet(ws, trimestres: list[str]) -> None:
    """Populate _Trimestres with one YYYY-Q[1-4] per row in column A, then hide it."""
    for i, tri in enumerate(trimestres, start=1):
        ws.cell(row=i, column=1, value=tri)
    ws.sheet_state = "hidden"


def _build_ventas_sheet(ws, df: pd.DataFrame) -> None:
    """Write raw ventas data as an openpyxl Table named TblVentasCCU."""
    columns = ["sucursal", "generico", "anio", "trimestre", "bultos"]

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    if not df.empty:
        for row_idx, row in enumerate(df[columns].itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    n_rows = len(df) + 1  # +1 for header
    last_col = get_column_letter(len(columns))
    table_ref = f"A1:{last_col}{max(n_rows, 2)}"

    tbl = Table(displayName="TblVentasCCU", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


def _build_cobertura_sheet(ws, df: pd.DataFrame) -> None:
    """Write raw cobertura data as an openpyxl Table named TblCoberturaCCU.

    Columns: sucursal, anio, trimestre, id_cliente,
             bultos, bultos_sin_regalos,
             bultos_aguas_danone, bultos_aguas_danone_sin_regalos,
             meses_con_compra
    """
    columns = [
        "sucursal", "anio", "trimestre", "id_cliente",
        "bultos", "bultos_sin_regalos",
        "bultos_aguas_danone", "bultos_aguas_danone_sin_regalos",
        "meses_con_compra",
    ]

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    if not df.empty:
        for row_idx, row in enumerate(df[columns].itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    n_rows = len(df) + 1
    last_col = get_column_letter(len(columns))
    table_ref = f"A1:{last_col}{max(n_rows, 2)}"

    tbl = Table(displayName="TblCoberturaCCU", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


def _build_reporte_sheet(
    ws, sucursales: list[str], trimestres: list[str]
) -> None:
    """Build the Reporte sheet with title, dropdown, headers, and formula rows."""
    n_trim = len(trimestres)
    last_trim = trimestres[-1] if trimestres else ""

    # A1: title
    title_cell = ws["A1"]
    title_cell.value = "Reporte General Badie"
    title_cell.font = Font(bold=True, size=14)

    # A2: label
    ws["A2"] = "Trimestre:"

    # B2: initial value + DataValidation dropdown
    ws["B2"] = last_trim
    dv = DataValidation(
        type="list",
        formula1=f"_Trimestres!$A$1:$A${n_trim}",
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("B2")

    # Row 3: group headers (merged cells)
    for col_start, col_end, title in _GROUP_HEADERS:
        ws.merge_cells(f"{col_start}3:{col_end}3")
        cell = ws[f"{col_start}3"]
        cell.value = title
        cell.fill = _GROUP_HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # Column widths
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Row 4: headers (with wrap text + taller row to fit 2 lines)
    ws.row_dimensions[4].height = 32
    for col_idx, header in enumerate(_REPORTE_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # Rows 5..(4+N): one per sucursal
    for i, sucursal in enumerate(sucursales):
        r = 5 + i
        _write_sucursal_row(ws, r, sucursal)

    # Total row immediately after sucursales
    n_sucursales = len(sucursales)
    if n_sucursales > 0:
        total_row = 5 + n_sucursales
        _write_total_row(ws, total_row, n_sucursales)
        _apply_conditional_formats(ws, n_sucursales)
        _apply_group_borders(ws, n_sucursales)


def _write_sucursal_row(ws, r: int, sucursal: str) -> None:
    """Write one data row for a sucursal: literal name in A, formulas in B-J."""
    ws.cell(row=r, column=1, value=sucursal)

    # B: Total CCU bultos for selected quarter
    ws.cell(row=r, column=2, value=_formula_total_ccu(r)).number_format = "#,##0"

    # C: Total CCU Año Anterior (mismo trimestre, año - 1)
    c_cell = ws.cell(row=r, column=3, value=_formula_total_ccu_anio_anterior(r))
    c_cell.number_format = "#,##0"

    # D: AA vs MMAA = (B - C) / C
    d_cell = ws.cell(row=r, column=4, value=_formula_variacion_yoy(r))
    d_cell.number_format = "0.0%"

    # E: % Cervezas
    e_cell = ws.cell(row=r, column=5, value=_formula_pct_generico(r, "CERVEZAS"))
    e_cell.number_format = "0.0%"

    # F: % Aguas Danone
    f_cell = ws.cell(row=r, column=6, value=_formula_pct_generico(r, "AGUAS DANONE"))
    f_cell.number_format = "0.0%"

    # G: % Multi CCU (Vinos + Sidras)
    g_cell = ws.cell(row=r, column=7, value=_formula_pct_multi_ccu(r))
    g_cell.number_format = "0.0%"

    # H: Cobertura Normal (>0 bultos)
    h_cell = ws.cell(row=r, column=8, value=_formula_cobertura(r, operator=">", threshold=0))
    h_cell.number_format = "#,##0"

    # I: Cobertura ≥3 Bultos (clients with bultos >= 3 in the quarter)
    i_cell = ws.cell(row=r, column=9, value=_formula_cobertura(r, operator=">=", threshold=3))
    i_cell.number_format = "#,##0"

    # J: Cobertura Promedio ≥1 Bulto/Mes (clients with avg monthly bultos >= 1)
    j_cell = ws.cell(row=r, column=10, value=_formula_cobertura_promedio(r, threshold=1))
    j_cell.number_format = "#,##0"

    # K: Cob. Normal s/regalos (bultos_sin_regalos > 0)
    k_cell = ws.cell(
        row=r, column=11,
        value=_formula_cobertura(r, operator=">", threshold=0, field="bultos_sin_regalos"),
    )
    k_cell.number_format = "#,##0"

    # L: Cob. ≥3 Bultos s/regalos
    l_cell = ws.cell(
        row=r, column=12,
        value=_formula_cobertura(r, operator=">=", threshold=3, field="bultos_sin_regalos"),
    )
    l_cell.number_format = "#,##0"

    # M: Cob. Promedio ≥1 Bulto/Mes s/regalos
    m_cell = ws.cell(
        row=r, column=13,
        value=_formula_cobertura_promedio(r, threshold=1, field="bultos_sin_regalos"),
    )
    m_cell.number_format = "#,##0"

    # N: Cob. <1 Bulto s/regalos (clientes con bultos_sin_regalos < 1)
    n_cell = ws.cell(
        row=r, column=14,
        value=_formula_cobertura(r, operator="<", threshold=1, field="bultos_sin_regalos"),
    )
    n_cell.number_format = "#,##0"

    # O: Cob. <1 Bulto c/regalos (clientes con bultos total < 1, incluyendo regalos)
    o_cell = ws.cell(
        row=r, column=15,
        value=_formula_cobertura(r, operator="<", threshold=1, field="bultos"),
    )
    o_cell.number_format = "#,##0"

    # P: AGUAS DANONE ≥3 Bultos (clientes con ≥3 bultos de AGUAS DANONE)
    p_cell = ws.cell(
        row=r, column=16,
        value=_formula_cobertura(r, operator=">=", threshold=3, field="bultos_aguas_danone"),
    )
    p_cell.number_format = "#,##0"

    # Q: AGUAS DANONE ≥3 Bultos s/regalos
    q_cell = ws.cell(
        row=r, column=17,
        value=_formula_cobertura(r, operator=">=", threshold=3, field="bultos_aguas_danone_sin_regalos"),
    )
    q_cell.number_format = "#,##0"


def _write_total_row(ws, r: int, n_sucursales: int) -> None:
    """Write a TOTAL row at row r summing/recalculating across all sucursales."""
    first_data_row = 5
    last_data_row = 5 + n_sucursales - 1

    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="D9E1F2")

    # A: label
    a_cell = ws.cell(row=r, column=1, value="TOTAL")
    a_cell.font = total_font
    a_cell.fill = total_fill

    # B: SUM Total CCU
    b_cell = ws.cell(row=r, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})")
    b_cell.number_format = "#,##0"

    # C: SUM Total CCU Año Anterior
    c_cell = ws.cell(row=r, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
    c_cell.number_format = "#,##0"

    # D: AA vs MMAA recalculated over totals
    d_cell = ws.cell(row=r, column=4, value=f'=IFERROR((B{r}-C{r})/C{r},"")')
    d_cell.number_format = "0.0%"

    # E, F, G: percentages recalculated over global totals (Option A)
    y = _year_expr()
    t = _trimestre_expr()
    for col_idx, generico_filter in [
        (5, '(TblVentasCCU[generico]="CERVEZAS")'),
        (6, '(TblVentasCCU[generico]="AGUAS DANONE")'),
        (7, '((TblVentasCCU[generico]="VINOS CCU")+(TblVentasCCU[generico]="SIDRAS Y LICORES"))'),
    ]:
        formula = (
            f'=IFERROR(SUMPRODUCT('
            f'(TblVentasCCU[anio]={y})'
            f'*(TblVentasCCU[trimestre]={t})'
            f'*{generico_filter}'
            f'*TblVentasCCU[bultos])/B{r},"")'
        )
        cell = ws.cell(row=r, column=col_idx, value=formula)
        cell.number_format = "0.0%"

    # H..Q: SUM cobertura (incluye coberturas, auditoría y aguas danone)
    for col_idx, col_letter in [
        (8, "H"), (9, "I"), (10, "J"),
        (11, "K"), (12, "L"), (13, "M"),
        (14, "N"), (15, "O"),
        (16, "P"), (17, "Q"),
    ]:
        formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell = ws.cell(row=r, column=col_idx, value=formula)
        cell.number_format = "#,##0"

    # Bold + fill all total cells
    for col_idx in range(1, 18):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill


def _apply_conditional_formats(ws, n_sucursales: int) -> None:
    """Apply conditional formatting to data rows (excluding TOTAL row)."""
    first = 5
    last = 5 + n_sucursales - 1

    # Mix Volumen (E, F, G): heatmap — light to deep blue
    for col in ("E", "F", "G"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="4472C4",
            ),
        )

    # AA vs MMAA (D): 3-color scale red-yellow-green
    ws.conditional_formatting.add(
        f"D{first}:D{last}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )

    # Cobertura "positiva" (H..M, P, Q): heatmap rojo→verde (más = mejor)
    for col in ("H", "I", "J", "K", "L", "M", "P", "Q"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # Auditoría (N, O): heatmap INVERTIDO verde→rojo (más = peor)
    for col in ("N", "O"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    # Total CCU + Año Anterior (B, C): data bars
    for col in ("B", "C"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            DataBarRule(start_type="min", end_type="max", color="638EC6"),
        )


def _apply_group_borders(ws, n_sucursales: int) -> None:
    """Apply thick vertical borders separating each group of columns.

    Groups: A (Sucursal) | B-D | E-G | H-J | K-M | N-O | P-Q
    Thick LEFT border on first column of each group after A: B, E, H, K, N, P
    Thick RIGHT border on the last column (Q) to close the last group
    Range: row 3 (group header) → row 5+n_sucursales (TOTAL row, inclusive)
    """
    first_row = 3
    last_row = 5 + n_sucursales  # inclusive of TOTAL row
    thick = Side(style="thick", color="000000")

    left_border_cols = [2, 5, 8, 11, 14, 16]  # B, E, H, K, N, P
    for col_idx in left_border_cols:
        for row in range(first_row, last_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            existing = cell.border
            cell.border = Border(
                left=thick,
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom,
            )

    # Closing border on column Q (col 17)
    for row in range(first_row, last_row + 1):
        cell = ws.cell(row=row, column=17)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=thick,
            top=existing.top,
            bottom=existing.bottom,
        )


# ---------------------------------------------------------------------------
# Formula helpers
# ---------------------------------------------------------------------------


def _year_expr() -> str:
    return "VALUE(LEFT($B$2,4))"


def _trimestre_expr() -> str:
    """Extracts the quarter digit (1-4) from B2 string formatted 'YYYY-Q[1-4]'."""
    return "VALUE(MID($B$2,7,1))"


def _formula_total_ccu(r: int) -> str:
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f"=SUMPRODUCT("
        f"(TblVentasCCU[sucursal]=$A{r})"
        f"*(TblVentasCCU[anio]={y})"
        f"*(TblVentasCCU[trimestre]={t})"
        f"*TblVentasCCU[bultos])"
    )


def _formula_pct_generico(r: int, generico: str) -> str:
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f'=IFERROR(SUMPRODUCT('
        f'(TblVentasCCU[sucursal]=$A{r})'
        f'*(TblVentasCCU[anio]={y})'
        f'*(TblVentasCCU[trimestre]={t})'
        f'*(TblVentasCCU[generico]="{generico}")'
        f'*TblVentasCCU[bultos])/B{r},"")'
    )


def _formula_pct_multi_ccu(r: int) -> str:
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f'=IFERROR(SUMPRODUCT('
        f'(TblVentasCCU[sucursal]=$A{r})'
        f'*(TblVentasCCU[anio]={y})'
        f'*(TblVentasCCU[trimestre]={t})'
        f'*((TblVentasCCU[generico]="VINOS CCU")+(TblVentasCCU[generico]="SIDRAS Y LICORES"))'
        f'*TblVentasCCU[bultos])/B{r},"")'
    )


def _formula_total_ccu_anio_anterior(r: int) -> str:
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f"=SUMPRODUCT("
        f"(TblVentasCCU[sucursal]=$A{r})"
        f"*(TblVentasCCU[anio]=({y}-1))"
        f"*(TblVentasCCU[trimestre]={t})"
        f"*TblVentasCCU[bultos])"
    )


def _formula_variacion_yoy(r: int) -> str:
    return f'=IFERROR((B{r}-C{r})/C{r},"")'


def _formula_cobertura_promedio(
    r: int, threshold: float, field: str = "bultos"
) -> str:
    """Count clients whose AVERAGE monthly value in the quarter (`field`/3)
    >= `threshold`, using TblCoberturaCCU.
    """
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f"=SUMPRODUCT("
        f"(TblCoberturaCCU[sucursal]=$A{r})"
        f"*(TblCoberturaCCU[anio]={y})"
        f"*(TblCoberturaCCU[trimestre]={t})"
        f"*((TblCoberturaCCU[{field}]/3)>={threshold})"
        f"*1)"
    )


def _formula_cobertura(
    r: int, operator: str, threshold: float, field: str = "bultos"
) -> str:
    """Count distinct clients matching `{field} {operator} {threshold}` for the
    selected sucursal/year/quarter, against TblCoberturaCCU.
    """
    y = _year_expr()
    t = _trimestre_expr()
    return (
        f"=SUMPRODUCT("
        f"(TblCoberturaCCU[sucursal]=$A{r})"
        f"*(TblCoberturaCCU[anio]={y})"
        f"*(TblCoberturaCCU[trimestre]={t})"
        f"*(TblCoberturaCCU[{field}]{operator}{threshold})"
        f"*1)"
    )


# ---------------------------------------------------------------------------
# Quarter list helper
# ---------------------------------------------------------------------------


def _generar_trimestres(desde: str, hasta: str) -> list[str]:
    """
    Return list of 'YYYY-Q[1-4]' strings covering the date range, inclusive of
    both endpoints' quarters.

    Args:
        desde: Date string (YYYY-MM-DD)
        hasta: Date string (YYYY-MM-DD)

    Returns:
        List of 'YYYY-Q[1-4]' strings in chronological order.
    """
    start = datetime.strptime(desde[:7], "%Y-%m")
    end = datetime.strptime(hasta[:7], "%Y-%m")

    start_q = (start.month - 1) // 3 + 1
    end_q = (end.month - 1) // 3 + 1

    trimestres: list[str] = []
    year = start.year
    q = start_q
    while (year < end.year) or (year == end.year and q <= end_q):
        trimestres.append(f"{year}-Q{q}")
        if q == 4:
            year += 1
            q = 1
        else:
            q += 1

    return trimestres
