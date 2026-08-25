"""
ResumenMensualService - Servicio para generacion de reportes de resumen mensual.

Orquesta el flujo completo: extraccion, procesamiento y generacion de Excel.
"""
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.base_processor import calcular_info_dias
from src.core.data_loader import DataLoader
from src.core.excel_updater import import_xlsx_as_sheet
from src.core.excel_writer import ExcelWriter, SheetStyle, ColumnFormat
from src.core.zonas import aplicar_zonas_virtuales
from src.services.base_service import BaseService
from src.services.resumen_mensual.processor import procesar_resumen_mensual

logger = logging.getLogger(__name__)

# Default genericos when config.genericos is None
_DEFAULT_GENERICOS = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES", "PERNOD RICARD"]

# Genericos lógicos cuya carga en fact_cupos está partida en sub-marcas.
# Cuando una hoja pide "VINOS", el cupo se reconstruye sumando ARIZU + CANCILLER.
_CUPOS_GENERICO_PARTS: dict[str, list[str]] = {
    "VINOS": ["ARIZU", "CANCILLER", "TORO"],
    "FRATELLI B": ["FRATELLI BRANCA"],
}

# Row labels for subtotals
_SUBTOTAL_CC = "SUBTOTAL CASA CENTRAL"
_SUC_SIN_DIRECTA = "SUCURSALES SIN DIRECTA"
_TOTAL_SIN_SMK = "TOTAL SIN SMK"

# CC-family sucursales (order matters for display)
_CC_FAMILY = ["CASA CENTRAL", "VALLE SALTA", "SUB DISTRIBUIDORES"]

# Subtotal row fills (one color per label) and matching font colors
_SUBTOTAL_FILLS = {
    _SUBTOTAL_CC:     "548235",  # green
    _SUC_SIN_DIRECTA: "7030A0",  # purple
    _TOTAL_SIN_SMK:   "FF0000",  # red
}
_SUBTOTAL_FONT_COLOR = "FFFFFF"  # white text on all 3 fills
# El semaforo (Tend vs Obj) SIEMPRE en negro: el relleno ya dice si esta bien o
# mal, y sobre los pasteles del ColorScaleRule cualquier otro color se pierde.
_SEMAFORO_FONT_COLOR = "000000"

# Header fill + font
_HEADER_FILL_COLOR = "1F4E78"  # dark blue
_HEADER_FONT_COLOR = "FFFFFF"  # white

# Sheet-wide font family
_FONT_NAME = "JetBrainsMono Nerd Font"


@dataclass
class ResumenMensualConfig:
    """Configuracion para el reporte de resumen mensual.

    Standalone dataclass; NO hereda de BaseReporteConfig para evitar que
    __post_init__ sobreescriba nombre_archivo=None con un valor por defecto.
    """
    fecha_desde: str           # "YYYY-MM-DD", primer dia del mes
    fecha_hasta: str           # "YYYY-MM-DD", ultimo dia con ventas (o fin de mes)
    genericos: list[str] | None = None
    nombre_archivo: str | None = None
    con_objetivo: bool = True  # True: uses gold.fact_cupos; set False if table is missing
    detalle_movimientos_path: str | None = None  # Path to source xlsx for Detalle Movimientos sheet
    detalle_movimientos_ma_path: str | None = None    # Mes anterior — imported as "Detalle Movimientos MA"
    detalle_movimientos_mmaa_path: str | None = None  # Mismo mes año anterior — imported as "Detalle Movimientos MMAA"
    categorias_deposito_path: str | None = None       # JSON con master-data para sheet "Categorias" + lookup en detalles
    # Genericos for which fact_ventas.id_documento = 'PRVTA' (facturas presupuesto)
    # is excluded from totals. When None, defaults to ["FRATELLI B"]; pass [] to disable.
    genericos_sin_prvta: list[str] | None = None
    # Mapping {generico: [marcas]} — for each entry, the sheet of that generico will
    # contain rows split by marca: one synthetic generico per listed marca, plus a
    # "{generico} (sin {marcas})" group for everything else. Each section gets its
    # own subtotals. Example: {"VINOS FINOS": ["QUARA"]}.
    marca_splits: dict[str, list[str]] | None = None
    # Cupos hardcodeados {sucursal: {generico: cupo}} — se concatenan al df_cupos
    # antes del merge. Útil para sucursales que no entran a fact_cupos (e.g. GUEMES).
    cupos_manuales: dict[str, dict[str, float]] | None = None


# Default list of genericos that exclude PRVTA documents (facturas presupuesto).
# Applied when ResumenMensualConfig.genericos_sin_prvta is None.
_DEFAULT_GENERICOS_SIN_PRVTA = ["FRATELLI B"]


@dataclass
class ResumenMensualResult:
    """Resultado de la generacion de un reporte de resumen mensual."""
    ruta_archivo: Path
    registros_procesados: int   # total de filas (sucursal, generico) en el archivo (suma de todas las hojas)
    sucursales: int             # cantidad de sucursales unicas en el resultado
    genericos_incluidos: list[str]
    hojas: list[str]            # nombres de hojas = nombres de genericos


def _nombre_reporte(df_dias: pd.DataFrame, fecha_hasta: str) -> str:
    """Genera nombre de archivo: 'Resumen {dd-mm-yyyy}' usando ultima fecha con ventas.

    Args:
        df_dias: DataFrame con columna 'fecha' (ventas de los ultimos dias habiles).
        fecha_hasta: Fecha limite del rango; se usa como fallback si df_dias esta vacio.

    Returns:
        Nombre de archivo sin extension.
    """
    if not df_dias.empty and "fecha" in df_dias.columns:
        ultima_fecha = pd.to_datetime(df_dias["fecha"]).max().strftime("%d-%m-%Y")
    else:
        ultima_fecha = pd.to_datetime(fecha_hasta).strftime("%d-%m-%Y")
    return f"Resumen - {ultima_fecha}"


def _segregar_directa_sucursales(df: pd.DataFrame) -> pd.DataFrame:
    """
    Renombra a 'DIRECTA SUCURSALES' las filas cuyo id_ruta == 100 y cuya
    sucursal no es 'CASA CENTRAL'.

    Debe llamarse ANTES del groupby que descarta id_ruta, y DESPUÉS de
    aplicar_zonas_virtuales (que ya separó VALLE SALTA y SUB DISTRIBUIDORES).

    Args:
        df: DataFrame con columnas 'sucursal', 'id_ruta' (y otras).

    Returns:
        Copia del DataFrame con las filas afectadas renombradas.
    """
    if "id_ruta" not in df.columns:
        return df
    df = df.copy()
    mask = (df["id_ruta"] == 100) & (df["sucursal"] != "CASA CENTRAL")
    df.loc[mask, "sucursal"] = "DIRECTA SUCURSALES"
    return df


def _crear_estilo_resumen(info_dias: dict, col_n1: str, col_n2: str) -> SheetStyle:
    """Crea el SheetStyle para las hojas del reporte de resumen mensual.

    Args:
        info_dias: Diccionario con Dias Habiles, Dias Transcurridos, Dias Faltantes.
        col_n1: Nombre de la columna del ultimo dia con ventas (ej: '28-02 Sabado').
        col_n2: Nombre de la columna del penultimo dia con ventas (ej: '27-02 Viernes').

    Returns:
        SheetStyle configurado para el reporte de resumen mensual.
    """
    _DASH_FMT = '#,##0;-#,##0;"-"'
    return SheetStyle(
        numeric_format="#,##0",
        column_formats={
            "Sucursal":        ColumnFormat(width=30.375, font_bold=True),
            "Generico":        ColumnFormat(width=14.125, font_bold=True),
            col_n1:            ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True),
            col_n2:            ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True),
            "Total Ventas":    ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True),
            "Tendencia":       ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True),
            # T-021: MMAA = dark red, MA = olive, Objetivo = light blue
            "MMAA":            ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True, font_color="C00000"),
            "MA":              ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True, font_color="808000"),
            "Objetivo":        ColumnFormat(number_format=_DASH_FMT, width=8.5, font_bold=True, font_color="4472C4"),
            "Tend vs Obj (%)": ColumnFormat(number_format="0.0%", width=8.5, font_bold=True),
        },
        summary_rows=info_dias,
        as_table=False,
    )


def _ordenar_e_inyectar_subtotales(df_hoja: pd.DataFrame) -> pd.DataFrame:
    """
    Ordena las filas del DataFrame de una hoja e inyecta las 3 filas de subtotales
    como filas con valores None (se rellenan con formulas SUM en post-write).

    Orden de salida:
        CASA CENTRAL → VALLE SALTA → SUB DISTRIBUIDORES
        → SUBTOTAL CASA CENTRAL
        → sucursales numeradas (alfa)
        → SUCURSALES SIN DIRECTA
        → DIRECTA SUCURSALES
        → TOTAL SIN SMK

    Args:
        df_hoja: DataFrame con columna 'Sucursal' y columnas numericas.

    Returns:
        DataFrame reordenado con filas de subtotales inyectadas (valores None
        en columnas numericas; 'Sucursal' = etiqueta del subtotal).
    """
    if df_hoja.empty:
        return df_hoja

    cols = list(df_hoja.columns)

    def _empty_row(label: str) -> dict:
        row = {c: None for c in cols}
        row["Sucursal"] = label
        if "Generico" in cols:
            # Keep the generico consistent so the filter in the loop still matches
            genericos = df_hoja["Generico"].dropna().unique()
            row["Generico"] = genericos[0] if len(genericos) == 1 else None
        return row

    suc_col = df_hoja["Sucursal"]

    # CC family (in fixed order)
    cc_rows = []
    for label in _CC_FAMILY:
        mask = suc_col == label
        if mask.any():
            cc_rows.append(df_hoja[mask])

    # Numbered sucursales: everything that is NOT cc_family, DIRECTA SUCURSALES, subtotal labels
    _special = set(_CC_FAMILY) | {"DIRECTA SUCURSALES", _SUBTOTAL_CC, _SUC_SIN_DIRECTA, _TOTAL_SIN_SMK}
    numbered_mask = ~suc_col.isin(_special)
    numbered = df_hoja[numbered_mask].sort_values("Sucursal")

    # DIRECTA SUCURSALES
    directa_mask = suc_col == "DIRECTA SUCURSALES"
    directa = df_hoja[directa_mask]

    # Build ordered list
    parts = []
    if cc_rows:
        parts.extend(cc_rows)
    subtotal_cc_row = pd.DataFrame([_empty_row(_SUBTOTAL_CC)])
    parts.append(subtotal_cc_row)
    if not numbered.empty:
        parts.append(numbered)
    suc_sin_directa_row = pd.DataFrame([_empty_row(_SUC_SIN_DIRECTA)])
    parts.append(suc_sin_directa_row)
    if not directa.empty:
        parts.append(directa)
    total_sin_smk_row = pd.DataFrame([_empty_row(_TOTAL_SIN_SMK)])
    parts.append(total_sin_smk_row)

    result = pd.concat(parts, ignore_index=True)
    return result


def _post_write_subtotals_and_heatmap(
    ws,
    df_hoja_ordered: pd.DataFrame,
    summary_rows_count: int,
    note: str | None = None,
):
    """
    Post-write step: resolves subtotal placeholder rows to real =SUM(...) formulas,
    applies bold+fill to subtotal rows, and adds a ColorScaleRule on Tend vs Obj (%).

    This function is called after ExcelWriter.add_sheet() returns the Worksheet.

    Args:
        ws: openpyxl Worksheet (returned by ExcelWriter.add_sheet)
        df_hoja_ordered: The ordered DataFrame (same order as what was written to the sheet,
                         including subtotal placeholder rows)
        summary_rows_count: Number of summary rows written before the header
                            (from SheetStyle.summary_rows). Used to compute row offsets.
        note: Optional label inserted as a new top row, merged across all columns
              with an amber highlight. Used to flag sheet-specific data caveats
              (e.g. PRVTA exclusion for FRATELLI B).
    """
    # Guard: if ws is not a real openpyxl Worksheet (e.g. a Mock in unit tests),
    # skip post-write silently.
    if not hasattr(ws, "iter_rows") or not callable(getattr(ws, "iter_rows", None)):
        return
    try:
        # Confirm it's a real worksheet by checking if iter_rows yields tuples of cells
        test_iter = ws.iter_rows(min_row=1, max_row=1)
        next(test_iter)  # will raise StopIteration on empty ws but not TypeError on Mock
    except (TypeError, AttributeError):
        return

    # -------------------------------------------------------------------
    # Optional caveat note: insert a merged label row at the top BEFORE
    # computing positions, so SUM formulas reference correctly-shifted rows.
    # -------------------------------------------------------------------
    note_offset = 0
    if note:
        ws.insert_rows(1)
        note_offset = 1

    # -------------------------------------------------------------------
    # Compute layout offsets (with note offset baked in if applicable)
    # -------------------------------------------------------------------
    # ExcelWriter._write_summary_rows writes `len(summary_rows) + 1` rows
    # before the header (the +1 is an empty separator row).
    header_row = note_offset + summary_rows_count + 1 + 1  # note + summary + separator + header
    data_start_row = header_row + 1
    n_data_rows = len(df_hoja_ordered)
    data_end_row = data_start_row + n_data_rows - 1

    # -------------------------------------------------------------------
    # Build column name → letter map from the actual header row in the sheet
    # -------------------------------------------------------------------
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    col_map = {}  # col_name -> column_letter
    for cell in header_cells:
        if cell.value:
            col_map[cell.value] = get_column_letter(cell.column)

    # -------------------------------------------------------------------
    # Numeric columns to include in SUM formulas (all except Sucursal, Generico)
    _NON_SUM_COLS = {"Sucursal", "Generico", "Tend vs Obj (%)"}
    sum_cols = [c for c in df_hoja_ordered.columns if c not in _NON_SUM_COLS]

    # -------------------------------------------------------------------
    # Split rows into sections by Generico value (one section per logical group).
    # Subtotals are computed per-section so multi-Generico sheets (marca_splits)
    # don't sum across groups.
    # -------------------------------------------------------------------
    n_cols = len(header_cells)
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    sections: list[tuple[str, int, int]] = []  # (generico_value, section_start_row, section_end_row)
    if "Generico" in df_hoja_ordered.columns:
        prev_gen = object()
        sec_start = data_start_row
        for df_idx, gen_value in enumerate(df_hoja_ordered["Generico"]):
            sheet_row = data_start_row + df_idx
            if gen_value != prev_gen and df_idx > 0:
                sections.append((prev_gen, sec_start, sheet_row - 1))
                sec_start = sheet_row
            prev_gen = gen_value
        sections.append((prev_gen, sec_start, data_end_row))
    else:
        sections.append((None, data_start_row, data_end_row))

    _NON_SUMMING_LABELS = set(_CC_FAMILY) | {
        "DIRECTA SUCURSALES", _SUBTOTAL_CC, _SUC_SIN_DIRECTA, _TOTAL_SIN_SMK
    }

    for _gen_value, sec_start_row, sec_end_row in sections:
        # Build label → row map for THIS section only
        section_rows_by_label: dict[str, int] = {}
        for sheet_row in range(sec_start_row, sec_end_row + 1):
            df_idx = sheet_row - data_start_row
            sucursal = df_hoja_ordered["Sucursal"].iloc[df_idx]
            if sucursal is not None:
                section_rows_by_label[sucursal] = sheet_row

        cc_group_labels = [l for l in _CC_FAMILY if l in section_rows_by_label]
        # Numbered labels in section: keep order, dedupe
        numbered_labels_unique: list[str] = []
        seen: set[str] = set()
        for sheet_row in range(sec_start_row, sec_end_row + 1):
            df_idx = sheet_row - data_start_row
            lbl = df_hoja_ordered["Sucursal"].iloc[df_idx]
            if lbl is None or lbl in _NON_SUMMING_LABELS:
                continue
            if lbl not in seen:
                seen.add(lbl)
                numbered_labels_unique.append(lbl)

        directa_labels = (
            ["DIRECTA SUCURSALES"] if "DIRECTA SUCURSALES" in section_rows_by_label else []
        )

        def _rows_for(labels):
            return sorted(
                section_rows_by_label[l] for l in labels if l in section_rows_by_label
            )

        subtotal_definitions = {
            _SUBTOTAL_CC: _rows_for(cc_group_labels),
            _SUC_SIN_DIRECTA: _rows_for(numbered_labels_unique),
            _TOTAL_SIN_SMK: (
                ([section_rows_by_label[_SUBTOTAL_CC]] if _SUBTOTAL_CC in section_rows_by_label else [])
                + ([section_rows_by_label[_SUC_SIN_DIRECTA]] if _SUC_SIN_DIRECTA in section_rows_by_label else [])
                + _rows_for(directa_labels)
            ),
        }

        for subtotal_label, source_rows in subtotal_definitions.items():
            if subtotal_label not in section_rows_by_label:
                continue
            subtotal_sheet_row = section_rows_by_label[subtotal_label]

            fill_color = _SUBTOTAL_FILLS.get(subtotal_label, "D9D9D9")
            subtotal_fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid",
            )
            subtotal_font = Font(name=_FONT_NAME, bold=True, color=_SUBTOTAL_FONT_COLOR)
            subtotal_font_semaforo = Font(
                name=_FONT_NAME, bold=True, color=_SEMAFORO_FONT_COLOR
            )

            for col_name in sum_cols:
                col_letter = col_map.get(col_name)
                if col_letter is None:
                    continue
                if not source_rows:
                    formula = None
                elif len(source_rows) == 1:
                    formula = f"=SUM({col_letter}{source_rows[0]})"
                else:
                    refs = ",".join(f"{col_letter}{r}" for r in source_rows)
                    formula = f"=SUM({refs})"

                if formula and col_letter:
                    cell = ws[f"{col_letter}{subtotal_sheet_row}"]
                    cell.value = formula

            tend_col = col_map.get("Tend vs Obj (%)")
            tend_col_num = col_map.get("Tendencia")
            obj_col_num = col_map.get("Objetivo")
            if tend_col and tend_col_num and obj_col_num:
                formula_tend = (
                    f"=IF({obj_col_num}{subtotal_sheet_row}=0,\"\","
                    f"{tend_col_num}{subtotal_sheet_row}/{obj_col_num}{subtotal_sheet_row})"
                )
                ws[f"{tend_col}{subtotal_sheet_row}"].value = formula_tend

            # La columna del semaforo lleva fuente NEGRA, no la blanca del
            # subtotal: el ColorScaleRule le repinta el relleno con los pasteles
            # (rojo FF6366 / amarillo FFEB9C / verde C6EFCE) y la blanca queda
            # invisible encima. Es la fila que mas se mira del informe.
            tend_col_idx = column_index_from_string(tend_col) if tend_col else None
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=subtotal_sheet_row, column=col_idx)
                cell.font = subtotal_font_semaforo if col_idx == tend_col_idx else subtotal_font
                cell.fill = subtotal_fill
                cell.border = thin_border

    # -------------------------------------------------------------------
    # Header styling: dark blue fill, white bold font, thin border
    # -------------------------------------------------------------------
    header_fill = PatternFill(
        start_color=_HEADER_FILL_COLOR, end_color=_HEADER_FILL_COLOR, fill_type="solid",
    )
    header_font = Font(name=_FONT_NAME, bold=True, color=_HEADER_FONT_COLOR)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # -------------------------------------------------------------------
    # Thin borders on every data cell (skip subtotal rows already styled)
    # -------------------------------------------------------------------
    subtotal_rows_set: set[int] = set()
    for sec_start_row, sec_end_row in [(s, e) for _g, s, e in sections]:
        for sheet_row in range(sec_start_row, sec_end_row + 1):
            df_idx = sheet_row - data_start_row
            sucursal = df_hoja_ordered["Sucursal"].iloc[df_idx]
            if sucursal in _SUBTOTAL_FILLS:
                subtotal_rows_set.add(sheet_row)
    for row_num in range(data_start_row, data_end_row + 1):
        if row_num in subtotal_rows_set:
            continue  # already styled above
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_num, column=col_idx).border = thin_border

    # -------------------------------------------------------------------
    # T-022: Apply ColorScaleRule on Tend vs Obj (%) column
    # -------------------------------------------------------------------
    tend_obj_col_letter = col_map.get("Tend vs Obj (%)")
    if tend_obj_col_letter and data_start_row <= data_end_row:
        heatmap_range = f"{tend_obj_col_letter}{data_start_row}:{tend_obj_col_letter}{data_end_row}"
        # Paleta pastel consistente con el reporte de rebotes (rechazos):
        # rojo FF6366 / amarillo FFEB9C / verde C6EFCE.
        color_scale = ColorScaleRule(
            start_type="num",   start_value=0,   start_color="FF6366",  # rojo pastel
            mid_type="num",     mid_value=1.0,   mid_color="FFEB9C",    # amarillo pastel
            end_type="num",     end_value=1.2,   end_color="C6EFCE",    # verde pastel
        )
        ws.conditional_formatting.add(heatmap_range, color_scale)

    # -------------------------------------------------------------------
    # Caveat note row was inserted at the start of this function (row 1).
    # Now style it: merge across, amber background, bold.
    # -------------------------------------------------------------------
    if note:
        last_col = max(n_cols, 1)
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=last_col
        )
        cell = ws.cell(row=1, column=1, value=note)
        cell.fill = PatternFill(
            start_color="FFE699", end_color="FFE699", fill_type="solid"
        )  # amber
        cell.font = Font(name=_FONT_NAME, bold=True, color="7F6000")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        ws.row_dimensions[1].height = 22

    # -------------------------------------------------------------------
    # Sheet-wide font family override (last pass — preserves bold/color/size)
    # -------------------------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            existing = cell.font
            cell.font = Font(
                name=_FONT_NAME,
                size=existing.size,
                bold=existing.bold,
                italic=existing.italic,
                color=existing.color,
            )


def _load_categorias_deposito(path: str) -> list[dict]:
    """
    Carga master-data de categorias_deposito desde JSON.

    Path absoluto o relativo (resuelto desde CWD). Cada entrada debe tener
    'codigo', 'concatenar' y 'division'. Retorna [] ante cualquier error.
    """
    p = Path(path)
    if not p.is_absolute():
        p = Path.cwd() / p
    if not p.exists():
        logger.warning("categorias_deposito source not found: %s — skipping", p)
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("categorias_deposito JSON parse failed (%s) — skipping", exc)
        return []
    if not isinstance(data, list):
        logger.warning(
            "categorias_deposito JSON must be array, got %s — skipping",
            type(data).__name__,
        )
        return []
    return data


def _build_categorias_sheet(workbook, entradas: list[dict]) -> None:
    """Crea hoja 'Categorias' con columnas codigo / concatenar / division."""
    if "Categorias" in workbook.sheetnames:
        del workbook["Categorias"]
    ws = workbook.create_sheet("Categorias")
    ws.append(["codigo", "concatenar", "division"])
    for entry in entradas:
        ws.append([
            entry.get("codigo"),
            entry.get("concatenar"),
            entry.get("division"),
        ])


def _add_lookup_columns_to_detalle(ws) -> None:
    """
    Agrega columnas 'Concatenar', 'División' y 'Generico' al final de la hoja:
      - Concatenar  =Serie & Descripción Depósito  (col C & col L)
      - División    =IFERROR(VLOOKUP(<concat>,Categorias!B:C,2,FALSE),"")
      - Generico    =IFERROR(VLOOKUP(<articulo>,dim_articulo!A:B,2,FALSE),"")
                    (Articulo en col R = col 18)
    """
    if ws.max_row < 2:
        return
    concat_col = ws.max_column + 1
    division_col = concat_col + 1
    generico_col = division_col + 1
    concat_letter = get_column_letter(concat_col)

    ws.cell(1, concat_col).value = "Concatenar"
    ws.cell(1, division_col).value = "División"
    ws.cell(1, generico_col).value = "Generico"

    for row in range(2, ws.max_row + 1):
        ws.cell(row, concat_col).value = f"=C{row}&L{row}"
        ws.cell(row, division_col).value = (
            f'=IFERROR(VLOOKUP({concat_letter}{row},Categorias!B:C,2,FALSE),"")'
        )
        ws.cell(row, generico_col).value = (
            f'=IFERROR(VLOOKUP(R{row},dim_articulo!A:B,2,FALSE),"")'
        )


def _add_division_totals(
    ws,
    generico_name: str,
    fecha_desde: str,
    info_dias: dict,
    divisions: list[str],
    detalle_sheet_actual: str | None,
    detalle_sheet_ma: str | None,
    detalle_sheet_mmaa: str | None,
) -> None:
    """
    Agrega al pie de una hoja de generico filas de totales por division (SUPERMERCADOS X).

    Cada fila usa SUMIFS contra las hojas de Detalle Movimientos correspondientes,
    filtrando por (Generico = nombre de la hoja, División = division).
    Las cols T (Bultos), AD (División) y AE (Generico) tienen que existir en los detalles.

    Columnas de la hoja:
        A=Sucursal | B=Generico | C=DIA1 | D=DIA2 | E=Total Ventas |
        F=Tendencia | G=MMAA | H=MA | I=Objetivo | J=Tend vs Obj %
    """
    # Buscar fila de header
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Sucursal":
            header_row = r
            break
    if header_row is None:
        return

    # Parsear los headers de DIA1 y DIA2 (formato "dd-mm DiaSemana")
    def _parse_day_header(value) -> tuple[int, int] | None:
        if not value:
            return None
        first = str(value).split()[0]
        parts = first.split("-")
        if len(parts) != 2:
            return None
        try:
            return int(parts[1]), int(parts[0])  # (month, day)
        except ValueError:
            return None

    year = int(fecha_desde[:4])
    dia1 = _parse_day_header(ws.cell(header_row, 3).value)
    dia2 = _parse_day_header(ws.cell(header_row, 4).value)

    dias_habiles = info_dias.get("Dias Habiles", 0)
    dias_trans = info_dias.get("Dias Transcurridos", 0)

    last_row = ws.max_row
    start_row = last_row + 2  # una fila vacia de separacion

    def _sumifs(sheet: str, with_date: tuple[int, int] | None, row: int) -> str:
        base = (
            f"=SUMIFS('{sheet}'!T:T,"
            f"'{sheet}'!AE:AE,B{row},"
            f"'{sheet}'!AD:AD,A{row}"
        )
        if with_date is not None:
            month, day = with_date
            base += f",'{sheet}'!F:F,DATE({year},{month},{day})"
        base += ")"
        return base

    for i, division in enumerate(divisions):
        r = start_row + i
        ws.cell(r, 1).value = division
        ws.cell(r, 2).value = generico_name
        if dia1 and detalle_sheet_actual:
            ws.cell(r, 3).value = _sumifs(detalle_sheet_actual, dia1, r)
        if dia2 and detalle_sheet_actual:
            ws.cell(r, 4).value = _sumifs(detalle_sheet_actual, dia2, r)
        if detalle_sheet_actual:
            ws.cell(r, 5).value = _sumifs(detalle_sheet_actual, None, r)
        if dias_trans > 0:
            ws.cell(r, 6).value = f"=E{r}/{dias_trans}*{dias_habiles}"
        if detalle_sheet_mmaa:
            ws.cell(r, 7).value = _sumifs(detalle_sheet_mmaa, None, r)
        if detalle_sheet_ma:
            ws.cell(r, 8).value = _sumifs(detalle_sheet_ma, None, r)
        # I (Objetivo) y J (Tend vs Obj %) quedan en blanco — no hay cupos por division.
        # Formato entero para filas de supermercados (sin redondear, solo visual)
        for col in range(3, 9):  # C..H (DIA1, DIA2, Total, Tend, MMAA, MA)
            ws.cell(r, col).number_format = '#,##0'


def _build_dim_articulo_sheet(workbook, df: pd.DataFrame) -> None:
    """
    Crea hoja 'dim_articulo' con id_articulo, generico, marca, descripcion.
    Usada como lookup table para los VLOOKUP en hojas de detalle.
    """
    if "dim_articulo" in workbook.sheetnames:
        del workbook["dim_articulo"]
    ws = workbook.create_sheet("dim_articulo")
    ws.append(["id_articulo", "generico", "marca", "des_articulo"])
    for _, row in df.iterrows():
        ws.append([
            row.get("id_articulo"),
            row.get("generico"),
            row.get("marca"),
            row.get("des_articulo"),
        ])


@dataclass
class _SheetSection:
    """One section within a logical sheet (a marca_split group or the whole generico)."""
    label: str              # human-readable label; equals the logical generico for unsplit sheets
    df: pd.DataFrame        # ordered rows + 3 injected subtotal rows


@dataclass
class _SheetStruct:
    """One logical sheet (one entry per generico in the final Excel / JSON)."""
    logical_generico: str               # sheet name, e.g. "VINOS FINOS"
    sections: list[_SheetSection] = field(default_factory=list)
    note: str | None = None             # amber PRVTA-exclusion note when applicable


class ResumenMensualService(BaseService):
    """
    Servicio para generacion de reportes de resumen mensual.

    Orquesta el flujo completo: extraccion de los 4 DataFrames, aplicacion de
    zonas virtuales, procesamiento y generacion del Excel con una hoja por generico.
    """

    SERVICE_SLUG = "resumen-mensual"
    GRANULARITY = "month"

    def _preparar_datos(
        self, config: ResumenMensualConfig
    ) -> tuple[pd.DataFrame, dict, str, str, pd.DataFrame]:
        """
        Extrae, transforma y procesa los datos para el reporte de resumen mensual.

        Shared pipeline consumed by both generar_reporte (Excel sink) and
        generar_datos (JSON sink). Returns all the artifacts needed to build
        either output without duplicating business logic.

        Args:
            config: Configuracion del reporte.

        Returns:
            Tuple of (df_resultado, info_dias, col_n1, col_n2, df_dias) where:
              - df_resultado: processed DataFrame with 10 columns in fixed order
              - info_dias: dict with Dias Habiles / Transcurridos / Faltantes
              - col_n1: dynamic name of the N-1 day column (e.g. "09-06 Martes")
              - col_n2: dynamic name of the N-2 day column (e.g. "08-06 Lunes")
              - df_dias: processed daily sales DataFrame (used for filename derivation)
        """
        # Normalizar genericos: lista vacia se trata como None (traer todos)
        genericos = config.genericos if config.genericos else None

        # PRVTA exclusion: None → default; explicit [] disables; explicit list overrides
        sin_prvta = (
            config.genericos_sin_prvta
            if config.genericos_sin_prvta is not None
            else list(_DEFAULT_GENERICOS_SIN_PRVTA)
        )

        marca_splits = config.marca_splits or {}

        # -----------------------------------------------------------------
        # 1. Fetch de los 4 DataFrames
        # -----------------------------------------------------------------
        df_ventas_mes = self.data_loader.get_ventas_resumen_mensual(
            config.fecha_desde,
            config.fecha_hasta,
            genericos,
            genericos_sin_prvta=sin_prvta,
            marca_splits=marca_splits or None,
        )
        df_dias = self.data_loader.get_ventas_ultimos_dias_habiles(
            config.fecha_desde,
            config.fecha_hasta,
            genericos,
            genericos_sin_prvta=sin_prvta,
            marca_splits=marca_splits or None,
        )

        try:
            df_ventas_ma = self.data_loader.get_ventas_mes_anterior(
                config.fecha_desde,
                genericos,
                genericos_sin_prvta=sin_prvta,
                marca_splits=marca_splits or None,
            )
        except Exception:
            df_ventas_ma = pd.DataFrame(columns=["sucursal", "generico", "cantidad"])

        try:
            df_ventas_aa = self.data_loader.get_ventas_mismo_mes_anio_anterior(
                config.fecha_desde,
                config.fecha_hasta,
                genericos,
                genericos_sin_prvta=sin_prvta,
                marca_splits=marca_splits or None,
            )
        except Exception:
            df_ventas_aa = pd.DataFrame(columns=["sucursal", "generico", "cantidad"])

        # -----------------------------------------------------------------
        # 2. Segregar DIRECTA SUCURSALES y aplicar zonas virtuales.
        #    IMPORTANTE: segregar PRIMERO (necesita id_ruta), zonas DESPUÉS
        #    (drops id_ruta). Orden inverso pierde la segregación.
        # -----------------------------------------------------------------
        df_ventas_mes = _segregar_directa_sucursales(df_ventas_mes)
        df_ventas_mes = aplicar_zonas_virtuales(df_ventas_mes)
        if not df_ventas_mes.empty:
            df_ventas_mes = df_ventas_mes.groupby(
                ["sucursal", "generico"], as_index=False
            )["cantidad"].sum()

        df_dias = _segregar_directa_sucursales(df_dias)
        df_dias = aplicar_zonas_virtuales(df_dias)
        if not df_dias.empty:
            df_dias = df_dias.groupby(
                ["sucursal", "generico", "fecha"], as_index=False
            )["cantidad"].sum()

        df_ventas_ma = _segregar_directa_sucursales(df_ventas_ma)
        df_ventas_ma = aplicar_zonas_virtuales(df_ventas_ma)
        if not df_ventas_ma.empty:
            df_ventas_ma = df_ventas_ma.groupby(
                ["sucursal", "generico"], as_index=False
            )["cantidad"].sum()

        df_ventas_aa = _segregar_directa_sucursales(df_ventas_aa)
        df_ventas_aa = aplicar_zonas_virtuales(df_ventas_aa)
        if not df_ventas_aa.empty:
            df_ventas_aa = df_ventas_aa.groupby(
                ["sucursal", "generico"], as_index=False
            )["cantidad"].sum()

        # -----------------------------------------------------------------
        # 3. Calcular info de dias habiles
        # -----------------------------------------------------------------
        info_dias = calcular_info_dias(config.fecha_desde, config.fecha_hasta)

        # -----------------------------------------------------------------
        # 4. Fetch cupos and pass to processor
        # -----------------------------------------------------------------
        periodo = config.fecha_desde[:7]  # "YYYY-MM-DD" → "YYYY-MM"
        genericos_for_cupos = genericos if genericos else _DEFAULT_GENERICOS

        # Expandir genericos lógicos a sub-marcas reales en fact_cupos
        # (e.g. "VINOS" → ["ARIZU", "CANCILLER"]). Mantener un mapeo inverso
        # para volver a renombrar después del query.
        # Si hay marca_splits para un genérico, el cupo "real" se renombra a
        # "{generico} (sin {marcas})" y cada marca listada se trae por separado.
        genericos_reales: list[str] = []
        map_real_a_logico: dict[str, str] = {}
        for g in genericos_for_cupos:
            partes = _CUPOS_GENERICO_PARTS.get(g, [g])
            marcas_split = (marca_splits or {}).get(g, [])
            if marcas_split:
                logico_sin = f"{g} (sin {', '.join(marcas_split)})"
                for parte in partes:
                    genericos_reales.append(parte)
                    map_real_a_logico[parte] = logico_sin
                for marca in marcas_split:
                    genericos_reales.append(marca)
                    map_real_a_logico[marca] = marca
            else:
                for parte in partes:
                    genericos_reales.append(parte)
                    map_real_a_logico[parte] = g

        try:
            df_cupos_raw = self.data_loader.get_cupos_resumen_mensual(
                periodo, genericos_reales
            )
        except Exception as exc:
            logger.warning("get_cupos_resumen_mensual failed (%s) — Objetivo column will be blank", exc)
            df_cupos_raw = pd.DataFrame(columns=["sucursal", "generico", "cupo"])

        # Renombrar sub-marcas al genérico lógico (ARIZU/CANCILLER → VINOS)
        if not df_cupos_raw.empty and "generico" in df_cupos_raw.columns:
            df_cupos_raw["generico"] = (
                df_cupos_raw["generico"].map(map_real_a_logico).fillna(df_cupos_raw["generico"])
            )

        # Apply segregar + zonas to cupos (segregar PRIMERO; ver nota arriba)
        if not df_cupos_raw.empty and "id_ruta" in df_cupos_raw.columns:
            df_cupos_raw = _segregar_directa_sucursales(df_cupos_raw)
            df_cupos_raw = aplicar_zonas_virtuales(df_cupos_raw)
            df_cupos_raw = df_cupos_raw.groupby(
                ["sucursal", "generico"], as_index=False
            )["cupo"].sum()
        elif not df_cupos_raw.empty:
            # No id_ruta present — just groupby to be safe
            df_cupos_raw = df_cupos_raw.groupby(
                ["sucursal", "generico"], as_index=False
            )["cupo"].sum()

        # Inyectar cupos manuales (sucursales que no se cargan en fact_cupos)
        if config.cupos_manuales:
            manual_rows = [
                {"sucursal": sucursal, "generico": generico, "cupo": cupo}
                for sucursal, gens in config.cupos_manuales.items()
                for generico, cupo in gens.items()
            ]
            if manual_rows:
                df_manual = pd.DataFrame(manual_rows)
                df_cupos_raw = pd.concat([df_cupos_raw, df_manual], ignore_index=True)
                df_cupos_raw = df_cupos_raw.groupby(
                    ["sucursal", "generico"], as_index=False
                )["cupo"].sum()

        # -----------------------------------------------------------------
        # 5. Procesar datos (una llamada para todos los genericos)
        # -----------------------------------------------------------------
        df_resultado = procesar_resumen_mensual(
            df_ventas_mes,
            df_dias,
            df_ventas_ma,
            df_ventas_aa,
            config.fecha_desde,
            config.fecha_hasta,
            config.con_objetivo,
            df_cupos=df_cupos_raw,
        )

        # Detectar nombres dinámicos de columnas N-1 y N-2 (posiciones 2 y 3)
        cols = list(df_resultado.columns)
        col_n1 = cols[2] if len(cols) > 2 else "Vtas Dia N-1"
        col_n2 = cols[3] if len(cols) > 3 else "Vtas Dia N-2"

        return df_resultado, info_dias, col_n1, col_n2, df_dias

    def _build_sheet_structs(
        self,
        df_resultado: pd.DataFrame,
        marca_splits: dict[str, list[str]],
        genericos_sin_prvta: list[str],
    ) -> list[_SheetStruct]:
        """
        Build the ordered sectioned structure consumed by both the Excel writer
        and the JSON serializer.

        For each logical generico, constructs a _SheetStruct with one _SheetSection
        per marca_split group (or a single section for unsplit genericos). Each
        section's DataFrame has already been through _ordenar_e_inyectar_subtotales().

        Args:
            df_resultado: Processed DataFrame from procesar_resumen_mensual (10 cols).
            marca_splits: Mapping {generico: [marcas]} defining per-sheet splits.
            genericos_sin_prvta: Genericos for which the PRVTA-exclusion note applies.

        Returns:
            Ordered list of _SheetStruct (one per logical generico present in df_resultado).
        """
        if df_resultado.empty:
            return []

        # Map de genericos sinteticos (resultado de marca_splits) -> generico logico
        synthetic_to_logical: dict[str, str] = {}
        for logical_gen, marcas in (marca_splits or {}).items():
            synthetic_to_logical[f"{logical_gen} (sin {', '.join(marcas)})"] = logical_gen
            for marca in marcas:
                synthetic_to_logical[marca] = logical_gen

        # Build ordered list of logical genericos, preserving original order from df_resultado
        seen_logical: set[str] = set()
        logical_genericos: list[str] = []
        for syn in df_resultado["Generico"].tolist():
            logical = synthetic_to_logical.get(syn, syn)
            if logical not in seen_logical:
                seen_logical.add(logical)
                logical_genericos.append(logical)

        structs: list[_SheetStruct] = []
        for logical_gen in logical_genericos:
            # Section order: "sin {marcas}" first, then each marca
            if logical_gen in (marca_splits or {}):
                marcas = marca_splits[logical_gen]
                section_order = [f"{logical_gen} (sin {', '.join(marcas)})"] + list(marcas)
                section_order = [s for s in section_order if s in df_resultado["Generico"].values]
            else:
                section_order = [logical_gen]

            sections: list[_SheetSection] = []
            for syn_gen in section_order:
                df_section = df_resultado[df_resultado["Generico"] == syn_gen].copy()
                if df_section.empty:
                    continue
                sections.append(_SheetSection(
                    label=syn_gen,
                    df=_ordenar_e_inyectar_subtotales(df_section),
                ))

            if not sections:
                continue

            # Compute PRVTA-exclusion note
            note: str | None = None
            if logical_gen in (genericos_sin_prvta or []):
                note = f"Nota: {logical_gen} excluye documentos PRVTA (facturas presupuesto)"

            structs.append(_SheetStruct(
                logical_generico=logical_gen,
                sections=sections,
                note=note,
            ))

        return structs

    def generar_datos(self, config: ResumenMensualConfig) -> dict:
        """
        Generate report data as a JSON-serializable dict (no file written).

        Runs the same extraction + processing pipeline as generar_reporte but
        returns the structured JSON contract instead of writing an Excel file.
        Retained as the structured-data path used by the view-vs-Excel oracle
        cross-check (tests/test_v_resumen_mensual_oracle.py). The former
        POST /resumen-mensual/datos endpoint and React frontend were removed
        when the report migrated to the Superset dashboard.

        Args:
            config: Configuracion del reporte.

        Returns:
            Dict matching the JSON contract (meta + sheets structure).
        """
        from src.services.resumen_mensual.serializer import to_datos_json

        df_resultado, info_dias, col_n1, col_n2, _df_dias = self._preparar_datos(config)

        marca_splits = config.marca_splits or {}
        sin_prvta_effective = (
            config.genericos_sin_prvta
            if config.genericos_sin_prvta is not None
            else list(_DEFAULT_GENERICOS_SIN_PRVTA)
        )

        structs = self._build_sheet_structs(df_resultado, marca_splits, sin_prvta_effective)

        return to_datos_json(
            structs,
            info_dias,
            col_n1,
            col_n2,
            con_objetivo=config.con_objetivo,
            fecha_desde=config.fecha_desde,
            fecha_hasta=config.fecha_hasta,
        )

    def generar_reporte(self, config: ResumenMensualConfig) -> ResumenMensualResult:
        """
        Genera un reporte de resumen mensual.

        Genera un archivo Excel con una hoja por cada generico presente en los datos.
        Cada hoja contiene: Vtas Dia N-1, N-2, Total Ventas, Tendencia,
        Ventas Mes Anterior, Ventas Mismo Mes AA, Objetivo, Tend vs Obj (%).

        Args:
            config: Configuracion del reporte.

        Returns:
            ResumenMensualResult con informacion del reporte generado.
        """
        df_resultado, info_dias, col_n1, col_n2, df_dias = self._preparar_datos(config)

        # -----------------------------------------------------------------
        # 6. Generar Excel: una hoja por generico
        # -----------------------------------------------------------------
        marca_splits = config.marca_splits or {}

        nombre = config.nombre_archivo or _nombre_reporte(df_dias, config.fecha_hasta)
        out = self._output_dir(config.fecha_desde)
        out.mkdir(parents=True, exist_ok=True)

        # Merge mode: find existing xlsx in the output folder
        existing_files = list(out.glob("*.xlsx"))
        if len(existing_files) > 1:
            raise RuntimeError(
                f"Found {len(existing_files)} xlsx files in {out}, expected at most 1. "
                f"Please clean up old files: {[str(p.name) for p in existing_files]}"
            )
        merge_target = existing_files[0] if existing_files else None

        # When merging, preserve the existing filename
        if merge_target:
            nombre = merge_target.stem

        writer = ExcelWriter(nombre, output_dir=out, merge_with=merge_target)

        style = _crear_estilo_resumen(info_dias, col_n1, col_n2)
        summary_rows_count = len(info_dias)  # used for row offset calculation

        # Resolve effective genericos_sin_prvta for note signal
        sin_prvta_effective = (
            config.genericos_sin_prvta
            if config.genericos_sin_prvta is not None
            else list(_DEFAULT_GENERICOS_SIN_PRVTA)
        )

        # Build the ordered, sectioned, subtotal-injected structure
        structs = self._build_sheet_structs(df_resultado, marca_splits, sin_prvta_effective)

        logical_genericos = [s.logical_generico for s in structs]

        for struct in structs:
            section_dfs = [sec.df for sec in struct.sections]
            df_hoja_ordered = pd.concat(section_dfs, ignore_index=True)
            sheet_name = struct.logical_generico[:31]  # Excel max 31 caracteres
            ws = writer.add_sheet(df_hoja_ordered, sheet_name=sheet_name, style=style)

            # T-020/T-022/T-023: post-write — resolve SUM formulas, heatmap, subtotal styling
            _post_write_subtotals_and_heatmap(
                ws, df_hoja_ordered, summary_rows_count, note=struct.note
            )

        # T-09: Import Detalle Movimientos sheets from external sources
        # 3 hojas: actual + mes anterior (MA) + mismo mes año anterior (MMAA).
        detalle_imports = [
            (config.detalle_movimientos_path, "Detalle Movimientos"),
            (config.detalle_movimientos_ma_path, "Detalle Movimientos MA"),
            (config.detalle_movimientos_mmaa_path, "Detalle Movimientos MMAA"),
        ]
        imported_detalles: list[str] = []
        for src_str, sheet_name in detalle_imports:
            if not src_str:
                continue
            src_path = Path(src_str)
            try:
                rows_imported = import_xlsx_as_sheet(
                    writer.workbook, src_path, sheet_name
                )
                imported_detalles.append(sheet_name)
                logger.info(
                    "%s imported: %d rows from %s", sheet_name, rows_imported, src_path
                )
            except FileNotFoundError:
                logger.warning(
                    "%s source not found: %s — skipping", sheet_name, src_path
                )
            except Exception as exc:
                logger.warning(
                    "%s import failed (%s) — skipping", sheet_name, exc
                )

        # T-10: dim_articulo + Categorias deposito → hojas lookup + columns en cada Detalle.
        if imported_detalles:
            # dim_articulo (siempre que haya al menos 1 detalle importado)
            try:
                df_dim = self.data_loader.get_dim_articulo()
                if not df_dim.empty:
                    _build_dim_articulo_sheet(writer.workbook, df_dim)
                    logger.info("dim_articulo sheet escrita: %d filas", len(df_dim))
            except Exception as exc:
                logger.warning("dim_articulo fetch failed (%s) — skipping", exc)

            # Categorias (solo si hay path configurado)
            if config.categorias_deposito_path:
                entradas = _load_categorias_deposito(config.categorias_deposito_path)
                if entradas:
                    _build_categorias_sheet(writer.workbook, entradas)

            # Lookup columns en cada hoja de detalle
            for sheet_name in imported_detalles:
                if sheet_name in writer.workbook.sheetnames:
                    _add_lookup_columns_to_detalle(writer.workbook[sheet_name])
            logger.info(
                "Lookup columns (Concatenar/División/Generico) aplicados a %d hojas de detalle",
                len(imported_detalles),
            )

            # T-11: Totales por division SUPERMERCADOS al pie de cada hoja de generico.
            # Solo si tenemos categorias_deposito (de donde sacamos la lista de divisiones)
            # y al menos detalle_movimientos del periodo actual importado.
            if config.categorias_deposito_path and "Detalle Movimientos" in imported_detalles:
                entradas_for_div = _load_categorias_deposito(config.categorias_deposito_path)
                if entradas_for_div:
                    # divisiones unicas, preservando orden de aparicion
                    seen = {}
                    for e in entradas_for_div:
                        d = e.get("division")
                        if d and d not in seen:
                            seen[d] = True
                    divisions = list(seen.keys())

                    excluded = {
                        "Sheet1", "dim_articulo", "Categorias",
                        *imported_detalles,
                    }
                    for sheet_name in list(writer.workbook.sheetnames):
                        if sheet_name in excluded:
                            continue
                        _add_division_totals(
                            writer.workbook[sheet_name],
                            generico_name=sheet_name,
                            fecha_desde=config.fecha_desde,
                            info_dias=info_dias,
                            divisions=divisions,
                            detalle_sheet_actual="Detalle Movimientos" if "Detalle Movimientos" in imported_detalles else None,
                            detalle_sheet_ma="Detalle Movimientos MA" if "Detalle Movimientos MA" in imported_detalles else None,
                            detalle_sheet_mmaa="Detalle Movimientos MMAA" if "Detalle Movimientos MMAA" in imported_detalles else None,
                        )
                    logger.info(
                        "Totales por division agregados a hojas de generico (divisiones: %s)",
                        divisions,
                    )

        ruta = writer.save()

        # -----------------------------------------------------------------
        # 6. Construir y retornar resultado
        # -----------------------------------------------------------------
        sucursales_unicas = (
            df_resultado["Sucursal"].nunique() if not df_resultado.empty else 0
        )

        return ResumenMensualResult(
            ruta_archivo=ruta,
            registros_procesados=len(df_resultado),
            sucursales=sucursales_unicas,
            genericos_incluidos=logical_genericos,
            hojas=[g[:31] for g in logical_genericos],
        )
