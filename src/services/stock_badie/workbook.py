"""Workbook builder for stock-badie: STOCK sheet skeleton, values, the core
live Excel formulas (RF-06/RF-07), the per-generico totals band (RF-08),
the TOTAL GENERAL row (RF-09), and number/conditional formatting (RF-10).

Consumes the wide DataFrame produced by ``processor.pivot_wide()`` — a
MultiIndex-columned frame with 4 identity columns, 14 sucursal blocks of
[Stock, VENTA, PEDIDO, ALCANCE] (PEDIDO/ALCANCE are ``None`` placeholders
there), and a Total block [Total, VENTA TOTAL, PEDIDO TOTAL, ALCANCE TOTAL]
(also placeholders). This module fills those placeholders with LIVE Excel
formulas — never computed in Python — so ``DiasStock`` stays an interactive
knob in the delivered workbook. Stock/VENTA cells are written as plain
values pulled straight from ``wide_df`` (never formulas, never rounded).

Row layout is COMPUTED (see ``compute_layout``), not hardcoded past row 3:
the per-generico band (RF-08) sits above the table header (user
requirement), so the header/data/TOTAL GENERAL rows shift depending on how
many distinct genericos and articulo rows a given run produces. Column
letters never move — only sucursal block/Total block ROW numbers do — so
every formula below is generated relative to ``compute_layout``'s output.
"""

from __future__ import annotations

from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

SHEET_NAME = "STOCK"

# ── Param cells (rows 1-2) ──────────────────────────────────────────────
DIAS_STOCK_ROW = 1
DIAS_VENTA_ROW = 2
PARAM_LABEL_COL = 1
PARAM_VALUE_COL = 2
DIAS_STOCK_CELL = f"{get_column_letter(PARAM_VALUE_COL)}{DIAS_STOCK_ROW}"
DIAS_VENTA_CELL = f"{get_column_letter(PARAM_VALUE_COL)}{DIAS_VENTA_ROW}"

# Row 3 is always a blank spacer below the params. The per-generico band
# (RF-08) starts right below it — this is the ONLY row number below the
# params that stays a fixed constant; everything else (header, data, TOTAL
# GENERAL) is derived via compute_layout() because the band's height
# depends on how many distinct genericos are present in a given run.
BAND_START_ROW = 4

_IDENTITY_HEADERS = ["idArticulo", "dsArticulo", "GENERICO", "MARCA"]
_N_IDENTITY = len(_IDENTITY_HEADERS)  # 4
_BLOCK_WIDTH = 4  # Stock, VENTA, PEDIDO, ALCANCE
_FIRST_BLOCK_COL = _N_IDENTITY + 1  # column 5 (E)
_GENERICO_COL = 3  # identity column C — also the per-generico band's SUMIFS criteria column

_TOTAL_HEADERS = ["Total", "VENTA TOTAL", "PEDIDO TOTAL", "ALCANCE TOTAL"]

# RF-10: accounting number format shared by every numeric cell (band, data
# Stock/Venta/Pedido/Alcance, Total block, TOTAL GENERAL). Identity columns
# (A/B/D) and the band's GENERICO label cell (C) stay General.
NUMBER_FORMAT = r'_-* #,##0_-;\-* #,##0_-;_-* "-"??_-;_-@_-'

# RF-10: the FAITHFUL Stock-vs-Pedido semaforo (NOT a 3-band alcance rule).
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# RF-09: TOTAL GENERAL row styling (bold + fill — project "totals row" rule).
_TOTAL_GENERAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_TOTAL_GENERAL_FONT = Font(bold=True)
_TOTAL_GENERAL_LABEL = "TOTAL GENERAL"

# ── Visual styling: colored headers + cell borders ──────────────────────
# Header palette. Identity columns and each sucursal's name cell get the
# dark tone; the VENTA/PEDIDO/ALCANCE sub-headers get the medium tone so a
# reader can tell block boundaries apart at a glance across 64 columns.
# The Total block gets the darkest tone to read as "this is the rollup".
_HEADER_FILL_DARK = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FILL_MEDIUM = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FILL_TOTAL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Per-generico band: a light accent so the summary zone reads as separate
# from the article table below it.
_BAND_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_BAND_LABEL_FONT = Font(bold=True)

# Param cells (DiasStock / DiasVenta) — DiasStock is the interactive knob,
# so it gets a distinct "edit me" accent.
_PARAM_LABEL_FONT = Font(bold=True)
_PARAM_VALUE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

_THIN_SIDE = Side(style="thin", color="B4B4B4")
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

# Column widths: identity columns carry text (the article description is
# long), the numeric block columns only ever hold formatted numbers.
_WIDTH_ID_ARTICULO = 12
_WIDTH_DS_ARTICULO = 42
_WIDTH_GENERICO = 20
_WIDTH_MARCA = 18
_WIDTH_NUMERIC = 13
_HEADER_ROW_HEIGHT = 30


@dataclass(frozen=True)
class SheetLayout:
    """Row numbers for the STOCK sheet, computed from how many distinct
    genericos and articulo rows a given run produces. Column letters are
    NEVER affected by this — only row numbers shift."""

    band_header_row: int
    band_start_row: int
    band_end_row: int
    header_row: int
    data_start_row: int
    data_end_row: int
    total_general_row: int

    @property
    def has_data(self) -> bool:
        return self.data_end_row >= self.data_start_row


def compute_layout(n_genericos: int, n_articulos: int) -> SheetLayout:
    """Compute the STOCK sheet's row layout.

    Row 1-2: DiasStock/DiasVenta params. Row 3: blank spacer (fixed,
    BAND_START_ROW - 1). band_header_row: the totals band carries its OWN
    copy of the table header, so the summary reads as a self-contained
    table instead of bare numbers whose columns are only labelled further
    down. band_start_row..band_end_row: one row per distinct generico — an
    EMPTY range when n_genericos == 0 (e.g. an empty universe), never a
    crash. One blank spacer row follows the band. header_row: table header.
    data_start_row..data_end_row: one row per articulo — an empty range
    when n_articulos == 0 (header-only sheet). total_general_row sits
    directly below the last data row (or directly below the header when
    there are no data rows, via the same ``data_end_row + 1`` formula — no
    special-casing needed).
    """
    band_header_row = BAND_START_ROW
    band_start_row = band_header_row + 1
    band_end_row = band_start_row + n_genericos - 1
    header_row = band_end_row + 2  # +1 blank spacer row between band and header
    data_start_row = header_row + 1
    data_end_row = data_start_row + n_articulos - 1
    total_general_row = data_end_row + 1

    return SheetLayout(
        band_header_row=band_header_row,
        band_start_row=band_start_row,
        band_end_row=band_end_row,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        total_general_row=total_general_row,
    )


def _sucursal_block_keys(columns: pd.MultiIndex) -> list[str]:
    """Ordered, de-duplicated list of sucursal block keys (MultiIndex level
    0), excluding the identity ("") and Total blocks. Order is inherited
    from wide_df's column order, which processor.pivot_wide() fixes to
    SUCURSAL_ORDER regardless of input row order."""
    keys: list[str] = []
    seen: set[str] = set()
    for level0, _level1 in columns:
        if level0 in ("", "Total") or level0 in seen:
            continue
        seen.add(level0)
        keys.append(level0)
    return keys


def _generico_labels(wide_df: pd.DataFrame) -> list[str]:
    """Deterministic (sorted), de-duplicated list of GENERICO labels present
    in wide_df, for the per-generico band (RF-08). Blank labels (pivot_wide
    coalesces missing/NaN GENERICO to "") are excluded — they would produce
    a meaningless band row with an empty SUMIFS criteria cell."""
    if wide_df.empty:
        return []
    labels = wide_df[("", "GENERICO")].tolist()
    return sorted({str(label) for label in labels if str(label) != ""})


def _apply_number_format_row(ws, row: int, columns: range) -> None:
    for col in columns:
        ws.cell(row=row, column=col).number_format = NUMBER_FORMAT


def _apply_borders_row(ws, row: int, last_col: int) -> None:
    """Thin border on every cell of `row` from column A through `last_col`."""
    for col in range(PARAM_LABEL_COL, last_col + 1):
        ws.cell(row=row, column=col).border = _CELL_BORDER


def _write_header_labels(
    ws,
    row: int,
    sucursales: list[str],
    block_col_of: dict[str, int],
    total_col: int,
) -> None:
    """Write the table header labels on `row`: identity columns, then each
    sucursal's name over its Stock column with VENTA/PEDIDO/ALCANCE beside
    it, then the Total block. Used for BOTH the per-generico band header
    and the article-table header so the two tables read identically."""
    for i, label in enumerate(_IDENTITY_HEADERS, 1):
        ws.cell(row=row, column=i, value=label)

    for sucursal in sucursales:
        col = block_col_of[sucursal]
        ws.cell(row=row, column=col, value=sucursal)
        ws.cell(row=row, column=col + 1, value="VENTA")
        ws.cell(row=row, column=col + 2, value="PEDIDO")
        ws.cell(row=row, column=col + 3, value="ALCANCE")

    for i, label in enumerate(_TOTAL_HEADERS):
        ws.cell(row=row, column=total_col + i, value=label)


def _style_header_row(
    ws,
    header_row: int,
    sucursales: list[str],
    block_col_of: dict[str, int],
    total_col: int,
) -> None:
    """Colored, bold, wrapped, centered header row.

    Identity columns and each sucursal's NAME cell get the dark tone; the
    VENTA/PEDIDO/ALCANCE sub-headers get the medium tone so block
    boundaries stay readable across 64 columns; the Total block gets the
    darkest tone to read as the rollup.
    """
    for col in range(PARAM_LABEL_COL, _N_IDENTITY + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _HEADER_FILL_DARK
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT

    for sucursal in sucursales:
        block_col = block_col_of[sucursal]
        # Stock column carries the sucursal name -> dark tone.
        name_cell = ws.cell(row=header_row, column=block_col)
        name_cell.fill = _HEADER_FILL_DARK
        name_cell.font = _HEADER_FONT
        name_cell.alignment = _HEADER_ALIGNMENT
        # VENTA / PEDIDO / ALCANCE sub-headers -> medium tone.
        for offset in (1, 2, 3):
            sub_cell = ws.cell(row=header_row, column=block_col + offset)
            sub_cell.fill = _HEADER_FILL_MEDIUM
            sub_cell.font = _HEADER_FONT
            sub_cell.alignment = _HEADER_ALIGNMENT

    for offset in range(_BLOCK_WIDTH):
        cell = ws.cell(row=header_row, column=total_col + offset)
        cell.fill = _HEADER_FILL_TOTAL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT

    ws.row_dimensions[header_row].height = _HEADER_ROW_HEIGHT


def _group_block_detail_columns(ws, sucursales: list[str], block_col_of: dict[str, int]) -> None:
    """Outline-group each sucursal's VENTA/PEDIDO/ALCANCE columns so they
    collapse to just the Stock column (user request).

    Collapsing all 14 groups turns the sheet into a 14-sucursal stock
    overview; expanding one shows that sucursal's detail. Left EXPANDED on
    open (hidden=False) — the outline controls are present so a reader
    collapses with one click without losing data on first open.
    """
    for sucursal in sucursales:
        block_col = block_col_of[sucursal]
        ws.column_dimensions.group(
            get_column_letter(block_col + 1),  # VENTA
            get_column_letter(block_col + 3),  # ALCANCE
            outline_level=1,
            hidden=False,
        )


def _group_top_rows(ws, header_row: int) -> None:
    """Outline-group every row ABOVE the table header (params, spacer, the
    per-generico band, spacer) so the whole summary zone collapses to a
    single control and the article table jumps to the top of the screen.

    Left EXPANDED on open (hidden=False) — the control is there, the reader
    collapses with one click.
    """
    if header_row <= 1:
        return
    ws.row_dimensions.group(1, header_row - 1, outline_level=1, hidden=False)


def _set_column_widths(ws, sucursales: list[str], block_col_of: dict[str, int], last_col: int) -> None:
    """Identity columns sized for their text; every numeric block column
    gets a uniform width wide enough for the accounting format."""
    ws.column_dimensions[get_column_letter(1)].width = _WIDTH_ID_ARTICULO
    ws.column_dimensions[get_column_letter(2)].width = _WIDTH_DS_ARTICULO
    ws.column_dimensions[get_column_letter(3)].width = _WIDTH_GENERICO
    ws.column_dimensions[get_column_letter(4)].width = _WIDTH_MARCA
    for col in range(_FIRST_BLOCK_COL, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = _WIDTH_NUMERIC


def _write_generico_band_row(
    ws,
    band_row: int,
    generico: str,
    blocks: list[tuple[int, int, int, int]],
    data_start: int,
    data_end: int,
) -> None:
    """Write one per-generico SUMIFS band row (RF-08): live totals for
    `generico`, scoped to the article data range, across every sucursal
    block AND the Total block. AlcanceGen is a ratio of the SUMIFS'd
    Stock/Venta cells, NEVER a SUMIFS of the per-article ALCANCE column —
    sum of ratios != ratio of sums (same correction as RF-07/RF-09)."""
    ws.cell(row=band_row, column=_GENERICO_COL, value=generico)

    generico_letter = get_column_letter(_GENERICO_COL)
    criteria_range = f"${generico_letter}${data_start}:${generico_letter}${data_end}"
    criteria_cell = f"${generico_letter}${band_row}"

    for stock_col, venta_col, pedido_col, alcance_col in blocks:
        stock_letter = get_column_letter(stock_col)
        venta_letter = get_column_letter(venta_col)
        pedido_letter = get_column_letter(pedido_col)

        stock_range = f"{stock_letter}{data_start}:{stock_letter}{data_end}"
        venta_range = f"{venta_letter}{data_start}:{venta_letter}{data_end}"
        pedido_range = f"{pedido_letter}{data_start}:{pedido_letter}{data_end}"

        ws.cell(
            row=band_row, column=stock_col,
            value=f"=SUMIFS({stock_range},{criteria_range},{criteria_cell})",
        )
        ws.cell(
            row=band_row, column=venta_col,
            value=f"=SUMIFS({venta_range},{criteria_range},{criteria_cell})",
        )
        ws.cell(
            row=band_row, column=pedido_col,
            value=f"=SUMIFS({pedido_range},{criteria_range},{criteria_cell})",
        )

        stock_gen_ref = f"{stock_letter}{band_row}"
        venta_gen_ref = f"{venta_letter}{band_row}"
        ws.cell(
            row=band_row, column=alcance_col,
            value=f"=IFERROR({stock_gen_ref}/({venta_gen_ref}/DiasVenta),0)",
        )


def _write_total_general_row(
    ws,
    total_general_row: int,
    blocks: list[tuple[int, int, int, int]],
    data_start: int,
    data_end: int,
) -> None:
    """Write the bottom TOTAL GENERAL row (RF-09): SUM over every article
    data row for Stock/Venta/Pedido — per sucursal block AND the Total
    block — and the corrected ratio (NOT a SUM of the ALCANCE cells) for
    every ALCANCE column. Styled distinctly (bold + fill)."""
    ws.cell(row=total_general_row, column=PARAM_LABEL_COL, value=_TOTAL_GENERAL_LABEL)

    for stock_col, venta_col, pedido_col, alcance_col in blocks:
        stock_letter = get_column_letter(stock_col)
        venta_letter = get_column_letter(venta_col)
        pedido_letter = get_column_letter(pedido_col)

        stock_range = f"{stock_letter}{data_start}:{stock_letter}{data_end}"
        venta_range = f"{venta_letter}{data_start}:{venta_letter}{data_end}"
        pedido_range = f"{pedido_letter}{data_start}:{pedido_letter}{data_end}"

        ws.cell(row=total_general_row, column=stock_col, value=f"=SUM({stock_range})")
        ws.cell(row=total_general_row, column=venta_col, value=f"=SUM({venta_range})")
        ws.cell(row=total_general_row, column=pedido_col, value=f"=SUM({pedido_range})")

        stock_sum_ref = f"{stock_letter}{total_general_row}"
        venta_sum_ref = f"{venta_letter}{total_general_row}"
        ws.cell(
            row=total_general_row, column=alcance_col,
            value=f"=IFERROR({stock_sum_ref}/({venta_sum_ref}/DiasVenta),0)",
        )

    last_col = blocks[-1][3] if blocks else PARAM_LABEL_COL
    for col in range(PARAM_LABEL_COL, last_col + 1):
        cell = ws.cell(row=total_general_row, column=col)
        cell.font = _TOTAL_GENERAL_FONT
        cell.fill = _TOTAL_GENERAL_FILL


def _apply_stock_vs_pedido_conditional_formatting(
    ws,
    stock_cols: list[int],
    data_start: int,
    data_end: int,
) -> None:
    """RF-10: the FAITHFUL semaforo. For each Stock column (14 per-sucursal
    + the Total-block Stock), over the DATA rows only: Stock < its block's
    Pedido (stock col + 2) -> RED; Stock > Pedido -> GREEN. Uses relative
    FormulaRule refs anchored at data_start so the rule auto-adjusts per row
    across the whole range."""
    for stock_col in stock_cols:
        pedido_col = stock_col + 2
        stock_letter = get_column_letter(stock_col)
        pedido_letter = get_column_letter(pedido_col)
        cell_range = f"{stock_letter}{data_start}:{stock_letter}{data_end}"

        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f"{stock_letter}{data_start}<{pedido_letter}{data_start}"],
                fill=_RED_FILL,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f"{stock_letter}{data_start}>{pedido_letter}{data_start}"],
                fill=_GREEN_FILL,
            ),
        )


def build_workbook(wide_df: pd.DataFrame, dias_venta: int, dias_stock: int = 15) -> Workbook:
    """Build the STOCK sheet: params, per-generico band, header, one row per
    article (Stock/VENTA values plus live PEDIDO/ALCANCE/Total-block
    formulas), the TOTAL GENERAL row, number formatting, and Stock-vs-Pedido
    conditional formatting.

    Args:
        wide_df: output of processor.pivot_wide() (MultiIndex columns).
        dias_venta: business days elapsed this month
            (processor.compute_dias_venta()) — written as a plain value into
            the DiasVenta param cell.
        dias_stock: target days of stock coverage — the interactive knob,
            written as a plain editable value into the DiasStock param cell
            (default 15).

    Returns:
        An in-memory openpyxl Workbook. Caller is responsible for saving.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    sucursales = _sucursal_block_keys(wide_df.columns)
    genericos = _generico_labels(wide_df)
    layout = compute_layout(n_genericos=len(genericos), n_articulos=len(wide_df))

    # ── Param cells + workbook-level named ranges ──
    ws.cell(row=DIAS_STOCK_ROW, column=PARAM_LABEL_COL, value="DiasStock:")
    ws.cell(row=DIAS_STOCK_ROW, column=PARAM_VALUE_COL, value=dias_stock)
    ws.cell(row=DIAS_VENTA_ROW, column=PARAM_LABEL_COL, value="DiasVenta:")
    ws.cell(row=DIAS_VENTA_ROW, column=PARAM_VALUE_COL, value=dias_venta)

    # DiasStock is the interactive knob — highlight it so it reads as editable.
    ws.cell(row=DIAS_STOCK_ROW, column=PARAM_LABEL_COL).font = _PARAM_LABEL_FONT
    ws.cell(row=DIAS_VENTA_ROW, column=PARAM_LABEL_COL).font = _PARAM_LABEL_FONT
    for param_row in (DIAS_STOCK_ROW, DIAS_VENTA_ROW):
        value_cell = ws.cell(row=param_row, column=PARAM_VALUE_COL)
        value_cell.font = _PARAM_LABEL_FONT
        value_cell.border = _CELL_BORDER
    ws.cell(row=DIAS_STOCK_ROW, column=PARAM_VALUE_COL).fill = _PARAM_VALUE_FILL

    _param_col_letter = get_column_letter(PARAM_VALUE_COL)
    wb.defined_names["DiasStock"] = DefinedName(
        "DiasStock", attr_text=f"'{SHEET_NAME}'!${_param_col_letter}${DIAS_STOCK_ROW}"
    )
    wb.defined_names["DiasVenta"] = DefinedName(
        "DiasVenta", attr_text=f"'{SHEET_NAME}'!${_param_col_letter}${DIAS_VENTA_ROW}"
    )

    # ── Column map + header rows ──
    block_col_of: dict[str, int] = {}  # sucursal -> first (Stock) column of its block
    col = _FIRST_BLOCK_COL
    for sucursal in sucursales:
        block_col_of[sucursal] = col
        col += _BLOCK_WIDTH
    total_col = col  # first column of the Total block

    # The same header is written TWICE: once above the per-generico totals
    # band and once above the article table, so each reads as a
    # self-contained table instead of bare numbers.
    _write_header_labels(ws, layout.band_header_row, sucursales, block_col_of, total_col)
    _write_header_labels(ws, layout.header_row, sucursales, block_col_of, total_col)

    total_stock_col = total_col
    total_venta_col = total_col + 1
    total_pedido_col = total_col + 2
    total_alcance_col = total_col + 3
    total_stock_letter = get_column_letter(total_stock_col)
    total_venta_letter = get_column_letter(total_venta_col)

    numeric_cols = range(_FIRST_BLOCK_COL, total_alcance_col + 1)

    # blocks used by the band row and the TOTAL GENERAL row: every sucursal
    # block plus the Total block, as (stock_col, venta_col, pedido_col,
    # alcance_col) tuples — same column letters the data rows use below.
    all_blocks = [
        (block_col_of[s], block_col_of[s] + 1, block_col_of[s] + 2, block_col_of[s] + 3)
        for s in sucursales
    ]
    all_blocks.append((total_stock_col, total_venta_col, total_pedido_col, total_alcance_col))

    # ── Data rows ──
    for offset, (_, row) in enumerate(wide_df.iterrows()):
        r = layout.data_start_row + offset

        ws.cell(row=r, column=1, value=row[("", "idArticulo")])
        ws.cell(row=r, column=2, value=row[("", "dsArticulo")])
        ws.cell(row=r, column=3, value=row[("", "GENERICO")])
        ws.cell(row=r, column=4, value=row[("", "MARCA")])

        stock_refs: list[str] = []
        venta_refs: list[str] = []
        pedido_refs: list[str] = []
        for sucursal in sucursales:
            block_col = block_col_of[sucursal]
            stock_col = block_col
            venta_col = block_col + 1
            pedido_col = block_col + 2
            alcance_col = block_col + 3

            stock_letter = get_column_letter(stock_col)
            venta_letter = get_column_letter(venta_col)
            stock_ref = f"{stock_letter}{r}"
            venta_ref = f"{venta_letter}{r}"

            # Stock/VENTA: plain DB values, never formulas, never rounded.
            ws.cell(row=r, column=stock_col, value=row[(sucursal, "Stock")])
            ws.cell(row=r, column=venta_col, value=row[(sucursal, "VENTA")])
            ws.cell(
                row=r, column=pedido_col,
                value=f"=MAX(({venta_ref}/DiasVenta)*DiasStock-{stock_ref},0)",
            )
            ws.cell(
                row=r, column=alcance_col,
                value=f"=IFERROR({stock_ref}/({venta_ref}/DiasVenta),0)",
            )

            stock_refs.append(stock_ref)
            venta_refs.append(venta_ref)
            pedido_refs.append(f"{get_column_letter(pedido_col)}{r}")

        total_stock_ref = f"{total_stock_letter}{r}"
        total_venta_ref = f"{total_venta_letter}{r}"

        ws.cell(row=r, column=total_stock_col, value=f"=SUM({','.join(stock_refs)})")
        ws.cell(row=r, column=total_venta_col, value=f"=SUM({','.join(venta_refs)})")
        ws.cell(row=r, column=total_pedido_col, value=f"=SUM({','.join(pedido_refs)})")
        # Ratio-of-sums correction (RF-07): NOT a SUM of the 14 per-sucursal
        # ALCANCE cells — sum of ratios != ratio of sums.
        ws.cell(
            row=r, column=total_alcance_col,
            value=f"=IFERROR({total_stock_ref}/({total_venta_ref}/DiasVenta),0)",
        )

        # RF-10: accounting number format on every numeric cell of this row.
        _apply_number_format_row(ws, r, numeric_cols)
        _apply_borders_row(ws, r, total_alcance_col)

    # ── Per-generico totals band (RF-08) ──
    for idx, generico in enumerate(genericos):
        band_row = layout.band_start_row + idx
        _write_generico_band_row(
            ws, band_row, generico, all_blocks, layout.data_start_row, layout.data_end_row,
        )
        _apply_number_format_row(ws, band_row, numeric_cols)
        _apply_borders_row(ws, band_row, total_alcance_col)
        for col in range(PARAM_LABEL_COL, total_alcance_col + 1):
            ws.cell(row=band_row, column=col).fill = _BAND_FILL
        ws.cell(row=band_row, column=_GENERICO_COL).font = _BAND_LABEL_FONT

    # ── TOTAL GENERAL row (RF-09) + conditional formatting (RF-10) ──
    # Both need at least one data row: an empty SUM(...)/conditional range
    # over a header-only sheet is meaningless and would emit a malformed
    # reversed range.
    if layout.has_data:
        _write_total_general_row(
            ws, layout.total_general_row, all_blocks, layout.data_start_row, layout.data_end_row,
        )
        _apply_number_format_row(ws, layout.total_general_row, numeric_cols)
        _apply_borders_row(ws, layout.total_general_row, total_alcance_col)

        stock_cols = [block_col_of[s] for s in sucursales] + [total_stock_col]
        _apply_stock_vs_pedido_conditional_formatting(
            ws, stock_cols, layout.data_start_row, layout.data_end_row,
        )

    # ── Visual styling: colored headers, borders, widths, groups, panes ──
    for hdr_row in (layout.band_header_row, layout.header_row):
        _style_header_row(ws, hdr_row, sucursales, block_col_of, total_col)
        _apply_borders_row(ws, hdr_row, total_alcance_col)
    _set_column_widths(ws, sucursales, block_col_of, total_alcance_col)
    _group_block_detail_columns(ws, sucursales, block_col_of)
    _group_top_rows(ws, layout.header_row)

    # Freeze the identity columns and everything above the first data row,
    # so scrolling a 64-column x N-row sheet keeps both the article name and
    # the sucursal headers on screen.
    ws.freeze_panes = f"{get_column_letter(_FIRST_BLOCK_COL)}{layout.data_start_row}"
    ws.sheet_view.showGridLines = False

    return wb
