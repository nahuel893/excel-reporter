"""Processor for stock-diario: pivots raw data and builds Excel with openpyxl."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import DATA_OUTPUT

logger = logging.getLogger(__name__)

# Styles
_BULTOS_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HTLS_FILL = PatternFill(fill_type="solid", fgColor="70AD47")
_BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_HEADER_FONT = Font(bold=True, size=8)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_NUMBER_FORMAT = "#,##0"
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

DESC_COLS = ["Codigo", "Articulo", "Marca", "Generico"]
DESC_WIDTH = 22
SUC_WIDTH = 9.5


def pivot_stock(df: pd.DataFrame, val_col: str, sucursales: list[str]) -> pd.DataFrame:
    """Pivot raw stock data to articles × sucursales matrix."""
    if df.empty:
        return pd.DataFrame()
    pivot = df.pivot_table(
        index=["generico", "marca", "des_articulo"],
        columns="sucursal",
        values=val_col,
        aggfunc="sum",
        fill_value=0,
    )
    # Reindex to ensure all sucursales present, fill missing with 0
    pivot = pivot.reindex(columns=sucursales, fill_value=0)
    return pivot.reset_index()


def build_excel(
    fecha_str: str,
    df: pd.DataFrame,
    output_dir: Path | None = None,
    nombre_prefijo: str = "Stock",
) -> Path:
    """Build the Stock Excel file for one date. Returns the file path."""
    output = output_dir or DATA_OUTPUT
    output.mkdir(parents=True, exist_ok=True)

    sucursales = sorted(df["sucursal"].unique().tolist())
    n_suc = len(sucursales)

    n_desc = len(DESC_COLS)  # 3
    bultos_start = n_desc + 1          # col 4
    bultos_end = n_desc + n_suc        # col 3+n
    htls_start = bultos_end + 1        # col 4+n
    htls_end = bultos_end + n_suc      # col 3+2n

    # Build article list (sorted) and lookup dicts.
    # dropna=False: keep articles whose generico/marca/des_articulo is NULL —
    # pandas' default (dropna=True) would silently drop them and their stock.
    articles = (
        df.groupby(["id_articulo", "generico", "marca", "des_articulo"], dropna=False)
        .size()
        .reset_index()[["id_articulo", "generico", "marca", "des_articulo"]]
        .sort_values(["des_articulo", "marca", "generico"])
        .reset_index(drop=True)
    )

    bultos_lookup: dict = {}
    htls_lookup: dict = {}
    for _, r in df.iterrows():
        key = (r["id_articulo"], r["sucursal"])
        # Keep the raw value — never int()-truncate. Fractional htls (and any
        # future fractional bultos) must survive; display is handled by the
        # Excel number_format only.
        bultos_lookup[key] = float(r["cant_bultos"]) if pd.notna(r["cant_bultos"]) else 0
        htls_lookup[key] = float(r["cant_htls"]) if pd.notna(r["cant_htls"]) else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    # ── Row 1: Banner ──
    if n_suc > 0:
        ws.merge_cells(start_row=1, start_column=bultos_start, end_row=1, end_column=bultos_end)
        cell_b = ws.cell(row=1, column=bultos_start, value="BULTOS")
        cell_b.font = _BANNER_FONT
        cell_b.fill = _BULTOS_FILL
        cell_b.alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=1, start_column=htls_start, end_row=1, end_column=htls_end)
        cell_h = ws.cell(row=1, column=htls_start, value="HTLs")
        cell_h.font = _BANNER_FONT
        cell_h.fill = _HTLS_FILL
        cell_h.alignment = Alignment(horizontal="center")

    # ── Row 2: Headers ──
    for i, label in enumerate(DESC_COLS, 1):
        cell = ws.cell(row=2, column=i, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _HEADER_ALIGNMENT

    for i, suc in enumerate(sucursales):
        # Bultos header
        cell = ws.cell(row=2, column=bultos_start + i, value=suc)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _HEADER_ALIGNMENT
        # HTLs header
        cell = ws.cell(row=2, column=htls_start + i, value=suc)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _HEADER_ALIGNMENT

    # ── Row 3+: Data ──
    for r_idx, (_, art) in enumerate(articles.iterrows(), 3):
        ws.cell(row=r_idx, column=1, value=art["id_articulo"])
        # Coalesce NULL text fields to "" so openpyxl never writes a bare NaN.
        ws.cell(row=r_idx, column=2, value="" if pd.isna(art["des_articulo"]) else art["des_articulo"])
        ws.cell(row=r_idx, column=3, value="" if pd.isna(art["marca"]) else art["marca"])
        ws.cell(row=r_idx, column=4, value="" if pd.isna(art["generico"]) else art["generico"])
        for s_idx, suc in enumerate(sucursales):
            key = (art["id_articulo"], suc)
            val_b = bultos_lookup.get(key, 0)
            cell = ws.cell(row=r_idx, column=bultos_start + s_idx, value=val_b)
            cell.number_format = _NUMBER_FORMAT
            val_h = htls_lookup.get(key, 0)
            cell = ws.cell(row=r_idx, column=htls_start + s_idx, value=val_h)
            cell.number_format = _NUMBER_FORMAT

    # ── Column widths ──
    for i in range(1, n_desc + 1):
        ws.column_dimensions[get_column_letter(i)].width = DESC_WIDTH
    for i in range(bultos_start, htls_end + 1):
        ws.column_dimensions[get_column_letter(i)].width = SUC_WIDTH

    # ── Freeze panes ──
    ws.freeze_panes = "E3"

    fecha_fmt = pd.to_datetime(fecha_str).strftime("%d-%m-%Y")
    nombre = f"{nombre_prefijo} - {fecha_fmt}.xlsx"
    ruta = output / nombre
    wb.save(ruta)
    return ruta
