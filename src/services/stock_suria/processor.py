"""Processor for stock-suria: builds 3-sheet Excel from matched article list and SURIA stock data."""

import logging
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Shared styles ──────────────────────────────────────────────────────────────
_BULTOS_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HTLS_FILL = PatternFill(fill_type="solid", fgColor="70AD47")
_BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)
_BANNER_ALIGNMENT = Alignment(horizontal="center")

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_HEADER_FONT = Font(bold=True, size=8)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_SIDE = Side(style="thin")
_THIN_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)

SUCURSALES = ["ABRA PAMPA", "HUMAHUACA", "JUJUY", "LA QUIACA", "MAIMARA", "PERICO"]

# Number formats
_FMT_BULTOS = "#,##0"
_FMT_HTLS = "#,##0.0"
_FMT_SIM = "0%"


def _apply_header_style(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = _HEADER_ALIGNMENT


def _write_banner(ws, start_col: int, end_col: int, row: int, value: str, fill: PatternFill) -> None:
    """Merge cells and write a banner (BULTOS / HTLs header)."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.font = _BANNER_FONT
    cell.fill = fill
    cell.alignment = _BANNER_ALIGNMENT


def _build_resumen_sheet(ws, config_data: dict) -> None:
    """Write the 'RESUMEN DEL MATCH' sheet (10 rows, no merges)."""
    resumen = config_data["resumen"]
    por_esquema = resumen["por_esquema"]
    matched = resumen["matched"]
    sin_match = resumen["sin_match"]
    total_activos = resumen["total_activos"]
    today_str = date.today().isoformat()

    rows = [
        ("RESUMEN DEL MATCH — Stock SURIA", None),
        (
            f"Generado: {today_str}  |  Archivo proveedor: articulos_coca.xlsx  |  BD: medallion_db_suria",
            None,
        ),
        (None, None),
        ("Tipo de match", "N articulos"),
        (f"Match por codigo '40'", por_esquema.get("40", 0)),
        (f"Match por codigo '400'", por_esquema.get("400", 0)),
        ("Match por codigo pelado", por_esquema.get("pelado", 0)),
        ("TOTAL MATCHEADOS", matched),
        ("Sin match (no existe en SURIA)", sin_match),
        ("TOTAL ACTIVOS PROVEEDOR", total_activos),
    ]

    for r_idx, (col_a, col_b) in enumerate(rows, 1):
        ws.cell(row=r_idx, column=1, value=col_a)
        ws.cell(row=r_idx, column=2, value=col_b)


def _build_stock_sheet(ws, config_data: dict, stock_data: dict, generico_map: dict) -> None:
    """Write the 'Stock SURIA' sheet (matched articles)."""
    articulos = config_data["articulos"]
    n_desc = 6
    n_suc = len(SUCURSALES)
    bultos_start = n_desc + 1       # col 7
    bultos_end = n_desc + n_suc     # col 12
    htls_start = bultos_end + 1     # col 13
    htls_end = bultos_end + n_suc   # col 18

    # Row 1: BULTOS / HTLs banners
    _write_banner(ws, bultos_start, bultos_end, 1, "BULTOS", _BULTOS_FILL)
    _write_banner(ws, htls_start, htls_end, 1, "HTLs", _HTLS_FILL)

    # Row 2: headers
    desc_headers = ["Cod Prov", "Desc Proveedor", "Cod SURIA", "Desc SURIA", "Marca", "Generico"]
    for col_idx, label in enumerate(desc_headers, 1):
        _apply_header_style(ws.cell(row=2, column=col_idx, value=label))

    for i, suc in enumerate(SUCURSALES):
        _apply_header_style(ws.cell(row=2, column=bultos_start + i, value=suc))
        _apply_header_style(ws.cell(row=2, column=htls_start + i, value=suc))

    # Row 3+: data — sorted by marca then desc_suria
    sorted_articles = sorted(
        articulos,
        key=lambda a: (a.get("marca") or "", a.get("desc_suria") or ""),
    )

    for r_idx, art in enumerate(sorted_articles, 3):
        id_art = art["id_articulo"]
        generico = generico_map.get(id_art)

        ws.cell(row=r_idx, column=1, value=art["cod_prov"])
        ws.cell(row=r_idx, column=2, value=art["desc_prov"])
        ws.cell(row=r_idx, column=3, value=id_art)
        ws.cell(row=r_idx, column=4, value=art["desc_suria"])
        ws.cell(row=r_idx, column=5, value=art["marca"])
        ws.cell(row=r_idx, column=6, value=generico)

        suc_data = stock_data.get(id_art, {})
        for s_idx, suc in enumerate(SUCURSALES):
            entry = suc_data.get(suc, {})
            bultos_val = entry.get("bultos", 0) if entry else 0
            htls_raw = entry.get("htls", 0) if entry else 0

            cell_b = ws.cell(row=r_idx, column=bultos_start + s_idx, value=bultos_val)
            cell_b.number_format = _FMT_BULTOS

            # None for zero HTLs (matches reference file behavior)
            htls_val = htls_raw if htls_raw else None
            cell_h = ws.cell(row=r_idx, column=htls_start + s_idx, value=htls_val)
            if htls_val is not None:
                cell_h.number_format = _FMT_HTLS

    # Freeze panes
    ws.freeze_panes = "G3"

    # Column widths
    col_widths = {1: 8, 2: 34, 3: 9, 4: 30, 5: 16, 6: 16}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(bultos_start, htls_end + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10


def _build_sin_match_sheet(ws, config_data: dict, stock_data: dict) -> None:
    """Write the 'ARTICULOS SIN MATCH POR CODIGO' sheet."""
    sin_match_list = config_data["sin_match"]
    n_desc = 5
    n_suc = len(SUCURSALES)
    bultos_start = n_desc + 1       # col 6
    bultos_end = n_desc + n_suc     # col 11
    htls_start = bultos_end + 1     # col 12
    htls_end = bultos_end + n_suc   # col 17

    # Row 1: banners
    # Descriptor section banner (A1:E1) — no fill (just text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_desc)
    cell_desc = ws.cell(row=1, column=1, value="ARTICULO MAS PARECIDO EN SURIA (otra presentacion)")
    cell_desc.font = Font(bold=True, size=12)
    cell_desc.alignment = _BANNER_ALIGNMENT

    _write_banner(ws, bultos_start, bultos_end, 1, "BULTOS (del parecido)", _BULTOS_FILL)
    _write_banner(ws, htls_start, htls_end, 1, "HTLs (del parecido)", _HTLS_FILL)

    # Row 2: headers
    desc_headers = [
        "Cod Prov",
        "Desc Proveedor",
        "ID SURIA parecido",
        "Desc SURIA parecido",
        "Sim %",
    ]
    for col_idx, label in enumerate(desc_headers, 1):
        _apply_header_style(ws.cell(row=2, column=col_idx, value=label))

    for i, suc in enumerate(SUCURSALES):
        _apply_header_style(ws.cell(row=2, column=bultos_start + i, value=suc))
        _apply_header_style(ws.cell(row=2, column=htls_start + i, value=suc))

    # Row 3+: data
    for r_idx, entry in enumerate(sin_match_list, 3):
        closest_id = entry["closest_id"]

        ws.cell(row=r_idx, column=1, value=entry["cod_prov"])
        ws.cell(row=r_idx, column=2, value=entry["desc_prov"])
        ws.cell(row=r_idx, column=3, value=closest_id)
        ws.cell(row=r_idx, column=4, value=entry["closest_desc"])

        sim_cell = ws.cell(row=r_idx, column=5, value=entry["closest_sim"])
        sim_cell.number_format = _FMT_SIM

        suc_data = stock_data.get(closest_id, {})
        for s_idx, suc in enumerate(SUCURSALES):
            suc_entry = suc_data.get(suc, {})
            bultos_val = suc_entry.get("bultos", 0) if suc_entry else 0
            htls_raw = suc_entry.get("htls", 0) if suc_entry else 0

            cell_b = ws.cell(row=r_idx, column=bultos_start + s_idx, value=bultos_val)
            cell_b.number_format = _FMT_BULTOS

            htls_val = htls_raw if htls_raw else None
            cell_h = ws.cell(row=r_idx, column=htls_start + s_idx, value=htls_val)
            if htls_val is not None:
                cell_h.number_format = _FMT_HTLS

    # Freeze panes
    ws.freeze_panes = "F3"

    # Column widths
    col_widths = {1: 8, 2: 34, 3: 14, 4: 30, 5: 8}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(bultos_start, htls_end + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10


def build_excel_todos(articulos: list[dict], fecha_str: str, output_dir: Path) -> Path:
    """Build the all-articles Stock SURIA Excel (single sheet, no provider match).

    Same BULTOS/HTLs × sucursal layout as the matched 'Stock SURIA' sheet, but the
    descriptor columns come straight from SURIA dim_articulo (no cod_prov/desc_prov)
    and every article with a stock record is included (unfiltered by the JSON list).

    Args:
        articulos: list of {id_articulo, des_suria, marca, generico,
                   suc: {sucursal_short: {"bultos": float, "htls": float}}}.
        fecha_str: date label (YYYY-MM-DD) used in the filename.
        output_dir: directory where the file is written.

    Returns:
        Path to the generated Excel file.
    """
    import pandas as pd

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Stock SURIA (todos)")

    n_desc = 4
    n_suc = len(SUCURSALES)
    bultos_start = n_desc + 1        # col 5
    bultos_end = n_desc + n_suc      # col 10
    htls_start = bultos_end + 1      # col 11
    htls_end = bultos_end + n_suc    # col 16

    _write_banner(ws, bultos_start, bultos_end, 1, "BULTOS", _BULTOS_FILL)
    _write_banner(ws, htls_start, htls_end, 1, "HTLs", _HTLS_FILL)

    for col_idx, label in enumerate(["Cod SURIA", "Desc SURIA", "Marca", "Generico"], 1):
        _apply_header_style(ws.cell(row=2, column=col_idx, value=label))
    for i, suc in enumerate(SUCURSALES):
        _apply_header_style(ws.cell(row=2, column=bultos_start + i, value=suc))
        _apply_header_style(ws.cell(row=2, column=htls_start + i, value=suc))

    sorted_articles = sorted(
        articulos, key=lambda a: (a.get("marca") or "", a.get("des_suria") or "")
    )
    for r_idx, art in enumerate(sorted_articles, 3):
        ws.cell(row=r_idx, column=1, value=art["id_articulo"])
        ws.cell(row=r_idx, column=2, value=art.get("des_suria"))
        ws.cell(row=r_idx, column=3, value=art.get("marca"))
        ws.cell(row=r_idx, column=4, value=art.get("generico"))

        suc_data = art.get("suc", {})
        for s_idx, suc in enumerate(SUCURSALES):
            entry = suc_data.get(suc, {})
            bultos_val = entry.get("bultos", 0) if entry else 0
            htls_raw = entry.get("htls", 0) if entry else 0

            cell_b = ws.cell(row=r_idx, column=bultos_start + s_idx, value=bultos_val)
            cell_b.number_format = _FMT_BULTOS
            htls_val = htls_raw if htls_raw else None
            cell_h = ws.cell(row=r_idx, column=htls_start + s_idx, value=htls_val)
            if htls_val is not None:
                cell_h.number_format = _FMT_HTLS

    ws.freeze_panes = "E3"
    for col_idx, width in {1: 9, 2: 34, 3: 16, 4: 16}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(bultos_start, htls_end + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    fecha_fmt = pd.to_datetime(fecha_str).strftime("%d-%m-%Y")
    output_path = output_dir / f"Stock SURIA completo - {fecha_fmt}.xlsx"
    wb.save(output_path)
    logger.info("Stock SURIA completo guardado: %s (%d articulos)", output_path, len(sorted_articles))
    return output_path


def build_excel(
    config_data: dict,
    stock_data: dict,
    generico_map: dict,
    fecha_str: str,
    output_dir: Path,
) -> Path:
    """Build the Stock SURIA Excel with 3 sheets and return the output path.

    Args:
        config_data: Full parsed JSON from stock_suria_articulos.json.
        stock_data: {id_articulo: {sucursal_name: {"bultos": float, "htls": float}}}
        generico_map: {id_articulo: des_generico}
        fecha_str: YYYY-MM-DD used for the filename date label.
        output_dir: Directory where the file will be written.

    Returns:
        Path to the generated Excel file.
    """
    import pandas as pd

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet and add named sheets in order
    wb.remove(wb.active)

    ws_resumen = wb.create_sheet("RESUMEN DEL MATCH")
    ws_stock = wb.create_sheet("Stock SURIA")
    ws_sin_match = wb.create_sheet("ARTICULOS SIN MATCH POR CODIGO")

    _build_resumen_sheet(ws_resumen, config_data)
    _build_stock_sheet(ws_stock, config_data, stock_data, generico_map)
    _build_sin_match_sheet(ws_sin_match, config_data, stock_data)

    fecha_fmt = pd.to_datetime(fecha_str).strftime("%d-%m-%Y")
    filename = f"Stock SURIA - {fecha_fmt}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    logger.info("Stock SURIA guardado: %s", output_path)
    return output_path
