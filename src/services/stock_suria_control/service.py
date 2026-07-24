"""
StockSuriaControlService — refreshes the per-deposito stock columns of a
user-maintained "Control de stock" SURIA base workbook.

The base file (data/input/...) is maintained by the user: they paste/edit the
article rows and the descriptive columns (Generico, Marca, Cod SURIA, Desc,
ACCION SEMANAL, OBSERVACIONES). This service ONLY overwrites the stock columns
(one per deposito), matching rows by "Cod SURIA" (== id_articulo in SURIA). All
other cells, sheets and formatting are preserved (key-matched in-place update).

SURIA stock lives in a separate database (DB_NAME_SURIA / medallion_db_suria),
so this service builds its own engine — the default DataLoader points at the
BADIE Salta gold DB, which has DIFFERENT depositos.
"""
import logging
import os
import shutil
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from src.services.base_service import BaseService

logger = logging.getLogger(__name__)

# --- Base workbook layout (Control de stock SURIA) ---
SHEET_NAME = "Stock SURIA"
HEADER_ROW = 2                 # headers live on row 2; data starts row 3
COD_HEADER = "Cod SURIA"       # article key column header (== id_articulo)

# Base-file stock column header -> SURIA id_deposito (confirmed vs gold.dim_deposito).
# Each Jujuy-region sucursal has exactly one deposito (1:1), so no fan-out.
DEPOSITO_ID = {
    "JUJUY": 9,
    "MAIMARA": 11,
    "LA QUIACA": 14,
    "HUMAHUACA": 12,
    "PERICO": 5,
}


@dataclass
class StockSuriaControlConfig:
    """Config for the Control-de-stock SURIA in-place refresh."""
    archivo_plantilla: str              # path to the user's base xlsx
    nombre_archivo: str | None = None   # output filename (copy mode); default = base stem
    fecha: str | None = None            # run date (YYYY-MM-DD) for the output dir; stock uses latest snapshot
    in_place: bool = False              # True: overwrite the base file; False: write a copy to data/output


@dataclass
class StockSuriaControlResult:
    ruta_archivo: Path
    fecha_stock: str | None
    filas_actualizadas: int
    articulos_sin_stock: list[int] = field(default_factory=list)


def _num(value):
    """Coerce a DB numeric to a native int/float exactly (never rounds).

    Whole numbers become int, fractional values stay float. None -> 0
    (a deposito absent from the snapshot means zero stock).
    """
    if value is None:
        return 0
    f = float(value)
    return int(f) if f == int(f) else f


class StockSuriaControlService(BaseService):
    """Fills the SURIA stock columns of a base workbook, matched by Cod SURIA."""

    SERVICE_SLUG = "stock-suria-control"
    GRANULARITY = "day"

    def _build_suria_engine(self):
        """Create a SQLAlchemy engine for the SURIA database (medallion_db_suria)."""
        host = os.environ.get("DB_HOST", "localhost")
        port = os.environ.get("DB_PORT", "5432")
        user = os.environ.get("DB_USER", "")
        password = os.environ.get("DB_PASSWORD", "")
        db_name = os.environ.get("DB_NAME_SURIA", "medallion_db_suria")
        url = f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{db_name}"
        return create_engine(url)

    def _fetch_stock(self, cod_list: list[int]) -> tuple[str | None, dict[int, dict[str, object]]]:
        """Return (fecha_stock, {cod_suria: {deposito_col: bultos}}) for the latest snapshot."""
        if not cod_list:
            return None, {}
        engine = self._build_suria_engine()
        # id_deposito / column names are internal constants -> safe to inline in VALUES.
        values = ", ".join(f"({dep_id}, '{col}')" for col, dep_id in DEPOSITO_ID.items())
        query = text(
            f"""
            WITH d(id_deposito, col) AS (VALUES {values})
            SELECT f.id_articulo        AS cod,
                   d.col                AS deposito,
                   SUM(f.cant_bultos)   AS bultos
            FROM   gold.fact_stock f
            JOIN   d ON d.id_deposito = f.id_deposito
            WHERE  f.date_stock = (SELECT MAX(date_stock) FROM gold.fact_stock)
              AND  f.id_articulo = ANY(:ids)
            GROUP BY f.id_articulo, d.col
            """
        )
        with engine.connect() as conn:
            row = conn.execute(text("SELECT MAX(date_stock) FROM gold.fact_stock")).fetchone()
            fecha_stock = row[0].isoformat() if row and row[0] else None
            rows = conn.execute(query, {"ids": cod_list}).fetchall()
        stock: dict[int, dict[str, object]] = {}
        for cod, deposito, bultos in rows:
            stock.setdefault(int(cod), {})[deposito] = bultos
        logger.info("SURIA stock snapshot %s: %d articulos con stock", fecha_stock, len(stock))
        return fecha_stock, stock

    def _resolve_dest(self, config: StockSuriaControlConfig, base: Path) -> Path:
        if config.in_place:
            return base
        fecha = config.fecha or date.today().isoformat()
        out_dir = self._output_dir(fecha)
        out_dir.mkdir(parents=True, exist_ok=True)
        nombre = config.nombre_archivo or base.stem
        dest = out_dir / f"{nombre}.xlsx"
        shutil.copy2(base, dest)
        return dest

    def generar_reporte(self, config: StockSuriaControlConfig) -> StockSuriaControlResult:
        base = Path(config.archivo_plantilla)
        if not base.exists():
            raise FileNotFoundError(f"Archivo base SURIA no encontrado: {base}")

        dest = self._resolve_dest(config, base)
        wb = load_workbook(dest)
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"El base no tiene la hoja '{SHEET_NAME}'. Hojas: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        # Map header text -> column index (from HEADER_ROW).
        header_idx: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=HEADER_ROW, column=c).value
            if v is not None:
                header_idx[str(v).strip()] = c

        if COD_HEADER not in header_idx:
            raise ValueError(f"No se encontro la columna '{COD_HEADER}' en la fila {HEADER_ROW}")
        cod_col = header_idx[COD_HEADER]

        dep_col: dict[str, int] = {}
        for col_name in DEPOSITO_ID:
            if col_name not in header_idx:
                raise ValueError(f"No se encontro la columna de deposito '{col_name}' en el base")
            dep_col[col_name] = header_idx[col_name]

        # Collect Cod SURIA per row.
        rows_by_cod: dict[int, list[int]] = {}
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            raw = ws.cell(row=r, column=cod_col).value
            if raw is None or str(raw).strip() == "":
                continue
            try:
                cod = int(raw)
            except (ValueError, TypeError):
                logger.warning("Fila %d: Cod SURIA no numerico (%r), se omite", r, raw)
                continue
            rows_by_cod.setdefault(cod, []).append(r)

        fecha_stock, stock = self._fetch_stock(list(rows_by_cod.keys()))

        filas_actualizadas = 0
        sin_stock: list[int] = []
        for cod, rows in rows_by_cod.items():
            data = stock.get(cod)
            if data is None:
                sin_stock.append(cod)
            for r in rows:
                for col_name, cidx in dep_col.items():
                    ws.cell(row=r, column=cidx).value = _num((data or {}).get(col_name))
                filas_actualizadas += 1

        wb.save(dest)
        logger.info(
            "Stock SURIA Control actualizado: %s (%d filas, %d articulos sin stock)",
            dest, filas_actualizadas, len(sin_stock),
        )
        return StockSuriaControlResult(
            ruta_archivo=dest,
            fecha_stock=fecha_stock,
            filas_actualizadas=filas_actualizadas,
            articulos_sin_stock=sorted(set(sin_stock)),
        )
