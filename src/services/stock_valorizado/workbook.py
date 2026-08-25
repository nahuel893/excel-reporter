"""Workbook construction for stock-valorizado.

Five sheets: the wide article grid, a per-sucursal summary, an ABC/Pareto
ranking, a generico x sucursal valuation matrix, and a Control sheet that
accounts for everything the main grid does not show.

Numbers are written at full float precision; every bit of presentation is
``number_format``. Nothing here rounds.
"""

from __future__ import annotations

from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.services.stock_valorizado.precios import EstadoListaPrecios
from src.services.stock_valorizado.processor import IDENTIDAD

FMT_BULTOS = "#,##0"
# No decimals by request. This is presentation only — the cell values stay at
# full float precision, nothing is rounded on the way in.
FMT_PLATA = '$ #,##0'
FMT_PCT = "0.0%"

# Fixed width for every money column, so the two valuation sheets line up.
ANCHO_PLATA = 14.5

_AZUL = "1F3864"
_AZUL_CLARO = "2E5B9A"
_GRIS = "D9D9D9"
_AMARILLO = "FFF2CC"
_BLANCO = "FFFFFF"

_HEADER_FONT = Font(bold=True, color=_BLANCO, size=10)
_TITULO_FONT = Font(bold=True, size=14, color=_AZUL)
_SUBTITULO_FONT = Font(italic=True, size=9, color="595959")
_TOTAL_FONT = Font(bold=True, size=10)
_SECCION_FONT = Font(bold=True, size=11, color=_AZUL)

_HEADER_FILL = PatternFill("solid", fgColor=_AZUL)
_HEADER_FILL_ALT = PatternFill("solid", fgColor=_AZUL_CLARO)
_TOTAL_FILL = PatternFill("solid", fgColor=_AMARILLO)

# Stale-price alert: red on white, impossible to skim past.
_ROJO = "C00000"
_ALERTA_FONT = Font(bold=True, size=11, color=_BLANCO)
_ALERTA_FILL = PatternFill("solid", fgColor=_ROJO)

_THIN = Side(style="thin", color=_GRIS)
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _titulo(
    ws: Worksheet, texto: str, subtitulo: str, ancho: int, *, alerta: bool = False
) -> None:
    """Title on row 1, subtitle on row 2.

    ``alerta`` repaints the subtitle as a red banner. The stale-price warning
    rides on row 2 rather than inserting a row of its own, so the sheet layout
    (mirror on 3, headers on 4-5, data from 6) never shifts.
    """
    ws.cell(row=1, column=1, value=texto).font = _TITULO_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ancho, 2))

    celda = ws.cell(row=2, column=1, value=subtitulo)
    celda.font = _ALERTA_FONT if alerta else _SUBTITULO_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ancho, 2))
    if alerta:
        celda.alignment = Alignment(vertical="center")
        ws.row_dimensions[2].height = 30
        for c in range(1, max(ancho, 2) + 1):
            ws.cell(row=2, column=c).fill = _ALERTA_FILL


def _formato_de(nombre: str) -> str | None:
    """Pick a number format from the column name — the sheets share vocabulary."""
    if nombre.startswith("%"):
        return FMT_PCT
    if "Valorizado" in nombre or "Valor" in nombre or nombre == "Total":
        return FMT_PLATA
    if "Bultos" in nombre or "Artículos" in nombre:
        return FMT_BULTOS
    return None


# Row 3 — between the subtitle and the header block. Always inside the frozen
# pane, so the mirrored totals never scroll away.
FILA_ESPEJO = 3


def _write_mirror_row(
    ws: Worksheet, total_row: int, n_ident: int, last_col: int, *, con_datos: bool
) -> None:
    """Echo the bottom TOTAL row at the top of the sheet.

    Each cell is a plain reference (``=E2128``) to its SUBTOTAL counterpart, so
    the mirror follows the autofilter for free and there is only one formula to
    get right.
    """
    etiqueta = ws.cell(row=FILA_ESPEJO, column=1, value="TOTAL VISIBLE")
    etiqueta.font = _TOTAL_FONT
    for c in range(1, last_col + 1):
        celda = ws.cell(row=FILA_ESPEJO, column=c)
        celda.fill = _TOTAL_FILL
        celda.font = _TOTAL_FONT
        celda.border = _BORDE
        if c > n_ident:
            columna = get_column_letter(c)
            celda.number_format = FMT_BULTOS if c % 2 == (n_ident + 1) % 2 else FMT_PLATA
            if con_datos:
                celda.value = f"={columna}{total_row}"


# ── main sheet ──────────────────────────────────────────────────────────


def _build_stock_sheet(
    ws: Worksheet,
    wide: pd.DataFrame,
    sucursales: list[str],
    fecha_stock: date,
    estado_precios: EstadoListaPrecios,
    *,
    titulo: str = "STOCK VALORIZADO",
    columna_precio: str = "Precio Base",
) -> None:
    n_ident = len(IDENTIDAD)
    last_col = n_ident + len(sucursales) * 2 + 2

    total_bultos = float(wide[("Total", "Total Bultos")].sum()) if len(wide) else 0.0
    total_plata = float(wide[("Total", "Total Valorizado")].sum()) if len(wide) else 0.0

    resumen = (
        f"TOTAL: {total_bultos:,.0f} bultos — $ {total_plata:,.0f}   |   "
        f"{columna_precio} por bulto   |   "
        f"{len(wide):,} artículos × {len(sucursales)} sucursales"
    )
    # When the list is stale the warning leads: the money below it is suspect,
    # so it should not be the first thing the eye lands on.
    subtitulo = (
        f"{estado_precios.leyenda}   ||   {resumen}"
        if estado_precios.vencida
        else f"{resumen}   |   {estado_precios.leyenda}"
    )

    _titulo(
        ws,
        f"{titulo} — {fecha_stock.strftime('%d-%m-%Y')}",
        subtitulo,
        last_col,
        alerta=estado_precios.vencida,
    )

    grupo_row, head_row = 4, 5

    for i, nombre in enumerate(IDENTIDAD, start=1):
        ws.merge_cells(start_row=grupo_row, start_column=i, end_row=head_row, end_column=i)
        celda = ws.cell(row=grupo_row, column=i, value=nombre)
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=head_row, column=i).fill = _HEADER_FILL

    col = n_ident + 1
    for idx, sucursal in enumerate(sucursales):
        fill = _HEADER_FILL if idx % 2 == 0 else _HEADER_FILL_ALT
        ws.merge_cells(start_row=grupo_row, start_column=col, end_row=grupo_row, end_column=col + 1)
        grupo = ws.cell(row=grupo_row, column=col, value=sucursal.replace("SUCURSAL ", ""))
        grupo.font = _HEADER_FONT
        grupo.fill = fill
        grupo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=grupo_row, column=col + 1).fill = fill
        for offset, etiqueta in enumerate(("Bultos", "$")):
            hoja = ws.cell(row=head_row, column=col + offset, value=etiqueta)
            hoja.font = _HEADER_FONT
            hoja.fill = fill
            hoja.alignment = Alignment(horizontal="center")
        col += 2

    ws.merge_cells(start_row=grupo_row, start_column=col, end_row=grupo_row, end_column=col + 1)
    total_head = ws.cell(row=grupo_row, column=col, value="TOTAL")
    total_head.font = _HEADER_FONT
    total_head.fill = _HEADER_FILL
    total_head.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=grupo_row, column=col + 1).fill = _HEADER_FILL
    for offset, etiqueta in enumerate(("Total Bultos", "Total Valorizado")):
        hoja = ws.cell(row=head_row, column=col + offset, value=etiqueta)
        hoja.font = _HEADER_FONT
        hoja.fill = _HEADER_FILL
        hoja.alignment = Alignment(horizontal="center", wrap_text=True)

    first_data = head_row + 1
    for r, (_, fila) in enumerate(wide.iterrows(), start=first_data):
        for c, clave in enumerate(wide.columns, start=1):
            celda = ws.cell(row=r, column=c, value=fila[clave])
            celda.border = _BORDE
            if c > n_ident:
                celda.number_format = FMT_BULTOS if c % 2 == (n_ident + 1) % 2 else FMT_PLATA

    last_row = first_data + len(wide) - 1 if len(wide) else head_row

    # TOTAL GENERAL sits below the data and outside the autofilter range, so a
    # filtered view never mixes it in with the article rows — and SUBTOTAL never
    # counts itself.
    #
    # SUBTOTAL(9, ...) rather than SUM: it recomputes over the rows the
    # autofilter leaves visible, which is the whole point of filtering a
    # 2000-row grid. Function 9 (not 109) tracks the filter only — a row hidden
    # by hand keeps counting, so nobody silently changes a total by dragging a
    # row border.
    total_row = last_row + 1
    etiqueta = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
    etiqueta.font = _TOTAL_FONT
    for c in range(1, last_col + 1):
        celda = ws.cell(row=total_row, column=c)
        celda.fill = _TOTAL_FILL
        celda.font = _TOTAL_FONT
        celda.border = _BORDE
        if c > n_ident:
            columna = get_column_letter(c)
            celda.number_format = FMT_BULTOS if c % 2 == (n_ident + 1) % 2 else FMT_PLATA
            if len(wide):
                celda.value = f"=SUBTOTAL(9,{columna}{first_data}:{columna}{last_row})"

    # Mirror of the totals on row 3, inside the frozen pane, so the filtered
    # figures stay on screen instead of living 2000 rows down. It references the
    # SUBTOTAL cells rather than repeating the formula — one source of truth.
    _write_mirror_row(ws, total_row, n_ident, last_col, con_datos=bool(len(wide)))

    if len(wide):
        ws.auto_filter.ref = f"A{head_row}:{get_column_letter(last_col)}{last_row}"
        barra = get_column_letter(last_col)
        ws.conditional_formatting.add(
            f"{barra}{first_data}:{barra}{last_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6"),
        )
    ws.freeze_panes = ws.cell(row=first_data, column=n_ident + 1)

    for i, ancho in enumerate((10, 38, 20, 20), start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for c in range(n_ident + 1, last_col + 1):
        es_bultos = c % 2 == (n_ident + 1) % 2
        ws.column_dimensions[get_column_letter(c)].width = 10 if es_bultos else ANCHO_PLATA

    # Collapsible per-sucursal blocks: the totals stay readable with the 28
    # detail columns folded away.
    #
    # Set outline_level per column rather than calling DimensionHolder.group():
    # group() DELETES the ColumnDimension of every column after the first and
    # stretches the first one across the range, which silently discards the
    # widths set just above (every money column inherited the Bultos width).
    # Excel groups contiguous columns sharing an outline level all the same.
    for idx in range(len(sucursales)):
        inicio = n_ident + 1 + idx * 2
        for c in (inicio, inicio + 1):
            ws.column_dimensions[get_column_letter(c)].outline_level = 1


# ── generic frame sheets ────────────────────────────────────────────────


def _write_frame(
    ws: Worksheet,
    df: pd.DataFrame,
    header_row: int,
    *,
    total_row: bool = False,
    formatos: dict[str, str] | None = None,
    total_overrides: dict[str, object] | None = None,
) -> int:
    """Write a flat frame with a styled header. Returns the last written row.

    ``total_overrides`` replaces the default column SUM for columns where a sum
    is wrong — a distinct count, or a ratio that must be recomputed from the
    total row rather than added up. A string value is treated as a formula and
    gets ``{first}``/``{last}``/``{total}`` interpolated; anything else is
    written verbatim.
    """
    formatos = formatos or {}

    for c, nombre in enumerate(df.columns, start=1):
        celda = ws.cell(row=header_row, column=c, value=str(nombre))
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = Alignment(horizontal="center", wrap_text=True)

    numericas = {
        c: formatos.get(str(nombre), _formato_de(str(nombre)))
        for c, nombre in enumerate(df.columns, start=1)
    }

    first = header_row + 1
    for r, (_, fila) in enumerate(df.iterrows(), start=first):
        for c, nombre in enumerate(df.columns, start=1):
            celda = ws.cell(row=r, column=c, value=fila[nombre])
            celda.border = _BORDE
            if numericas[c]:
                celda.number_format = numericas[c]

    last = first + len(df) - 1 if len(df) else header_row

    if total_row and len(df):
        fila_total = last + 1
        ws.cell(row=fila_total, column=1, value="TOTAL GENERAL")
        for c, nombre in enumerate(df.columns, start=1):
            celda = ws.cell(row=fila_total, column=c)
            celda.fill = _TOTAL_FILL
            celda.font = _TOTAL_FONT
            celda.border = _BORDE
            if c > 1 and pd.api.types.is_numeric_dtype(df[nombre]):
                columna = get_column_letter(c)
                override = (total_overrides or {}).get(str(nombre))
                if override is None:
                    celda.value = f"=SUM({columna}{first}:{columna}{last})"
                elif isinstance(override, str):
                    celda.value = override.format(first=first, last=last, total=fila_total)
                else:
                    celda.value = override
                celda.number_format = numericas[c] or FMT_BULTOS
        last = fila_total

    for c, nombre in enumerate(df.columns, start=1):
        if numericas[c] == FMT_PLATA:
            ws.column_dimensions[get_column_letter(c)].width = ANCHO_PLATA
            continue
        largo = max([len(str(nombre))] + [len(str(v)) for v in df[nombre].head(200)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(largo + 3, 12), 42)

    return last


def _build_resumen_sucursal(
    ws: Worksheet, resumen: pd.DataFrame, fecha_stock: date, articulos_con_stock: int
) -> None:
    _titulo(
        ws,
        "RESUMEN POR SUCURSAL",
        f"Capital inmovilizado al {fecha_stock.strftime('%d-%m-%Y')}",
        len(resumen.columns),
    )
    last = _write_frame(
        ws, resumen, header_row=4, total_row=True,
        total_overrides={
            # Summing the per-sucursal counts would multiply every article that
            # sits in more than one sucursal; the total is a distinct count.
            "Artículos con Stock": articulos_con_stock,
            # An average of averages is not an average — recompute from the
            # total row's own money and bultos.
            "Valor Promedio x Bulto": "=IFERROR(D{total}/B{total},0)",
        },
    )
    if len(resumen):
        ws.conditional_formatting.add(
            f"D5:D{last - 1}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6"),
        )
    ws.freeze_panes = "A5"


def _build_abc(ws: Worksheet, abc: pd.DataFrame, fecha_stock: date) -> None:
    if len(abc):
        conteo = abc["Clase"].value_counts()
        reparto = " | ".join(
            f"{clase}: {int(conteo.get(clase, 0)):,} artículos" for clase in ("A", "B", "C")
        )
    else:
        reparto = "sin artículos con valorización positiva"
    _titulo(
        ws,
        "ABC / PARETO — dónde está el capital",
        f"A ≤ 80% del dinero, B ≤ 95%, C el resto. {reparto}. "
        f"Excluye artículos sin valorización positiva (ver Control).",
        len(abc.columns),
    )
    last = _write_frame(ws, abc, header_row=4)
    if len(abc):
        ws.auto_filter.ref = f"A4:{get_column_letter(len(abc.columns))}{last}"
    ws.freeze_panes = "A5"


def _build_matriz(ws: Worksheet, matriz: pd.DataFrame, fecha_stock: date) -> None:
    _titulo(
        ws,
        "GENÉRICO × SUCURSAL — valorizado",
        f"Pesos inmovilizados por genérico y sucursal al {fecha_stock.strftime('%d-%m-%Y')}",
        len(matriz.columns),
    )
    # Every column but "Genérico" is money; the sucursal names carry no keyword
    # for _formato_de to latch onto, so state it explicitly.
    last = _write_frame(
        ws, matriz, header_row=4, total_row=True,
        formatos={str(c): FMT_PLATA for c in matriz.columns if str(c) != "Genérico"},
    )
    if len(matriz) and len(matriz.columns) > 1:
        rango = f"B5:{get_column_letter(len(matriz.columns) - 1)}{last - 1}"
        ws.conditional_formatting.add(
            rango,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            ),
        )
    ws.freeze_panes = "B5"


def _build_control(
    ws: Worksheet,
    secciones: dict[str, pd.DataFrame],
    fecha_stock: date,
    estado_precios: EstadoListaPrecios,
) -> None:
    _titulo(
        ws,
        "CONTROL",
        f"Todo lo que la hoja principal no muestra tal cual. "
        f"Snapshot de stock: {fecha_stock.strftime('%d-%m-%Y')}.   |   "
        f"{estado_precios.leyenda}",
        8,
        alerta=estado_precios.vencida,
    )

    # Price-list status first: it is the one thing that invalidates every other
    # number on every other sheet.
    veredicto = (
        f"VENCIDA — hay que re-exportarla del ERP (máximo {estado_precios.max_dias} días)"
        if estado_precios.vencida
        else f"vigente (máximo {estado_precios.max_dias} días)"
    )
    encabezado = ws.cell(row=4, column=1, value="Estado de la lista de precios")
    encabezado.font = _SECCION_FONT
    filas_estado = [
        ("Archivo", estado_precios.nombre),
        ("Última actualización", estado_precios.mtime.strftime("%d-%m-%Y %H:%M")),
        ("Antigüedad (días)", estado_precios.dias),
        ("Estado", veredicto),
        ("Se carga", "A MANO: exportar del ERP y reemplazar el archivo"),
    ]
    for offset, (clave, valor) in enumerate(filas_estado, start=5):
        etiqueta = ws.cell(row=offset, column=1, value=clave)
        etiqueta.font = _TOTAL_FONT
        celda = ws.cell(row=offset, column=2, value=valor)
        if clave == "Estado" and estado_precios.vencida:
            celda.font = Font(bold=True, color=_ROJO)

    fila = 5 + len(filas_estado) + 1
    for titulo, df in secciones.items():
        celda = ws.cell(row=fila, column=1, value=titulo)
        celda.font = _SECCION_FONT
        fila += 1
        if df.empty:
            ws.cell(row=fila, column=1, value="— sin casos —").font = _SUBTITULO_FONT
            fila += 3
            continue
        fila = _write_frame(ws, df, header_row=fila) + 3

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40


# ── entry point ─────────────────────────────────────────────────────────


def build_workbook(
    wide: pd.DataFrame,
    sucursales: list[str],
    resumen: pd.DataFrame,
    abc: pd.DataFrame,
    matriz: pd.DataFrame,
    control: dict[str, pd.DataFrame],
    *,
    fecha_stock: date,
    estado_precios: EstadoListaPrecios,
    wide_final: pd.DataFrame | None = None,
) -> Workbook:
    """Assemble the workbook.

    ``wide_final`` adds a second copy of the main grid valued at Precio Final
    (Precio Base * 1.21 + Imp. Internos). The analytics sheets stay on Precio
    Base — they answer "where is the capital", and mixing two price bases in
    one ABC ranking would make it unreadable.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Stock Valorizado"
    _build_stock_sheet(ws, wide, sucursales, fecha_stock, estado_precios)

    if wide_final is not None:
        _build_stock_sheet(
            wb.create_sheet("Stock Valorizado Final"),
            wide_final, sucursales, fecha_stock, estado_precios,
            titulo="STOCK VALORIZADO (PRECIO FINAL)",
            columna_precio="Precio Final (Base × 1,21 + internos)",
        )

    articulos_con_stock = (
        int((wide[("Total", "Total Bultos")] != 0).sum()) if len(wide) else 0
    )
    _build_resumen_sucursal(
        wb.create_sheet("Resumen Sucursal"), resumen, fecha_stock, articulos_con_stock
    )
    _build_abc(wb.create_sheet("ABC Pareto"), abc, fecha_stock)
    _build_matriz(wb.create_sheet("Generico x Sucursal"), matriz, fecha_stock)
    _build_control(wb.create_sheet("Control"), control, fecha_stock, estado_precios)

    return wb
