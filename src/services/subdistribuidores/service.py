"""SubdistribuidoresService - Reporte de ventas para subdistribuidores (ruta 93)."""

import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from src.core.data_loader import DataLoader
from src.services.base_service import BaseService, BaseReporteConfig, BaseReporteResult
from src.services.subdistribuidores.processor import procesar_subdistribuidores

logger = logging.getLogger(__name__)


@dataclass
class SubdistribuidoresConfig(BaseReporteConfig):
    """Configuracion para el reporte de subdistribuidores.

    Hereda fecha_desde, fecha_hasta, nombre_archivo de BaseReporteConfig.
    No requiere genericos ni otros filtros — ruta 93 es fija.
    """
    pass


@dataclass
class SubdistribuidoresResult(BaseReporteResult):
    """Resultado del reporte de subdistribuidores."""
    clientes: int = 0
    filas_bultos: int = 0
    hojas: list[str] = None

    def __post_init__(self):
        if self.hojas is None:
            self.hojas = []


class SubdistribuidoresService(BaseService):
    """Genera el reporte de ventas para subdistribuidores (ruta 93)."""

    SERVICE_SLUG = "subdistribuidores"
    GRANULARITY = "month"

    def _thin_border(self) -> Border:
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _build_bultos_sheet(self, wb: Workbook, df_bultos: pd.DataFrame) -> None:
        """Construye la hoja Bultos (detalle a nivel de articulo)."""
        ws = wb.create_sheet(title="Bultos")

        # Headers
        headers = ["Cliente", "Fantasia", "Razon Social", "Generico", "Marca", "Articulo", "Bultos"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 35
        ws.column_dimensions["G"].width = 12

        if df_bultos.empty:
            return

        # Data rows
        for row_idx, row in enumerate(df_bultos.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._thin_border()
                if col_idx == 7:
                    cell.number_format = "#,##0"

    def _build_totales_sheet(self, wb: Workbook, df_totales: pd.DataFrame) -> None:
        """Construye la hoja Totales (agregados por nivel jerarquico)."""
        ws = wb.create_sheet(title="Totales")

        # Headers
        headers = ["Nivel", "Cliente", "Fantasia", "Razon Social", "Generico", "Marca", "Bultos"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 12

        if df_totales.empty:
            return

        # Data rows
        for row_idx, row in enumerate(df_totales.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._thin_border()
                if col_idx == 7:
                    cell.number_format = "#,##0"
                # Bold total rows (marcadores de nivel)
                if col_idx == 1:
                    cell.font = Font(bold=True)

    def _build_excel(
        self, df_bultos: pd.DataFrame, df_totales: pd.DataFrame, nombre: str, output_dir: Path
    ) -> Path:
        """Construye el workbook y lo persiste a disco."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        self._build_bultos_sheet(wb, df_bultos)
        self._build_totales_sheet(wb, df_totales)

        output_dir.mkdir(parents=True, exist_ok=True)
        ruta = output_dir / f"{nombre}.xlsx"
        wb.save(ruta)
        return ruta

    def generar_reporte(self, config: SubdistribuidoresConfig) -> SubdistribuidoresResult:
        """
        Genera el reporte de subdistribuidores.

        Args:
            config: SubdistribuidoresConfig con fecha_desde, fecha_hasta, nombre_archivo

        Returns:
            SubdistribuidoresResult con ruta_archivo y metricas
        """
        # 1. Obtener datos crudos
        df_ventas = self.data_loader.get_ventas_subdistribuidores(
            config.fecha_desde, config.fecha_hasta
        )

        if df_ventas.empty:
            logger.warning(
                "No se encontraron ventas para subdistribuidores (ruta 93) "
                "en el periodo %s - %s",
                config.fecha_desde,
                config.fecha_hasta,
            )

        # 2. Procesar: generar Bultos y Totales
        df_bultos, df_totales = procesar_subdistribuidores(df_ventas)

        # 3. Nombre de archivo
        nombre = config.nombre_archivo or f"Subdistribuidores {config.fecha_hasta[:7]}"
        out_dir = self._output_dir(config.fecha_desde)

        # 4. Generar Excel
        ruta = self._build_excel(df_bultos, df_totales, nombre, out_dir)

        # 5. Contar metricas
        clientes = df_ventas["id_cliente"].nunique() if not df_ventas.empty else 0

        return SubdistribuidoresResult(
            ruta_archivo=ruta,
            registros_procesados=len(df_bultos),
            clientes=clientes,
            filas_bultos=len(df_bultos),
            hojas=["Bultos", "Totales"],
        )

    def run(self, config: SubdistribuidoresConfig) -> SubdistribuidoresResult:
        return self.generar_reporte(config)