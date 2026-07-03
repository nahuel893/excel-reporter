"""
VentasService - Servicio para generacion de reportes de ventas.

Orquesta el flujo completo: extraccion, procesamiento y generacion de Excel.
"""
from dataclasses import dataclass, replace
from pathlib import Path
import pandas as pd

from config.settings import COLUMN_NAMES
from src.core.base_processor import calcular_info_dias
from src.core.data_loader import DataLoader
from src.core.excel_writer import ExcelWriter, SheetStyle, ColumnFormat, ColumnGroup
from src.core.excel_slicers import agregar_slicers, slicers_disponibles
from src.core.zonas import aplicar_zonas_virtuales as _aplicar_zonas_virtuales, expandir_sucursales as _expandir_sucursales
from src.services.base_service import BaseService
from src.services.ventas.processor import completar_combinaciones, procesar_ventas, procesar_ventas_diarias

# Columnas para slicers en reporte de ventas
SLICER_COLUMNS = [
    COLUMN_NAMES["sucursal"],
    COLUMN_NAMES["generico"],
    COLUMN_NAMES["marca"],
]

# Configuracion base de formatos para ventas
VENTAS_COLUMN_FORMATS = {
    COLUMN_NAMES["cant_generico"]: ColumnFormat(number_format='#,##0', width=15, font_bold=True),
    COLUMN_NAMES["tend_generico"]: ColumnFormat(number_format='#,##0', width=15, font_bold=True),
    COLUMN_NAMES["monto_generico"]: ColumnFormat(number_format='$ #,##0', width=15, font_bold=True, header_fill_color="0055A4"),
    COLUMN_NAMES["desc_generico"]: ColumnFormat(number_format='$ #,##0', width=15, font_bold=True, header_fill_color="0055A4"),
    COLUMN_NAMES["desc_pct_generico"]: ColumnFormat(number_format='0.0%', width=11, font_bold=True),
    COLUMN_NAMES["cob_generico"]: ColumnFormat(number_format='#,##0', width=13, font_bold=True, header_fill_color="60497A"),
    COLUMN_NAMES["cupo_generico"]: ColumnFormat(number_format='#,##0', width=13, font_bold=True),
    COLUMN_NAMES["cupo_vs_tend_generico"]: ColumnFormat(number_format='0.0%', width=13, font_bold=True),
    COLUMN_NAMES["total_marca"]: ColumnFormat(number_format='#,##0', width=11, font_bold=True),
    COLUMN_NAMES["mmaa_marca"]: ColumnFormat(number_format='#,##0', width=11, font_bold=True),
    COLUMN_NAMES["var_mmaa_marca"]: ColumnFormat(number_format='0.0%', width=9, font_bold=True),
    COLUMN_NAMES["tend_marca"]: ColumnFormat(number_format='#,##0', width=11, font_bold=True),
    COLUMN_NAMES["monto_marca"]: ColumnFormat(number_format='$ #,##0', width=15, font_bold=True, header_fill_color="0055A4"),
    COLUMN_NAMES["desc_marca"]: ColumnFormat(number_format='$ #,##0', width=15, font_bold=True, header_fill_color="0055A4"),
    COLUMN_NAMES["desc_pct_marca"]: ColumnFormat(number_format='0.0%', width=11, font_bold=True),
    COLUMN_NAMES["cob_marca"]: ColumnFormat(number_format='#,##0', width=13, font_bold=True, header_fill_color="60497A"),
    COLUMN_NAMES["cupo_marca"]: ColumnFormat(number_format='#,##0', width=13, font_bold=True),
    COLUMN_NAMES["cupo_vs_tend_marca"]: ColumnFormat(number_format='0.0%', width=13, font_bold=True),
}

# Columnas relacionadas con monto/descuentos. Cuando con_montos=False, se omiten
# tanto del DataFrame final como del estilo (ni siquiera ocultas — directamente
# no aparecen en el archivo).
_MONEY_COLUMNS = (
    COLUMN_NAMES["monto_generico"],
    COLUMN_NAMES["desc_generico"],
    COLUMN_NAMES["desc_pct_generico"],
    COLUMN_NAMES["monto_marca"],
    COLUMN_NAMES["desc_marca"],
    COLUMN_NAMES["desc_pct_marca"],
)

# Color de fondo por generico (toda la fila — bloque entero del generico).
# Genericos sin entrada quedan sin color de fondo.
_GENERICO_FILL_COLORS: dict[str, str] = {
    "AGUAS DANONE": "DDEBF7",       # celeste claro
    "CERVEZAS": "FCE4D6",            # peach claro
    "SIDRAS Y LICORES": "E2EFDA",   # verde menta
    "VINOS CCU": "F4CCCC",           # rojo palido
}

# Genericos cuyo color de fondo es oscuro y necesitan font blanco.
_GENERICO_WHITE_FONT: set[str] = set()

# Zonas virtuales locales al servicio ventas: SUB DISTRIBUIDORES (ruta 93) se
# absorbe dentro de VALLE SALTA en vez de mostrarse como zona aparte. Otros
# servicios (ej. resumen-mensual) siguen usando el mapeo global de settings.
_ZONAS_VIRTUALES_VENTAS: dict[str, dict] = {
    "VALLE SALTA": {
        "sucursal_real": "CASA CENTRAL",
        "rutas": [81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 118, 119, 120, 122],
    },
}


def _fechas_a_periodos(fecha_desde: str, fecha_hasta: str) -> list[str]:
    """Convierte un rango de fechas en lista de primeros dias de mes cubiertos."""
    desde = pd.to_datetime(fecha_desde).replace(day=1)
    hasta = pd.to_datetime(fecha_hasta)
    periodos = pd.date_range(desde, hasta, freq="MS")
    return [p.strftime("%Y-%m-%d") for p in periodos]


def _crear_estilo_ventas(
    columnas_dias: list[str],
    info_dias: dict[str, int],
    dias_visibles: int = 2,
    con_montos: bool = True,
    unidad: str = "bultos",
) -> SheetStyle:
    """
    Crea el estilo para el reporte de ventas con grupos de columnas.

    Args:
        columnas_dias: Lista de nombres de columnas de dias
        info_dias: Diccionario con info de dias habiles para mostrar en encabezado
        dias_visibles: Cantidad de dias al final que no se agrupan (default: 2)
        con_montos: Si False, omite columnas de monto/descuento del estilo y subtotales.
        unidad: "bultos" (enteros) o "htls" (con 2 decimales).

    Returns:
        SheetStyle configurado con el grupo de dias y filas de resumen
    """
    groups = []

    # Solo agrupar si hay mas dias que los visibles
    if len(columnas_dias) > dias_visibles:
        # Agrupar desde el primer dia hasta (total - dias_visibles)
        start_col = columnas_dias[0]
        end_col = columnas_dias[-(dias_visibles + 1)]
        groups.append(ColumnGroup(start_col=start_col, end_col=end_col, collapsed=True))

    # En HTLs los valores son decimales — formato con 2 decimales para no truncar.
    fmt_cantidad = '#,##0.00' if unidad == "htls" else '#,##0'

    # Formato de columnas: base + dias con ancho fijo
    column_formats = dict(VENTAS_COLUMN_FORMATS)
    if unidad == "htls":
        # Sobrescribir number_format en columnas de cantidad/tendencia/cupo (no en cobertura)
        for key in ("cant_generico", "tend_generico", "total_marca", "mmaa_marca",
                    "tend_marca", "cupo_generico", "cupo_marca"):
            col_name = COLUMN_NAMES[key]
            existing = column_formats.get(col_name)
            if existing is not None:
                column_formats[col_name] = replace(existing, number_format=fmt_cantidad)
    if not con_montos:
        for col in _MONEY_COLUMNS:
            column_formats.pop(col, None)
    for col_dia in columnas_dias:
        column_formats[col_dia] = ColumnFormat(number_format=fmt_cantidad, width=9.3)

    # Columnas con subtotal: todas las numéricas menos cobertura, Var% y ratios de cupo
    subtotal_cols = [
        COLUMN_NAMES["cant_generico"],
        COLUMN_NAMES["tend_generico"],
        COLUMN_NAMES["monto_generico"],
        COLUMN_NAMES["desc_generico"],
        # Cobertura (Generico) excluida
        COLUMN_NAMES["cupo_generico"],
        # Cupo vs Tend (Generico) excluida (ratio)
        *columnas_dias,
        COLUMN_NAMES["total_marca"],
        COLUMN_NAMES["tend_marca"],
        COLUMN_NAMES["cupo_marca"],
        # Cupo vs Tend (Marca) excluida (ratio)
        COLUMN_NAMES["mmaa_marca"],
        # Var% excluida
        # Cobertura (Marca) excluida
        COLUMN_NAMES["monto_marca"],
        COLUMN_NAMES["desc_marca"],
    ]
    if not con_montos:
        subtotal_cols = [c for c in subtotal_cols if c not in _MONEY_COLUMNS]

    return SheetStyle(
        column_formats=column_formats,
        column_groups=groups,
        summary_rows=info_dias,
        subtotal_columns=subtotal_cols,
    )


# Unidades disponibles
UNIDAD_BULTOS = "bultos"
UNIDAD_HTLS = "htls"

# Mapeo unidad -> columna de cantidad en el DataFrame
_COL_CANTIDAD = {
    UNIDAD_BULTOS: "cantidad",
    UNIDAD_HTLS: "cantidad_htls",
}

_UNIDADES = [
    (UNIDAD_BULTOS, "Ventas Bultos"),
    (UNIDAD_HTLS, "Ventas HTLs"),
]

# Mapeo (columna%, numerador, denominador) — keys de COLUMN_NAMES.
# Solo se escriben formulas para columnas presentes en la hoja (graceful skip
# cuando con_montos=False u otra columna falta).
_PCT_SUBTOTAL_FORMULAS: list[tuple[str, str, str]] = [
    ("cupo_vs_tend_generico", "tend_generico", "cupo_generico"),
    ("cupo_vs_tend_marca", "tend_marca", "cupo_marca"),
    ("var_mmaa_marca", "total_marca", "mmaa_marca"),
    ("desc_pct_generico", "desc_generico", "monto_generico"),
    ("desc_pct_marca", "desc_marca", "monto_marca"),
]


def _aplicar_bordes_blancos(ruta: Path, sheet_names: list[str]) -> None:
    """
    Pinta bordes blancos thin a todas las celdas del area de la tabla
    (desde la fila "Subtotales" hasta la ultima fila), para separar visualmente
    cada celda sobre los fondos coloreados.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side

    side = Side(style="thin", color="FFFFFF")
    border = Border(left=side, right=side, top=side, bottom=side)
    wb = load_workbook(ruta)
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Subtotales":
                start_row = r
                break
        if start_row is None:
            continue
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).border = border
    wb.save(ruta)


def _aplicar_estilo_fila_subtotal(ruta: Path, sheet_names: list[str]) -> None:
    """
    Pinta la fila "Subtotales" (la que tiene SUBTOTAL(109,...) y los % derivados)
    con un fondo morado oscuro y fuente blanca bold a lo largo de TODA la row.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    fill = PatternFill(start_color="4A235A", end_color="4A235A", fill_type="solid")
    wb = load_workbook(ruta)
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        subtotal_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Subtotales":
                subtotal_row = r
                break
        if subtotal_row is None:
            continue
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(subtotal_row, c)
            cell.fill = fill
            existing = cell.font
            cell.font = Font(
                name=existing.name,
                size=existing.size,
                bold=True,
                italic=existing.italic,
                color="FFFFFF",
            )
    wb.save(ruta)


def _aplicar_porcentajes_subtotal(ruta: Path, sheet_names: list[str]) -> None:
    """
    Inyecta formulas de porcentaje en la fila "Subtotales" (la que ya tiene
    SUBTOTAL(109,...) en las columnas sumables) para las columnas %.

    Cada formula es =IFERROR(num_subtotal / den_subtotal, 0) — usa las celdas
    de subtotal de la misma fila para que cambie cuando el filtrado autosuma
    distinto. Si la columna numerador o denominador no existe (ej: con_montos=False),
    salta esa formula silenciosamente.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(ruta)
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        subtotal_row = None
        header_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v == "Subtotales":
                subtotal_row = r
            elif v == "Sucursal":
                header_row = r
                break
        if subtotal_row is None or header_row is None:
            continue
        col_idx_by_name: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                col_idx_by_name[str(v)] = c
        for pct_key, num_key, den_key in _PCT_SUBTOTAL_FORMULAS:
            pct_name = COLUMN_NAMES.get(pct_key)
            num_name = COLUMN_NAMES.get(num_key)
            den_name = COLUMN_NAMES.get(den_key)
            if not pct_name or pct_name not in col_idx_by_name:
                continue
            if not num_name or num_name not in col_idx_by_name:
                continue
            if not den_name or den_name not in col_idx_by_name:
                continue
            pct_idx = col_idx_by_name[pct_name]
            num_letter = get_column_letter(col_idx_by_name[num_name])
            den_letter = get_column_letter(col_idx_by_name[den_name])
            cell = ws.cell(subtotal_row, pct_idx)
            cell.value = f"=IFERROR({num_letter}{subtotal_row}/{den_letter}{subtotal_row},0)"
            cell.number_format = "0.0%"
    wb.save(ruta)


def _aplicar_colores_por_generico(ruta: Path, sheet_names: list[str]) -> None:
    """
    Pinta cada fila de datos con el color asociado a su generico (col B).

    Recorre las hojas indicadas, encuentra el header buscando "Sucursal" en col A,
    y para cada fila de datos siguiente colorea TODA la fila segun el mapping
    `_GENERICO_FILL_COLORS`. Si el generico esta en `_GENERICO_WHITE_FONT`, usa
    fuente blanca preservando el resto del estilo.

    Genericos sin entrada en el mapping no se modifican.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    wb = load_workbook(ruta)
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Sucursal":
                header_row = r
                break
        if header_row is None:
            continue
        gen_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(header_row, c).value == "Generico":
                gen_col = c
                break
        if gen_col is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            gen = ws.cell(r, gen_col).value
            if not gen or gen not in _GENERICO_FILL_COLORS:
                continue
            color = _GENERICO_FILL_COLORS[gen]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            white_font = gen in _GENERICO_WHITE_FONT
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.fill = fill
                if white_font:
                    existing = cell.font
                    cell.font = Font(
                        name=existing.name,
                        size=existing.size,
                        bold=existing.bold,
                        italic=existing.italic,
                        color="FFFFFF",
                    )
    wb.save(ruta)


# Hojas de cobertura
_COB_GENERICO_SHEET = "Cobertura Generico"
_COB_MARCA_SHEET = "Cobertura Marca"

# Hoja extra "SUB DISTRIBUIDORES": solo se agrega al reporte de este supervisor.
# Combina sub-distribuidores de CASA CENTRAL (id_lista_precio=11, todos suc 1) +
# sub-distribuidores del interior (id_lista_precio=12) de SUS sucursales.
# Se abre por Origen (Casa Central / sucursal) + generico/marca.
_SUBDIST_SUPERVISOR = "Adrian Garcia"
_SUBDIST_SHEET = "SUB DISTRIBUIDORES"
# DIRECTIVA: la hoja SUB DISTRIBUIDORES SIEMPRE muestra solo genericos CCU,
# sin importar lo que pida el config. Fijo aca para que no se filtre de mas.
_SUBDIST_GENERICOS_CCU = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES"]

_COB_GENERICO_COLUMNS = {
    "sucursal": COLUMN_NAMES["sucursal"],
    "generico": COLUMN_NAMES["generico"],
    "clientes_compradores": "Clientes Compradores",
}

_COB_MARCA_COLUMNS = {
    "sucursal": COLUMN_NAMES["sucursal"],
    "marca": COLUMN_NAMES["marca"],
    "clientes_compradores": "Clientes Compradores",
}

_COBERTURA_STYLE = SheetStyle(
    column_formats={
        "Clientes Compradores": ColumnFormat(number_format="#,##0", width=20),
    },
    as_table=True,
    table_style="TableStyleMedium9",
)


def _preparar_cobertura(
    df_cob: pd.DataFrame | None,
    sort_cols: list[str],
    rename_map: dict[str, str],
) -> pd.DataFrame | None:
    """Ordena y renombra un DataFrame de cobertura para escribirlo como hoja Excel."""
    if df_cob is None or df_cob.empty:
        return None
    df = df_cob.sort_values(sort_cols).reset_index(drop=True)
    df = df[list(rename_map.keys())]
    return df.rename(columns=rename_map)


@dataclass
class ReporteVentasConfig:
    """Configuracion para generar un reporte de ventas."""
    fecha_desde: str
    fecha_hasta: str
    genericos: list[str] | None = None
    nombre_archivo: str | None = None
    con_slicers: bool = True
    con_cobertura: bool = True
    con_montos: bool = True

    def __post_init__(self):
        pass  # El nombre se genera en el servicio a partir de la ultima fecha real de ventas


@dataclass
class ReporteVentasResult:
    """Resultado de la generacion de un reporte."""
    ruta_archivo: Path
    registros_ventas: int
    registros_procesados: int
    sucursales: int
    genericos_incluidos: list[str]
    hojas: list[str] = None
    slicers_agregados: bool = False
    supervisor: str | None = None


def _nombre_reporte(
    df_ventas: pd.DataFrame,
    fecha_hasta: str,
    supervisor: str | None = None,
    nombre_explicito: str | None = None,
) -> str:
    """
    Genera el nombre de archivo para el reporte.

    Formato: 'Ventas {supervisor} - {dd-mm-yyyy}' o 'Ventas - {dd-mm-yyyy}'.
    La fecha es la ultima fecha con ventas reales; si no hay datos, usa fecha_hasta.

    Args:
        df_ventas: DataFrame con columna 'fecha' de las ventas
        fecha_hasta: Fecha limite del rango (fallback si df_ventas esta vacio)
        supervisor: Nombre del supervisor (None para reporte global)
        nombre_explicito: Si el usuario especifico un nombre custom, lo usa tal cual.
    """
    if nombre_explicito:
        return nombre_explicito

    if not df_ventas.empty and "fecha" in df_ventas.columns:
        ultima_fecha = pd.to_datetime(df_ventas["fecha"]).max().strftime("%d-%m-%Y")
    else:
        ultima_fecha = pd.to_datetime(fecha_hasta).strftime("%d-%m-%Y")

    if supervisor:
        return f"Ventas {supervisor} - {ultima_fecha}"
    return f"Ventas - {ultima_fecha}"


class VentasService(BaseService):
    """
    Servicio para generacion de reportes de ventas.

    Orquesta el flujo completo: extraccion, procesamiento y generacion de Excel.
    """

    SERVICE_SLUG = "ventas"
    GRANULARITY = "month"

    def _fetch_data(self, config: ReporteVentasConfig) -> tuple:
        """
        Extrae todos los datos necesarios para el reporte.

        Usa get_ventas_diarias_con_ruta() para incluir id_ruta y poder splitear
        zonas virtuales (ej: CASA CENTRAL → CASA CENTRAL + VALLE SALTA).

        Returns:
            (df_ventas, df_sucursales, df_articulos, df_cob_generico, df_cob_marca, df_mmaa, df_cupos, info_dias)
        """
        df_ventas = self.data_loader.get_ventas_diarias_con_ruta(
            config.fecha_desde,
            config.fecha_hasta,
            config.genericos
        )
        df_ventas = _aplicar_zonas_virtuales(zonas_config=_ZONAS_VIRTUALES_VENTAS, df=df_ventas)

        df_sucursales = self.data_loader.get_sucursales()
        # Agregar zonas virtuales a la lista de sucursales (solo si la sucursal real existe)
        for zona_nombre, zona_config in _ZONAS_VIRTUALES_VENTAS.items():
            if (
                zona_config["sucursal_real"] in df_sucursales["sucursal"].values
                and zona_nombre not in df_sucursales["sucursal"].values
            ):
                df_sucursales = pd.concat(
                    [df_sucursales, pd.DataFrame({"sucursal": [zona_nombre]})],
                    ignore_index=True,
                )

        df_articulos = self.data_loader.get_articulos(config.genericos)
        info_dias = calcular_info_dias(config.fecha_desde, config.fecha_hasta)

        df_cob_generico = None
        df_cob_marca = None

        if config.con_cobertura:
            periodos = _fechas_a_periodos(config.fecha_desde, config.fecha_hasta)
            try:
                df_cg = self.data_loader.get_cobertura_preventista_generico(periodos=periodos)
                if not df_cg.empty:
                    df_cg = _aplicar_zonas_virtuales(zonas_config=_ZONAS_VIRTUALES_VENTAS, df=df_cg)
                    df_cob_generico = (
                        df_cg.groupby(["sucursal", "generico"])["clientes_compradores"]
                        .sum()
                        .reset_index()
                    )
            except Exception:
                pass

            try:
                df_cm = self.data_loader.get_cobertura_preventista_marca(periodos=periodos)
                if not df_cm.empty:
                    df_cm = _aplicar_zonas_virtuales(zonas_config=_ZONAS_VIRTUALES_VENTAS, df=df_cm)
                    df_cob_marca = (
                        df_cm.groupby(["sucursal", "marca"])["clientes_compradores"]
                        .sum()
                        .reset_index()
                    )
            except Exception:
                pass

        df_mmaa = pd.DataFrame(columns=["sucursal", "generico", "marca", "cantidad", "cantidad_htls"])
        try:
            df_mmaa_raw = self.data_loader.get_ventas_historico_mmaa(
                config.fecha_desde, config.fecha_hasta, config.genericos
            )
            if not df_mmaa_raw.empty:
                df_mmaa_raw = _aplicar_zonas_virtuales(zonas_config=_ZONAS_VIRTUALES_VENTAS, df=df_mmaa_raw)
                df_mmaa = df_mmaa_raw.groupby(
                    ["sucursal", "generico", "marca"], as_index=False
                )[["cantidad", "cantidad_htls"]].sum()
        except Exception:
            pass

        df_cupos = pd.DataFrame(columns=["sucursal", "cupo_generico", "cupo"])
        try:
            periodo = pd.to_datetime(config.fecha_desde).strftime("%Y-%m")
            df_cupos_raw = self.data_loader.get_cupos(periodo)
            if not df_cupos_raw.empty:
                df_cupos_raw = _aplicar_zonas_virtuales(zonas_config=_ZONAS_VIRTUALES_VENTAS, df=df_cupos_raw)
                df_cupos = df_cupos_raw.groupby(["sucursal", "cupo_generico"], as_index=False)["cupo"].sum()
        except Exception:
            pass

        return df_ventas, df_sucursales, df_articulos, df_cob_generico, df_cob_marca, df_mmaa, df_cupos, info_dias

    def _build_workbook(
        self,
        nombre_archivo: str,
        fecha_desde: str,
        fecha_hasta: str,
        df_ventas: pd.DataFrame,
        df_sucursales: pd.DataFrame,
        df_articulos: pd.DataFrame,
        df_cob_generico: pd.DataFrame | None,
        df_cob_marca: pd.DataFrame | None,
        df_mmaa: pd.DataFrame | None,
        df_cupos: pd.DataFrame | None,
        info_dias: dict,
        con_slicers: bool,
        output_dir: Path | None = None,
        con_montos: bool = True,
        supervisor: str | None = None,
        subdist_sucursales: list[str] | None = None,
    ) -> tuple[Path, int, bool, list[str]]:
        """
        Genera el archivo Excel con hojas de ventas y cobertura.

        Returns:
            (ruta_archivo, total_procesados, slicers_ok, hojas)
        """
        writer = ExcelWriter(nombre_archivo, output_dir=output_dir)
        total_procesados = 0
        hojas = []

        for unidad, sheet_label in _UNIDADES:
            col_cantidad = _COL_CANTIDAD[unidad]
            df_procesado = procesar_ventas_diarias(
                df_ventas,
                fecha_desde,
                fecha_hasta,
                df_sucursales,
                df_articulos,
                col_cantidad=col_cantidad,
                df_cob_generico=df_cob_generico,
                df_cob_marca=df_cob_marca,
                df_mmaa=df_mmaa,
                df_cupos=df_cupos,
            )

            if not con_montos:
                df_procesado = df_procesado.drop(
                    columns=[c for c in _MONEY_COLUMNS if c in df_procesado.columns]
                )

            # Detectar columnas de dias (entre Marca y Total)
            columnas = list(df_procesado.columns)
            idx_marca = columnas.index(COLUMN_NAMES["marca"])
            idx_total = columnas.index(COLUMN_NAMES["total_marca"])
            columnas_dias = columnas[idx_marca + 1:idx_total]

            style = _crear_estilo_ventas(columnas_dias, info_dias, con_montos=con_montos, unidad=unidad)
            writer.add_sheet(df_procesado, sheet_name=sheet_label, style=style)
            total_procesados += len(df_procesado)
            hojas.append(sheet_label)

        # Hojas de cobertura
        df_cob_gen_sheet = _preparar_cobertura(
            df_cob_generico, ["sucursal", "generico"], _COB_GENERICO_COLUMNS
        )
        if df_cob_gen_sheet is not None:
            writer.add_sheet(df_cob_gen_sheet, sheet_name=_COB_GENERICO_SHEET, style=_COBERTURA_STYLE)
            hojas.append(_COB_GENERICO_SHEET)

        df_cob_marca_sheet = _preparar_cobertura(
            df_cob_marca, ["sucursal", "marca"], _COB_MARCA_COLUMNS
        )
        if df_cob_marca_sheet is not None:
            writer.add_sheet(df_cob_marca_sheet, sheet_name=_COB_MARCA_SHEET, style=_COBERTURA_STYLE)
            hojas.append(_COB_MARCA_SHEET)

        # Hoja extra SUB DISTRIBUIDORES — solo para Adrian Garcia.
        # CASA CENTRAL (lista 11) + sus sucursales del interior (lista 12),
        # abierto por origen/generico/marca.
        if supervisor and supervisor.strip().lower() == _SUBDIST_SUPERVISOR.lower():
            df_subdist = self.data_loader.get_ventas_subdistribuidores_sheet(
                fecha_desde, fecha_hasta,
                sucursales_interior=subdist_sucursales or [],
                genericos=_SUBDIST_GENERICOS_CCU,  # directiva: siempre solo CCU
            )
            if df_subdist is not None and not df_subdist.empty:
                df_subdist = df_subdist.rename(columns={
                    "origen": "Origen",
                    "razon_social": "Razon Social",
                    "fantasia": "Fantasia",
                    "generico": "Generico", "marca": "Marca",
                    "bultos": "Bultos", "htls": "HTLs",
                })
                style_subdist = SheetStyle(
                    numeric_format="#,##0",
                    column_formats={
                        "Origen": ColumnFormat(width=24, font_bold=True),
                        "Razon Social": ColumnFormat(width=32),
                        "Fantasia": ColumnFormat(width=26),
                        "Generico": ColumnFormat(width=18),
                        "Marca": ColumnFormat(width=22),
                        "Bultos": ColumnFormat(number_format="#,##0", width=12),
                        "HTLs": ColumnFormat(number_format="#,##0.0", width=12),
                    },
                    as_table=True,
                    table_style="TableStyleMedium9",
                )
                writer.add_sheet(df_subdist, sheet_name=_SUBDIST_SHEET, style=style_subdist)
                hojas.append(_SUBDIST_SHEET)

        ruta = writer.save()

        sheet_labels = [s for _, s in _UNIDADES]
        _aplicar_porcentajes_subtotal(ruta, sheet_labels)
        _aplicar_estilo_fila_subtotal(ruta, sheet_labels)
        _aplicar_colores_por_generico(ruta, sheet_labels)
        _aplicar_bordes_blancos(ruta, sheet_labels)

        slicers_ok = False
        if con_slicers and slicers_disponibles():
            for _, sheet_label in _UNIDADES:
                nombre_tabla = f"Tabla_{sheet_label.replace(' ', '_')}"
                agregar_slicers(ruta, nombre_tabla, SLICER_COLUMNS)
            slicers_ok = True

        return ruta, total_procesados, slicers_ok, hojas

    def generar_reporte(self, config: ReporteVentasConfig) -> ReporteVentasResult:
        """
        Genera un reporte de ventas completo con desglose diario.

        Genera un archivo Excel con dos hojas: Ventas Bultos y Ventas HTLs.
        Incluye columnas de cobertura (Generico y Marca) cruzando con tablas de cobertura.

        Args:
            config: Configuracion del reporte.

        Returns:
            ReporteVentasResult con informacion del reporte generado.
        """
        df_ventas, df_sucursales, df_articulos, df_cob_gen, df_cob_marca, df_mmaa, df_cupos, info_dias = (
            self._fetch_data(config)
        )

        nombre = _nombre_reporte(df_ventas, config.fecha_hasta, nombre_explicito=config.nombre_archivo)

        out = self._output_dir(config.fecha_desde)
        out.mkdir(parents=True, exist_ok=True)

        ruta, total_procesados, slicers_ok, hojas = self._build_workbook(
            nombre,
            config.fecha_desde,
            config.fecha_hasta,
            df_ventas,
            df_sucursales,
            df_articulos,
            df_cob_gen,
            df_cob_marca,
            df_mmaa,
            df_cupos,
            info_dias,
            config.con_slicers,
            output_dir=out,
            con_montos=config.con_montos,
        )

        genericos_incluidos = (
            df_articulos["generico"].unique().tolist() if not df_articulos.empty else []
        )

        return ReporteVentasResult(
            ruta_archivo=ruta,
            registros_ventas=len(df_ventas),
            registros_procesados=total_procesados,
            sucursales=len(df_sucursales),
            genericos_incluidos=genericos_incluidos,
            hojas=hojas,
            slicers_agregados=slicers_ok,
        )

    def generar_reporte_supervisores(
        self,
        config: ReporteVentasConfig,
        supervisores: dict[str, list[str]],
    ) -> list[ReporteVentasResult]:
        """
        Genera un archivo Excel por supervisor, filtrado por sus sucursales.

        Realiza una sola consulta a BD y luego filtra los datos por supervisor.

        Args:
            config: Configuracion base del reporte (fechas, genericos, etc.)
            supervisores: Mapeo {nombre_supervisor: [lista_de_sucursales]}

        Returns:
            Lista de ReporteVentasResult, uno por supervisor.
        """
        # Una sola consulta para todos los supervisores
        df_ventas, _, df_articulos, df_cob_gen, df_cob_marca, df_mmaa, df_cupos, info_dias = (
            self._fetch_data(config)
        )

        results = []
        for supervisor, sucursales_list in supervisores.items():
            # Expandir zonas virtuales (CASA CENTRAL → CASA CENTRAL + VALLE SALTA)
            sucursales_expandidas = _expandir_sucursales(sucursales_list, zonas_config=_ZONAS_VIRTUALES_VENTAS)

            # Filtrar datos al universo de este supervisor
            df_ventas_sup = df_ventas[df_ventas["sucursal"].isin(sucursales_expandidas)]
            df_sucursales_sup = pd.DataFrame({"sucursal": sucursales_expandidas})

            df_cob_gen_sup = None
            if df_cob_gen is not None:
                df_cob_gen_sup = df_cob_gen[df_cob_gen["sucursal"].isin(sucursales_expandidas)]

            df_cob_marca_sup = None
            if df_cob_marca is not None:
                df_cob_marca_sup = df_cob_marca[df_cob_marca["sucursal"].isin(sucursales_expandidas)]

            df_mmaa_sup = df_mmaa[df_mmaa["sucursal"].isin(sucursales_expandidas)] if not df_mmaa.empty else df_mmaa

            df_cupos_sup = df_cupos[df_cupos["sucursal"].isin(sucursales_expandidas)] if not df_cupos.empty else df_cupos

            # Nombre de archivo: "Ventas {supervisor} - {ultima_fecha}"
            nombre = _nombre_reporte(df_ventas_sup, config.fecha_hasta, supervisor=supervisor)

            out = self._output_dir(config.fecha_desde)
            out.mkdir(parents=True, exist_ok=True)

            ruta, total_procesados, slicers_ok, hojas = self._build_workbook(
                nombre,
                config.fecha_desde,
                config.fecha_hasta,
                df_ventas_sup,
                df_sucursales_sup,
                df_articulos,
                df_cob_gen_sup,
                df_cob_marca_sup,
                df_mmaa_sup,
                df_cupos_sup,
                info_dias,
                config.con_slicers,
                output_dir=out,
                con_montos=config.con_montos,
                supervisor=supervisor,
                subdist_sucursales=sucursales_list,
            )

            genericos_incluidos = (
                df_articulos["generico"].unique().tolist() if not df_articulos.empty else []
            )

            results.append(ReporteVentasResult(
                ruta_archivo=ruta,
                registros_ventas=len(df_ventas_sup),
                registros_procesados=total_procesados,
                sucursales=len(sucursales_expandidas),
                genericos_incluidos=genericos_incluidos,
                hojas=hojas,
                slicers_agregados=slicers_ok,
                supervisor=supervisor,
            ))

        return results

    def obtener_ventas(
        self,
        fecha_desde: str,
        fecha_hasta: str,
        genericos: list[str] | None = None
    ) -> pd.DataFrame:
        """
        Obtiene datos de ventas procesados sin generar Excel.

        Args:
            fecha_desde: Fecha inicio formato 'YYYY-MM-DD'
            fecha_hasta: Fecha fin formato 'YYYY-MM-DD'
            genericos: Lista de genericos a filtrar.

        Returns:
            DataFrame con ventas procesadas.
        """
        df_ventas = self.data_loader.get_ventas(fecha_desde, fecha_hasta, genericos)
        df_sucursales = self.data_loader.get_sucursales()
        df_articulos = self.data_loader.get_articulos(genericos)

        df_completo = completar_combinaciones(df_ventas, df_sucursales, df_articulos)

        return procesar_ventas(df_completo, fecha_desde, fecha_hasta)

    def listar_genericos_disponibles(self) -> list[str]:
        """Obtiene lista de genericos disponibles en la base de datos."""
        df = self.data_loader.get_articulos()
        return sorted(df["generico"].unique().tolist())

    def listar_sucursales(self) -> list[str]:
        """Obtiene lista de sucursales disponibles."""
        df = self.data_loader.get_sucursales()
        return df["sucursal"].tolist()
