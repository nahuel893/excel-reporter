"""Processor for ventas-articulo-diario: builds daily sales Excel with openpyxl."""

import calendar
import logging
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config.settings import DATA_OUTPUT

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="2D6A2E")
_TOTAL_FONT = Font(bold=True, color="FFFFFF")

_DOMINGO_FILL = PatternFill(fill_type="solid", fgColor="F2DCDB")
_CON_VENTA_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
_SIN_VENTA_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")

_THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

_BULTOS_FORMAT = "#,##0"
_DATE_FORMAT = "DD/MM/YYYY"

# ── Spanish day/month names ───────────────────────────────────────────────────

DIAS_SEMANA_ABBR = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

MESES_COMPLETOS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

MESES_ABREVIADOS = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

# Excel sheet name forbidden chars
_FORBIDDEN_CHARS = re.compile(r'[\[\]:*?/\\]')


def _sheet_name(articulo_nombre: str, anio: int, mes: int) -> str:
    """Build a valid Excel sheet name (max 31 chars)."""
    mes_abbr = MESES_ABREVIADOS[mes]
    suffix = f" - {mes_abbr} {anio}"
    max_art_len = 31 - len(suffix)
    clean = _FORBIDDEN_CHARS.sub("", articulo_nombre)
    return f"{clean[:max_art_len]}{suffix}"


def build_excel(
    anio: int,
    mes: int,
    articulo_nombre: str,
    id_articulo: int,
    id_sucursal: int,
    ventas_por_fecha: dict,
    nombre_archivo: str,
    output_dir: Path | None = None,
) -> Path:
    """Build the daily-sales Excel for one article+month. Returns the file path.

    Args:
        anio: Year (e.g. 2026)
        mes: Month (1-12)
        articulo_nombre: Human-readable article description (sheet title)
        id_articulo: Article ID (for metadata only)
        id_sucursal: Sucursal ID (for metadata only)
        ventas_por_fecha: {date: float} — only dates with sales (> 0)
        nombre_archivo: Base filename without extension
        output_dir: Directory to write the file. Defaults to DATA_OUTPUT.

    Returns:
        Path to the generated .xlsx file.
    """
    output = output_dir or DATA_OUTPUT
    output.mkdir(parents=True, exist_ok=True)

    mes_nombre = MESES_COMPLETOS[mes]
    _, days_in_month = calendar.monthrange(anio, mes)

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(articulo_nombre, anio, mes)

    # ── Column widths ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    # ── Row 1: Merged title ──────────────────────────────────────
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"{articulo_nombre} (id {id_articulo}) — Sucursal {id_sucursal} — {mes_nombre} {anio}"
    )
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: Empty ─────────────────────────────────────────────
    # intentionally blank

    # ── Row 3: Headers ───────────────────────────────────────────
    headers = ["Día", "Fecha", "Bultos"]
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # ── Rows 4+: One row per calendar day ───────────────────────
    total_bultos = 0.0
    dias_con_venta = 0

    for day in range(1, days_in_month + 1):
        row = day + 3  # row 4 = day 1
        d = date(anio, mes, day)
        weekday = d.weekday()  # 0=Mon … 6=Sun
        abbr = DIAS_SEMANA_ABBR[weekday]
        is_sunday = weekday == 6

        bultos = ventas_por_fecha.get(d)
        has_sales = bultos is not None and bultos > 0

        # Determine fill
        if is_sunday:
            fill = _DOMINGO_FILL
        elif has_sales:
            fill = _CON_VENTA_FILL
        else:
            fill = _SIN_VENTA_FILL

        # Col A: Día
        cell_a = ws.cell(row=row, column=1, value=f"{day} {abbr}")
        cell_a.fill = fill
        cell_a.border = _THIN_BORDER
        cell_a.alignment = Alignment(horizontal="center")

        # Col B: Fecha (date value, formatted)
        cell_b = ws.cell(row=row, column=2, value=d)
        cell_b.number_format = _DATE_FORMAT
        cell_b.fill = fill
        cell_b.border = _THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")

        # Col C: Bultos — None when 0 (blank), float otherwise
        if has_sales:
            cell_c = ws.cell(row=row, column=3, value=float(bultos))
            cell_c.number_format = _BULTOS_FORMAT
            total_bultos += float(bultos)
            dias_con_venta += 1
        else:
            cell_c = ws.cell(row=row, column=3, value=None)
        cell_c.fill = fill
        cell_c.border = _THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")

    # ── TOTAL row ────────────────────────────────────────────────
    total_row = days_in_month + 4  # one past last day row

    cell_t_a = ws.cell(row=total_row, column=1, value="TOTAL")
    cell_t_a.fill = _TOTAL_FILL
    cell_t_a.font = _TOTAL_FONT
    cell_t_a.border = _THIN_BORDER
    cell_t_a.alignment = Alignment(horizontal="center")

    cell_t_b = ws.cell(row=total_row, column=2, value=f"{dias_con_venta} días con venta")
    cell_t_b.fill = _TOTAL_FILL
    cell_t_b.font = _TOTAL_FONT
    cell_t_b.border = _THIN_BORDER

    cell_t_c = ws.cell(row=total_row, column=3, value=float(total_bultos))
    cell_t_c.number_format = _BULTOS_FORMAT
    cell_t_c.fill = _TOTAL_FILL
    cell_t_c.font = _TOTAL_FONT
    cell_t_c.border = _THIN_BORDER
    cell_t_c.alignment = Alignment(horizontal="center")

    # ── Save ─────────────────────────────────────────────────────
    ruta = output / f"{nombre_archivo}.xlsx"
    wb.save(str(ruta))
    logger.info("Ventas articulo generado: %s (%d dias con venta)", ruta.name, dias_con_venta)
    return ruta
