"""VentasCoberPreventistaMarcaService — ventas + cobertura por preventista y supervisor.

Para una marca y una sucursal, en un rango de dias, produce una hoja con dos bloques:

- POR PREVENTISTA: vendedor | supervisor | bultos | cobertura (clientes)
- POR SUPERVISOR:  supervisor | bultos | cobertura (clientes)

Cada bloque cierra con una fila TOTAL GENERAL (convencion del proyecto).

Con `incluir_mes_anterior` cada bloque muestra DOS periodos lado a lado: el mes
anterior cerrado primero y el rango pedido despues. La ventana anterior se DERIVA
de `fecha_desde` (ver src.core.periodos), nunca se escribe en el config.

Notas clave:
- Cobertura = clientes compradores DISTINTOS (cantidades_total > 0). NO es aditiva:
  la suma por supervisor puede superar el total, porque un cliente atendido por dos
  supervisores se cuenta una sola vez en el total. Por eso se calcula por separado en
  cada nivel desde el grano (vendedor, cliente).
- Tampoco es aditiva ENTRE PERIODOS: un cliente que compro en julio y en agosto
  cuenta 1 en cada mes, nunca 2. Cada columna de cobertura se cuenta contra su
  propio periodo, desde su propio dataframe.
- El match vendedor->supervisor sale del SUPERVISOR_VENDOR_MAP curado (dim_vendedor.
  supervisor NO es confiable); se excluye al gerente (GFARAH). Vendedores no mapeados
  (ej. DIRECTA) caen en "SIN SUPERVISOR".
- El acceso a datos usa clave compuesta (id_vendedor + id_sucursal) — ver DataLoader.
"""
import logging
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.output_paths import service_output_dir
from src.core.periodos import etiqueta_mes, rango_mes_anterior
from src.services.base_service import BaseService
from src.services.rebotes.constants import SUPERVISOR_VENDOR_MAP

logger = logging.getLogger(__name__)

GERENTE_KEY = "GFARAH"
SIN_SUPERVISOR = "SIN SUPERVISOR"
ID_SUCURSAL_CASA_CENTRAL = 1

HEADER_FILL, HEADER_FONT = "90A4AE", "FFFFFF"
SECTION_FILL, SECTION_FONT = "546E7A", "FFFFFF"
TOTAL_FILL = "FFE08A"  # ámbar — fila TOTAL GENERAL
_FMT_BULTOS, _FMT_COB = "#,##0.00", "#,##0"
_SHEET_TITLE = "Ventas Cob x Preventista"


@dataclass
class VentasCoberPreventistaMarcaConfig:
    """Config del informe.

    `objetivo_cobertura` agrega una columna Objetivo por bloque, calculada como
    un porcentaje de la cobertura de OTRA marca de referencia:

        {"marca": "SALTA", "pct_anterior": 0.20, "pct_actual": 0.25,
         "base_actual": "anterior" | "propio"}

    `base_actual` decide contra que mes se mide el bloque actual — el negocio lo
    fijo en el mes anterior, pero puede cambiar al propio sin tocar codigo.

    `clausula_gatillo` agrega una fila al pie con ese objetivo de VOLUMEN y el
    porcentaje que cada mes lleva alcanzado.
    """
    marca: str
    fecha_desde: str
    fecha_hasta: str
    id_sucursal: int = ID_SUCURSAL_CASA_CENTRAL
    nombre_archivo: str | None = None
    incluir_mes_anterior: bool = False
    objetivo_cobertura: dict | None = None
    clausula_gatillo: float | None = None


@dataclass
class VentasCoberPreventistaMarcaResult:
    """`total_bultos`/`cobertura_total` son siempre los del periodo pedido.

    Los campos `_prev` solo traen valor cuando se pidio el mes anterior.
    """
    ruta_archivo: Path
    preventistas: int
    total_bultos: float
    cobertura_total: int
    fecha_desde: str
    fecha_hasta: str
    total_bultos_prev: float | None = None
    cobertura_prev: int | None = None
    etiqueta_prev: str | None = None


@dataclass
class _Periodo:
    """Un periodo ya agregado, listo para volcarse a la hoja.

    Se guarda agregado por nivel (no crudo) porque la cobertura de cada nivel se
    cuenta desde el grano cliente y no se puede rederivar sumando el nivel de
    abajo.
    """
    etiqueta: str
    by_vend: pd.DataFrame
    by_sup: pd.DataFrame
    total_bultos: float
    cobertura_total: int
    objetivo_total: int | None = None


def _vendor_to_supervisor() -> dict[str, str]:
    """Invierte SUPERVISOR_VENDOR_MAP → {vendedor_upper: supervisor}, sin el gerente."""
    result: dict[str, str] = {}
    for sup, vendors in SUPERVISOR_VENDOR_MAP.items():
        if sup == GERENTE_KEY:
            continue
        for v in vendors:
            result[v.upper()] = sup
    return result


def _thin() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _agg(df: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    """Suma bultos + cuenta clientes distintos con compra > 0, agrupado por `keys`."""
    g = df.groupby(keys, as_index=False).agg(bultos=("bultos", "sum"))
    cob = (
        df[df["bultos"] > 0].groupby(keys)["id_cliente"].nunique().rename("cobertura")
    )
    out = g.merge(cob, on=keys, how="left")
    out["cobertura"] = out["cobertura"].fillna(0).astype(int)
    return out.sort_values("bultos", ascending=False)


def _redondear(valor: float) -> int:
    """Redondeo medio-arriba, como la funcion ROUND de Excel.

    El `round()` de Python usa banker's rounding: `round(2.5)` devuelve 2. Un
    objetivo que el usuario recalcula a mano en Excel le daria 3 y no cerraria
    contra el informe.
    """
    return int(Decimal(str(valor)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


_MEDIDAS_BASE = ["bultos", "cobertura"]
_FMT_MEDIDA = {"bultos": _FMT_BULTOS, "cobertura": _FMT_COB, "objetivo": _FMT_COB}


def _medidas(periodos: list["_Periodo"]) -> list[str]:
    """Medidas por bloque: se agrega `objetivo` solo si el config lo pidio."""
    extra = ["objetivo"] if "objetivo" in periodos[0].by_vend.columns else []
    return _MEDIDAS_BASE + extra


def _combinar(periodos: list["_Periodo"], keys: list[str], atributo: str) -> pd.DataFrame:
    """Outer-join one aggregation level across periods, one column pair each.

    Columns come out as ``{i}|bultos`` / ``{i}|cobertura`` keyed by POSITION, not
    by label: two periods could resolve to the same month name (a re-run with an
    odd window) and a label collision would silently overwrite a column.

    Outer, not inner: a preventista who sold in only one of the months must still
    appear, with 0 in the other. Those are precisely the rows worth reading.
    """
    medidas = _medidas(periodos)
    wide: pd.DataFrame | None = None
    for i, p in enumerate(periodos):
        parcial = getattr(p, atributo)[keys + medidas].rename(
            columns={m: f"{i}|{m}" for m in medidas}
        )
        wide = parcial if wide is None else wide.merge(parcial, on=keys, how="outer")
    assert wide is not None  # `periodos` nunca viene vacio
    wide = wide.fillna(0.0)
    for i in range(len(periodos)):
        for m in medidas:
            if m != "bultos":
                wide[f"{i}|{m}"] = wide[f"{i}|{m}"].astype(int)
    # Orden estable por el primer periodo (el mes cerrado), desempatando por el
    # siguiente y por nombre para que no baile entre corridas.
    orden = [f"{i}|bultos" for i in range(len(periodos))]
    return wide.sort_values(orden + keys, ascending=[False] * len(orden) + [True] * len(keys),
                            kind="mergesort")


class VentasCoberPreventistaMarcaService(BaseService):
    SERVICE_SLUG = "ventas-cober-preventista-marca"
    GRANULARITY = "month"

    def _traer(
        self, marca: str, config: VentasCoberPreventistaMarcaConfig,
        desde: str, hasta: str,
    ) -> pd.DataFrame:
        """Query one window at client grain and normalize vendedor/supervisor."""
        raw = self.data_loader.get_ventas_cobertura_por_vendedor(
            marca=marca,
            fecha_desde=desde,
            fecha_hasta=hasta,
            id_sucursal=config.id_sucursal,
        )
        if raw is None or raw.empty:
            logger.warning(
                "Sin ventas para marca=%s fechas=%s..%s suc=%s",
                marca, desde, hasta, config.id_sucursal,
            )
            raw = pd.DataFrame(columns=["vendedor", "id_cliente", "bultos"])

        raw = raw.copy()
        raw["vendedor"] = raw["vendedor"].fillna("(sin vendedor)")
        raw["bultos"] = raw["bultos"].fillna(0.0)
        raw["supervisor"] = raw["vendedor"].str.upper().map(_vendor_to_supervisor()).fillna(SIN_SUPERVISOR)
        return raw

    def _cobertura_de_referencia(
        self, config: VentasCoberPreventistaMarcaConfig, desde: str, hasta: str,
    ) -> dict:
        """Cobertura de la marca de referencia, por vendedor / supervisor / total.

        Usa el MISMO umbral que el resto del informe (`> 0`, dentro de `_agg`).
        Medir el objetivo con un umbral y el logro con otro compararia contra dos
        varas distintas e inflaria el cumplimiento.
        """
        raw = self._traer(config.objetivo_cobertura["marca"], config, desde, hasta)
        return {
            "vendedor": _agg(raw, ["vendedor"]).set_index("vendedor")["cobertura"].to_dict(),
            "supervisor": _agg(raw, ["supervisor"]).set_index("supervisor")["cobertura"].to_dict(),
            "total": int(raw[raw["bultos"] > 0]["id_cliente"].nunique()) if not raw.empty else 0,
        }

    @staticmethod
    def _aplicar_objetivo(p: _Periodo, base: dict, pct: float) -> None:
        """Escribe la columna `objetivo` en los dos niveles del periodo."""
        p.by_vend["objetivo"] = [
            _redondear(base["vendedor"].get(v, 0) * pct) for v in p.by_vend["vendedor"]
        ]
        p.by_sup["objetivo"] = [
            _redondear(base["supervisor"].get(s, 0) * pct) for s in p.by_sup["supervisor"]
        ]
        p.objetivo_total = _redondear(base["total"] * pct)

    def _periodo(
        self, config: VentasCoberPreventistaMarcaConfig,
        desde: str, hasta: str, etiqueta: str,
    ) -> _Periodo:
        """Query one window and aggregate it at both levels independently."""
        raw = self._traer(config.marca, config, desde, hasta)
        return _Periodo(
            etiqueta=etiqueta,
            by_vend=_agg(raw, ["vendedor", "supervisor"]),
            by_sup=_agg(raw, ["supervisor"]),
            total_bultos=float(raw["bultos"].sum()),
            # Distintos DENTRO de este periodo: nunca se cruza con el otro mes.
            cobertura_total=int(raw[raw["bultos"] > 0]["id_cliente"].nunique()),
        )

    def generar_reporte(
        self, config: VentasCoberPreventistaMarcaConfig
    ) -> VentasCoberPreventistaMarcaResult:
        actual = self._periodo(
            config, config.fecha_desde, config.fecha_hasta,
            etiqueta_mes(config.fecha_desde),
        )

        if config.incluir_mes_anterior:
            prev_desde, prev_hasta = rango_mes_anterior(config.fecha_desde)
            previo = self._periodo(config, prev_desde, prev_hasta, etiqueta_mes(prev_desde))
            # Mes cerrado primero: es la referencia contra la que se lee el parcial.
            periodos = [previo, actual]
        else:
            prev_desde = prev_hasta = None
            previo = None
            periodos = [actual]

        if config.objetivo_cobertura:
            cfg = config.objetivo_cobertura
            if previo is not None:
                self._aplicar_objetivo(
                    previo,
                    self._cobertura_de_referencia(config, prev_desde, prev_hasta),
                    cfg.get("pct_anterior", 0.0),
                )
            # El bloque actual se mide contra el mes anterior salvo que se pida el
            # propio; sin mes anterior cargado, cae al propio para no quedar en 0.
            usa_anterior = cfg.get("base_actual", "anterior") == "anterior" and previo is not None
            ventana = (prev_desde, prev_hasta) if usa_anterior else (config.fecha_desde, config.fecha_hasta)
            self._aplicar_objetivo(
                actual,
                self._cobertura_de_referencia(config, *ventana),
                cfg.get("pct_actual", 0.0),
            )

        nombre = config.nombre_archivo or f"Ventas y Cobertura {config.marca} por Preventista"
        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha_desde, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(config, periodos, ruta)

        return VentasCoberPreventistaMarcaResult(
            ruta_archivo=ruta,
            preventistas=len(_combinar(periodos, ["vendedor", "supervisor"], "by_vend")),
            total_bultos=actual.total_bultos,
            cobertura_total=actual.cobertura_total,
            fecha_desde=config.fecha_desde,
            fecha_hasta=config.fecha_hasta,
            total_bultos_prev=previo.total_bultos if previo else None,
            cobertura_prev=previo.cobertura_total if previo else None,
            etiqueta_prev=previo.etiqueta if previo else None,
        )

    def _build_workbook(self, config, periodos: list[_Periodo], ruta: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = _SHEET_TITLE
        border = _thin()
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        section_fill = PatternFill("solid", fgColor=SECTION_FILL)
        total_fill = PatternFill("solid", fgColor=TOTAL_FILL)

        n = len(periodos)
        multi = n > 1
        medidas = _medidas(periodos)
        ancho = len(medidas)                       # columnas por bloque de mes
        ultima_col = 2 + ancho * n

        def col_de(i: int, medida: str) -> int:
            """Columna de una medida dentro del bloque del periodo `i`."""
            return 3 + ancho * i + medidas.index(medida)

        fecha_txt = (
            config.fecha_desde if config.fecha_desde == config.fecha_hasta
            else f"{config.fecha_desde} a {config.fecha_hasta}"
        )
        if config.incluir_mes_anterior:
            prev_desde, prev_hasta = rango_mes_anterior(config.fecha_desde)
            fecha_txt = f"{prev_desde} a {prev_hasta}  ·  {fecha_txt}"
        ws["A1"] = f"Ventas y Cobertura — {config.marca}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = f"Sucursal: {config.id_sucursal}  |  {fecha_txt}  |  Cobertura = clientes compradores"
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        def section(r: int, title: str, primera: str, segunda: str) -> int:
            ws.cell(r, 1, title).font = Font(bold=True, color=SECTION_FONT, size=12)
            for c in range(1, ultima_col + 1):
                ws.cell(r, c).fill = section_fill
            r += 1
            if multi:
                # Fila de grupo: el nombre del mes cubre todas sus medidas.
                for i, p in enumerate(periodos):
                    col = col_de(i, medidas[0])
                    ws.merge_cells(start_row=r, start_column=col,
                                   end_row=r, end_column=col + ancho - 1)
                    cell = ws.cell(r, col, p.etiqueta)
                    cell.alignment = Alignment(horizontal="center")
                    for c in range(col, col + ancho):
                        ws.cell(r, c).fill = header_fill
                        ws.cell(r, c).font = Font(bold=True, color=HEADER_FONT)
                        ws.cell(r, c).border = border
                for c in (1, 2):
                    ws.cell(r, c).fill = header_fill
                    ws.cell(r, c).border = border
                r += 1
            headers = [primera, segunda] + [m.capitalize() for m in medidas] * n
            for c, h in enumerate(headers, 1):
                cell = ws.cell(r, c, h)
                cell.fill = header_fill
                cell.font = Font(bold=True, color=HEADER_FONT)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            return r + 1

        def measure_cells(r: int, valores: list[dict]) -> None:
            for i, vals in enumerate(valores):
                for m in medidas:
                    v = vals.get(m)
                    if v is None:
                        continue
                    c = ws.cell(r, col_de(i, m),
                                float(v) if m == "bultos" else int(v))
                    c.number_format = _FMT_MEDIDA[m]
                    c.border = border
                    c.alignment = Alignment(horizontal="right")

        def total_row(r: int) -> int:
            ws.cell(r, 1, "TOTAL GENERAL").font = Font(bold=True)
            for c in range(1, ultima_col + 1):
                ws.cell(r, c).fill = total_fill
                ws.cell(r, c).border = border
                ws.cell(r, c).font = Font(bold=True)
            # Cada periodo aporta SU total de clientes distintos, contado desde el
            # grano. No se suma entre meses ni entre niveles.
            measure_cells(r, [{"bultos": p.total_bultos, "cobertura": p.cobertura_total,
                               "objetivo": p.objetivo_total} for p in periodos])
            for i in range(n):
                for m in medidas:
                    ws.cell(r, col_de(i, m)).fill = total_fill
            return r + 2

        def gatillo_row(r: int) -> int:
            """Objetivo de VOLUMEN al pie, con lo alcanzado por cada mes.

            Compara contra `bultos`, no contra cobertura: es un piso de cajas.
            """
            meta = float(config.clausula_gatillo)
            ws.cell(r, 1, "CLAUSULA GATILLO").font = Font(bold=True)
            for c in range(1, ultima_col + 1):
                ws.cell(r, c).fill = total_fill
                ws.cell(r, c).border = border
                ws.cell(r, c).font = Font(bold=True)
            for i, p in enumerate(periodos):
                cm = ws.cell(r, col_de(i, "bultos"), meta)
                cm.number_format = _FMT_BULTOS
                cm.alignment = Alignment(horizontal="right")
                cp = ws.cell(r, col_de(i, "cobertura"),
                             p.total_bultos / meta if meta else None)
                cp.number_format = "0.0%"
                cp.alignment = Alignment(horizontal="right")
                for m in medidas:
                    ws.cell(r, col_de(i, m)).fill = total_fill
                    ws.cell(r, col_de(i, m)).font = Font(bold=True)
            return r + 2

        def volcar(r: int, wide: pd.DataFrame, cols_texto: list[str]) -> int:
            for _, row in wide.iterrows():
                for c, key in enumerate(cols_texto, 1):
                    ws.cell(r, c, row[key] if key else "").border = border
                measure_cells(r, [{m: row[f"{i}|{m}"] for m in medidas} for i in range(n)])
                r += 1
            return r

        r = 4
        r = section(r, "POR PREVENTISTA", "Vendedor", "Supervisor")
        r = volcar(r, _combinar(periodos, ["vendedor", "supervisor"], "by_vend"),
                   ["vendedor", "supervisor"])
        r = total_row(r)
        # La clausula va debajo de los preventistas, separada por la fila en
        # blanco que ya deja `total_row`.
        if config.clausula_gatillo:
            r = gatillo_row(r)

        r = section(r, "POR SUPERVISOR", "Supervisor", "")
        r = volcar(r, _combinar(periodos, ["supervisor"], "by_sup"), ["supervisor", ""])
        r = total_row(r)

        anchos = [26, 16] + [12] * (ancho * n)
        for c, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{6 if multi else 5}"
        wb.save(ruta)

    def run(self, config):
        return self.generar_reporte(config)
