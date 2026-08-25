"""VentasMarcaService — cantidad vendida por marca de un generico.

Reporte simple: una fila por marca del generico con la cantidad vendida (bultos)
en un rango de dias, ordenado de mayor a menor, con una fila TOTAL GENERAL
(convencion del proyecto: todo informe lleva fila de totales).

Con `incluir_mes_anterior` la hoja muestra DOS periodos lado a lado: el mes
anterior cerrado primero y el rango pedido despues. La ventana anterior se
DERIVA de `fecha` (ver src.core.periodos), nunca se escribe en el config: el
daily patchea fechas pero no el resto del JSON, asi que un mes hardcodeado se
desincroniza solo al cambiar de mes.
"""
import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.output_paths import service_output_dir
from src.core.periodos import etiqueta_mes, periodo_mes, rango_mes_anterior
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)

HEADER_FILL, HEADER_FONT = "4472C4", "FFFFFF"
TOTAL_FILL = "FFE08A"  # ámbar — fila TOTAL GENERAL
ID_SUCURSAL_CASA_CENTRAL = 1


@dataclass
class VentasMarcaConfig:
    """Config del reporte de ventas por marca.

    Args:
        generico: Nombre exacto del generico (ej. 'PERNOD RICARD').
        fecha: Dia desde (YYYY-MM-DD).
        fecha_hasta: Dia hasta (YYYY-MM-DD). None → mismo dia que `fecha`.
        id_sucursal: Sucursal a filtrar (default 1 = CASA CENTRAL).
        nombre_archivo: nombre de salida (sin extension).
        incluir_mes_anterior: si True agrega una columna con el mes anterior
            completo, derivado de `fecha`, a la izquierda del periodo pedido.
    """
    generico: str
    fecha: str
    fecha_hasta: str | None = None
    id_sucursal: int = ID_SUCURSAL_CASA_CENTRAL
    nombre_archivo: str | None = None
    incluir_mes_anterior: bool = False


@dataclass
class VentasMarcaResult:
    """Resultado del reporte.

    `total_bultos`/`cobertura_total` son siempre los del periodo pedido. Los
    campos `_prev` solo traen valor cuando se pidio el mes anterior.
    """
    ruta_archivo: Path
    marcas: int
    total_bultos: float
    fecha_desde: str
    fecha_hasta: str
    cobertura_total: int = 0
    total_bultos_prev: float | None = None
    cobertura_prev: int | None = None
    etiqueta_prev: str | None = None


@dataclass
class _Periodo:
    """Un periodo ya agregado, con su cobertura contada aparte.

    Se guarda el total de cobertura junto al agregado por marca porque NO se
    puede rederivar sumando las marcas: un cliente que compro dos marcas cuenta
    una sola vez en el total.
    """
    etiqueta: str
    por_marca: pd.DataFrame
    total_bultos: float
    cobertura_total: int


def _combinar(periodos: list[_Periodo]) -> pd.DataFrame:
    """Outer-join the periods on ``marca``, one column pair per period.

    Columns come out as ``{i}|bultos`` / ``{i}|cobertura`` keyed by POSITION, not
    by label: two periods could resolve to the same month name and a label
    collision would silently overwrite a column.

    Outer, not inner: a marca sold in only one of the months must still show up,
    with 0 in the other. An inner join would hide exactly the movements worth
    looking at — the marcas that appeared or disappeared.
    """
    wide: pd.DataFrame | None = None
    for i, p in enumerate(periodos):
        parcial = p.por_marca.rename(
            columns={"bultos": f"{i}|bultos", "cobertura": f"{i}|cobertura"}
        )
        wide = parcial if wide is None else wide.merge(parcial, on="marca", how="outer")
    assert wide is not None  # `periodos` nunca viene vacio
    wide = wide.fillna(0.0)
    for i in range(len(periodos)):
        wide[f"{i}|cobertura"] = wide[f"{i}|cobertura"].astype(int)
    # Orden estable por el primer periodo (el mes cerrado), desempatando por el
    # siguiente y por nombre para que no baile entre corridas.
    orden = [f"{i}|bultos" for i in range(len(periodos))]
    return wide.sort_values(orden + ["marca"],
                            ascending=[False] * len(orden) + [True],
                            kind="mergesort")


def _thin_border() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


class VentasMarcaService(BaseService):
    """Genera el reporte de cantidad vendida por marca de un generico."""

    SERVICE_SLUG = "ventas-marca"
    GRANULARITY = "month"

    def _fetch(self, config: VentasMarcaConfig, desde: str, hasta: str) -> pd.DataFrame:
        """Query the sales volume of one window: ``marca`` / ``bultos``."""
        df = self.data_loader.get_ventas_por_marca(
            generico=config.generico,
            fecha_desde=desde,
            fecha_hasta=hasta,
            id_sucursal=config.id_sucursal,
        )
        if df is None or df.empty:
            logger.warning(
                "Sin ventas para generico=%s fechas=%s..%s suc=%s",
                config.generico, desde, hasta, config.id_sucursal,
            )
            return pd.DataFrame({"marca": pd.Series(dtype="object"),
                                 "bultos": pd.Series(dtype="float")})
        df = df.copy()
        df["marca"] = df["marca"].fillna("(sin marca)")
        df["bultos"] = df["bultos"].fillna(0.0)
        return df[["marca", "bultos"]]

    def _periodo(
        self, config: VentasMarcaConfig, desde: str, hasta: str, etiqueta: str
    ) -> _Periodo:
        """Volume from `fact_ventas`, coverage from the ETL's `cob_*` tables.

        Las dos fuentes son distintas a proposito: el volumen es del fact, la
        cobertura es la definicion oficial del negocio y no se recalcula. El
        total del generico se lee de `cob_sucursal_generico` en vez de sumar las
        marcas, porque entre marcas la cobertura NO es aditiva (el mismo cliente
        compra varias). Entre rutas y preventistas si lo es, pero ese nivel no
        entra en este informe.
        """
        ventas = self._fetch(config, desde, hasta)
        periodo = periodo_mes(desde)

        cob = self.data_loader.get_cobertura_marca_de_generico(
            generico=config.generico, periodo=periodo, id_sucursal=config.id_sucursal,
        )
        if cob is None or cob.empty:
            cob = pd.DataFrame({"marca": pd.Series(dtype="object"),
                                "cobertura": pd.Series(dtype="int64")})

        # Outer: una marca puede tener venta sin fila de cobertura, o al reves si
        # la ventana de dias no cubre el mes entero. Ninguna de las dos se pierde.
        por_marca = ventas.merge(cob[["marca", "cobertura"]], on="marca", how="outer")
        por_marca["bultos"] = por_marca["bultos"].fillna(0.0)
        por_marca["cobertura"] = por_marca["cobertura"].fillna(0).astype(int)

        return _Periodo(
            etiqueta=etiqueta,
            por_marca=por_marca,
            total_bultos=float(ventas["bultos"].sum()) if not ventas.empty else 0.0,
            cobertura_total=self.data_loader.get_cobertura_generico(
                generico=config.generico, periodo=periodo, id_sucursal=config.id_sucursal,
            ),
        )

    def generar_reporte(self, config: VentasMarcaConfig) -> VentasMarcaResult:
        fecha_desde = config.fecha
        fecha_hasta = config.fecha_hasta or config.fecha

        actual = self._periodo(config, fecha_desde, fecha_hasta, etiqueta_mes(fecha_desde))

        if config.incluir_mes_anterior:
            prev_desde, prev_hasta = rango_mes_anterior(fecha_desde)
            previo = self._periodo(config, prev_desde, prev_hasta, etiqueta_mes(prev_desde))
            # Mes cerrado primero: es la referencia contra la que se lee el parcial.
            periodos = [previo, actual]
        else:
            previo = None
            periodos = [actual]

        wide = _combinar(periodos)

        nombre = config.nombre_archivo or f"Venta por Marca {config.generico} - {config.fecha}"
        out_dir = service_output_dir(self.SERVICE_SLUG, config.fecha, granularity="month")
        out_dir.mkdir(parents=True, exist_ok=True)
        ruta = out_dir / f"{nombre}.xlsx"

        self._build_workbook(config, fecha_desde, fecha_hasta, wide, periodos, ruta)

        return VentasMarcaResult(
            ruta_archivo=ruta,
            marcas=len(wide),
            total_bultos=actual.total_bultos,
            cobertura_total=actual.cobertura_total,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            total_bultos_prev=previo.total_bultos if previo else None,
            cobertura_prev=previo.cobertura_total if previo else None,
            etiqueta_prev=previo.etiqueta if previo else None,
        )

    def _build_workbook(
        self, config: VentasMarcaConfig, fecha_desde: str, fecha_hasta: str,
        df: pd.DataFrame, periodos: list[_Periodo], ruta: Path,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Venta x Marca"
        border = _thin_border()
        header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        tot_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")

        n = len(periodos)
        multi = n > 1
        ultima_col = 1 + 2 * n

        ws.column_dimensions["A"].width = 26
        for j in range(2, ultima_col + 1):
            ws.column_dimensions[get_column_letter(j)].width = 13

        ws["A1"] = f"Venta por Marca — {config.generico}"
        ws["A1"].font = Font(bold=True, size=13)
        fecha_txt = fecha_desde if fecha_desde == fecha_hasta else f"{fecha_desde} a {fecha_hasta}"
        if config.incluir_mes_anterior:
            prev_desde, prev_hasta = rango_mes_anterior(fecha_desde)
            fecha_txt = f"{prev_desde} a {prev_hasta}  ·  {fecha_txt}"
        ws["A2"] = (
            f"Fecha: {fecha_txt}  |  Sucursal: {config.id_sucursal}  |  "
            f"Bultos vendidos  ·  Cobertura = clientes compradores"
        )
        ws["A2"].font = Font(italic=True, size=10, color="546E7A")

        r = 4
        if multi:
            # Fila de grupo: el nombre del mes cubre sus dos medidas.
            for i, p in enumerate(periodos):
                col = 2 + 2 * i
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
                ws.cell(r, col, p.etiqueta).alignment = Alignment(horizontal="center")
                for c in (col, col + 1):
                    ws.cell(r, c).fill = header_fill
                    ws.cell(r, c).font = Font(bold=True, color=HEADER_FONT)
                    ws.cell(r, c).border = border
            ws.cell(r, 1).fill = header_fill
            ws.cell(r, 1).border = border
            r += 1

        for j, h in enumerate(["Marca", *(["Bultos", "Cobertura"] * n)], 1):
            hc = ws.cell(r, j, h)
            hc.fill = header_fill
            hc.font = Font(bold=True, color=HEADER_FONT)
            hc.alignment = Alignment(horizontal="center")
            hc.border = border
        r += 1

        def medidas(fila: int, valores: list[tuple[float, int]]) -> None:
            for i, (bultos, cob) in enumerate(valores):
                col = 2 + 2 * i
                cb = ws.cell(fila, col, float(bultos))
                cb.number_format = "#,##0.00"
                cb.border = border
                cb.alignment = Alignment(horizontal="right")
                cc = ws.cell(fila, col + 1, int(cob))
                cc.number_format = "#,##0"
                cc.border = border
                cc.alignment = Alignment(horizontal="right")

        for _, fila in df.iterrows():
            ca = ws.cell(r, 1, fila["marca"])
            ca.border = border
            ca.font = Font(bold=True)
            medidas(r, [(fila[f"{i}|bultos"], fila[f"{i}|cobertura"]) for i in range(n)])
            r += 1

        # TOTAL GENERAL (convencion: todo informe lleva fila de totales).
        # Los bultos se suman; la cobertura NO. Cada periodo aporta su propio
        # conteo de clientes distintos, calculado desde el grano — sumar las
        # marcas contaria dos veces al cliente que compro dos marcas.
        cta = ws.cell(r, 1, "TOTAL GENERAL")
        cta.font = Font(bold=True)
        cta.fill = tot_fill
        cta.border = border
        medidas(r, [(p.total_bultos, p.cobertura_total) for p in periodos])
        for c in range(2, ultima_col + 1):
            ws.cell(r, c).fill = tot_fill
            ws.cell(r, c).font = Font(bold=True)

        wb.save(ruta)

    def run(self, config: VentasMarcaConfig) -> VentasMarcaResult:
        return self.generar_reporte(config)
