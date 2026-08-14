"""VolumenCoberturaService — volumen y cobertura de un generico, por sucursal y por mes.

Cuatro hojas:

    Resumen           una fila por sucursal, un bloque de columnas por mes
                      (Bultos / HL / Cobertura) y el acumulado de la ventana.
                      Es la que se captura como imagen.
    Por Marca         un bloque por sucursal con sus marcas adentro y un subtotal
                      al pie, mas el consolidado de todas al final.
    Sucursal x Marca  matriz de bultos del acumulado: que marca entro donde.
    Criterio          como se conto cada numero, para auditarlo sin preguntar.

El servicio es generico a proposito: recibe el `generico` por config. Se creo
para PERNOD RICARD, que entro al interior en julio-2026, pero sirve igual para
cualquier otro sin tocar codigo.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.core.output_paths import service_output_dir
from src.services.base_service import BaseService

from .constants import (
    ETIQUETA_TOTAL,
    RUTA_DIRECTA,
    ID_FUERZA_VENTAS_PREVENTA,
    RUTAS_EXCLUIDAS,
    UMBRAL_COBERTURA,
    etiqueta_mes,
)
from .processor import (
    construir_bloques,
    construir_bloques_supervisor,
    construir_tabla,
    fila_total,
    matriz_sucursal_marca,
    universo_marcas,
    meses_con_movimiento,
    validar_particion,
)

logger = logging.getLogger(__name__)

# --- Paleta. Roles, no colores sueltos (skills/formato-excel-badie). ---------
HEADER_FILL = "2E75B6"      # encabezado de columna
BANDA_FILL = "DDEBF7"       # banda que agrupa el bloque de cada mes
ACUM_FILL = "FFF2CC"        # bloque acumulado: es un subtotal, no un mes mas
SUBTOTAL_FILL = "D9E1F2"    # subtotal de cada bloque de sucursal
TOTAL_FILL = "FFE08A"       # TOTAL GENERAL
ZEBRA_FILL = "F7F9FC"
# Rol "dato faltante": la marca NO llego a esa sucursal. Es informacion, no un
# hueco — por eso se dibuja y no se omite.
FALTANTE_FILL = "F4F6FA"
FALTANTE_FUENTE = "A3B0C4"
FUENTE_HEADER = "FFFFFF"
FUENTE_SUBTITULO = "546E7A"
BORDE = Side(style="thin", color="D9D9D9")
# Grueso solo en los limites de cada bloque de mes. Con todo del mismo grosor el
# ojo se cruza de mes al recorrer una fila de 12 columnas numericas.
BORDE_BLOQUE = Side(style="medium", color="8EA9DB")

FMT_BULTOS = "#,##0"
FMT_HL = "#,##0.00"
FMT_COBERTURA = "#,##0"
FMT_PCT = "0.0%"


@dataclass
class VolumenCoberturaConfig:
    fecha_desde: str
    fecha_hasta: str
    generico: str
    nombre_archivo: str
    sucursales_excluidas: list[int] = field(default_factory=list)
    supervisores_sucursales: dict[str, list[str]] = field(default_factory=dict)
    incluir_directa: bool = False
    output_dir: str | None = None


@dataclass
class VolumenCoberturaResult:
    ruta_archivo: Path
    meses: list[str]
    sucursales: int
    bultos: float
    hectolitros: float
    cobertura: int
    articulos_sin_factor: list[int]


class VolumenCoberturaService(BaseService):
    SERVICE_SLUG = "volumen-cobertura"

    def __init__(self, data_loader: DataLoader | None = None):
        self._data_loader = data_loader

    @property
    def data_loader(self) -> DataLoader:
        if self._data_loader is None:
            self._data_loader = DataLoader()
        return self._data_loader

    # --- datos --------------------------------------------------------------

    def _traer(self, config: VolumenCoberturaConfig) -> tuple[pd.DataFrame, pd.DataFrame, list[int]]:
        # El MISMO conjunto de rutas para ventas y para padron: son numerador y
        # denominador del % s/ padron.
        rutas = [r for r in RUTAS_EXCLUIDAS
                 if not (config.incluir_directa and r == RUTA_DIRECTA)]
        ventas = self.data_loader.get_ventas_generico_cliente_mes(
            generico=config.generico,
            fecha_desde=config.fecha_desde,
            fecha_hasta=config.fecha_hasta,
            id_fuerza_ventas=ID_FUERZA_VENTAS_PREVENTA,
            sucursales_excluidas=config.sucursales_excluidas,
            rutas_excluidas=rutas,
        )
        # El padron sale del MISMO universo de rutas que las ventas, o el peso
        # sobre padron compara un numerador con un denominador de otra cosa.
        padron = self.data_loader.get_padron_activo(rutas_excluidas=rutas)
        if config.sucursales_excluidas and not padron.empty:
            padron = padron[~padron["id_sucursal"].isin(config.sucursales_excluidas)]

        sin_factor = self.data_loader.get_articulos_sin_factor_hl(config.generico)
        ids_sin_factor = (
            [int(x) for x in sin_factor["id_articulo"]] if not sin_factor.empty else []
        )
        if ids_sin_factor:
            logger.warning(
                "%s: %d articulos sin factor de hectolitros — la columna HL sale corta: %s",
                config.generico, len(ids_sin_factor), ids_sin_factor[:10],
            )
        return ventas, padron, ids_sin_factor

    # --- excel --------------------------------------------------------------

    def _titulo(self, ws, texto: str, subtitulo: str, ancho: int) -> None:
        ws.cell(1, 1, texto).font = Font(bold=True, size=14)
        c = ws.cell(2, 1, subtitulo)
        c.font = Font(italic=True, size=10, color=FUENTE_SUBTITULO)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    def _bloques(self, meses: list[str], con_padron: bool) -> list[tuple[str, list[tuple[str, str, str]]]]:
        """(etiqueta del bloque, [(titulo, clave, number_format)]).

        Cada columna declara su formato al lado del nombre. Decidir el formato
        por posicion (``if j == col_acum + 1``) se rompe en silencio en cuanto
        alguien mete una columna al medio.
        """
        bloques: list[tuple[str, list[tuple[str, str, str]]]] = []
        for mes in meses:
            bloques.append((
                etiqueta_mes(mes),
                [
                    ("Bultos", f"bultos_{mes}", FMT_BULTOS),
                    ("HL", f"hl_{mes}", FMT_HL),
                    ("Cob.", f"cob_{mes}", FMT_COBERTURA),
                ],
            ))
        acum = [
            ("Bultos", "bultos_acum", FMT_BULTOS),
            ("HL", "hl_acum", FMT_HL),
            ("Cob.", "cob_acum", FMT_COBERTURA),
        ]
        if con_padron:
            acum += [("Padron", "padron", FMT_COBERTURA), ("% s/ Padron", "pct_padron", FMT_PCT)]
        bloques.append(("ACUMULADO", acum))
        return bloques

    def _mapa_bordes(self, bloques_col) -> dict[int, tuple[Side, Side]]:
        """columna -> (lado izquierdo, lado derecho).

        El borde grueso cae en el primer y el ultimo campo de cada bloque de
        mes, y encierra tambien la columna de etiquetas.
        """
        mapa: dict[int, tuple[Side, Side]] = {1: (BORDE_BLOQUE, BORDE_BLOQUE)}
        col = 2
        for _, cols in bloques_col:
            ini, fin = col, col + len(cols) - 1
            for j in range(ini, fin + 1):
                mapa[j] = (
                    BORDE_BLOQUE if j == ini else BORDE,
                    BORDE_BLOQUE if j == fin else BORDE,
                )
            col = fin + 1
        return mapa

    @staticmethod
    def _borde(mapa: dict[int, tuple[Side, Side]], col: int,
               arriba: Side = BORDE, abajo: Side = BORDE) -> Border:
        izq, der = mapa.get(col, (BORDE, BORDE))
        return Border(left=izq, right=der, top=arriba, bottom=abajo)

    def _hoja_tabla(
        self, wb: Workbook, nombre: str, titulo: str, subtitulo: str,
        etiqueta_dim: str, dimension: str,
        tabla: pd.DataFrame, total: dict, meses: list[str], con_padron: bool,
    ):
        ws = wb.create_sheet(nombre)
        bloques = self._bloques(meses, con_padron)
        n_cols = 1 + sum(len(cols) for _, cols in bloques)
        self._titulo(ws, titulo, subtitulo, n_cols)

        fila_banda, fila_header, fila_datos = 4, 5, 6
        mapa = self._mapa_bordes(bloques)

        # Banda de agrupacion: un bloque por mes + el acumulado.
        col = 2
        for etiqueta, cols in bloques:
            ini, fin = col, col + len(cols) - 1
            ws.merge_cells(start_row=fila_banda, start_column=ini, end_row=fila_banda, end_column=fin)
            c = ws.cell(fila_banda, ini, etiqueta)
            c.fill = PatternFill("solid", fgColor=ACUM_FILL if etiqueta == "ACUMULADO" else BANDA_FILL)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center")
            # El borde va celda por celda: en una combinada, openpyxl solo
            # dibuja el de la esquina y el bloque queda abierto a la derecha.
            for j in range(ini, fin + 1):
                ws.cell(fila_banda, j).border = self._borde(mapa, j, arriba=BORDE_BLOQUE)
            col = fin + 1
        ws.cell(fila_banda, 1).border = self._borde(mapa, 1, arriba=BORDE_BLOQUE)

        # Encabezados
        planas: list[tuple[str, str, str]] = [(etiqueta_dim, dimension, "")]
        for _, cols in bloques:
            planas += cols
        for j, (titulo_col, _, _) in enumerate(planas, start=1):
            c = ws.cell(fila_header, j, titulo_col)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(bold=True, color=FUENTE_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = self._borde(mapa, j)

        # Datos
        r = fila_datos
        for i, (_, registro) in enumerate(tabla.iterrows()):
            for j, (_, clave, fmt) in enumerate(planas, start=1):
                # El valor se guarda con todos sus decimales; el number_format
                # es de presentacion. Redondear aca romperia la conciliacion
                # contra la base.
                c = ws.cell(r, j, registro.get(clave))
                if fmt:
                    c.number_format = fmt
                if i % 2:
                    c.fill = PatternFill("solid", fgColor=ZEBRA_FILL)
                c.border = self._borde(mapa, j)
            r += 1

        # TOTAL GENERAL
        for j, (_, clave, fmt) in enumerate(planas, start=1):
            c = ws.cell(r, j, total.get(clave))
            if fmt:
                c.number_format = fmt
            c.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            c.font = Font(bold=True)
            c.border = self._borde(mapa, j, arriba=BORDE_BLOQUE, abajo=BORDE_BLOQUE)

        ws.column_dimensions["A"].width = 30
        for j in range(2, len(planas) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 12
        ws.freeze_panes = ws.cell(fila_datos, 2)
        ws.auto_filter.ref = f"A{fila_header}:{get_column_letter(len(planas))}{r - 1}"
        return ws

    def _hoja_bloques(
        self, wb: Workbook, nombre: str, titulo: str, subtitulo: str,
        bloques: list[tuple[str, pd.DataFrame, dict]],
        consolidado: tuple[pd.DataFrame, dict] | None,
        total: dict, meses: list[str],
        etiqueta_dim: str = "Marca", dimension: str = "marca",
        con_padron: bool = False,
    ):
        """Bloques con subtotal al pie: sucursal->marcas, o supervisor->sucursales.

        `consolidado` es opcional: en la hoja por marca cierra con el peso de
        cada marca en el total; en la de supervisores no hace falta, porque la
        fila de abajo de cada bloque ya es la sucursal.
        """
        ws = wb.create_sheet(nombre)
        bloques_col = self._bloques(meses, con_padron=con_padron)
        planas: list[tuple[str, str, str]] = [(etiqueta_dim, dimension, "")]
        for _, cols in bloques_col:
            planas += cols
        self._titulo(ws, titulo, subtitulo, len(planas))

        fila_banda, fila_header = 4, 5
        mapa = self._mapa_bordes(bloques_col)
        col = 2
        for etiqueta, cols in bloques_col:
            ini, fin = col, col + len(cols) - 1
            ws.merge_cells(start_row=fila_banda, start_column=ini, end_row=fila_banda, end_column=fin)
            c = ws.cell(fila_banda, ini, etiqueta)
            c.fill = PatternFill("solid", fgColor=ACUM_FILL if etiqueta == "ACUMULADO" else BANDA_FILL)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center")
            # Celda por celda: en una combinada openpyxl solo dibuja el borde de
            # la esquina y el bloque queda abierto a la derecha.
            for j in range(ini, fin + 1):
                ws.cell(fila_banda, j).border = self._borde(mapa, j, arriba=BORDE_BLOQUE)
            col = fin + 1
        ws.cell(fila_banda, 1).border = self._borde(mapa, 1, arriba=BORDE_BLOQUE)

        for j, (titulo_col, _, _) in enumerate(planas, start=1):
            c = ws.cell(fila_header, j, titulo_col)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(bold=True, color=FUENTE_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = self._borde(mapa, j)

        def _escribir(registro, r: int, fill: str | None, negrita: bool,
                      abajo: Side = BORDE) -> None:
            # Fila sin una sola venta en la ventana: la marca no llego a esta
            # sucursal. Se resalta en gris en vez de omitirla, asi el hueco se
            # ve al lado de las que si entraron.
            vacia = fill is None and not registro.get("bultos_acum")
            for j, (_, clave, fmt) in enumerate(planas, start=1):
                c = ws.cell(r, j, registro.get(clave))
                if fmt:
                    c.number_format = fmt
                if fill:
                    c.fill = PatternFill("solid", fgColor=fill)
                elif vacia:
                    c.fill = PatternFill("solid", fgColor=FALTANTE_FILL)
                if negrita:
                    c.font = Font(bold=True)
                elif vacia:
                    c.font = Font(color=FALTANTE_FUENTE, italic=True)
                c.border = self._borde(mapa, j, abajo=abajo)

        r = fila_header + 1
        for etiqueta, filas, subtotal in bloques:
            # Cabecera de sucursal: banda celeste que abre el bloque.
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(planas))
            c = ws.cell(r, 1, etiqueta)
            c.fill = PatternFill("solid", fgColor=BANDA_FILL)
            c.font = Font(bold=True)
            for j in range(1, len(planas) + 1):
                ws.cell(r, j).border = self._borde(mapa, j, arriba=BORDE_BLOQUE)
            r += 1
            for _, registro in filas.iterrows():
                _escribir(registro, r, None, False)
                r += 1
            _escribir(subtotal, r, SUBTOTAL_FILL, True, abajo=BORDE_BLOQUE)
            r += 2  # una fila en blanco entre bloques

        # Consolidado: responde "cuanto pesa cada marca en el total", que se
        # pierde si solo se abre por sucursal.
        if consolidado is not None:
            filas_cons, _ = consolidado
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(planas))
            c = ws.cell(r, 1, "CONSOLIDADO — TODAS LAS SUCURSALES")
            c.fill = PatternFill("solid", fgColor=ACUM_FILL)
            c.font = Font(bold=True)
            for j in range(1, len(planas) + 1):
                ws.cell(r, j).border = self._borde(mapa, j, arriba=BORDE_BLOQUE)
            r += 1
            for _, registro in filas_cons.iterrows():
                _escribir(registro, r, None, False)
                r += 1
        _escribir(total, r, TOTAL_FILL, True, abajo=BORDE_BLOQUE)

        ws.column_dimensions["A"].width = 34
        for j in range(2, len(planas) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 12
        # Coordenada como string, no ws.cell(): la primera fila de datos es la
        # cabecera del primer bloque, que esta combinada, y openpyxl no acepta
        # una MergedCell en freeze_panes.
        ws.freeze_panes = f"B{fila_header + 1}"
        return ws

    def _hoja_matriz(self, wb: Workbook, matriz: pd.DataFrame, subtitulo: str):
        ws = wb.create_sheet("Sucursal x Marca")
        self._titulo(ws, "Bultos acumulados por sucursal y marca", subtitulo, len(matriz.columns) + 1)
        fila_header = 4

        ws.cell(fila_header, 1, "Sucursal")
        for j, marca in enumerate(matriz.columns, start=2):
            ws.cell(fila_header, j, marca)
        for j in range(1, len(matriz.columns) + 2):
            c = ws.cell(fila_header, j)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(bold=True, color=FUENTE_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(left=BORDE, right=BORDE, top=BORDE_BLOQUE, bottom=BORDE)

        r = fila_header + 1
        for suc, registro in matriz.iterrows():
            ws.cell(r, 1, suc).border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)
            for j, marca in enumerate(matriz.columns, start=2):
                valor = float(registro[marca])
                c = ws.cell(r, j, valor)
                c.number_format = FMT_BULTOS
                c.border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)
                # Gris claro = la marca NO llego a esa sucursal. Es informacion,
                # no un hueco: distingue "no se distribuye" de "se vende poco".
                if valor == 0:
                    c.fill = PatternFill("solid", fgColor="F4F6FA")
                    c.font = Font(color="A3B0C4")
            r += 1

        ws.cell(r, 1, ETIQUETA_TOTAL).font = Font(bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=TOTAL_FILL)
        ws.cell(r, 1).border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE_BLOQUE)
        for j, marca in enumerate(matriz.columns, start=2):
            c = ws.cell(r, j, float(matriz[marca].sum()))
            c.number_format = FMT_BULTOS
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            c.border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE_BLOQUE)

        ws.column_dimensions["A"].width = 30
        for j in range(2, len(matriz.columns) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 13
        ws.freeze_panes = ws.cell(fila_header + 1, 2)
        return ws

    def _hoja_criterio(self, wb: Workbook, config: VolumenCoberturaConfig,
                       meses: list[str], sin_factor: list[int]) -> None:
        ws = wb.create_sheet("Criterio")
        lineas = [
            ("Generico", config.generico),
            ("Ventana pedida", f"{config.fecha_desde} a {config.fecha_hasta}"),
            ("Meses con movimiento", ", ".join(etiqueta_mes(m) for m in meses) or "(ninguno)"),
            ("", ""),
            ("Cobertura", f"clientes DISTINTOS con neto > {UMBRAL_COBERTURA:g} en el corte"),
            ("Clave del cliente", "(id_cliente, id_sucursal) — el id se reusa entre sucursales"),
            ("Orden del calculo", "se totaliza por cliente DENTRO del corte y recien ahi se filtra"),
            ("Cobertura acumulada", "se mide sobre la ventana completa; NO es la suma de los meses"),
            ("Cobertura entre sucursales", "SI es aditiva: el TOTAL suma las filas"),
            ("", ""),
            ("Bultos", "cantidades_total de fact_ventas, neto (anulado = false)"),
            ("Hectolitros", "cantidades_total x dim_articulo.factor_hectolitros, POR ARTICULO"),
            ("Articulos sin factor HL", str(sin_factor) if sin_factor else "ninguno — los HL cierran"),
            ("", ""),
            ("Fuerza de ventas", f"id_fuerza_ventas = {ID_FUERZA_VENTAS_PREVENTA} (preventa)"),
            ("", "con este filtro el conteo reproduce gold.cob_sucursal_* exacto"),
            ("Ruta DIRECTA (100)", "INCLUIDA" if config.incluir_directa else "EXCLUIDA"),
            ("", "DIRECTA son entregas sin visita de preventista. Incluida, el informe"),
            ("", "mide cuanto salio de la sucursal; excluida, que hizo la fuerza de venta."),
            ("", "El flag mueve ventas Y padron: son numerador y denominador del %."),
            ("Otras rutas excluidas", "CHOPERAS (200) en CASA CENTRAL"),
            ("Sucursales excluidas", str(config.sucursales_excluidas) or "ninguna"),
            ("", ""),
            ("Contra gold.cob_*", (
                "con DIRECTA incluida este calculo reproduce cob_sucursal_generico "
                "EXACTO en un mes cerrado (julio-2026 PERNOD: 726 contra 726)."
                if config.incluir_directa else
                "cob_sucursal_generico NO excluye DIRECTA, asi que da un poco mas "
                "que este informe (julio-2026 PERNOD: 726 contra 712)."
            )),
            ("", "Un mes ABIERTO difiere igual: cob_* es la foto que dejo la ultima"),
            ("", "corrida del ETL y fact_ventas ya tiene las ventas posteriores."),
            ("Padron", "gold.dim_cliente con anulado = false, mismas rutas excluidas"),
            ("", "es SCD tipo 1: foto de HOY, no del mes medido"),
        ]
        ws.cell(1, 1, "Como se calculo cada numero").font = Font(bold=True, size=14)
        for i, (etiqueta, valor) in enumerate(lineas, start=3):
            ws.cell(i, 1, etiqueta).font = Font(bold=True)
            ws.cell(i, 2, valor)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 95

    # --- orquestacion -------------------------------------------------------

    def generar_reporte(self, config: VolumenCoberturaConfig) -> VolumenCoberturaResult:
        ventas, padron, sin_factor = self._traer(config)
        if ventas.empty:
            raise ValueError(
                f"{config.generico}: sin ventas entre {config.fecha_desde} y "
                f"{config.fecha_hasta} con el criterio del informe. "
                "Revisar el generico, la ventana y las sucursales excluidas."
            )

        meses = meses_con_movimiento(ventas)
        etiquetas = ", ".join(etiqueta_mes(m) for m in meses)
        base = (
            f"Cobertura = clientes distintos con neto > 0 | Ventana {config.fecha_desde} a "
            f"{config.fecha_hasta} | Meses con movimiento: {etiquetas} | "
            f"El acumulado se mide sobre la ventana completa, NO es la suma de los meses"
        )

        por_suc = construir_tabla(ventas, padron, "des_sucursal")
        total_suc = fila_total(ventas, padron, "des_sucursal")

        # El mapa de supervisores tiene que ser una PARTICION. Si no lo es, los
        # subtotales suman dos veces la misma sucursal y el informe miente sin
        # que se note. Preferimos romper aca.
        mapa_sup = config.supervisores_sucursales or {}
        if mapa_sup:
            repetidas, huerfanas = validar_particion(ventas, mapa_sup)
            if repetidas:
                raise ValueError(
                    "supervisores_sucursales no es una particion: "
                    f"{repetidas} figuran en mas de un supervisor. Los subtotales "
                    "sumarian dos veces esas sucursales. Revisar el mapa: en "
                    "configs/ventas.json hay un supervisor que tiene TODAS las "
                    "sucursales y es el supraconjunto de los demas."
                )
            if huerfanas:
                logger.warning(
                    "sucursales sin supervisor en el mapa: %s — van al bloque "
                    "SIN SUPERVISOR", huerfanas,
                )
        por_marca = construir_tabla(ventas, pd.DataFrame(), "marca")
        total_marca = fila_total(ventas, pd.DataFrame(), "marca")

        wb = Workbook()
        wb.remove(wb.active)
        if mapa_sup:
            self._hoja_bloques(
                wb, "Resumen",
                f"{config.generico} — volumen y cobertura por supervisor y sucursal",
                base + " | Cada bloque es un supervisor; el subtotal suma sus "
                       "sucursales, que es la unica dimension donde la cobertura ES aditiva",
                construir_bloques_supervisor(ventas, padron, mapa_sup),
                None, total_suc, meses,
                etiqueta_dim="Supervisor / Sucursal", dimension="des_sucursal",
                con_padron=True,
            )
        else:
            self._hoja_tabla(
                wb, "Resumen", f"{config.generico} — volumen y cobertura por sucursal",
                base, "Sucursal", "des_sucursal", por_suc, total_suc, meses, con_padron=True,
            )
        self._hoja_bloques(
            wb, "Por Marca",
            f"{config.generico} — volumen y cobertura por sucursal y marca",
            base + " | Gris = la marca no llego a esa sucursal | La cobertura NO se suma "
                   "entre marcas: el mismo cliente compra varias, por eso cada "
                   "subtotal se recalcula en vez de sumar la columna",
            construir_bloques(ventas, "des_sucursal", "marca", universo_marcas(ventas)),
            (por_marca, total_marca), total_marca, meses,
            etiqueta_dim="Marca", dimension="marca", con_padron=False,
        )
        self._hoja_matriz(
            wb, matriz_sucursal_marca(ventas),
            base + " | Gris = la marca no llego a esa sucursal",
        )
        self._hoja_criterio(wb, config, meses, sin_factor)

        output_dir = (
            Path(config.output_dir) if config.output_dir
            else service_output_dir(self.SERVICE_SLUG, config.fecha_desde, "month")
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        ruta = output_dir / f"{config.nombre_archivo}.xlsx"
        wb.save(str(ruta))

        return VolumenCoberturaResult(
            ruta_archivo=ruta,
            meses=meses,
            sucursales=len(por_suc),
            bultos=float(total_suc["bultos_acum"]),
            hectolitros=float(total_suc["hl_acum"]),
            cobertura=int(total_suc["cob_acum"]),
            articulos_sin_factor=sin_factor,
        )
