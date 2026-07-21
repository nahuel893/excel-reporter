"""RED tests — S3.1: BASE control workbook writer (RF-10, RF-11).

Covers ``writers/base_control.py``:

  - 6-sheet output: FACT_NET, ART-ACCION, CLIENTE-FECHA, ACC-GEN, wapi,
    Reconciliacion — exact names + order (RF-10).
  - EVERY sheet ends with a distinctly-styled (bold + FFE08A fill)
    TOTAL GENERAL row (RF-10). The 4 pivot frames arrive with the row
    already baked in (pivots.py); base_control.py appends + styles it for
    the wapi-derived table and appends + styles it for the reconciliation
    sheet (which the sheet literally ENDS with).
  - Reconciliation sums the CORRECT named Facturacion-Neta/Descuentos
    columns (never a column-letter/AZ-AX-style drift bug) grouped by
    SUCURSAL, with a per-row AND overall tolerance flag (RF-11).
  - MIN/MAX date-range checks (aexcel vs wapi vs the configured period).
  - Multi-price terna audit section: terna key, candidate prices/Bonific,
    the deterministically-picked value, and the pick reason (RF-11 audit
    surface for the RF-12 diff harness).
  - Unresolved-rows sections fed by the RF-04/RF-05 flagged subsets.

Strict-TDD: written before ``writers/base_control.py`` exists (import fails
RED).
"""
from __future__ import annotations

import datetime as dt

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.services.acciones_comerciales.constants import (
    COL_PRECIO_FINAL,
    COL_PRECIO_FINAL_COMPROBANTE,
)
from src.services.acciones_comerciales.gold_source import MultiPriceTerna
from src.services.acciones_comerciales.writers.base_control import (
    SHEET_ACC_GEN,
    SHEET_ART_ACCION,
    SHEET_CLIENTE_FECHA,
    SHEET_FACT_NET,
    SHEET_RECONCILIACION,
    SHEET_WAPI,
    TOTAL_GENERAL_LABEL,
    _CLASIFICACION_DESFASAJE,
    _CLASIFICACION_REVISAR,
    _PRECIO_COMPARISON_HEADERS,
    _UNRESOLVED_PRECIO_HEADERS,
    _build_precio_comparison_rows,
    _build_unresolved_precio_rows,
    ReconciliationInputs,
    build_base_control_workbook,
)

_TOTAL_FILL_RGB = "00FFE08A"


# ─────────────────────────────────────────────────────────────────────────
# fixtures
# ─────────────────────────────────────────────────────────────────────────


def _pivot_with_total(rows: list[dict], total: dict) -> pd.DataFrame:
    """Mimic a pivots.py frame that already carries its own TOTAL GENERAL
    row (base_control.py only STYLES it for the 4 pivot sheets, never
    appends a second one)."""
    return pd.concat([pd.DataFrame(rows), pd.DataFrame([total])], ignore_index=True)


def _fact_net_df() -> pd.DataFrame:
    return _pivot_with_total(
        [
            {
                "Sucursal": "1 - CASA CENTRAL",
                "Código": 900,
                "Descripción_2": "ART UNO",
                "Descripción_3": "MARCA UNO",
                "Descripción_12": "CERVEZAS",
                "Suma de Facturacion Neta": 1000.0,
                "Suma de Descuentos": 100.0,
                "Suma de Campo1": 0.1,
            }
        ],
        {
            "Sucursal": TOTAL_GENERAL_LABEL,
            "Código": "",
            "Descripción_2": "",
            "Descripción_3": "",
            "Descripción_12": "",
            "Suma de Facturacion Neta": 1000.0,
            "Suma de Descuentos": 100.0,
            "Suma de Campo1": 0.1,
        },
    )


def _simple_descuento_pivot_df(label_col: str) -> pd.DataFrame:
    return _pivot_with_total(
        [{label_col: "1 - CASA CENTRAL", "Suma de Descuento": 90.0}],
        {label_col: TOTAL_GENERAL_LABEL, "Suma de Descuento": 90.0},
    )


def _acc_gen_df() -> pd.DataFrame:
    return _pivot_with_total(
        [
            {
                "SUCURSAL": "1 - CASA CENTRAL",
                "Acción": "ACC1",
                "Descripción Acción": "MVB PROMO",
                "mvb": "MVB",
                "(en blanco)": "",
                "CERVEZAS": 90.0,
                "AGUAS DANONE": 0.0,
                "VINOS CCU": 0.0,
                "PERNOD RICARD": 0.0,
                "SIDRAS Y LICORES": 0.0,
            }
        ],
        {
            "SUCURSAL": TOTAL_GENERAL_LABEL,
            "Acción": "",
            "Descripción Acción": "",
            "mvb": "",
            "(en blanco)": "",
            "CERVEZAS": 90.0,
            "AGUAS DANONE": 0.0,
            "VINOS CCU": 0.0,
            "PERNOD RICARD": 0.0,
            "SIDRAS Y LICORES": 0.0,
        },
    )


def _aexcel_data_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Sucursal": "1 - CASA CENTRAL",
                "Descripción Período": "2026-07-01",
                "Facturacion Neta": 1000.0,
                "Descuentos": 100.0,
            },
            {
                "Sucursal": "2 - SUCURSAL CAFAYATE",
                "Descripción Período": "2026-07-15",
                "Facturacion Neta": 500.0,
                "Descuentos": 50.0,
            },
        ]
    )


def _wapi_enriched_df() -> pd.DataFrame:
    # RF-04's SUCURSAL is a BARE name (spec scenario: SUCURSAL = "CASA
    # CENTRAL", no id prefix) — a DIFFERENT convention from aexcel's own
    # "{id} - {DESC}" Sucursal field used above.
    #
    # Decision 19: both diagnostic price columns are also present here.
    # Row 1 (CASA CENTRAL) has matching terna/comprobante prices (within
    # tolerance) — excluded from the COMPARACION PRECIO section. Row 2
    # (SUCURSAL CAFAYATE) DIFFERS by 5.0 — the one row that section must
    # surface.
    return pd.DataFrame(
        [
            {
                "Fecha": "2026-07-02",
                "SUCURSAL": "CASA CENTRAL",
                "Cantidad": 10.0,
                "Total2": 500.0,
                "Descuento": 90.0,
                "Comprobante": "FCVTAA000300850740",
                "Cod. Cliente": 100,
                "Artículo Distribuidora": 900,
                COL_PRECIO_FINAL: 50.0,
                COL_PRECIO_FINAL_COMPROBANTE: 50.0,
            },
            {
                "Fecha": "2026-07-20",
                "SUCURSAL": "SUCURSAL CAFAYATE",
                "Cantidad": 5.0,
                "Total2": 250.0,
                "Descuento": 60.0,
                "Comprobante": "FCVTAA000300850999",
                "Cod. Cliente": 101,
                "Artículo Distribuidora": 901,
                COL_PRECIO_FINAL: 60.0,
                COL_PRECIO_FINAL_COMPROBANTE: 65.0,
            },
        ]
    )


def _multi_price_ternas() -> list[MultiPriceTerna]:
    return [
        MultiPriceTerna(
            fecha="2026-07-01",
            id_cliente=100,
            id_articulo=900,
            candidate_precios=[50.0, 55.0],
            candidate_bonific=[0.1],
            picked_precio=55.0,
            picked_bonific=0.1,
            pick_reason="greatest Cantidades Totales; tie-break greatest Precio; tie-break lowest _id_linea (ctid-equivalent)",
        )
    ]


def _unresolved_sucursal_df() -> pd.DataFrame:
    return pd.DataFrame([{"Cod. Cliente": 999, "Fecha": "2026-07-03", "Artículo Distribuidora": 901}])


def _unresolved_precio_df() -> pd.DataFrame:
    # Row 1: Fecha == fecha_hasta ("2026-07-31") -> "dia en curso" (expected lag).
    # Row 2: Fecha < fecha_hasta ("2026-07-04") -> "dia cerrado" (real problem).
    # Deliberately ordered en-curso-first in the source to prove the writer
    # SORTS dia-cerrado rows first, not merely preserves insertion order.
    return pd.DataFrame(
        [
            {
                "Fecha": "2026-07-31",
                "Comprobante": "B-002",
                "Cod. Cliente": 200,
                "Razón Social": "CLIENTE DOS SA",
                "Artículo Distribuidora": 903,
                "Descripción Acción": "PROMO DOS",
            },
            {
                "Fecha": "2026-07-04",
                "Comprobante": "A-001",
                "Cod. Cliente": 100,
                "Razón Social": "CLIENTE UNO SA",
                "Artículo Distribuidora": 902,
                "Descripción Acción": "PROMO UNO",
            },
        ]
    )


def _build_workbook_path(tmp_path):
    reconciliation = ReconciliationInputs(
        aexcel_data=_aexcel_data_df(),
        wapi_enriched=_wapi_enriched_df(),
        multi_price_ternas=_multi_price_ternas(),
        unresolved_sucursal=_unresolved_sucursal_df(),
        unresolved_precio=_unresolved_precio_df(),
        fecha_desde="2026-07-01",
        fecha_hasta="2026-07-31",
    )
    return build_base_control_workbook(
        nombre_archivo="BASE control TEST",
        output_dir=tmp_path,
        fact_net=_fact_net_df(),
        art_accion=_simple_descuento_pivot_df("SUCURSAL"),
        cliente_fecha=_simple_descuento_pivot_df("SUCURSAL"),
        acc_gen=_acc_gen_df(),
        wapi_enriched=_wapi_enriched_df(),
        reconciliation=reconciliation,
    )


def _find_row(ws, value, col: int = 1) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, col).value == value:
            return row
    raise AssertionError(f"row with {value!r} in col {col} not found in sheet {ws.title!r}")


# ─────────────────────────────────────────────────────────────────────────
# sheet structure (RF-10)
# ─────────────────────────────────────────────────────────────────────────


class TestSheetStructure:
    def test_six_sheets_exact_names_and_order(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        assert wb.sheetnames == [
            SHEET_FACT_NET,
            SHEET_ART_ACCION,
            SHEET_CLIENTE_FECHA,
            SHEET_ACC_GEN,
            SHEET_WAPI,
            SHEET_RECONCILIACION,
        ]

    @pytest.mark.parametrize(
        "sheet_name",
        [SHEET_FACT_NET, SHEET_ART_ACCION, SHEET_CLIENTE_FECHA, SHEET_ACC_GEN, SHEET_WAPI, SHEET_RECONCILIACION],
    )
    def test_total_general_row_ends_every_sheet_distinctly_styled(self, tmp_path, sheet_name):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[sheet_name]
        last_row = ws.max_row
        label_cell = ws.cell(last_row, 1)
        assert label_cell.value == TOTAL_GENERAL_LABEL
        assert label_cell.font.bold is True
        assert label_cell.fill.start_color.rgb == _TOTAL_FILL_RGB


# ─────────────────────────────────────────────────────────────────────────
# wapi sheet — summable Total2/Descuento (RF-10)
# ─────────────────────────────────────────────────────────────────────────


class TestWapiSheetTotals:
    def test_wapi_total_general_sums_total2_and_descuento(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_WAPI]
        headers = [c.value for c in ws[1]]
        last_row = ws.max_row
        total2_col = headers.index("Total2") + 1
        descuento_col = headers.index("Descuento") + 1
        assert ws.cell(last_row, total2_col).value == 750.0
        assert ws.cell(last_row, descuento_col).value == 150.0


# ─────────────────────────────────────────────────────────────────────────
# reconciliation sheet — correct FN/Descuentos columns, tolerance (RF-11)
# ─────────────────────────────────────────────────────────────────────────


class TestReconciliationFacturacionDescuentos:
    def test_reconciliation_sums_correct_named_columns_per_sucursal(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        # aexcel's "1 - CASA CENTRAL" is normalized (prefix stripped) to
        # match wapi's bare "CASA CENTRAL" before grouping.
        row = _find_row(ws, "CASA CENTRAL")
        values = [ws.cell(row, c).value for c in range(1, 7)]
        # SUCURSAL | Facturacion Neta (aexcel) | Descuentos (aexcel) | Descuento (wapi) | Delta | tolerancia
        assert values[1] == 1000.0
        assert values[2] == 100.0
        assert values[3] == 90.0
        assert values[4] == pytest.approx(10.0)
        assert values[5] == "NO"  # delta 10.0 exceeds the 0.01 tolerance

    def test_reconciliation_ends_with_total_general_over_delta_columns(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]
        last_row = ws.max_row
        values = [ws.cell(last_row, c).value for c in range(1, 7)]
        assert values[0] == TOTAL_GENERAL_LABEL
        assert values[1] == 1500.0  # 1000 + 500
        assert values[2] == 150.0  # 100 + 50
        assert values[3] == 150.0  # 90 + 60
        assert values[4] == pytest.approx(0.0, abs=1e-9)  # deltas net to ~0
        assert values[5] == "SI"  # grand-total delta IS within tolerance


class TestReconciliationDateRangeChecks:
    def test_min_max_date_range_per_source(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        ax_row = _find_row(ws, "aexcel (gold)")
        assert ws.cell(ax_row, 2).value == "2026-07-01"
        assert ws.cell(ax_row, 3).value == "2026-07-15"

        wapi_row = _find_row(ws, "wapi")
        assert ws.cell(wapi_row, 2).value == "2026-07-02"
        assert ws.cell(wapi_row, 3).value == "2026-07-20"


class TestReconciliationMultiPriceTernaSection:
    def test_multi_price_terna_row_present_with_pick_reason(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        header_row = _find_row(ws, "Fecha", col=1)
        # header row for the multi-price section must declare all 8 columns
        headers = [ws.cell(header_row, c).value for c in range(1, 9)]
        assert headers == [
            "Fecha",
            "Cod. Cliente",
            "Codigo Articulo",
            "Precios candidatos",
            "Bonific candidatos",
            "Precio elegido",
            "Bonific elegido",
            "Motivo de eleccion",
        ]

        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2026-07-01"
        assert ws.cell(data_row, 2).value == 100
        assert ws.cell(data_row, 3).value == 900
        assert "50.0" in ws.cell(data_row, 4).value
        assert "55.0" in ws.cell(data_row, 4).value
        assert ws.cell(data_row, 6).value == 55.0
        assert ws.cell(data_row, 7).value == 0.1
        assert "Cantidades Totales" in ws.cell(data_row, 8).value


class TestReconciliationUnresolvedSections:
    def test_unresolved_sucursal_rows_present(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        row = _find_row(ws, 999, col=1)
        assert ws.cell(row, 2).value == "2026-07-03"

    def test_unresolved_precio_section_has_seven_column_header(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        title_row = _find_row(ws, "FILAS NO RESUELTAS - PRECIO (RF-05)")
        header_row = title_row + 1
        headers = [ws.cell(header_row, c).value for c in range(1, 8)]
        assert headers == [
            "Fecha",
            "Comprobante",
            "Cod. Cliente",
            "Razón Social",
            "Artículo Distribuidora",
            "Descripción Acción",
            "Clasificacion",
        ]

    def test_unresolved_precio_rows_present_richer_columns_and_classified(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        title_row = _find_row(ws, "FILAS NO RESUELTAS - PRECIO (RF-05)")
        header_row = title_row + 1

        # dia-cerrado row (Fecha "2026-07-04" < fecha_hasta "2026-07-31") is
        # sorted FIRST even though it was the SECOND row in the source df.
        first_data_row = header_row + 1
        values = [ws.cell(first_data_row, c).value for c in range(1, 8)]
        assert values == [
            "2026-07-04",
            "A-001",
            100,
            "CLIENTE UNO SA",
            902,
            "PROMO UNO",
            "REVISAR (dia cerrado)",
        ]

        # dia-en-curso row (Fecha "2026-07-31" == fecha_hasta) sorts SECOND.
        second_data_row = header_row + 2
        values = [ws.cell(second_data_row, c).value for c in range(1, 8)]
        assert values == [
            "2026-07-31",
            "B-002",
            200,
            "CLIENTE DOS SA",
            903,
            "PROMO DOS",
            "Desfasaje de carga (dia en curso)",
        ]


# ─────────────────────────────────────────────────────────────────────────
# Decision 19 — COMPARACION PRECIO: TERNA vs COMPROBANTE section
# (BASE-control-ONLY parallel-run comparison; placed BEFORE the final
# FACTURACION/DESCUENTOS block, which must remain the sheet's last section).
# ─────────────────────────────────────────────────────────────────────────


class TestReconciliationPrecioComparisonSection:
    def test_comparison_section_lists_only_the_differing_row_with_count(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        title_row = None
        for r in range(1, ws.max_row + 1):
            value = ws.cell(r, 1).value
            if isinstance(value, str) and value.startswith("COMPARACION PRECIO"):
                title_row = r
                break
        assert title_row is not None, "COMPARACION PRECIO section title not found"
        assert "1 filas difieren de 2 con ambos precios" in ws.cell(title_row, 1).value

        header_row = title_row + 1
        headers = [ws.cell(header_row, c).value for c in range(1, 7)]
        assert headers == _PRECIO_COMPARISON_HEADERS
        assert headers == [
            "Comprobante",
            "Cod. Cliente",
            "Artículo Distribuidora",
            "PRECIO FINAL (terna)",
            "PRECIO FINAL (comprobante)",
            "Delta",
        ]

        data_row = header_row + 1
        values = [ws.cell(data_row, c).value for c in range(1, 7)]
        assert values[0] == "FCVTAA000300850999"
        assert values[1] == 101
        assert values[2] == 901
        assert values[3] == 60.0
        assert values[4] == 65.0
        assert values[5] == pytest.approx(-5.0)

    def test_section_sits_before_the_final_facturacion_descuentos_block(self, tmp_path):
        path = _build_workbook_path(tmp_path)
        wb = load_workbook(str(path))
        ws = wb[SHEET_RECONCILIACION]

        comparison_title_row = None
        for r in range(1, ws.max_row + 1):
            value = ws.cell(r, 1).value
            if isinstance(value, str) and value.startswith("COMPARACION PRECIO"):
                comparison_title_row = r
                break
        facturacion_title_row = _find_row(ws, "RECONCILIACION FACTURACION / DESCUENTOS (RF-11)")

        assert comparison_title_row < facturacion_title_row
        # RF-10: the sheet's LAST row must still be the FACTURACION/DESCUENTOS
        # TOTAL GENERAL row — the new section must never displace it.
        assert ws.cell(ws.max_row, 1).value == TOTAL_GENERAL_LABEL
        assert facturacion_title_row < ws.max_row


class TestBuildPrecioComparisonRows:
    def test_only_rows_with_both_prices_and_delta_over_tolerance_are_listed(self):
        df = pd.DataFrame(
            [
                {
                    "Comprobante": "A",
                    "Cod. Cliente": 1,
                    "Artículo Distribuidora": 900,
                    COL_PRECIO_FINAL: 50.0,
                    COL_PRECIO_FINAL_COMPROBANTE: 50.0,
                },
                {
                    "Comprobante": "B",
                    "Cod. Cliente": 2,
                    "Artículo Distribuidora": 901,
                    COL_PRECIO_FINAL: 60.0,
                    COL_PRECIO_FINAL_COMPROBANTE: 65.0,
                },
                {
                    "Comprobante": "C",
                    "Cod. Cliente": 3,
                    "Artículo Distribuidora": 902,
                    COL_PRECIO_FINAL: float("nan"),
                    COL_PRECIO_FINAL_COMPROBANTE: 10.0,
                },
            ]
        )
        rows, both_present = _build_precio_comparison_rows(df, 0.01)
        assert both_present == 2  # A and B have both prices; C is excluded (NaN terna)
        assert len(rows) == 1
        assert rows[0][0] == "B"
        assert rows[0][-1] == pytest.approx(-5.0)

    def test_within_tolerance_excluded(self):
        df = pd.DataFrame(
            [
                {
                    "Comprobante": "A",
                    "Cod. Cliente": 1,
                    "Artículo Distribuidora": 900,
                    COL_PRECIO_FINAL: 50.0,
                    COL_PRECIO_FINAL_COMPROBANTE: 50.005,
                }
            ]
        )
        rows, both_present = _build_precio_comparison_rows(df, 0.01)
        assert both_present == 1
        assert rows == []

    def test_missing_required_columns_returns_empty(self):
        df = pd.DataFrame([{"Fecha": "2026-07-01"}])
        rows, both_present = _build_precio_comparison_rows(df, 0.01)
        assert rows == []
        assert both_present == 0

    def test_empty_dataframe_returns_empty(self):
        rows, both_present = _build_precio_comparison_rows(pd.DataFrame(), 0.01)
        assert rows == []
        assert both_present == 0

    def test_headers_constant_matches_expected_order(self):
        assert _PRECIO_COMPARISON_HEADERS == [
            "Comprobante",
            "Cod. Cliente",
            "Artículo Distribuidora",
            "PRECIO FINAL (terna)",
            "PRECIO FINAL (comprobante)",
            "Delta",
        ]


# ─────────────────────────────────────────────────────────────────────────
# _build_unresolved_precio_rows — direct unit coverage of the classifier
# and the row builder (Fecha/fecha_hasta type handling, sort order,
# defensive missing-column handling).
# ─────────────────────────────────────────────────────────────────────────


def _precio_row(**overrides) -> dict:
    base = {
        "Fecha": "2026-07-04",
        "Comprobante": "A-001",
        "Cod. Cliente": 100,
        "Razón Social": "CLIENTE UNO SA",
        "Artículo Distribuidora": 902,
        "Descripción Acción": "PROMO UNO",
    }
    base.update(overrides)
    return base


class TestBuildUnresolvedPrecioRowsClassification:
    def test_fecha_equals_fecha_hasta_string_is_dia_en_curso(self):
        df = pd.DataFrame([_precio_row(Fecha="2026-07-31")])
        rows = _build_unresolved_precio_rows(df, "2026-07-31")
        assert rows[0][-1] == _CLASIFICACION_DESFASAJE

    def test_fecha_before_fecha_hasta_string_is_dia_cerrado(self):
        df = pd.DataFrame([_precio_row(Fecha="2026-07-04")])
        rows = _build_unresolved_precio_rows(df, "2026-07-31")
        assert rows[0][-1] == _CLASIFICACION_REVISAR

    def test_fecha_hasta_none_classifies_everything_revisar(self):
        df = pd.DataFrame([_precio_row(Fecha="2026-07-31")])
        rows = _build_unresolved_precio_rows(df, None)
        assert rows[0][-1] == _CLASIFICACION_REVISAR

    def test_pandas_timestamp_fecha_compares_on_date_only(self):
        df = pd.DataFrame([_precio_row(Fecha=pd.Timestamp("2026-07-31 08:30:00"))])
        rows = _build_unresolved_precio_rows(df, "2026-07-31")
        assert rows[0][-1] == _CLASIFICACION_DESFASAJE
        # the Fecha CELL VALUE itself stays unchanged (RF-23 — no reformatting)
        assert rows[0][0] == pd.Timestamp("2026-07-31 08:30:00")

    def test_python_datetime_fecha_vs_date_fecha_hasta(self):
        df = pd.DataFrame([_precio_row(Fecha=dt.datetime(2026, 7, 31, 23, 59))])
        rows = _build_unresolved_precio_rows(df, dt.date(2026, 7, 31))
        assert rows[0][-1] == _CLASIFICACION_DESFASAJE

    def test_missing_columns_default_to_none_defensively(self):
        df = pd.DataFrame([{"Fecha": "2026-07-04"}])
        rows = _build_unresolved_precio_rows(df, "2026-07-31")
        row = rows[0]
        assert row[0] == "2026-07-04"
        assert row[1] is None  # Comprobante missing
        assert row[2] is None  # Cod. Cliente missing
        assert row[-1] == _CLASIFICACION_REVISAR

    def test_empty_dataframe_returns_no_rows(self):
        assert _build_unresolved_precio_rows(pd.DataFrame(), "2026-07-31") == []

    def test_sort_puts_dia_cerrado_before_dia_en_curso(self):
        df = pd.DataFrame(
            [
                _precio_row(Fecha="2026-07-31", Comprobante="B-002"),  # dia en curso
                _precio_row(Fecha="2026-07-04", Comprobante="A-001"),  # dia cerrado
            ]
        )
        rows = _build_unresolved_precio_rows(df, "2026-07-31")
        assert rows[0][-1] == _CLASIFICACION_REVISAR
        assert rows[1][-1] == _CLASIFICACION_DESFASAJE

    def test_headers_constant_matches_expected_order(self):
        assert _UNRESOLVED_PRECIO_HEADERS == [
            "Fecha",
            "Comprobante",
            "Cod. Cliente",
            "Razón Social",
            "Artículo Distribuidora",
            "Descripción Acción",
            "Clasificacion",
        ]
