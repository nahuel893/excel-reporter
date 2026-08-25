"""RED tests — S4.1: parallel diff harness (RF-12, Decision 1 + Decision 14).

Covers ``src/services/acciones_comerciales/diff.py``:

  - Like-for-like period-scope filtering: out-of-scope accumulated backup
    rows (dated outside the configured window) are EXCLUDED from the diff,
    never surfaced as missing/extra (RF-12 "Like-for-like period scope").
  - EXACT $0.01 comparison tolerance: a delta of a cent or less is treated
    as a match; anything greater is a real, surfaced delta. The tolerance is
    a comparison mechanism ONLY — underlying float values are never rounded
    (RF-23).
  - Every non-zero delta classified as EXACTLY one of ``baseline-defect``
    (attributable to a specific known manual-flow bug — stale BD:BE / BG:BH
    snapshots, the 6-row "es CCU?" map, tabla_control AZ/AX drift) or
    ``real-divergence`` (unexplained, needs investigation), with a reason
    string that names the bug (RF-12).
  - The RF-01 terna->precio (and ->Bonific) pick rule is validated
    EMPIRICALLY against the real ``aexcel.xlsx`` file: a pick mismatch is
    SURFACED, never silently accepted (Decision 14).
  - Machine-readable (JSON) + xlsx + human-readable summary outputs.

Strict-TDD: written before ``diff.py`` exists (import fails RED). The aexcel
fixture is built PROGRAMMATICALLY inline (openpyxl) into ``tmp_path`` rather
than committing a static binary — the established pattern in this suite
(S1.3/S1.5/S3 all build their xlsx/xls fixtures inline).
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.services.acciones_comerciales.constants import (
    ART_ACCION_ROW_FIELDS,
    CLIENTE_FECHA_ROW_FIELDS,
    GENERICOS_ORDER,
)
from src.services.acciones_comerciales.diff import (
    DEFAULT_TOLERANCE,
    KEY_FIELDS_BY_PIVOT,
    VALUE_COLUMNS_BY_PIVOT,
    DiffReport,
    KnownDefectContext,
    TernaPriceMismatch,
    diff_pivot,
    filter_period_scope,
    load_defect_context,
    load_pivots_from_workbook,
    read_aexcel_export,
    run_diff,
    run_diff_step,
    validate_terna_precios,
    write_diff_report,
)

# ─────────────────────────────────────────────────────────────────────────
# frame builders (mimic pivots.py output: flat frame + a TOTAL GENERAL row)
# ─────────────────────────────────────────────────────────────────────────

_CF_COLS = CLIENTE_FECHA_ROW_FIELDS + ["Suma de Descuento"]
_AA_COLS = ART_ACCION_ROW_FIELDS + ["Suma de Descuento"]
_ACC_GEN_KEYS = ["SUCURSAL", "Acción", "Descripción Acción", "mvb"]


def _with_total_general(df: pd.DataFrame, value_cols: list[str]) -> pd.DataFrame:
    """Append a pivots.py-style TOTAL GENERAL row (label in the first column,
    grand totals in the value columns) — diff_pivot must strip it before
    comparing so it never contaminates the key set."""
    total = {c: "" for c in df.columns}
    total[df.columns[0]] = "TOTAL GENERAL"
    for vc in value_cols:
        total[vc] = df[vc].sum()
    return pd.concat([df, pd.DataFrame([total], columns=df.columns)], ignore_index=True)


def _cf_row(
    *,
    fecha="2026-07-01",
    sucursal="CASA CENTRAL",
    cliente=100,
    razon="CLIENTE UNO",
    articulo=900,
    descripcion="ART UNO",
    calibre="CERVEZAS",
    accion="ACC1",
    descripcion_accion="MVB PROMO",
    descuento=90.0,
) -> dict:
    return {
        "Fecha": fecha,
        "SUCURSAL": sucursal,
        "Cod. Cliente": cliente,
        "Razón Social": razon,
        "Artículo Distribuidora": articulo,
        "Descripción": descripcion,
        "Calibre": calibre,
        "Acción": accion,
        "Descripción Acción": descripcion_accion,
        "Suma de Descuento": descuento,
    }


def _cf_frame(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=_CF_COLS)
    return _with_total_general(df, ["Suma de Descuento"])


def _aa_row(
    *,
    sucursal="CASA CENTRAL",
    articulo=900,
    descripcion="ART UNO",
    accion="ACC1",
    descripcion_accion="MVB PROMO",
    mvb="MVB",
    descuento=90.0,
) -> dict:
    return {
        "SUCURSAL": sucursal,
        "Artículo Distribuidora": articulo,
        "Descripción": descripcion,
        "Acción": accion,
        "Descripción Acción": descripcion_accion,
        "mvb": mvb,
        "Suma de Descuento": descuento,
    }


def _aa_frame(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=_AA_COLS)
    return _with_total_general(df, ["Suma de Descuento"])


def _acc_gen_row(
    *,
    sucursal="CASA CENTRAL",
    accion="ACC1",
    descripcion_accion="MVB PROMO",
    mvb="MVB",
    generico_values: dict | None = None,
) -> dict:
    row = {
        "SUCURSAL": sucursal,
        "Acción": accion,
        "Descripción Acción": descripcion_accion,
        "mvb": mvb,
        "(en blanco)": "",
    }
    for g in GENERICOS_ORDER:
        row[g] = (generico_values or {}).get(g, 0.0)
    return row


def _acc_gen_frame(rows: list[dict]) -> pd.DataFrame:
    cols = _ACC_GEN_KEYS + ["(en blanco)"] + GENERICOS_ORDER
    df = pd.DataFrame(rows, columns=cols)
    return _with_total_general(df, GENERICOS_ORDER)


# ─────────────────────────────────────────────────────────────────────────
# aexcel fixture builder (real xlsx, built inline)
# ─────────────────────────────────────────────────────────────────────────

_AEXCEL_HEADERS = ["Descripción Período", "Cod. Cliente", "Código", "Precio", "Bonific"]


def _write_aexcel_fixture(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "aexcel"
    for c, h in enumerate(_AEXCEL_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    for r_off, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=1 + r_off, column=c, value=v)
    wb.save(path)


def _generated_ternas(rows: list[dict]) -> pd.DataFrame:
    """A generated terna->precio frame the way gold_source produces it."""
    return pd.DataFrame(rows, columns=["Descripción Período", "Cod. Cliente", "Código", "Precio", "Bonific"])


# ─────────────────────────────────────────────────────────────────────────
# constants
# ─────────────────────────────────────────────────────────────────────────


class TestDiffConstants:
    def test_default_tolerance_is_one_cent(self):
        assert DEFAULT_TOLERANCE == 0.01

    def test_key_and_value_column_maps_cover_all_four_pivots(self):
        for pivot in ("FACT_NET", "ART-ACCION", "CLIENTE-FECHA", "ACC-GEN"):
            assert pivot in KEY_FIELDS_BY_PIVOT
            assert pivot in VALUE_COLUMNS_BY_PIVOT
        assert VALUE_COLUMNS_BY_PIVOT["ACC-GEN"] == GENERICOS_ORDER


# ─────────────────────────────────────────────────────────────────────────
# period-scope filtering (RF-12 like-for-like)
# ─────────────────────────────────────────────────────────────────────────


class TestPeriodScopeFilter:
    def test_filter_period_scope_excludes_out_of_scope_rows(self):
        df = pd.DataFrame(
            {"Fecha": ["2026-07-01", "2026-07-16", "2026-07-20"], "v": [1, 2, 3]}
        )
        out = filter_period_scope(df, "Fecha", "2026-07-01", "2026-07-16")
        assert list(out["Fecha"]) == ["2026-07-01", "2026-07-16"]

    def test_diff_pivot_excludes_accumulated_backup_rows(self):
        """BASE covers Jul 1-16; the backup engine table holds an accumulated
        Jul 20 row. Only the Jul 1-16 scope is compared — the Jul 20 backup
        row is excluded, NOT reported as a backup-only/extra delta."""
        base = _cf_frame([_cf_row(fecha="2026-07-01"), _cf_row(fecha="2026-07-16")])
        backup = _cf_frame(
            [
                _cf_row(fecha="2026-07-01"),
                _cf_row(fecha="2026-07-16"),
                _cf_row(fecha="2026-07-20", descuento=999.0),  # accumulated, out of scope
            ]
        )
        rows = diff_pivot(
            "CLIENTE-FECHA",
            base,
            backup,
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-16",
            context=KnownDefectContext(),
        )
        assert all("2026-07-20" not in tuple(r.key) for r in rows)
        assert rows == []  # everything in-scope matches


# ─────────────────────────────────────────────────────────────────────────
# exact $0.01 tolerance (RF-12 / RF-23)
# ─────────────────────────────────────────────────────────────────────────


class TestExactCentTolerance:
    def test_sub_cent_delta_is_not_flagged(self):
        base = _cf_frame([_cf_row(descuento=100.0)])
        backup = _cf_frame([_cf_row(descuento=100.005)])
        rows = diff_pivot(
            "CLIENTE-FECHA", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=KnownDefectContext(),
        )
        assert rows == []

    def test_over_cent_delta_is_flagged_and_value_not_rounded(self):
        base = _cf_frame([_cf_row(descuento=100.123456)])
        backup = _cf_frame([_cf_row(descuento=100.20)])
        rows = diff_pivot(
            "CLIENTE-FECHA", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=KnownDefectContext(),
        )
        assert len(rows) == 1
        assert rows[0].column == "Suma de Descuento"
        # underlying values preserved verbatim (no rounding/truncation, RF-23)
        assert rows[0].base_value == 100.123456
        assert rows[0].backup_value == 100.20


# ─────────────────────────────────────────────────────────────────────────
# classification: baseline-defect vs real-divergence (RF-12)
# ─────────────────────────────────────────────────────────────────────────


class TestClassification:
    def test_stale_bg_bh_sucursal_defect(self):
        """A client added after the manual BG:BH SUCURSAL snapshot froze
        (row 72,759) is missing from the backup -> baseline-defect."""
        base = _cf_frame([_cf_row(cliente=99999, descuento=90.0)])
        backup = _cf_frame([_cf_row(cliente=100, descuento=90.0)])
        ctx = KnownDefectContext(stale_sucursal_clients=frozenset({99999}))
        rows = diff_pivot(
            "CLIENTE-FECHA", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=ctx,
        )
        defect = [r for r in rows if 99999 in tuple(r.key)]
        assert defect and defect[0].classification == "baseline-defect"
        assert "BG:BH" in defect[0].reason

    def test_stale_bd_be_precio_defect(self):
        """A terna added after the manual BD:BE PRECIO snapshot froze shows a
        Descuento delta -> baseline-defect referencing BD:BE."""
        base = _cf_frame([_cf_row(cliente=88888, descuento=120.0)])
        backup = _cf_frame([_cf_row(cliente=88888, descuento=90.0)])
        ctx = KnownDefectContext(stale_precio_clients=frozenset({88888}))
        rows = diff_pivot(
            "CLIENTE-FECHA", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=ctx,
        )
        assert rows and rows[0].classification == "baseline-defect"
        assert "BD:BE" in rows[0].reason

    def test_six_row_es_ccu_map_defect(self):
        """A genérico the 6-row 'es CCU?' map failed to classify shows an
        ACC-GEN column delta -> baseline-defect referencing the es CCU map."""
        base = _acc_gen_frame([_acc_gen_row(generico_values={"PERNOD RICARD": 500.0})])
        backup = _acc_gen_frame([_acc_gen_row(generico_values={"PERNOD RICARD": 0.0})])
        ctx = KnownDefectContext(es_ccu_defect_generics=frozenset({"PERNOD RICARD"}))
        rows = diff_pivot(
            "ACC-GEN", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=ctx,
        )
        pr = [r for r in rows if r.column == "PERNOD RICARD"]
        assert pr and pr[0].classification == "baseline-defect"
        assert "es CCU" in pr[0].reason

    def test_az_ax_column_drift_defect(self):
        """A value column the backup's tabla_control summed from the drifted
        AZ/AX columns -> baseline-defect referencing AZ/AX."""
        base = _aa_frame([_aa_row(descuento=90.0)])
        backup = _aa_frame([_aa_row(descuento=45.0)])
        ctx = KnownDefectContext(az_ax_drift_columns=frozenset({"Suma de Descuento"}))
        rows = diff_pivot(
            "ART-ACCION", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=ctx,
        )
        assert rows and rows[0].classification == "baseline-defect"
        assert "AZ/AX" in rows[0].reason

    def test_unexplained_delta_is_real_divergence(self):
        """No known-bug context matches -> real-divergence, surfaced (never
        silently absorbed)."""
        base = _cf_frame([_cf_row(descuento=120.0)])
        backup = _cf_frame([_cf_row(descuento=90.0)])
        rows = diff_pivot(
            "CLIENTE-FECHA", base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31", context=KnownDefectContext(),
        )
        assert rows and rows[0].classification == "real-divergence"
        assert "real-divergence" in rows[0].reason


# ─────────────────────────────────────────────────────────────────────────
# terna->precio empirical validator vs the real aexcel (Decision 14)
# ─────────────────────────────────────────────────────────────────────────


class TestTernaPriceValidator:
    def test_read_aexcel_export_reads_terna_precio(self, tmp_path):
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 50.0, 0.1]])
        df = read_aexcel_export(path)
        assert list(df.columns) == ["Descripción Período", "Cod. Cliente", "Código", "Precio", "Bonific"]
        assert df.iloc[0]["Precio"] == 50.0

    def test_matching_pick_yields_no_mismatch(self, tmp_path):
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 50.0, 0.1]])
        aexcel = read_aexcel_export(path)
        generated = _generated_ternas([{"Descripción Período": "2026-07-01", "Cod. Cliente": 100, "Código": 900, "Precio": 50.0, "Bonific": 0.1}])
        assert validate_terna_precios(generated, aexcel) == []

    def test_precio_mismatch_is_surfaced_not_silently_accepted(self, tmp_path):
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 55.0, 0.1]])  # real price 55
        aexcel = read_aexcel_export(path)
        generated = _generated_ternas([{"Descripción Período": "2026-07-01", "Cod. Cliente": 100, "Código": 900, "Precio": 50.0, "Bonific": 0.1}])
        mismatches = validate_terna_precios(generated, aexcel)
        assert len(mismatches) == 1
        m = mismatches[0]
        assert isinstance(m, TernaPriceMismatch)
        assert m.kind == "precio"
        assert m.generated_precio == 50.0
        assert m.aexcel_precio == 55.0

    def test_bonific_mismatch_is_surfaced(self, tmp_path):
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 50.0, 0.25]])  # real bonif .25
        aexcel = read_aexcel_export(path)
        generated = _generated_ternas([{"Descripción Período": "2026-07-01", "Cod. Cliente": 100, "Código": 900, "Precio": 50.0, "Bonific": 0.1}])
        mismatches = validate_terna_precios(generated, aexcel)
        assert [m.kind for m in mismatches] == ["bonific"]

    def test_terna_missing_in_aexcel_is_surfaced(self, tmp_path):
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 50.0, 0.1]])
        aexcel = read_aexcel_export(path)
        generated = _generated_ternas([{"Descripción Período": "2026-07-01", "Cod. Cliente": 777, "Código": 901, "Precio": 60.0, "Bonific": 0.2}])
        mismatches = validate_terna_precios(generated, aexcel)
        assert [m.kind for m in mismatches] == ["missing-in-aexcel"]


# ─────────────────────────────────────────────────────────────────────────
# run_diff + report outputs (JSON + xlsx + human summary)
# ─────────────────────────────────────────────────────────────────────────


class TestRunDiffAndReport:
    def _base_and_backup(self):
        base = {
            "CLIENTE-FECHA": _cf_frame([_cf_row(cliente=99999, descuento=90.0), _cf_row(cliente=100, descuento=120.0)]),
            "ART-ACCION": _aa_frame([_aa_row(descuento=90.0)]),
        }
        backup = {
            "CLIENTE-FECHA": _cf_frame([_cf_row(cliente=100, descuento=90.0)]),
            "ART-ACCION": _aa_frame([_aa_row(descuento=90.0)]),
        }
        return base, backup

    def test_run_diff_classifies_and_collects_terna_mismatches(self, tmp_path):
        base, backup = self._base_and_backup()
        path = tmp_path / "aexcel.xlsx"
        _write_aexcel_fixture(path, [["2026-07-01", 100, 900, 55.0, 0.1]])
        aexcel = read_aexcel_export(path)
        generated = _generated_ternas([{"Descripción Período": "2026-07-01", "Cod. Cliente": 100, "Código": 900, "Precio": 50.0, "Bonific": 0.1}])

        report = run_diff(
            base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31",
            context=KnownDefectContext(stale_sucursal_clients=frozenset({99999})),
            generated_ternas=generated,
            aexcel=aexcel,
        )
        assert isinstance(report, DiffReport)
        # the 99999 base-only row is a baseline-defect; the client-100
        # descuento delta (120 vs 90) is a real-divergence.
        assert any(r.classification == "baseline-defect" for r in report.rows)
        assert any(r.classification == "real-divergence" for r in report.rows)
        assert len(report.terna_mismatches) == 1

    def test_write_diff_report_emits_json_xlsx_and_summary(self, tmp_path):
        base, backup = self._base_and_backup()
        report = run_diff(
            base, backup,
            fecha_desde="2026-07-01", fecha_hasta="2026-07-31",
            context=KnownDefectContext(stale_sucursal_clients=frozenset({99999})),
        )
        paths = write_diff_report(report, tmp_path, stem="diff_test")

        assert paths["json"].exists() and paths["xlsx"].exists() and paths["summary"].exists()

        payload = json.loads(paths["json"].read_text(encoding="utf-8"))
        assert "rows" in payload and "summary" in payload
        assert payload["summary"]["baseline_defect"] >= 1

        wb = load_workbook(paths["xlsx"])
        # every generated sheet must end in a distinctly-styled TOTAL GENERAL row
        for name in wb.sheetnames:
            ws = wb[name]
            col_a = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
            assert "TOTAL GENERAL" in col_a

        summary_text = paths["summary"].read_text(encoding="utf-8")
        assert "baseline-defect" in summary_text and "real-divergence" in summary_text


# ─────────────────────────────────────────────────────────────────────────
# backup loader + defect-context loader + optional service/CLI step (S4.3)
# ─────────────────────────────────────────────────────────────────────────


class TestBackupLoaderAndDiffStep:
    def _write_backup_workbook(self, path: Path) -> None:
        base_cf = _cf_frame([_cf_row(cliente=100, descuento=90.0)])
        base_aa = _aa_frame([_aa_row(descuento=90.0)])
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            base_cf.to_excel(xw, sheet_name="CLIENTE-FECHA", index=False)
            base_aa.to_excel(xw, sheet_name="ART-ACCION", index=False)

    def test_load_pivots_from_workbook(self, tmp_path):
        wb_path = tmp_path / "backup.xlsx"
        self._write_backup_workbook(wb_path)
        frames = load_pivots_from_workbook(wb_path)
        assert "CLIENTE-FECHA" in frames and "ART-ACCION" in frames
        assert "Suma de Descuento" in frames["CLIENTE-FECHA"].columns

    def test_load_defect_context_from_json(self, tmp_path):
        cfg = tmp_path / "known_defects.json"
        cfg.write_text(json.dumps({"stale_sucursal_clients": [99999], "az_ax_drift_columns": ["Suma de Descuento"]}), encoding="utf-8")
        ctx = load_defect_context(cfg)
        assert 99999 in ctx.stale_sucursal_clients
        assert "Suma de Descuento" in ctx.az_ax_drift_columns

    def test_load_defect_context_missing_file_is_empty(self, tmp_path):
        ctx = load_defect_context(tmp_path / "nope.json")
        assert ctx.stale_sucursal_clients == frozenset()

    def test_run_diff_step_writes_report_next_to_base_output(self, tmp_path):
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir()
        self._write_backup_workbook(backup_dir / "backup.xlsx")

        base_frames = {
            "CLIENTE-FECHA": _cf_frame([_cf_row(cliente=100, descuento=120.0)]),  # 120 vs backup 90 -> divergence
            "ART-ACCION": _aa_frame([_aa_row(descuento=90.0)]),
        }
        output_dir = tmp_path / "out"
        output_dir.mkdir()

        paths = run_diff_step(
            base_frames=base_frames,
            backup_dir=backup_dir,
            aexcel_path=None,
            generated_ternas=None,
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            output_dir=output_dir,
        )
        assert paths is not None
        assert paths["json"].parent == output_dir
        assert paths["json"].exists() and paths["xlsx"].exists()

    def test_run_diff_step_returns_none_when_no_backup_workbook(self, tmp_path):
        empty_dir = tmp_path / "backups"
        empty_dir.mkdir()
        out = tmp_path / "out"
        out.mkdir()
        paths = run_diff_step(
            base_frames={"ART-ACCION": _aa_frame([_aa_row()])},
            backup_dir=empty_dir,
            aexcel_path=None,
            generated_ternas=None,
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            output_dir=out,
        )
        assert paths is None


# ─────────────────────────────────────────────────────────────────────────
# RF-23 — no rounding/int-cast anywhere in diff.py
# ─────────────────────────────────────────────────────────────────────────


class TestNoRoundingInSource:
    def test_diff_source_has_no_rounding_or_int_cast(self):
        import re

        src = (
            Path(__file__).resolve().parents[2]
            / "src" / "services" / "acciones_comerciales" / "diff.py"
        ).read_text(encoding="utf-8")
        forbidden = ["round(", "astype(int", "astype('int", 'astype("int', "astype(np.int", "np.floor", "np.ceil", "np.trunc", "//"]
        for token in forbidden:
            assert token not in src, f"diff.py contains forbidden token {token!r}"
        assert re.compile(r"(?<![A-Za-z_.])int\s*\(").search(src) is None, "diff.py contains a bare int() call"
