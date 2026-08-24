"""RED tests — S1.3: wapi.xlsx ingestion, CCU variant only (RF-02).

Covers:
  - Only the ``Wapi`` sheet is read (Wapi_R2/Wapi_Branca ignored even if
    present), header at Excel row 8 (pandas header=7).
  - 21-column raw contract, in order.
  - Calibre IS the business genérico dimension — passed through verbatim,
    never renamed/reinterpreted.
  - Missing-file abort.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.services.acciones_comerciales.readers.wapi import (
    WAPI_HEADER_ROW,
    WAPI_RAW_COLUMNS,
    WAPI_SHEET_NAME,
    read_wapi,
)

# Banner rows above the real header — 7 rows (Excel rows 1-7), header at row 8.
_BANNER_ROWS = 7


def _make_wapi_fixture(tmp_path: Path, extra_sheets: bool = True) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = WAPI_SHEET_NAME

    for r in range(1, _BANNER_ROWS + 1):
        ws.cell(row=r, column=1, value=f"banner row {r}")

    header_row = _BANNER_ROWS + 1  # Excel row 8
    for c, h in enumerate(WAPI_RAW_COLUMNS, 1):
        ws.cell(row=header_row, column=c, value=h)

    data = [
        date(2026, 7, 1), "0001-00001234", "AGRUP-1", 12345, "CLIENTE UNO",
        "DIRECCION 1", "ART-001", "DESCRIPCION ART 1", "MARCA UNO", "CERVEZAS",
        10, 100.5, 1005.0, 1, 5.0, 50.25, 0.1, 954.75, "ACC-01",
        "MVB DESCUENTO", "DIST-001",
    ]
    for c, v in enumerate(data, 1):
        ws.cell(row=header_row + 1, column=c, value=v)

    if extra_sheets:
        ws_r2 = wb.create_sheet("Wapi_R2")
        ws_r2.cell(row=1, column=1, value="SHOULD NOT BE READ (FEDESUR)")
        ws_branca = wb.create_sheet("Wapi_Branca")
        ws_branca.cell(row=1, column=1, value="SHOULD NOT BE READ (BRANCA)")

    path = tmp_path / "wapi.xlsx"
    wb.save(path)
    return path


class TestReadWapi:
    def test_reads_only_wapi_sheet_ignoring_r2_and_branca(self, tmp_path):
        path = _make_wapi_fixture(tmp_path)

        df = read_wapi(path)

        assert list(df.columns) == WAPI_RAW_COLUMNS
        assert len(df) == 1

    def test_header_row_is_excel_row_8(self):
        assert WAPI_HEADER_ROW == 7  # pandas 0-indexed header -> Excel row 8

    def test_21_column_raw_contract(self):
        assert len(WAPI_RAW_COLUMNS) == 21
        assert WAPI_RAW_COLUMNS[0] == "Fecha"
        assert WAPI_RAW_COLUMNS[-1] == "Artículo Distribuidora"
        assert "Calibre" in WAPI_RAW_COLUMNS
        assert "PRECIO FINAL " not in WAPI_RAW_COLUMNS  # derived col — not part of the raw wapi contract

    def test_calibre_treated_as_generico_passthrough(self, tmp_path):
        """Calibre is the business genérico dimension — read verbatim,
        never remapped/renamed by the reader."""
        path = _make_wapi_fixture(tmp_path)

        df = read_wapi(path)

        assert df.iloc[0]["Calibre"] == "CERVEZAS"

    def test_data_row_values_survive(self, tmp_path):
        path = _make_wapi_fixture(tmp_path)

        df = read_wapi(path)

        row = df.iloc[0]
        assert row["Cod. Cliente"] == 12345
        assert row["Razón Social"] == "CLIENTE UNO"
        assert row["Acción"] == "ACC-01"
        assert row["Descripción Acción"] == "MVB DESCUENTO"
        assert row["Artículo Distribuidora"] == "DIST-001"

    def test_missing_file_aborts(self, tmp_path):
        missing = tmp_path / "does_not_exist_wapi.xlsx"

        with pytest.raises(FileNotFoundError):
            read_wapi(missing)

    def test_ignores_r2_and_branca_even_without_being_asked(self, tmp_path):
        """Sanity: sibling FEDESUR/BRANCA sheets never leak into the result,
        even indirectly (e.g. via a stray column)."""
        path = _make_wapi_fixture(tmp_path, extra_sheets=True)

        df = read_wapi(path)

        for col in df.columns:
            assert "SHOULD NOT BE READ" not in str(df[col].iloc[0] if len(df) else "")
