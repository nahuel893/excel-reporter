"""RED tests — S1.7/S3.3-S3.5: Phase-1 CLI/service end-to-end (RF-22).

Covers:
  - AccionesComercialesConfig defaults (escribir_informe=False — RF-13).
  - AccionesComercialesService's FULL Phase-1 orchestration (S3): gold
    aexcel-equivalent -> wapi ingestion -> derived-column enrichment -> the
    4 pivots -> the 6-sheet BASE control workbook. Writes ONLY under
    data/output/acciones-comerciales/{YYYY-MM}/ (service_output_dir
    convention) and never touches any external file (informe_path) when
    escribir_informe is False (default).
  - main.py wiring: REPORT_HANDLERS["acciones-comerciales"], ReportConfig
    accepts the new "tipo", merge_filters carries the new custom filtros
    through, and the full `python main.py --config <file>` path (global
    --config, RF-22's literal scenario) dispatches end-to-end with zero
    external-file writes.
  - S3.5: the full CLI run against a FAKE DataLoader (fixture aexcel/
    sucursal data, no real gold) + a real fixture wapi.xlsx, producing a
    real xlsx in a tmp dir — asserts 6-sheet structure, TOTAL GENERAL
    placement, and that PRECIO FINAL actually resolves through the terna
    wiring (the S3 "CRITICAL WIRING GOTCHA": Fecha/Descripción Período
    dtype normalization).

Gold/DataLoader access is ALWAYS faked via ``monkeypatch.setattr(DataLoader,
...)`` at the class level — no test in this module ever touches a real
database connection (S1.10/RF-25 CI discipline: zero ``@integration``-marked
tests, zero live DB dependency).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.core.data_loader import DataLoader
from src.services.acciones_comerciales.config import AccionesComercialesConfig
from src.services.acciones_comerciales.readers.wapi import WAPI_RAW_COLUMNS, WAPI_SHEET_NAME
from src.services.acciones_comerciales.service import (
    AccionesComercialesResult,
    AccionesComercialesService,
)

_BANNER_ROWS = 7  # wapi.xlsx: 7 banner rows above the real header (Excel row 8)


def _write_backup_pivot_workbook(path: Path) -> None:
    """A minimal manual-backup workbook shaped like the BASE control (one
    ART-ACCION pivot sheet, header in row 1) — enough for the optional diff
    step (S4.3) to load and compare against the just-built BASE frames."""
    from src.services.acciones_comerciales.constants import ART_ACCION_ROW_FIELDS

    cols = ART_ACCION_ROW_FIELDS + ["Suma de Descuento"]
    df = pd.DataFrame(
        [
            {
                "SUCURSAL": "CASA CENTRAL",
                "Artículo Distribuidora": 900,
                "Descripción": "ART UNO",
                "Acción": "ACC1",
                "Descripción Acción": "MVB PROMO",
                "mvb": "MVB",
                "Suma de Descuento": 45.0,  # diverges from BASE -> a surfaced delta
            }
        ],
        columns=cols,
    )
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="ART-ACCION", index=False)


# ─────────────────────────────────────────────────────────────────────────
# shared fixture builders — fake gold + real wapi.xlsx (no live DB, ever)
# ─────────────────────────────────────────────────────────────────────────


def _write_wapi_fixture(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = WAPI_SHEET_NAME
    for r in range(1, _BANNER_ROWS + 1):
        ws.cell(row=r, column=1, value=f"banner row {r}")
    header_row = _BANNER_ROWS + 1
    for c, h in enumerate(WAPI_RAW_COLUMNS, 1):
        ws.cell(row=header_row, column=c, value=h)
    for r_offset, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=header_row + r_offset, column=c, value=v)
    wb.save(path)


def _wapi_row(
    fecha="2026-07-01",
    cod_cliente=100,
    razon="CLIENTE UNO",
    calibre="CERVEZAS",
    cantidad=10.0,
    cantidad_sin_cargo=0.0,
    descuento_pct=10.0,
    accion="ACC1",
    descripcion_accion="MVB PROMO",
    articulo_distribuidora=900,
) -> list:
    return [
        fecha, "FC-1", "", cod_cliente, razon, "", "CMQ1", "ART UNO",
        "MARCA UNO", calibre, cantidad, 45.0, 450.0, cantidad_sin_cargo,
        descuento_pct, 0.0, 0.0, 0.0, accion, descripcion_accion,
        articulo_distribuidora,
    ]


def _aexcel_line(
    fecha="2026-07-01",
    cliente=100,
    articulo=900,
    precio=50.0,
    bonific=0.1,
    cantidad=10.0,
    facturacion=450.0,
    descuentos=45.0,
    id_linea=1,
    sucursal="1 - CASA CENTRAL",
) -> dict:
    return {
        "_id_linea": id_linea,
        "Descripción Período": fecha,
        "Cod. Cliente": cliente,
        "Descripción": "CLIENTE UNO",
        "Sucursal": sucursal,
        "Código": articulo,
        "Descripción_2": "ART UNO",
        "Descripción_3": "MARCA UNO",
        "Descripción_12": "CERVEZAS",
        "Precio": precio,
        "Bonific": bonific,
        "Cantidades Totales": cantidad,
        "Facturacion Neta": facturacion,
        "Descuentos": descuentos,
    }


def _patch_gold(
    monkeypatch,
    aexcel_lines: list[dict],
    sucursal_rows: list[tuple],
    comprobante_lines: list[dict] | None = None,
) -> None:
    """Patch DataLoader.get_aexcel_equivalent/get_clientes_sucursal/
    get_comprobante_precio at the CLASS level — the service's default
    (uninjected) DataLoader never touches a real engine/connection,
    satisfying the zero-live-DB rule.

    ``comprobante_lines`` (Decision 19, optional) — fake rows for the
    comprobante-keyed diagnostic price lookup. Defaults to empty (no
    diagnostic matches), which is a no-op for every pre-existing test."""
    aexcel_df = pd.DataFrame(aexcel_lines)
    sucursal_df = pd.DataFrame(sucursal_rows, columns=["Cod. Cliente", "Sucursal"])
    comprobante_df = pd.DataFrame(
        comprobante_lines or [],
        columns=["Comprobante", "Código", "Precio", "Cantidades Totales"],
    )

    def _fake_aexcel(self, fecha_desde, fecha_hasta):
        return aexcel_df.copy()

    def _fake_sucursal(self):
        return sucursal_df.copy()

    def _fake_comprobante(self, fecha_desde, fecha_hasta):
        return comprobante_df.copy()

    monkeypatch.setattr(DataLoader, "get_aexcel_equivalent", _fake_aexcel)
    monkeypatch.setattr(DataLoader, "get_clientes_sucursal", _fake_sucursal)
    monkeypatch.setattr(DataLoader, "get_comprobante_precio", _fake_comprobante)


def _default_gold_and_wapi(tmp_path, monkeypatch) -> Path:
    """Wire the common happy-path fixture: one aexcel terna, one matching
    wapi row (same fecha/cliente/articulo terna, so PRECIO FINAL resolves),
    SUCURSAL fresh lookup mapping client 100 -> bare 'CASA CENTRAL' (RF-04
    format, matching configs/acciones_comerciales_zonas.json)."""
    _patch_gold(
        monkeypatch,
        aexcel_lines=[_aexcel_line()],
        sucursal_rows=[(100, "CASA CENTRAL")],
    )
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    _write_wapi_fixture(input_dir / "wapi.xlsx", [_wapi_row()])
    return input_dir


# ─────────────────────────────────────────────────────────────────────────
# config.py — AccionesComercialesConfig
# ─────────────────────────────────────────────────────────────────────────


class TestAccionesComercialesConfig:
    def test_escribir_informe_defaults_false(self, tmp_path):
        config = AccionesComercialesConfig(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            input_dir=str(tmp_path),
        )

        assert config.escribir_informe is False

    def test_wapi_and_compras_path_properties(self, tmp_path):
        config = AccionesComercialesConfig(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            input_dir=str(tmp_path),
        )

        assert config.wapi_path == tmp_path / "wapi.xlsx"
        assert config.compras_path == tmp_path / "compras.xls"


# ─────────────────────────────────────────────────────────────────────────
# service.py — AccionesComercialesService full Phase-1 orchestration (S3)
# ─────────────────────────────────────────────────────────────────────────


class TestAccionesComercialesServiceFullPipeline:
    def test_service_slug_and_granularity(self):
        assert AccionesComercialesService.SERVICE_SLUG == "acciones-comerciales"
        assert AccionesComercialesService.GRANULARITY == "month"

    def test_writes_base_control_under_conventional_output_dir(self, tmp_path, monkeypatch):
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        expected_dir = tmp_path / "out" / "acciones-comerciales" / "2026-07"
        assert result.ruta_archivo.parent == expected_dir
        assert result.ruta_archivo.exists()
        assert isinstance(result, AccionesComercialesResult)

    def test_zero_external_file_writes_when_escribir_informe_false(self, tmp_path, monkeypatch):
        """RF-13/RF-22: default run (flag OFF) touches nothing outside
        data/output — the informe_path is never created/modified."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)
        informe_path = tmp_path / "INFO - ACCIONES BADIE JULIO 2026.xlsm"

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
                escribir_informe=False,
                informe_path=str(informe_path),
            )
            service.generar_reporte(config)

        assert not informe_path.exists()
        # source input files remain byte-identical (RF-24) — untouched here.
        assert (input_dir / "wapi.xlsx").exists()

    def test_output_has_six_sheets_each_ending_in_total_general(self, tmp_path, monkeypatch):
        """RF-10: 6 sheets, every sheet ends with a TOTAL GENERAL row."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert wb.sheetnames == [
            "FACT_NET",
            "ART-ACCION",
            "CLIENTE-FECHA",
            "ACC-GEN",
            "wapi",
            "Reconciliacion",
        ]
        for name in wb.sheetnames:
            ws = wb[name]
            values_col_a = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
            assert "TOTAL GENERAL" in values_col_a

    def test_missing_wapi_file_aborts_with_no_partial_output(self, tmp_path, monkeypatch):
        """RF-02 scenario: missing wapi.xlsx aborts before any write."""
        _patch_gold(monkeypatch, aexcel_lines=[_aexcel_line()], sucursal_rows=[(100, "CASA CENTRAL")])
        input_dir = tmp_path / "input"
        input_dir.mkdir()  # wapi.xlsx deliberately NOT created

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            with pytest.raises(FileNotFoundError):
                service.generar_reporte(config)

        # no partial xlsx output — the output dir may exist as scaffolding
        # (mkdir is safe/idempotent) but must contain no written file.
        out_dir = tmp_path / "out"
        assert not list(out_dir.rglob("*.xlsx"))


# ─────────────────────────────────────────────────────────────────────────
# S3.5 — full pipeline wiring: PRECIO FINAL actually resolves through the
# terna dtype-normalization gotcha
# ─────────────────────────────────────────────────────────────────────────


class TestFullPipelineWiring:
    def test_precio_final_resolves_via_terna_match_across_dtype_forms(self, tmp_path, monkeypatch):
        """The S3 CRITICAL WIRING GOTCHA: aexcel's 'Descripción Período' and
        wapi's 'Fecha' must be normalized to the SAME form before the terna
        lookup, or every row lands in unresolved_precio. This fixture uses
        a python ``date`` object on the wapi side and a plain ISO string on
        the aexcel side to prove the normalization actually happens."""
        from datetime import date

        _patch_gold(
            monkeypatch,
            aexcel_lines=[_aexcel_line(fecha="2026-07-01", cliente=100, articulo=900, precio=73.33)],
            sucursal_rows=[(100, "CASA CENTRAL")],
        )
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_wapi_fixture(
            input_dir / "wapi.xlsx",
            [_wapi_row(fecha=date(2026, 7, 1), cod_cliente=100, articulo_distribuidora=900)],
        )

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        ws = wb["wapi"]
        headers = [c.value for c in ws[1]]
        precio_col = headers.index("PRECIO FINAL ") + 1
        # row 2 = the single data row (row 1 = headers)
        assert ws.cell(2, precio_col).value == 73.33

    def test_reconciliation_sheet_reflects_wired_totals(self, tmp_path, monkeypatch):
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        ws = wb["Reconciliacion"]
        col_a_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert "CASA CENTRAL" in col_a_values  # aexcel prefix stripped to match wapi's bare SUCURSAL
        assert "TOTAL GENERAL" in col_a_values
        # the sheet ends with the reconciliation's own TOTAL GENERAL row.
        assert ws.cell(ws.max_row, 1).value == "TOTAL GENERAL"


# ─────────────────────────────────────────────────────────────────────────
# Decision 19 — comprobante-based diagnostic price, wired end-to-end
# through the service (BASE-control ONLY parallel-run comparison).
# ─────────────────────────────────────────────────────────────────────────


class TestPrecioComprobanteDiagnosticWiring:
    def test_diagnostic_column_populates_and_terna_precio_final_stays_authoritative(
        self, tmp_path, monkeypatch
    ):
        """The comprobante lookup feeds a TRAILING diagnostic column without
        disturbing the terna-based PRECIO FINAL (the wapi row's Comprobante
        is "FC-1"/Artículo Distribuidora 900, matching the fake comprobante
        line; its diagnostic price DIFFERS from the terna price, exercising
        the COMPARACION PRECIO section too)."""
        _patch_gold(
            monkeypatch,
            aexcel_lines=[_aexcel_line(precio=73.33)],
            sucursal_rows=[(100, "CASA CENTRAL")],
            comprobante_lines=[
                {"Comprobante": "FC-1", "Código": 900, "Precio": 80.0, "Cantidades Totales": 10.0}
            ],
        )
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_wapi_fixture(input_dir / "wapi.xlsx", [_wapi_row(articulo_distribuidora=900)])

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        ws = wb["wapi"]
        headers = [c.value for c in ws[1]]
        precio_terna_col = headers.index("PRECIO FINAL ") + 1
        precio_comp_col = headers.index("PRECIO FINAL (comprobante)") + 1
        assert ws.cell(2, precio_terna_col).value == 73.33  # terna value UNCHANGED
        assert ws.cell(2, precio_comp_col).value == 80.0

        recon_ws = wb["Reconciliacion"]
        title_row = None
        for r in range(1, recon_ws.max_row + 1):
            value = recon_ws.cell(r, 1).value
            if isinstance(value, str) and value.startswith("COMPARACION PRECIO"):
                title_row = r
                break
        assert title_row is not None
        assert "1 filas difieren de 1 con ambos precios" in recon_ws.cell(title_row, 1).value
        # the sheet still ends with the FACTURACION/DESCUENTOS TOTAL GENERAL.
        assert recon_ws.cell(recon_ws.max_row, 1).value == "TOTAL GENERAL"

    def test_empty_comprobante_data_leaves_diagnostic_column_all_blank(self, tmp_path, monkeypatch):
        """Default fixture (no comprobante_lines) — the diagnostic column
        must still exist (all-NaN) and never crash the pipeline."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        ws = wb["wapi"]
        headers = [c.value for c in ws[1]]
        assert "PRECIO FINAL (comprobante)" in headers
        precio_comp_col = headers.index("PRECIO FINAL (comprobante)") + 1
        assert ws.cell(2, precio_comp_col).value in (None, "")


# ─────────────────────────────────────────────────────────────────────────
# main.py wiring — REPORT_HANDLERS, config models, merge_filters, full CLI
# ─────────────────────────────────────────────────────────────────────────


class TestMainWiring:
    def test_report_handlers_registers_acciones_comerciales(self):
        import main as main_module

        assert (
            main_module.REPORT_HANDLERS.get("acciones-comerciales")
            == "_run_acciones_comerciales_report"
        )

    def test_report_config_accepts_acciones_comerciales_tipo(self):
        from src.config.models import ReportConfig

        raw = {
            "tipo": "acciones-comerciales",
            "filtros": {
                "fecha_desde": "2026-07-01",
                "fecha_hasta": "2026-07-31",
                "input_dir": "/tmp/does-not-matter",
            },
            "reportes": [{"nombre": "BASE control TEST"}],
        }
        cfg = ReportConfig.model_validate(raw)

        assert cfg.tipo == "acciones-comerciales"
        assert cfg.filtros.input_dir == "/tmp/does-not-matter"
        assert cfg.filtros.escribir_informe is False

    def test_merge_filters_carries_new_fields(self):
        from src.config.models import GlobalFilters
        from src.config.resolver import merge_filters

        global_f = GlobalFilters(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            input_dir="/tmp/acciones-input",
            escribir_informe=True,
            informe_path="/tmp/informe.xlsm",
            esperar_wapi_fresco=True,
            wapi_cobertura_requerida="habil_anterior",
        )

        merged = merge_filters(global_f, None)

        assert merged["input_dir"] == "/tmp/acciones-input"
        assert merged["escribir_informe"] is True
        assert merged["informe_path"] == "/tmp/informe.xlsm"
        assert merged["esperar_wapi_fresco"] is True
        assert merged["wapi_cobertura_requerida"] == "habil_anterior"

    def test_merge_filters_carries_backup_dir_and_aexcel_path(self):
        """S4.3: the optional parallel-diff step reads a config-driven backup
        dir + real aexcel path — both must survive the global->merged merge."""
        from src.config.models import GlobalFilters
        from src.config.resolver import merge_filters

        global_f = GlobalFilters(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            input_dir="/tmp/acciones-input",
            backup_dir="/tmp/backups/acciones-comerciales-2026-07-16",
            aexcel_path="/tmp/acciones-input/aexcel.xlsx",
        )

        merged = merge_filters(global_f, None)

        assert merged["backup_dir"] == "/tmp/backups/acciones-comerciales-2026-07-16"
        assert merged["aexcel_path"] == "/tmp/acciones-input/aexcel.xlsx"

    def test_run_acciones_comerciales_report_returns_path_and_meta(self, tmp_path, monkeypatch):
        import main as main_module

        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        report = type("Report", (), {"nombre": "BASE control TEST"})()
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "input_dir": str(input_dir),
            "escribir_informe": False,
            "informe_path": None,
            "esperar_wapi_fresco": False,
            "wapi_cobertura_requerida": None,
        }

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            artifacts = main_module._run_acciones_comerciales_report(report, merged)

        assert len(artifacts) == 1
        path, meta = artifacts[0]
        assert isinstance(path, Path)
        assert path.exists()
        assert meta["nombre"] == "BASE control TEST"

    def test_full_cli_global_config_flag_dispatches_end_to_end(self, tmp_path, monkeypatch):
        """RF-22 literal scenario: `python main.py --config <file>` with
        Phase-2 flag OFF writes BASE control under data/output and touches
        nothing else."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)
        informe_path = tmp_path / "INFO - ACCIONES BADIE JULIO 2026.xlsm"

        config_path = tmp_path / "acciones_comerciales.json"
        config_path.write_text(
            json.dumps(
                {
                    "tipo": "acciones-comerciales",
                    "filtros": {
                        "fecha_desde": "2026-07-01",
                        "fecha_hasta": "2026-07-31",
                        "input_dir": str(input_dir),
                        "escribir_informe": False,
                        "informe_path": str(informe_path),
                        "enviar_email": False,
                        "enviar_whatsapp": False,
                    },
                    "reportes": [{"nombre": "BASE control TEST"}],
                }
            ),
            encoding="utf-8",
        )

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            monkeypatch.setattr(sys, "argv", ["main.py", "--config", str(config_path)])
            import main

            result = main.main()

        assert result == 0
        expected_dir = tmp_path / "out" / "acciones-comerciales" / "2026-07"
        assert expected_dir.is_dir()
        assert any(expected_dir.glob("*.xlsx"))
        assert not informe_path.exists()


# ─────────────────────────────────────────────────────────────────────────
# S4.3 — optional parallel-diff step wired behind config.backup_dir
# ─────────────────────────────────────────────────────────────────────────


class TestDiffStepWiring:
    def test_backup_dir_defaults_none_and_no_diff_report(self, tmp_path, monkeypatch):
        """Default run (no backup_dir) never produces a diff report — the
        diff step is strictly opt-in."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
            )
            result = service.generar_reporte(config)

        assert config.backup_dir is None
        assert result.diff_report_paths is None
        out_dir = result.ruta_archivo.parent
        assert not list(out_dir.glob("diff_acciones_comerciales.*"))

    def test_service_writes_diff_report_next_to_base_when_backup_dir_set(self, tmp_path, monkeypatch):
        """S4.3/RF-12: with backup_dir pointing at a backup workbook, the
        service runs the diff and writes the JSON+xlsx report next to the
        BASE output."""
        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)
        backup_dir = tmp_path / "backups" / "acciones-comerciales-2026-07-16"
        backup_dir.mkdir(parents=True)
        _write_backup_pivot_workbook(backup_dir / "backup.xlsx")

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AccionesComercialesService()
            config = AccionesComercialesConfig(
                fecha_desde="2026-07-01",
                fecha_hasta="2026-07-31",
                input_dir=str(input_dir),
                nombre_archivo="BASE control TEST",
                backup_dir=str(backup_dir),
            )
            result = service.generar_reporte(config)

        out_dir = result.ruta_archivo.parent
        assert result.diff_report_paths is not None
        assert (out_dir / "diff_acciones_comerciales.json").exists()
        assert (out_dir / "diff_acciones_comerciales.xlsx").exists()
        assert (out_dir / "diff_acciones_comerciales.txt").exists()

    def test_run_report_handler_passes_backup_dir_through(self, tmp_path, monkeypatch):
        """main._run_acciones_comerciales_report forwards backup_dir/aexcel_path
        from the merged filtros into the service config."""
        import main as main_module

        input_dir = _default_gold_and_wapi(tmp_path, monkeypatch)
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir()
        _write_backup_pivot_workbook(backup_dir / "backup.xlsx")

        report = type("Report", (), {"nombre": "BASE control TEST"})()
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "input_dir": str(input_dir),
            "backup_dir": str(backup_dir),
            "aexcel_path": None,
        }

        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            artifacts = main_module._run_acciones_comerciales_report(report, merged)

        path, _meta = artifacts[0]
        assert (path.parent / "diff_acciones_comerciales.json").exists()
