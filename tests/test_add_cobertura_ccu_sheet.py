"""Tests for the 'Cobertura CCU' sheet builder (add_cobertura_ccu_sheet.py).

Regression focus: the previous 'Cober Nueva' sheet hardcoded '1 - CASA CENTRAL'
filters and had broken cross-block references. This sheet is a clean, from-
scratch replacement scoped to GUEMES ('16 - SUCURSAL GUEMES'), so every test
below double-checks the sucursal literal and the absence of any leftover
CASA CENTRAL / Cober Nueva reference.

Also covers three post-review fixes: the generico OBJ rollup must include
marcas that are not rendered as columns (OBJ_ONLY_MARCAS), a third
preventista (DIRECTA), a blank instead of a fake 0% when OBJ is 0, a backup
filename recognized by the avances-service backup guard, validate-before-
backup ordering in `main()`, and golden-formula assertions strong enough to
catch realistic mutations (wrong column, dropped rollup half, etc).
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import scripts.add_cobertura_ccu_sheet as ccu
from src.services.avances.service import _is_backup_name

SUC = "16 - SUCURSAL GUEMES"
SHEET = "Cobertura CCU"

CONFIG_PATH = Path(__file__).resolve().parents[1] / "configs" / "avances_guemes.json"


def _make_workbook():
    """Minimal GUEMES-like workbook with stub COBER / CuposCober sheets.

    CuposCober carries one real data row scoped to GUEMES so that `main()`'s
    sucursal-presence validation passes by default; tests that need to
    exercise the abort path build their own fixture instead.
    """
    wb = Workbook()
    wb.active.title = "Placeholder"

    cober = wb.create_sheet("COBER")
    cober.append(["Sucursal", "Descripcion Vendedor", "Ruta", "Descripcion_Marca", "Numero_Clientes"])

    cupos = wb.create_sheet("CuposCober")
    cupos.append([None, "Ruta", "Descripción Vendedor", "MARCA", "ZONA", None, None, "CUPO "])
    cupos.append([None, 1, "JORGE RAMOS", "SALTA", SUC, None, None, 10])

    return wb


def _make_workbook_missing_cuposcober():
    wb = Workbook()
    wb.active.title = "Placeholder"
    cober = wb.create_sheet("COBER")
    cober.append(["Sucursal", "Descripcion Vendedor", "Ruta", "Descripcion_Marca", "Numero_Clientes"])
    return wb


def _make_workbook_without_guemes_in_cupos():
    """COBER + CuposCober both present, but column E of CuposCober never
    mentions GUEMES — simulates running the script against the wrong
    workbook (e.g. AVANCE BADIE)."""
    wb = Workbook()
    wb.active.title = "Placeholder"
    cober = wb.create_sheet("COBER")
    cober.append(["Sucursal", "Descripcion Vendedor", "Ruta", "Descripcion_Marca", "Numero_Clientes"])
    cupos = wb.create_sheet("CuposCober")
    cupos.append([None, "Ruta", "Descripción Vendedor", "MARCA", "ZONA", None, None, "CUPO "])
    cupos.append([None, 1, "SOME VENDEDOR", "SALTA", "1 - CASA CENTRAL", None, None, 10])
    return wb


def _make_target_file(tmp_path, builder=_make_workbook):
    path = tmp_path / "AVANCE GUEMES.xlsx"
    builder().save(path)
    return path


def _all_formula_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                yield cell


def _data_query_formula_cells(ws):
    """PDV/OBJ cells that query COBER/CuposCober directly (SUMIFS).

    Excludes the derived Faltan (subtraction) and % (division) cells, which
    reference other cells in the same row rather than querying sucursal data
    directly — they have no sucursal literal to scope.
    """
    for cell in _all_formula_cells(ws):
        if "SUMIFS(" in cell.value:
            yield cell


def _extract_array_literal(formula: str) -> list[str]:
    match = re.search(r"\{([^}]*)\}", formula)
    assert match, f"no array literal found in {formula!r}"
    return [item.strip('"') for item in match.group(1).split(";")]


def test_sheet_created_at_index_0():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)

    assert SHEET in wb.sheetnames
    assert wb.sheetnames.index(SHEET) == 0


def test_rebuild_is_idempotent():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws_first = wb[SHEET]
    first_dims = (ws_first.max_row, ws_first.max_column)
    first_values = [[c.value for c in row] for row in ws_first.iter_rows()]

    # Sheet count before/after must stay 1 duplicate call.
    ccu.build_cobertura_ccu(wb)

    assert wb.sheetnames.count(SHEET) == 1
    ws_second = wb[SHEET]
    second_dims = (ws_second.max_row, ws_second.max_column)
    second_values = [[c.value for c in row] for row in ws_second.iter_rows()]

    assert first_dims == second_dims
    assert first_values == second_values


def test_all_data_query_formulas_scope_to_guemes_never_casa_central():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    formula_cells = list(_data_query_formula_cells(ws))
    assert formula_cells, "expected at least one SUMIFS formula cell"

    for cell in formula_cells:
        assert SUC in cell.value, f"{cell.coordinate} missing sucursal scope: {cell.value!r}"
        assert "CASA CENTRAL" not in cell.value, f"{cell.coordinate} leaked CASA CENTRAL: {cell.value!r}"


def test_no_formula_anywhere_references_casa_central():
    """Even derived (Faltan/%) cells must never leak a CASA CENTRAL literal."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    for cell in _all_formula_cells(ws):
        assert "CASA CENTRAL" not in cell.value, f"{cell.coordinate}: {cell.value!r}"


def test_no_formula_references_cober_nueva():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert "Cober Nueva" not in cell.value, (
                    f"{cell.coordinate} references 'Cober Nueva': {cell.value!r}"
                )


def test_generico_total_pdv_uses_generico_grain_not_sum_of_marcas():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    generico_total_pdv_formulas = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("=SUMIFS(COBER!$L:$L"):
                generico_total_pdv_formulas.append(value)

    assert generico_total_pdv_formulas, "expected at least one generico-total PDV formula"
    for formula in generico_total_pdv_formulas:
        # Golden-string check (not a substring trick): must be a SUMIFS read
        # of the generico grain, never a SUM/rollup of the marca cells.
        assert formula.startswith("=SUMIFS(COBER!$L:$L,COBER!$H:$H,"), formula
        assert "COBER!$K:$K" in formula


def test_total_guemes_row_omits_preventista_criteria():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    total_rows = []
    for row_cells in ws.iter_rows():
        if row_cells[0].value == "TOTAL GUEMES":
            total_rows.append(row_cells[0].row)

    assert len(total_rows) == len(ccu.SECTIONS)

    for row_number in total_rows:
        for cell in ws[row_number]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                assert "COBER!$B:$B" not in cell.value, cell.value
                assert "CuposCober!$C:$C" not in cell.value, cell.value


def test_section_and_row_structure():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    preventista_names = [
        row[0].value
        for row in ws.iter_rows()
        if row[0].value in ccu.PREVENTISTAS
    ]
    # Each of the 5 sections has exactly len(PREVENTISTAS) data rows.
    assert len(preventista_names) == len(ccu.SECTIONS) * len(ccu.PREVENTISTAS)

    total_row_count = sum(
        1 for row in ws.iter_rows() if row[0].value == "TOTAL GUEMES"
    )
    assert total_row_count == len(ccu.SECTIONS)


def test_full_sport_is_a_displayed_aguas_danone_column():
    """FULL SPORT is a real AGUAS DANONE marca (gold.dim_articulo) and must be
    shown as its own column, not hidden as an OBJ-only rollup contributor."""
    aguas = next(
        marcas for _title, generico, marcas, _ in ccu.SECTIONS
        if generico == "AGUAS DANONE"
    )
    assert "FULL SPORT" in aguas
    assert "FULL SPORT" not in ccu.OBJ_ONLY_MARCAS.get("AGUAS DANONE", [])

    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]
    marca_headers = {cell.value for row in ws.iter_rows() for cell in row}
    assert "FULL SPORT" in marca_headers


def test_cervezas_emits_exactly_one_total_column_group():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    total_cervezas_headers = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "TOTAL CERVEZAS"
    ]
    assert len(total_cervezas_headers) == 1


def test_cli_main_missing_file_returns_error(tmp_path):
    missing = tmp_path / "does-not-exist.xlsx"
    assert ccu.main(str(missing)) == 1
    assert not missing.exists()
    # No stray backup/output file must appear in the tmp dir either.
    assert list(tmp_path.iterdir()) == []


def test_cli_main_creates_backup_before_writing(tmp_path):
    path = _make_target_file(tmp_path)

    assert ccu.main(str(path)) == 0

    backups = list(tmp_path.glob("AVANCE GUEMES.backup-*.xlsx"))
    assert len(backups) == 1

    out = load_workbook(path)
    assert SHEET in out.sheetnames


def test_cli_main_is_idempotent(tmp_path):
    path = _make_target_file(tmp_path)

    assert ccu.main(str(path)) == 0
    assert ccu.main(str(path)) == 0

    out = load_workbook(path)
    assert out.sheetnames.count(SHEET) == 1


# --- F2: DIRECTA is a third preventista -------------------------------------


def test_preventistas_includes_directa():
    assert ccu.PREVENTISTAS == ["JORGE RAMOS", "TALLO GABRIELA", "DIRECTA"]


# --- F5: backup filename must be caught by the service's backup guard ------


def test_backup_filename_is_recognized_by_avances_service_backup_guard(tmp_path):
    path = _make_target_file(tmp_path)

    assert ccu.main(str(path)) == 0

    backups = [p for p in tmp_path.iterdir() if p != path]
    assert len(backups) == 1
    assert _is_backup_name(backups[0].stem), (
        f"backup stem {backups[0].stem!r} is not recognized as a backup by "
        "_is_backup_name — it would pollute next-month seeding"
    )


# --- F6: validate before writing the backup ---------------------------------


def test_cli_main_aborts_before_backup_when_required_sheets_missing(tmp_path):
    path = _make_target_file(tmp_path, builder=_make_workbook_missing_cuposcober)

    assert ccu.main(str(path)) == 1

    assert [p.name for p in tmp_path.iterdir()] == [path.name]
    out = load_workbook(path)
    assert SHEET not in out.sheetnames


def test_cli_main_aborts_before_backup_when_sucursal_not_found(tmp_path):
    path = _make_target_file(tmp_path, builder=_make_workbook_without_guemes_in_cupos)

    assert ccu.main(str(path)) == 1

    assert [p.name for p in tmp_path.iterdir()] == [path.name]
    out = load_workbook(path)
    assert SHEET not in out.sheetnames


def test_cli_main_succeeds_when_sucursal_present_in_cuposcober(tmp_path):
    path = _make_target_file(tmp_path)

    assert ccu.main(str(path)) == 0

    out = load_workbook(path)
    assert SHEET in out.sheetnames


# --- F4: capture range must track the layout --------------------------------


def test_capture_range_matches_actual_sheet_extent():
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    config = json.loads(CONFIG_PATH.read_text())
    capture_entries = config["reportes"][0]["capture_images"]
    ccu_entry = next(e for e in capture_entries if e["hoja"] == SHEET)

    expected_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    assert ccu_entry["rango"] == expected_range


# --- F7.1: golden formula assertions for anchor cells -----------------------


def test_golden_formulas_marca_group_first_data_row():
    """Section 1 (CERVEZAS 1/2): top_row=1, header_row=2, subheader_row=3,
    first_data_row=4. JORGE RAMOS is the first preventista row. SALTA is the
    first marca (group_index 0 -> column B)."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    assert ws["A4"].value == "JORGE RAMOS"
    assert ws["B2"].value == "SALTA"

    assert ws["B4"].value == (
        '=SUMIFS(COBER!$E:$E,COBER!$A:$A,"16 - SUCURSAL GUEMES",'
        'COBER!$B:$B,$A4,COBER!$D:$D,B$2)'
    )
    assert ws["C4"].value == (
        '=SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"16 - SUCURSAL GUEMES",'
        'CuposCober!$C:$C,$A4,CuposCober!$D:$D,B$2)'
    )
    assert ws["D4"].value == "=C4-B4"
    assert ws["E4"].value == '=IF(C4=0,"",B4/C4)'
    # Percentage columns display exactly 2 decimals; the value is never rounded.
    assert ws["E4"].number_format == "0.00%"
    assert ws["B4"].number_format == "#,##0"


def test_golden_formulas_marca_group_total_guemes_row():
    """Same marca group (SALTA), TOTAL GUEMES row of section 1 -> row 7."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    assert ws["A7"].value == "TOTAL GUEMES"

    assert ws["B7"].value == (
        '=SUMIFS(COBER!$E:$E,COBER!$A:$A,"16 - SUCURSAL GUEMES",'
        'COBER!$D:$D,B$2)'
    )
    assert ws["C7"].value == (
        '=SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"16 - SUCURSAL GUEMES",'
        'CuposCober!$D:$D,B$2)'
    )
    assert ws["D7"].value == "=C7-B7"
    assert ws["E7"].value == '=IF(C7=0,"",B7/C7)'


def _find_total_generico_header(ws, generico: str):
    """Locate the "TOTAL {generico}" header cell without hardcoding its
    row/column — section layout can shift when a section is inserted."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == f"TOTAL {generico}":
                return cell
    raise AssertionError(f"no 'TOTAL {generico}' header found")


def test_golden_formulas_generico_total_group_first_data_row():
    """The TOTAL CERVEZAS group now lives in the third CERVEZAS section
    (CERVEZAS (3/3)). Locate it dynamically instead of hardcoding a fixed
    column/row, since inserting a section shifts everything after it."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    header_cell = _find_total_generico_header(ws, "CERVEZAS")
    pdv_col = get_column_letter(header_cell.column)
    obj_col = get_column_letter(header_cell.column + 1)
    faltan_col = get_column_letter(header_cell.column + 2)
    pct_col = get_column_letter(header_cell.column + 3)
    data_row = header_cell.row + 2  # header_row -> subheader_row -> first data row

    assert ws[f"A{data_row}"].value == "JORGE RAMOS"

    expected_marcas = ccu.GENERICO_MARCAS["CERVEZAS"]
    assert len(expected_marcas) == 17
    array_literal = "{" + ";".join(f'"{m}"' for m in expected_marcas) + "}"

    pdv_formula = ws[f"{pdv_col}{data_row}"].value
    obj_formula = ws[f"{obj_col}{data_row}"].value

    assert pdv_formula == (
        '=SUMIFS(COBER!$L:$L,COBER!$H:$H,"16 - SUCURSAL GUEMES",'
        f'COBER!$I:$I,$A{data_row},COBER!$K:$K,"CERVEZAS")'
    )
    assert obj_formula == (
        '=SUM(SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"16 - SUCURSAL GUEMES",'
        f'CuposCober!$C:$C,$A{data_row},CuposCober!$D:$D,{array_literal}))'
    )
    assert ws[f"{faltan_col}{data_row}"].value == f"={obj_col}{data_row}-{pdv_col}{data_row}"
    assert ws[f"{pct_col}{data_row}"].value == (
        f'=IF({obj_col}{data_row}=0,"",{pdv_col}{data_row}/{obj_col}{data_row})'
    )
    assert len(_extract_array_literal(obj_formula)) == 17


def test_golden_formulas_generico_total_group_total_guemes_row():
    """Same generico-total group (CERVEZAS), TOTAL GUEMES row of its section,
    located dynamically."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    header_cell = _find_total_generico_header(ws, "CERVEZAS")
    pdv_col = get_column_letter(header_cell.column)
    obj_col = get_column_letter(header_cell.column + 1)
    faltan_col = get_column_letter(header_cell.column + 2)
    pct_col = get_column_letter(header_cell.column + 3)
    data_row = header_cell.row + 2
    total_row = data_row + len(ccu.PREVENTISTAS)

    assert ws[f"A{total_row}"].value == "TOTAL GUEMES"

    expected_marcas = ccu.GENERICO_MARCAS["CERVEZAS"]
    assert len(expected_marcas) == 17
    array_literal = "{" + ";".join(f'"{m}"' for m in expected_marcas) + "}"

    pdv_formula = ws[f"{pdv_col}{total_row}"].value
    obj_formula = ws[f"{obj_col}{total_row}"].value

    assert pdv_formula == (
        '=SUMIFS(COBER!$L:$L,COBER!$H:$H,"16 - SUCURSAL GUEMES",'
        'COBER!$K:$K,"CERVEZAS")'
    )
    assert obj_formula == (
        '=SUM(SUMIFS(CuposCober!$H:$H,CuposCober!$E:$E,"16 - SUCURSAL GUEMES",'
        f'CuposCober!$D:$D,{array_literal}))'
    )
    assert ws[f"{faltan_col}{total_row}"].value == f"={obj_col}{total_row}-{pdv_col}{total_row}"
    assert ws[f"{pct_col}{total_row}"].value == (
        f'=IF({obj_col}{total_row}=0,"",{pdv_col}{total_row}/{obj_col}{total_row})'
    )
    assert len(_extract_array_literal(obj_formula)) == 17


# --- promoted marcas: OBJ_ONLY_MARCAS is now empty --------------------------


def test_promoted_marcas_are_displayed_columns_and_obj_only_is_empty():
    """The 7 marcas previously hidden in OBJ_ONLY_MARCAS (verified generico
    via gold.dim_articulo) must now be their own displayed columns, not an
    OBJ-only rollup contributor."""
    promoted = ["GROLSCH", "IGUANA", "ISENBECK", "WARSTEINER", "NORTE", "PALERMO", "CONTROL C"]

    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]
    marca_headers = {cell.value for row in ws.iter_rows() for cell in row}

    for marca in promoted:
        assert marca in marca_headers, f"{marca} is not displayed as a column header"

    assert ccu.OBJ_ONLY_MARCAS == {}


# --- F1 + F7.2: generico OBJ rollup must cover ALL marcas of the generico ---


def test_generico_obj_rollup_matches_displayed_marcas_plus_obj_only_marcas():
    """For each generico with a total group, the array literal in the
    TOTAL {generico} OBJ formula must equal the union of displayed marcas
    across ALL sections of that generico (both CERVEZAS halves) plus
    OBJ_ONLY_MARCAS. This must fail if someone drops a rollup half or
    overwrites GENERICO_MARCAS instead of extending it."""
    wb = _make_workbook()
    ccu.build_cobertura_ccu(wb)
    ws = wb[SHEET]

    displayed_by_generico: dict[str, set[str]] = {}
    has_total_group: set[str] = set()
    for _title, generico, marcas, emit_generico_total in ccu.SECTIONS:
        displayed_by_generico.setdefault(generico, set()).update(marcas)
        if emit_generico_total:
            has_total_group.add(generico)

    assert has_total_group, "expected at least one generico with a total group"

    for generico in has_total_group:
        expected = displayed_by_generico[generico] | set(ccu.OBJ_ONLY_MARCAS.get(generico, []))

        header_cell = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == f"TOTAL {generico}":
                    header_cell = cell
        assert header_cell is not None, f"no 'TOTAL {generico}' header found"

        obj_col = header_cell.column + 1  # PDV, OBJ, Faltan, %
        data_row = header_cell.row + 2  # header_row -> subheader_row -> first data row
        obj_formula = ws.cell(row=data_row, column=obj_col).value
        actual = set(_extract_array_literal(obj_formula))

        assert actual == expected, f"{generico}: rollup {actual} != expected {expected}"
