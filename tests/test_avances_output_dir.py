"""Tests for the new avances workflow: base + output_dir per period.

Spec:
- archivo_plantilla is the BASE template, READ-ONLY (input dir, never modified)
- Output goes to data/output/avances/{YYYY-MM}/{nombre_archivo}.xlsx
- Snapshot of the base is saved alongside the output (same folder, original filename)
- Re-running the same period updates the existing output (preserves user customizations)
  but always refreshes the base snapshot to reflect the base used in this run
"""
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.core.data_loader import DataLoader
from src.services.avances.service import AvancesConfig, AvancesService


def _make_base(path: Path) -> None:
    """Create a minimal valid base xlsx with required gold sheets + a USER sheet
    that must NOT be deleted by avances."""
    wb = Workbook()
    ws_fv = wb.active
    ws_fv.title = "gold fact_ventas"
    cols = [
        "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
        "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
        "anulado", "cantidades_total", "bonificacion",
    ]
    for i, h in enumerate(cols, 1):
        ws_fv.cell(row=1, column=i, value=h)

    ws_da = wb.create_sheet("gold dim_articulo")
    for i, h in enumerate([
        "id_articulo", "des_articulo", "marca", "generico", "calibre",
        "proveedor", "unidad_negocio", "factor_hectolitros",
    ], 1):
        ws_da.cell(row=1, column=i, value=h)

    ws_dc = wb.create_sheet("gold dim_cliente")
    ws_dc.cell(row=1, column=1, value="TITULO")
    for i, h in enumerate([
        "id_cliente", "fantasia", "razon_social", "des_sucursal", "id_sucursal",
        "id_ruta_fv1", "des_personal_fv1", "id_ruta_fv4", "des_personal_fv4",
    ], 1):
        ws_dc.cell(row=2, column=i, value=h)

    ws_cg = wb.create_sheet("gold cob_preventista_generico")
    for i, h in enumerate([
        "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
        "id_sucursal", "ds_sucursal", "generico", "clientes_compradores",
        "volumen_total",
    ], 1):
        ws_cg.cell(row=1, column=i, value=h)

    ws_cm = wb.create_sheet("gold cob_preventista_marca")
    for i, h in enumerate([
        "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
        "id_sucursal", "ds_sucursal", "marca", "clientes_compradores",
        "volumen_total",
    ], 1):
        ws_cm.cell(row=1, column=i, value=h)

    # USER-OWNED sheet with formulas — must survive any avances run
    ws_user = wb.create_sheet("MI ANALISIS")
    ws_user.cell(row=1, column=1, value="formula")
    ws_user.cell(row=1, column=2, value="=SUM(1+1)")

    wb.save(str(path))


def _mock_loader() -> MagicMock:
    """Loader returning empty DFs with the schemas avances expects."""
    loader = MagicMock(spec=DataLoader)
    loader.get_fact_ventas_raw.return_value = pd.DataFrame(columns=[
        "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
        "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
        "anulado", "cantidades_total", "bonificacion",
    ])
    loader.get_dim_articulo_raw.return_value = pd.DataFrame(columns=[
        "id_articulo", "des_articulo", "marca", "generico", "calibre",
        "proveedor", "unidad_negocio", "factor_hectolitros",
    ])
    loader.get_dim_cliente_raw.return_value = pd.DataFrame(columns=[
        "id_cliente", "fantasia", "razon_social", "des_sucursal", "id_sucursal",
        "id_ruta_fv1", "des_personal_fv1", "id_ruta_fv4", "des_personal_fv4",
    ])
    loader.get_cob_preventista_generico_raw.return_value = pd.DataFrame(columns=[
        "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
        "id_sucursal", "ds_sucursal", "generico", "clientes_compradores",
        "volumen_total",
    ])
    loader.get_cob_preventista_marca_raw.return_value = pd.DataFrame(columns=[
        "id", "periodo", "id_fuerza_ventas", "id_vendedor", "id_ruta",
        "id_sucursal", "ds_sucursal", "marca", "clientes_compradores",
        "volumen_total",
    ])
    return loader


class TestAvancesOutputDir:
    """The new workflow writes to data/output/avances/{YYYY-MM}/, not in-place."""

    def test_output_lands_in_period_folder(self, tmp_path):
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)
        output_root = tmp_path / "out"

        service = AvancesService(data_loader=_mock_loader())
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            nombre_archivo="AVANCE BRANCA - ABRIL 2026",
            output_dir=output_root / "avances" / "2026-04",
        )
        result = service.generar_reporte(config)

        expected = output_root / "avances" / "2026-04" / "AVANCE BRANCA - ABRIL 2026.xlsx"
        assert result.ruta_archivo == expected
        assert expected.exists()

    def test_base_in_input_is_not_modified(self, tmp_path):
        """The input/base file must be byte-identical after a run."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)
        original_bytes = base.read_bytes()
        output_root = tmp_path / "out"

        service = AvancesService(data_loader=_mock_loader())
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            nombre_archivo="AVANCE BRANCA - ABRIL 2026",
            output_dir=output_root / "avances" / "2026-04",
        )
        service.generar_reporte(config)

        assert base.read_bytes() == original_bytes, (
            "Base file in input/ must never be modified by avances service"
        )

    def test_base_snapshot_saved_alongside_output(self, tmp_path):
        """A copy of the base lives in the same period folder as the output."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)
        output_root = tmp_path / "out"

        service = AvancesService(data_loader=_mock_loader())
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            nombre_archivo="AVANCE BRANCA - ABRIL 2026",
            output_dir=output_root / "avances" / "2026-04",
        )
        service.generar_reporte(config)

        snapshot = output_root / "avances" / "2026-04" / "AVANCE BRANCA.xlsx"
        assert snapshot.exists(), "Base snapshot must be saved in the period folder"
        # Snapshot must match the base byte-for-byte (it's a literal copy)
        assert snapshot.read_bytes() == base.read_bytes()

    def test_user_sheets_preserved_on_first_run(self, tmp_path):
        """User-added sheets in the base survive into the generated output."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)
        output_root = tmp_path / "out"

        service = AvancesService(data_loader=_mock_loader())
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            nombre_archivo="AVANCE BRANCA - ABRIL 2026",
            output_dir=output_root / "avances" / "2026-04",
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "MI ANALISIS" in wb.sheetnames
        assert wb["MI ANALISIS"].cell(1, 1).value == "formula"

    def test_regeneration_preserves_user_edits_in_output(self, tmp_path):
        """A second run on the same period preserves user edits made to the output file."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)
        output_root = tmp_path / "out"
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            nombre_archivo="AVANCE BRANCA - ABRIL 2026",
            output_dir=output_root / "avances" / "2026-04",
        )
        service = AvancesService(data_loader=_mock_loader())

        # First run
        result = service.generar_reporte(config)
        # User adds a sheet to the OUTPUT file (not the base)
        wb = load_workbook(str(result.ruta_archivo))
        wb.create_sheet("USER ADDED POST-GEN")
        wb["USER ADDED POST-GEN"].cell(1, 1).value = "important"
        wb.save(str(result.ruta_archivo))

        # Second run on same period
        service.generar_reporte(config)

        # User's post-gen sheet must still be there
        wb2 = load_workbook(str(result.ruta_archivo))
        assert "USER ADDED POST-GEN" in wb2.sheetnames
        assert wb2["USER ADDED POST-GEN"].cell(1, 1).value == "important"

    def test_default_output_dir_uses_service_output_dir_helper(self, tmp_path):
        """When output_dir is None, derive from service_output_dir('avances', fecha_desde)."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)

        # Patch DATA_OUTPUT so service_output_dir returns under tmp_path
        with patch("config.settings.DATA_OUTPUT", tmp_path / "out"):
            service = AvancesService(data_loader=_mock_loader())
            config = AvancesConfig(
                archivo_plantilla=str(base),
                fecha_desde="2026-05-01",
                fecha_hasta="2026-05-31",
                nombre_archivo="AVANCE BRANCA - MAYO 2026",
            )
            result = service.generar_reporte(config)

        expected_period = tmp_path / "out" / "avances" / "2026-05"
        assert result.ruta_archivo == expected_period / "AVANCE BRANCA - MAYO 2026.xlsx"
        assert (expected_period / "AVANCE BRANCA.xlsx").exists()  # snapshot

    def test_nombre_archivo_required(self, tmp_path):
        """nombre_archivo must be supplied — output filename is not derivable from base alone."""
        base = tmp_path / "AVANCE BRANCA.xlsx"
        _make_base(base)

        service = AvancesService(data_loader=_mock_loader())
        # Missing nombre_archivo
        config = AvancesConfig(
            archivo_plantilla=str(base),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            output_dir=tmp_path / "out",
        )
        try:
            service.generar_reporte(config)
        except ValueError as e:
            assert "nombre_archivo" in str(e).lower()
        else:
            raise AssertionError("Should have raised ValueError for missing nombre_archivo")
