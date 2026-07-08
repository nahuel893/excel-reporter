"""Tests para AvancesService."""
import pytest
import pandas as pd
from pathlib import Path
from unittest.mock import MagicMock
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.services.avances.service import (
    AvancesService,
    AvancesConfig,
    SHEET_CONFIGS_BRANCA,
    SHEET_CONFIGS_BADIE,
    PLANTILLA_SHEET_CONFIGS,
)


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_template(tmp_path: Path) -> Path:
    """Create a minimal template Excel with the 3 gold sheets."""
    wb = Workbook()

    # gold fact_ventas — header at row 1
    ws_fv = wb.active
    ws_fv.title = "gold fact_ventas"
    fact_ventas_cols = [
        "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
        "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
        "anulado", "cantidades_total", "bonificacion",
    ]
    for col_idx, h in enumerate(fact_ventas_cols, 1):
        ws_fv.cell(row=1, column=col_idx, value=h)
    # Some old data rows
    ws_fv.cell(row=2, column=1, value=999)
    ws_fv.cell(row=2, column=2, value=888)
    # Formula column after data columns
    ws_fv.cell(row=1, column=len(fact_ventas_cols) + 1, value="Formula")
    ws_fv.cell(row=2, column=len(fact_ventas_cols) + 1, value="=A2*2")

    # gold dim_articulo — header at row 1
    ws_art = wb.create_sheet("gold dim_articulo")
    dim_articulo_cols = [
        "id_articulo", "des_articulo", "marca", "generico", "calibre",
        "proveedor", "unidad_negocio", "factor_hectolitros",
    ]
    for col_idx, h in enumerate(dim_articulo_cols, 1):
        ws_art.cell(row=1, column=col_idx, value=h)
    ws_art.cell(row=2, column=1, value=1001)

    # gold dim_cliente — header at row 2 (row 1 is a title row)
    ws_cli = wb.create_sheet("gold dim_cliente")
    ws_cli.cell(row=1, column=1, value="TITULO")
    dim_cliente_cols = [
        "id_cliente", "fantasia", "razon_social", "des_sucursal", "id_sucursal",
        "id_ruta_fv1", "des_personal_fv1", "id_ruta_fv4", "des_personal_fv4",
    ]
    for col_idx, h in enumerate(dim_cliente_cols, 1):
        ws_cli.cell(row=2, column=col_idx, value=h)
    ws_cli.cell(row=3, column=1, value=5000)

    path = tmp_path / "plantilla.xlsx"
    wb.save(str(path))
    return path


def _make_fact_ventas_df():
    return pd.DataFrame({
        "id_cliente": [1, 2],
        "id_articulo": [101, 102],
        "id_vendedor": [10, 11],
        "id_sucursal": [1, 2],
        "fecha_comprobante": pd.to_datetime(["2026-04-01", "2026-04-02"]),
        "id_documento": [1001, 1002],
        "letra": ["A", "B"],
        "serie": ["0001", "0002"],
        "nro_doc": [500001, 500002],
        "anulado": [0, 0],
        "cantidades_total": [10.0, 20.0],
        "bonificacion": [0.0, 5.0],
    })


def _make_dim_articulo_df():
    return pd.DataFrame({
        "id_articulo": [101, 102],
        "des_articulo": ["BRAHMA LATA 473", "QUILMES 1L"],
        "marca": ["BRAHMA", "QUILMES"],
        "generico": ["CERVEZAS", "CERVEZAS"],
        "calibre": ["473ML", "1L"],
        "proveedor": ["AB INBEV", "AB INBEV"],
        "unidad_negocio": ["CERVEZAS", "CERVEZAS"],
        "factor_hectolitros": [0.00473, 0.01],
    })


def _make_dim_cliente_df():
    return pd.DataFrame({
        "id_cliente": [1, 2],
        "fantasia": ["ALMACEN SAN JUAN", "SUPER NORTE"],
        "razon_social": ["ALMACEN SAN JUAN SRL", "SUPER NORTE SA"],
        "des_sucursal": ["CASA CENTRAL", "CAFAYATE"],
        "id_sucursal": [1, 2],
        "id_ruta_fv1": [81, 10],
        "des_personal_fv1": ["Juan", "Ana"],
        "id_ruta_fv4": [81, 10],
        "des_personal_fv4": ["Juan", "Ana"],
    })


def _make_cob_prev_generico_df():
    return pd.DataFrame({
        "id": [1, 2],
        "periodo": pd.to_datetime(["2026-04-01", "2026-04-01"]),
        "id_fuerza_ventas": [1, 1],
        "id_vendedor": [10, 11],
        "id_ruta": [81, 10],
        "id_sucursal": [1, 1],
        "ds_sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
        "generico": ["CERVEZAS", "AGUAS DANONE"],
        "clientes_compradores": [50, 30],
        "volumen_total": [100.5, 50.3],
    })


def _make_cob_prev_marca_df():
    return pd.DataFrame({
        "id": [1, 2],
        "periodo": pd.to_datetime(["2026-04-01", "2026-04-01"]),
        "id_fuerza_ventas": [1, 1],
        "id_vendedor": [10, 11],
        "id_ruta": [81, 10],
        "id_sucursal": [1, 1],
        "ds_sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
        "marca": ["QUILMES", "BRAHMA"],
        "clientes_compradores": [40, 25],
        "volumen_total": [80.0, 45.0],
    })


def _make_mock_loader():
    loader = MagicMock(spec=DataLoader)
    loader.get_fact_ventas_raw.return_value = _make_fact_ventas_df()
    loader.get_dim_articulo_raw.return_value = _make_dim_articulo_df()
    loader.get_dim_cliente_raw.return_value = _make_dim_cliente_df()
    loader.get_cob_preventista_generico_raw.return_value = _make_cob_prev_generico_df()
    loader.get_cob_preventista_marca_raw.return_value = _make_cob_prev_marca_df()
    return loader


# ── Tests ─────────────────────────────────────────────────────────────────────


class TestAvancesServiceHappyPath:
    def test_output_written_to_period_folder(self, tmp_path):
        """Template with 3 gold sheets + mocked DataLoader → output written to
        output_dir, registros_por_hoja has correct counts, formula columns preserved."""
        plantilla = _make_template(tmp_path)
        mock_loader = _make_mock_loader()
        out_dir = tmp_path / "out" / "avances" / "2026-04"

        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(plantilla),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-15",
            nombre_archivo="AVANCE TEST - ABRIL 2026",
            output_dir=out_dir,
        )

        result = service.generar_reporte(config)

        # Result points to the OUTPUT path (not the base in input)
        assert result.ruta_archivo == out_dir / "AVANCE TEST - ABRIL 2026.xlsx"
        assert result.ruta_archivo.exists()
        # Base in input is untouched
        assert plantilla.exists()

        # registros_por_hoja has correct counts for all 5 gold-* sheets
        assert result.registros_por_hoja["gold fact_ventas"] == 2
        assert result.registros_por_hoja["gold dim_articulo"] == 2
        assert result.registros_por_hoja["gold dim_cliente"] == 2
        assert result.registros_por_hoja["gold cob_preventista_generico"] == 2
        assert result.registros_por_hoja["gold cob_preventista_marca"] == 2

        # Formula column in fact_ventas untouched
        wb = load_workbook(str(result.ruta_archivo))
        ws = wb["gold fact_ventas"]
        formula_col_idx = len([
            "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
            "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
            "anulado", "cantidades_total", "bonificacion",
        ]) + 1
        assert ws.cell(row=2, column=formula_col_idx).value == "=A2*2"

    def test_loader_methods_called_with_filter_params(self, tmp_path):
        """Service MUST pass id_sucursal + id_fuerza_ventas to the loader methods."""
        plantilla = _make_template(tmp_path)
        mock_loader = _make_mock_loader()

        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(plantilla),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            id_sucursal=1,
            id_fuerza_ventas=4,
            nombre_archivo="AVANCE TEST - ABRIL 2026",
            output_dir=tmp_path / "out",
        )
        service.generar_reporte(config)

        # fact_ventas: fechas + id_sucursal
        fv_call = mock_loader.get_fact_ventas_raw.call_args
        assert fv_call.kwargs == {
            "fecha_desde": "2026-04-01",
            "fecha_hasta": "2026-04-30",
            "id_sucursal": 1,
        }

        # dim_articulo: no params — full dimension table (NOT filtered)
        da_call = mock_loader.get_dim_articulo_raw.call_args
        assert da_call.kwargs == {}
        assert da_call.args == ()

        # dim_cliente: id_sucursal only
        dc_call = mock_loader.get_dim_cliente_raw.call_args
        assert dc_call.kwargs == {"id_sucursal": 1}

        # cob_preventista_generico: fechas + FV + sucursal
        cg_call = mock_loader.get_cob_preventista_generico_raw.call_args
        assert cg_call.kwargs == {
            "fecha_desde": "2026-04-01",
            "fecha_hasta": "2026-04-30",
            "id_fuerza_ventas": 4,
            "id_sucursal": 1,
        }

        # cob_preventista_marca: same
        cm_call = mock_loader.get_cob_preventista_marca_raw.call_args
        assert cm_call.kwargs == {
            "fecha_desde": "2026-04-01",
            "fecha_hasta": "2026-04-30",
            "id_fuerza_ventas": 4,
            "id_sucursal": 1,
        }

    def test_base_in_input_unchanged_after_run(self, tmp_path):
        """The base template in input must be byte-identical after a run."""
        plantilla = _make_template(tmp_path)
        original_bytes = plantilla.read_bytes()
        mock_loader = _make_mock_loader()

        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(plantilla),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-15",
            nombre_archivo="AVANCE TEST - ABRIL 2026",
            output_dir=tmp_path / "out" / "avances" / "2026-04",
        )
        service.generar_reporte(config)

        assert plantilla.read_bytes() == original_bytes


class TestAvancesServiceErrors:
    def test_missing_template_raises(self, tmp_path):
        """Config points to nonexistent file → FileNotFoundError."""
        service = AvancesService(data_loader=MagicMock(spec=DataLoader))
        config = AvancesConfig(
            archivo_plantilla=str(tmp_path / "no_existe.xlsx"),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-15",
            nombre_archivo="ignored",
            output_dir=tmp_path / "out",
        )

        with pytest.raises(FileNotFoundError):
            service.generar_reporte(config)


class TestAvancesServiceMissingSheet:
    def test_missing_sheet_creates_it(self, tmp_path, caplog):
        """Template missing sheets → creates them, writes data, all 5 in result."""
        # Build template with only 2 of the 5 sheets
        wb = Workbook()
        ws_fv = wb.active
        ws_fv.title = "gold fact_ventas"
        for col_idx, h in enumerate([
            "id_cliente", "id_articulo", "id_vendedor", "id_sucursal",
            "fecha_comprobante", "id_documento", "letra", "serie", "nro_doc",
            "anulado", "cantidades_total", "bonificacion",
        ], 1):
            ws_fv.cell(row=1, column=col_idx, value=h)

        ws_art = wb.create_sheet("gold dim_articulo")
        for col_idx, h in enumerate([
            "id_articulo", "des_articulo", "marca", "generico", "calibre",
            "proveedor", "unidad_negocio", "factor_hectolitros",
        ], 1):
            ws_art.cell(row=1, column=col_idx, value=h)

        # gold dim_cliente, gold cob_preventista_generico, gold cob_preventista_marca omitted
        plantilla = tmp_path / "plantilla_incompleta.xlsx"
        wb.save(str(plantilla))

        mock_loader = _make_mock_loader()
        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(plantilla),
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-15",
            nombre_archivo="AVANCE TEST INCOMPLETO",
            output_dir=tmp_path / "out" / "avances" / "2026-04",
        )

        import logging
        with caplog.at_level(logging.INFO):
            result = service.generar_reporte(config)

        # All 5 sheets processed (3 created on the fly)
        assert len(result.registros_por_hoja) == 5
        assert "gold dim_cliente" in result.registros_por_hoja
        assert "gold cob_preventista_generico" in result.registros_por_hoja
        assert "gold cob_preventista_marca" in result.registros_por_hoja
        # Log says sheets were created
        assert any("not found, creating" in record.message for record in caplog.records)


# ── Registry unit tests (tasks 1.8 – 1.12) ────────────────────────────────────


class TestPlantillaRegistry:
    """Pure unit tests for PLANTILLA_SHEET_CONFIGS registry and AvancesConfig defaults.

    No file I/O — registry assertions only.
    """

    def test_registry_branca_resolves(self):
        """PLANTILLA_SHEET_CONFIGS["branca"] must be the exact SHEET_CONFIGS_BRANCA object."""
        assert PLANTILLA_SHEET_CONFIGS["branca"] is SHEET_CONFIGS_BRANCA

    def test_registry_badie_resolves(self):
        """PLANTILLA_SHEET_CONFIGS["badie"] must be the exact SHEET_CONFIGS_BADIE object."""
        assert PLANTILLA_SHEET_CONFIGS["badie"] is SHEET_CONFIGS_BADIE

    def test_registry_unknown_raises(self):
        """Accessing a non-existent key must raise KeyError — no silent fallback."""
        with pytest.raises(KeyError):
            _ = PLANTILLA_SHEET_CONFIGS["bogus"]

    def test_avances_config_tipo_plantilla_default(self):
        """AvancesConfig without tipo_plantilla must default to 'branca'."""
        config = AvancesConfig(fecha_desde="2026-06-01", fecha_hasta="2026-06-30")
        assert config.tipo_plantilla == "branca"

    def test_avances_config_tipo_plantilla_badie(self):
        """AvancesConfig with tipo_plantilla='badie' must store that value."""
        config = AvancesConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            tipo_plantilla="badie",
        )
        assert config.tipo_plantilla == "badie"

    def test_avances_config_tipo_plantilla_literal_includes_guemes(self):
        """RF-07: AvancesConfig.tipo_plantilla Literal must include 'guemes'
        (dataclass Literal — static type-hint consistency, not runtime-enforced;
        the runtime gate is GlobalFilters in src/config/models.py)."""
        import typing
        from dataclasses import fields

        tipo_field = next(f for f in fields(AvancesConfig) if f.name == "tipo_plantilla")
        assert "guemes" in typing.get_args(tipo_field.type)


# ── Badie integration tests (tasks 2.5–2.6) ───────────────────────────────────

BADIE_FIXTURE = Path(__file__).parent / "fixtures" / "avance_badie_minimal.xlsx"


def _make_badie_ventas_df():
    """DataFrame shape matching get_fact_ventas_pivot_badie — descriptive headers."""
    return pd.DataFrame({
        "Sucursal": ["1 - CASA CENTRAL", "1 - CASA CENTRAL", "1 - CASA CENTRAL"],
        "Descripcion Período": pd.to_datetime(["2026-05-01", "2026-05-02", "2026-05-03"]),
        "Descripcion Vendedor": ["AGUIRRE ETHEL", "AGUIRRE ETHEL", "GOMEZ JUAN"],
        "Ruta": [6, 6, 7],
        "Descripcion_Ruta": ["AGUIRRE ETHEL LUJU", "AGUIRRE ETHEL LUJU", "GOMEZ ZONA SUR"],
        "Descripcion_Marca": ["ARIZU", "BRAHMA", "QUILMES"],
        "GENERICO": ["VINOS", "CERVEZAS", "CERVEZAS"],
        "Código_Articulo": [821016, 102, 103],
        "Descripcion_Articulo": ["ARIZU BLANCO 1000 * 12", "BRAHMA 473", "QUILMES 1L"],
        "Cantidades Totales": [10.0, 20.0, 30.0],
    })


def _make_badie_cob_gen_df():
    """DataFrame shape matching get_cob_preventista_generico_pivot_badie."""
    return pd.DataFrame({
        "Sucursal": ["1 - CASA CENTRAL", "1 - CASA CENTRAL", "1 - CASA CENTRAL"],
        "Descripcion Vendedor": ["AGUIRRE ETHEL", "AGUIRRE ETHEL", "GOMEZ JUAN"],
        "Ruta": [6, 6, 7],
        "GENERICO": ["CERVEZAS", "AGUAS DANONE", "VINOS"],
        "Numero_Clientes": [50, 30, 20],
    })


def _make_badie_cob_marca_df():
    """DataFrame shape matching get_cob_preventista_marca_pivot_badie."""
    return pd.DataFrame({
        "Sucursal": ["1 - CASA CENTRAL", "1 - CASA CENTRAL", "1 - CASA CENTRAL"],
        "Descripcion Vendedor": ["AGUIRRE ETHEL", "AGUIRRE ETHEL", "GOMEZ JUAN"],
        "Ruta": [6, 6, 7],
        "Descripcion_Marca": ["QUILMES", "BRAHMA", "ARIZU"],
        "Numero_Clientes": [40, 25, 15],
    })


def _make_badie_cupos_volumen_df():
    """DataFrame shape matching get_cupos_volumen_badie."""
    return pd.DataFrame({
        "Código": [6, 6, 7],
        "Descripción": ["AGUIRRE ETHEL", "AGUIRRE ETHEL", "GOMEZ JUAN"],
        "PREVENTISTA": ["AGUIRRE ETHEL LUJU", "AGUIRRE ETHEL LUJU", "GOMEZ ZONA SUR"],
        "GENERICO": ["CERVEZAS", "AGUAS DANONE", "CERVEZAS"],
        "DESAGREGADO": ["CERVEZAS", "AGUAS DANONE", "CERVEZAS"],
        "Cupo ": [800.5, 600.0, 1100.25],
    })


def _make_badie_cupos_cob_gen_df():
    """DataFrame shape matching get_cupos_cobertura_generico_badie (post-swap)."""
    return pd.DataFrame({
        "Ruta": [1, 1, 2],
        "Preventista": ["ROBLES ORLANDO", "ROBLES ORLANDO", "AGUIRRE ETHEL"],
        "Generico": ["CERVEZAS", "AGUAS DANONE", "VINOS CCU"],
        "ZONA": ["1 - CASA CENTRAL", "1 - CASA CENTRAL", "1 - CASA CENTRAL"],
        "CUPO ": [67.6, 30.5, 12.3],
    })


def _make_badie_cupos_cob_marca_df():
    """DataFrame shape matching get_cupos_cobertura_marca_badie (post-swap)."""
    return pd.DataFrame({
        "Ruta": [1, 1, 2],
        "Descripción Vendedor": ["ROBLES ORLANDO", "ROBLES ORLANDO", "AGUIRRE ETHEL"],
        "MARCA": ["GROLSCH", "HEINEKEN", "ARIZU"],
        "ZONA": ["1 - CASA CENTRAL", "1 - CASA CENTRAL", "1 - CASA CENTRAL"],
        "CUPO ": [0.89, 19.31, 0.74],
    })


def _make_badie_mock_loader():
    loader = MagicMock(spec=DataLoader)
    loader.get_fact_ventas_pivot_badie.return_value = _make_badie_ventas_df()
    loader.get_cob_preventista_generico_pivot_badie.return_value = _make_badie_cob_gen_df()
    loader.get_cob_preventista_marca_pivot_badie.return_value = _make_badie_cob_marca_df()
    loader.get_cupos_volumen_badie.return_value = _make_badie_cupos_volumen_df()
    loader.get_cupos_cobertura_generico_badie.return_value = _make_badie_cupos_cob_gen_df()
    loader.get_cupos_cobertura_marca_badie.return_value = _make_badie_cupos_cob_marca_df()
    return loader


class TestBadieRoundTrip:
    """Integration tests for tipo_plantilla='badie' using the minimal fixture."""

    def test_badie_round_trip_row_counts(self, tmp_path):
        """generar_reporte with tipo_plantilla='badie' must refresh all 3 data sheets.

        registros_por_hoja must have exactly 3 keys matching the Badie sheet names,
        each with the correct row count from the mocked DataLoader.

        Uses fecha_desde in 2019 to avoid picking up real previous-month output.
        """
        assert BADIE_FIXTURE.exists(), f"Fixture not found: {BADIE_FIXTURE}"

        mock_loader = _make_badie_mock_loader()
        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(BADIE_FIXTURE),
            fecha_desde="2019-05-01",
            fecha_hasta="2019-05-31",
            tipo_plantilla="badie",
            id_sucursal=1,
            id_fuerza_ventas=1,
            nombre_archivo="AVANCE BADIE - MAYO 2019",
            output_dir=tmp_path / "out",
        )

        result = service.generar_reporte(config)

        # Exactly 6 sheets refreshed: 3 data + 3 cupos
        assert len(result.registros_por_hoja) == 6
        for sheet in ("pivot_python", "cober_gen", "cober_marca",
                      "CuposVolumen", "CuposCoberGen", "CuposCober"):
            assert sheet in result.registros_por_hoja, f"Missing sheet: {sheet}"
            assert result.registros_por_hoja[sheet] == 3, f"{sheet}: expected 3 rows"

    def test_badie_round_trip_sample_values(self, tmp_path):
        """Data written to pivot_python and cober_gen must reflect the mocked DF values.

        Uses fecha_desde in 2019 to avoid picking up real previous-month output.
        """
        assert BADIE_FIXTURE.exists(), f"Fixture not found: {BADIE_FIXTURE}"

        mock_loader = _make_badie_mock_loader()
        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(BADIE_FIXTURE),
            fecha_desde="2019-05-01",
            fecha_hasta="2019-05-31",
            tipo_plantilla="badie",
            id_sucursal=1,
            id_fuerza_ventas=1,
            nombre_archivo="AVANCE BADIE - MAYO 2019",
            output_dir=tmp_path / "out",
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))

        # pivot_python: Cantidades Totales col J, row 2 should be 10.0 (first mock row)
        ws_piv = wb["pivot_python"]
        cantidades_col = None
        for cell in ws_piv[1]:
            if cell.value == "Cantidades Totales":
                cantidades_col = cell.column
                break
        assert cantidades_col is not None, "Header 'Cantidades Totales' not found in pivot_python"
        assert ws_piv.cell(row=2, column=cantidades_col).value == 10.0

        # cober_gen: GENERICO col E, row 2 should be "CERVEZAS"
        ws_cg = wb["cober_gen"]
        generico_col = None
        for cell in ws_cg[1]:
            if cell.value == "GENERICO":
                generico_col = cell.column
                break
        assert generico_col is not None, "Header 'GENERICO' not found in cober_gen"
        assert ws_cg.cell(row=2, column=generico_col).value == "CERVEZAS"

        # cober_marca: Descripcion_Marca col E, row 2 should be "QUILMES"
        ws_cm = wb["cober_marca"]
        marca_col = None
        for cell in ws_cm[1]:
            if cell.value == "Descripcion_Marca":
                marca_col = cell.column
                break
        assert marca_col is not None, "Header 'Descripcion_Marca' not found in cober_marca"
        assert ws_cm.cell(row=2, column=marca_col).value == "QUILMES"

        wb.close()

    def test_badie_avance_formula_sheet_untouched(self, tmp_path):
        """The Avance formula sheet must exist and its formulas must be preserved.

        Note: formula EVALUATION requires Excel — we only verify the formula text
        is preserved (cell.value starts with '=').

        Uses fecha_desde in 2019 to avoid picking up any real previous-month output
        from data/output/avances/ — the resolver falls back to archivo_plantilla.
        """
        assert BADIE_FIXTURE.exists(), f"Fixture not found: {BADIE_FIXTURE}"

        mock_loader = _make_badie_mock_loader()
        service = AvancesService(data_loader=mock_loader)
        config = AvancesConfig(
            archivo_plantilla=str(BADIE_FIXTURE),
            fecha_desde="2019-05-01",
            fecha_hasta="2019-05-31",
            tipo_plantilla="badie",
            id_sucursal=1,
            id_fuerza_ventas=1,
            nombre_archivo="AVANCE BADIE - MAYO 2019",
            output_dir=tmp_path / "out",
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo), data_only=False)

        # Avance sheet must exist
        assert "Avance" in wb.sheetnames

        # All 3 formula cells in Avance must start with "=" (formula text preserved)
        ws_avance = wb["Avance"]
        for col_idx in (1, 2, 3):
            cell_val = ws_avance.cell(row=1, column=col_idx).value
            assert isinstance(cell_val, str) and cell_val.startswith("="), (
                f"Avance!{col_idx}1 expected formula, got {cell_val!r}"
            )

        wb.close()

    @pytest.mark.parametrize("tipo_plantilla,expected_sheets", [
        ("branca", {"gold fact_ventas", "gold dim_articulo", "gold dim_cliente",
                    "gold cob_preventista_generico", "gold cob_preventista_marca"}),
        ("badie", {"pivot_python", "cober_gen", "cober_marca",
                    "CuposVolumen", "CuposCoberGen", "CuposCober"}),
    ])
    def test_registry_dispatch_correct_sheets(self, tmp_path, tipo_plantilla, expected_sheets):
        """Registry dispatch must route to the correct sheet set per tipo_plantilla."""
        from unittest.mock import patch, call

        # For branca, use the branca fixture helper; for badie use the badie fixture
        if tipo_plantilla == "branca":
            fixture = _make_template(tmp_path)
            loader = _make_mock_loader()
        else:
            fixture = BADIE_FIXTURE
            loader = _make_badie_mock_loader()

        service = AvancesService(data_loader=loader)
        # Use 2019 dates to avoid picking up real previous-month output from
        # data/output/avances/ — resolver falls back to archivo_plantilla.
        config = AvancesConfig(
            archivo_plantilla=str(fixture),
            fecha_desde="2019-05-01",
            fecha_hasta="2019-05-31",
            tipo_plantilla=tipo_plantilla,
            nombre_archivo=f"AVANCE TEST {tipo_plantilla.upper()} - MAYO 2019",
            output_dir=tmp_path / f"out_{tipo_plantilla}",
        )

        result = service.generar_reporte(config)
        assert set(result.registros_por_hoja.keys()) == expected_sheets


class TestMainForwardsTipoPlantilla:
    """Verify _run_avances_report forwards tipo_plantilla from merged dict."""

    def test_main_forwards_tipo_plantilla(self, tmp_path):
        """When merged contains tipo_plantilla='badie', AvancesConfig receives it."""
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch, MagicMock

        # Minimal report entry (nombre required by _run_avances_report)
        report = ReportEntry(nombre="AVANCE BADIE - JUNIO 2026")

        merged = {
            "tipo_plantilla": "badie",
            "fecha_desde": "2026-06-01",
            "fecha_hasta": "2026-06-30",
            "id_sucursal": 1,
            "id_fuerza_ventas": 1,
        }

        captured_configs: list[AvancesConfig] = []

        def fake_generar_reporte(self_svc, config: AvancesConfig):
            captured_configs.append(config)
            result = MagicMock()
            result.ruta_archivo = tmp_path / "output.xlsx"
            result.registros_por_hoja = {}
            return result

        with patch("src.services.avances.service.AvancesService.generar_reporte", fake_generar_reporte):
            main_module._run_avances_report(report, merged)

        assert len(captured_configs) == 1
        assert captured_configs[0].tipo_plantilla == "badie"

    def test_main_defaults_tipo_plantilla_to_branca(self, tmp_path):
        """When merged has no tipo_plantilla key, AvancesConfig defaults to 'branca'."""
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch, MagicMock

        report = ReportEntry(nombre="AVANCE BRANCA - JUNIO 2026")

        merged = {
            "fecha_desde": "2026-06-01",
            "fecha_hasta": "2026-06-30",
        }

        captured_configs: list[AvancesConfig] = []

        def fake_generar_reporte(self_svc, config: AvancesConfig):
            captured_configs.append(config)
            result = MagicMock()
            result.ruta_archivo = tmp_path / "output.xlsx"
            result.registros_por_hoja = {}
            return result

        with patch("src.services.avances.service.AvancesService.generar_reporte", fake_generar_reporte):
            main_module._run_avances_report(report, merged)

        assert len(captured_configs) == 1
        assert captured_configs[0].tipo_plantilla == "branca"


# ── dias sheet holidays (feriados) ────────────────────────────────────────────

from datetime import date


def _make_wb_with_dias_sheet():
    """In-memory workbook seeded like the real avance-badie 'dias' sheet.

    B2/B3 hold NETWORKDAYS.INTL formulas referencing the hardcoded $H$2:$H$5
    holiday range. H2..H5 hold four stale June dates (leftovers from a prior
    month). Seeding rows BELOW the new month's holiday count lets tests prove
    that _update_dias_feriados actually CLEARS stale rows, not just that it
    writes the new ones.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "dias"
    ws["B2"] = "=NETWORKDAYS.INTL(F6,F7,11,$H$2:$H$5)"
    ws["B3"] = "=NETWORKDAYS.INTL(F6,F8,11,$H$2:$H$5)"
    ws["H2"] = date(2026, 6, 15)
    ws["H3"] = date(2026, 6, 20)
    ws["H4"] = date(2026, 6, 25)
    ws["H5"] = date(2026, 6, 27)
    return wb


class TestUpdateDiasFeriados:
    def test_populates_current_month_holidays(self):
        """For July 2026 the H column holds the two July holidays and the old
        June dates are gone; formula range is widened to $H$2:$H$10."""
        service = AvancesService(data_loader=MagicMock(spec=DataLoader))
        wb = _make_wb_with_dias_sheet()

        feriados = service._update_dias_feriados(wb, "2026-07-01")

        ws = wb["dias"]
        # First July holiday in H2
        assert ws["H2"].value == date(2026, 7, 9)
        assert ws["H3"].value == date(2026, 7, 10)
        # H4/H5 were seeded with stale June dates; they must be CLEARED because
        # July has only two holidays. This proves stale rows below the new
        # holiday count are wiped (not merely that empty rows stay empty).
        assert ws["H4"].value is None
        assert ws["H5"].value is None
        # Formula range widened so any month's holiday count fits
        assert "$H$2:$H$10" in ws["B2"].value
        assert "$H$2:$H$10" in ws["B3"].value
        # Return value carries (date, motivo) for the applied holidays
        assert any(
            f == date(2026, 7, 9) and "Independencia" in m for f, m in feriados
        )

    def test_no_dias_sheet_returns_empty_and_no_raise(self):
        """A workbook without a 'dias' sheet is a no-op returning []."""
        service = AvancesService(data_loader=MagicMock(spec=DataLoader))
        wb = Workbook()
        wb.active.title = "otra_hoja"

        result = service._update_dias_feriados(wb, "2026-07-01")

        assert result == []

    def test_zero_holiday_month_clears_range_and_returns_empty(self, monkeypatch):
        """A month with no holidays must clear the whole H2:H10 range and return [].

        Every real 2026 Salta month has >=1 holiday, so we monkeypatch
        feriados_del_mes to return [] and confirm the pre-seeded stale dates
        (H2..H5) are wiped and nothing is written back.
        """
        import src.services.avances.service as avances_service

        monkeypatch.setattr(
            avances_service, "feriados_del_mes", lambda anio, mes: []
        )

        service = AvancesService(data_loader=MagicMock(spec=DataLoader))
        wb = _make_wb_with_dias_sheet()  # H2..H5 pre-seeded with stale June dates

        result = service._update_dias_feriados(wb, "2026-07-01")

        ws = wb["dias"]
        # Entire holiday range H2..H10 must be empty.
        for row in range(2, 11):
            assert ws.cell(row=row, column=8).value is None, f"H{row} not cleared"
        assert result == []


class TestAvancesResultFeriadosField:
    def test_result_has_feriados_aplicados_default_empty(self):
        """AvancesResult must expose feriados_aplicados defaulting to []."""
        from src.services.avances.service import AvancesResult

        result = AvancesResult(ruta_archivo=Path("x.xlsx"), registros_por_hoja={})
        assert result.feriados_aplicados == []


class TestNotificarFeriadosConfig:
    """notificar_feriados_a threads from GlobalFilters through merge_filters."""

    def test_global_filters_accepts_notificar_feriados_a(self):
        from src.config.models import GlobalFilters

        gf = GlobalFilters(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            notificar_feriados_a="Nahuel Aguirre",
        )
        assert gf.notificar_feriados_a == "Nahuel Aguirre"

    def test_global_filters_notificar_feriados_a_defaults_none(self):
        from src.config.models import GlobalFilters

        gf = GlobalFilters(fecha_desde="2026-07-01", fecha_hasta="2026-07-31")
        assert gf.notificar_feriados_a is None

    def test_merge_filters_threads_notificar_feriados_a(self):
        from src.config.models import GlobalFilters
        from src.config.resolver import merge_filters

        gf = GlobalFilters(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-31",
            notificar_feriados_a="Nahuel Aguirre",
        )
        merged = merge_filters(gf, None)
        assert merged["notificar_feriados_a"] == "Nahuel Aguirre"


class TestNotificaFeriadosWiring:
    """main._run_avances_report sends the holidays notification, guarded."""

    def test_resolve_target_raw_phone_used_directly(self):
        import main as main_module

        assert main_module._resolve_feriados_target("5493875000000") == "5493875000000"

    def test_resolve_target_contact_name_resolves_to_telefono(self, monkeypatch):
        """A contact NAME resolves to that contact's telefono via load_contacts."""
        import main as main_module
        from types import SimpleNamespace

        fake_contacts = {"Walter Vilte": SimpleNamespace(telefono="5493875111222")}
        monkeypatch.setattr(
            "src.config.resolver.load_contacts", lambda path: fake_contacts
        )

        assert main_module._resolve_feriados_target("Walter Vilte") == "5493875111222"

    def test_resolve_target_unresolvable_name_returns_none(self, monkeypatch):
        """A name absent from the catalog resolves to None (best-effort)."""
        import main as main_module

        monkeypatch.setattr(
            "src.config.resolver.load_contacts", lambda path: {}
        )

        assert main_module._resolve_feriados_target("Ghost Contact") is None

    def _fake_result(self, tmp_path):
        from unittest.mock import MagicMock

        result = MagicMock()
        result.ruta_archivo = tmp_path / "output.xlsx"
        result.registros_por_hoja = {}
        result.feriados_aplicados = [(date(2026, 7, 9), "Día de la Independencia")]
        return result

    def test_notifica_via_whatsapp_cuando_configurado(self, tmp_path):
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "tipo_plantilla": "badie",
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "id_sucursal": 1,
            "id_fuerza_ventas": 1,
            "notificar_feriados_a": "5493875000000",
        }

        sent: dict = {}

        class FakeClient:
            def __init__(self, url):
                sent["url"] = url

            def send_text(self, target="", text="", group_name=None):
                sent["target"] = target
                sent["text"] = text
                return {"success": True}

        fake_result = self._fake_result(tmp_path)

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", FakeClient):
            main_module._run_avances_report(report, merged)

        assert sent.get("target") == "5493875000000"
        assert "Independencia" in sent.get("text", "")

    def test_notificacion_fallida_no_rompe_generacion(self, tmp_path):
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "notificar_feriados_a": "5493875000000",
        }

        class BoomClient:
            def __init__(self, url):
                pass

            def send_text(self, *a, **k):
                raise ConnectionError("no service")

        fake_result = self._fake_result(tmp_path)

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", BoomClient):
            artifacts = main_module._run_avances_report(report, merged)

        # Generation still returns the artifact despite the notification failure.
        assert len(artifacts) == 1

    def test_no_notifica_cuando_no_configurado(self, tmp_path):
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
        }

        fake_result = self._fake_result(tmp_path)
        called = {"n": 0}

        class FakeClient:
            def __init__(self, url):
                called["n"] += 1

            def send_text(self, *a, **k):
                called["n"] += 1

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", FakeClient):
            main_module._run_avances_report(report, merged)

        assert called["n"] == 0

    # ── Fix 1: notification honors --test-mode and --no-delivery ──────────────

    def test_test_mode_redirects_to_test_contact(self, tmp_path):
        """In test mode the notification goes to the TEST_CONTACT (Nahuel),
        never to the configured real supervisor."""
        import main as main_module
        from src.config.resolver import TEST_CONTACT_NAME
        from src.config.models import ReportEntry
        from unittest.mock import patch
        from types import SimpleNamespace

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "notificar_feriados_a": "Real Supervisor Name",
            "test_mode": True,
        }

        # Test contact and real contact resolve to DISTINCT phones so we can
        # assert the redirect actually happened.
        fake_contacts = {
            TEST_CONTACT_NAME: SimpleNamespace(telefono="5493875000001"),
            "Real Supervisor Name": SimpleNamespace(telefono="5493875999999"),
        }

        sent: dict = {}

        class FakeClient:
            def __init__(self, url):
                pass

            def send_text(self, target="", text="", group_name=None):
                sent["target"] = target
                return {"success": True}

        fake_result = self._fake_result(tmp_path)

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", FakeClient), patch(
            "src.config.resolver.load_contacts", lambda path: fake_contacts
        ):
            main_module._run_avances_report(report, merged)

        assert sent.get("target") == "5493875000001"  # test contact (Nahuel)
        assert sent.get("target") != "5493875999999"  # NOT the real supervisor

    def test_no_delivery_suppresses_notification(self, tmp_path):
        """A --no-delivery run never calls send_text."""
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "notificar_feriados_a": "X",
            "no_delivery": True,
        }

        called = {"n": 0}

        class FakeClient:
            def __init__(self, url):
                pass

            def send_text(self, *a, **k):
                called["n"] += 1

        fake_result = self._fake_result(tmp_path)

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", FakeClient):
            main_module._run_avances_report(report, merged)

        assert called["n"] == 0

    def test_normal_run_sends_to_resolved_contact_phone(self, tmp_path):
        """Without test_mode/no_delivery the notification sends to the
        configured contact's resolved phone."""
        import main as main_module
        from src.config.models import ReportEntry
        from unittest.mock import patch
        from types import SimpleNamespace

        report = ReportEntry(nombre="AVANCE BADIE - JULIO 2026")
        merged = {
            "fecha_desde": "2026-07-01",
            "fecha_hasta": "2026-07-31",
            "notificar_feriados_a": "X",
        }

        fake_contacts = {"X": SimpleNamespace(telefono="5493875123456")}

        sent: dict = {}

        class FakeClient:
            def __init__(self, url):
                pass

            def send_text(self, target="", text="", group_name=None):
                sent["target"] = target
                return {"success": True}

        fake_result = self._fake_result(tmp_path)

        with patch(
            "src.services.avances.service.AvancesService.generar_reporte",
            lambda self_svc, config: fake_result,
        ), patch("src.core.whatsapp_client.WhatsAppClient", FakeClient), patch(
            "src.config.resolver.load_contacts", lambda path: fake_contacts
        ):
            main_module._run_avances_report(report, merged)

        assert sent.get("target") == "5493875123456"


class TestResolveBasePicksLatestNonBackup:
    """_resolve_base must seed from the previous month's REAL latest output,
    never a stale '.bak'/'_backup' copy — even when the backup sorts first
    alphabetically ('.bak' < '.xlsx')."""

    def _config(self):
        return AvancesConfig(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-08",
            nombre_archivo="AVANCE BRANCA - JULIO 2026",
        )

    def test_ignores_bak_and_picks_most_recent_real_file(self, tmp_path, monkeypatch):
        import os
        import time
        import src.services.avances.service as svc_mod

        prev_dir = tmp_path / "2026-06"
        prev_dir.mkdir()
        bak = prev_dir / "AVANCE BRANCA - JUNIO 2026.bak.20260606-1423.xlsx"
        real = prev_dir / "AVANCE BRANCA - JUNIO 2026.xlsx"
        mayo = prev_dir / "AVANCE BRANCA - MAYO 2026.xlsx"
        for p in (bak, real, mayo):
            p.write_bytes(b"x")
        old = time.time() - 100_000
        os.utime(bak, (old, old))
        os.utime(mayo, (old, old))  # `real` is the newest

        monkeypatch.setattr(svc_mod, "service_output_dir", lambda *a, **k: prev_dir)

        service = AvancesService(data_loader=MagicMock())
        base = service._resolve_base(self._config(), tmp_path / "2026-07")
        assert base == real, f"picked {base.name!r}, expected the real non-backup June file"

    def test_excludes_underscore_backup_and_misnamed_backup(self, tmp_path, monkeypatch):
        import src.services.avances.service as svc_mod

        prev_dir = tmp_path / "2026-06"
        prev_dir.mkdir()
        real = prev_dir / "AVANCE BRANCA - JUNIO 2026.xlsx"
        bkp = prev_dir / "AVANCE BRANCA - JUNIO 2026_backup-20260704.xlsx"
        mis = prev_dir / "AVANCE BRANCA - JUNIO 2026_misnamed-backup-20260706.xlsx"
        for p in (real, bkp, mis):
            p.write_bytes(b"x")

        monkeypatch.setattr(svc_mod, "service_output_dir", lambda *a, **k: prev_dir)

        service = AvancesService(data_loader=MagicMock())
        base = service._resolve_base(self._config(), tmp_path / "2026-07")
        assert base == real


class TestResolveBaseFirstRunNoSiblingFallback:
    """RF-06: when name_prefix has no match in the previous-month dir (a new
    report's first run, sharing the output dir with badie/branca siblings),
    _resolve_base must NOT fall back to a sibling report's file — it must
    proceed to archivo_plantilla."""

    def test_no_prefix_match_falls_through_to_archivo_plantilla(self, tmp_path, monkeypatch):
        import src.services.avances.service as svc_mod

        prev_dir = tmp_path / "2026-06"
        prev_dir.mkdir()
        # Only sibling reports' outputs exist — nothing matches "AVANCE GUEMES"
        badie = prev_dir / "AVANCE BADIE - JUNIO 2026.xlsx"
        branca = prev_dir / "AVANCE BRANCA - JUNIO 2026.xlsx"
        for p in (badie, branca):
            p.write_bytes(b"x")

        monkeypatch.setattr(svc_mod, "service_output_dir", lambda *a, **k: prev_dir)

        plantilla = tmp_path / "AVANCE GUEMES.xlsx"
        plantilla.write_bytes(b"template")

        service = AvancesService(data_loader=MagicMock())
        config = AvancesConfig(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-08",
            archivo_plantilla=str(plantilla),
            nombre_archivo="AVANCE GUEMES - JULIO 2026",
        )
        base = service._resolve_base(config, tmp_path / "2026-07")

        assert base == plantilla, (
            f"expected archivo_plantilla fallback, got sibling file {base!r}"
        )

    def test_badie_prefix_match_unaffected_regression(self, tmp_path, monkeypatch):
        """Regression: badie resolution with a matching prefix in the prev
        month must remain exactly as before this fix."""
        import src.services.avances.service as svc_mod

        prev_dir = tmp_path / "2026-06"
        prev_dir.mkdir()
        badie = prev_dir / "AVANCE BADIE - JUNIO 2026.xlsx"
        branca = prev_dir / "AVANCE BRANCA - JUNIO 2026.xlsx"
        for p in (badie, branca):
            p.write_bytes(b"x")

        monkeypatch.setattr(svc_mod, "service_output_dir", lambda *a, **k: prev_dir)

        service = AvancesService(data_loader=MagicMock())
        config = AvancesConfig(
            fecha_desde="2026-07-01",
            fecha_hasta="2026-07-08",
            nombre_archivo="AVANCE BADIE - JULIO 2026",
        )
        base = service._resolve_base(config, tmp_path / "2026-07")

        assert base == badie
