"""Tests for the cobertura-cupos report.

The load-bearing rule here: the generic's TOTAL row is a SEPARATE query against
the generic grain, never the sum of the marca rows above it. Coverage counts
DISTINCT clients and one client buys several marcas — summing the rows would
count them once per marca.
"""
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.services.cobertura_cupos import (
    CoberturaCuposConfig,
    CoberturaCuposService,
    GENERICOS_CCU,
    Zona,
    zonas_por_defecto,
)

# Real shape: PERNOD julio-2026 son 721 clientes, pero sus marcas suman 1083.
MARCAS_ACTUAL = pd.DataFrame({
    "marca": ["ABSOLUT", "CHIVAS", "SOMETHING"],
    "cobertura": [500, 400, 183],
})
TOTAL_REAL_ACTUAL = 721
MARCAS_ANTERIOR = pd.DataFrame({
    "marca": ["ABSOLUT", "OTRA VIEJA"],
    "cobertura": [300, 90],
})
TOTAL_REAL_ANTERIOR = 350

# El maestro tiene una marca mas que las que vendieron: SIN VENTAS nunca tuvo
# cobertura y aun asi debe figurar, con 0, para poder asignarle cupo.
UNIVERSO = pd.DataFrame({
    "marca": ["ABSOLUT", "CHIVAS", "OTRA VIEJA", "SIN VENTAS", "SOMETHING"],
})


def _loader(marcas_actual=MARCAS_ACTUAL, marcas_anterior=MARCAS_ANTERIOR,
            universo=UNIVERSO):
    loader = MagicMock(spec=DataLoader)

    def marcas(generico, periodo, **kw):
        return marcas_actual if periodo.startswith("2026") else marcas_anterior

    def total(generico, periodo, **kw):
        return TOTAL_REAL_ACTUAL if periodo.startswith("2026") else TOTAL_REAL_ANTERIOR

    loader.get_cobertura_marca_de_generico_zona.side_effect = marcas
    loader.get_cobertura_generico_zona.side_effect = total
    loader.get_marcas_de_generico.return_value = universo
    return loader


def _run(tmp_path, loader=None, **kwargs):
    loader = loader or _loader()
    kwargs.setdefault("genericos", ["PERNOD RICARD"])
    kwargs.setdefault("zonas", [Zona("CASA CENTRAL", 1)])
    config = CoberturaCuposConfig(fecha_desde="2026-08-04", **kwargs)
    with patch("src.services.cobertura_cupos.service.service_output_dir",
               return_value=tmp_path):
        return CoberturaCuposService(data_loader=loader).generar_reporte(config), loader


def _filas(ruta, hoja=None):
    wb = load_workbook(ruta)
    ws = wb[hoja] if hoja else wb.worksheets[0]
    return [[ws.cell(r, c).value for c in range(1, 5)] for r in range(1, ws.max_row + 1)]


# ---------------------------------------------------------------------------
# Períodos: derivados de fecha_desde, nunca escritos a mano
# ---------------------------------------------------------------------------


class TestPeriodos:
    def test_periodo_actual_es_el_mes_anterior_cerrado(self):
        cfg = CoberturaCuposConfig(fecha_desde="2026-08-04")
        assert cfg.periodos[0] == "2026-07-01"

    def test_segunda_columna_es_el_mismo_mes_del_anio_anterior(self):
        """El mes EN CURSO un anio atras (agosto-2025), no el interanual del mes
        cerrado (julio-2025): el cupo se reparte contra el agosto que paso."""
        cfg = CoberturaCuposConfig(fecha_desde="2026-08-04")
        assert cfg.periodos == ["2026-07-01", "2025-08-01"]

    def test_derivacion_sigue_a_fecha_desde_en_el_rollover_de_anio(self):
        cfg = CoberturaCuposConfig(fecha_desde="2026-01-15")
        assert cfg.periodos == ["2025-12-01", "2025-01-01"]

    def test_nombre_archivo_default_lleva_el_mes_del_dato(self):
        cfg = CoberturaCuposConfig(fecha_desde="2026-08-04")
        assert "JULIO 2026" in cfg.nombre_archivo


# ---------------------------------------------------------------------------
# Defaults de config
# ---------------------------------------------------------------------------


class TestDefaults:
    def test_genericos_default_son_los_cinco_ccu(self):
        cfg = CoberturaCuposConfig(fecha_desde="2026-08-04")
        assert cfg.genericos == list(GENERICOS_CCU)
        assert "FRATELLI B" not in cfg.genericos

    def test_zonas_default_son_las_tres_pedidas(self):
        cfg = CoberturaCuposConfig(fecha_desde="2026-08-04")
        assert [z.nombre for z in cfg.zonas] == [
            "CASA CENTRAL", "VALLE SALTA", "SUCURSAL GUEMES",
        ]

    def test_casa_central_excluye_las_rutas_de_valle_salta(self):
        """Las zonas son DISJUNTAS: sin la resta, los clientes de VALLE SALTA
        se contarían dos veces entre las hojas."""
        casa, valle, _ = zonas_por_defecto()
        assert casa.rutas_excluidas, "CASA CENTRAL debe restar sus zonas virtuales"
        assert set(valle.rutas_incluidas) <= set(casa.rutas_excluidas)

    def test_guemes_es_la_sucursal_entera(self):
        _, _, guemes = zonas_por_defecto()
        assert guemes.id_sucursal == 16
        assert guemes.rutas_incluidas is None


# ---------------------------------------------------------------------------
# LA regla: el total del genérico no es la suma de sus marcas
# ---------------------------------------------------------------------------


class TestTotalNoEsSumaDeMarcas:
    def test_total_viene_de_su_propia_consulta(self, tmp_path):
        result, loader = _run(tmp_path)
        assert loader.get_cobertura_generico_zona.called, (
            "el total tiene que salir del grano de genérico, no de sumar marcas"
        )
        filas = _filas(result.ruta_archivo)
        total = next(f for f in filas if str(f[0]).startswith("TOTAL "))
        assert total[1] == TOTAL_REAL_ACTUAL

    def test_total_difiere_de_la_suma_de_las_filas(self, tmp_path):
        """500+400+183 = 1083 contra los 721 reales. Si el informe mostrara
        1083 estaría contando al mismo cliente una vez por marca."""
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        marcas = [f for f in filas if f[0] in list(MARCAS_ACTUAL["marca"])]
        suma_filas = sum(f[1] for f in marcas)
        total = next(f for f in filas if str(f[0]).startswith("TOTAL "))
        assert suma_filas == 1083
        assert total[1] == 721
        assert total[1] != suma_filas

    def test_total_del_periodo_anterior_tambien_es_consulta_aparte(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        total = next(f for f in filas if str(f[0]).startswith("TOTAL "))
        assert total[2] == TOTAL_REAL_ANTERIOR
        assert total[2] != 300 + 90

    def test_una_fila_total_por_generico(self, tmp_path):
        result, _ = _run(tmp_path, genericos=["CERVEZAS", "VINOS CCU"])
        filas = _filas(result.ruta_archivo)
        totales = [f[0] for f in filas if str(f[0]).startswith("TOTAL ")]
        assert totales == ["TOTAL CERVEZAS", "TOTAL VINOS CCU"]


# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------


class TestLayout:
    def test_encabezados_nombran_los_dos_meses_y_el_cupo(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        header = next(f for f in filas if f[0] == "Marca")
        assert header == ["Marca", "Cob JULIO 2026", "Cob AGOSTO 2025", "Cupo"]

    def test_columna_cupo_queda_vacia(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        marcas = [f for f in filas if f[0] in list(MARCAS_ACTUAL["marca"])]
        assert marcas, "deberia haber filas de marca"
        assert all(f[3] is None for f in marcas), "el cupo se carga a mano"

    def test_marcas_ordenadas_por_cobertura_descendente(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        marcas = [f for f in filas if f[0] in list(MARCAS_ACTUAL["marca"])]
        assert [f[0] for f in marcas] == ["ABSOLUT", "CHIVAS", "SOMETHING"]

    def test_marca_solo_del_anio_anterior_aparece_con_cero(self, tmp_path):
        """Outer join: una baja de portafolio es justo lo que hay que ver."""
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        vieja = next(f for f in filas if f[0] == "OTRA VIEJA")
        assert vieja[1] == 0
        assert vieja[2] == 90

    def test_marca_nueva_aparece_con_cero_en_el_anio_anterior(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        nueva = next(f for f in filas if f[0] == "CHIVAS")
        assert nueva[2] == 0

    def test_una_hoja_por_zona(self, tmp_path):
        result, _ = _run(
            tmp_path,
            zonas=[Zona("CASA CENTRAL", 1), Zona("VALLE SALTA", 1, rutas_incluidas=(81, 82))],
        )
        wb = load_workbook(result.ruta_archivo)
        assert wb.sheetnames == ["CASA CENTRAL", "VALLE SALTA"]
        assert result.zonas == ["CASA CENTRAL", "VALLE SALTA"]

    def test_aparecen_todas_las_marcas_del_generico(self, tmp_path):
        """El eje de filas es el maestro, no las marcas con cobertura: hay que
        poder asignarle cupo a una marca que hoy no vende en esa zona."""
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        en_hoja = {f[0] for f in filas}
        assert set(UNIVERSO["marca"]) <= en_hoja

    def test_marca_sin_cobertura_aparece_en_cero_y_con_cupo_editable(self, tmp_path):
        result, _ = _run(tmp_path)
        filas = _filas(result.ruta_archivo)
        sin_ventas = next(f for f in filas if f[0] == "SIN VENTAS")
        assert sin_ventas[1] == 0 and sin_ventas[2] == 0
        assert sin_ventas[3] is None

    def test_las_tres_zonas_traen_las_mismas_marcas(self, tmp_path):
        """Filas comparables entre hojas: sin esto no se pueden leer en paralelo."""
        result, _ = _run(
            tmp_path,
            zonas=[Zona("CASA CENTRAL", 1), Zona("SUCURSAL GUEMES", 16)],
        )
        marcas_por_hoja = [
            {f[0] for f in _filas(result.ruta_archivo, hoja) if f[0] in set(UNIVERSO["marca"])}
            for hoja in ("CASA CENTRAL", "SUCURSAL GUEMES")
        ]
        assert marcas_por_hoja[0] == marcas_por_hoja[1] == set(UNIVERSO["marca"])

    def test_marca_con_cobertura_fuera_del_maestro_no_se_pierde(self, tmp_path):
        """Si el maestro no la tiene pero vendió, igual tiene que figurar."""
        universo_corto = pd.DataFrame({"marca": ["ABSOLUT"]})
        result, _ = _run(tmp_path, loader=_loader(universo=universo_corto))
        filas = _filas(result.ruta_archivo)
        assert any(f[0] == "ABSOLUT" for f in filas)

    def test_la_hoja_avisa_que_el_total_no_es_la_suma(self, tmp_path):
        result, _ = _run(tmp_path)
        ws = load_workbook(result.ruta_archivo).worksheets[0]
        assert "NO es la suma" in ws["A2"].value


# ---------------------------------------------------------------------------
# Wiring de zonas hacia el loader — la clave compuesta
# ---------------------------------------------------------------------------


class TestZonaLlegaAlLoader:
    def test_rutas_incluidas_llegan_junto_con_la_sucursal(self, tmp_path):
        """id_ruta se reusa entre sucursales: filtrar solo por ruta produce
        fan-out con clientes de otras sucursales."""
        _, loader = _run(
            tmp_path, zonas=[Zona("VALLE SALTA", 1, rutas_incluidas=(81, 82, 83))]
        )
        _, kwargs = loader.get_cobertura_marca_de_generico_zona.call_args
        assert kwargs["id_sucursal"] == 1
        assert kwargs["rutas_incluidas"] == [81, 82, 83]

    def test_rutas_excluidas_llegan_al_loader(self, tmp_path):
        _, loader = _run(
            tmp_path, zonas=[Zona("CASA CENTRAL", 1, rutas_excluidas=(81, 93))]
        )
        _, kwargs = loader.get_cobertura_generico_zona.call_args
        assert kwargs["rutas_excluidas"] == [81, 93]
        assert kwargs["rutas_incluidas"] is None

    def test_los_dos_periodos_se_consultan(self, tmp_path):
        _, loader = _run(tmp_path)
        periodos = {
            c.kwargs["periodo"]
            for c in loader.get_cobertura_marca_de_generico_zona.call_args_list
        }
        assert periodos == {"2026-07-01", "2025-08-01"}


# ---------------------------------------------------------------------------
# Bordes
# ---------------------------------------------------------------------------


class TestBordes:
    def test_generico_sin_datos_no_rompe(self, tmp_path):
        """PERNOD no existe en julio-2025: la columna del año anterior da 0."""
        loader = _loader(marcas_anterior=pd.DataFrame(columns=["marca", "cobertura"]))
        loader.get_cobertura_generico_zona.side_effect = (
            lambda generico, periodo, **kw: 721 if periodo.startswith("2026") else 0
        )
        result, _ = _run(tmp_path, loader=loader)
        filas = _filas(result.ruta_archivo)
        total = next(f for f in filas if str(f[0]).startswith("TOTAL "))
        assert total[1] == 721
        assert total[2] == 0

    def test_zona_sin_cobertura_igual_lista_todas_las_marcas_en_cero(self, tmp_path):
        """Una zona que no vendió nada del genérico no queda en blanco: muestra
        el universo completo en 0, con su casillero de cupo listo."""
        vacio = pd.DataFrame(columns=["marca", "cobertura"])
        loader = _loader(marcas_actual=vacio, marcas_anterior=vacio)
        loader.get_cobertura_generico_zona.side_effect = lambda **kw: 0
        result, _ = _run(tmp_path, loader=loader)
        assert result.ruta_archivo.exists()
        assert result.filas_marca == len(UNIVERSO)
        filas = _filas(result.ruta_archivo)
        marcas = [f for f in filas if f[0] in set(UNIVERSO["marca"])]
        assert all(f[1] == 0 and f[2] == 0 for f in marcas)

    def test_sin_universo_ni_cobertura_no_rompe(self, tmp_path):
        vacio = pd.DataFrame(columns=["marca", "cobertura"])
        loader = _loader(
            marcas_actual=vacio, marcas_anterior=vacio,
            universo=pd.DataFrame(columns=["marca"]),
        )
        loader.get_cobertura_generico_zona.side_effect = lambda **kw: 0
        result, _ = _run(tmp_path, loader=loader)
        assert result.ruta_archivo.exists()
        assert result.filas_marca == 0

    def test_nombre_de_hoja_se_recorta_al_limite_de_excel(self, tmp_path):
        largo = "ZONA CON UN NOMBRE EXCESIVAMENTE LARGO QUE NO ENTRA"
        result, _ = _run(tmp_path, zonas=[Zona(largo, 1)])
        assert load_workbook(result.ruta_archivo).sheetnames == [largo[:31]]


# ---------------------------------------------------------------------------
# El filtro de rutas del DataLoader
# ---------------------------------------------------------------------------


class TestFiltroRutas:
    def _dl(self):
        return DataLoader(engine=MagicMock())

    def test_sin_rutas_no_agrega_filtro(self):
        sql, params = self._dl()._filtro_rutas("cpm", None, None)
        assert sql == ""
        assert params == {}

    def test_incluidas_y_excluidas_se_combinan(self):
        sql, params = self._dl()._filtro_rutas("cpm", [81, 82], [93])
        assert "cpm.id_ruta = ANY(:rutas_in)" in sql
        assert "NOT (cpm.id_ruta = ANY(:rutas_out))" in sql
        assert params == {"rutas_in": [81, 82], "rutas_out": [93]}

    def test_lista_incluida_vacia_no_devuelve_la_sucursal_entera(self):
        """Degradar a 'todas' devolvería la sucursal disfrazada de zona."""
        sql, params = self._dl()._filtro_rutas("cpm", [], None)
        assert sql == "AND FALSE"
        assert params == {}
