"""Tests for the cobertura service — histórico de cobertura por período.

Cobertura = DISTINCT clients. It is additive across rutas/preventistas (a client
belongs to exactly one ruta) but NEVER across marcas, genéricos or periods. The
pivot keeps marca/generico in the index and the period in the columns, so its
``aggfunc="sum"`` can only ever collapse rows that differ by preventista — the
one axis where the sum is legitimate.
"""
import json
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.core.periodos import periodo_meses_atras
from src.services.base_service import BaseService
from src.services.cobertura import (
    TIPO_PREVENTISTA_GENERICO,
    TIPO_PREVENTISTA_MARCA,
    TIPO_SUCURSAL_MARCA,
    CoberturaService,
    ReporteCoberturaConfig,
)
from src.services.cobertura.processor import (
    SIN_DATO,
    SIN_DATO_NUMERICO,
    procesar_cobertura,
)

INDEX_GENERICO = ["sucursal", "vendedor", "id_ruta", "generico"]


# ---------------------------------------------------------------------------
# periodo_meses_atras — period derivation helper
# ---------------------------------------------------------------------------


class TestPeriodoMesesAtras:
    def test_one_month_back(self):
        assert periodo_meses_atras("2026-08-04", 1) == "2026-07-01"

    def test_thirteen_months_back_is_same_month_previous_year(self):
        assert periodo_meses_atras("2026-08-04", 13) == "2025-07-01"

    def test_zero_is_own_month(self):
        assert periodo_meses_atras("2026-08-04", 0) == "2026-08-01"

    def test_year_rollover(self):
        assert periodo_meses_atras("2026-01-15", 1) == "2025-12-01"

    def test_twelve_months_back_is_same_month(self):
        assert periodo_meses_atras("2026-03-31", 12) == "2025-03-01"

    def test_negative_raises(self):
        with pytest.raises(ValueError, match="meses"):
            periodo_meses_atras("2026-08-04", -1)

    def test_invalid_date_raises(self):
        with pytest.raises(ValueError):
            periodo_meses_atras("04-08-2026", 1)


# ---------------------------------------------------------------------------
# ReporteCoberturaConfig — periods DERIVED from fecha_desde, never hardcoded
# ---------------------------------------------------------------------------


class TestConfigPeriodDerivation:
    def test_default_compares_previous_month_against_same_month_last_year(self):
        cfg = ReporteCoberturaConfig(fecha_desde="2026-08-04")
        assert cfg.periodos == ["2025-07-01", "2026-07-01"]

    def test_periods_come_out_chronological_regardless_of_input_order(self):
        cfg = ReporteCoberturaConfig(fecha_desde="2026-08-04", meses_atras=[1, 13])
        assert cfg.periodos == ["2025-07-01", "2026-07-01"]

    def test_custom_offsets(self):
        cfg = ReporteCoberturaConfig(fecha_desde="2026-08-04", meses_atras=[1])
        assert cfg.periodos == ["2026-07-01"]

    def test_duplicate_offsets_collapse(self):
        """A repeated period would make pivot_table sum a month onto itself."""
        cfg = ReporteCoberturaConfig(fecha_desde="2026-08-04", meses_atras=[1, 1, 13])
        assert cfg.periodos == ["2025-07-01", "2026-07-01"]

    def test_derivation_follows_fecha_desde_across_month_rollover(self):
        """The daily patches fecha_desde only — the window must follow it alone."""
        cfg = ReporteCoberturaConfig(fecha_desde="2026-01-09")
        assert cfg.periodos == ["2024-12-01", "2025-12-01"]

    def test_explicit_periodos_override_derivation(self):
        cfg = ReporteCoberturaConfig(
            fecha_desde="2026-08-04", periodos=["2025-02-01", "2026-01-01"]
        )
        assert cfg.periodos == ["2025-02-01", "2026-01-01"]

    def test_empty_meses_atras_raises(self):
        with pytest.raises(ValueError, match="meses_atras"):
            ReporteCoberturaConfig(fecha_desde="2026-08-04", meses_atras=[])

    def test_invalid_tipo_raises(self):
        with pytest.raises(ValueError, match="tipo"):
            ReporteCoberturaConfig(fecha_desde="2026-08-04", tipo="no_existe")

    def test_default_nombre_archivo_includes_tipo(self):
        cfg = ReporteCoberturaConfig(fecha_desde="2026-08-04", tipo=TIPO_SUCURSAL_MARCA)
        assert cfg.tipo in cfg.nombre_archivo


# ---------------------------------------------------------------------------
# procesar_cobertura — the pivot
# ---------------------------------------------------------------------------


def _raw_dos_periodos():
    """Same preventista/generico measured in two periods."""
    return pd.DataFrame({
        "periodo": [date(2025, 2, 1), date(2026, 1, 1)],
        "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
        "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY"],
        "id_ruta": [81, 81],
        "generico": ["CERVEZAS", "CERVEZAS"],
        "clientes_compradores": [40, 55],
    })


class TestProcesarCobertura:
    def test_periods_become_columns(self):
        out = procesar_cobertura(_raw_dos_periodos(), INDEX_GENERICO)
        assert len(out) == 1
        assert out.iloc[0]["Feb 2025"] == 40
        assert out.iloc[0]["Ene 2026"] == 55

    def test_column_labels_are_short_month_year(self):
        out = procesar_cobertura(_raw_dos_periodos(), INDEX_GENERICO)
        assert list(out.columns) == INDEX_GENERICO + ["Feb 2025", "Ene 2026"]

    def test_columns_are_chronological_not_alphabetical(self):
        """'Ene' sorts before 'Feb' alphabetically — chronologically it does not.

        Reading two periods side by side only works if the older one is on the
        left; an alphabetical sort silently swaps them.
        """
        out = procesar_cobertura(_raw_dos_periodos(), INDEX_GENERICO)
        periodos = [c for c in out.columns if c not in INDEX_GENERICO]
        assert periodos == ["Feb 2025", "Ene 2026"]

    def test_row_with_null_index_value_is_not_dropped(self):
        """dim_vendedor is LEFT JOINed: a miss leaves vendedor NULL.

        pivot_table drops rows whose index key is NaN, so that cobertura would
        vanish from the report with no error at all.
        """
        raw = _raw_dos_periodos()
        raw.loc[len(raw)] = [date(2026, 1, 1), "CASA CENTRAL", None, 99, "CERVEZAS", 7]
        out = procesar_cobertura(raw, INDEX_GENERICO)
        assert len(out) == 2
        assert out[["Feb 2025", "Ene 2026"]].to_numpy().sum() == 40 + 55 + 7

    def test_null_text_key_gets_the_text_sentinel(self):
        """Desde pandas 3.0 las columnas de texto son `str`, no `object`: elegir
        el relleno por `dtype == object` manda los nombres faltantes al -1."""
        raw = _raw_dos_periodos()
        raw.loc[len(raw)] = [date(2026, 1, 1), "CASA CENTRAL", None, 99, "CERVEZAS", 7]
        out = procesar_cobertura(raw, INDEX_GENERICO)
        assert SIN_DATO in out["vendedor"].tolist()

    def test_null_numeric_key_gets_the_numeric_sentinel(self):
        raw = pd.DataFrame({
            "periodo": [date(2026, 1, 1), date(2026, 1, 1)],
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "vendedor": ["LORENA TARITOLAY", "OTRO"],
            "id_ruta": pd.Series([81, None], dtype="float64"),
            "generico": ["CERVEZAS", "CERVEZAS"],
            "clientes_compradores": [40, 7],
        })
        out = procesar_cobertura(raw, INDEX_GENERICO)
        assert SIN_DATO_NUMERICO in out["id_ruta"].tolist()

    def test_does_not_sum_across_marcas(self):
        """Two marcas on one ruta must stay on two rows — a client buying both
        would otherwise be counted twice in a single cell."""
        raw = pd.DataFrame({
            "periodo": [date(2026, 1, 1), date(2026, 1, 1)],
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY"],
            "id_ruta": [81, 81],
            "marca": ["SALTA", "SCHNEIDER"],
            "clientes_compradores": [30, 25],
        })
        index_marca = ["sucursal", "vendedor", "id_ruta", "marca"]
        out = procesar_cobertura(raw, index_marca)
        assert len(out) == 2
        assert sorted(out["Ene 2026"].tolist()) == [25, 30]

    def test_sums_across_preventistas_of_the_same_ruta(self):
        """Additive axis: each client belongs to one ruta, so two preventistas
        rotating on that ruta within a period do not share clients."""
        raw = pd.DataFrame({
            "periodo": [date(2026, 1, 1), date(2026, 1, 1)],
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY"],
            "id_ruta": [81, 81],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "clientes_compradores": [30, 25],
        })
        out = procesar_cobertura(raw, INDEX_GENERICO)
        assert len(out) == 1
        assert out.iloc[0]["Ene 2026"] == 55

    def test_missing_period_fills_zero_not_blank(self):
        raw = _raw_dos_periodos()
        raw.loc[len(raw)] = [date(2026, 1, 1), "SUCURSAL PERICO", "OTRO", 5, "CERVEZAS", 9]
        out = procesar_cobertura(raw, INDEX_GENERICO)
        perico = out[out["sucursal"] == "SUCURSAL PERICO"].iloc[0]
        assert perico["Feb 2025"] == 0

    def test_empty_dataframe_returns_empty(self):
        empty = pd.DataFrame(columns=["periodo", "sucursal", "clientes_compradores"])
        assert procesar_cobertura(empty, ["sucursal"]).empty


# ---------------------------------------------------------------------------
# CoberturaService — loader wiring, output path, totals row
# ---------------------------------------------------------------------------


def _raw_por_tipo(tipo: str) -> pd.DataFrame:
    base = {
        "periodo": [date(2025, 7, 1), date(2026, 7, 1)],
        "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
        "clientes_compradores": [40, 55],
    }
    if tipo == TIPO_SUCURSAL_MARCA:
        return pd.DataFrame({**base, "marca": ["SALTA", "SALTA"]})
    apertura = "generico" if tipo == TIPO_PREVENTISTA_GENERICO else "marca"
    valor = "CERVEZAS" if apertura == "generico" else "SALTA"
    return pd.DataFrame({
        **base,
        "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY"],
        "id_ruta": [81, 81],
        apertura: [valor, valor],
    })


def _run(tmp_path, tipo, **kwargs):
    loader = MagicMock(spec=DataLoader)
    for metodo in (
        "get_cobertura_preventista_generico",
        "get_cobertura_preventista_marca",
        "get_cobertura_sucursal_marca",
    ):
        getattr(loader, metodo).return_value = _raw_por_tipo(tipo)
    config = ReporteCoberturaConfig(
        fecha_desde="2026-08-04", tipo=tipo, con_slicers=False, **kwargs
    )
    with patch("src.services.cobertura.service.service_output_dir", return_value=tmp_path):
        result = CoberturaService(data_loader=loader).generar_reporte(config)
    return result, loader


LOADER_POR_TIPO = {
    TIPO_PREVENTISTA_GENERICO: "get_cobertura_preventista_generico",
    TIPO_PREVENTISTA_MARCA: "get_cobertura_preventista_marca",
    TIPO_SUCURSAL_MARCA: "get_cobertura_sucursal_marca",
}


class TestCoberturaService:
    @pytest.mark.parametrize("tipo", list(LOADER_POR_TIPO))
    def test_each_tipo_calls_its_own_loader(self, tmp_path, tipo):
        _, loader = _run(tmp_path, tipo)
        for otro_tipo, metodo in LOADER_POR_TIPO.items():
            llamado = getattr(loader, metodo).called
            assert llamado is (otro_tipo == tipo)

    @pytest.mark.parametrize("tipo", list(LOADER_POR_TIPO))
    def test_derived_periods_reach_the_loader(self, tmp_path, tipo):
        _, loader = _run(tmp_path, tipo)
        _, kwargs = getattr(loader, LOADER_POR_TIPO[tipo]).call_args
        assert kwargs["periodos"] == ["2025-07-01", "2026-07-01"]

    @pytest.mark.parametrize("tipo", list(LOADER_POR_TIPO))
    def test_file_is_written(self, tmp_path, tipo):
        result, _ = _run(tmp_path, tipo)
        assert result.ruta_archivo.exists()
        assert result.ruta_archivo.parent == tmp_path
        assert result.tipo == tipo

    def test_sucursales_filter_reaches_the_loader(self, tmp_path):
        _, loader = _run(tmp_path, TIPO_SUCURSAL_MARCA, sucursales=["CASA CENTRAL"])
        _, kwargs = loader.get_cobertura_sucursal_marca.call_args
        assert kwargs["sucursales"] == ["CASA CENTRAL"]

    def test_output_dir_uses_service_slug_and_month_of_fecha_desde(self, tmp_path):
        loader = MagicMock(spec=DataLoader)
        loader.get_cobertura_sucursal_marca.return_value = _raw_por_tipo(TIPO_SUCURSAL_MARCA)
        config = ReporteCoberturaConfig(
            fecha_desde="2026-08-04", tipo=TIPO_SUCURSAL_MARCA, con_slicers=False
        )
        with patch("src.services.cobertura.service.service_output_dir",
                   return_value=tmp_path) as spy:
            CoberturaService(data_loader=loader).generar_reporte(config)
        assert spy.call_args.args[0] == "cobertura"
        assert spy.call_args.args[1] == "2026-08-04"
        assert spy.call_args.kwargs["granularity"] == "month"

    def test_empty_result_does_not_crash(self, tmp_path):
        loader = MagicMock(spec=DataLoader)
        loader.get_cobertura_sucursal_marca.return_value = pd.DataFrame(
            columns=["periodo", "sucursal", "marca", "clientes_compradores"]
        )
        config = ReporteCoberturaConfig(
            fecha_desde="2026-08-04", tipo=TIPO_SUCURSAL_MARCA, con_slicers=False
        )
        with patch("src.services.cobertura.service.service_output_dir", return_value=tmp_path):
            result = CoberturaService(data_loader=loader).generar_reporte(config)
        assert result.ruta_archivo.exists()
        assert result.registros_procesados == 0


class TestTotalGeneralRow:
    def _rows(self, ruta):
        ws = load_workbook(ruta).active
        return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]

    def test_report_has_a_total_general_row(self, tmp_path):
        result, _ = _run(tmp_path, TIPO_PREVENTISTA_GENERICO)
        rows = self._rows(result.ruta_archivo)
        assert any(r[0] == "TOTAL GENERAL" for r in rows)

    def test_total_sums_each_period_column(self, tmp_path):
        result, _ = _run(tmp_path, TIPO_PREVENTISTA_GENERICO)
        rows = self._rows(result.ruta_archivo)
        total = next(r for r in rows if r[0] == "TOTAL GENERAL")
        # 4 index columns, then one cell per period.
        assert total[4:6] == [40, 55]

    def test_total_row_warns_it_is_not_distinct_clients(self, tmp_path):
        """The grand total crosses genéricos/marcas, so it is a sum of coverages,
        not a client count. The row has to say so or it will be misread."""
        result, _ = _run(tmp_path, TIPO_PREVENTISTA_GENERICO)
        rows = self._rows(result.ruta_archivo)
        total = next(r for r in rows if r[0] == "TOTAL GENERAL")
        assert "no son clientes" in str(total[1]).lower()

    def test_total_row_is_outside_the_excel_table(self, tmp_path):
        """Inside the table it would be dragged around by sorting/filtering."""
        result, _ = _run(tmp_path, TIPO_PREVENTISTA_GENERICO)
        ws = load_workbook(result.ruta_archivo).active
        tabla = next(iter(ws.tables.values()))
        ultima_fila_tabla = int(tabla.ref.split(":")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        assert ws.max_row > ultima_fila_tabla


# ---------------------------------------------------------------------------
# Cableado config-driven: los filtros nuevos tienen que sobrevivir el merge.
# Olvidar la rama `if ... is not None` de merge_filters deja el flag en su
# default sin producir ningun error — asi se rompio `solo_con_cargo`.
# ---------------------------------------------------------------------------


class TestConfigWiring:
    def test_registry_maps_tipo_to_handler(self):
        import main

        assert main.REPORT_HANDLERS["cobertura"] == "_run_cobertura_report"
        assert callable(main._run_cobertura_report)

    def test_config_file_loads(self):
        from src.config.resolver import load_report_config

        cfg = load_report_config(Path("configs/cobertura.json"))
        assert cfg.tipo == "cobertura"
        assert len(cfg.reportes) == 3

    def test_config_file_hardcodes_no_periods(self):
        """Los periodos se derivan de fecha_desde; un mes escrito a mano se
        desincroniza solo cuando el daily patchea la fecha."""
        crudo = json.loads(Path("configs/cobertura.json").read_text(encoding="utf-8"))
        assert "periodos" not in json.dumps(crudo)

    def test_apertura_survives_merge_filters(self):
        from src.config.resolver import load_report_config, merge_filters

        cfg = load_report_config(Path("configs/cobertura.json"))
        aperturas = [
            merge_filters(cfg.filtros, r.filtros)["apertura_cobertura"]
            for r in cfg.reportes
        ]
        assert aperturas == [
            TIPO_PREVENTISTA_GENERICO,
            TIPO_PREVENTISTA_MARCA,
            TIPO_SUCURSAL_MARCA,
        ]

    def test_meses_atras_survives_merge_filters(self):
        from src.config.models import GlobalFilters, ReportFilters
        from src.config.resolver import merge_filters

        merged = merge_filters(
            GlobalFilters(fecha_desde="2026-08-01", fecha_hasta="2026-08-31"),
            ReportFilters(meses_atras=[25, 13, 1]),
        )
        assert merged["meses_atras"] == [25, 13, 1]

    def test_defaults_are_none_so_the_service_decides(self):
        from src.config.models import GlobalFilters
        from src.config.resolver import merge_filters

        merged = merge_filters(
            GlobalFilters(fecha_desde="2026-08-01", fecha_hasta="2026-08-31"), None
        )
        assert merged["apertura_cobertura"] is None
        assert merged["meses_atras"] is None

    def test_handler_builds_config_from_merged_filters(self, tmp_path):
        import main
        from src.config.models import GlobalFilters, ReportFilters
        from src.config.resolver import merge_filters

        merged = merge_filters(
            GlobalFilters(fecha_desde="2026-08-04", fecha_hasta="2026-08-31",
                          con_slicers=False),
            ReportFilters(apertura_cobertura=TIPO_SUCURSAL_MARCA, meses_atras=[1]),
        )
        report = MagicMock()
        report.nombre = "Cobertura Test"

        loader = MagicMock(spec=DataLoader)
        loader.get_cobertura_sucursal_marca.return_value = _raw_por_tipo(TIPO_SUCURSAL_MARCA)
        with patch("src.services.cobertura.service.service_output_dir", return_value=tmp_path), \
             patch("src.services.cobertura.CoberturaService.__init__",
                   lambda self, *a, **k: BaseService.__init__(self, loader)):
            artifacts = main._run_cobertura_report(report, merged)

        assert len(artifacts) == 1
        ruta, metadata = artifacts[0]
        assert ruta.name == "Cobertura Test.xlsx"
        assert metadata["nombre"] == "Cobertura Test"
        _, kwargs = loader.get_cobertura_sucursal_marca.call_args
        assert kwargs["periodos"] == ["2026-07-01"]

    def test_handler_reports_invalid_apertura_instead_of_raising(self):
        import main

        report = MagicMock()
        report.nombre = "Cobertura Test"
        merged = {"fecha_desde": "2026-08-04", "apertura_cobertura": "no_existe",
                  "con_slicers": False}
        assert main._run_cobertura_report(report, merged) == []

    def test_handler_rejects_explicit_empty_meses_atras(self):
        """Un `[]` escrito a mano tiene que fallar ruidosamente, no caer al
        default: un filtro que se ignora en silencio es como se rompió
        `solo_con_cargo`."""
        import main

        report = MagicMock()
        report.nombre = "Cobertura Test"
        merged = {"fecha_desde": "2026-08-04", "meses_atras": [], "con_slicers": False}
        assert main._run_cobertura_report(report, merged) == []
