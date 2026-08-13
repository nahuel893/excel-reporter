"""Tests for the cupo-desagregado processor (pure logic, no DB, no Excel writes)."""
import pandas as pd
import pytest
from datetime import date
from openpyxl import Workbook

from src.services.cupo_desagregado.processor import (
    Vendedor,
    agregar_historia,
    clasificar_categoria,
    construir_mapa_vendedores,
    construir_rutas_por_vendedor,
    distribuir_cupos,
    hoja_del_mes,
    leer_cupos,
    periodo_historia,
    split_proporcional,
    validar,
)


# ---------------------------------------------------------------------------
# split_proporcional
# ---------------------------------------------------------------------------
class TestSplitProporcional:
    def test_reparte_segun_pesos(self):
        assert split_proporcional(100.0, [3.0, 1.0]) == [75.0, 25.0]

    def test_ultima_ruta_absorbe_el_residuo(self):
        partes = split_proporcional(10.0, [1.0, 1.0, 1.0])
        assert partes == [3.33, 3.33, 3.34]
        assert sum(partes) == 10.0

    def test_sin_historia_reparte_parejo(self):
        assert split_proporcional(100.0, [0.0, 0.0]) == [50.0, 50.0]

    def test_cupo_cero_devuelve_ceros(self):
        assert split_proporcional(0.0, [5.0, 1.0]) == [0.0, 0.0]

    def test_sin_rutas_devuelve_lista_vacia(self):
        assert split_proporcional(100.0, []) == []

    def test_pesos_negativos_se_tratan_como_sin_historia(self):
        assert split_proporcional(100.0, [-1.0, -1.0]) == [50.0, 50.0]


# ---------------------------------------------------------------------------
# clasificar_categoria
# ---------------------------------------------------------------------------
class TestClasificarCategoria:
    @pytest.mark.parametrize("marca", ["SALTA", "SCHNEIDER", "NORTE"])
    def test_marcas_del_grupo_salta(self, marca):
        assert clasificar_categoria("CERVEZAS", marca, False) == "SALTA"

    @pytest.mark.parametrize("marca", ["HEINEKEN", "IMPERIAL", "MILLER"])
    def test_marcas_propias(self, marca):
        assert clasificar_categoria("CERVEZAS", marca, False) == marca

    @pytest.mark.parametrize("marca", ["AMSTEL", "GROLSCH", "IGUANA", "WARSTEINER"])
    def test_multicerveza(self, marca):
        assert clasificar_categoria("CERVEZAS", marca, False) == "MULTICERVEZA"

    def test_agua_danone_por_generico(self):
        assert clasificar_categoria("AGUAS DANONE", "VILLAVICENCIO", False) == "AGUA DANONE"

    def test_fernet_por_flag_de_articulo(self):
        # No existe marca 'FERNET' en dim_articulo: se detecta por des_articulo.
        assert clasificar_categoria("FRATELLI B", "FRATELLI BRANCA", True) == "FERNET"

    def test_vinos_por_generico(self):
        # La marca 'VINOS' tiene cero ventas; el generico es el que clasifica.
        assert clasificar_categoria("VINOS", "CUALQUIERA", False) == "VINOS"

    def test_r2_por_marca(self):
        assert clasificar_categoria("ENERGIZANTES", "R2", False) == "R2"

    def test_cerveza_gana_sobre_fernet(self):
        # El orden importa: una cerveza nunca cae en FERNET aunque el flag venga True.
        assert clasificar_categoria("CERVEZAS", "SALTA", True) == "SALTA"

    def test_generico_no_mapeado_devuelve_none(self):
        assert clasificar_categoria("GASEOSAS", "PEPSI", False) is None

    def test_tolera_nulos(self):
        assert clasificar_categoria(None, None, False) is None

    def test_tolera_nan_de_pandas(self):
        # dim_articulo tiene marcas/genericos nulos: llegan como float('nan').
        import numpy as np

        assert clasificar_categoria(np.nan, np.nan, False) is None
        assert clasificar_categoria("CERVEZAS", np.nan, False) is None
        assert clasificar_categoria(np.nan, "R2", False) == "R2"


# ---------------------------------------------------------------------------
# agregar_historia
# ---------------------------------------------------------------------------
class TestAgregarHistoria:
    def test_agrupa_por_sucursal_ruta_categoria(self):
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": 1, "generico": "CERVEZAS",
             "marca": "SALTA", "es_fernet": False, "cantidad": 10.0},
            {"id_sucursal": 3, "id_ruta": 1, "generico": "CERVEZAS",
             "marca": "SCHNEIDER", "es_fernet": False, "cantidad": 5.0},
            {"id_sucursal": 3, "id_ruta": 2, "generico": "CERVEZAS",
             "marca": "HEINEKEN", "es_fernet": False, "cantidad": 7.0},
        ])
        historia = agregar_historia(df)
        assert historia[(3, 1, "SALTA")] == 15.0
        assert historia[(3, 2, "HEINEKEN")] == 7.0

    def test_descarta_filas_sin_categoria(self):
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": 1, "generico": "GASEOSAS",
             "marca": "PEPSI", "es_fernet": False, "cantidad": 99.0},
        ])
        assert agregar_historia(df) == {}

    def test_dataframe_vacio(self):
        df = pd.DataFrame(
            columns=["id_sucursal", "id_ruta", "generico", "marca", "es_fernet", "cantidad"]
        )
        assert agregar_historia(df) == {}

    def test_descarta_ventas_sin_ruta(self):
        # Un cliente puede tener preventista y no tener ruta: esa venta no se
        # puede imputar a ninguna ruta.
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": None, "generico": "CERVEZAS",
             "marca": "SALTA", "es_fernet": False, "cantidad": 10.0},
            {"id_sucursal": 3, "id_ruta": 1, "generico": "CERVEZAS",
             "marca": "SALTA", "es_fernet": False, "cantidad": 4.0},
        ])
        assert agregar_historia(df) == {(3, 1, "SALTA"): 4.0}

    def test_cantidad_nula_no_rompe(self):
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": 1, "generico": "CERVEZAS",
             "marca": "SALTA", "es_fernet": False, "cantidad": None},
        ])
        assert agregar_historia(df) == {(3, 1, "SALTA"): 0.0}


# ---------------------------------------------------------------------------
# construir_* (DataFrame -> dict lookups)
# ---------------------------------------------------------------------------
class TestConstruirLookups:
    def test_mapa_vendedores_normaliza_nombre(self):
        df = pd.DataFrame([
            {"id_vendedor": 10, "des_vendedor": "  perez juan  ", "id_sucursal": 3},
        ])
        assert construir_mapa_vendedores(df) == {("PEREZ JUAN", 3): 10}

    def test_rutas_por_vendedor_agrupa_por_clave_compuesta(self):
        # REGLA DE ORO: id_ruta se reusa entre sucursales -> clave (vendedor, sucursal).
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": 1, "des_ruta": "DIAS LU-JU", "id_vendedor": 10},
            {"id_sucursal": 4, "id_ruta": 1, "des_ruta": "VILLA LU-JU", "id_vendedor": 10},
        ])
        rutas = construir_rutas_por_vendedor(df)
        assert rutas[(10, 3)] == [(1, "DIAS LU-JU")]
        assert rutas[(10, 4)] == [(1, "VILLA LU-JU")]

    def test_rutas_sin_descripcion_usan_fallback(self):
        df = pd.DataFrame([
            {"id_sucursal": 3, "id_ruta": 7, "des_ruta": None, "id_vendedor": 10},
        ])
        assert construir_rutas_por_vendedor(df)[(10, 3)] == [(7, "RUTA 7")]


# ---------------------------------------------------------------------------
# distribuir_cupos
# ---------------------------------------------------------------------------
def _vendedor(nombre="PEREZ JUAN", sucursal="CAFAYATE", id_sucursal=3, **cupos):
    base = {c: 0.0 for c in
            ["CERVEZAS", "SALTA", "HEINEKEN", "IMPERIAL", "MILLER",
             "MULTICERVEZA", "AGUA DANONE", "FERNET", "VINOS", "R2"]}
    base.update(cupos)
    return Vendedor(nombre=nombre, sucursal=sucursal, id_sucursal=id_sucursal, cupos=base)


class TestDistribuirCupos:
    def test_reparte_proporcional_a_la_historia_de_cada_ruta(self):
        v = _vendedor(SALTA=100.0)
        rutas = {(10, 3): [(1, "RUTA A"), (2, "RUTA B")]}
        mapa = {("PEREZ JUAN", 3): 10}
        historia = {(3, 1, "SALTA"): 30.0, (3, 2, "SALTA"): 10.0}

        resultado = distribuir_cupos([v], rutas, mapa, historia)

        assert [f["vals"]["SALTA"] for f in resultado.filas] == [75.0, 25.0]
        assert [f["codigo"] for f in resultado.filas] == [1, 2]

    def test_cervezas_es_la_suma_de_las_cinco_marcas(self):
        v = _vendedor(SALTA=100.0, HEINEKEN=50.0, IMPERIAL=10.0,
                      MILLER=5.0, MULTICERVEZA=1.0, CERVEZAS=166.0)
        rutas = {(10, 3): [(1, "RUTA A"), (2, "RUTA B")]}
        mapa = {("PEREZ JUAN", 3): 10}
        historia = {(3, 1, "SALTA"): 1.0, (3, 2, "SALTA"): 1.0}

        resultado = distribuir_cupos([v], rutas, mapa, historia)

        for fila in resultado.filas:
            esperado = sum(fila["vals"][m] for m in
                           ["SALTA", "HEINEKEN", "IMPERIAL", "MILLER", "MULTICERVEZA"])
            assert fila["vals"]["CERVEZAS"] == pytest.approx(esperado, abs=0.01)
        total = sum(f["vals"]["CERVEZAS"] for f in resultado.filas)
        assert total == pytest.approx(166.0, abs=0.01)

    def test_vendedor_sin_rutas_recibe_fila_unica(self):
        v = _vendedor(nombre="AQUINO GUSTAVO", SALTA=80.0)
        resultado = distribuir_cupos([v], {}, {}, {})

        assert len(resultado.filas) == 1
        fila = resultado.filas[0]
        assert fila["ruta"] == "SIN RUTA ASIGNADA"
        assert fila["codigo"] is None
        assert fila["vals"]["SALTA"] == 80.0
        assert resultado.sin_ruta == ["AQUINO GUSTAVO"]

    def test_categoria_sin_historia_se_reparte_pareja(self):
        v = _vendedor(MULTICERVEZA=10.0)
        rutas = {(10, 3): [(1, "RUTA A"), (2, "RUTA B")]}
        mapa = {("PEREZ JUAN", 3): 10}

        resultado = distribuir_cupos([v], rutas, mapa, {})

        assert [f["vals"]["MULTICERVEZA"] for f in resultado.filas] == [5.0, 5.0]
        assert "PEREZ JUAN / MULTICERVEZA" in resultado.sin_historia

    def test_nombre_override_resuelve_el_id_vendedor(self):
        v = _vendedor(nombre="CRUZ GABRIEL ARNALDO", id_sucursal=14, SALTA=10.0)
        rutas = {(167, 14): [(5, "RUTA X")]}

        resultado = distribuir_cupos(
            [v], rutas, {}, {},
            nombre_overrides={("CRUZ GABRIEL ARNALDO", 14): 167},
        )

        assert resultado.filas[0]["ruta"] == "RUTA X"
        assert resultado.filas[0]["vals"]["SALTA"] == 10.0

    def test_rutas_override_usa_historia_de_otra_sucursal(self):
        # LAMAS tiene cupo en ABRA PAMPA (13) pero opera rutas de LA QUIACA (14).
        v = _vendedor(nombre="LAMAS SEBASTIAN", sucursal="ABRA PAMPA",
                      id_sucursal=13, SALTA=100.0)
        historia = {(14, 14, "SALTA"): 3.0, (14, 15, "SALTA"): 1.0}

        resultado = distribuir_cupos(
            [v], {}, {}, historia,
            rutas_override={("LAMAS SEBASTIAN", 13): [
                (14, 14, "LAMAS MA-VI"), (14, 15, "LAMAS MI-SA"),
            ]},
        )

        assert [f["ruta"] for f in resultado.filas] == ["LAMAS MA-VI", "LAMAS MI-SA"]
        assert [f["vals"]["SALTA"] for f in resultado.filas] == [75.0, 25.0]
        assert resultado.sin_ruta == []

    def test_la_suma_de_rutas_iguala_el_cupo_al_centavo(self):
        v = _vendedor(SALTA=1000.01, AGUA_DANONE=0.0)
        rutas = {(10, 3): [(1, "A"), (2, "B"), (3, "C")]}
        mapa = {("PEREZ JUAN", 3): 10}
        historia = {(3, 1, "SALTA"): 1.0, (3, 2, "SALTA"): 1.0, (3, 3, "SALTA"): 1.0}

        resultado = distribuir_cupos([v], rutas, mapa, historia)

        assert sum(f["vals"]["SALTA"] for f in resultado.filas) == 1000.01


# ---------------------------------------------------------------------------
# validar
# ---------------------------------------------------------------------------
class TestValidar:
    def test_sin_diferencias_devuelve_dict_vacio(self):
        v = _vendedor(SALTA=100.0)
        rutas = {(10, 3): [(1, "A"), (2, "B")]}
        mapa = {("PEREZ JUAN", 3): 10}
        historia = {(3, 1, "SALTA"): 1.0, (3, 2, "SALTA"): 3.0}

        resultado = distribuir_cupos([v], rutas, mapa, historia)
        assert validar(resultado.filas, [v]) == {}

    def test_detecta_una_ruta_que_no_cierra(self):
        v = _vendedor(SALTA=100.0)
        filas = [{"sucursal": "CAFAYATE", "vendedor": "PEREZ JUAN", "codigo": 1,
                  "ruta": "A", "vals": {c: 0.0 for c in
                                        ["CERVEZAS", "SALTA", "HEINEKEN", "IMPERIAL",
                                         "MILLER", "MULTICERVEZA", "AGUA DANONE",
                                         "FERNET", "VINOS", "R2"]}}]
        filas[0]["vals"]["SALTA"] = 90.0

        errores = validar(filas, [v])
        assert errores["PEREZ JUAN/SALTA"] == 10.0


# ---------------------------------------------------------------------------
# Periodo / hoja del mes
# ---------------------------------------------------------------------------
class TestPeriodo:
    def test_hoja_del_mes(self):
        assert hoja_del_mes("2026-07-01") == "JULIO"
        assert hoja_del_mes("2026-08-15") == "AGOSTO"

    def test_historia_es_el_mes_anterior_completo(self):
        assert periodo_historia("2026-07-01") == (date(2026, 6, 1), date(2026, 7, 1))

    def test_historia_cruza_el_anio(self):
        assert periodo_historia("2026-01-20") == (date(2025, 12, 1), date(2026, 1, 1))


# ---------------------------------------------------------------------------
# leer_cupos
# ---------------------------------------------------------------------------
def _crear_fuente(tmp_path, hoja="JULIO", filas=None):
    """Construye un xlsx con el layout del bloque 'Objetivo' (filas 66-109)."""
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    for offset, (nombre, sucursal, cupos) in enumerate(filas or []):
        row = 66 + offset
        ws[f"A{row}"] = nombre
        ws[f"B{row}"] = sucursal
        for col, valor in cupos.items():
            ws[f"{col}{row}"] = valor
    path = tmp_path / "Objetivo.xlsx"
    wb.save(path)
    return path


class TestLeerCupos:
    def test_lee_nombre_sucursal_y_cupos(self, tmp_path):
        path = _crear_fuente(tmp_path, filas=[
            ("PEREZ JUAN", "CAFAYATE", {"D": 166.0, "G": 100.0, "J": 50.0, "V": 20.0}),
        ])
        vendedores = leer_cupos(path, "JULIO")

        assert len(vendedores) == 1
        v = vendedores[0]
        assert v.nombre == "PEREZ JUAN"
        assert v.sucursal == "CAFAYATE"
        assert v.id_sucursal == 3
        assert v.cupos["CERVEZAS"] == 166.0
        assert v.cupos["SALTA"] == 100.0
        assert v.cupos["HEINEKEN"] == 50.0
        assert v.cupos["AGUA DANONE"] == 20.0
        assert v.cupos["R2"] == 0.0

    def test_saltea_filas_vacias_y_totales(self, tmp_path):
        path = _crear_fuente(tmp_path, filas=[
            ("PEREZ JUAN", "CAFAYATE", {"G": 10.0}),
            (None, None, {}),
            ("TOTAL", "CAFAYATE", {"G": 10.0}),
        ])
        assert [v.nombre for v in leer_cupos(path, "JULIO")] == ["PEREZ JUAN"]

    def test_hoja_inexistente_da_error_claro(self, tmp_path):
        path = _crear_fuente(tmp_path, hoja="JULIO", filas=[])
        with pytest.raises(ValueError, match="AGOSTO"):
            leer_cupos(path, "AGOSTO")

    def test_sucursal_desconocida_da_error_claro(self, tmp_path):
        path = _crear_fuente(tmp_path, filas=[
            ("PEREZ JUAN", "MARTE", {"G": 10.0}),
        ])
        with pytest.raises(ValueError, match="MARTE"):
            leer_cupos(path, "JULIO")
