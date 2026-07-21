"""Tests for DescuentosService — con_lista_precio flag (Walter Vilte variant)."""
from unittest.mock import MagicMock

import pandas as pd
from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.services.descuentos.service import DescuentosService


def _df():
    return pd.DataFrame({
        "sucursal": ["1 - CASA CENTRAL", "3 - CAFAYATE"],
        "generico": ["CERVEZAS", "AGUAS DANONE"],
        "marca": ["SALTA", "LEVITE"],
        "lista_precio": ["LISTA SALTA MAYORISTA", "INTERIOR MINORISTA"],
        "importe_neto_sin_desc": [1000.0, 500.0],
        "bonificacion_pesos": [150.0, 40.0],
    })


def test_con_lista_precio_true_has_both_sheets(tmp_path):
    svc = DescuentosService(data_loader=MagicMock(spec=DataLoader))
    out = tmp_path / "d.xlsx"
    svc._build_workbook(_df(), out, "2026-06-01 a 2026-06-30", con_lista_precio=True)
    wb = load_workbook(out)
    assert wb.sheetnames == ["normal", "lista_precio"]


def test_con_lista_precio_false_omits_lista_sheet(tmp_path):
    """Variante Walter Vilte: sin la hoja lista_precio."""
    svc = DescuentosService(data_loader=MagicMock(spec=DataLoader))
    out = tmp_path / "d_wv.xlsx"
    svc._build_workbook(_df(), out, "2026-06-01 a 2026-06-30", con_lista_precio=False)
    wb = load_workbook(out)
    assert wb.sheetnames == ["normal"]


def test_default_keeps_lista_precio(tmp_path):
    svc = DescuentosService(data_loader=MagicMock(spec=DataLoader))
    out = tmp_path / "d_def.xlsx"
    svc._build_workbook(_df(), out, "periodo")  # default con_lista_precio=True
    assert load_workbook(out).sheetnames == ["normal", "lista_precio"]
