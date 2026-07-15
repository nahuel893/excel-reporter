"""Tests para ExcelManager."""
import shutil
import subprocess
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.core.excel_manager import ExcelManager, _col_letter_to_index, _parse_range

REAL_XLSX_CANDIDATES = [
    Path("data/output/avances/2026-07/AVANCE BADIE - JULIO 2026.xlsx"),
    Path("data/input/avances/AVANCE BADIE.xlsx"),
]


def _real_xlsx_path() -> Path | None:
    for candidate in REAL_XLSX_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


class TestColLetterToIndex:
    def test_single_letters(self):
        assert _col_letter_to_index("A") == 1
        assert _col_letter_to_index("Z") == 26

    def test_double_letters(self):
        assert _col_letter_to_index("AA") == 27
        assert _col_letter_to_index("AZ") == 52

    def test_case_insensitive(self):
        assert _col_letter_to_index("a") == _col_letter_to_index("A")


class TestParseRange:
    def test_simple_range(self):
        col1, row1, col2, row2 = _parse_range("A1:H20")
        assert col1 == 1
        assert row1 == 1
        assert col2 == 8
        assert row2 == 20

    def test_double_letter_column(self):
        col1, row1, col2, row2 = _parse_range("AA1:AB5")
        assert col1 == 27
        assert col2 == 28

    def test_invalid_range_raises(self):
        with pytest.raises(ValueError, match="Rango invalido"):
            _parse_range("A1")


class TestExcelManagerInit:
    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ExcelManager(tmp_path / "no_existe.xlsx")

    def test_accepts_existing_file(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx)
        mgr = ExcelManager(xlsx)
        assert mgr.ruta_excel == xlsx


class TestExcelManagerCaptureRange:
    def _make_xlsx(self, tmp_path: Path, sheet_name: str = "Hoja1") -> Path:
        xlsx = tmp_path / "reporte.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws["A1"] = "Header"
        ws["B1"] = "Value"
        ws["A2"] = 100
        ws["B2"] = 200
        wb.save(xlsx)
        return xlsx

    def test_missing_sheet_raises_value_error(self, tmp_path):
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with pytest.raises(ValueError, match="Hoja 'NoExiste' no encontrada"):
            mgr.capture_range("NoExiste", "A1:B2")

    def test_pillow_not_installed_raises_import_error(self, tmp_path):
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        # soffice debe estar disponible (o mockeado) para llegar al import de Pillow
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                with patch.dict("sys.modules", {"PIL": None, "PIL.Image": None, "PIL.ImageDraw": None, "PIL.ImageFont": None}):
                    with pytest.raises((ImportError, ModuleNotFoundError)):
                        mgr.capture_range("Hoja1", "A1:B2")

    def test_soffice_not_found_raises_runtime_error(self, tmp_path):
        """Sin LibreOffice en el PATH -> RuntimeError antes de intentar renderizar."""
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value=None):
            with pytest.raises(RuntimeError, match="LibreOffice no encontrado"):
                mgr.capture_range("Hoja1", "A1:B2")

    def test_capture_range_generates_png(self, tmp_path):
        """Verifica que genera un archivo PNG real."""
        try:
            from PIL import Image
        except ImportError:
            pytest.skip("Pillow not installed")

        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                out_path = mgr.capture_range("Hoja1", "A1:B2", output_dir=tmp_path)

        assert out_path.exists()
        assert out_path.suffix == ".png"
        img = Image.open(out_path)
        assert img.width > 0
        assert img.height > 0

    def test_capture_range_auto_detect(self, tmp_path):
        """Verifica auto-deteccion de bordes gruesos."""
        try:
            from PIL import Image
        except ImportError:
            pytest.skip("Pillow not installed")

        from openpyxl.styles import Border, Side

        xlsx = tmp_path / "bordered.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        thick = Side(style="thick")
        # Poner bordes gruesos en A1:B3
        for r in range(1, 4):
            for c in range(1, 3):
                cell = ws.cell(row=r, column=c, value=f"R{r}C{c}")
                cell.border = Border(
                    top=thick if r == 1 else None,
                    bottom=thick if r == 3 else None,
                    left=thick if c == 1 else None,
                    right=thick if c == 2 else None,
                )
        wb.save(xlsx)

        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                out_path = mgr.capture_range("Hoja1", "auto", output_dir=tmp_path)
        assert out_path.exists()

    def test_auto_detect_no_borders_raises(self, tmp_path):
        """Sin bordes gruesos y rango 'auto' -> ValueError."""
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with pytest.raises(ValueError, match="No se detectaron bordes gruesos"):
            mgr.capture_range("Hoja1", "auto")


class TestPrepareSheetForExportPrintArea:
    """FAST unit tests — pure openpyxl, no LibreOffice involved."""

    def _make_xlsx(self, tmp_path: Path, sheet_name: str = "Hoja1") -> Path:
        xlsx = tmp_path / "source.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws["A1"] = "Header"
        wb.save(xlsx)
        return xlsx

    def test_sets_print_area_when_range_addr_provided(self, tmp_path):
        xlsx = self._make_xlsx(tmp_path, "Hoja1")
        work_dir = tmp_path / "work"
        work_dir.mkdir()

        out_path = ExcelManager._prepare_sheet_for_export(
            xlsx, "Hoja1", work_dir, range_addr="B2:E10"
        )

        wb = openpyxl.load_workbook(out_path)
        ws = wb["Hoja1"]
        assert ws.print_area == "'Hoja1'!$B$2:$E$10"
        wb.close()

    def test_no_print_area_set_when_range_addr_is_none(self, tmp_path):
        xlsx = self._make_xlsx(tmp_path, "Hoja1")
        work_dir = tmp_path / "work"
        work_dir.mkdir()

        out_path = ExcelManager._prepare_sheet_for_export(xlsx, "Hoja1", work_dir)

        wb = openpyxl.load_workbook(out_path)
        ws = wb["Hoja1"]
        assert not ws.print_area
        wb.close()

    def test_no_print_area_set_when_range_addr_omitted_legacy_call(self, tmp_path):
        """Legacy 3-positional-arg call (no range_addr at all) must keep
        rendering the whole sheet, exactly like before this change."""
        xlsx = self._make_xlsx(tmp_path, "Hoja1")
        work_dir = tmp_path / "work"
        work_dir.mkdir()

        out_path = ExcelManager._prepare_sheet_for_export(xlsx, "Hoja1", work_dir)

        wb = openpyxl.load_workbook(out_path)
        ws = wb["Hoja1"]
        assert not ws.print_area
        # Fit-to-page/other setup must remain untouched by the new param.
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 1
        wb.close()


class TestCaptureRangeThreadsRangeAddr:
    """Verifies capture_range() only forwards the resolved range_addr into
    _prepare_sheet_for_export when crop=True; the default crop=False path is
    byte-for-byte legacy (range_addr=None -> no print_area -> whole sheet).
    Fast — mocks LibreOffice/soffice."""

    def _make_xlsx(self, tmp_path: Path, sheet_name: str = "Hoja1") -> Path:
        xlsx = tmp_path / "reporte.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws["A1"] = "Header"
        wb.save(xlsx)
        return xlsx

    def _make_bordered_xlsx(self, tmp_path: Path) -> Path:
        from openpyxl.styles import Border, Side

        xlsx = tmp_path / "bordered.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        thick = Side(style="thick")
        for r in range(1, 4):
            for c in range(1, 3):
                cell = ws.cell(row=r, column=c, value=f"R{r}C{c}")
                cell.border = Border(
                    top=thick if r == 1 else None,
                    bottom=thick if r == 3 else None,
                    left=thick if c == 1 else None,
                    right=thick if c == 2 else None,
                )
        wb.save(xlsx)
        return xlsx

    def test_explicit_range_addr_is_forwarded_when_crop_true(self, tmp_path):
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                with patch.object(
                    ExcelManager, "_prepare_sheet_for_export", return_value=xlsx
                ) as mock_prepare:
                    mgr.capture_range("Hoja1", "A1:B2", output_dir=tmp_path, crop=True)

        mock_prepare.assert_called_once()
        _, kwargs = mock_prepare.call_args
        assert kwargs.get("range_addr") == "A1:B2"

    def test_default_crop_false_passes_none_range_addr(self, tmp_path):
        """Backward-compat guarantee for branca/schneider: even when a concrete
        range is given, the DEFAULT (crop=False) path must pass range_addr=None
        so _prepare_sheet_for_export never sets print_area (whole-sheet render,
        byte-for-byte pre-PR2 behavior)."""
        xlsx = self._make_xlsx(tmp_path)
        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                with patch.object(
                    ExcelManager, "_prepare_sheet_for_export", return_value=xlsx
                ) as mock_prepare:
                    mgr.capture_range("Hoja1", "A1:B2", output_dir=tmp_path)

        mock_prepare.assert_called_once()
        _, kwargs = mock_prepare.call_args
        assert kwargs.get("range_addr") is None

    def test_auto_sentinel_forwards_resolved_concrete_range_when_crop_true(self, tmp_path):
        """When range_addr='auto' and crop=True, the concrete detected range
        (not the literal string 'auto') must reach _prepare_sheet_for_export."""
        xlsx = self._make_bordered_xlsx(tmp_path)

        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                with patch.object(
                    ExcelManager, "_prepare_sheet_for_export", return_value=xlsx
                ) as mock_prepare:
                    mgr.capture_range("Hoja1", "auto", output_dir=tmp_path, crop=True)

        mock_prepare.assert_called_once()
        _, kwargs = mock_prepare.call_args
        assert kwargs.get("range_addr") == "A1:B3"

    def test_auto_sentinel_default_crop_false_passes_none_range_addr(self, tmp_path):
        """'auto' is still resolved (for the filename slug) but with crop=False
        no print_area is set — range_addr=None reaches _prepare_sheet_for_export."""
        xlsx = self._make_bordered_xlsx(tmp_path)

        mgr = ExcelManager(xlsx)
        with patch.object(ExcelManager, "_find_soffice", return_value="/usr/bin/soffice"):
            with patch.object(ExcelManager, "_recalc_with_libreoffice", return_value=xlsx):
                with patch.object(
                    ExcelManager, "_prepare_sheet_for_export", return_value=xlsx
                ) as mock_prepare:
                    mgr.capture_range("Hoja1", "auto", output_dir=tmp_path)

        mock_prepare.assert_called_once()
        _, kwargs = mock_prepare.call_args
        assert kwargs.get("range_addr") is None


class TestExcelManagerDetectBorderedRange:
    def test_detects_thick_borders(self, tmp_path):
        from openpyxl.styles import Border, Side

        xlsx = tmp_path / "borders.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"

        thick = Side(style="medium")
        ws.cell(row=2, column=2).border = Border(top=thick)
        ws.cell(row=5, column=4).border = Border(bottom=thick)
        wb.save(xlsx)

        mgr = ExcelManager(xlsx)
        result = mgr.detect_bordered_range("Test")
        assert result == "B2:D5"

    def test_no_borders_returns_none(self, tmp_path):
        xlsx = tmp_path / "no_borders.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Test"
        wb.save(xlsx)

        mgr = ExcelManager(xlsx)
        assert mgr.detect_bordered_range("Test") is None

    def test_missing_sheet_returns_none(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx)

        mgr = ExcelManager(xlsx)
        assert mgr.detect_bordered_range("NoExiste") is None


class TestExcelManagerGetDataframe:
    def _make_xlsx_with_data(self, tmp_path: Path) -> Path:
        xlsx = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        ws["A1"] = "Nombre"
        ws["B1"] = "Valor"
        ws["A2"] = "Alpha"
        ws["B2"] = 10
        ws["A3"] = "Beta"
        ws["B3"] = 20
        wb.save(xlsx)
        return xlsx

    def test_returns_dataframe_with_headers(self, tmp_path):
        xlsx = self._make_xlsx_with_data(tmp_path)
        mgr = ExcelManager(xlsx)
        df = mgr.get_dataframe("Datos", "A1:B3")
        assert list(df.columns) == ["Nombre", "Valor"]
        assert len(df) == 2
        assert df.iloc[0]["Nombre"] == "Alpha"
        assert df.iloc[1]["Valor"] == 20

    def test_missing_sheet_raises_value_error(self, tmp_path):
        xlsx = self._make_xlsx_with_data(tmp_path)
        mgr = ExcelManager(xlsx)
        with pytest.raises(ValueError, match="NoExiste"):
            mgr.get_dataframe("NoExiste", "A1:B3")

    def test_empty_range_returns_empty_dataframe(self, tmp_path):
        xlsx = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vacio"
        wb.save(xlsx)
        mgr = ExcelManager(xlsx)
        df = mgr.get_dataframe("Vacio", "A5:B4")
        assert df.empty


@pytest.mark.skipif(
    _real_xlsx_path() is None
    or shutil.which("soffice") is None
    or shutil.which("pdftoppm") is None,
    reason="Real AVANCE BADIE xlsx or soffice/pdftoppm not present locally",
)
class TestPrintAreaCropRenderIntegration:
    """GUARDED integration test — proves ws.print_area cropping actually
    shrinks the rendered CONTENT via a real LibreOffice + pdftoppm pipeline.

    `_prepare_sheet_for_export` always forces fitToWidth=1/fitToHeight=1
    (fit-to-one-page, to avoid LibreOffice cutting wide tables at A4
    width), so the RAW PDF page canvas is a FIXED physical page size
    regardless of print_area — comparing raw page pixel dimensions is a
    false signal (verified empirically: both came out byte-identical,
    1100x850). The real content-bounding-box only emerges after the same
    whitespace autocrop capture_range applies as its own step 5, which is
    what this test measures.

    SLOW (several minutes): openpyxl's non-read_only load of the real
    ~7.7MB, 21-sheet production workbook takes ~110-125s on its own (all
    styles are parsed eagerly; openpyxl has no partial-sheet load in
    write-capable mode — see RangeRecognizer's PR1 notes), and this test
    needs two such loads (one per render) plus two real PDF exports and
    two pdftoppm rasterizations. Kept out of the fast unit-test path by
    the skipif guard above; only runs when the gitignored real file AND
    both external binaries are present locally.
    """

    def _render_content_dimensions(self, tmp_path: Path, real_xlsx: Path, range_addr, label: str):
        """Builds an export xlsx via the real _prepare_sheet_for_export, then
        runs it through the real PDF export + pdftoppm rasterization steps
        (mirroring capture_range's steps 3-4), and finally applies the SAME
        numpy whitespace autocrop as capture_range's step 5, returning the
        (width, height) of the trimmed CONTENT bounding box.

        NOTE: the raw (pre-trim) PDF page is a FIXED physical page size
        (`_prepare_sheet_for_export` always sets fitToWidth=1/fitToHeight=1
        to avoid LibreOffice cutting wide tables at A4 width) — it does NOT
        shrink regardless of print_area. Measuring raw page pixels showed
        cropped and whole-sheet renders as byte-identical (1100x850 in a
        manual run), which is misleading: the real "did the crop work"
        signal only appears after trimming the white margin LibreOffice
        pads the page with, exactly like the PNG capture_range actually
        hands to callers. Deliberately skips capture_range's own
        LibreOffice-recalc step (unrelated to print_area geometry), since
        page/content geometry doesn't depend on live formula values."""
        import numpy as np
        from PIL import Image

        soffice = shutil.which("soffice")
        pdftoppm_bin = shutil.which("pdftoppm")

        work_dir = tmp_path / f"work_{label}"
        work_dir.mkdir()
        export_path = ExcelManager._prepare_sheet_for_export(
            real_xlsx, "Avance", work_dir, range_addr=range_addr
        )

        pdf_dir = tmp_path / f"pdf_{label}"
        pdf_dir.mkdir()
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(pdf_dir), str(export_path)],
            capture_output=True, text=True, timeout=120,
        )
        assert result.returncode == 0, f"soffice PDF export failed: {result.stderr}"
        pdf_candidates = list(pdf_dir.glob("*.pdf"))
        assert pdf_candidates, "LibreOffice did not produce a PDF"

        png_dir = tmp_path / f"png_{label}"
        png_dir.mkdir()
        png_prefix = png_dir / "page"
        result = subprocess.run(
            [pdftoppm_bin, "-r", "100", "-png", "-f", "1", "-l", "1", str(pdf_candidates[0]), str(png_prefix)],
            capture_output=True, text=True, timeout=120,
        )
        assert result.returncode == 0, f"pdftoppm rasterization failed: {result.stderr}"
        png_candidates = list(png_dir.glob("page-*.png"))
        assert png_candidates, "pdftoppm did not produce a PNG"

        img = Image.open(png_candidates[0])

        # Whitespace autocrop — same logic as capture_range's step 5.
        arr = np.array(img.convert("RGB"))
        non_white = np.any(arr < 250, axis=2)
        rows = np.any(non_white, axis=1)
        cols = np.any(non_white, axis=0)
        if rows.any() and cols.any():
            rmin, rmax = np.where(rows)[0][[0, -1]]
            cmin, cmax = np.where(cols)[0][[0, -1]]
            return (cmax + 1 - cmin), (rmax + 1 - rmin)
        return img.width, img.height

    def test_cropped_region_render_smaller_than_whole_sheet(self, tmp_path):
        real_xlsx = _real_xlsx_path()

        # Avance's first card region (see ground truth in RangeRecognizer's
        # golden tests): A1:AR18 — one of four cards on the sheet. Passing a
        # concrete range_addr to _prepare_sheet_for_export is exactly what
        # capture_range(..., crop=True) forwards, so this validates real
        # cropping at the geometry level.
        cropped_w, cropped_h = self._render_content_dimensions(
            tmp_path, real_xlsx, "A1:AR18", "cropped"
        )
        # range_addr=None mirrors the default crop=False (legacy whole-sheet)
        # path — no print_area is set.
        whole_w, whole_h = self._render_content_dimensions(
            tmp_path, real_xlsx, None, "whole"
        )

        assert cropped_w > 0 and cropped_h > 0
        assert whole_w > 0 and whole_h > 0

        cropped_area = cropped_w * cropped_h
        whole_area = whole_w * whole_h
        assert cropped_area < whole_area * 0.5, (
            f"Expected cropped render ({cropped_w}x{cropped_h}) to be "
            f"meaningfully smaller than whole-sheet render ({whole_w}x{whole_h})"
        )
