"""Tests for image renderer protocol, factory, and backend wrappers."""
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook


class TestProtocolAndFactory:
    def test_image_renderer_is_runtime_checkable(self):
        from src.core.excel_renderers import ImageRenderer

        class FakeRenderer:
            name = "fake"

            def render(self, xlsx_path, sheet, range_addr, output_dir, dpi=300):
                return Path("x.png")

        assert isinstance(FakeRenderer(), ImageRenderer)

    def test_object_missing_methods_is_not_renderer(self):
        from src.core.excel_renderers import ImageRenderer

        class NotARenderer:
            name = "nope"

        assert not isinstance(NotARenderer(), ImageRenderer)

    def test_factory_returns_libreoffice(self):
        from src.core.excel_renderers import get_renderer
        from src.core.excel_renderers.libreoffice_renderer import LibreOfficeRenderer

        r = get_renderer("libreoffice")
        assert isinstance(r, LibreOfficeRenderer)
        assert r.name == "libreoffice"

    def test_factory_returns_html_playwright(self):
        from src.core.excel_renderers import get_renderer
        from src.core.excel_renderers.html_playwright_renderer import (
            HtmlPlaywrightRenderer,
        )

        r = get_renderer("html_playwright")
        assert isinstance(r, HtmlPlaywrightRenderer)
        assert r.name == "html_playwright"

    def test_factory_unknown_raises_valueerror(self):
        from src.core.excel_renderers import get_renderer

        with pytest.raises(ValueError, match="Unknown renderer"):
            get_renderer("does-not-exist")


class TestLibreOfficeRenderer:
    def test_name(self):
        from src.core.excel_renderers.libreoffice_renderer import LibreOfficeRenderer

        assert LibreOfficeRenderer().name == "libreoffice"

    def test_delegates_to_excel_manager(self, tmp_path):
        """Wrapper must call ExcelManager.capture_range with same args."""
        from src.core.excel_renderers.libreoffice_renderer import LibreOfficeRenderer

        xlsx = tmp_path / "x.xlsx"
        xlsx.write_bytes(b"")
        expected_png = tmp_path / "result.png"

        with patch(
            "src.core.excel_renderers.libreoffice_renderer.ExcelManager"
        ) as MockManager:
            instance = MockManager.return_value
            instance.capture_range.return_value = expected_png

            result = LibreOfficeRenderer().render(
                xlsx, "Sheet1", "A1:B2", tmp_path, dpi=200
            )

        MockManager.assert_called_once_with(xlsx)
        instance.capture_range.assert_called_once_with(
            sheet_name="Sheet1",
            range_addr="A1:B2",
            output_dir=tmp_path,
            dpi=200,
            crop=False,
        )
        assert result == expected_png

    def test_forwards_crop_flag(self, tmp_path):
        """crop must be threaded through to ExcelManager.capture_range."""
        from src.core.excel_renderers.libreoffice_renderer import LibreOfficeRenderer

        xlsx = tmp_path / "x.xlsx"
        xlsx.write_bytes(b"")
        expected_png = tmp_path / "result.png"

        with patch(
            "src.core.excel_renderers.libreoffice_renderer.ExcelManager"
        ) as MockManager:
            instance = MockManager.return_value
            instance.capture_range.return_value = expected_png

            LibreOfficeRenderer().render(
                xlsx, "Sheet1", "A1:B2", tmp_path, dpi=200, crop=True
            )

        instance.capture_range.assert_called_once_with(
            sheet_name="Sheet1",
            range_addr="A1:B2",
            output_dir=tmp_path,
            dpi=200,
            crop=True,
        )


class TestHtmlPlaywrightRenderer:
    def test_name(self):
        from src.core.excel_renderers.html_playwright_renderer import (
            HtmlPlaywrightRenderer,
        )

        assert HtmlPlaywrightRenderer().name == "html_playwright"

    def test_missing_sheet_raises(self, tmp_path):
        from src.core.excel_renderers.html_playwright_renderer import (
            HtmlPlaywrightRenderer,
        )

        xlsx = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.active.title = "Hoja1"
        wb.save(str(xlsx))

        with pytest.raises(ValueError, match="Sheet"):
            HtmlPlaywrightRenderer().render(
                xlsx, "NoExiste", "A1:B2", tmp_path
            )

    def test_missing_xlsx2html_raises_runtime_error(self, tmp_path):
        from src.core.excel_renderers.html_playwright_renderer import (
            HtmlPlaywrightRenderer,
        )

        xlsx = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.active.title = "Hoja1"
        wb.save(str(xlsx))

        with patch.dict("sys.modules", {"xlsx2html": None}):
            with pytest.raises(RuntimeError, match="xlsx2html"):
                HtmlPlaywrightRenderer().render(
                    xlsx, "Hoja1", "A1:B2", tmp_path
                )

    def test_render_produces_png(self, tmp_path):
        """Real render smoke test — skip if deps missing."""
        pytest.importorskip("xlsx2html")
        pytest.importorskip("playwright")

        from src.core.excel_renderers.html_playwright_renderer import (
            HtmlPlaywrightRenderer,
        )

        xlsx = tmp_path / "mini.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "DATA"
        ws["A1"] = "Producto"
        ws["B1"] = "Cantidad"
        ws["A2"] = "Cerveza"
        ws["B2"] = 100
        wb.save(str(xlsx))

        png = HtmlPlaywrightRenderer().render(
            xlsx, "DATA", "A1:B2", tmp_path, dpi=150
        )

        assert png.exists()
        assert png.suffix == ".png"
        with open(png, "rb") as f:
            assert f.read(4) == b"\x89PNG"
        # Filename follows convention
        assert png.name == "mini_DATA_A1_B2.png"
