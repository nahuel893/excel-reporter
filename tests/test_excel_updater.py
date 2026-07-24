"""Tests para replace_sheet_data y _coerce_value de src.core.excel_updater."""
import pytest
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

from src.core.excel_updater import replace_sheet_data, _coerce_value


# ── Helper ────────────────────────────────────────────────────────────────────


def _make_wb(headers, data, table_name=None, header_row=1, sheet_name="Sheet1"):
    """Create a Workbook with headers and data, optionally as an Excel table."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Write headers
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    # Write data
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    # Add table if requested
    if table_name:
        end_col = get_column_letter(len(headers))
        end_row = header_row + len(data)
        ref = f"A{header_row}:{end_col}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        ws.add_table(table)
    return wb


# ── Tests para replace_sheet_data ─────────────────────────────────────────────


class TestReplaceSheetData:
    def test_basic_write(self):
        """3 cols, 3 rows → replace with 3 new rows → cell values match."""
        headers = ["A", "B", "C"]
        old_data = [(1, 2, 3), (4, 5, 6), (7, 8, 9)]
        wb = _make_wb(headers, old_data)

        df = pd.DataFrame({"A": [10, 20, 30], "B": [11, 21, 31], "C": [12, 22, 32]})
        rows_written = replace_sheet_data(wb, "Sheet1", df, ["A", "B", "C"])

        ws = wb["Sheet1"]
        assert rows_written == 3
        assert ws.cell(2, 1).value == 10
        assert ws.cell(3, 2).value == 21
        assert ws.cell(4, 3).value == 32

    def test_fewer_rows_clears_old(self):
        """5 rows → replace with 2 → rows 3-5 are None."""
        headers = ["X", "Y"]
        old_data = [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5)]
        wb = _make_wb(headers, old_data)

        df = pd.DataFrame({"X": [10, 20], "Y": [11, 21]})
        replace_sheet_data(wb, "Sheet1", df, ["X", "Y"])

        ws = wb["Sheet1"]
        # Rows 4 and 5 (index 3-4 in original) should be None
        assert ws.cell(4, 1).value is None
        assert ws.cell(5, 1).value is None
        assert ws.cell(6, 1).value is None

    def test_more_rows(self):
        """2 rows → replace with 5 → all 5 written correctly."""
        headers = ["N"]
        old_data = [(1,), (2,)]
        wb = _make_wb(headers, old_data)

        df = pd.DataFrame({"N": [10, 20, 30, 40, 50]})
        rows_written = replace_sheet_data(wb, "Sheet1", df, ["N"])

        ws = wb["Sheet1"]
        assert rows_written == 5
        assert ws.cell(2, 1).value == 10
        assert ws.cell(6, 1).value == 50

    def test_table_ref_resized(self):
        """Table ref 'A1:C5' → write 3 rows → ref becomes 'A1:C4'."""
        headers = ["A", "B", "C"]
        old_data = [(1, 2, 3), (4, 5, 6), (7, 8, 9), (10, 11, 12)]
        wb = _make_wb(headers, old_data, table_name="MiTabla")

        df = pd.DataFrame({"A": [100, 200, 300], "B": [101, 201, 301], "C": [102, 202, 302]})
        replace_sheet_data(wb, "Sheet1", df, ["A", "B", "C"])

        ws = wb["Sheet1"]
        table = list(ws.tables.values())[0]
        assert table.ref == "A1:C4"

    def test_header_row_2(self):
        """Headers at row 2, data starts row 3 → works correctly."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Row 1: some title
        ws.cell(row=1, column=1, value="TITULO")
        # Row 2: headers
        ws.cell(row=2, column=1, value="Nombre")
        ws.cell(row=2, column=2, value="Valor")
        # Row 3: old data
        ws.cell(row=3, column=1, value="viejo")
        ws.cell(row=3, column=2, value=99)

        df = pd.DataFrame({"Nombre": ["nuevo"], "Valor": [42]})
        rows_written = replace_sheet_data(wb, "Sheet1", df, ["Nombre", "Valor"], header_row=2)

        assert rows_written == 1
        assert ws.cell(3, 1).value == "nuevo"
        assert ws.cell(3, 2).value == 42

    def test_missing_column_raises_valueerror(self):
        """data_columns includes col not in headers → ValueError."""
        headers = ["A", "B"]
        wb = _make_wb(headers, [(1, 2)])

        df = pd.DataFrame({"A": [10], "B": [20], "C_MISSING": [30]})
        with pytest.raises(ValueError, match="header"):
            replace_sheet_data(wb, "Sheet1", df, ["A", "B", "C_MISSING"])

    def test_missing_df_column_raises_valueerror(self):
        """data_columns includes col not in DataFrame → ValueError."""
        headers = ["A", "B", "C"]
        wb = _make_wb(headers, [(1, 2, 3)])

        df = pd.DataFrame({"A": [10], "B": [20]})  # C missing from df
        with pytest.raises(ValueError, match="DataFrame"):
            replace_sheet_data(wb, "Sheet1", df, ["A", "B", "C"])

    def test_missing_sheet_raises_keyerror(self):
        """sheet_name not in wb → KeyError."""
        wb = Workbook()
        wb.active.title = "Sheet1"

        df = pd.DataFrame({"A": [1]})
        with pytest.raises(KeyError):
            replace_sheet_data(wb, "NoExiste", df, ["A"])

    def test_empty_dataframe(self):
        """Write empty df → all data cleared, returns 0."""
        headers = ["A", "B"]
        old_data = [(1, 2), (3, 4), (5, 6)]
        wb = _make_wb(headers, old_data)

        df = pd.DataFrame({"A": pd.Series([], dtype=object), "B": pd.Series([], dtype=object)})
        rows_written = replace_sheet_data(wb, "Sheet1", df, ["A", "B"])

        ws = wb["Sheet1"]
        assert rows_written == 0
        # Old data should be cleared
        assert ws.cell(2, 1).value is None
        assert ws.cell(3, 1).value is None

    def test_preserves_formula_columns(self):
        """wb has 4 cols, replace only first 2 → col 3-4 values untouched."""
        headers = ["A", "B", "Formula1", "Formula2"]
        old_data = [(1, 2, "=A2+B2", "=A2*B2"), (3, 4, "=A3+B3", "=A3*B3")]
        wb = _make_wb(headers, old_data)

        df = pd.DataFrame({"A": [10, 20], "B": [11, 21]})
        replace_sheet_data(wb, "Sheet1", df, ["A", "B"])

        ws = wb["Sheet1"]
        # New data written to A, B
        assert ws.cell(2, 1).value == 10
        assert ws.cell(2, 2).value == 11
        # Formula columns untouched
        assert ws.cell(2, 3).value == "=A2+B2"
        assert ws.cell(2, 4).value == "=A2*B2"
        assert ws.cell(3, 3).value == "=A3+B3"

    def test_numeric_columns_resets_inherited_date_format(self):
        """A key column that inherits a date number_format from the template is
        reset to a plain numeric format when listed in numeric_columns."""
        headers = ["Código", "Valor"]
        wb = _make_wb(headers, [(1, 100)])
        ws = wb["Sheet1"]
        # Template inherited a date format on the integer key column.
        ws.cell(2, 1).number_format = "[$-C0A]dd\\-mmm\\-yy;@"

        df = pd.DataFrame({"Código": [6, 61], "Valor": [1, 2]})
        replace_sheet_data(
            wb, "Sheet1", df, ["Código", "Valor"], numeric_columns=["Código"]
        )

        ws = wb["Sheet1"]
        assert ws.cell(2, 1).number_format == "0"
        assert ws.cell(3, 1).number_format == "0"
        assert ws.cell(2, 1).value == 6
        assert isinstance(ws.cell(2, 1).value, int)
        # Non-numeric column keeps whatever format it had (General here).
        assert ws.cell(2, 2).number_format != "0" or ws.cell(2, 2).value == 1

    def test_numeric_columns_survive_openpyxl_roundtrip(self):
        """REGRESSION: integer keys written into a date-formatted column must
        keep their exact value across an openpyxl save/reload cycle.

        Without numeric_columns, openpyxl reserializes the int as a date on
        reload (and the 1900 leap-year bug shifts serials >= 60 by a day),
        silently breaking downstream exact-match VLOOKUPs. See CuposVolumen!
        Código -> AvanceR lookup (avance-badie cerveza cupo 82.574 vs 83.000).
        """
        import io
        from openpyxl import load_workbook

        headers = ["Código", "Valor"]
        wb = _make_wb(headers, [(1, 1)])
        wb["Sheet1"].cell(2, 1).number_format = "[$-C0A]dd\\-mmm\\-yy;@"

        df = pd.DataFrame({"Código": [61, 62, 63], "Valor": [1, 2, 3]})
        replace_sheet_data(
            wb, "Sheet1", df, ["Código", "Valor"], numeric_columns=["Código"]
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ws2 = load_workbook(buf)["Sheet1"]
        assert ws2.cell(2, 1).value == 61
        assert ws2.cell(3, 1).value == 62
        assert ws2.cell(4, 1).value == 63


# ── Tests para _coerce_value ───────────────────────────────────────────────────


class TestCoerceValue:
    def test_none_returns_none(self):
        assert _coerce_value(None) is None

    def test_bool_preserved(self):
        assert _coerce_value(True) is True
        assert _coerce_value(False) is False

    def test_numpy_integer(self):
        val = _coerce_value(np.int64(42))
        assert val == 42
        assert isinstance(val, int)

    def test_numpy_float(self):
        val = _coerce_value(np.float64(3.14))
        assert abs(val - 3.14) < 1e-9
        assert isinstance(val, float)

    def test_numpy_nan_returns_none(self):
        assert _coerce_value(np.float64(np.nan)) is None

    def test_native_float_nan_returns_none(self):
        assert _coerce_value(float("nan")) is None

    def test_pandas_timestamp(self):
        ts = pd.Timestamp("2026-04-15")
        result = _coerce_value(ts)
        from datetime import datetime
        assert isinstance(result, datetime)
        assert result.year == 2026
        assert result.month == 4
        assert result.day == 15

    def test_pandas_nat_returns_none(self):
        assert _coerce_value(pd.NaT) is None

    def test_string_preserved(self):
        assert _coerce_value("hello") == "hello"

    def test_native_int_passthrough(self):
        assert _coerce_value(7) == 7
