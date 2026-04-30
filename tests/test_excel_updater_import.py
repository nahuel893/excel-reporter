"""
Tests for import_xlsx_as_sheet helper (T-04).

Verifies that the helper:
- Copies values from the first sheet of a source xlsx into a target workbook
- Replaces an existing target sheet with the same name
- Returns the correct data row count
- Raises FileNotFoundError when source is missing
- Only imports the first sheet when source has multiple sheets
"""
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.core.excel_updater import import_xlsx_as_sheet


def _make_source_xlsx(path: Path, sheets: dict[str, list[list]]) -> Path:
    """Helper: create a source xlsx with given sheet names and rows."""
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    return path


class TestImportXlsxAsSheet:
    """T-04: import_xlsx_as_sheet function."""

    def test_import_copies_values_from_first_sheet(self, tmp_path):
        """T-04: Values from first source sheet appear in target wb under target_sheet_name."""
        src = _make_source_xlsx(
            tmp_path / "source.xlsx",
            {"Datos": [["Name", "Value"], ["Alice", 100], ["Bob", 200]]},
        )
        target_wb = Workbook()
        target_wb.active.title = "Existing"

        rows_written = import_xlsx_as_sheet(target_wb, src, "Detalle Movimientos")

        assert "Detalle Movimientos" in target_wb.sheetnames
        ws = target_wb["Detalle Movimientos"]
        # Header row
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Value"
        # Data rows
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 100
        assert ws.cell(3, 1).value == "Bob"
        assert ws.cell(3, 2).value == 200

    def test_import_replaces_existing_target_sheet(self, tmp_path):
        """T-04: If target_sheet_name already exists, it is removed and recreated with source data."""
        src = _make_source_xlsx(
            tmp_path / "source.xlsx",
            {"Datos": [["NewCol"], ["fresh"]]},
        )
        target_wb = Workbook()
        # Pre-populate target with old data
        ws_old = target_wb.active
        ws_old.title = "Detalle Movimientos"
        ws_old["A1"] = "OldHeader"
        ws_old["A2"] = "old_value"

        import_xlsx_as_sheet(target_wb, src, "Detalle Movimientos")

        ws = target_wb["Detalle Movimientos"]
        assert ws.cell(1, 1).value == "NewCol"
        assert ws.cell(2, 1).value == "fresh"
        # Old data should be gone
        assert ws.cell(1, 2).value is None

    def test_import_returns_data_row_count(self, tmp_path):
        """T-04: Returns number of data rows (total rows minus header row)."""
        src = _make_source_xlsx(
            tmp_path / "source.xlsx",
            {"Datos": [["Header"], ["r1"], ["r2"], ["r3"], ["r4"], ["r5"]]},
        )
        target_wb = Workbook()

        rows = import_xlsx_as_sheet(target_wb, src, "Dest")

        assert rows == 5  # 6 total rows - 1 header = 5 data rows

    def test_import_source_not_found_raises_filenotfound(self, tmp_path):
        """T-04: FileNotFoundError raised when source_path does not exist."""
        target_wb = Workbook()
        nonexistent = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError, match="nonexistent"):
            import_xlsx_as_sheet(target_wb, nonexistent, "Dest")

    def test_import_uses_first_sheet_only_when_multiple(self, tmp_path):
        """T-04: Only first sheet is imported even when source has multiple sheets."""
        src = _make_source_xlsx(
            tmp_path / "source.xlsx",
            {
                "FirstSheet": [["first_header"], ["first_data"]],
                "SecondSheet": [["second_header"], ["second_data"]],
            },
        )
        target_wb = Workbook()

        import_xlsx_as_sheet(target_wb, src, "Result")

        ws = target_wb["Result"]
        assert ws.cell(1, 1).value == "first_header"
        assert ws.cell(2, 1).value == "first_data"
        # Second sheet data should NOT be in result (only 2 rows)
        assert ws.cell(3, 1).value is None

    def test_import_returns_zero_for_header_only_source(self, tmp_path):
        """T-04: Returns 0 when source has only a header row and no data."""
        src = _make_source_xlsx(
            tmp_path / "source.xlsx",
            {"Empty": [["Header1", "Header2"]]},
        )
        target_wb = Workbook()

        rows = import_xlsx_as_sheet(target_wb, src, "Dest")

        assert rows == 0
