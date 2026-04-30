"""
Tests for ExcelWriter merge mode (T-01, T-02, T-03).

Verifies that ExcelWriter can:
- Load an existing xlsx as merge target (T-01)
- Replace managed sheets while preserving user sheets (T-02)
- Save back to the merge_with path (T-03)
"""
import pytest
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.core.excel_writer import ExcelWriter, SheetStyle


# ---------------------------------------------------------------------------
# T-01: merge_with parameter — load vs. fresh workbook
# ---------------------------------------------------------------------------

class TestExcelWriterMergeInit:
    """T-01: ExcelWriter __init__ with merge_with parameter."""

    def test_merge_with_none_creates_fresh_workbook(self, tmp_path):
        """T-01: merge_with=None → fresh Workbook created (default Sheet present)."""
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=None)
        assert writer.workbook is not None
        # Fresh workbook has one sheet named "Sheet"
        assert "Sheet" in writer.workbook.sheetnames
        assert writer._merge_mode is False

    def test_merge_with_existing_path_loads_workbook(self, tmp_path):
        """T-01: merge_with=existing_path → workbook loaded with original sheets."""
        # Create a source xlsx with two sheets
        src = tmp_path / "existing.xlsx"
        wb = Workbook()
        wb.active.title = "A"
        wb.create_sheet(title="B")
        wb.save(str(src))

        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=src)
        assert set(writer.workbook.sheetnames) == {"A", "B"}
        assert writer._merge_mode is True
        assert writer._merge_path == src

    def test_merge_with_nonexistent_path_falls_back_to_fresh(self, tmp_path):
        """T-01: merge_with=non-existent path → fresh workbook (defensive fallback)."""
        nonexistent = tmp_path / "does_not_exist.xlsx"
        writer = ExcelWriter("fallback", output_dir=tmp_path, merge_with=nonexistent)
        assert writer._merge_mode is False
        # Fresh workbook should still be functional
        assert "Sheet" in writer.workbook.sheetnames


# ---------------------------------------------------------------------------
# T-02: add_sheet behavior in merge mode
# ---------------------------------------------------------------------------

class TestExcelWriterAddSheetMergeMode:
    """T-02: add_sheet replaces managed sheets and preserves user sheets."""

    def _make_existing_xlsx(self, tmp_path: Path, sheets: dict[str, list]) -> Path:
        """Helper: create xlsx with given sheet names and simple data rows."""
        src = tmp_path / "existing.xlsx"
        wb = Workbook()
        first = True
        for sheet_name, rows in sheets.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(str(src))
        return src

    def test_add_sheet_replaces_existing_in_merge_mode(self, tmp_path):
        """T-02: add_sheet removes old content and inserts fresh df data for same sheet name."""
        src = self._make_existing_xlsx(tmp_path, {"CERVEZAS": [["old", "data"], [1, 2]]})

        df = pd.DataFrame({"Col1": ["new"], "Col2": [99]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=src)
        writer.add_sheet(df, sheet_name="CERVEZAS", style=style)

        wb = writer.workbook
        ws = wb["CERVEZAS"]
        # Header row should be Col1, Col2 (not old/data)
        header_vals = [ws.cell(1, c).value for c in range(1, 3)]
        assert header_vals == ["Col1", "Col2"]
        # Data row should have new values
        data_vals = [ws.cell(2, c).value for c in range(1, 3)]
        assert data_vals == ["new", 99]

    def test_add_sheet_preserves_unrelated_sheets_in_merge_mode(self, tmp_path):
        """T-02: add_sheet leaves user sheets intact (cells + existence)."""
        src = self._make_existing_xlsx(
            tmp_path,
            {
                "MyAnalysis": [["formula", "sheet"], ["=A1+1", 10]],
                "CERVEZAS": [["old", "header"]],
            },
        )

        df = pd.DataFrame({"Col1": ["new"]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=src)
        writer.add_sheet(df, sheet_name="CERVEZAS", style=style)

        wb = writer.workbook
        assert "MyAnalysis" in wb.sheetnames
        ws_user = wb["MyAnalysis"]
        # User data preserved
        assert ws_user.cell(1, 1).value == "formula"
        assert ws_user.cell(2, 2).value == 10

    def test_add_sheet_preserves_tab_order_on_replacement(self, tmp_path):
        """T-02 (RF-09): replacing a sheet keeps it at its original tab index."""
        src = self._make_existing_xlsx(
            tmp_path,
            {
                "CERVEZAS": [["old"]],
                "MiAnalisis": [["formula"]],
                "AGUAS DANONE": [["old2"]],
            },
        )
        # Capture original order before re-run
        wb_before = load_workbook(str(src))
        order_before = list(wb_before.sheetnames)
        wb_before.close()

        df = pd.DataFrame({"Col1": ["new"]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=src)
        writer.add_sheet(df, sheet_name="CERVEZAS", style=style)
        writer.add_sheet(df, sheet_name="AGUAS DANONE", style=style)

        # Order must be preserved exactly — managed sheets stay in their original positions
        assert writer.workbook.sheetnames == order_before, (
            f"Tab order changed: was {order_before}, now {writer.workbook.sheetnames}"
        )

    def test_add_sheet_creates_new_when_name_not_present_in_merge_mode(self, tmp_path):
        """T-02: add_sheet adds new sheet when the name does not exist in merged wb."""
        src = self._make_existing_xlsx(tmp_path, {"MyAnalysis": [["x", "y"]]})

        df = pd.DataFrame({"Col1": ["value"]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=src)
        writer.add_sheet(df, sheet_name="CERVEZAS", style=style)

        wb = writer.workbook
        assert "CERVEZAS" in wb.sheetnames
        assert "MyAnalysis" in wb.sheetnames

    def test_add_sheet_fresh_mode_first_sheet_uses_active(self, tmp_path):
        """T-02: fresh mode (no merge_with) first add_sheet renames the active sheet."""
        df = pd.DataFrame({"A": [1, 2]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("test_fresh", output_dir=tmp_path)
        writer.add_sheet(df, sheet_name="CERVEZAS", style=style)

        assert "CERVEZAS" in writer.workbook.sheetnames
        # The default "Sheet" should be gone (renamed)
        assert "Sheet" not in writer.workbook.sheetnames

    def test_add_sheet_fresh_mode_subsequent_creates_new(self, tmp_path):
        """T-02: fresh mode second add_sheet creates a second sheet."""
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("test_fresh_multi", output_dir=tmp_path)
        writer.add_sheet(df1, sheet_name="First", style=style)
        writer.add_sheet(df2, sheet_name="Second", style=style)

        assert set(writer.workbook.sheetnames) == {"First", "Second"}


# ---------------------------------------------------------------------------
# T-03: save() honors merge_path
# ---------------------------------------------------------------------------

class TestExcelWriterSaveMergeMode:
    """T-03: save() writes to merge_with path in merge mode, to output_dir otherwise."""

    def test_save_merge_mode_writes_to_merge_path(self, tmp_path):
        """T-03: save() returns existing_path and writes file there (not nombre_archivo)."""
        existing = tmp_path / "Custom Name.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet"
        wb.save(str(existing))

        other_dir = tmp_path / "other"
        other_dir.mkdir()

        writer = ExcelWriter("ignored_name", output_dir=other_dir, merge_with=existing)
        result = writer.save()

        assert result == existing
        assert existing.exists()
        # File in other_dir should NOT have been created
        assert not (other_dir / "ignored_name.xlsx").exists()

    def test_save_fresh_mode_writes_to_output_dir(self, tmp_path):
        """T-03: save() in fresh mode writes to output_dir / nombre_archivo.xlsx."""
        writer = ExcelWriter("my_report", output_dir=tmp_path)
        result = writer.save()

        assert result == tmp_path / "my_report.xlsx"
        assert result.exists()

    def test_save_merge_mode_file_contains_added_sheet(self, tmp_path):
        """T-03: save() in merge mode persists the new sheet data to the merge_path file."""
        existing = tmp_path / "target.xlsx"
        wb = Workbook()
        wb.active.title = "OldSheet"
        wb.save(str(existing))

        df = pd.DataFrame({"X": [42]})
        style = SheetStyle(as_table=False, summary_rows={})
        writer = ExcelWriter("ignored", output_dir=tmp_path, merge_with=existing)
        writer.add_sheet(df, sheet_name="NewData", style=style)
        writer.save()

        # Reload and verify
        wb2 = load_workbook(str(existing))
        assert "NewData" in wb2.sheetnames
        # Header and data cell
        assert wb2["NewData"].cell(1, 1).value == "X"
        assert wb2["NewData"].cell(2, 1).value == 42
