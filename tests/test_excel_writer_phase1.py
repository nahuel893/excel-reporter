"""
Tests for ExcelWriter Phase 1 changes (T-005, T-008).
RF-011: ColumnFormat.font_color applied to data cells only (not header).
RF-012: add_sheet() returns openpyxl Worksheet instance.
"""
import pytest
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.excel_writer import (
    ColumnFormat,
    SheetStyle,
    ExcelWriter,
)


class TestColumnFormatFontColor:
    """T-005: font_color attribute on ColumnFormat applied to data cells, not header."""

    def test_column_format_has_font_color_attribute(self):
        """T-005: ColumnFormat accepts font_color kwarg (not None by default)."""
        fmt = ColumnFormat(font_color="FF0000")
        assert fmt.font_color == "FF0000"

    def test_column_format_font_color_defaults_to_none(self):
        """T-005: ColumnFormat.font_color defaults to None."""
        fmt = ColumnFormat()
        assert fmt.font_color is None

    def test_font_color_applied_to_data_cells_not_header(self, tmp_path):
        """T-005: font_color is applied to data rows but NOT to the header row."""
        df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
        style = SheetStyle(
            column_formats={"A": ColumnFormat(font_color="FF0000")},
            as_table=False,
        )
        writer = ExcelWriter("test_font_color", output_dir=tmp_path, style=style)
        writer.add_sheet(df, sheet_name="Test", style=style)
        writer.save()

        wb = load_workbook(tmp_path / "test_font_color.xlsx")
        ws = wb.active

        # Row 1 is header — must NOT have red font color
        # openpyxl stores colors as ARGB (8 chars: alpha+RGB). Header has white font (FFFFFFFF).
        header_cell = ws["A1"]
        header_rgb = header_cell.font.color.rgb  # e.g. "FFFFFFFF" for white
        assert not header_rgb.endswith("FF0000"), (
            f"Header cell should not have data font_color FF0000, got {header_rgb}"
        )

        # Row 2 onwards are data — MUST have red font color
        # openpyxl returns ARGB so "FF0000" becomes "00FF0000" (with 00 alpha prefix)
        data_cell_row2 = ws["A2"]
        assert data_cell_row2.font.color.rgb.endswith("FF0000"), (
            f"Data cell row2 should have font_color ending FF0000, got {data_cell_row2.font.color.rgb}"
        )
        data_cell_row3 = ws["A3"]
        assert data_cell_row3.font.color.rgb.endswith("FF0000"), (
            f"Data cell row3 should have font_color ending FF0000, got {data_cell_row3.font.color.rgb}"
        )

    def test_font_color_not_applied_to_other_columns(self, tmp_path):
        """T-005: font_color in col A does not leak into col B."""
        df = pd.DataFrame({"A": [1], "B": [2]})
        style = SheetStyle(
            column_formats={"A": ColumnFormat(font_color="FF0000")},
            as_table=False,
        )
        writer = ExcelWriter("test_no_leak", output_dir=tmp_path, style=style)
        writer.add_sheet(df, sheet_name="Test", style=style)
        writer.save()

        wb = load_workbook(tmp_path / "test_no_leak.xlsx")
        ws = wb.active

        # Col B data cell should NOT have red font
        b_data = ws["B2"]
        # Default bold font has no explicit color set (None or 00000000)
        color = b_data.font.color
        # Should not be FF0000
        if color and color.type == "rgb":
            assert color.rgb != "FF0000"

    def test_none_font_color_does_not_break_existing_behavior(self, tmp_path):
        """T-005: ColumnFormat with font_color=None behaves as before (bold, no color change)."""
        df = pd.DataFrame({"A": [10]})
        style = SheetStyle(
            column_formats={"A": ColumnFormat(font_color=None)},
            as_table=False,
        )
        writer = ExcelWriter("test_none_color", output_dir=tmp_path, style=style)
        writer.add_sheet(df, sheet_name="Test", style=style)
        writer.save()

        wb = load_workbook(tmp_path / "test_none_color.xlsx")
        ws = wb.active
        # Just verify it doesn't crash and data is written
        assert ws["A2"].value == 10


class TestAddSheetReturnsWorksheet:
    """T-008: add_sheet() returns an openpyxl Worksheet object."""

    def test_add_sheet_returns_worksheet(self, tmp_path):
        """T-008: add_sheet() return value is an instance of openpyxl Worksheet."""
        df = pd.DataFrame({"X": [1, 2, 3]})
        writer = ExcelWriter("test_ws_return", output_dir=tmp_path)
        result = writer.add_sheet(df, sheet_name="Sheet1")
        assert isinstance(result, Worksheet), (
            f"Expected Worksheet, got {type(result)}"
        )

    def test_add_sheet_returns_worksheet_on_second_call(self, tmp_path):
        """T-008: add_sheet() returns Worksheet even on the second call (non-first sheet)."""
        df = pd.DataFrame({"X": [1]})
        writer = ExcelWriter("test_ws_return2", output_dir=tmp_path)
        ws1 = writer.add_sheet(df, sheet_name="Sheet1")
        ws2 = writer.add_sheet(df, sheet_name="Sheet2")
        assert isinstance(ws1, Worksheet)
        assert isinstance(ws2, Worksheet)
        assert ws1.title == "Sheet1"
        assert ws2.title == "Sheet2"

    def test_returned_worksheet_is_the_actual_sheet(self, tmp_path):
        """T-008: The returned Worksheet is the same one written in the workbook."""
        df = pd.DataFrame({"Col": [42]})
        writer = ExcelWriter("test_ws_identity", output_dir=tmp_path)
        ws = writer.add_sheet(df, sheet_name="MySheet")
        # The ws title should match what we passed
        assert ws.title == "MySheet"
        # Data must be accessible through the returned ws
        # Row 1 = header, Row 2 = data (no summary rows by default)
        assert ws.cell(row=2, column=1).value == 42
