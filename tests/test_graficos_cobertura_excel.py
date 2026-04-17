"""Tests for excel_builder — builds resumen.xlsx."""
import pandas as pd
from openpyxl import load_workbook

from src.services.graficos_cobertura.excel_builder import build_resumen_xlsx


class TestBuildResumenXlsx:
    """RF-015: resumen.xlsx with generico sheets + mensual sheets + comparativo."""

    def _simple_gen_sheet(self):
        return pd.DataFrame({
            "Zona": ["NOA NORTE", "NOA NORTE"],
            "Mes": ["Ene", "Feb"],
            "SALTA": [100, 120],
            "Total 2026": [500, 550],
        })

    def _simple_mensual_sheet(self):
        return pd.DataFrame({
            "Zona": ["NOA NORTE", "SALTA CAPITAL"],
            "CERVEZAS": [500, 300],
            "AGUAS SABORIZADAS": [200, 150],
        })

    def _simple_comparativo(self):
        return pd.DataFrame({
            "Zona": ["NOA NORTE", "NOA NORTE"],
            "Anio": [2025, 2026],
            "SALTA": [100, 150],
            "HEINEKEN": [80, 90],
        })

    def test_creates_xlsx_file(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={"CERVEZAS": self._simple_gen_sheet()},
            sheets_mensuales={"Marzo 2026": self._simple_mensual_sheet()},
            sheet_comparativo=self._simple_comparativo(),
        )
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_has_sheet_per_generico(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={
                "CERVEZAS": self._simple_gen_sheet(),
                "VINOS CCU": self._simple_gen_sheet(),
            },
            sheets_mensuales={},
            sheet_comparativo=None,
        )
        wb = load_workbook(out)
        assert "CERVEZAS" in wb.sheetnames
        assert "VINOS CCU" in wb.sheetnames

    def test_truncates_long_sheet_name_to_31_chars(self, tmp_path):
        long_name = "A" * 40
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={long_name: self._simple_gen_sheet()},
            sheets_mensuales={},
            sheet_comparativo=None,
        )
        wb = load_workbook(out)
        # Excel truncates to 31
        assert any(len(n) == 31 for n in wb.sheetnames)

    def test_has_mensual_sheets(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={},
            sheets_mensuales={
                "Ene 2026": self._simple_mensual_sheet(),
                "Feb 2026": self._simple_mensual_sheet(),
            },
            sheet_comparativo=None,
        )
        wb = load_workbook(out)
        assert "Ene 2026" in wb.sheetnames
        assert "Feb 2026" in wb.sheetnames

    def test_comparativo_sheet_when_provided(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={"CERVEZAS": self._simple_gen_sheet()},
            sheets_mensuales={},
            sheet_comparativo=self._simple_comparativo(),
        )
        wb = load_workbook(out)
        assert any("Comparativo" in n or "Comp" in n for n in wb.sheetnames)

    def test_comparativo_absent_when_none(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={"CERVEZAS": self._simple_gen_sheet()},
            sheets_mensuales={},
            sheet_comparativo=None,
        )
        wb = load_workbook(out)
        assert not any("Comp" in n for n in wb.sheetnames)

    def test_data_rows_written(self, tmp_path):
        out = build_resumen_xlsx(
            output_path=tmp_path / "resumen.xlsx",
            sheets_por_generico={"CERVEZAS": self._simple_gen_sheet()},
            sheets_mensuales={},
            sheet_comparativo=None,
        )
        wb = load_workbook(out)
        ws = wb["CERVEZAS"]
        # Header row + 2 data rows
        assert ws.max_row == 3
        # First data row NOA NORTE / Ene / 100 / 500
        assert ws.cell(2, 1).value == "NOA NORTE"
        assert ws.cell(2, 3).value == 100
