"""Tests for the historico-cliente CLI wrapper (scripts/historico_cliente_cli.py).

The wrapper exists so an agent can run the report from one command instead of
composing a JSON config by hand. Everything the config gets wrong silently —
the CCU universe, the composite client key, the capture range — is decided here
and covered by these tests.
"""
import importlib.util
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest

_CLI_PATH = Path(__file__).resolve().parent.parent / "scripts" / "historico_cliente_cli.py"


def _cli():
    """Import the wrapper by path — scripts/ is not a package."""
    spec = importlib.util.spec_from_file_location("historico_cliente_cli", _CLI_PATH)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ── Ventana de fechas ────────────────────────────────────────────────────────

def test_ventana_sin_argumentos_cubre_anio_pasado_mas_el_actual():
    """Sin --meses/--anios/--desde: 1-ene del año pasado hasta hoy."""
    cli = _cli()
    desde, hasta = cli.ventana(hoy=date(2026, 8, 12))
    assert desde == "2025-01-01"
    assert hasta == "2026-08-12"


def test_ventana_por_defecto_cubre_12_meses_hasta_hoy():
    cli = _cli()
    desde, hasta = cli.ventana(meses=12, hoy=date(2026, 8, 11))
    # 12 meses inclusive: septiembre 2025 .. agosto 2026
    assert desde == "2025-09-01"
    assert hasta == "2026-08-11"


def test_ventana_respeta_meses_distintos():
    cli = _cli()
    desde, hasta = cli.ventana(meses=3, hoy=date(2026, 8, 11))
    assert desde == "2026-06-01"
    assert hasta == "2026-08-11"


def test_ventana_cruza_el_anio():
    cli = _cli()
    desde, _ = cli.ventana(meses=6, hoy=date(2026, 2, 15))
    assert desde == "2025-09-01"


def test_ventana_acepta_fechas_explicitas():
    cli = _cli()
    assert cli.ventana(desde="2024-01-01", hasta="2026-12-31") == (
        "2024-01-01", "2026-12-31",
    )


def test_ventana_por_anios_arranca_en_enero_del_primero():
    """`--anios 2024 2025 2026` cubre desde el 1 de enero del más viejo."""
    cli = _cli()
    desde, _ = cli.ventana(anios=[2025, 2024, 2026], hoy=date(2026, 8, 12))
    assert desde == "2024-01-01"


def test_ventana_por_anios_no_proyecta_al_futuro():
    """Un año en curso corta hoy, no el 31 de diciembre."""
    cli = _cli()
    _, hasta = cli.ventana(anios=[2024, 2025, 2026], hoy=date(2026, 8, 12))
    assert hasta == "2026-08-12"


def test_ventana_por_anios_pasados_llega_a_fin_de_anio():
    cli = _cli()
    _, hasta = cli.ventana(anios=[2024, 2025], hoy=date(2026, 8, 12))
    assert hasta == "2025-12-31"


# ── Config ───────────────────────────────────────────────────────────────────

def test_config_usa_el_universo_ccu_y_modo_agrupado():
    """The report only means anything with the grouped + full-universe combo."""
    cli = _cli()
    cfg = cli.construir_config(
        clientes=[(7255, 1)], desde="2025-08-01", hasta="2026-07-31",
        solo_con_cargo=False, nombre="X",
    )
    assert cfg.agrupar_por_generico is True
    assert cfg.marcas_completas is True
    assert cfg.genericos_universo == [
        "CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES",
    ]


def test_config_lleva_la_clave_compuesta():
    """id_cliente alone is not a key — id_sucursal must travel with it."""
    cli = _cli()
    cfg = cli.construir_config(
        clientes=[(7255, 3)], desde="2025-08-01", hasta="2026-07-31",
        solo_con_cargo=False, nombre="X",
    )
    assert cfg.clientes == [{"id_cliente": 7255, "id_sucursal": 3}]


def test_config_admite_varios_clientes_en_orden():
    """One sheet per client, in the order given."""
    cli = _cli()
    cfg = cli.construir_config(
        clientes=[(30158, 4), (30056, 4), (30253, 4)],
        desde="2024-01-01", hasta="2026-08-12",
        solo_con_cargo=True, nombre="X",
    )
    assert cfg.clientes == [
        {"id_cliente": 30158, "id_sucursal": 4},
        {"id_cliente": 30056, "id_sucursal": 4},
        {"id_cliente": 30253, "id_sucursal": 4},
    ]


def test_config_propaga_solo_con_cargo():
    cli = _cli()
    cfg = cli.construir_config(
        clientes=[(1, 1)], desde="2025-08-01", hasta="2026-07-31",
        solo_con_cargo=True, nombre="X",
    )
    assert cfg.solo_con_cargo is True


# ── Parseo de códigos ────────────────────────────────────────────────────────

def test_parse_cliente_codigo_solo():
    cli = _cli()
    assert cli.parse_cliente("30158") == (30158, None)


def test_parse_cliente_con_sucursal():
    """`ID:SUC` fija la sucursal de ese código sin afectar a los otros."""
    cli = _cli()
    assert cli.parse_cliente("30158:4") == (30158, 4)


def test_parse_cliente_basura_falla():
    cli = _cli()
    with pytest.raises(ValueError):
        cli.parse_cliente("no-es-un-codigo")


# ── Resolución de cliente ────────────────────────────────────────────────────

def _loader_con(df: pd.DataFrame):
    loader = MagicMock()
    loader.execute_query.return_value = df
    return loader


def test_resolver_cliente_devuelve_el_unico_match():
    cli = _cli()
    loader = _loader_con(pd.DataFrame({
        "id_cliente": [7255], "id_sucursal": [1], "nombre": ["EL ENCUENTRO"],
    }))
    assert cli.resolver_cliente(loader, 7255) == (7255, 1, "EL ENCUENTRO")


def test_resolver_cliente_exige_sucursal_si_hay_varias():
    """The same id_cliente exists in several sucursales — never guess."""
    cli = _cli()
    loader = _loader_con(pd.DataFrame({
        "id_cliente": [100, 100], "id_sucursal": [1, 3],
        "nombre": ["UNO", "OTRO"],
    }))
    with pytest.raises(cli.ClienteAmbiguo) as exc:
        cli.resolver_cliente(loader, 100)
    assert "1" in str(exc.value) and "3" in str(exc.value)


def test_resolver_cliente_con_sucursal_explicita_desambigua():
    cli = _cli()
    loader = _loader_con(pd.DataFrame({
        "id_cliente": [100, 100], "id_sucursal": [1, 3],
        "nombre": ["UNO", "OTRO"],
    }))
    assert cli.resolver_cliente(loader, 100, id_sucursal=3) == (100, 3, "OTRO")


def test_resolver_cliente_inexistente_falla_claro():
    cli = _cli()
    with pytest.raises(cli.ClienteNoEncontrado):
        cli.resolver_cliente(_loader_con(pd.DataFrame()), 999999)


# ── Rango de captura ─────────────────────────────────────────────────────────

def test_rango_se_deriva_de_la_hoja_no_se_hardcodea(tmp_path):
    """The wrapper must ask the sheet for its range, not carry a constant."""
    cli = _cli()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws.cell(row=9, column=4, value="y")  # usa hasta D9

    assert cli.rango_de(ws) == "A1:D9"


def test_falla_si_una_columna_desborda(tmp_path):
    """A column that would render as ### aborts the run instead of sending."""
    cli = _cli()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"

    with patch.object(cli, "columnas_desbordadas", return_value=["C", "F"]):
        with pytest.raises(cli.CapturaInvalida) as exc:
            cli.validar_captura(ws)
    assert "C" in str(exc.value) and "F" in str(exc.value)
