"""
Tests for HistoricoClienteService.
"""
import logging
import pandas as pd
import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch

from src.core.data_loader import DataLoader
from src.services.historico_cliente import (
    HistoricoClienteConfig,
    HistoricoClienteService,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_two_client_df():
    """Two clients, two months, marcas as row_key."""
    return pd.DataFrame({
        "id_cliente": [1, 1, 2, 2],
        "id_sucursal": [1, 1, 1, 1],
        "nombre_cliente": ["Cliente A", "Cliente A", "Cliente B", "Cliente B"],
        "row_key": ["BRANCA", "BRANCA", "BRANCA", "BRAHMA"],
        "mes": ["2026-01", "2026-02", "2026-01", "2026-02"],
        "bultos": [10.0, 20.0, 5.0, 15.0],
    })


def _base_config(**kwargs):
    defaults = dict(
        fecha_desde="2026-01-01",
        fecha_hasta="2026-02-28",
        clientes=[
            {"id_cliente": 1, "id_sucursal": 1},
            {"id_cliente": 2, "id_sucursal": 1},
        ],
        marcas=["BRANCA", "BRAHMA"],
        nombre_archivo="Test",
    )
    defaults.update(kwargs)
    return HistoricoClienteConfig(**defaults)


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_happy_path_marcas(tmp_path):
    """Two clients with marcas filter → 2 sheets generated, file exists."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_two_client_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config()

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    assert result.ruta_archivo.exists()
    assert len(result.sheets_generated) == 2


def test_happy_path_articulos(tmp_path):
    """Articulos filter → row labels use 'Articulo' column name."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1, 1],
        "id_sucursal": [1, 1],
        "nombre_cliente": ["Cliente A", "Cliente A"],
        "row_key": ["12345 CERVEZA", "67890 AGUA"],
        "mes": ["2026-01", "2026-02"],
        "bultos": [10.0, 20.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        articulos=[12345, 67890],
        marcas=None,
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    # Verify the sheet was generated
    assert result.ruta_archivo.exists()
    assert len(result.sheets_generated) == 1

    # Verify the column label is "Articulo" (not "Marca")
    from openpyxl import load_workbook
    wb = load_workbook(result.ruta_archivo)
    ws = wb.active
    headers = _headers(ws)
    assert "Articulo" in headers
    assert "Marca" not in headers


def test_both_filters_raises(tmp_path):
    """Config with both articulos and marcas → ValueError before any DB call."""
    loader = MagicMock(spec=DataLoader)

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(articulos=[1, 2], marcas=["BRANCA"])

    with pytest.raises(ValueError, match="articulos"):
        service.generar_reporte(config)

    loader.get_ventas_historico_cliente.assert_not_called()


def test_neither_filter_raises(tmp_path):
    """Config with neither articulos nor marcas → ValueError."""
    loader = MagicMock(spec=DataLoader)

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(articulos=None, marcas=None)

    with pytest.raises(ValueError):
        service.generar_reporte(config)


def test_client_empty_skipped(tmp_path, caplog):
    """Client with no data in mock → 1 sheet generated, 1 warning logged."""
    loader = MagicMock(spec=DataLoader)
    # Only client 1 has data; client 2 is absent from the DataFrame
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1],
        "id_sucursal": [1],
        "nombre_cliente": ["Cliente A"],
        "row_key": ["BRANCA"],
        "mes": ["2026-01"],
        "bultos": [10.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config()  # still requests 2 clients

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        with caplog.at_level(logging.WARNING, logger="src.services.historico_cliente.service"):
            result = service.generar_reporte(config)

    assert len(result.sheets_generated) == 1
    assert any("sin datos" in record.message for record in caplog.records)


def test_all_months_covered(tmp_path):
    """Mock returns data for 2 of 4 months → pivot has all 4 columns, missing filled with 0."""
    loader = MagicMock(spec=DataLoader)
    # Data only for Jan and Mar; Feb and Apr are missing
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1, 1],
        "id_sucursal": [1, 1],
        "nombre_cliente": ["Cliente A", "Cliente A"],
        "row_key": ["BRANCA", "BRANCA"],
        "mes": ["2026-01", "2026-03"],
        "bultos": [10.0, 30.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        fecha_desde="2026-01-01",
        fecha_hasta="2026-04-30",
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    assert result.ruta_archivo.exists()

    from openpyxl import load_workbook
    wb = load_workbook(result.ruta_archivo)
    ws = wb.active
    headers = _headers(ws)

    for month in ("2026-01", "2026-02", "2026-03", "2026-04"):
        assert month in headers, f"Missing month column: {month}"

    # Find Feb column and verify its value is 0
    feb_col = headers.index("2026-02") + 1  # 1-based
    data_row = 2  # first data row
    assert ws.cell(row=data_row, column=feb_col).value == 0


def _make_grouped_df():
    """One client, two genericos (CERVEZAS: SALTA, HEINEKEN; VINOS: TORO), two months.

    Row grain matches loader output in ``agrupar_por_generico`` mode:
    columns include ``generico`` and ``row_key`` (= marca).
    """
    return pd.DataFrame({
        "id_cliente": [1] * 6,
        "id_sucursal": [1] * 6,
        "nombre_cliente": ["Cli A"] * 6,
        "generico": ["CERVEZAS", "CERVEZAS", "CERVEZAS", "CERVEZAS", "VINOS", "VINOS"],
        "row_key": ["SALTA", "SALTA", "HEINEKEN", "HEINEKEN", "TORO", "TORO"],
        "mes": ["2026-01", "2026-02", "2026-01", "2026-02", "2026-01", "2026-02"],
        "bultos": [100.0, 120.0, 30.0, 40.0, 10.0, 5.0],
    })


def _header_row(ws) -> int:
    """Worksheet row holding the column headers.

    Grouped sheets carry a title band above the table, so the header is not on
    row 1; the marca/articulo sheets still are. Locating it by content keeps
    both paths working from one helper.
    """
    # Grouped sheets put the label in column 2 (column 1 is Genérico); the
    # marca/articulo sheets put it in column 1.
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            if ws.cell(row=r, column=c).value in ("Marca", "Articulo"):
                return r
    raise AssertionError("no encontre la fila de encabezados")


def _headers(ws) -> list:
    fila = _header_row(ws)
    return [ws.cell(row=fila, column=c).value for c in range(1, ws.max_column + 1)]


def _read_marca_column(ruta):
    """Return the values of the 'Marca' column (data rows only) of the active sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(ruta)
    ws = wb.active
    fila_hdr = _header_row(ws)
    headers = _headers(ws)
    marca_col = headers.index("Marca") + 1
    total_col = headers.index("Total") + 1
    rows = []
    for r in range(fila_hdr + 1, ws.max_row + 1):
        rows.append((ws.cell(row=r, column=marca_col).value, ws.cell(row=r, column=total_col).value))
    return rows


def test_grouped_by_generico_adds_subtotals(tmp_path):
    """agrupar_por_generico=True → per-generico subtotal rows + grand total row."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    assert result.ruta_archivo.exists()

    # Loader must be told to group by generico
    _, kwargs = loader.get_ventas_historico_cliente.call_args
    assert kwargs.get("agrupar_por_generico") is True

    rows = _read_marca_column(result.ruta_archivo)
    labels = {label for label, _ in rows}
    assert "TOTAL CERVEZAS" in labels
    assert "TOTAL VINOS" in labels
    assert "TOTAL GENERAL" in labels

    # Subtotal + grand-total values are correct
    totals = dict(rows)
    assert totals["TOTAL CERVEZAS"] == 290
    assert totals["TOTAL VINOS"] == 15
    assert totals["TOTAL GENERAL"] == 305


def test_grouped_mode_no_filter_required(tmp_path):
    """In grouped mode, neither marcas nor articulos is required (shows all marcas)."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)  # must NOT raise

    assert len(result.sheets_generated) == 1


def test_grouped_mode_generico_ordered_by_total(tmp_path):
    """Genericos are ordered by descending total; CERVEZAS (290) before VINOS (15)."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    labels = [label for label, _ in _read_marca_column(result.ruta_archivo)]
    assert labels.index("TOTAL CERVEZAS") < labels.index("TOTAL VINOS")
    # Grand total is the last row
    assert labels[-1] == "TOTAL GENERAL"


def test_marcas_completas_fills_universe(tmp_path):
    """marcas_completas=True → marcas del universo no compradas aparecen con Total 0."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()
    # Universe adds QUILMES (CERVEZAS) + ANDES (VINOS) the client never bought.
    loader.get_marca_universe.return_value = pd.DataFrame({
        "generico": ["CERVEZAS", "CERVEZAS", "CERVEZAS", "VINOS", "VINOS"],
        "marca": ["SALTA", "HEINEKEN", "QUILMES", "TORO", "ANDES"],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
        marcas_completas=True,
        genericos_universo=["CERVEZAS", "VINOS"],
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    loader.get_marca_universe.assert_called_once_with(["CERVEZAS", "VINOS"])

    rows = _read_marca_column(result.ruta_archivo)
    labels = {label for label, _ in rows}
    totals = dict(rows)
    # Never-bought marcas are present with 0
    assert "QUILMES" in labels and totals["QUILMES"] == 0
    assert "ANDES" in labels and totals["ANDES"] == 0
    # Bought marcas + subtotals unchanged by the zero-fill
    assert totals["TOTAL CERVEZAS"] == 290
    assert totals["TOTAL VINOS"] == 15
    assert totals["TOTAL GENERAL"] == 305


def _grouped_workbook(tmp_path):
    """Generate a grouped-mode report and return its active worksheet."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
    )

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    return load_workbook(result.ruta_archivo).active


def test_grouped_usa_fuente_agrandada(tmp_path):
    """Every cell of the grouped sheet is written at the report font size.

    The sheet is read as a WhatsApp image, so the default 11pt is too small
    once the capture is scaled down on a phone screen.
    """
    from src.services.historico_cliente.service import _FONT_SIZE

    ws = _grouped_workbook(tmp_path)
    assert _FONT_SIZE > 11, "el punto del cambio es agrandar sobre el default"

    # Scoped to the table: the title band above it is deliberately larger.
    for row in ws.iter_rows(min_row=_header_row(ws), max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            assert cell.font.size == _FONT_SIZE, (
                f"celda {cell.coordinate} quedo en {cell.font.size}pt"
            )


def test_grouped_usa_etiqueta_de_mes_corta(tmp_path):
    """Grouped headers read MM/YY, not YYYY-MM.

    Twelve month columns dominate the sheet width; two characters saved on each
    header is what buys back the room the bigger font costs.
    """
    ws = _grouped_workbook(tmp_path)
    headers = _headers(ws)

    assert "01/26" in headers and "02/26" in headers
    assert "2026-01" not in headers


def test_etiqueta_de_mes_larga_sobrevive_fuera_del_modo_agrupado(tmp_path):
    """The marcas/articulos path keeps the full YYYY-MM header."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_two_client_df()

    service = HistoricoClienteService(data_loader=loader)
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(_base_config())

    from openpyxl import load_workbook
    ws = load_workbook(result.ruta_archivo).active
    headers = _headers(ws)
    assert "2026-01" in headers


def test_grouped_columnas_de_mes_compactas(tmp_path):
    """Small-value month columns sit at the floor width, not auto-fit padding."""
    from openpyxl.utils import get_column_letter

    from src.services.historico_cliente.service import _WIDTH_MES_MIN

    ws = _grouped_workbook(tmp_path)
    headers = _headers(ws)

    meses = [h for h in headers if isinstance(h, str) and len(h) == 5 and h[2] == "/"]
    assert meses, "la hoja deberia tener columnas de mes MM/YY"

    # With only small values the column is sized by its own "MM/YY" header and
    # nothing more — never padded out the way auto-fit did.
    for mes in meses:
        letter = get_column_letter(headers.index(mes) + 1)
        ancho = ws.column_dimensions[letter].width
        assert _WIDTH_MES_MIN <= ancho <= _WIDTH_MES_MIN + 1, (
            f"{mes}: ancho {ancho} fuera del rango compacto"
        )


def test_mes_con_valor_grande_ensancha_su_columna(tmp_path):
    """A month column widens to fit its own largest value.

    Regression: with a fixed width of 7, the bold TOTAL GENERAL row rendered
    "132.87" as ### in the LibreOffice capture — the number silently vanished
    from the image while the xlsx was fine.
    """
    from openpyxl.utils import get_column_letter

    from src.services.historico_cliente.service import _FONT_SIZE, _WIDTH_MES_MIN

    loader = MagicMock(spec=DataLoader)
    # 2026-01 carries a 4-digit value; 2026-02 stays tiny.
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1, 1],
        "id_sucursal": [1, 1],
        "nombre_cliente": ["Cli A", "Cli A"],
        "generico": ["CERVEZAS", "CERVEZAS"],
        "row_key": ["SALTA", "SALTA"],
        "mes": ["2026-01", "2026-02"],
        # Se muestra "123,457": mas ancho que el header "01/26", que es lo que
        # obliga a la columna a crecer. Con #,##0 un valor de 4 digitos ya entra
        # en el ancho del header y no discriminaria nada.
        "bultos": [123456.78, 2.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None, articulos=None, agrupar_por_generico=True,
    )
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    ws = load_workbook(result.ruta_archivo).active
    headers = _headers(ws)

    ancho_grande = ws.column_dimensions[get_column_letter(headers.index("01/26") + 1)].width
    ancho_chico = ws.column_dimensions[get_column_letter(headers.index("02/26") + 1)].width

    # Se MUESTRA "123,457" (7 caracteres, formato #,##0); debe entrar completo.
    assert ancho_grande >= len("123,457") * (_FONT_SIZE / 11)
    # The quiet month is not dragged along with it.
    assert ancho_chico <= _WIDTH_MES_MIN + 1
    assert ancho_chico < ancho_grande


def test_grouped_columnas_de_texto_entran_sin_cortarse(tmp_path):
    """Genérico/Marca must be wide enough for their longest value at _FONT_SIZE.

    Regression: at 19/24 units the 14pt text clipped — "AGUAS DANONE" ran into
    the marca column and "TOTAL SIDRAS Y LICORES" lost its tail in the capture.
    Excel width units track the 11pt default, so a 14pt string needs 14/11 as
    many of them.
    """
    from openpyxl.utils import get_column_letter

    from src.services.historico_cliente.service import _FONT_SIZE

    # "SIDRAS Y LICORES" is the longest real generico, and its subtotal label
    # "TOTAL SIDRAS Y LICORES" (22 chars) is the widest string on the sheet.
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1, 1],
        "id_sucursal": [1, 1],
        "nombre_cliente": ["Cli A", "Cli A"],
        "generico": ["SIDRAS Y LICORES", "AGUAS DANONE"],
        "row_key": ["SAENZ BRIONES", "VILLAVICENCIO"],
        "mes": ["2026-01", "2026-01"],
        "bultos": [1.5, 2.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None,
        articulos=None,
        agrupar_por_generico=True,
    )
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    ws = load_workbook(result.ruta_archivo).active
    headers = _headers(ws)
    escala = _FONT_SIZE / 11

    for col_name in ("Genérico", "Marca"):
        idx = headers.index(col_name) + 1
        letter = get_column_letter(idx)
        mas_largo = max(
            len(str(ws.cell(row=r, column=idx).value or ""))
            for r in range(_header_row(ws), ws.max_row + 1)
        )
        requerido = mas_largo * escala
        assert ws.column_dimensions[letter].width >= requerido, (
            f"{col_name}: ancho {ws.column_dimensions[letter].width} < "
            f"{requerido:.1f} que necesita '{mas_largo} chars' a {_FONT_SIZE}pt"
        )


def test_grouped_filas_resumen_conservan_negrita_y_fuente(tmp_path):
    """Subtotal/grand-total rows keep bold AND pick up the bigger font size."""
    from src.services.historico_cliente.service import _FONT_SIZE

    ws = _grouped_workbook(tmp_path)
    headers = _headers(ws)
    marca_col = headers.index("Marca") + 1

    resumen = [
        r for r in range(_header_row(ws) + 1, ws.max_row + 1)
        if str(ws.cell(row=r, column=marca_col).value or "").startswith("TOTAL")
    ]
    assert resumen, "deberia haber filas TOTAL"

    for r in resumen:
        cell = ws.cell(row=r, column=marca_col)
        assert cell.font.bold is True, f"fila {r} perdio la negrita"
        assert cell.font.size == _FONT_SIZE, f"fila {r} quedo en {cell.font.size}pt"


def test_grouped_todas_las_celdas_llevan_borde(tmp_path):
    """Every cell of the table carries a thin border on all four sides.

    Without a grid the capture reads as loose floating numbers; the border is
    what makes a 15-column month matrix scannable.
    """
    ws = _grouped_workbook(tmp_path)

    for row in ws.iter_rows(min_row=_header_row(ws), max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            for lado in ("left", "right", "top", "bottom"):
                assert getattr(cell.border, lado).style is not None, (
                    f"celda {cell.coordinate} sin borde {lado}"
                )


def test_grouped_paleta_corporativa_azul(tmp_path):
    """Header, subtotal and grand-total use the blue palette — no amber left."""
    from src.services.historico_cliente.service import (
        _GRAND_FILL, _HEADER_FILL, _SUBTOTAL_FILL,
    )

    ws = _grouped_workbook(tmp_path)
    fila_hdr = _header_row(ws)
    headers = _headers(ws)
    marca_col = headers.index("Marca") + 1

    def _fill(cell):
        return (cell.fill.start_color.rgb or "")[-6:]

    assert _fill(ws.cell(row=fila_hdr, column=1)) == _HEADER_FILL
    assert ws.cell(row=fila_hdr, column=1).font.color.rgb[-6:] == "FFFFFF"

    subtotales = [
        r for r in range(fila_hdr + 1, ws.max_row + 1)
        if str(ws.cell(row=r, column=marca_col).value or "").startswith("TOTAL ")
        and ws.cell(row=r, column=marca_col).value != "TOTAL GENERAL"
    ]
    assert subtotales
    for r in subtotales:
        assert _fill(ws.cell(row=r, column=marca_col)) == _SUBTOTAL_FILL

    grand = next(
        r for r in range(fila_hdr + 1, ws.max_row + 1)
        if ws.cell(row=r, column=marca_col).value == "TOTAL GENERAL"
    )
    assert _fill(ws.cell(row=grand, column=marca_col)) == _GRAND_FILL

    # The old amber grand-total is gone.
    todos = {
        _fill(ws.cell(row=r, column=c))
        for r in range(fila_hdr, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    }
    assert "FFE08A" not in todos


def test_grouped_usa_fuente_corporativa(tmp_path):
    """One typeface across the whole sheet, declared by name.

    Arial is picked because LibreOffice on the render host substitutes it with
    Liberation Sans (metric-identical) while Excel on Windows uses real Arial —
    so the column widths hold in both. Calibri instead falls back to Noto Sans
    here, which is NOT metric-compatible.
    """
    from src.services.historico_cliente.service import _FONT_NAME

    ws = _grouped_workbook(tmp_path)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            assert cell.font.name == _FONT_NAME, (
                f"celda {cell.coordinate} usa {cell.font.name}"
            )


def test_grouped_lleva_titulo_con_el_cliente(tmp_path):
    """A title band above the table names the client and the period.

    The image gets forwarded on WhatsApp beyond the original caption, so it has
    to identify itself.
    """
    ws = _grouped_workbook(tmp_path)
    fila_hdr = _header_row(ws)
    assert fila_hdr > 1, "el titulo deberia ir arriba del encabezado"

    banda = " ".join(
        str(ws.cell(row=r, column=c).value or "")
        for r in range(1, fila_hdr)
        for c in range(1, ws.max_column + 1)
    )
    assert "Cli A" in banda, "el titulo deberia nombrar al cliente"
    assert "2026-01" in banda and "2026-02" in banda, "y el periodo cubierto"


def _subtitulo(ws) -> str:
    """Text of the subtitle line, above the header band."""
    return " ".join(
        str(ws.cell(row=r, column=c).value or "")
        for r in range(1, _header_row(ws))
        for c in range(1, ws.max_column + 1)
    )


def test_subtitulo_declara_que_excluye_los_sin_cargo(tmp_path):
    """With solo_con_cargo the sheet must say so on its face.

    The number changes materially (gifts can be >15% of a month), and the image
    travels beyond the message that requested it, so the basis has to be stated
    inside the capture — not only in the WhatsApp caption.
    """
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = _make_grouped_df()

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None, articulos=None,
        agrupar_por_generico=True, solo_con_cargo=True,
    )
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    texto = _subtitulo(load_workbook(result.ruta_archivo).active).lower()
    assert "con cargo" in texto
    assert "sin cargo" in texto or "bonificad" in texto


def test_subtitulo_no_habla_de_cargos_cuando_suma_todo(tmp_path):
    """Default run includes gifts, so it must NOT claim to exclude them."""
    ws = _grouped_workbook(tmp_path)  # solo_con_cargo=False
    assert "con cargo" not in _subtitulo(ws).lower()


def _multianio_workbook(tmp_path):
    """Grouped sheet spanning 2024-2026 with one marca and known monthly values."""
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1] * 4,
        "id_sucursal": [1] * 4,
        "nombre_cliente": ["Cli A"] * 4,
        "generico": ["CERVEZAS"] * 4,
        "row_key": ["SALTA"] * 4,
        "mes": ["2024-01", "2024-03", "2025-05", "2026-02"],
        "bultos": [10.0, 5.0, 20.0, 7.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        fecha_desde="2024-01-01", fecha_hasta="2026-03-31",
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None, articulos=None, agrupar_por_generico=True,
    )
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    return load_workbook(result.ruta_archivo).active


def test_totales_por_anio_una_columna_por_anio(tmp_path):
    """A window spanning several years gets one total column per year."""
    ws = _multianio_workbook(tmp_path)
    headers = _headers(ws)
    for anio in ("2024", "2025", "2026"):
        assert f"Total {anio}" in headers, f"falta la columna del {anio}"


def test_total_del_anio_va_pegado_a_sus_meses(tmp_path):
    """Each year total sits right after that year's last month, not at the end.

    Reading a year as a block is the point; a column parked at the far right
    forces the eye to travel across 30+ month columns to pair them up.
    """
    headers = _headers(_multianio_workbook(tmp_path))
    assert headers[headers.index("Total 2024") - 1] == "12/24"
    assert headers[headers.index("Total 2025") - 1] == "12/25"
    # 2026 se corta en marzo por la ventana pedida
    assert headers[headers.index("Total 2026") - 1] == "03/26"


def test_total_del_anio_suma_solo_ese_anio(tmp_path):
    """10 + 5 en 2024, 20 en 2025, 7 en 2026 — cada uno en su columna."""
    ws = _multianio_workbook(tmp_path)
    headers = _headers(ws)
    fila = _header_row(ws) + 1  # primera fila de datos: SALTA

    def _valor(col):
        return ws.cell(row=fila, column=headers.index(col) + 1).value

    assert _valor("Total 2024") == 15
    assert _valor("Total 2025") == 20
    assert _valor("Total 2026") == 7
    assert _valor("Total") == 42


def test_totales_por_anio_tambien_en_las_filas_resumen(tmp_path):
    """Subtotal and grand-total rows carry the year columns too."""
    ws = _multianio_workbook(tmp_path)
    headers = _headers(ws)
    col = headers.index("Total 2024") + 1
    marca_col = headers.index("Marca") + 1

    grand = next(
        r for r in range(_header_row(ws) + 1, ws.max_row + 1)
        if ws.cell(row=r, column=marca_col).value == "TOTAL GENERAL"
    )
    assert ws.cell(row=grand, column=col).value == 15


def test_columna_de_anio_se_distingue_de_los_meses(tmp_path):
    """Year totals get their own tint — otherwise they vanish among 30+ months."""
    from src.services.historico_cliente.service import (
        _TOTAL_ANIO_FILL, _TOTAL_COL_FILL,
    )

    ws = _multianio_workbook(tmp_path)
    headers = _headers(ws)
    fila = _header_row(ws) + 1

    def _fill(col):
        celda = ws.cell(row=fila, column=headers.index(col) + 1)
        return (celda.fill.start_color.rgb or "")[-6:]

    assert _fill("Total 2024") == _TOTAL_ANIO_FILL
    assert _fill("01/24") != _TOTAL_ANIO_FILL
    # y tampoco se confunde con el Total general de la derecha
    assert _TOTAL_ANIO_FILL != _TOTAL_COL_FILL


def test_un_solo_anio_no_agrega_columna_redundante(tmp_path):
    """Within one calendar year the grand Total already is the year total."""
    ws = _grouped_workbook(tmp_path)  # ventana 2026-01..2026-02
    assert not [h for h in _headers(ws) if str(h).startswith("Total 20")]


def test_numeros_se_muestran_sin_decimales(tmp_path):
    """Every numeric cell carries the 0-decimal format."""
    ws = _grouped_workbook(tmp_path)
    headers = _headers(ws)
    fila = _header_row(ws) + 1

    for col in range(3, len(headers) + 1):  # de la primera columna de mes en adelante
        celda = ws.cell(row=fila, column=col)
        assert celda.number_format == "#,##0", (
            f"{celda.coordinate} quedo en {celda.number_format!r}"
        )


def test_el_valor_guardado_conserva_sus_decimales(tmp_path):
    """Displaying integers must NOT round the stored number.

    The cell keeps the exact figure from the database; only its display is
    integer. Rounding in Python would make the xlsx disagree with the source
    and the column totals stop adding up.
    """
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1], "id_sucursal": [1], "nombre_cliente": ["Cli A"],
        "generico": ["CERVEZAS"], "row_key": ["SALTA"],
        "mes": ["2026-01"], "bultos": [10.416667],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(
        clientes=[{"id_cliente": 1, "id_sucursal": 1}],
        marcas=None, articulos=None, agrupar_por_generico=True,
    )
    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    from openpyxl import load_workbook
    ws = load_workbook(result.ruta_archivo).active
    headers = _headers(ws)
    valor = ws.cell(row=_header_row(ws) + 1, column=headers.index("01/26") + 1).value

    assert valor == pytest.approx(10.416667), f"el valor se redondeo a {valor}"
    assert valor != 10, "no debe guardarse el entero"


def test_rango_captura_cubre_toda_la_hoja(tmp_path):
    """The capture range is derived from the sheet, never hardcoded.

    A fixed range is how the TOTAL GENERAL row silently disappears from the
    image: the xlsx stays right, LibreOffice just crops what the range leaves
    out and nothing errors.
    """
    from openpyxl.utils import get_column_letter

    from src.services.historico_cliente.service import rango_captura

    ws = _grouped_workbook(tmp_path)
    esperado = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    assert rango_captura(ws) == esperado


def test_rango_captura_incluye_la_fila_total_general(tmp_path):
    """The last data row — TOTAL GENERAL — must fall inside the range."""
    import re

    from src.services.historico_cliente.service import rango_captura

    ws = _grouped_workbook(tmp_path)
    ultima = int(re.search(r"(\d+)$", rango_captura(ws)).group(1))

    fila_total = max(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=2).value == "TOTAL GENERAL"
    )
    assert fila_total <= ultima


def test_columnas_desbordadas_detecta_ancho_insuficiente(tmp_path):
    """A column too narrow for its own content is reported.

    LibreOffice renders an overflowing number as ### in the capture, so this
    has to be caught before the image is sent, not after.
    """
    from src.services.historico_cliente.service import columnas_desbordadas

    ws = _grouped_workbook(tmp_path)
    assert columnas_desbordadas(ws) == []

    ws.column_dimensions["C"].width = 1  # a un mes no le entra ni un digito
    assert "C" in columnas_desbordadas(ws)


def test_nombre_cliente_ignora_fantasia_vacia():
    """`fantasia` empty-string must fall back to razon_social, not to ''.

    Regression: plain COALESCE only skips NULL. Clients loaded with an empty
    `fantasia` (e.g. 201160 GONZALEZ ELBA CONCEPCION) returned '', which the
    service then replaced with the `id-sucursal` fallback, so the sheet was
    named "201160-1" instead of the client's name.
    """
    loader = DataLoader.__new__(DataLoader)  # no DB connection needed
    captured = {}

    def _fake(query, params=None):
        captured["sql"] = query
        return pd.DataFrame()

    loader.execute_query = _fake
    loader.get_ventas_historico_cliente(
        fecha_desde="2026-01-01",
        fecha_hasta="2026-01-31",
        clientes=[{"id_cliente": 201160, "id_sucursal": 1}],
        agrupar_por_generico=True,
    )

    assert "NULLIF(TRIM(dc.fantasia), '')" in captured["sql"]


def test_long_client_name_truncated(tmp_path):
    """Client with 40-char nombre_cliente → sheet name is exactly 31 chars."""
    long_name = "A" * 40
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_historico_cliente.return_value = pd.DataFrame({
        "id_cliente": [1],
        "id_sucursal": [1],
        "nombre_cliente": [long_name],
        "row_key": ["BRANCA"],
        "mes": ["2026-01"],
        "bultos": [10.0],
    })

    service = HistoricoClienteService(data_loader=loader)
    config = _base_config(clientes=[{"id_cliente": 1, "id_sucursal": 1}])

    with patch("src.services.historico_cliente.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    assert len(result.sheets_generated) == 1
    assert len(result.sheets_generated[0]) == 31
