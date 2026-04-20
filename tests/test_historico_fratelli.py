"""
Tests for historico_fratelli service.
"""
import pandas as pd
import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch
from openpyxl import Workbook

import config.settings as _settings
from src.core.data_loader import DataLoader
from src.services.historico_fratelli.service import (
    HistoricoFratelliConfig,
    HistoricoFratelliService,
    _pivot_por_mes,
    _pivot_por_dimension,
    _write_section_title,
    _write_year_block_simple,
    _write_year_block_with_rows,
    MESES,
    ANIOS,
)


def _make_test_data():
    return pd.DataFrame({
        "anio": [2024, 2024, 2024, 2025, 2025, 2026],
        "mes": [1, 2, 1, 1, 3, 6],
        "marca": ["FRATELLI", "FRATELLI", "OTRA", "FRATELLI", "OTRA", "FRATELLI"],
        "id_lista_precio": [1, 1, 2, 1, 2, 1],
        "cantidad": [100, 200, 50, 300, 75, 150],
        "descuentos": [500, 1000, 250, 1500, 375, 750],
    })


# ── TestPivotPorMes ─────────────────────────────────────────

class TestPivotPorMes:
    def test_aggregates_by_year_and_month(self):
        df = _make_test_data()
        result = _pivot_por_mes(df)
        assert set(result.keys()) == {2024, 2025, 2026}
        assert len(result[2024]) == 12
        assert len(result[2025]) == 12
        assert len(result[2026]) == 12

    def test_correct_values(self):
        df = _make_test_data()
        result = _pivot_por_mes(df)
        # 2024 Jan: 100 (FRATELLI) + 50 (OTRA) = 150
        assert result[2024][0] == 150
        # 2024 Feb: 200
        assert result[2024][1] == 200
        # 2025 Jan: 300
        assert result[2025][0] == 300
        # 2026 Jun: 150
        assert result[2026][5] == 150

    def test_missing_months_are_zero(self):
        df = pd.DataFrame({
            "anio": [2024, 2024],
            "mes": [1, 2],
            "marca": ["A", "A"],
            "id_lista_precio": [1, 1],
            "cantidad": [100, 200],
        })
        result = _pivot_por_mes(df)
        # Months 3-12 should be 0 for 2024
        for i in range(2, 12):
            assert result[2024][i] == 0

    def test_empty_dataframe(self):
        df = pd.DataFrame(columns=["anio", "mes", "marca", "id_lista_precio", "cantidad"])
        result = _pivot_por_mes(df)
        assert result == {2024: [0] * 12, 2025: [0] * 12, 2026: [0] * 12}


# ── TestPivotPorDimension ───────────────────────────────────

class TestPivotPorDimension:
    def test_pivots_by_marca(self):
        df = _make_test_data()
        result = _pivot_por_dimension(df, "marca")
        df_2024 = result[2024]
        assert not df_2024.empty
        # Should have marca column + 12 month columns = 13 cols
        assert len(df_2024.columns) == 13
        # Should have 2 unique marcas in 2024
        assert len(df_2024) == 2
        assert "marca" in df_2024.columns

    def test_missing_months_filled_zero(self):
        df = pd.DataFrame({
            "anio": [2024],
            "mes": [1],
            "marca": ["A"],
            "id_lista_precio": [1],
            "cantidad": [100],
        })
        result = _pivot_por_dimension(df, "marca")
        df_2024 = result[2024]
        assert not df_2024.empty
        # Month 2 (index 2) should be 0
        assert df_2024.iloc[0][2] == 0

    def test_empty_year(self):
        df = _make_test_data()
        # 2026 has only 1 entry with marca FRATELLI
        result = _pivot_por_dimension(df, "marca")
        df_2026 = result[2026]
        assert not df_2026.empty
        assert len(df_2026) == 1

    def test_empty_dataframe(self):
        df = pd.DataFrame(columns=["anio", "mes", "marca", "id_lista_precio", "cantidad"])
        result = _pivot_por_dimension(df, "marca")
        for anio in ANIOS:
            assert result[anio].empty


# ── TestWriteYearBlockSimple ────────────────────────────────

class TestWriteYearBlockSimple:
    def test_writes_year_and_months(self):
        wb = Workbook()
        ws = wb.active
        values = list(range(12))
        next_row = _write_year_block_simple(ws, 1, 2024, values)
        # Row 1: year label
        assert ws.cell(row=1, column=1).value == 2024
        # Row 2: month headers
        assert ws.cell(row=2, column=1).value == "Ene"
        assert ws.cell(row=2, column=12).value == "Dic"
        # next_row should be > 3 (1 year + 1 headers + 1 data + 1 spacing)
        assert next_row > 3

    def test_values_written(self):
        wb = Workbook()
        ws = wb.active
        values = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120]
        _write_year_block_simple(ws, 1, 2024, values)
        # Row 3: data
        for i, val in enumerate(values):
            assert ws.cell(row=3, column=i + 1).value == val


# ── TestWriteYearBlockWithRows ──────────────────────────────

class TestWriteYearBlockWithRows:
    def _make_pivot_df(self):
        """Create a simple pivot DataFrame for testing."""
        rows = [
            {"marca": "FRATELLI", **{m: m * 10 for m in range(1, 13)}},
            {"marca": "OTRA", **{m: m * 5 for m in range(1, 13)}},
        ]
        return pd.DataFrame(rows, columns=["marca"] + list(range(1, 13)))

    def test_writes_header_and_rows(self):
        wb = Workbook()
        ws = wb.active
        df = self._make_pivot_df()
        next_row = _write_year_block_with_rows(ws, 1, 2024, df, "marca")
        # Row 1: year label
        assert ws.cell(row=1, column=1).value == 2024
        # Row 2: header row with dim_col name
        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=2, column=2).value == "Ene"
        # Row 3: first data row
        assert ws.cell(row=3, column=1).value == "FRATELLI"
        # next_row > 4 (year + header + 2 data rows + spacing)
        assert next_row > 4

    def test_empty_df_writes_just_header(self):
        wb = Workbook()
        ws = wb.active
        next_row = _write_year_block_with_rows(ws, 1, 2024, pd.DataFrame(), "marca")
        # Row 1: year label
        assert ws.cell(row=1, column=1).value == 2024
        # Row 2: header
        assert ws.cell(row=2, column=2).value == "Ene"
        # No data rows written
        assert ws.cell(row=3, column=1).value is None
        # next_row returned correctly
        assert next_row > 2


# ── TestHistoricoFratelliService ────────────────────────────

class TestHistoricoFratelliService:
    def test_generates_excel_with_3_sections(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_historico_fratelli.return_value = _make_test_data()

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            service = HistoricoFratelliService(data_loader=mock_loader)
            config = HistoricoFratelliConfig(nombre_archivo="test_historico")
            result = service.generar_reporte(config)

        assert result.ruta_archivo.exists()
        assert result.ruta_archivo.suffix == ".xlsx"
        mock_loader.get_ventas_historico_fratelli.assert_called_once()

    def test_empty_data_produces_valid_excel(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_historico_fratelli.return_value = pd.DataFrame(
            columns=["anio", "mes", "marca", "id_lista_precio", "cantidad", "descuentos"]
        )

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            service = HistoricoFratelliService(data_loader=mock_loader)
            config = HistoricoFratelliConfig(nombre_archivo="test_empty")
            result = service.generar_reporte(config)

        assert result.ruta_archivo.exists()
        assert result.registros_procesados == 0

    def test_result_fields(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_historico_fratelli.return_value = _make_test_data()

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            service = HistoricoFratelliService(data_loader=mock_loader)
            config = HistoricoFratelliConfig(nombre_archivo="test_fields")
            result = service.generar_reporte(config)

        assert isinstance(result.ruta_archivo, Path)
        assert isinstance(result.hojas, list)
        assert len(result.hojas) > 0
        assert result.registros_procesados == len(_make_test_data())
