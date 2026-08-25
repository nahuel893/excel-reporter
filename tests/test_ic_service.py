"""Tests del servicio y del constructor del libro.

Sin base de datos: los analisis se reemplazan por AnalysisResult armados a mano.
Lo que se prueba aca es la orquestacion y el armado del Excel, no la estadistica.
"""
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.services.inteligencia_comercial import service as svc
from src.services.inteligencia_comercial.contracts import (
    Alert,
    AnalysisResult,
    Headline,
)
from src.services.inteligencia_comercial.excel_builder import SheetSpec, build_workbook
from src.services.inteligencia_comercial.service import (
    InteligenciaComercialConfig,
    InteligenciaComercialService,
)


def _resultado(nombre="Prueba", con_tabla=True, failed=False):
    tablas = {}
    if con_tabla:
        tablas["principal"] = pd.DataFrame(
            {
                "sucursal": ["CASA CENTRAL", "ORAN", "TOTAL GENERAL"],
                "bultos": [100.0, 50.0, 150.0],
                "pct_neto": [0.6667, 0.3333, 1.0],
            }
        )
    return AnalysisResult(
        name=nombre,
        tables=tablas,
        headlines=[Headline("Bultos", 150.0, "#,##0", delta=0.12, note="vs periodo previo")],
        alerts=[Alert("critica", "Algo pasa", "Detalle con un numero: 150 bultos", 150.0)],
        notes=[f"Nota de metodologia de {nombre}."],
        failed=failed,
    )


class TestConfig:
    def test_deriva_el_nombre_de_archivo(self):
        config = InteligenciaComercialConfig(fecha_hasta="2026-07-30")
        assert config.nombre_archivo == "Inteligencia Comercial - 2026-07-30"

    def test_rechaza_una_fecha_mal_formada(self):
        with pytest.raises(ValueError, match="YYYY-MM-DD"):
            InteligenciaComercialConfig(fecha_hasta="30/07/2026")

    def test_rechaza_una_historia_mas_corta_que_la_ventana(self):
        with pytest.raises(ValueError, match="meses_historia"):
            InteligenciaComercialConfig(
                fecha_hasta="2026-07-30", meses_ventana=24, meses_historia=12
            )

    def test_rechaza_un_modulo_inexistente(self):
        with pytest.raises(ValueError, match="modulos desconocidos"):
            InteligenciaComercialConfig(fecha_hasta="2026-07-30", modulos=["inventado"])

    def test_acepta_un_subconjunto_de_modulos(self):
        config = InteligenciaComercialConfig(fecha_hasta="2026-07-30", modulos=["clientes"])
        assert config.modulos == ["clientes"]


class TestOrquestacion:
    def test_un_analisis_que_falla_no_tumba_el_reporte(self, tmp_path):
        # el resto del informe tiene que entregarse igual: es la razon de ser
        # de que build() devuelva failed en vez de levantar
        bueno, malo = MagicMock(), MagicMock()
        bueno.build.return_value = _resultado("bueno")
        malo.build.side_effect = RuntimeError("la BD se cayo")

        servicio = InteligenciaComercialService(data_loader=MagicMock())
        with patch.dict(svc.MODULOS, {"clientes": bueno, "demanda": malo}, clear=True), \
             patch.object(InteligenciaComercialService, "_output_dir", return_value=tmp_path):
            resultado = servicio.generar_reporte(
                InteligenciaComercialConfig(fecha_hasta="2026-07-30",
                                            modulos=["clientes", "demanda"])
            )

        assert resultado.analisis_ok == ["clientes"]
        assert resultado.analisis_fallidos == ["demanda"]
        assert resultado.ruta_archivo.exists()

    def test_corre_solo_los_modulos_pedidos(self, tmp_path):
        uno, dos = MagicMock(), MagicMock()
        uno.build.return_value = _resultado("uno")
        dos.build.return_value = _resultado("dos")

        servicio = InteligenciaComercialService(data_loader=MagicMock())
        with patch.dict(svc.MODULOS, {"clientes": uno, "demanda": dos}, clear=True), \
             patch.object(InteligenciaComercialService, "_output_dir", return_value=tmp_path):
            servicio.generar_reporte(
                InteligenciaComercialConfig(fecha_hasta="2026-07-30", modulos=["clientes"])
            )

        uno.build.assert_called_once()
        dos.build.assert_not_called()

    def test_el_contexto_recibe_la_ventana_configurada(self, tmp_path):
        modulo = MagicMock()
        modulo.build.return_value = _resultado()
        servicio = InteligenciaComercialService(data_loader=MagicMock())
        with patch.dict(svc.MODULOS, {"clientes": modulo}, clear=True), \
             patch.object(InteligenciaComercialService, "_output_dir", return_value=tmp_path):
            servicio.generar_reporte(
                InteligenciaComercialConfig(
                    fecha_hasta="2026-07-30", meses_ventana=6, meses_historia=18,
                    modulos=["clientes"],
                )
            )
        ctx = modulo.build.call_args[0][0]
        assert ctx.fecha_hasta == "2026-07-30"
        assert ctx.meses_ventana == 6
        assert ctx.meses_historia == 18

    def test_el_slug_de_salida_es_el_del_dominio(self):
        assert InteligenciaComercialService.SERVICE_SLUG == "inteligencia-comercial"


class TestVentanaDelContexto:
    def test_desde_retrocede_los_meses_pedidos(self):
        from src.core.data_loader import DataLoader
        from src.services.inteligencia_comercial.contracts import AnalysisContext

        ctx = AnalysisContext(data_loader=MagicMock(spec=DataLoader),
                              fecha_hasta="2026-07-30")
        assert ctx.desde(12).startswith("2025-07")
        assert ctx.desde(24).startswith("2024-07")
        assert ctx.desde(1).startswith("2026-06")

    def test_desde_cruza_bien_el_cambio_de_anio(self):
        from src.core.data_loader import DataLoader
        from src.services.inteligencia_comercial.contracts import AnalysisContext

        ctx = AnalysisContext(data_loader=MagicMock(spec=DataLoader),
                              fecha_hasta="2026-02-15")
        assert ctx.desde(3).startswith("2025-11")


class TestConstruccionDelLibro:
    def _libro(self, tmp_path, results=None, specs=None, reconciliaciones=None):
        results = results or {"a": _resultado("Analisis A")}
        specs = specs if specs is not None else [
            SheetSpec(sheet="Datos", analysis="a", table="principal",
                      title="Titulo", subtitle="Subtitulo")
        ]
        destino = tmp_path / "libro.xlsx"
        build_workbook(results, specs, destino, "Periodo de prueba", "01/01/2026",
                       reconciliaciones=reconciliaciones)
        return load_workbook(destino)

    def test_siempre_hay_portada_y_metodologia(self, tmp_path):
        wb = self._libro(tmp_path)
        assert wb.sheetnames[0] == "Portada"
        assert "Metodologia" in wb.sheetnames

    def test_una_tabla_vacia_no_genera_hoja(self, tmp_path):
        # una hoja vacia se lee como "no hay nada que ver" y no como "fallo el analisis"
        wb = self._libro(tmp_path, results={"a": _resultado(con_tabla=False)})
        assert wb.sheetnames == ["Portada", "Metodologia"]

    def test_una_hoja_que_apunta_a_un_analisis_ausente_se_omite(self, tmp_path):
        wb = self._libro(
            tmp_path,
            specs=[SheetSpec(sheet="Fantasma", analysis="no_existe", table="x",
                             title="T")],
        )
        assert "Fantasma" not in wb.sheetnames

    def test_el_nombre_de_hoja_se_recorta_a_31_caracteres(self, tmp_path):
        largo = "Un nombre de hoja larguisimo que Excel no soporta"
        wb = self._libro(
            tmp_path,
            specs=[SheetSpec(sheet=largo, analysis="a", table="principal", title="T")],
        )
        assert largo[:31] in wb.sheetnames
        assert all(len(n) <= 31 for n in wb.sheetnames)

    def test_las_alertas_llegan_a_la_portada(self, tmp_path):
        wb = self._libro(tmp_path)
        texto = " ".join(
            str(c.value) for row in wb["Portada"].iter_rows() for c in row if c.value
        )
        assert "CRITICA" in texto
        assert "Algo pasa" in texto

    def test_las_notas_llegan_a_metodologia(self, tmp_path):
        wb = self._libro(tmp_path)
        texto = " ".join(
            str(c.value) for row in wb["Metodologia"].iter_rows() for c in row if c.value
        )
        assert "Nota de metodologia de Analisis A" in texto

    def test_el_indice_apunta_a_hojas_que_existen(self, tmp_path):
        wb = self._libro(tmp_path)
        destinos = [
            c.hyperlink.location.split("!")[0].strip("'")
            for row in wb["Portada"].iter_rows() for c in row if c.hyperlink
        ]
        assert destinos
        for destino in destinos:
            assert destino in wb.sheetnames

    def test_un_decorador_que_falla_no_tumba_el_libro(self, tmp_path):
        # la tabla ya esta escrita y es el dato que importa; un grafico roto
        # no puede costar el informe entero
        def explota(ws, block, result):
            raise RuntimeError("grafico roto")

        wb = self._libro(
            tmp_path,
            specs=[SheetSpec(sheet="Datos", analysis="a", table="principal",
                             title="T", decorate=explota)],
        )
        assert "Datos" in wb.sheetnames
        assert wb["Datos"].max_row > 1

    def test_las_reconciliaciones_se_publican(self, tmp_path):
        wb = self._libro(tmp_path, reconciliaciones=["La brecha es de 13.122 htl"])
        texto = " ".join(
            str(c.value) for row in wb["Metodologia"].iter_rows() for c in row if c.value
        )
        assert "13.122 htl" in texto
        # write_section rotula en mayusculas
        assert "CONCILIACION ENTRE ANALISIS" in texto

    def test_un_analisis_fallido_se_avisa_en_la_portada(self, tmp_path):
        wb = self._libro(tmp_path, results={"a": _resultado("Roto", failed=True)})
        texto = " ".join(
            str(c.value) for row in wb["Portada"].iter_rows() for c in row if c.value
        )
        assert "no pudieron ejecutarse" in texto

    def test_los_formatos_explicitos_ganan(self, tmp_path):
        wb = self._libro(
            tmp_path,
            specs=[SheetSpec(sheet="Datos", analysis="a", table="principal",
                             title="T", formats={"pct_neto": "0.00%"})],
        )
        ws = wb["Datos"]
        fila = next(r for r in range(1, 12) if ws.cell(row=r, column=1).value == "sucursal")
        assert ws.cell(row=fila + 1, column=3).number_format == "0.00%"


class TestConciliacion:
    def test_declara_la_brecha_entre_el_puente_y_la_serie(self):
        clientes = AnalysisResult(
            name="clientes",
            tables={"puente": pd.DataFrame({
                "Movimiento": ["Nuevos", "TOTAL GENERAL"],
                "Htl actual": [100.0, 382420.0],
            })},
        )
        demanda = AnalysisResult(
            name="demanda",
            tables={"serie_mensual": pd.DataFrame({
                "mes": pd.date_range("2025-08-01", periods=12, freq="MS"),
                "TOTAL GENERAL - Htl": [32961.9] * 12,
            })},
        )
        notas = InteligenciaComercialService._conciliar(
            {"clientes": clientes, "demanda": demanda}
        )
        assert len(notas) == 1
        assert "382.420" in notas[0]
        assert "ventana movil" in notas[0]

    def test_sin_los_dos_analisis_no_inventa_nada(self):
        assert InteligenciaComercialService._conciliar({}) == []

    def test_con_tablas_vacias_no_rompe(self):
        vacio = AnalysisResult(name="x", tables={})
        assert InteligenciaComercialService._conciliar(
            {"clientes": vacio, "demanda": vacio}
        ) == []
