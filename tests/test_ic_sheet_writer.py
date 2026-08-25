"""Tests del escritor de hojas: formato inferido, fila de total y bloques."""
import pandas as pd
import pytest
from openpyxl import Workbook

from src.services.inteligencia_comercial import sheet_writer as sw
from src.services.inteligencia_comercial.excel_style import (
    FMT_DATE,
    FMT_DEC2,
    FMT_INT,
    FMT_KEY,
    FMT_MONEY,
    FMT_PCT1,
)


class TestInferenciaDeFormato:
    def test_las_columnas_de_id_no_llevan_separador_de_miles(self):
        # con #,##0.00 un id 1 se lee "1,00" y los VLOOKUP aguas abajo dejan de matchear
        assert sw.infer_format("id_cliente") == FMT_KEY
        assert sw.infer_format("id_ruta") == FMT_KEY
        assert sw.infer_format("codigo") == FMT_KEY

    def test_los_porcentajes_se_detectan(self):
        assert sw.infer_format("pct_neto") == FMT_PCT1
        assert sw.infer_format("tasa_descuento") == FMT_PCT1
        assert sw.infer_format("% clientes") == FMT_PCT1
        assert sw.infer_format("penetracion") == FMT_PCT1

    def test_los_montos_se_detectan(self):
        assert sw.infer_format("monetario_neto") == FMT_MONEY
        assert sw.infer_format("bruto") == FMT_MONEY
        assert sw.infer_format("ticket_medio") == FMT_MONEY
        assert sw.infer_format("exceso_vs_mediana") == FMT_MONEY

    def test_los_conteos_se_detectan(self):
        assert sw.infer_format("clientes") == FMT_INT
        assert sw.infer_format("bultos") == FMT_INT
        assert sw.infer_format("entregas") == FMT_INT

    def test_las_metricas_continuas_llevan_decimales(self):
        assert sw.infer_format("lift") == FMT_DEC2
        assert sw.infer_format("gap_p90") == FMT_DEC2
        assert sw.infer_format("z_robusto") == FMT_DEC2

    def test_las_fechas_se_detectan(self):
        assert sw.infer_format("fecha") == FMT_DATE
        assert sw.infer_format("periodo") == FMT_DATE

    def test_una_columna_desconocida_cae_en_entero(self):
        assert sw.infer_format("cualquier_cosa") == FMT_INT

    def test_el_porcentaje_gana_sobre_monto_cuando_ambos_aplican(self):
        # "margen_pct" contiene tanto un patron de % como ninguno de monto: debe ser %
        assert sw.infer_format("margen_pct") == FMT_PCT1


class TestFilaDeTotal:
    def _df(self):
        return pd.DataFrame(
            {
                "sucursal": ["CASA CENTRAL", "ORAN", "METAN"],
                "bultos": [100.0, 50.0, 25.0],
                "bruto": [1000.0, 400.0, 100.0],
                "descuento": [150.0, 20.0, 5.0],
            }
        )

    def test_agrega_una_fila_rotulada(self):
        out = sw.add_total_row(self._df())
        assert len(out) == 4
        assert out.iloc[-1]["sucursal"] == sw.TOTAL_LABEL

    def test_suma_las_medidas(self):
        out = sw.add_total_row(self._df())
        assert out.iloc[-1]["bultos"] == pytest.approx(175.0)
        assert out.iloc[-1]["bruto"] == pytest.approx(1500.0)

    def test_las_tasas_se_recalculan_no_se_promedian(self):
        # el promedio de tasas por sucursal (15%, 5%, 5% -> 8.33%) NO es la tasa
        # de la empresa (175/1500 = 11.67%); promediarlas es un error clasico
        df = self._df()
        df["tasa"] = df["descuento"] / df["bruto"]
        out = sw.add_total_row(
            df, recompute={"tasa": lambda d: d["descuento"].sum() / d["bruto"].sum()}
        )
        assert out.iloc[-1]["tasa"] == pytest.approx(175.0 / 1500.0)
        assert out.iloc[-1]["tasa"] != pytest.approx(df["tasa"].mean())

    def test_un_dataframe_vacio_no_rompe(self):
        assert sw.add_total_row(pd.DataFrame()).empty

    def test_las_columnas_de_texto_quedan_vacias(self):
        df = self._df()
        df["nota"] = ["a", "b", "c"]
        out = sw.add_total_row(df)
        assert out.iloc[-1]["nota"] is None


class TestEscrituraDeBloque:
    def _ws(self):
        return Workbook().active

    def _df(self):
        return pd.DataFrame(
            {
                "sucursal": ["CASA CENTRAL", "ORAN", sw.TOTAL_LABEL],
                "id_ruta": [81, 12, None],
                "bultos": [100.0, 50.0, 150.0],
                "pct_neto": [0.6, 0.4, 1.0],
            }
        )

    def test_devuelve_las_coordenadas_del_bloque(self):
        ws = self._ws()
        block = sw.write_dataframe(ws, self._df(), row=5)
        assert block["header_row"] == 5
        assert block["first_row"] == 6
        assert block["last_row"] == 8
        assert block["written"] == 3
        assert block["dropped"] == 0

    def test_escribe_el_encabezado_y_los_datos(self):
        ws = self._ws()
        sw.write_dataframe(ws, self._df(), row=5)
        assert ws.cell(row=5, column=1).value == "sucursal"
        assert ws.cell(row=6, column=1).value == "CASA CENTRAL"
        assert ws.cell(row=6, column=3).value == pytest.approx(100.0)

    def test_aplica_el_formato_numerico_inferido(self):
        ws = self._ws()
        sw.write_dataframe(ws, self._df(), row=1)
        assert ws.cell(row=2, column=2).number_format == FMT_KEY
        assert ws.cell(row=2, column=4).number_format == FMT_PCT1

    def test_los_overrides_ganan_sobre_la_inferencia(self):
        ws = self._ws()
        sw.write_dataframe(ws, self._df(), row=1, formats={"bultos": FMT_MONEY})
        assert ws.cell(row=2, column=3).number_format == FMT_MONEY

    def test_la_fila_de_total_queda_resaltada(self):
        ws = self._ws()
        sw.write_dataframe(ws, self._df(), row=1)
        assert ws.cell(row=4, column=1).font.bold is True

    def test_informa_las_filas_truncadas(self):
        # una tabla truncada en silencio se lee como completa: el corte debe ser visible
        ws = self._ws()
        df = pd.DataFrame({"a": [str(i) for i in range(100)], "bultos": range(100)})
        block = sw.write_dataframe(ws, df, row=1, max_rows=10)
        assert block["written"] == 10
        assert block["dropped"] == 90

    def test_el_truncado_conserva_la_fila_de_total(self):
        # la fila de total es la ultima: un head() se la come justo en las tablas
        # largas, que son las que mas necesitan un total
        ws = self._ws()
        df = pd.DataFrame(
            {"a": [str(i) for i in range(99)] + [sw.TOTAL_LABEL],
             "bultos": list(range(99)) + [4851]}
        )
        block = sw.write_dataframe(ws, df, row=1, max_rows=10)
        assert block["written"] == 11          # 10 filas + el total
        assert block["dropped"] == 89
        ultima = ws.cell(row=block["last_row"], column=1).value
        assert ultima == sw.TOTAL_LABEL
        # y sigue siendo el total del universo completo, no del recorte
        assert ws.cell(row=block["last_row"], column=2).value == 4851

    def test_sin_truncado_no_duplica_el_total(self):
        ws = self._ws()
        block = sw.write_dataframe(ws, self._df(), row=1, max_rows=99)
        assert block["written"] == 3
        assert block["dropped"] == 0

    def test_un_dataframe_vacio_escribe_un_aviso_y_no_rompe(self):
        ws = self._ws()
        block = sw.write_dataframe(ws, pd.DataFrame(), row=3)
        assert block["written"] == 0
        assert "Sin datos" in str(ws.cell(row=3, column=1).value)

    def test_los_nulos_se_escriben_como_celda_vacia(self):
        ws = self._ws()
        sw.write_dataframe(ws, self._df(), row=1)
        assert ws.cell(row=4, column=2).value is None

    def test_helpers_de_rango(self):
        ws = self._ws()
        block = sw.write_dataframe(ws, self._df(), row=5)
        assert sw.column_letter(block, "bultos") == "C"
        assert sw.data_range(block, "bultos") == "C6:C8"
        assert sw.data_range(block, "bultos", include_header=True) == "C5:C8"

    def test_los_escalares_numpy_se_desempaquetan(self):
        # openpyxl escribe np.float64 como texto si no se desempaqueta
        import numpy as np

        ws = self._ws()
        df = pd.DataFrame({"a": ["x"], "bultos": [np.float64(3.5)]})
        sw.write_dataframe(ws, df, row=1)
        assert isinstance(ws.cell(row=2, column=2).value, float)
