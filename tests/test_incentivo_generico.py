"""Tests for the generico-based cobertura incentive (e.g. PERNOD RICARD).

Reads pre-computed coverage from gold.cob_preventista_generico, one column per
the target generico, sections Por Vendedor + Por Supervisor. Objetivo is optional
— when absent, the Obj/% columns are hidden.
"""
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.services.incentivo_generico import (
    IncentivoGenericoConfig,
    IncentivoGenericoService,
)


# ── Loader query ─────────────────────────────────────────────────────────────

def test_loader_computes_daily_coverage_medallion_formula():
    """Coverage is computed from fact_ventas at daily grain, replicating medallion:
    the vendor is the client's assigned preventista (id_personal_fv1), a client
    counts when SUM(cantidades_total) > 0, and clientes = COUNT(DISTINCT cliente).
    """
    loader = DataLoader()
    cap: dict = {}

    def fake(q, params=None):
        cap["q"] = q
        cap["p"] = params
        return pd.DataFrame(columns=["vendedor", "clientes", "volumen"])

    loader.execute_query = MagicMock(side_effect=fake)
    loader.get_cobertura_generico_por_vendedor(
        generico="PERNOD RICARD", fecha_desde="2026-07-03", fecha_hasta="2026-07-03",
        id_sucursal=1, id_fuerza_ventas=1,
    )
    q = cap["q"].lower()
    assert "gold.fact_ventas" in q
    assert "id_personal_fv1" in q                 # vendor = client's assigned preventista
    assert "count(distinct" in q                   # distinct clients
    assert "having sum(fv.cantidades_total) > 0" in q
    assert "fecha_comprobante::date between" in q  # daily range, not month
    assert cap["p"]["generico"] == "PERNOD RICARD"
    assert cap["p"]["fecha_desde"] == "2026-07-03"
    assert cap["p"]["fecha_hasta"] == "2026-07-03"
    assert cap["p"]["id_sucursal"] == 1


def test_loader_fuerza4_uses_fv4_columns():
    """id_fuerza_ventas=4 → uses id_personal_fv4 instead of fv1."""
    loader = DataLoader()
    cap: dict = {}
    loader.execute_query = MagicMock(side_effect=lambda q, params=None: cap.update(q=q) or pd.DataFrame())
    loader.get_cobertura_generico_por_vendedor(
        generico="PERNOD RICARD", fecha_desde="2026-07-03", fecha_hasta="2026-07-03",
        id_sucursal=1, id_fuerza_ventas=4,
    )
    assert "id_personal_fv4" in cap["q"].lower()


# ── Service ──────────────────────────────────────────────────────────────────

def _loader_with_rows():
    """ROBLES ORLANDO + SEBASTIAN PIZARRO map to FGUANTAY; DIRECTA must be hidden."""
    loader = MagicMock(spec=DataLoader)
    loader.get_cobertura_generico_por_vendedor.return_value = pd.DataFrame({
        "vendedor": ["ROBLES ORLANDO", "SEBASTIAN PIZARRO", "DIRECTA"],
        "clientes": [25, 24, 1],
    })
    return loader


def test_single_day_uses_same_desde_hasta(tmp_path):
    """fecha without fecha_hasta → the loader gets a one-day range (desde==hasta)."""
    loader = _loader_with_rows()
    service = IncentivoGenericoService(data_loader=loader)
    config = IncentivoGenericoConfig(generico="PERNOD RICARD", fecha="2026-07-03")

    with patch("src.services.incentivo_generico.service.service_output_dir", return_value=tmp_path):
        service.generar_reporte(config)

    _, kwargs = loader.get_cobertura_generico_por_vendedor.call_args
    assert kwargs["fecha_desde"] == "2026-07-03"
    assert kwargs["fecha_hasta"] == "2026-07-03"
    assert kwargs["generico"] == "PERNOD RICARD"


def test_sin_objetivo_oculta_columnas(tmp_path):
    """objetivo=None → no Obj./% columns, just the clients count per actor."""
    loader = _loader_with_rows()
    service = IncentivoGenericoService(data_loader=loader)
    config = IncentivoGenericoConfig(generico="PERNOD RICARD", fecha="2026-07-03", objetivo=None)

    with patch("src.services.incentivo_generico.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    ws = load_workbook(result.ruta_archivo).active
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "ROBLES ORLANDO" in texts
    assert any("PERNOD RICARD" in t for t in texts)
    # Supervisor from the curated map (code), not dim_vendedor's full name
    assert "FGUANTAY" in texts
    assert "FACUNDO GONGORA" not in texts
    # DIRECTA is hidden
    assert "DIRECTA" not in texts
    # Objetivo columns hidden
    assert "Obj." not in texts


def test_con_objetivo_muestra_columnas(tmp_path):
    """objetivo set → Obj. and % columns appear."""
    loader = _loader_with_rows()
    service = IncentivoGenericoService(data_loader=loader)
    config = IncentivoGenericoConfig(generico="PERNOD RICARD", fecha="2026-07-03", objetivo=10)

    with patch("src.services.incentivo_generico.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    ws = load_workbook(result.ruta_archivo).active
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "Obj." in texts


def test_total_general_row(tmp_path):
    """A 'TOTAL GENERAL' row shows the overall clients (ROBLES 25 + PIZARRO 24 = 49)."""
    loader = _loader_with_rows()
    service = IncentivoGenericoService(data_loader=loader)
    config = IncentivoGenericoConfig(generico="PERNOD RICARD", fecha="2026-07-03")

    with patch("src.services.incentivo_generico.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    ws = load_workbook(result.ruta_archivo).active
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "TOTAL GENERAL" in texts
    nums = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))]
    assert 49 in nums  # 25 + 24 (DIRECTA excluded)


def test_supervisor_totals_aggregate_vendors(tmp_path):
    """Por Supervisor section sums each supervisor's vendors' clients."""
    loader = _loader_with_rows()
    service = IncentivoGenericoService(data_loader=loader)
    config = IncentivoGenericoConfig(generico="PERNOD RICARD", fecha="2026-07-03")

    with patch("src.services.incentivo_generico.service.service_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)

    ws = load_workbook(result.ruta_archivo).active
    # FACUNDO GONGORA total = 25 + 24 = 49 must appear as a numeric cell
    nums = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))]
    assert 49 in nums
