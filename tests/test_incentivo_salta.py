"""Tests del incentivo preventa SALTA.

Tres cosas que este informe hace distinto al resto y que estos tests fijan:

- La cobertura cuenta neto POSITIVO (`> 0`). Antes usaba umbral 0.5 bultos.
- Los cupos se LEEN del xlsx de objetivos y nunca se recalculan.
- Un bloque cuyo mes todavia no empezo queda en BLANCO, no en cero.
"""
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.core.data_loader import DataLoader
from src.services.incentivo_salta.objetivos import leer_objetivos
from src.services.incentivo_salta.processor import (
    contar_cobertura,
    cobertura_total,
    mes_ya_empezo,
    ventana_del_mes,
)
from src.services.incentivo_salta.service import (
    IncentivoSaltaConfig,
    IncentivoSaltaService,
)


# ── Ventana y vigencia ───────────────────────────────────────────────────────

class TestVentanaDelMes:
    def test_mes_cerrado_devuelve_el_mes_entero(self):
        assert ventana_del_mes("2026-07", "2026-08-04") == ("2026-07-01", "2026-07-31")

    def test_mes_en_curso_se_recorta_en_hasta(self):
        assert ventana_del_mes("2026-08", "2026-08-04") == ("2026-08-01", "2026-08-04")

    def test_sin_hasta_devuelve_el_mes_entero(self):
        assert ventana_del_mes("2026-02") == ("2026-02-01", "2026-02-28")

    def test_febrero_bisiesto(self):
        assert ventana_del_mes("2028-02") == ("2028-02-01", "2028-02-29")


class TestMesYaEmpezo:
    @pytest.mark.parametrize("mes,hoy,esperado", [
        ("2026-08", "2026-08-04", True),    # el mes en curso cuenta
        ("2026-07", "2026-08-04", True),
        ("2026-09", "2026-08-04", False),   # todavia no llego
        ("2026-09", "2026-09-01", True),
    ])
    def test_vigencia(self, mes, hoy, esperado):
        assert mes_ya_empezo(mes, hoy) is esperado


# ── Conteo de cobertura ──────────────────────────────────────────────────────

def _ventas():
    """Los cuatro tienen neto positivo, asi que los cuatro cuentan.

    El cliente 3 lleva 0.25 bultos: con el umbral viejo de media caja quedaba
    afuera; con `> 0` entra.
    """
    return pd.DataFrame({
        "id_cliente": [1, 2, 3, 4],
        "id_sucursal": [1, 1, 1, 1],
        "preventista": ["LORENA", "LORENA", "LORENA", "NAHUEL"],
        "sabor": ["NEGRA"] * 4,
        "calibre": ["1000"] * 4,
        "bultos": [3.0, 0.5, 0.25, 2.0],
    })


class TestContarCobertura:
    def test_cuenta_todo_neto_positivo(self):
        """Criterio 2026-08-19: cualquier compra neta positiva cuenta.

        Con el umbral viejo de 0.5 el cliente 3 (0.25 bultos) quedaba afuera y
        LORENA daba 2.
        """
        assert contar_cobertura(_ventas(), "NEGRA", "1000") == {"LORENA": 3, "NAHUEL": 1}

    def test_una_fraccion_chica_cuenta(self):
        d = _ventas()[_ventas().id_cliente == 3]
        assert contar_cobertura(d, "NEGRA", "1000") == {"LORENA": 1}

    def test_neto_cero_exacto_NO_cuenta(self):
        """El filtro es estricto: `> 0`, no `>= 0`.

        Si fuera `>=`, el que compro y devolvio todo entraria como cubierto.
        """
        d = pd.DataFrame({
            "id_cliente": [9, 9], "id_sucursal": [1, 1],
            "preventista": ["LORENA", "LORENA"], "sabor": ["NEGRA"] * 2,
            "calibre": ["1000"] * 2, "bultos": [4.0, -4.0],
        })
        assert contar_cobertura(d, "NEGRA", "1000") == {}

    def test_totaliza_por_cliente_antes_de_filtrar(self):
        """Dos compras de 0.3 suman 0.6 y el cliente cuenta; linea a linea no."""
        d = pd.DataFrame({
            "id_cliente": [9, 9], "id_sucursal": [1, 1],
            "preventista": ["LORENA", "LORENA"], "sabor": ["NEGRA"] * 2,
            "calibre": ["1000"] * 2, "bultos": [0.3, 0.3],
        })
        assert contar_cobertura(d, "NEGRA", "1000") == {"LORENA": 1}

    def test_devolucion_que_cancela_la_compra_deja_al_cliente_afuera(self):
        d = pd.DataFrame({
            "id_cliente": [9, 9], "id_sucursal": [1, 1],
            "preventista": ["LORENA", "LORENA"], "sabor": ["NEGRA"] * 2,
            "calibre": ["1000"] * 2, "bultos": [5.0, -5.0],
        })
        assert contar_cobertura(d, "NEGRA", "1000") == {}

    def test_el_corte_de_sabor_y_calibre_separa(self):
        assert contar_cobertura(_ventas(), "BLANCA (rubia)", "1000") == {}

    def test_df_vacio(self):
        assert contar_cobertura(pd.DataFrame(), "NEGRA", "1000") == {}


class TestCoberturaTotal:
    def test_cuenta_desde_el_grano_no_sumando_preventistas(self):
        # 4 y no 3: con el criterio `> 0` el cliente de 0.25 bultos tambien entra.
        assert cobertura_total(_ventas(), "NEGRA", "1000") == 4

    def test_mismo_cliente_en_dos_preventistas_cuenta_una_vez(self):
        d = pd.DataFrame({
            "id_cliente": [7, 7], "id_sucursal": [1, 1],
            "preventista": ["LORENA", "NAHUEL"], "sabor": ["NEGRA"] * 2,
            "calibre": ["1000"] * 2, "bultos": [3.0, 3.0],
        })
        assert cobertura_total(d, "NEGRA", "1000") == 1
        assert sum(contar_cobertura(d, "NEGRA", "1000").values()) == 2


# ── Lectura del archivo de objetivos ─────────────────────────────────────────

def _xlsx_objetivos(tmp_path, cupos=((115, 170), (51, 84))):
    from datetime import datetime
    wb = Workbook(); ws = wb.active
    ws["A4"] = "INCENTIVO AGOSTO"
    ws["B5"] = "SALTA NEGRA"; ws["E5"] = "SALTA RUBIA"
    ws["B6"] = "1000 cc";     ws["E6"] = "1200 cc"
    ws["B7"] = "Cupo"; ws["C7"] = datetime(2026, 8, 1); ws["D7"] = "%"
    ws["E7"] = "Cupo"; ws["F7"] = datetime(2026, 8, 1); ws["G7"] = "%"
    for i, (neg, rub) in enumerate(cupos, 8):
        ws.cell(i, 1, f"PREVENTISTA {i - 7}")
        ws.cell(i, 2, neg); ws.cell(i, 5, rub)
    ws.cell(8 + len(cupos), 1, "TOTAL GENERAL")
    ruta = tmp_path / "objetivos.xlsx"
    wb.save(ruta)
    return ruta


class TestLeerObjetivos:
    def test_descubre_los_bloques_y_sus_cupos(self, tmp_path):
        bloques = leer_objetivos(_xlsx_objetivos(tmp_path))
        assert len(bloques) == 2
        neg, rub = bloques
        assert (neg.grupo, neg.sabor, neg.calibre, neg.mes) == (
            "INCENTIVO AGOSTO", "NEGRA", "1000", "2026-08")
        assert (rub.sabor, rub.calibre) == ("BLANCA (rubia)", "1200")
        assert neg.cupos == {"PREVENTISTA 1": 115.0, "PREVENTISTA 2": 51.0}
        assert neg.cupo_total == 166.0

    def test_corta_en_la_fila_de_total(self, tmp_path):
        bloques = leer_objetivos(_xlsx_objetivos(tmp_path))
        assert "TOTAL GENERAL" not in bloques[0].cupos

    def test_falla_si_un_bloque_no_tiene_fecha(self, tmp_path):
        wb = Workbook(); ws = wb.active
        ws["B7"] = "Cupo"; ws["C7"] = "no es fecha"
        ws["A8"] = "X"; ws["B8"] = 10
        ruta = tmp_path / "roto.xlsx"; wb.save(ruta)
        with pytest.raises(ValueError, match="fecha de mes"):
            leer_objetivos(ruta)

    def test_falla_si_no_hay_fila_de_medidas(self, tmp_path):
        wb = Workbook(); wb.active["A1"] = "nada"
        ruta = tmp_path / "vacio.xlsx"; wb.save(ruta)
        with pytest.raises(ValueError, match="fila de medidas"):
            leer_objetivos(ruta)


# ── Servicio ─────────────────────────────────────────────────────────────────

def _xlsx_dos_meses(tmp_path):
    """Un bloque de agosto y otro de septiembre, para probar la vigencia."""
    from datetime import datetime
    wb = Workbook(); ws = wb.active
    ws["A4"] = "INCENTIVO AGOSTO"; ws["E4"] = "INCENTIVO SEPTIEMBRE"
    ws["B5"] = "SALTA NEGRA";      ws["E5"] = "SALTA NEGRA"
    ws["B6"] = "1000 cc";          ws["E6"] = "1000 cc"
    ws["B7"] = "Cupo"; ws["C7"] = datetime(2026, 8, 1); ws["D7"] = "%"
    ws["E7"] = "Cupo"; ws["F7"] = datetime(2026, 9, 1); ws["G7"] = "%"
    for i, nombre in enumerate(("LORENA", "NAHUEL", "DIRECTA"), 8):
        ws.cell(i, 1, nombre); ws.cell(i, 2, 10); ws.cell(i, 5, 20)
    ws.cell(11, 1, "TOTAL GENERAL")
    ruta = tmp_path / "objetivos.xlsx"; wb.save(ruta)
    return ruta


def _generar(tmp_path, *, fecha_hasta="2026-08-04", excluir=("DIRECTA",)):
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_cliente_sabor_mes.return_value = _ventas()
    service = IncentivoSaltaService(data_loader=loader)
    config = IncentivoSaltaConfig(
        fecha_hasta=fecha_hasta,
        objetivos_path=str(_xlsx_dos_meses(tmp_path)),
        excluir_vendedores=list(excluir),
    )
    with patch("src.services.incentivo_salta.service.service_output_dir",
               return_value=tmp_path):
        return service.generar_reporte(config), loader


def _celdas(ruta):
    ws = load_workbook(ruta).active
    return ws, {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)}


def test_solo_consulta_los_meses_que_ya_empezaron(tmp_path):
    result, loader = _generar(tmp_path)
    ventanas = {(c.kwargs["fecha_desde"], c.kwargs["fecha_hasta"])
                for c in loader.get_ventas_cliente_sabor_mes.call_args_list}
    assert ventanas == {("2026-08-01", "2026-08-04")}   # septiembre no se pide
    assert result.bloques == 2 and len(result.bloques_activos) == 1


def test_el_bloque_futuro_queda_en_blanco_no_en_cero(tmp_path):
    result, _ = _generar(tmp_path)
    ws, filas = _celdas(result.ruta_archivo)
    r = filas["LORENA"]
    # 3 y no 2 desde el cambio de criterio a `> 0` (2026-08-19).
    assert ws.cell(r, 3).value == 3       # agosto: cobertura cargada
    assert ws.cell(r, 6).value is None    # septiembre: en blanco
    assert ws.cell(r, 7).value is None    # y sin formula de %


def test_los_cupos_salen_del_archivo_y_no_se_recalculan(tmp_path):
    result, _ = _generar(tmp_path)
    ws, filas = _celdas(result.ruta_archivo)
    assert ws.cell(filas["LORENA"], 2).value == 10
    assert ws.cell(filas["NAHUEL"], 5).value == 20


def test_excluye_a_los_que_no_participan(tmp_path):
    result, _ = _generar(tmp_path, excluir=("DIRECTA",))
    _, filas = _celdas(result.ruta_archivo)
    assert "DIRECTA" not in filas
    assert result.preventistas == 2


def test_el_porcentaje_es_formula_viva(tmp_path):
    result, _ = _generar(tmp_path)
    ws, filas = _celdas(result.ruta_archivo)
    assert ws.cell(filas["LORENA"], 4).value == '=IF(B7=0,"",C7/B7)'


def test_lleva_fila_de_total_y_de_avance(tmp_path):
    result, _ = _generar(tmp_path)
    ws, filas = _celdas(result.ruta_archivo)
    assert "TOTAL GENERAL" in filas and "% AVANCE" in filas
    assert ws.cell(filas["TOTAL GENERAL"], 2).value.startswith("=SUM(")


def test_semaforo_en_las_columnas_de_porcentaje(tmp_path):
    result, _ = _generar(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    rangos = {str(r.sqref): r.rules for r in ws.conditional_formatting}
    assert any(s.startswith("D") for s in rangos)
    regla = next(iter(rangos.values()))[0]
    assert regla.iconSet.iconSet == "3TrafficLights1"
    # Umbrales fijos del negocio, no percentiles.
    assert [c.val for c in regla.iconSet.cfvo] == [0.0, 0.5, 1.0]


class TestRenombreDePreventistas:
    """Un preventista renombrado en el maestro no puede salir en cero.

    dim_vendedor es SCD tipo 1: al renombrarlo se pierde el nombre anterior,
    pero el xlsx de objetivos lo mantiene otra persona y se queda con el viejo.
    Sin traducir, la fila aparece con el nombre viejo y TODO en cero — peor que
    si no apareciera, porque parece un preventista que no vendio nada.
    """

    def test_el_nombre_viejo_se_traduce_al_actual(self):
        from src.core.vendedores import nombre_actual
        assert nombre_actual("DARIO LUPATY") == "LUCIANO GUZMAN"

    def test_no_distingue_mayusculas_ni_espacios(self):
        from src.core.vendedores import nombre_actual
        assert nombre_actual("  dario lupaty  ") == "LUCIANO GUZMAN"

    def test_un_nombre_sin_renombre_pasa_igual(self):
        from src.core.vendedores import nombre_actual
        assert nombre_actual("GUANCA LUIS") == "GUANCA LUIS"

    def test_none_no_rompe(self):
        from src.core.vendedores import nombre_actual
        assert nombre_actual(None) == ""

    def test_los_cupos_del_xlsx_salen_con_el_nombre_actual(self):
        """El contrato de punta a punta contra el archivo real."""
        from pathlib import Path
        from src.services.incentivo_salta.objetivos import leer_objetivos

        bloques = leer_objetivos(Path("configs/objetivos_incentivo_salta.xlsx"))
        todos = {v for b in bloques for v in b.cupos}
        assert "LUCIANO GUZMAN" in todos
        assert "DARIO LUPATY" not in todos
