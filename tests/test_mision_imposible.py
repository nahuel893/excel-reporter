"""Tests para MisionImposibleService."""
import pandas as pd
import pytest
from unittest.mock import MagicMock, patch
from pathlib import Path

import config.settings as _settings
from src.core.data_loader import DataLoader
from src.services.mision_imposible.service import (
    MisionImposibleConfig,
    MisionImposibleService,
    _aplicar_zonas_categoria,
    _pivotear_categoria,
    _preparar_hoja,
    _COB_PREV_GEN_COLUMNS,
    _COB_SUC_GEN_COLUMNS,
)


class TestPrepararHoja:
    def test_ordena_y_renombra(self):
        df = pd.DataFrame({
            "periodo": ["2026-03-01", "2026-03-01"],
            "sucursal": ["SUC B", "SUC A"],
            "vendedor": ["Juan", "Ana"],
            "generico": ["CERVEZAS", "AGUAS"],
            "clientes_compradores": [100, 200],
            "volumen_total": [500.0, 300.0],
        })
        result = _preparar_hoja(
            df, ["sucursal", "vendedor", "generico"], _COB_PREV_GEN_COLUMNS
        )
        assert list(result.columns) == [
            "Periodo", "Sucursal", "Vendedor", "Generico",
            "Clientes Compradores", "Volumen Total",
        ]
        assert result.iloc[0]["Sucursal"] == "SUC A"

    def test_none_returns_none(self):
        assert _preparar_hoja(None, [], {}) is None

    def test_empty_returns_none(self):
        df = pd.DataFrame()
        assert _preparar_hoja(df, [], {}) is None

    def test_excludes_extra_columns(self):
        df = pd.DataFrame({
            "periodo": ["2026-03-01"],
            "sucursal": ["SUC A"],
            "generico": ["CERVEZAS"],
            "clientes_compradores": [100],
            "volumen_total": [500.0],
            "id_vendedor": [99],
        })
        result = _preparar_hoja(
            df, ["sucursal", "generico"], _COB_SUC_GEN_COLUMNS
        )
        assert "id_vendedor" not in result.columns


class TestMisionImposibleConfig:
    def test_defaults(self):
        cfg = MisionImposibleConfig(fecha_desde="2026-03-01", fecha_hasta="2026-03-31")
        assert cfg.genericos is None
        assert cfg.nombre_archivo is None

    def test_with_genericos(self):
        cfg = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            genericos=["CERVEZAS"],
        )
        assert cfg.genericos == ["CERVEZAS"]


class TestMisionImposibleService:
    def _make_prev_generico_df(self):
        return pd.DataFrame({
            "periodo": ["2026-03-01"] * 4,
            "sucursal": ["SUC A", "SUC A", "SUC B", "SUC B"],
            "id_vendedor": [1, 1, 2, 2],
            "vendedor": ["Juan", "Juan", "Ana", "Ana"],
            "generico": ["CERVEZAS", "AGUAS", "CERVEZAS", "AGUAS"],
            "clientes_compradores": [100, 50, 80, 40],
            "volumen_total": [500.0, 200.0, 400.0, 150.0],
        })

    def _make_prev_marca_df(self):
        return pd.DataFrame({
            "periodo": ["2026-03-01"] * 2,
            "sucursal": ["SUC A", "SUC B"],
            "id_vendedor": [1, 2],
            "vendedor": ["Juan", "Ana"],
            "marca": ["QUILMES", "BRAHMA"],
            "clientes_compradores": [60, 30],
            "volumen_total": [300.0, 150.0],
        })

    def _make_suc_generico_df(self):
        return pd.DataFrame({
            "periodo": ["2026-03-01"] * 2,
            "sucursal": ["SUC A", "SUC B"],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "clientes_compradores": [180, 80],
            "volumen_total": [900.0, 400.0],
        })

    def _make_suc_marca_df(self):
        return pd.DataFrame({
            "periodo": ["2026-03-01"] * 2,
            "sucursal": ["SUC A", "SUC B"],
            "marca": ["QUILMES", "BRAHMA"],
            "clientes_compradores": [60, 30],
            "volumen_total": [300.0, 150.0],
        })

    def test_generates_4_sheets(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_cobertura_preventista_generico.return_value = self._make_prev_generico_df()
        mock_loader.get_cobertura_preventista_marca.return_value = self._make_prev_marca_df()
        mock_loader.get_cobertura_sucursal_generico.return_value = self._make_suc_generico_df()
        mock_loader.get_cobertura_sucursal_marca.return_value = self._make_suc_marca_df()

        service = MisionImposibleService(data_loader=mock_loader)
        config = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            nombre_archivo="test_mision",
        )

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        assert len(result.hojas) == 4
        import openpyxl
        wb = openpyxl.load_workbook(str(result.ruta_archivo))
        assert len(wb.sheetnames) == 4

    def test_filters_by_genericos(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_cobertura_preventista_generico.return_value = self._make_prev_generico_df()
        mock_loader.get_cobertura_preventista_marca.return_value = self._make_prev_marca_df()
        mock_loader.get_cobertura_sucursal_generico.return_value = self._make_suc_generico_df()
        mock_loader.get_cobertura_sucursal_marca.return_value = self._make_suc_marca_df()

        service = MisionImposibleService(data_loader=mock_loader)
        config = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            genericos=["CERVEZAS"],
            nombre_archivo="test_filtro",
        )

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        # Verify the prev generico sheet only has CERVEZAS rows
        import openpyxl
        wb = openpyxl.load_workbook(str(result.ruta_archivo))
        ws = wb["Cob Preventista Generico"]
        # Column 4 is Generico (after Periodo, Sucursal, Vendedor)
        generico_col = 4
        for row in range(2, ws.max_row + 1):
            assert ws.cell(row, generico_col).value == "CERVEZAS"

    def test_applies_zonas_virtuales(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_cobertura_preventista_generico.return_value = self._make_prev_generico_df()
        mock_loader.get_cobertura_preventista_marca.return_value = pd.DataFrame()
        mock_loader.get_cobertura_sucursal_generico.return_value = pd.DataFrame()
        mock_loader.get_cobertura_sucursal_marca.return_value = pd.DataFrame()

        service = MisionImposibleService(data_loader=mock_loader)
        config = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            nombre_archivo="test_zonas",
        )

        with (
            patch.object(_settings, "DATA_OUTPUT", tmp_path),
            patch("src.services.mision_imposible.service.aplicar_zonas_virtuales") as mock_zonas,
        ):
            mock_zonas.return_value = self._make_prev_generico_df()
            service.generar_reporte(config)

        mock_zonas.assert_called_once()

    def test_empty_data_produces_no_sheets(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_cobertura_preventista_generico.return_value = pd.DataFrame()
        mock_loader.get_cobertura_preventista_marca.return_value = pd.DataFrame()
        mock_loader.get_cobertura_sucursal_generico.return_value = pd.DataFrame()
        mock_loader.get_cobertura_sucursal_marca.return_value = pd.DataFrame()

        service = MisionImposibleService(data_loader=mock_loader)
        config = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            nombre_archivo="test_empty",
        )

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        assert result.hojas == []

    def test_result_fields(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_cobertura_preventista_generico.return_value = self._make_prev_generico_df()
        mock_loader.get_cobertura_preventista_marca.return_value = self._make_prev_marca_df()
        mock_loader.get_cobertura_sucursal_generico.return_value = self._make_suc_generico_df()
        mock_loader.get_cobertura_sucursal_marca.return_value = self._make_suc_marca_df()

        service = MisionImposibleService(data_loader=mock_loader)
        config = MisionImposibleConfig(
            fecha_desde="2026-03-01",
            fecha_hasta="2026-03-31",
            nombre_archivo="test_result",
        )

        with patch.object(_settings, "DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        assert result.sucursales == 2
        assert "CERVEZAS" in result.genericos_incluidos
        assert "AGUAS" in result.genericos_incluidos
        assert result.registros_procesados > 0


# ---------------------------------------------------------------------------
# Categorias: _aplicar_zonas_categoria y _pivotear_categoria
# ---------------------------------------------------------------------------


def _make_ventas_categoria_df():
    """DataFrame simulando lo que devuelve get_ventas_mision_imposible_categorias."""
    return pd.DataFrame({
        "sucursal": ["CASA CENTRAL", "CASA CENTRAL", "CASA CENTRAL", "SUC A"],
        "id_ruta": [81, 81, 5, 10],  # ruta 81 es VALLE SALTA
        "id_vendedor": [1, 1, 2, 3],
        "vendedor": ["Juan", "Juan", "Ana", "Pedro"],
        "id_cliente": [100, 100, 200, 300],
        "cliente": ["Cli A", "Cli A", "Cli B", "Cli C"],
        "id_articulo": [1001, 1002, 1001, 1002],
        "des_articulo": ["Quilmes Lata", "Quilmes Botella", "Quilmes Lata", "Quilmes Botella"],
        "marca": ["QUILMES", "QUILMES", "QUILMES", "QUILMES"],
        "cantidad": [50, 30, 100, 75],
    })


class TestAplicarZonasCategoria:
    def test_renames_zona_virtual(self):
        df = _make_ventas_categoria_df()
        result = _aplicar_zonas_categoria(df)
        assert "zona_virtual" in result.columns
        assert "sucursal" not in result.columns

    def test_splits_casa_central(self):
        df = _make_ventas_categoria_df()
        result = _aplicar_zonas_categoria(df)
        zonas = result["zona_virtual"].unique()
        assert "VALLE SALTA" in zonas
        assert "CASA CENTRAL" in zonas

    def test_ruta_81_becomes_valle_salta(self):
        df = _make_ventas_categoria_df()
        result = _aplicar_zonas_categoria(df)
        valle_rows = result[result["zona_virtual"] == "VALLE SALTA"]
        assert len(valle_rows) == 2  # 2 rows with id_ruta=81
        assert all(valle_rows["id_ruta"] == 81)

    def test_creates_articulo_desc(self):
        df = _make_ventas_categoria_df()
        result = _aplicar_zonas_categoria(df)
        assert "articulo_desc" in result.columns
        assert "1001 - Quilmes Lata" in result["articulo_desc"].values

    def test_does_not_modify_original(self):
        df = _make_ventas_categoria_df()
        original_cols = list(df.columns)
        _aplicar_zonas_categoria(df)
        assert list(df.columns) == original_cols


class TestPivotearCategoria:
    def test_pivot_structure(self):
        df = _aplicar_zonas_categoria(_make_ventas_categoria_df())
        result = _pivotear_categoria(df)

        assert "zona_virtual" in result.columns
        assert "id_ruta" in result.columns
        assert "id_vendedor" in result.columns
        assert "vendedor" in result.columns
        assert "id_cliente" in result.columns
        assert "cliente" in result.columns

    def test_has_total_column(self):
        df = _aplicar_zonas_categoria(_make_ventas_categoria_df())
        result = _pivotear_categoria(df)
        total_cols = [c for c in result.columns if "Total" in str(c)]
        assert len(total_cols) > 0

    def test_total_sums_correctly(self):
        df = _aplicar_zonas_categoria(_make_ventas_categoria_df())
        result = _pivotear_categoria(df)
        total_col = [c for c in result.columns if "Total" in str(c)][0]
        # VALLE SALTA client 100 has 50 + 30 = 80 (both articles)
        valle_row = result[
            (result["zona_virtual"] == "VALLE SALTA") & (result["id_cliente"] == 100)
        ]
        assert valle_row[total_col].values[0] == 80

    def test_fill_value_zero(self):
        df = _aplicar_zonas_categoria(_make_ventas_categoria_df())
        result = _pivotear_categoria(df)
        # SUC A client 300 only has Quilmes Botella, Quilmes Lata should be 0
        suc_a_row = result[result["zona_virtual"] == "SUC A"]
        lata_col = [c for c in result.columns if "Quilmes Lata" in str(c)]
        if lata_col:
            assert suc_a_row[lata_col[0]].values[0] == 0

    def test_multiindex_flattened(self):
        df = _aplicar_zonas_categoria(_make_ventas_categoria_df())
        result = _pivotear_categoria(df)
        # Columns should be flattened strings, not tuples
        for col in result.columns:
            assert isinstance(col, str)


class TestProcesarCategoriaIntegration:
    def test_empty_data_returns_none(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_mision_imposible_categorias.return_value = pd.DataFrame()

        service = MisionImposibleService(data_loader=mock_loader)
        result = service._procesar_categoria("2026-03-01", "2026-03-31", [1001])
        assert result is None

    def test_returns_dataframe_with_data(self, tmp_path):
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_mision_imposible_categorias.return_value = (
            _make_ventas_categoria_df()
        )

        service = MisionImposibleService(data_loader=mock_loader)
        result = service._procesar_categoria("2026-03-01", "2026-03-31", [1001, 1002])
        assert result is not None
        assert len(result) > 0
        assert "zona_virtual" in result.columns

    def test_error_propagates(self):
        """Errors in _procesar_categoria propagate (not silently swallowed)."""
        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_ventas_mision_imposible_categorias.side_effect = RuntimeError("DB down")

        service = MisionImposibleService(data_loader=mock_loader)
        with pytest.raises(RuntimeError, match="DB down"):
            service._procesar_categoria("2026-03-01", "2026-03-31", [1001])
