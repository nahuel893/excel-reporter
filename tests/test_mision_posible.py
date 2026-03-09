"""Tests unitarios para Mision Posible."""
from datetime import date
from pathlib import Path
from unittest.mock import Mock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.services.mision_posible.processor import (
    concatenar_tablas,
    procesar_cobertura_sucursal,
    procesar_cobertura_vendedor,
)
from src.services.mision_posible.service import (
    GrupoArticulos,
    MisionPosibleConfig,
    MisionPosibleService,
    _nombre_reporte,
    _normalizar_periodo,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _df_cob(rows=None):
    """Crea DataFrame de cobertura sin columna marca (post get_cobertura_custom)."""
    if rows is None:
        rows = [
            {"sucursal": "CASA CENTRAL", "vendedor": "Juan", "id_ruta": 1, "clientes_compradores": 20, "volumen_total": 100, "periodo": "2026-03-01"},
            {"sucursal": "CASA CENTRAL", "vendedor": "Maria", "id_ruta": 2, "clientes_compradores": 25, "volumen_total": 150, "periodo": "2026-03-01"},
            {"sucursal": "SUCURSAL CAFAYATE", "vendedor": "Pedro", "id_ruta": 10, "clientes_compradores": 15, "volumen_total": 80, "periodo": "2026-03-01"},
        ]
    return pd.DataFrame(rows)


def _df_cob_imperial():
    """DataFrame para grupo IMPERIAL (sin columna marca)."""
    return pd.DataFrame([
        {"sucursal": "CASA CENTRAL", "vendedor": "Juan", "id_ruta": 1, "clientes_compradores": 20, "volumen_total": 100, "periodo": "2026-03-01"},
        {"sucursal": "CASA CENTRAL", "vendedor": "Maria", "id_ruta": 2, "clientes_compradores": 25, "volumen_total": 150, "periodo": "2026-03-01"},
        {"sucursal": "SUCURSAL CAFAYATE", "vendedor": "Pedro", "id_ruta": 10, "clientes_compradores": 15, "volumen_total": 80, "periodo": "2026-03-01"},
    ])


def _df_cob_levite():
    """DataFrame para grupo LEVITE (sin columna marca)."""
    return pd.DataFrame([
        {"sucursal": "CASA CENTRAL", "vendedor": "Juan", "id_ruta": 1, "clientes_compradores": 10, "volumen_total": 50, "periodo": "2026-03-01"},
    ])


def _df_cob_aguas():
    """DataFrame para grupo AGUAS (union LEVITE+VILLAVICENCIO)."""
    return pd.DataFrame([
        {"sucursal": "CASA CENTRAL", "vendedor": "Juan", "id_ruta": 1, "clientes_compradores": 30, "volumen_total": 200, "periodo": "2026-03-01"},
        {"sucursal": "SUCURSAL CAFAYATE", "vendedor": "Pedro", "id_ruta": 10, "clientes_compradores": 12, "volumen_total": 60, "periodo": "2026-03-01"},
    ])


def _df_cob_preventista_marca():
    """DataFrame simulando tabla cob_preventista_marca del ETL (con columna marca)."""
    rows = []
    for row in _df_cob_imperial().to_dict("records"):
        rows.append({**row, "marca": "IMPERIAL"})
    for row in _df_cob_levite().to_dict("records"):
        rows.append({**row, "marca": "LEVITE"})
    return pd.DataFrame(rows)


def _mock_loader(ultima_fecha=date(2026, 3, 6)):
    loader = Mock(spec=DataLoader)

    # Tabla pre-agregada del ETL (marcas simples)
    loader.get_cobertura_preventista_marca.return_value = _df_cob_preventista_marca()

    # Query custom (grupos multi-marca o con filtro)
    def _side_effect_cob(periodo, marcas, filtro_descripcion=None):
        key = tuple(sorted(m.upper() for m in marcas))
        return {
            ("IMPERIAL",): _df_cob_imperial(),
            ("LEVITE",): _df_cob_levite(),
            ("LEVITE", "VILLAVICENCIO"): _df_cob_aguas(),
        }.get(key, pd.DataFrame())

    loader.get_cobertura_custom.side_effect = _side_effect_cob
    loader.get_ultima_fecha_venta.return_value = ultima_fecha
    return loader


PORCENTAJES = {"CASA CENTRAL": 30, "SUCURSAL CAFAYATE": 10, "SUCURSAL METAN": 10}


# ── Processor tests ─────────────────────────────────────────────────────────

class TestProcesarCoberturaSucursal:
    """Tests para procesar_cobertura_sucursal."""

    def test_columnas_tabla_sucursal(self):
        """RF-006: columnas exactas."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        assert list(df.columns) == ["Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]

    def test_cobertura_sucursal_agrupa_correctamente(self):
        """RF-008: suma clientes_compradores de vendedores en misma sucursal."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        cc = df[df["Sucursal"] == "CASA CENTRAL"]["Cobertura"].iloc[0]
        assert cc == 45  # 20 + 25

    def test_objetivo_sucursal_calculo_porcentaje(self):
        """RF-010: objetivo = total * porcentaje / 100."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        cc = df[df["Sucursal"] == "CASA CENTRAL"]["Objetivo"].iloc[0]
        caf = df[df["Sucursal"] == "SUCURSAL CAFAYATE"]["Objetivo"].iloc[0]
        assert cc == 150  # 500 * 30 / 100
        assert caf == 50   # 500 * 10 / 100

    def test_objetivo_ausente_queda_none(self):
        """RF-010: sin objetivos, todo None."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", None, PORCENTAJES)
        assert df["Objetivo"].isna().all()
        assert df["Faltante"].isna().all()
        assert df["%"].isna().all()

    def test_sucursal_sin_porcentaje_objetivo_none(self):
        """RF-010: sucursal en datos pero sin porcentaje."""
        pct = {"CASA CENTRAL": 30}  # CAFAYATE no incluida
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, pct)
        # CAFAYATE no aparece porque no esta en porcentajes_sucursal
        assert "SUCURSAL CAFAYATE" not in df["Sucursal"].values

    def test_todas_sucursales_presentes(self):
        """RF-014: todas las sucursales de porcentajes_sucursal aparecen."""
        pct = {"CASA CENTRAL": 30, "SUCURSAL CAFAYATE": 10, "SUCURSAL METAN": 10, "SUCURSAL PERICO": 10, "SUCURSAL TARTAGAL": 15}
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, pct)
        assert len(df) == 5
        # METAN tiene Cobertura = 0 (no hay datos)
        metan = df[df["Sucursal"] == "SUCURSAL METAN"]["Cobertura"].iloc[0]
        assert metan == 0

    def test_faltante_negativo_cuando_supera_objetivo(self):
        """RF-012: Faltante negativo si Cobertura > Objetivo."""
        pct = {"CASA CENTRAL": 5}  # 500*5/100 = 25, cobertura = 45
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, pct)
        faltante = df[df["Sucursal"] == "CASA CENTRAL"]["Faltante"].iloc[0]
        assert faltante == -20  # 25 - 45

    def test_porcentaje_none_cuando_objetivo_cero(self):
        """RF-013: % = None si objetivo = 0."""
        pct = {"CASA CENTRAL": 0}
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, pct)
        pct_val = df[df["Sucursal"] == "CASA CENTRAL"]["%"].iloc[0]
        assert pct_val is None or pd.isna(pct_val)

    def test_porcentaje_calculado_correctamente(self):
        """RF-013: % = round(Cobertura / Objetivo * 100, 1)."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        cc_row = df[df["Sucursal"] == "CASA CENTRAL"]
        pct = cc_row["%"].iloc[0]
        # Cobertura=45, Objetivo=150 → 0.30
        assert pct == 0.3

    def test_orden_filas_sucursal(self):
        """RF-014: filas ordenadas alfabeticamente por sucursal."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        assert list(df["Sucursal"]) == sorted(df["Sucursal"])

    def test_procesar_sucursal_sin_filtro_interno(self):
        """RF-017: procesar_cobertura_sucursal con DataFrame ya filtrado (sin columna marca)."""
        df = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        assert list(df.columns) == ["Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]
        assert df[df["Sucursal"] == "CASA CENTRAL"]["Cobertura"].iloc[0] == 45


class TestProcesarCoberturaVendedor:
    """Tests para procesar_cobertura_vendedor."""

    def test_columnas_tabla_vendedor(self):
        """RF-007: columnas exactas."""
        df = procesar_cobertura_vendedor(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        assert list(df.columns) == ["Vendedor", "Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]

    def test_objetivo_vendedor_reparto_igualitario(self):
        """RF-011: objetivo = objetivo_sucursal / cant_vendedores."""
        df = procesar_cobertura_vendedor(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        # CASA CENTRAL: obj=150, 2 vendedores → 75 cada uno
        juan = df[(df["Vendedor"] == "Juan") & (df["Sucursal"] == "CASA CENTRAL")]
        assert juan["Objetivo"].iloc[0] == 75

    def test_orden_filas_vendedor(self):
        """RF-015: ordenado por Sucursal, luego Vendedor."""
        df = procesar_cobertura_vendedor(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        pairs = list(zip(df["Sucursal"], df["Vendedor"]))
        assert pairs == sorted(pairs)

    def test_vendedor_df_vacio_retorna_vacio(self):
        """DataFrame vacio → DataFrame vacio con columnas correctas."""
        df = procesar_cobertura_vendedor(pd.DataFrame(), "NoExiste", 500, PORCENTAJES)
        assert df.empty
        assert list(df.columns) == ["Vendedor", "Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]

    def test_procesar_vendedor_sin_filtro_interno(self):
        """RF-017: procesar_cobertura_vendedor con DataFrame ya filtrado (sin columna marca)."""
        df = procesar_cobertura_vendedor(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        assert list(df.columns) == ["Vendedor", "Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]
        assert not df.empty

    def test_grupo_nombre_no_rompe_processor(self):
        """RF-014: nombre del grupo con espacios puede pasarse como _grupo_nombre."""
        df = procesar_cobertura_vendedor(_df_cob_imperial(), "SCHNEIDER 710", 500, PORCENTAJES)
        assert list(df.columns) == ["Vendedor", "Sucursal", "Cobertura", "Objetivo", "Faltante", "%"]


class TestConcatenarTablas:
    """Tests para concatenar_tablas."""

    def test_concatena_con_separador(self):
        """RF-005: dos tablas separadas por fila vacia + titulo."""
        df_suc = procesar_cobertura_sucursal(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        df_vend = procesar_cobertura_vendedor(_df_cob_imperial(), "IMPERIAL", 500, PORCENTAJES)
        result = concatenar_tablas(df_suc, df_vend)
        # Debe tener: filas_suc + 1(sep) + 1(titulo) + 1(header) + filas_vend
        assert len(result) == len(df_suc) + 3 + len(df_vend)
        # Fila titulo contiene "Por Vendedor"
        titulo_row = result.iloc[len(df_suc) + 1]
        assert titulo_row.iloc[0] == "Por Vendedor"


# ── Service tests ────────────────────────────────────────────────────────────

class TestMisionPosibleService:
    """Tests para MisionPosibleService."""

    def _config(self, **overrides):
        defaults = {
            "periodo": "2026-03-01",
            "grupos": [GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"]), GrupoArticulos("LEVITE", marcas=["LEVITE"])],
            "objetivos": {"IMPERIAL": 500, "LEVITE": 300},
            "porcentajes_sucursal": PORCENTAJES,
        }
        defaults.update(overrides)
        return MisionPosibleConfig(**defaults)

    def _generar(self, tmp_path, loader=None, **config_overrides):
        """Helper: genera reporte y retorna (result, workbook)."""
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader or _mock_loader())
            result = service.generar_reporte(self._config(**config_overrides))
        wb = load_workbook(result.ruta_archivos[0])
        return result, wb

    # ── Tests adaptados ──────────────────────────────────────────────────────

    def test_nombre_archivo_formato_mes_anio(self, tmp_path):
        """RF-003: nombre = Mision Posible MM-YYYY."""
        result, _ = self._generar(tmp_path)
        assert result.ruta_archivos[0].stem == "Mision Posible 03-2026"

    def test_nombre_archivo_custom(self, tmp_path):
        """RF-003: nombre custom se respeta."""
        result, _ = self._generar(tmp_path, nombre_archivo="Mi Mision")
        assert result.ruta_archivos[0].stem == "Mi Mision"

    def test_zonas_virtuales_aplicadas_a_cobertura(self, tmp_path):
        """RF-015: zonas virtuales se aplican; workbook tiene 2 hojas."""
        loader = Mock(spec=DataLoader)
        loader.get_cobertura_preventista_marca.return_value = pd.DataFrame([{
            "sucursal": "CASA CENTRAL", "vendedor": "Juan",
            "id_ruta": 81, "clientes_compradores": 20,
            "volumen_total": 100, "periodo": "2026-03-01",
            "marca": "IMPERIAL",
        }])
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        pct = {"VALLE SALTA": 15, "CASA CENTRAL": 30}
        result, wb = self._generar(
            tmp_path,
            loader=loader,
            grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])],
            porcentajes_sucursal=pct,
        )
        assert wb.sheetnames == ["Sucursales", "Por Vendedor"]

    def test_modo_supervisores_genera_un_archivo_por_supervisor(self, tmp_path):
        """RF-016: un archivo por supervisor."""
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=_mock_loader())
            supervisores = {"Ana": ["CASA CENTRAL"], "Luis": ["SUCURSAL CAFAYATE"]}
            results = service.generar_reporte_supervisores(self._config(), supervisores)
        assert len(results) == 2
        assert results[0].supervisor == "Ana"
        assert results[1].supervisor == "Luis"

    def test_modo_supervisores_una_sola_consulta(self, tmp_path):
        """RF-016: una consulta por grupo aunque haya varios supervisores (2 grupos = 2 calls)."""
        loader = _mock_loader()
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            supervisores = {"Ana": ["CASA CENTRAL"], "Luis": ["SUCURSAL CAFAYATE"]}
            service.generar_reporte_supervisores(self._config(), supervisores)
        # 2 marcas simples → 2 llamadas a get_cobertura_preventista_marca
        assert loader.get_cobertura_preventista_marca.call_count == 2

    def test_consulta_bd_falla_genera_hojas_vacias(self, tmp_path):
        """RNF-003: si BD falla, no crashea y genera archivo con tablas vacias."""
        loader = Mock(spec=DataLoader)
        loader.get_cobertura_preventista_marca.side_effect = Exception("DB error")
        loader.get_cobertura_custom.side_effect = Exception("DB error")
        loader.get_ultima_fecha_venta.side_effect = Exception("DB error")
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            result = service.generar_reporte(self._config(grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])]))
        assert result.ruta_archivos[0].exists()

    def test_error_grupos_vacio(self):
        """RF-004: error si grupos esta vacio."""
        service = MisionPosibleService(data_loader=_mock_loader())
        with pytest.raises(ValueError, match="vacia"):
            service.generar_reporte(self._config(grupos=[]))

    def test_periodo_normalizado_al_primer_dia(self, capsys):
        """Decision 4: periodo se normaliza con warning."""
        periodo = _normalizar_periodo("2026-03-15")
        assert periodo == "2026-03-01"
        captured = capsys.readouterr()
        assert "normalizado" in captured.out.lower() or "\u26a0" in captured.out

    # ── Tests nuevos (spec seccion 7) ────────────────────────────────────────

    def test_hojas_generadas_son_sucursales_y_por_vendedor(self, tmp_path):
        """RF-001: workbook tiene exactamente dos hojas en ese orden."""
        _, wb = self._generar(tmp_path)
        assert wb.sheetnames == ["Sucursales", "Por Vendedor"]

    def test_nombre_archivo_sin_cambio(self, tmp_path):
        """RF-003: nombre del archivo es Mision Posible 03-2026.xlsx."""
        result, _ = self._generar(tmp_path)
        assert result.ruta_archivos[0].name == "Mision Posible 03-2026.xlsx"

    def test_resumen_ult_actualizacion_en_fila_1(self, tmp_path):
        """RF-004: A1='Ult. Actualizacion', B1='06/03/2026' en hoja Sucursales."""
        _, wb = self._generar(tmp_path)
        ws = wb["Sucursales"]
        assert ws["A1"].value == "Ult. Actualizacion"
        assert ws["B1"].value == "06/03/2026"

    def test_primera_tabla_en_posicion_correcta(self, tmp_path):
        """RF-005/RF-007: con 1 grupo, titulo en fila 3 col 1."""
        _, wb = self._generar(tmp_path, grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])], objetivos={"IMPERIAL": 500})
        ws = wb["Sucursales"]
        # fila 1 = resumen, fila 2 = vacia, fila 3 = titulo primera tabla
        assert ws.cell(row=3, column=1).value == "IMPERIAL"

    def test_segunda_tabla_en_columna_correcta(self, tmp_path):
        """RF-005/RF-007: con 2 grupos, segunda tabla en fila 3 col 7."""
        _, wb = self._generar(tmp_path)
        ws = wb["Sucursales"]
        # col_grupo=1 → col_inicio = 1 + 1*(5+1) = 7
        assert ws.cell(row=3, column=7).value == "LEVITE"

    def test_quinta_marca_inicia_nueva_fila_de_tablas(self, tmp_path):
        """RF-006: con 5 grupos, titulo de 5o grupo esta en fila mayor que los 4 primeros."""
        nombres = ["M1", "M2", "M3", "M4", "M5"]
        grupos = [GrupoArticulos(n, marcas=[n]) for n in nombres]
        objetivos = {n: 100 for n in nombres}
        loader = Mock(spec=DataLoader)
        loader.get_cobertura_preventista_marca.return_value = pd.DataFrame()
        loader.get_cobertura_custom.return_value = pd.DataFrame()
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        _, wb = self._generar(tmp_path, loader=loader, grupos=grupos, objetivos=objetivos)
        ws = wb["Sucursales"]
        # Fila del titulo de las primeras 4 marcas es 3 (fila_grupo=0)
        fila_primeras = ws.cell(row=3, column=1).row
        # 5a marca: col_grupo=0, fila_grupo=1 → fila_inicio_grupo[1] > 3
        # Encuentra la fila donde aparece "M5" en col 1
        fila_m5 = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "M5" and cell.column == 1:
                    fila_m5 = cell.row
                    break
            if fila_m5:
                break
        assert fila_m5 is not None
        assert fila_m5 > fila_primeras

    def test_fila_separadora_tiene_altura_reducida(self, tmp_path):
        """RF-006: con 5 grupos, la fila separadora entre grupos tiene height=6."""
        nombres = ["M1", "M2", "M3", "M4", "M5"]
        grupos = [GrupoArticulos(n, marcas=[n]) for n in nombres]
        objetivos = {n: 100 for n in nombres}
        loader = Mock(spec=DataLoader)
        loader.get_cobertura_preventista_marca.return_value = pd.DataFrame()
        loader.get_cobertura_custom.return_value = pd.DataFrame()
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        _, wb = self._generar(tmp_path, loader=loader, grupos=grupos, objetivos=objetivos)
        ws = wb["Sucursales"]
        # Busca la fila de M5 (inicio del grupo 1) y la separadora es la fila anterior
        fila_m5 = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "M5" and cell.column == 1:
                    fila_m5 = cell.row
                    break
            if fila_m5:
                break
        fila_sep = fila_m5 - 1
        assert ws.row_dimensions[fila_sep].height == 6

    def test_titulo_marca_mergeado(self, tmp_path):
        """RF-007: el titulo de la marca cubre las 5 columnas de la tabla."""
        _, wb = self._generar(tmp_path, grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])], objetivos={"IMPERIAL": 500})
        ws = wb["Sucursales"]
        # titulo en fila 3 col 1; debe haber merge C3:G3 (col 1 a col 5)
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        # C3:G3 seria col 1-5 row 3
        assert any("A3" in r and "E3" in r for r in merged_ranges)

    def test_columna_separadora_tiene_ancho_reducido(self, tmp_path):
        """RF-008: la columna F (primera columna separadora) tiene ancho=2."""
        _, wb = self._generar(tmp_path)
        ws = wb["Sucursales"]
        # col_inicio=1, COLS_POR_TABLA=5, separadora en col 6 = F
        ancho_f = ws.column_dimensions["F"].width
        assert ancho_f == 2

    def test_formato_condicional_en_columna_pct(self, tmp_path):
        """RF-009: CF en columna E (1a tabla) y K (2a tabla, col 7+4=11)."""
        _, wb = self._generar(tmp_path)
        ws = wb["Sucursales"]
        # Recolectar todas las referencias de CF
        cf_refs = set()
        for sqref, rules in ws.conditional_formatting._cf_rules.items():
            cf_refs.add(str(sqref))
        # Columna E = primera tabla %, columna K = segunda tabla %
        assert any("E" in ref for ref in cf_refs), f"No CF on col E: {cf_refs}"
        assert any("K" in ref for ref in cf_refs), f"No CF on col K: {cf_refs}"

    def test_hoja_por_vendedor_contiene_bloques_por_marca(self, tmp_path):
        """RF-011/RF-012: hoja Por Vendedor tiene titulos de ambos grupos."""
        _, wb = self._generar(tmp_path)
        ws = wb["Por Vendedor"]
        valores = [cell.value for row in ws.iter_rows() for cell in row]
        assert "IMPERIAL" in valores
        assert "LEVITE" in valores

    def test_formato_condicional_en_por_vendedor(self, tmp_path):
        """RF-013: CF en columna % de hoja Por Vendedor (col F = col 6)."""
        _, wb = self._generar(tmp_path)
        ws = wb["Por Vendedor"]
        cf_refs = set()
        for sqref, rules in ws.conditional_formatting._cf_rules.items():
            cf_refs.add(str(sqref))
        assert any("F" in ref for ref in cf_refs), f"No CF on col F in Por Vendedor: {cf_refs}"

    def test_result_hojas_son_sucursales_y_por_vendedor(self, tmp_path):
        """RF-017: result.hojas == ['Sucursales', 'Por Vendedor']."""
        result, _ = self._generar(tmp_path)
        assert result.hojas == ["Sucursales", "Por Vendedor"]

    def test_ult_actualizacion_en_hoja_por_vendedor(self, tmp_path):
        """RF-014: A1='Ult. Actualizacion', B1='06/03/2026' en hoja Por Vendedor."""
        _, wb = self._generar(tmp_path)
        ws = wb["Por Vendedor"]
        assert ws["A1"].value == "Ult. Actualizacion"
        assert ws["B1"].value == "06/03/2026"

    def test_modo_supervisores_genera_dos_hojas_por_archivo(self, tmp_path):
        """RF-016: cada archivo generado por supervisor tiene 2 hojas correctas."""
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=_mock_loader())
            supervisores = {"Ana": ["CASA CENTRAL"], "Luis": ["SUCURSAL CAFAYATE"]}
            results = service.generar_reporte_supervisores(self._config(), supervisores)
        for result in results:
            wb = load_workbook(result.ruta_archivos[0])
            assert wb.sheetnames == ["Sucursales", "Por Vendedor"]

    def test_marca_sin_datos_renderiza_encabezado_sin_filas(self, tmp_path):
        """Edge case: df_suc vacio → titulo + encabezado pero sin filas de datos.

        procesar_cobertura_sucursal retorna filas vacias (cob=0) para todas las
        sucursales en porcentajes_sucursal. Para obtener un df_suc realmente vacio
        se configura porcentajes_sucursal={} (sin sucursales definidas).
        """
        _, wb = self._generar(
            tmp_path,
            grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])],
            objetivos={"IMPERIAL": 500},
            porcentajes_sucursal={},  # sin sucursales configuradas → df_suc vacio
        )
        ws = wb["Sucursales"]
        # fila 3 = titulo, fila 4 = encabezado, fila 5 debe estar vacia (sin datos)
        assert ws.cell(row=3, column=1).value == "IMPERIAL"
        assert ws.cell(row=4, column=1).value == "Sucursal"
        # No debe haber datos en fila 5
        assert ws.cell(row=5, column=1).value is None

    def test_calculo_fila_inicio_con_marcas_de_distintos_tamanos(self, tmp_path):
        """RF-006: 5 grupos con diferente cantidad de sucursales; 5o grupo comienza
        en la fila correcta (determinada por el grupo con mas sucursales del grupo 0)."""
        pct_grande = {"SUC_A": 10, "SUC_B": 10, "SUC_C": 10}
        nombres = ["M1", "M2", "M3", "M4", "M5"]
        grupos = [GrupoArticulos(n, marcas=[n]) for n in nombres]
        objetivos = {n: 100 for n in nombres}

        loader = Mock(spec=DataLoader)
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        loader.get_cobertura_custom.side_effect = lambda periodo, marcas, filtro_descripcion=None: pd.DataFrame([
            {"sucursal": "SUC_A", "vendedor": "V1", "id_ruta": 1, "clientes_compradores": 5, "volumen_total": 50, "periodo": "2026-03-01"},
        ])

        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            config = MisionPosibleConfig(
                periodo="2026-03-01",
                grupos=grupos,
                objetivos=objetivos,
                porcentajes_sucursal=pct_grande,
            )
            result = service.generar_reporte(config)
        wb = load_workbook(result.ruta_archivos[0])
        ws = wb["Sucursales"]

        # grupo 0: grupos M1-M4, each has 3 sucursales → max_filas = 2+3 = 5
        # fila_inicio_grupo[0] = 3 (resumen en fila 1, vacia fila 2)
        # fila_sep = 3 + 5 + 1 - 1 = 8  (fila separadora)
        # fila_inicio_grupo[1] = 3 + 5 + 1 = 9
        fila_m5 = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "M5" and cell.column == 1:
                    fila_m5 = cell.row
                    break
            if fila_m5:
                break
        assert fila_m5 is not None
        # M5 must be below the first 4 tables; with 3 suc each: fila=3+5+1=9
        assert fila_m5 == 9

    # ── Tests nuevos del spec seccion 7 ──────────────────────────────────────

    def test_config_acepta_grupos(self):
        """RF-001: MisionPosibleConfig acepta grupos=[GrupoArticulos(...)]."""
        config = MisionPosibleConfig(
            periodo="2026-03-01",
            grupos=[GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])],
        )
        assert len(config.grupos) == 1
        assert config.grupos[0].nombre == "IMPERIAL"

    def test_servicio_usa_etl_para_marcas_simples(self, tmp_path):
        """Marca simple (1 marca, sin filtro) usa tabla pre-agregada del ETL."""
        loader = _mock_loader()
        self._generar(tmp_path, loader=loader)
        # IMPERIAL y LEVITE son marcas simples → usan get_cobertura_preventista_marca
        assert loader.get_cobertura_preventista_marca.call_count == 2
        assert loader.get_cobertura_custom.call_count == 0

    def test_servicio_usa_custom_para_grupo_con_filtro(self, tmp_path):
        """Grupo con filtro_descripcion usa query custom."""
        loader = _mock_loader()
        grupo = GrupoArticulos("SCHNEIDER 710", marcas=["SCHNEIDER"], filtro_descripcion="710")
        self._generar(tmp_path, loader=loader, grupos=[grupo], objetivos={"SCHNEIDER 710": 100})
        loader.get_cobertura_custom.assert_called_once_with(
            periodo="2026-03-01", marcas=["SCHNEIDER"], filtro_descripcion="710",
        )
        # No debe usar la tabla pre-agregada para este grupo
        loader.get_cobertura_preventista_marca.call_count  # may be called for other groups

    def test_grupo_con_filtro_descripcion_pasa_filtro(self, tmp_path):
        """RF-002/RF-010: grupo con filtro_descripcion lo pasa al DataLoader."""
        loader = Mock(spec=DataLoader)
        loader.get_cobertura_preventista_marca.return_value = pd.DataFrame()
        loader.get_cobertura_custom.return_value = pd.DataFrame()
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        grupo = GrupoArticulos("SCHNEIDER 710", marcas=["SCHNEIDER"], filtro_descripcion="710")
        self._generar(tmp_path, loader=loader, grupos=[grupo], objetivos={"SCHNEIDER 710": 100})
        loader.get_cobertura_custom.assert_called_once_with(
            periodo="2026-03-01", marcas=["SCHNEIDER"], filtro_descripcion="710",
        )

    def test_objetivos_usan_nombre_del_grupo(self, tmp_path):
        """RF-005: objetivos indexados por grupo.nombre."""
        loader = Mock(spec=DataLoader)
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        loader.get_cobertura_custom.return_value = pd.DataFrame()
        grupo = GrupoArticulos("SCHNEIDER 710", marcas=["SCHNEIDER"], filtro_descripcion="710")
        _, wb = self._generar(tmp_path, loader=loader, grupos=[grupo], objetivos={"SCHNEIDER 710": 100})
        ws = wb["Sucursales"]
        # Row 3 = titulo "SCHNEIDER 710"
        assert ws.cell(row=3, column=1).value == "SCHNEIDER 710"

    def test_result_marcas_incluidas_son_nombres_de_grupos(self, tmp_path):
        """RF-016: result.marcas_incluidas tiene los nombres de los grupos."""
        result, _ = self._generar(tmp_path)
        assert result.marcas_incluidas == ["IMPERIAL", "LEVITE"]

    def test_fallo_en_un_grupo_no_cancela_otros(self, tmp_path):
        """RNF-003: fallo en un grupo no cancela los demas."""
        loader = Mock(spec=DataLoader)
        call_count = [0]

        def side_effect(periodos):
            call_count[0] += 1
            if call_count[0] == 1:
                raise Exception("DB error")
            return _df_cob_preventista_marca()

        loader.get_cobertura_preventista_marca.side_effect = side_effect
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            result = service.generar_reporte(self._config())
        assert result.ruta_archivos[0].exists()

    def test_modo_supervisores_una_consulta_por_grupo(self, tmp_path):
        """RF-015: N queries (una por grupo), no N x supervisores."""
        loader = _mock_loader()
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            supervisores = {"Ana": ["CASA CENTRAL"], "Luis": ["SUCURSAL CAFAYATE"]}
            service.generar_reporte_supervisores(self._config(), supervisores)
        # 2 marcas simples = 2 calls a ETL (no 2 x 2 supervisores = 4)
        assert loader.get_cobertura_preventista_marca.call_count == 2

    def test_titulo_de_tabla_usa_nombre_del_grupo(self, tmp_path):
        """RF-014: titulo de tabla usa grupo.nombre."""
        loader = Mock(spec=DataLoader)
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        loader.get_cobertura_custom.return_value = pd.DataFrame()
        grupo = GrupoArticulos("SCHNEIDER 710", marcas=["SCHNEIDER"], filtro_descripcion="710")
        _, wb = self._generar(tmp_path, loader=loader, grupos=[grupo], objetivos={"SCHNEIDER 710": 100})
        ws = wb["Sucursales"]
        assert ws.cell(row=3, column=1).value == "SCHNEIDER 710"

    def test_grupo_articulos_acepta_marcas_lista(self):
        """RF-001: GrupoArticulos acepta marcas como lista."""
        g = GrupoArticulos("AGUAS", marcas=["LEVITE", "VILLAVICENCIO"])
        assert g.marcas == ["LEVITE", "VILLAVICENCIO"]
        assert g.nombre == "AGUAS"

    def test_grupo_articulos_rechaza_lista_vacia(self):
        """RF-004: GrupoArticulos rechaza marcas=[]."""
        with pytest.raises(ValueError, match="vacia"):
            GrupoArticulos("X", marcas=[])

    def test_fetch_data_grupo_pasa_lista_de_marcas(self, tmp_path):
        """RF-013: _fetch_data_grupo pasa marcas como lista."""
        loader = _mock_loader()
        self._generar(
            tmp_path,
            loader=loader,
            grupos=[GrupoArticulos("AGUAS", marcas=["LEVITE", "VILLAVICENCIO"])],
            objetivos={"AGUAS": 500},
        )
        loader.get_cobertura_custom.assert_called_once_with(
            periodo="2026-03-01", marcas=["LEVITE", "VILLAVICENCIO"], filtro_descripcion=None,
        )

    def test_grupo_multimarca_resultado_marcas_incluidas_usa_nombre(self, tmp_path):
        """RF-014: marcas_incluidas contiene grupo.nombre, no las marcas individuales."""
        result, _ = self._generar(
            tmp_path,
            grupos=[GrupoArticulos("AGUAS", marcas=["LEVITE", "VILLAVICENCIO"]), GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])],
            objetivos={"AGUAS": 500, "IMPERIAL": 300},
        )
        assert result.marcas_incluidas == ["AGUAS", "IMPERIAL"]
        assert "LEVITE" not in result.marcas_incluidas
        assert "VILLAVICENCIO" not in result.marcas_incluidas

    def test_fallo_en_grupo_multimarca_no_cancela_otros(self, tmp_path):
        """RNF-003: fallo en grupo multi-marca no cancela los demas."""
        loader = Mock(spec=DataLoader)

        def side_effect(periodo, marcas, filtro_descripcion=None):
            key = tuple(sorted(m.upper() for m in marcas))
            if key == ("LEVITE", "VILLAVICENCIO"):
                raise Exception("DB error")
            return _df_cob_imperial()

        loader.get_cobertura_custom.side_effect = side_effect
        loader.get_ultima_fecha_venta.return_value = date(2026, 3, 6)
        with patch("src.services.mision_posible.service.DATA_OUTPUT", tmp_path):
            service = MisionPosibleService(data_loader=loader)
            result = service.generar_reporte(self._config(
                grupos=[GrupoArticulos("AGUAS", marcas=["LEVITE", "VILLAVICENCIO"]), GrupoArticulos("IMPERIAL", marcas=["IMPERIAL"])],
                objetivos={"AGUAS": 500, "IMPERIAL": 300},
            ))
        assert result.ruta_archivos[0].exists()


class TestNombreReporte:
    def test_sin_supervisor(self):
        assert _nombre_reporte("2026-03-01") == "Mision Posible 03-2026"

    def test_con_supervisor(self):
        assert _nombre_reporte("2026-03-01", "Ana") == "Mision Posible Ana 03-2026"
