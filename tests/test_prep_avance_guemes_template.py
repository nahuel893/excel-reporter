"""Tests for the one-time AVANCE GUEMES template-prep script (header-only edits)."""
from openpyxl import Workbook, load_workbook

import scripts.prep_avance_guemes_template as prep


def _make_template(path, *, legend="GENERICO", data_col="GENERICO"):
    """Minimal CuposVolumen-only workbook mimicking the pre-prep GUEMES template:
    data col C and legend col P both headed 'GENERICO', no CuposCoberGen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CuposVolumen"
    headers = ["Código", "Descripción", data_col, "DESAGREGADO", "Cupo "]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=1, column=16, value=legend)  # col P
    wb.save(path)


def test_prep_renames_legend_and_creates_cobergen(tmp_path):
    path = tmp_path / "AVANCE GUEMES.xlsx"
    _make_template(path)

    assert prep.main(str(path)) == 0

    out = load_workbook(path)
    ws = out["CuposVolumen"]
    assert ws.cell(row=1, column=3).value == "GENERICO"          # data col C intact
    assert ws.cell(row=1, column=16).value == "GENERICO_LEGEND"  # legend renamed
    assert "CuposCoberGen" in out.sheetnames
    cg = out["CuposCoberGen"]
    assert [cg.cell(row=1, column=c).value for c in range(1, 7)] == [
        None, "Ruta", "Preventista", "Generico", "ZONA", "CUPO "
    ]


def test_prep_is_idempotent(tmp_path):
    path = tmp_path / "AVANCE GUEMES.xlsx"
    _make_template(path)

    assert prep.main(str(path)) == 0
    assert prep.main(str(path)) == 0  # second run: no-op, still succeeds

    out = load_workbook(path)
    assert out["CuposVolumen"].cell(row=1, column=16).value == "GENERICO_LEGEND"
    # CuposCoberGen not duplicated
    assert out.sheetnames.count("CuposCoberGen") == 1


def test_prep_aborts_when_data_col_not_generico(tmp_path):
    """If col C is not the GENERICO data column, the template is not what we
    expect — abort rather than corrupt it."""
    path = tmp_path / "AVANCE GUEMES.xlsx"
    _make_template(path, data_col="OTRA_COSA")

    assert prep.main(str(path)) == 1
    # legend header must be left untouched on abort
    assert load_workbook(path)["CuposVolumen"].cell(row=1, column=16).value == "GENERICO"


def test_prep_missing_file_returns_error(tmp_path):
    assert prep.main(str(tmp_path / "does-not-exist.xlsx")) == 1
