"""Tests for RangeRecognizer — maximal medium-bordered rectangle detection."""
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from src.core.range_recognizer import RangeRecognizer

REAL_XLSX_CANDIDATES = [
    Path("data/output/avances/2026-07/AVANCE BADIE - JULIO 2026.xlsx"),
    Path("data/input/avances/AVANCE BADIE.xlsx"),
]


def _real_xlsx_path() -> Path | None:
    for candidate in REAL_XLSX_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


def _set_box_border(ws, r1: int, c1: int, r2: int, c2: int, style: str = "medium") -> None:
    """Sets an outline border around [r1,c1]..[r2,c2], mirroring how Excel
    applies an 'outside border' to a selection: each side is set only on the
    boundary cells that touch it."""
    side = Side(style=style)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            cell.border = Border(
                top=side if r == r1 else existing.top,
                bottom=side if r == r2 else existing.bottom,
                left=side if c == c1 else existing.left,
                right=side if c == c2 else existing.right,
            )


def _set_full_grid_border(ws, r1: int, c1: int, r2: int, c2: int, style: str = "medium") -> None:
    """Borders EVERY cell in [r1,c1]..[r2,c2] on all four sides — mirrors a
    fully gridlined data table (every cell boundary is a wall), which is
    what produces nested sub-rectangles inside an outer block."""
    side = Side(style=style)
    border = Border(top=side, bottom=side, left=side, right=side)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


class TestSingleBorderedRect:
    def test_single_bordered_rect(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        xlsx = tmp_path / "single.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["B2:D5"]


class TestReadingOrder:
    def test_two_stacked_rects_reading_order(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 4)  # A1:D3
        _set_box_border(ws, 5, 1, 7, 4)  # A5:D7 (row 4 gap, no border)
        xlsx = tmp_path / "stacked.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:D3", "A5:D7"]

    def test_2x2_grid_reading_order(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 3)  # A1:C3 top-left
        _set_box_border(ws, 1, 5, 3, 7)  # E1:G3 top-right
        _set_box_border(ws, 5, 1, 7, 3)  # A5:C7 bottom-left
        _set_box_border(ws, 5, 5, 7, 7)  # E5:G7 bottom-right
        xlsx = tmp_path / "grid2x2.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == [
            "A1:C3", "E1:G3", "A5:C7", "E5:G7",
        ]

    def test_two_side_by_side_plus_one_below(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 2)  # A1:B3 (left)
        _set_box_border(ws, 1, 4, 3, 5)  # D1:E3 (right, side by side)
        _set_box_border(ws, 5, 1, 7, 5)  # A5:E7 (below, spans both)
        xlsx = tmp_path / "side_by_side_below.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == [
            "A1:B3", "D1:E3", "A5:E7",
        ]


class TestMaximalReduction:
    def test_nested_gridlines_only_outer_maximal_returned(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Fully gridlined 4x4 block: every cell bordered on all sides ->
        # ~100 closed sub-rectangles exist, but only the outer 4x4 block
        # is maximal (cannot grow further and stay closed).
        _set_full_grid_border(ws, 1, 1, 4, 4)
        xlsx = tmp_path / "nested.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:D4"]

    def test_adjacent_rects_sharing_a_wall_both_returned(self, tmp_path):
        """Two boxes of DIFFERENT heights, directly touching (no gap
        column) at the B/C boundary. Because their heights differ, their
        combined bounding box can never itself be a valid closed rectangle
        (the shorter box has no wall data at the taller box's extra row),
        so containment-based maximality correctly keeps both as distinct
        regions instead of collapsing them into one union."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 4, 2)  # A1:B4 (4 rows tall)
        _set_box_border(ws, 1, 3, 3, 4)  # C1:D3 (3 rows tall, shares the B/C wall)
        xlsx = tmp_path / "adjacent.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:B4", "C1:D3"]


class TestContainmentMaximalitySynthetic:
    def test_subdivided_cards_collapse_to_outer_and_gap_siblings_stay_distinct(self, tmp_path):
        """Committed regression guard for the containment-based maximality
        decision — the risky replacement of the old grow-by-one strategy.
        The golden tests skip when the gitignored real file is absent, so
        this synthetic case (no real file) pins the same shape grow-by-one
        failed on, keeping CI enforcement.

        Card #1 is fully gridlined: its interior is subdivided into mini-cells
        by internal walls, mimicking Cober Nueva's BZ19:CW32-style widget
        subdivision. A naive grow-by-one would fragment such a card into every
        interior sub-rectangle; containment-maximality must collapse the whole
        thing into the single outer card. Card #2 is a sibling separated by a
        FULL gap row (row 6) AND gap column (column 6), which breaks wall
        continuity so no accidental union can close — proving gap-separated
        siblings stay distinct."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # OUTER card #1: fully gridlined 5x5 -> perimeter + interior mini-cells
        _set_full_grid_border(ws, 1, 1, 5, 5)   # A1:E5
        # OUTER card #2: gap row 6 AND gap column 6 separate it from card #1
        _set_full_grid_border(ws, 7, 7, 11, 11)  # G7:K11
        xlsx = tmp_path / "subdivided_gap_siblings.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        # Exactly the two OUTER cards, in reading order — no interior piece.
        assert recognizer.detect_ranges("Sheet1") == ["A1:E5", "G7:K11"]


class TestAdjacencyOrWall:
    def test_wall_declared_on_only_one_neighbor_side(self, tmp_path):
        """The box's right wall is declared ONLY via the LEFT side of the
        cells in the column outside the box (column C) — the box's own
        right side (column B) never carries a border. Detection must still
        close the box using the adjacency-OR rule."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        medium = Side(style="medium")
        ws.cell(row=1, column=1).border = Border(top=medium, left=medium)
        ws.cell(row=1, column=2).border = Border(top=medium)
        ws.cell(row=2, column=1).border = Border(left=medium, bottom=medium)
        ws.cell(row=2, column=2).border = Border(bottom=medium)
        ws.cell(row=1, column=3).border = Border(left=medium)
        ws.cell(row=2, column=3).border = Border(left=medium)
        xlsx = tmp_path / "adjacency_or.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:B2"]


class TestEdgeClamp:
    def test_rect_touching_row1_and_col_a(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 3)  # A1:C3 — touches the sheet's top-left corner
        xlsx = tmp_path / "edge_clamp.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:C3"]


class TestMergedCells:
    def test_merged_cell_border_only_on_anchor_still_closes_box(self, tmp_path):
        """A1:B1 is merged; the merge's border is carried only on the
        top-left anchor cell (A1) — the swallowed cell (B1) has no border
        of its own. Detection must route the query through the merge index
        to still find the box's continuous top edge."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        medium = Side(style="medium")

        ws.merge_cells("A1:B1")
        ws.cell(row=1, column=1).border = Border(top=medium, left=medium)
        ws.cell(row=1, column=3).border = Border(top=medium, right=medium)
        ws.cell(row=2, column=1).border = Border(left=medium)
        ws.cell(row=2, column=3).border = Border(right=medium)
        ws.cell(row=3, column=1).border = Border(left=medium, bottom=medium)
        ws.cell(row=3, column=2).border = Border(bottom=medium)
        ws.cell(row=3, column=3).border = Border(bottom=medium, right=medium)

        xlsx = tmp_path / "merged_anchor_only.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["A1:C3"]


class TestDetectAll:
    def test_detect_all_returns_ranges_per_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Uno"
        _set_box_border(ws1, 1, 1, 2, 2)  # A1:B2

        ws2 = wb.create_sheet("Dos")
        _set_box_border(ws2, 1, 1, 3, 3)  # A1:C3

        xlsx = tmp_path / "multi_sheet.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        result = recognizer.detect_all()
        assert result["Uno"] == ["A1:B2"]
        assert result["Dos"] == ["A1:C3"]


class TestRangeRecognizerInit:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            RangeRecognizer(tmp_path / "does_not_exist.xlsx")

    def test_missing_sheet_raises(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        xlsx = tmp_path / "empty.xlsx"
        wb.save(xlsx)
        recognizer = RangeRecognizer(xlsx)
        with pytest.raises(ValueError, match="NoExiste"):
            recognizer.detect_ranges("NoExiste")


class TestDocumentedUnionLimitation:
    def test_same_height_touching_cards_report_union_characterization(self, tmp_path):
        """CHARACTERIZATION test pinning the KNOWN, documented limitation in
        the module docstring ("Known limitation"): two SAME-height cards that
        directly touch (no gap) and whose union also forms a valid closed
        rectangle collapse into that union — the two individual pieces are
        dropped.

        This intentionally asserts the CURRENT documented behavior, NOT a
        desired one, so any future change to the containment rule becomes
        regression-visible here. Do NOT 'fix' this by changing the algorithm:
        the target templates always gap-separate distinct sibling cards, so
        this edge case is out of scope and safe for them."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 2)  # A1:B3 (3 rows tall)
        _set_box_border(ws, 1, 3, 3, 4)  # C1:D3 (same height, touches at B/C)
        xlsx = tmp_path / "union_limitation.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        # Documented limitation: the closed union wins, both pieces dropped.
        assert recognizer.detect_ranges("Sheet1") == ["A1:D3"]


class TestBorderStyles:
    @pytest.mark.parametrize("style", ["thick", "double"])
    def test_non_medium_wall_styles_detected(self, tmp_path, style):
        """DEFAULT_BORDER_STYLES accepts ("medium", "thick", "double"), but
        only 'medium' is exercised elsewhere. This covers the 'thick' and
        'double' membership paths through _is_wall."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4, style=style)  # B2:D5
        xlsx = tmp_path / f"style_{style}.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["B2:D5"]


class TestContextManager:
    def test_with_block_releases_cached_workbook(self, tmp_path):
        """The context-manager protocol must close/release the cached
        workbook on exit, while the explicit close() path keeps working."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 2, 2)  # A1:B2
        xlsx = tmp_path / "ctx.xlsx"
        wb.save(xlsx)

        with RangeRecognizer(xlsx) as recognizer:
            assert recognizer.detect_ranges("Sheet1") == ["A1:B2"]
            assert recognizer._workbook is not None  # cached inside the block
        # __exit__ must have released the cached workbook.
        assert recognizer._workbook is None


class TestCaptionExtraction:
    def test_caption_is_first_non_empty_string_cell_top_to_bottom_left_to_right(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        # First (reading-order) cell in the region carries a NUMBER, not a
        # string — it must be skipped. The caption is the first STRING cell.
        ws.cell(row=2, column=2, value=42)          # B2 — numeric, skipped
        ws.cell(row=2, column=3, value="GFLORES")    # C2 — first string cell
        ws.cell(row=3, column=2, value="ignored")    # later in reading order
        xlsx = tmp_path / "caption_scan.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1") == [
            ("B2:D5", "GFLORES"),
        ]

    def test_caption_falls_back_to_a1_when_region_has_no_text(self, tmp_path):
        """A text-less region on the DEFAULT-scan path falls back to its own
        A1 address (never None), so it still gets a unique, non-empty caption."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        ws.cell(row=2, column=2, value=42)  # numeric only, no text anywhere
        ws.cell(row=3, column=3, value=100)
        xlsx = tmp_path / "caption_empty.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1") == [
            ("B2:D5", "B2:D5"),
        ]

    def test_two_text_less_regions_get_distinct_captions(self, tmp_path):
        """Two bordered regions that BOTH have no text cell must still get
        DISTINCT captions — each falls back to its own (unique) A1 address
        instead of both collapsing to None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 2, 2)  # A1:B2 (text-less)
        _set_box_border(ws, 1, 4, 2, 5)  # D1:E2 (text-less, side by side)
        # No string cells anywhere in either region.
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=4, value=2)
        xlsx = tmp_path / "caption_two_textless.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        pairs = recognizer.detect_ranges_with_captions("Sheet1")
        assert pairs == [("A1:B2", "A1:B2"), ("D1:E2", "D1:E2")]
        captions = [c for _, c in pairs]
        assert len(set(captions)) == len(captions)  # distinct
        assert all(c is not None for c in captions)

    def test_multiple_regions_get_independent_captions_in_reading_order(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 2)  # A1:B3
        _set_box_border(ws, 5, 1, 7, 2)  # A5:B7
        ws.cell(row=1, column=1, value="Region Uno")
        ws.cell(row=5, column=1, value="Region Dos")
        xlsx = tmp_path / "caption_multi.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1") == [
            ("A1:B3", "Region Uno"),
            ("A5:B7", "Region Dos"),
        ]

    def test_caption_anchor_override_bypasses_scan(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        ws.cell(row=2, column=2, value="First Scan Cell")  # would win by default scan
        ws.cell(row=3, column=3, value="Anchor Cell")       # relative offset row+1,col+1 -> C3
        xlsx = tmp_path / "caption_anchor.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        # caption_anchor="B2" is relative to each region's top-left corner
        # (row offset 1, col offset 1 -> region's row1+1, col1+1 -> row3,col3).
        assert recognizer.detect_ranges_with_captions("Sheet1", caption_anchor="B2") == [
            ("B2:D5", "Anchor Cell"),
        ]

    def test_caption_anchor_cell_with_no_text_returns_none(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        ws.cell(row=2, column=2, value="First Scan Cell")  # ignored, anchor wins
        xlsx = tmp_path / "caption_anchor_empty.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1", caption_anchor="B2") == [
            ("B2:D5", None),
        ]

    def test_caption_anchor_outside_region_bounds_returns_none(self, tmp_path):
        """An anchor offset that lands PAST the region's bottom/right edge must
        NOT silently read a foreign cell — it is clamped/rejected, returning
        None instead of the out-of-region value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5 (rows 2-5, cols 2-4)
        # Foreign cell OUTSIDE the region — must never be picked up.
        ws.cell(row=11, column=2, value="Foreign Cell")
        xlsx = tmp_path / "caption_anchor_oob.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        # caption_anchor "A10" -> offset (9, 0) -> target row 2+9=11 > region
        # max row 5 -> outside bounds -> None (NOT "Foreign Cell").
        assert recognizer.detect_ranges_with_captions("Sheet1", caption_anchor="A10") == [
            ("B2:D5", None),
        ]

    def test_colliding_first_cells_are_disambiguated_with_next_text_cell(self, tmp_path):
        """Two sibling regions whose first-scanned cell holds the IDENTICAL
        generic label (mirrors real templates like Cober Nueva's repeated
        "Cobertura Cervezas 1" band header). The second region's caption
        must extend with its own next distinct text cell instead of
        colliding with the first region's caption."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 3, 3)  # A1:C3
        _set_box_border(ws, 1, 5, 3, 7)  # E1:G3 (side by side, same row band)
        ws.cell(row=1, column=1, value="Cobertura Cervezas 1")
        ws.cell(row=1, column=2, value="SALTA")
        ws.cell(row=1, column=5, value="Cobertura Cervezas 1")
        ws.cell(row=1, column=6, value="BIECKERT")
        xlsx = tmp_path / "caption_collision.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1") == [
            ("A1:C3", "Cobertura Cervezas 1"),
            ("E1:G3", "Cobertura Cervezas 1 — BIECKERT"),
        ]

    def test_fully_identical_region_text_falls_back_to_a1_address(self, tmp_path):
        """Degenerate case: BOTH regions have byte-identical text content
        (no distinguishing cell anywhere). The second region's caption must
        still be unique, falling back to appending its own A1 address."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 1, 1, 2, 2)  # A1:B2
        _set_box_border(ws, 1, 4, 2, 5)  # D1:E2 (side by side)
        ws.cell(row=1, column=1, value="Same Label")
        ws.cell(row=1, column=4, value="Same Label")
        xlsx = tmp_path / "caption_fallback.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges_with_captions("Sheet1") == [
            ("A1:B2", "Same Label"),
            ("D1:E2", "Same Label (D1:E2)"),
        ]

    def test_detect_ranges_unchanged_alongside_captions(self, tmp_path):
        """detect_ranges() must keep returning bare A1 strings, unaffected
        by the new detect_ranges_with_captions() method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _set_box_border(ws, 2, 2, 5, 4)  # B2:D5
        ws.cell(row=2, column=2, value="Caption")
        xlsx = tmp_path / "caption_no_regression.xlsx"
        wb.save(xlsx)

        recognizer = RangeRecognizer(xlsx)
        assert recognizer.detect_ranges("Sheet1") == ["B2:D5"]
        assert recognizer.detect_ranges_with_captions("Sheet1") == [("B2:D5", "Caption")]


@pytest.mark.skipif(_real_xlsx_path() is None, reason="Real AVANCE BADIE xlsx not present locally")
class TestGoldenRealFile:
    """Ground-truth check against the real production template.
    Guarded: only runs when the (gitignored) real file exists locally.

    Uses a single class-scoped RangeRecognizer instance so the (slow,
    non-read_only) workbook load happens only once for both sheets."""

    @pytest.fixture(scope="class")
    def recognizer(self):
        xlsx = _real_xlsx_path()
        rr = RangeRecognizer(xlsx)
        yield rr
        rr.close()

    def test_avance_sheet_ground_truth(self, recognizer):
        assert recognizer.detect_ranges("Avance") == [
            "A1:AR18", "A20:AR33", "A35:AR48", "A50:AR57",
        ]

    def test_cober_nueva_sheet_ground_truth(self, recognizer):
        assert recognizer.detect_ranges("Cober Nueva") == [
            "A2:R17", "T2:AW17", "AY2:BT3", "BZ2:CS3", "CY2:DR3",
            "AY5:BX17", "BZ5:CW17", "CY5:DV17",
            "A19:R32", "T19:AW32", "AY19:BX32", "BZ19:CW32", "CY19:DV32",
            "A34:R47", "BZ34:CS47",
            "A49:R55", "T49:AW55", "AY49:BX55", "BZ49:CW55", "CY49:DV55",
        ]

    def test_avance_default_scan_captions_are_first_text_cells(self, recognizer):
        """The DEFAULT scan (no anchor) returns each region's first non-empty
        text cell in reading order. These are the HONEST current values — they
        are NOT meaningful supervisor labels (note 'c' and the plain 'SALTA'
        band header). Meaningful per-card captions REQUIRE caption_anchor
        (see test_avance_supervisor_captions_via_anchor; wired in PR4). Pinning
        the real values here makes any regression in the scan/disambiguation
        (including the A1 fallback for text-less regions) visible."""
        pairs = recognizer.detect_ranges_with_captions("Avance")
        assert pairs == [
            ("A1:AR18", "Días Hábiles:"),
            ("A20:AR33", "SALTA"),
            ("A35:AR48", "c"),
            ("A50:AR57", "SALTA — SALTA CAUTIVA1"),
        ]

    def test_avance_supervisor_captions_via_anchor(self, recognizer):
        """LOAD-BEARING: caption_anchor="C3" is offset (row+2, col+2) relative
        to each card's top-left. Every supervisor card starts at column A, so
        the anchor resolves to column C, row (top+2) — the cell that actually
        holds the supervisor code. This pins the REAL behavior PR4 relies on:
        the 3 supervisor cards must yield FGUANTAY, VCHAPUR, GFARAH."""
        pairs = recognizer.detect_ranges_with_captions("Avance", caption_anchor="C3")
        by_range = dict(pairs)
        assert by_range["A20:AR33"] == "FGUANTAY"
        assert by_range["A35:AR48"] == "VCHAPUR"
        assert by_range["A50:AR57"] == "GFARAH"
        # The header card (A1:AR18) has no text at the C3 anchor -> None.
        assert by_range["A1:AR18"] is None

    def test_cober_nueva_captions_are_distinct_and_non_empty(self, recognizer):
        pairs = recognizer.detect_ranges_with_captions("Cober Nueva")
        assert len(pairs) == 20
        captions = [c for _, c in pairs]
        assert all(c is not None and c.strip() != "" for c in captions), captions
        assert len(set(captions)) == len(captions), captions
