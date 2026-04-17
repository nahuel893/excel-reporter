import pytest
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import tempfile
from unittest.mock import patch

from src.core.excel_writer import generar_excel, ColumnFormat, SheetStyle


class TestGenerarExcel:
    """Tests para generar_excel."""

    @pytest.fixture
    def df_ejemplo(self):
        """DataFrame de ejemplo para tests."""
        return pd.DataFrame({
            "Sucursal": ["SUC1", "SUC1"],
            "Generico": ["CERVEZAS", "CERVEZAS"],
            "Cantidad (Generico)": [150, None],
            "Tendencia (Generico)": [200, None],
            "Monto (Generico)": [7500, None],
            "Marca": ["CORONA", "HEINEKEN"],
            "Cantidad (Marca)": [100, 50],
            "Tendencia (Marca)": [130, 70],
            "Monto (Marca)": [5000, 2500]
        })

    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """Directorio temporal para output."""
        return tmp_path

    def test_genera_archivo_xlsx(self, df_ejemplo, temp_output_dir):
        """Verifica que se genera un archivo .xlsx."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test_reporte")

        assert ruta.exists()
        assert ruta.suffix == ".xlsx"

    def test_nombre_archivo_correcto(self, df_ejemplo, temp_output_dir):
        """Verifica el nombre del archivo generado."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "ventas_2026-01-01_2026-01-31")

        assert ruta.name == "ventas_2026-01-01_2026-01-31.xlsx"

    def test_hoja_tiene_nombre_datos(self, df_ejemplo, temp_output_dir):
        """Verifica que la hoja se llama 'Datos' por defecto."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        wb = load_workbook(ruta)
        assert "Datos" in wb.sheetnames

    def test_encabezados_correctos(self, df_ejemplo, temp_output_dir):
        """Verifica que los encabezados son correctos."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        wb = load_workbook(ruta)
        ws = wb.active

        # Primera fila debe ser encabezados
        headers = [cell.value for cell in ws[1]]
        assert headers == list(df_ejemplo.columns)

    def test_datos_correctos(self, df_ejemplo, temp_output_dir):
        """Verifica que los datos se escriben correctamente."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        wb = load_workbook(ruta)
        ws = wb.active

        # Segunda fila (primer registro de datos)
        assert ws.cell(row=2, column=1).value == "SUC1"
        assert ws.cell(row=2, column=6).value == "CORONA"
        assert ws.cell(row=2, column=7).value == 100

    def test_encabezados_en_negrita(self, df_ejemplo, temp_output_dir):
        """Verifica que los encabezados estan en negrita."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        wb = load_workbook(ruta)
        ws = wb.active

        for cell in ws[1]:
            assert cell.font.bold is True

    def test_encabezados_centrados(self, df_ejemplo, temp_output_dir):
        """Verifica que los encabezados estan centrados."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        wb = load_workbook(ruta)
        ws = wb.active

        for cell in ws[1]:
            assert cell.alignment.horizontal == "center"

    def test_crea_directorio_si_no_existe(self, df_ejemplo, tmp_path):
        """Verifica que crea el directorio de salida si no existe."""
        nuevo_dir = tmp_path / "nuevo" / "directorio"

        with patch("src.core.excel_writer.DATA_OUTPUT", nuevo_dir):
            ruta = generar_excel(df_ejemplo, "test")

        assert nuevo_dir.exists()
        assert ruta.exists()

    def test_dataframe_vacio(self, temp_output_dir):
        """Verifica comportamiento con DataFrame vacio."""
        df_vacio = pd.DataFrame(columns=["Col1", "Col2"])

        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_vacio, "test_vacio")

        assert ruta.exists()

        wb = load_workbook(ruta)
        ws = wb.active
        # Solo debe tener encabezados
        assert ws.max_row == 1

    def test_retorna_path(self, df_ejemplo, temp_output_dir):
        """Verifica que retorna un objeto Path."""
        with patch("src.core.excel_writer.DATA_OUTPUT", temp_output_dir):
            ruta = generar_excel(df_ejemplo, "test")

        assert isinstance(ruta, Path)


class TestColumnFormatHidden:
    """Tests para la opcion hidden de ColumnFormat."""

    @pytest.fixture
    def df_ejemplo(self):
        return pd.DataFrame({
            "Sucursal": ["SUC1", "SUC2"],
            "Monto": [1000, 2000],
            "Cantidad": [10, 20],
        })

    def test_hidden_oculta_columna(self, df_ejemplo, tmp_path):
        """Una columna marcada hidden=True debe quedar oculta en el Excel."""
        style = SheetStyle(
            column_formats={
                "Monto": ColumnFormat(number_format="$ #,##0", hidden=True),
            }
        )

        with patch("src.core.excel_writer.DATA_OUTPUT", tmp_path):
            ruta = generar_excel(df_ejemplo, "test_hidden", style=style)

        wb = load_workbook(ruta)
        ws = wb.active
        # "Monto" es la columna B
        assert ws.column_dimensions["B"].hidden is True
        # "Sucursal" (A) y "Cantidad" (C) NO estan ocultas
        assert ws.column_dimensions["A"].hidden is False
        assert ws.column_dimensions["C"].hidden is False

    def test_hidden_default_false(self, df_ejemplo, tmp_path):
        """Sin setear hidden, la columna queda visible."""
        style = SheetStyle(
            column_formats={
                "Monto": ColumnFormat(number_format="$ #,##0"),
            }
        )

        with patch("src.core.excel_writer.DATA_OUTPUT", tmp_path):
            ruta = generar_excel(df_ejemplo, "test_visible", style=style)

        wb = load_workbook(ruta)
        ws = wb.active
        assert ws.column_dimensions["B"].hidden is False
