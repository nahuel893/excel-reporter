"""
Unit tests for reporte-general-badie processor:
  - _generar_trimestres
  - build_workbook (sheet structure, formulas, formats, DataValidation)
"""

import pandas as pd
import pytest
from openpyxl.worksheet.datavalidation import DataValidation

from src.services.reporte_general_badie.processor import (
    _generar_trimestres,
    build_workbook,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_ventas_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL", "SUCURSAL CAFAYATE"],
            "generico": ["CERVEZAS", "AGUAS DANONE", "CERVEZAS"],
            "anio": [2026, 2026, 2026],
            "trimestre": [2, 2, 2],
            "bultos": [1000, 200, 500],
        }
    )


def _make_cob_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "anio": [2026, 2026],
            "trimestre": [2, 2],
            "id_cliente": [101, 102],
            "bultos": [5, 3],
            "bultos_sin_regalos": [5, 2],
            "bultos_aguas_danone": [4, 1],
            "bultos_aguas_danone_sin_regalos": [4, 1],
            "meses_con_compra": [3, 2],
        }
    )


SUCURSALES = ["CASA CENTRAL", "SUCURSAL CAFAYATE"]
TRIMESTRES = ["2026-Q1", "2026-Q2"]


@pytest.fixture
def wb():
    return build_workbook(
        sucursales=SUCURSALES,
        df_ventas=_make_ventas_df(),
        df_cob=_make_cob_df(),
        trimestres=TRIMESTRES,
    )


# ---------------------------------------------------------------------------
# _generar_trimestres
# ---------------------------------------------------------------------------


class TestGenerarTrimestres:
    def test_single_quarter(self):
        result = _generar_trimestres("2026-04-01", "2026-04-30")
        assert result == ["2026-Q2"]

    def test_range_within_same_year(self):
        result = _generar_trimestres("2026-01-01", "2026-09-30")
        assert result == ["2026-Q1", "2026-Q2", "2026-Q3"]

    def test_range_crossing_year_boundary(self):
        result = _generar_trimestres("2025-10-01", "2026-03-31")
        assert result == ["2025-Q4", "2026-Q1"]

    def test_multi_year_range_count(self):
        # 2024: Q1..Q4 (4) + 2025: Q1..Q4 (4) + 2026: Q1..Q2 (2) = 10
        result = _generar_trimestres("2024-01-01", "2026-04-30")
        assert len(result) == 10

    def test_first_and_last_entries(self):
        result = _generar_trimestres("2024-01-01", "2026-04-30")
        assert result[0] == "2024-Q1"
        assert result[-1] == "2026-Q2"

    def test_q4_to_q1(self):
        result = _generar_trimestres("2024-12-01", "2025-01-31")
        assert result == ["2024-Q4", "2025-Q1"]

    def test_uses_only_year_month_from_date(self):
        # Day part of dates should be ignored
        result = _generar_trimestres("2026-03-15", "2026-04-01")
        assert result == ["2026-Q1", "2026-Q2"]


# ---------------------------------------------------------------------------
# Sheet order and presence
# ---------------------------------------------------------------------------


class TestSheetStructure:
    def test_sheet_count(self, wb):
        assert len(wb.sheetnames) == 4

    def test_sheet_order(self, wb):
        assert wb.sheetnames == [
            "Reporte", "VentasCCU", "CoberturaCCU", "_Trimestres",
        ]

    def test_reporte_is_first(self, wb):
        assert wb.sheetnames[0] == "Reporte"

    def test_trimestres_is_last(self, wb):
        assert wb.sheetnames[-1] == "_Trimestres"

    def test_trimestres_is_hidden(self, wb):
        ws = wb["_Trimestres"]
        assert ws.sheet_state == "hidden"

    def test_reporte_is_visible(self, wb):
        ws = wb["Reporte"]
        assert ws.sheet_state != "hidden"

    def test_ventas_ccu_is_visible(self, wb):
        ws = wb["VentasCCU"]
        assert ws.sheet_state != "hidden"

    def test_cobertura_is_visible(self, wb):
        ws = wb["CoberturaCCU"]
        assert ws.sheet_state != "hidden"


# ---------------------------------------------------------------------------
# _Trimestres sheet content
# ---------------------------------------------------------------------------


class TestTrimestresSheet:
    def test_values_match_input(self, wb):
        ws = wb["_Trimestres"]
        values = [ws.cell(row=i + 1, column=1).value for i in range(len(TRIMESTRES))]
        assert values == TRIMESTRES

    def test_count(self, wb):
        ws = wb["_Trimestres"]
        count = sum(
            1 for row in ws.iter_rows(min_col=1, max_col=1) if row[0].value is not None
        )
        assert count == len(TRIMESTRES)


# ---------------------------------------------------------------------------
# Reporte sheet — static cells
# ---------------------------------------------------------------------------


class TestReporteStaticCells:
    def test_a1_value(self, wb):
        ws = wb["Reporte"]
        assert ws["A1"].value == "Reporte General Badie"

    def test_a1_is_bold(self, wb):
        ws = wb["Reporte"]
        assert ws["A1"].font.bold is True

    def test_a1_font_size(self, wb):
        ws = wb["Reporte"]
        assert ws["A1"].font.size == 14

    def test_a2_label(self, wb):
        ws = wb["Reporte"]
        assert ws["A2"].value == "Trimestre:"

    def test_b2_initial_value_is_last_quarter(self, wb):
        ws = wb["Reporte"]
        assert ws["B2"].value == TRIMESTRES[-1]

    def test_row4_headers(self, wb):
        ws = wb["Reporte"]
        expected = [
            "Sucursal", "Total CCU", "Total CCU Año Anterior", "AA vs MMAA",
            "% Cerveza", "% Aguas Danone", "% Multi CCU",
            "Cobertura Normal", "Cobertura ≥3 Bultos",
            "Cobertura Prom. ≥1 Bulto/Mes",
            "Cob. Normal (s/regalos)",
            "Cob. ≥3 Bultos (s/regalos)",
            "Cob. Prom. ≥1 Bulto/Mes (s/regalos)",
            "Cob. <1 Bulto (s/regalos)",
            "Cob. <1 Bulto (c/regalos)",
            "AGUAS DANONE ≥3 Bultos",
            "AGUAS DANONE ≥3 Bultos (s/regalos)",
        ]
        for col_idx, expected_header in enumerate(expected, start=1):
            assert ws.cell(row=4, column=col_idx).value == expected_header

    def test_row4_headers_have_fill(self, wb):
        ws = wb["Reporte"]
        for col_idx in range(1, 18):
            cell = ws.cell(row=4, column=col_idx)
            assert "1F4E79" in cell.fill.fgColor.rgb.upper()

    def test_row4_headers_are_bold_white(self, wb):
        ws = wb["Reporte"]
        for col_idx in range(1, 18):
            cell = ws.cell(row=4, column=col_idx)
            assert cell.font.bold is True
            assert "FFFFFF" in cell.font.color.rgb.upper()


# ---------------------------------------------------------------------------
# Reporte sheet — DataValidation
# ---------------------------------------------------------------------------


class TestReporteDataValidation:
    def _get_dv(self, wb) -> DataValidation | None:
        ws = wb["Reporte"]
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list":
                return dv
        return None

    def test_b2_has_list_data_validation(self, wb):
        dv = self._get_dv(wb)
        assert dv is not None, "No list DataValidation found on Reporte sheet"

    def test_dv_formula_references_trimestres_sheet(self, wb):
        dv = self._get_dv(wb)
        assert "_Trimestres!" in dv.formula1

    def test_dv_formula_references_correct_range(self, wb):
        dv = self._get_dv(wb)
        expected_end = f"$A${len(TRIMESTRES)}"
        assert expected_end in dv.formula1

    def test_dv_show_dropdown_is_false(self, wb):
        dv = self._get_dv(wb)
        assert dv.showDropDown is False


# ---------------------------------------------------------------------------
# Reporte sheet — sucursal rows
# ---------------------------------------------------------------------------


class TestReporteSucursalRows:
    def test_a5_is_literal_string_not_formula(self, wb):
        ws = wb["Reporte"]
        val = ws["A5"].value
        assert isinstance(val, str)
        assert not val.startswith("=")

    def test_sucursal_names_in_column_a(self, wb):
        ws = wb["Reporte"]
        names = [ws.cell(row=5 + i, column=1).value for i in range(len(SUCURSALES))]
        assert names == SUCURSALES

    def test_b5_formula_contains_sumproduct(self, wb):
        ws = wb["Reporte"]
        assert "SUMPRODUCT" in ws["B5"].value

    def test_b5_formula_references_tbl_ventas(self, wb):
        ws = wb["Reporte"]
        assert "TblVentasCCU" in ws["B5"].value

    def test_b5_formula_references_trimestre(self, wb):
        ws = wb["Reporte"]
        assert "trimestre" in ws["B5"].value

    def test_c5_formula_is_total_year_prior(self, wb):
        ws = wb["Reporte"]
        # C5 = Total CCU Año Anterior — SUMPRODUCT over TblVentasCCU with year-1
        assert "TblVentasCCU" in ws["C5"].value
        assert "-1" in ws["C5"].value

    def test_d5_formula_references_b_and_c(self, wb):
        ws = wb["Reporte"]
        # D5 = AA vs MMAA = (B-C)/C
        assert "B5" in ws["D5"].value
        assert "C5" in ws["D5"].value

    def test_e5_formula_contains_cervezas(self, wb):
        ws = wb["Reporte"]
        assert "CERVEZAS" in ws["E5"].value

    def test_f5_formula_contains_aguas_danone(self, wb):
        ws = wb["Reporte"]
        assert "AGUAS DANONE" in ws["F5"].value

    def test_g5_formula_contains_vinos_ccu(self, wb):
        ws = wb["Reporte"]
        assert "VINOS CCU" in ws["G5"].value

    def test_g5_formula_contains_sidras_y_licores(self, wb):
        ws = wb["Reporte"]
        assert "SIDRAS Y LICORES" in ws["G5"].value

    def test_h5_formula_references_tbl_cobertura(self, wb):
        ws = wb["Reporte"]
        assert "TblCoberturaCCU" in ws["H5"].value

    def test_h5_formula_threshold_gt_zero(self, wb):
        ws = wb["Reporte"]
        assert ">0" in ws["H5"].value

    def test_i5_formula_threshold_ge_three(self, wb):
        ws = wb["Reporte"]
        assert ">=3" in ws["I5"].value

    def test_j5_formula_uses_average_division(self, wb):
        ws = wb["Reporte"]
        # J5 = clientes con promedio mensual ≥ 1 → bultos/3 >= 1
        assert "/3" in ws["J5"].value
        assert ">=1" in ws["J5"].value
        assert "TblCoberturaCCU" in ws["J5"].value

    def test_j5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["J5"].number_format == "#,##0"

    def test_b5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["B5"].number_format == "#,##0"

    def test_c5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["C5"].number_format == "#,##0"

    def test_d5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["D5"].number_format == "0.0%"

    def test_e5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["E5"].number_format == "0.0%"

    def test_f5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["F5"].number_format == "0.0%"

    def test_g5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["G5"].number_format == "0.0%"

    def test_h5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["H5"].number_format == "#,##0"

    def test_i5_number_format(self, wb):
        ws = wb["Reporte"]
        assert ws["I5"].number_format == "#,##0"

    def test_row_count_matches_sucursales(self, wb):
        ws = wb["Reporte"]
        data_rows = [
            ws.cell(row=5 + i, column=1).value for i in range(len(SUCURSALES))
        ]
        assert all(v is not None for v in data_rows)


# ---------------------------------------------------------------------------
# VentasCCU and CoberturaCCU sheets — table presence
# ---------------------------------------------------------------------------


class TestRawDataSheets:
    def test_ventas_ccu_has_table(self, wb):
        ws = wb["VentasCCU"]
        assert len(ws.tables) > 0
        assert "TblVentasCCU" in ws.tables

    def test_cobertura_has_table(self, wb):
        ws = wb["CoberturaCCU"]
        assert len(ws.tables) > 0
        assert "TblCoberturaCCU" in ws.tables

    def test_ventas_ccu_headers(self, wb):
        ws = wb["VentasCCU"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(5)]
        assert headers == ["sucursal", "generico", "anio", "trimestre", "bultos"]

    def test_cobertura_headers(self, wb):
        ws = wb["CoberturaCCU"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(5)]
        assert headers == ["sucursal", "anio", "trimestre", "id_cliente", "bultos"]

    def test_ventas_row_count(self, wb):
        ws = wb["VentasCCU"]
        data_count = sum(
            1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=1)
            if row[0].value is not None
        )
        assert data_count == 3

    def test_cobertura_row_count(self, wb):
        ws = wb["CoberturaCCU"]
        data_count = sum(
            1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=1)
            if row[0].value is not None
        )
        assert data_count == 2


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    def test_empty_sucursales_produces_no_data_rows(self):
        empty_cob = pd.DataFrame(columns=[
            "sucursal", "anio", "trimestre", "id_cliente",
            "bultos", "bultos_sin_regalos",
            "bultos_aguas_danone", "bultos_aguas_danone_sin_regalos",
            "meses_con_compra",
        ])
        wb = build_workbook(
            sucursales=[],
            df_ventas=pd.DataFrame(columns=["sucursal", "generico", "anio", "trimestre", "bultos"]),
            df_cob=empty_cob,
            trimestres=["2026-Q2"],
        )
        ws = wb["Reporte"]
        assert ws["A5"].value is None

    def test_empty_df_ventas_still_creates_table(self):
        empty_cob = pd.DataFrame(columns=[
            "sucursal", "anio", "trimestre", "id_cliente",
            "bultos", "bultos_sin_regalos",
            "bultos_aguas_danone", "bultos_aguas_danone_sin_regalos",
            "meses_con_compra",
        ])
        wb = build_workbook(
            sucursales=["CASA CENTRAL"],
            df_ventas=pd.DataFrame(columns=["sucursal", "generico", "anio", "trimestre", "bultos"]),
            df_cob=empty_cob,
            trimestres=["2026-Q2"],
        )
        ws = wb["VentasCCU"]
        assert "TblVentasCCU" in ws.tables
