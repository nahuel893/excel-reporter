"""
Tests for ResumenMensualService merge mode (T-07, T-08, T-09).

Verifies:
- T-07: ResumenMensualConfig accepts detalle_movimientos_path field
- T-08: Service uses existing xlsx as merge target; creates fresh on first run;
        raises on multiple xlsx; preserves user sheets
- T-09: Service imports Detalle Movimientos sheet from source path
"""
import logging
import pytest
import pandas as pd
from pathlib import Path
from unittest.mock import Mock, patch
from openpyxl import Workbook, load_workbook

from src.core.data_loader import DataLoader
from src.services.resumen_mensual import (
    ResumenMensualConfig,
    ResumenMensualResult,
    ResumenMensualService,
)


# ---------------------------------------------------------------------------
# Shared test helpers
# ---------------------------------------------------------------------------

def _df_ventas_mes(sucursales=None, genericos=None):
    sucursales = sucursales or ["SUC1"]
    genericos = genericos or ["CERVEZAS"]
    return pd.DataFrame({
        "sucursal": sucursales,
        "generico": genericos,
        "id_ruta": [1] * len(sucursales),
        "cantidad": [100] * len(sucursales),
    })


def _df_dias():
    return pd.DataFrame({
        "sucursal": ["SUC1"],
        "generico": ["CERVEZAS"],
        "fecha": pd.to_datetime(["2026-04-28"]),
        "id_ruta": [1],
        "cantidad": [10],
    })


def _df_vazio():
    return pd.DataFrame(columns=["sucursal", "generico", "cantidad"])


def _mock_loader(genericos=None):
    """Create a minimal mock DataLoader that returns enough data to run generar_reporte."""
    loader = Mock(spec=DataLoader)
    loader.get_ventas_resumen_mensual.return_value = _df_ventas_mes(
        genericos=genericos or ["CERVEZAS"]
    )
    loader.get_ventas_ultimos_dias_habiles.return_value = _df_dias()
    loader.get_ventas_mes_anterior.return_value = _df_vazio()
    loader.get_ventas_mismo_mes_anio_anterior.return_value = _df_vazio()
    loader.get_cupos_resumen_mensual.return_value = pd.DataFrame(
        columns=["sucursal", "generico", "cupo"]
    )
    return loader


def _make_xlsx(path: Path, sheets: dict[str, list]) -> Path:
    """Create xlsx at path with given {sheet_name: [[row], ...]} data."""
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# T-07: ResumenMensualConfig new field
# ---------------------------------------------------------------------------

class TestResumenMensualConfigField:
    """T-07: ResumenMensualConfig.detalle_movimientos_path field."""

    def test_config_accepts_detalle_movimientos_path(self):
        """T-07: Config created with detalle_movimientos_path → field accessible."""
        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            detalle_movimientos_path="/x.xlsx",
        )
        assert config.detalle_movimientos_path == "/x.xlsx"

    def test_config_detalle_movimientos_path_defaults_none(self):
        """T-07: Config without detalle_movimientos_path → defaults to None."""
        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
        )
        assert config.detalle_movimientos_path is None


# ---------------------------------------------------------------------------
# T-08: Service merge target discovery
# ---------------------------------------------------------------------------

class TestResumenMensualMergeTarget:
    """T-08: Service finds existing xlsx and uses as merge target."""

    def test_first_run_no_existing_file_creates_new(self, tmp_path):
        """T-08: Empty output folder → new file created with generico sheets."""
        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        # Override _output_dir to use tmp_path
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
        )
        result = service.generar_reporte(config)

        # A file should have been created
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1
        wb = load_workbook(str(xlsx_files[0]))
        assert "CERVEZAS" in wb.sheetnames

    def test_rerun_with_existing_file_replaces_generico_sheets(self, tmp_path):
        """T-08: Existing xlsx in folder → generico sheet replaced with new data."""
        # Pre-create an xlsx with old CERVEZAS data
        existing = _make_xlsx(
            tmp_path / "Resumen - 01-04-2026.xlsx",
            {"CERVEZAS": [["OldCol"], ["old_value"]]},
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
        )
        service.generar_reporte(config)

        # Same file path should still exist (merge mode)
        assert existing.exists()
        wb = load_workbook(str(existing))
        ws = wb["CERVEZAS"]
        # Should NOT contain the old header
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "OldCol" not in header_row

    def test_rerun_preserves_user_sheets(self, tmp_path):
        """T-08: Re-run preserves sheets not managed by the service."""
        existing = _make_xlsx(
            tmp_path / "Resumen - 01-04-2026.xlsx",
            {
                "CERVEZAS": [["Col"], ["val"]],
                "MiAnalisis": [["formula_header"], ["=A1+1"]],
            },
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
        )
        service.generar_reporte(config)

        wb = load_workbook(str(existing))
        assert "MiAnalisis" in wb.sheetnames
        ws_user = wb["MiAnalisis"]
        assert ws_user.cell(1, 1).value == "formula_header"

    def test_multiple_xlsx_in_folder_raises(self, tmp_path):
        """T-08: Two xlsx in folder → RuntimeError with both filenames in message."""
        _make_xlsx(tmp_path / "File1.xlsx", {"Sheet1": [["a"]]})
        _make_xlsx(tmp_path / "File2.xlsx", {"Sheet2": [["b"]]})

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
        )
        with pytest.raises(RuntimeError, match="File1.xlsx|File2.xlsx"):
            service.generar_reporte(config)

    def test_existing_file_path_preserved(self, tmp_path):
        """T-08: save() returns the existing file path (not a new one)."""
        existing = _make_xlsx(
            tmp_path / "Custom Name.xlsx",
            {"CERVEZAS": [["Col"], ["val"]]},
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
        )
        result = service.generar_reporte(config)

        assert result.ruta_archivo == existing


# ---------------------------------------------------------------------------
# T-09: Service imports Detalle Movimientos sheet
# ---------------------------------------------------------------------------

class TestResumenMensualDetalleMovimientos:
    """T-09: Service imports Detalle Movimientos from detalle_movimientos_path."""

    def _make_source(self, tmp_path: Path) -> Path:
        """Create a minimal detalle_movimientos.xlsx source."""
        src = tmp_path / "detalle_movimientos.xlsx"
        _make_xlsx(
            src,
            {
                "Datos": [
                    ["Artículo", "Fecha Mvto", "Fecha Stock", "Bultos"],
                    ["BEER", pd.Timestamp("2026-04-10").to_pydatetime(), pd.Timestamp("2026-04-10").to_pydatetime(), 500],
                ]
            },
        )
        return src

    def _make_source_with_dates(self, path: Path, fechas: list[str]) -> Path:
        rows = [["Artículo", "Fecha Mvto", "Fecha Stock", "Bultos"]]
        for i, fecha in enumerate(fechas, start=1):
            dt = pd.Timestamp(fecha).to_pydatetime()
            rows.append([f"SKU-{i}", dt, dt, 10])
        return _make_xlsx(path, {"Datos": rows})

    def _make_source_with_dates_split(self, path: Path, fechas_mvto: list[str], fechas_stock: list[str]) -> Path:
        rows = [["Artículo", "Fecha Mvto", "Fecha Stock", "Bultos"]]
        for i, (fecha_mvto, fecha_stock) in enumerate(zip(fechas_mvto, fechas_stock), start=1):
            dt_mvto = pd.Timestamp(fecha_mvto).to_pydatetime()
            dt_stock = pd.Timestamp(fecha_stock).to_pydatetime()
            rows.append([f"SKU-{i}", dt_mvto, dt_stock, 10])
        return _make_xlsx(path, {"Datos": rows})

    def test_import_detalle_movimientos_when_path_set_and_exists(self, tmp_path):
        """T-09: detalle_movimientos_path set + file exists → 'Detalle Movimientos' sheet created."""
        src = self._make_source(tmp_path)

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(src),
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" in wb.sheetnames
        ws = wb["Detalle Movimientos"]
        assert ws.cell(1, 1).value == "Artículo"
        assert ws.cell(2, 2).value == pd.Timestamp("2026-04-10").to_pydatetime()
        assert ws.cell(2, 4).value == 500

    def test_skip_detalle_movimientos_when_path_none(self, tmp_path):
        """T-09: detalle_movimientos_path=None → no 'Detalle Movimientos' sheet created."""
        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=None,
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" not in wb.sheetnames

    def test_skip_detalle_movimientos_when_source_missing(self, tmp_path, caplog):
        """T-09: source file not found → no error, WARNING logged, no sheet created."""
        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(tmp_path / "nonexistent.xlsx"),
        )
        with caplog.at_level(logging.WARNING):
            result = service.generar_reporte(config)  # must NOT raise

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" not in wb.sheetnames
        assert any("nonexistent" in msg.lower() or "not found" in msg.lower() or "skipping" in msg.lower()
                   for msg in caplog.messages)

    def test_existing_detalle_movimientos_preserved_when_source_missing(self, tmp_path):
        """T-09: pre-existing 'Detalle Movimientos' sheet preserved when source file is missing."""
        # Pre-create xlsx with a Detalle Movimientos sheet
        existing = _make_xlsx(
            tmp_path / "Resumen - 01-04-2026.xlsx",
            {
                "CERVEZAS": [["Col"], ["v"]],
                "Detalle Movimientos": [["ArtiHeader"], ["preserved_row"]],
            },
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-04-01",
            fecha_hasta="2026-04-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(tmp_path / "nonexistent_source.xlsx"),
        )
        service.generar_reporte(config)  # must NOT raise

        wb = load_workbook(str(existing))
        # The old sheet should still be there, untouched
        assert "Detalle Movimientos" in wb.sheetnames
        ws = wb["Detalle Movimientos"]
        assert ws.cell(1, 1).value == "ArtiHeader"
        assert ws.cell(2, 1).value == "preserved_row"

    def test_import_detalle_movimientos_accepts_only_expected_periods(self, tmp_path):
        """Actual, MA y MMAA importan cuando cada archivo contiene solo su mes esperado."""
        src_dir = tmp_path / "sources"
        src_dir.mkdir()
        src_actual = self._make_source_with_dates(
            src_dir / "detalle_junio.xlsx", ["2026-06-04", "2026-06-18"]
        )
        src_ma = self._make_source_with_dates(
            src_dir / "detalle_mayo.xlsx", ["2026-05-02", "2026-05-30"]
        )
        src_mmaa = self._make_source_with_dates(
            src_dir / "detalle_junio_2025.xlsx", ["2025-06-01", "2025-06-29"]
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(src_actual),
            detalle_movimientos_ma_path=str(src_ma),
            detalle_movimientos_mmaa_path=str(src_mmaa),
        )
        result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" in wb.sheetnames
        assert "Detalle Movimientos MA" in wb.sheetnames
        assert "Detalle Movimientos MMAA" in wb.sheetnames

    def test_import_detalle_movimientos_when_actual_contains_mixed_months(self, tmp_path, caplog):
        src_actual = self._make_source_with_dates(
            tmp_path / "detalle_wrong_actual.xlsx", ["2026-05-31", "2026-06-01"]
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(src_actual),
        )

        with caplog.at_level(logging.WARNING):
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" in wb.sheetnames
        assert any("expected only 2026-06" in msg for msg in caplog.messages)

    def test_import_detalle_movimientos_ma_when_contains_mixed_months(self, tmp_path, caplog):
        src_ma = self._make_source_with_dates(
            tmp_path / "detalle_wrong_ma.xlsx", ["2026-04-30", "2026-05-01"]
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_ma_path=str(src_ma),
        )

        with caplog.at_level(logging.WARNING):
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos MA" in wb.sheetnames
        assert any("expected only 2026-05" in msg for msg in caplog.messages)

    def test_import_detalle_movimientos_mmaa_when_contains_mixed_months(self, tmp_path, caplog):
        src_mmaa = self._make_source_with_dates(
            tmp_path / "detalle_wrong_mmaa.xlsx", ["2025-05-31", "2025-06-01"]
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_mmaa_path=str(src_mmaa),
        )

        with caplog.at_level(logging.WARNING):
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos MMAA" in wb.sheetnames
        assert any("expected only 2025-06" in msg for msg in caplog.messages)

    def test_validate_detalle_uses_fecha_mvto_only(self, tmp_path):
        """Fecha Stock fuera de mes no invalida si Fecha Mvto está en el mes correcto."""
        src_actual = self._make_source_with_dates_split(
            tmp_path / "detalle_fecha_mvto_ref.xlsx",
            ["2026-06-04", "2026-06-18"],
            ["2026-07-01", "2026-07-02"],
        )

        loader = _mock_loader()
        service = ResumenMensualService(data_loader=loader)
        service._output_dir = lambda date_str: tmp_path

        config = ResumenMensualConfig(
            fecha_desde="2026-06-01",
            fecha_hasta="2026-06-30",
            genericos=["CERVEZAS"],
            detalle_movimientos_path=str(src_actual),
        )

        result = service.generar_reporte(config)
        wb = load_workbook(str(result.ruta_archivo))
        assert "Detalle Movimientos" in wb.sheetnames
