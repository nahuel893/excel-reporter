"""Tests for the FRATELLI B PRVTA exclusion in resumen-mensual.

Spec:
- For genericos in ``genericos_sin_prvta``, queries must EXCLUDE rows whose
  fact_ventas.id_documento = 'PRVTA' (facturas presupuesto).
- Other genericos are unaffected (PRVTA included as before).
- Default for resumen-mensual: ``genericos_sin_prvta=["FRATELLI B"]``.
- The FRATELLI B sheet gets a visible label noting the exclusion.
"""
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import Engine

from src.core.data_loader import DataLoader
from src.services.resumen_mensual.service import (
    ResumenMensualConfig,
    ResumenMensualService,
)


def _loader_with_query_capture():
    """DataLoader whose execute_query records (query, params) per call."""
    loader = DataLoader(engine=MagicMock(spec=Engine))
    captured = []

    def _capture(query: str, params: dict | None = None):
        captured.append((query, params or {}))
        return pd.DataFrame()

    loader.execute_query = Mock(side_effect=_capture)
    loader._captured = captured  # for tests to inspect
    return loader


class TestGetVentasResumenMensualPrvtaFilter:
    """Data-loader level: SQL must conditionally exclude PRVTA documents."""

    def test_no_prvta_filter_when_param_absent(self):
        loader = _loader_with_query_capture()
        loader.get_ventas_resumen_mensual("2026-05-01", "2026-05-31", ["CERVEZAS"])

        sql, _params = loader._captured[-1]
        assert "PRVTA" not in sql.upper()

    def test_prvta_filter_applied_only_to_listed_generics(self):
        loader = _loader_with_query_capture()
        loader.get_ventas_resumen_mensual(
            "2026-05-01", "2026-05-31",
            genericos=["CERVEZAS", "FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
        )

        sql, params = loader._captured[-1]
        # The SQL must reference id_documento and PRVTA
        assert "id_documento" in sql
        assert "PRVTA" in sql
        # FRATELLI B must be among the params for the PRVTA-excluded list
        sin_prvta_values = {v for k, v in params.items() if k.startswith("sp_")}
        assert "FRATELLI B" in sin_prvta_values
        assert "CERVEZAS" not in sin_prvta_values

    def test_prvta_filter_uses_or_disjunction(self):
        """Logic must be: (generic NOT IN sin_prvta_list) OR (id_documento != 'PRVTA').

        That is: rows whose generic is in the list AND id_documento = 'PRVTA' get excluded;
        all others pass through.
        """
        loader = _loader_with_query_capture()
        loader.get_ventas_resumen_mensual(
            "2026-05-01", "2026-05-31",
            genericos=["CERVEZAS", "FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
        )

        sql, _ = loader._captured[-1]
        # Look for the OR disjunction pattern
        sql_upper = sql.upper()
        assert "NOT IN" in sql_upper or "OR" in sql_upper

    def test_ultimos_dias_habiles_supports_genericos_sin_prvta(self):
        loader = _loader_with_query_capture()
        loader.get_ventas_ultimos_dias_habiles(
            "2026-05-01", "2026-05-31",
            genericos=["FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
        )
        sql, _ = loader._captured[-1]
        assert "PRVTA" in sql

    def test_mes_anterior_propagates_genericos_sin_prvta(self):
        loader = _loader_with_query_capture()
        loader.get_ventas_mes_anterior(
            "2026-05-01",
            genericos=["FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
        )
        sql, _ = loader._captured[-1]
        assert "PRVTA" in sql

    def test_mismo_mes_anio_anterior_propagates_genericos_sin_prvta(self):
        loader = _loader_with_query_capture()
        loader.get_ventas_mismo_mes_anio_anterior(
            "2026-05-01", "2026-05-31",
            genericos=["FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
        )
        sql, _ = loader._captured[-1]
        assert "PRVTA" in sql


class TestServicePassesGenericosSinPrvta:
    """Service-level: passes genericos_sin_prvta through to all 4 queries."""

    def _make_loader(self):
        loader = Mock(spec=DataLoader)
        loader.get_ventas_resumen_mensual.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "id_ruta", "cantidad"]
        )
        loader.get_ventas_ultimos_dias_habiles.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "fecha", "id_ruta", "cantidad"]
        )
        loader.get_ventas_mes_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_ventas_mismo_mes_anio_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_cupos_resumen_mensual.side_effect = Exception("no cupos")
        return loader

    def test_default_excludes_prvta_for_fratelli_b(self, tmp_path):
        """When config.genericos_sin_prvta is None, default is ['FRATELLI B']."""
        loader = self._make_loader()
        with patch("src.services.resumen_mensual.service.ExcelWriter") as mock_w, \
             patch("config.settings.DATA_OUTPUT", tmp_path):
            mock_w.return_value.save.return_value = Path("/tmp/x.xlsx")
            service = ResumenMensualService(data_loader=loader)
            config = ResumenMensualConfig(
                fecha_desde="2026-05-01",
                fecha_hasta="2026-05-31",
                genericos=["CERVEZAS", "FRATELLI B"],
            )
            service.generar_reporte(config)

        kwargs = loader.get_ventas_resumen_mensual.call_args.kwargs
        # genericos_sin_prvta must reach the loader
        assert kwargs.get("genericos_sin_prvta") == ["FRATELLI B"]

    def test_explicit_empty_list_disables_filter(self, tmp_path):
        """Setting genericos_sin_prvta=[] explicitly disables the PRVTA filter."""
        loader = self._make_loader()
        with patch("src.services.resumen_mensual.service.ExcelWriter") as mock_w, \
             patch("config.settings.DATA_OUTPUT", tmp_path):
            mock_w.return_value.save.return_value = Path("/tmp/x.xlsx")
            service = ResumenMensualService(data_loader=loader)
            config = ResumenMensualConfig(
                fecha_desde="2026-05-01",
                fecha_hasta="2026-05-31",
                genericos=["FRATELLI B"],
                genericos_sin_prvta=[],
            )
            service.generar_reporte(config)

        kwargs = loader.get_ventas_resumen_mensual.call_args.kwargs
        assert kwargs.get("genericos_sin_prvta") == []

    def test_passed_to_all_four_queries(self, tmp_path):
        loader = self._make_loader()
        with patch("src.services.resumen_mensual.service.ExcelWriter") as mock_w, \
             patch("config.settings.DATA_OUTPUT", tmp_path):
            mock_w.return_value.save.return_value = Path("/tmp/x.xlsx")
            service = ResumenMensualService(data_loader=loader)
            config = ResumenMensualConfig(
                fecha_desde="2026-05-01",
                fecha_hasta="2026-05-31",
                genericos=["FRATELLI B"],
                genericos_sin_prvta=["FRATELLI B"],
            )
            service.generar_reporte(config)

        for method in (
            loader.get_ventas_resumen_mensual,
            loader.get_ventas_ultimos_dias_habiles,
            loader.get_ventas_mes_anterior,
            loader.get_ventas_mismo_mes_anio_anterior,
        ):
            assert method.call_args.kwargs.get("genericos_sin_prvta") == ["FRATELLI B"], (
                f"{method._mock_name} did not receive genericos_sin_prvta"
            )


class TestFratelliBSheetLabel:
    """The FRATELLI B sheet must show a visible label about PRVTA exclusion."""

    def _full_loader(self):
        """Loader returning small but valid data so the sheet actually gets written."""
        loader = Mock(spec=DataLoader)
        loader.get_ventas_resumen_mensual.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL"],
            "generico": ["FRATELLI B"],
            "id_ruta": [1],
            "cantidad": [100.0],
        })
        loader.get_ventas_ultimos_dias_habiles.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL"],
            "generico": ["FRATELLI B"],
            "fecha": pd.to_datetime(["2026-05-31"]),
            "id_ruta": [1],
            "cantidad": [50.0],
        })
        loader.get_ventas_mes_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_ventas_mismo_mes_anio_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_cupos_resumen_mensual.side_effect = Exception("no cupos")
        return loader

    def test_fratelli_b_sheet_has_prvta_label(self, tmp_path):
        loader = self._full_loader()
        service = ResumenMensualService(data_loader=loader)
        config = ResumenMensualConfig(
            fecha_desde="2026-05-01",
            fecha_hasta="2026-05-31",
            genericos=["FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
            nombre_archivo="resumen_test",
        )
        # Patch DATA_OUTPUT to tmp_path so file lands there
        with patch("config.settings.DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        assert "FRATELLI B" in wb.sheetnames
        ws = wb["FRATELLI B"]
        # Search first 5 rows for a cell mentioning PRVTA
        found = False
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell_value in row:
                if cell_value and "PRVTA" in str(cell_value).upper():
                    found = True
                    break
            if found:
                break
        assert found, "FRATELLI B sheet must include a label mentioning PRVTA"

    def test_other_sheets_have_no_prvta_label(self, tmp_path):
        """Only FRATELLI B gets the label; other genericos must not have it."""
        loader = Mock(spec=DataLoader)
        loader.get_ventas_resumen_mensual.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "generico": ["CERVEZAS", "FRATELLI B"],
            "id_ruta": [1, 1],
            "cantidad": [200.0, 50.0],
        })
        loader.get_ventas_ultimos_dias_habiles.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "CASA CENTRAL"],
            "generico": ["CERVEZAS", "FRATELLI B"],
            "fecha": pd.to_datetime(["2026-05-31", "2026-05-31"]),
            "id_ruta": [1, 1],
            "cantidad": [100.0, 25.0],
        })
        loader.get_ventas_mes_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_ventas_mismo_mes_anio_anterior.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "cantidad"]
        )
        loader.get_cupos_resumen_mensual.side_effect = Exception("no cupos")

        service = ResumenMensualService(data_loader=loader)
        config = ResumenMensualConfig(
            fecha_desde="2026-05-01",
            fecha_hasta="2026-05-31",
            genericos=["CERVEZAS", "FRATELLI B"],
            genericos_sin_prvta=["FRATELLI B"],
            nombre_archivo="resumen_multi",
        )
        with patch("config.settings.DATA_OUTPUT", tmp_path):
            result = service.generar_reporte(config)

        wb = load_workbook(str(result.ruta_archivo))
        ws_cerv = wb["CERVEZAS"]
        for row in ws_cerv.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell_value in row:
                if cell_value and "PRVTA" in str(cell_value).upper():
                    pytest.fail("CERVEZAS sheet must NOT include PRVTA label")
