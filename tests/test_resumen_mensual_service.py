"""
Integration tests for ResumenMensualService — Phase 3 (T-022).

Covers:
- T-019: cupos fetched from DataLoader and passed to processor
- T-020: subtotal rows injected; worksheet has SUM formulas at correct positions
- T-021: font_color applied to MMAA, MA, Objetivo columns
- T-022: as_table=False; ColorScaleRule applied to Tend vs Obj % column
- T-023: subtotal rows have bold font + fill

These tests follow Strict TDD: written BEFORE the implementation (Phase 3).
"""
import io
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock, call

import pandas as pd
import pytest
from openpyxl import load_workbook
# Note: ColorScaleRule is a factory function in openpyxl, not a class.
# We check rule.type == 'colorScale' instead of isinstance(rule, ColorScaleRule).

from src.services.resumen_mensual import (
    ResumenMensualConfig,
    ResumenMensualService,
)
from src.core.data_loader import DataLoader


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_GENERICOS_DEFAULT = ["CERVEZAS", "AGUAS DANONE", "VINOS CCU", "SIDRAS Y LICORES"]

FECHA_DESDE = "2026-04-01"
FECHA_HASTA = "2026-04-25"


def _df_ventas_mes(sucursal="CASA CENTRAL", generico="CERVEZAS", cantidad=1000, id_ruta=1):
    return pd.DataFrame({
        "sucursal": [sucursal],
        "generico": [generico],
        "id_ruta": [id_ruta],
        "cantidad": [cantidad],
    })


def _df_dias(sucursal="CASA CENTRAL", generico="CERVEZAS", cantidad=50, id_ruta=1):
    return pd.DataFrame({
        "sucursal": [sucursal],
        "generico": [generico],
        "fecha": pd.to_datetime(["2026-04-24"]),
        "id_ruta": [id_ruta],
        "cantidad": [cantidad],
    })


def _df_vacio():
    return pd.DataFrame(columns=["sucursal", "generico", "cantidad"])


def _df_cupos(sucursal="CASA CENTRAL", generico="CERVEZAS", cupo=2000):
    return pd.DataFrame({
        "sucursal": [sucursal],
        "generico": [generico],
        "id_ruta": [1],
        "cupo": [cupo],
    })


def _make_loader(cupos_df=None, cupos_raises=False):
    """Create a mocked DataLoader with sane defaults."""
    loader = Mock(spec=DataLoader)
    loader.get_ventas_resumen_mensual.return_value = _df_ventas_mes()
    loader.get_ventas_ultimos_dias_habiles.return_value = _df_dias()
    loader.get_ventas_mes_anterior.return_value = _df_vacio()
    loader.get_ventas_mismo_mes_anio_anterior.return_value = _df_vacio()
    if cupos_raises:
        loader.get_cupos_resumen_mensual.side_effect = Exception("DB error")
    else:
        loader.get_cupos_resumen_mensual.return_value = (
            cupos_df if cupos_df is not None else _df_cupos()
        )
    return loader


def _run_service(loader, config=None, tmp_path=None):
    """Run service with a temp output dir, return the path."""
    if config is None:
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
    if tmp_path is None:
        import tempfile
        tmp_path = Path(tempfile.mkdtemp())

    service = ResumenMensualService(data_loader=loader)
    with patch.object(service, "_output_dir", return_value=tmp_path):
        result = service.generar_reporte(config)
    return result


# ===========================================================================
# T-019: Cupos fetch wired
# ===========================================================================

class TestCuposFetch:
    """T-019: cupos are fetched from DataLoader and passed to processor."""

    def test_cupos_fetched_with_correct_periodo(self, tmp_path):
        """get_cupos_resumen_mensual is called with periodo='2026-04' derived from fecha_desde."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        _run_service(loader, config, tmp_path)
        loader.get_cupos_resumen_mensual.assert_called_once()
        args, kwargs = loader.get_cupos_resumen_mensual.call_args
        periodo_arg = args[0] if args else kwargs.get("periodo")
        assert periodo_arg == "2026-04"

    def test_cupos_fetched_with_default_genericos_when_none(self, tmp_path):
        """When config.genericos is None, cupos fetch uses the default CCU genericos."""
        from src.services.resumen_mensual.service import _DEFAULT_GENERICOS

        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
            genericos=None,
        )
        _run_service(loader, config, tmp_path)
        loader.get_cupos_resumen_mensual.assert_called_once()
        args, kwargs = loader.get_cupos_resumen_mensual.call_args
        genericos_arg = args[1] if len(args) > 1 else kwargs.get("genericos")
        assert isinstance(genericos_arg, list)
        assert genericos_arg == _DEFAULT_GENERICOS
        assert "PERNOD RICARD" in genericos_arg

    def test_cupos_fetched_with_explicit_genericos(self, tmp_path):
        """When config.genericos is provided, cupos fetch uses that list."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
            genericos=["CERVEZAS"],
        )
        _run_service(loader, config, tmp_path)
        loader.get_cupos_resumen_mensual.assert_called_once()
        args, kwargs = loader.get_cupos_resumen_mensual.call_args
        genericos_arg = args[1] if len(args) > 1 else kwargs.get("genericos")
        assert genericos_arg == ["CERVEZAS"]

    def test_objetivo_populated_when_cupos_available(self, tmp_path):
        """When cupos data exists, the Objetivo column is populated in the output xlsx."""
        loader = _make_loader(cupos_df=_df_cupos(cupo=2000))
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        # Find the Objetivo column index
        header_row = None
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if "Objetivo" in values:
                header_row = row
                break
        assert header_row is not None, "Header row with 'Objetivo' not found"
        obj_col_idx = [c.value for c in header_row].index("Objetivo") + 1
        # Find a data row with CASA CENTRAL / CERVEZAS
        data_rows = []
        for row in ws.iter_rows(min_row=header_row[0].row + 1):
            vals = [c.value for c in row]
            if vals and vals[0] not in (None, "", "Subtotales"):
                data_rows.append(row)
        assert len(data_rows) > 0, "No data rows found"
        # The Objetivo cell for our data row should be non-null (2000)
        obj_val = data_rows[0][obj_col_idx - 1].value
        assert obj_val is not None, f"Objetivo should be populated but got None"
        assert obj_val == 2000.0 or obj_val == 2000

    def test_cupos_db_error_yields_blank_objetivo(self, tmp_path):
        """If cupos fetch raises an exception, the report is generated with blank Objetivo."""
        loader = _make_loader(cupos_raises=True)
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        # Must not raise
        result = _run_service(loader, config, tmp_path)
        assert result.ruta_archivo.exists()
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        # Find Objetivo column
        header_row = None
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if "Objetivo" in values:
                header_row = row
                break
        assert header_row is not None
        obj_col_idx = [c.value for c in header_row].index("Objetivo") + 1
        # All data rows should have None/null in Objetivo
        for row in ws.iter_rows(min_row=header_row[0].row + 1):
            vals = [c.value for c in row]
            if vals and vals[0] not in (None, ""):
                obj_val = row[obj_col_idx - 1].value
                # Should be None or a formula referencing blank — not a number
                assert obj_val is None or (isinstance(obj_val, str) and obj_val.startswith("=")), \
                    f"Expected None or formula but got {obj_val!r}"


# ===========================================================================
# T-020: Subtotal rows (SUM formulas)
# ===========================================================================

class TestSubtotalRows:
    """T-020: Subtotal rows with SUM formulas are injected into the sheet."""

    def _setup_loader_multi_suc(self):
        """Setup a loader with CASA CENTRAL + SUCURSAL CAFAYATE data."""
        loader = Mock(spec=DataLoader)
        loader.get_ventas_resumen_mensual.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "SUCURSAL CAFAYATE"],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "id_ruta": [1, 1],
            "cantidad": [500, 300],
        })
        loader.get_ventas_ultimos_dias_habiles.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "SUCURSAL CAFAYATE"],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "fecha": pd.to_datetime(["2026-04-24", "2026-04-24"]),
            "id_ruta": [1, 1],
            "cantidad": [50, 30],
        })
        loader.get_ventas_mes_anterior.return_value = _df_vacio()
        loader.get_ventas_mismo_mes_anio_anterior.return_value = _df_vacio()
        loader.get_cupos_resumen_mensual.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "id_ruta", "cupo"]
        )
        return loader

    def test_subtotal_casa_central_row_present(self, tmp_path):
        """SUBTOTAL CASA CENTRAL row appears in the sheet."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        sucursal_values = [row[0].value for row in ws.iter_rows() if row[0].value]
        assert "SUBTOTAL CASA CENTRAL" in sucursal_values, \
            f"SUBTOTAL CASA CENTRAL not found in col A. Values: {sucursal_values}"

    def test_sucursales_sin_directa_row_present(self, tmp_path):
        """SUCURSALES SIN DIRECTA row appears in the sheet."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        sucursal_values = [row[0].value for row in ws.iter_rows() if row[0].value]
        assert "SUCURSALES SIN DIRECTA" in sucursal_values, \
            f"SUCURSALES SIN DIRECTA not found in col A. Values: {sucursal_values}"

    def test_total_sin_smk_row_present(self, tmp_path):
        """TOTAL SIN SMK row appears in the sheet."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        sucursal_values = [row[0].value for row in ws.iter_rows() if row[0].value]
        assert "TOTAL SIN SMK" in sucursal_values, \
            f"TOTAL SIN SMK not found in col A. Values: {sucursal_values}"

    def test_subtotal_rows_contain_sum_formulas(self, tmp_path):
        """Subtotal rows contain =SUM(...) formulas in at least one numeric column."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo, data_only=False)
        ws = wb.active
        subtotal_labels = {"SUBTOTAL CASA CENTRAL", "SUCURSALES SIN DIRECTA", "TOTAL SIN SMK"}
        found_sum_formula = False
        for row in ws.iter_rows():
            if row[0].value in subtotal_labels:
                for cell in row[1:]:  # skip col A (the label)
                    if isinstance(cell.value, str) and cell.value.upper().startswith("=SUM("):
                        found_sum_formula = True
                        break
            if found_sum_formula:
                break
        assert found_sum_formula, "No =SUM(...) formula found in any subtotal row"


# ===========================================================================
# T-021: font_color for MMAA / MA / Objetivo
# ===========================================================================

class TestFontColor:
    """T-021: MMAA, MA, Objetivo data cells have correct font colors."""

    def _get_col_letter(self, ws, col_name):
        """Find the column letter by header name."""
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if col_name in values:
                idx = values.index(col_name)
                from openpyxl.utils import get_column_letter
                return get_column_letter(idx + 1), row[0].row
        return None, None

    def test_mmaa_data_cells_have_red_font(self, tmp_path):
        """MMAA column data cells have a red font color."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        col_letter, header_row_num = self._get_col_letter(ws, "MMAA")
        assert col_letter is not None, "MMAA column not found"
        # Check first data cell (row after header)
        data_cell = ws[f"{col_letter}{header_row_num + 1}"]
        color = data_cell.font.color
        assert color is not None
        # Red family: "C00000" (dark red) stored as ARGB: "FFC00000" or "00C00000"
        color_rgb = color.rgb if color.type == "rgb" else ""
        assert color_rgb.upper().endswith("C00000"), \
            f"Expected red font on MMAA data cell, got {color_rgb!r}"

    def test_ma_data_cells_have_olive_font(self, tmp_path):
        """MA column data cells have an olive font color."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        col_letter, header_row_num = self._get_col_letter(ws, "MA")
        assert col_letter is not None, "MA column not found"
        data_cell = ws[f"{col_letter}{header_row_num + 1}"]
        color = data_cell.font.color
        assert color is not None
        # Olive: "808000"
        color_rgb = color.rgb if color.type == "rgb" else ""
        assert color_rgb.upper().endswith("808000"), \
            f"Expected olive font on MA data cell, got {color_rgb!r}"

    def test_objetivo_data_cells_have_blue_font(self, tmp_path):
        """Objetivo column data cells have a light blue font color."""
        loader = _make_loader(cupos_df=_df_cupos(cupo=2000))
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        col_letter, header_row_num = self._get_col_letter(ws, "Objetivo")
        assert col_letter is not None, "Objetivo column not found"
        data_cell = ws[f"{col_letter}{header_row_num + 1}"]
        color = data_cell.font.color
        assert color is not None
        # Light blue: "4472C4"
        color_rgb = color.rgb if color.type == "rgb" else ""
        assert color_rgb.upper().endswith("4472C4"), \
            f"Expected blue font on Objetivo data cell, got {color_rgb!r}"

    def test_mmaa_header_does_not_have_red_font(self, tmp_path):
        """MMAA column header cell does NOT have the red font (headers use white on dark background)."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        col_letter, header_row_num = self._get_col_letter(ws, "MMAA")
        assert col_letter is not None
        header_cell = ws[f"{col_letter}{header_row_num}"]
        color = header_cell.font.color
        # Header should be white (FFFFFFFF or 00FFFFFF)
        color_rgb = color.rgb if color is not None and color.type == "rgb" else ""
        assert color_rgb.upper().endswith("FFFFFF"), \
            f"Header should be white, got {color_rgb!r}"


# ===========================================================================
# T-022: Heatmap on Tend vs Obj % + as_table=False
# ===========================================================================

class TestHeatmapAndTableMode:
    """T-022: ColorScaleRule applied to Tend vs Obj % column; as_table=False."""

    def test_as_table_false_no_tables_in_sheet(self, tmp_path):
        """With as_table=False, the worksheet has no Excel tables."""
        loader = _make_loader()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        assert len(ws.tables) == 0, \
            f"Expected no tables (as_table=False) but found: {list(ws.tables.keys())}"

    def test_color_scale_rule_applied_to_tend_vs_obj(self, tmp_path):
        """A ColorScaleRule (type='colorScale') is applied to the Tend vs Obj % column range."""
        loader = _make_loader(cupos_df=_df_cupos(cupo=2000))
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        # In openpyxl, ColorScaleRule is a factory — actual rules are Rule objects
        # with rule.type == 'colorScale'. _cf_rules keys are ConditionalFormatting objects.
        has_color_scale = False
        for cf_obj, cf_list in ws.conditional_formatting._cf_rules.items():
            for rule in cf_list:
                if getattr(rule, "type", None) == "colorScale":
                    has_color_scale = True
                    break
        assert has_color_scale, \
            "Expected a colorScale rule in conditional_formatting but none found"

    def test_color_scale_range_covers_tend_vs_obj_column(self, tmp_path):
        """The colorScale rule range covers the Tend vs Obj % column."""
        loader = _make_loader(cupos_df=_df_cupos(cupo=2000))
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        # Find the Tend vs Obj % column letter
        tend_col_letter = None
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if "Tend vs Obj (%)" in values:
                from openpyxl.utils import get_column_letter
                idx = values.index("Tend vs Obj (%)")
                tend_col_letter = get_column_letter(idx + 1)
                break
        assert tend_col_letter is not None, "Tend vs Obj (%) column not found"
        # Verify at least one CF range uses that column letter
        # _cf_rules keys are ConditionalFormatting objects — str() gives e.g. "J5:J6"
        found = False
        for cf_obj, cf_list in ws.conditional_formatting._cf_rules.items():
            cf_range_str = str(cf_obj)
            if tend_col_letter.upper() in cf_range_str.upper():
                for rule in cf_list:
                    if getattr(rule, "type", None) == "colorScale":
                        found = True
                        break
        assert found, \
            f"colorScale rule not found on column {tend_col_letter}. CF ranges: {[str(k) for k in ws.conditional_formatting._cf_rules.keys()]}"


# ===========================================================================
# T-023: Subtotal row visual styling (bold + fill)
# ===========================================================================

class TestSubtotalStyling:
    """T-023: Subtotal rows have bold font and fill."""

    def _setup_loader_multi_suc(self):
        """Setup a loader with CASA CENTRAL + SUCURSAL CAFAYATE data."""
        loader = Mock(spec=DataLoader)
        loader.get_ventas_resumen_mensual.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "SUCURSAL CAFAYATE"],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "id_ruta": [1, 1],
            "cantidad": [500, 300],
        })
        loader.get_ventas_ultimos_dias_habiles.return_value = pd.DataFrame({
            "sucursal": ["CASA CENTRAL", "SUCURSAL CAFAYATE"],
            "generico": ["CERVEZAS", "CERVEZAS"],
            "fecha": pd.to_datetime(["2026-04-24", "2026-04-24"]),
            "id_ruta": [1, 1],
            "cantidad": [50, 30],
        })
        loader.get_ventas_mes_anterior.return_value = _df_vacio()
        loader.get_ventas_mismo_mes_anio_anterior.return_value = _df_vacio()
        loader.get_cupos_resumen_mensual.return_value = pd.DataFrame(
            columns=["sucursal", "generico", "id_ruta", "cupo"]
        )
        return loader

    def _find_subtotal_rows(self, ws):
        """Find rows with subtotal labels, return list of row tuples."""
        subtotal_labels = {"SUBTOTAL CASA CENTRAL", "SUCURSALES SIN DIRECTA", "TOTAL SIN SMK"}
        result = []
        for row in ws.iter_rows():
            if row[0].value in subtotal_labels:
                result.append(row)
        return result

    def test_subtotal_rows_have_bold_font(self, tmp_path):
        """Each subtotal row's col A cell is bold."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        subtotal_rows = self._find_subtotal_rows(ws)
        assert len(subtotal_rows) >= 1, "No subtotal rows found"
        for row in subtotal_rows:
            label = row[0].value
            assert row[0].font.bold, f"Expected bold on col A of row '{label}'"

    def test_subtotal_rows_have_fill(self, tmp_path):
        """Each subtotal row's col A cell has a non-default fill."""
        loader = self._setup_loader_multi_suc()
        config = ResumenMensualConfig(
            fecha_desde=FECHA_DESDE,
            fecha_hasta=FECHA_HASTA,
            con_objetivo=True,
        )
        result = _run_service(loader, config, tmp_path)
        wb = load_workbook(result.ruta_archivo)
        ws = wb.active
        subtotal_rows = self._find_subtotal_rows(ws)
        assert len(subtotal_rows) >= 1, "No subtotal rows found"
        for row in subtotal_rows:
            label = row[0].value
            fill = row[0].fill
            assert fill.fill_type not in (None, "none"), \
                f"Expected a fill on subtotal row '{label}' but got fill_type={fill.fill_type!r}"
