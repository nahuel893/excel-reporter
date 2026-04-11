"""
Tests for stock_diario service — strict TDD.

Test IDs:
  TS-001: Single date, valid data → 1 file created
  TS-002: Date range → multiple files
  TS-003: Empty date skipped, in fechas_sin_datos
  TS-004: Banner row structure (BULTOS merged, HTLs merged, correct fills)
  TS-005: Header row content
  TS-007: Sucursal alphabetical ordering
  TS-008: Zero values shown as 0
  TS-009: Frozen panes at D3
  TS-010: pivot_stock correctness
"""

import pandas as pd
import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch, call
from openpyxl import load_workbook

from src.core.data_loader import DataLoader


# ── Fixtures ──────────────────────────────────────────────────────────────────


def _make_stock_df():
    """Minimal valid stock DataFrame with 3 articles × 2 sucursales."""
    return pd.DataFrame({
        "generico": ["CERVEZAS", "CERVEZAS", "AGUAS", "AGUAS", "VINOS", "VINOS"],
        "marca": ["BRAHMA", "BRAHMA", "VILLA DEL SUR", "VILLA DEL SUR", "NAVARRO", "NAVARRO"],
        "des_articulo": ["BRAHMA LATA 473", "BRAHMA LATA 473", "VDS 500ML", "VDS 500ML", "NAV 750ML", "NAV 750ML"],
        "sucursal": ["CAFAYATE", "CASA CENTRAL", "CAFAYATE", "CASA CENTRAL", "CAFAYATE", "CASA CENTRAL"],
        "cant_bultos": [10, 20, 5, 15, 8, 12],
        "cant_htls": [100, 200, 50, 150, 80, 120],
    })


def _make_single_article_df():
    """One article, one sucursal."""
    return pd.DataFrame({
        "generico": ["CERVEZAS"],
        "marca": ["BRAHMA"],
        "des_articulo": ["BRAHMA LATA 473"],
        "sucursal": ["CAFAYATE"],
        "cant_bultos": [10],
        "cant_htls": [100],
    })


# ── TS-010: pivot_stock ───────────────────────────────────────────────────────


class TestPivotStock:
    def test_basic_pivot(self):
        from src.services.stock_diario.processor import pivot_stock

        df = _make_stock_df()
        sucursales = ["CAFAYATE", "CASA CENTRAL"]
        result = pivot_stock(df, "cant_bultos", sucursales)

        assert not result.empty
        assert "generico" in result.columns
        assert "marca" in result.columns
        assert "des_articulo" in result.columns
        assert "CAFAYATE" in result.columns
        assert "CASA CENTRAL" in result.columns

    def test_correct_values(self):
        from src.services.stock_diario.processor import pivot_stock

        df = _make_stock_df()
        sucursales = ["CAFAYATE", "CASA CENTRAL"]
        result = pivot_stock(df, "cant_bultos", sucursales)

        brahma_row = result[result["des_articulo"] == "BRAHMA LATA 473"].iloc[0]
        assert brahma_row["CAFAYATE"] == 10
        assert brahma_row["CASA CENTRAL"] == 20

    def test_missing_sucursal_filled_zero(self):
        from src.services.stock_diario.processor import pivot_stock

        df = _make_single_article_df()
        sucursales = ["CAFAYATE", "SALTA"]
        result = pivot_stock(df, "cant_bultos", sucursales)

        assert "SALTA" in result.columns
        assert result.iloc[0]["SALTA"] == 0

    def test_empty_df_returns_empty(self):
        from src.services.stock_diario.processor import pivot_stock

        df = pd.DataFrame()
        result = pivot_stock(df, "cant_bultos", ["CAFAYATE"])
        assert result.empty

    def test_htls_pivot(self):
        from src.services.stock_diario.processor import pivot_stock

        df = _make_stock_df()
        sucursales = ["CAFAYATE", "CASA CENTRAL"]
        result = pivot_stock(df, "cant_htls", sucursales)

        brahma_row = result[result["des_articulo"] == "BRAHMA LATA 473"].iloc[0]
        assert brahma_row["CAFAYATE"] == 100
        assert brahma_row["CASA CENTRAL"] == 200


# ── TS-001: Single date, valid data → 1 file ──────────────────────────────────


class TestStockDiarioServiceSingleDate:
    def test_generates_one_file(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService
        from src.services.stock_diario.processor import build_excel

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-10", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 1
        assert len(result.fechas_sin_datos) == 0
        mock_loader.get_stock_diario.assert_called_once_with("2026-04-10", None)

    def test_generated_file_exists(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-10", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        for ruta in result.archivos_generados:
            assert Path(ruta).exists()

    def test_file_named_with_date(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-10", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        assert "10-04-2026" in result.archivos_generados[0].name


# ── TS-002: Date range → multiple files ───────────────────────────────────────


class TestStockDiarioServiceDateRange:
    def test_generates_one_file_per_day(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-08", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 3
        assert mock_loader.get_stock_diario.call_count == 3

    def test_correct_dates_queried(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-08", fecha_hasta="2026-04-10")
            service.generar_reporte(config)

        calls = mock_loader.get_stock_diario.call_args_list
        fechas = [c.args[0] for c in calls]
        assert fechas == ["2026-04-08", "2026-04-09", "2026-04-10"]


# ── TS-003: Empty date skipped ────────────────────────────────────────────────


class TestStockDiarioEmptyDate:
    def test_empty_date_in_fechas_sin_datos(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        # Day 1: data, Day 2: empty, Day 3: data
        mock_loader.get_stock_diario.side_effect = [
            _make_stock_df(),
            pd.DataFrame(),
            _make_stock_df(),
        ]

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-08", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 2
        assert len(result.fechas_sin_datos) == 1
        assert "2026-04-09" in result.fechas_sin_datos

    def test_all_empty_produces_no_files(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = pd.DataFrame()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(fecha_desde="2026-04-10", fecha_hasta="2026-04-10")
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 0
        assert "2026-04-10" in result.fechas_sin_datos


# ── TS-004: Banner row structure ──────────────────────────────────────────────


class TestBannerRowStructure:
    def _build(self, tmp_path):
        from src.services.stock_diario.processor import build_excel
        ruta = build_excel("2026-04-10", _make_stock_df(), output_dir=tmp_path)
        return load_workbook(ruta)

    def test_bultos_banner_text(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        # Row 1 should contain "BULTOS" in column 4 (after 3 desc cols)
        bultos_cell = ws.cell(row=1, column=4)
        assert bultos_cell.value == "BULTOS"

    def test_htls_banner_text(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        df = _make_stock_df()
        n_suc = len(df["sucursal"].unique())
        htls_start = 4 + n_suc
        htls_cell = ws.cell(row=1, column=htls_start)
        assert htls_cell.value == "HTLs"

    def test_bultos_banner_fill_color(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        bultos_cell = ws.cell(row=1, column=4)
        # Should have a solid fill (non-empty)
        assert bultos_cell.fill is not None
        assert bultos_cell.fill.fill_type == "solid"
        # openpyxl stores colors without alpha as 00RRGGBB; strip leading 2 chars for check
        assert bultos_cell.fill.fgColor.rgb[-6:] == "4472C4"

    def test_htls_banner_fill_color(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        df = _make_stock_df()
        n_suc = len(df["sucursal"].unique())
        htls_start = 4 + n_suc
        htls_cell = ws.cell(row=1, column=htls_start)
        assert htls_cell.fill is not None
        assert htls_cell.fill.fill_type == "solid"
        # openpyxl stores colors without alpha as 00RRGGBB; strip leading 2 chars for check
        assert htls_cell.fill.fgColor.rgb[-6:] == "70AD47"

    def test_banner_cells_are_bold(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        bultos_cell = ws.cell(row=1, column=4)
        assert bultos_cell.font.bold is True


# ── TS-005: Header row content ────────────────────────────────────────────────


class TestHeaderRowContent:
    def _build(self, tmp_path):
        from src.services.stock_diario.processor import build_excel
        ruta = build_excel("2026-04-10", _make_stock_df(), output_dir=tmp_path)
        return load_workbook(ruta)

    def test_desc_col_headers(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Articulo"
        assert ws.cell(row=2, column=2).value == "Marca"
        assert ws.cell(row=2, column=3).value == "Generico"

    def test_sucursal_headers_appear_in_bultos(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        df = _make_stock_df()
        sucursales = sorted(df["sucursal"].unique().tolist())
        for i, suc in enumerate(sucursales):
            cell = ws.cell(row=2, column=4 + i)
            assert cell.value == suc

    def test_sucursal_headers_appear_in_htls(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb.active
        df = _make_stock_df()
        sucursales = sorted(df["sucursal"].unique().tolist())
        n_suc = len(sucursales)
        htls_start = 4 + n_suc
        for i, suc in enumerate(sucursales):
            cell = ws.cell(row=2, column=htls_start + i)
            assert cell.value == suc


# ── TS-007: Sucursal alphabetical ordering ────────────────────────────────────


class TestSucursalAlphabeticalOrdering:
    def test_headers_alphabetical(self, tmp_path):
        from src.services.stock_diario.processor import build_excel

        df = pd.DataFrame({
            "generico": ["CERVEZAS"] * 3,
            "marca": ["BRAHMA"] * 3,
            "des_articulo": ["ART1"] * 3,
            "sucursal": ["ZARAGOZA", "ALFARO", "METAN"],
            "cant_bultos": [1, 2, 3],
            "cant_htls": [10, 20, 30],
        })
        ruta = build_excel("2026-04-10", df, output_dir=tmp_path)
        wb = load_workbook(ruta)
        ws = wb.active

        assert ws.cell(row=2, column=4).value == "ALFARO"
        assert ws.cell(row=2, column=5).value == "METAN"
        assert ws.cell(row=2, column=6).value == "ZARAGOZA"


# ── TS-008: Zero values shown as 0 ───────────────────────────────────────────


class TestZeroValues:
    def test_missing_combination_shows_zero(self, tmp_path):
        from src.services.stock_diario.processor import build_excel

        # Only CAFAYATE has stock for ART1; SALTA has none
        df = pd.DataFrame({
            "generico": ["CERVEZAS", "CERVEZAS"],
            "marca": ["BRAHMA", "BRAHMA"],
            "des_articulo": ["ART1", "ART1"],
            "sucursal": ["CAFAYATE", "SALTA"],
            "cant_bultos": [10, 0],
            "cant_htls": [100, 0],
        })
        ruta = build_excel("2026-04-10", df, output_dir=tmp_path)
        wb = load_workbook(ruta)
        ws = wb.active

        sucursales = ["CAFAYATE", "SALTA"]
        # Data starts at row 3
        salta_idx = sucursales.index("SALTA")
        cell = ws.cell(row=3, column=4 + salta_idx)
        assert cell.value == 0


# ── TS-009: Frozen panes at D3 ────────────────────────────────────────────────


class TestFrozenPanes:
    def test_frozen_at_d3(self, tmp_path):
        from src.services.stock_diario.processor import build_excel

        ruta = build_excel("2026-04-10", _make_stock_df(), output_dir=tmp_path)
        wb = load_workbook(ruta)
        ws = wb.active
        assert ws.freeze_panes == "D3"


# ── TS-011: Data correctness in cells ─────────────────────────────────────────


class TestDataCellValues:
    def test_article_data_in_rows(self, tmp_path):
        from src.services.stock_diario.processor import build_excel

        df = pd.DataFrame({
            "generico": ["CERVEZAS"],
            "marca": ["BRAHMA"],
            "des_articulo": ["BRAHMA LATA 473"],
            "sucursal": ["CAFAYATE"],
            "cant_bultos": [42],
            "cant_htls": [420],
        })
        ruta = build_excel("2026-04-10", df, output_dir=tmp_path)
        wb = load_workbook(ruta)
        ws = wb.active

        # Row 3 = first data row (order: Articulo, Marca, Generico)
        assert ws.cell(row=3, column=1).value == "BRAHMA LATA 473"
        assert ws.cell(row=3, column=2).value == "BRAHMA"
        assert ws.cell(row=3, column=3).value == "CERVEZAS"
        # Bultos value (col 4 = first sucursal)
        assert ws.cell(row=3, column=4).value == 42
        # HTLs value (col 4 + 1 suc = col 5)
        assert ws.cell(row=3, column=5).value == 420

    def test_multiple_articles_in_correct_rows(self, tmp_path):
        from src.services.stock_diario.processor import build_excel

        df = pd.DataFrame({
            "generico": ["AGUAS", "CERVEZAS"],
            "marca": ["VDS", "BRAHMA"],
            "des_articulo": ["VDS 500ML", "BRAHMA LATA"],
            "sucursal": ["CAFAYATE", "CAFAYATE"],
            "cant_bultos": [5, 10],
            "cant_htls": [50, 100],
        })
        ruta = build_excel("2026-04-10", df, output_dir=tmp_path)
        wb = load_workbook(ruta)
        ws = wb.active

        # Should be alphabetically sorted by des_articulo, marca, generico
        # Col 1 = Articulo: BRAHMA LATA < VDS 500ML
        articulos_in_excel = [
            ws.cell(row=3, column=1).value,
            ws.cell(row=4, column=1).value,
        ]
        assert articulos_in_excel[0] == "BRAHMA LATA"
        assert articulos_in_excel[1] == "VDS 500ML"


# ── Genericos filter ────────────────────────────────────────────────────────


class TestStockDiarioGenericosFilter:
    def test_passes_genericos_to_data_loader(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(
                fecha_desde="2026-04-10",
                fecha_hasta="2026-04-10",
                genericos=["CERVEZAS", "AGUAS"],
            )
            service.generar_reporte(config)

        mock_loader.get_stock_diario.assert_called_once_with(
            "2026-04-10", ["CERVEZAS", "AGUAS"]
        )


# ── Supervisores split ──────────────────────────────────────────────────────


class TestStockDiarioSucursalesFilter:
    def test_filters_by_sucursales(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        df = pd.DataFrame({
            "generico": ["CERVEZAS"] * 3,
            "marca": ["BRAHMA"] * 3,
            "des_articulo": ["ART1"] * 3,
            "sucursal": ["CAFAYATE", "METAN", "ORAN"],
            "cant_bultos": [10, 20, 30],
            "cant_htls": [100, 200, 300],
        })

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = df

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(
                fecha_desde="2026-04-10",
                fecha_hasta="2026-04-10",
                sucursales=["ORAN"],
            )
            result = service.generar_reporte(config)

        wb = load_workbook(result.archivos_generados[0])
        ws = wb.active
        # Only 1 sucursal → bultos col 4, htls col 5
        assert ws.cell(row=2, column=4).value == "ORAN"
        assert ws.cell(row=2, column=5).value == "ORAN"

    def test_supervisor_name_in_filename(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = _make_stock_df()

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(
                fecha_desde="2026-04-10",
                fecha_hasta="2026-04-10",
                sucursales=["CAFAYATE", "METAN", "SALTA"],
                supervisor="Walter Vilte",
            )
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 1
        assert "Stock Walter Vilte" in result.archivos_generados[0].name
        assert "10-04-2026" in result.archivos_generados[0].name

    def test_no_matching_sucursales_produces_empty(self, tmp_path):
        from src.services.stock_diario.service import StockDiarioConfig, StockDiarioService

        df = pd.DataFrame({
            "generico": ["CERVEZAS"],
            "marca": ["BRAHMA"],
            "des_articulo": ["ART1"],
            "sucursal": ["CAFAYATE"],
            "cant_bultos": [10],
            "cant_htls": [100],
        })

        mock_loader = MagicMock(spec=DataLoader)
        mock_loader.get_stock_diario.return_value = df

        with patch("src.services.stock_diario.processor.DATA_OUTPUT", tmp_path):
            service = StockDiarioService(data_loader=mock_loader)
            config = StockDiarioConfig(
                fecha_desde="2026-04-10",
                fecha_hasta="2026-04-10",
                sucursales=["ORAN"],  # no data for ORAN
            )
            result = service.generar_reporte(config)

        assert len(result.archivos_generados) == 0
        assert "2026-04-10" in result.fechas_sin_datos
