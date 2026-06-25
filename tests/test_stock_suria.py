"""
Tests for stock_suria service — strict TDD.

Test IDs:
  TSS-001: configs/stock_suria_articulos.json has 170 matched + 17 sin_match
  TSS-002: Every matched entry has id_articulo + esquema
  TSS-003: Resumen totals add up (matched + sin_match == total_activos)
  TSS-004: REPORT_HANDLERS["stock-suria"] registered
  TSS-005: build_excel produces 3 sheets with correct names
  TSS-006: RESUMEN sheet has 10 rows; row 1 col A starts with "RESUMEN DEL MATCH"
  TSS-007: Stock SURIA sheet freeze_panes == "G3"
  TSS-008: Stock SURIA sheet row 2 col 1 = "Cod Prov", col 7 = "ABRA PAMPA"
  TSS-009: Stock SURIA sheet banner row 1 col 7 = "BULTOS", col 13 = "HTLs"
  TSS-010: Sin Match sheet freeze_panes == "F3"
  TSS-011: Sin Match sheet row 2 col 5 = "Sim %", col 6 = "ABRA PAMPA"
  TSS-012: Sim% cell number_format = "0%"
  TSS-013: No rounding — float bultos stored as-is
  TSS-014: Stock SURIA sheet has 172 rows (170 data + 2 header rows)
  TSS-015: HTL zero stored as None
"""

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_minimal_config():
    """Minimal config_data with 3 articulos and 2 sin_match for unit tests."""
    return {
        "resumen": {
            "total_activos": 5,
            "matched": 3,
            "sin_match": 2,
            "por_esquema": {"40": 2, "400": 1, "pelado": 0},
        },
        "articulos": [
            {
                "cod_prov": 101,
                "id_articulo": 400101,
                "esquema": "400",
                "desc_prov": "Coca 350ml",
                "desc_suria": "COCA COLA 350",
                "marca": "COCA COLA",
            },
            {
                "cod_prov": 102,
                "id_articulo": 400102,
                "esquema": "400",
                "desc_prov": "Fanta 500ml",
                "desc_suria": "FANTA 500",
                "marca": "FANTA",
            },
            {
                "cod_prov": 103,
                "id_articulo": 40103,
                "esquema": "40",
                "desc_prov": "Sprite 1L",
                "desc_suria": "SPRITE 1000",
                "marca": "SPRITE",
            },
        ],
        "sin_match": [
            {
                "cod_prov": 201,
                "desc_prov": "Schweppes 1.5L",
                "closest_id": 400201,
                "closest_desc": "SCHWEPPES 2.25L",
                "closest_sim": 0.5,
            },
            {
                "cod_prov": 202,
                "desc_prov": "Fanta Lata 220ml",
                "closest_id": 400202,
                "closest_desc": "FANTA LATA 354",
                "closest_sim": 0.3,
            },
        ],
    }


def _make_stock_data():
    """Minimal stock_data dict for 3 matched articles + 2 closest articles."""
    return {
        400101: {"JUJUY": {"bultos": 12.0, "htls": 144.0}},
        400102: {"ABRA PAMPA": {"bultos": 5.0, "htls": 0.0}},
        40103: {},
        # closest IDs for sin_match
        400201: {"PERICO": {"bultos": 3.0, "htls": 36.0}},
        400202: {},
    }


def _make_generico_map():
    return {400101: "GASEOSAS", 400102: "GASEOSAS", 40103: "GASEOSAS"}


# ── TSS-001 / TSS-002 / TSS-003: Config JSON integrity ───────────────────────


class TestConfigJsonIntegrity:
    """Validates that the frozen article match file matches expected shape."""

    @pytest.fixture(scope="class")
    def config_data(self):
        config_path = Path("configs/stock_suria_articulos.json")
        with open(config_path, encoding="utf-8") as fh:
            return json.load(fh)

    def test_matched_count(self, config_data):
        """TSS-001a: 170 matched articles."""
        assert len(config_data["articulos"]) == 170

    def test_sin_match_count(self, config_data):
        """TSS-001b: 17 sin_match articles."""
        assert len(config_data["sin_match"]) == 17

    def test_every_matched_has_id_articulo(self, config_data):
        """TSS-002a: every matched entry has id_articulo."""
        for art in config_data["articulos"]:
            assert "id_articulo" in art, f"Missing id_articulo: {art}"

    def test_every_matched_has_esquema(self, config_data):
        """TSS-002b: every matched entry has esquema."""
        for art in config_data["articulos"]:
            assert "esquema" in art, f"Missing esquema: {art}"

    def test_resumen_totals_add_up(self, config_data):
        """TSS-003: matched + sin_match == total_activos."""
        resumen = config_data["resumen"]
        assert resumen["matched"] + resumen["sin_match"] == resumen["total_activos"]

    def test_resumen_por_esquema_adds_up(self, config_data):
        """TSS-003b: sum of por_esquema values == matched."""
        resumen = config_data["resumen"]
        esquema_total = sum(resumen["por_esquema"].values())
        assert esquema_total == resumen["matched"]


# ── TSS-004: REPORT_HANDLERS registration ─────────────────────────────────────


class TestReportHandlerRegistration:
    def test_handler_registered(self):
        """TSS-004: REPORT_HANDLERS["stock-suria"] == "_run_stock_suria_report"."""
        import main

        assert main.REPORT_HANDLERS.get("stock-suria") == "_run_stock_suria_report"

    def test_handler_callable_exists(self):
        """TSS-004b: the handler function is defined in main module."""
        import main

        assert callable(getattr(main, "_run_stock_suria_report", None))


# ── TSS-005: 3 sheets with correct names ──────────────────────────────────────


class TestBuildExcelSheets:
    def test_three_sheets(self, tmp_path):
        """TSS-005: build_excel produces exactly 3 sheets."""
        from src.services.stock_suria.processor import build_excel

        config_data = _make_minimal_config()
        stock_data = _make_stock_data()
        generico_map = _make_generico_map()

        path = build_excel(config_data, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        assert len(wb.sheetnames) == 3

    def test_sheet_names(self, tmp_path):
        """TSS-005b: sheet names are exactly as specified."""
        from src.services.stock_suria.processor import build_excel

        config_data = _make_minimal_config()
        stock_data = _make_stock_data()
        generico_map = _make_generico_map()

        path = build_excel(config_data, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "RESUMEN DEL MATCH",
            "Stock SURIA",
            "ARTICULOS SIN MATCH POR CODIGO",
        ]


# ── TSS-006: RESUMEN sheet structure ─────────────────────────────────────────


class TestResumenSheet:
    @pytest.fixture
    def ws_resumen(self, tmp_path):
        from src.services.stock_suria.processor import build_excel

        path = build_excel(
            _make_minimal_config(), _make_stock_data(), _make_generico_map(), "2026-06-25", tmp_path
        )
        return load_workbook(path)["RESUMEN DEL MATCH"]

    def test_row_count(self, ws_resumen):
        """TSS-006a: RESUMEN sheet has exactly 10 rows."""
        assert ws_resumen.max_row == 10

    def test_row1_starts_with_resumen(self, ws_resumen):
        """TSS-006b: row 1 col A starts with 'RESUMEN DEL MATCH'."""
        val = ws_resumen.cell(1, 1).value
        assert val is not None
        assert val.startswith("RESUMEN DEL MATCH")

    def test_row4_header(self, ws_resumen):
        """TSS-006c: row 4 is the column header row."""
        assert ws_resumen.cell(4, 1).value == "Tipo de match"
        assert ws_resumen.cell(4, 2).value == "N articulos"

    def test_row8_total_matcheados(self, ws_resumen):
        """TSS-006d: row 8 is TOTAL MATCHEADOS with matched count."""
        assert ws_resumen.cell(8, 1).value == "TOTAL MATCHEADOS"
        assert ws_resumen.cell(8, 2).value == 3  # minimal config has 3

    def test_row10_total_activos(self, ws_resumen):
        """TSS-006e: row 10 is TOTAL ACTIVOS PROVEEDOR."""
        assert ws_resumen.cell(10, 1).value == "TOTAL ACTIVOS PROVEEDOR"
        assert ws_resumen.cell(10, 2).value == 5  # minimal config total_activos


# ── TSS-007 / TSS-008 / TSS-009: Stock SURIA sheet structure ─────────────────


class TestStockSuriaSheet:
    @pytest.fixture
    def ws_stock(self, tmp_path):
        from src.services.stock_suria.processor import build_excel

        path = build_excel(
            _make_minimal_config(), _make_stock_data(), _make_generico_map(), "2026-06-25", tmp_path
        )
        return load_workbook(path)["Stock SURIA"]

    def test_freeze_panes(self, ws_stock):
        """TSS-007: freeze_panes == 'G3'."""
        assert ws_stock.freeze_panes == "G3"

    def test_row2_col1_cod_prov(self, ws_stock):
        """TSS-008a: row 2 col 1 = 'Cod Prov'."""
        assert ws_stock.cell(2, 1).value == "Cod Prov"

    def test_row2_col7_abra_pampa(self, ws_stock):
        """TSS-008b: row 2 col 7 = 'ABRA PAMPA' (first sucursal for BULTOS)."""
        assert ws_stock.cell(2, 7).value == "ABRA PAMPA"

    def test_row2_col13_abra_pampa_htls(self, ws_stock):
        """TSS-008c: row 2 col 13 = 'ABRA PAMPA' (first sucursal for HTLs)."""
        assert ws_stock.cell(2, 13).value == "ABRA PAMPA"

    def test_banner_bultos(self, ws_stock):
        """TSS-009a: banner row 1 col 7 = 'BULTOS'."""
        assert ws_stock.cell(1, 7).value == "BULTOS"

    def test_banner_htls(self, ws_stock):
        """TSS-009b: banner row 1 col 13 = 'HTLs'."""
        assert ws_stock.cell(1, 13).value == "HTLs"

    def test_banner_bultos_fill(self, ws_stock):
        """TSS-009c: BULTOS banner has blue fill #4472C4."""
        fill = ws_stock.cell(1, 7).fill
        assert fill.fgColor.rgb.upper().endswith("4472C4")

    def test_banner_htls_fill(self, ws_stock):
        """TSS-009d: HTLs banner has green fill #70AD47."""
        fill = ws_stock.cell(1, 13).fill
        assert fill.fgColor.rgb.upper().endswith("70AD47")

    def test_total_rows(self, tmp_path):
        """TSS-014: Stock SURIA sheet has 172 rows with 170-article full config."""
        from src.services.stock_suria.processor import build_excel

        # Use the real config file to test the actual 170 articles
        config_path = Path("configs/stock_suria_articulos.json")
        with open(config_path, encoding="utf-8") as fh:
            full_config = json.load(fh)

        # Empty stock for all articles (just testing structure)
        matched_ids = [a["id_articulo"] for a in full_config["articulos"]]
        generico_map = {id_art: "GASEOSAS" for id_art in matched_ids}
        stock_data: dict = {}

        path = build_excel(full_config, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stock SURIA"]
        # 2 header rows + 170 data rows = 172
        assert ws.max_row == 172


# ── TSS-010 / TSS-011 / TSS-012: Sin Match sheet structure ───────────────────


class TestSinMatchSheet:
    @pytest.fixture
    def ws_sin_match(self, tmp_path):
        from src.services.stock_suria.processor import build_excel

        path = build_excel(
            _make_minimal_config(), _make_stock_data(), _make_generico_map(), "2026-06-25", tmp_path
        )
        return load_workbook(path)["ARTICULOS SIN MATCH POR CODIGO"]

    def test_freeze_panes(self, ws_sin_match):
        """TSS-010: freeze_panes == 'F3'."""
        assert ws_sin_match.freeze_panes == "F3"

    def test_row2_col5_sim_pct(self, ws_sin_match):
        """TSS-011a: row 2 col 5 = 'Sim %'."""
        assert ws_sin_match.cell(2, 5).value == "Sim %"

    def test_row2_col6_abra_pampa(self, ws_sin_match):
        """TSS-011b: row 2 col 6 = 'ABRA PAMPA' (first sucursal for BULTOS)."""
        assert ws_sin_match.cell(2, 6).value == "ABRA PAMPA"

    def test_sim_pct_number_format(self, ws_sin_match):
        """TSS-012: Sim % cell number_format == '0%'."""
        # Row 3, col 5 is the first data row's Sim % cell
        cell = ws_sin_match.cell(3, 5)
        assert cell.number_format == "0%"

    def test_sim_pct_value_stored_as_float(self, ws_sin_match):
        """TSS-012b: Sim % value is the raw float (0.5), not 50."""
        cell = ws_sin_match.cell(3, 5)
        # first sin_match entry has closest_sim=0.5
        assert cell.value == pytest.approx(0.5)


# ── TSS-013: No rounding rule ─────────────────────────────────────────────────


class TestNoRounding:
    def test_float_bultos_stored_as_float(self, tmp_path):
        """TSS-013: bultos float value stored without int() conversion."""
        from src.services.stock_suria.processor import build_excel

        config_data = _make_minimal_config()
        # Use a non-integer bultos value to prove no rounding
        stock_data = {
            400101: {"JUJUY": {"bultos": 12.7, "htls": 0.0}},
            400102: {},
            40103: {},
            400201: {},
            400202: {},
        }
        generico_map = _make_generico_map()

        path = build_excel(config_data, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stock SURIA"]

        # Find the row for article 400101 (sorted by marca, desc_suria)
        # COCA COLA sorts before FANTA and SPRITE
        # Row 3 = first data row
        found = False
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 3).value == 400101:
                # JUJUY is index 2 in SUCURSALES (0-based), so col = 7 + 2 = 9
                jujuy_col = 9  # col G=7 (ABRA PAMPA), H=8 (HUMAHUACA), I=9 (JUJUY)
                cell = ws.cell(r, jujuy_col)
                assert cell.value == pytest.approx(12.7), (
                    f"Expected 12.7, got {cell.value} — float was incorrectly converted"
                )
                found = True
                break

        assert found, "Article 400101 not found in sheet"


# ── TSS-015: HTL zero stored as None ──────────────────────────────────────────


class TestHtlZeroAsNone:
    def test_zero_htl_stored_as_none(self, tmp_path):
        """TSS-015: HTL value of 0 is stored as None (matches reference file)."""
        from src.services.stock_suria.processor import build_excel

        config_data = _make_minimal_config()
        stock_data = {
            400101: {"JUJUY": {"bultos": 10.0, "htls": 0.0}},  # htls=0 → None
            400102: {},
            40103: {},
            400201: {},
            400202: {},
        }
        generico_map = _make_generico_map()

        path = build_excel(config_data, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stock SURIA"]

        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 3).value == 400101:
                # JUJUY HTL col = 13 (ABRA PAMPA) + 2 (JUJUY offset) = 15
                jujuy_htl_col = 15
                cell = ws.cell(r, jujuy_htl_col)
                assert cell.value is None, (
                    f"Expected None for 0 HTL, got {cell.value}"
                )
                return

        pytest.fail("Article 400101 not found in sheet")

    def test_nonzero_htl_stored_as_value(self, tmp_path):
        """TSS-015b: non-zero HTL stored as actual float."""
        from src.services.stock_suria.processor import build_excel

        config_data = _make_minimal_config()
        stock_data = {
            400101: {"JUJUY": {"bultos": 10.0, "htls": 120.0}},
            400102: {},
            40103: {},
            400201: {},
            400202: {},
        }
        generico_map = _make_generico_map()

        path = build_excel(config_data, stock_data, generico_map, "2026-06-25", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stock SURIA"]

        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 3).value == 400101:
                jujuy_htl_col = 15
                cell = ws.cell(r, jujuy_htl_col)
                assert cell.value == pytest.approx(120.0)
                return

        pytest.fail("Article 400101 not found in sheet")
