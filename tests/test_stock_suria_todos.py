"""Tests for the all-articles Stock SURIA variant (build_excel_todos)."""
from openpyxl import load_workbook

from src.services.stock_suria.processor import build_excel_todos, SUCURSALES


def _articulos():
    return [
        {"id_articulo": 406914, "des_suria": "SCHNEIDER 710 * 24 LATAS",
         "marca": "SCHNEIDER", "generico": "CERVEZAS",
         "suc": {"LA QUIACA": {"bultos": 12, "htls": 1.5}, "JUJUY": {"bultos": 0, "htls": 0}}},
        {"id_articulo": 421071, "des_suria": "SCHNEIDER 473 * 24 LATAS",
         "marca": "SCHNEIDER", "generico": "CERVEZAS",
         "suc": {"PERICO": {"bultos": 30, "htls": 4.0}}},
        {"id_articulo": 100, "des_suria": "AGUA X",
         "marca": "AGUA", "generico": "AGUAS", "suc": {}},  # sin stock en ninguna
    ]


def test_build_excel_todos_all_articles_present(tmp_path):
    path = build_excel_todos(_articulos(), "2026-07-17", tmp_path)
    assert path.exists()
    ws = load_workbook(path).active
    assert ws.title == "Stock SURIA (todos)"
    # header row 2: descriptor + 2 blocks of sucursales
    assert [ws.cell(2, c).value for c in range(1, 5)] == ["Cod SURIA", "Desc SURIA", "Marca", "Generico"]
    # 3 articles -> rows 3,4,5
    cods = [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]
    assert set(cods) == {406914, 421071, 100}


def test_build_excel_todos_stock_values_by_sucursal(tmp_path):
    path = build_excel_todos(_articulos(), "2026-07-17", tmp_path)
    ws = load_workbook(path).active
    n_desc = 4
    la_quiaca_col = n_desc + 1 + SUCURSALES.index("LA QUIACA")  # bultos block
    perico_col = n_desc + 1 + SUCURSALES.index("PERICO")
    # find the SCHNEIDER 710 row (id 406914) and the 473 row (421071)
    rows = {ws.cell(r, 1).value: r for r in range(3, ws.max_row + 1)}
    assert ws.cell(rows[406914], la_quiaca_col).value == 12
    assert ws.cell(rows[421071], perico_col).value == 30
    # the zero-stock article shows 0 across all bultos columns
    for i in range(len(SUCURSALES)):
        assert ws.cell(rows[100], n_desc + 1 + i).value == 0
