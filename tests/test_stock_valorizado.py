"""Tests for the stock-valorizado service.

Covers the price-list loader, the universe policy (one test per case agreed in
docs/superpowers/specs/2026-08-07-stock-valorizado-design.md section 3), the
wide pivot, the analytics frames, and the service orchestration.
"""

from datetime import date, datetime, timedelta
from unittest.mock import Mock

import pandas as pd
import pytest
from openpyxl.utils import get_column_letter

from src.services.stock_valorizado.config import StockValorizadoConfig
from src.services.stock_valorizado.precios import (
    COLUMNA_ID,
    COLUMNA_PRECIO,
    COLUMNA_PRECIO_FINAL,
    cargar_lista_precios,
    estado_lista_precios,
)
from src.services.stock_valorizado.processor import (
    NO_VENDIBLES,
    SIN_CLASIFICAR,
    abc_pareto,
    build_universe,
    generico_x_sucursal,
    ordenar_sucursales,
    pivot_wide,
    resumen_sucursal,
)
from src.services.stock_valorizado.service import StockValorizadoService


# ── fixtures ────────────────────────────────────────────────────────────


def _stock_row(id_articulo, sucursal, bultos, generico="CERVEZAS", marca="SALTA"):
    return {
        "id_articulo": id_articulo,
        "generico": generico,
        "marca": marca,
        "des_articulo": f"ART {id_articulo}",
        "sucursal": sucursal,
        "cant_bultos": bultos,
        "cant_htls": 0.0,
    }


@pytest.fixture
def stock_df():
    """Two articles across two sucursales, plus the edge cases."""
    return pd.DataFrame(
        [
            _stock_row(1, "CASA CENTRAL", 10.0),
            _stock_row(1, "SUCURSAL METAN", 5.0),
            _stock_row(2, "CASA CENTRAL", 3.0, generico="AGUAS DANONE", marca="VILLAVICENCIO"),
            _stock_row(2, "SUCURSAL METAN", 0.0, generico="AGUAS DANONE", marca="VILLAVICENCIO"),
            # non-sellable generico -> excluded
            _stock_row(3, "CASA CENTRAL", 999.0, generico="ENVASES CCU", marca="CCU"),
            # NULL generico -> kept as SIN CLASIFICAR
            _stock_row(4, "CASA CENTRAL", 7.0, generico=None, marca=None),
            # zero price with real stock -> kept at $0
            _stock_row(5, "CASA CENTRAL", 4.0),
            # negative stock -> kept with negative valuation
            _stock_row(6, "SUCURSAL METAN", -2.0),
        ]
    )


@pytest.fixture
def precios_df():
    return pd.DataFrame(
        {
            "id_articulo": [1, 2, 3, 4, 5, 6],
            "precio_base": [100.5, 200.0, 50.0, 0.0, 0.0, 300.25],
            # Precio Final = base * 1.21 + internos; el articulo 5 tiene base 0
            # y final != 0 (solo internos), igual que 15 casos reales.
            "precio_final": [121.605, 242.0, 60.5, 0.0, 30.0, 363.3025],
        }
    )


@pytest.fixture
def lista_precios_xlsx(tmp_path):
    """A minimal but structurally faithful ERP price-list export."""
    path = tmp_path / "lista_precios.xlsx"
    pd.DataFrame(
        {
            "Unnamed: 0": [None, None],
            COLUMNA_ID: [1011, 1012],
            "Descripción": ["CHOPP SALTA BLANCO * 30 LST", "CHOPP SALTA BLANCO * 50 LTS."],
            "Presentación": [1, 1],
            COLUMNA_PRECIO: [102981.380, 171633.350],
            COLUMNA_PRECIO_FINAL: [136297.975, 227160.283],
        }
    ).to_excel(path, index=False)
    return path


# ── cargar_lista_precios ────────────────────────────────────────────────


class TestCargarListaPrecios:
    def test_devuelve_id_y_precio_normalizados(self, lista_precios_xlsx):
        df = cargar_lista_precios(lista_precios_xlsx)

        assert list(df.columns) == ["id_articulo", "precio_base", "precio_final"]
        assert df["id_articulo"].tolist() == [1011, 1012]
        assert df["precio_base"].tolist() == [102981.380, 171633.350]
        assert df["precio_final"].tolist() == [136297.975, 227160.283]

    def test_no_redondea_el_precio(self, tmp_path):
        path = tmp_path / "p.xlsx"
        pd.DataFrame({COLUMNA_ID: [1], COLUMNA_PRECIO: [1234.56789],
                      COLUMNA_PRECIO_FINAL: [1493.8]}).to_excel(path, index=False)

        df = cargar_lista_precios(path)

        assert df["precio_base"].iloc[0] == 1234.56789

    def test_archivo_faltante_falla_fuerte(self, tmp_path):
        with pytest.raises(FileNotFoundError, match="lista de precios"):
            cargar_lista_precios(tmp_path / "no-existe.xlsx")

    def test_columna_faltante_falla_fuerte(self, tmp_path):
        path = tmp_path / "p.xlsx"
        pd.DataFrame({COLUMNA_ID: [1], "Otra": [2]}).to_excel(path, index=False)

        with pytest.raises(ValueError, match=COLUMNA_PRECIO):
            cargar_lista_precios(path)

    def test_articulo_duplicado_falla_fuerte(self, tmp_path):
        path = tmp_path / "p.xlsx"
        pd.DataFrame({COLUMNA_ID: [1, 1], COLUMNA_PRECIO: [10.0, 20.0],
                      COLUMNA_PRECIO_FINAL: [12.1, 24.2]}).to_excel(path, index=False)

        with pytest.raises(ValueError, match="duplicad"):
            cargar_lista_precios(path)

    def test_precio_no_numerico_falla_fuerte(self, tmp_path):
        path = tmp_path / "p.xlsx"
        pd.DataFrame({COLUMNA_ID: [1], COLUMNA_PRECIO: ["s/precio"],
                      COLUMNA_PRECIO_FINAL: [1.0]}).to_excel(path, index=False)

        with pytest.raises(ValueError, match="numeric|numéric"):
            cargar_lista_precios(path)


# ── alerta de lista de precios vencida ──────────────────────────────────


def _envejecer(path, dias):
    """Backdate a file's mtime — age is read from mtime, which is what actually
    changes when the export is replaced."""
    import os

    ts = (datetime.now() - timedelta(days=dias)).timestamp()
    os.utime(path, (ts, ts))
    return path


class TestEstadoListaPrecios:
    def test_lista_fresca_no_esta_vencida(self, lista_precios_xlsx):
        estado = estado_lista_precios(lista_precios_xlsx, max_dias=30)

        assert estado.dias == 0
        assert estado.vencida is False
        assert "ATENCIÓN" not in estado.leyenda

    def test_justo_en_el_limite_todavia_no_vence(self, lista_precios_xlsx):
        estado = estado_lista_precios(_envejecer(lista_precios_xlsx, 30), max_dias=30)

        assert estado.dias == 30
        assert estado.vencida is False

    def test_un_dia_pasado_el_limite_vence(self, lista_precios_xlsx):
        estado = estado_lista_precios(_envejecer(lista_precios_xlsx, 31), max_dias=30)

        assert estado.dias == 31
        assert estado.vencida is True

    def test_la_leyenda_vencida_dice_que_hacer(self, lista_precios_xlsx):
        estado = estado_lista_precios(_envejecer(lista_precios_xlsx, 120), max_dias=30)

        leyenda = estado.leyenda
        assert "DESACTUALIZADA" in leyenda
        assert "120 días" in leyenda
        assert "A MANO" in leyenda  # el punto: nadie la actualiza sola
        assert "regenerá" in leyenda

    def test_el_umbral_es_configurable(self, lista_precios_xlsx):
        viejo = _envejecer(lista_precios_xlsx, 10)

        assert estado_lista_precios(viejo, max_dias=30).vencida is False
        assert estado_lista_precios(viejo, max_dias=7).vencida is True

    def test_avisa_por_log_cuando_vence(self, lista_precios_xlsx, caplog):
        import logging

        with caplog.at_level(logging.WARNING):
            estado_lista_precios(_envejecer(lista_precios_xlsx, 90), max_dias=30)

        assert any("VENCIDA" in m for m in caplog.messages)


# ── universe policy (design spec section 3) ─────────────────────────────


class TestPoliticaDeUniverso:
    def test_excluye_genericos_no_vendibles(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        assert 3 not in universe["id_articulo"].tolist()

    def test_exclusion_es_case_e_espacio_insensible(self, precios_df):
        stock = pd.DataFrame([_stock_row(3, "CASA CENTRAL", 5.0, generico="  envases ccu ")])

        universe = build_universe(stock, precios_df, genericos_excluidos=["ENVASES CCU"])

        assert universe.empty

    def test_generico_null_sobrevive_como_sin_clasificar(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        fila = universe.loc[universe["id_articulo"] == 4].iloc[0]
        assert fila["generico"] == SIN_CLASIFICAR
        assert fila["cant_bultos"] == 7.0

    def test_precio_cero_sobrevive_valorizado_en_cero(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        fila = universe.loc[universe["id_articulo"] == 5].iloc[0]
        assert fila["cant_bultos"] == 4.0
        assert fila["valorizado"] == 0.0

    def test_stock_negativo_conserva_valorizacion_negativa(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        fila = universe.loc[universe["id_articulo"] == 6].iloc[0]
        assert fila["cant_bultos"] == -2.0
        assert fila["valorizado"] == -600.50

    def test_filas_en_cero_permanecen(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        cero = universe.loc[(universe["id_articulo"] == 2) & (universe["sucursal"] == "SUCURSAL METAN")]
        assert len(cero) == 1
        assert cero.iloc[0]["cant_bultos"] == 0.0

    def test_valorizado_es_bultos_por_precio_sin_redondeo(self, precios_df):
        stock = pd.DataFrame([_stock_row(1, "CASA CENTRAL", 3.0)])

        universe = build_universe(stock, precios_df)

        assert universe["valorizado"].iloc[0] == 3.0 * 100.5

    def test_articulo_sin_precio_no_desaparece(self):
        """A stock article absent from the price list keeps its bultos at $0."""
        stock = pd.DataFrame([_stock_row(99, "CASA CENTRAL", 8.0)])
        precios = pd.DataFrame(
            {"id_articulo": [1], "precio_base": [10.0], "precio_final": [12.1]}
        )

        universe = build_universe(stock, precios)

        assert universe["cant_bultos"].iloc[0] == 8.0
        assert universe["precio_base"].iloc[0] == 0.0
        assert universe["valorizado"].iloc[0] == 0.0
        assert universe["precio_final"].iloc[0] == 0.0
        assert universe["valorizado_final"].iloc[0] == 0.0

    def test_bultos_nan_se_coalescen_a_cero(self, precios_df):
        stock = pd.DataFrame([_stock_row(1, "CASA CENTRAL", float("nan"))])

        universe = build_universe(stock, precios_df)

        assert universe["cant_bultos"].iloc[0] == 0.0
        assert universe["valorizado"].iloc[0] == 0.0


# ── valorización a Precio Final ─────────────────────────────────────────


class TestPrecioFinal:
    def test_falta_la_columna_precio_final_falla_fuerte(self, tmp_path):
        """El informe valoriza dos veces; un archivo con una sola columna es
        el export equivocado y no puede pasar en silencio."""
        path = tmp_path / "p.xlsx"
        pd.DataFrame({COLUMNA_ID: [1], COLUMNA_PRECIO: [10.0]}).to_excel(path, index=False)

        with pytest.raises(ValueError, match=COLUMNA_PRECIO_FINAL):
            cargar_lista_precios(path)

    def test_valorizado_final_es_bultos_por_precio_final(self, precios_df):
        stock = pd.DataFrame([_stock_row(1, "CASA CENTRAL", 3.0)])

        universe = build_universe(stock, precios_df)

        assert universe["valorizado_final"].iloc[0] == 3.0 * 121.605

    def test_base_cero_con_final_no_cero_se_valoriza_en_la_hoja_final(self, stock_df, precios_df):
        """Artículo 5: base 0, final 30 — invisible en la hoja base, visible en
        la final. Es el caso de los 15 artículos reales con solo internos."""
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        fila = universe.loc[universe["id_articulo"] == 5].iloc[0]
        assert fila["valorizado"] == 0.0
        assert fila["valorizado_final"] == 4.0 * 30.0

    def test_pivot_wide_usa_la_columna_de_valor_pedida(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        base = pivot_wide(universe)
        final = pivot_wide(universe, valor_col="valorizado_final")

        fila_base = base.loc[base[("", "Artículo")] == 1].iloc[0]
        fila_final = final.loc[final[("", "Artículo")] == 1].iloc[0]
        assert fila_base[("CASA CENTRAL", "Valorizado")] == pytest.approx(10.0 * 100.5)
        assert fila_final[("CASA CENTRAL", "Valorizado")] == pytest.approx(10.0 * 121.605)

    def test_las_dos_hojas_comparten_los_bultos(self, stock_df, precios_df):
        """Cambia el precio, no el stock: los bultos tienen que ser idénticos."""
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        base = pivot_wide(universe).sort_values(("", "Artículo"))
        final = pivot_wide(universe, valor_col="valorizado_final").sort_values(("", "Artículo"))

        assert base[("Total", "Total Bultos")].sum() == final[("Total", "Total Bultos")].sum()


# ── ordenar_sucursales ──────────────────────────────────────────────────


class TestOrdenarSucursales:
    def test_casa_central_va_primero(self):
        orden = ordenar_sucursales(["SUCURSAL METAN", "CASA CENTRAL", "SUCURSAL ORAN"])

        assert orden[0] == "CASA CENTRAL"

    def test_sucursal_desconocida_va_al_final_alfabetica(self):
        orden = ordenar_sucursales(["SUCURSAL NUEVA", "CASA CENTRAL", "SUCURSAL METAN"])

        assert orden == ["CASA CENTRAL", "SUCURSAL METAN", "SUCURSAL NUEVA"]


# ── pivot_wide ──────────────────────────────────────────────────────────


class TestPivotWide:
    def test_dos_columnas_por_sucursal(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        wide = pivot_wide(universe)

        for sucursal in ("CASA CENTRAL", "SUCURSAL METAN"):
            assert (sucursal, "Bultos") in wide.columns
            assert (sucursal, "Valorizado") in wide.columns

    def test_una_fila_por_articulo(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        wide = pivot_wide(universe)

        assert len(wide) == universe["id_articulo"].nunique()

    def test_articulo_ausente_en_sucursal_lee_cero_no_nan(self, precios_df):
        stock = pd.DataFrame(
            [
                _stock_row(1, "CASA CENTRAL", 10.0),
                _stock_row(2, "SUCURSAL METAN", 4.0),
            ]
        )
        universe = build_universe(stock, precios_df)

        wide = pivot_wide(universe)

        fila = wide.loc[wide[("", "Artículo")] == 1].iloc[0]
        assert fila[("SUCURSAL METAN", "Bultos")] == 0.0
        assert fila[("SUCURSAL METAN", "Valorizado")] == 0.0

    def test_totales_de_fila_cierran_contra_las_sucursales(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        wide = pivot_wide(universe)

        fila = wide.loc[wide[("", "Artículo")] == 1].iloc[0]
        assert fila[("Total", "Total Bultos")] == 15.0
        assert fila[("Total", "Total Valorizado")] == pytest.approx(15.0 * 100.5)

    def test_ordenado_por_valorizado_descendente(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        wide = pivot_wide(universe)

        valores = wide[("Total", "Total Valorizado")].tolist()
        assert valores == sorted(valores, reverse=True)

    def test_identidad_conserva_descripcion_y_marca(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        wide = pivot_wide(universe)

        fila = wide.loc[wide[("", "Artículo")] == 1].iloc[0]
        assert fila[("", "Descripción")] == "ART 1"
        assert fila[("", "Marca")] == "SALTA"


# ── analytics sheets ────────────────────────────────────────────────────


class TestResumenSucursal:
    def test_totaliza_bultos_y_plata_por_sucursal(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        resumen = resumen_sucursal(universe)

        cc = resumen.loc[resumen["Sucursal"] == "CASA CENTRAL"].iloc[0]
        assert cc["Bultos"] == 10.0 + 3.0 + 7.0 + 4.0
        assert cc["Valorizado"] == pytest.approx(10.0 * 100.5 + 3.0 * 200.0)

    def test_participacion_suma_uno(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        resumen = resumen_sucursal(universe)

        assert resumen["% Capital"].sum() == pytest.approx(1.0)

    def test_ordenado_por_valorizado_descendente(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        resumen = resumen_sucursal(universe)

        assert resumen["Valorizado"].tolist() == sorted(resumen["Valorizado"].tolist(), reverse=True)


class TestAbcPareto:
    def test_clasifica_a_b_c_por_valor_acumulado(self):
        wide = pd.DataFrame(
            {
                ("", "Artículo"): [1, 2, 3],
                ("", "Descripción"): ["A", "B", "C"],
                ("", "Genérico"): ["G", "G", "G"],
                ("", "Marca"): ["M", "M", "M"],
                ("Total", "Total Bultos"): [1.0, 1.0, 1.0],
                ("Total", "Total Valorizado"): [800.0, 150.0, 50.0],
            }
        )
        wide.columns = pd.MultiIndex.from_tuples(wide.columns)

        abc = abc_pareto(wide)

        assert abc["Clase"].tolist() == ["A", "B", "C"]
        assert abc["% Acumulado"].tolist() == pytest.approx([0.80, 0.95, 1.00])

    def test_ignora_articulos_sin_valor(self):
        wide = pd.DataFrame(
            {
                ("", "Artículo"): [1, 2],
                ("", "Descripción"): ["A", "B"],
                ("", "Genérico"): ["G", "G"],
                ("", "Marca"): ["M", "M"],
                ("Total", "Total Bultos"): [1.0, 1.0],
                ("Total", "Total Valorizado"): [100.0, 0.0],
            }
        )
        wide.columns = pd.MultiIndex.from_tuples(wide.columns)

        abc = abc_pareto(wide)

        assert abc["Artículo"].tolist() == [1]


class TestGenericoXSucursal:
    def test_matriz_valorizada_por_generico_y_sucursal(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        matriz = generico_x_sucursal(universe)

        cervezas = matriz.loc[matriz["Genérico"] == "CERVEZAS"].iloc[0]
        assert cervezas["CASA CENTRAL"] == pytest.approx(10.0 * 100.5)
        assert cervezas["SUCURSAL METAN"] == pytest.approx(5.0 * 100.5 - 2.0 * 300.25)

    def test_incluye_columna_total(self, stock_df, precios_df):
        universe = build_universe(stock_df, precios_df, genericos_excluidos=NO_VENDIBLES)

        matriz = generico_x_sucursal(universe)

        assert "Total" in matriz.columns
        fila = matriz.loc[matriz["Genérico"] == "AGUAS DANONE"].iloc[0]
        assert fila["Total"] == pytest.approx(3.0 * 200.0)


# ── service orchestration ───────────────────────────────────────────────


class TestStockValorizadoService:
    def test_falla_si_no_hay_snapshot_de_stock(self, lista_precios_xlsx):
        loader = Mock()
        loader.get_ultima_fecha_stock.return_value = None
        service = StockValorizadoService(data_loader=loader)

        with pytest.raises(ValueError, match="fact_stock"):
            service.generar_reporte(
                StockValorizadoConfig(lista_precios_path=str(lista_precios_xlsx))
            )

    def test_genera_el_archivo_y_reporta_totales(self, tmp_path, stock_df, monkeypatch):
        lista = tmp_path / "precios.xlsx"
        pd.DataFrame(
            {
                COLUMNA_ID: [1, 2, 3, 4, 5, 6],
                COLUMNA_PRECIO: [100.5, 200.0, 50.0, 0.0, 0.0, 300.25],
                COLUMNA_PRECIO_FINAL: [121.605, 242.0, 60.5, 0.0, 30.0, 363.3025],
            }
        ).to_excel(lista, index=False)

        loader = Mock()
        loader.get_ultima_fecha_stock.return_value = date(2026, 8, 7)
        loader.get_stock_diario.return_value = stock_df

        service = StockValorizadoService(data_loader=loader)
        monkeypatch.setattr(service, "_output_dir", lambda _fecha: tmp_path / "out")

        result = service.generar_reporte(
            StockValorizadoConfig(lista_precios_path=str(lista))
        )

        assert result.archivo_generado.exists()
        assert result.fecha_stock == date(2026, 8, 7)
        assert result.n_articulos == 5  # article 3 excluded (ENVASES CCU)
        assert result.total_bultos == 10.0 + 5.0 + 3.0 + 0.0 + 7.0 + 4.0 - 2.0
        assert result.total_valorizado == pytest.approx(
            15.0 * 100.5 + 3.0 * 200.0 - 2.0 * 300.25
        )

    def test_el_workbook_tiene_las_cinco_hojas(self, tmp_path, stock_df, monkeypatch):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        wb = load_workbook(ruta)

        assert wb.sheetnames == [
            "Stock Valorizado", "Stock Valorizado Final", "Resumen Sucursal",
            "ABC Pareto", "Generico x Sucursal", "Control",
        ]

    def test_hoja_principal_tiene_dos_columnas_por_sucursal_y_total_general(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        ws = load_workbook(ruta)["Stock Valorizado"]

        assert [ws.cell(5, c).value for c in range(5, 9)] == ["Bultos", "$", "Bultos", "$"]
        assert ws.cell(ws.max_row, 1).value == "TOTAL GENERAL"
        assert ws.cell(ws.max_row, 5).value.startswith("=SUBTOTAL(9,E6:")
        # The autofilter must stop above TOTAL GENERAL, or filtering folds the
        # total in with the article rows.
        assert ws.auto_filter.ref.endswith(str(ws.max_row - 1))

    def test_total_de_articulos_con_stock_no_se_duplica_por_sucursal(
        self, tmp_path, stock_df, monkeypatch
    ):
        """Article 1 sits in two sucursales; the total must count it once."""
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        ws = load_workbook(ruta)["Resumen Sucursal"]

        total_row = ws.max_row
        por_sucursal = sum(ws.cell(r, 3).value for r in range(5, total_row))
        # 5 articles carry stock; the naive sum over sucursales would say more.
        assert ws.cell(total_row, 3).value == 5
        assert por_sucursal > 5

    def test_valor_promedio_del_total_no_es_suma_de_promedios(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        ws = load_workbook(ruta)["Resumen Sucursal"]

        assert ws.cell(ws.max_row, 6).value == "=IFERROR(D19/B19,0)".replace(
            "19", str(ws.max_row)
        )

    def test_matriz_generico_sucursal_tiene_formato_moneda(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        ws = load_workbook(ruta)["Generico x Sucursal"]

        assert ws.cell(5, 2).number_format == "$ #,##0"

    def test_columnas_de_plata_sin_decimales_y_ancho_14_5(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        wb = load_workbook(ruta)

        for hoja in ("Stock Valorizado", "Stock Valorizado Final"):
            ws = wb[hoja]
            # F = primera columna de plata (4 de identidad + Bultos + $)
            assert ws.cell(6, 6).number_format == "$ #,##0", hoja
            assert ws.column_dimensions["F"].width == 14.5, hoja
        # y tambien en las hojas de analitica
        assert wb["Resumen Sucursal"].column_dimensions["D"].width == 14.5

    def test_el_total_usa_subtotal_para_seguir_al_filtro(
        self, tmp_path, stock_df, monkeypatch
    ):
        """SUM ignora el autofiltro; SUBTOTAL(9,...) no. El total tiene que
        moverse cuando el usuario filtra."""
        from openpyxl import load_workbook

        wb = load_workbook(self._generar(tmp_path, stock_df, monkeypatch))

        for hoja in ("Stock Valorizado", "Stock Valorizado Final"):
            ws = wb[hoja]
            total_row = ws.max_row
            formula = ws.cell(total_row, 5).value
            assert formula == f"=SUBTOTAL(9,E6:E{total_row - 1})", hoja
            assert "SUM(" not in formula, hoja

    def test_fila_espejo_arriba_apunta_al_subtotal(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        wb = load_workbook(self._generar(tmp_path, stock_df, monkeypatch))

        for hoja in ("Stock Valorizado", "Stock Valorizado Final"):
            ws = wb[hoja]
            total_row = ws.max_row
            ultima = get_column_letter(ws.max_column)
            assert ws.cell(3, 1).value == "TOTAL VISIBLE"
            assert ws.cell(3, 5).value == f"=E{total_row}", hoja
            assert ws.cell(3, ws.max_column).value == f"={ultima}{total_row}", hoja
            # tiene que quedar dentro del panel congelado para verse siempre
            assert ws.freeze_panes == "E6", hoja

    def test_los_rangos_se_acoplan_al_largo_de_la_tabla(self, tmp_path, monkeypatch):
        """Dos corridas con distinta cantidad de articulos: los rangos y la
        referencia del espejo tienen que seguir el largo real."""
        from openpyxl import load_workbook

        chico = pd.DataFrame([_stock_row(1, "CASA CENTRAL", 5.0)])
        grande = pd.DataFrame(
            [_stock_row(i, "CASA CENTRAL", float(i)) for i in range(1, 7)]
        )

        filas = {}
        for etiqueta, df in (("chico", chico), ("grande", grande)):
            ws = load_workbook(
                self._generar(tmp_path / etiqueta, df, monkeypatch)
            )["Stock Valorizado"]
            total_row = ws.max_row
            filas[etiqueta] = total_row
            assert ws.cell(total_row, 5).value == f"=SUBTOTAL(9,E6:E{total_row - 1})"
            assert ws.cell(3, 5).value == f"=E{total_row}"

        assert filas["chico"] != filas["grande"]

    def test_el_autofiltro_deja_afuera_la_fila_de_total(
        self, tmp_path, stock_df, monkeypatch
    ):
        """Si el total entra en el rango del filtro, SUBTOTAL se cuenta a si
        mismo y el filtro lo trata como un articulo mas."""
        from openpyxl import load_workbook

        ws = load_workbook(self._generar(tmp_path, stock_df, monkeypatch))["Stock Valorizado"]

        ultima = get_column_letter(ws.max_column)
        assert ws.auto_filter.ref == f"A5:{ultima}{ws.max_row - 1}"

    def test_lista_fresca_no_pinta_alerta(self, tmp_path, stock_df, monkeypatch):
        from openpyxl import load_workbook

        wb = load_workbook(self._generar(tmp_path, stock_df, monkeypatch))

        for hoja in ("Stock Valorizado", "Stock Valorizado Final", "Control"):
            ws = wb[hoja]
            assert "DESACTUALIZADA" not in (ws["A2"].value or ""), hoja
            assert ws["A2"].fill.fgColor.rgb not in ("FFC00000", "00C00000"), hoja

    def test_lista_vencida_pinta_banner_rojo_en_las_dos_hojas(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        wb = load_workbook(
            self._generar(tmp_path, stock_df, monkeypatch, dias_lista=95)
        )

        for hoja in ("Stock Valorizado", "Stock Valorizado Final"):
            ws = wb[hoja]
            assert "DESACTUALIZADA" in ws["A2"].value, hoja
            assert "95 días" in ws["A2"].value, hoja
            # el aviso va primero: la plata de abajo es sospechosa
            assert ws["A2"].value.startswith("⚠"), hoja
            assert "C00000" in str(ws["A2"].fill.fgColor.rgb), hoja

    def test_control_reporta_el_estado_de_la_lista(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ws = load_workbook(
            self._generar(tmp_path, stock_df, monkeypatch, dias_lista=95)
        )["Control"]

        assert ws.cell(4, 1).value == "Estado de la lista de precios"
        etiquetas = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(5, 10)}
        assert etiquetas["Antigüedad (días)"] == 95
        assert "VENCIDA" in etiquetas["Estado"]
        assert "A MANO" in etiquetas["Se carga"]

    def test_el_resultado_expone_la_bandera_de_vencida(
        self, tmp_path, stock_df, monkeypatch
    ):
        """main.py y cualquier consumidor tienen que poder avisar sin re-leer
        el archivo."""
        fresco = self._resultado(tmp_path / "f", stock_df, monkeypatch)
        vencido = self._resultado(tmp_path / "v", stock_df, monkeypatch, dias_lista=95)

        assert fresco.lista_precios_vencida is False
        assert fresco.lista_precios_dias == 0
        assert vencido.lista_precios_vencida is True
        assert vencido.lista_precios_dias == 95

    def test_los_bloques_por_sucursal_siguen_siendo_colapsables(
        self, tmp_path, stock_df, monkeypatch
    ):
        """El agrupamiento no puede volver a comerse los anchos: openpyxl
        `group()` borra las dimensiones del rango, por eso vamos con
        outline_level por columna."""
        from openpyxl import load_workbook

        ws = load_workbook(self._generar(tmp_path, stock_df, monkeypatch))["Stock Valorizado"]

        assert ws.column_dimensions["E"].outline_level == 1
        assert ws.column_dimensions["F"].outline_level == 1
        assert ws.column_dimensions["F"].width == 14.5
        # las de identidad y las de Total quedan fuera del grupo
        assert ws.column_dimensions["A"].outline_level == 0

    def test_la_hoja_final_valoriza_con_precio_final(
        self, tmp_path, stock_df, monkeypatch
    ):
        from openpyxl import load_workbook

        ruta = self._generar(tmp_path, stock_df, monkeypatch)
        wb = load_workbook(ruta)

        def total_de(hoja):
            ws = wb[hoja]
            col = ws.max_column
            return [ws.cell(r, col).value for r in range(6, ws.max_row)]

        # Los bultos son los mismos; la plata no.
        base, final = wb["Stock Valorizado"], wb["Stock Valorizado Final"]
        assert base.max_row == final.max_row
        assert sum(total_de("Stock Valorizado Final")) > sum(total_de("Stock Valorizado"))
        assert "PRECIO FINAL" in final["A1"].value

    @classmethod
    def _generar(cls, tmp_path, stock_df, monkeypatch, dias_lista=0):
        return cls._resultado(tmp_path, stock_df, monkeypatch, dias_lista).archivo_generado

    @staticmethod
    def _resultado(tmp_path, stock_df, monkeypatch, dias_lista=0):
        tmp_path.mkdir(parents=True, exist_ok=True)
        lista = tmp_path / "precios.xlsx"
        pd.DataFrame(
            {
                COLUMNA_ID: [1, 2, 3, 4, 5, 6],
                COLUMNA_PRECIO: [100.5, 200.0, 50.0, 0.0, 0.0, 300.25],
                COLUMNA_PRECIO_FINAL: [121.605, 242.0, 60.5, 0.0, 30.0, 363.3025],
            }
        ).to_excel(lista, index=False)
        if dias_lista:
            _envejecer(lista, dias_lista)

        loader = Mock()
        loader.get_ultima_fecha_stock.return_value = date(2026, 8, 7)
        loader.get_stock_diario.return_value = stock_df

        service = StockValorizadoService(data_loader=loader)
        monkeypatch.setattr(service, "_output_dir", lambda _fecha: tmp_path / "out")
        return service.generar_reporte(
            StockValorizadoConfig(lista_precios_path=str(lista))
        )

    def test_usa_fecha_stock_del_config_si_viene(self, tmp_path, stock_df, monkeypatch):
        lista = tmp_path / "precios.xlsx"
        pd.DataFrame({COLUMNA_ID: [1], COLUMNA_PRECIO: [10.0],
                      COLUMNA_PRECIO_FINAL: [12.1]}).to_excel(lista, index=False)

        loader = Mock()
        loader.get_stock_diario.return_value = stock_df
        service = StockValorizadoService(data_loader=loader)
        monkeypatch.setattr(service, "_output_dir", lambda _fecha: tmp_path / "out")

        service.generar_reporte(
            StockValorizadoConfig(lista_precios_path=str(lista), fecha_stock="2026-07-31")
        )

        loader.get_ultima_fecha_stock.assert_not_called()
        loader.get_stock_diario.assert_called_once_with("2026-07-31", None)
