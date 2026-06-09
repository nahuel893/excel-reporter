"""Tests for src/services/ventas_articulo/processor.py — build_excel."""

import calendar
from datetime import date

import openpyxl
import pytest


def _build(tmp_path, anio, mes, ventas_por_fecha=None, articulo_nombre="TEST ART"):
    """Helper: call build_excel and return (wb, ws, path)."""
    from src.services.ventas_articulo.processor import build_excel

    ruta = build_excel(
        anio=anio,
        mes=mes,
        articulo_nombre=articulo_nombre,
        id_articulo=99,
        id_sucursal=1,
        ventas_por_fecha=ventas_por_fecha or {},
        nombre_archivo="test_output",
        output_dir=tmp_path,
    )
    wb = openpyxl.load_workbook(str(ruta))
    ws = wb.active
    return wb, ws, ruta


class TestWorkbookStructure:
    def test_workbook_has_correct_structure(self, tmp_path):
        """One sheet; title in A1; row 2 empty; headers in row 3; 30 data rows; TOTAL at row 34."""
        wb, ws, _ = _build(tmp_path, 2026, 4)

        # One sheet
        assert len(wb.sheetnames) == 1

        # Title cell A1 contains the article name, id, sucursal, month and year
        title = str(ws["A1"].value)
        assert "TEST ART" in title
        assert "id 99" in title
        assert "Sucursal 1" in title
        assert "Abril" in title
        assert "2026" in title

        # Row 2 empty
        row2_vals = [ws.cell(row=2, column=c).value for c in range(1, 4)]
        assert all(v is None for v in row2_vals)

        # Row 3: headers
        assert ws.cell(row=3, column=1).value == "Día"
        assert ws.cell(row=3, column=2).value == "Fecha"
        assert ws.cell(row=3, column=3).value == "Bultos"

        # 30 data rows for April (rows 4..33)
        for r in range(4, 34):
            assert ws.cell(row=r, column=1).value is not None

        # TOTAL row at 34
        total_val = ws.cell(row=34, column=1).value
        assert total_val == "TOTAL"


class TestSundayFill:
    def test_sunday_rows_have_pink_fill(self, tmp_path):
        """April 2026 Sundays (5, 12, 19, 26) → rows 8, 15, 22, 29 have pink fill."""
        _, ws, _ = _build(tmp_path, 2026, 4)

        # Apr 5 = row 4 + 4 = row 8
        sunday_rows = [8, 15, 22, 29]
        for row in sunday_rows:
            fill = ws.cell(row=row, column=1).fill
            rgb = fill.fgColor.rgb
            assert "FFCDD2" in rgb, f"Row {row} expected pink but got {rgb}"


class TestConVentaFill:
    def test_day_with_sales_has_light_blue_fill(self, tmp_path):
        """Apr 2 (Thursday) with sales → row 5 fill is light blue D9E2F3."""
        ventas = {date(2026, 4, 2): 100.0}
        _, ws, _ = _build(tmp_path, 2026, 4, ventas_por_fecha=ventas)

        # Apr 2 = row 4 + 1 = row 5
        fill = ws.cell(row=5, column=1).fill
        rgb = fill.fgColor.rgb
        assert "BBDEFB" in rgb, f"Expected light blue but got {rgb}"


class TestSinVentaFill:
    def test_day_without_sales_has_gray_fill(self, tmp_path):
        """Empty sales dict + non-Sunday row → fill is gray F2F2F2."""
        _, ws, _ = _build(tmp_path, 2026, 4)

        # Apr 1 = row 4 (Wednesday — not Sunday, no sales)
        fill = ws.cell(row=4, column=1).fill
        rgb = fill.fgColor.rgb
        assert "ECEFF1" in rgb, f"Expected gray but got {rgb}"


class TestTotalRow:
    def test_total_row_content_and_green_fill(self, tmp_path):
        """Total row: A='TOTAL', C=150.0 (float), fill green, font bold+white."""
        ventas = {date(2026, 4, 2): 100.0, date(2026, 4, 3): 50.0}
        _, ws, _ = _build(tmp_path, 2026, 4, ventas_por_fecha=ventas)

        total_row = 34
        assert ws.cell(row=total_row, column=1).value == "TOTAL"
        assert ws.cell(row=total_row, column=3).value == 150.0

        fill = ws.cell(row=total_row, column=1).fill
        rgb = fill.fgColor.rgb
        assert "A5D6A7" in rgb, f"Expected green fill but got {rgb}"

        font = ws.cell(row=total_row, column=1).font
        assert font.bold is True
        assert "1B5E20" in font.color.rgb


class TestBultosBlank:
    def test_bultos_blank_when_zero(self, tmp_path):
        """Days with no sales → Bultos cell value is None."""
        _, ws, _ = _build(tmp_path, 2026, 4)

        # Apr 1 = row 4, no sales
        assert ws.cell(row=4, column=3).value is None


class TestFebruaryDayCount:
    def test_february_28_days_row_count(self, tmp_path):
        """February 2025 (28 days) → data rows 4-31, TOTAL at row 32."""
        _, ws, _ = _build(tmp_path, 2025, 2)

        # Row 31 should have a day value (day 28)
        assert ws.cell(row=31, column=1).value is not None
        # Row 32 is TOTAL
        assert ws.cell(row=32, column=1).value == "TOTAL"
        # Row 33 should be empty (no day 29)
        assert ws.cell(row=33, column=1).value is None
