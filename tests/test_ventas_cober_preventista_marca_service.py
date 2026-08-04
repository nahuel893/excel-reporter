"""Tests for VentasCoberPreventistaMarcaService."""
import pandas as pd
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.services.ventas_cober_preventista_marca import (
    VentasCoberPreventistaMarcaConfig,
    VentasCoberPreventistaMarcaService,
)


def _raw():
    """Client 1 bought from FGUANTAY (LORENA) AND VCHAPUR (NAHUEL) → distinct-count test."""
    return pd.DataFrame({
        "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY", "NAHUEL RUEDA", "DIRECTA"],
        "id_cliente": [1, 2, 1, 3],
        "bultos": [10.0, 5.0, 8.0, 3.0],
    })


def _sheet_rows(ruta):
    ws = load_workbook(ruta).active
    return [[ws.cell(r, c).value for c in range(1, 5)] for r in range(1, ws.max_row + 1)]


def _make(tmp_path, raw):
    loader = MagicMock(spec=DataLoader)
    loader.get_ventas_cobertura_por_vendedor.return_value = raw
    service = VentasCoberPreventistaMarcaService(data_loader=loader)
    config = VentasCoberPreventistaMarcaConfig(
        marca="FULL SPORT", fecha_desde="2026-07-01", fecha_hasta="2026-07-31", id_sucursal=1,
    )
    with patch("src.services.ventas_cober_preventista_marca.service.service_output_dir",
               return_value=tmp_path):
        return service.generar_reporte(config), loader


def test_composite_key_query_args(tmp_path):
    result, loader = _make(tmp_path, _raw())
    _, kwargs = loader.get_ventas_cobertura_por_vendedor.call_args
    assert kwargs["marca"] == "FULL SPORT"
    assert kwargs["id_sucursal"] == 1
    assert result.ruta_archivo.exists()


def test_supervisor_mapping_and_totals(tmp_path):
    result, _ = _make(tmp_path, _raw())
    rows = _sheet_rows(result.ruta_archivo)
    # supervisor of each vendedor row
    vend = {r[0]: r[1] for r in rows if r[0] in ("LORENA TARITOLAY", "NAHUEL RUEDA", "DIRECTA")}
    assert vend["LORENA TARITOLAY"] == "FGUANTAY"
    assert vend["NAHUEL RUEDA"] == "VCHAPUR"
    assert vend["DIRECTA"] == "SIN SUPERVISOR"
    # per-preventista total bultos
    assert result.total_bultos == 26.0


def test_cobertura_is_not_additive(tmp_path):
    """Sum of supervisor coberturas (2+1+1=4) exceeds the true total (3 distinct clients)."""
    result, _ = _make(tmp_path, _raw())
    assert result.cobertura_total == 3  # clients {1,2,3}, client 1 shared across supervisors
    rows = _sheet_rows(result.ruta_archivo)
    # supervisor-section cobertura (col 4) for FGUANTAY should be 2 (clients 1,2)
    sup_cob = {r[0]: r[3] for r in rows if r[0] in ("FGUANTAY", "VCHAPUR", "SIN SUPERVISOR")}
    assert sup_cob["FGUANTAY"] == 2
    assert sup_cob["VCHAPUR"] == 1


def test_has_two_total_general_rows(tmp_path):
    result, _ = _make(tmp_path, _raw())
    rows = _sheet_rows(result.ruta_archivo)
    totals = [r for r in rows if r[0] == "TOTAL GENERAL"]
    assert len(totals) == 2  # one per section (convención: fila de totales)


def test_empty_raw_does_not_crash(tmp_path):
    result, _ = _make(tmp_path, pd.DataFrame(columns=["vendedor", "id_cliente", "bultos"]))
    assert result.ruta_archivo.exists()
    assert result.total_bultos == 0.0
    assert result.cobertura_total == 0


# ── Dos periodos (mes anterior + mes corriente) ──────────────────────────────

def _raw_julio():
    """Julio: clientes 1 y 2 a LORENA, cliente 1 tambien a NAHUEL, 3 a DIRECTA."""
    return pd.DataFrame({
        "vendedor": ["LORENA TARITOLAY", "LORENA TARITOLAY", "NAHUEL RUEDA", "DIRECTA"],
        "id_cliente": [1, 2, 1, 3],
        "bultos": [10.0, 5.0, 8.0, 3.0],
    })


def _raw_agosto():
    """Agosto: solo LORENA vendio, al mismo cliente 1. NAHUEL y DIRECTA no aparecen."""
    return pd.DataFrame({
        "vendedor": ["LORENA TARITOLAY"],
        "id_cliente": [1],
        "bultos": [4.0],
    })


def _make_dos_periodos(tmp_path):
    loader = MagicMock(spec=DataLoader)

    def por_periodo(*, marca, fecha_desde, fecha_hasta, id_sucursal):
        return _raw_agosto() if fecha_desde.startswith("2026-08") else _raw_julio()

    loader.get_ventas_cobertura_por_vendedor.side_effect = por_periodo
    service = VentasCoberPreventistaMarcaService(data_loader=loader)
    config = VentasCoberPreventistaMarcaConfig(
        marca="FULL SPORT",
        fecha_desde="2026-08-01",
        fecha_hasta="2026-08-03",
        id_sucursal=1,
        incluir_mes_anterior=True,
    )
    with patch("src.services.ventas_cober_preventista_marca.service.service_output_dir",
               return_value=tmp_path):
        return service.generar_reporte(config), loader


def _grid(ruta, ancho=6):
    ws = load_workbook(ruta).active
    return [[ws.cell(r, c).value for c in range(1, ancho + 1)] for r in range(1, ws.max_row + 1)]


def test_dos_periodos_deriva_la_ventana_anterior(tmp_path):
    """The July window is DERIVED from fecha_desde, never read from the config."""
    _, loader = _make_dos_periodos(tmp_path)
    ventanas = {
        (c.kwargs["fecha_desde"], c.kwargs["fecha_hasta"])
        for c in loader.get_ventas_cobertura_por_vendedor.call_args_list
    }
    assert ventanas == {("2026-07-01", "2026-07-31"), ("2026-08-01", "2026-08-03")}


def test_dos_periodos_encabezado_de_dos_niveles(tmp_path):
    result, _ = _make_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    # fila 5: grupo de mes (anterior primero); fila 6: medidas
    assert ws.cell(5, 3).value == "JULIO 2026"
    assert ws.cell(5, 5).value == "AGOSTO 2026"
    assert [ws.cell(6, c).value for c in range(1, 7)] == [
        "Vendedor", "Supervisor", "Bultos", "Cobertura", "Bultos", "Cobertura",
    ]


def test_dos_periodos_vendedor_sin_venta_en_agosto_queda_en_cero(tmp_path):
    result, _ = _make_dos_periodos(tmp_path)
    filas = {r[0]: r for r in _grid(result.ruta_archivo) if r[0] == "NAHUEL RUEDA"}
    assert filas["NAHUEL RUEDA"][2:6] == [8.0, 1, 0.0, 0]


def test_dos_periodos_cobertura_no_se_suma_entre_meses(tmp_path):
    """Cliente 1 compro en julio Y en agosto: cuenta 1 en cada mes, nunca 2."""
    result, _ = _make_dos_periodos(tmp_path)
    grid = _grid(result.ruta_archivo)
    lorena = next(r for r in grid if r[0] == "LORENA TARITOLAY")
    assert lorena[2:6] == [15.0, 2, 4.0, 1]   # jul: 2 clientes | ago: 1 cliente

    totales = [r for r in grid if r[0] == "TOTAL GENERAL"]
    # Julio: clientes distintos {1,2,3} = 3 (NO 2+1+1=4, que seria sumar niveles)
    assert totales[0][3] == 3
    # Agosto: solo cliente 1
    assert totales[0][5] == 1


def test_dos_periodos_totales_de_bultos_por_columna(tmp_path):
    result, _ = _make_dos_periodos(tmp_path)
    total_prev = next(r for r in _grid(result.ruta_archivo) if r[0] == "TOTAL GENERAL")
    assert total_prev[2] == 26.0   # 10 + 5 + 8 + 3 en julio
    assert total_prev[4] == 4.0    # solo LORENA en agosto


def test_un_solo_periodo_mantiene_el_layout_original(tmp_path):
    """Regression: sin el flag, el encabezado sigue en la fila 5 con 4 columnas."""
    result, loader = _make(tmp_path, _raw())
    ws = load_workbook(result.ruta_archivo).active
    assert [ws.cell(5, c).value for c in range(1, 5)] == [
        "Vendedor", "Supervisor", "Bultos", "Cobertura",
    ]
    assert loader.get_ventas_cobertura_por_vendedor.call_count == 1


# ── Objetivo de cobertura + clausula gatillo ─────────────────────────────────

def _raw_marca_base(mes: str):
    """Cobertura de la marca de referencia (SALTA). Julio mas ancha que agosto."""
    if mes.startswith("2026-08"):
        return pd.DataFrame({
            "vendedor": ["LORENA TARITOLAY"] * 2,
            "id_cliente": [1, 2],
            "bultos": [5.0, 5.0],
        })
    return pd.DataFrame({                       # julio: 10 pdv LORENA, 5 NAHUEL
        "vendedor": ["LORENA TARITOLAY"] * 10 + ["NAHUEL RUEDA"] * 5,
        "id_cliente": list(range(1, 11)) + list(range(11, 16)),
        "bultos": [3.0] * 15,
    })


def _make_objetivo(tmp_path, *, base_actual="anterior", gatillo=1800.0):
    loader = MagicMock(spec=DataLoader)

    def por_marca(*, marca, fecha_desde, fecha_hasta, id_sucursal):
        if marca == "SALTA":
            return _raw_marca_base(fecha_desde)
        return _raw_agosto() if fecha_desde.startswith("2026-08") else _raw_julio()

    loader.get_ventas_cobertura_por_vendedor.side_effect = por_marca
    service = VentasCoberPreventistaMarcaService(data_loader=loader)
    config = VentasCoberPreventistaMarcaConfig(
        marca="FULL SPORT",
        fecha_desde="2026-08-01", fecha_hasta="2026-08-04", id_sucursal=1,
        incluir_mes_anterior=True,
        objetivo_cobertura={"marca": "SALTA", "pct_anterior": 0.20,
                            "pct_actual": 0.25, "base_actual": base_actual},
        clausula_gatillo=gatillo,
    )
    with patch("src.services.ventas_cober_preventista_marca.service.service_output_dir",
               return_value=tmp_path):
        return service.generar_reporte(config), loader


def _fila(ws, etiqueta, col=1):
    return next(r for r in range(1, ws.max_row + 1) if ws.cell(r, col).value == etiqueta)


def test_objetivo_agrega_una_columna_por_bloque(tmp_path):
    result, _ = _make_objetivo(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    assert [ws.cell(6, c).value for c in range(1, 9)] == [
        "Vendedor", "Supervisor", "Bultos", "Cobertura", "Objetivo",
        "Bultos", "Cobertura", "Objetivo",
    ]


def test_objetivo_del_mes_anterior_es_el_pct_de_su_propio_mes(tmp_path):
    """Julio: 20% de la cobertura SALTA de julio (10 pdv LORENA -> 2)."""
    result, _ = _make_objetivo(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "LORENA TARITOLAY")
    assert ws.cell(r, 5).value == 2            # round(10 * 0.20)


def test_objetivo_del_mes_actual_usa_la_cobertura_del_mes_anterior(tmp_path):
    """Agosto: 25% de la cobertura SALTA de JULIO (10 pdv -> 3), no de agosto (2)."""
    result, _ = _make_objetivo(tmp_path, base_actual="anterior")
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "LORENA TARITOLAY")
    assert ws.cell(r, 8).value == 3            # round(10 * 0.25), no round(2 * 0.25)


def test_objetivo_del_mes_actual_puede_usar_su_propio_mes(tmp_path):
    """`base_actual='propio'` cambia la referencia a agosto (2 pdv -> 1)."""
    result, _ = _make_objetivo(tmp_path, base_actual="propio")
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "LORENA TARITOLAY")
    assert ws.cell(r, 8).value == 1            # round(2 * 0.25)


def test_objetivo_usa_umbral_mayor_a_cero_no_medio_bulto(tmp_path):
    """Este informe cuenta con `> 0`; el 0.5 es criterio de otro informe."""
    loader = MagicMock(spec=DataLoader)

    def por_marca(*, marca, fecha_desde, fecha_hasta, id_sucursal):
        if marca == "SALTA":
            # 4 pdv de LORENA, dos de ellos por debajo de medio bulto
            return pd.DataFrame({
                "vendedor": ["LORENA TARITOLAY"] * 4,
                "id_cliente": [1, 2, 3, 4],
                "bultos": [3.0, 3.0, 0.25, 0.1],
            })
        return _raw_julio()

    loader.get_ventas_cobertura_por_vendedor.side_effect = por_marca
    service = VentasCoberPreventistaMarcaService(data_loader=loader)
    config = VentasCoberPreventistaMarcaConfig(
        marca="FULL SPORT", fecha_desde="2026-07-01", fecha_hasta="2026-07-31",
        objetivo_cobertura={"marca": "SALTA", "pct_anterior": 0.5,
                            "pct_actual": 0.5, "base_actual": "propio"},
    )
    with patch("src.services.ventas_cober_preventista_marca.service.service_output_dir",
               return_value=tmp_path):
        result = service.generar_reporte(config)
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "LORENA TARITOLAY")
    assert ws.cell(r, 5).value == 2            # round(4 * 0.5); con >=0.5 seria round(2*0.5)=1


def test_clausula_gatillo_compara_el_volumen_de_cada_mes(tmp_path):
    result, _ = _make_objetivo(tmp_path, gatillo=1800.0)
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "CLAUSULA GATILLO")
    assert ws.cell(r, 3).value == 1800.0                        # julio
    assert abs(ws.cell(r, 4).value - 26.0 / 1800.0) < 1e-9      # % sobre el volumen
    assert ws.cell(r, 6).value == 1800.0                        # agosto
    assert abs(ws.cell(r, 7).value - 4.0 / 1800.0) < 1e-9


def test_clausula_gatillo_deja_una_fila_en_blanco_encima(tmp_path):
    result, _ = _make_objetivo(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    r = _fila(ws, "CLAUSULA GATILLO")
    assert all(ws.cell(r - 1, c).value in (None, "") for c in range(1, 9))


def test_sin_objetivo_ni_gatillo_el_layout_no_cambia(tmp_path):
    """Regression: los informes que no los configuran quedan igual que antes."""
    result, _ = _make_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    assert [ws.cell(6, c).value for c in range(1, 7)] == [
        "Vendedor", "Supervisor", "Bultos", "Cobertura", "Bultos", "Cobertura",
    ]
    assert ws.max_column == 6
    assert not any(ws.cell(r, 1).value == "CLAUSULA GATILLO"
                   for r in range(1, ws.max_row + 1))
