"""Tests for the ventas-marca report (volume + coverage by marca for one generico).

Rows = marcas of the generico. Each period contributes two measures:

- **bultos**: sales volume, from `gold.fact_ventas`.
- **cobertura**: buying clients, from the ETL's `gold.cob_*` tables.

The two sources are deliberately different. Coverage is a business definition
that lives in the ETL and is NEVER recomputed here: recounting clients from the
fact gives a different number and drags in rules (bonificados, anulados) the ETL
already settled.

The TOTAL row's coverage comes from `cob_sucursal_generico`, not from summing the
marcas. Coverage IS additive across rutas and preventistas (a client belongs to
one route), but NOT across marcas — the same client buys several.
"""
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook

from src.core.data_loader import DataLoader
from src.services.ventas_marca import VentasMarcaConfig, VentasMarcaService


# ── Loader queries ───────────────────────────────────────────────────────────

def _capture(loader, retorno):
    cap: dict = {}

    def fake(q, params=None):
        cap["q"] = q
        cap["p"] = params
        return retorno

    loader.execute_query = MagicMock(side_effect=fake)
    return cap


def test_ventas_query_sums_volume_by_marca():
    loader = DataLoader()
    cap = _capture(loader, pd.DataFrame(columns=["marca", "bultos"]))
    loader.get_ventas_por_marca(
        generico="PERNOD RICARD", fecha_desde="2026-07-03", fecha_hasta="2026-07-03",
        id_sucursal=1,
    )
    q = cap["q"].lower()
    assert "gold.fact_ventas" in q
    assert "sum(f.cantidades_total)" in q
    assert "fecha_comprobante::date between" in q
    assert cap["p"]["generico"] == "PERNOD RICARD"
    assert cap["p"]["id_sucursal"] == 1


def test_ventas_query_does_not_compute_coverage():
    """Volumen y nada mas: la cobertura no se recalcula desde el fact."""
    loader = DataLoader()
    cap = _capture(loader, pd.DataFrame(columns=["marca", "bultos"]))
    loader.get_ventas_por_marca(
        generico="PERNOD RICARD", fecha_desde="2026-07-01", fecha_hasta="2026-07-31",
    )
    q = cap["q"].lower()
    assert "id_cliente" not in q
    assert "count(distinct" not in q


def test_cobertura_marca_reads_the_etl_table():
    loader = DataLoader()
    cap = _capture(loader, pd.DataFrame(columns=["marca", "cobertura"]))
    loader.get_cobertura_marca_de_generico(
        generico="PERNOD RICARD", periodo="2026-07-01", id_sucursal=1,
    )
    q = cap["q"].lower()
    assert "gold.cob_sucursal_marca" in q
    assert "clientes_compradores" in q
    assert "gold.fact_ventas" not in q            # nunca desde el fact
    assert cap["p"] == {"generico": "PERNOD RICARD", "periodo": "2026-07-01",
                        "id_sucursal": 1}


def test_cobertura_generico_reads_the_aggregated_table_not_a_sum():
    """El total sale de cob_sucursal_generico, no de sumar cob_sucursal_marca."""
    loader = DataLoader()
    cap = _capture(loader, pd.DataFrame({"cobertura": [721]}))
    total = loader.get_cobertura_generico(
        generico="PERNOD RICARD", periodo="2026-07-01", id_sucursal=1,
    )
    q = cap["q"].lower()
    assert "gold.cob_sucursal_generico" in q
    assert "sum(" not in q                        # ya viene agregada
    assert total == 721


def test_cobertura_generico_sin_fila_devuelve_cero():
    loader = DataLoader()
    _capture(loader, pd.DataFrame(columns=["cobertura"]))
    assert loader.get_cobertura_generico(
        generico="PERNOD RICARD", periodo="2030-01-01", id_sucursal=1) == 0


# ── Service ──────────────────────────────────────────────────────────────────

def _loader(ventas_por_periodo, cob_por_periodo, total_por_periodo):
    """Fake loader keyed by period start ('2026-07' / '2026-08')."""
    loader = MagicMock(spec=DataLoader)

    def _key(fecha):
        return fecha[:7]

    loader.get_ventas_por_marca.side_effect = (
        lambda *, generico, fecha_desde, fecha_hasta, id_sucursal=1:
        ventas_por_periodo[_key(fecha_desde)]
    )
    loader.get_cobertura_marca_de_generico.side_effect = (
        lambda *, generico, periodo, id_sucursal=1: cob_por_periodo[_key(periodo)]
    )
    loader.get_cobertura_generico.side_effect = (
        lambda *, generico, periodo, id_sucursal=1: total_por_periodo[_key(periodo)]
    )
    return loader


def _loader_un_periodo():
    return _loader(
        ventas_por_periodo={"2026-07": pd.DataFrame({
            "marca": ["CUSENIER", "ABSOLUT", None],
            "bultos": [27.04, 18.17, 1.5],
        })},
        cob_por_periodo={"2026-07": pd.DataFrame({
            "marca": ["CUSENIER", "ABSOLUT"],
            "cobertura": [345, 23],
        })},
        total_por_periodo={"2026-07": 721},
    )


def _generar(tmp_path, loader, config):
    service = VentasMarcaService(data_loader=loader)
    with patch("src.services.ventas_marca.service.service_output_dir", return_value=tmp_path):
        return service.generar_reporte(config)


def _cfg(**kw):
    base = dict(generico="PERNOD RICARD", fecha="2026-07-03")
    base.update(kw)
    return VentasMarcaConfig(**base)


def test_builds_marca_rows_and_total(tmp_path):
    result = _generar(tmp_path, _loader_un_periodo(), _cfg())
    ws = load_workbook(result.ruta_archivo).active
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "CUSENIER" in texts
    assert "ABSOLUT" in texts
    assert "TOTAL GENERAL" in texts          # totals-row convention
    assert "(sin marca)" in texts            # NULL marca handled

    nums = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))]
    assert any(abs(n - 46.71) < 1e-6 for n in nums)  # 27.04 + 18.17 + 1.5, NOT rounded


def test_un_periodo_muestra_bultos_y_cobertura_de_la_tabla(tmp_path):
    result = _generar(tmp_path, _loader_un_periodo(), _cfg())
    ws = load_workbook(result.ruta_archivo).active
    assert [ws.cell(4, c).value for c in range(1, 4)] == ["Marca", "Bultos", "Cobertura"]
    filas = {
        ws.cell(r, 1).value: (ws.cell(r, 2).value, ws.cell(r, 3).value)
        for r in range(5, ws.max_row + 1)
    }
    assert filas["CUSENIER"] == (27.04, 345)
    assert filas["ABSOLUT"] == (18.17, 23)
    assert filas["(sin marca)"] == (1.5, 0)   # sin fila en la tabla → 0, no desaparece


def test_consulta_la_cobertura_por_periodo_mensual(tmp_path):
    """La tabla cob_* es mensual: un rango de dias se traduce a su mes."""
    loader = _loader_un_periodo()
    _generar(tmp_path, loader, _cfg(fecha="2026-07-03", fecha_hasta="2026-07-15"))
    _, kwargs = loader.get_cobertura_marca_de_generico.call_args
    assert kwargs["periodo"] == "2026-07-01"
    _, kwargs_tot = loader.get_cobertura_generico.call_args
    assert kwargs_tot["periodo"] == "2026-07-01"


def test_total_de_cobertura_no_es_la_suma_de_las_marcas(tmp_path):
    """345 + 23 = 368, pero el total real es 721: sale de la tabla agregada."""
    result = _generar(tmp_path, _loader_un_periodo(), _cfg())
    ws = load_workbook(result.ruta_archivo).active
    fila = next(r for r in range(5, ws.max_row + 1)
                if ws.cell(r, 1).value == "TOTAL GENERAL")
    assert ws.cell(fila, 3).value == 721
    assert ws.cell(fila, 3).value != 368


def test_single_day_uses_same_desde_hasta(tmp_path):
    loader = _loader_un_periodo()
    _generar(tmp_path, loader, _cfg())
    _, kwargs = loader.get_ventas_por_marca.call_args
    assert kwargs["fecha_desde"] == "2026-07-03"
    assert kwargs["fecha_hasta"] == "2026-07-03"
    assert kwargs["generico"] == "PERNOD RICARD"


# ── Dos periodos (mes anterior + mes corriente) ──────────────────────────────

def _loader_dos_periodos():
    """July closed vs. August partial, with the real shape of the PERNOD data:
    a mass bonification blows BUHERO's August coverage past its July one."""
    return _loader(
        ventas_por_periodo={
            "2026-07": pd.DataFrame({"marca": ["BUHERO", "ABSOLUT"],
                                     "bultos": [222.08, 31.92]}),
            "2026-08": pd.DataFrame({"marca": ["BUHERO"], "bultos": [84.33]}),
        },
        cob_por_periodo={
            "2026-07": pd.DataFrame({"marca": ["BUHERO", "ABSOLUT"],
                                     "cobertura": [245, 23]}),
            "2026-08": pd.DataFrame({"marca": ["BUHERO"], "cobertura": [939]}),
        },
        total_por_periodo={"2026-07": 721, "2026-08": 951},
    )


def _generar_dos_periodos(tmp_path):
    loader = _loader_dos_periodos()
    config = _cfg(fecha="2026-08-01", fecha_hasta="2026-08-03", incluir_mes_anterior=True)
    return _generar(tmp_path, loader, config), loader


def test_dos_periodos_deriva_la_ventana_anterior(tmp_path):
    """The July window is DERIVED from fecha, never read from the config."""
    _, loader = _generar_dos_periodos(tmp_path)
    ventanas = {
        (c.kwargs["fecha_desde"], c.kwargs["fecha_hasta"])
        for c in loader.get_ventas_por_marca.call_args_list
    }
    assert ventanas == {("2026-07-01", "2026-07-31"), ("2026-08-01", "2026-08-03")}
    periodos = {c.kwargs["periodo"] for c in loader.get_cobertura_generico.call_args_list}
    assert periodos == {"2026-07-01", "2026-08-01"}


def test_dos_periodos_encabezado_de_dos_niveles(tmp_path):
    result, _ = _generar_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    assert ws.cell(4, 2).value == "JULIO 2026"      # mes cerrado primero
    assert ws.cell(4, 4).value == "AGOSTO 2026"
    assert [ws.cell(5, c).value for c in range(1, 6)] == [
        "Marca", "Bultos", "Cobertura", "Bultos", "Cobertura",
    ]


def test_dos_periodos_marca_ausente_en_un_mes_queda_en_cero(tmp_path):
    result, _ = _generar_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    filas = {
        ws.cell(r, 1).value: [ws.cell(r, c).value for c in range(2, 6)]
        for r in range(6, ws.max_row + 1)
    }
    assert filas["BUHERO"] == [222.08, 245, 84.33, 939]
    assert filas["ABSOLUT"] == [31.92, 23, 0.0, 0]   # sin agosto, no desaparece


def test_dos_periodos_cada_mes_trae_su_propia_cobertura(tmp_path):
    """Cada columna sale de la tabla de SU periodo; no se cruzan ni se suman."""
    result, _ = _generar_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    fila = next(r for r in range(6, ws.max_row + 1)
                if ws.cell(r, 1).value == "TOTAL GENERAL")
    assert ws.cell(fila, 3).value == 721   # julio
    assert ws.cell(fila, 5).value == 951   # agosto
    assert ws.cell(fila, 3).value + ws.cell(fila, 5).value != ws.cell(fila, 5).value


def test_dos_periodos_total_de_bultos_por_columna_sin_redondeo(tmp_path):
    result, _ = _generar_dos_periodos(tmp_path)
    ws = load_workbook(result.ruta_archivo).active
    fila = next(r for r in range(6, ws.max_row + 1)
                if ws.cell(r, 1).value == "TOTAL GENERAL")
    assert abs(ws.cell(fila, 2).value - 254.0) < 1e-9   # 222.08 + 31.92
    assert abs(ws.cell(fila, 4).value - 84.33) < 1e-9


def test_un_solo_periodo_no_agrega_columnas_de_otro_mes(tmp_path):
    """Regression: sin el flag, la hoja no crece a dos periodos."""
    loader = _loader_un_periodo()
    result = _generar(tmp_path, loader, _cfg())
    ws = load_workbook(result.ruta_archivo).active
    assert ws.max_column == 3
    assert loader.get_ventas_por_marca.call_count == 1
